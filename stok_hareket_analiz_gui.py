"""
Stok Hareket Analiz Raporu GUI
Son X yılda hareketi olan ürünlerin aylık hareket analizi
Eşdeğer grup bazlı raporlama ve alt toplamlar
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import logging
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from decimal import Decimal
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class StokHareketAnalizGUI:
    """Stok Hareket Analiz Raporu Penceresi"""

    # Hareket tipleri tanımları
    HAREKET_TIPLERI = {
        'GIRIS': [
            ('FATURA_GIRIS', 'Faturalı Giriş'),
            ('TAKAS_GIRIS', 'Takas Giriş'),
        ],
        'CIKIS': [
            ('FATURA_CIKIS', 'Faturalı Çıkış'),
            ('TAKAS_CIKIS', 'Takas Çıkış'),
            ('IADE', 'İade'),
            ('RECETE_SATIS', 'Reçeteli Satış'),
            ('ELDEN_SATIS', 'Parakende Satış'),
        ]
    }

    # Renkler
    RENK_GIRIS = '#C8E6C9'  # Açık yeşil
    RENK_CIKIS = '#FFCDD2'  # Açık pembe
    RENK_ALT_TOPLAM = '#B0BEC5'  # Koyu gri
    RENK_GRUP_BASLIK = '#78909C'  # Daha koyu gri

    def __init__(self, parent):
        self.parent = parent
        self.parent.title("Stok Hareket Analiz Raporu")
        self.parent.geometry("1700x850")

        self.db = None
        self.veriler = []
        self.islenenmis_veriler = []  # Gruplandırılmış ve sıralanmış veriler

        # Checkbox değişkenleri
        self.hareket_vars = {}

        # Parametre değişkenleri
        self.yil_sayisi = tk.IntVar(value=2)
        self.ay_sayisi = tk.IntVar(value=6)

        # Ürün tipi çoklu seçim için
        self.urun_tipi_vars = {}  # {tip_adi: BooleanVar}
        self.urun_tipleri_listesi = []  # Veritabanından doldurulacak

        # Sıralama durumu
        self.sort_column = None
        self.sort_reverse = False

        # Sütun filtreleri {sutun_id: set(secili_degerler)}
        self.sutun_filtreleri = {}
        self.filtre_popup = None

        self._arayuz_olustur()
        self._baglanti_kur()

    def _arayuz_olustur(self):
        """Ana arayüzü oluştur"""
        # Ana frame
        main_frame = ttk.Frame(self.parent, padding=5)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Üst panel - Filtreler ve Ayarlar
        self._filtre_panel_olustur(main_frame)

        # Orta panel - Tablo
        self._tablo_olustur(main_frame)

        # Alt panel - Status bar
        self._status_bar_olustur(main_frame)

    def _filtre_panel_olustur(self, parent):
        """Filtre panelini oluştur"""
        filtre_frame = ttk.LabelFrame(parent, text="Parametreler ve Hareket Tipleri", padding=10)
        filtre_frame.pack(fill=tk.X, pady=(0, 5))

        # Satır 1 - Temel parametreler
        row1 = ttk.Frame(filtre_frame)
        row1.pack(fill=tk.X, pady=5)

        # X Yıl
        ttk.Label(row1, text="Son Kaç Yıl:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        yil_combo = ttk.Combobox(row1, textvariable=self.yil_sayisi, width=5, state="readonly")
        yil_combo['values'] = [1, 2, 3, 4, 5]
        yil_combo.set(2)
        yil_combo.pack(side=tk.LEFT, padx=(0, 20))

        # Y Ay
        ttk.Label(row1, text="Aylık Detay (Son Kaç Ay):", font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        ay_combo = ttk.Combobox(row1, textvariable=self.ay_sayisi, width=5, state="readonly")
        ay_combo['values'] = [3, 6, 9, 12, 18, 24, 36]
        ay_combo.set(6)
        ay_combo.pack(side=tk.LEFT, padx=(0, 20))

        # Ürün Tipi (Çoklu Seçim)
        ttk.Label(row1, text="Ürün Tipi:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 5))

        # Menubutton ile çoklu seçim
        self.urun_tipi_menubutton = tk.Menubutton(
            row1,
            text="Tümü Seçili",
            relief=tk.RAISED,
            width=20,
            font=('Arial', 9)
        )
        self.urun_tipi_menu = tk.Menu(self.urun_tipi_menubutton, tearoff=0)
        self.urun_tipi_menubutton["menu"] = self.urun_tipi_menu
        self.urun_tipi_menubutton.pack(side=tk.LEFT, padx=(0, 10))

        # Tümünü Seç / Temizle butonları
        ttk.Button(row1, text="Tümü", command=self._tum_tipleri_sec, width=5).pack(side=tk.LEFT, padx=(0, 2))
        ttk.Button(row1, text="Hiçbiri", command=self._tum_tipleri_temizle, width=6).pack(side=tk.LEFT, padx=(0, 20))

        # Sorgula ve Excel butonları
        ttk.Button(row1, text="Sorgula", command=self.sorgula, width=15).pack(side=tk.LEFT, padx=(20, 5))
        ttk.Button(row1, text="Excel Aktar", command=self.excel_aktar, width=15).pack(side=tk.LEFT, padx=(0, 5))

        # Satır 2 - Hareket tipleri checkboxları
        row2 = ttk.Frame(filtre_frame)
        row2.pack(fill=tk.X, pady=10)

        # GİRİŞ hareket tipleri
        giris_frame = ttk.LabelFrame(row2, text="GİRİŞ Hareketleri", padding=5)
        giris_frame.pack(side=tk.LEFT, padx=(0, 20))

        for kod, ad in self.HAREKET_TIPLERI['GIRIS']:
            var = tk.BooleanVar(value=False)
            self.hareket_vars[kod] = var
            cb = ttk.Checkbutton(giris_frame, text=ad, variable=var)
            cb.pack(side=tk.LEFT, padx=10)

        # ÇIKIŞ hareket tipleri
        cikis_frame = ttk.LabelFrame(row2, text="ÇIKIŞ Hareketleri", padding=5)
        cikis_frame.pack(side=tk.LEFT, padx=(0, 20))

        for kod, ad in self.HAREKET_TIPLERI['CIKIS']:
            # Reçeteli ve Parakende varsayılan olarak seçili
            varsayilan = kod in ['RECETE_SATIS', 'ELDEN_SATIS']
            var = tk.BooleanVar(value=varsayilan)
            self.hareket_vars[kod] = var
            cb = ttk.Checkbutton(cikis_frame, text=ad, variable=var)
            cb.pack(side=tk.LEFT, padx=10)

        # Hızlı seçim butonları
        hizli_frame = ttk.Frame(row2)
        hizli_frame.pack(side=tk.LEFT, padx=20)

        ttk.Button(hizli_frame, text="Tüm Girişler", command=self._tum_girisler_sec, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(hizli_frame, text="Tüm Çıkışlar", command=self._tum_cikislar_sec, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(hizli_frame, text="Temizle", command=self._secimi_temizle, width=10).pack(side=tk.LEFT, padx=2)

        # Satır 3 - Tablo filtresi
        row3 = ttk.Frame(filtre_frame)
        row3.pack(fill=tk.X, pady=5)

        ttk.Label(row3, text="Tablo Filtresi:").pack(side=tk.LEFT, padx=(0, 5))
        self.tablo_filtre_entry = ttk.Entry(row3, width=40)
        self.tablo_filtre_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.tablo_filtre_entry.bind('<KeyRelease>', self._tablo_filtrele)

        ttk.Button(row3, text="Filtreyi Temizle", command=self._filtreyi_temizle, width=15).pack(side=tk.LEFT)

        # Sütun filtrelerini temizle butonu
        ttk.Button(row3, text="Sütun Filtrelerini Temizle", command=self._tum_sutun_filtrelerini_temizle, width=22).pack(side=tk.LEFT, padx=(10, 0))

        # Filtre bilgi etiketi
        self.sutun_filtre_label = ttk.Label(row3, text="", foreground="blue")
        self.sutun_filtre_label.pack(side=tk.LEFT, padx=(10, 0))

    def _tablo_olustur(self, parent):
        """Tabloyu oluştur"""
        tablo_frame = ttk.Frame(parent)
        tablo_frame.pack(fill=tk.BOTH, expand=True)

        # Dinamik sütunlar oluşturulacak - başlangıçta temel sütunlar
        self.temel_sutunlar = [
            ("Yon", 60),
            ("UrunTipi", 100),
            ("UrunAdi", 280),
            ("Esdeger", 90),
            ("Stok", 60),
            ("TopCikis", 70),
            ("CikisAdet", 70),
            ("AylikOrt", 70),
            ("AylikBitis", 80),
        ]

        self.temel_basliklar = {
            "Yon": "Yön",
            "UrunTipi": "Ürün Tipi",
            "UrunAdi": "Ürün Adı",
            "Esdeger": "Eşdeğer",
            "Stok": "Stok",
            "TopCikis": "Top.Çıkış",
            "CikisAdet": "Çıkış Adet",
            "AylikOrt": "Aylık Ort",
            "AylikBitis": "Ay.Bitiş",
        }

        # Treeview oluştur - sütunlar sorguda dinamik olarak güncellenecek
        self.tree = ttk.Treeview(tablo_frame, columns=[], show='headings', height=25)

        # Scrollbarlar
        vsb = ttk.Scrollbar(tablo_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tablo_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Grid
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tablo_frame.columnconfigure(0, weight=1)
        tablo_frame.rowconfigure(0, weight=1)

        # Renk etiketleri
        self.tree.tag_configure('giris', background=self.RENK_GIRIS)
        self.tree.tag_configure('cikis', background=self.RENK_CIKIS)
        self.tree.tag_configure('alt_toplam', background=self.RENK_ALT_TOPLAM, font=('Arial', 9, 'bold'))
        self.tree.tag_configure('grup_baslik', background=self.RENK_GRUP_BASLIK, font=('Arial', 10, 'bold'))

        # Sağ tık menüsü için binding
        self.tree.bind('<Button-3>', self._sutun_filtre_menu)

    def _sutun_filtre_menu(self, event):
        """Sağ tık ile sütun filtre menüsünü göster"""
        # Tıklanan bölgeyi kontrol et
        region = self.tree.identify_region(event.x, event.y)
        if region == 'heading':
            # Tıklanan sütunu bul
            column = self.tree.identify_column(event.x)
            if column:
                col_index = int(column.replace('#', '')) - 1
                if col_index >= 0 and col_index < len(self.tree['columns']):
                    col_id = self.tree['columns'][col_index]
                    self._sutun_filtre_popup_ac(col_id, event.x_root, event.y_root)

    def _sutun_filtre_popup_ac(self, col_id, x, y):
        """Sütun filtre popup penceresini aç"""
        # Önceki popup varsa kapat
        if self.filtre_popup:
            self.filtre_popup.destroy()

        # Sütundaki benzersiz değerleri al
        degerler = set()
        for veri in self.islenenmis_veriler:
            if veri.get('satir_tipi') in ['giris', 'cikis']:  # Sadece veri satırları
                val = veri.get(col_id, '')
                if val != '' and val is not None:
                    degerler.add(str(val))

        if not degerler:
            return

        # Popup penceresi
        self.filtre_popup = tk.Toplevel(self.parent)
        self.filtre_popup.title(f"Filtre: {self.aktif_basliklar.get(col_id, col_id)}")
        self.filtre_popup.geometry(f"+{x}+{y}")
        self.filtre_popup.resizable(False, False)
        self.filtre_popup.transient(self.parent)
        self.filtre_popup.grab_set()

        # Ana frame
        main_frame = ttk.Frame(self.filtre_popup, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Üst butonlar
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Button(btn_frame, text="Tümünü Seç", width=12,
                   command=lambda: self._filtre_tumunu_sec(col_id, True)).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Temizle", width=12,
                   command=lambda: self._filtre_tumunu_sec(col_id, False)).pack(side=tk.LEFT, padx=2)

        # Arama kutusu
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=5)
        ttk.Label(search_frame, text="Ara:").pack(side=tk.LEFT)
        self.filtre_search_var = tk.StringVar()
        self.filtre_search_var.trace('w', lambda *args: self._filtre_arama_guncelle(col_id))
        search_entry = ttk.Entry(search_frame, textvariable=self.filtre_search_var, width=25)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Liste frame (scrollable)
        liste_frame = ttk.Frame(main_frame)
        liste_frame.pack(fill=tk.BOTH, expand=True)

        # Canvas ve scrollbar
        canvas = tk.Canvas(liste_frame, width=250, height=300)
        scrollbar = ttk.Scrollbar(liste_frame, orient="vertical", command=canvas.yview)
        self.filtre_checkbox_frame = ttk.Frame(canvas)

        self.filtre_checkbox_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.filtre_checkbox_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Mouse wheel binding
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Checkbox değişkenleri
        self.filtre_vars = {}
        mevcut_filtre = self.sutun_filtreleri.get(col_id, None)

        # Değerleri sırala (sayısal ise sayısal, değilse alfabetik)
        try:
            sirali_degerler = sorted(degerler, key=lambda x: float(x) if x.replace('.', '').replace('-', '').isdigit() else float('inf'))
        except:
            sirali_degerler = sorted(degerler)

        for deger in sirali_degerler:
            var = tk.BooleanVar(value=True if mevcut_filtre is None else deger in mevcut_filtre)
            self.filtre_vars[deger] = var
            cb = ttk.Checkbutton(self.filtre_checkbox_frame, text=deger, variable=var)
            cb.pack(anchor=tk.W, padx=5, pady=1)

        self.filtre_tum_checkboxlar = list(self.filtre_checkbox_frame.winfo_children())

        # Alt butonlar
        alt_btn_frame = ttk.Frame(main_frame)
        alt_btn_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(alt_btn_frame, text="Uygula", width=12,
                   command=lambda: self._filtre_uygula(col_id)).pack(side=tk.LEFT, padx=2)
        ttk.Button(alt_btn_frame, text="Filtreyi Kaldır", width=12,
                   command=lambda: self._filtre_kaldir(col_id)).pack(side=tk.LEFT, padx=2)
        ttk.Button(alt_btn_frame, text="İptal", width=12,
                   command=self.filtre_popup.destroy).pack(side=tk.RIGHT, padx=2)

        # Pencere kapanırken mousewheel binding'i kaldır
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            self.filtre_popup.destroy()
            self.filtre_popup = None

        self.filtre_popup.protocol("WM_DELETE_WINDOW", on_close)

    def _filtre_arama_guncelle(self, col_id):
        """Filtre listesinde arama yap"""
        arama = self.filtre_search_var.get().upper()
        for cb in self.filtre_tum_checkboxlar:
            text = cb.cget('text').upper()
            if arama in text:
                cb.pack(anchor=tk.W, padx=5, pady=1)
            else:
                cb.pack_forget()

    def _filtre_tumunu_sec(self, col_id, sec):
        """Tüm filtreleri seç veya temizle"""
        for var in self.filtre_vars.values():
            var.set(sec)

    def _filtre_uygula(self, col_id):
        """Sütun filtresini uygula"""
        secili_degerler = {deger for deger, var in self.filtre_vars.items() if var.get()}

        if len(secili_degerler) == len(self.filtre_vars):
            # Tümü seçiliyse filtre yok demek
            if col_id in self.sutun_filtreleri:
                del self.sutun_filtreleri[col_id]
        else:
            self.sutun_filtreleri[col_id] = secili_degerler

        self._filtreleri_uygula()
        self.filtre_popup.destroy()
        self.filtre_popup = None

    def _filtre_kaldir(self, col_id):
        """Sütun filtresini kaldır"""
        if col_id in self.sutun_filtreleri:
            del self.sutun_filtreleri[col_id]
        self._filtreleri_uygula()
        self.filtre_popup.destroy()
        self.filtre_popup = None

    def _filtreleri_uygula(self):
        """Tüm sütun filtrelerini uygula ve tabloyu güncelle"""
        if not self.sutun_filtreleri:
            # Filtre yoksa orijinal verileri göster
            self._tabloyu_guncelle(self.islenenmis_veriler)
            self._filtre_durumunu_guncelle()
            return

        # Filtrelenmiş verileri oluştur
        filtreli = []
        for veri in self.islenenmis_veriler:
            satir_tipi = veri.get('satir_tipi')

            # Grup başlıkları ve alt toplamlar için özel işlem
            if satir_tipi in ['grup_baslik', 'alt_toplam']:
                filtreli.append(veri)
                continue

            # Veri satırlarını filtrele
            dahil_et = True
            for col_id, secili_degerler in self.sutun_filtreleri.items():
                val = str(veri.get(col_id, ''))
                if val not in secili_degerler:
                    dahil_et = False
                    break

            if dahil_et:
                filtreli.append(veri)

        # Boş grupları temizle
        temiz_filtreli = self._bos_gruplari_temizle(filtreli)
        self._tabloyu_guncelle(temiz_filtreli)
        self._filtre_durumunu_guncelle()

    def _bos_gruplari_temizle(self, veriler):
        """Veri satırı olmayan grup başlıklarını ve alt toplamlarını kaldır"""
        sonuc = []
        i = 0
        while i < len(veriler):
            veri = veriler[i]
            if veri.get('satir_tipi') == 'grup_baslik':
                # Bu grubun veri satırları var mı kontrol et
                grup_verileri = []
                j = i + 1
                while j < len(veriler) and veriler[j].get('satir_tipi') != 'grup_baslik':
                    if veriler[j].get('satir_tipi') in ['giris', 'cikis']:
                        grup_verileri.append(veriler[j])
                    j += 1

                # Grup verisi varsa ekle
                if grup_verileri:
                    sonuc.append(veri)  # Grup başlığı
                    for gv in grup_verileri:
                        sonuc.append(gv)
                    # Alt toplamları da ekle
                    k = i + 1
                    while k < len(veriler) and veriler[k].get('satir_tipi') != 'grup_baslik':
                        if veriler[k].get('satir_tipi') == 'alt_toplam':
                            sonuc.append(veriler[k])
                        k += 1
                i = j
            else:
                sonuc.append(veri)
                i += 1

        return sonuc

    def _filtre_durumunu_guncelle(self):
        """Aktif filtre sayısını status bar ve label'da göster"""
        filtre_sayisi = len(self.sutun_filtreleri)
        if filtre_sayisi > 0:
            filtre_sutunlar = [self.aktif_basliklar.get(col, col) for col in self.sutun_filtreleri.keys()]
            self.sutun_filtre_label.config(text=f"Filtreli sütunlar: {', '.join(filtre_sutunlar)}")
        else:
            self.sutun_filtre_label.config(text="")

    def _tum_sutun_filtrelerini_temizle(self):
        """Tüm sütun filtrelerini temizle"""
        self.sutun_filtreleri.clear()
        self._tabloyu_guncelle(self.islenenmis_veriler)
        self._filtre_durumunu_guncelle()

    def _status_bar_olustur(self, parent):
        """Status bar oluştur"""
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill=tk.X, pady=(5, 0))

        self.status_label = ttk.Label(status_frame, text="Hazır", font=('Arial', 9))
        self.status_label.pack(side=tk.LEFT)

        self.kayit_label = ttk.Label(status_frame, text="", font=('Arial', 9))
        self.kayit_label.pack(side=tk.RIGHT)

    def _baglanti_kur(self):
        """Veritabanı bağlantısı kur"""
        try:
            from botanik_db import BotanikDB
            self.db = BotanikDB()
            if self.db.baglan():
                self.status_label.config(text="Veritabanı bağlantısı başarılı")
                # Ürün tiplerini yükle
                self._urun_tiplerini_yukle()
            else:
                self.status_label.config(text="Veritabanı bağlantı hatası!")
        except Exception as e:
            logger.error(f"Veritabanı bağlantı hatası: {e}")
            self.status_label.config(text=f"Hata: {e}")

    def _urun_tiplerini_yukle(self):
        """Veritabanından ürün tiplerini yükle ve menüyü oluştur"""
        try:
            if self.db:
                tipler = self.db.urun_tipleri_getir()
                self.urun_tipleri_listesi = [t['UrunTipAdi'] for t in tipler]

                # Menüyü temizle
                self.urun_tipi_menu.delete(0, tk.END)
                self.urun_tipi_vars.clear()

                # Her tip için checkbox ekle
                for tip_adi in self.urun_tipleri_listesi:
                    var = tk.BooleanVar(value=True)  # Varsayılan olarak hepsi seçili
                    self.urun_tipi_vars[tip_adi] = var
                    self.urun_tipi_menu.add_checkbutton(
                        label=tip_adi,
                        variable=var,
                        command=self._urun_tipi_secim_guncelle
                    )

                self._urun_tipi_secim_guncelle()
        except Exception as e:
            logger.error(f"Ürün tipleri yükleme hatası: {e}")

    def _urun_tipi_secim_guncelle(self):
        """Seçili ürün tiplerini göster"""
        secili_tipler = [tip for tip, var in self.urun_tipi_vars.items() if var.get()]
        toplam = len(self.urun_tipi_vars)
        secili = len(secili_tipler)

        if secili == 0:
            self.urun_tipi_menubutton.config(text="Hiçbiri Seçili")
        elif secili == toplam:
            self.urun_tipi_menubutton.config(text="Tümü Seçili")
        elif secili <= 2:
            self.urun_tipi_menubutton.config(text=", ".join(secili_tipler[:2]))
        else:
            self.urun_tipi_menubutton.config(text=f"{secili} tip seçili")

    def _tum_tipleri_sec(self):
        """Tüm ürün tiplerini seç"""
        for var in self.urun_tipi_vars.values():
            var.set(True)
        self._urun_tipi_secim_guncelle()

    def _tum_tipleri_temizle(self):
        """Tüm ürün tipi seçimlerini temizle"""
        for var in self.urun_tipi_vars.values():
            var.set(False)
        self._urun_tipi_secim_guncelle()

    def _tum_girisler_sec(self):
        """Tüm giriş hareketlerini seç"""
        for kod, _ in self.HAREKET_TIPLERI['GIRIS']:
            self.hareket_vars[kod].set(True)

    def _tum_cikislar_sec(self):
        """Tüm çıkış hareketlerini seç"""
        for kod, _ in self.HAREKET_TIPLERI['CIKIS']:
            self.hareket_vars[kod].set(True)

    def _secimi_temizle(self):
        """Tüm seçimleri temizle"""
        for var in self.hareket_vars.values():
            var.set(False)

    def _filtreyi_temizle(self):
        """Tablo filtresini temizle"""
        self.tablo_filtre_entry.delete(0, tk.END)
        self._tabloyu_guncelle(self.islenenmis_veriler)

    def _tablo_filtrele(self, event=None):
        """Tablodaki verileri filtrele"""
        filtre = self.tablo_filtre_entry.get().strip().upper()
        if not filtre:
            self._tabloyu_guncelle(self.islenenmis_veriler)
            return

        # Sadece ürün satırlarını filtrele, grup başlıkları ve alt toplamları koru
        filtreli_veriler = []
        aktif_grup = None
        grup_urunleri = []

        for veri in self.islenenmis_veriler:
            if veri.get('satir_tipi') == 'grup_baslik':
                # Önceki grubun ürünleri varsa ekle
                if aktif_grup and grup_urunleri:
                    filtreli_veriler.append(aktif_grup)
                    filtreli_veriler.extend(grup_urunleri)
                    # Alt toplamları da ekle
                    for v in self.islenenmis_veriler:
                        if v.get('satir_tipi') == 'alt_toplam' and v.get('EsdegerId') == aktif_grup.get('EsdegerId'):
                            filtreli_veriler.append(v)

                aktif_grup = veri
                grup_urunleri = []

            elif veri.get('satir_tipi') in ['giris', 'cikis']:
                urun_adi = str(veri.get('UrunAdi', '')).upper()
                if filtre in urun_adi:
                    grup_urunleri.append(veri)

        # Son grubu ekle
        if aktif_grup and grup_urunleri:
            filtreli_veriler.append(aktif_grup)
            filtreli_veriler.extend(grup_urunleri)
            for v in self.islenenmis_veriler:
                if v.get('satir_tipi') == 'alt_toplam' and v.get('EsdegerId') == aktif_grup.get('EsdegerId'):
                    filtreli_veriler.append(v)

        self._tabloyu_guncelle(filtreli_veriler)

    def _sutunlari_guncelle(self, ay_sayisi):
        """Aylık sütunları dinamik olarak güncelle"""
        # Sütun listesi
        sutunlar = list(self.temel_sutunlar)
        basliklar = dict(self.temel_basliklar)

        # Aylık sütunları ekle
        bugun = datetime.now()
        for i in range(ay_sayisi):
            ay_basi = (bugun - relativedelta(months=i)).replace(day=1)
            ay_adi = ay_basi.strftime('%b %y')  # Ör: Oca 24

            col_id = f"Ay{i+1}"
            sutunlar.append((col_id, 60))
            basliklar[col_id] = ay_adi

        # Treeview sütunlarını güncelle
        col_ids = [c[0] for c in sutunlar]
        self.tree['columns'] = col_ids

        for col_id, width in sutunlar:
            header_text = basliklar.get(col_id, col_id)
            self.tree.heading(col_id, text=header_text, command=lambda c=col_id: self._siralama_yap(c))
            self.tree.column(col_id, width=width, minwidth=40)

        self.aktif_sutunlar = sutunlar
        self.aktif_basliklar = basliklar

    def sorgula(self):
        """Veritabanından verileri sorgula"""
        # Seçili hareket tiplerini al
        secili_tipler = [kod for kod, var in self.hareket_vars.items() if var.get()]

        if not secili_tipler:
            messagebox.showwarning("Uyarı", "En az bir hareket tipi seçmelisiniz!")
            return

        # Giriş ve çıkış tiplerini ayır
        giris_tipleri = [t for t in secili_tipler if t in ['FATURA_GIRIS', 'TAKAS_GIRIS']]
        cikis_tipleri = [t for t in secili_tipler if t not in ['FATURA_GIRIS', 'TAKAS_GIRIS']]

        self.status_label.config(text="Sorgulanıyor...")
        self.parent.update()

        # Mevcut sütun filtrelerini temizle
        self.sutun_filtreleri.clear()
        self._filtre_durumunu_guncelle()

        def sorgu_thread():
            try:
                yil = self.yil_sayisi.get()
                ay = self.ay_sayisi.get()

                # Seçili ürün tiplerini al
                secili_urun_tipleri = [tip for tip, var in self.urun_tipi_vars.items() if var.get()]

                if not secili_urun_tipleri:
                    self.parent.after(0, lambda: messagebox.showwarning("Uyarı", "En az bir ürün tipi seçmelisiniz!"))
                    return

                # Verileri al
                from botanik_db import BotanikDB
                db = BotanikDB()
                if not db.baglan():
                    self.parent.after(0, lambda: messagebox.showerror("Hata", "Veritabanına bağlanılamadı!"))
                    return

                # Ham verileri al - özel SQL sorgusu ile
                veriler = self._verileri_getir(db, yil, ay, secili_tipler, secili_urun_tipleri)
                db.kapat()

                # Verileri işle ve grupla
                self.veriler = veriler
                islenenmis = self._verileri_isle(veriler, giris_tipleri, cikis_tipleri, ay)

                # UI güncelle
                self.parent.after(0, lambda: self._sorgu_tamamlandi(islenenmis, ay))

            except Exception as e:
                logger.error(f"Sorgu hatası: {e}")
                self.parent.after(0, lambda: self._sorgu_hatasi(str(e)))

        thread = threading.Thread(target=sorgu_thread)
        thread.start()

    def _verileri_getir(self, db, yil_sayisi, ay_sayisi, hareket_tipleri, urun_tipleri=None):
        """Veritabanından ham verileri getir"""
        from datetime import datetime
        from dateutil.relativedelta import relativedelta

        bugun = datetime.now()
        baslangic_tarih = (bugun - relativedelta(years=yil_sayisi)).strftime('%Y-%m-%d')

        # Ürün tipi filtresi (çoklu seçim)
        urun_tipi_filtre = ""
        if urun_tipleri and len(urun_tipleri) > 0:
            # Tüm tipler seçili değilse filtrele
            tipler_temiz = [f"'{tip.replace(chr(39), chr(39)+chr(39))}'" for tip in urun_tipleri]
            urun_tipi_filtre = f"AND ut.UrunTipAdi IN ({', '.join(tipler_temiz)})"

        # Hareket sorguları
        hareket_bloklari = []

        if 'FATURA_GIRIS' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                fs.FSUrunId as UrunId,
                fs.FSUrunAdet as Adet,
                CAST(fg.FGFaturaTarihi as date) as Tarih,
                'FATURA_GIRIS' as HareketTipi,
                'GIRIS' as Yon
            FROM FaturaSatir fs
            JOIN FaturaGiris fg ON fs.FSFGId = fg.FGId
            WHERE fg.FGSilme = 0 AND fg.FGFaturaTarihi >= '{baslangic_tarih}'
            """)

        if 'FATURA_CIKIS' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                fcs.FSUrunId as UrunId,
                fcs.FSUrunAdet as Adet,
                CAST(fc.FGFaturaTarihi as date) as Tarih,
                'FATURA_CIKIS' as HareketTipi,
                'CIKIS' as Yon
            FROM FaturaCikisSatir fcs
            JOIN FaturaCikis fc ON fcs.FSFGId = fc.FGId
            WHERE fc.FGSilme = 0 AND fc.FGFaturaTarihi >= '{baslangic_tarih}'
            """)

        if 'TAKAS_GIRIS' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                ts.TSUrunId as UrunId,
                ts.TSUrunAdedi as Adet,
                CAST(t.TakasTarihi as date) as Tarih,
                'TAKAS_GIRIS' as HareketTipi,
                'GIRIS' as Yon
            FROM TakasSatir ts
            JOIN Takas t ON ts.TSTakasId = t.TakasId
            WHERE t.TakasSilme = 0 AND ts.TSSilme = 0 AND t.TakasYonu = 1
            AND t.TakasTarihi >= '{baslangic_tarih}'
            """)

        if 'TAKAS_CIKIS' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                ts.TSUrunId as UrunId,
                ts.TSUrunAdedi as Adet,
                CAST(t.TakasTarihi as date) as Tarih,
                'TAKAS_CIKIS' as HareketTipi,
                'CIKIS' as Yon
            FROM TakasSatir ts
            JOIN Takas t ON ts.TSTakasId = t.TakasId
            WHERE t.TakasSilme = 0 AND ts.TSSilme = 0 AND t.TakasYonu = 0
            AND t.TakasTarihi >= '{baslangic_tarih}'
            """)

        if 'IADE' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                it.ITUrunId as UrunId,
                it.ITUrunAdet as Adet,
                CAST(it.ITKayitTarihi as date) as Tarih,
                'IADE' as HareketTipi,
                'CIKIS' as Yon
            FROM IadeTakip it
            WHERE it.ITKayitTarihi >= '{baslangic_tarih}'
            """)

        if 'RECETE_SATIS' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                ri.RIUrunId as UrunId,
                ri.RIAdet as Adet,
                CAST(ra.RxKayitTarihi as date) as Tarih,
                'RECETE_SATIS' as HareketTipi,
                'CIKIS' as Yon
            FROM ReceteIlaclari ri
            JOIN ReceteAna ra ON ri.RIRxId = ra.RxId
            WHERE ra.RxSilme = 0 AND ri.RISilme = 0 AND (ri.RIIade = 0 OR ri.RIIade IS NULL)
            AND ra.RxKayitTarihi >= '{baslangic_tarih}'
            """)

        if 'ELDEN_SATIS' in hareket_tipleri:
            hareket_bloklari.append(f"""
            SELECT
                ei.RIUrunId as UrunId,
                ei.RIAdet as Adet,
                CAST(ea.RxKayitTarihi as date) as Tarih,
                'ELDEN_SATIS' as HareketTipi,
                'CIKIS' as Yon
            FROM EldenIlaclari ei
            JOIN EldenAna ea ON ei.RIRxId = ea.RxId
            WHERE ea.RxSilme = 0 AND ei.RISilme = 0 AND (ei.RIIade = 0 OR ei.RIIade IS NULL)
            AND ea.RxKayitTarihi >= '{baslangic_tarih}'
            """)

        if not hareket_bloklari:
            return []

        hareket_union = " UNION ALL ".join(hareket_bloklari)

        # Aylık toplamlar için CASE ifadeleri
        aylik_cases = []
        for i in range(ay_sayisi):
            ay_basi = (bugun - relativedelta(months=i)).replace(day=1)
            if i == 0:
                ay_sonu = bugun
            else:
                ay_sonu = (ay_basi + relativedelta(months=1)) - relativedelta(days=1)

            baslangic = ay_basi.strftime('%Y-%m-%d')
            bitis = ay_sonu.strftime('%Y-%m-%d')

            aylik_cases.append(f"""
                SUM(CASE WHEN Tarih >= '{baslangic}' AND Tarih <= '{bitis}' THEN Adet ELSE 0 END) as Ay{i+1}
            """)

        aylik_kolonlar = ",\n".join(aylik_cases)

        sql = f"""
        ;WITH TumHareketler AS (
            {hareket_union}
        ),
        UrunHareketOzet AS (
            SELECT
                UrunId,
                Yon,
                SUM(Adet) as ToplamAdet,
                COUNT(*) as ToplamIslem,
                {aylik_kolonlar}
            FROM TumHareketler
            GROUP BY UrunId, Yon
        )

        SELECT
            u.UrunId,
            u.UrunAdi,
            COALESCE(ut.UrunTipAdi, 'Belirsiz') as UrunTipi,
            u.UrunEsdegerId as EsdegerId,
            (u.UrunStokDepo + u.UrunStokRaf + u.UrunStokAcik) as Stok,
            ho.Yon,
            ho.ToplamAdet,
            ho.ToplamIslem,
            {', '.join([f'ho.Ay{i+1}' for i in range(ay_sayisi)])}
        FROM Urun u
        LEFT JOIN UrunTip ut ON u.UrunUrunTipId = ut.UrunTipId
        JOIN UrunHareketOzet ho ON u.UrunId = ho.UrunId
        WHERE u.UrunSilme = 0
        {urun_tipi_filtre}
        ORDER BY u.UrunEsdegerId, u.UrunAdi, ho.Yon DESC
        """

        return db.sorgu_calistir(sql)

    def _verileri_isle(self, veriler, giris_tipleri, cikis_tipleri, ay_sayisi):
        """Ham verileri işle, grupla ve formatla"""
        if not veriler:
            return []

        # Eşdeğer gruplarına göre düzenle
        esdeger_gruplari = {}
        for veri in veriler:
            esdeger_id = veri.get('EsdegerId') or 0  # None ise 0 kullan
            if esdeger_id not in esdeger_gruplari:
                esdeger_gruplari[esdeger_id] = {'GIRIS': [], 'CIKIS': []}

            yon = veri.get('Yon', 'CIKIS')
            esdeger_gruplari[esdeger_id][yon].append(veri)

        # İşlenmiş verileri oluştur
        islenenmis = []
        hem_giris_hem_cikis = bool(giris_tipleri) and bool(cikis_tipleri)

        for esdeger_id in sorted(esdeger_gruplari.keys()):
            grup = esdeger_gruplari[esdeger_id]

            # Gruptaki benzersiz ürün sayısını hesapla (giriş ve çıkış birleşik)
            urun_idler = set()
            for veri in grup['GIRIS']:
                urun_idler.add(veri.get('UrunId'))
            for veri in grup['CIKIS']:
                urun_idler.add(veri.get('UrunId'))

            urun_sayisi = len(urun_idler)

            # Eşdeğersiz (EsdegerId=0) veya tek ürünlü grup ise: basit görünüm
            if esdeger_id == 0 or urun_sayisi <= 1:
                # Her ürün için giriş+çıkış alt alta göster
                # Önce tüm ürünleri topla (hem giriş hem çıkış)
                urun_verileri = {}
                for veri in grup['GIRIS']:
                    urun_id = veri.get('UrunId')
                    if urun_id not in urun_verileri:
                        urun_verileri[urun_id] = {'GIRIS': None, 'CIKIS': None, 'UrunAdi': veri.get('UrunAdi', '')}
                    urun_verileri[urun_id]['GIRIS'] = veri
                for veri in grup['CIKIS']:
                    urun_id = veri.get('UrunId')
                    if urun_id not in urun_verileri:
                        urun_verileri[urun_id] = {'GIRIS': None, 'CIKIS': None, 'UrunAdi': veri.get('UrunAdi', '')}
                    urun_verileri[urun_id]['CIKIS'] = veri

                # Ürün adına göre sırala ve satırları ekle
                for urun_id in sorted(urun_verileri.keys(), key=lambda x: urun_verileri[x]['UrunAdi']):
                    urun = urun_verileri[urun_id]
                    # Giriş satırı
                    if giris_tipleri and urun['GIRIS']:
                        satir = self._satir_olustur(urun['GIRIS'], 'giris', ay_sayisi)
                        islenenmis.append(satir)
                    # Çıkış satırı
                    if cikis_tipleri and urun['CIKIS']:
                        satir = self._satir_olustur(urun['CIKIS'], 'cikis', ay_sayisi)
                        islenenmis.append(satir)

            else:
                # Birden fazla ürün olan eşdeğer grubu: grup başlığı + alt toplamlar
                esdeger_str = f"Eşdeğer #{esdeger_id}"
                baslik = {
                    'satir_tipi': 'grup_baslik',
                    'EsdegerId': esdeger_id,
                    'Yon': '',
                    'UrunTipi': '',
                    'UrunAdi': f'═══ {esdeger_str} ═══',
                    'Esdeger': '',
                    'Stok': '',
                    'TopCikis': '',
                    'CikisAdet': '',
                    'AylikOrt': '',
                    'AylikBitis': '',
                }
                # Ay kolonlarını ekle
                for i in range(ay_sayisi):
                    baslik[f'Ay{i+1}'] = ''
                islenenmis.append(baslik)

                # Giriş satırları (eğer giriş seçildiyse)
                if giris_tipleri and grup['GIRIS']:
                    giris_toplam_adet = 0
                    giris_toplam_islem = 0
                    giris_aylik_toplamlar = [0] * ay_sayisi

                    for veri in sorted(grup['GIRIS'], key=lambda x: x.get('UrunAdi', '')):
                        satir = self._satir_olustur(veri, 'giris', ay_sayisi)
                        islenenmis.append(satir)

                        giris_toplam_adet += veri.get('ToplamAdet', 0) or 0
                        giris_toplam_islem += veri.get('ToplamIslem', 0) or 0
                        for i in range(ay_sayisi):
                            giris_aylik_toplamlar[i] += veri.get(f'Ay{i+1}', 0) or 0

                    # Giriş alt toplam (birden fazla giriş varsa)
                    if len(grup['GIRIS']) > 1:
                        alt_toplam = self._alt_toplam_olustur(
                            'GİRİŞ TOPLAM', giris_toplam_adet, giris_toplam_islem,
                            giris_aylik_toplamlar, ay_sayisi, esdeger_id, 'giris'
                        )
                        islenenmis.append(alt_toplam)

                # Çıkış satırları (eğer çıkış seçildiyse)
                if cikis_tipleri and grup['CIKIS']:
                    cikis_toplam_adet = 0
                    cikis_toplam_islem = 0
                    cikis_aylik_toplamlar = [0] * ay_sayisi
                    cikis_stok_toplam = 0

                    for veri in sorted(grup['CIKIS'], key=lambda x: x.get('UrunAdi', '')):
                        satir = self._satir_olustur(veri, 'cikis', ay_sayisi)
                        islenenmis.append(satir)

                        cikis_toplam_adet += veri.get('ToplamAdet', 0) or 0
                        cikis_toplam_islem += veri.get('ToplamIslem', 0) or 0
                        cikis_stok_toplam += veri.get('Stok', 0) or 0
                        for i in range(ay_sayisi):
                            cikis_aylik_toplamlar[i] += veri.get(f'Ay{i+1}', 0) or 0

                    # Çıkış alt toplam (birden fazla çıkış varsa)
                    if len(grup['CIKIS']) > 1:
                        # Aylık ortalama ve bitiş hesapla
                        aylik_ort = cikis_toplam_adet / ay_sayisi if ay_sayisi > 0 else 0
                        aylik_bitis = cikis_stok_toplam / aylik_ort if aylik_ort > 0 else 0

                        alt_toplam = self._alt_toplam_olustur(
                            'ÇIKIŞ TOPLAM', cikis_toplam_adet, cikis_toplam_islem,
                            cikis_aylik_toplamlar, ay_sayisi, esdeger_id, 'cikis',
                            stok=cikis_stok_toplam, aylik_ort=aylik_ort, aylik_bitis=aylik_bitis
                        )
                        islenenmis.append(alt_toplam)

        return islenenmis

    def _satir_olustur(self, veri, satir_tipi, ay_sayisi):
        """Tek bir veri satırı oluştur"""
        toplam_adet = veri.get('ToplamAdet', 0) or 0
        toplam_islem = veri.get('ToplamIslem', 0) or 0
        stok = veri.get('Stok', 0) or 0

        # Aylık ortalama
        aylik_ort = toplam_adet / ay_sayisi if ay_sayisi > 0 else 0

        # Aylık bitiş (stok / aylık ortalama)
        aylik_bitis = stok / aylik_ort if aylik_ort > 0 else 0

        esdeger_id = veri.get('EsdegerId')
        esdeger_str = f"#{esdeger_id}" if esdeger_id else "-"

        satir = {
            'satir_tipi': satir_tipi,
            'EsdegerId': esdeger_id,
            'UrunId': veri.get('UrunId'),
            'Yon': 'GİRİŞ' if satir_tipi == 'giris' else 'ÇIKIŞ',
            'UrunTipi': veri.get('UrunTipi', ''),
            'UrunAdi': veri.get('UrunAdi', ''),
            'Esdeger': esdeger_str,
            'Stok': stok,
            'TopCikis': toplam_adet,
            'CikisAdet': toplam_islem,
            'AylikOrt': round(aylik_ort, 1),
            'AylikBitis': round(aylik_bitis, 1),
        }

        # Aylık kolonları ekle
        for i in range(ay_sayisi):
            satir[f'Ay{i+1}'] = veri.get(f'Ay{i+1}', 0) or 0

        return satir

    def _alt_toplam_olustur(self, etiket, toplam_adet, toplam_islem, aylik_toplamlar, ay_sayisi,
                            esdeger_id, yon_tipi, stok=None, aylik_ort=None, aylik_bitis=None):
        """Alt toplam satırı oluştur"""
        if aylik_ort is None:
            aylik_ort = toplam_adet / ay_sayisi if ay_sayisi > 0 else 0
        if aylik_bitis is None:
            aylik_bitis = 0

        satir = {
            'satir_tipi': 'alt_toplam',
            'EsdegerId': esdeger_id,
            'Yon': '',
            'UrunTipi': '',
            'UrunAdi': f'  └─ {etiket}',
            'Esdeger': '',
            'Stok': stok if stok is not None else '',
            'TopCikis': toplam_adet,
            'CikisAdet': toplam_islem,
            'AylikOrt': round(aylik_ort, 1),
            'AylikBitis': round(aylik_bitis, 1) if aylik_bitis else '',
        }

        for i in range(ay_sayisi):
            satir[f'Ay{i+1}'] = aylik_toplamlar[i]

        return satir

    def _sorgu_tamamlandi(self, veriler, ay_sayisi):
        """Sorgu tamamlandığında UI güncelle"""
        self.islenenmis_veriler = veriler

        # Sütunları güncelle
        self._sutunlari_guncelle(ay_sayisi)

        # Tabloyu güncelle
        self._tabloyu_guncelle(veriler)

        # Kayıt sayısı
        urun_sayisi = len([v for v in veriler if v.get('satir_tipi') in ['giris', 'cikis']])
        grup_sayisi = len([v for v in veriler if v.get('satir_tipi') == 'grup_baslik'])

        self.status_label.config(text="Sorgu tamamlandı")
        self.kayit_label.config(text=f"{urun_sayisi} ürün satırı, {grup_sayisi} eşdeğer grubu")

    def _sorgu_hatasi(self, hata_mesaji):
        """Sorgu hatası durumunda"""
        self.status_label.config(text=f"Hata: {hata_mesaji}")
        messagebox.showerror("Sorgu Hatası", hata_mesaji)

    def _tabloyu_guncelle(self, veriler):
        """Tabloyu verilerle güncelle"""
        # Mevcut verileri temizle
        for item in self.tree.get_children():
            self.tree.delete(item)

        if not veriler:
            return

        # Verileri ekle
        for veri in veriler:
            satir_tipi = veri.get('satir_tipi', 'cikis')

            # Değerleri al
            values = []
            for col_id, _ in self.aktif_sutunlar:
                val = veri.get(col_id, '')
                if isinstance(val, float):
                    val = round(val, 1)
                values.append(val if val != '' else '')

            # Tag belirle
            if satir_tipi == 'grup_baslik':
                tag = 'grup_baslik'
            elif satir_tipi == 'alt_toplam':
                tag = 'alt_toplam'
            elif satir_tipi == 'giris':
                tag = 'giris'
            else:
                tag = 'cikis'

            self.tree.insert('', 'end', values=values, tags=(tag,))

    def _siralama_yap(self, column):
        """Sütuna göre sırala"""
        # Grup yapısını koruyarak sıralama karmaşık olacağı için
        # şimdilik basit sıralama devre dışı
        pass

    def excel_aktar(self):
        """Verileri Excel'e aktar (XLSX formatında)"""
        if not self.islenenmis_veriler:
            messagebox.showwarning("Uyarı", "Aktarılacak veri yok!")
            return

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx"), ("CSV Dosyası", "*.csv"), ("Tüm Dosyalar", "*.*")],
            title="Excel Olarak Kaydet"
        )

        if not dosya_yolu:
            return

        try:
            # CSV mi XLSX mi?
            if dosya_yolu.lower().endswith('.csv'):
                self._csv_aktar(dosya_yolu)
            else:
                self._xlsx_aktar(dosya_yolu)

            messagebox.showinfo("Başarılı", f"Veriler başarıyla aktarıldı:\n{dosya_yolu}")

        except Exception as e:
            logger.error(f"Excel aktarım hatası: {e}")
            messagebox.showerror("Hata", f"Aktarım hatası: {e}")

    def _csv_aktar(self, dosya_yolu):
        """CSV formatında aktar"""
        with open(dosya_yolu, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')

            # Başlık satırı
            basliklar = [self.aktif_basliklar.get(col[0], col[0]) for col in self.aktif_sutunlar]
            writer.writerow(basliklar)

            # Veri satırları
            for veri in self.islenenmis_veriler:
                satir = []
                for col_id, _ in self.aktif_sutunlar:
                    val = veri.get(col_id, '')
                    if isinstance(val, float):
                        val = round(val, 1)
                    satir.append(val)
                writer.writerow(satir)

    def _xlsx_aktar(self, dosya_yolu):
        """XLSX formatında aktar (renkli ve formatlı)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Stok Hareket Analiz"

        # Stiller
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        baslik_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        giris_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        cikis_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        alt_toplam_fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
        alt_toplam_font = Font(bold=True, size=10)
        grup_baslik_fill = PatternFill(start_color="78909C", end_color="78909C", fill_type="solid")
        grup_baslik_font = Font(bold=True, size=11, color="FFFFFF")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Başlık satırı
        basliklar = [self.aktif_basliklar.get(col[0], col[0]) for col in self.aktif_sutunlar]
        for col_idx, baslik in enumerate(basliklar, 1):
            cell = ws.cell(row=1, column=col_idx, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = baslik_alignment
            cell.border = thin_border

        # Veri satırları
        for row_idx, veri in enumerate(self.islenenmis_veriler, 2):
            satir_tipi = veri.get('satir_tipi', '')

            for col_idx, (col_id, _) in enumerate(self.aktif_sutunlar, 1):
                val = veri.get(col_id, '')
                if isinstance(val, float):
                    val = round(val, 1)

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                # Satır tipine göre renklendirme
                if satir_tipi == 'giris':
                    cell.fill = giris_fill
                elif satir_tipi == 'cikis':
                    cell.fill = cikis_fill
                elif satir_tipi == 'alt_toplam':
                    cell.fill = alt_toplam_fill
                    cell.font = alt_toplam_font
                elif satir_tipi == 'grup_baslik':
                    cell.fill = grup_baslik_fill
                    cell.font = grup_baslik_font

                # Sayısal kolonlar sağa hizalı
                if col_idx > 4 and isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

        # Sütun genişlikleri
        sutun_genislikleri = {
            "Yon": 8,
            "UrunTipi": 15,
            "UrunAdi": 40,
            "Esdeger": 12,
            "Stok": 10,
            "TopCikis": 12,
            "CikisAdet": 12,
            "AylikOrt": 12,
            "AylikBitis": 12,
        }

        for col_idx, (col_id, _) in enumerate(self.aktif_sutunlar, 1):
            genislik = sutun_genislikleri.get(col_id, 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = genislik

        # Başlık satırını sabitle
        ws.freeze_panes = 'A2'

        # Kaydet
        wb.save(dosya_yolu)


# Test
if __name__ == "__main__":
    root = tk.Tk()
    app = StokHareketAnalizGUI(root)
    root.mainloop()
