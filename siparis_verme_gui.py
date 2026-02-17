"""
Sipariş Verme Modülü GUI - v2
Stok hareket analizi bazlı sipariş hazırlama sistemi
Eşdeğer gruplu görünüm, manuel giriş, MF/Zam analizi
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import logging
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from decimal import Decimal
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tkcalendar import DateEntry

# Tema yönetimi
try:
    from tema_yonetimi import get_tema, renkler, renk, koyu_mu
    TEMA_YUKLENDI = True
except ImportError:
    TEMA_YUKLENDI = False

logger = logging.getLogger(__name__)

# Loading indicator
try:
    from loading_indicator import LoadingIndicator
    LOADING_INDICATOR_YUKLENDI = True
except ImportError:
    LOADING_INDICATOR_YUKLENDI = False

# Matplotlib (grafik icin)
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.ticker import FuncFormatter
    import numpy as np
    MATPLOTLIB_YUKLENDI = True
except ImportError:
    MATPLOTLIB_YUKLENDI = False

# Siparis Database (kalici kayit)
try:
    from siparis_db import get_siparis_db
    SIPARIS_DB_YUKLENDI = True
except ImportError:
    SIPARIS_DB_YUKLENDI = False
    logger.warning("siparis_db modulu yuklenemedi, siparisler kaydedilmeyecek")


class SiparisVermeGUI:
    """Sipariş Verme Modülü Penceresi - v2"""

    VARSAYILAN_URUN_TIPLERI = ['İLAÇ', 'PASİF İLAÇ', 'SERUMLAR']

    def __init__(self, parent):
        self.parent = parent
        self.parent.title("Sipariş Verme Modülü")
        self.parent.geometry("1700x950")

        # Tema renklerini yükle
        self._tema_renklerini_yukle()

        self.db = None
        self.tum_veriler = []           # Tum islenmis veriler
        self.gorunen_veriler = []       # Filtrelenmis gorunen veriler
        self.kesin_siparis_listesi = [] # Kesinlesmis siparisler
        self.secili_urun = None         # Detay paneli icin secili urun

        # Siparis calismasi (oturum) yonetimi
        self.siparis_db = None
        self.aktif_calisma = None       # Aktif calisma bilgisi {id, ad, ...}

        # Parametre değişkenleri
        self.sene_sayisi = tk.IntVar(value=1)
        self.ay_sayisi = tk.IntVar(value=6)
        self.beklenen_zam_orani = tk.DoubleVar(value=0.0)

        # Checkbox değişkenleri
        self.hedef_tarih_aktif = tk.BooleanVar(value=False)
        self.min_stok_aktif = tk.BooleanVar(value=True)
        self.min_stok_kaynagi = tk.StringVar(value='botanik')  # 'botanik' veya 'yerel'
        self.zam_aktif = tk.BooleanVar(value=False)
        self.yeterlileri_gizle = tk.BooleanVar(value=False)

        # Zam stratejisi seçimi
        # pareto: %80 kazanç noktası (dengeli)
        # optimum: Maksimum mutlak kazanç (agresif)
        # verimlilik: Maksimum ROI (sermaye verimli)
        self.zam_stratejisi = tk.StringVar(value="pareto")

        # Faiz parametreleri
        self.mevduat_faizi = tk.DoubleVar(value=40.0)  # Yıllık %
        self.kredi_faizi = tk.DoubleVar(value=50.0)    # Yıllık %
        self.faiz_turu = tk.StringVar(value="mevduat") # "mevduat" veya "kredi"
        self.depo_vadesi = tk.IntVar(value=75)         # Gün

        # Ürün tipi seçimi
        self.urun_tipi_vars = {}
        self.urun_tipleri_listesi = []

        # Dinamik sütunlar
        self.aktif_sutunlar = []
        self.aktif_basliklar = {}

        # Manuel giriş değerleri {urun_id: miktar}
        self.manuel_miktarlar = {}
        self.manuel_mf_girisler = {}  # {urun_id: "5+1"}

        self._arayuz_olustur()
        self._baglanti_kur()
        self._siparis_db_baslat()

    def _siparis_db_baslat(self):
        """Siparis database'ini baslat ve aktif calismayi yukle"""
        if not SIPARIS_DB_YUKLENDI:
            return

        try:
            self.siparis_db = get_siparis_db()

            # Aktif calisma var mi?
            aktif = self.siparis_db.aktif_calisma_getir()

            if aktif:
                # Mevcut calismayi yukle
                self.aktif_calisma = aktif
                self._calisma_siparislerini_yukle()
                self._calisma_bilgisini_guncelle()
            else:
                # Yeni calisma olustur
                self._yeni_calisma_olustur()

        except Exception as e:
            logger.error(f"Siparis DB baslama hatasi: {e}")

    def _yeni_calisma_olustur(self, ad=None):
        """Yeni siparis calismasi olustur"""
        if not self.siparis_db:
            return

        try:
            # Parametreleri kaydet
            parametreler = {
                'ay_sayisi': self.ay_sayisi.get(),
                'mevduat_faizi': self.mevduat_faizi.get(),
                'depo_vadesi': self.depo_vadesi.get(),
            }

            calisma_id = self.siparis_db.yeni_calisma_olustur(ad=ad, parametreler=parametreler)
            if calisma_id:
                self.aktif_calisma = self.siparis_db.aktif_calisma_getir()
                self.kesin_siparis_listesi = []
                self._kesin_liste_guncelle()
                self._calisma_bilgisini_guncelle()
        except Exception as e:
            logger.error(f"Yeni calisma olusturma hatasi: {e}")

    def _calisma_siparislerini_yukle(self):
        """Aktif calismanin siparislerini yukle"""
        if not self.siparis_db or not self.aktif_calisma:
            return

        try:
            siparisler = self.siparis_db.calisma_siparisleri_getir(self.aktif_calisma['id'])

            self.kesin_siparis_listesi = []
            for s in siparisler:
                siparis = {
                    'db_id': s['id'],
                    'UrunId': s.get('urun_id'),
                    'UrunAdi': s.get('urun_adi', ''),
                    'Barkod': s.get('barkod', ''),
                    'Miktar': s.get('miktar', 0),
                    'MF': s.get('mf', ''),
                    'Toplam': s.get('toplam', 0),
                    'Stok': s.get('stok', 0),
                    'AylikOrt': s.get('aylik_ort', 0),
                }
                # Depo sonuclari
                if s.get('DepoSonuclari'):
                    for depo, deger in s['DepoSonuclari'].items():
                        siparis[depo] = deger
                self.kesin_siparis_listesi.append(siparis)

            self._kesin_liste_guncelle()
        except Exception as e:
            logger.error(f"Siparisler yukleme hatasi: {e}")

    def _calisma_bilgisini_guncelle(self):
        """Calisma bilgisi etiketini guncelle"""
        if hasattr(self, 'calisma_label') and self.aktif_calisma:
            siparis_sayisi = len(self.kesin_siparis_listesi)
            self.calisma_label.config(
                text=f"Calisma: {self.aktif_calisma['ad']} | {siparis_sayisi} urun"
            )

    def _tema_renklerini_yukle(self):
        """Tema renklerini yükle - koyu/açık tema desteği"""
        if TEMA_YUKLENDI:
            tema = get_tema()
            tema.ttk_stili_uygula()
            r = tema.renkler
            koyu = tema.aktif_tema == "koyu"
        else:
            koyu = True  # Varsayılan koyu tema
            r = {}

        # Tema bazlı renkler
        if koyu:
            # Koyu tema
            self.R_BG = r.get("bg", "#1E1E1E")
            self.R_BG_SECONDARY = r.get("bg_secondary", "#2D2D2D")
            self.R_BG_TERTIARY = r.get("bg_tertiary", "#3D3D3D")
            self.R_FG = r.get("fg", "#FFFFFF")
            self.R_FG_SECONDARY = r.get("fg_secondary", "#B0B0B0")
            self.R_ACCENT = r.get("accent", "#1976D2")
            self.R_ACCENT_HOVER = r.get("accent_hover", "#1565C0")
            self.R_SUCCESS = r.get("success", "#4CAF50")
            self.R_SUCCESS_BG = r.get("success_bg", "#1B3D1B")
            self.R_WARNING = r.get("warning", "#FF9800")
            self.R_WARNING_BG = r.get("warning_bg", "#3D2E00")
            self.R_ERROR = r.get("error", "#F44336")
            self.R_ERROR_BG = r.get("error_bg", "#3D1B1B")
            self.R_INFO = r.get("info", "#2196F3")
            self.R_INFO_BG = r.get("info_bg", "#1B2D3D")
            self.R_BORDER = r.get("border", "#404040")
            self.R_INPUT_BG = r.get("input_bg", "#3D3D3D")
            self.R_INPUT_FG = r.get("input_fg", "#FFFFFF")
            self.R_HEADER_BG = r.get("header_bg", "#0D2137")
            self.R_TABLE_HEADER_BG = r.get("table_header_bg", "#1565C0")
            self.R_TABLE_ROW_BG = r.get("table_row_bg", "#2D2D2D")
            self.R_TABLE_ROW_ALT = r.get("table_row_alt_bg", "#353535")
            self.R_TABLE_SELECTED = r.get("table_selected_bg", "#1976D2")
            self.R_FRAME_BG = r.get("frame_bg", "#252525")
            self.R_FRAME_FG = r.get("frame_fg", "#90CAF9")
            # Özel renkler
            self.R_GRUP_BASLIK = "#3949AB"  # İndigo koyu
            self.R_GRUP_ICERIK = "#283593"  # Daha koyu indigo
            self.R_ALT_TOPLAM = "#5C6BC0"   # Orta indigo
            self.R_SIPARIS_GEREK = "#5D3A3A"  # Koyu kırmızı
            self.R_YETERLI = "#2E5A2E"       # Koyu yeşil
            self.R_TEK_SATIR = "#2D2D2D"     # Koyu gri
            self.R_PARAM_BG = "#2D2D30"      # Parametre panel arka plan
            self.R_GRP_DATA = "#1E3A5F"      # Veri grubu arka plan (koyu mavi)
            self.R_GRP_DATA_FG = "#90CAF9"   # Veri grubu yazı
            self.R_GRP_HEDEF = "#1E4D2B"     # Hedef tarih grubu (koyu yeşil)
            self.R_GRP_HEDEF_FG = "#81C784"  # Hedef tarih yazı
            self.R_GRP_ZAM = "#4D3319"       # Zam grubu (koyu turuncu)
            self.R_GRP_ZAM_FG = "#FFB74D"    # Zam yazı
            self.R_GRP_MIN = "#3D2152"       # Min stok grubu (koyu mor)
            self.R_GRP_MIN_FG = "#CE93D8"    # Min stok yazı
            self.R_GRP_FAIZ = "#194D4D"      # Faiz grubu (koyu cyan)
            self.R_GRP_FAIZ_FG = "#80DEEA"   # Faiz yazı
            self.R_MF_BG = "#1B3D1B"         # MF analiz arka plan
            self.R_MF_FG = "#81C784"         # MF analiz yazı
        else:
            # Açık tema
            self.R_BG = r.get("bg", "#F5F5F5")
            self.R_BG_SECONDARY = r.get("bg_secondary", "#FFFFFF")
            self.R_BG_TERTIARY = r.get("bg_tertiary", "#E8E8E8")
            self.R_FG = r.get("fg", "#212121")
            self.R_FG_SECONDARY = r.get("fg_secondary", "#757575")
            self.R_ACCENT = r.get("accent", "#1976D2")
            self.R_ACCENT_HOVER = r.get("accent_hover", "#1565C0")
            self.R_SUCCESS = r.get("success", "#388E3C")
            self.R_SUCCESS_BG = r.get("success_bg", "#E8F5E9")
            self.R_WARNING = r.get("warning", "#F57C00")
            self.R_WARNING_BG = r.get("warning_bg", "#FFF3E0")
            self.R_ERROR = r.get("error", "#D32F2F")
            self.R_ERROR_BG = r.get("error_bg", "#FFEBEE")
            self.R_INFO = r.get("info", "#1976D2")
            self.R_INFO_BG = r.get("info_bg", "#E3F2FD")
            self.R_BORDER = r.get("border", "#E0E0E0")
            self.R_INPUT_BG = r.get("input_bg", "#FFFFFF")
            self.R_INPUT_FG = r.get("input_fg", "#212121")
            self.R_HEADER_BG = r.get("header_bg", "#1976D2")
            self.R_TABLE_HEADER_BG = r.get("table_header_bg", "#1976D2")
            self.R_TABLE_ROW_BG = r.get("table_row_bg", "#FFFFFF")
            self.R_TABLE_ROW_ALT = r.get("table_row_alt_bg", "#F5F5F5")
            self.R_TABLE_SELECTED = r.get("table_selected_bg", "#BBDEFB")
            self.R_FRAME_BG = r.get("frame_bg", "#FAFAFA")
            self.R_FRAME_FG = r.get("frame_fg", "#1565C0")
            # Özel renkler (açık tema)
            self.R_GRUP_BASLIK = "#5C6BC0"  # İndigo
            self.R_GRUP_ICERIK = "#E8EAF6"  # Açık indigo
            self.R_ALT_TOPLAM = "#9FA8DA"   # Orta indigo
            self.R_SIPARIS_GEREK = "#FFCDD2"  # Açık kırmızı
            self.R_YETERLI = "#C8E6C9"       # Açık yeşil
            self.R_TEK_SATIR = "#FFFFFF"     # Beyaz
            self.R_PARAM_BG = "#ECEFF1"      # Parametre panel arka plan
            self.R_GRP_DATA = "#E3F2FD"      # Veri grubu arka plan (açık mavi)
            self.R_GRP_DATA_FG = "#1565C0"   # Veri grubu yazı
            self.R_GRP_HEDEF = "#E8F5E9"     # Hedef tarih grubu (açık yeşil)
            self.R_GRP_HEDEF_FG = "#2E7D32"  # Hedef tarih yazı
            self.R_GRP_ZAM = "#FFF3E0"       # Zam grubu (açık turuncu)
            self.R_GRP_ZAM_FG = "#E65100"    # Zam yazı
            self.R_GRP_MIN = "#F3E5F5"       # Min stok grubu (açık mor)
            self.R_GRP_MIN_FG = "#7B1FA2"    # Min stok yazı
            self.R_GRP_FAIZ = "#E0F7FA"      # Faiz grubu (açık cyan)
            self.R_GRP_FAIZ_FG = "#00838F"   # Faiz yazı
            self.R_MF_BG = "#E8F5E9"         # MF analiz arka plan
            self.R_MF_FG = "#2E7D32"         # MF analiz yazı

        # Ana pencere arkaplan rengi
        self.parent.configure(bg=self.R_BG)

    def _arayuz_olustur(self):
        """Ana arayüzü oluştur"""
        main_frame = ttk.Frame(self.parent, padding=3)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Üst panel - Parametreler (2 satır)
        self._parametre_panel_olustur(main_frame)

        # Orta bölüm - Ana DataGrid + Detay Paneli
        self._orta_bolum_olustur(main_frame)

        # Alt bölüm - Kesin Sipariş Listesi
        self._kesin_liste_olustur(main_frame)

        # Status bar
        self._status_bar_olustur(main_frame)

    def _parametre_panel_olustur(self, parent):
        """Parametre panelini oluştur - 2 satır düzenli tasarım"""
        param_frame = tk.Frame(parent, bg=self.R_PARAM_BG, relief='raised', bd=1)
        param_frame.pack(fill=tk.X, pady=(0, 5))

        # ═══════════════════════════════════════════════════════════
        # SATIR 1: Parametre Grupları (4 Ana Grup)
        # ═══════════════════════════════════════════════════════════
        row1 = tk.Frame(param_frame, bg=self.R_PARAM_BG)
        row1.pack(fill=tk.X, padx=8, pady=(5, 3))

        # ─────────────────────────────────────────────────────────
        # GRUP 1: VERİ KAYNAĞI (Analiz edilecek veriler)
        # ─────────────────────────────────────────────────────────
        grp_veri = tk.LabelFrame(row1, text=" 📊 Veri Kaynağı ", font=('Arial', 9, 'bold'),
                                  bg=self.R_GRP_DATA, fg=self.R_GRP_DATA_FG, padx=6, pady=4)
        grp_veri.pack(side=tk.LEFT, padx=(0, 12), ipady=2)

        # Hareket Süresi - Stok hareket analiz süresi
        tk.Label(grp_veri, text="Stok Hareket Süresi:", font=('Arial', 9), bg=self.R_GRP_DATA, fg=self.R_FG).pack(side=tk.LEFT)
        sene_combo = ttk.Combobox(grp_veri, textvariable=self.sene_sayisi, width=3, state="readonly", font=('Arial', 10))
        sene_combo['values'] = [1, 2, 3]
        sene_combo.pack(side=tk.LEFT, padx=(3, 0))
        tk.Label(grp_veri, text="yıl", font=('Arial', 9), bg=self.R_GRP_DATA, fg=self.R_FG).pack(side=tk.LEFT, padx=(3, 12))

        # Aylık Ortalama - Aylık gidiş hesaplama süresi
        tk.Label(grp_veri, text="Aylık Gidiş Ortalaması:", font=('Arial', 9), bg=self.R_GRP_DATA, fg=self.R_FG).pack(side=tk.LEFT)
        ay_combo = ttk.Combobox(grp_veri, textvariable=self.ay_sayisi, width=3, state="readonly", font=('Arial', 10))
        ay_combo['values'] = [3, 6, 9, 12]
        ay_combo.pack(side=tk.LEFT, padx=(3, 0))
        tk.Label(grp_veri, text="ay", font=('Arial', 9), bg=self.R_GRP_DATA, fg=self.R_FG).pack(side=tk.LEFT, padx=(3, 12))

        # Ürün Tipi
        tk.Label(grp_veri, text="Tip:", font=('Arial', 9), bg=self.R_GRP_DATA, fg=self.R_FG).pack(side=tk.LEFT)
        self.urun_tipi_menubutton = tk.Menubutton(
            grp_veri, text="Seçiniz", relief=tk.RAISED, width=8, font=('Arial', 9),
            bg=self.R_INPUT_BG, fg=self.R_INPUT_FG
        )
        self.urun_tipi_menu = tk.Menu(self.urun_tipi_menubutton, tearoff=0, font=('Arial', 9),
                                       bg=self.R_INPUT_BG, fg=self.R_INPUT_FG)
        self.urun_tipi_menubutton["menu"] = self.urun_tipi_menu
        self.urun_tipi_menubutton.pack(side=tk.LEFT, padx=(2, 3))

        tk.Button(grp_veri, text="Tümü", command=self._tum_tipleri_sec, font=('Arial', 8),
                  bg=self.R_ACCENT, fg='white', width=4, relief='flat').pack(side=tk.LEFT, padx=1)
        tk.Button(grp_veri, text="Vars", command=self._varsayilan_tipleri_sec, font=('Arial', 8),
                  bg=self.R_ACCENT, fg='white', width=4, relief='flat').pack(side=tk.LEFT)

        # ─────────────────────────────────────────────────────────
        # GRUP 2: SİPARİŞ HESAPLAMA (Hedef ve Min Stok)
        # ─────────────────────────────────────────────────────────
        grp_siparis = tk.LabelFrame(row1, text=" 🎯 Sipariş Hesaplama ", font=('Arial', 9, 'bold'),
                                     bg=self.R_GRP_HEDEF, fg=self.R_GRP_HEDEF_FG, padx=6, pady=4)
        grp_siparis.pack(side=tk.LEFT, padx=(0, 12), ipady=2)

        # Hedef Tarih
        self.hedef_check = tk.Checkbutton(
            grp_siparis, text="Hedef:", variable=self.hedef_tarih_aktif,
            bg=self.R_GRP_HEDEF, fg=self.R_FG, font=('Arial', 9), activebackground=self.R_GRP_HEDEF,
            selectcolor=self.R_BG_SECONDARY, command=self._hedef_tarih_toggle
        )
        self.hedef_check.pack(side=tk.LEFT)

        bugun = datetime.now()
        ay_son_gun = calendar.monthrange(bugun.year, bugun.month)[1]
        self.hedef_tarih_entry = DateEntry(
            grp_siparis, width=10, background=self.R_ACCENT, foreground='white',
            borderwidth=1, date_pattern='yyyy-mm-dd', font=('Arial', 9),
            year=bugun.year, month=bugun.month, day=ay_son_gun
        )
        self.hedef_tarih_entry.pack(side=tk.LEFT, padx=(0, 3))
        tk.Button(grp_siparis, text="Ay Sonu", command=self._ay_sonu_sec, font=('Arial', 8),
                  bg=self.R_SUCCESS, fg='white', relief='flat').pack(side=tk.LEFT, padx=(0, 10))

        # Min Stok
        self.min_check = tk.Checkbutton(
            grp_siparis, text="Min Stok:", variable=self.min_stok_aktif,
            bg=self.R_GRP_HEDEF, fg=self.R_FG, font=('Arial', 9), activebackground=self.R_GRP_HEDEF,
            selectcolor=self.R_BG_SECONDARY
        )
        self.min_check.pack(side=tk.LEFT)

        tk.Radiobutton(
            grp_siparis, text="Botanik", variable=self.min_stok_kaynagi, value="botanik",
            bg=self.R_GRP_HEDEF, fg=self.R_FG, font=('Arial', 8), activebackground=self.R_GRP_HEDEF,
            selectcolor=self.R_BG_SECONDARY
        ).pack(side=tk.LEFT)
        tk.Radiobutton(
            grp_siparis, text="Yerel", variable=self.min_stok_kaynagi, value="yerel",
            bg=self.R_GRP_HEDEF, fg=self.R_FG, font=('Arial', 8), activebackground=self.R_GRP_HEDEF,
            selectcolor=self.R_BG_SECONDARY
        ).pack(side=tk.LEFT)

        tk.Button(
            grp_siparis, text="Hesapla", command=self._min_stok_analiz_ac,
            bg=self.R_GRP_MIN_FG, fg='white', font=('Arial', 8),
            relief='flat'
        ).pack(side=tk.LEFT, padx=(5, 0))

        # ─────────────────────────────────────────────────────────
        # GRUP 3: ZAM ANALİZİ (Zam optimizasyonu)
        # ─────────────────────────────────────────────────────────
        grp_zam = tk.LabelFrame(row1, text=" 📈 Zam Analizi ", font=('Arial', 9, 'bold'),
                                 bg=self.R_GRP_ZAM, fg=self.R_GRP_ZAM_FG, padx=6, pady=4)
        grp_zam.pack(side=tk.LEFT, padx=(0, 12), ipady=2)

        self.zam_check = tk.Checkbutton(
            grp_zam, text="Aktif", variable=self.zam_aktif,
            bg=self.R_GRP_ZAM, fg=self.R_FG, font=('Arial', 9), activebackground=self.R_GRP_ZAM,
            selectcolor=self.R_BG_SECONDARY, command=self._zam_toggle
        )
        self.zam_check.pack(side=tk.LEFT, padx=(0, 5))

        tk.Label(grp_zam, text="Oran:", font=('Arial', 9), bg=self.R_GRP_ZAM, fg=self.R_FG).pack(side=tk.LEFT)
        tk.Label(grp_zam, text="%", font=('Arial', 9), bg=self.R_GRP_ZAM, fg=self.R_FG).pack(side=tk.LEFT)
        zam_entry = ttk.Entry(grp_zam, textvariable=self.beklenen_zam_orani, width=4, font=('Arial', 9))
        zam_entry.pack(side=tk.LEFT, padx=(1, 8))

        tk.Label(grp_zam, text="Tarih:", font=('Arial', 9), bg=self.R_GRP_ZAM, fg=self.R_FG).pack(side=tk.LEFT)
        self.zam_tarih_entry = DateEntry(
            grp_zam, width=10, background=self.R_WARNING, foreground='white',
            borderwidth=1, date_pattern='yyyy-mm-dd', font=('Arial', 9)
        )
        self.zam_tarih_entry.pack(side=tk.LEFT, padx=(2, 8))

        tk.Label(grp_zam, text="Strateji:", font=('Arial', 9), bg=self.R_GRP_ZAM, fg=self.R_FG).pack(side=tk.LEFT)
        self.strateji_combo = ttk.Combobox(
            grp_zam, textvariable=self.zam_stratejisi, width=8, font=('Arial', 9),
            state='readonly', values=['pareto', 'optimum', 'verimlilik']
        )
        self.strateji_combo.pack(side=tk.LEFT, padx=(2, 0))
        self.strateji_combo.bind('<<ComboboxSelected>>', self._strateji_bilgi_goster)

        # ─────────────────────────────────────────────────────────
        # GRUP 4: FİNANSAL PARAMETRELER (Faiz hesaplamaları)
        # ─────────────────────────────────────────────────────────
        grp_finans = tk.LabelFrame(row1, text=" 💰 Finansal ", font=('Arial', 9, 'bold'),
                                    bg=self.R_GRP_FAIZ, fg=self.R_GRP_FAIZ_FG, padx=6, pady=4)
        grp_finans.pack(side=tk.LEFT, ipady=2)

        tk.Label(grp_finans, text="M%", font=('Arial', 9), bg=self.R_GRP_FAIZ, fg=self.R_FG).pack(side=tk.LEFT)
        mevduat_entry = ttk.Entry(grp_finans, textvariable=self.mevduat_faizi, width=3, font=('Arial', 9))
        mevduat_entry.pack(side=tk.LEFT, padx=(1, 4))

        tk.Label(grp_finans, text="K%", font=('Arial', 9), bg=self.R_GRP_FAIZ, fg=self.R_FG).pack(side=tk.LEFT)
        kredi_entry = ttk.Entry(grp_finans, textvariable=self.kredi_faizi, width=3, font=('Arial', 9))
        kredi_entry.pack(side=tk.LEFT, padx=(1, 4))

        tk.Radiobutton(
            grp_finans, text="M", variable=self.faiz_turu, value="mevduat",
            bg=self.R_GRP_FAIZ, fg=self.R_FG, font=('Arial', 9), activebackground=self.R_GRP_FAIZ,
            selectcolor=self.R_BG_SECONDARY
        ).pack(side=tk.LEFT)
        tk.Radiobutton(
            grp_finans, text="K", variable=self.faiz_turu, value="kredi",
            bg=self.R_GRP_FAIZ, fg=self.R_FG, font=('Arial', 9), activebackground=self.R_GRP_FAIZ,
            selectcolor=self.R_BG_SECONDARY
        ).pack(side=tk.LEFT, padx=(0, 4))

        tk.Label(grp_finans, text="Vade", font=('Arial', 9), bg=self.R_GRP_FAIZ, fg=self.R_FG).pack(side=tk.LEFT)
        vade_entry = ttk.Entry(grp_finans, textvariable=self.depo_vadesi, width=3, font=('Arial', 9))
        vade_entry.pack(side=tk.LEFT, padx=(1, 0))
        tk.Label(grp_finans, text="g", font=('Arial', 9), bg=self.R_GRP_FAIZ, fg=self.R_FG).pack(side=tk.LEFT, padx=(1, 0))

        # ═══════════════════════════════════════════════════════════
        # SATIR 2: İşlem Butonları
        # ═══════════════════════════════════════════════════════════
        row2 = tk.Frame(param_frame, bg=self.R_PARAM_BG)
        row2.pack(fill=tk.X, padx=8, pady=(3, 5))

        # Sol taraf butonları
        self.gizle_btn = tk.Button(
            row2, text="Yeterlileri Gizle: KAPALI", command=self._toggle_yeterlileri_gizle,
            bg=self.R_FG_SECONDARY, fg='white', font=('Arial', 9, 'bold'),
            relief='raised', bd=2, padx=10, pady=3
        )
        self.gizle_btn.pack(side=tk.LEFT, padx=(0, 12))

        tk.Button(
            row2, text="Önerileri Kopyala", command=self._tum_onerileri_kopyala,
            bg=self.R_GRP_MIN_FG, fg='white', font=('Arial', 9), relief='raised', bd=2, padx=10, pady=3
        ).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(
            row2, text="Seçilileri Ekle", command=self._secilileri_kesin_listeye_ekle,
            bg=self.R_SUCCESS, fg='white', font=('Arial', 9), relief='raised', bd=2, padx=10, pady=3
        ).pack(side=tk.LEFT, padx=(0, 8))

        self.excel_btn = tk.Button(
            row2, text="Excel'e Aktar", command=self.excel_aktar,
            bg=self.R_WARNING, fg='white', font=('Arial', 9, 'bold'),
            relief='raised', bd=2, padx=10, pady=3
        )
        self.excel_btn.pack(side=tk.LEFT)

        # Ortada bilgi etiketi
        self.hesaplama_label = tk.Label(
            row2, text="", font=('Arial', 9, 'bold'), bg=self.R_PARAM_BG, fg=self.R_ACCENT
        )
        self.hesaplama_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=15)

        # Sağda ana buton
        self.getir_btn = tk.Button(
            row2, text="  VERİLERİ GETİR  ", command=self.verileri_getir,
            bg=self.R_ACCENT, fg='white', font=('Arial', 11, 'bold'),
            relief='raised', bd=2, padx=20, pady=4
        )
        self.getir_btn.pack(side=tk.RIGHT)

    def _orta_bolum_olustur(self, parent):
        """Orta bölüm - Ana DataGrid + Detay Paneli"""
        # PanedWindow ile bölünebilir
        self.orta_paned = tk.PanedWindow(parent, orient=tk.HORIZONTAL, sashwidth=4, bg=self.R_BORDER)
        self.orta_paned.pack(fill=tk.BOTH, expand=True, pady=(0, 3))

        # Sol: Ana DataGrid
        self._ana_grid_olustur()

        # Sağ: Detay Paneli
        self._detay_panel_olustur()

    def _ana_grid_olustur(self):
        """Ana DataGrid - gruplu görünüm (genişletilmiş)"""
        grid_frame = tk.Frame(self.orta_paned, bg=self.R_BG_SECONDARY, relief='sunken', bd=1)
        self.orta_paned.add(grid_frame, minsize=1000, stretch='always')

        # Başlık
        header = tk.Frame(grid_frame, bg=self.R_TABLE_HEADER_BG, height=26)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="📊 STOK VE SİPARİŞ ANALİZİ", bg=self.R_TABLE_HEADER_BG, fg='white',
                font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=10)

        # Arama kutusu (Ctrl+F)
        self.arama_frame = tk.Frame(header, bg=self.R_TABLE_HEADER_BG)
        self.arama_frame.pack(side=tk.RIGHT, padx=5)

        tk.Label(self.arama_frame, text="🔍", bg=self.R_TABLE_HEADER_BG, fg='white', font=('Arial', 10)).pack(side=tk.LEFT)
        self.arama_var = tk.StringVar()
        self.arama_var.trace('w', self._arama_yap)
        self.arama_entry = ttk.Entry(self.arama_frame, textvariable=self.arama_var, width=20, font=('Arial', 9))
        self.arama_entry.pack(side=tk.LEFT, padx=3)

        self.arama_sonuc_label = tk.Label(self.arama_frame, text="", bg=self.R_TABLE_HEADER_BG, fg='yellow', font=('Arial', 8))
        self.arama_sonuc_label.pack(side=tk.LEFT, padx=3)

        tk.Button(self.arama_frame, text="▲", font=('Arial', 8), width=2,
                  command=self._onceki_eslesme).pack(side=tk.LEFT, padx=1)
        tk.Button(self.arama_frame, text="▼", font=('Arial', 8), width=2,
                  command=self._sonraki_eslesme).pack(side=tk.LEFT, padx=1)
        tk.Button(self.arama_frame, text="✕", font=('Arial', 8), width=2,
                  command=self._arama_temizle).pack(side=tk.LEFT, padx=1)

        # Ctrl+F kısayolu
        self.parent.bind('<Control-f>', self._arama_odaklan)
        self.parent.bind('<Control-F>', self._arama_odaklan)
        self.arama_entry.bind('<Return>', lambda e: self._sonraki_eslesme())
        self.arama_entry.bind('<Escape>', lambda e: self._arama_temizle())

        # Arama sonuçları için değişkenler
        self.arama_eslesmeler = []
        self.arama_index = 0

        # Treeview için frame
        tree_container = tk.Frame(grid_frame, bg=self.R_BG_SECONDARY)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # Sütun tanımları
        # NOT: Stok ve Aylık sütunları vurgulu gösterilecek (daha geniş)
        self.ana_sutunlar = [
            ("Tur", "", 25),
            ("UrunAdi", "Urun Adi", 220),
            ("Stok", "★STOK★", 65),  # Vurgulu - genişletildi
            ("Min", "Min", 35),
            ("Sart1", "Sart1", 50),
            ("Sart2", "Sart2", 50),
            ("Sart3", "Sart3", 50),
        ]

        # Aylık sütunlar dinamik eklenecek
        self.aylik_sutunlar = []

        self.son_sutunlar = [
            ("AylikOrt", "★AYLIK★", 70),  # Vurgulu - genişletildi
            ("GunlukOrt", "Gun", 45),
            ("AyBitis", "AyBitis", 55),
            ("Oneri", "SİPARİŞ", 75),  # Vurgulu ve genis
            ("OneriAy", "OneriAy", 60),
            ("YeniAyBitis", "YeniBitis", 65),
            ("Manuel", "Manuel", 55),
            ("MF", "MF", 50),
        ]

        # Treeview
        self.ana_tree = ttk.Treeview(tree_container, show='headings', selectmode='extended')

        # Scrollbarlar
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.ana_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.ana_tree.xview)
        self.ana_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Grid yerleşimi
        tree_container.columnconfigure(0, weight=1)
        tree_container.rowconfigure(0, weight=1)
        self.ana_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        # Treeview stil ayarları - daha büyük font (11pt)
        style = ttk.Style()
        style.configure('Siparis.Treeview', font=('Arial', 11), rowheight=30)
        style.configure('Siparis.Treeview.Heading', font=('Arial', 11, 'bold'))
        self.ana_tree.configure(style='Siparis.Treeview')

        # Tag'ler - sipariş gerektiren satırlar bold ve vurgulu
        self.ana_tree.tag_configure('grup_baslik', background=self.R_GRUP_BASLIK, foreground='white',
                                     font=('Arial', 11, 'bold'))
        # Sübvansiyon tag'leri - stok yeterliliğine göre renkler
        self.ana_tree.tag_configure('grup_baslik_yeterli', background='#2E7D32', foreground='white',
                                     font=('Arial', 11, 'bold'))  # Koyu yeşil - stok yeterli
        self.ana_tree.tag_configure('grup_baslik_iyi', background='#558B2F', foreground='white',
                                     font=('Arial', 11, 'bold'))  # Yeşil - %80+
        self.ana_tree.tag_configure('grup_baslik_orta', background='#F9A825', foreground='black',
                                     font=('Arial', 11, 'bold'))  # Sarı - %50-80
        self.ana_tree.tag_configure('grup_baslik_dusuk', background='#D84315', foreground='white',
                                     font=('Arial', 11, 'bold'))  # Kırmızı - %50 altı
        self.ana_tree.tag_configure('grup_satir', background=self.R_GRUP_ICERIK, foreground=self.R_FG,
                                     font=('Arial', 11))
        self.ana_tree.tag_configure('grup_satir_siparis', background=self.R_SIPARIS_GEREK, foreground=self.R_FG,
                                     font=('Arial', 12, 'bold'))
        self.ana_tree.tag_configure('alt_toplam', background=self.R_ALT_TOPLAM, foreground=self.R_FG,
                                     font=('Arial', 11, 'bold'))
        self.ana_tree.tag_configure('tek_satir', background=self.R_TEK_SATIR, foreground=self.R_FG,
                                     font=('Arial', 11))
        self.ana_tree.tag_configure('tek_satir_siparis', background=self.R_SIPARIS_GEREK, foreground=self.R_FG,
                                     font=('Arial', 12, 'bold'))
        self.ana_tree.tag_configure('arama_eslesme', background=self.R_WARNING, foreground=self.R_FG,
                                     font=('Arial', 12, 'bold'))

        # Seçim olayı
        self.ana_tree.bind('<<TreeviewSelect>>', self._satir_secildi)
        self.ana_tree.bind('<Double-1>', self._satir_cift_tiklandi)
        self.ana_tree.bind('<Button-1>', self._sutun_tiklandi)

        # Tooltip (ilaç ismi gösterimi)
        self._tooltip = None
        self._tooltip_after_id = None
        self.ana_tree.bind('<Motion>', self._tree_motion)
        self.ana_tree.bind('<Leave>', self._tooltip_gizle)

    def _tree_motion(self, event):
        """Mouse hareket edince UrunAdi sütunundaysa tooltip göster"""
        # Önceki zamanlayıcıyı iptal et
        if self._tooltip_after_id:
            self.ana_tree.after_cancel(self._tooltip_after_id)
            self._tooltip_after_id = None

        # Tooltip'i gizle (yeni pozisyonda tekrar açılacak)
        self._tooltip_gizle()

        region = self.ana_tree.identify_region(event.x, event.y)
        if region != 'cell':
            return

        col = self.ana_tree.identify_column(event.x)
        col_id = self.ana_tree.column(col, 'id') if col else None
        if col_id != 'UrunAdi':
            return

        item = self.ana_tree.identify_row(event.y)
        if not item:
            return

        values = self.ana_tree.item(item, 'values')
        if not values or len(values) < 2:
            return

        urun_adi = str(values[1])
        if not urun_adi or len(urun_adi) < 15:
            return

        # 300ms gecikme ile tooltip göster
        self._tooltip_after_id = self.ana_tree.after(
            300, lambda: self._tooltip_goster(event.x_root, event.y_root, urun_adi))

    def _tooltip_goster(self, x, y, text):
        """Küçük tooltip etiketi göster"""
        self._tooltip_gizle()
        tw = tk.Toplevel(self.ana_tree)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x + 12}+{y + 12}")
        tw.attributes('-topmost', True)
        label = tk.Label(tw, text=text, bg='#FFFFDD', fg='#333',
                         font=('Arial', 10), relief='solid', bd=1, padx=4, pady=2)
        label.pack()
        self._tooltip = tw

    def _tooltip_gizle(self, event=None):
        """Tooltip'i kapat"""
        if self._tooltip:
            self._tooltip.destroy()
            self._tooltip = None

    def _detay_panel_olustur(self):
        """Sağ taraftaki detay paneli - scroll'lu tasarım"""
        detay_frame = tk.Frame(self.orta_paned, bg=self.R_BG_SECONDARY, relief='sunken', bd=1)
        self.orta_paned.add(detay_frame, minsize=300, width=375, stretch='never')

        # Başlık
        header = tk.Frame(detay_frame, bg=self.R_GRUP_BASLIK, height=26)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="📝 İLAÇ DETAY", bg=self.R_GRUP_BASLIK, fg='white',
                font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=10)

        # Scrollable içerik alanı
        self.detay_canvas = tk.Canvas(detay_frame, bg=self.R_BG_SECONDARY, highlightthickness=0)
        self.detay_scrollbar = ttk.Scrollbar(detay_frame, orient='vertical', command=self.detay_canvas.yview)
        self.detay_content = tk.Frame(self.detay_canvas, bg=self.R_BG_SECONDARY)

        self.detay_canvas.configure(yscrollcommand=self.detay_scrollbar.set)
        self.detay_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.detay_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.detay_canvas_window = self.detay_canvas.create_window((0, 0), window=self.detay_content, anchor='nw')

        # Canvas ve içerik boyut güncellemeleri
        self.detay_content.bind('<Configure>', self._detay_scroll_configure)
        self.detay_canvas.bind('<Configure>', lambda e: self.detay_canvas.itemconfig(
            self.detay_canvas_window, width=e.width))

        # Mouse wheel scroll
        self.detay_canvas.bind('<Enter>', lambda e: self.detay_canvas.bind_all('<MouseWheel>', self._detay_mouse_scroll))
        self.detay_canvas.bind('<Leave>', lambda e: self.detay_canvas.unbind_all('<MouseWheel>'))

        # Placeholder
        self.detay_placeholder = tk.Label(
            self.detay_content,
            text="Detay görmek için\nbir ilaç satırı seçin",
            bg=self.R_BG_SECONDARY, fg=self.R_FG_SECONDARY, font=('Arial', 11, 'italic'),
            justify='center'
        )
        self.detay_placeholder.pack(expand=True)

    def _detay_scroll_configure(self, event=None):
        """Detay paneli scroll region güncelle"""
        self.detay_canvas.configure(scrollregion=self.detay_canvas.bbox('all'))

    def _detay_mouse_scroll(self, event):
        """Detay paneli mouse wheel scroll"""
        self.detay_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _kesin_liste_olustur(self, parent):
        """Alt kısımdaki kesin sipariş listesi - depo bilgileri ile"""
        kesin_frame = tk.Frame(parent, bg=self.R_BG_SECONDARY, relief='sunken', bd=1, height=170)
        kesin_frame.pack(fill=tk.X, pady=(0, 3))
        kesin_frame.pack_propagate(False)

        # Başlık
        header = tk.Frame(kesin_frame, bg=self.R_SUCCESS, height=30)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(header, text="KESIN SIPARIS LISTESI", bg=self.R_SUCCESS, fg='white',
                font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=10)

        # Calisma bilgisi etiketi
        self.calisma_label = tk.Label(header, text="Calisma: -", bg=self.R_SUCCESS, fg='yellow',
                                       font=('Arial', 9, 'bold'))
        self.calisma_label.pack(side=tk.LEFT, padx=15)

        # Yeni Calisma butonu
        tk.Button(header, text="Yeni Calisma", command=self._yeni_calisma_dialog,
                 bg=self.R_ACCENT, fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.LEFT, padx=5, pady=2)

        # Calisma Yukle butonu
        tk.Button(header, text="Calisma Yukle", command=self._calisma_yukle_dialog,
                 bg=self.R_GRP_MIN_FG, fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.LEFT, padx=2, pady=2)

        # Kaydet ve Kapat butonu
        tk.Button(header, text="Kaydet/Kapat", command=self._calisma_kaydet_kapat,
                 bg='#6A1B9A', fg='white', font=('Arial', 8, 'bold'), relief='flat', padx=5
                 ).pack(side=tk.LEFT, padx=2, pady=2)

        # Arşiv butonu
        tk.Button(header, text="Arşiv", command=self._calisma_arsiv_goster,
                 bg='#5D4037', fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.LEFT, padx=2, pady=2)

        # Depolarda Ara butonu
        self.depo_ara_btn = tk.Button(
            header, text="🔍 DEPOLARDA ARA", command=self._depolarda_ara,
            bg=self.R_ACCENT, fg='white', font=('Arial', 9, 'bold'), relief='raised', padx=10
        )
        self.depo_ara_btn.pack(side=tk.LEFT, padx=20, pady=2)

        # Durum etiketi
        self.depo_durum_label = tk.Label(header, text="", bg=self.R_SUCCESS, fg='yellow',
                                          font=('Arial', 9))
        self.depo_durum_label.pack(side=tk.LEFT, padx=10)

        tk.Button(header, text="Listeyi Temizle", command=self._kesin_listeyi_temizle,
                 bg=self.R_ERROR, fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.RIGHT, padx=5, pady=2)

        tk.Button(header, text="Seciliyi Sil", command=self._kesin_seciliyi_sil,
                 bg=self.R_WARNING, fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.RIGHT, padx=2, pady=2)

        tk.Button(header, text="Miktari Duzenle", command=self._kesin_miktari_duzenle,
                 bg=self.R_GRP_MIN_FG, fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.RIGHT, padx=2, pady=2)

        tk.Button(header, text="Excel'e Aktar", command=self._kesin_listeyi_excel_aktar,
                 bg=self.R_WARNING, fg='white', font=('Arial', 8), relief='flat', padx=5
                 ).pack(side=tk.RIGHT, padx=2, pady=2)

        # Treeview
        tree_frame = tk.Frame(kesin_frame, bg=self.R_BG_SECONDARY)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # Depo sütunları ile genişletilmiş sütunlar
        columns = [
            ("UrunAdi", "Ürün Adı", 250),
            ("Barkod", "Barkod", 100),
            ("Miktar", "Miktar", 55),
            ("MF", "MF", 50),
            ("Toplam", "Toplam", 55),
            ("Selcuk", "Selçuk", 80),
            ("Alliance", "Alliance", 80),
            ("Sancak", "Sancak", 80),
            ("Iskoop", "İskoop", 80),
            ("Farmazon", "Farmazon", 80),
        ]

        self.kesin_tree = ttk.Treeview(tree_frame, columns=[c[0] for c in columns],
                                       show='headings', height=5)
        for col_id, baslik, width in columns:
            self.kesin_tree.heading(col_id, text=baslik)
            self.kesin_tree.column(col_id, width=width, minwidth=40)

        # Depo sütunları için renkli tag'ler
        self.kesin_tree.tag_configure('stok_var', background=self.R_YETERLI, foreground=self.R_FG)
        self.kesin_tree.tag_configure('stok_yok', background=self.R_SIPARIS_GEREK, foreground=self.R_FG)
        self.kesin_tree.tag_configure('mf_var', background=self.R_INFO_BG, foreground=self.R_FG)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.kesin_tree.yview)
        self.kesin_tree.configure(yscrollcommand=vsb.set)

        self.kesin_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # F ve F1 tuşu ile barkod kopyalama
        self.kesin_tree.bind('<F1>', self._barkod_kopyala)
        self.kesin_tree.bind('<f>', self._barkod_kopyala)
        self.kesin_tree.bind('<F>', self._barkod_kopyala)

        # F2 tuşu ile barkod yapıştırma (uygulama genelinde)
        self.parent.bind_all('<F2>', self._barkod_yapistir)

        # Klavye kısayolu etiketi
        kisayol_label = tk.Label(tree_frame, text="F/F1: Kopyala | F2: Yapıştır", font=('Arial', 7),
                                 fg='#666', bg=self.R_BG)
        kisayol_label.place(relx=1.0, y=0, anchor='ne')

    def _barkod_kopyala(self, event=None):
        """Seçili satırın barkodunu panoya kopyala (F1)"""
        secili = self.kesin_tree.selection()
        if not secili:
            self.status_label.config(text="⚠ Barkod kopyalamak için bir satır seçin")
            return

        item = secili[0]
        idx = self.kesin_tree.index(item)

        if idx >= len(self.kesin_siparis_listesi):
            return

        siparis = self.kesin_siparis_listesi[idx]
        barkod = siparis.get('Barkod', '')

        if barkod:
            # Panoya kopyala
            self.parent.clipboard_clear()
            self.parent.clipboard_append(barkod)

            # Görsel geri bildirim
            urun_adi = siparis.get('UrunAdi', '')[:25]
            self.status_label.config(text=f"📋 Barkod kopyalandı: {barkod} ({urun_adi})")
        else:
            self.status_label.config(text="⚠ Bu ürünün barkodu yok!")

    def _barkod_yapistir(self, event=None):
        """F2: Panodaki barkodu aktif metin alanına yapıştır"""
        try:
            barkod = self.parent.clipboard_get()
        except:
            self.status_label.config(text="⚠ Panoda barkod yok!")
            return

        if not barkod:
            self.status_label.config(text="⚠ Panoda barkod yok!")
            return

        # Aktif widget'a yapıştır (Entry veya Text alanı)
        focused = self.parent.focus_get()
        if focused and isinstance(focused, (tk.Entry, ttk.Entry)):
            focused.delete(0, tk.END)
            focused.insert(0, barkod)
            self.status_label.config(text=f"📋 Barkod yapıştırıldı: {barkod}")
        elif focused and isinstance(focused, tk.Text):
            focused.insert(tk.INSERT, barkod)
            self.status_label.config(text=f"📋 Barkod yapıştırıldı: {barkod}")
        else:
            self.status_label.config(text=f"📋 Barkod panoda: {barkod} (bir metin alanına tıklayıp F2)")

    def _status_bar_olustur(self, parent):
        """Status bar"""
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill=tk.X)

        self.status_label = ttk.Label(status_frame, text="Hazır", font=('Arial', 9))
        self.status_label.pack(side=tk.LEFT)

        self.kayit_label = ttk.Label(status_frame, text="", font=('Arial', 9))
        self.kayit_label.pack(side=tk.RIGHT)

    # ═══════════════════════════════════════════════════════════════════════
    # VERİTABANI İŞLEMLERİ
    # ═══════════════════════════════════════════════════════════════════════

    def _baglanti_kur(self):
        """Veritabanı bağlantısı kur"""
        try:
            from botanik_db import BotanikDB
            self.db = BotanikDB()
            if self.db.baglan():
                self.status_label.config(text="Veritabanı bağlantısı başarılı")
                self._urun_tiplerini_yukle()
            else:
                self.status_label.config(text="Veritabanı bağlantı hatası!")
        except Exception as e:
            logger.error(f"Veritabanı bağlantı hatası: {e}")
            self.status_label.config(text=f"Hata: {e}")

    def _urun_tiplerini_yukle(self):
        """Ürün tiplerini yükle"""
        try:
            if self.db:
                tipler = self.db.urun_tipleri_getir()
                self.urun_tipleri_listesi = [t['UrunTipAdi'] for t in tipler]

                self.urun_tipi_menu.delete(0, tk.END)
                self.urun_tipi_vars.clear()

                for tip_adi in self.urun_tipleri_listesi:
                    varsayilan = tip_adi in self.VARSAYILAN_URUN_TIPLERI
                    var = tk.BooleanVar(value=varsayilan)
                    self.urun_tipi_vars[tip_adi] = var
                    self.urun_tipi_menu.add_checkbutton(
                        label=tip_adi, variable=var,
                        command=self._urun_tipi_secim_guncelle
                    )
                self._urun_tipi_secim_guncelle()
        except Exception as e:
            logger.error(f"Ürün tipleri yükleme hatası: {e}")

    def _urun_tipi_secim_guncelle(self):
        """Seçili ürün tiplerini göster"""
        secili = [t for t, v in self.urun_tipi_vars.items() if v.get()]
        toplam = len(self.urun_tipi_vars)

        if len(secili) == 0:
            self.urun_tipi_menubutton.config(text="Hiçbiri")
        elif len(secili) == toplam:
            self.urun_tipi_menubutton.config(text="Tümü")
        elif len(secili) <= 2:
            self.urun_tipi_menubutton.config(text=", ".join(secili[:2]))
        else:
            self.urun_tipi_menubutton.config(text=f"{len(secili)} tip")

    def _tum_tipleri_sec(self):
        for v in self.urun_tipi_vars.values():
            v.set(True)
        self._urun_tipi_secim_guncelle()

    def _varsayilan_tipleri_sec(self):
        for t, v in self.urun_tipi_vars.items():
            v.set(t in self.VARSAYILAN_URUN_TIPLERI)
        self._urun_tipi_secim_guncelle()

    def _ay_sonu_sec(self):
        bugun = datetime.now()
        ay_son_gun = calendar.monthrange(bugun.year, bugun.month)[1]
        self.hedef_tarih_entry.set_date(date(bugun.year, bugun.month, ay_son_gun))

    def _hedef_tarih_toggle(self):
        """Hedef tarih checkbox değiştiğinde - zam'ı kapat"""
        if self.hedef_tarih_aktif.get():
            self.zam_aktif.set(False)

    def _zam_toggle(self):
        """Zam checkbox değiştiğinde - hedef tarihi kapat"""
        if self.zam_aktif.get():
            self.hedef_tarih_aktif.set(False)

    def _strateji_bilgi_goster(self, event=None):
        """Seçilen strateji hakkında bilgi göster"""
        strateji = self.zam_stratejisi.get()
        bilgiler = {
            'pareto': "PARETO (%80 Kazanç)\n\nYatırımın %80 kazancına ulaşılan nokta.\nDengeli yaklaşım - sermaye verimli kullanılır.",
            'optimum': "OPTİMUM (Maks. Kazanç)\n\nMutlak en yüksek kazanç noktası.\nAgresif yaklaşım - daha fazla sermaye bağlar.",
            'verimlilik': "VERİMLİLİK (Maks. ROI)\n\nBirim yatırım başına en yüksek getiri.\nKonservatif yaklaşım - küçük ama verimli alım."
        }
        # Status bar'da bilgi göster
        self.status_label.config(text=bilgiler.get(strateji, ""))

    def _aktif_faiz_getir(self):
        """Seçili faiz türüne göre aktif faiz oranını döndür"""
        if self.faiz_turu.get() == "mevduat":
            return self.mevduat_faizi.get()
        else:
            return self.kredi_faizi.get()

    def _depocu_fiyat_hesapla(self, psf, iskonto_kamu) -> float:
        """
        Depocu fiyatı (bize geliş fiyatı) hesapla.

        Formül: PSF × 0.71 × 1.10 × (1 - IskontoKamu/100)

        Args:
            psf: Perakende Satış Fiyatı (UrunFiyatEtiket)
            iskonto_kamu: Kamu iskonto yüzdesi (UrunIskontoKamu)

        Returns:
            Depocu fiyatı (KDV dahil, iskontolu)
        """
        # Decimal tipini float'a çevir
        psf = float(psf) if psf else 0
        iskonto_kamu = float(iskonto_kamu) if iskonto_kamu else 0

        if psf <= 0:
            return 0

        depocu_kdv_haric = psf * 0.71
        depocu_kdv_dahil = depocu_kdv_haric * 1.10
        depocu_iskontolu = depocu_kdv_dahil * (1 - iskonto_kamu / 100)

        return round(depocu_iskontolu, 2)

    # ═══════════════════════════════════════════════════════════════════════
    # NPV HESAPLAMA FONKSİYONLARI
    # ═══════════════════════════════════════════════════════════════════════

    def _npv_hesapla(self, alinan: int, mf: int, maliyet: float, aylik_ort: float,
                    faiz_yillik: float, depo_vade: int,
                    fatura_tarihi: date, zam_tarihi: date, zam_orani: float,
                    mevcut_stok: int = 0, kalan_gun: int = 15) -> tuple:
        """
        NPV hesaplama - MF'li vs MF'siz alım karşılaştırması

        Returns:
            (npv_mfsiz, npv_mfli, net_kazanc)
        """
        if alinan <= 0 or maliyet <= 0 or aylik_ort <= 0:
            return (0, 0, 0)

        aylik_faiz = (faiz_yillik / 100) / 12
        depo_vade_ay = depo_vade / 30

        toplam_gelen = alinan + (mf or 0)

        # İlk ay sarfı (kalan güne göre)
        gunluk_sarf = aylik_ort / 30
        ilk_ay_sarf = kalan_gun * gunluk_sarf

        # Zam ayı hesapla
        zam_ay_sonra = 999
        if zam_tarihi and fatura_tarihi and zam_orani > 0:
            if isinstance(fatura_tarihi, date):
                zam_ay_sonra = (zam_tarihi.year - fatura_tarihi.year) * 12 + \
                               (zam_tarihi.month - fatura_tarihi.month)

        # ===== MEVCUT STOK KAYBI HESAPLA =====
        npv_mevcut_ayri = 0
        kalan_mevcut = mevcut_stok
        ay = 0

        while kalan_mevcut > 0 and ay < 120:
            if ay == 0:
                bu_ay = min(ilk_ay_sarf, kalan_mevcut)
            else:
                bu_ay = min(aylik_ort, kalan_mevcut)

            if bu_ay > 0:
                if zam_orani > 0 and ay >= zam_ay_sonra:
                    fiyat = maliyet * (1 + zam_orani / 100)
                else:
                    fiyat = maliyet
                npv_mevcut_ayri += (bu_ay * fiyat) / ((1 + aylik_faiz) ** (ay + 1 + depo_vade_ay))
                kalan_mevcut -= bu_ay

            ay += 1

        npv_mevcut_toplu = mevcut_stok * maliyet
        mevcut_stok_kaybi = npv_mevcut_toplu - npv_mevcut_ayri

        # ===== YENİ ALIM NPV HESAPLAMASI =====
        # Senaryo A: MF'siz - her ay ihtiyaç kadar al
        npv_mfsiz = 0
        kalan_ihtiyac = toplam_gelen
        kalan_mevcut_sim = mevcut_stok
        ay = 0

        while kalan_ihtiyac > 0 and ay < 120:
            if ay == 0:
                bu_ay_sarf = ilk_ay_sarf
            else:
                bu_ay_sarf = aylik_ort

            # Önce mevcut stoktan harca
            mevcut_kullanim = min(kalan_mevcut_sim, bu_ay_sarf)
            kalan_mevcut_sim -= mevcut_kullanim

            # Kalan sarfı yeni alımdan karşıla
            yeni_kullanim = min(bu_ay_sarf - mevcut_kullanim, kalan_ihtiyac)

            if yeni_kullanim > 0:
                if zam_orani > 0 and ay >= zam_ay_sonra:
                    fiyat = maliyet * (1 + zam_orani / 100)
                else:
                    fiyat = maliyet

                odeme = yeni_kullanim * fiyat
                iskonto_faktor = (1 + aylik_faiz) ** (ay + 1 + depo_vade_ay)
                npv_mfsiz += odeme / iskonto_faktor

                kalan_ihtiyac -= yeni_kullanim

            ay += 1

        # Senaryo B: Toplu ödeme (MF'li) - sadece alınan kadar ödenir
        odenen_para = alinan * maliyet
        npv_mfli = odenen_para / ((1 + aylik_faiz) ** (1 + depo_vade_ay))

        # Net kazanç
        yeni_alim_kazanc = npv_mfsiz - npv_mfli
        net_kazanc = yeni_alim_kazanc - mevcut_stok_kaybi

        return (round(npv_mfsiz, 2), round(npv_mfli, 2), round(net_kazanc, 2))

    def _zam_oncesi_optimum_hesapla(self, maliyet: float, aylik_ort: float,
                                     mevcut_stok: int, zam_tarihi: date, zam_orani: float,
                                     faiz_yillik: float, depo_vade: int) -> dict:
        """
        Zam öncesi optimum alım miktarını hesapla.

        Mantık: Belli bir süre boyunca harcanacak miktarlar için
        aylık mı yoksa toplu mu almak daha avantajlı?
        Her iki senaryonun NPV'sini (bugünkü değer) hesapla ve karşılaştır.

        Karar Noktaları:
        - verimlilik: Maksimum ROI (Kazanç/Yatırım) noktası
        - pareto: %80 kazanca ulaşılan nokta (Pareto prensibi)
        - optimum: Maksimum mutlak kazanç noktası (Pik)
        - maksimum: Karlılığın sıfıra düştüğü nokta (Sınır)

        Returns:
            {
                'verimlilik': int,        # Maksimum ROI noktası
                'verimlilik_roi': float,  # ROI yüzdesi
                'verimlilik_kazanc': float,
                'pareto': int,            # %80 kazanç noktası
                'pareto_kazanc': float,
                'optimum': int,           # Maksimum kazanç noktası
                'kazanc_optimum': float,
                'maksimum': int,          # Sınır (karlılık sıfır)
                'kazanc_maksimum': float
            }
        """
        bos_sonuc = {
            'verimlilik': 0, 'verimlilik_roi': 0, 'verimlilik_kazanc': 0,
            'pareto': 0, 'pareto_kazanc': 0,
            'optimum': 0, 'kazanc_optimum': 0,
            'maksimum': 0, 'kazanc_maksimum': 0
        }

        if not zam_tarihi or zam_orani <= 0 or aylik_ort <= 0:
            return bos_sonuc

        bugun = date.today()
        zam_gun = (zam_tarihi - bugun).days
        if zam_gun <= 0:
            return bos_sonuc

        # Gün bazlı faiz hesabı (smooth hesaplama için)
        aylik_faiz = (faiz_yillik / 100) / 12
        gunluk_faiz = (1 + aylik_faiz) ** (1/30) - 1
        gunluk_sarf = aylik_ort / 30

        # 12 aya kadar test et
        max_test = int(aylik_ort * 12)
        if max_test < 10:
            max_test = 100

        # Tüm miktarlar için kazanç hesapla
        kazanclar = []
        for test_miktar in range(0, max_test + 1):
            if test_miktar == 0:
                kazanclar.append(0)
                continue

            kalan_mevcut = mevcut_stok
            kalan_yeni = test_miktar
            npv_aylik = 0

            gun = 0
            while kalan_yeni > 0 and gun < 720:
                harcanan = gunluk_sarf
                mevcut_harcanan = min(kalan_mevcut, harcanan)
                kalan_mevcut -= mevcut_harcanan
                yeni_harcanan = min(harcanan - mevcut_harcanan, kalan_yeni)

                if yeni_harcanan > 0:
                    if gun < zam_gun:
                        fiyat = maliyet
                    else:
                        fiyat = maliyet * (1 + zam_orani / 100)

                    # Gün bazlı iskonto
                    odeme_gun = gun + 30 + depo_vade
                    odeme = yeni_harcanan * fiyat
                    iskonto = (1 + gunluk_faiz) ** odeme_gun
                    npv_aylik += odeme / iskonto
                    kalan_yeni -= yeni_harcanan
                gun += 1

            # Toplu alım NPV
            npv_toplu = (test_miktar * maliyet) / ((1 + gunluk_faiz) ** (30 + depo_vade))
            kazanc = npv_aylik - npv_toplu
            kazanclar.append(kazanc)

        # ===== KARAR NOKTALARI =====

        # 1. OPTİMUM (Pik) - Maksimum kazanç
        optimum_miktar = 0
        en_iyi_kazanc = 0
        for i, k in enumerate(kazanclar):
            if k > en_iyi_kazanc:
                en_iyi_kazanc = k
                optimum_miktar = i

        # 2. MAKSİMUM (Sınır) - Son pozitif kazanç noktası
        maksimum_miktar = 0
        for i, k in enumerate(kazanclar):
            if k > 0:
                maksimum_miktar = i

        # 3. PARETO (%80) - %80 kazanca ulaşılan ilk nokta
        hedef_80 = en_iyi_kazanc * 0.80
        pareto_miktar = 0
        pareto_kazanc = 0
        for i, k in enumerate(kazanclar):
            if k >= hedef_80:
                pareto_miktar = i
                pareto_kazanc = k
                break

        # 4. VERİMLİLİK (Max ROI) - Kazanç/Yatırım maksimum
        verimlilik_miktar = 0
        verimlilik_roi = 0
        verimlilik_kazanc = 0
        for i, k in enumerate(kazanclar):
            if i > 0 and k > 0:
                yatirim = i * maliyet
                roi = (k / yatirim) * 100
                if roi > verimlilik_roi:
                    verimlilik_roi = roi
                    verimlilik_miktar = i
                    verimlilik_kazanc = k

        return {
            'verimlilik': verimlilik_miktar,
            'verimlilik_roi': round(verimlilik_roi, 1),
            'verimlilik_kazanc': round(verimlilik_kazanc, 2),
            'pareto': pareto_miktar,
            'pareto_kazanc': round(pareto_kazanc, 2),
            'optimum': optimum_miktar,
            'kazanc_optimum': round(en_iyi_kazanc, 2),
            'maksimum': maksimum_miktar,
            'kazanc_maksimum': round(kazanclar[maksimum_miktar] if maksimum_miktar > 0 else 0, 2)
        }

    # ═══════════════════════════════════════════════════════════════════════
    # KARLILIK GRAFİĞİ FONKSİYONLARI
    # ═══════════════════════════════════════════════════════════════════════

    def _karlilik_grafigi_goster(self):
        """Seçili ilaç(lar) için karlılık grafiği göster"""
        if not MATPLOTLIB_YUKLENDI:
            messagebox.showerror("Hata", "Matplotlib yüklü değil!\npip install matplotlib numpy")
            return

        # Seçili satırları al
        secili = self.tablo.selection()
        if not secili:
            messagebox.showwarning("Uyarı", "Lütfen tabloda bir ilaç seçin.")
            return

        # Parametreleri al
        try:
            zam_orani = float(self.beklenen_zam_orani.get())
            faiz_yillik = self._aktif_faiz_getir()
            depo_vade = int(self.depo_vadesi.get())
            zam_tarihi = self.zam_tarih_entry.get_date()
        except:
            messagebox.showerror("Hata", "Zam/Faiz parametrelerini kontrol edin.")
            return

        if zam_orani <= 0:
            messagebox.showwarning("Uyarı", "Zam oranı 0'dan büyük olmalı.")
            return

        # Seçili ilaçları topla
        ilaclar = []
        for item_id in secili[:4]:  # Max 4 ilaç
            values = self.tablo.item(item_id, 'values')
            if values:
                ilaclar.append({
                    'UrunAdi': values[0],
                    'Stok': int(values[1]) if values[1] else 0,
                    'AylikOrt': float(values[2]) if values[2] else 0,
                    'DepocuFiyat': float(values[5]) if values[5] else 0
                })

        if not ilaclar:
            messagebox.showwarning("Uyarı", "Seçili ilaç verisi alınamadı.")
            return

        # Grafik çiz
        if len(ilaclar) == 1:
            dosya = self._tek_ilac_grafik_ciz(ilaclar[0], zam_orani, zam_tarihi, faiz_yillik, depo_vade)
        else:
            dosya = self._coklu_ilac_grafik_ciz(ilaclar, zam_orani, zam_tarihi, faiz_yillik, depo_vade)

        if dosya:
            import os
            os.startfile(dosya)

    def _tek_ilac_grafik_goster(self, urun: dict):
        """Detay panelinden tek ilaç için grafik göster"""
        if not MATPLOTLIB_YUKLENDI:
            messagebox.showerror("Hata", "Matplotlib yüklü değil!\npip install matplotlib numpy")
            return

        # Parametreleri al
        try:
            zam_orani = float(self.beklenen_zam_orani.get())
            faiz_yillik = self._aktif_faiz_getir()
            depo_vade = int(self.depo_vadesi.get())
            zam_tarihi = self.zam_tarih_entry.get_date()
        except Exception as e:
            messagebox.showerror("Hata", f"Parametre hatası: {e}")
            return

        if zam_orani <= 0:
            messagebox.showwarning("Uyarı", "Zam oranı 0'dan büyük olmalı.")
            return

        ilac = {
            'UrunAdi': urun.get('UrunAdi', ''),
            'Stok': urun.get('Stok', 0),
            'AylikOrt': urun.get('AylikOrt', 0),
            'DepocuFiyat': urun.get('DepocuFiyat', 0)
        }

        if ilac['DepocuFiyat'] <= 0 or ilac['AylikOrt'] <= 0:
            messagebox.showwarning("Uyarı", "Depocu fiyat veya aylık ortalama eksik.")
            return

        dosya = self._tek_ilac_grafik_ciz(ilac, zam_orani, zam_tarihi, faiz_yillik, depo_vade)

        if dosya:
            import os
            os.startfile(dosya)

    def _roi_grafik_goster(self, urun: dict):
        """ROI bazlı grafik göster - Her 100 TL'de kazanç"""
        if not MATPLOTLIB_YUKLENDI:
            messagebox.showerror("Hata", "Matplotlib yüklü değil!")
            return

        try:
            zam_orani = float(self.beklenen_zam_orani.get())
            faiz_yillik = self._aktif_faiz_getir()
            depo_vade = int(self.depo_vadesi.get())
            zam_tarihi = self.zam_tarih_entry.get_date()
        except Exception as e:
            messagebox.showerror("Hata", f"Parametre hatası: {e}")
            return

        if zam_orani <= 0:
            messagebox.showwarning("Uyarı", "Zam oranı 0'dan büyük olmalı.")
            return

        ilac = {
            'UrunAdi': urun.get('UrunAdi', ''),
            'Stok': urun.get('Stok', 0),
            'AylikOrt': urun.get('AylikOrt', 0),
            'DepocuFiyat': urun.get('DepocuFiyat', 0)
        }

        if ilac['DepocuFiyat'] <= 0 or ilac['AylikOrt'] <= 0:
            messagebox.showwarning("Uyarı", "Depocu fiyat veya aylık ortalama eksik.")
            return

        dosya = self._roi_grafik_ciz(ilac, zam_orani, zam_tarihi, faiz_yillik, depo_vade)

        if dosya:
            import os
            os.startfile(dosya)

    def _roi_verisi_hesapla(self, maliyet: float, aylik_ort: float, mevcut_stok: int,
                            zam_tarihi, zam_orani: float, faiz_yillik: float, depo_vade: int):
        """ROI bazlı karlılık hesabı"""
        bugun = date.today()
        zam_gun = (zam_tarihi - bugun).days

        if zam_gun <= 0 or aylik_ort <= 0 or maliyet <= 0:
            return None

        aylik_faiz = (faiz_yillik / 100) / 12
        gunluk_faiz = (1 + aylik_faiz) ** (1/30) - 1
        gunluk_sarf = aylik_ort / 30

        max_miktar = int(aylik_ort * 12)
        if max_miktar < 50:
            max_miktar = 200

        miktarlar, kazanclar, yatirimlar, roi_ler, marjinal_roi = [], [], [], [], []

        for test_miktar in range(0, max_miktar + 1):
            if test_miktar == 0:
                miktarlar.append(0)
                kazanclar.append(0)
                yatirimlar.append(0)
                roi_ler.append(0)
                marjinal_roi.append(0)
                continue

            kalan_mevcut = mevcut_stok
            kalan_yeni = test_miktar
            npv_aylik = 0

            gun = 0
            while kalan_yeni > 0 and gun < 720:
                harcanan = gunluk_sarf
                mevcut_harcanan = min(kalan_mevcut, harcanan)
                kalan_mevcut -= mevcut_harcanan
                yeni_harcanan = min(harcanan - mevcut_harcanan, kalan_yeni)

                if yeni_harcanan > 0:
                    fiyat = maliyet if gun < zam_gun else maliyet * (1 + zam_orani / 100)
                    ay_sonu = ((gun // 30) + 1) * 30
                    odeme_gun = ay_sonu + depo_vade
                    iskonto = (1 + gunluk_faiz) ** odeme_gun
                    npv_aylik += (yeni_harcanan * fiyat) / iskonto
                    kalan_yeni -= yeni_harcanan
                gun += 1

            npv_toplu = (test_miktar * maliyet) / ((1 + gunluk_faiz) ** (30 + depo_vade))
            kazanc = npv_aylik - npv_toplu
            yatirim = test_miktar * maliyet

            miktarlar.append(test_miktar)
            kazanclar.append(kazanc)
            yatirimlar.append(yatirim)
            roi = (kazanc / yatirim * 100) if yatirim > 0 else 0
            roi_ler.append(roi)

            if test_miktar > 1:
                ek_kazanc = kazanc - kazanclar[-2]
                m_roi = (ek_kazanc / maliyet * 100) if maliyet > 0 else 0
                marjinal_roi.append(m_roi)
            else:
                marjinal_roi.append(roi)

        # Kritik noktalar
        roi_arr = np.array(roi_ler)
        kazanc_arr = np.array(kazanclar)

        roi_arr_copy = roi_arr.copy()
        roi_arr_copy[0] = -999
        max_roi_idx = int(np.argmax(roi_arr_copy))

        max_roi_val = roi_ler[max_roi_idx]
        yarim_roi_idx = max_roi_idx
        for i in range(max_roi_idx, len(roi_ler)):
            if roi_ler[i] <= max_roi_val / 2:
                yarim_roi_idx = i
                break

        sifir_roi_idx = len(roi_ler) - 1
        for i in range(max_roi_idx, len(roi_ler)):
            if roi_ler[i] <= 0:
                sifir_roi_idx = i
                break

        optimum_idx = int(np.argmax(kazanc_arr))
        pareto_hedef = kazanclar[optimum_idx] * 0.80
        pareto_idx = next((i for i, k in enumerate(kazanclar) if k >= pareto_hedef), 0)

        return {
            'miktarlar': miktarlar, 'kazanclar': kazanclar, 'yatirimlar': yatirimlar,
            'roi_ler': roi_ler, 'marjinal_roi': marjinal_roi, 'mevcut_stok': mevcut_stok,
            'kritik': {
                'max_roi': {'m': miktarlar[max_roi_idx], 'k': kazanclar[max_roi_idx],
                           'y': yatirimlar[max_roi_idx], 'r': roi_ler[max_roi_idx], 'i': max_roi_idx},
                'yarim_roi': {'m': miktarlar[yarim_roi_idx], 'k': kazanclar[yarim_roi_idx],
                             'y': yatirimlar[yarim_roi_idx], 'r': roi_ler[yarim_roi_idx], 'i': yarim_roi_idx},
                'pareto': {'m': miktarlar[pareto_idx], 'k': kazanclar[pareto_idx],
                          'y': yatirimlar[pareto_idx], 'r': roi_ler[pareto_idx], 'i': pareto_idx},
                'optimum': {'m': miktarlar[optimum_idx], 'k': kazanclar[optimum_idx],
                           'y': yatirimlar[optimum_idx], 'r': roi_ler[optimum_idx], 'i': optimum_idx},
                'sifir_roi': {'m': miktarlar[sifir_roi_idx], 'k': kazanclar[sifir_roi_idx],
                             'y': yatirimlar[sifir_roi_idx], 'r': roi_ler[sifir_roi_idx], 'i': sifir_roi_idx}
            }
        }

    def _roi_grafik_ciz(self, ilac: dict, zam_orani: float, zam_tarihi,
                        faiz: float, depo_vade: int) -> str:
        """ROI bazlı grafik çiz"""
        import tempfile
        import os

        sonuc = self._roi_verisi_hesapla(
            ilac['DepocuFiyat'], ilac['AylikOrt'], ilac['Stok'],
            zam_tarihi, zam_orani, faiz, depo_vade
        )

        if not sonuc:
            messagebox.showwarning("Uyarı", f"{ilac['UrunAdi']} için ROI hesaplanamadı.")
            return None

        zam_gun = (zam_tarihi - date.today()).days

        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle(f"ROI Bazli Zam Analizi - Her 100 TL Yatirimda Kazanc\n{ilac['UrunAdi']}",
                     fontsize=14, fontweight='bold', y=0.98)

        m = sonuc['miktarlar']
        k = sonuc['kazanclar']
        y = sonuc['yatirimlar']
        r = sonuc['roi_ler']
        mr = sonuc['marjinal_roi']
        kn = sonuc['kritik']
        stok = sonuc['mevcut_stok']

        colors = {'max_roi': '#FF9800', 'yarim_roi': '#9C27B0', 'pareto': '#4CAF50',
                  'optimum': '#2196F3', 'sifir_roi': '#F44336'}

        # Dinamik eksen ölçeklendirmesi - ilacın kritik noktalarına göre
        x_max = int(kn['sifir_roi']['m'] * 1.15)  # Negatif noktasının %15 ötesine kadar
        if x_max < 10:
            x_max = 20
        y_kazanc_max = kn['optimum']['k'] * 1.2  # Tepe kazancın %20 üstü
        y_kazanc_min = min(k[:x_max+1]) if x_max < len(k) else min(k)
        if y_kazanc_min > 0:
            y_kazanc_min = -y_kazanc_max * 0.1

        # Grafik 1: ROI Eğrisi
        ax1 = axes[0, 0]
        ax1.fill_between(m[:x_max+1], r[:x_max+1], alpha=0.3, color='green', where=[x >= 0 for x in r[:x_max+1]])
        ax1.fill_between(m[:x_max+1], r[:x_max+1], alpha=0.3, color='red', where=[x < 0 for x in r[:x_max+1]])
        ax1.plot(m[:x_max+1], r[:x_max+1], 'b-', linewidth=2.5)
        ax1.axhline(y=0, color='black', linestyle='-', linewidth=1)
        for n, v in kn.items():
            if v['m'] <= x_max:
                ax1.scatter(v['m'], v['r'], c=colors[n], s=150, marker='o', zorder=5, edgecolors='black')
                ax1.axvline(x=v['m'], color=colors[n], linestyle=':', linewidth=1, alpha=0.5)
        if stok > 0 and stok <= x_max:
            ax1.axvline(x=stok, color='gray', linestyle='--', linewidth=2, alpha=0.7)
        ax1.set_xlim(0, x_max)
        ax1.set_xlabel('Siparis Miktari (adet)')
        ax1.set_ylabel('ROI - Her 100 TL\'de Kazanc (%)')
        ax1.set_title('Siparis Miktarina Gore ROI', fontweight='bold')
        ax1.grid(True, alpha=0.3)

        # Grafik 2: Yatırım vs Kazanç (dinamik ölçekli)
        ax2 = axes[0, 1]
        y_limited = y[:x_max+1]
        k_limited = k[:x_max+1]
        ax2.fill_between(y_limited, k_limited, alpha=0.3, color='green', where=[x >= 0 for x in k_limited])
        ax2.fill_between(y_limited, k_limited, alpha=0.3, color='red', where=[x < 0 for x in k_limited])
        ax2.plot(y_limited, k_limited, 'b-', linewidth=2.5)
        ax2.axhline(y=0, color='black', linestyle='-', linewidth=1)
        for n, v in kn.items():
            if v['m'] <= x_max:
                ax2.scatter(v['y'], v['k'], c=colors[n], s=150, marker='o', zorder=5, edgecolors='black')
        ax2.set_ylim(y_kazanc_min * 1.1, y_kazanc_max)
        ax2.set_xlabel('Yatirim (TL)')
        ax2.set_ylabel('Net Kazanc (TL)')
        ax2.set_title('Yatirim vs Kazanc', fontweight='bold')
        ax2.grid(True, alpha=0.3)
        ax2.xaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x/1000:.0f}K'))
        ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))

        # Grafik 3: Marjinal ROI (dinamik ölçekli)
        ax3 = axes[1, 0]
        mr_limited = mr[:x_max+1]
        ax3.fill_between(m[:x_max+1], mr_limited, alpha=0.3, color='orange', where=[x >= 0 for x in mr_limited])
        ax3.fill_between(m[:x_max+1], mr_limited, alpha=0.3, color='red', where=[x < 0 for x in mr_limited])
        ax3.plot(m[:x_max+1], mr_limited, 'orange', linewidth=2)
        ax3.axhline(y=0, color='black', linestyle='-', linewidth=1)
        for n, v in kn.items():
            if v['i'] < len(mr) and v['m'] <= x_max:
                ax3.scatter(v['m'], mr[v['i']], c=colors[n], s=100, marker='o', zorder=5, edgecolors='black')
        ax3.set_xlim(0, x_max)
        ax3.set_xlabel('Siparis Miktari (adet)')
        ax3.set_ylabel('Marjinal ROI (%)')
        ax3.set_title('Marjinal ROI (Her Ek Birim icin)', fontweight='bold')
        ax3.grid(True, alpha=0.3)

        # Grafik 4: Özet Tablo
        ax4 = axes[1, 1]
        ax4.axis('off')
        tablo = f"""
┌─────────────────────────────────────────────────┐
│ {ilac['UrunAdi'][:45]:<45} │
├─────────────────────────────────────────────────┤
│ Depocu: {ilac['DepocuFiyat']:>8.2f} TL  Aylik: {ilac['AylikOrt']:>6.0f} ad. │
│ Stok: {stok:>6} ad.  Zam: %{zam_orani:<4.0f} ({zam_gun} gun)  │
│ Faiz: %{faiz:<4.0f} (yillik)                       │
├─────────────────────────────────────────────────┤
│ KRITIK NOKTALAR (Her 100 TL'de Kazanc)          │
├─────────────────────────────────────────────────┤
│ Maks ROI:  +{kn['max_roi']['m']:>4} ad. = {kn['max_roi']['k']:>7,.0f} TL [%{kn['max_roi']['r']:>5.2f}] │
│            Yatirim: {kn['max_roi']['y']:>10,.0f} TL             │
│            Her 100 TL'de {kn['max_roi']['r']:.2f} TL kazanc     │
├─────────────────────────────────────────────────┤
│ ROI %50:   +{kn['yarim_roi']['m']:>4} ad. = {kn['yarim_roi']['k']:>7,.0f} TL [%{kn['yarim_roi']['r']:>5.2f}] │
│ Pareto:    +{kn['pareto']['m']:>4} ad. = {kn['pareto']['k']:>7,.0f} TL [%{kn['pareto']['r']:>5.2f}] │
│ Tepe:      +{kn['optimum']['m']:>4} ad. = {kn['optimum']['k']:>7,.0f} TL [%{kn['optimum']['r']:>5.2f}] │
│ ROI=0:     +{kn['sifir_roi']['m']:>4} ad. = {kn['sifir_roi']['k']:>7,.0f} TL [%{kn['sifir_roi']['r']:>5.2f}] │
└─────────────────────────────────────────────────┘
"""
        ax4.text(0.05, 0.95, tablo, transform=ax4.transAxes, fontsize=10,
                 verticalalignment='top', fontfamily='monospace',
                 bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.9))

        # Legend
        labels = {'max_roi': f"Maks ROI %{kn['max_roi']['r']:.1f}", 'yarim_roi': f"ROI %50",
                  'pareto': 'Pareto %80', 'optimum': 'Tepe', 'sifir_roi': 'ROI=0'}
        legend_el = [mpatches.Patch(color=colors[n], label=labels[n]) for n in colors]
        fig.legend(handles=legend_el, loc='lower center', ncol=5, fontsize=9, bbox_to_anchor=(0.5, 0.01))

        plt.tight_layout()
        plt.subplots_adjust(bottom=0.08, top=0.93, hspace=0.25, wspace=0.2)

        dosya = os.path.join(tempfile.gettempdir(), f"roi_{ilac['UrunAdi'][:20].replace(' ', '_')}.png")
        plt.savefig(dosya, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return dosya

    def _karlilik_verisi_hesapla(self, maliyet: float, aylik_ort: float, mevcut_stok: int,
                                  zam_tarihi, zam_orani: float, faiz_yillik: float, depo_vade: int):
        """Karlılık verilerini hesapla (grafik için)"""
        bugun = date.today()
        zam_gun = (zam_tarihi - bugun).days

        if zam_gun <= 0 or aylik_ort <= 0 or maliyet <= 0:
            return None

        aylik_faiz = (faiz_yillik / 100) / 12
        gunluk_faiz = (1 + aylik_faiz) ** (1/30) - 1
        gunluk_sarf = aylik_ort / 30

        max_miktar = int(aylik_ort * 12)
        if max_miktar < 50:
            max_miktar = 200

        miktarlar, kazanclar, roi_ler = [], [], []

        for test_miktar in range(0, max_miktar + 1):
            if test_miktar == 0:
                miktarlar.append(0)
                kazanclar.append(0)
                roi_ler.append(0)
                continue

            kalan_mevcut = mevcut_stok
            kalan_yeni = test_miktar
            npv_aylik = 0

            gun = 0
            while kalan_yeni > 0 and gun < 720:
                harcanan = gunluk_sarf
                mevcut_harcanan = min(kalan_mevcut, harcanan)
                kalan_mevcut -= mevcut_harcanan
                yeni_harcanan = min(harcanan - mevcut_harcanan, kalan_yeni)

                if yeni_harcanan > 0:
                    fiyat = maliyet if gun < zam_gun else maliyet * (1 + zam_orani / 100)
                    ay_sonu = ((gun // 30) + 1) * 30
                    odeme_gun = ay_sonu + depo_vade
                    iskonto = (1 + gunluk_faiz) ** odeme_gun
                    npv_aylik += (yeni_harcanan * fiyat) / iskonto
                    kalan_yeni -= yeni_harcanan
                gun += 1

            npv_toplu = (test_miktar * maliyet) / ((1 + gunluk_faiz) ** (30 + depo_vade))
            kazanc = npv_aylik - npv_toplu

            miktarlar.append(test_miktar)
            kazanclar.append(kazanc)
            yatirim = test_miktar * maliyet
            roi_ler.append((kazanc / yatirim * 100) if yatirim > 0 else 0)

        # Marjinal kazançlar
        marjinal = [0] + [kazanclar[i] - kazanclar[i-1] for i in range(1, len(kazanclar))]

        # Kritik noktalar
        kazanc_arr = np.array(kazanclar)
        roi_arr = np.array(roi_ler)

        optimum_idx = np.argmax(kazanc_arr)
        roi_arr[0] = -999
        max_roi_idx = np.argmax(roi_arr)

        pareto_hedef = kazanclar[optimum_idx] * 0.80
        pareto_idx = next((i for i, k in enumerate(kazanclar) if k >= pareto_hedef), 0)

        negatif_idx = max((i for i, k in enumerate(kazanclar) if k > 0), default=0)
        if negatif_idx < len(kazanclar) - 1:
            negatif_idx += 1

        # Azalan verimlilik
        max_marj = max(marjinal[1:]) if len(marjinal) > 1 else 0
        azalan_idx = next((i for i, m in enumerate(marjinal) if m == max_marj), 1)
        yarim_marj = max_marj * 0.5
        azalan_yarim_idx = next((i for i in range(azalan_idx, len(marjinal)) if marjinal[i] <= yarim_marj), azalan_idx)

        return {
            'miktarlar': miktarlar, 'kazanclar': kazanclar, 'roi_ler': roi_ler,
            'marjinal': marjinal, 'mevcut_stok': mevcut_stok,
            'kritik': {
                'max_roi': {'m': miktarlar[max_roi_idx], 'k': kazanclar[max_roi_idx], 'r': roi_ler[max_roi_idx], 'i': max_roi_idx},
                'azalan': {'m': miktarlar[azalan_yarim_idx], 'k': kazanclar[azalan_yarim_idx], 'mr': marjinal[azalan_yarim_idx], 'i': azalan_yarim_idx},
                'pareto': {'m': miktarlar[pareto_idx], 'k': kazanclar[pareto_idx], 'i': pareto_idx},
                'optimum': {'m': miktarlar[optimum_idx], 'k': kazanclar[optimum_idx], 'i': optimum_idx},
                'negatif': {'m': miktarlar[negatif_idx], 'k': kazanclar[negatif_idx], 'i': negatif_idx}
            }
        }

    def _tek_ilac_grafik_ciz(self, ilac: dict, zam_orani: float, zam_tarihi,
                             faiz: float, depo_vade: int) -> str:
        """Tek ilaç için detaylı grafik çiz"""
        import tempfile
        import os

        sonuc = self._karlilik_verisi_hesapla(
            ilac['DepocuFiyat'], ilac['AylikOrt'], ilac['Stok'],
            zam_tarihi, zam_orani, faiz, depo_vade
        )

        if not sonuc:
            messagebox.showwarning("Uyarı", f"{ilac['UrunAdi']} için karlılık hesaplanamadı.\nStok fazla veya zam tarihi geçmiş olabilir.")
            return None

        zam_gun = (zam_tarihi - date.today()).days

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 10), height_ratios=[2.5, 1])
        fig.suptitle(f"Zam Oncesi Siparis Karlilik Analizi - {ilac['UrunAdi']}", fontsize=14, fontweight='bold', x=0.99, ha='right', y=0.98)

        m, k, r, mr = sonuc['miktarlar'], sonuc['kazanclar'], sonuc['roi_ler'], sonuc['marjinal']
        kn = sonuc['kritik']
        stok = sonuc['mevcut_stok']

        colors = {'max_roi': '#FF9800', 'azalan': '#9C27B0', 'pareto': '#4CAF50', 'optimum': '#2196F3', 'negatif': '#F44336'}
        markers = {'max_roi': 'D', 'azalan': 'p', 'pareto': 's', 'optimum': '^', 'negatif': 'X'}
        labels = {'max_roi': 'Maks. ROI', 'azalan': 'Azalan Verim', 'pareto': 'Pareto (%80)', 'optimum': 'Tepe (Pik)', 'negatif': 'Negatife Donus'}

        # Dinamik eksen ölçeklendirmesi - ilacın kritik noktalarına göre
        x_max = int(kn['negatif']['m'] * 1.15)  # Negatif noktasının %15 ötesine kadar
        if x_max < 10:
            x_max = min(20, len(m) - 1)
        x_max = min(x_max, len(m) - 1)
        y_max = kn['optimum']['k'] * 1.25  # Tepe kazancın %25 üstü
        y_min = min(k[:x_max+1]) if x_max < len(k) else min(k)
        if y_min > -y_max * 0.1:
            y_min = -y_max * 0.15

        # Üst grafik (dinamik ölçekli)
        m_lim, k_lim = m[:x_max+1], k[:x_max+1]
        ax1.fill_between(m_lim, k_lim, alpha=0.3, color='green', where=[x >= 0 for x in k_lim])
        ax1.fill_between(m_lim, k_lim, alpha=0.3, color='red', where=[x < 0 for x in k_lim])
        ax1.plot(m_lim, k_lim, 'b-', linewidth=2.5, label='Kazanc Egrisi')
        ax1.axhline(y=0, color='black', linestyle='-', linewidth=1)

        if stok > 0 and stok <= x_max:
            ax1.axvline(x=stok, color='gray', linestyle='--', linewidth=2, alpha=0.7)
            ax1.text(stok + 2, y_max * 0.85, f'Mevcut Stok: {stok}', fontsize=10, color='gray', rotation=90, va='top')

        # Her nokta için sabit yön offset'leri (çakışma olmasın, başlıklara binmesin)
        # Sıra: max_roi (sol-üst), azalan (sağ), pareto (sol-alt), optimum (sol-üst), negatif (sağ-alt)
        offsets = {
            'max_roi': (-x_max * 0.12, y_max * 0.18, -0.3),   # Sol üst
            'azalan': (x_max * 0.15, y_max * 0.10, 0.3),      # Sağ
            'pareto': (-x_max * 0.12, -y_max * 0.18, 0.3),    # Sol alt
            'optimum': (-x_max * 0.15, y_max * 0.15, -0.3),   # Sol üst (başlığa binmesin)
            'negatif': (x_max * 0.10, -y_max * 0.22, -0.3)    # Sağ alt
        }

        for n, v in kn.items():
            if v['m'] <= x_max:
                ax1.scatter(v['m'], v['k'], c=colors[n], s=200, marker=markers[n], zorder=5, edgecolors='black', linewidths=1.5)
                ax1.axvline(x=v['m'], color=colors[n], linestyle=':', linewidth=1, alpha=0.5)
                extra = f"\nROI: %{v.get('r', 0):.1f}" if 'r' in v else (f"\nMarj: {v.get('mr', 0):.2f}" if 'mr' in v else "")

                x_off, y_off, rad = offsets.get(n, (0, y_max * 0.15, 0))

                ax1.annotate(f"{labels[n]}\n{v['m']} ad.\n{v['k']:,.0f} TL{extra}", xy=(v['m'], v['k']),
                            xytext=(v['m'] + x_off, v['k'] + y_off), fontsize=9, ha='center', fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.4', facecolor=colors[n], alpha=0.8, edgecolor='black'),
                            arrowprops=dict(arrowstyle='->', color=colors[n], lw=2, connectionstyle=f'arc3,rad={rad}'))

        ax1.set_xlim(0, x_max)
        ax1.set_ylim(y_min, y_max)
        ax1.set_xlabel('Siparis Miktari (adet)', fontsize=12)
        ax1.set_ylabel('Net Kazanc (TL)', fontsize=12)
        ax1.set_title('Siparis Miktarina Gore Net Kazanc', fontsize=12, loc='right', pad=5)
        ax1.grid(True, alpha=0.3)
        ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))

        # Alt grafik (dinamik ölçekli)
        r_lim, mr_lim = r[:x_max+1], mr[:x_max+1]
        ax2_twin = ax2.twinx()
        ax2.plot(m_lim, r_lim, 'g-', linewidth=2, label='ROI (%)')
        ax2.set_ylabel('ROI (%)', color='green', fontsize=11)
        ax2.tick_params(axis='y', labelcolor='green')
        ax2.axhline(y=0, color='black', linestyle='-', linewidth=0.5)

        ax2_twin.plot(m_lim, mr_lim, 'orange', linewidth=2, linestyle='--', label='Marjinal (TL/ad)')
        ax2_twin.set_ylabel('Marjinal Kazanc (TL/ad)', color='orange', fontsize=11)
        ax2_twin.tick_params(axis='y', labelcolor='orange')
        ax2_twin.axhline(y=0, color='orange', linestyle=':', linewidth=0.5, alpha=0.5)

        for n, v in kn.items():
            if v['i'] < len(r) and v['m'] <= x_max:
                ax2.scatter(v['m'], r[v['i']], c=colors[n], s=100, marker=markers[n], zorder=5, edgecolors='black')

        ax2.set_xlim(0, x_max)
        ax2.set_xlabel('Siparis Miktari (adet)', fontsize=12)
        ax2.set_title('ROI ve Marjinal Kazanc Egrileri', fontsize=12, loc='right', pad=5)
        ax2.grid(True, alpha=0.3)
        ax2.legend(['ROI (%)', 'Marjinal (TL/ad)'], loc='upper right')

        # Bilgi kutusu
        bilgi = (f"Ilac: {ilac['UrunAdi']}\nDepocu: {ilac['DepocuFiyat']:,.2f} TL\n"
                f"Aylik: {ilac['AylikOrt']:.1f} ad.\nStok: {stok} ad.\n{'─'*25}\n"
                f"Zam: %{zam_orani} | {zam_gun} gun\nFaiz: %{faiz}\n{'─'*25}\n"
                f"Maks.ROI: {kn['max_roi']['m']} ad. ({kn['max_roi']['k']:,.0f} TL)\n"
                f"Azalan: {kn['azalan']['m']} ad.\nPareto: {kn['pareto']['m']} ad.\n"
                f"Tepe: {kn['optimum']['m']} ad. ({kn['optimum']['k']:,.0f} TL)\nSinir: {kn['negatif']['m']} ad.")
        fig.text(0.01, 0.01, bilgi, fontsize=9, va='bottom', bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.9), family='monospace')

        legend_el = [mpatches.Patch(color=colors[n], label=labels[n]) for n in colors]
        fig.legend(handles=legend_el, loc='lower right', fontsize=9, bbox_to_anchor=(0.99, 0.01))

        plt.tight_layout()
        plt.subplots_adjust(bottom=0.22, top=0.92, hspace=0.25)

        dosya = os.path.join(tempfile.gettempdir(), f"karlilik_{ilac['UrunAdi'][:20].replace(' ', '_')}.png")
        plt.savefig(dosya, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return dosya

    def _coklu_ilac_grafik_ciz(self, ilaclar: list, zam_orani: float, zam_tarihi,
                               faiz: float, depo_vade: int) -> str:
        """Birden fazla ilaç için karşılaştırmalı grafik"""
        import tempfile
        import os

        zam_gun = (zam_tarihi - date.today()).days

        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        axes = axes.flatten()
        fig.suptitle(f"Zam Oncesi Karlilik | Zam: %{zam_orani} | {zam_gun} gun | Faiz: %{faiz}", fontsize=12, fontweight='bold', x=0.99, ha='right', y=0.98)

        colors = {'max_roi': '#FF9800', 'azalan': '#9C27B0', 'pareto': '#4CAF50', 'optimum': '#2196F3', 'negatif': '#F44336'}
        markers = {'max_roi': 'D', 'azalan': 'p', 'pareto': 's', 'optimum': '^', 'negatif': 'X'}
        labels = {'max_roi': 'Maks. ROI', 'azalan': 'Azalan Verim', 'pareto': 'Pareto (%80)', 'optimum': 'Tepe', 'negatif': 'Negatif'}

        for i, ilac in enumerate(ilaclar[:4]):
            ax = axes[i]
            sonuc = self._karlilik_verisi_hesapla(ilac['DepocuFiyat'], ilac['AylikOrt'], ilac['Stok'], zam_tarihi, zam_orani, faiz, depo_vade)

            if not sonuc:
                ax.text(0.5, 0.5, f"{ilac['UrunAdi']}\nHesaplanamadi", ha='center', va='center', fontsize=12)
                ax.set_title(ilac['UrunAdi'][:35], fontsize=10, fontweight='bold', loc='right')
                continue

            m, k, kn, stok = sonuc['miktarlar'], sonuc['kazanclar'], sonuc['kritik'], sonuc['mevcut_stok']

            ax.fill_between(m, k, alpha=0.3, color='green', where=[x >= 0 for x in k])
            ax.fill_between(m, k, alpha=0.3, color='red', where=[x < 0 for x in k])
            ax.plot(m, k, 'b-', linewidth=2)
            ax.axhline(y=0, color='black', linewidth=0.5)

            if stok > 0 and stok < max(m):
                ax.axvline(x=stok, color='gray', linestyle='--', linewidth=2, alpha=0.7)
                ax.text(stok, max(k) * 0.8, f' {stok}', fontsize=8, color='gray', rotation=90, va='top')

            x_max = max(m) if m else 100
            y_max = max(k) if k else 100

            # Sabit yön offset'leri (çakışma olmasın, başlıklara binmesin)
            offsets = {
                'max_roi': (-x_max * 0.10, y_max * 0.15, -0.25),
                'azalan': (x_max * 0.12, y_max * 0.08, 0.25),
                'pareto': (-x_max * 0.10, -y_max * 0.15, 0.25),
                'optimum': (-x_max * 0.12, y_max * 0.12, -0.25),  # Sol üst (başlığa binmesin)
                'negatif': (x_max * 0.08, -y_max * 0.18, -0.25)
            }

            for n, v in kn.items():
                ax.scatter(v['m'], v['k'], c=colors[n], s=120, marker=markers[n], zorder=5, edgecolors='black')
                x_off, y_off, rad = offsets.get(n, (0, y_max * 0.12, 0))
                ax.annotate(f"{labels[n]}\n{v['m']} ad.\n{v['k']:,.0f} TL", xy=(v['m'], v['k']),
                           xytext=(v['m'] + x_off, v['k'] + y_off), fontsize=8, ha='center',
                           bbox=dict(boxstyle='round,pad=0.2', facecolor=colors[n], alpha=0.6),
                           arrowprops=dict(arrowstyle='->', color=colors[n], lw=1, connectionstyle=f'arc3,rad={rad}'))

            ax.text(0.02, 0.98, f"Stok: {stok}\nAylik: {ilac['AylikOrt']:.0f}\nFiyat: {ilac['DepocuFiyat']:.1f}",
                   transform=ax.transAxes, fontsize=9, va='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
            ax.set_title(ilac['UrunAdi'][:35], fontsize=10, fontweight='bold', loc='right')
            ax.set_xlabel('Siparis (adet)', fontsize=10)
            ax.set_ylabel('Kazanc (TL)', fontsize=10)
            ax.grid(True, alpha=0.3)
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))

        for i in range(len(ilaclar), 4):
            axes[i].axis('off')

        legend_el = [mpatches.Patch(color=colors[n], label=labels[n]) for n in colors] + [mpatches.Patch(color='gray', alpha=0.5, label='Mevcut Stok')]
        fig.legend(handles=legend_el, loc='lower center', ncol=3, fontsize=9, bbox_to_anchor=(0.5, 0.01))

        plt.tight_layout()
        plt.subplots_adjust(bottom=0.10, top=0.92, hspace=0.3)

        dosya = os.path.join(tempfile.gettempdir(), "karlilik_karsilastirma.png")
        plt.savefig(dosya, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return dosya

    def _mf_optimum_hesapla(self, mf_sart: str, maliyet: float, aylik_ort: float,
                            mevcut_stok: int, faiz_yillik: float, depo_vade: int,
                            hedef_gun: int) -> dict:
        """
        MF şartına göre optimum alım miktarını hesapla.

        Örnek: 5+1 şartı için:
        - 5 al 1 bedava = 6 ürün için 5 ürün parası
        - Ama 6 ay stok tutmak faiz maliyeti
        - Hangisi karlı?

        Returns:
            {
                'mfli_al': bool,       # MF'li almak karlı mı?
                'alinan': int,         # Alınacak miktar
                'bedava': int,         # Bedava gelecek
                'toplam': int,         # Toplam stok
                'kazanc': float,       # Net kazanç/kayıp
                'oneri': str           # "5+1 al" veya "2+0 al (MF'siz)"
            }
        """
        if not mf_sart or '+' not in mf_sart:
            return {'mfli_al': False, 'alinan': 0, 'bedava': 0, 'toplam': 0, 'kazanc': 0, 'oneri': ''}

        try:
            parcalar = mf_sart.split('+')
            mf_alinan = int(parcalar[0])
            mf_bedava = int(parcalar[1])
        except:
            return {'mfli_al': False, 'alinan': 0, 'bedava': 0, 'toplam': 0, 'kazanc': 0, 'oneri': ''}

        if mf_alinan <= 0 or mf_bedava <= 0 or maliyet <= 0 or aylik_ort <= 0:
            return {'mfli_al': False, 'alinan': 0, 'bedava': 0, 'toplam': 0, 'kazanc': 0, 'oneri': ''}

        gunluk_sarf = aylik_ort / 30
        hedef_ihtiyac = max(0, int(gunluk_sarf * hedef_gun) - mevcut_stok)

        # Hedef ihtiyacı karşılayacak MF'siz miktar
        mfsiz_miktar = hedef_ihtiyac

        # MF'li alım için kaç set gerekiyor?
        mf_toplam = mf_alinan + mf_bedava
        set_sayisi = max(1, (hedef_ihtiyac + mf_toplam - 1) // mf_toplam)
        mfli_alinan = set_sayisi * mf_alinan
        mfli_bedava = set_sayisi * mf_bedava
        mfli_toplam = mfli_alinan + mfli_bedava

        # NPV karşılaştırması
        bugun = date.today()
        kalan_gun = hedef_gun

        # MF'siz NPV
        npv_mfsiz, _, _ = self._npv_hesapla(
            alinan=mfsiz_miktar, mf=0, maliyet=maliyet, aylik_ort=aylik_ort,
            faiz_yillik=faiz_yillik, depo_vade=depo_vade,
            fatura_tarihi=bugun, zam_tarihi=None, zam_orani=0,
            mevcut_stok=mevcut_stok, kalan_gun=kalan_gun
        )

        # MF'li NPV
        _, npv_mfli, _ = self._npv_hesapla(
            alinan=mfli_alinan, mf=mfli_bedava, maliyet=maliyet, aylik_ort=aylik_ort,
            faiz_yillik=faiz_yillik, depo_vade=depo_vade,
            fatura_tarihi=bugun, zam_tarihi=None, zam_orani=0,
            mevcut_stok=mevcut_stok, kalan_gun=kalan_gun
        )

        # Kazanç hesapla (pozitif = MF'li karlı)
        kazanc = npv_mfsiz - npv_mfli

        if kazanc > 0:
            return {
                'mfli_al': True,
                'alinan': mfli_alinan,
                'bedava': mfli_bedava,
                'toplam': mfli_toplam,
                'kazanc': round(kazanc, 2),
                'oneri': f"{mfli_alinan}+{mfli_bedava} al ({kazanc:.0f}₺ karlı)"
            }
        else:
            return {
                'mfli_al': False,
                'alinan': mfsiz_miktar,
                'bedava': 0,
                'toplam': mfsiz_miktar,
                'kazanc': round(kazanc, 2),
                'oneri': f"{mfsiz_miktar}+0 al (MF'siz daha karlı)"
            }

    def _hedef_gun_hesapla(self):
        """Hedef tarihe kalan gün sayısı.

        NOT: Zam tarihi burada hedef olarak KULLANILMAZ.
        Zam optimizasyonu zam tarihini ayrıca alır (_zam_oncesi_optimum_hesapla).
        Hedef gün, temel ihtiyaç hesaplaması için kullanılır (ay sonu veya kullanıcı hedefi).
        Zam aktif olup tarih bugün/yarın gibi yakın olduğunda hedef_gun=0 olması
        tüm sipariş önerilerini sıfırlıyordu - bu düzeltildi.
        """
        try:
            if self.hedef_tarih_aktif.get():
                hedef = self.hedef_tarih_entry.get_date()
            else:
                # Varsayılan: Ay sonu
                bugun = datetime.now()
                ay_son_gun = calendar.monthrange(bugun.year, bugun.month)[1]
                hedef = date(bugun.year, bugun.month, ay_son_gun)

            bugun = date.today()
            fark = (hedef - bugun).days
            return max(0, fark)
        except:
            return 0

    # ═══════════════════════════════════════════════════════════════════════
    # VERİ GETİRME
    # ═══════════════════════════════════════════════════════════════════════

    def verileri_getir(self):
        """Veritabanından verileri getir"""
        secili_tipler = [t for t, v in self.urun_tipi_vars.items() if v.get()]

        if not secili_tipler:
            messagebox.showwarning("Uyarı", "En az bir ürün tipi seçmelisiniz!")
            return

        self.status_label.config(text="Veriler getiriliyor...")
        self.parent.update()

        # Loading indicator başlat
        loading = None
        if LOADING_INDICATOR_YUKLENDI:
            loading = LoadingIndicator(self.parent, "Stok verileri sorgulanıyor...", stil="progress")
            loading.show()

        def sorgu_thread():
            try:
                from botanik_db import BotanikDB
                db = BotanikDB()
                if not db.baglan():
                    self.parent.after(0, lambda: self._sorgu_tamamlandi(loading, hata="Veritabanına bağlanılamadı!"))
                    return

                sene = self.sene_sayisi.get()
                ay = self.ay_sayisi.get()

                # Ana verileri getir
                veriler = self._verileri_getir_sql(db, sene, ay, secili_tipler)

                # MF şartlarını getir
                mf_sartlari = self._mf_sartlari_getir(db)

                db.kapat()

                # Verileri işle
                islenenmis = self._verileri_isle(veriler, mf_sartlari, ay)

                self.parent.after(0, lambda: self._sorgu_tamamlandi(loading, islenenmis, ay))

            except Exception as e:
                logger.error(f"Sorgu hatası: {e}")
                import traceback
                traceback.print_exc()
                self.parent.after(0, lambda: self._sorgu_tamamlandi(loading, hata=str(e)))

        thread = threading.Thread(target=sorgu_thread)
        thread.start()

    def _sorgu_tamamlandi(self, loading, veriler=None, ay=None, hata=None):
        """Sorgu tamamlandığında çağrılır"""
        # Loading'i kapat
        if loading:
            loading.hide()

        if hata:
            messagebox.showerror("Hata", hata)
        elif veriler is not None:
            self._veriler_yuklendi(veriler, ay)

    def _verileri_getir_sql(self, db, sene_sayisi, ay_sayisi, urun_tipleri):
        """SQL sorgusu - X sene hareket görmüş + Y ay aylık dağılım"""
        bugun = datetime.now()

        # X sene içinde hareket görmüş ürünler için
        hareket_baslangic = (bugun - relativedelta(years=sene_sayisi)).strftime('%Y-%m-%d')

        # Y ay için aylık dağılım
        aylik_baslangic = (bugun - relativedelta(months=ay_sayisi)).replace(day=1).strftime('%Y-%m-%d')

        tipler_sql = ', '.join([f"'{t}'" for t in urun_tipleri])

        # Aylık CASE ifadeleri
        aylik_cases = []
        for i in range(ay_sayisi):
            ay_tarihi = bugun - relativedelta(months=i)
            ay_basi = ay_tarihi.replace(day=1)
            if i == 0:
                ay_sonu = bugun
            else:
                ay_sonu = (ay_basi + relativedelta(months=1)) - relativedelta(days=1)

            baslangic_str = ay_basi.strftime('%Y-%m-%d')
            bitis_str = ay_sonu.strftime('%Y-%m-%d')

            aylik_cases.append(f"""
                SUM(CASE WHEN Tarih >= '{baslangic_str}' AND Tarih <= '{bitis_str}' THEN Adet ELSE 0 END) as Ay_{i}
            """)

        aylik_kolonlar = ",\n".join(aylik_cases)

        sql = f"""
        ;WITH HareketliUrunler AS (
            -- Son X sene içinde hareket görmüş ürünler
            SELECT DISTINCT ri.RIUrunId as UrunId
            FROM ReceteIlaclari ri
            JOIN ReceteAna ra ON ri.RIRxId = ra.RxId
            WHERE ra.RxSilme = 0 AND ri.RISilme = 0
            AND ra.RxKayitTarihi >= '{hareket_baslangic}'

            UNION

            SELECT DISTINCT ei.RIUrunId as UrunId
            FROM EldenIlaclari ei
            JOIN EldenAna ea ON ei.RIRxId = ea.RxId
            WHERE ea.RxSilme = 0 AND ei.RISilme = 0
            AND ea.RxKayitTarihi >= '{hareket_baslangic}'
        ),
        CikisVerileri AS (
            SELECT ri.RIUrunId as UrunId, ri.RIAdet as Adet, CAST(ra.RxKayitTarihi as date) as Tarih
            FROM ReceteIlaclari ri
            JOIN ReceteAna ra ON ri.RIRxId = ra.RxId
            WHERE ra.RxSilme = 0 AND ri.RISilme = 0
            AND (ri.RIIade = 0 OR ri.RIIade IS NULL)
            AND ra.RxKayitTarihi >= '{aylik_baslangic}'

            UNION ALL

            SELECT ei.RIUrunId as UrunId, ei.RIAdet as Adet, CAST(ea.RxKayitTarihi as date) as Tarih
            FROM EldenIlaclari ei
            JOIN EldenAna ea ON ei.RIRxId = ea.RxId
            WHERE ea.RxSilme = 0 AND ei.RISilme = 0
            AND (ei.RIIade = 0 OR ei.RIIade IS NULL)
            AND ea.RxKayitTarihi >= '{aylik_baslangic}'
        ),
        UrunAylikOzet AS (
            SELECT UrunId, SUM(Adet) as ToplamCikis,
                {aylik_kolonlar}
            FROM CikisVerileri
            GROUP BY UrunId
        )

        SELECT
            u.UrunId,
            u.UrunAdi,
            COALESCE(ut.UrunTipAdi, 'Belirsiz') as UrunTipi,
            u.UrunEsdegerId as EsdegerId,
            -- Stok: Sadece İLAÇ(1) ve PASİF İLAÇ(16) için Karekod, diğerleri için EskiStok
            CASE
                WHEN u.UrunUrunTipId IN (1, 16) THEN
                    (SELECT COUNT(*) FROM Karekod kk WHERE kk.KKUrunId = u.UrunId AND kk.KKDurum = 1)
                ELSE
                    (COALESCE(u.UrunStokDepo,0) + COALESCE(u.UrunStokRaf,0) + COALESCE(u.UrunStokAcik,0))
            END as Stok,
            COALESCE(u.UrunMinimum, 0) as MinStok,
            COALESCE(ao.ToplamCikis, 0) as ToplamCikis,
            -- Fiyat bilgileri (Depocu fiyat hesaplaması için)
            COALESCE(u.UrunFiyatEtiket, 0) as PSF,
            COALESCE(u.UrunIskontoKamu, 0) as IskontoKamu,
            -- Barkod (Barkod tablosundan)
            (SELECT TOP 1 b.BarkodAdi FROM Barkod b WHERE b.BarkodUrunId = u.UrunId AND b.BarkodSilme = 0) as Barkod,
            {', '.join([f'COALESCE(ao.Ay_{i}, 0) as Ay_{i}' for i in range(ay_sayisi)])}
        FROM Urun u
        INNER JOIN HareketliUrunler hu ON u.UrunId = hu.UrunId
        LEFT JOIN UrunTip ut ON u.UrunUrunTipId = ut.UrunTipId
        LEFT JOIN UrunAylikOzet ao ON u.UrunId = ao.UrunId
        WHERE u.UrunSilme = 0
        AND ut.UrunTipAdi IN ({tipler_sql})
        ORDER BY u.UrunEsdegerId, u.UrunAdi
        """

        return db.sorgu_calistir(sql)

    def _mf_sartlari_getir(self, db):
        """Son 1 yılda alınan MF şartlarını getir (5+1 formatında)"""
        bugun = datetime.now()
        baslangic = (bugun - relativedelta(years=1)).strftime('%Y-%m-%d')

        sql = f"""
        SELECT
            fs.FSUrunId as UrunId,
            CAST(fs.FSUrunAdet as int) as Adet,
            CAST(fs.FSUrunMf as int) as MF,
            fg.FGFaturaTarihi as Tarih
        FROM FaturaSatir fs
        JOIN FaturaGiris fg ON fs.FSFGId = fg.FGId
        WHERE fg.FGSilme = 0
        AND fg.FGFaturaTarihi >= '{baslangic}'
        AND fs.FSUrunMf > 0
        ORDER BY fs.FSUrunId, fg.FGFaturaTarihi DESC
        """

        sonuclar = db.sorgu_calistir(sql)

        # Her ürün için en fazla 3 farklı şart
        mf_sartlari = {}
        for row in sonuclar:
            urun_id = row['UrunId']
            adet = row['Adet'] or 0
            mf = row['MF'] or 0

            if mf > 0:
                sart = f"{adet}+{mf}"
                if urun_id not in mf_sartlari:
                    mf_sartlari[urun_id] = []

                # Aynı şart yoksa ekle
                if sart not in mf_sartlari[urun_id] and len(mf_sartlari[urun_id]) < 3:
                    mf_sartlari[urun_id].append(sart)

        return mf_sartlari

    def _verileri_isle(self, veriler, mf_sartlari, ay_sayisi):
        """Verileri işle ve hesapla"""
        if not veriler:
            return []

        hedef_gun = self._hedef_gun_hesapla()
        min_stok_aktif = self.min_stok_aktif.get()
        min_stok_kaynagi = self.min_stok_kaynagi.get()  # 'botanik' veya 'yerel'
        zam_aktif = self.zam_aktif.get()
        zam_orani = self.beklenen_zam_orani.get() / 100 if zam_aktif else 0
        zam_stratejisi = self.zam_stratejisi.get() if zam_aktif else 'pareto'

        # Yerel min stok verilerini cache'e al (kaynak 'yerel' ise)
        yerel_min_stok_cache = {}
        if min_stok_kaynagi == 'yerel':
            try:
                from siparis_db import get_siparis_db
                siparis_db = get_siparis_db()
                yerel_liste = siparis_db.min_stok_listesi_getir()
                yerel_min_stok_cache = {v['urun_id']: v['min_onerilen'] for v in yerel_liste}
                logger.info(f"Yerel min stok cache: {len(yerel_min_stok_cache)} kayıt")
            except Exception as e:
                logger.warning(f"Yerel min stok yüklenemedi: {e}")

        # Efektif ay sayısı hesaplama (12 ay için düzeltme)
        # Sorun: 12 ay seçilince içinde bulunulan ay kısmi, toplam 11.X ay oluyor
        # Çözüm: Efektif ay sayısı = (tam aylar) + (bu ayın geçen günü / 30)
        bugun = datetime.now()
        ay_basi = bugun.replace(day=1)
        bu_ay_gecen_gun = (bugun - ay_basi).days + 1  # Bugün dahil

        # Efektif ay: (ay_sayisi - 1) tam ay + bu ayın oranı
        efektif_ay_sayisi = (ay_sayisi - 1) + (bu_ay_gecen_gun / 30)

        islenenmis = []

        for veri in veriler:
            urun_id = veri.get('UrunId')
            stok = veri.get('Stok', 0) or 0

            # Min stok kaynağına göre değeri al
            if min_stok_kaynagi == 'yerel' and urun_id in yerel_min_stok_cache:
                min_stok = yerel_min_stok_cache.get(urun_id, 0) or 0
            else:
                min_stok = veri.get('MinStok', 0) or 0

            toplam_cikis = veri.get('ToplamCikis', 0) or 0

            # Fiyat bilgileri
            psf = veri.get('PSF', 0) or 0
            iskonto_kamu = veri.get('IskontoKamu', 0) or 0
            depocu_fiyat = self._depocu_fiyat_hesapla(psf, iskonto_kamu)

            # Aylık ve günlük ortalama - efektif ay sayısı kullan
            aylik_ort = toplam_cikis / efektif_ay_sayisi if efektif_ay_sayisi > 0 else 0
            gunluk_ort = aylik_ort / 30

            # Ay bitiş (stok kaç ay yeter)
            # KURAL 1: Stok=0 → Ay Bitiş=0 (stok yok)
            # KURAL 2: Stok>0 ve Gidiş=0 → Ay Bitiş=999 (sonsuz stok)
            # KURAL 3: Stok>0 ve Gidiş>0 → Normal hesaplama
            if stok == 0:
                ay_bitis = 0
            elif aylik_ort == 0:
                ay_bitis = 999  # Stok var ama gidiş yok = sonsuz
            else:
                ay_bitis = stok / aylik_ort

            # Zam analizi (NPV bazlı) - zam aktif ise öncelikli hesapla
            zam_oneri = None
            if zam_aktif and zam_orani > 0 and depocu_fiyat > 0 and aylik_ort > 0:
                try:
                    zam_tarihi = self.zam_tarih_entry.get_date()
                    faiz = self._aktif_faiz_getir()
                    vade = self.depo_vadesi.get()

                    zam_oneri = self._zam_oncesi_optimum_hesapla(
                        maliyet=depocu_fiyat,
                        aylik_ort=aylik_ort,
                        mevcut_stok=stok,
                        zam_tarihi=zam_tarihi,
                        zam_orani=zam_orani * 100,  # Yüzde olarak
                        faiz_yillik=faiz,
                        depo_vade=vade
                    )
                except Exception as e:
                    logger.warning(f"Zam hesaplama hatası: {e}")

            # Sipariş önerisi hesaplama
            if zam_aktif and zam_oneri:
                # ZAM AKTİF: Seçilen stratejiye göre hesapla
                strateji = zam_stratejisi

                if strateji == 'verimlilik' and zam_oneri.get('verimlilik', 0) > 0:
                    # Maksimum ROI noktası - en verimli sermaye kullanımı
                    secilen_miktar = zam_oneri['verimlilik']
                elif strateji == 'optimum' and zam_oneri.get('optimum', 0) > 0:
                    # Maksimum mutlak kazanç noktası - agresif
                    secilen_miktar = zam_oneri['optimum']
                elif zam_oneri.get('pareto', 0) > 0:
                    # Pareto (%80 kazanç) - dengeli (varsayılan)
                    secilen_miktar = zam_oneri['pareto']
                else:
                    secilen_miktar = 0

                # NOT: secilen_miktar zaten YENİ ALIM miktarıdır.
                # Algoritma mevcut stoğu hesaba katarak optimal yeni alımı döndürür.
                # Stok tekrar çıkarılMAMALI!
                oneri = secilen_miktar
            else:
                # NORMAL MOD: Hedef güne kadar gereken temel ihtiyaç
                temel_ihtiyac = gunluk_ort * hedef_gun
                oneri = max(0, temel_ihtiyac - stok)

            # Minimum stok kontrolü (her durumda uygula)
            if min_stok_aktif and min_stok > 0 and stok < min_stok:
                min_eksik = min_stok - stok
                oneri = max(oneri, min_eksik)

            # MF şartları
            sartlar = mf_sartlari.get(urun_id, [])

            esdeger_id = veri.get('EsdegerId')

            # Öneri kaç aylık stok?
            oneri_int = int(round(oneri, 0))
            if aylik_ort > 0:
                oneri_ay = oneri_int / aylik_ort
                yeni_ay_bitis = (stok + oneri_int) / aylik_ort
            else:
                oneri_ay = 0 if oneri_int == 0 else 999
                yeni_ay_bitis = 999 if stok > 0 or oneri_int > 0 else 0

            satir = {
                'UrunId': urun_id,
                'UrunAdi': veri.get('UrunAdi', ''),
                'UrunTipi': veri.get('UrunTipi', ''),
                'EsdegerId': esdeger_id,
                'Barkod': veri.get('Barkod', ''),
                'Stok': stok,
                'MinStok': min_stok,
                'ToplamCikis': toplam_cikis,
                'AylikOrt': round(aylik_ort, 1),
                'GunlukOrt': round(gunluk_ort, 2),
                'AyBitis': round(ay_bitis, 1) if ay_bitis < 100 else "∞",
                'HedefGun': hedef_gun,
                'Oneri': oneri_int,
                'OneriAy': round(oneri_ay, 1) if oneri_ay < 100 else "∞",
                'YeniAyBitis': round(yeni_ay_bitis, 1) if yeni_ay_bitis < 100 else "∞",
                'Manuel': self.manuel_miktarlar.get(urun_id, 0),
                'MF': self.manuel_mf_girisler.get(urun_id, ''),
                'Sart1': sartlar[0] if len(sartlar) > 0 else '',
                'Sart2': sartlar[1] if len(sartlar) > 1 else '',
                'Sart3': sartlar[2] if len(sartlar) > 2 else '',
                'ZamOneri': zam_oneri,  # Zam optimizasyon bilgisi
                'DepocuFiyat': depocu_fiyat,  # Bize geliş fiyatı
                'PSF': psf,  # Perakende satış fiyatı
            }

            # Aylık verileri ekle
            for i in range(ay_sayisi):
                satir[f'Ay_{i}'] = veri.get(f'Ay_{i}', 0) or 0

            islenenmis.append(satir)

        return islenenmis

    def _veriler_yuklendi(self, veriler, ay_sayisi):
        """Veriler yüklendiğinde UI güncelle"""
        self.tum_veriler = veriler

        # Sütunları güncelle
        self._sutunlari_guncelle(ay_sayisi)

        # Filtreleme uygula ve tabloyu güncelle
        self._filtreleme_uygula()

        # Bilgi güncelle
        hedef_gun = self._hedef_gun_hesapla()
        siparis_gereken = len([v for v in veriler if v.get('Oneri', 0) > 0])
        toplam_kutu = sum(v.get('Oneri', 0) for v in veriler if v.get('Oneri', 0) > 0)
        toplam_tutar = sum(v.get('Oneri', 0) * v.get('DepocuFiyat', 0) for v in veriler if v.get('Oneri', 0) > 0)
        aktif_faiz = self._aktif_faiz_getir()
        faiz_tur = "M" if self.faiz_turu.get() == "mevduat" else "K"

        if self.zam_aktif.get():
            zam_orani = self.beklenen_zam_orani.get()
            strateji = self.zam_stratejisi.get()
            strateji_adi = {'pareto': 'Pareto', 'optimum': 'Maks.Kazanç', 'verimlilik': 'Maks.ROI'}.get(strateji, strateji)
            mod_bilgi = f"ZAM: %{zam_orani:.0f} | {strateji_adi} | {hedef_gun} gün"
        else:
            mod_bilgi = f"Hedef: {hedef_gun} gün"

        self.hesaplama_label.config(
            text=f"{mod_bilgi} | Faiz: %{aktif_faiz:.0f}({faiz_tur}) | {len(veriler)} ürün | {siparis_gereken} kalem ({toplam_kutu} kutu) | {toplam_tutar:,.0f} ₺"
        )

        self.status_label.config(text="Veriler yüklendi")
        self.kayit_label.config(text=f"{len(veriler)} kayıt")

    def _sutunlari_guncelle(self, ay_sayisi):
        """Sütunları dinamik güncelle"""
        # Aylık sütunları oluştur
        bugun = datetime.now()
        self.aylik_sutunlar = []
        for i in range(ay_sayisi - 1, -1, -1):
            ay_tarihi = bugun - relativedelta(months=i)
            ay_adi = ay_tarihi.strftime('%b')[:3]
            self.aylik_sutunlar.append((f"Ay_{i}", ay_adi, 40))

        # Tüm sütunları birleştir
        tum_sutunlar = self.ana_sutunlar + self.aylik_sutunlar + self.son_sutunlar

        col_ids = [c[0] for c in tum_sutunlar]
        self.ana_tree['columns'] = col_ids

        # Vurgulu sutunlar: Stok, Aylik satislar, Siparis
        vurgulu_sutunlar = {'Stok', 'AylikOrt', 'Oneri'}
        aylik_sutun_ids = {f"Ay_{i}" for i in range(12)}

        for col_id, baslik, width in tum_sutunlar:
            # Vurgulu basliklar
            if col_id == 'Stok':
                gosterim = "[ STOK ]"
            elif col_id == 'AylikOrt':
                gosterim = "[ AYLIK ]"
            elif col_id == 'Oneri':
                gosterim = ">> SIPARIS <<"
            elif col_id in aylik_sutun_ids:
                gosterim = f"◆{baslik}"  # Aylik sutunlar mavi karo ile
            else:
                gosterim = baslik

            self.ana_tree.heading(col_id, text=gosterim)
            self.ana_tree.column(col_id, width=width, minwidth=25)

        self.aktif_sutunlar = tum_sutunlar

    def _filtreleme_uygula(self):
        """Yeterlileri gizle filtresini uygula ve tabloyu güncelle"""
        if not self.tum_veriler:
            self.gorunen_veriler = []
            self._tabloyu_guncelle()
            return

        if not self.yeterlileri_gizle.get():
            self.gorunen_veriler = self.tum_veriler
        else:
            # Eşdeğer gruplarını analiz et
            esdeger_gruplari = {}
            for veri in self.tum_veriler:
                eid = veri.get('EsdegerId') or 0
                if eid not in esdeger_gruplari:
                    esdeger_gruplari[eid] = []
                esdeger_gruplari[eid].append(veri)

            # Filtreleme
            gorunen = []
            for eid, urunler in esdeger_gruplari.items():
                grup_siparis_var = any(u.get('Oneri', 0) > 0 for u in urunler)

                if eid == 0:
                    # Eşdeğersiz - sadece sipariş gerekenleri göster
                    for u in urunler:
                        if u.get('Oneri', 0) > 0:
                            gorunen.append(u)
                else:
                    # Eşdeğerli grup - grupta herhangi biri sipariş gerektiriyorsa tümünü göster
                    if grup_siparis_var:
                        gorunen.extend(urunler)

            self.gorunen_veriler = gorunen

        self._tabloyu_guncelle()

    def _tabloyu_guncelle(self):
        """Ana tabloyu gruplu şekilde güncelle"""
        # Mevcut satırları temizle
        for item in self.ana_tree.get_children():
            self.ana_tree.delete(item)

        if not self.gorunen_veriler:
            return

        # Eşdeğer gruplarına ayır
        esdeger_gruplari = {}
        for veri in self.gorunen_veriler:
            eid = veri.get('EsdegerId') or 0
            if eid not in esdeger_gruplari:
                esdeger_gruplari[eid] = []
            esdeger_gruplari[eid].append(veri)

        # Her grubun toplam parasal değerini hesapla (Oneri × DepocuFiyat)
        def grup_parasal_deger(urunler):
            return sum(u.get('Oneri', 0) * u.get('DepocuFiyat', 0) for u in urunler)

        # Grup içi sıralama: bireysel parasal değere göre (yüksekten düşüğe)
        def birey_parasal_deger(urun):
            return urun.get('Oneri', 0) * urun.get('DepocuFiyat', 0)

        for eid, urunler in esdeger_gruplari.items():
            urunler.sort(key=birey_parasal_deger, reverse=True)

        # Grupları parasal değere göre sırala (yüksekten düşüğe)
        # EsdegerId=0 olanlar (eşdeğersiz) ayrı tutulur ve bireysel sıralanır
        esdegerli_gruplar = [(eid, urunler) for eid, urunler in esdeger_gruplari.items() if eid != 0]
        esdegersiz_urunler = esdeger_gruplari.get(0, [])

        # Eşdeğerli grupları toplam parasal değere göre sırala
        esdegerli_gruplar.sort(key=lambda x: grup_parasal_deger(x[1]), reverse=True)

        # Eşdeğersiz ürünleri bireysel parasal değere göre sırala
        esdegersiz_urunler.sort(key=birey_parasal_deger, reverse=True)

        # Tüm grupları birleştir: önce eşdeğerliler (parasal değere göre), sonra eşdeğersizler (bireysel değere göre)
        sirali_gruplar = esdegerli_gruplar + [(0, esdegersiz_urunler)] if esdegersiz_urunler else esdegerli_gruplar

        for eid, urunler in sirali_gruplar:
            if eid == 0:
                # Eşdeğersiz - tek satırlar
                for urun in urunler:
                    self._tek_satir_ekle(urun)
            elif len(urunler) == 1:
                # Tek ürünlü grup - tek satır olarak göster
                self._tek_satir_ekle(urunler[0])
            else:
                # Çoklu eşdeğer grubu
                self._grup_ekle(eid, urunler)

    def _grup_ekle(self, esdeger_id, urunler):
        """Eşdeğer grubu ekle - başlık + satırlar + alt toplam"""
        # Grup özet bilgileri
        grup_stok = sum(u.get('Stok', 0) for u in urunler)
        grup_oneri = sum(u.get('Oneri', 0) for u in urunler)
        grup_aylik = sum(u.get('AylikOrt', 0) for u in urunler)
        grup_tutar = sum(u.get('Oneri', 0) * u.get('DepocuFiyat', 0) for u in urunler)

        # Grup ihtiyacı hesapla (stok + öneri = toplam ihtiyaç değil, öneri zaten ihtiyaç-stok)
        # Gerçek ihtiyaç = her ürünün (Stok + Oneri) toplamı değil
        # Daha doğrusu: İhtiyaç = Stok yetmezliği olan ürünlerin eksik miktarları toplamı
        # Basitleştirilmiş: grup_ihtiyac = grup_stok + grup_oneri (hedef stok seviyesi)
        # VEYA: AylikOrt * kalan ay sayısı

        # Hedef tarihine kadar toplam ihtiyaç (aylık ort * kalan süre)
        try:
            hedef_tarih = self.hedef_tarih.get_date()
            bugun = date.today()
            kalan_gun = max(0, (hedef_tarih - bugun).days)
        except:
            kalan_gun = 30  # Varsayılan 1 ay

        # Grup toplam ihtiyaç = Aylık ort * (kalan gün / 30)
        grup_ihtiyac = grup_aylik * (kalan_gun / 30) if grup_aylik > 0 else 0

        # Sübvansiyon analizi
        # Stok, ihtiyacın ne kadarını karşılıyor?
        if grup_ihtiyac > 0:
            subvansiyon_orani = grup_stok / grup_ihtiyac
        else:
            subvansiyon_orani = 999 if grup_stok > 0 else 0

        # Sübvansiyon etiketi - stok sütununa yazılacak
        subv_tag = "grup_baslik"
        if subvansiyon_orani >= 1.0:
            subv_stok_str = f"✓%{subvansiyon_orani*100:.0f}"
            subv_tag = "grup_baslik_yeterli"
        elif subvansiyon_orani >= 0.8:
            subv_stok_str = f"%{subvansiyon_orani*100:.0f}"
            subv_tag = "grup_baslik_iyi"
        elif subvansiyon_orani >= 0.5:
            subv_stok_str = f"%{subvansiyon_orani*100:.0f}"
            subv_tag = "grup_baslik_orta"
        elif grup_ihtiyac > 0:
            subv_stok_str = f"%{subvansiyon_orani*100:.0f}"
            subv_tag = "grup_baslik_dusuk"
        else:
            subv_stok_str = ""

        # Grup başlık satırı - sübvansiyon oranı STOK sütununda (index 2)
        tutar_str = f" ({grup_tutar:,.0f}₺)" if grup_tutar > 0 else ""
        # Sütun yapısı: [simge, ad, STOK, min, sart1, sart2, sart3, aylar..., aylik, gun, aybitis, oneri...]
        baslik_values = ['═', f"══ GRUP #{esdeger_id}{tutar_str} ══", subv_stok_str] + [''] * (len(self.aktif_sutunlar) - 3)
        self.ana_tree.insert('', 'end', values=baslik_values, tags=(subv_tag,))

        # Grup üyeleri
        for urun in urunler:
            values = self._satir_degerleri_olustur(urun, '├')
            tag = 'grup_satir_siparis' if urun.get('Oneri', 0) > 0 else 'grup_satir'
            self.ana_tree.insert('', 'end', iid=f"urun_{urun['UrunId']}", values=values, tags=(tag,))

        # Alt toplam satırı
        alt_values = ['└', f"─── TOPLAM ───", grup_stok, '', '', '', '']

        # Aylık toplamlar
        for col_id, _, _ in self.aylik_sutunlar:
            ay_toplam = sum(u.get(col_id, 0) for u in urunler)
            alt_values.append(ay_toplam)

        alt_values.extend([round(grup_aylik, 1), '', '', grup_oneri, '', '', '', ''])
        self.ana_tree.insert('', 'end', values=alt_values, tags=('alt_toplam',))

    def _tek_satir_ekle(self, urun):
        """Tek satır (eşdeğersiz) ekle"""
        values = self._satir_degerleri_olustur(urun, '●')
        tag = 'tek_satir_siparis' if urun.get('Oneri', 0) > 0 else 'tek_satir'
        self.ana_tree.insert('', 'end', iid=f"urun_{urun['UrunId']}", values=values, tags=(tag,))

    def _satir_degerleri_olustur(self, urun, tur_simge):
        """Satır değerlerini oluştur"""
        values = [
            tur_simge,
            urun.get('UrunAdi', ''),
            urun.get('Stok', 0),
            urun.get('MinStok', 0) or '',
            urun.get('Sart1', ''),
            urun.get('Sart2', ''),
            urun.get('Sart3', ''),
        ]

        # Aylık değerler
        for col_id, _, _ in self.aylik_sutunlar:
            values.append(urun.get(col_id, 0))

        # Öneri değerini vurgulu göster
        oneri_val = urun.get('Oneri', 0)
        oneri_gosterim = f"► {oneri_val}" if oneri_val > 0 else ""

        values.extend([
            urun.get('AylikOrt', 0),
            urun.get('GunlukOrt', 0),
            urun.get('AyBitis', ''),
            oneri_gosterim,
            urun.get('OneriAy', ''),
            urun.get('YeniAyBitis', ''),
            urun.get('Manuel', '') or '',
            urun.get('MF', ''),
        ])

        return values

    def _sorgu_hatasi(self, hata):
        self.status_label.config(text=f"Hata: {hata}")
        messagebox.showerror("Hata", hata)

    # ═══════════════════════════════════════════════════════════════════════
    # KULLANICI ETKİLEŞİMLERİ
    # ═══════════════════════════════════════════════════════════════════════

    def _toggle_yeterlileri_gizle(self):
        """Yeterlileri gizle toggle"""
        self.yeterlileri_gizle.set(not self.yeterlileri_gizle.get())

        if self.yeterlileri_gizle.get():
            self.gizle_btn.config(text="Yeterlileri Gizle: AÇIK", bg=self.R_SUCCESS)
        else:
            self.gizle_btn.config(text="Yeterlileri Gizle: KAPALI", bg=self.R_FG_SECONDARY)

        self._filtreleme_uygula()

    def _satir_secildi(self, event):
        """Satır seçildiğinde detay panelini güncelle"""
        selection = self.ana_tree.selection()
        if not selection:
            return

        item_id = selection[0]
        if item_id.startswith('urun_'):
            urun_id = int(item_id.replace('urun_', ''))
            urun = next((u for u in self.tum_veriler if u['UrunId'] == urun_id), None)
            if urun:
                self._detay_paneli_guncelle(urun)

    def _satir_cift_tiklandi(self, event):
        """Çift tıklama - öneriyi kesin sipariş listesine ekle"""
        selection = self.ana_tree.selection()
        if not selection:
            return

        item_id = selection[0]
        if item_id.startswith('urun_'):
            urun_id = int(item_id.replace('urun_', ''))
            urun = next((u for u in self.tum_veriler if u['UrunId'] == urun_id), None)
            if urun:
                oneri = urun.get('Oneri', 0)
                if oneri > 0:
                    self._kesin_listeye_ekle_hizli(urun, oneri, '')
                    self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:30]} → {oneri} adet kesin listeye eklendi")
                else:
                    self.status_label.config(text=f"⚠ {urun.get('UrunAdi', '')[:30]} için sipariş önerisi yok")

    def _sutun_tiklandi(self, event):
        """Sütun tıklaması - Oneri/MF sütunlarına tıklama işlemleri"""
        # Tıklanan bölgeyi tespit et
        region = self.ana_tree.identify_region(event.x, event.y)
        if region != 'cell':
            return

        # Tıklanan sütunu tespit et
        column = self.ana_tree.identify_column(event.x)
        column_id = self.ana_tree.column(column, 'id') if column else None

        # Tıklanan satırı tespit et
        item_id = self.ana_tree.identify_row(event.y)
        if not item_id or not item_id.startswith('urun_'):
            return

        urun_id = int(item_id.replace('urun_', ''))
        urun = next((u for u in self.tum_veriler if u['UrunId'] == urun_id), None)
        if not urun:
            return

        # Oneri sütununa tıklandıysa kesin listeye ekle
        if column_id == 'Oneri':
            oneri = urun.get('Oneri', 0)
            if oneri > 0:
                self._kesin_listeye_ekle_hizli(urun, oneri, urun.get('MF', ''))
                self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:30]} → {oneri} adet (öneri tıklandı)")
            return

        # MF Şartı sütununa tıklandıysa (Sart1, Sart2, Sart3)
        if column_id in ('Sart1', 'Sart2', 'Sart3'):
            sart_degeri = urun.get(column_id, '')
            if not sart_degeri or '+' not in sart_degeri:
                return

            mevcut_mf = urun.get('MF', '')

            if mevcut_mf == sart_degeri:
                # İkinci tıklama: Aynı MF zaten seçili → kesin siparişe ekle
                try:
                    parts = sart_degeri.split('+')
                    alinan = int(parts[0])
                except:
                    return
                self._kesin_listeye_ekle_hizli(urun, alinan, sart_degeri)
                self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:25]} → {sart_degeri} MF ile kesin siparişe eklendi")
            else:
                # İlk tıklama: MF'yi sipariş önerisine kaydet + detay kutucuklarını doldur
                urun['MF'] = sart_degeri
                self.manuel_mf_girisler[urun_id] = sart_degeri
                # Grid'deki MF sütununu güncelle
                self.ana_tree.set(item_id, 'MF', sart_degeri)
                # Detay paneldeki kutucukları doldur (Adet=alinan, MF=bedava)
                self._mf_kutucuklari_doldur(sart_degeri)
                self.status_label.config(text=f"📝 {urun.get('UrunAdi', '')[:25]} → MF: {sart_degeri} (tekrar tıkla → kesin sipariş)")
            return

    def _detay_paneli_guncelle(self, urun):
        """Detay panelini güncelle - çok kompakt tasarım"""
        self.secili_urun = urun

        # Placeholder'ı kaldır
        for widget in self.detay_content.winfo_children():
            widget.destroy()

        # ═══ ÜRÜN BİLGİLERİ (Tek satır) ═══
        info_frame = tk.LabelFrame(self.detay_content, text=urun.get('UrunAdi', '')[:40], font=('Arial', 10, 'bold'), padx=3, pady=2)
        info_frame.pack(fill=tk.X, pady=(0, 2))

        depocu_fiyat = urun.get('DepocuFiyat', 0)
        stok = urun.get('Stok', 0)
        min_stok = urun.get('MinStok', 0)
        aylik_ort = urun.get('AylikOrt', 0)
        oneri = urun.get('Oneri', 0)

        # Tek satırda tüm bilgiler
        row1 = tk.Frame(info_frame, bg='#E3F2FD')
        row1.pack(fill=tk.X)
        tk.Label(row1, text=f"S:{stok} M:{min_stok} A:{aylik_ort:.0f} F:{depocu_fiyat:.0f}₺", font=('Arial', 10), bg='#E3F2FD').pack(side=tk.LEFT, padx=2)
        tk.Button(row1, text=f"Öneri:{oneri}", font=('Arial', 10, 'bold'), bg='#1976D2', fg='white', cursor='hand2', relief='flat', padx=3,
                  command=lambda: self._oneri_kesin_listeye_ekle(urun)).pack(side=tk.LEFT, padx=2)
        tk.Button(row1, text="-", width=2, font=('Arial', 10, 'bold'), bg='#F44336', fg='white',
                  command=lambda: self._oneri_azalt_ve_ekle(urun)).pack(side=tk.LEFT, padx=1)
        tk.Button(row1, text="+", width=2, font=('Arial', 10, 'bold'), bg='#4CAF50', fg='white',
                  command=lambda: self._oneri_artir_ve_ekle(urun)).pack(side=tk.LEFT, padx=1)

        # Zam Optimizasyonu Bilgisi (varsa)
        zam_oneri = urun.get('ZamOneri')
        stok = urun.get('Stok', 0)
        aylik_ort = urun.get('AylikOrt', 1) or 1

        # Zam Optimizasyonu frame - her zaman göster
        zam_frame = tk.LabelFrame(self.detay_content, text="Zam Optimizasyonu", font=('Arial', 10, 'bold'), bg='#FFF3E0', padx=3, pady=2)
        zam_frame.pack(fill=tk.X, pady=2)

        if zam_oneri and (zam_oneri.get('optimum', 0) > 0 or zam_oneri.get('maksimum', 0) > 0):
            verimlilik = zam_oneri.get('verimlilik', 0)
            verimlilik_roi = zam_oneri.get('verimlilik_roi', 0)
            verimlilik_kazanc = zam_oneri.get('verimlilik_kazanc', 0)
            pareto = zam_oneri.get('pareto', 0)
            pareto_kazanc = zam_oneri.get('pareto_kazanc', 0)
            optimum = zam_oneri.get('optimum', 0)
            kazanc_optimum = zam_oneri.get('kazanc_optimum', 0)
            maksimum = zam_oneri.get('maksimum', 0)

            zam_bilgi = [
                ("ROI", verimlilik, verimlilik_kazanc, '#2E7D32'),
                ("Pareto", pareto, pareto_kazanc, '#F57C00'),
                ("Tepe", optimum, kazanc_optimum, '#1565C0'),
            ]

            for label, miktar, kazanc, color in zam_bilgi:
                if miktar > 0:
                    row = tk.Frame(zam_frame, bg='#FFF3E0')
                    row.pack(fill=tk.X, pady=0)
                    tk.Label(row, text=f"{label}:{miktar}ad={kazanc:.0f}TL", font=('Arial', 9), bg='#FFF3E0', fg=color).pack(side=tk.LEFT)
                    tk.Button(row, text="Seç", font=('Arial', 8), bg=color, fg='white', padx=2, pady=0,
                              command=lambda m=miktar, u=urun: self._zam_miktari_sec(u, m)).pack(side=tk.RIGHT)

        # Grafik Butonları - HER ZAMAN GÖRÜNÜR
        grafik_row = tk.Frame(zam_frame, bg='#FFF3E0')
        grafik_row.pack(fill=tk.X, pady=2)
        tk.Button(grafik_row, text="Karlılık", font=('Arial', 9), bg='#7B1FA2', fg='white',
                  command=lambda u=urun: self._tek_ilac_grafik_goster(u)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 1))
        tk.Button(grafik_row, text="ROI", font=('Arial', 9), bg='#00796B', fg='white',
                  command=lambda u=urun: self._roi_grafik_goster(u)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(1, 0))

        # MF Karlılık Analizi (Kompakt)
        mf_analiz_frame = tk.LabelFrame(self.detay_content, text="MF Analizi", font=('Arial', 10, 'bold'), bg='#E8F5E9', padx=3, pady=2)
        mf_analiz_frame.pack(fill=tk.X, pady=2)

        # MF'siz + Geçmiş MF aynı satırda
        mf_row1 = tk.Frame(mf_analiz_frame, bg='#E8F5E9')
        mf_row1.pack(fill=tk.X)

        self.mf_mfsiz_var = tk.BooleanVar(value=True)
        tk.Checkbutton(mf_row1, text="MF'siz", variable=self.mf_mfsiz_var, bg='#E8F5E9', font=('Arial', 9, 'bold'),
                      fg='#1565C0', activebackground='#E8F5E9').pack(side=tk.LEFT)

        # Geçmiş MF Şartları (Tıklanabilir Butonlar)
        sartlar = [urun.get('Sart1', ''), urun.get('Sart2', ''), urun.get('Sart3', '')]
        gecmis_sartlar = [s for s in sartlar if s]

        self.mf_gecmis_vars = []
        for sart in gecmis_sartlar:
            var = tk.BooleanVar(value=True)
            self.mf_gecmis_vars.append((var, sart))
            btn = tk.Button(mf_row1, text=sart, font=('Arial', 9, 'bold'), bg='#81C784', fg='#1B5E20', relief='raised', padx=4, cursor='hand2',
                           command=lambda s=sart, u=urun: self._mf_buton_tiklandi(u, s))
            btn.pack(side=tk.LEFT, padx=2)

        # Manuel MF + Sipariş miktarı aynı satırda
        mf_row2 = tk.Frame(mf_analiz_frame, bg='#E8F5E9')
        mf_row2.pack(fill=tk.X, pady=2)

        self.mf_manuel_entries = []
        for i in range(5):
            entry = ttk.Entry(mf_row2, width=6, font=('Arial', 10))
            entry.pack(side=tk.LEFT, padx=1)
            self.mf_manuel_entries.append(entry)

        tk.Label(mf_row2, text="Sip:", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(3, 0))
        self.mf_siparis_entry = ttk.Entry(mf_row2, width=4, font=('Arial', 10))
        self.mf_siparis_entry.pack(side=tk.LEFT, padx=1)
        varsayilan_siparis = urun.get('Oneri', 0) or urun.get('AylikOrt', 10)
        self.mf_siparis_entry.insert(0, str(int(varsayilan_siparis)))

        # Zam parametreleri satırı
        mf_zam_row = tk.Frame(mf_analiz_frame, bg='#E8F5E9')
        mf_zam_row.pack(fill=tk.X, pady=2)

        self.mf_zam_aktif = tk.BooleanVar(value=self.zam_aktif.get() if hasattr(self, 'zam_aktif') else False)
        tk.Checkbutton(mf_zam_row, text="Zam:", variable=self.mf_zam_aktif, bg='#E8F5E9',
                      font=('Arial', 9, 'bold'), fg='#E65100', activebackground='#E8F5E9').pack(side=tk.LEFT)

        tk.Label(mf_zam_row, text="%", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT)
        self.mf_zam_orani = ttk.Entry(mf_zam_row, width=4, font=('Arial', 10))
        self.mf_zam_orani.pack(side=tk.LEFT, padx=1)
        # Ana panelden varsayılan al
        try:
            varsayilan_zam = self.beklenen_zam_orani.get() if hasattr(self, 'beklenen_zam_orani') else '15'
        except:
            varsayilan_zam = '15'
        self.mf_zam_orani.insert(0, varsayilan_zam)

        tk.Label(mf_zam_row, text="Tarih:", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(5, 0))
        self.mf_zam_tarih = DateEntry(mf_zam_row, width=8, background='#E65100',
                                       foreground='white', borderwidth=2, font=('Arial', 9),
                                       date_pattern='dd.mm.yyyy', locale='tr_TR')
        self.mf_zam_tarih.pack(side=tk.LEFT, padx=1)
        # Ana panelden varsayılan tarih al
        try:
            if hasattr(self, 'zam_tarih_entry'):
                self.mf_zam_tarih.set_date(self.zam_tarih_entry.get_date())
        except:
            pass

        # Hesapla + Temizle butonları
        mf_row3 = tk.Frame(mf_analiz_frame, bg='#E8F5E9')
        mf_row3.pack(fill=tk.X, pady=2)

        tk.Button(mf_row3, text="HESAPLA", font=('Arial', 10, 'bold'), bg='#388E3C', fg='white', padx=8,
                  command=lambda u=urun: self._mf_karsilastirma_hesapla(u)).pack(side=tk.LEFT)
        tk.Button(mf_row3, text="Grafik", font=('Arial', 9), bg='#1565C0', fg='white',
                  command=lambda u=urun: self._mf_kombine_grafik(u)).pack(side=tk.LEFT, padx=3)
        tk.Button(mf_row3, text="Temizle", font=('Arial', 9), bg='#757575', fg='white',
                  command=self._mf_sonuc_temizle).pack(side=tk.LEFT, padx=3)

        # MF Sonuç frame
        self.mf_sonuc_frame = tk.Frame(mf_analiz_frame, bg='#E8F5E9')
        self.mf_sonuc_frame.pack(fill=tk.X, pady=2)

        # ═══ MANUEL SİPARİŞ (Tek satır) ═══
        manuel_frame = tk.LabelFrame(self.detay_content, text="Sipariş Girişi", font=('Arial', 10, 'bold'), bg='#ECEFF1', padx=3, pady=2)
        manuel_frame.pack(fill=tk.X, pady=2)

        giris_row = tk.Frame(manuel_frame, bg='#ECEFF1')
        giris_row.pack(fill=tk.X)

        tk.Label(giris_row, text="Adet:", font=('Arial', 10), bg='#ECEFF1').pack(side=tk.LEFT)
        self.manuel_entry = ttk.Entry(giris_row, width=5, font=('Arial', 10))
        self.manuel_entry.pack(side=tk.LEFT, padx=2)
        self.manuel_entry.insert(0, str(urun.get('Manuel', '') or ''))

        tk.Label(giris_row, text="MF:", font=('Arial', 10), bg='#ECEFF1').pack(side=tk.LEFT, padx=(5, 0))
        self.mf_entry = ttk.Entry(giris_row, width=7, font=('Arial', 10))
        self.mf_entry.pack(side=tk.LEFT, padx=2)
        self.mf_entry.insert(0, urun.get('MF', ''))

        # Butonlar aynı satırda
        tk.Button(giris_row, text="Al", command=lambda: self._oneriyi_al(urun),
                 bg='#7B1FA2', fg='white', font=('Arial', 9)).pack(side=tk.LEFT, padx=2)
        tk.Button(giris_row, text="Hes", command=lambda: self._mf_manuel_hesapla(urun),
                 bg='#00796B', fg='white', font=('Arial', 9)).pack(side=tk.LEFT, padx=1)
        tk.Button(giris_row, text="Kay", command=lambda: self._manuel_kaydet(urun),
                 bg='#1976D2', fg='white', font=('Arial', 9, 'bold')).pack(side=tk.LEFT, padx=1)

        # Kesin listeye ekle butonu (kompakt)
        tk.Button(self.detay_content, text="KESİN LİSTEYE EKLE",
                 command=lambda: self._kesin_listeye_ekle(urun),
                 bg='#388E3C', fg='white', font=('Arial', 11, 'bold'),
                 cursor='hand2', relief='raised', bd=2).pack(fill=tk.X, pady=3)

        # Scroll region güncelle
        self.detay_content.update_idletasks()
        self._detay_scroll_configure()

    def _manuel_artir(self, urun):
        """Manuel miktarı artır"""
        try:
            mevcut = int(self.manuel_entry.get() or 0)
        except:
            mevcut = 0
        self.manuel_entry.delete(0, tk.END)
        self.manuel_entry.insert(0, str(mevcut + 1))

    def _manuel_azalt(self, urun):
        """Manuel miktarı azalt"""
        try:
            mevcut = int(self.manuel_entry.get() or 0)
        except:
            mevcut = 0
        self.manuel_entry.delete(0, tk.END)
        self.manuel_entry.insert(0, str(max(0, mevcut - 1)))

    def _oneriyi_al(self, urun):
        """Öneriyi manuel alana kopyala"""
        oneri = urun.get('Oneri', 0)
        self.manuel_entry.delete(0, tk.END)
        self.manuel_entry.insert(0, str(oneri))

    def _zam_miktari_sec(self, urun, miktar):
        """Zam optimizasyonu karar noktasından miktar seç"""
        # NOT: miktar zaten optimal YENİ ALIM miktarıdır.
        # Algoritma stoğu hesaba katarak hesaplar, tekrar çıkarmıyoruz.
        self.manuel_entry.delete(0, tk.END)
        self.manuel_entry.insert(0, str(miktar))

    def _mf_sonuc_temizle(self):
        """MF sonuç alanını temizle"""
        for widget in self.mf_sonuc_frame.winfo_children():
            widget.destroy()

    def _mf_kutucuklari_doldur(self, mf_sart):
        """MF şartını parse edip detay paneldeki kutucukları doldur.
        Örn: '10+1' → Adet kutucuğuna 10, MF kutucuğuna 1 yazar.
        """
        if not mf_sart or '+' not in mf_sart:
            return
        try:
            parts = mf_sart.split('+')
            alinan = int(parts[0])
            bedava = int(parts[1])
        except (ValueError, IndexError):
            return

        # Sipariş Girişi kutucukları (Adet + MF)
        if hasattr(self, 'manuel_entry'):
            self.manuel_entry.delete(0, tk.END)
            self.manuel_entry.insert(0, str(alinan))
        if hasattr(self, 'mf_entry'):
            self.mf_entry.delete(0, tk.END)
            self.mf_entry.insert(0, str(bedava))

        # MF Analiz bölümündeki sipariş miktarı
        if hasattr(self, 'mf_siparis_entry'):
            self.mf_siparis_entry.delete(0, tk.END)
            self.mf_siparis_entry.insert(0, str(alinan))

    def _mf_buton_tiklandi(self, urun, mf_sart):
        """Detay paneldeki MF butonuna tıklama.
        İlk tık: kutucukları doldur (Adet=alinan, MF=bedava).
        İkinci tık (aynı MF seçiliyse): kesin siparişe ekle.
        """
        mevcut_mf = urun.get('MF', '')
        if mevcut_mf == mf_sart:
            # İkinci tıklama → kesin siparişe ekle
            self._mf_kesin_listeye_ekle(urun, mf_sart)
        else:
            # İlk tıklama → kutucukları doldur + MF'yi kaydet
            urun['MF'] = mf_sart
            urun_id = urun.get('UrunId')
            if urun_id:
                self.manuel_mf_girisler[urun_id] = mf_sart
                # Grid'deki MF sütununu güncelle
                item_id = f"urun_{urun_id}"
                try:
                    self.ana_tree.set(item_id, 'MF', mf_sart)
                except:
                    pass
            self._mf_kutucuklari_doldur(mf_sart)
            self.status_label.config(text=f"📝 {urun.get('UrunAdi', '')[:25]} → MF: {mf_sart} (tekrar tıkla → kesin sipariş)")

    def _oneri_kesin_listeye_ekle(self, urun):
        """Öneriye tıklayınca kesin siparişe ekle"""
        miktar = urun.get('Oneri', 0)
        if miktar > 0:
            self._kesin_listeye_ekle_hizli(urun, miktar, '')
            self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:25]} → {miktar} adet eklendi")

    def _oneri_artir_ve_ekle(self, urun):
        """Öneriyi +1 artırıp kesin siparişe ekle"""
        miktar = urun.get('Oneri', 0) + 1
        self._kesin_listeye_ekle_hizli(urun, miktar, '')
        self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:25]} → {miktar} adet eklendi (+1)")

    def _oneri_azalt_ve_ekle(self, urun):
        """Öneriyi -1 azaltıp kesin siparişe ekle"""
        miktar = max(0, urun.get('Oneri', 0) - 1)
        if miktar > 0:
            self._kesin_listeye_ekle_hizli(urun, miktar, '')
            self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:25]} → {miktar} adet eklendi (-1)")
        else:
            self.status_label.config(text="⚠ Miktar 0 olamaz")

    def _mf_kesin_listeye_ekle(self, urun, mf_sart):
        """MF şartına göre kesin siparişe ekle"""
        try:
            # MF parse: "5+1" → alinan=5
            parts = mf_sart.split('+')
            if len(parts) == 2:
                alinan = int(parts[0])
                self._kesin_listeye_ekle_hizli(urun, alinan, mf_sart)
                self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:25]} → {mf_sart} MF eklendi")
        except Exception as e:
            self.status_label.config(text=f"⚠ MF parse hatası: {mf_sart}")

    def _kesin_listeye_ekle_hizli(self, urun, miktar, mf=''):
        """Hızlı kesin listeye ekleme (tek tıklama için)"""
        if miktar <= 0:
            return

        # MF toplam hesapla
        toplam = miktar
        if mf and '+' in mf:
            try:
                parts = mf.split('+')
                mf_ek = int(parts[1])
                toplam = miktar + mf_ek
            except:
                pass

        siparis_data = {
            'UrunId': urun.get('UrunId'),
            'UrunAdi': urun.get('UrunAdi', ''),
            'Barkod': urun.get('Barkod', ''),
            'Miktar': miktar,
            'MF': mf,
            'Toplam': toplam,
            'Stok': urun.get('Stok', 0),
            'AylikOrt': urun.get('AylikOrt', 0),
        }

        # Mevcut listede aynı ürün var mı?
        for i, mevcut in enumerate(self.kesin_siparis_listesi):
            if mevcut.get('UrunId') == urun.get('UrunId'):
                # Güncelle ve kullanıcıyı bilgilendir
                siparis_data['db_id'] = mevcut.get('db_id')
                self.kesin_siparis_listesi[i] = siparis_data
                if hasattr(self, 'siparis_db') and self.siparis_db and self.aktif_calisma:
                    if mevcut.get('db_id'):
                        self.siparis_db.siparis_guncelle(mevcut['db_id'], miktar=miktar, mf=mf, toplam=toplam)
                self._kesin_liste_guncelle()
                urun_adi = urun.get('UrunAdi', '')[:25]
                self.status_label.config(text=f"⚠ {urun_adi} zaten listede - güncellendi ({miktar} ad. {mf})")
                return

        # Yeni ekleme
        if hasattr(self, 'siparis_db') and self.siparis_db and self.aktif_calisma:
            db_id = self.siparis_db.siparis_ekle(self.aktif_calisma, siparis_data)
            if db_id:
                siparis_data['db_id'] = db_id

        self.kesin_siparis_listesi.append(siparis_data)
        self._kesin_liste_guncelle()
        self._calisma_bilgisini_guncelle()

    def _mevcut_stok_fazlaligi_hesapla(self, mevcut_stok: int, aylik_ort: float,
                                        maliyet: float, faiz_yillik: float, depo_vade: int) -> dict:
        """
        Mevcut stoğun ay sonuna sarkma maliyetini hesapla.

        Mantık: Ay sonuna kadar ihtiyaçtan fazla olan stok, gereksiz yere erken alınmış demektir.
        Bu fazlalık için finansman maliyeti oluşur (SGK ödemesi 1 ay geç gelecek).

        Returns:
            {
                'fazla_adet': int,      # Ay sonuna sarkan adet
                'fazla_maliyet': float, # Finansman maliyeti (negatif değer)
                'ay_sonu_gun': int      # Ay sonuna kalan gün
            }
        """
        bugun = date.today()
        ay_son_gun = calendar.monthrange(bugun.year, bugun.month)[1]
        ay_sonuna_kalan = ay_son_gun - bugun.day

        if ay_sonuna_kalan <= 0 or aylik_ort <= 0:
            return {'fazla_adet': 0, 'fazla_maliyet': 0, 'ay_sonu_gun': 0}

        # Ay sonuna kadar ihtiyaç
        gunluk_sarf = aylik_ort / 30
        ay_sonu_ihtiyac = gunluk_sarf * ay_sonuna_kalan

        # Fazla stok (sarkan miktar)
        fazla_adet = max(0, mevcut_stok - ay_sonu_ihtiyac)

        if fazla_adet <= 0:
            return {'fazla_adet': 0, 'fazla_maliyet': 0, 'ay_sonu_gun': ay_sonuna_kalan}

        # Bu fazla stok için 1 aylık finansman maliyeti
        # (SGK ödemesi 30 gün geç gelecek, depo ödemesi zamanında yapılacak)
        aylik_faiz = (faiz_yillik / 100) / 12
        fazla_maliyet = fazla_adet * maliyet * aylik_faiz

        return {
            'fazla_adet': int(fazla_adet),
            'fazla_maliyet': round(fazla_maliyet, 2),
            'ay_sonu_gun': ay_sonuna_kalan
        }

    def _npv_mf_hesapla(self, alinan: int, bedava: int, maliyet: float, aylik_ort: float,
                        mevcut_stok: int, faiz_yillik: float, depo_vade: int,
                        zam_gun: int = 0, zam_orani: float = 0) -> dict:
        """
        MF alımı için detaylı NPV hesabı.

        Karşılaştırma: Toplu MF'li alım vs ay be ay alım

        Returns:
            {
                'npv_toplu': float,     # Toplu alım NPV (ödenen)
                'npv_aylik': float,     # Ay be ay alım NPV (ödenmesi gereken)
                'kazanc': float,        # npv_aylik - npv_toplu (pozitif = karlı)
                'stok_ay': float,       # Kaç aylık stok
                'zam_kazanc': float     # Zam farkından kazanç
            }
        """
        if alinan <= 0 or maliyet <= 0 or aylik_ort <= 0:
            return {'npv_toplu': 0, 'npv_aylik': 0, 'kazanc': 0, 'stok_ay': 0, 'zam_kazanc': 0, 'mf_kazanc': 0, 'stok_maliyet': 0}

        toplam_gelen = alinan + bedava
        toplam_stok = mevcut_stok + toplam_gelen

        # Gün bazlı faiz hesabı
        aylik_faiz = (faiz_yillik / 100) / 12
        gunluk_faiz = (1 + aylik_faiz) ** (1/30) - 1
        gunluk_sarf = aylik_ort / 30

        # Kaç aylık stok
        stok_ay = toplam_stok / aylik_ort if aylik_ort > 0 else 0

        # ===== SENARYO A: AY BE AY ALIM (MF'siz, ihtiyaç kadar) =====
        # Her gün ihtiyaç kadar alınsa ne ödenirdi?
        npv_aylik = 0
        kalan_mevcut = mevcut_stok
        kalan_yeni = toplam_gelen  # MF'li alımla gelen toplam

        gun = 0
        while kalan_yeni > 0 and gun < 720:  # Max 2 yıl
            harcanan = gunluk_sarf

            # Önce mevcut stoktan harca
            mevcut_harcanan = min(kalan_mevcut, harcanan)
            kalan_mevcut -= mevcut_harcanan

            # Kalan ihtiyacı yeni alımdan karşıla
            yeni_harcanan = min(harcanan - mevcut_harcanan, kalan_yeni)

            if yeni_harcanan > 0:
                # Zam kontrolü
                if zam_orani > 0 and gun >= zam_gun:
                    fiyat = maliyet * (1 + zam_orani / 100)
                else:
                    fiyat = maliyet

                # O gün alınsa, ay sonunda senet kesilir, depo vadesi sonra ödenir
                # Grafik ile tutarlılık için ay sonu hesabı
                ay_sonu = ((gun // 30) + 1) * 30
                odeme_gun = ay_sonu + depo_vade
                odeme = yeni_harcanan * fiyat
                iskonto = (1 + gunluk_faiz) ** odeme_gun
                npv_aylik += odeme / iskonto

                kalan_yeni -= yeni_harcanan

            gun += 1

        # ===== SENARYO B: TOPLU ALIM (MF'li) =====
        # Bugün toplu alım, ay sonunda senet, depo vadesi sonra ödeme
        # Not: Grafik ile tutarlılık için 30 + depo_vade kullanılıyor
        npv_toplu = (alinan * maliyet) / ((1 + gunluk_faiz) ** (30 + depo_vade))

        # ===== MF KAZANCI =====
        # MF avantajı = bedava ürünlerin değeri (bugüne indirgenmiş)
        mf_kazanc = 0
        if bedava > 0:
            # Bedava ürünler ortalama satış gününde satılacak
            ort_satis_gun = (toplam_stok / gunluk_sarf / 2) if gunluk_sarf > 0 else 30
            iskonto_mf = (1 + gunluk_faiz) ** ort_satis_gun
            mf_kazanc = (bedava * maliyet) / iskonto_mf

        # ===== ZAM KAZANCI =====
        zam_kazanc = 0
        if zam_orani > 0 and zam_gun > 0:
            # Zam sonrası satılacak miktarın zam farkı
            zam_oncesi_tuketim = min(toplam_stok, zam_gun * gunluk_sarf)
            zam_sonrasi_miktar = max(0, toplam_stok - zam_oncesi_tuketim)

            if zam_sonrasi_miktar > 0:
                zam_fark = maliyet * (zam_orani / 100)
                ort_satis_gun = zam_gun + (zam_sonrasi_miktar / gunluk_sarf / 2) if gunluk_sarf > 0 else zam_gun
                iskonto = (1 + gunluk_faiz) ** ort_satis_gun
                zam_kazanc = (zam_sonrasi_miktar * zam_fark) / iskonto

        # ===== STOK FİNANSMAN MALİYETİ =====
        # Stok maliyeti = Erken ödeme nedeniyle paranın bağlı kalma maliyeti
        # Ay sonu ihtiyacının ötesinde stok tutmanın maliyeti
        # Hesaplama: Toplam ödeme × (ortalama stok gün / 365) × faiz
        ay_sonu_ihtiyac = 30 * gunluk_sarf  # Bu ay satılacak miktar
        fazla_stok = max(0, toplam_gelen - ay_sonu_ihtiyac)  # Ay sonuna sarkacak stok

        stok_maliyet = 0
        if fazla_stok > 0:
            # Fazla stoğun ortalama tutulma süresi (gün)
            # Fazla stok, ay sonundan itibaren tüketilecek
            fazla_stok_gun = (fazla_stok / gunluk_sarf / 2) if gunluk_sarf > 0 else 30
            # Finansman maliyeti = Stok değeri × gün × günlük faiz
            stok_maliyet = fazla_stok * maliyet * fazla_stok_gun * gunluk_faiz

        # ===== NET KAZANÇ =====
        # Net = MF Avantajı + Zam Avantajı - Stok Maliyeti
        kazanc = mf_kazanc + zam_kazanc - stok_maliyet

        return {
            'npv_toplu': round(npv_toplu, 2),
            'npv_aylik': round(npv_aylik, 2),
            'kazanc': round(kazanc, 2),
            'stok_ay': round(stok_ay, 1),
            'zam_kazanc': round(zam_kazanc, 2),
            'mf_kazanc': round(mf_kazanc, 2),
            'stok_maliyet': round(stok_maliyet, 2)  # Her zaman pozitif (maliyet)
        }

    def _bolge_analizi_hesapla(self, miktar: int, urun: dict) -> dict:
        """
        Verilen miktar için hangi bölgede olduğunu hesapla.

        Bölgeler:
        1. Sıfır - Maks ROI arası (Verimli Başlangıç)
        2. Maks ROI - Eğim Azalması arası (Azalan Verim)
        3. Eğim Azalması - Pareto arası (Dengeli)
        4. Pareto - Tepe arası (Agresif)
        5. Tepe - Negatif Dönüş arası (Riskli)
        6. Negatif Dönüş sonrası (Zarar)

        Returns:
            dict: bolge_no, bolge_adi, renk, aciklama
        """
        try:
            faiz = self._aktif_faiz_getir()
            depo_vade = self.depo_vadesi.get()
            zam_aktif = self.zam_aktif.get()
            zam_orani = float(self.beklenen_zam_orani.get()) if zam_aktif else 0
            zam_tarihi = self.zam_tarih_entry.get_date() if zam_aktif else None
        except:
            return {'bolge_no': 0, 'bolge_adi': '?', 'renk': '#757575', 'aciklama': 'Hesaplanamadı'}

        if not zam_aktif or zam_orani <= 0:
            return {'bolge_no': 0, 'bolge_adi': '-', 'renk': '#757575', 'aciklama': 'Zam analizi kapalı'}

        # ROI verisini hesapla
        sonuc = self._roi_verisi_hesapla(
            maliyet=urun.get('DepocuFiyat', 0),
            aylik_ort=urun.get('AylikOrt', 0),
            mevcut_stok=urun.get('Stok', 0),
            zam_tarihi=zam_tarihi,
            zam_orani=zam_orani,
            faiz_yillik=faiz,
            depo_vade=depo_vade
        )

        if not sonuc:
            return {'bolge_no': 0, 'bolge_adi': '?', 'renk': '#757575', 'aciklama': 'Hesaplanamadı'}

        kn = sonuc['kritik']
        max_roi_m = kn['max_roi']['m']
        yarim_roi_m = kn['yarim_roi']['m']  # Eğim azalması noktası
        pareto_m = kn['pareto']['m']
        optimum_m = kn['optimum']['m']  # Tepe noktası
        sifir_roi_m = kn['sifir_roi']['m']  # Negatife dönüş

        # Bölge belirleme
        if miktar <= 0:
            return {'bolge_no': 0, 'bolge_adi': '-', 'renk': '#757575', 'aciklama': 'Miktar yok'}
        elif miktar <= max_roi_m:
            return {'bolge_no': 1, 'bolge_adi': 'B1', 'renk': '#4CAF50',
                    'aciklama': f'Verimli (0-{max_roi_m})'}
        elif miktar <= yarim_roi_m:
            return {'bolge_no': 2, 'bolge_adi': 'B2', 'renk': '#8BC34A',
                    'aciklama': f'Azalan Verim ({max_roi_m}-{yarim_roi_m})'}
        elif miktar <= pareto_m:
            return {'bolge_no': 3, 'bolge_adi': 'B3', 'renk': '#FFC107',
                    'aciklama': f'Dengeli ({yarim_roi_m}-{pareto_m})'}
        elif miktar <= optimum_m:
            return {'bolge_no': 4, 'bolge_adi': 'B4', 'renk': '#FF9800',
                    'aciklama': f'Agresif ({pareto_m}-{optimum_m})'}
        elif miktar <= sifir_roi_m:
            return {'bolge_no': 5, 'bolge_adi': 'B5', 'renk': '#FF5722',
                    'aciklama': f'Riskli ({optimum_m}-{sifir_roi_m})'}
        else:
            return {'bolge_no': 6, 'bolge_adi': 'B6', 'renk': '#F44336',
                    'aciklama': f'Zarar ({sifir_roi_m}+)'}

    def _mf_karsilastirma_hesapla(self, urun):
        """
        MF karlılık analizi - NPV bazlı detaylı hesaplama.

        Prensipler:
        1. Ay be ay alım vs toplu MF'li alım karşılaştırması (NPV)
        2. Mevcut stok fazlalığı her satırın net karından düşülür
        3. Zam varsa, zam kazancı da hesaba katılır
        """

        # Sipariş miktarını al
        try:
            siparis_miktar = int(self.mf_siparis_entry.get().strip() or 0)
        except:
            siparis_miktar = 0

        if siparis_miktar <= 0:
            messagebox.showwarning("Uyarı", "Sipariş miktarı girin!")
            return

        # MF listesi oluştur
        mf_listesi = []

        # 1. MF'siz alım (referans)
        mfsiz_dahil = hasattr(self, 'mf_mfsiz_var') and self.mf_mfsiz_var.get()

        # 2. Geçmiş MF'lerden seçilenleri al
        if hasattr(self, 'mf_gecmis_vars'):
            for var, sart in self.mf_gecmis_vars:
                if var.get():
                    mf_listesi.append(sart)

        # 3. Manuel girişlerden al
        if hasattr(self, 'mf_manuel_entries'):
            for entry in self.mf_manuel_entries:
                mf_text = entry.get().strip()
                if mf_text and '+' in mf_text:
                    if mf_text not in mf_listesi:
                        mf_listesi.append(mf_text)

        if not mf_listesi and not mfsiz_dahil:
            messagebox.showwarning("Uyarı", "En az bir seçenek işaretleyin veya MF şartı girin!")
            return

        # Parametreleri al - MF bölümündeki zam ayarlarından
        try:
            faiz = self._aktif_faiz_getir()
            depo_vade = self.depo_vadesi.get()
            # MF bölümündeki zam parametrelerini kullan
            zam_aktif = hasattr(self, 'mf_zam_aktif') and self.mf_zam_aktif.get()
            zam_orani = float(self.mf_zam_orani.get()) if zam_aktif and hasattr(self, 'mf_zam_orani') else 0
            zam_tarihi = self.mf_zam_tarih.get_date() if zam_aktif and hasattr(self, 'mf_zam_tarih') else None
        except:
            faiz, depo_vade, zam_orani, zam_tarihi = 45, 75, 0, None

        aylik_ort = urun.get('AylikOrt', 0)
        stok = urun.get('Stok', 0)
        maliyet = urun.get('DepocuFiyat', 0)

        if maliyet <= 0:
            messagebox.showwarning("Uyarı", "Ürün fiyat bilgisi bulunamadı!")
            return

        if aylik_ort <= 0:
            messagebox.showwarning("Uyarı", "Aylık ortalama satış bilgisi yok!")
            return

        # Zam gün hesabı
        zam_gun = 0
        if zam_tarihi and zam_orani > 0:
            bugun = date.today()
            zam_gun = max(0, (zam_tarihi - bugun).days)

        # Mevcut stok fazlalığı hesapla
        stok_fazlaligi = self._mevcut_stok_fazlaligi_hesapla(
            mevcut_stok=stok,
            aylik_ort=aylik_ort,
            maliyet=maliyet,
            faiz_yillik=faiz,
            depo_vade=depo_vade
        )
        fazla_adet = stok_fazlaligi['fazla_adet']
        fazla_maliyet = stok_fazlaligi['fazla_maliyet']

        # Sonuç frame'i temizle
        for widget in self.mf_sonuc_frame.winfo_children():
            widget.destroy()

        # Bilgi satırı
        bilgi_row = tk.Frame(self.mf_sonuc_frame, bg='#E8F5E9')
        bilgi_row.pack(fill=tk.X, pady=(0, 3))
        bilgi_text = f"Stok: {stok} | Sipariş: {siparis_miktar} | Aylık: {aylik_ort:.0f} | Faiz: %{faiz:.0f}"
        if zam_gun > 0:
            bilgi_text += f" | Zam: {zam_gun}g %{zam_orani:.0f}"
        tk.Label(bilgi_row, text=bilgi_text, font=('Arial', 9), bg='#E8F5E9', fg='#555').pack(anchor='w')

        # Mevcut stok fazlalığı uyarısı (varsa)
        if fazla_adet > 0:
            fazla_row = tk.Frame(self.mf_sonuc_frame, bg='#FFCDD2')
            fazla_row.pack(fill=tk.X, pady=(0, 3))
            fazla_text = f"⚠ Mevcut Stok Fazlası: {fazla_adet} ad. → -{fazla_maliyet:.0f}₺ (her satırdan düşülür)"
            tk.Label(fazla_row, text=fazla_text, font=('Arial', 9, 'bold'),
                    bg='#FFCDD2', fg='#C62828').pack(anchor='w', padx=2)

        # Formül açıklama satırı
        formul_row = tk.Frame(self.mf_sonuc_frame, bg='#FFF9C4')
        formul_row.pack(fill=tk.X, pady=(0, 2))
        formul_text = "Net = MF Av.(+) + Zam Av.(+) - Stok Mal.(-)"
        tk.Label(formul_row, text=formul_text, font=('Arial', 8, 'italic'),
                bg='#FFF9C4', fg='#5D4037').pack(anchor='w', padx=2)

        # Başlık satırı
        baslik_row = tk.Frame(self.mf_sonuc_frame, bg='#C8E6C9')
        baslik_row.pack(fill=tk.X)
        tk.Label(baslik_row, text="MF", font=('Arial', 8, 'bold'), width=6, bg='#C8E6C9', anchor='w').pack(side=tk.LEFT, padx=1)
        tk.Label(baslik_row, text="Al", font=('Arial', 8, 'bold'), width=3, bg='#C8E6C9').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="Top", font=('Arial', 8, 'bold'), width=3, bg='#C8E6C9').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="+MF", font=('Arial', 8, 'bold'), width=5, bg='#C8E6C9', fg='#1B5E20').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="+Zam", font=('Arial', 8, 'bold'), width=5, bg='#C8E6C9', fg='#E65100').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="-Stok", font=('Arial', 8, 'bold'), width=5, bg='#C8E6C9', fg='#C62828').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="=Net", font=('Arial', 8, 'bold'), width=5, bg='#C8E6C9').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="ROI", font=('Arial', 8, 'bold'), width=4, bg='#C8E6C9').pack(side=tk.LEFT)
        tk.Label(baslik_row, text="Ay", font=('Arial', 8, 'bold'), width=3, bg='#C8E6C9').pack(side=tk.LEFT)

        sonuclar = []

        # MF'siz alım hesabı (referans)
        if mfsiz_dahil:
            npv_sonuc = self._npv_mf_hesapla(
                alinan=siparis_miktar, bedava=0, maliyet=maliyet, aylik_ort=aylik_ort,
                mevcut_stok=stok, faiz_yillik=faiz, depo_vade=depo_vade,
                zam_gun=zam_gun, zam_orani=zam_orani
            )

            # Net kar = NPV kazancı - mevcut stok fazlalığı maliyeti
            net_kar = npv_sonuc['kazanc'] - fazla_maliyet
            yatirim = siparis_miktar * maliyet
            roi = (net_kar / yatirim * 100) if yatirim > 0 else 0

            sonuclar.append({
                'mf': "MF'siz", 'al': siparis_miktar, 'bedava': 0,
                'toplam': siparis_miktar, 'stok_ay': npv_sonuc['stok_ay'],
                'kar': net_kar, 'roi': roi,
                'zam_kazanc': npv_sonuc['zam_kazanc'],
                'mf_kazanc': npv_sonuc.get('mf_kazanc', 0),
                'stok_maliyet': npv_sonuc.get('stok_maliyet', 0),
                'is_referans': True
            })

        # MF'li alımlar
        for mf_sart in mf_listesi:
            try:
                parcalar = mf_sart.split('+')
                mf_al = int(parcalar[0])
                mf_bedava = int(parcalar[1])

                # Sipariş miktarına göre kaç set MF alınacak
                set_sayisi = siparis_miktar // mf_al
                if set_sayisi == 0:
                    set_sayisi = 1  # En az 1 set

                al = set_sayisi * mf_al
                bedava = set_sayisi * mf_bedava
                toplam = al + bedava

                npv_sonuc = self._npv_mf_hesapla(
                    alinan=al, bedava=bedava, maliyet=maliyet, aylik_ort=aylik_ort,
                    mevcut_stok=stok, faiz_yillik=faiz, depo_vade=depo_vade,
                    zam_gun=zam_gun, zam_orani=zam_orani
                )

                # Net kar = NPV kazancı - mevcut stok fazlalığı maliyeti
                net_kar = npv_sonuc['kazanc'] - fazla_maliyet
                yatirim = al * maliyet
                roi = (net_kar / yatirim * 100) if yatirim > 0 else 0

                sonuclar.append({
                    'mf': mf_sart, 'al': al, 'bedava': bedava,
                    'toplam': toplam, 'stok_ay': npv_sonuc['stok_ay'],
                    'kar': net_kar, 'roi': roi,
                    'zam_kazanc': npv_sonuc['zam_kazanc'],
                    'mf_kazanc': npv_sonuc.get('mf_kazanc', 0),
                    'stok_maliyet': npv_sonuc.get('stok_maliyet', 0),
                    'is_referans': False
                })

            except Exception as e:
                continue

        # Sonuçları sırala (ROI'ye göre)
        sonuclar.sort(key=lambda x: x['roi'], reverse=True)

        # Sonuçları göster
        for i, s in enumerate(sonuclar):
            if s['is_referans']:
                renk = '#BBDEFB'  # Mavi (referans)
            elif s['roi'] > 0:
                renk = '#A5D6A7'  # Yeşil (karlı)
            else:
                renk = '#FFCDD2'  # Kırmızı (zararlı)

            row = tk.Frame(self.mf_sonuc_frame, bg=renk)
            row.pack(fill=tk.X)

            # MF Şartı (tıklanabilir - kesin siparişe ekle)
            mf_text = s['mf']
            if i == 0:
                mf_text += "★"

            mf_btn = tk.Button(row, text=mf_text, font=('Arial', 8, 'bold' if i == 0 else 'normal'),
                              width=6, bg=renk, anchor='w', relief='flat', cursor='hand2',
                              command=lambda al=s['al'], mf=s['mf'], u=urun: self._mf_sonuc_kesin_ekle(u, al, mf))
            mf_btn.pack(side=tk.LEFT, padx=1)

            # Al
            tk.Label(row, text=str(s['al']), font=('Arial', 8), width=3, bg=renk).pack(side=tk.LEFT)

            # Toplam
            tk.Label(row, text=str(s['toplam']), font=('Arial', 8), width=3, bg=renk).pack(side=tk.LEFT)

            # MF Avantajı
            mf_av = s.get('mf_kazanc', 0)
            tk.Label(row, text=f"{mf_av:.0f}", font=('Arial', 8),
                    width=5, bg=renk, fg='#1B5E20').pack(side=tk.LEFT)

            # Zam Avantajı
            zam_av = s.get('zam_kazanc', 0)
            tk.Label(row, text=f"{zam_av:.0f}", font=('Arial', 8),
                    width=5, bg=renk, fg='#E65100').pack(side=tk.LEFT)

            # Stok Maliyeti (her zaman maliyet, kırmızı renk)
            stok_m = s.get('stok_maliyet', 0)
            stok_text = f"{stok_m:.0f}" if stok_m > 0 else "0"
            tk.Label(row, text=stok_text, font=('Arial', 8),
                    width=5, bg=renk, fg='#C62828').pack(side=tk.LEFT)

            # Net Kar
            kar_renk = 'green' if s['kar'] > 0 else ('red' if s['kar'] < 0 else 'black')
            tk.Label(row, text=f"{s['kar']:.0f}", font=('Arial', 8, 'bold'),
                    width=5, bg=renk, fg=kar_renk).pack(side=tk.LEFT)

            # ROI %
            tk.Label(row, text=f"%{s['roi']:.0f}", font=('Arial', 8, 'bold'),
                    width=4, bg=renk, fg=kar_renk).pack(side=tk.LEFT)

            # MF için stok ay sayısına göre risk göstergesi
            stok_ay = s.get('stok_ay', 0)
            if stok_ay <= 1:
                risk_text, risk_renk, risk_aciklama = '1ay', '#4CAF50', f'{stok_ay:.1f} aylık stok - Güvenli'
            elif stok_ay <= 2:
                risk_text, risk_renk, risk_aciklama = '2ay', '#8BC34A', f'{stok_ay:.1f} aylık stok - Normal'
            elif stok_ay <= 3:
                risk_text, risk_renk, risk_aciklama = '3ay', '#FFC107', f'{stok_ay:.1f} aylık stok - Dikkat'
            elif stok_ay <= 4:
                risk_text, risk_renk, risk_aciklama = '4ay', '#FF9800', f'{stok_ay:.1f} aylık stok - Riskli'
            elif stok_ay <= 6:
                risk_text, risk_renk, risk_aciklama = '5-6', '#FF5722', f'{stok_ay:.1f} aylık stok - Yüksek Risk'
            else:
                risk_text, risk_renk, risk_aciklama = '6+', '#F44336', f'{stok_ay:.1f} aylık stok - Çok Riskli'

            risk_label = tk.Label(row, text=risk_text, font=('Arial', 8, 'bold'),
                                   width=3, bg=risk_renk, fg='white')
            risk_label.pack(side=tk.LEFT, padx=1)

            # Tooltip için risk açıklaması
            def show_tooltip(event, aciklama=risk_aciklama):
                tooltip = tk.Toplevel()
                tooltip.wm_overrideredirect(True)
                tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
                tk.Label(tooltip, text=aciklama, bg='#FFFFE0', relief='solid', bd=1,
                        font=('Arial', 8)).pack()
                tooltip.after(2000, tooltip.destroy)
            risk_label.bind('<Enter>', show_tooltip)

        # Scroll region güncelle (yeni içerik eklendi)
        self.detay_content.update_idletasks()
        self._detay_scroll_configure()

    def _mf_sonuc_kesin_ekle(self, urun, miktar, mf_sart):
        """MF analiz sonucunu kesin siparişe ekle"""
        if miktar <= 0:
            return

        mf = mf_sart if mf_sart != "MF'siz" else ''
        self._kesin_listeye_ekle_hizli(urun, miktar, mf)
        self.status_label.config(text=f"✓ {urun.get('UrunAdi', '')[:25]} → {miktar} ad. ({mf_sart}) kesin listeye eklendi")

    def _mf_kombine_grafik(self, urun):
        """MF + Zam kombine grafiği - Zam eğrisi üzerine MF noktaları"""
        import tempfile
        import os
        import traceback

        try:
            # Parametreleri al - MF bölümündeki zam ayarlarından
            faiz = self._aktif_faiz_getir()
            depo_vade = self.depo_vadesi.get()
            zam_aktif = hasattr(self, 'mf_zam_aktif') and self.mf_zam_aktif.get()
            zam_orani = float(self.mf_zam_orani.get()) if zam_aktif and hasattr(self, 'mf_zam_orani') else 0
            zam_tarihi = self.mf_zam_tarih.get_date() if zam_aktif and hasattr(self, 'mf_zam_tarih') else None

            if not zam_aktif or zam_orani <= 0:
                messagebox.showwarning("Uyarı", "MF kombine grafiği için zam analizi aktif olmalı.\nMF Analiz bölümünde Zam oranı ve tarihi girin.")
                return

            maliyet = urun.get('DepocuFiyat', 0)
            aylik_ort = urun.get('AylikOrt', 0)
            stok = urun.get('Stok', 0)

            if maliyet <= 0 or aylik_ort <= 0:
                messagebox.showwarning("Uyarı", "Depocu fiyat veya aylık ortalama eksik.")
                return

            # Zam eğrisi verilerini hesapla
            sonuc = self._karlilik_verisi_hesapla(maliyet, aylik_ort, stok, zam_tarihi, zam_orani, faiz, depo_vade)
            if not sonuc:
                messagebox.showwarning("Uyarı", "Karlılık hesaplanamadı.")
                return

            zam_gun = (zam_tarihi - date.today()).days

            # MF listesini oluştur
            mf_listesi = []
            if hasattr(self, 'mf_mfsiz_var') and self.mf_mfsiz_var.get():
                mf_listesi.append("MF'siz")

            # Standart MF baremleri
            for btn_text in ['5+1', '10+2', '10+3', '20+5', '20+7', '50+15', '50+20', '100+30', '100+50']:
                if btn_text not in mf_listesi:
                    mf_listesi.append(btn_text)

            # Manuel MF
            if hasattr(self, 'mf_manuel_entries'):
                for entry in self.mf_manuel_entries:
                    mf = entry.get().strip()
                    if mf and '+' in mf and mf not in mf_listesi:
                        mf_listesi.append(mf)

            # Sipariş miktarı
            try:
                siparis_miktar = int(self.mf_siparis_entry.get()) if hasattr(self, 'mf_siparis_entry') else max(1, int(aylik_ort))
            except:
                siparis_miktar = max(1, int(aylik_ort))

            # MF noktalarını hesapla
            mf_noktalar = []
            for mf_sart in mf_listesi:
                try:
                    if mf_sart == "MF'siz":
                        al, bedava, toplam = siparis_miktar, 0, siparis_miktar
                    else:
                        parcalar = mf_sart.split('+')
                        mf_al = int(parcalar[0])
                        mf_bedava = int(parcalar[1])
                        set_sayisi = max(1, siparis_miktar // mf_al)
                        al = set_sayisi * mf_al
                        bedava = set_sayisi * mf_bedava
                        toplam = al + bedava

                    # NPV hesapla
                    npv_sonuc = self._npv_mf_hesapla(
                        alinan=al, bedava=bedava, maliyet=maliyet, aylik_ort=aylik_ort,
                        mevcut_stok=stok, faiz_yillik=faiz, depo_vade=depo_vade,
                        zam_gun=zam_gun, zam_orani=zam_orani
                    )

                    yatirim = al * maliyet
                    roi = (npv_sonuc['kazanc'] / yatirim * 100) if yatirim > 0 else 0

                    mf_noktalar.append({
                        'mf': mf_sart, 'toplam': toplam, 'al': al,
                        'kar': npv_sonuc['kazanc'], 'roi': roi
                    })
                except:
                    continue

            if not mf_noktalar:
                messagebox.showwarning("Uyarı", "Hesaplanacak MF noktası bulunamadı.")
                return

            # Grafik çiz
            m, k, kn = sonuc['miktarlar'], sonuc['kazanclar'], sonuc['kritik']

            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10), height_ratios=[2, 1])
            fig.suptitle(f"MF + Zam Kombine Analiz - {urun['UrunAdi'][:50]}", fontsize=12, fontweight='bold', x=0.99, ha='right', y=0.98)

            # Dinamik ölçeklendirme
            x_max = int(kn['negatif']['m'] * 1.2)
            x_max = max(x_max, max([n['toplam'] for n in mf_noktalar]) + 10)
            x_max = min(x_max, len(m) - 1)
            y_max = kn['optimum']['k'] * 1.3
            y_min = min(k[:x_max+1]) if x_max < len(k) else min(k)
            if y_min > -y_max * 0.1:
                y_min = -y_max * 0.15

            m_lim, k_lim = m[:x_max+1], k[:x_max+1]

            # Zam eğrisi
            ax1.fill_between(m_lim, k_lim, alpha=0.2, color='green', where=[x >= 0 for x in k_lim])
            ax1.fill_between(m_lim, k_lim, alpha=0.2, color='red', where=[x < 0 for x in k_lim])
            ax1.plot(m_lim, k_lim, 'b-', linewidth=2, alpha=0.7, label='Zam Eğrisi')
            ax1.axhline(y=0, color='black', linestyle='-', linewidth=1)

            # Kritik noktalar
            colors_kn = {'max_roi': '#FF9800', 'azalan': '#9C27B0', 'pareto': '#4CAF50', 'optimum': '#2196F3', 'negatif': '#F44336'}
            for n, v in kn.items():
                if v['m'] <= x_max:
                    ax1.axvline(x=v['m'], color=colors_kn[n], linestyle=':', linewidth=1, alpha=0.5)
                    ax1.scatter(v['m'], v['k'], c=colors_kn[n], s=80, marker='o', zorder=4, alpha=0.6)

            # MF noktaları
            mf_colors = plt.cm.Set1(np.linspace(0, 1, len(mf_noktalar)))
            for i, nokta in enumerate(mf_noktalar):
                toplam = nokta['toplam']
                mf_kar = nokta['kar']

                ax1.scatter(toplam, mf_kar, c=[mf_colors[i]], s=250, marker='s', zorder=6,
                           edgecolors='black', linewidths=2)

                y_off = y_max * 0.08 * (1 if i % 2 == 0 else -1)
                x_off = x_max * 0.02 * (1 if i % 3 == 0 else -1)
                ax1.annotate(f"{nokta['mf']}\n{toplam}ad\n{mf_kar:,.0f}₺",
                            xy=(toplam, mf_kar), xytext=(toplam + x_off, mf_kar + y_off),
                            fontsize=8, ha='center', fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.3', facecolor=mf_colors[i], alpha=0.8, edgecolor='black'),
                            arrowprops=dict(arrowstyle='->', color='black', lw=1))

            ax1.set_xlim(0, x_max)
            ax1.set_ylim(y_min, y_max)
            ax1.set_xlabel('Toplam Miktar (adet)', fontsize=11)
            ax1.set_ylabel('Net Kazanç (TL)', fontsize=11)
            ax1.set_title('Zam Eğrisi + MF Noktaları', fontsize=11, loc='right')
            ax1.grid(True, alpha=0.3)
            ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))

            # Alt grafik: Bar chart
            mf_adlar = [n['mf'] for n in mf_noktalar]
            mf_karlar = [n['kar'] for n in mf_noktalar]
            mf_roiler = [n['roi'] for n in mf_noktalar]

            x_pos = np.arange(len(mf_adlar))
            bar_colors = ['#4CAF50' if kar > 0 else '#F44336' for kar in mf_karlar]

            bars = ax2.bar(x_pos, mf_karlar, color=bar_colors, edgecolor='black', alpha=0.8)
            ax2.axhline(y=0, color='black', linewidth=1)

            for i, (bar, roi) in enumerate(zip(bars, mf_roiler)):
                height = bar.get_height()
                y_pos = height + (abs(y_max) * 0.02 if height >= 0 else -abs(y_max) * 0.05)
                ax2.annotate(f'%{roi:.1f}', xy=(bar.get_x() + bar.get_width() / 2, y_pos),
                            ha='center', va='bottom' if height >= 0 else 'top',
                            fontsize=9, fontweight='bold', color='#1565C0')

            ax2.set_xticks(x_pos)
            ax2.set_xticklabels(mf_adlar, rotation=45, ha='right', fontsize=9)
            ax2.set_ylabel('Net Kazanç (TL)', fontsize=11)
            ax2.set_title('MF Baremleri Karşılaştırması (ROI ile)', fontsize=11, loc='right')
            ax2.grid(True, alpha=0.3, axis='y')
            ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))

            # Bilgi kutusu
            bilgi = (f"Stok: {stok} | Aylık: {aylik_ort:.0f} | Fiyat: {maliyet:.1f}₺\n"
                    f"Zam: %{zam_orani} | {zam_gun} gün | Faiz: %{faiz}")
            fig.text(0.01, 0.01, bilgi, fontsize=9, va='bottom',
                    bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.9))

            plt.tight_layout()
            plt.subplots_adjust(bottom=0.12, top=0.94, hspace=0.35)

            dosya = os.path.join(tempfile.gettempdir(), f"mf_kombine_{urun['UrunAdi'][:20].replace(' ', '_')}.png")
            plt.savefig(dosya, dpi=150, bbox_inches='tight')
            plt.close(fig)

            os.startfile(dosya)

        except Exception as e:
            messagebox.showerror("Grafik Hatası", f"MF kombine grafik oluşturulamadı:\n{e}\n\n{traceback.format_exc()}")

    def _mf_miktar_sec(self, urun, miktar):
        """MF miktarını manuel girişe kopyala"""
        self.manuel_entry.delete(0, tk.END)
        self.manuel_entry.insert(0, str(miktar))

    def _mf_sart_hesapla(self, urun, mf_sart):
        """Geçmiş MF şartını hesapla ve sonucu göster"""
        self._mf_hesapla_ve_goster(urun, mf_sart)

    def _mf_manuel_hesapla(self, urun):
        """Manuel girilen MF şartını hesapla"""
        mf_sart = self.mf_entry.get().strip()
        if not mf_sart or '+' not in mf_sart:
            messagebox.showwarning("Uyarı", "Lütfen geçerli bir MF şartı girin (örn: 5+1)")
            return
        self._mf_hesapla_ve_goster(urun, mf_sart)

    def _mf_hesapla_ve_goster(self, urun, mf_sart):
        """MF hesaplaması yap ve sonucu göster"""
        # Parametreleri al
        faiz = self._aktif_faiz_getir()
        vade = self.depo_vadesi.get()
        hedef_gun = self._hedef_gun_hesapla()
        aylik_ort = urun.get('AylikOrt', 0)
        stok = urun.get('Stok', 0)

        # Depocu fiyatı (bize geliş fiyatı)
        maliyet = urun.get('DepocuFiyat', 0)
        if maliyet <= 0:
            messagebox.showwarning("Uyarı", "Ürün fiyat bilgisi bulunamadı!")
            return

        # MF optimizasyonu hesapla
        sonuc = self._mf_optimum_hesapla(
            mf_sart=mf_sart,
            maliyet=maliyet,
            aylik_ort=aylik_ort,
            mevcut_stok=stok,
            faiz_yillik=faiz,
            depo_vade=vade,
            hedef_gun=hedef_gun
        )

        # Sonuç frame'ini temizle ve yeniden oluştur
        for widget in self.mf_sonuc_frame.winfo_children():
            widget.destroy()

        if sonuc and sonuc.get('oneri'):
            # Sonuç kutusu
            if sonuc['mfli_al']:
                bg_renk = '#E8F5E9'  # Yeşil - MF'li karlı
                fg_renk = '#2E7D32'
            else:
                bg_renk = '#FFEBEE'  # Kırmızı - MF'siz daha karlı
                fg_renk = '#C62828'

            sonuc_label = tk.LabelFrame(self.mf_sonuc_frame, text=f"MF Analizi: {mf_sart}", bg=bg_renk)
            sonuc_label.pack(fill=tk.X, pady=3)

            tk.Label(
                sonuc_label, text=sonuc['oneri'],
                font=('Arial', 9, 'bold'), bg=bg_renk, fg=fg_renk
            ).pack(padx=5, pady=3)

            # Detay bilgileri
            detay_text = f"Alınan: {sonuc['alinan']} | Bedava: {sonuc['bedava']} | Toplam: {sonuc['toplam']}"
            tk.Label(
                sonuc_label, text=detay_text,
                font=('Arial', 8), bg=bg_renk
            ).pack(padx=5, pady=1)

            # "Uygula" butonu
            tk.Button(
                sonuc_label, text="Bu Öneriyi Uygula",
                command=lambda: self._mf_oneri_uygula(urun, sonuc),
                bg='#1976D2', fg='white', font=('Arial', 8)
            ).pack(pady=3)

            # Scroll region güncelle (yeni içerik eklendi)
            self.detay_content.update_idletasks()
            self._detay_scroll_configure()

    def _mf_oneri_uygula(self, urun, sonuc):
        """MF önerisini manuel alanlara uygula (Adet=alinan, MF=bedava)"""
        self.manuel_entry.delete(0, tk.END)
        self.manuel_entry.insert(0, str(sonuc['alinan']))

        self.mf_entry.delete(0, tk.END)
        if sonuc['bedava'] > 0:
            self.mf_entry.insert(0, str(sonuc['bedava']))

        # MF sipariş entry'sini de güncelle
        if hasattr(self, 'mf_siparis_entry'):
            self.mf_siparis_entry.delete(0, tk.END)
            self.mf_siparis_entry.insert(0, str(sonuc['alinan']))

        self.status_label.config(text=f"MF önerisi uygulandı: Adet={sonuc['alinan']} MF={sonuc['bedava']}")

    def _manuel_kaydet(self, urun):
        """Manuel girişleri kaydet"""
        try:
            miktar = int(self.manuel_entry.get() or 0)
        except:
            miktar = 0

        mf_raw = self.mf_entry.get().strip()

        # MF kutucuğunda sadece bedava sayısı varsa (ör: "1") tam formatı oluştur
        if mf_raw and '+' not in mf_raw and miktar > 0:
            try:
                bedava = int(mf_raw)
                mf = f"{miktar}+{bedava}"
            except:
                mf = mf_raw
        else:
            mf = mf_raw

        urun_id = urun['UrunId']
        self.manuel_miktarlar[urun_id] = miktar
        self.manuel_mf_girisler[urun_id] = mf

        urun['Manuel'] = miktar
        urun['MF'] = mf

        self._tabloyu_guncelle()
        self.status_label.config(text=f"{urun.get('UrunAdi', '')} güncellendi")

    # ═══════════════════════════════════════════════════════════════════════
    # ARAMA FONKSİYONLARI (Ctrl+F)
    # ═══════════════════════════════════════════════════════════════════════

    def _arama_odaklan(self, event=None):
        """Ctrl+F ile arama kutusuna odaklan"""
        self.arama_entry.focus_set()
        self.arama_entry.select_range(0, tk.END)
        return "break"

    def _arama_yap(self, *args):
        """Arama metnine göre eşleşmeleri bul"""
        arama = self.arama_var.get().strip().lower()

        # Önceki vurgulamaları temizle
        for item in self.ana_tree.get_children():
            tags = list(self.ana_tree.item(item, 'tags'))
            if 'arama_eslesme' in tags:
                tags.remove('arama_eslesme')
                self.ana_tree.item(item, tags=tags)

        self.arama_eslesmeler = []
        self.arama_index = 0

        if not arama or len(arama) < 2:
            self.arama_sonuc_label.config(text="")
            return

        # Tüm satırları ara
        for item in self.ana_tree.get_children():
            values = self.ana_tree.item(item, 'values')
            if values and len(values) > 1:
                urun_adi = str(values[1]).lower()
                if arama in urun_adi:
                    self.arama_eslesmeler.append(item)
                    # Vurgula
                    tags = list(self.ana_tree.item(item, 'tags'))
                    tags.append('arama_eslesme')
                    self.ana_tree.item(item, tags=tags)

        # Sonuç göster
        if self.arama_eslesmeler:
            self.arama_sonuc_label.config(text=f"1/{len(self.arama_eslesmeler)}")
            self._eslesmeye_git(0)
        else:
            self.arama_sonuc_label.config(text="0 sonuç")

    def _eslesmeye_git(self, index):
        """Belirtilen eşleşmeye git"""
        if not self.arama_eslesmeler:
            return

        if index < 0:
            index = len(self.arama_eslesmeler) - 1
        elif index >= len(self.arama_eslesmeler):
            index = 0

        self.arama_index = index
        item = self.arama_eslesmeler[index]

        # Satırı görünür yap ve seç
        self.ana_tree.see(item)
        self.ana_tree.selection_set(item)
        self.ana_tree.focus(item)

        self.arama_sonuc_label.config(text=f"{index + 1}/{len(self.arama_eslesmeler)}")

    def _sonraki_eslesme(self, event=None):
        """Sonraki eşleşmeye git"""
        self._eslesmeye_git(self.arama_index + 1)

    def _onceki_eslesme(self, event=None):
        """Önceki eşleşmeye git"""
        self._eslesmeye_git(self.arama_index - 1)

    def _arama_temizle(self, event=None):
        """Aramayı temizle"""
        self.arama_var.set("")
        self.arama_sonuc_label.config(text="")
        self.arama_eslesmeler = []
        self.arama_index = 0

        # Vurgulamaları temizle
        for item in self.ana_tree.get_children():
            tags = list(self.ana_tree.item(item, 'tags'))
            if 'arama_eslesme' in tags:
                tags.remove('arama_eslesme')
                self.ana_tree.item(item, tags=tags)

    def _kesin_listeye_ekle(self, urun):
        """Urunu kesin listeye ekle ve database'e kaydet (mükerrer kontrolü ile)"""
        try:
            miktar = int(self.manuel_entry.get() or 0)
        except:
            miktar = urun.get('Oneri', 0)

        if miktar <= 0:
            messagebox.showwarning("Uyari", "Lutfen bir miktar girin!")
            return

        # Mükerrer kontrolü
        urun_id = urun['UrunId']
        for mevcut in self.kesin_siparis_listesi:
            if mevcut.get('UrunId') == urun_id:
                urun_adi = urun.get('UrunAdi', '')[:30]
                messagebox.showwarning("Mükerrer Uyarısı",
                    f"Bu ilaç zaten kesin sipariş listesinde!\n\n"
                    f"{urun_adi}\n"
                    f"Mevcut: {mevcut.get('Miktar', 0)} ad. {mevcut.get('MF', '')}\n\n"
                    f"Düzenlemek için listeden seçip 'Düzenle' butonunu kullanın.")
                return

        mf_raw = self.mf_entry.get().strip()

        # MF'den toplam hesapla
        # mf_raw "10+1" formatında veya sadece bedava sayısı "1" olabilir
        toplam = miktar
        if mf_raw and '+' in mf_raw:
            # Tam format: "10+1"
            mf = mf_raw
            try:
                mf_ek = int(mf_raw.split('+')[1])
                toplam = miktar + mf_ek
            except:
                pass
        elif mf_raw:
            # Sadece bedava sayısı: "1" → tam MF: "miktar+bedava"
            try:
                bedava = int(mf_raw)
                mf = f"{miktar}+{bedava}"
                toplam = miktar + bedava
            except:
                mf = mf_raw
        else:
            mf = ''

        siparis_data = {
            'UrunId': urun['UrunId'],
            'UrunAdi': urun.get('UrunAdi', ''),
            'Barkod': urun.get('Barkod', ''),
            'Miktar': miktar,
            'MF': mf,
            'Toplam': toplam,
            'Stok': urun.get('Stok', 0),
            'AylikOrt': urun.get('AylikOrt', 0),
        }

        # Database'e kaydet
        db_id = None
        if self.siparis_db and self.aktif_calisma:
            db_id = self.siparis_db.siparis_ekle(self.aktif_calisma['id'], siparis_data)

        siparis_data['db_id'] = db_id
        self.kesin_siparis_listesi.append(siparis_data)

        self._kesin_liste_guncelle()
        self._calisma_bilgisini_guncelle()
        self.status_label.config(text=f"{urun.get('UrunAdi', '')} kesin listeye eklendi")

    def _tum_onerileri_kopyala(self):
        """Tüm önerileri manuel alanlara kopyala"""
        for urun in self.tum_veriler:
            oneri = urun.get('Oneri', 0)
            if oneri > 0:
                urun_id = urun['UrunId']
                self.manuel_miktarlar[urun_id] = oneri
                urun['Manuel'] = oneri

        self._tabloyu_guncelle()
        self.status_label.config(text="Tüm öneriler manuel alanlara kopyalandı")

    def _secilileri_kesin_listeye_ekle(self):
        """Seçili satırları kesin listeye ekle (mükerrer kontrolü ile)"""
        selection = self.ana_tree.selection()
        eklenen = 0
        atlanan_mukerrer = 0
        mevcut_urun_idler = {s.get('UrunId') for s in self.kesin_siparis_listesi}

        for item_id in selection:
            if item_id.startswith('urun_'):
                urun_id = int(item_id.replace('urun_', ''))
                urun = next((u for u in self.tum_veriler if u['UrunId'] == urun_id), None)
                if urun:
                    # Mükerrer kontrolü
                    if urun_id in mevcut_urun_idler:
                        atlanan_mukerrer += 1
                        continue

                    miktar = urun.get('Manuel', 0) or urun.get('Oneri', 0)
                    if miktar > 0:
                        mf = urun.get('MF', '')
                        toplam = miktar
                        if mf and '+' in mf:
                            try:
                                mf_ek = int(mf.split('+')[1])
                                toplam = miktar + mf_ek
                            except:
                                pass

                        self.kesin_siparis_listesi.append({
                            'UrunId': urun['UrunId'],
                            'UrunAdi': urun.get('UrunAdi', ''),
                            'Barkod': urun.get('Barkod', ''),
                            'Miktar': miktar,
                            'MF': mf,
                            'Toplam': toplam,
                            'Stok': urun.get('Stok', 0),
                            'AylikOrt': urun.get('AylikOrt', 0),
                            'Selcuk': '',
                            'Alliance': '',
                            'Sancak': '',
                            'Iskoop': '',
                            'Farmazon': ''
                        })
                        eklenen += 1
                        mevcut_urun_idler.add(urun_id)

        self._kesin_liste_guncelle()
        durum = f"{eklenen} ürün kesin listeye eklendi"
        if atlanan_mukerrer > 0:
            durum += f" | ⚠ {atlanan_mukerrer} ürün zaten listede (atlandı)"
        self.status_label.config(text=durum)

    def _kesin_liste_guncelle(self):
        """Kesin sipariş listesini güncelle"""
        for item in self.kesin_tree.get_children():
            self.kesin_tree.delete(item)

        for siparis in self.kesin_siparis_listesi:
            self.kesin_tree.insert('', 'end', values=(
                siparis.get('UrunAdi') or '',
                siparis.get('Barkod') or '',  # None ise boş string
                siparis.get('Miktar') or 0,
                siparis.get('MF') or '',
                siparis.get('Toplam') or 0,
                siparis.get('Selcuk') or '',
                siparis.get('Alliance') or '',
                siparis.get('Sancak') or '',
                siparis.get('Iskoop') or '',
                siparis.get('Farmazon') or ''
            ))

    def _kesin_listeyi_temizle(self):
        """Kesin listeyi temizle (database dahil)"""
        if messagebox.askyesno("Onay", "Kesin siparis listesi temizlensin mi?"):
            # Database'den sil
            if self.siparis_db and self.aktif_calisma:
                self.siparis_db.calisma_siparislerini_temizle(self.aktif_calisma['id'])

            self.kesin_siparis_listesi = []
            self._kesin_liste_guncelle()
            self._calisma_bilgisini_guncelle()

    def _kesin_seciliyi_sil(self):
        """Secili siparis kaydini sil (database dahil)"""
        secili = self.kesin_tree.selection()
        if not secili:
            messagebox.showwarning("Uyari", "Lutfen silmek icin bir satir secin!")
            return

        # Secili satirlarin indexlerini bul
        silinecekler = []
        for item in secili:
            idx = self.kesin_tree.index(item)
            silinecekler.append(idx)

        # Onay al
        adet = len(silinecekler)
        if not messagebox.askyesno("Onay", f"{adet} adet siparis silinsin mi?"):
            return

        # Buyukten kucuge sil (index kaymamasi icin)
        for idx in sorted(silinecekler, reverse=True):
            if idx < len(self.kesin_siparis_listesi):
                siparis = self.kesin_siparis_listesi[idx]
                # Database'den sil
                db_id = siparis.get('db_id')
                if db_id and self.siparis_db:
                    self.siparis_db.siparis_sil(db_id)
                del self.kesin_siparis_listesi[idx]

        self._kesin_liste_guncelle()
        self._calisma_bilgisini_guncelle()

    def _kesin_miktari_duzenle(self):
        """Secili siparisin miktarini duzenle (database dahil) - Miktar, MF ve Muadil Çevirme"""
        secili = self.kesin_tree.selection()
        if not secili or len(secili) != 1:
            messagebox.showwarning("Uyari", "Lutfen duzenlemek icin tek bir satir secin!")
            return

        item = secili[0]
        idx = self.kesin_tree.index(item)

        if idx >= len(self.kesin_siparis_listesi):
            return

        siparis = self.kesin_siparis_listesi[idx]
        mevcut_miktar = siparis.get('Miktar', 0)
        mevcut_mf = siparis.get('MF', '')
        urun_adi = siparis.get('UrunAdi', '')
        urun_id = siparis.get('UrunId')

        # Özel dialog penceresi
        dialog = tk.Toplevel(self.parent)
        dialog.title("Sipariş Düzenle")
        dialog.geometry("420x480")
        dialog.resizable(False, False)
        dialog.transient(self.parent)
        dialog.grab_set()

        # Ana frame
        main_frame = tk.Frame(dialog, padx=15, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Ürün adı
        urun_label = tk.Label(main_frame, text=urun_adi[:50], font=('Arial', 10, 'bold'),
                              wraplength=380, fg='#1565C0')
        urun_label.pack(pady=(5, 10))

        # Miktar ve MF çerçevesi
        duzenle_frame = tk.LabelFrame(main_frame, text=" Miktar & MF ", font=('Arial', 9, 'bold'),
                                      padx=10, pady=8)
        duzenle_frame.pack(fill=tk.X, pady=(0, 10))

        # Miktar girişi
        miktar_row = tk.Frame(duzenle_frame)
        miktar_row.pack(fill=tk.X, pady=3)
        tk.Label(miktar_row, text="Sipariş Miktarı:", font=('Arial', 10), width=15, anchor='e').pack(side=tk.LEFT)
        miktar_entry = ttk.Entry(miktar_row, width=10, font=('Arial', 11))
        miktar_entry.pack(side=tk.LEFT, padx=5)
        miktar_entry.insert(0, str(mevcut_miktar))

        # MF girişi
        mf_row = tk.Frame(duzenle_frame)
        mf_row.pack(fill=tk.X, pady=3)
        tk.Label(mf_row, text="MF Şartı (ör: 5+1):", font=('Arial', 10), width=15, anchor='e').pack(side=tk.LEFT)
        mf_entry = ttk.Entry(mf_row, width=10, font=('Arial', 11))
        mf_entry.pack(side=tk.LEFT, padx=5)
        mf_entry.insert(0, mevcut_mf)

        # ═══════════════════════════════════════════════════════════════════
        # MUADİLİNE ÇEVİR BÖLÜMÜ
        # ═══════════════════════════════════════════════════════════════════
        muadil_frame = tk.LabelFrame(main_frame, text=" 🔄 Muadiline Çevir ", font=('Arial', 9, 'bold'),
                                     padx=10, pady=8, bg='#FFF3E0')
        muadil_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Muadil kartları için scrollable frame
        muadil_canvas = tk.Canvas(muadil_frame, bg='#FFF3E0', highlightthickness=0, height=180)
        muadil_scrollbar = ttk.Scrollbar(muadil_frame, orient='vertical', command=muadil_canvas.yview)
        muadil_cards_frame = tk.Frame(muadil_canvas, bg='#FFF3E0')

        muadil_canvas.configure(yscrollcommand=muadil_scrollbar.set)
        muadil_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        muadil_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        muadil_canvas.create_window((0, 0), window=muadil_cards_frame, anchor='nw')
        muadil_cards_frame.bind('<Configure>',
            lambda e: muadil_canvas.configure(scrollregion=muadil_canvas.bbox('all')))

        # Seçili muadil değişkeni
        secili_muadil = {'data': None}

        def muadil_sec(muadil_data, card_frame):
            """Muadil kartını seç"""
            # Önceki seçimi temizle
            for widget in muadil_cards_frame.winfo_children():
                widget.configure(bg='#FFF8E1')
                for child in widget.winfo_children():
                    try:
                        child.configure(bg='#FFF8E1')
                    except:
                        pass

            # Yeni seçimi işaretle
            card_frame.configure(bg='#C8E6C9')
            for child in card_frame.winfo_children():
                try:
                    child.configure(bg='#C8E6C9')
                except:
                    pass

            secili_muadil['data'] = muadil_data
            urun_label.config(text=f"➜ {muadil_data['UrunAdi'][:50]}", fg='#2E7D32')

        def muadilleri_yukle():
            """Muadil ilaçları veritabanından getir ve kartları oluştur"""
            # Temizle
            for widget in muadil_cards_frame.winfo_children():
                widget.destroy()

            if not self.db:
                tk.Label(muadil_cards_frame, text="Veritabanı bağlantısı yok",
                        font=('Arial', 9), bg='#FFF3E0', fg='#C62828').pack(pady=20)
                return

            if not urun_id:
                tk.Label(muadil_cards_frame, text="Ürün ID bulunamadı\n(Eski sipariş olabilir)",
                        font=('Arial', 9), bg='#FFF3E0', fg='#E65100').pack(pady=20)
                return

            try:
                # Önce ilacın eşdeğer grubunu bul
                sql_esdeger = f"""
                SELECT UrunEsdegerId FROM Urun WHERE UrunId = {urun_id} AND UrunSilme = 0
                """
                sonuc = self.db.sorgu_calistir(sql_esdeger)
                if not sonuc or not sonuc[0].get('UrunEsdegerId'):
                    tk.Label(muadil_cards_frame, text="Bu ilacın eşdeğer grubu yok",
                            font=('Arial', 9), bg='#FFF3E0', fg='#666').pack(pady=20)
                    return

                esdeger_id = sonuc[0]['UrunEsdegerId']

                # Aynı gruptaki tüm ilaçları getir (kendisi hariç)
                sql_muadiller = f"""
                SELECT
                    u.UrunId,
                    u.UrunAdi,
                    CASE WHEN u.UrunUrunTipId IN (1, 16) THEN
                        (SELECT COUNT(*) FROM Karekod kk WHERE kk.KKUrunId = u.UrunId AND kk.KKDurum = 1)
                    ELSE (COALESCE(u.UrunStokDepo,0) + COALESCE(u.UrunStokRaf,0) + COALESCE(u.UrunStokAcik,0))
                    END AS Stok,
                    COALESCE(u.UrunFiyatEtiket, 0) as PSF,
                    COALESCE(u.UrunIskontoKamu, 0) as Iskonto,
                    (SELECT TOP 1 b.BarkodAdi FROM Barkod b WHERE b.BarkodUrunId = u.UrunId AND b.BarkodSilme = 0) as Barkod
                FROM Urun u
                WHERE u.UrunEsdegerId = {esdeger_id}
                AND u.UrunSilme = 0
                AND u.UrunId != {urun_id}
                ORDER BY u.UrunAdi
                """
                muadiller = self.db.sorgu_calistir(sql_muadiller)

                if not muadiller:
                    tk.Label(muadil_cards_frame, text="Başka muadil bulunamadı",
                            font=('Arial', 9), bg='#FFF3E0', fg='#666').pack(pady=20)
                    return

                # Muadil kartlarını oluştur
                for muadil in muadiller:
                    stok = int(muadil.get('Stok') or 0)
                    psf = float(muadil.get('PSF') or 0)
                    iskonto = float(muadil.get('Iskonto') or 0)
                    depocu = psf * 0.71 * 1.10 * (1 - iskonto / 100) if psf > 0 else 0

                    # Kart frame
                    card = tk.Frame(muadil_cards_frame, bg='#FFF8E1', relief='raised', bd=1)
                    card.pack(fill=tk.X, padx=5, pady=3)

                    # Sol: İlaç bilgileri
                    info_frame = tk.Frame(card, bg='#FFF8E1')
                    info_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)

                    ilac_adi = muadil['UrunAdi'][:35] + ".." if len(muadil['UrunAdi']) > 37 else muadil['UrunAdi']
                    tk.Label(info_frame, text=ilac_adi, font=('Arial', 9, 'bold'),
                            bg='#FFF8E1', anchor='w').pack(anchor='w')

                    # Stok ve fiyat bilgisi
                    stok_renk = '#2E7D32' if stok > 0 else '#C62828'
                    bilgi_text = f"Stok: {stok}  |  Fiyat: {depocu:.0f}₺"
                    tk.Label(info_frame, text=bilgi_text, font=('Arial', 8),
                            bg='#FFF8E1', fg=stok_renk).pack(anchor='w')

                    # Sağ: Seç butonu
                    btn = tk.Button(card, text="Seç", font=('Arial', 9, 'bold'),
                                   bg='#FF9800', fg='white', width=5,
                                   command=lambda m=muadil, c=card: muadil_sec(m, c))
                    btn.pack(side=tk.RIGHT, padx=5, pady=5)

                    # Karta tıklama
                    card.bind('<Button-1>', lambda e, m=muadil, c=card: muadil_sec(m, c))

            except Exception as e:
                logger.error(f"Muadil yükleme hatası: {e}")
                tk.Label(muadil_cards_frame, text=f"Hata: {str(e)[:40]}",
                        font=('Arial', 9), bg='#FFF3E0', fg='#C62828').pack(pady=20)

        # Muadilleri yükle butonu (ilk yüklemede çağrılacak)
        muadilleri_yukle()

        # ═══════════════════════════════════════════════════════════════════
        # BUTONLAR
        # ═══════════════════════════════════════════════════════════════════
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(pady=10)

        def kaydet():
            try:
                yeni_miktar = int(miktar_entry.get().strip() or 0)
                yeni_mf = mf_entry.get().strip()

                if yeni_miktar <= 0:
                    messagebox.showwarning("Uyarı", "Miktar 0'dan büyük olmalı!")
                    return

                # Toplam hesapla
                if yeni_mf and '+' in yeni_mf:
                    try:
                        parcalar = yeni_mf.split('+')
                        mf_al = int(parcalar[0])
                        mf_bedava = int(parcalar[1])
                        set_sayisi = yeni_miktar // mf_al
                        if set_sayisi == 0:
                            set_sayisi = 1
                        toplam = (set_sayisi * mf_al) + (set_sayisi * mf_bedava)
                    except:
                        toplam = yeni_miktar
                else:
                    toplam = yeni_miktar

                # Muadile çevirme kontrolü
                if secili_muadil['data']:
                    muadil = secili_muadil['data']
                    siparis['UrunId'] = muadil['UrunId']
                    siparis['UrunAdi'] = muadil['UrunAdi']
                    siparis['Barkod'] = muadil.get('Barkod', '')
                    siparis['Stok'] = int(muadil.get('Stok') or 0)

                siparis['Miktar'] = yeni_miktar
                siparis['MF'] = yeni_mf
                siparis['Toplam'] = toplam

                # Database'i guncelle
                db_id = siparis.get('db_id')
                if db_id and self.siparis_db:
                    guncelleme_data = {
                        'miktar': yeni_miktar,
                        'mf': yeni_mf,
                        'toplam': toplam
                    }
                    if secili_muadil['data']:
                        guncelleme_data['urun_id'] = siparis['UrunId']
                        guncelleme_data['urun_adi'] = siparis['UrunAdi']
                        guncelleme_data['barkod'] = siparis.get('Barkod', '')
                        guncelleme_data['stok'] = siparis.get('Stok', 0)

                    self.siparis_db.siparis_guncelle(db_id, **guncelleme_data)

                self._kesin_liste_guncelle()
                dialog.destroy()

                if secili_muadil['data']:
                    messagebox.showinfo("Başarılı",
                        f"Sipariş muadiline çevrildi:\n{secili_muadil['data']['UrunAdi'][:40]}")

            except Exception as e:
                messagebox.showerror("Hata", f"Güncelleme hatası: {e}")

        tk.Button(btn_frame, text="Kaydet", command=kaydet, bg='#4CAF50', fg='white',
                  font=('Arial', 10, 'bold'), width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="İptal", command=dialog.destroy, bg='#757575', fg='white',
                  font=('Arial', 10), width=10).pack(side=tk.LEFT, padx=5)

        # Enter tuşu ile kaydet
        dialog.bind('<Return>', lambda e: kaydet())
        miktar_entry.focus_set()

    def _yeni_calisma_dialog(self):
        """Yeni siparis calismasi olusturma dialogu"""
        if not self.siparis_db:
            messagebox.showerror("Hata", "Siparis veritabani baglantisi yok!")
            return

        # Mevcut calismada siparis varsa uyar
        if self.kesin_siparis_listesi:
            if not messagebox.askyesno("Onay",
                    f"Mevcut calismada {len(self.kesin_siparis_listesi)} siparis var.\n"
                    "Yeni calisma olusturulursa mevcut calisma kapatilacak.\n\n"
                    "Devam edilsin mi?"):
                return

            # Mevcut calismayi kapat
            if self.aktif_calisma:
                self.siparis_db.calisma_kapat(self.aktif_calisma['id'])

        # Yeni calisma adi sor
        from tkinter import simpledialog
        bugun = datetime.now().strftime("%d.%m.%Y")
        varsayilan_ad = f"{bugun} Siparis Calismasi"

        ad = simpledialog.askstring(
            "Yeni Calisma",
            "Calisma adi:",
            initialvalue=varsayilan_ad
        )

        if ad:
            self._yeni_calisma_olustur(ad=ad)
            messagebox.showinfo("Basarili", f"Yeni calisma olusturuldu:\n{ad}")

    def _calisma_yukle_dialog(self):
        """Onceki siparis calismasini yukleme dialogu"""
        if not self.siparis_db:
            messagebox.showerror("Hata", "Siparis veritabani baglantisi yok!")
            return

        # Calisma listesini getir
        calismalar = self.siparis_db.calisma_listesi_getir(limit=20)

        if not calismalar:
            messagebox.showinfo("Bilgi", "Kayitli siparis calismasi bulunamadi.")
            return

        # Secim penceresi olustur
        dialog = tk.Toplevel(self.parent)
        dialog.title("Calisma Yukle")
        dialog.geometry("500x400")
        dialog.transient(self.parent)
        dialog.grab_set()

        # Baslik
        tk.Label(dialog, text="Onceki Siparis Calismalari", font=('Arial', 12, 'bold'),
                bg='#1976D2', fg='white', pady=10).pack(fill=tk.X)

        # Liste
        liste_frame = tk.Frame(dialog)
        liste_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('ad', 'tarih', 'siparis', 'durum')
        tree = ttk.Treeview(liste_frame, columns=columns, show='headings', height=12)
        tree.heading('ad', text='Calisma Adi')
        tree.heading('tarih', text='Tarih')
        tree.heading('siparis', text='Siparis')
        tree.heading('durum', text='Durum')
        tree.column('ad', width=200)
        tree.column('tarih', width=120)
        tree.column('siparis', width=60)
        tree.column('durum', width=80)

        vsb = ttk.Scrollbar(liste_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Calismalari listele
        for c in calismalar:
            tarih = c.get('olusturma_tarihi', '')[:16]
            tree.insert('', 'end', values=(
                c.get('ad', ''),
                tarih,
                c.get('siparis_sayisi', 0),
                c.get('durum', '')
            ), tags=(str(c['id']),))

        # Butonlar
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        def yukle():
            secili = tree.selection()
            if not secili:
                messagebox.showwarning("Uyari", "Lutfen bir calisma secin!")
                return

            calisma_id = int(tree.item(secili[0], 'tags')[0])

            # Mevcut calismayi kapat
            if self.aktif_calisma and self.aktif_calisma['id'] != calisma_id:
                if self.kesin_siparis_listesi:
                    self.siparis_db.calisma_kapat(self.aktif_calisma['id'])

            # Secilen calismayi aktif yap
            self.siparis_db.calisma_guncelle(calisma_id, durum='aktif')
            self.aktif_calisma = {'id': calisma_id, 'ad': tree.item(secili[0], 'values')[0]}

            # Siparisleri yukle
            self._calisma_siparislerini_yukle()
            self._calisma_bilgisini_guncelle()

            dialog.destroy()
            messagebox.showinfo("Basarili", f"Calisma yuklendi: {self.aktif_calisma['ad']}")

        def sil():
            secili = tree.selection()
            if not secili:
                messagebox.showwarning("Uyari", "Lutfen bir calisma secin!")
                return

            calisma_id = int(tree.item(secili[0], 'tags')[0])
            calisma_ad = tree.item(secili[0], 'values')[0]

            if messagebox.askyesno("Onay", f"'{calisma_ad}' calismasi silinsin mi?\nTum siparisler de silinecek!"):
                self.siparis_db.calisma_sil(calisma_id)
                tree.delete(secili[0])

                # Eger aktif calisma silindiyse
                if self.aktif_calisma and self.aktif_calisma['id'] == calisma_id:
                    self.aktif_calisma = None
                    self.kesin_siparis_listesi = []
                    self._kesin_liste_guncelle()
                    self._yeni_calisma_olustur()

        tk.Button(btn_frame, text="Yukle", command=yukle,
                 bg='#388E3C', fg='white', font=('Arial', 10, 'bold'), width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Sil", command=sil,
                 bg='#C62828', fg='white', font=('Arial', 10), width=10).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Kapat", command=dialog.destroy,
                 font=('Arial', 10), width=10).pack(side=tk.RIGHT, padx=5)

    def _calisma_kaydet_kapat(self):
        """Aktif çalışmayı kaydet ve kapat"""
        if not self.siparis_db or not self.aktif_calisma:
            messagebox.showwarning("Uyarı", "Aktif sipariş çalışması yok!")
            return

        kalem = len(self.kesin_siparis_listesi)
        kutu = sum(s.get('Toplam') or s.get('Miktar') or 0 for s in self.kesin_siparis_listesi)

        mesaj = f"Çalışma kapatılacak:\n\n"
        mesaj += f"Çalışma: {self.aktif_calisma['ad']}\n"
        mesaj += f"Kalem: {kalem}\n"
        mesaj += f"Toplam Kutu: {kutu}\n\n"
        mesaj += "Onaylıyor musunuz?"

        if not messagebox.askyesno("Çalışmayı Kaydet ve Kapat", mesaj):
            return

        # Çalışmayı kapat
        self.siparis_db.calisma_kapat(self.aktif_calisma['id'])

        # Yeni çalışma oluştur
        self.aktif_calisma = None
        self.kesin_siparis_listesi = []
        self._kesin_liste_guncelle()
        self._yeni_calisma_olustur()

        messagebox.showinfo("Başarılı", "Çalışma kaydedildi ve kapatıldı.\nYeni çalışma oluşturuldu.")

    def _calisma_arsiv_goster(self):
        """Eski sipariş çalışmalarını arşiv olarak göster"""
        if not self.siparis_db:
            messagebox.showerror("Hata", "Sipariş veritabanı bağlantısı yok!")
            return

        # Arşiv listesini getir
        arsiv = self.siparis_db.arsiv_listesi_getir(limit=50)

        if not arsiv:
            messagebox.showinfo("Bilgi", "Arşivde sipariş çalışması bulunamadı.")
            return

        # Arşiv penceresi
        dialog = tk.Toplevel(self.parent)
        dialog.title("Sipariş Çalışmaları Arşivi")
        dialog.geometry("750x500")
        dialog.transient(self.parent)
        dialog.grab_set()

        # Başlık
        tk.Label(dialog, text="SİPARİŞ ÇALIŞMALARI ARŞİVİ", font=('Arial', 14, 'bold'),
                bg='#5D4037', fg='white', pady=12).pack(fill=tk.X)

        # Tablo
        tablo_frame = tk.Frame(dialog)
        tablo_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('ad', 'baslangic', 'bitis', 'durum', 'kalem', 'kutu')
        tree = ttk.Treeview(tablo_frame, columns=columns, show='headings', height=18)
        tree.heading('ad', text='Çalışma Adı')
        tree.heading('baslangic', text='Başlangıç')
        tree.heading('bitis', text='Bitiş')
        tree.heading('durum', text='Durum')
        tree.heading('kalem', text='Kalem')
        tree.heading('kutu', text='Kutu')

        tree.column('ad', width=200)
        tree.column('baslangic', width=130)
        tree.column('bitis', width=130)
        tree.column('durum', width=90)
        tree.column('kalem', width=60)
        tree.column('kutu', width=70)

        vsb = ttk.Scrollbar(tablo_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Verileri ekle
        for c in arsiv:
            baslangic = c.get('olusturma_tarihi', '')[:16] if c.get('olusturma_tarihi') else '-'
            bitis = c.get('bitis_tarihi', '')[:16] if c.get('bitis_tarihi') else '-'
            durum = c.get('durum', '')
            durum_text = 'Tamamlandı' if durum == 'tamamlandi' else ('Aktif' if durum == 'aktif' else durum)

            tree.insert('', 'end', values=(
                c.get('ad', ''),
                baslangic,
                bitis,
                durum_text,
                c.get('kalem', 0),
                c.get('kutu', 0)
            ), tags=(str(c['id']),))

        # Renklendirme
        tree.tag_configure('aktif', background='#E8F5E9')
        tree.tag_configure('tamamlandi', background='#ECEFF1')

        # Özet bilgi
        toplam_kalem = sum(c.get('kalem', 0) for c in arsiv)
        toplam_kutu = sum(c.get('kutu', 0) for c in arsiv)
        tamamlanan = sum(1 for c in arsiv if c.get('durum') == 'tamamlandi')

        ozet_frame = tk.Frame(dialog, bg='#ECEFF1', pady=5)
        ozet_frame.pack(fill=tk.X, padx=10)
        tk.Label(ozet_frame, text=f"Toplam: {len(arsiv)} çalışma | {tamamlanan} tamamlandı | {toplam_kalem} kalem | {toplam_kutu} kutu",
                font=('Arial', 10), bg='#ECEFF1').pack()

        # Butonlar
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        def yukle_secili():
            secili = tree.selection()
            if not secili:
                messagebox.showwarning("Uyarı", "Lütfen bir çalışma seçin!")
                return

            calisma_id = int(tree.item(secili[0], 'tags')[0])
            calisma_ad = tree.item(secili[0], 'values')[0]

            # Mevcut çalışmayı kapat
            if self.aktif_calisma and self.aktif_calisma['id'] != calisma_id:
                if self.kesin_siparis_listesi:
                    self.siparis_db.calisma_kapat(self.aktif_calisma['id'])

            # Seçilen çalışmayı aktif yap
            self.siparis_db.calisma_guncelle(calisma_id, durum='aktif')
            self.aktif_calisma = {'id': calisma_id, 'ad': calisma_ad}

            # Siparişleri yükle
            self._calisma_siparislerini_yukle()
            self._calisma_bilgisini_guncelle()

            dialog.destroy()
            messagebox.showinfo("Başarılı", f"Çalışma yüklendi: {calisma_ad}")

        tk.Button(btn_frame, text="Seçili Çalışmayı Yükle", command=yukle_secili,
                 bg='#388E3C', fg='white', font=('Arial', 10, 'bold'), width=20).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Kapat", command=dialog.destroy,
                 font=('Arial', 10), width=10).pack(side=tk.RIGHT, padx=5)

    def _kesin_listeyi_excel_aktar(self):
        """Kesin listeyi Excel'e aktar"""
        if not self.kesin_siparis_listesi:
            messagebox.showwarning("Uyari", "Kesin siparis listesi bos!")
            return

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            title="Kesin Sipariş Listesini Kaydet"
        )

        if not dosya_yolu:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Kesin Sipariş"

            headers = ["Ürün Adı", "Barkod", "Miktar", "MF", "Toplam", "Selçuk", "Alliance", "Sancak", "İskoop", "Farmazon"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h).font = Font(bold=True)

            for row_idx, siparis in enumerate(self.kesin_siparis_listesi, 2):
                ws.cell(row=row_idx, column=1, value=siparis.get('UrunAdi', ''))
                ws.cell(row=row_idx, column=2, value=siparis.get('Barkod', ''))
                ws.cell(row=row_idx, column=3, value=siparis.get('Miktar', 0))
                ws.cell(row=row_idx, column=4, value=siparis.get('MF', ''))
                ws.cell(row=row_idx, column=5, value=siparis.get('Toplam', 0))
                ws.cell(row=row_idx, column=6, value=siparis.get('Selcuk', ''))
                ws.cell(row=row_idx, column=7, value=siparis.get('Alliance', ''))
                ws.cell(row=row_idx, column=8, value=siparis.get('Sancak', ''))
                ws.cell(row=row_idx, column=9, value=siparis.get('Iskoop', ''))
                ws.cell(row=row_idx, column=10, value=siparis.get('Farmazon', ''))

            # Sütun genişliklerini ayarla
            ws.column_dimensions['A'].width = 40  # Ürün Adı
            ws.column_dimensions['B'].width = 15  # Barkod
            ws.column_dimensions['C'].width = 8   # Miktar
            ws.column_dimensions['D'].width = 8   # MF
            ws.column_dimensions['E'].width = 8   # Toplam
            ws.column_dimensions['F'].width = 12  # Selçuk
            ws.column_dimensions['G'].width = 12  # Alliance
            ws.column_dimensions['H'].width = 12  # Sancak
            ws.column_dimensions['I'].width = 12  # İskoop
            ws.column_dimensions['J'].width = 12  # Farmazon

            wb.save(dosya_yolu)
            messagebox.showinfo("Başarılı", f"Liste kaydedildi:\n{dosya_yolu}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatası: {e}")

    def _depolarda_ara(self):
        """Kesin listedeki ürünleri depolarda ara"""
        if not self.kesin_siparis_listesi:
            messagebox.showwarning("Uyarı", "Kesin sipariş listesi boş!")
            return

        # Barkodu olmayan ürünleri kontrol et
        barkodlu_urunler = [s for s in self.kesin_siparis_listesi if s.get('Barkod')]
        if not barkodlu_urunler:
            messagebox.showwarning("Uyarı", "Listede barkodu olan ürün yok!")
            return

        # Butonları devre dışı bırak
        self.depo_ara_btn.config(state='disabled', text="Aranıyor...")
        self.depo_durum_label.config(text=f"0/{len(barkodlu_urunler)} ürün arandı")

        def arama_thread():
            try:
                # BotSiparis modüllerini import et
                import sys
                import os

                # BotSiparis kök klasörünü path'e ekle
                bot_root = os.path.join(os.path.dirname(__file__), "BotSiparis - Kopya")
                bot_src = os.path.join(bot_root, "src")

                # Hem kök hem src'yi path'e ekle (sıra önemli)
                for p in [bot_root, bot_src]:
                    if p not in sys.path:
                        sys.path.insert(0, p)

                from src.depolar.selcuk import SelcukDepo
                from src.depolar.alliance import AllianceDepo
                from src.depolar.sancak import SancakDepo
                from src.depolar.iskoop import IskoopDepo
                from src.depolar.farmazon import FarmazonDepo
                from dotenv import load_dotenv

                # .env dosyasını yükle
                env_path = os.path.join(os.path.dirname(__file__), "BotSiparis - Kopya", ".env")
                load_dotenv(env_path)

                # Depo sırası: Selçuk, Alliance, Sancak, İskoop, Farmazon
                depolar = []

                # Selçuk - login(hesap_kodu, username, password)
                if os.environ.get('SELCUK_ENABLED', 'false').lower() == 'true':
                    depolar.append({
                        'key': 'Selcuk',
                        'class': SelcukDepo,
                        'login_args': (
                            os.environ.get('SELCUK_HESAP_KODU'),
                            os.environ.get('SELCUK_USERNAME'),
                            os.environ.get('SELCUK_PASSWORD')
                        )
                    })

                # Alliance - login(eczane_kodu, username, password)
                if os.environ.get('ALLIANCE_ENABLED', 'false').lower() == 'true':
                    depolar.append({
                        'key': 'Alliance',
                        'class': AllianceDepo,
                        'login_args': (
                            os.environ.get('ALLIANCE_ECZANE_KODU'),
                            os.environ.get('ALLIANCE_USERNAME'),
                            os.environ.get('ALLIANCE_PASSWORD')
                        )
                    })

                # Sancak - login(username, password)
                if os.environ.get('SANCAK_ENABLED', 'false').lower() == 'true':
                    depolar.append({
                        'key': 'Sancak',
                        'class': SancakDepo,
                        'login_args': (
                            os.environ.get('SANCAK_USERNAME'),
                            os.environ.get('SANCAK_PASSWORD')
                        )
                    })

                # İskoop - login(username, password)
                if os.environ.get('ISKOOP_ENABLED', 'false').lower() == 'true':
                    depolar.append({
                        'key': 'Iskoop',
                        'class': IskoopDepo,
                        'login_args': (
                            os.environ.get('ISKOOP_USERNAME'),
                            os.environ.get('ISKOOP_PASSWORD')
                        )
                    })

                # Farmazon - login(username, password)
                if os.environ.get('FARMAZON_ENABLED', 'false').lower() == 'true':
                    depolar.append({
                        'key': 'Farmazon',
                        'class': FarmazonDepo,
                        'login_args': (
                            os.environ.get('FARMAZON_USERNAME'),
                            os.environ.get('FARMAZON_PASSWORD')
                        )
                    })

                if not depolar:
                    self.parent.after(0, lambda: messagebox.showwarning(
                        "Uyarı", "Hiçbir depo etkinleştirilmemiş!\n.env dosyasını kontrol edin."
                    ))
                    return

                # Her depo için driver başlat ve login ol
                aktif_depolar = []
                for depo_info in depolar:
                    try:
                        depo = depo_info['class']()
                        if depo.init_driver():
                            if depo.login(*depo_info['login_args']):
                                aktif_depolar.append({'key': depo_info['key'], 'depo': depo})
                                logger.info(f"{depo_info['key']} deposuna giriş başarılı")
                            else:
                                logger.warning(f"{depo_info['key']} deposuna giriş başarısız")
                    except Exception as e:
                        logger.error(f"{depo_info['key']} depo hatası: {e}")

                if not aktif_depolar:
                    self.parent.after(0, lambda: messagebox.showerror(
                        "Hata", "Hiçbir depoya giriş yapılamadı!"
                    ))
                    return

                # Her ürün için arama yap
                for idx, siparis in enumerate(self.kesin_siparis_listesi):
                    barkod = siparis.get('Barkod')
                    if not barkod:
                        continue

                    # Durum güncelle
                    self.parent.after(0, lambda i=idx, t=len(barkodlu_urunler):
                        self.depo_durum_label.config(text=f"{i+1}/{t} aranıyor...")
                    )

                    # Her depoda ara
                    for depo_bilgi in aktif_depolar:
                        depo_key = depo_bilgi['key']
                        depo = depo_bilgi['depo']

                        try:
                            # Barkodu ara
                            if depo.search_barcode(str(barkod)):
                                # Stok durumu kontrol et
                                sonuc = depo.check_stock_status()

                                if sonuc.get('stok_var'):
                                    # MF varsa göster
                                    mf_sart = sonuc.get('sart', '')
                                    if mf_sart:
                                        siparis[depo_key] = f"✓ {mf_sart}"
                                    else:
                                        siparis[depo_key] = "✓ Var"
                                elif sonuc.get('mesaj') == 'Depoyu Ara':
                                    siparis[depo_key] = "📞 Ara"
                                else:
                                    siparis[depo_key] = "✗ Yok"
                            else:
                                siparis[depo_key] = "- Yok"
                        except Exception as e:
                            logger.error(f"{depo_key} arama hatası ({barkod}): {e}")
                            siparis[depo_key] = "! Hata"

                    # Listeyi güncelle
                    self.parent.after(0, self._kesin_liste_guncelle)

                # Tarayıcıları kapat
                for depo_bilgi in aktif_depolar:
                    try:
                        depo_bilgi['depo'].close()
                    except:
                        pass

                # Tamamlandı
                self.parent.after(0, lambda: self._depo_arama_tamamlandi(len(barkodlu_urunler)))

            except Exception as e:
                logger.error(f"Depo arama hatası: {e}")
                self.parent.after(0, lambda: messagebox.showerror("Hata", f"Depo arama hatası:\n{e}"))
            finally:
                self.parent.after(0, lambda: self.depo_ara_btn.config(state='normal', text="🔍 DEPOLARDA ARA"))

        # Thread başlat
        thread = threading.Thread(target=arama_thread, daemon=True)
        thread.start()

    def _depo_arama_tamamlandi(self, toplam):
        """Depo araması tamamlandığında"""
        self.depo_durum_label.config(text=f"✓ {toplam} ürün arandı")
        self._kesin_liste_guncelle()

    def _min_stok_analiz_ac(self):
        """Minimum Stok Analizi penceresini aç"""
        from min_stok_analiz import tum_ilaclari_analiz_et, toplu_min_stok_guncelle

        # Yeni pencere
        analiz_win = tk.Toplevel(self.parent)
        analiz_win.title("Minimum Stok Analizi - Tüm İlaçlar")
        analiz_win.geometry("1400x700")
        analiz_win.configure(bg='#ECEFF1')

        # Üst panel - Parametreler
        param_frame = tk.Frame(analiz_win, bg='#E3F2FD', relief='raised', bd=1)
        param_frame.pack(fill=tk.X, padx=10, pady=10)

        # Hareket Süresi - TURUNCU VURGULU (en önemli parametre)
        hareket_frame = tk.Frame(param_frame, bg='#FFE0B2', relief='ridge', bd=1, padx=5, pady=2)
        hareket_frame.pack(side=tk.LEFT, padx=(10, 15))
        tk.Label(hareket_frame, text="Hareket Süresi:", font=('Arial', 10, 'bold'),
                bg='#FFE0B2', fg='#E65100').pack(side=tk.LEFT, padx=(5, 5))
        hareket_yili_var = tk.IntVar(value=3)
        hareket_combo = ttk.Combobox(hareket_frame, textvariable=hareket_yili_var, width=4, state='readonly')
        hareket_combo['values'] = [1, 2, 3, 5, 10]
        hareket_combo.pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(hareket_frame, text="yıl", font=('Arial', 10), bg='#FFE0B2').pack(side=tk.LEFT, padx=(0, 5))

        # Analiz dönemi
        tk.Label(param_frame, text="Analiz Dönemi:", font=('Arial', 10, 'bold'),
                bg='#E3F2FD').pack(side=tk.LEFT, padx=(10, 5))
        ay_var = tk.IntVar(value=12)
        ay_combo = ttk.Combobox(param_frame, textvariable=ay_var, width=5, state='readonly')
        ay_combo['values'] = [6, 12, 18, 24]
        ay_combo.pack(side=tk.LEFT, padx=(0, 5))
        tk.Label(param_frame, text="ay", font=('Arial', 10), bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 15))

        # Kar marjı
        tk.Label(param_frame, text="Kar Marjı %:", font=('Arial', 10, 'bold'),
                bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 5))
        kar_var = tk.DoubleVar(value=22)
        ttk.Entry(param_frame, textvariable=kar_var, width=5).pack(side=tk.LEFT, padx=(0, 15))

        # Faiz
        tk.Label(param_frame, text="Mevduat Faizi %:", font=('Arial', 10, 'bold'),
                bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 5))
        faiz_var = tk.DoubleVar(value=self.mevduat_faizi.get())
        ttk.Entry(param_frame, textvariable=faiz_var, width=5).pack(side=tk.LEFT, padx=(0, 20))

        # Sadece stoklu
        stoklu_var = tk.BooleanVar(value=True)
        tk.Checkbutton(param_frame, text="Sadece Stoklu İlaçlar", variable=stoklu_var,
                      bg='#E3F2FD', font=('Arial', 10)).pack(side=tk.LEFT, padx=(0, 20))

        # Durum etiketi
        durum_label = tk.Label(param_frame, text="Hareket süresi içinde stok/satış/alış hareketi olan ilaçlar analiz edilir",
                              font=('Arial', 9), bg='#E3F2FD', fg='#666')
        durum_label.pack(side=tk.RIGHT, padx=10)

        # Tablo çerçevesi
        tablo_frame = tk.Frame(analiz_win, bg='#ECEFF1')
        tablo_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Treeview
        columns = ('adi', 'stok', 'mevcut_min', 'aylik', 'talep', 'parti', 'cv', 'adi_col', 'sinif',
                  'min_bil', 'min_fin', 'min_oner')
        tree = ttk.Treeview(tablo_frame, columns=columns, show='headings', height=20)

        tree.heading('adi', text='İlaç Adı')
        tree.heading('stok', text='Stok')
        tree.heading('mevcut_min', text='Mevcut Min')
        tree.heading('aylik', text='Aylık Ort')
        tree.heading('talep', text='Talep Sayısı')
        tree.heading('parti', text='Ort Parti')
        tree.heading('cv', text='CV')
        tree.heading('adi_col', text='ADI (gün)')
        tree.heading('sinif', text='Sınıf')
        tree.heading('min_bil', text='Min (Bilim)')
        tree.heading('min_fin', text='Min (Finans)')
        tree.heading('min_oner', text='ÖNERİLEN')

        tree.column('adi', width=250, anchor='w')
        tree.column('stok', width=60, anchor='center')
        tree.column('mevcut_min', width=70, anchor='center')
        tree.column('aylik', width=70, anchor='center')
        tree.column('talep', width=70, anchor='center')
        tree.column('parti', width=70, anchor='center')
        tree.column('cv', width=50, anchor='center')
        tree.column('adi_col', width=70, anchor='center')
        tree.column('sinif', width=100, anchor='center')
        tree.column('min_bil', width=70, anchor='center')
        tree.column('min_fin', width=70, anchor='center')
        tree.column('min_oner', width=80, anchor='center')

        # Tag renkleri
        tree.tag_configure('degisecek', background='#FFF9C4')  # Sarı - değişecek
        tree.tag_configure('artacak', background='#FFCDD2')    # Kırmızı - artacak
        tree.tag_configure('azalacak', background='#C8E6C9')   # Yeşil - azalacak

        # Scrollbar
        scrollbar_y = ttk.Scrollbar(tablo_frame, orient='vertical', command=tree.yview)
        scrollbar_x = ttk.Scrollbar(tablo_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        # Alt butonlar
        btn_frame = tk.Frame(analiz_win, bg='#ECEFF1')
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        # Analiz sonuçlarını sakla
        analiz_sonuclari = []

        def analiz_yap():
            nonlocal analiz_sonuclari

            if not self.db:
                messagebox.showerror("Hata", "Veritabanı bağlantısı yok!")
                return

            durum_label.config(text="Analiz yapılıyor...")
            analiz_win.update()

            def progress_cb(current, total):
                durum_label.config(text=f"Analiz: {current}/{total} ilaç...")
                analiz_win.update()

            try:
                analiz_sonuclari = tum_ilaclari_analiz_et(
                    self.db,
                    ay_sayisi=ay_var.get(),
                    kar_marji=kar_var.get() / 100,
                    yillik_faiz=faiz_var.get() / 100,
                    sadece_stoklu=stoklu_var.get(),
                    hareket_yili=hareket_yili_var.get(),
                    progress_callback=progress_cb
                )

                # Tabloyu temizle ve doldur
                for item in tree.get_children():
                    tree.delete(item)

                degisecek_sayisi = 0
                for s in analiz_sonuclari:
                    mevcut = s['MevcutMin']
                    onerilen = s['MinOnerilen']

                    # Tag belirle
                    if mevcut != onerilen:
                        degisecek_sayisi += 1
                        if onerilen > mevcut:
                            tag = 'artacak'
                        else:
                            tag = 'azalacak'
                    else:
                        tag = ''

                    tree.insert('', 'end', values=(
                        s['UrunAdi'][:40],
                        s['Stok'],
                        mevcut,
                        s['AylikOrt'],
                        s['TalepSayisi'],
                        s['OrtParti'],
                        s['CV'],
                        s['ADI'],
                        s['Sinif'],
                        s['MinBilimsel'],
                        s['MinFinansal'],
                        s['MinOnerilen']
                    ), tags=(tag,))

                hareket_bilgi = f"Son {hareket_yili_var.get()} yıl hareket görmüş"
                durum_label.config(text=f"✓ {len(analiz_sonuclari)} ilaç ({hareket_bilgi}), {degisecek_sayisi} değişecek")

            except Exception as e:
                messagebox.showerror("Hata", f"Analiz hatası: {e}")
                durum_label.config(text=f"Hata: {e}")

        def tumu_uygula():
            """Analiz sonuçlarını yerel SQLite veritabanına kaydet"""
            if not analiz_sonuclari:
                messagebox.showwarning("Uyarı", "Önce analiz yapın!")
                return

            # Değişecekleri bul
            guncellemeler = []
            for s in analiz_sonuclari:
                if s['MevcutMin'] != s['MinOnerilen']:
                    guncellemeler.append(s)

            if not guncellemeler:
                messagebox.showinfo("Bilgi", "Değişiklik gerektiren ilaç yok!")
                return

            cevap = messagebox.askyesno(
                "Yerel Tabloya Kaydet",
                f"{len(guncellemeler)} ilacın minimum stok analizi yerel veritabanına kaydedilecek.\n\n"
                "NOT: Bu işlem Botanik EOS'a YAZMAZ!\n"
                "Sadece siparis_calismalari.db dosyasına kaydeder.\n\n"
                "Devam edilsin mi?"
            )

            if not cevap:
                return

            durum_label.config(text="Kaydediliyor...")
            analiz_win.update()

            def progress_cb(current, total):
                durum_label.config(text=f"Kaydediliyor: {current}/{total}...")
                analiz_win.update()

            try:
                from siparis_db import get_siparis_db
                siparis_db = get_siparis_db()
                basarili, hata = siparis_db.min_stok_toplu_kaydet(guncellemeler, progress_cb)
                durum_label.config(text=f"✓ {basarili} kaydedildi, {hata} hata")

                if hata == 0:
                    messagebox.showinfo(
                        "Başarılı",
                        f"{basarili} ilacın minimum stok analizi yerel tabloya kaydedildi!\n\n"
                        "Konum: siparis_calismalari.db"
                    )
            except Exception as e:
                messagebox.showerror("Hata", f"Kaydetme hatası: {e}")
                durum_label.config(text=f"Hata: {e}")

        tk.Button(
            btn_frame, text="ANALİZ YAP", command=analiz_yap,
            bg='#1976D2', fg='white', font=('Arial', 11, 'bold'),
            relief='raised', bd=2, padx=20, pady=5
        ).pack(side=tk.LEFT, padx=(0, 20))

        tk.Button(
            btn_frame, text="TABLOYA KAYDET (Yerel DB)", command=tumu_uygula,
            bg='#4CAF50', fg='white', font=('Arial', 11, 'bold'),
            relief='raised', bd=2, padx=20, pady=5
        ).pack(side=tk.LEFT, padx=(0, 20))

        # Açıklama
        tk.Label(
            btn_frame,
            text="Sarı: Değişecek | Kırmızı: Artacak | Yeşil: Azalacak",
            font=('Arial', 9), bg='#ECEFF1', fg='#666'
        ).pack(side=tk.RIGHT, padx=10)

    def excel_aktar(self):
        """Ana tabloyu Excel'e aktar"""
        if not self.tum_veriler:
            messagebox.showwarning("Uyarı", "Aktarılacak veri yok!")
            return

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            title="Excel Olarak Kaydet"
        )

        if not dosya_yolu:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sipariş Analizi"

            # Başlıklar
            headers = [c[1] for c in self.aktif_sutunlar]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h).font = Font(bold=True)

            # Veriler
            for row_idx, veri in enumerate(self.tum_veriler, 2):
                col = 1
                for col_id, _, _ in self.aktif_sutunlar:
                    if col_id == 'Tur':
                        ws.cell(row=row_idx, column=col, value='')
                    else:
                        ws.cell(row=row_idx, column=col, value=veri.get(col_id, ''))
                    col += 1

            wb.save(dosya_yolu)
            messagebox.showinfo("Başarılı", f"Veriler aktarıldı:\n{dosya_yolu}")
        except Exception as e:
            messagebox.showerror("Hata", f"Aktarım hatası: {e}")


# Test
if __name__ == "__main__":
    root = tk.Tk()
    app = SiparisVermeGUI(root)
    root.mainloop()
