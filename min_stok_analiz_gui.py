"""
Minimum Stok Analizi GUI Modulu
Donem bazli frekans, finansal basabas ve ROP yontemleriyle minimum stok hesaplama
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import logging
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from min_stok_analiz import (
    tum_ilaclari_analiz_et,
    toplu_min_stok_guncelle,
    basabas_noktasi_hesapla
)
from siparis_db import get_siparis_db

logger = logging.getLogger(__name__)

TR_AY_KISA = ['Oca', 'Sub', 'Mar', 'Nis', 'May', 'Haz',
               'Tem', 'Agu', 'Eyl', 'Eki', 'Kas', 'Ara']

TEMEL_SUTUNLAR = ('urun_id', 'barkod', 'adi', 'tip', 'stok', 'mevcut_min',
                  'aylik', 'talep', 'parti', 'cv', 'adi_col', 'sinif',
                  'min_bil', 'min_fin', 'min_oner')

TEMEL_BASLIK = {
    'urun_id':   ('Urun ID',      60, 'center'),
    'barkod':    ('Barkod',      120, 'w'),
    'adi':       ('Ilac Adi',    260, 'w'),
    'tip':       ('Urun Tipi',    90, 'center'),
    'stok':      ('Stok',         55, 'center'),
    'mevcut_min':('Mevcut Min',   70, 'center'),
    'aylik':     ('Aylik Ort',    65, 'center'),
    'talep':     ('Talep Sayisi', 70, 'center'),
    'parti':     ('Ort Parti',    65, 'center'),
    'cv':        ('CV',           45, 'center'),
    'adi_col':   ('ADI (gun)',    65, 'center'),
    'sinif':     ('Sinif',        95, 'center'),
    'min_bil':   ('Min (Bilim)',  70, 'center'),
    'min_fin':   ('Min (Finans)', 70, 'center'),
    'min_oner':  ('ONERILEN',     75, 'center'),
}

SAYISAL_OLMAYAN = {'barkod', 'adi', 'tip', 'sinif'}


class MinStokAnalizGUI:
    """Minimum Stok Analizi Ana Penceresi"""

    def __init__(self, parent, ana_menu_callback=None):
        self.parent = parent
        self.ana_menu_callback = ana_menu_callback

        self.parent.title("Minimum Stok Analizi")
        self.parent.geometry("1700x900")
        self.parent.configure(bg='#ECEFF1')

        self.db = None
        self._db_baglan()

        self.analiz_sonuclari = []
        self.analiz_satirlari = []   # filtre/siralama kaynagi
        self.aktif_filtreler = {}
        self.siralama_durum = {'col': None, 'yon': 'asc'}
        self.sutun_temel_baslik = {}
        self.taban_durum = {'metin': ''}

        # Degiskenler
        self.ay_var         = tk.IntVar(value=12)
        self.kar_marji      = tk.DoubleVar(value=22)
        self.yillik_faiz    = tk.DoubleVar(value=40)
        self.sadece_stoklu  = tk.BooleanVar(value=False)
        self.hareket_yili   = tk.IntVar(value=2)
        self.hesaplama_modu = tk.StringVar(value='frekans')
        self.servis_var     = tk.DoubleVar(value=95.0)
        self.tedarik_var    = tk.IntVar(value=0)
        self.inceleme_var   = tk.IntVar(value=1)

        self._arayuz_olustur()

    def _db_baglan(self):
        try:
            from botanik_db import BotanikDB
            self.db = BotanikDB()
            if not self.db.baglan():
                self.db = None
        except Exception as e:
            logger.error(f"DB baglanti hatasi: {e}")
            self.db = None

    # ── Arayuz ──────────────────────────────────────────────────────────────

    def _arayuz_olustur(self):
        header = tk.Frame(self.parent, bg='#1976D2', height=45)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="Minimum Stok Analizi",
                 bg='#1976D2', fg='white', font=('Arial', 13, 'bold')
                 ).pack(side=tk.LEFT, padx=20, pady=8)

        self._parametre_panel_olustur()
        self._rop_panel_olustur()
        self._tablo_olustur()
        self._buton_panel_olustur()

    def _parametre_panel_olustur(self):
        pf = tk.Frame(self.parent, bg='#E3F2FD', relief='ridge', bd=2)
        pf.pack(fill=tk.X, padx=10, pady=(5, 0))

        # Satir 1 — parametreler
        r1 = tk.Frame(pf, bg='#E3F2FD')
        r1.pack(fill=tk.X, padx=10, pady=(8, 2))

        tk.Label(r1, text="Analiz Donemi:", font=('Arial', 10, 'bold'), bg='#E3F2FD').pack(side=tk.LEFT)
        self.ay_combo = ttk.Combobox(r1, textvariable=self.ay_var, width=5, state='readonly')
        self.ay_combo['values'] = [6, 12, 18, 24]
        self.ay_combo.pack(side=tk.LEFT, padx=(4, 2))
        tk.Label(r1, text="ay", font=('Arial', 10), bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(r1, text="Hareket Suresi:", font=('Arial', 10, 'bold'),
                 bg='#FFECB3', fg='#E65100').pack(side=tk.LEFT)
        hc = ttk.Combobox(r1, textvariable=self.hareket_yili, width=4, state='readonly')
        hc['values'] = [0, 1, 2, 3, 5, 10]
        hc.pack(side=tk.LEFT, padx=(4, 2))
        tk.Label(r1, text="yil", font=('Arial', 10), bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(r1, text="Kar Marji:", font=('Arial', 10, 'bold'), bg='#E3F2FD').pack(side=tk.LEFT)
        ttk.Entry(r1, textvariable=self.kar_marji, width=5).pack(side=tk.LEFT, padx=(4, 2))
        tk.Label(r1, text="%", font=('Arial', 10), bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 15))

        tk.Label(r1, text="Faiz:", font=('Arial', 10, 'bold'), bg='#E3F2FD').pack(side=tk.LEFT)
        ttk.Entry(r1, textvariable=self.yillik_faiz, width=5).pack(side=tk.LEFT, padx=(4, 2))
        tk.Label(r1, text="%", font=('Arial', 10), bg='#E3F2FD').pack(side=tk.LEFT, padx=(0, 15))

        self.basabas_label = tk.Label(r1, text="Basabas: --", font=('Arial', 9, 'bold'),
                                      bg='#FFF9C4', fg='#F57F17', padx=6, pady=3)
        self.basabas_label.pack(side=tk.LEFT, padx=(0, 20))

        tk.Checkbutton(r1, text="Sadece Stoklu", variable=self.sadece_stoklu,
                       bg='#E3F2FD', font=('Arial', 9)).pack(side=tk.LEFT, padx=(0, 20))

        tk.Button(r1, text="ANALIZ YAP", command=self._analiz_yap,
                  bg='#1976D2', fg='white', font=('Arial', 11, 'bold'),
                  relief='raised', bd=2, padx=20, pady=4).pack(side=tk.LEFT, padx=(0, 10))

        # Satir 2 — mod + durum
        r2 = tk.Frame(pf, bg='#E3F2FD')
        r2.pack(fill=tk.X, padx=10, pady=(2, 6))

        mod_f = tk.Frame(r2, bg='#E8EAF6', relief='ridge', bd=1, padx=5, pady=2)
        mod_f.pack(side=tk.LEFT)
        tk.Label(mod_f, text="Hesaplama:", font=('Arial', 9, 'bold'), bg='#E8EAF6').pack(side=tk.LEFT, padx=(2, 5))
        tk.Radiobutton(mod_f, text="Frekans Bazli", variable=self.hesaplama_modu,
                       value='frekans', bg='#E8EAF6', font=('Arial', 9)).pack(side=tk.LEFT, padx=(0, 5))
        tk.Radiobutton(mod_f, text="Finansal Basabas", variable=self.hesaplama_modu,
                       value='finansal', bg='#E8EAF6', font=('Arial', 9)).pack(side=tk.LEFT, padx=(0, 5))
        tk.Radiobutton(mod_f, text="ROP (Bilimsel)", variable=self.hesaplama_modu,
                       value='rop', bg='#E8EAF6', font=('Arial', 9)).pack(side=tk.LEFT, padx=(0, 2))

        self.durum_label = tk.Label(r2, text="Hazir", font=('Arial', 9), bg='#E3F2FD', fg='#666')
        self.durum_label.pack(side=tk.RIGHT, padx=10)

        self.kar_marji.trace_add('write', self._basabas_guncelle)
        self.yillik_faiz.trace_add('write', self._basabas_guncelle)
        self.hesaplama_modu.trace_add('write', self._mod_degisti)
        self._basabas_guncelle()

    def _rop_panel_olustur(self):
        self.rop_frame = tk.Frame(self.parent, bg='#E8F5E9', relief='ridge', bd=1)
        tk.Label(self.rop_frame, text="ROP Parametreleri:", font=('Arial', 9, 'bold'),
                 bg='#E8F5E9', fg='#2E7D32').pack(side=tk.LEFT, padx=(10, 10), pady=5)
        tk.Label(self.rop_frame, text="Servis Seviyesi:", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(5, 3))
        sc = ttk.Combobox(self.rop_frame, textvariable=self.servis_var, width=5, state='readonly')
        sc['values'] = [90.0, 95.0, 97.5, 99.0]
        sc.pack(side=tk.LEFT, padx=(0, 3), pady=5)
        tk.Label(self.rop_frame, text="%", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(self.rop_frame, text="Tedarik Suresi:", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(5, 3))
        ttk.Spinbox(self.rop_frame, from_=0, to=30, textvariable=self.tedarik_var, width=3).pack(side=tk.LEFT, padx=(0, 3), pady=5)
        tk.Label(self.rop_frame, text="gun", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(self.rop_frame, text="Inceleme Periyodu:", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(5, 3))
        ttk.Spinbox(self.rop_frame, from_=1, to=30, textvariable=self.inceleme_var, width=3).pack(side=tk.LEFT, padx=(0, 3), pady=5)
        tk.Label(self.rop_frame, text="gun", font=('Arial', 9), bg='#E8F5E9').pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(self.rop_frame, text="PP=Tedarik+Inceleme | SS=Z*σ*√PP | Min=⌈d*PP+SS⌉",
                 font=('Arial', 8), bg='#E8F5E9', fg='#666').pack(side=tk.RIGHT, padx=10, pady=5)

    def _mod_degisti(self, *args):
        modu = self.hesaplama_modu.get()
        if modu == 'rop':
            self.rop_frame.pack(fill=tk.X, padx=10, pady=(0, 5), before=self.tablo_frame)
        else:
            self.rop_frame.pack_forget()
        if modu == 'finansal':
            self.ay_combo.config(state='disabled')
            self.ay_var.set(24)
        else:
            self.ay_combo.config(state='readonly')

    def _basabas_guncelle(self, *args):
        try:
            kar  = self.kar_marji.get() / 100
            faiz = self.yillik_faiz.get() / 100
            bb   = basabas_noktasi_hesapla(kar, faiz)
            self.basabas_label.config(text=f"Basabas: {bb:.1f} ay")
        except Exception:
            self.basabas_label.config(text="Basabas: --")

    # ── Tablo ───────────────────────────────────────────────────────────────

    def _tablo_olustur(self):
        self.tablo_frame = tk.Frame(self.parent, bg='#ECEFF1')
        self.tablo_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.tree = ttk.Treeview(self.tablo_frame, columns=TEMEL_SUTUNLAR, show='headings', height=22)

        self.tree.tag_configure('degisecek', background='#FFF9C4')
        self.tree.tag_configure('artacak',   background='#FFCDD2')
        self.tree.tag_configure('azalacak',  background='#C8E6C9')

        sb_y = ttk.Scrollbar(self.tablo_frame, orient='vertical',   command=self.tree.yview)
        sb_x = ttk.Scrollbar(self.tablo_frame, orient='horizontal',  command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        sb_x.pack(side=tk.BOTTOM, fill=tk.X)
        sb_y.pack(side=tk.RIGHT,  fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._sutunlari_kur(self.ay_var.get())

    def _sutunlari_kur(self, analiz_ay):
        ay_sutunlar = tuple(f'ay_{i}' for i in range(analiz_ay))
        self.tree.configure(columns=TEMEL_SUTUNLAR + ay_sutunlar)
        self.sutun_temel_baslik.clear()

        for col, (etiket, gen, hiza) in TEMEL_BASLIK.items():
            self.sutun_temel_baslik[col] = etiket
            self.tree.heading(col, text=etiket, command=lambda c=col: self._baslik_tikla(c))
            self.tree.column(col, width=gen, anchor=hiza)

        bugun = datetime.now()
        for i in range(analiz_ay):
            col = f'ay_{i}'
            ay_tarihi = bugun - relativedelta(months=i)
            etiket = f"{TR_AY_KISA[ay_tarihi.month - 1]}{str(ay_tarihi.year)[2:]}"
            self.sutun_temel_baslik[col] = etiket
            self.tree.heading(col, text=etiket, command=lambda c=col: self._baslik_tikla(c))
            self.tree.column(col, width=48, anchor='center')

    # ── Filtre / Siralama ───────────────────────────────────────────────────

    def _hucre_sayi(self, v):
        try:
            return float(str(v).replace(',', '.')) if v not in (None, '') else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _basliklari_guncelle(self):
        for col, temel in self.sutun_temel_baslik.items():
            ek = ''
            if col in self.aktif_filtreler:
                ek += ' ⛛'
            if self.siralama_durum['col'] == col:
                ek += ' ▲' if self.siralama_durum['yon'] == 'asc' else ' ▼'
            try:
                self.tree.heading(col, text=temel + ek)
            except tk.TclError:
                pass

    def _satir_gecer(self, values, cols):
        for col, spec in self.aktif_filtreler.items():
            if col not in cols:
                continue
            idx  = cols.index(col)
            val  = values[idx] if idx < len(values) else ''
            sval = '' if val is None else str(val)
            izinli = spec.get('izinli')
            if izinli is not None and sval not in izinli:
                return False
            mn, mx = spec.get('min'), spec.get('max')
            if mn is not None or mx is not None:
                num = self._hucre_sayi(val)
                if mn is not None and num < mn:
                    return False
                if mx is not None and num > mx:
                    return False
        return True

    def _tabloyu_yenile(self):
        cols    = list(self.tree['columns'])
        satirlar = [r for r in self.analiz_satirlari if self._satir_gecer(r['values'], cols)]
        scol = self.siralama_durum['col']
        if scol and scol in cols:
            idx     = cols.index(scol)
            sayisal = scol not in SAYISAL_OLMAYAN
            def anahtar(r):
                v = r['values'][idx] if idx < len(r['values']) else ''
                return self._hucre_sayi(v) if sayisal else str(v).lower()
            satirlar = sorted(satirlar, key=anahtar,
                              reverse=(self.siralama_durum['yon'] == 'desc'))
        for item in self.tree.get_children():
            self.tree.delete(item)
        for r in satirlar:
            self.tree.insert('', 'end', values=r['values'], tags=(r['tag'],) if r['tag'] else ())
        self._basliklari_guncelle()
        toplam = len(self.analiz_satirlari)
        self.kayit_label.config(text=f"{len(satirlar)} / {toplam} kayit")

    def _baslik_tikla(self, col):
        if not self.analiz_satirlari:
            return
        self._filtre_popup(col, self.tree.winfo_pointerx(), self.tree.winfo_pointery())

    def _filtre_popup(self, col, x, y):
        cols = list(self.tree['columns'])
        if col not in cols:
            return
        idx     = cols.index(col)
        sayisal = col not in SAYISAL_OLMAYAN
        baslik  = self.sutun_temel_baslik.get(col, col)

        ham = []
        for r in self.analiz_satirlari:
            v = r['values'][idx] if idx < len(r['values']) else ''
            ham.append('' if v is None else str(v))
        if sayisal:
            benzersiz = sorted(set(ham), key=lambda s: self._hucre_sayi(s))
        else:
            benzersiz = sorted(set(ham), key=lambda s: s.lower())

        spec    = self.aktif_filtreler.get(col, {})
        izinli0 = spec.get('izinli')
        secili  = set(benzersiz) if izinli0 is None else set(izinli0)

        pop = tk.Toplevel(self.parent)
        pop.title(f"{baslik} - Filtre / Sirala")
        pop.transient(self.parent)
        pop.configure(bg='#FAFAFA')
        pop.geometry(f"270x460+{x}+{y}")
        pop.resizable(False, True)

        # Siralama butonlari
        sf = tk.Frame(pop, bg='#FAFAFA')
        sf.pack(fill=tk.X, padx=8, pady=(8, 2))
        def sirala(yon):
            self.siralama_durum['col'] = col
            self.siralama_durum['yon'] = yon
            self._tabloyu_yenile()
            pop.destroy()
        tk.Button(sf, text='▲ A→Z (Artan)', command=lambda: sirala('asc'),
                  bg='#E3F2FD', relief='groove', font=('Arial', 8)
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        tk.Button(sf, text='▼ Z→A (Azalan)', command=lambda: sirala('desc'),
                  bg='#E3F2FD', relief='groove', font=('Arial', 8)
                  ).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)

        ttk.Separator(pop, orient='horizontal').pack(fill=tk.X, padx=8, pady=4)

        # Sayi filtresi (yalnizca sayisal sutunlar)
        min_var = tk.StringVar(value='' if spec.get('min') is None else str(spec['min']))
        max_var = tk.StringVar(value='' if spec.get('max') is None else str(spec['max']))
        if sayisal:
            nf = tk.LabelFrame(pop, text='Sayi filtresi', bg='#FAFAFA', font=('Arial', 8, 'bold'))
            nf.pack(fill=tk.X, padx=8, pady=(0, 4))
            row = tk.Frame(nf, bg='#FAFAFA')
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text='Min:', bg='#FAFAFA').pack(side=tk.LEFT)
            ttk.Entry(row, textvariable=min_var, width=8).pack(side=tk.LEFT, padx=(2, 8))
            tk.Label(row, text='Max:', bg='#FAFAFA').pack(side=tk.LEFT)
            ttk.Entry(row, textvariable=max_var, width=8).pack(side=tk.LEFT, padx=2)

        # Degerler: arama kutusu + checkbox listesi (Excel AutoFilter stili)
        vf = tk.LabelFrame(pop, text='Filtrele', bg='#FAFAFA', font=('Arial', 8, 'bold'))
        vf.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 4))

        arama_var = tk.StringVar()
        arama_e = ttk.Entry(vf, textvariable=arama_var)
        arama_e.pack(fill=tk.X, padx=4, pady=(4, 2))

        # Canvas + scrollbar icin checkbox frame
        cb_outer = tk.Frame(vf, bg='#FAFAFA')
        cb_outer.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        cb_canvas = tk.Canvas(cb_outer, bg='#FAFAFA', highlightthickness=0)
        cb_sb = ttk.Scrollbar(cb_outer, orient='vertical', command=cb_canvas.yview)
        cb_canvas.configure(yscrollcommand=cb_sb.set)
        cb_sb.pack(side=tk.RIGHT, fill=tk.Y)
        cb_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cb_frame = tk.Frame(cb_canvas, bg='#FAFAFA')
        cb_frame_id = cb_canvas.create_window((0, 0), window=cb_frame, anchor='nw')
        def _cb_resize(e):
            cb_canvas.configure(scrollregion=cb_canvas.bbox('all'))
            cb_canvas.itemconfig(cb_frame_id, width=cb_canvas.winfo_width())
        cb_frame.bind('<Configure>', _cb_resize)
        def _mousewheel(e):
            cb_canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        cb_canvas.bind('<MouseWheel>', _mousewheel)
        cb_frame.bind('<MouseWheel>', _mousewheel)

        # Checkbox degiskenleri: value → BooleanVar
        cb_vars = {}
        tumu_var = tk.BooleanVar()
        gosterilen = []

        def _tumu_guncelle():
            tum_secili = all(cb_vars[v].get() for v in gosterilen if v in cb_vars)
            tumu_var.set(tum_secili)

        def _tumu_toggle():
            val = tumu_var.get()
            for v in gosterilen:
                if v in cb_vars:
                    cb_vars[v].set(val)
                    if val:
                        secili.add(v)
                    else:
                        secili.discard(v)

        def _cb_toggle(v):
            if cb_vars[v].get():
                secili.add(v)
            else:
                secili.discard(v)
            _tumu_guncelle()

        tumu_cb = tk.Checkbutton(
            cb_frame, text='(Tümünü Seç)', variable=tumu_var,
            command=_tumu_toggle, bg='#FAFAFA', anchor='w',
            font=('Arial', 9, 'bold'), fg='#1565C0',
            activebackground='#E3F2FD')
        tumu_cb.pack(fill=tk.X, padx=4, pady=(2, 0))
        ttk.Separator(cb_frame, orient='horizontal').pack(fill=tk.X, padx=4, pady=2)

        def doldur(*_):
            nonlocal gosterilen
            f = arama_var.get().lower()
            gosterilen = [v for v in benzersiz if f in v.lower()]
            for w in list(cb_frame.winfo_children())[2:]:
                w.destroy()
            for v in gosterilen:
                if v not in cb_vars:
                    cb_vars[v] = tk.BooleanVar(value=(v in secili))
                else:
                    cb_vars[v].set(v in secili)
                label = v if v != '' else '(bos)'
                cb = tk.Checkbutton(
                    cb_frame, text=label, variable=cb_vars[v],
                    command=lambda _v=v: _cb_toggle(_v),
                    bg='#FAFAFA', anchor='w', font=('Arial', 9),
                    activebackground='#E3F2FD')
                cb.pack(fill=tk.X, padx=(16, 4), pady=1)
                cb.bind('<MouseWheel>', _mousewheel)
            _tumu_guncelle()
            cb_canvas.update_idletasks()
            cb_canvas.configure(scrollregion=cb_canvas.bbox('all'))

        arama_var.trace_add('write', doldur)
        doldur()

        # Alt butonlar
        bf = tk.Frame(pop, bg='#FAFAFA')
        bf.pack(fill=tk.X, padx=8, pady=6)
        def uygula():
            yeni = {}
            if set(secili) != set(benzersiz):
                yeni['izinli'] = set(secili)
            def parse(s):
                s = s.strip().replace(',', '.')
                try:   return float(s) if s != '' else None
                except ValueError: return None
            mn, mx = parse(min_var.get()), parse(max_var.get())
            if mn is not None: yeni['min'] = mn
            if mx is not None: yeni['max'] = mx
            if yeni:
                self.aktif_filtreler[col] = yeni
            else:
                self.aktif_filtreler.pop(col, None)
            self._tabloyu_yenile()
            pop.destroy()
        def temizle():
            self.aktif_filtreler.pop(col, None)
            self._tabloyu_yenile()
            pop.destroy()
        tk.Button(bf, text='Uygula',  command=uygula,  bg='#4CAF50', fg='white',
                  font=('Arial', 9, 'bold')).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        tk.Button(bf, text='Temizle', command=temizle, bg='#FFB74D',
                  font=('Arial', 9)).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        tk.Button(bf, text='Iptal',   command=pop.destroy,
                  font=('Arial', 9)).pack(side=tk.LEFT, padx=2)
        pop.grab_set()

    def _filtreleri_temizle(self):
        self.aktif_filtreler.clear()
        self.siralama_durum['col'] = None
        self.siralama_durum['yon'] = 'asc'
        self._tabloyu_yenile()

    # ── Alt butonlar ────────────────────────────────────────────────────────

    def _buton_panel_olustur(self):
        bf = tk.Frame(self.parent, bg='#ECEFF1')
        bf.pack(fill=tk.X, padx=10, pady=8)

        tk.Button(bf, text="TABLOYA KAYDET (Yerel DB)", command=self._tabloya_kaydet,
                  bg='#4CAF50', fg='white', font=('Arial', 11, 'bold'),
                  relief='raised', bd=2, padx=20, pady=5).pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(bf, text="EXCEL'E AKTAR", command=self._excel_aktar,
                  bg='#FF6F00', fg='white', font=('Arial', 11, 'bold'),
                  relief='raised', bd=2, padx=20, pady=5).pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(bf, text="Filtreleri Temizle", command=self._filtreleri_temizle,
                  bg='#FFB74D', font=('Arial', 10), relief='raised', bd=2,
                  padx=12, pady=5).pack(side=tk.LEFT, padx=(0, 10))

        tk.Button(bf, text="Ozet Bilgi", command=self._ozet_goster,
                  bg='#5C6BC0', fg='white', font=('Arial', 10, 'bold'),
                  relief='raised', bd=2, padx=12, pady=5).pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(bf, text="NOT: Botanik EOS'a yazma YASAK — sadece yerel DB veya Excel.",
                 font=('Arial', 9, 'italic'), bg='#ECEFF1', fg='#1565C0').pack(side=tk.LEFT)

        self.kayit_label = tk.Label(bf, text="", font=('Arial', 10, 'bold'),
                                    bg='#ECEFF1', fg='#37474F')
        self.kayit_label.pack(side=tk.RIGHT, padx=10)

    # ── Analiz ──────────────────────────────────────────────────────────────

    def _analiz_yap(self):
        if not self.db:
            messagebox.showerror("Hata", "Veritabani baglantisi yok!")
            return

        self.durum_label.config(text="Analiz yapiliyor...")
        self.parent.update()

        def progress_cb(current, total):
            self.durum_label.config(text=f"Analiz: {current}/{total} ilac...")
            self.parent.update()

        def _thread():
            try:
                sonuclar = tum_ilaclari_analiz_et(
                    self.db,
                    ay_sayisi=self.ay_var.get(),
                    kar_marji=self.kar_marji.get() / 100,
                    yillik_faiz=self.yillik_faiz.get() / 100,
                    sadece_stoklu=self.sadece_stoklu.get(),
                    hareket_yili=self.hareket_yili.get(),
                    hesaplama_modu=self.hesaplama_modu.get(),
                    progress_callback=progress_cb,
                    servis_seviyesi=self.servis_var.get(),
                    tedarik_suresi=self.tedarik_var.get(),
                    inceleme_periyodu=self.inceleme_var.get()
                )
                self.analiz_sonuclari = sonuclar
                self.parent.after(0, self._tabloyu_doldur)
            except Exception as e:
                self.parent.after(0, lambda: messagebox.showerror("Hata", f"Analiz hatasi: {e}"))
                self.parent.after(0, lambda: self.durum_label.config(text="Hata!"))

        threading.Thread(target=_thread, daemon=True).start()

    def _tabloyu_doldur(self):
        modu = self.hesaplama_modu.get()
        analiz_ay = 24 if modu == 'finansal' else self.ay_var.get()

        self._sutunlari_kur(analiz_ay)

        # Mod bazli sutun basliklari
        if modu == 'rop':
            self.sutun_temel_baslik['min_bil'] = 'Em.Stok (SS)'
            self.sutun_temel_baslik['min_fin'] = 'ROP'
        elif modu == 'frekans':
            self.sutun_temel_baslik['min_bil'] = 'Min (Frekans)'
            self.sutun_temel_baslik['min_fin'] = '-'

        self.aktif_filtreler.clear()
        self.siralama_durum['col'] = None
        self.siralama_durum['yon'] = 'asc'
        self.analiz_satirlari.clear()

        degisecek = 0
        for s in self.analiz_sonuclari:
            mevcut   = s['MevcutMin']
            onerilen = s['MinOnerilen']

            if mevcut != onerilen:
                degisecek += 1
                tag = 'artacak' if onerilen > mevcut else 'azalacak'
            else:
                tag = ''

            dokum    = s.get('AylikDokum', []) or []
            ay_deger = [
                (int(round(dokum[i])) if i < len(dokum) and dokum[i] else '')
                for i in range(analiz_ay)
            ]

            self.analiz_satirlari.append({
                'values': (
                    s['UrunId'], s.get('UrunBarkodu', '') or '',
                    s['UrunAdi'], s.get('UrunTipi', '') or '',
                    s['Stok'], mevcut, s['AylikOrt'],
                    s['TalepSayisi'], s['OrtParti'], s['CV'], s['ADI'],
                    s['Sinif'], s['MinBilimsel'], s['MinFinansal'],
                    s['MinOnerilen'], *ay_deger
                ),
                'tag': tag
            })

        hareket_bilgi = f"son {self.hareket_yili.get()} yil hareket gormush" if self.hareket_yili.get() > 0 else "tum ilaclar"
        self.taban_durum['metin'] = (f"Tamamlandi: {len(self.analiz_sonuclari)} ilac ({hareket_bilgi}), "
                                      f"{degisecek} degisecek")
        self.durum_label.config(text=self.taban_durum['metin'])
        self._tabloyu_yenile()

    # ── Kaydet / Excel ──────────────────────────────────────────────────────

    def _tabloya_kaydet(self):
        if not self.analiz_sonuclari:
            messagebox.showwarning("Uyari", "Once analiz yapin!")
            return

        guncellemeler = [s for s in self.analiz_sonuclari if s['MevcutMin'] != s['MinOnerilen']]
        if not guncellemeler:
            messagebox.showinfo("Bilgi", "Degisiklik gerektiren ilac yok!")
            return

        if not messagebox.askyesno(
            "Yerel Tabloya Kaydet",
            f"{len(guncellemeler)} ilacin minimum stok analizi yerel veritabanina kaydedilecek.\n\n"
            "NOT: Bu islem Botanik EOS'a YAZMAZ!\n"
            "Sadece siparis_calismalari.db dosyasina kaydeder.\n\nDevam edilsin mi?"
        ):
            return

        self.durum_label.config(text="Kaydediliyor...")
        self.parent.update()

        def progress_cb(current, total):
            self.durum_label.config(text=f"Kaydediliyor: {current}/{total}...")
            self.parent.update()

        try:
            siparis_db = get_siparis_db()
            basarili, hata = siparis_db.min_stok_toplu_kaydet(guncellemeler, progress_cb)
            self.durum_label.config(text=f"{basarili} kaydedildi, {hata} hata")
            if hata == 0:
                messagebox.showinfo("Basarili",
                    f"{basarili} ilacin minimum stok analizi yerel tabloya kaydedildi!\n\n"
                    "Konum: siparis_calismalari.db")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme hatasi: {e}")
            self.durum_label.config(text="Kayit hatasi!")

    def _ozet_goster(self):
        if not self.analiz_sonuclari:
            messagebox.showwarning("Uyari", "Once analiz yapin!")
            return

        # Min dagılımı say
        from collections import Counter
        min_dagılım = Counter()
        sifir_ama_satisli = 0

        for s in self.analiz_sonuclari:
            m = s['MinOnerilen']
            t = s['TalepSayisi']
            if m > 0:
                min_dagılım[m] += 1
            elif t >= 1:
                sifir_ama_satisli += 1

        toplam_belirl = sum(min_dagılım.values())

        pop = tk.Toplevel(self.parent)
        pop.title("Ozet Bilgi — Minimum Stok Dagilimi")
        pop.configure(bg='#ECEFF1')
        pop.resizable(False, False)

        # Baslik
        tk.Label(pop, text="Minimum Stok Ozeti", font=('Arial', 13, 'bold'),
                 bg='#1976D2', fg='white', padx=20, pady=8).pack(fill=tk.X)

        # Tablo cercevesi
        tf = tk.Frame(pop, bg='#ECEFF1')
        tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Basliklar
        for ci, (metin, gen) in enumerate([("Min Miktar", 120), ("Ilac Adeti", 100), ("Oran (%)", 90)]):
            tk.Label(tf, text=metin, font=('Arial', 10, 'bold'),
                     bg='#1976D2', fg='white', width=gen//8, padx=8, pady=4,
                     relief='flat').grid(row=0, column=ci, padx=2, pady=(0, 4), sticky='ew')

        satirlar_sorted = sorted(min_dagılım.items())
        for ri, (deger, adet) in enumerate(satirlar_sorted, 1):
            oran = adet / len(self.analiz_sonuclari) * 100 if self.analiz_sonuclari else 0
            bg = '#E3F2FD' if ri % 2 == 0 else 'white'
            tk.Label(tf, text=str(deger), font=('Arial', 10, 'bold'),
                     bg=bg, padx=8, pady=3, anchor='center').grid(row=ri, column=0, padx=2, pady=1, sticky='ew')
            tk.Label(tf, text=str(adet), font=('Arial', 10),
                     bg=bg, padx=8, pady=3, anchor='center').grid(row=ri, column=1, padx=2, pady=1, sticky='ew')
            tk.Label(tf, text=f"%{oran:.1f}", font=('Arial', 10),
                     bg=bg, padx=8, pady=3, anchor='center').grid(row=ri, column=2, padx=2, pady=1, sticky='ew')

        # Ozet satırlar
        sf = tk.Frame(pop, bg='#E8F5E9', relief='ridge', bd=1)
        sf.pack(fill=tk.X, padx=20, pady=(0, 10))

        def satir(sf, etiket, deger, renk='#1B5E20'):
            fr = tk.Frame(sf, bg='#E8F5E9')
            fr.pack(fill=tk.X, padx=10, pady=2)
            tk.Label(fr, text=etiket, font=('Arial', 10), bg='#E8F5E9',
                     anchor='w').pack(side=tk.LEFT)
            tk.Label(fr, text=str(deger), font=('Arial', 10, 'bold'),
                     bg='#E8F5E9', fg=renk).pack(side=tk.RIGHT)

        satir(sf, "Toplam analiz edilen ilac:", len(self.analiz_sonuclari))
        satir(sf, "Minimum belirlenen ilac:", toplam_belirl, '#1B5E20')
        satir(sf, "Min=0 (en az 1 satis var ama minimum belirlenemedi):",
              sifir_ama_satisli, '#B71C1C')

        tk.Button(pop, text="Kapat", command=pop.destroy,
                  bg='#757575', fg='white', font=('Arial', 10),
                  padx=20, pady=4).pack(pady=(0, 15))

        pop.update_idletasks()
        pop.grab_set()

    def _excel_aktar(self):
        if not self.analiz_satirlari:
            messagebox.showwarning("Uyari", "Once analiz yapin!")
            return

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyasi", "*.xlsx")],
            title="Minimum Stok Analizi — Excel Olarak Kaydet",
            initialfilename=f"min_stok_analiz_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not dosya_yolu:
            return

        try:
            cols   = list(self.tree['columns'])
            baslik = [self.sutun_temel_baslik.get(c, c) for c in cols]

            wb = Workbook()
            ws = wb.active
            ws.title = "Minimum Stok Analizi"

            header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for ci, h in enumerate(baslik, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            fill_art = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            fill_aza = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

            for ri, r in enumerate(self.analiz_satirlari, 2):
                for ci, v in enumerate(r['values'], 1):
                    ws.cell(row=ri, column=ci, value=v)
                if r['tag'] == 'artacak':
                    for ci in range(1, len(r['values']) + 1):
                        ws.cell(row=ri, column=ci).fill = fill_art
                elif r['tag'] == 'azalacak':
                    for ci in range(1, len(r['values']) + 1):
                        ws.cell(row=ri, column=ci).fill = fill_aza

            # Kolon genislikleri
            kolon_gen = {
                'urun_id':    10, 'barkod': 18, 'adi': 50, 'tip': 14,
                'stok': 8, 'mevcut_min': 10, 'aylik': 10, 'talep': 12,
                'parti': 10, 'cv': 8, 'adi_col': 10, 'sinif': 14,
                'min_bil': 12, 'min_fin': 12, 'min_oner': 12,
            }
            for ci, col in enumerate(cols, 1):
                gen = kolon_gen.get(col, 12)
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = gen
            wb.save(dosya_yolu)
            messagebox.showinfo("Basarili", f"Excel dosyasi kaydedildi:\n{dosya_yolu}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel kaydetme hatasi: {e}")


def min_stok_analiz_ac(parent=None, ana_menu_callback=None):
    """Minimum Stok Analizi modulunu ac"""
    if parent is None:
        root = tk.Tk()
        MinStokAnalizGUI(root)
        root.mainloop()
    else:
        MinStokAnalizGUI(parent, ana_menu_callback)


if __name__ == "__main__":
    min_stok_analiz_ac()
