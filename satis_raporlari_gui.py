"""
Satış Raporları Modülü (Faz 1)

İki tarih arası satışları periyot bazlı (Günlük/Haftalık/Aylık/3Aylık/Yıllık)
ve kırılım bazlı (Tümü / Reçeteli / Elden / Kurum) raporlar.

Metrikler: Reçete sayısı, Satış sayısı, Kutu sayısı, TL tutarı.

Veri kaynağı: Botanik EOS (SADECE SELECT - botanik_db.BotanikDB üzerinden).
Faz 2 (nöbet) ve Faz 3 (endeks) sonraki sürümlerde eklenecek.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
import threading
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Tuple

try:
    from tkcalendar import DateEntry
    TKCALENDAR_AVAILABLE = True
except ImportError:
    TKCALENDAR_AVAILABLE = False

logger = logging.getLogger(__name__)


def _kayit_saatsiz_mi(kayit) -> bool:
    """Kayıt zamanı saat bileşeni 00:00:00 ise True (Botanik EOS'ta bazı
    kayıtlar tarih olarak yazılır → DATEDIFF sahte fark üretir). Bu durumda
    fark güvenilmez."""
    if kayit is None:
        return False
    if not hasattr(kayit, 'hour'):
        return True  # date type — saat yok demek
    return (kayit.hour == 0 and kayit.minute == 0
            and getattr(kayit, 'second', 0) == 0)


def _saniye_okunabilir(sn) -> str:
    """Saniyeyi 'Ng Xsa Ydk Zsn' okunabilir formata çevir.

    Örnekler:
      45    → '45sn'
      120   → '2dk'
      5400  → '1sa 30dk'
      90000 → '1g 1sa'
      262800→ '3g 1sa'
      -7200 → '-2sa'
    """
    if sn is None:
        return '?'
    try:
        sn = int(sn)
    except (TypeError, ValueError):
        return str(sn)
    if sn == 0:
        return '0sn'
    sgn = '-' if sn < 0 else ''
    a = abs(sn)
    gun = a // 86400
    sa = (a % 86400) // 3600
    dk = (a % 3600) // 60
    ksn = a % 60
    parts = []
    if gun:
        parts.append(f"{gun}g")
    if sa:
        parts.append(f"{sa}sa")
    if dk:
        parts.append(f"{dk}dk")
    # Saniye: sadece gün/saat yoksa göster (büyük değerlerde gürültü olmasın)
    if ksn and not (gun or sa):
        parts.append(f"{ksn}sn")
    if not parts:
        parts.append(f"{a}sn")
    return sgn + " ".join(parts)


class SatisRaporlariGUI:
    """Satış Raporları penceresi (Faz 1)."""

    PERIYOT_SECENEKLERI = [
        ('gunluk',   'Günlük'),
        ('haftalik', 'Haftalık'),
        ('aylik',    'Aylık'),
        ('3aylik',   '3 Aylık'),
        ('yillik',   'Yıllık'),
    ]

    KIRILIM_SECENEKLERI = [
        ('tumu',   'Tümü (Reçete + Elden)'),
        ('recete', 'Reçeteli (Kurum)'),
        ('elden',  'Elden (Parakende)'),
        ('kurum',  'Belirli Kurum'),
    ]

    METRIK_TANIMLARI = [
        ('recete_sayisi', 'Reçete Sayısı', 'ReceteSayisi'),
        ('satis_sayisi',  'Satış Sayısı',  'SatisSayisi'),
        ('kutu_sayisi',   'Kutu Sayısı',   'KutuSayisi'),
        ('tl_tutar',      'TL Tutar',      'TLTutar'),
    ]

    # Sütun grupları — her birinde 4 metrik: Satış Adedi / Kalem Sayısı / Kutu Adedi / TL Tutarı
    # Format: grup_kodu, etiket, renk, [(alan_kodu, baslik, db_alan, tip), ...]
    SUTUN_GRUPLARI = [
        ('receteli', 'Reçeteli', '#E3F2FD', [
            ('rec_sayi',  'Satış Adedi',  'ReceteliSayisi', 'int'),
            ('rec_kalem', 'Kalem Sayısı', 'ReceteliKalem',  'int'),
            ('rec_kutu',  'Kutu Adedi',   'ReceteliKutu',   'float'),
            ('rec_tl',    'TL Tutarı',    'ReceteliTL',     'tl'),
        ]),
        ('elden', 'Elden', '#FFF3E0', [
            ('eld_sayi',  'Satış Adedi',  'EldenSayisi', 'int'),
            ('eld_kalem', 'Kalem Sayısı', 'EldenKalem',  'int'),
            ('eld_kutu',  'Kutu Adedi',   'EldenKutu',   'float'),
            ('eld_tl',    'TL Tutarı',    'EldenTL',     'tl'),
        ]),
        ('genel', 'Genel Toplam', '#F1F8E9', [
            ('gen_sayi',  'Satış Adedi',  'SatisSayisi', 'int'),
            ('gen_kalem', 'Kalem Sayısı', 'KalemSayisi', 'int'),
            ('gen_kutu',  'Kutu Adedi',   'KutuSayisi',  'float'),
            ('gen_tl',    'TL Tutarı',    'TLTutar',     'tl'),
        ]),
    ]

    def __init__(self, root, ana_menu_callback=None):
        self.root = root
        self.ana_menu_callback = ana_menu_callback

        self.root.title("Satış Raporları")
        try:
            self.root.state('zoomed')
        except Exception:
            self.root.geometry("1400x800")

        self.bg_color = '#F5F5F5'
        self.header_color = '#1565C0'
        self.filter_bg = '#E3F2FD'
        self.fg_color = 'white'
        self.root.configure(bg=self.bg_color)

        self.db = None
        self.kurum_listesi: List[Dict] = []
        self.son_rapor: List[Dict] = []

        # Filtre değişkenleri
        self.periyot_var = tk.StringVar(value='aylik')
        self.kirilim_var = tk.StringVar(value='tumu')
        self.kurum_var = tk.StringVar(value='')
        self.metrik_vars = {m[0]: tk.BooleanVar(value=True) for m in self.METRIK_TANIMLARI}

        # Sütun grupları toggle — Reçeteli/Elden/Genel hepsi açık başlar
        self.grup_vars = {g[0]: tk.BooleanVar(value=True) for g in self.SUTUN_GRUPLARI}

        # Nöbet modu
        self.nobet_modu_var = tk.BooleanVar(value=False)
        self.nobet_takvim = None  # NobetTakvimi instance (lazy)
        self.son_nobet_shiftleri: Dict = {}  # {shift_tarihi: info}

        # Endeks bazlı görünüm
        self.endeks_db = None  # EndeksDB instance (lazy)
        self.endeks_secim_var = tk.StringVar(value='TL (varsayılan)')
        # endeks_secim_var: legacy tek-endeks combobox (artık popup ile yönetiliyor;
        # geriye uyum için ilk seçili endeksi yansıtacak şekilde tutulur)

        # Çoklu endeks sütun seçimi (yeni)
        self.aktif_endeks_listesi: List[Dict] = []
        # Format: [{'kod': 'E:usd', 'id': 1, 'tip': 'endeks'|'sepet',
        #          'birim': 'USD', 'ad': 'Amerikan Doları'}, ...]

        # Sıralama durumu
        self._son_siralama_sutun = None
        self._son_siralama_ters = False

        self._arayuz_olustur()
        self._baglanti_kur()

    # ------------------------------------------------------------------
    # Arayüz
    # ------------------------------------------------------------------
    def _arayuz_olustur(self):
        self._header_olustur()
        self._filtre_panel_olustur()
        self._tablo_alani_olustur()
        self._status_bar_olustur()

    def _header_olustur(self):
        header = tk.Frame(self.root, bg=self.header_color, height=55)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header, text="📊 Satış Raporları",
            font=("Segoe UI", 14, "bold"), bg=self.header_color, fg=self.fg_color
        ).pack(side="left", padx=15, pady=12)

        btn_style = {'font': ("Segoe UI", 9), 'cursor': 'hand2', 'bd': 0, 'padx': 12, 'pady': 4}
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white',
                  activebackground='#B71C1C', command=self._kapat, **btn_style
                  ).pack(side="right", padx=10, pady=12)
        tk.Button(header, text="📥 Excel", bg='#2E7D32', fg='white',
                  activebackground='#1B5E20', command=self._excel_aktar, **btn_style
                  ).pack(side="right", padx=5, pady=12)
        tk.Button(header, text="🔎 Detay Görüntüle", bg='#FF8F00', fg='white',
                  activebackground='#E65100', command=self._donem_detayini_goster, **btn_style
                  ).pack(side="right", padx=5, pady=12)
        tk.Button(header, text="⏱ Toplu Zaman Anomalisi", bg='#5E35B1', fg='white',
                  activebackground='#4527A0', command=self._toplu_anomali_ac, **btn_style
                  ).pack(side="right", padx=5, pady=12)
        tk.Button(header, text="🔬 Logger Forensik", bg='#00838F', fg='white',
                  activebackground='#006064', command=self._logger_forensik_ac, **btn_style
                  ).pack(side="right", padx=5, pady=12)
        tk.Button(header, text="📊 Birleşik Zaman Analizi", bg='#6A1B9A', fg='white',
                  activebackground='#4A148C', command=self._birlesik_zaman_ac, **btn_style
                  ).pack(side="right", padx=5, pady=12)
        tk.Button(header, text="🔍 Sorgula", bg='#0277BD', fg='white',
                  activebackground='#01579B', command=self._sorgula, **btn_style
                  ).pack(side="right", padx=5, pady=12)

    def _filtre_panel_olustur(self):
        ana_filtre = tk.Frame(self.root, bg=self.filter_bg)
        ana_filtre.pack(fill="x", padx=5, pady=5)

        # --- Satır 1: Tarih aralığı + hızlı tarih butonları ---
        satir1 = tk.Frame(ana_filtre, bg=self.filter_bg)
        satir1.pack(fill="x", pady=5, padx=10)

        self._lbl(satir1, "Başlangıç:").pack(side="left")
        if TKCALENDAR_AVAILABLE:
            self.baslangic_tarih = DateEntry(
                satir1, width=12, date_pattern='dd.mm.yyyy',
                background='#1565C0', foreground='white', headersbackground='#1565C0',
                selectbackground='#1976D2', selectforeground='white',
                normalbackground='white', normalforeground='black',
                weekendbackground='#FFCDD2', weekendforeground='black',
                locale='tr_TR'
            )
            self.baslangic_tarih.set_date(datetime.now() - timedelta(days=30))
        else:
            self.baslangic_tarih = ttk.Entry(satir1, width=12)
            self.baslangic_tarih.insert(0, (datetime.now() - timedelta(days=30)).strftime("%d.%m.%Y"))
        self.baslangic_tarih.pack(side="left", padx=(5, 15))

        self._lbl(satir1, "Bitiş:").pack(side="left")
        if TKCALENDAR_AVAILABLE:
            self.bitis_tarih = DateEntry(
                satir1, width=12, date_pattern='dd.mm.yyyy',
                background='#1565C0', foreground='white', headersbackground='#1565C0',
                selectbackground='#1976D2', selectforeground='white',
                normalbackground='white', normalforeground='black',
                weekendbackground='#FFCDD2', weekendforeground='black',
                locale='tr_TR'
            )
            self.bitis_tarih.set_date(datetime.now())
        else:
            self.bitis_tarih = ttk.Entry(satir1, width=12)
            self.bitis_tarih.insert(0, datetime.now().strftime("%d.%m.%Y"))
        self.bitis_tarih.pack(side="left", padx=(5, 20))

        for txt, gun in [("Bu Ay", -1), ("30 Gün", 30), ("90 Gün", 90), ("1 Yıl", 365), ("Tümü", -2)]:
            tk.Button(satir1, text=txt, font=("Segoe UI", 8), bg='#90CAF9',
                      command=lambda g=gun: self._hizli_tarih(g), bd=0, padx=8
                      ).pack(side="left", padx=2)

        # --- Satır 2: Periyot radio ---
        satir2 = tk.Frame(ana_filtre, bg=self.filter_bg)
        satir2.pack(fill="x", pady=5, padx=10)

        self._lbl(satir2, "Periyot:").pack(side="left", padx=(0, 10))
        for kod, etiket in self.PERIYOT_SECENEKLERI:
            tk.Radiobutton(
                satir2, text=etiket, variable=self.periyot_var, value=kod,
                bg=self.filter_bg, font=("Segoe UI", 9)
            ).pack(side="left", padx=4)

        # --- Satır 3: Kırılım radio + kurum dropdown ---
        satir3 = tk.Frame(ana_filtre, bg=self.filter_bg)
        satir3.pack(fill="x", pady=5, padx=10)

        self._lbl(satir3, "Kırılım:").pack(side="left", padx=(0, 10))
        for kod, etiket in self.KIRILIM_SECENEKLERI:
            tk.Radiobutton(
                satir3, text=etiket, variable=self.kirilim_var, value=kod,
                bg=self.filter_bg, font=("Segoe UI", 9),
                command=self._kirilim_degisti
            ).pack(side="left", padx=4)

        self._lbl(satir3, "Kurum:").pack(side="left", padx=(15, 5))
        self.kurum_combo = ttk.Combobox(satir3, textvariable=self.kurum_var,
                                        width=35, state="disabled")
        self.kurum_combo.pack(side="left")

        # --- Satır 4: Sütun grupları toggle (Reçeteli / Elden / Genel) ---
        satir4 = tk.Frame(ana_filtre, bg=self.filter_bg)
        satir4.pack(fill="x", pady=5, padx=10)

        self._lbl(satir4, "Sütun Grupları:").pack(side="left", padx=(0, 10))
        for grup_kod, grup_et, grup_renk, _ in self.SUTUN_GRUPLARI:
            tk.Checkbutton(
                satir4, text=f"{grup_et}", variable=self.grup_vars[grup_kod],
                bg=grup_renk, font=("Segoe UI", 9, "bold"),
                padx=8, pady=2, relief='ridge', bd=1,
                command=self._grup_degisti
            ).pack(side="left", padx=4)

        tk.Label(satir4, text="(Reç./Eld./Genel — her birinde Sayı + Kutu + TL)",
                 bg=self.filter_bg, font=("Segoe UI", 8, "italic"),
                 fg='#37474F').pack(side="left", padx=15)

        # --- Satır 5: Nöbet modu ---
        satir5 = tk.Frame(ana_filtre, bg=self.filter_bg)
        satir5.pack(fill="x", pady=5, padx=10)

        tk.Checkbutton(
            satir5, text="🌙 Sadece Nöbet Satışları (mesai dışı yoğun günler)",
            variable=self.nobet_modu_var, bg=self.filter_bg,
            font=("Segoe UI", 9, "bold"), fg='#5D4037',
            command=self._nobet_modu_degisti
        ).pack(side="left", padx=6)

        self.nobet_bilgi_lbl = tk.Label(
            satir5, text="",
            bg=self.filter_bg, font=("Segoe UI", 9, "italic"), fg='#37474F'
        )
        self.nobet_bilgi_lbl.pack(side="left", padx=12)

        # --- Satır 6: Endeks bazlı görünüm ---
        satir6 = tk.Frame(ana_filtre, bg=self.filter_bg)
        satir6.pack(fill="x", pady=5, padx=10)

        self._lbl(satir6, "📈 Endeks Bazlı Görünüm:").pack(side="left", padx=(0, 8))

        self.endeks_sec_btn = tk.Button(
            satir6, text="📋 Endeks Sütunları (0 seçili)",
            font=("Segoe UI", 9, "bold"),
            bg='#00838F', fg='white', bd=0, padx=12, pady=3, cursor='hand2',
            command=self._endeks_sutunlari_sec_dialog
        )
        self.endeks_sec_btn.pack(side="left", padx=4)

        tk.Button(satir6, text="⚙ Endeks Ayarları", font=("Segoe UI", 9, "bold"),
                  bg='#5D4037', fg='white', bd=0, padx=10, pady=2, cursor='hand2',
                  command=self._endeks_ayarlari_ac).pack(side="left", padx=8)

        tk.Button(satir6, text="📊 Grafik", font=("Segoe UI", 9, "bold"),
                  bg='#1976D2', fg='white', bd=0, padx=10, pady=2, cursor='hand2',
                  command=self._grafik_ac).pack(side="left", padx=4)

        self.endeks_bilgi_lbl = tk.Label(
            satir6, text="(seçilen her endeks için TL ÷ endeks değeri sütunu eklenir)",
            bg=self.filter_bg, font=("Segoe UI", 9, "italic"), fg='#37474F'
        )
        self.endeks_bilgi_lbl.pack(side="left", padx=12)

    def _tablo_alani_olustur(self):
        tablo_frame = tk.Frame(self.root, bg=self.bg_color)
        tablo_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # Treeview - sütunlar dinamik, sorgu sonrası kurulur
        self.tree = ttk.Treeview(tablo_frame, columns=[], show='headings', height=25)

        vsb = ttk.Scrollbar(tablo_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tablo_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tablo_frame.columnconfigure(0, weight=1)
        tablo_frame.rowconfigure(0, weight=1)

        self.tree.tag_configure('toplam', background='#FFE082', font=('Segoe UI', 9, 'bold'))
        self.tree.tag_configure('cift', background='#F5F5F5')
        self.tree.tag_configure('tek', background='white')

        # Çift-tıklayınca o dönemin detay popup'ı açılır (butona basmakla aynı)
        self.tree.bind('<Double-1>', lambda e: self._donem_detayini_goster())

        self._sutunlari_kur()

    def _status_bar_olustur(self):
        self.status_bar = tk.Frame(self.root, bg='#37474F', height=24)
        self.status_bar.pack(fill="x", side="bottom")
        self.status_bar.pack_propagate(False)

        self.status_label = tk.Label(
            self.status_bar, text="Hazır.",
            bg='#37474F', fg='white', font=('Segoe UI', 9),
            anchor='w', padx=10
        )
        self.status_label.pack(side="left", fill="x", expand=True)

        self.progress = ttk.Progressbar(self.status_bar, mode='indeterminate', length=120)
        # Sorgu sırasında pack edilir

    # ------------------------------------------------------------------
    # Yardımcılar
    # ------------------------------------------------------------------
    def _lbl(self, parent, text):
        return tk.Label(parent, text=text, bg=self.filter_bg, font=("Segoe UI", 9, "bold"))

    def _status(self, mesaj: str, hata: bool = False):
        self.status_label.config(text=mesaj, fg='#FFCDD2' if hata else 'white')

    def _kapat(self):
        try:
            self.root.destroy()
        finally:
            if self.ana_menu_callback:
                try:
                    self.ana_menu_callback()
                except Exception:
                    pass

    def _hizli_tarih(self, gun: int):
        bugun = datetime.now()
        if gun == -1:  # Bu ay
            bas = bugun.replace(day=1)
            bit = bugun
        elif gun == -2:  # Tümü (eczane açılışı)
            bas = datetime(2017, 5, 23)
            bit = bugun
        else:
            bas = bugun - timedelta(days=gun)
            bit = bugun
        self._tarih_uygula(self.baslangic_tarih, bas)
        self._tarih_uygula(self.bitis_tarih, bit)

    def _tarih_uygula(self, widget, dt: datetime):
        if TKCALENDAR_AVAILABLE and hasattr(widget, 'set_date'):
            widget.set_date(dt)
        else:
            widget.delete(0, tk.END)
            widget.insert(0, dt.strftime("%d.%m.%Y"))

    def _tarih_oku(self, widget) -> Optional[date]:
        try:
            if TKCALENDAR_AVAILABLE and hasattr(widget, 'get_date'):
                return widget.get_date()
            else:
                metin = widget.get().strip()
                return datetime.strptime(metin, "%d.%m.%Y").date()
        except Exception as e:
            logger.error(f"Tarih okuma hatası: {e}")
            return None

    def _kirilim_degisti(self):
        if self.kirilim_var.get() == 'kurum':
            self.kurum_combo.config(state="readonly")
            if not self.kurum_listesi:
                self._kurum_listesini_yukle()
        else:
            self.kurum_combo.config(state="disabled")

    def _metrik_degisti(self):
        # En az bir metrik açık olmalı
        if not any(v.get() for v in self.metrik_vars.values()):
            self.metrik_vars['tl_tutar'].set(True)
        self._sutunlari_kur()
        if self.son_rapor:
            self._tabloyu_doldur(self.son_rapor)

    def _grup_degisti(self):
        # En az bir grup açık olmalı
        if not any(v.get() for v in self.grup_vars.values()):
            self.grup_vars['genel'].set(True)
        self._sutunlari_kur()
        if self.son_rapor:
            self._tabloyu_doldur(self.son_rapor)

    def _nobet_modu_degisti(self):
        if self.nobet_modu_var.get():
            self.nobet_bilgi_lbl.config(
                text="Nöbet eşiği: mesai dışı ≥10 satış VE ≥4 farklı saat. Sorgula'ya basın."
            )
            self._takvim_yukle()
        else:
            self.nobet_bilgi_lbl.config(text="")
            self.son_nobet_shiftleri = {}

    def _takvim_yukle(self):
        if self.nobet_takvim is None:
            try:
                from nobet_takvimi import get_nobet_takvimi
                self.nobet_takvim = get_nobet_takvimi()
                tatil_sayisi = len(self.nobet_takvim._tatil_seti())
                self._status(f"Nöbet takvimi yüklendi ({tatil_sayisi} resmi tatil).")
            except Exception as e:
                logger.error(f"Nöbet takvimi yüklenemedi: {e}", exc_info=True)
                messagebox.showerror("Hata", f"Nöbet takvimi yüklenemedi:\n{e}")
                self.nobet_modu_var.set(False)

    # ------------------------------------------------------------------
    # Endeks bazlı görünüm
    # ------------------------------------------------------------------
    def _endeks_db_yukle(self):
        if self.endeks_db is not None:
            return True
        try:
            from endeksler_db import get_endeks_db
            self.endeks_db = get_endeks_db()
            return True
        except Exception as e:
            logger.error(f"EndeksDB yüklenemedi: {e}", exc_info=True)
            self._status(f"EndeksDB hatası: {e}", hata=True)
            return False

    def _endeks_listesini_yenile(self):
        """Endeks Ayarları kapatıldığında çağrılır — silinen endeksleri seçim
        listesinden temizler. Yeni eklenen endeksler popup açıldığında görünür.
        """
        if not self._endeks_db_yukle():
            return
        try:
            endeksler = self.endeks_db.endeksleri_getir()
            sepetler = self.endeks_db.sepet_listesi()
        except Exception as e:
            logger.error(f"Endeks listesi alınamadı: {e}")
            return

        # Mevcut seçimlerden artık var olmayanları temizle
        mevcut_endeks_ids = {e['id'] for e in endeksler}
        mevcut_sepet_ids = {s['id'] for s in sepetler}
        yeni_liste = []
        for sec in self.aktif_endeks_listesi:
            if sec['tip'] == 'endeks' and sec['id'] in mevcut_endeks_ids:
                yeni_liste.append(sec)
            elif sec['tip'] == 'sepet' and sec['id'] in mevcut_sepet_ids:
                yeni_liste.append(sec)
        self.aktif_endeks_listesi = yeni_liste
        self._endeks_secim_butonu_guncelle()
        self._status(f"Endeks listesi: {len(endeksler)} endeks, {len(sepetler)} sepet.")

    def _endeks_secim_butonu_guncelle(self):
        """Endeks Sütunları butonunun başlığını güncel seçim sayısı ile yeniler."""
        n = len(self.aktif_endeks_listesi)
        if hasattr(self, 'endeks_sec_btn'):
            self.endeks_sec_btn.config(text=f"📋 Endeks Sütunları ({n} seçili)")

    def _endeks_sutunlari_sec_dialog(self):
        """Modal popup — tabloya eklenecek endeks/sepet sütunlarını seçtirir.

        Tamam'a basıldığında self.aktif_endeks_listesi güncellenir ve
        tablo + sütunlar yeniden kurulur.
        """
        if not self._endeks_db_yukle():
            return

        try:
            endeksler = self.endeks_db.endeksleri_getir()
            sepetler = self.endeks_db.sepet_listesi()
        except Exception as e:
            logger.error(f"Endeks listesi alınamadı: {e}")
            messagebox.showerror("Hata", f"Endeks listesi alınamadı:\n{e}")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Endeks Sütun Seçimi")
        dlg.geometry("560x650")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.configure(bg='#ECEFF1')

        # Header
        hdr = tk.Frame(dlg, bg='#00838F', height=42)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📋 Tabloya Eklenecek Endeks/Sepet Sütunları",
                 font=("Segoe UI", 11, "bold"), bg='#00838F', fg='white'
                 ).pack(side='left', padx=12, pady=8)

        # Açıklama
        tk.Label(dlg, text="Seçilen her endeks için tabloya 'TL ÷ endeks' sütunu eklenir.",
                 font=("Segoe UI", 9, "italic"), bg='#ECEFF1', fg='#37474F'
                 ).pack(pady=(8, 4), padx=12, anchor='w')

        # Scrollable frame
        outer = tk.Frame(dlg, bg='white', relief='solid', bd=1)
        outer.pack(fill='both', expand=True, padx=8, pady=4)

        canvas = tk.Canvas(outer, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        scrollable = tk.Frame(canvas, bg='white')
        scrollable.bind('<Configure>',
                        lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable, anchor='nw',
                             width=520)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Mevcut seçimlerin ID setleri (tip'e göre)
        secili_endeks_ids = {s['id'] for s in self.aktif_endeks_listesi if s['tip'] == 'endeks'}
        secili_sepet_ids = {s['id'] for s in self.aktif_endeks_listesi if s['tip'] == 'sepet'}

        check_vars: List[Tuple[tk.BooleanVar, Dict, str]] = []

        # Kategori bazlı grupla
        kategori_etiketleri = {
            'para':    ('💱 Para (Döviz/Altın)',  '#E1F5FE'),
            'ucret':   ('💼 Ücret',                '#FFF9C4'),
            'mal':     ('⛽ Mal',                  '#FFE0B2'),
            'ilac':    ('💊 İlaç',                 '#E8F5E9'),
            'kira_vs': ('🏠 Diğer (Kira vs.)',     '#F3E5F5'),
        }
        kategoriler = sorted(set(e['kategori'] for e in endeksler))
        # Bilinen kategorileri önce göster
        oncelikli = ['para', 'ucret', 'mal', 'kira_vs', 'ilac']
        kategoriler = [k for k in oncelikli if k in kategoriler] + \
                      [k for k in kategoriler if k not in oncelikli]

        for kategori in kategoriler:
            etiket, renk = kategori_etiketleri.get(kategori, (kategori, '#E0E0E0'))
            tk.Label(scrollable, text=etiket, font=("Segoe UI", 10, "bold"),
                     bg=renk, anchor='w', padx=8, pady=4
                     ).pack(fill='x', pady=(8, 2))
            for e in endeksler:
                if e['kategori'] != kategori:
                    continue
                var = tk.BooleanVar(value=(e['id'] in secili_endeks_ids))
                check_vars.append((var, dict(e), 'endeks'))
                tk.Checkbutton(scrollable, variable=var,
                               text=f"  {e['ad']}  ({e['birim'] or '?'})",
                               anchor='w', bg='white', font=("Segoe UI", 9)
                               ).pack(fill='x', padx=14)

        if sepetler:
            tk.Label(scrollable, text="🧺 Sepetler", font=("Segoe UI", 10, "bold"),
                     bg='#C8E6C9', anchor='w', padx=8, pady=4
                     ).pack(fill='x', pady=(8, 2))
            for s in sepetler:
                var = tk.BooleanVar(value=(s['id'] in secili_sepet_ids))
                check_vars.append((var, dict(s), 'sepet'))
                tk.Checkbutton(scrollable, variable=var,
                               text=f"  {s['ad']}",
                               anchor='w', bg='white', font=("Segoe UI", 9)
                               ).pack(fill='x', padx=14)

        # Buton barı
        btn_frame = tk.Frame(dlg, bg='#ECEFF1', pady=8)
        btn_frame.pack(fill='x', side='bottom')

        def tamam():
            yeni = []
            for var, item, tip in check_vars:
                if not var.get():
                    continue
                if tip == 'endeks':
                    yeni.append({
                        'tip': 'endeks',
                        'id': item['id'],
                        'kod': item['kod'],
                        'birim': item['birim'] or '?',
                        'ad': item['ad'],
                    })
                else:
                    yeni.append({
                        'tip': 'sepet',
                        'id': item['id'],
                        'kod': f"sepet_{item['id']}",
                        'birim': 'sepet',
                        'ad': item['ad'],
                    })
            self.aktif_endeks_listesi = yeni
            # Legacy combobox değerini güncelle (ilk seçili endeks)
            if yeni:
                ilk = yeni[0]
                if ilk['tip'] == 'endeks':
                    self.endeks_secim_var.set(f"E:{ilk['kod']} — {ilk['ad']} ({ilk['birim']})")
                else:
                    self.endeks_secim_var.set(f"S:{ilk['id']} — SEPET: {ilk['ad']}")
            else:
                self.endeks_secim_var.set('TL (varsayılan)')
            self._endeks_secim_butonu_guncelle()
            self._sutunlari_kur()
            if self.son_rapor:
                self._tabloyu_doldur(self.son_rapor)
            self._status(f"{len(yeni)} endeks sütunu aktif.")
            dlg.destroy()

        def temizle():
            for var, _, _ in check_vars:
                var.set(False)

        def hepsini():
            for var, _, _ in check_vars:
                var.set(True)

        tk.Button(btn_frame, text='⊞ Hepsini Seç', command=hepsini,
                  bg='#78909C', fg='white', bd=0, padx=12, pady=4,
                  font=("Segoe UI", 9)).pack(side='left', padx=8)
        tk.Button(btn_frame, text='⊟ Temizle', command=temizle,
                  bg='#FFB74D', fg='white', bd=0, padx=12, pady=4,
                  font=("Segoe UI", 9)).pack(side='left', padx=4)

        tk.Button(btn_frame, text='İptal', command=dlg.destroy,
                  bg='#CFD8DC', fg='#37474F', bd=0, padx=14, pady=4,
                  font=("Segoe UI", 9)).pack(side='right', padx=8)
        tk.Button(btn_frame, text='✓ Tamam', command=tamam,
                  bg='#2E7D32', fg='white', bd=0, padx=18, pady=4,
                  font=("Segoe UI", 10, "bold")).pack(side='right', padx=4)

    def _endeks_ayarlari_ac(self):
        """Endeks Ayarları yönetim penceresini Toplevel olarak aç.

        Pencere kapanınca endeks listesini otomatik yenile (kullanıcı yeni
        endeks eklediyse Satış Raporları dropdown'ı güncellensin).
        """
        try:
            from endeks_ayarlari_gui import EndeksAyarlariGUI
            pencere = tk.Toplevel(self.root)
            EndeksAyarlariGUI(
                pencere,
                ana_menu_callback=self._endeks_listesini_yenile
            )
            self._status("Endeks Ayarları açıldı.")
        except ImportError as e:
            logger.error(f"Endeks Ayarları import hatası: {e}")
            messagebox.showerror("Hata", f"Endeks Ayarları modülü yüklenemedi:\n{e}")
        except Exception as e:
            logger.error(f"Endeks Ayarları açma hatası: {e}", exc_info=True)
            messagebox.showerror("Hata", f"Endeks Ayarları açılamadı:\n{e}")

    def _grafik_ac(self):
        """Endeks/ciro grafik penceresini Toplevel olarak aç.

        Mevcut tarih aralığı ve seçili endeksi grafik penceresine taşır.
        """
        try:
            from endeks_grafik_gui import EndeksGrafikPenceresi
        except ImportError as e:
            logger.error(f"Grafik modülü import hatası: {e}")
            messagebox.showerror(
                "Grafik Modülü Yüklenemedi",
                f"endeks_grafik_gui modülü yüklenemedi:\n{e}\n\n"
                "matplotlib yüklü mü? pip install matplotlib"
            )
            return

        # Mevcut filtre değerlerini yansıt
        bas_d = self._tarih_oku(self.baslangic_tarih)
        bit_d = self._tarih_oku(self.bitis_tarih)
        bas = bas_d.isoformat() if bas_d else None
        bit = bit_d.isoformat() if bit_d else None
        endeks_secim_txt = self.endeks_secim_var.get()

        # Endeks dropdown'unda görünen format ile grafik penceresindeki uyumlu mu?
        # Satış Raporları'nda 'E:usd — Amerikan Doları (USD)' formatı
        # Grafik penceresinde aynı format kullanılıyor — direkt geç.
        try:
            EndeksGrafikPenceresi(
                self.root,
                baslangic_tarih=bas,
                bitis_tarih=bit,
                ilk_endeks_secim=endeks_secim_txt if endeks_secim_txt and endeks_secim_txt != 'TL (varsayılan)' else None,
            )
            self._status("Grafik penceresi açıldı.")
        except Exception as e:
            logger.error(f"Grafik açma hatası: {e}", exc_info=True)
            messagebox.showerror("Hata", f"Grafik penceresi açılamadı:\n{e}")

    def _endeks_secimi_degisti(self, _evt=None):
        # Yeniden sütun yapısı ve tablo
        self._sutunlari_kur()
        if self.son_rapor:
            self._tabloyu_doldur(self.son_rapor)

    def _aktif_endeks_id_ve_birim(self):
        """Şu an seçili olan endeks/sepet için (endeks_id|sepet_id, birim, tip, ad) döndürür.

        tip: 'tl' | 'endeks' | 'sepet'
        Hiçbir endeks seçili değilse ('TL') → tip='tl', id=None
        """
        secim = self.endeks_secim_var.get()
        if not secim or secim.startswith('TL'):
            return None, '₺', 'tl', 'TL'

        if secim.startswith('E:'):
            kod = secim.split(' — ')[0][2:].strip()
            if self.endeks_db is None:
                return None, '?', 'tl', 'TL'
            endeks_id = self.endeks_db._endeks_id_kod(kod)
            if endeks_id is None:
                return None, '?', 'tl', 'TL'
            c = self.endeks_db.conn.cursor()
            row = c.execute("SELECT ad, birim FROM endeks_tanim WHERE id=?", (endeks_id,)).fetchone()
            return endeks_id, (row['birim'] if row else '?'), 'endeks', (row['ad'] if row else kod)

        if secim.startswith('S:'):
            sid_str = secim.split(' — ')[0][2:].strip()
            try:
                sepet_id = int(sid_str)
            except ValueError:
                return None, '?', 'tl', 'TL'
            c = self.endeks_db.conn.cursor()
            row = c.execute("SELECT ad FROM endeks_sepet WHERE id=?", (sepet_id,)).fetchone()
            return sepet_id, 'sepet', 'sepet', (row['ad'] if row else 'Sepet')

        return None, '₺', 'tl', 'TL'

    def _endeks_degeri_donem(self, donem: date, endeks_kimlik, tip: str) -> Optional[float]:
        """Dönem için endeks değerini bul.

        Dönem aralığı periyot'a göre belirlenir (örn. aylik için ay başı→ay sonu).
        Sepet ise sepet_donem_ortalama, endeks ise donem_ortalama.
        """
        if endeks_kimlik is None or self.endeks_db is None:
            return None
        per = self.periyot_var.get()
        # Dönem bitiş tarihi
        if per == 'gunluk':
            bas = donem
            bit = donem
        elif per == 'haftalik':
            bas = donem
            bit = donem + timedelta(days=6)
        elif per == 'aylik':
            bas = donem
            if donem.month == 12:
                bit = date(donem.year + 1, 1, 1) - timedelta(days=1)
            else:
                bit = date(donem.year, donem.month + 1, 1) - timedelta(days=1)
        elif per == '3aylik':
            bas = donem
            son_ay = donem.month + 2
            if son_ay > 12:
                bit = date(donem.year + 1, son_ay - 12 + 1, 1) - timedelta(days=1)
            else:
                bit = date(donem.year, son_ay + 1, 1) - timedelta(days=1)
        elif per == 'yillik':
            bas = donem
            bit = date(donem.year + 1, 1, 1) - timedelta(days=1)
        else:
            bas = bit = donem

        if tip == 'endeks':
            return self.endeks_db.donem_ortalama(endeks_kimlik, bas, bit)
        elif tip == 'sepet':
            return self.endeks_db.sepet_donem_ortalama(endeks_kimlik, bas, bit)
        return None

    def _sutunlari_kur(self):
        sutunlar = ['Donem']
        basliklar = {'Donem': 'Dönem'}
        genislikler = {'Donem': 130}

        # Yeni: 3 grup (Reçeteli/Elden/Genel) — her birinin 3 alt sütunu
        # _alan_kodu_to_db_alan: alan kodundan db alanına mapping (TOPLAM hesabı için)
        self._alan_kodu_map: Dict[str, Tuple[str, str]] = {}  # kod → (db_alan, tip)
        for grup_kod, grup_et, grup_renk, alanlar in self.SUTUN_GRUPLARI:
            if not self.grup_vars[grup_kod].get():
                continue
            for alan_kod, alan_baslik, db_alan, tip in alanlar:
                sutunlar.append(alan_kod)
                # Başlığı 2 satırlı yap: "[Grup]\n alan adı"
                basliklar[alan_kod] = f"{grup_et}\n{alan_baslik}"
                if tip == 'tl':
                    genislikler[alan_kod] = 110
                else:
                    genislikler[alan_kod] = 80
                self._alan_kodu_map[alan_kod] = (db_alan, tip)

        # Kasaya etkiyen tutar sütunu (Reçeteli RIToplam + Elden RIToplam = TLTutar)
        # Hep göster — TL toplamına eşit olsa da semantik olarak ayrı.
        sutunlar.append('kasa_tl')
        basliklar['kasa_tl'] = '💵 Kasa ₺\n(Hasta payı + Elden)'
        genislikler['kasa_tl'] = 140

        # Endeks bazlı görünüm — çoklu sütun (her seçili endeks/sepet için ayrı)
        # self._aktif_endeks_kolonlari: tablo doldurma + Excel için sıralı liste
        self._aktif_endeks_kolonlari: List[Dict] = []
        if self.grup_vars['genel'].get():
            for sec in self.aktif_endeks_listesi:
                sutun_id = f"endeks_{sec['tip']}_{sec['id']}"
                sutunlar.append(sutun_id)
                basliklar[sutun_id] = f"TL ÷\n{sec['ad']}"
                genislikler[sutun_id] = 140
                self._aktif_endeks_kolonlari.append({
                    'sutun_id': sutun_id,
                    'tip': sec['tip'],
                    'id': sec['id'],
                    'birim': sec['birim'],
                    'ad': sec['ad'],
                })

        # Legacy attribute'ları ilk seçimle doldur (geriye uyumluluk — eski yerlerde)
        if self._aktif_endeks_kolonlari:
            ilk = self._aktif_endeks_kolonlari[0]
            self._aktif_endeks_kimlik = ilk['id']
            self._aktif_endeks_tip = ilk['tip']
            self._aktif_endeks_birim = ilk['birim']
            self._aktif_endeks_ad = ilk['ad']
        else:
            self._aktif_endeks_kimlik = None
            self._aktif_endeks_tip = 'tl'
            self._aktif_endeks_birim = '₺'
            self._aktif_endeks_ad = 'TL'

        self.tree['columns'] = sutunlar
        for s in sutunlar:
            self.tree.heading(s, text=basliklar[s],
                              command=lambda c=s: self._sirala(c))
            self.tree.column(s, width=genislikler[s], anchor='center')

        # Treeview heading'ler 2 satırlı görsünler (varsayılan tek satır)
        try:
            style = ttk.Style()
            style.configure('Treeview.Heading', font=('Segoe UI', 9, 'bold'))
            style.configure('Treeview', rowheight=24)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # DB bağlantısı + sorgu
    # ------------------------------------------------------------------
    def _baglanti_kur(self):
        def _calis():
            try:
                from botanik_db import get_botanik_db
                self.db = get_botanik_db()
                if not self.db.baglan():
                    self.root.after(0, lambda: self._status("Botanik EOS bağlantısı kurulamadı.", hata=True))
                    return
                self.root.after(0, lambda: self._status(
                    "Botanik EOS bağlandı. Tarih aralığı ve filtreler seçili, Sorgula'ya basın."
                ))
            except Exception as e:
                logger.error(f"DB bağlantı hatası: {e}", exc_info=True)
                self.root.after(0, lambda: self._status(f"DB bağlantı hatası: {e}", hata=True))

        threading.Thread(target=_calis, daemon=True).start()

    def _kurum_listesini_yukle(self):
        def _calis():
            try:
                if self.db is None:
                    return
                self.kurum_listesi = self.db.kurum_listesi_getir()
                degerler = [
                    f"{k['KurumId']} - {k['KurumAdi']}" for k in self.kurum_listesi
                ]
                self.root.after(0, lambda: self.kurum_combo.config(values=degerler))
                if degerler:
                    self.root.after(0, lambda: self.kurum_combo.set(degerler[0]))
            except Exception as e:
                logger.error(f"Kurum listesi hatası: {e}")

        threading.Thread(target=_calis, daemon=True).start()

    def _sorgula(self):
        bas = self._tarih_oku(self.baslangic_tarih)
        bit = self._tarih_oku(self.bitis_tarih)
        if not bas or not bit:
            messagebox.showwarning("Tarih hatası",
                                   "Tarih biçimi GG.AA.YYYY olmalı (ör. 01.05.2025).")
            return
        if bas > bit:
            messagebox.showwarning("Tarih hatası", "Başlangıç tarihi bitişten büyük olamaz.")
            return

        periyot = self.periyot_var.get()
        kirilim = self.kirilim_var.get()
        kurum_id = None
        if kirilim == 'kurum':
            secim = self.kurum_var.get().strip()
            if not secim:
                messagebox.showwarning("Kurum seçimi", "Lütfen bir kurum seçin.")
                return
            try:
                kurum_id = int(secim.split(' - ', 1)[0])
            except Exception:
                messagebox.showerror("Hata", f"Kurum ID parse edilemedi: {secim}")
                return

        nobet_modu = self.nobet_modu_var.get()
        if nobet_modu and self.nobet_takvim is None:
            self._takvim_yukle()
            if self.nobet_takvim is None:
                return  # yüklenemedi, hata zaten gösterildi

        # Async sorgu
        mod_etiket = "NÖBET MODU" if nobet_modu else "normal"
        self._status(f"Sorgu çalıştırılıyor... ({bas} → {bit}, periyot={periyot}, kırılım={kirilim}, {mod_etiket})")
        self.progress.pack(side="right", padx=10)
        self.progress.start(10)

        def _calis():
            try:
                if self.db is None:
                    raise RuntimeError("Veritabanı bağlantısı yok.")

                if nobet_modu:
                    # Saat detaylı satışlar → Python'da nöbet shift'i hesapla → periyot grupla
                    detay = self.db.satis_zaman_detay_getir(
                        baslangic_tarih=bas,
                        bitis_tarih=bit,
                        kirilim=kirilim,
                        kurum_id=kurum_id,
                    )
                    sonuc = self._nobet_periyot_grupla(detay, periyot)
                    self.root.after(0, lambda: self._sorgu_bitti(sonuc))
                else:
                    sonuc = self.db.satis_raporu_getir(
                        baslangic_tarih=bas,
                        bitis_tarih=bit,
                        kirilim=kirilim,
                        kurum_id=kurum_id,
                        periyot=periyot,
                    )
                    self.root.after(0, lambda: self._sorgu_bitti(sonuc))
            except Exception as e:
                logger.error(f"Satış raporu sorgu hatası: {e}", exc_info=True)
                self.root.after(0, lambda: self._sorgu_hata(str(e)))

        threading.Thread(target=_calis, daemon=True).start()

    def _nobet_periyot_grupla(self, satislar: List[Dict], periyot: str) -> List[Dict]:
        """Saat detaylı satışları nöbet shift'lerine ata, onaylı shift'ler için
        periyot bazlı toplam metrikleri üret. Reçeteli/Elden ayrımı dahil.

        Returns:
            [{'Donem': date, 'ReceteliSayisi', 'ReceteliKutu', 'ReceteliTL',
              'EldenSayisi', 'EldenKutu', 'EldenTL',
              'SatisSayisi', 'KutuSayisi', 'TLTutar', 'ReceteSayisi'(legacy)}, ...]
        """
        if self.nobet_takvim is None or not satislar:
            return []

        shiftler = self.nobet_takvim.nobet_shiftlerini_tespit_et(satislar)
        self.son_nobet_shiftleri = shiftler

        onayli_satislar = []
        for shift_tarihi, info in shiftler.items():
            if not info['onayli_nobet']:
                continue
            for s in info['satislar']:
                onayli_satislar.append({
                    **s,
                    'shift_tarihi': shift_tarihi,
                    'shift_tipi': info['tip'],
                })

        if not onayli_satislar:
            return []

        from collections import defaultdict
        grup_metrik = defaultdict(lambda: {
            'recete_rx': set(),
            'elden_rx': set(),
            'recete_kalem': 0,
            'recete_kutu': 0.0,
            'recete_tl': 0.0,
            'elden_kalem': 0,
            'elden_kutu': 0.0,
            'elden_tl': 0.0,
        })

        for s in onayli_satislar:
            donem = self._periyot_donem(s['shift_tarihi'], periyot)
            kaynak = s['kaynak']
            rx_id = s['rx_id']
            kalem = int(s.get('kalem_sayisi') or 0)
            kutu = float(s.get('adet') or 0)
            tl = float(s.get('tutar') or 0)
            g = grup_metrik[donem]
            if kaynak == 'RECETE':
                g['recete_rx'].add(rx_id)
                g['recete_kalem'] += kalem
                g['recete_kutu'] += kutu
                g['recete_tl'] += tl
            else:  # ELDEN
                g['elden_rx'].add(rx_id)
                g['elden_kalem'] += kalem
                g['elden_kutu'] += kutu
                g['elden_tl'] += tl

        sonuc = []
        for donem in sorted(grup_metrik.keys()):
            g = grup_metrik[donem]
            recete_sayisi = len(g['recete_rx'])
            elden_sayisi = len(g['elden_rx'])
            sonuc.append({
                'Donem': donem,
                # Reçeteli kırılım
                'ReceteliSayisi': recete_sayisi,
                'ReceteliKalem': g['recete_kalem'],
                'ReceteliKutu': g['recete_kutu'],
                'ReceteliTL': g['recete_tl'],
                # Elden kırılım
                'EldenSayisi': elden_sayisi,
                'EldenKalem': g['elden_kalem'],
                'EldenKutu': g['elden_kutu'],
                'EldenTL': g['elden_tl'],
                # Genel toplam
                'SatisSayisi': recete_sayisi + elden_sayisi,
                'KalemSayisi': g['recete_kalem'] + g['elden_kalem'],
                'KutuSayisi': g['recete_kutu'] + g['elden_kutu'],
                'TLTutar': g['recete_tl'] + g['elden_tl'],
                'ReceteSayisi': recete_sayisi,
            })
        return sonuc

    @staticmethod
    def _periyot_donem(d: date, periyot: str) -> date:
        """Bir tarihi periyodun başlangıç tarihine çevirir (SQL DATEFROMPARTS muadili)."""
        if periyot == 'gunluk':
            return d
        elif periyot == 'haftalik':
            # Pazartesi başlangıçlı hafta
            return d - timedelta(days=d.weekday())
        elif periyot == 'aylik':
            return date(d.year, d.month, 1)
        elif periyot == '3aylik':
            ceyrek_bas_ay = ((d.month - 1) // 3) * 3 + 1
            return date(d.year, ceyrek_bas_ay, 1)
        elif periyot == 'yillik':
            return date(d.year, 1, 1)
        return d

    def _sorgu_bitti(self, sonuc: List[Dict]):
        self.progress.stop()
        self.progress.pack_forget()
        self.son_rapor = sonuc
        self._sutunlari_kur()
        self._tabloyu_doldur(sonuc)

        if self.nobet_modu_var.get():
            # Nöbet özet bilgisi
            onayli = sum(1 for v in self.son_nobet_shiftleri.values() if v['onayli_nobet'])
            top_shift = len(self.son_nobet_shiftleri)
            tip_dagilim = {}
            for v in self.son_nobet_shiftleri.values():
                if v['onayli_nobet']:
                    tip_dagilim[v['tip']] = tip_dagilim.get(v['tip'], 0) + 1
            tip_str = ", ".join(f"{tipi}: {sayi}" for tipi, sayi in tip_dagilim.items()) or "—"
            self.nobet_bilgi_lbl.config(
                text=f"Toplam {top_shift} shift incelendi, {onayli} onaylı nöbet. ({tip_str})"
            )
            if sonuc:
                self._status(f"NÖBET: {onayli} onaylı nöbet shift'i, {len(sonuc)} dönem satırı.")
            else:
                self._status("NÖBET: Tarih aralığında onaylı nöbet bulunamadı (eşik altında).")
        else:
            if sonuc:
                self._status(f"{len(sonuc)} satır döndü. Toplam satır eklendi.")
            else:
                self._status("Sorgu çalıştı ama hiç satır dönmedi (tarih aralığında veri yok).")

    def _sorgu_hata(self, mesaj: str):
        self.progress.stop()
        self.progress.pack_forget()
        self._status(f"Sorgu hatası: {mesaj}", hata=True)
        messagebox.showerror("Sorgu hatası", mesaj)

    # ------------------------------------------------------------------
    # Tablo + sıralama + export
    # ------------------------------------------------------------------
    def _tabloyu_doldur(self, sonuc: List[Dict]):
        for c in self.tree.get_children():
            self.tree.delete(c)

        if not sonuc:
            return

        endeks_kolonlari = getattr(self, '_aktif_endeks_kolonlari', [])

        # Aktif alan kodları (grup_vars'a göre)
        aktif_alanlar = list(self._alan_kodu_map.keys())  # sıralı

        # Toplam akümülatör (her aktif alan için)
        toplam = {alan_kod: 0.0 for alan_kod in aktif_alanlar}
        toplam_endeksler = {kol['sutun_id']: 0.0 for kol in endeks_kolonlari}

        toplam_kasa = 0.0
        for idx, r in enumerate(sonuc):
            donem = r.get('Donem')
            donem_str = self._donem_format(donem)

            satir = [donem_str]
            for alan_kod in aktif_alanlar:
                db_alan, tip = self._alan_kodu_map[alan_kod]
                deger = r.get(db_alan) or 0
                satir.append(self._alan_format(deger, tip))
                toplam[alan_kod] += float(deger or 0)

            # Kasaya etkiyen tutar: Reçeteli + Elden RIToplam = TLTutar
            kasa_deg = float(r.get('TLTutar') or 0)
            satir.append(self._alan_format(kasa_deg, 'tl'))
            toplam_kasa += kasa_deg

            # Endeks sütunları — her seçili endeks/sepet için TL ÷ endeks
            tl_deg = float(r.get('TLTutar') or 0)
            for kol in endeks_kolonlari:
                e_deg = self._endeks_degeri_donem(
                    donem if isinstance(donem, date) else None,
                    kol['id'], kol['tip']
                )
                if e_deg and e_deg > 0:
                    bolum = tl_deg / e_deg
                    satir.append(self._endeks_format(bolum, kol['birim']))
                    toplam_endeksler[kol['sutun_id']] += bolum
                else:
                    satir.append("—")

            tag = 'cift' if idx % 2 == 0 else 'tek'
            self.tree.insert('', 'end', values=satir, tags=(tag,))

        # TOPLAM satırı
        toplam_satir = ['TOPLAM']
        for alan_kod in aktif_alanlar:
            db_alan, tip = self._alan_kodu_map[alan_kod]
            toplam_satir.append(self._alan_format(toplam[alan_kod], tip))
        toplam_satir.append(self._alan_format(toplam_kasa, 'tl'))
        for kol in endeks_kolonlari:
            toplam_satir.append(self._endeks_format(
                toplam_endeksler[kol['sutun_id']], kol['birim']
            ))
        self.tree.insert('', 'end', values=toplam_satir, tags=('toplam',))

    @staticmethod
    def _alan_format(deger, tip: str) -> str:
        try:
            if tip == 'tl':
                return f"{float(deger):,.2f} ₺".replace(',', 'X').replace('.', ',').replace('X', '.')
            elif tip == 'float':
                return f"{float(deger):,.0f}".replace(',', '.')
            else:  # int
                return f"{int(deger):,}".replace(',', '.')
        except Exception:
            return str(deger)

    @staticmethod
    def _endeks_format(deger, birim: str) -> str:
        try:
            son_birim = '' if birim == 'sepet' else birim
            return f"{float(deger):,.2f} {son_birim}".replace(',', 'X').replace('.', ',').replace('X', '.').strip()
        except Exception:
            return str(deger)

    def _donem_format(self, donem) -> str:
        if donem is None:
            return ''
        if isinstance(donem, str):
            try:
                donem = datetime.strptime(donem.split(' ')[0], '%Y-%m-%d').date()
            except Exception:
                return donem
        periyot = self.periyot_var.get()
        if periyot == 'gunluk':
            return donem.strftime('%d.%m.%Y (%a)')
        elif periyot == 'haftalik':
            return f"Hafta: {donem.strftime('%d.%m.%Y')}"
        elif periyot == 'aylik':
            return donem.strftime('%Y - %B')
        elif periyot == '3aylik':
            ay = donem.month
            ceyrek = (ay - 1) // 3 + 1
            return f"{donem.year} - Q{ceyrek}"
        elif periyot == 'yillik':
            return donem.strftime('%Y')
        return str(donem)

    def _sayi_format(self, deger, metrik_kodu: str) -> str:
        try:
            if metrik_kodu == 'tl_tutar':
                return f"{float(deger):,.2f} ₺".replace(',', 'X').replace('.', ',').replace('X', '.')
            elif metrik_kodu == 'kutu_sayisi':
                return f"{float(deger):,.0f}".replace(',', '.')
            else:
                return f"{int(deger):,}".replace(',', '.')
        except Exception:
            return str(deger)

    def _sirala(self, sutun: str):
        if not self.son_rapor:
            return
        ters = (self._son_siralama_sutun == sutun) and not self._son_siralama_ters
        self._son_siralama_sutun = sutun
        self._son_siralama_ters = ters

        if sutun == 'Donem':
            anahtar = lambda r: r.get('Donem') or date.min
        elif sutun == 'endeks_tutar':
            endeks_kimlik = getattr(self, '_aktif_endeks_kimlik', None)
            endeks_tip = getattr(self, '_aktif_endeks_tip', 'tl')

            def _endeks_anahtar(r):
                tl = float(r.get('TLTutar') or 0)
                d = r.get('Donem')
                endeks_deg = self._endeks_degeri_donem(
                    d if isinstance(d, date) else None,
                    endeks_kimlik, endeks_tip,
                )
                if endeks_deg and endeks_deg > 0:
                    return tl / endeks_deg
                return 0.0
            anahtar = _endeks_anahtar
        else:
            # Yeni grup alan kodları → db_alan eşleme
            alan_kod_map = getattr(self, '_alan_kodu_map', {})
            db_alan, _ = alan_kod_map.get(sutun, ('TLTutar', 'tl'))
            anahtar = lambda r: r.get(db_alan) or 0

        sirali = sorted(self.son_rapor, key=anahtar, reverse=ters)
        self.son_rapor = sirali
        self._tabloyu_doldur(sirali)

    def _excel_aktar(self):
        if not self.son_rapor:
            messagebox.showinfo("Boş rapor", "Önce sorgu çalıştırın.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("openpyxl yok", "Excel aktarımı için openpyxl gerekli.")
            return

        dosya = filedialog.asksaveasfilename(
            title="Satış Raporu Excel'e Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"satis_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not dosya:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Satış Raporu"

        beyaz = Font(bold=True, color='FFFFFF')
        siyah_bold = Font(bold=True)
        donem_fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
        toplam_fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
        merkez = Alignment(horizontal='center', vertical='center')

        # Satır 1: Grup başlıkları (merge edilmiş hücreler), Satır 2: alt başlıklar
        # Sütun A: Dönem (2 satır merge)
        ws.cell(row=1, column=1, value='Dönem').font = beyaz
        ws.cell(row=1, column=1).fill = donem_fill
        ws.cell(row=1, column=1).alignment = merkez
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        col = 2
        # Grup başlıklarını yaz
        for grup_kod, grup_et, grup_renk_hex, alanlar in self.SUTUN_GRUPLARI:
            if not self.grup_vars[grup_kod].get():
                continue
            renk = grup_renk_hex.lstrip('#').upper()
            # Hex -> openpyxl fill (FFRRGGBB)
            grup_fill = PatternFill(start_color=renk, end_color=renk, fill_type='solid')
            start_col = col
            for alan_kod, alan_baslik, db_alan, tip in alanlar:
                c = ws.cell(row=2, column=col, value=alan_baslik)
                c.font = siyah_bold
                c.fill = grup_fill
                c.alignment = merkez
                col += 1
            end_col = col - 1
            # Grup başlığı (1. satır)
            gc = ws.cell(row=1, column=start_col, value=grup_et)
            gc.font = siyah_bold
            gc.fill = grup_fill
            gc.alignment = merkez
            if start_col < end_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        # Kasa sütunu başlığı (grup'ların hemen sonrası)
        kasa_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9',
                                 fill_type='solid')
        kasa_col = col
        kc = ws.cell(row=1, column=kasa_col, value='💵 Kasa ₺')
        kc.font = siyah_bold
        kc.fill = kasa_fill
        kc.alignment = merkez
        ws.merge_cells(start_row=1, start_column=kasa_col,
                       end_row=2, end_column=kasa_col)
        col = kasa_col + 1

        # Endeks sütunları başlıkları
        endeks_kolonlari = getattr(self, '_aktif_endeks_kolonlari', [])
        endeks_fill = PatternFill(start_color='B2EBF2', end_color='B2EBF2',
                                   fill_type='solid')
        endeks_col_start = col
        for kol in endeks_kolonlari:
            ec = ws.cell(row=1, column=col,
                         value=f"TL ÷ {kol['ad']} ({kol['birim']})")
            ec.font = siyah_bold
            ec.fill = endeks_fill
            ec.alignment = merkez
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=2, end_column=col)
            col += 1

        # Veri satırları
        toplam = {alan_kod: 0.0 for alan_kod in self._alan_kodu_map.keys()}
        toplam_kasa = 0.0
        toplam_endeksler = {kol['sutun_id']: 0.0 for kol in endeks_kolonlari}

        for row_idx, r in enumerate(self.son_rapor, 3):
            donem = r.get('Donem')
            ws.cell(row=row_idx, column=1, value=self._donem_format(donem))
            col = 2
            for alan_kod in self._alan_kodu_map.keys():
                db_alan, tip = self._alan_kodu_map[alan_kod]
                deger = r.get(db_alan) or 0
                c = ws.cell(row=row_idx, column=col)
                if tip == 'tl':
                    c.value = float(deger)
                    c.number_format = '#,##0.00 ₺'
                elif tip == 'float':
                    c.value = float(deger)
                    c.number_format = '#,##0'
                else:
                    c.value = int(deger)
                    c.number_format = '#,##0'
                toplam[alan_kod] += float(deger or 0)
                col += 1
            # Kasa sütunu
            kasa_d = float(r.get('TLTutar') or 0)
            kc = ws.cell(row=row_idx, column=kasa_col, value=kasa_d)
            kc.number_format = '#,##0.00 ₺'
            toplam_kasa += kasa_d

            # Endeks sütunları
            tl_deg = float(r.get('TLTutar') or 0)
            ekol = endeks_col_start
            for kol in endeks_kolonlari:
                e_deg = self._endeks_degeri_donem(
                    donem if isinstance(donem, date) else None,
                    kol['id'], kol['tip']
                )
                ec = ws.cell(row=row_idx, column=ekol)
                if e_deg and e_deg > 0:
                    bolum = tl_deg / e_deg
                    ec.value = bolum
                    ec.number_format = '#,##0.00'
                    toplam_endeksler[kol['sutun_id']] += bolum
                else:
                    ec.value = None
                ekol += 1

        # TOPLAM satırı
        top_row = len(self.son_rapor) + 3
        ws.cell(row=top_row, column=1, value='TOPLAM').font = siyah_bold
        ws.cell(row=top_row, column=1).fill = toplam_fill
        col = 2
        for alan_kod in self._alan_kodu_map.keys():
            db_alan, tip = self._alan_kodu_map[alan_kod]
            c = ws.cell(row=top_row, column=col, value=toplam[alan_kod])
            c.font = siyah_bold
            c.fill = toplam_fill
            if tip == 'tl':
                c.number_format = '#,##0.00 ₺'
            elif tip == 'float':
                c.number_format = '#,##0'
            else:
                c.value = int(toplam[alan_kod])
                c.number_format = '#,##0'
            col += 1
        # Kasa toplamı
        kc = ws.cell(row=top_row, column=kasa_col, value=toplam_kasa)
        kc.font = siyah_bold
        kc.fill = toplam_fill
        kc.number_format = '#,##0.00 ₺'
        # Endeks toplamları
        ekol = endeks_col_start
        for kol in endeks_kolonlari:
            ec = ws.cell(row=top_row, column=ekol,
                          value=toplam_endeksler[kol['sutun_id']])
            ec.font = siyah_bold
            ec.fill = toplam_fill
            ec.number_format = '#,##0.00'
            ekol += 1
        col = endeks_col_start + len(endeks_kolonlari)

        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 22
        for col_idx in range(2, col):
            # openpyxl 1-indexed; chr ile A,B,... AA, AB için utility kullan
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        try:
            wb.save(dosya)
            self._status(f"Excel'e kaydedildi: {dosya}")
            messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{dosya}")
        except Exception as e:
            logger.error(f"Excel kaydetme hatası: {e}", exc_info=True)
            messagebox.showerror("Hata", f"Excel kaydedilemedi:\n{e}")


    # ------------------------------------------------------------------
    # Toplu Zaman Anomalisi
    # ------------------------------------------------------------------
    def _toplu_anomali_ac(self):
        """Geniş aralıkta RxIslemTarihi vs RxKayitTarihi anomali tarayıcısını aç."""
        if self.db is None:
            messagebox.showerror("DB hazır değil",
                "Veritabanı bağlantısı henüz kurulmadı. Birkaç saniye sonra deneyin.")
            return
        TopluZamanAnomaliPopup(parent=self.root, db=self.db)

    def _logger_forensik_ac(self):
        """Logger tablosu ile manuel tarih değişikliği + PC sapma analizi."""
        if self.db is None:
            messagebox.showerror("DB hazır değil",
                "Veritabanı bağlantısı henüz kurulmadı.")
            return
        LoggerForensikPopup(parent=self.root, db=self.db)

    def _birlesik_zaman_ac(self):
        """Üç zaman analizini tek pencerede gösterir:
        - RxIslemTarihi vs RxKayitTarihi (Zaman Tutarlılığı / Toplu Anomali)
        - RxIslemTarihi vs Logger.LoggerTarihi (Logger Forensik)
        - RxKontrolTarihi vs RxIslemTarihi (kontrol gecikme)
        """
        if self.db is None:
            messagebox.showerror("DB hazır değil",
                "Veritabanı bağlantısı henüz kurulmadı.")
            return
        BirlesikZamanAnaliziPopup(parent=self.root, db=self.db)

    # ------------------------------------------------------------------
    # Dönem Detay Popup
    # ------------------------------------------------------------------
    def _donem_detayini_goster(self):
        """Seçili dönem için ayrı pencerede tüm satışların detayını göster."""
        # 1. DB hazır mı?
        if self.db is None:
            messagebox.showerror("DB hazır değil",
                "Veritabanı bağlantısı henüz kurulmadı. Birkaç saniye sonra deneyin.")
            return
        # 2. Önce sorgu çalışmış mı (son_rapor dolu mu)?
        if not self.son_rapor:
            messagebox.showinfo("Sorgu yok",
                "Önce tarih + filtre seçip 🔍 Sorgula butonuna basın. "
                "Tabloda satır olmadan detay görüntülenemez.")
            return
        # 3. Satır seçimi var mı?
        sec = self.tree.selection()
        if not sec:
            messagebox.showinfo("Seçim yok",
                "Önce tablodan bir dönem satırı seçin "
                "(satıra tek tıkla mavi yap, sonra butona bas — "
                "veya satıra çift-tıkla).")
            return
        sec_id = sec[0]
        vals = self.tree.item(sec_id, 'values')
        if not vals:
            messagebox.showerror("Satır boş",
                "Seçili satır boş görünüyor. Başka satır seçip tekrar deneyin.")
            return
        donem_str = vals[0]
        if donem_str == 'TOPLAM':
            messagebox.showinfo("Toplam satırı",
                "Toplam satırının detayı yok. Lütfen bir dönem satırı seçin.")
            return

        # Seçili satırın indeksi → son_rapor'daki dönem
        try:
            idx = self.tree.index(sec_id)
        except Exception as e:
            messagebox.showerror("İndex hatası", f"Satır indeksi alınamadı:\n{e}")
            return
        if idx >= len(self.son_rapor):
            messagebox.showerror(
                "Satır eşleşmedi",
                f"Bu satır rapor verisiyle eşleşmiyor (idx={idx}, "
                f"rapor satır sayısı={len(self.son_rapor)}). "
                f"Lütfen tekrar Sorgula'ya bas ve tekrar dene."
            )
            return
        donem = self.son_rapor[idx].get('Donem')
        if isinstance(donem, datetime):
            donem = donem.date()
        elif not isinstance(donem, date):
            donem_str_raw = str(donem).split(' ')[0].strip()
            for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y/%m/%d', '%d/%m/%Y'):
                try:
                    donem = datetime.strptime(donem_str_raw, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                messagebox.showerror(
                    "Hata",
                    f"Dönem parse edilemedi: {donem!r} (beklenen format: "
                    f"YYYY-MM-DD, DD.MM.YYYY)"
                )
                return

        # Dönem aralığı
        per = self.periyot_var.get()
        if per == 'gunluk':
            bas, bit = donem, donem
        elif per == 'haftalik':
            bas = donem
            bit = donem + timedelta(days=6)
        elif per == 'aylik':
            bas = donem
            if donem.month == 12:
                bit = date(donem.year + 1, 1, 1) - timedelta(days=1)
            else:
                bit = date(donem.year, donem.month + 1, 1) - timedelta(days=1)
        elif per == '3aylik':
            bas = donem
            son_ay = donem.month + 2
            if son_ay > 12:
                bit = date(donem.year + 1, son_ay - 12 + 1, 1) - timedelta(days=1)
            else:
                bit = date(donem.year, son_ay + 1, 1) - timedelta(days=1)
        elif per == 'yillik':
            bas = donem
            bit = date(donem.year + 1, 1, 1) - timedelta(days=1)
        else:
            bas = bit = donem

        kirilim = self.kirilim_var.get()
        kurum_id = None
        if kirilim == 'kurum':
            secim = self.kurum_var.get().strip()
            try:
                kurum_id = int(secim.split(' - ', 1)[0])
            except Exception:
                pass

        # PERFORMANS FIX (2026-05-18): Nöbet modunda günlük satır için sadece
        # shift'in DATETIME sınırını çek (önceki gün 19:00 → ertesi gün 09:00 = 38 saat).
        # Eski yöntem 3 günlük tüm günü çekiyordu (yavaş). Yeni: dar saat aralığı.
        nobet_aktif = self.nobet_modu_var.get()
        sorgu_bas, sorgu_bit = bas, bit  # default: date
        if nobet_aktif and per == 'gunluk':
            # Gece shift sınırı: önceki gun 19:00 → bas+1 gun 09:00 (38 saat)
            # Hem önceki gunun gece nöbet sonu sabah + bu günün gece nöbeti dahil
            sorgu_bas = datetime.combine(bas - timedelta(days=1),
                                          datetime.min.time()).replace(hour=19)
            sorgu_bit = datetime.combine(bas + timedelta(days=1),
                                          datetime.min.time()).replace(hour=9)
        elif nobet_aktif and per == 'haftalik':
            # Hafta: 1 gün margin (datetime değil — daha geniş)
            sorgu_bas = bas - timedelta(days=1)
            sorgu_bit = bit + timedelta(days=1)

        DonemDetayPopup(
            parent=self.root,
            db=self.db,
            nobet_takvim=self.nobet_takvim,
            donem_baslangic=sorgu_bas,
            donem_bitis=sorgu_bit,
            donem_etiketi=donem_str,
            kirilim=kirilim,
            kurum_id=kurum_id,
            nobet_modu_aktif=nobet_aktif,
            # Nöbet shift sınırı için orijinal dönem (margin'siz)
            shift_filtre_tarih=bas if nobet_aktif and per == 'gunluk' else None,
        )


class DonemDetayPopup:
    """Belirli bir dönemin TÜM satışlarını listeleyen detay (raw) penceresi.

    Detay penceresi raw görünümdür — shoulder filtresi veya nöbet eşik filtresi
    uygulanmaz. Dönem aralığındaki her satış görünür; kullanıcı T1/T2 datetime
    aralıkları + VEYA/VE bağlacı ile kendi filtresini kurar.

    Her satışın shift_tipi (mesai/gece/pazar/tatil) bilgi olarak gösterilir
    ama filtrelemeye karışmaz — kullanıcı kendi belirler.
    """

    OPERATOR_SECENEKLERI = [
        ('sadece_t1', 'SADECE T1'),
        ('veya',      'T1 VEYA T2 (her ikisi de)'),
        ('ve',        'T1 VE T2 (kesişim — nadiren işe yarar)'),
        ('sadece_t2', 'SADECE T2'),
    ]

    def __init__(self, parent, db, nobet_takvim, donem_baslangic, donem_bitis,
                 donem_etiketi: str, kirilim: str = 'tumu',
                 kurum_id: Optional[int] = None,
                 nobet_modu_aktif: bool = False,
                 shift_filtre_tarih=None):
        self.parent = parent
        self.db = db
        self.nobet_takvim = nobet_takvim
        if self.nobet_takvim is None:
            try:
                from nobet_takvimi import get_nobet_takvimi
                self.nobet_takvim = get_nobet_takvimi()
            except Exception:
                self.nobet_takvim = None
        self.nobet_modu_aktif = nobet_modu_aktif
        # SQL sorgusu için margin'li aralık
        self.donem_bas = donem_baslangic
        self.donem_bit = donem_bitis
        # Görüntüleme + shift filtresi için orijinal hedef tarih (None ise margin yok)
        self.shift_filtre_tarih = shift_filtre_tarih
        self.donem_etiketi = donem_etiketi
        self.kirilim = kirilim
        self.kurum_id = kurum_id
        self.son_detay: List[Dict] = []
        # Shoulder activity maps (gösterim amaçlı)
        self._sabah_derin: Dict = {}
        self._aksam_derin: Dict = {}

        self.top = tk.Toplevel(parent)
        self.top.title(f"Dönem Detayı (RAW) — {donem_etiketi}")
        # Ekrana sığacak boyut + tam genişlik için zoomed dene
        self.top.geometry("1300x680")
        try:
            self.top.state('zoomed')  # Windows'ta tam ekran
        except Exception:
            pass
        self.top.transient(parent)

        try:
            self._arayuz()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"DonemDetayPopup._arayuz HATA: {e}\n{tb}")
            print(f"[Detay Popup] _arayuz HATA: {e}\n{tb}")
            try:
                tk.Label(
                    self.top,
                    text=f"❌ Arayüz oluşturulamadı:\n\n{e}\n\n"
                         f"(Tam hata için terminal/log)",
                    fg='white', bg='#B71C1C',
                    font=("Segoe UI", 11, "bold"),
                    padx=20, pady=20, justify='left', wraplength=900,
                ).pack(fill="both", expand=True, padx=20, pady=20)
            except Exception:
                pass
            return
        try:
            self._verileri_yukle()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"DonemDetayPopup._verileri_yukle HATA: {e}\n{tb}")
            print(f"[Detay Popup] _verileri_yukle HATA: {e}\n{tb}")
            try:
                self._yukleme_hatasi(str(e))
            except Exception:
                pass

    def _arayuz(self):
        # Header
        header = tk.Frame(self.top, bg='#FF8F00', height=42)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(
            header,
            text=f"🔎 Dönem Detayı (RAW) — {self.donem_etiketi}   "
                 f"({self.donem_bas} → {self.donem_bit})",
            font=("Segoe UI", 11, "bold"), bg='#FF8F00', fg='white'
        ).pack(side="left", padx=15, pady=8)

        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self.top.destroy).pack(side="right", padx=10, pady=6)

        tk.Button(header, text="📥 Excel", bg='#2E7D32', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._excel_aktar).pack(side="right", padx=5, pady=6)

        tk.Button(header, text="🕒 Zaman Tutarlılığı", bg='#5E35B1', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._zaman_tutarlilik_ac).pack(side="right", padx=5, pady=6)

        # YÜKLEME BAR'I — header'ın hemen altında, BÜYÜK ve GÖRÜNÜR
        self.loading_frame = tk.Frame(self.top, bg='#FFF8E1', height=40)
        self.loading_frame.pack(fill="x", padx=0, pady=0)
        self.loading_frame.pack_propagate(False)
        self.loading_lbl = tk.Label(
            self.loading_frame,
            text="⏳ Veri çekiliyor...",
            bg='#FFF8E1', fg='#5D4037',
            font=("Segoe UI", 10, "bold"), anchor='w', padx=15
        )
        self.loading_lbl.pack(side="left", fill="both", expand=True)
        self.progress_top = ttk.Progressbar(
            self.loading_frame, mode='indeterminate', length=250
        )
        self.progress_top.pack(side="right", padx=15, pady=8)

        # Üst panel: özet + saat dağılımı
        ozet_frame = ttk.LabelFrame(self.top, text="Özet ve Saat Dağılımı", padding=5)
        ozet_frame.pack(fill="x", padx=5, pady=5)

        self.ozet_lbl = tk.Label(
            ozet_frame, text="Yükleniyor...",
            font=("Segoe UI", 10), anchor='w', justify='left'
        )
        self.ozet_lbl.pack(side="left", padx=10, fill="x", expand=True)

        # Saat dağılımı (text tabanlı çubuk grafik — 4 metrik: Satış/Reçete/Kutu/TL)
        self.saat_text = tk.Text(
            ozet_frame, height=14, width=78, font=("Consolas", 9),
            bg='#FAFAFA', relief='flat', wrap='none'
        )
        self.saat_text.pack(side="right", padx=10, pady=5)
        # Dağılım çubuğu için hangi metriği baz alacak — combobox
        # NOT: ttk.LabelFrame.cget('background') TclError fırlatabiliyor;
        # sabit renk kullan.
        bar_metrik_frame = tk.Frame(ozet_frame, bg='#F5F5F5')
        bar_metrik_frame.pack(side="right", padx=5, pady=5, anchor="n")
        tk.Label(bar_metrik_frame, text="Dağılım çubuğu:",
                 bg='#F5F5F5', font=("Segoe UI", 8)).pack(anchor="w")
        self.bar_metrik_var = tk.StringVar(value='satis')
        for kod, etiket in [('satis', 'Satış'), ('recete', 'Reçete'),
                             ('kutu', 'Kutu'), ('tl', 'TL')]:
            tk.Radiobutton(
                bar_metrik_frame, text=etiket, variable=self.bar_metrik_var, value=kod,
                bg='#F5F5F5', font=("Segoe UI", 8), command=self._ozet_guncelle
            ).pack(anchor="w")

        # === Filtre bloğu (T1 ve T2 datetime aralıkları + bağlaç) ===
        filt_block = ttk.LabelFrame(
            self.top,
            text="Filtre: İki tarih+saat+saniye aralığı (T1 ve T2) — VEYA/VE bağlacı",
            padding=8
        )
        filt_block.pack(fill="x", padx=5, pady=4)

        # Kaynak filtre
        kaynak_row = tk.Frame(filt_block)
        kaynak_row.pack(fill="x", pady=2)
        tk.Label(kaynak_row, text="Kaynak:", font=("Segoe UI", 9, "bold")
                 ).pack(side="left", padx=4)
        self.kaynak_filtre = ttk.Combobox(
            kaynak_row, values=["Tümü", "RECETE", "ELDEN"], width=10, state="readonly"
        )
        self.kaynak_filtre.set("Tümü")
        self.kaynak_filtre.pack(side="left", padx=2)
        self.kaynak_filtre.bind('<<ComboboxSelected>>', lambda e: self._tabloyu_doldur())

        # Bağlaç
        tk.Label(kaynak_row, text="    Bağlaç:", font=("Segoe UI", 9, "bold")
                 ).pack(side="left", padx=(20, 4))
        self.operator_var = tk.StringVar(value='sadece_t1')
        operator_etiketler = [et for _, et in self.OPERATOR_SECENEKLERI]
        self.operator_combo = ttk.Combobox(
            kaynak_row, values=operator_etiketler, width=32, state="readonly"
        )
        self.operator_combo.set(operator_etiketler[0])
        self.operator_combo.pack(side="left", padx=2)

        # Preset butonları
        preset_frame = tk.Frame(kaynak_row)
        preset_frame.pack(side="right", padx=10)
        tk.Label(preset_frame, text="Hızlı preset:", font=("Segoe UI", 9, "italic")
                 ).pack(side="left", padx=4)
        ttk.Button(preset_frame, text="🌐 Tüm dönem",
                   command=lambda: self._preset_uygula('tum')).pack(side="left", padx=2)
        ttk.Button(preset_frame, text="🌙 Gece nöbeti (00-09 + 19-24)",
                   command=lambda: self._preset_uygula('gece')).pack(side="left", padx=2)
        ttk.Button(preset_frame, text="☀ Mesai (09-19)",
                   command=lambda: self._preset_uygula('mesai')).pack(side="left", padx=2)

        # T1 satırı
        t1_row = tk.Frame(filt_block)
        t1_row.pack(fill="x", pady=4)
        tk.Label(t1_row, text="T1:", font=("Segoe UI", 10, "bold"), fg='#1565C0'
                 ).pack(side="left", padx=(4, 8))
        tk.Label(t1_row, text="Başlangıç (YYYY-MM-DD HH:MM:SS):"
                 ).pack(side="left", padx=2)
        self.t1_bas_e = ttk.Entry(t1_row, width=22)
        self.t1_bas_e.pack(side="left", padx=2)
        tk.Label(t1_row, text="→  Bitiş:").pack(side="left", padx=8)
        self.t1_bit_e = ttk.Entry(t1_row, width=22)
        self.t1_bit_e.pack(side="left", padx=2)

        # T2 satırı
        t2_row = tk.Frame(filt_block)
        t2_row.pack(fill="x", pady=4)
        tk.Label(t2_row, text="T2:", font=("Segoe UI", 10, "bold"), fg='#D84315'
                 ).pack(side="left", padx=(4, 8))
        tk.Label(t2_row, text="Başlangıç (YYYY-MM-DD HH:MM:SS):"
                 ).pack(side="left", padx=2)
        self.t2_bas_e = ttk.Entry(t2_row, width=22)
        self.t2_bas_e.pack(side="left", padx=2)
        tk.Label(t2_row, text="→  Bitiş:").pack(side="left", padx=8)
        self.t2_bit_e = ttk.Entry(t2_row, width=22)
        self.t2_bit_e.pack(side="left", padx=2)

        ttk.Button(t2_row, text="Filtreyi Uygula", command=self._tabloyu_doldur, width=16
                   ).pack(side="left", padx=20)

        self.satir_say_lbl = tk.Label(t2_row, text="", fg='#1976D2',
                                       font=("Segoe UI", 9, "italic"))
        self.satir_say_lbl.pack(side="right", padx=15)

        # Varsayılan preset:
        # - nöbet modu + shift_filtre_tarih varsa: shift sınırına göre T1/T2
        #   (örn. hedef = 2026-03-26 → T1 = 2026-03-26 19:00 → 2026-03-27 09:00)
        # - nöbet modu (genel) → gece preset (T1=sabah, T2=akşam, VEYA)
        # - Normal → Tüm dönem
        if self.nobet_modu_aktif and self.shift_filtre_tarih is not None:
            self._preset_uygula('nobet_shift')
        elif self.nobet_modu_aktif:
            self._preset_uygula('gece')
        else:
            self._preset_uygula('tum')

        # Tablo + alt detay paneli — yatay paned window (üst=satışlar, alt=seçili satış detayı)
        paned = ttk.PanedWindow(self.top, orient='vertical')
        paned.pack(fill="both", expand=True, padx=5, pady=5)

        tab_frame = tk.Frame(paned)
        paned.add(tab_frame, weight=3)

        cols = ('zaman', 'kaynak', 'rxid', 'kalem', 'kutu', 'tl', 'kasa',
                'shift_tipi')
        bash = {'zaman': 'Zaman', 'kaynak': 'Kaynak', 'rxid': 'Rx Id',
                'kalem': 'Kalem', 'kutu': 'Kutu', 'tl': 'TL',
                'kasa': '💵 Kasa ₺',
                'shift_tipi': 'Shift Tipi'}
        widths = {'zaman': 180, 'kaynak': 90, 'rxid': 100, 'kalem': 60,
                  'kutu': 70, 'tl': 120, 'kasa': 130, 'shift_tipi': 100}

        self.tree = ttk.Treeview(tab_frame, columns=cols, show='headings', height=14)
        for c in cols:
            self.tree.heading(c, text=bash[c])
            self.tree.column(c, width=widths[c], anchor='center')
        self.tree.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(tab_frame, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.tag_configure('recete', background='#E3F2FD')
        self.tree.tag_configure('elden', background='#FFF3E0')

        # Satır seçilince alt detay panelini güncelle (debounce'lu)
        self._detay_debounce_id = None
        self.tree.bind('<<TreeviewSelect>>', self._secili_satir_degisti)

        # Alt: seçili satış detayı paneli
        self._detay_paneli_olustur(paned)
        paned.add(self._detay_dis_frame, weight=2)

        # Status + progress
        bar = tk.Frame(self.top, bg='#37474F', height=24)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        self.status_lbl = tk.Label(bar, text="Hazırlanıyor...",
                                    bg='#37474F', fg='white',
                                    font=("Segoe UI", 9), anchor='w', padx=10)
        self.status_lbl.pack(side="left", fill="x", expand=True)
        self.progress = ttk.Progressbar(bar, mode='indeterminate', length=180)
        # pack() sorgu sırasında

    def _verileri_yukle(self):
        """RAW görünüm — shoulder/nöbet filtresi YOK, dönem aralığındaki tüm satışlar."""
        import time as _time

        def _calis():
            t0 = _time.time()
            try:
                # Log: tam SQL parametreleri (console + status için)
                print(f"[Detay Popup] SQL çağrısı: bas={self.donem_bas}, "
                      f"bit={self.donem_bit}, kirilim={self.kirilim}")
                detay = self.db.satis_zaman_detay_getir(
                    baslangic_tarih=self.donem_bas,
                    bitis_tarih=self.donem_bit,
                    kirilim=self.kirilim,
                    kurum_id=self.kurum_id,
                )
                t_sql = _time.time() - t0
                print(f"[Detay Popup] SQL bitti: {len(detay)} satır, {t_sql:.1f}sn")

                if self.nobet_takvim is not None:
                    try:
                        self._sabah_derin, self._aksam_derin = \
                            self.nobet_takvim._shoulder_aktivite_haritalari(detay)
                    except Exception:
                        self._sabah_derin = {}
                        self._aksam_derin = {}

                for s in detay:
                    s['shift_tipi'] = self._satis_tipini_belirle(s.get('tarih'))
                self.son_detay = detay
                self._sql_suresi = t_sql

                self.top.after(0, self._verileri_yuklendi)
            except Exception as e:
                t_err = _time.time() - t0
                logger.error(f"Detay yükleme hatası ({t_err:.1f}sn sonra): {e}",
                             exc_info=True)
                print(f"[Detay Popup] HATA ({t_err:.1f}sn): {e}")
                self.top.after(0, lambda: self._yukleme_hatasi(str(e)))

        # Loading: ÜST loading bar (büyük, görünür)
        self.progress_top.start(10)
        bas_str = (self.donem_bas.strftime('%Y-%m-%d %H:%M')
                    if hasattr(self.donem_bas, 'hour')
                    else str(self.donem_bas))
        bit_str = (self.donem_bit.strftime('%Y-%m-%d %H:%M')
                    if hasattr(self.donem_bit, 'hour')
                    else str(self.donem_bit))
        self.loading_lbl.config(
            text=f"⏳ Botanik EOS'tan veri çekiliyor: {bas_str} → {bit_str}  "
                 f"({self.kirilim}) — sorgu sürerken bu mesaj kalır."
        )
        # Alt status bar da güncelle (yedek)
        self.progress.pack(side="right", padx=10)
        self.progress.start(10)
        self._status(f"⏳ Sorgu çalışıyor: {bas_str} → {bit_str}")
        threading.Thread(target=_calis, daemon=True).start()

    def _verileri_yuklendi(self):
        """Veri yüklendi — progress kapat, tabloyu doldur, durum güncelle."""
        self.progress.stop()
        self.progress.pack_forget()
        self.progress_top.stop()
        n = len(self.son_detay)
        sure = getattr(self, '_sql_suresi', 0)
        if n == 0:
            mesaj = (f"⚠ Bu aralıkta HİÇ SATIŞ YOK ({sure:.1f}sn'de yanıt). "
                     f"Tarih/kırılım uyumsuz olabilir.")
            self.loading_lbl.config(text=mesaj, bg='#FFCDD2', fg='#B71C1C')
            self.loading_frame.config(bg='#FFCDD2')
            self._status(mesaj, hata=True)
            return
        mesaj = (f"✓ {n} satış çekildi ({sure:.1f}sn). Filtre uygulanıyor — "
                 f"sonuç altta gösterilecek.")
        self.loading_lbl.config(text=mesaj, bg='#C8E6C9', fg='#1B5E20')
        self.loading_frame.config(bg='#C8E6C9')
        self._status(mesaj)
        # 3 saniye sonra loading bar'ı gizle
        self.top.after(3000, self._loading_gizle)
        self._tabloyu_doldur()
        self._ozet_guncelle()

    def _loading_gizle(self):
        try:
            self.loading_frame.pack_forget()
        except Exception:
            pass

    def _yukleme_hatasi(self, mesaj: str):
        self.progress.stop()
        self.progress.pack_forget()
        self.progress_top.stop()
        sure = getattr(self, '_sql_suresi', 0)
        kisa = mesaj[:200]
        self.loading_lbl.config(
            text=f"❌ HATA ({sure:.1f}sn): {kisa}",
            bg='#FFCDD2', fg='#B71C1C'
        )
        self.loading_frame.config(bg='#FFCDD2')
        self._status(f"❌ Sorgu hatası: {kisa}", hata=True)
        messagebox.showerror("Sorgu hatası", f"Veri çekilemedi:\n{mesaj}")

    def _satis_tipini_belirle(self, dt) -> str:
        """Bilgi amaçlı: bu satış hangi shift kategorisindedir?
        'mesai' | 'mesai(shoulder)' | 'gece' | 'pazar' | 'tatil' | '-'
        """
        if not hasattr(dt, 'hour') or self.nobet_takvim is None:
            return '-'
        try:
            # Shoulder mı?
            if self.nobet_takvim._shoulder_mesai_mi(dt, self._sabah_derin, self._aksam_derin):
                return 'mesai(shoulder)'
            sh = self.nobet_takvim.nobet_shift(dt)
            if sh is None:
                return 'mesai'
            return sh['tip']
        except Exception:
            return '-'

    # ------------------------------------------------------------------
    # Preset ve filtre
    # ------------------------------------------------------------------
    def _operator_kod(self) -> str:
        secim = self.operator_combo.get()
        for kod, et in self.OPERATOR_SECENEKLERI:
            if et == secim:
                return kod
        return 'sadece_t1'

    def _preset_uygula(self, tip: str):
        """T1/T2 alanlarını doldurmak için hızlı seçenek.

        - 'tum': T1 = donem_bas 00:00:00 → donem_bit 23:59:59, T2 boş
        - 'gece': T1 = donem_bas 00:00:00 → 08:59:59 sabah, T2 = donem_bit 19:00:00 → 23:59:59 akşam, VEYA
        - 'mesai': T1 = donem_bas 09:00:00 → donem_bit 18:59:59, T2 boş
        - 'nobet_shift': hedef_gun gece nöbeti shift sınırı:
            T1 = hedef_gun 00:00:00 → 08:59:59 (Pazar nöbeti'nin sabah uzantısı)
            T2 = hedef_gun 19:00:00 → (hedef_gun+1) 08:59:59 (akşam→ertesi sabah)
            VEYA — shift'in tam aralığını kapsar.
        """
        def _doldur(entry, metin):
            entry.delete(0, tk.END)
            entry.insert(0, metin)

        if tip == 'tum':
            _doldur(self.t1_bas_e, f"{self.donem_bas} 00:00:00")
            _doldur(self.t1_bit_e, f"{self.donem_bit} 23:59:59")
            _doldur(self.t2_bas_e, "")
            _doldur(self.t2_bit_e, "")
            self.operator_combo.set(self.OPERATOR_SECENEKLERI[0][1])
        elif tip == 'gece':
            _doldur(self.t1_bas_e, f"{self.donem_bas} 00:00:00")
            _doldur(self.t1_bit_e, f"{self.donem_bas} 08:59:59")
            _doldur(self.t2_bas_e, f"{self.donem_bit} 19:00:00")
            _doldur(self.t2_bit_e, f"{self.donem_bit} 23:59:59")
            self.operator_combo.set(self.OPERATOR_SECENEKLERI[1][1])
        elif tip == 'mesai':
            _doldur(self.t1_bas_e, f"{self.donem_bas} 09:00:00")
            _doldur(self.t1_bit_e, f"{self.donem_bit} 18:59:59")
            _doldur(self.t2_bas_e, "")
            _doldur(self.t2_bit_e, "")
            self.operator_combo.set(self.OPERATOR_SECENEKLERI[0][1])
        elif tip == 'nobet_shift':
            # Nöbet shift'inin tam aralığı: hedef_gun
            from datetime import timedelta
            hedef = self.shift_filtre_tarih
            ertesi = hedef + timedelta(days=1)
            # Sabah uzantısı: hedef_gun 00-08:59 (önceki gece nöbetinin son saatleri,
            # ya da pazar/tatil ise hedef günün ilk saatleri)
            _doldur(self.t1_bas_e, f"{hedef} 00:00:00")
            _doldur(self.t1_bit_e, f"{hedef} 08:59:59")
            # Akşam → ertesi sabah: hedef 19:00 → ertesi 08:59
            _doldur(self.t2_bas_e, f"{hedef} 19:00:00")
            _doldur(self.t2_bit_e, f"{ertesi} 08:59:59")
            self.operator_combo.set(self.OPERATOR_SECENEKLERI[1][1])  # VEYA

        if hasattr(self, 'son_detay') and self.son_detay:
            self._tabloyu_doldur()

    @staticmethod
    def _datetime_parse(metin: str):
        metin = (metin or "").strip()
        if not metin:
            return None
        # Birkaç farklı format dene
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                    "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M",
                    "%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(metin, fmt)
            except ValueError:
                continue
        return None

    def _t_araliklarini_oku(self):
        """T1 ve T2 datetime aralıklarını oku. (t1_bas, t1_bit, t2_bas, t2_bit) tuple."""
        t1_bas = self._datetime_parse(self.t1_bas_e.get())
        t1_bit = self._datetime_parse(self.t1_bit_e.get())
        t2_bas = self._datetime_parse(self.t2_bas_e.get())
        t2_bit = self._datetime_parse(self.t2_bit_e.get())
        return t1_bas, t1_bit, t2_bas, t2_bit

    def _filtre_uygula(self) -> List[Dict]:
        kaynak = self.kaynak_filtre.get()
        operator = self._operator_kod()
        t1_bas, t1_bit, t2_bas, t2_bit = self._t_araliklarini_oku()

        def _in_aralik(dt, bas, bit):
            if dt is None or bas is None or bit is None:
                return False
            return bas <= dt <= bit

        sonuc = []
        for s in self.son_detay:
            if kaynak != "Tümü" and s.get('kaynak') != kaynak:
                continue
            dt = s.get('tarih')
            if dt is None:
                continue
            in_t1 = _in_aralik(dt, t1_bas, t1_bit)
            in_t2 = _in_aralik(dt, t2_bas, t2_bit)
            if operator == 'sadece_t1':
                if not in_t1:
                    continue
            elif operator == 'sadece_t2':
                if not in_t2:
                    continue
            elif operator == 'veya':
                if not (in_t1 or in_t2):
                    continue
            elif operator == 've':
                if not (in_t1 and in_t2):
                    continue
            sonuc.append(s)
        return sonuc

    # ------------------------------------------------------------------
    # Alt detay paneli (seçili satışın reçete/elden kalem detayı)
    # ------------------------------------------------------------------
    def _detay_paneli_olustur(self, parent):
        """Alt paned'e seçili satışın detayını gösteren paneli kur."""
        self._detay_dis_frame = ttk.LabelFrame(
            parent, text="🔍 Seçili Satış Detayı (üstte bir satıra tıklayın)",
            padding=4,
        )

        # Üst: başlık alanı (hasta/doktor/kurum/tarih)
        self._detay_baslik_frame = tk.Frame(self._detay_dis_frame, bg='#ECEFF1')
        self._detay_baslik_frame.pack(fill="x", padx=2, pady=(2, 4))

        self._detay_baslik_lbl = tk.Label(
            self._detay_baslik_frame,
            text="(Üstteki tablodan bir satır seçin — reçete/elden detayı burada görünür)",
            bg='#ECEFF1', fg='#37474F',
            font=("Segoe UI", 9), anchor='w', justify='left',
            padx=8, pady=6, wraplength=1200,
        )
        self._detay_baslik_lbl.pack(fill="x")

        # Alt: kalem tablosu
        kalem_dis = tk.Frame(self._detay_dis_frame)
        kalem_dis.pack(fill="both", expand=True, padx=2, pady=2)

        det_cols = ('barkod', 'urun', 'adet', 'etiket', 'kurum',
                    'fark', 'iskonto', 'toplam', 'iade')
        det_bash = {
            'barkod': 'Barkod', 'urun': 'Ürün',
            'adet': 'Adet', 'etiket': 'Etiket ₺', 'kurum': 'Kurum ₺',
            'fark': 'Fiyat Farkı ₺', 'iskonto': 'İskonto ₺',
            'toplam': 'Toplam ₺', 'iade': 'İade',
        }
        det_widths = {
            'barkod': 110, 'urun': 320,
            'adet': 60, 'etiket': 90, 'kurum': 90,
            'fark': 90, 'iskonto': 80,
            'toplam': 100, 'iade': 50,
        }
        det_anchor = {'urun': 'w', 'iade': 'center'}

        self._detay_tree = ttk.Treeview(
            kalem_dis, columns=det_cols, show='headings', height=6,
        )
        for c in det_cols:
            self._detay_tree.heading(c, text=det_bash[c])
            self._detay_tree.column(c, width=det_widths[c],
                                     anchor=det_anchor.get(c, 'e'))
        self._detay_tree.pack(side="left", fill="both", expand=True)

        det_sb = ttk.Scrollbar(kalem_dis, orient="vertical",
                                command=self._detay_tree.yview)
        det_sb.pack(side="right", fill="y")
        self._detay_tree.configure(yscrollcommand=det_sb.set)

        self._detay_tree.tag_configure('iade', background='#FFEBEE',
                                        foreground='#B71C1C')
        self._detay_tree.tag_configure('toplam', background='#FFF8E1',
                                        font=('Segoe UI', 9, 'bold'))

    def _secili_satir_degisti(self, _event=None):
        """Treeview select event — debounce'lu (200ms) detay yükle."""
        # Önceki bekleyen çağrıyı iptal et
        if self._detay_debounce_id is not None:
            try:
                self.top.after_cancel(self._detay_debounce_id)
            except Exception:
                pass
            self._detay_debounce_id = None
        self._detay_debounce_id = self.top.after(200, self._secili_detayi_yukle)

    def _secili_detayi_yukle(self):
        """Seçili satırın rx_id'sini al, DB'den detay çek, alt paneli doldur."""
        self._detay_debounce_id = None
        sec = self.tree.selection()
        if not sec:
            return
        vals = self.tree.item(sec[0], 'values')
        if not vals or len(vals) < 3:
            return
        # Tree kolon sırası: zaman, kaynak, rxid, kalem, kutu, tl, shift_tipi
        kaynak = (vals[1] or '').strip().upper()
        rx_id_str = (vals[2] or '').strip()

        # TOPLAM veya filtre uyarı satırı — atla
        if vals[0] == 'TOPLAM' or kaynak not in ('RECETE', 'ELDEN') or not rx_id_str:
            self._detay_baslik_lbl.config(
                text="(Bu satır bir satışa karşılık gelmiyor — TOPLAM veya bilgi satırı.)"
            )
            self._detay_tree_temizle()
            return

        self._detay_baslik_lbl.config(
            text=f"⏳ Yükleniyor: {kaynak} RxId={rx_id_str}..."
        )
        self._detay_tree_temizle()

        # DB sorgusu thread'de — UI bloklamasın
        def _calis():
            try:
                detay = self.db.satis_kalem_detay_getir(kaynak, rx_id_str)
                self.top.after(0, lambda: self._detay_goster(kaynak, rx_id_str, detay))
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Satış detayı yüklenirken hata: {e}\n{tb}")
                self.top.after(0, lambda: self._detay_baslik_lbl.config(
                    text=f"❌ Detay yüklenemedi: {e}"
                ))
        threading.Thread(target=_calis, daemon=True).start()

    def _detay_tree_temizle(self):
        for c in self._detay_tree.get_children():
            self._detay_tree.delete(c)

    def _detay_goster(self, kaynak: str, rx_id_str: str, detay: Dict):
        """DB sonucunu alt panele yaz."""
        baslik = detay.get('baslik') if detay else None
        kalemler = detay.get('kalemler') if detay else []

        if not baslik:
            baslik_hata = detay.get('baslik_hata') if detay else None
            kalem_hata = detay.get('kalem_hata') if detay else None
            mesaj = f"⚠ {kaynak} RxId={rx_id_str} bulunamadı veya silinmiş."
            if baslik_hata:
                mesaj += f"\n❌ Başlık SQL hatası: {baslik_hata}"
                logger.error(f"satis_kalem_detay_getir baslik_hata: {baslik_hata}")
            if kalem_hata:
                mesaj += f"\n❌ Kalem SQL hatası: {kalem_hata}"
                logger.error(f"satis_kalem_detay_getir kalem_hata: {kalem_hata}")
            self._detay_baslik_lbl.config(text=mesaj)
            self._detay_tree_temizle()
            return

        # Başlık metnini oluştur (RECETE / ELDEN farklı alanlar)
        def _g(d, k, varsayilan='—'):
            v = d.get(k)
            if v is None or v == '':
                return varsayilan
            return v

        islem = _g(baslik, 'RxIslemTarihi')
        kayit = _g(baslik, 'RxKayitTarihi')
        islem_str = (islem.strftime('%d.%m.%Y %H:%M:%S')
                     if hasattr(islem, 'strftime') else str(islem))
        kayit_str = (kayit.strftime('%d.%m.%Y %H:%M:%S')
                     if hasattr(kayit, 'strftime') else str(kayit))

        hasta_adi = _g(baslik, 'HastaAdi')
        hasta_tc = _g(baslik, 'HastaTCKN')

        # Toplam: kalemlerden hesapla (iadeleri çıkart)
        toplam_tl = sum(
            float(k.get('toplam', 0) or 0)
            for k in (kalemler or []) if not k.get('iade')
        )
        iade_tl = sum(
            float(k.get('toplam', 0) or 0)
            for k in (kalemler or []) if k.get('iade')
        )

        def _tl_fmt(v):
            return (f"{float(v):,.2f} ₺"
                    .replace(',', 'X').replace('.', ',').replace('X', '.'))

        if kaynak == 'RECETE':
            recete_no = _g(baslik, 'RxEReceteNo')
            doktor = _g(baslik, 'DoktorAdi')
            tesis = _g(baslik, 'TesisAdi')
            kurum = _g(baslik, 'KurumAdi')

            satirlar = [
                f"📋 REÇETE  •  RxId={baslik.get('RxId')}  •  "
                f"E-Reçete: {recete_no}",
                f"İşlem: {islem_str}    |    Kayıt: {kayit_str}",
                f"Hasta: {hasta_adi}  ({hasta_tc})",
                f"Doktor: {doktor}",
                f"Tesis: {tesis}    |    Kurum: {kurum}",
                f"Toplam: {_tl_fmt(toplam_tl)}"
                + (f"    İade: {_tl_fmt(iade_tl)}" if iade_tl else ""),
            ]
        else:  # ELDEN
            belge_no = _g(baslik, 'RxBelgeNo')
            satirlar = [
                f"🛒 ELDEN (Parakende)  •  RxId={baslik.get('RxId')}  •  "
                f"Belge: {belge_no}",
                f"İşlem: {islem_str}    |    Kayıt: {kayit_str}",
                f"Müşteri: {hasta_adi}"
                + (f"  ({hasta_tc})" if hasta_tc and hasta_tc != '—' else ''),
                f"Toplam: {_tl_fmt(toplam_tl)}"
                + (f"    İade: {_tl_fmt(iade_tl)}" if iade_tl else ""),
            ]

        self._detay_baslik_lbl.config(text="\n".join(satirlar))

        # Kalemleri tabloya yaz
        self._detay_tree_temizle()
        if not kalemler:
            self._detay_tree.insert('', 'end', values=(
                '', '(Bu satışta görünür kalem yok)', '', '', '', '', '', '', ''
            ))
            return

        def _f(v):
            try:
                return f"{float(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except Exception:
                return str(v) if v is not None else ''

        toplam_adet = 0.0
        toplam_etiket = 0.0
        toplam_kurum = 0.0
        toplam_fark = 0.0
        toplam_iskonto = 0.0
        toplam_top = 0.0
        for k in kalemler:
            iade_str = '⟲ İADE' if k.get('iade') else ''
            tag = ('iade',) if k.get('iade') else ()
            self._detay_tree.insert('', 'end', values=(
                k.get('barkod', ''),
                k.get('urun_adi', ''),
                f"{k.get('adet', 0):.0f}",
                _f(k.get('etiket_fiyat', 0)),
                _f(k.get('kurum_fiyat')) if k.get('kurum_fiyat') is not None else '—',
                _f(k.get('fiyat_farki')) if k.get('fiyat_farki') is not None else '—',
                _f(k.get('iskonto', 0)),
                _f(k.get('toplam', 0)),
                iade_str,
            ), tags=tag)
            if not k.get('iade'):
                toplam_adet += float(k.get('adet', 0) or 0)
                toplam_etiket += float(k.get('etiket_fiyat', 0) or 0) * float(k.get('adet', 0) or 0)
                if k.get('kurum_fiyat') is not None:
                    toplam_kurum += float(k.get('kurum_fiyat') or 0) * float(k.get('adet', 0) or 0)
                if k.get('fiyat_farki') is not None:
                    toplam_fark += float(k.get('fiyat_farki') or 0)
                toplam_iskonto += float(k.get('iskonto', 0) or 0)
                toplam_top += float(k.get('toplam', 0) or 0)

        self._detay_tree.insert('', 'end', values=(
            '', f'TOPLAM ({len(kalemler)} kalem)',
            f"{toplam_adet:.0f}",
            _f(toplam_etiket),
            _f(toplam_kurum) if toplam_kurum > 0 else '—',
            _f(toplam_fark) if toplam_fark > 0 else '—',
            _f(toplam_iskonto),
            _f(toplam_top),
            '',
        ), tags=('toplam',))

    def _tabloyu_doldur(self):
        for c in self.tree.get_children():
            self.tree.delete(c)
        filtre_sonuc = self._filtre_uygula()
        toplam_kalem = 0
        toplam_kutu = 0.0
        toplam_tl = 0.0
        for s in sorted(filtre_sonuc, key=lambda x: x.get('tarih') or datetime.min):
            dt = s.get('tarih')
            zaman = dt.strftime('%d.%m.%Y %H:%M:%S') if hasattr(dt, 'strftime') else str(dt)
            kaynak = s.get('kaynak', '')
            rx = s.get('rx_id', '')
            kalem = int(s.get('kalem_sayisi', 0) or 0)
            kutu = float(s.get('adet', 0) or 0)
            tl = float(s.get('tutar', 0) or 0)
            # Kasa etkisi: RECETE & ELDEN için RIToplam (= tl). Kasaya giren
            # nakit/kart tutarı. Şimdilik tl ile aynı (kullanıcı seçimi).
            kasa = tl
            shift_tipi = s.get('shift_tipi', '-')
            tag = 'recete' if kaynak == 'RECETE' else 'elden'
            self.tree.insert('', 'end', values=(
                zaman, kaynak, rx,
                f"{kalem}",
                f"{kutu:.0f}",
                f"{tl:,.2f} ₺".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"{kasa:,.2f} ₺".replace(',', 'X').replace('.', ',').replace('X', '.'),
                shift_tipi,
            ), tags=(tag,))
            toplam_kalem += kalem
            toplam_kutu += kutu
            toplam_tl += tl

        if filtre_sonuc:
            self.tree.insert('', 'end', values=(
                'TOPLAM', '', f"({len(filtre_sonuc)} satış)",
                f"{toplam_kalem}",
                f"{toplam_kutu:.0f}",
                f"{toplam_tl:,.2f} ₺".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"{toplam_tl:,.2f} ₺".replace(',', 'X').replace('.', ',').replace('X', '.'),
                '',
            ), tags=())

        self.satir_say_lbl.config(
            text=f"Filtre sonrası: {len(filtre_sonuc)} satış / "
                 f"Toplam: {len(self.son_detay)}"
        )

        # Filtre 0 sonuç verdi ama veri var → kullanıcıyı bilgilendir
        if len(self.son_detay) > 0 and len(filtre_sonuc) == 0:
            self.tree.insert('', 'end', values=(
                '', '', 'FİLTRE 0 SATIŞ DÖNDÜRDÜ', '', '',
                'Filtreyi gevşet veya', '"🌐 Tüm dönem" preset', '',
            ), tags=())
            self._status(
                f"⚠ {len(self.son_detay)} satış yüklendi ama filtre HİÇBİRİNİ "
                f"kapsamıyor. T1/T2 zaman aralığını kontrol et veya "
                f"\"🌐 Tüm dönem\" presetine bas.",
                hata=True
            )
        elif len(filtre_sonuc) > 0:
            self._status(
                f"✓ {len(filtre_sonuc)} satış gösteriliyor "
                f"(toplam yüklenen: {len(self.son_detay)})"
            )

    def _ozet_guncelle(self):
        from collections import defaultdict, Counter
        saat_d = defaultdict(lambda: {'satis': 0, 'recete': 0, 'elden': 0,
                                       'kutu': 0.0, 'tl': 0.0})
        tipler = Counter()
        toplam_recete = 0
        toplam_elden = 0
        toplam_tl = 0.0
        toplam_kutu = 0.0

        for s in self.son_detay:
            dt = s.get('tarih')
            if not hasattr(dt, 'hour'):
                continue
            h = dt.hour
            saat_d[h]['satis'] += 1
            saat_d[h]['kutu'] += float(s.get('adet', 0) or 0)
            saat_d[h]['tl'] += float(s.get('tutar', 0) or 0)
            if s.get('kaynak') == 'RECETE':
                saat_d[h]['recete'] += 1
                toplam_recete += 1
            else:
                saat_d[h]['elden'] += 1
                toplam_elden += 1
            toplam_tl += float(s.get('tutar', 0) or 0)
            toplam_kutu += float(s.get('adet', 0) or 0)
            if s.get('shift_tipi') and s.get('shift_tipi') != '-':
                tipler[s['shift_tipi']] += 1

        ozet_metin = (
            f"📊 ÖZET ({self.donem_etiketi})\n"
            f"  Toplam satış: {toplam_recete + toplam_elden} "
            f"(Reçete: {toplam_recete}, Elden: {toplam_elden})\n"
            f"  Toplam kutu: {toplam_kutu:,.0f}\n"
            f"  Toplam TL:  {toplam_tl:,.2f} ₺"
        )
        if tipler:
            ozet_metin += f"\n  Shift tipleri: {dict(tipler)}"
        self.ozet_lbl.config(text=ozet_metin)

        # Saat dağılımı text grafiği (4 metrik genişletilmiş)
        self.saat_text.delete('1.0', tk.END)
        if not saat_d:
            self.saat_text.insert(tk.END, "(saat dağılımı yok)")
            return

        bar_metrik = self.bar_metrik_var.get() if hasattr(self, 'bar_metrik_var') else 'satis'
        max_bar = max(d[bar_metrik] for d in saat_d.values()) or 1

        self.saat_text.insert(
            tk.END,
            f"{'Saat':<6} {'Satış':>6} {'Reçete':>7} {'Elden':>6} {'Kutu':>6} {'TL':>12}   Dağılım ({bar_metrik})\n"
        )
        self.saat_text.insert(tk.END, "-" * 75 + "\n")
        toplam_satir = {'satis': 0, 'recete': 0, 'elden': 0, 'kutu': 0.0, 'tl': 0.0}
        for h in sorted(saat_d.keys()):
            d = saat_d[h]
            bar_uzunluk = int((d[bar_metrik] / max_bar) * 18) if max_bar > 0 else 0
            bar = "█" * bar_uzunluk
            self.saat_text.insert(
                tk.END,
                f"{h:02d}:00  {d['satis']:>6} {d['recete']:>7} {d['elden']:>6} "
                f"{d['kutu']:>6.0f} {d['tl']:>11,.2f}   {bar}\n"
            )
            toplam_satir['satis'] += d['satis']
            toplam_satir['recete'] += d['recete']
            toplam_satir['elden'] += d['elden']
            toplam_satir['kutu'] += d['kutu']
            toplam_satir['tl'] += d['tl']

        self.saat_text.insert(tk.END, "-" * 75 + "\n")
        self.saat_text.insert(
            tk.END,
            f"{'TOPLAM':<6} {toplam_satir['satis']:>6} {toplam_satir['recete']:>7} "
            f"{toplam_satir['elden']:>6} {toplam_satir['kutu']:>6.0f} "
            f"{toplam_satir['tl']:>11,.2f}\n"
        )

    def _excel_aktar(self):
        if not self.son_detay:
            messagebox.showinfo("Boş", "Önce veriler yüklenmeli.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("openpyxl yok", "Excel export için openpyxl gerekli.")
            return

        # Dosya adı
        from tkinter import filedialog as fd
        dosya = fd.asksaveasfilename(
            title="Detay raporu kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"detay_{self.donem_etiketi.replace(' ','_').replace('.','-')}.xlsx"
        )
        if not dosya:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Detay"
        basliklar = ['Zaman', 'Kaynak', 'Rx Id', 'Kalem', 'Kutu', 'TL',
                     'Kasa ₺', 'Shift Tipi']
        bf = Font(bold=True, color='FFFFFF')
        bg = PatternFill(start_color='FF8F00', end_color='FF8F00', fill_type='solid')
        for ci, b in enumerate(basliklar, 1):
            c = ws.cell(row=1, column=ci, value=b)
            c.font = bf
            c.fill = bg
            c.alignment = Alignment(horizontal='center')

        filtre_sonuc = self._filtre_uygula()
        for ri, s in enumerate(sorted(filtre_sonuc, key=lambda x: x.get('tarih') or datetime.min), 2):
            dt = s.get('tarih')
            ws.cell(row=ri, column=1, value=dt.strftime('%Y-%m-%d %H:%M:%S') if hasattr(dt, 'strftime') else str(dt))
            ws.cell(row=ri, column=2, value=s.get('kaynak'))
            ws.cell(row=ri, column=3, value=s.get('rx_id'))
            ws.cell(row=ri, column=4, value=int(s.get('kalem_sayisi', 0) or 0))
            ws.cell(row=ri, column=5, value=float(s.get('adet', 0) or 0))
            tutar_v = float(s.get('tutar', 0) or 0)
            tl_c = ws.cell(row=ri, column=6, value=tutar_v)
            tl_c.number_format = '#,##0.00 ₺'
            kasa_c = ws.cell(row=ri, column=7, value=tutar_v)
            kasa_c.number_format = '#,##0.00 ₺'
            ws.cell(row=ri, column=8, value=s.get('shift_tipi', '-'))

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 14

        try:
            wb.save(dosya)
            self._status(f"Excel kaydedildi: {dosya}")
            messagebox.showinfo("Tamam", f"Excel dosyası kaydedildi:\n{dosya}")
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def _status(self, m: str, hata: bool = False):
        self.status_lbl.config(text=m, fg='#FFCDD2' if hata else 'white')

    def _zaman_tutarlilik_ac(self):
        """Mevcut T1/T2 aralığı için Zaman Tutarlılık penceresini aç."""
        t1_bas, t1_bit, t2_bas, t2_bit = self._t_araliklarini_oku()
        operator = self._operator_kod()

        # Analiz aralığını belirle (T1 + T2 birleşik kapsama)
        adaylar = []
        if t1_bas and t1_bit and operator in ('sadece_t1', 'veya', 've'):
            adaylar.extend([t1_bas, t1_bit])
        if t2_bas and t2_bit and operator in ('sadece_t2', 'veya', 've'):
            adaylar.extend([t2_bas, t2_bit])

        if not adaylar:
            messagebox.showinfo("Aralık yok",
                "T1 veya T2 aralığı girilmemiş. Önce filtre tarih+saatlerini doldurun.")
            return

        analiz_bas = min(adaylar)
        analiz_bit = max(adaylar)

        ZamanTutarlilikPopup(
            parent=self.top,
            db=self.db,
            analiz_bas=analiz_bas,
            analiz_bit=analiz_bit,
            etiket=self.donem_etiketi,
        )


class ZamanTutarlilikPopup:
    """Bir datetime aralığında satışların zaman damgalarını analiz eden pencere.

    4 analiz:
    1. RxIslemTarihi vs RxKayitTarihi karşılaştırma + fark istatistikleri
    2. Otomatik anomali sinyali yorumu (UTC sapması / toplu kayıt / vs.)
    3. RxId ardışıklık (toplu insert tespiti)
    4. ±14 gün kıyas tablosu (anormal yoğunluk var mı?)
    """

    def __init__(self, parent, db, analiz_bas: datetime, analiz_bit: datetime,
                 etiket: str = ""):
        self.parent = parent
        self.db = db
        self.analiz_bas = analiz_bas
        self.analiz_bit = analiz_bit
        self.etiket = etiket
        self.son_kayitlar: List[Dict] = []
        self.son_kiyas: List[Dict] = []
        # Sıralama durumu (tıklama ile değişir)
        self._son_sirala_sutun: Optional[str] = None
        self._son_sirala_ters: bool = False

        self.top = tk.Toplevel(parent)
        self.top.title(f"Zaman Tutarlılık Raporu — {etiket}")
        self.top.geometry("1400x800")
        self.top.transient(parent)

        self._arayuz()
        self._verileri_yukle()

    def _arayuz(self):
        header = tk.Frame(self.top, bg='#5E35B1', height=44)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(
            header,
            text=f"🕒 Zaman Tutarlılık Raporu — {self.analiz_bas} → {self.analiz_bit}",
            font=("Segoe UI", 11, "bold"), bg='#5E35B1', fg='white'
        ).pack(side="left", padx=15, pady=10)

        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self.top.destroy).pack(side="right", padx=10, pady=6)
        tk.Button(header, text="📥 Excel", bg='#2E7D32', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._excel_aktar).pack(side="right", padx=5, pady=6)
        tk.Button(header, text="🔎 Detay Gör", bg='#FF8F00', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._secili_detayi_ac).pack(side="right", padx=5, pady=6)

        # Kıyas gün sayısı + yeniden çalıştır butonu
        ctrl_frame = tk.Frame(self.top, bg='#F5F5F5')
        ctrl_frame.pack(fill="x", padx=5, pady=4)
        tk.Label(ctrl_frame, text="Kıyas penceresi (±gün):",
                 bg='#F5F5F5', font=("Segoe UI", 9)).pack(side="left", padx=4)
        self.kiyas_gun_var = tk.IntVar(value=14)
        ttk.Spinbox(ctrl_frame, from_=3, to=90, increment=1,
                    textvariable=self.kiyas_gun_var, width=5).pack(side="left", padx=2)
        ttk.Button(ctrl_frame, text="🔄 Yenile", command=self._verileri_yukle, width=10
                   ).pack(side="left", padx=8)

        self.status_lbl = tk.Label(ctrl_frame, text="", bg='#F5F5F5',
                                    font=("Segoe UI", 9, "italic"), fg='#1976D2')
        self.status_lbl.pack(side="right", padx=10)

        # Üst panel: özet + sinyal yorumu (yan yana)
        ust = ttk.LabelFrame(self.top, text="ÖZET & SİNYAL YORUMU", padding=8)
        ust.pack(fill="x", padx=5, pady=3)

        self.ozet_text = tk.Text(ust, height=8, width=60, font=("Consolas", 9),
                                  bg='#FAFAFA', relief='flat', wrap='word')
        self.ozet_text.pack(side="left", fill="both", expand=True, padx=4)

        self.sinyal_text = tk.Text(ust, height=8, width=60, font=("Segoe UI", 9),
                                    bg='#FFF8E1', relief='flat', wrap='word')
        self.sinyal_text.pack(side="right", fill="both", expand=True, padx=4)

        # Orta: kayıt tablosu
        kayit_frame = ttk.LabelFrame(
            self.top,
            text="1. KAYITLAR — RxIslemTarihi vs RxKayitTarihi",
            padding=4
        )
        kayit_frame.pack(fill="both", expand=True, padx=5, pady=3)

        cols = ('no', 'kaynak', 'rxid', 'islem', 'kayit', 'fark_sn', 'fark')
        bash = {'no': '#', 'kaynak': 'Kaynak', 'rxid': 'RxId',
                'islem': 'İşlem Tarihi', 'kayit': 'Kayıt Tarihi',
                'fark_sn': 'Fark (sn)', 'fark': 'Fark (okunabilir)'}
        widths = {'no': 40, 'kaynak': 70, 'rxid': 80,
                  'islem': 200, 'kayit': 200,
                  'fark_sn': 170, 'fark': 140}
        self.kayit_tree = ttk.Treeview(kayit_frame, columns=cols, show='headings', height=10)
        for c in cols:
            # Sıralanabilir başlık (no/sn dahil tümü)
            self.kayit_tree.heading(
                c, text=bash[c],
                command=lambda col=c: self._sirala(col),
            )
            self.kayit_tree.column(c, width=widths[c], anchor='center')
        self.kayit_tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(kayit_frame, orient="vertical", command=self.kayit_tree.yview)
        sb.pack(side="right", fill="y")
        self.kayit_tree.configure(yscrollcommand=sb.set)
        self.kayit_tree.tag_configure('negatif', background='#FFCDD2')
        self.kayit_tree.tag_configure('buyuk', background='#FFE082')
        self.kayit_tree.tag_configure('gun_anomali', background='#BBDEFB')
        self.kayit_tree.tag_configure('normal', background='white')
        self.kayit_tree.tag_configure('saatsiz', background='#ECEFF1',
                                       foreground='#546E7A')
        # Çift-tıkla → Detay Gör (butona basmakla aynı)
        self.kayit_tree.bind('<Double-1>', self._secili_detayi_ac)

        # Alt: RxId ardışıklık + kıyas tablosu (yan yana)
        alt = tk.Frame(self.top)
        alt.pack(fill="x", padx=5, pady=3)

        rxid_frame = ttk.LabelFrame(alt, text="2. RxId ARDIŞIKLIK", padding=4)
        rxid_frame.pack(side="left", fill="both", expand=True, padx=2)
        self.rxid_text = tk.Text(rxid_frame, height=6, font=("Consolas", 9),
                                  bg='#FAFAFA', relief='flat', wrap='word')
        self.rxid_text.pack(fill="both", expand=True)

        kiyas_frame = ttk.LabelFrame(alt, text="3. ±N GÜN KIYAS (saat aralığı)", padding=4)
        kiyas_frame.pack(side="right", fill="both", expand=True, padx=2)
        self.kiyas_text = tk.Text(kiyas_frame, height=10, font=("Consolas", 9),
                                   bg='#FAFAFA', relief='flat', wrap='none')
        self.kiyas_text.pack(fill="both", expand=True)

    def _set_status(self, m: str):
        self.status_lbl.config(text=m)

    def _sirala(self, sutun_kod: str):
        """Kayıt tablosunu seçilen sütuna göre sırala. Aynı sütuna ikinci
        tıkta yön ters çevrilir. Sıralama sonrası _raporu_doldur tekrar
        çağrılır — istatistikler değişmez, sadece sıra değişir."""
        if not self.son_kayitlar:
            return
        ters = (self._son_sirala_sutun == sutun_kod) and not self._son_sirala_ters
        self._son_sirala_sutun = sutun_kod
        self._son_sirala_ters = ters

        def _safe(v, default):
            return v if v is not None else default

        anahtarlar = {
            'no':      lambda r: 0,  # no görsel; sıralama anlamlı değil → değişmez
            'kaynak':  lambda r: r.get('Kaynak') or '',
            'rxid':    lambda r: _safe(r.get('RxId'), 0),
            'islem':   lambda r: _safe(r.get('IslemTarihi'), datetime.min),
            'kayit':   lambda r: _safe(r.get('KayitTarihi'), datetime.min),
            'fark_sn': lambda r: _safe(r.get('FarkSn'), 0),
            'fark':    lambda r: _safe(r.get('FarkSn'), 0),
        }
        key_fn = anahtarlar.get(sutun_kod, anahtarlar['islem'])
        try:
            self.son_kayitlar = sorted(self.son_kayitlar, key=key_fn, reverse=ters)
        except TypeError:
            # Bazı tarihler date / bazıları datetime ise karşılaştırma patlayabilir
            self.son_kayitlar = sorted(
                self.son_kayitlar,
                key=lambda r: str(key_fn(r)),
                reverse=ters,
            )
        # Başlığa ok simgesi koy
        self._sirala_baslik_isaretle(sutun_kod, ters)
        self._raporu_doldur()

    def _sirala_baslik_isaretle(self, aktif_sutun: str, ters: bool):
        """Sıralanan sütun başlığına ↓ veya ↑ ok ekle, diğerlerinden kaldır."""
        basliklar_orj = {'no': '#', 'kaynak': 'Kaynak', 'rxid': 'RxId',
                         'islem': 'İşlem Tarihi', 'kayit': 'Kayıt Tarihi',
                         'fark_sn': 'Fark (sn)', 'fark': 'Fark (okunabilir)'}
        ok = ' ↓' if ters else ' ↑'
        for c, orj in basliklar_orj.items():
            yeni = orj + (ok if c == aktif_sutun else '')
            try:
                self.kayit_tree.heading(c, text=yeni)
            except Exception:
                pass

    def _secili_detayi_ac(self, _event=None):
        """Seçili satırın RECETE/ELDEN detayını ayrı pencerede aç."""
        sec = self.kayit_tree.selection()
        if not sec:
            messagebox.showinfo("Seçim yok",
                "Önce 1. KAYITLAR tablosundan bir satır seçin "
                "(satıra tek tıkla mavi yap, sonra butona bas — "
                "veya satıra çift-tıkla).")
            return
        vals = self.kayit_tree.item(sec[0], 'values')
        # Kolon sırası: no, kaynak, rxid, islem, kayit, fark_sn, fark
        if not vals or len(vals) < 3:
            return
        kaynak = (vals[1] or '').strip().upper()
        rx_id = (vals[2] or '').strip()
        if kaynak not in ('RECETE', 'ELDEN') or not rx_id:
            messagebox.showinfo("Geçersiz satır",
                "Bu satır bir RECETE/ELDEN kaydına karşılık gelmiyor.")
            return
        SatisDetayPopup(parent=self.top, db=self.db,
                        kaynak=kaynak, rx_id=rx_id)

    def _verileri_yukle(self):
        self._set_status("Sorgular çalıştırılıyor...")

        def _calis():
            try:
                kayitlar = self.db.satis_zaman_tutarlilik_getir(
                    self.analiz_bas, self.analiz_bit
                )
                hedef_tarih = self.analiz_bas.date()
                kiyas_gun = self.kiyas_gun_var.get()
                kiyas_bas = hedef_tarih - timedelta(days=kiyas_gun)
                kiyas_bit = hedef_tarih + timedelta(days=kiyas_gun)
                kiyas = self.db.gun_saat_yogunlugu_getir(kiyas_bas, kiyas_bit)
                self.son_kayitlar = kayitlar
                self.son_kiyas = kiyas
                self.top.after(0, self._raporu_doldur)
            except Exception as e:
                logger.error(f"Tutarlılık sorgu hatası: {e}", exc_info=True)
                self.top.after(0, lambda: self._set_status(f"Hata: {e}"))

        threading.Thread(target=_calis, daemon=True).start()

    def _raporu_doldur(self):
        from collections import defaultdict, Counter

        kayitlar = self.son_kayitlar
        # Kayıt tablosu
        for c in self.kayit_tree.get_children():
            self.kayit_tree.delete(c)

        fark_degerleri = []
        fark_negatif = 0
        fark_buyuk = 0
        saatsiz_say = 0
        fark_sabit_aralikta = defaultdict(int)

        for i, k in enumerate(kayitlar, 1):
            islem = k['IslemTarihi']
            kayit = k['KayitTarihi']
            fark = k['FarkSn']

            def _dt_fmt(d):
                # datetime → tarih + saat; date (saat yok) → "tarih (saat yok)"
                if d is None:
                    return ''
                if hasattr(d, 'hour'):  # datetime
                    return d.strftime('%Y-%m-%d %H:%M:%S')
                if hasattr(d, 'strftime'):  # date (saat alanı yok)
                    return d.strftime('%Y-%m-%d') + ' (saat yok)'
                return str(d)

            islem_s = _dt_fmt(islem)
            kayit_s = _dt_fmt(kayit)

            # FarkTipi DB'den gelir: 'sn' veya 'gun'
            # 'gun' = kayıt saatsiz, fark gün bazlı (× 86400)
            fark_tipi = k.get('FarkTipi') or 'sn'

            tag = 'normal'
            if fark is None:
                fark_str = "?"
                fark_sn_str = "?"
            elif fark_tipi == 'gun':
                # Gün bazlı fark (kayıt saatsiz)
                saatsiz_say += 1
                gun = int(round(fark / 86400))
                sgn = '-' if gun < 0 else ''
                fark_sn_str = f"{gun} gün (kayıt saatsiz)"
                if gun == 0:
                    fark_str = "aynı gün (saatsiz)"
                    tag = 'saatsiz'
                else:
                    fark_str = f"{sgn}{abs(gun)} gün"
                    if gun < 0:
                        tag = 'negatif'
                        fark_str += " (NEGATİF)"
                    elif gun >= 1:
                        tag = 'gun_anomali'
                # İstatistik: gün-bazlı farkı saniye olarak ekleme (anlamsız);
                # ayrı sayaçta tut
            else:
                fark_degerleri.append(fark)
                # Fark (sn) sütunu: ham saniye + parantez içinde okunabilir
                if abs(fark) >= 60:
                    fark_sn_str = f"{fark} ({_saniye_okunabilir(fark)})"
                else:
                    fark_sn_str = f"{fark}"

                if fark < 0:
                    fark_negatif += 1
                    fark_str = f"{_saniye_okunabilir(fark)} (NEGATİF)"
                    tag = 'negatif'
                elif fark > 600:
                    fark_buyuk += 1
                    fark_str = _saniye_okunabilir(fark)
                    tag = 'buyuk'
                else:
                    fark_str = _saniye_okunabilir(fark)
                # Saat olarak yuvarla
                saat_fark = round(fark / 3600.0, 1)
                fark_sabit_aralikta[saat_fark] += 1

            self.kayit_tree.insert('', 'end', values=(
                i, k['Kaynak'], k['RxId'], islem_s, kayit_s, fark_sn_str, fark_str
            ), tags=(tag,))

        # Özet
        self.ozet_text.delete('1.0', tk.END)
        self.ozet_text.insert(tk.END, f"Aralık: {self.analiz_bas}\n  → {self.analiz_bit}\n\n")
        self.ozet_text.insert(tk.END, f"Toplam kayıt: {len(kayitlar)}\n")
        if saatsiz_say:
            self.ozet_text.insert(tk.END,
                f"⏱ Kayıt saatsiz (saat=00:00:00): {saatsiz_say} kayıt "
                f"— bu kayıtlarda fark gün bazında hesaplandı "
                f"(işlem saati yok sayıldı). Aşağıdaki saniye istatistikleri "
                f"sadece saatli kayıtlara ait.\n")
        if fark_degerleri:
            ort = sum(fark_degerleri) / len(fark_degerleri)
            med = sorted(fark_degerleri)[len(fark_degerleri) // 2]
            mn = min(fark_degerleri)
            mx = max(fark_degerleri)
            self.ozet_text.insert(tk.END,
                f"\nİşlem ↔ Kayıt farkı (saniye):\n"
                f"  Ortalama: {ort:.0f} ({ort/60:+.1f} dk)\n"
                f"  Medyan:   {med}\n"
                f"  Min:      {mn}\n  Max:      {mx}\n"
                f"\nNegatif fark: {fark_negatif} kayıt\n"
                f">10 dk fark: {fark_buyuk} kayıt\n"
            )
            self.ozet_text.insert(tk.END, "\nSabit fark dağılımı (~saat):\n")
            for sf, count in sorted(fark_sabit_aralikta.items()):
                yuzde = (count / len(kayitlar)) * 100
                self.ozet_text.insert(tk.END,
                    f"  {sf:>+6.1f} sa : {count} kayıt ({yuzde:.0f}%)\n")

        # Sinyal yorumu
        self.sinyal_text.delete('1.0', tk.END)
        self.sinyal_text.insert(tk.END, "SİNYAL YORUMU\n")
        self.sinyal_text.insert(tk.END, "=" * 40 + "\n\n")
        if not fark_degerleri:
            self.sinyal_text.insert(tk.END, "Yorumlanacak veri yok.\n")
        else:
            ort_fark = sum(fark_degerleri) / len(fark_degerleri)
            mx = max(fark_degerleri) if fark_degerleri else 0
            mn = min(fark_degerleri) if fark_degerleri else 0
            med = sorted(fark_degerleri)[len(fark_degerleri) // 2]
            farkli_saat_sayisi = len(set(fark_sabit_aralikta.keys()))

            # Saat dilimi sapması: tüm farklar ~aynı saat değerinde
            if (abs(ort_fark) > 3 * 3600 and farkli_saat_sayisi <= 2 and
                fark_negatif == 0):
                self.sinyal_text.insert(tk.END,
                    f"⚠ SAAT DİLİMİ SAPMASI\n"
                    f"Tüm farklar ~{ort_fark/3600:.1f} saat civarında.\n"
                    f"Muhtemelen server UTC, TR'ye çevrilmemiş.\n"
                    f"Gerçek satış saati ~{ort_fark/3600:.0f} saat sonrası olabilir.\n\n"
                )
            # Toplu/geç kayıt
            if mx > 24 * 3600 and med < 600:
                self.sinyal_text.insert(tk.END,
                    f"⚠ TOPLU/GEÇ KAYIT\n"
                    f"Bazı kayıtlar çok geç girilmiş (max {mx/3600:.1f} saat).\n"
                    f"Eczacı sonradan reçete kaydı yapmış olabilir.\n\n"
                )
            # Negatif fark = sistem saati bozuk
            if fark_negatif > len(kayitlar) * 0.3:
                self.sinyal_text.insert(tk.END,
                    f"⚠ SİSTEM SAATİ BOZUK?\n"
                    f"{fark_negatif} kayıtta kayıt zamanı işlem zamanından önce.\n"
                    f"Eczane bilgisayarının saati ileri olabilir.\n\n"
                )
            # Tutarlı
            if -10 <= med <= 60 and fark_negatif == 0 and fark_buyuk == 0:
                self.sinyal_text.insert(tk.END,
                    f"✓ ZAMAN DAMGALARI TUTARLI\n"
                    f"Tüm kayıtlar işlem anında DB'ye yazılmış (fark <1dk).\n"
                    f"İşlem tarihi GERÇEK satış zamanını yansıtıyor.\n"
                    f"Eğer bu saat anormal görünüyorsa: gerçek nöbet/anomali\n"
                    f"olabilir, saat hatası DEĞİL.\n\n"
                )

        # RxId ardışıklık
        self.rxid_text.delete('1.0', tk.END)
        recete_idler = sorted([k['RxId'] for k in kayitlar if k['Kaynak'] == 'RECETE'])
        elden_idler = sorted([k['RxId'] for k in kayitlar if k['Kaynak'] == 'ELDEN'])
        for ad, ids in [('Reçete', recete_idler), ('Elden', elden_idler)]:
            if not ids:
                self.rxid_text.insert(tk.END, f"{ad}: kayıt yok\n\n")
                continue
            self.rxid_text.insert(tk.END,
                f"{ad}: {len(ids)} kayıt | Min={ids[0]} Max={ids[-1]} | "
                f"Aralık genişliği={ids[-1]-ids[0]}\n"
            )
            if len(ids) > 1:
                ardisik = sum(1 for i in range(1, len(ids)) if ids[i] - ids[i-1] == 1)
                yuzde = (ardisik / (len(ids) - 1)) * 100
                self.rxid_text.insert(tk.END,
                    f"  Ardışık (Δ=1) çift: {ardisik}/{len(ids)-1} ({yuzde:.0f}%)\n"
                )
                if yuzde > 80:
                    self.rxid_text.insert(tk.END,
                        f"  ⚠ Çok ardışık → toplu kayıt veya hızlı arka arkaya işlem\n\n"
                    )
                elif yuzde > 50:
                    self.rxid_text.insert(tk.END,
                        f"  → Genelde ardışık, bazı boşluklar var (normal)\n\n"
                    )
                else:
                    self.rxid_text.insert(tk.END,
                        f"  → ID'ler dağılmış, başka reçeteler arasına serpişmiş\n"
                        f"    (NORMAL TRAFİK sinyali)\n\n"
                    )

        # ±N gün kıyas tablosu
        self.kiyas_text.delete('1.0', tk.END)
        hedef_tarih = self.analiz_bas.date()
        saat_bas = self.analiz_bas.hour
        saat_bit = self.analiz_bit.hour if self.analiz_bit.hour >= saat_bas else 23
        gun_saat = defaultdict(lambda: defaultdict(int))
        for r in self.son_kiyas:
            gun_saat[r['Tarih']][r['Saat']] = r['SatisSayisi']

        baslik = "Tarih       "
        for h in range(saat_bas, saat_bit + 1):
            baslik += f"{h:02d}".rjust(4)
        baslik += "  | Top.\n"
        self.kiyas_text.insert(tk.END, baslik)
        self.kiyas_text.insert(tk.END, "-" * (12 + 4 * (saat_bit - saat_bas + 1) + 9) + "\n")

        hedef_top = 0
        diger_toplar = []
        for d in sorted(gun_saat.keys()):
            prefix = '>' if d == hedef_tarih else ' '
            satir = f"{prefix} {d}  "
            gun_top = 0
            for h in range(saat_bas, saat_bit + 1):
                sayi = gun_saat[d].get(h, 0)
                gun_top += sayi
                satir += f"{sayi if sayi else '·':>4}"
            mark = '*' if d == hedef_tarih else ''
            satir += f"  | {gun_top}{mark}\n"
            self.kiyas_text.insert(tk.END, satir)
            if d == hedef_tarih:
                hedef_top = gun_top
            else:
                diger_toplar.append(gun_top)

        if diger_toplar:
            ort_diger = sum(diger_toplar) / len(diger_toplar)
            self.kiyas_text.insert(tk.END, "\n")
            self.kiyas_text.insert(tk.END,
                f"Hedef gün: {hedef_top}  vs  diğer günler ort: {ort_diger:.1f}\n")
            if hedef_top > 0 and ort_diger < 0.5:
                self.kiyas_text.insert(tk.END,
                    "→ Bu saatte diğer günler boş, sadece hedef günde yoğun.\n"
                    "  NÖBET veya ANORMAL bir gün olabilir.\n"
                )
            elif ort_diger > 0 and 0.5 <= hedef_top / max(ort_diger, 1) <= 2.0:
                self.kiyas_text.insert(tk.END,
                    "→ Hedef gün diğer günlerle BENZER.\n"
                    "  Bu saatler normalde de aktif → NORMAL TRAFİK\n"
                    "  (örn. saat dilimi sapması olası — gerçek saatler mesai içi).\n"
                )
            elif hedef_top > ort_diger * 5:
                self.kiyas_text.insert(tk.END,
                    "→ Hedef gün diğer günlerden ÇOK YÜKSEK.\n"
                    "  Toplu kayıt veya özel bir gün şüphesi.\n"
                )

        self._set_status(f"Hazır. {len(kayitlar)} kayıt analiz edildi.")

    def _excel_aktar(self):
        if not self.son_kayitlar:
            messagebox.showinfo("Boş", "Veri yok.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("openpyxl yok", "Excel için openpyxl gerekli.")
            return

        dosya = filedialog.asksaveasfilename(
            title="Zaman Tutarlılık Raporu",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"zaman_tutarlilik_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not dosya:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Tutarlılık"
        beyaz = Font(bold=True, color='FFFFFF')
        mor = PatternFill(start_color='5E35B1', end_color='5E35B1', fill_type='solid')
        basliklar = ['#', 'Kaynak', 'RxId', 'IslemTarihi', 'KayitTarihi', 'FarkSn']
        for ci, b in enumerate(basliklar, 1):
            c = ws.cell(row=1, column=ci, value=b)
            c.font = beyaz
            c.fill = mor
            c.alignment = Alignment(horizontal='center')
        for ri, k in enumerate(self.son_kayitlar, 2):
            ws.cell(row=ri, column=1, value=ri-1)
            ws.cell(row=ri, column=2, value=k['Kaynak'])
            ws.cell(row=ri, column=3, value=k['RxId'])
            islem = k['IslemTarihi']
            kayit = k['KayitTarihi']
            ws.cell(row=ri, column=4, value=islem.strftime('%Y-%m-%d %H:%M:%S') if hasattr(islem, 'strftime') else str(islem))
            ws.cell(row=ri, column=5, value=kayit.strftime('%Y-%m-%d %H:%M:%S') if hasattr(kayit, 'strftime') else str(kayit))
            ws.cell(row=ri, column=6, value=k['FarkSn'])
        for col_letter, w in zip(['A','B','C','D','E','F'], [6, 10, 12, 22, 22, 12]):
            ws.column_dimensions[col_letter].width = w
        try:
            wb.save(dosya)
            messagebox.showinfo("Tamam", f"Excel kaydedildi:\n{dosya}")
        except Exception as e:
            messagebox.showerror("Hata", str(e))


class SatisDetayPopup:
    """Tek bir satışın (RECETE/ELDEN) ayrı pencerede detayı.

    Zaman Tutarlılığı ve Toplu Zaman Anomalisi pencerelerinden 'Detay Gör'
    butonu ile açılır. DonemDetayPopup'taki alt panel detayıyla aynı
    görseli sunar ama bağımsız pencerede.
    """

    def __init__(self, parent, db, kaynak: str, rx_id):
        self.parent = parent
        self.db = db
        self.kaynak = (kaynak or '').upper().strip()
        self.rx_id = rx_id

        self.top = tk.Toplevel(parent)
        self.top.title(f"Satış Detayı — {self.kaynak} RxId={rx_id}")
        self.top.geometry("1200x650")
        self.top.transient(parent)

        try:
            self._arayuz()
            self._yukle()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"SatisDetayPopup HATA: {e}\n{tb}")
            print(f"[Satış Detayı] HATA: {e}\n{tb}")
            try:
                tk.Label(self.top, text=f"❌ Açılamadı: {e}",
                         fg='white', bg='#B71C1C', padx=20, pady=20,
                         font=("Segoe UI", 11, "bold")
                         ).pack(fill="both", expand=True)
            except Exception:
                pass

    def _arayuz(self):
        # Header
        renk = '#1565C0' if self.kaynak == 'RECETE' else '#E65100'
        header = tk.Frame(self.top, bg=renk, height=44)
        header.pack(fill="x")
        header.pack_propagate(False)
        ikon = '📋' if self.kaynak == 'RECETE' else '🛒'
        tk.Label(
            header,
            text=f"{ikon} {self.kaynak} Satış Detayı  —  RxId = {self.rx_id}",
            font=("Segoe UI", 11, "bold"), bg=renk, fg='white',
        ).pack(side="left", padx=15, pady=10)
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self.top.destroy).pack(side="right", padx=10, pady=6)
        tk.Button(header, text="🕐 Zaman Dökümü", bg='#5E35B1', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._zaman_dokumu_ac).pack(side="right", padx=5, pady=6)

        # Başlık alanı (hasta/doktor/kurum/teşhis)
        self.baslik_frame = tk.Frame(self.top, bg='#ECEFF1')
        self.baslik_frame.pack(fill="x", padx=4, pady=4)
        self.baslik_lbl = tk.Label(
            self.baslik_frame, text="⏳ Yükleniyor...",
            bg='#ECEFF1', fg='#37474F',
            font=("Segoe UI", 10), anchor='w', justify='left',
            padx=12, pady=10, wraplength=1150,
        )
        self.baslik_lbl.pack(fill="x")

        # Kalem tablosu
        kalem_dis = tk.Frame(self.top)
        kalem_dis.pack(fill="both", expand=True, padx=4, pady=4)

        cols = ('barkod', 'urun', 'adet', 'etiket', 'kurum',
                'fark', 'iskonto', 'toplam', 'iade')
        bash = {
            'barkod': 'Barkod', 'urun': 'Ürün',
            'adet': 'Adet', 'etiket': 'Etiket ₺', 'kurum': 'Kurum ₺',
            'fark': 'Fiyat Farkı ₺', 'iskonto': 'İskonto ₺',
            'toplam': 'Toplam ₺', 'iade': 'İade',
        }
        widths = {
            'barkod': 120, 'urun': 380,
            'adet': 60, 'etiket': 100, 'kurum': 100,
            'fark': 110, 'iskonto': 90,
            'toplam': 120, 'iade': 60,
        }
        anchors = {'urun': 'w', 'iade': 'center'}

        self.tree = ttk.Treeview(kalem_dis, columns=cols,
                                  show='headings', height=15)
        for c in cols:
            self.tree.heading(c, text=bash[c])
            self.tree.column(c, width=widths[c],
                              anchor=anchors.get(c, 'e'))
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(kalem_dis, orient="vertical",
                            command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.tag_configure('iade', background='#FFEBEE',
                                 foreground='#B71C1C')
        self.tree.tag_configure('toplam', background='#FFF8E1',
                                 font=('Segoe UI', 9, 'bold'))

    def _zaman_dokumu_ac(self):
        """Bu kaydın TÜM zaman alanlarını + sunucu/PC saatini gösteren mini
        popup açar. Saat sapması/kayma tespiti için kullanılır."""
        diag = tk.Toplevel(self.top)
        diag.title(f"🕐 Zaman Dökümü — {self.kaynak} RxId={self.rx_id}")
        diag.geometry("900x600")
        diag.transient(self.top)

        header = tk.Frame(diag, bg='#5E35B1', height=42)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header,
                 text=f"🕐 Zaman Dökümü — {self.kaynak} RxId={self.rx_id}",
                 font=("Segoe UI", 11, "bold"), bg='#5E35B1', fg='white',
                 ).pack(side="left", padx=15, pady=10)
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=diag.destroy).pack(side="right", padx=10, pady=6)

        info = tk.Label(diag,
            text="Tüm tarih/saat alanları + DB sunucusunun şu anki saati + "
                 "uygulamanın çalıştığı PC'nin şu anki saati. Saniyelik farklar "
                 "normal — saatler/günler kayma şüphesi yaratır.",
            font=("Segoe UI", 9, "italic"), fg='#37474F', bg='#FFF8E1',
            anchor='w', justify='left', padx=12, pady=8, wraplength=850,
        )
        info.pack(fill="x")

        durum = tk.Label(diag, text="⏳ DB'den çekiliyor...",
                         font=("Segoe UI", 10, "italic"), fg='#1565C0', pady=6)
        durum.pack(fill="x")

        txt_frame = tk.Frame(diag)
        txt_frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(txt_frame, font=("Consolas", 9), bg='#FAFAFA', wrap='none')
        vsb = ttk.Scrollbar(txt_frame, orient="vertical", command=txt.yview)
        hsb = ttk.Scrollbar(txt_frame, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        txt.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        txt_frame.grid_columnconfigure(0, weight=1)
        txt_frame.grid_rowconfigure(0, weight=1)

        def _calis():
            try:
                d = self.db.tek_kayit_zaman_dokumu(self.kaynak, self.rx_id)
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Zaman dökümü hatası: {e}\n{tb}")
                diag.after(0, lambda: durum.config(
                    text=f"❌ Hata: {e}", fg='#B71C1C'))
                return
            diag.after(0, lambda: _yaz(d))

        def _yaz(d):
            durum.config(text="✓ Tamam.", fg='#2E7D32')
            txt.delete('1.0', tk.END)

            if d.get('hata'):
                txt.insert(tk.END, f"❌ {d['hata']}\n")
                return

            def _fmt(v):
                if v is None:
                    return '(NULL)'
                if hasattr(v, 'strftime'):
                    if hasattr(v, 'hour'):
                        try:
                            return v.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                        except Exception:
                            return v.strftime('%Y-%m-%d %H:%M:%S')
                    return v.strftime('%Y-%m-%d') + ' (saat yok)'
                return str(v)

            def _fark(a, b):
                """a ve b datetime → fark saniye cinsinden."""
                try:
                    if a is None or b is None:
                        return None
                    if not hasattr(a, 'hour'):
                        from datetime import datetime as _dt
                        a = _dt(a.year, a.month, a.day)
                    if not hasattr(b, 'hour'):
                        from datetime import datetime as _dt
                        b = _dt(b.year, b.month, b.day)
                    return int((a - b).total_seconds())
                except Exception:
                    return None

            # Referans: İşlem Tarihi
            ref = d['degerler'].get('RxIslemTarihi')

            txt.insert(tk.END,
                f"═══════════════════════════════════════════════════════\n")
            txt.insert(tk.END,
                f"KAYIT BİLGİSİ: {d['tablo_adi']}.RxId = {d['rx_id']}\n")
            txt.insert(tk.END,
                f"═══════════════════════════════════════════════════════\n")
            txt.insert(tk.END,
                f"Referans (kıyas için): RxIslemTarihi = {_fmt(ref)}\n\n")

            # 1. Kayıttaki tüm zaman alanları
            txt.insert(tk.END,
                f"───────────────────────────────────────────────────────\n")
            txt.insert(tk.END,
                f"1. KAYITTAKİ ZAMAN ALANLARI (DB'den okunan)\n")
            txt.insert(tk.END,
                f"───────────────────────────────────────────────────────\n")
            txt.insert(tk.END,
                f"   {'Kolon Adı':<25} {'Tip':<14} {'Değer':<26} {'Fark (İşlem ref)'}\n")
            for kol in d.get('kolonlar', []):
                ad = kol.get('COLUMN_NAME', '')
                tip = kol.get('DATA_TYPE', '')
                deg = d['degerler'].get(ad)
                fark = _fark(deg, ref) if ad != 'RxIslemTarihi' else 0
                fark_str = ''
                if fark is None:
                    fark_str = '—'
                elif fark == 0:
                    fark_str = '0sn (referans)' if ad == 'RxIslemTarihi' else '0sn'
                else:
                    fark_str = (f"{fark:+,} sn  ({_saniye_okunabilir(fark)})"
                                if abs(fark) >= 60 else f"{fark:+,} sn")
                txt.insert(tk.END,
                    f"   {ad:<25} {tip:<14} {_fmt(deg):<26} {fark_str}\n")

            # 2. Şu anki saatler (kıyas için)
            txt.insert(tk.END,
                f"\n───────────────────────────────────────────────────────\n")
            txt.insert(tk.END,
                f"2. ŞU ANKİ SAATLER (kayıt zamanından bağımsız — "
                f"şu anki kayma tespiti)\n")
            txt.insert(tk.END,
                f"───────────────────────────────────────────────────────\n")
            sn_now = d.get('sunucu_now')
            sn_get = d.get('sunucu_now_getdate')
            sn_utc = d.get('sunucu_now_utc')
            pc_now = d.get('pc_now')
            pc_utc = d.get('pc_now_utc')
            txt.insert(tk.END,
                f"   {'Kaynak':<32} {'Değer':<26} {'Fark (PC vs)'}\n")
            txt.insert(tk.END,
                f"   {'SQL Server SYSDATETIME() (yerel)':<32} {_fmt(sn_now):<26} "
                f"{(_fark(sn_now, pc_now)):+,d} sn\n"
                if sn_now and pc_now else
                f"   {'SQL Server SYSDATETIME() (yerel)':<32} {_fmt(sn_now)}\n")
            txt.insert(tk.END,
                f"   {'SQL Server GETDATE()':<32} {_fmt(sn_get)}\n")
            txt.insert(tk.END,
                f"   {'SQL Server GETUTCDATE() (UTC)':<32} {_fmt(sn_utc)}\n")
            txt.insert(tk.END,
                f"   {'Python (uygulama PC) datetime.now()':<32} {_fmt(pc_now)}\n")
            txt.insert(tk.END,
                f"   {'Python datetime.utcnow()':<32} {_fmt(pc_utc)}\n")

            # 3. Yorum
            txt.insert(tk.END,
                f"\n───────────────────────────────────────────────────────\n")
            txt.insert(tk.END,
                f"3. YORUM\n")
            txt.insert(tk.END,
                f"───────────────────────────────────────────────────────\n")
            sn_pc_fark = _fark(sn_now, pc_now) if sn_now else None
            if sn_pc_fark is None:
                txt.insert(tk.END, "• SQL Server saati okunamadı.\n")
            elif abs(sn_pc_fark) < 5:
                txt.insert(tk.END,
                    f"• SQL Server ↔ PC saat farkı: {sn_pc_fark:+d} sn "
                    f"(normal, <5sn).\n")
            elif abs(sn_pc_fark) < 60:
                txt.insert(tk.END,
                    f"• ⚠ SQL Server ↔ PC saat farkı: {sn_pc_fark:+d} sn — "
                    f"NTP senkron sapması olabilir.\n")
            else:
                txt.insert(tk.END,
                    f"• ❌ SQL Server ↔ PC saat farkı: {sn_pc_fark:+d} sn — "
                    f"BÜYÜK FARK! PC saati yanlış kalibre olmuş olabilir.\n")

            # Kayıttaki alanlar arasındaki tutarsızlık
            alanlar = [(k, d['degerler'].get(k))
                        for k in (kol.get('COLUMN_NAME') for kol in d.get('kolonlar', []))]
            tarih_alanlar = [(a, v) for a, v in alanlar if v is not None]
            if len(tarih_alanlar) >= 2:
                farklar = []
                for a, v in tarih_alanlar:
                    if a == 'RxIslemTarihi':
                        continue
                    f = _fark(v, ref)
                    if f is not None:
                        farklar.append((a, f))
                buyuk = [a for a, f in farklar if abs(f) > 24 * 3600]
                if buyuk:
                    txt.insert(tk.END,
                        f"• ⚠ Kayıt içi alanlarda büyük tutarsızlık "
                        f"(>24sa) var: {', '.join(buyuk)}.\n"
                        f"  Bu kayıt için saat kayması/manuel düzenleme şüphesi.\n")
                else:
                    txt.insert(tk.END,
                        "• Kayıt içi alanlar tutarlı (24 saat içinde "
                        "birbirlerinden).\n")

        import threading
        threading.Thread(target=_calis, daemon=True).start()

    def _yukle(self):
        def _calis():
            try:
                detay = self.db.satis_kalem_detay_getir(self.kaynak, self.rx_id)
                self.top.after(0, lambda: self._goster(detay))
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Satış detayı yüklenemedi: {e}\n{tb}")
                print(f"[Satış Detayı] HATA: {e}\n{tb}")
                self.top.after(0, lambda: self.baslik_lbl.config(
                    text=f"❌ Yüklenemedi: {e}"))
        threading.Thread(target=_calis, daemon=True).start()

    def _goster(self, detay: Dict):
        baslik = detay.get('baslik') if detay else None
        kalemler = detay.get('kalemler') if detay else []

        if not baslik:
            baslik_hata = detay.get('baslik_hata') if detay else None
            kalem_hata = detay.get('kalem_hata') if detay else None
            mesaj = (f"⚠ {self.kaynak} RxId={self.rx_id} bulunamadı "
                     f"veya silinmiş.")
            if baslik_hata:
                mesaj += f"\n❌ Başlık SQL hatası: {baslik_hata}"
            if kalem_hata:
                mesaj += f"\n❌ Kalem SQL hatası: {kalem_hata}"
            self.baslik_lbl.config(text=mesaj)
            return

        def _g(d, k, varsayilan='—'):
            v = d.get(k)
            if v is None or v == '':
                return varsayilan
            return v

        islem = _g(baslik, 'RxIslemTarihi')
        kayit = _g(baslik, 'RxKayitTarihi')
        islem_str = (islem.strftime('%d.%m.%Y %H:%M:%S')
                     if hasattr(islem, 'strftime') else str(islem))
        kayit_str = (kayit.strftime('%d.%m.%Y %H:%M:%S')
                     if hasattr(kayit, 'strftime') else str(kayit))

        hasta_adi = _g(baslik, 'HastaAdi')
        hasta_tc = _g(baslik, 'HastaTCKN')

        toplam_tl = sum(float(k.get('toplam', 0) or 0)
                         for k in (kalemler or []) if not k.get('iade'))
        iade_tl = sum(float(k.get('toplam', 0) or 0)
                       for k in (kalemler or []) if k.get('iade'))

        def _tl(v):
            return (f"{float(v):,.2f} ₺"
                    .replace(',', 'X').replace('.', ',').replace('X', '.'))

        if self.kaynak == 'RECETE':
            recete_no = _g(baslik, 'RxEReceteNo')
            doktor = _g(baslik, 'DoktorAdi')
            tesis = _g(baslik, 'TesisAdi')
            kurum = _g(baslik, 'KurumAdi')
            satirlar = [
                f"📋 REÇETE  •  RxId={baslik.get('RxId')}  •  "
                f"E-Reçete: {recete_no}",
                f"İşlem: {islem_str}    |    Kayıt: {kayit_str}",
                f"Hasta: {hasta_adi}  ({hasta_tc})",
                f"Doktor: {doktor}",
                f"Tesis: {tesis}    |    Kurum: {kurum}",
                f"Toplam: {_tl(toplam_tl)}"
                + (f"    İade: {_tl(iade_tl)}" if iade_tl else ""),
            ]
        else:  # ELDEN
            belge_no = _g(baslik, 'RxBelgeNo')
            satirlar = [
                f"🛒 ELDEN (Parakende)  •  RxId={baslik.get('RxId')}  •  "
                f"Belge: {belge_no}",
                f"İşlem: {islem_str}    |    Kayıt: {kayit_str}",
                f"Müşteri: {hasta_adi}"
                + (f"  ({hasta_tc})" if hasta_tc and hasta_tc != '—' else ''),
                f"Toplam: {_tl(toplam_tl)}"
                + (f"    İade: {_tl(iade_tl)}" if iade_tl else ""),
            ]

        self.baslik_lbl.config(text="\n".join(satirlar))

        # Kalemleri tabloya yaz
        for c in self.tree.get_children():
            self.tree.delete(c)

        def _f(v):
            try:
                return (f"{float(v):,.2f}"
                        .replace(',', 'X').replace('.', ',').replace('X', '.'))
            except Exception:
                return str(v) if v is not None else ''

        toplam_adet = 0.0
        toplam_top = 0.0
        for k in kalemler or []:
            iade_str = '⟲ İADE' if k.get('iade') else ''
            tag = ('iade',) if k.get('iade') else ()
            self.tree.insert('', 'end', values=(
                k.get('barkod', ''),
                k.get('urun_adi', ''),
                f"{k.get('adet', 0):.0f}",
                _f(k.get('etiket_fiyat', 0)),
                _f(k.get('kurum_fiyat')) if k.get('kurum_fiyat') is not None else '—',
                _f(k.get('fiyat_farki')) if k.get('fiyat_farki') is not None else '—',
                _f(k.get('iskonto', 0)),
                _f(k.get('toplam', 0)),
                iade_str,
            ), tags=tag)
            if not k.get('iade'):
                toplam_adet += float(k.get('adet', 0) or 0)
                toplam_top += float(k.get('toplam', 0) or 0)

        if kalemler:
            self.tree.insert('', 'end', values=(
                '', f'TOPLAM ({len(kalemler)} kalem)',
                f"{toplam_adet:.0f}", '', '', '', '',
                _f(toplam_top), '',
            ), tags=('toplam',))
        else:
            self.tree.insert('', 'end', values=(
                '', '(Bu satışta görünür kalem yok)', '', '', '', '', '', '', ''
            ))


class TopluZamanAnomaliPopup:
    """Geniş aralıkta RxIslemTarihi vs RxKayitTarihi farkı eşik üstündeki
    satışları tarayan adli analiz penceresi.

    Kullanım: Satış Raporları başlık çubuğundaki "⏱ Toplu Zaman Anomalisi"
    butonu. SQL düzeyinde filtre → milyon kayıtlı tabloyu Python'a çekmez.
    """

    ESIK_PRESETLERI = [
        ('0',      '0 sn — fark olan TÜM kayıtlar (debug)'),
        ('60',     '1 dk'),
        ('600',    '10 dk (varsayılan)'),
        ('1800',   '30 dk'),
        ('3600',   '1 saat'),
        ('10800',  '3 saat'),
        ('86400',  '24 saat (sadece çok büyük)'),
    ]

    def __init__(self, parent, db):
        self.parent = parent
        self.db = db
        self.sonuclar: List[Dict] = []
        # Sıralama durumu
        self._son_sirala_sutun: Optional[str] = None
        self._son_sirala_ters: bool = False

        self.top = tk.Toplevel(parent)
        self.top.title("⏱ Toplu Zaman Anomalisi (RxIslemTarihi vs RxKayitTarihi)")
        self.top.geometry("1400x800")
        try:
            self.top.state('zoomed')
        except Exception:
            pass
        self.top.transient(parent)

        try:
            self._arayuz()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"TopluZamanAnomaliPopup._arayuz HATA: {e}\n{tb}")
            print(f"[Toplu Anomali] _arayuz HATA: {e}\n{tb}")
            try:
                tk.Label(
                    self.top,
                    text=f"❌ Arayüz oluşturulamadı:\n\n{e}",
                    fg='white', bg='#B71C1C',
                    font=("Segoe UI", 11, "bold"),
                    padx=20, pady=20, justify='left', wraplength=900,
                ).pack(fill="both", expand=True, padx=20, pady=20)
            except Exception:
                pass

    def _arayuz(self):
        # Header
        header = tk.Frame(self.top, bg='#5E35B1', height=44)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(
            header,
            text="⏱ Toplu Zaman Anomalisi Tarayıcısı  "
                 "(geniş aralıkta DATEDIFF(işlem, kayıt) > eşik veya negatif)",
            font=("Segoe UI", 11, "bold"), bg='#5E35B1', fg='white',
        ).pack(side="left", padx=15, pady=10)
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self.top.destroy).pack(side="right", padx=10, pady=6)
        tk.Button(header, text="📥 Excel", bg='#2E7D32', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._excel_aktar).pack(side="right", padx=5, pady=6)
        tk.Button(header, text="🔎 Detay Gör", bg='#FF8F00', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._secili_detayi_ac).pack(side="right", padx=5, pady=6)
        tk.Button(header, text="🩺 Tanılama", bg='#00838F', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._tanilama_ac).pack(side="right", padx=5, pady=6)

        # Filtre/parametre paneli
        filt = ttk.LabelFrame(self.top, text="Tarama Parametreleri", padding=8)
        filt.pack(fill="x", padx=5, pady=4)

        # Satır 1: tarih aralığı
        s1 = tk.Frame(filt)
        s1.pack(fill="x", pady=3)
        tk.Label(s1, text="Başlangıç (YYYY-MM-DD):",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        self.bas_entry = ttk.Entry(s1, width=14)
        # Varsayılan: eczane açılışı (2017-05-23)
        self.bas_entry.insert(0, "2017-05-23")
        self.bas_entry.pack(side="left", padx=4)

        tk.Label(s1, text="→  Bitiş:",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        self.bit_entry = ttk.Entry(s1, width=14)
        self.bit_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.bit_entry.pack(side="left", padx=4)

        # Hızlı preset
        tk.Label(s1, text="    Hızlı:", font=("Segoe UI", 9, "italic")
                 ).pack(side="left", padx=(15, 4))
        for et, gun in [("Son 90 gün", 90), ("Son 1 yıl", 365),
                         ("Son 3 yıl", 365*3), ("Tümü (2017→)", -1)]:
            ttk.Button(s1, text=et, width=14,
                       command=lambda g=gun: self._tarih_preset(g)
                       ).pack(side="left", padx=2)

        # Satır 2: eşik + tara butonu
        s2 = tk.Frame(filt)
        s2.pack(fill="x", pady=3)
        tk.Label(s2, text="Eşik (saniye):",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        self.esik_var = tk.StringVar(value='600')
        self.esik_combo = ttk.Combobox(
            s2, textvariable=self.esik_var, width=30, state="readonly",
            values=[f"{kod} — {et}" for kod, et in self.ESIK_PRESETLERI],
        )
        self.esik_combo.set(f"{self.ESIK_PRESETLERI[0][0]} — {self.ESIK_PRESETLERI[0][1]}")
        self.esik_combo.pack(side="left", padx=4)

        tk.Label(s2, text="  (|fark| > eşik VEYA fark < 0 olan kayıtlar listelenir)",
                 font=("Segoe UI", 8, "italic"), fg='#37474F'
                 ).pack(side="left", padx=4)

        self.tara_btn = ttk.Button(s2, text="▶ Taramayı Başlat", width=20,
                                    command=self._taramayi_baslat)
        self.tara_btn.pack(side="right", padx=8)

        # Açıklama satırı: hassasiyet mantığı
        s3 = tk.Frame(filt)
        s3.pack(fill="x", pady=3)
        tk.Label(
            s3,
            text="ℹ Kayıt saatli ise → saniye bazlı fark.  "
                 "Kayıt saatsiz (saat=00:00:00) ise → gün bazlı fark "
                 "(işlem saati yok sayılır, aynı gün ise anomali değil).",
            font=("Segoe UI", 8, "italic"), fg='#37474F',
        ).pack(side="left", padx=4)

        # Satır 4: filtre checkbox'ları (tarama sonrası, client-side)
        s4 = tk.Frame(filt)
        s4.pack(fill="x", pady=(2, 1))
        self.donusum_haric_var = tk.BooleanVar(value=False)
        self.antibiyotik_haric_var = tk.BooleanVar(value=False)
        self.is_bankasi_haric_var = tk.BooleanVar(value=False)
        self.bez_haric_var = tk.BooleanVar(value=False)
        self.cezaevi_haric_var = tk.BooleanVar(value=False)
        tk.Label(s4, text="Filtreler:",
                 font=("Segoe UI", 9, "bold"), fg='#37474F'
                 ).pack(side="left", padx=(4, 8))
        ttk.Checkbutton(
            s4, text="🔄 Perakende→Reçete dönüşümlerini gizle",
            variable=self.donusum_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s4, text="🦠 Antibiyotik gizle (ATC J)",
            variable=self.antibiyotik_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s4, text="🏦 İş Bankası gizle",
            variable=self.is_bankasi_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s4, text="🩹 Bez gizle",
            variable=self.bez_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s4, text="🔒 Cezaevi gizle",
            variable=self.cezaevi_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)

        # İlerleme + durum
        durum = tk.Frame(self.top, bg='#F5F5F5', height=32)
        durum.pack(fill="x")
        durum.pack_propagate(False)
        self.durum_lbl = tk.Label(
            durum, text="Hazır. Tarih ve eşik seçip 'Taramayı Başlat'a basın.",
            bg='#F5F5F5', fg='#1565C0', font=("Segoe UI", 9, "italic"),
            anchor='w', padx=10,
        )
        self.durum_lbl.pack(side="left", fill="x", expand=True)
        self.progress = ttk.Progressbar(durum, mode='indeterminate', length=200)
        self.progress.pack(side="right", padx=10, pady=5)

        # Üst özet
        ozet = ttk.LabelFrame(self.top, text="ÖZET", padding=6)
        ozet.pack(fill="x", padx=5, pady=3)
        self.ozet_text = tk.Text(ozet, height=7, font=("Consolas", 9),
                                  bg='#FAFAFA', relief='flat', wrap='word')
        self.ozet_text.pack(fill="x")
        self.ozet_text.insert('1.0', "(tarama henüz yapılmadı)")

        # Sonuç tablosu
        tab = ttk.LabelFrame(self.top, text="ANOMALİ SATIŞLAR (|fark| azalan)",
                              padding=4)
        tab.pack(fill="both", expand=True, padx=5, pady=3)

        cols = ('no', 'kaynak', 'rxid', 'islem', 'kayit', 'kayit_fark',
                'kontrol', 'kontrol_fark',
                'donusum', 'antibiyotik', 'isbankasi', 'bez', 'cezaevi')
        bash = {'no': '#', 'kaynak': 'Kaynak', 'rxid': 'RxId',
                'islem': 'İşlem Tarihi', 'kayit': 'Kayıt Tarihi',
                'kayit_fark': 'Kayıt Fark (gün)',
                'kontrol': 'Kontrol Tarihi',
                'kontrol_fark': 'Kontrol Fark (sn)',
                'donusum': '🔄', 'antibiyotik': '🦠',
                'isbankasi': '🏦', 'bez': '🩹', 'cezaevi': '🔒'}
        widths = {'no': 50, 'kaynak': 70, 'rxid': 80,
                  'islem': 180, 'kayit': 140,
                  'kayit_fark': 140,
                  'kontrol': 180, 'kontrol_fark': 180,
                  'donusum': 45, 'antibiyotik': 45,
                  'isbankasi': 45, 'bez': 45, 'cezaevi': 45}
        anchors = {'kaynak': 'center', 'rxid': 'center', 'no': 'center',
                   'donusum': 'center', 'antibiyotik': 'center',
                   'isbankasi': 'center', 'bez': 'center',
                   'cezaevi': 'center'}

        self.tree = ttk.Treeview(tab, columns=cols, show='headings', height=18)
        # Başlık başlıklarını sakla (sıralama oku eklemek için)
        self._anomali_basliklar_orj = dict(bash)
        for c in cols:
            self.tree.heading(
                c, text=bash[c],
                command=lambda col=c: self._sirala(col),
            )
            self.tree.column(c, width=widths[c], anchor=anchors.get(c, 'w'))
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(tab, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.tag_configure('negatif', background='#FFCDD2')
        self.tree.tag_configure('cok_buyuk', background='#FFAB91')
        self.tree.tag_configure('buyuk', background='#FFE082')
        self.tree.tag_configure('gun_anomali', background='#BBDEFB')
        self.tree.tag_configure('normal', background='white')
        self.tree.tag_configure('saatsiz', background='#ECEFF1',
                                 foreground='#546E7A')
        # Çift-tıkla → Detay Gör
        self.tree.bind('<Double-1>', self._secili_detayi_ac)

    def _secili_detayi_ac(self, _event=None):
        """Seçili anomali satırının RECETE/ELDEN detayını ayrı pencerede aç."""
        sec = self.tree.selection()
        if not sec:
            messagebox.showinfo("Seçim yok",
                "Önce ANOMALİ SATIŞLAR tablosundan bir satır seçin "
                "(satıra tek tıkla mavi yap, sonra butona bas — "
                "veya satıra çift-tıkla).")
            return
        vals = self.tree.item(sec[0], 'values')
        # Kolon sırası: no, kaynak, rxid, islem, kayit, fark_sn, fark
        if not vals or len(vals) < 3:
            return
        kaynak = (vals[1] or '').strip().upper()
        rx_id = (vals[2] or '').strip()
        if kaynak not in ('RECETE', 'ELDEN') or not rx_id:
            messagebox.showinfo("Geçersiz satır",
                "Bu satır bir RECETE/ELDEN kaydına karşılık gelmiyor.")
            return
        SatisDetayPopup(parent=self.top, db=self.db,
                        kaynak=kaynak, rx_id=rx_id)

    def _tanilama_ac(self):
        """RxKayitTarihi kolonunun gerçek yapısını analiz eden tanılama
        penceresi. Hangi kayıtlar saatli, hangileri saatsiz; BotanikEOS UI'da
        görünen saat hangi kolondan geliyor onu netleştirir."""
        diag_top = tk.Toplevel(self.top)
        diag_top.title("🩺 RxKayitTarihi Tanılama")
        diag_top.geometry("1100x700")
        diag_top.transient(self.top)

        header = tk.Frame(diag_top, bg='#00838F', height=40)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header,
                 text="🩺 RxKayitTarihi Yapı Tanılaması",
                 font=("Segoe UI", 11, "bold"), bg='#00838F', fg='white',
                 ).pack(side="left", padx=15, pady=10)
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=diag_top.destroy).pack(side="right", padx=10, pady=6)

        durum = tk.Label(diag_top, text="⏳ Diagnostik sorgular çalışıyor...",
                         font=("Segoe UI", 10, "italic"), fg='#1565C0',
                         pady=8)
        durum.pack(fill="x")

        txt_frame = tk.Frame(diag_top)
        txt_frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(txt_frame, font=("Consolas", 9), bg='#FAFAFA', wrap='none')
        vsb = ttk.Scrollbar(txt_frame, orient="vertical", command=txt.yview)
        hsb = ttk.Scrollbar(txt_frame, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        txt.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        txt_frame.grid_columnconfigure(0, weight=1)
        txt_frame.grid_rowconfigure(0, weight=1)

        def _calis():
            try:
                d = self.db.kayit_tarihi_diagnostik()
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Tanılama hatası: {e}\n{tb}")
                diag_top.after(0, lambda: durum.config(
                    text=f"❌ Hata: {e}", fg='#B71C1C'))
                return
            diag_top.after(0, lambda: _yaz(d))

        def _yaz(d):
            durum.config(text="✓ Diagnostik tamam.", fg='#2E7D32')
            txt.delete('1.0', tk.END)

            # 1. Kolon listesi
            txt.insert(tk.END, "═══════════════════════════════════════════\n")
            txt.insert(tk.END, "1. ReceteAna tablosunun ZAMAN kolonları\n")
            txt.insert(tk.END, "═══════════════════════════════════════════\n")
            for k in d.get('recete_kolonlar', []):
                txt.insert(tk.END,
                    f"  • {k.get('COLUMN_NAME'):<35} {k.get('DATA_TYPE')}\n")

            txt.insert(tk.END, "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END, "2. EldenAna tablosunun ZAMAN kolonları\n")
            txt.insert(tk.END, "═══════════════════════════════════════════\n")
            for k in d.get('elden_kolonlar', []):
                txt.insert(tk.END,
                    f"  • {k.get('COLUMN_NAME'):<35} {k.get('DATA_TYPE')}\n")

            # 2. Sayım
            txt.insert(tk.END, "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END, "3. RxKayitTarihi'nde saat var/yok DAĞILIMI\n")
            txt.insert(tk.END, "═══════════════════════════════════════════\n")
            r_saatli = d.get('recete_kayit_saatli', '?')
            r_saatsiz = d.get('recete_kayit_saatsiz', '?')
            r_top = d.get('recete_toplam', '?')
            txt.insert(tk.END,
                f"  ReceteAna toplam: {r_top}\n"
                f"    Saatli (HH:MM:SS > 00:00:00):  {r_saatli}\n"
                f"    Saatsiz (HH:MM:SS = 00:00:00): {r_saatsiz}\n")
            e_saatli = d.get('elden_kayit_saatli', '?')
            e_saatsiz = d.get('elden_kayit_saatsiz', '?')
            e_top = d.get('elden_toplam', '?')
            txt.insert(tk.END,
                f"\n  EldenAna toplam: {e_top}\n"
                f"    Saatli:  {e_saatli}\n"
                f"    Saatsiz: {e_saatsiz}\n")

            # 3. Örnekler
            for kaynak, baslik in [('recete', 'ReceteAna'),
                                    ('elden', 'EldenAna')]:
                for tip in ['saatli', 'saatsiz']:
                    txt.insert(tk.END,
                        f"\n═══════════════════════════════════════════\n")
                    txt.insert(tk.END,
                        f"4. {baslik} — {tip.upper()} örnek (TOP 5)\n")
                    txt.insert(tk.END,
                        f"═══════════════════════════════════════════\n")
                    rows = d.get(f'{kaynak}_ornek_{tip}', []) or []
                    if not rows:
                        txt.insert(tk.END, f"  (örnek yok)\n")
                        continue
                    txt.insert(tk.END,
                        f"  {'RxId':>8}  {'İşlem (DB raw)':<22}  "
                        f"{'Kayıt (DB raw)':<22}  H  M  S\n")
                    for r in rows:
                        txt.insert(tk.END,
                            f"  {r.get('RxId'):>8}  "
                            f"{(r.get('IslemStr') or ''):<22}  "
                            f"{(r.get('KayitStr') or ''):<22}  "
                            f"{r.get('KayitSaat'):>2} {r.get('KayitDk'):>2} "
                            f"{r.get('KayitSn'):>2}\n")

            # 5. Bağımsız sunucu-tarafı timestamp kolonları (TÜM DB)
            txt.insert(tk.END,
                "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END,
                "5. BAĞIMSIZ SUNUCU TIMESTAMP KOLONLARI\n"
                "   (DEFAULT GETDATE() vb. ile otomatik yazan kolonlar —\n"
                "   PC saat sapması tespitinde kullanılabilir)\n")
            txt.insert(tk.END,
                "═══════════════════════════════════════════\n")
            bts = d.get('bagimsiz_timestamp_kolonlari', []) or []
            if not bts:
                txt.insert(tk.END,
                    "  ❌ HİÇ YOK. Botanik DB'de hiçbir kolonda DEFAULT\n"
                    "  GETDATE/SYSDATETIME yok — yani saat sapmasını geriye\n"
                    "  dönük tespit edebileceğimiz bir DB-tarafı kaynak YOK.\n")
            else:
                txt.insert(tk.END,
                    f"  {'Tablo':<30} {'Kolon':<25} {'Tip':<14} Default\n")
                for k in bts:
                    txt.insert(tk.END,
                        f"  {k.get('TABLE_NAME'):<30} "
                        f"{k.get('COLUMN_NAME'):<25} "
                        f"{k.get('DATA_TYPE'):<14} "
                        f"{k.get('COLUMN_DEFAULT')}\n")

            # 6. Audit / log tablo adayları
            txt.insert(tk.END,
                "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END,
                "6. AUDIT / LOG / HAREKET TABLO ADAYLARI\n")
            txt.insert(tk.END,
                "═══════════════════════════════════════════\n")
            audit = d.get('audit_tablo_adaylari', []) or []
            if not audit:
                txt.insert(tk.END, "  (uygun adlı tablo yok)\n")
            else:
                for t in audit:
                    txt.insert(tk.END, f"  • {t.get('TABLE_NAME')}\n")
                txt.insert(tk.END,
                    "\n  ⚠ Bu tablolar İŞLEM kayıtları tutuyor olabilir;\n"
                    "  zaman damgaları varsa BotanikEOS işlem tarihiyle\n"
                    "  karşılaştırılabilir. Listede ilgi çekici bir tablo\n"
                    "  görürsen söyle, içeriği analiz edelim.\n")

            # 7. Şu anki SQL Server vs PC saat farkı
            txt.insert(tk.END,
                "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END,
                "7. ŞU ANKİ SAAT KARŞILAŞTIRMASI (canlı)\n")
            txt.insert(tk.END,
                "═══════════════════════════════════════════\n")
            sz = d.get('sunucu_zaman') or {}
            sn = sz.get('SunucuSysdatetime')
            pcn = d.get('pc_zaman')
            sn_fmt = (sn.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                       if hasattr(sn, 'strftime') else str(sn))
            pcn_fmt = (pcn.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                        if hasattr(pcn, 'strftime') else str(pcn))
            txt.insert(tk.END, f"  SQL Server (SYSDATETIME): {sn_fmt}\n")
            txt.insert(tk.END, f"  PC (Python datetime.now): {pcn_fmt}\n")
            if hasattr(sn, 'year') and hasattr(pcn, 'year'):
                try:
                    fark = int((sn - pcn).total_seconds())
                    if abs(fark) < 5:
                        ikon = '✓'
                    elif abs(fark) < 60:
                        ikon = '⚠'
                    else:
                        ikon = '❌'
                    txt.insert(tk.END,
                        f"  {ikon} Fark: {fark:+d} sn "
                        f"({_saniye_okunabilir(fark)})\n")
                except Exception:
                    pass

            # 8. LOG TABLOLARI ANALİZİ — senkron, hızlı sys.partitions ile
            txt.insert(tk.END,
                "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END,
                "8. LOG TABLOLARI — yapı + örnekler\n"
                "   (BotanikLog/Logger/ReceteHesapLog/MedulaDokumLog)\n"
                "   Trigger varsa → sunucu-tarafı insert zamanı yazılmış olabilir\n")
            txt.insert(tk.END,
                "═══════════════════════════════════════════\n")
            try:
                log_data = self.db.log_tablo_analizi()
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"log_tablo_analizi hatası: {e}\n{tb}")
                txt.insert(tk.END, f"  ❌ Hata: {e}\n")
                log_data = {}

            for tablo, info in (log_data or {}).items():
                say = info.get('satir_sayisi')
                say_str = f"{say:,}" if say is not None else '?'
                txt.insert(tk.END,
                    f"\n  ┌─ {tablo}  ({say_str} satır)\n")
                if info.get('hata'):
                    txt.insert(tk.END, f"  │  ⚠ {info['hata']}\n")
                    continue
                # Trigger durumu
                if info.get('trigger_var_mi'):
                    triglar = info.get('triggerlar', [])
                    ad_listesi = [t.get('TriggerAdi') for t in triglar]
                    txt.insert(tk.END,
                        f"  │  🔧 TRIGGER VAR ({len(triglar)}): "
                        f"{', '.join(ad_listesi)}\n"
                        f"  │     → insert zamanı sunucu tarafından yazılıyor olabilir!\n")
                else:
                    txt.insert(tk.END,
                        f"  │  (trigger yok)\n")
                # Datetime kolonlar
                dt_kolonlar = [k for k in info.get('kolonlar', [])
                                if str(k.get('DATA_TYPE') or '').lower()
                                in ('datetime', 'datetime2', 'datetimeoffset',
                                    'smalldatetime', 'date', 'time')]
                if dt_kolonlar:
                    txt.insert(tk.END, f"  │  📅 Zaman kolonları:\n")
                    for k in dt_kolonlar:
                        varsay = k.get('COLUMN_DEFAULT')
                        isaret = ' ⭐ DEFAULT' if varsay else ''
                        txt.insert(tk.END,
                            f"  │    • {k.get('COLUMN_NAME'):<28} "
                            f"{k.get('DATA_TYPE'):<14}"
                            f"{isaret} "
                            f"{varsay or ''}\n")
                else:
                    txt.insert(tk.END,
                        f"  │  (DATETIME kolon yok — zaman kıyası için kullanılamaz)\n")
                # Örnek satırlar
                ornekler = info.get('ornekler') or []
                if ornekler:
                    txt.insert(tk.END,
                        f"  │  📋 Son 5 örnek satır (ilk 6 kolon):\n")
                    kolon_adlari = [k.get('COLUMN_NAME')
                                     for k in info.get('kolonlar', [])][:6]
                    txt.insert(tk.END,
                        f"  │    " + " | ".join(
                            f"{(a or '')[:18]:<18}" for a in kolon_adlari) + "\n")
                    for r in ornekler:
                        satir = []
                        for ka in kolon_adlari:
                            v = r.get(ka)
                            if v is None:
                                s = 'NULL'
                            elif hasattr(v, 'strftime') and hasattr(v, 'hour'):
                                s = v.strftime('%Y-%m-%d %H:%M:%S')
                            elif hasattr(v, 'strftime'):
                                s = v.strftime('%Y-%m-%d')
                            else:
                                s = str(v)[:18]
                            satir.append(f"{s[:18]:<18}")
                        txt.insert(tk.END,
                            f"  │    " + " | ".join(satir) + "\n")
                txt.insert(tk.END, f"  └─\n")

            txt.insert(tk.END,
                "\n  💡 Yukarıda 🔧 TRIGGER VAR ya da ⭐ DEFAULT işaretli bir\n"
                "  DATETIME kolon görüyorsan, o BAĞIMSIZ bir zaman kaynağı\n"
                "  OLABİLİR. Söyle, oradan PC saat sapması tespitini deneyelim.\n")

            # 9. LOGGER tablosunun yapısı + cross-check için keşif
            txt.insert(tk.END,
                "\n═══════════════════════════════════════════\n")
            txt.insert(tk.END,
                "9. LOGGER TABLOSU CROSS-CHECK KEŞFİ\n"
                "   (ReceteAna.RxIslemTarihi vs Logger.LoggerTarihi\n"
                "    karşılaştırması için gerekli bilgiler)\n")
            txt.insert(tk.END,
                "═══════════════════════════════════════════\n")
            try:
                lg = self.db.logger_yapisini_incele()
            except Exception as e:
                txt.insert(tk.END, f"  ❌ Hata: {e}\n")
                lg = {}

            # 9a. Tüm kolonlar
            txt.insert(tk.END, "\n  📋 Logger TÜM kolonları:\n")
            for k in lg.get('kolonlar', []):
                txt.insert(tk.END,
                    f"    • {k.get('COLUMN_NAME'):<28} {k.get('DATA_TYPE')}\n")

            # 9b. IslemTuru + AltTuru dağılımı
            txt.insert(tk.END, "\n  📊 LoggerIslemTuru × AltTuru dağılımı (TOP 20):\n")
            dag = lg.get('islem_turu_dagilimi') or []
            if not dag:
                txt.insert(tk.END,
                    f"    (sorgu başarısız: {lg.get('dagilim_hata', '?')})\n")
            else:
                txt.insert(tk.END,
                    f"    {'IslemTuru':<12} {'AltTuru':<10} Adet\n")
                for r in dag:
                    txt.insert(tk.END,
                        f"    {str(r.get('LoggerIslemTuru') or 'NULL'):<12} "
                        f"{str(r.get('LoggerIslemAltTuru') or 'NULL'):<10} "
                        f"{r.get('Sayi', 0):,}\n")

            # 9c. Son reçete eşleşmesi
            r_rx = lg.get('ornek_recete_rxid')
            r_islem = lg.get('ornek_recete_islem')
            r_eslesme = lg.get('ornek_recete_eslesme') or []
            islem_str = (r_islem.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                          if hasattr(r_islem, 'strftime') else str(r_islem))
            txt.insert(tk.END,
                f"\n  🔍 Son Reçete: RxId={r_rx},  "
                f"RxIslemTarihi={islem_str}\n")
            txt.insert(tk.END,
                f"     Logger'da eşleşen satırlar ({len(r_eslesme)}):\n")
            if not r_eslesme:
                txt.insert(tk.END,
                    f"     (eşleşme yok — LoggerId veya LoggerIslemId "
                    f"RxId ile eşleşmiyor)\n")
            else:
                for r in r_eslesme[:10]:
                    lt = r.get('LoggerTarihi')
                    lt_str = (lt.strftime('%Y-%m-%d %H:%M:%S')
                               if hasattr(lt, 'strftime') else str(lt))
                    # Tahmini fark
                    fark = ''
                    if hasattr(lt, 'hour') and hasattr(r_islem, 'hour'):
                        try:
                            f = int((lt - r_islem).total_seconds())
                            fark = f" → fark: {f:+,} sn"
                        except Exception:
                            pass
                    txt.insert(tk.END,
                        f"       LoggerIslemId={r.get('LoggerIslemId')}, "
                        f"LoggerId={r.get('LoggerId')}, "
                        f"Turu={r.get('LoggerIslemTuru')}/"
                        f"{r.get('LoggerIslemAltTuru')}, "
                        f"Tarih={lt_str}{fark}\n")

            # 9d. Son elden eşleşmesi
            e_rx = lg.get('ornek_elden_rxid')
            e_islem = lg.get('ornek_elden_islem')
            e_eslesme = lg.get('ornek_elden_eslesme') or []
            islem_str_e = (e_islem.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                            if hasattr(e_islem, 'strftime') else str(e_islem))
            txt.insert(tk.END,
                f"\n  🔍 Son Elden: RxId={e_rx},  "
                f"RxIslemTarihi={islem_str_e}\n")
            txt.insert(tk.END,
                f"     Logger'da eşleşen satırlar ({len(e_eslesme)}):\n")
            if not e_eslesme:
                txt.insert(tk.END,
                    f"     (eşleşme yok — LoggerId veya LoggerIslemId "
                    f"RxId ile eşleşmiyor)\n")
            else:
                for r in e_eslesme[:10]:
                    lt = r.get('LoggerTarihi')
                    lt_str = (lt.strftime('%Y-%m-%d %H:%M:%S')
                               if hasattr(lt, 'strftime') else str(lt))
                    fark = ''
                    if hasattr(lt, 'hour') and hasattr(e_islem, 'hour'):
                        try:
                            f = int((lt - e_islem).total_seconds())
                            fark = f" → fark: {f:+,} sn"
                        except Exception:
                            pass
                    txt.insert(tk.END,
                        f"       LoggerIslemId={r.get('LoggerIslemId')}, "
                        f"LoggerId={r.get('LoggerId')}, "
                        f"Turu={r.get('LoggerIslemTuru')}/"
                        f"{r.get('LoggerIslemAltTuru')}, "
                        f"Tarih={lt_str}{fark}\n")

            txt.insert(tk.END,
                "\n  💡 Bu bilgiyi paylaşırsan cross-check sorgusunu yazabilirim.\n"
                "     Özellikle bilmem gereken: LoggerId nedir? LoggerIslemTuru\n"
                "     hangi değer reçete, hangisi elden satış?\n")

            txt.insert(tk.END,
                "\n═══════════════════════════════════════════\n"
                "YORUMLAMA\n"
                "═══════════════════════════════════════════\n"
                "• Kolon tipinde DATETIME görüyorsan ama saatsiz örnekler\n"
                "  varsa: bu kayıtlara veri girilirken saat yazılmamış.\n"
                "• Kolon tipi DATE ise: hiçbir kayıtta saat bilgisi yok.\n"
                "• 'Saatli' örneklerde H/M/S sıfırdan büyük olmalı; eğer\n"
                "  hepsi 0 ise BotanikEOS UI'da gördüğün saat BAŞKA bir\n"
                "  kolondan (örn. RxKayitSaat) geliyor — listede onu ara.\n"
            )

        import threading
        threading.Thread(target=_calis, daemon=True).start()

    def _tarih_preset(self, gun: int):
        bugun = datetime.now()
        if gun == -1:  # Tümü
            bas = datetime(2017, 5, 23)
        else:
            bas = bugun - timedelta(days=gun)
        self.bas_entry.delete(0, tk.END)
        self.bas_entry.insert(0, bas.strftime('%Y-%m-%d'))
        self.bit_entry.delete(0, tk.END)
        self.bit_entry.insert(0, bugun.strftime('%Y-%m-%d'))

    def _esik_oku(self) -> int:
        s = self.esik_var.get() or ''
        try:
            return int(s.split(' — ')[0].strip())
        except Exception:
            try:
                return int(s.strip())
            except Exception:
                return 600

    def _tarih_oku(self, widget) -> Optional[datetime]:
        metin = (widget.get() or '').strip()
        for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(metin, fmt)
            except ValueError:
                continue
        return None

    def _taramayi_baslat(self):
        bas = self._tarih_oku(self.bas_entry)
        bit = self._tarih_oku(self.bit_entry)
        if bas is None or bit is None:
            messagebox.showerror("Tarih hatası",
                "Başlangıç ve bitiş tarihini 'YYYY-MM-DD' (veya 'DD.MM.YYYY') "
                "formatında girin.")
            return
        if bas > bit:
            messagebox.showerror("Tarih hatası",
                "Başlangıç bitişten sonra olamaz.")
            return
        # Bitiş tarihinin sonuna kadar (23:59:59)
        bit = bit.replace(hour=23, minute=59, second=59)
        esik_sn = self._esik_oku()

        # UI: disable, progress
        self.tara_btn.config(state='disabled')
        self.progress.start(10)
        self.durum_lbl.config(
            text=f"⏳ DB taranıyor: {bas.date()} → {bit.date()}  "
                 f"(eşik {esik_sn}sn) — bu işlem dakikalar sürebilir.",
            fg='#C62828',
        )
        for c in self.tree.get_children():
            self.tree.delete(c)
        self.ozet_text.delete('1.0', tk.END)
        self.ozet_text.insert('1.0', "⏳ Tarama sürüyor...")

        import time as _time
        t0 = _time.time()

        def _calis():
            try:
                rows = self.db.zaman_anomali_tarama(
                    bas, bit, esik_sn=esik_sn,
                )
                dt = _time.time() - t0
                self.top.after(0, lambda: self._sonuclari_goster(
                    rows, bas, bit, esik_sn, dt))
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Anomali taraması hatası: {e}\n{tb}")
                self.top.after(0, lambda: self._hata(str(e)))

        threading.Thread(target=_calis, daemon=True).start()

    def _hata(self, mesaj: str):
        self.progress.stop()
        self.tara_btn.config(state='normal')
        self.durum_lbl.config(text=f"❌ Hata: {mesaj}", fg='#B71C1C')
        self.ozet_text.delete('1.0', tk.END)
        self.ozet_text.insert('1.0', f"Tarama hatası: {mesaj}")
        messagebox.showerror("Tarama hatası", mesaj)

    def _filtreleri_uygula(self):
        """Checkbox değişince yeniden render. self.sonuclar bozulmaz,
        yalnızca tabloya gösterilen kayıt seti filtrelenir."""
        self._tabloyu_doldur_anomali()

    def _tabloyu_doldur_anomali(self):
        """self.sonuclar'ı tabloya yaz; checkbox filtreleri uygulanır."""
        def _dt_fmt(d):
            if d is None:
                return ''
            if hasattr(d, 'hour'):
                return d.strftime('%Y-%m-%d %H:%M:%S')
            if hasattr(d, 'strftime'):
                return d.strftime('%Y-%m-%d') + ' (saat yok)'
            return str(d)

        for c in self.tree.get_children():
            self.tree.delete(c)

        don_haric = bool(getattr(self, 'donusum_haric_var',
                                  tk.BooleanVar(value=False)).get())
        anti_haric = bool(getattr(self, 'antibiyotik_haric_var',
                                   tk.BooleanVar(value=False)).get())
        bank_haric = bool(getattr(self, 'is_bankasi_haric_var',
                                   tk.BooleanVar(value=False)).get())
        bez_haric = bool(getattr(self, 'bez_haric_var',
                                  tk.BooleanVar(value=False)).get())
        cez_haric = bool(getattr(self, 'cezaevi_haric_var',
                                  tk.BooleanVar(value=False)).get())

        idx = 0
        for k in (self.sonuclar or []):
            if don_haric and k.get('DonusumMu'):
                continue
            if anti_haric and k.get('AntibiyotikMu'):
                continue
            if bank_haric and k.get('IsBankasiMu'):
                continue
            if bez_haric and k.get('BezMidir'):
                continue
            if cez_haric and k.get('CezaeviMidir'):
                continue
            idx += 1
            i = idx
            self._tabloyu_satir_ekle(i, k, _dt_fmt)

    def _tabloyu_satir_ekle(self, i: int, k: Dict, _dt_fmt):
        """Tek bir kaydı tabloya ekler — _tabloyu_doldur_anomali'den çağrılır."""
        islem = k.get('IslemTarihi')
        kayit = k.get('KayitTarihi')
        kontrol = k.get('KontrolTarihi')
        fark = k.get('FarkSn')
        fark_tipi = k.get('FarkTipi') or 'sn'
        kontrol_fark = k.get('KontrolFarkSn')

        if fark is None:
            return
        islem_s = _dt_fmt(islem)
        kayit_s = _dt_fmt(kayit)
        kontrol_s = _dt_fmt(kontrol) if kontrol is not None else '—'

        if fark_tipi == 'gun':
            gun = int(round(fark / 86400))
            sgn = '-' if gun < 0 else ''
            kayit_fark_str = f"{sgn}{abs(gun)} gün"
            kayit_negatif = (gun < 0)
            kayit_buyuk = abs(gun) >= 7
            kayit_anomali = (gun != 0)
        else:
            kayit_fark_str = _saniye_okunabilir(fark)
            kayit_negatif = (fark < 0)
            kayit_buyuk = abs(fark) > 24 * 3600
            kayit_anomali = (abs(fark) > 600 or fark < 0)

        kontrol_anomali = kontrol_negatif = kontrol_buyuk = False
        if kontrol_fark is None:
            kontrol_fark_str = '—'
        else:
            if abs(kontrol_fark) >= 60:
                kontrol_fark_str = f"{kontrol_fark} ({_saniye_okunabilir(kontrol_fark)})"
            else:
                kontrol_fark_str = str(kontrol_fark)
            kontrol_anomali = (kontrol_fark < 0 or kontrol_fark > 600)
            kontrol_negatif = (kontrol_fark < 0)
            kontrol_buyuk = (abs(kontrol_fark) > 24 * 3600)

        tag = 'normal'
        if kayit_negatif or kontrol_negatif:
            tag = 'negatif'
        elif kayit_buyuk or kontrol_buyuk:
            tag = 'cok_buyuk'
        elif fark_tipi == 'gun' and kayit_anomali:
            tag = 'gun_anomali'
        elif kontrol_anomali:
            tag = 'buyuk'

        if kayit_negatif:
            kayit_fark_str += " (NEGATİF)"
        if kontrol_negatif:
            kontrol_fark_str += " (NEGATİF)"

        don_str = '🔄' if k.get('DonusumMu') else ''
        anti_str = '🦠' if k.get('AntibiyotikMu') else ''
        bank_str = '🏦' if k.get('IsBankasiMu') else ''
        bez_str = '🩹' if k.get('BezMidir') else ''
        cez_str = '🔒' if k.get('CezaeviMidir') else ''

        self.tree.insert('', 'end', values=(
            i, k.get('Kaynak'), k.get('RxId'), islem_s, kayit_s,
            kayit_fark_str, kontrol_s, kontrol_fark_str,
            don_str, anti_str, bank_str, bez_str, cez_str,
        ), tags=(tag,))

    def _sirala(self, sutun_kod: str):
        """Anomali tablosunu sütuna göre sırala. Aynı sütuna ikinci tıkta
        yön ters. self.sonuclar listesi yeniden düzenlenir; özet/istatistik
        değişmez."""
        if not self.sonuclar:
            return
        ters = (self._son_sirala_sutun == sutun_kod) and not self._son_sirala_ters
        self._son_sirala_sutun = sutun_kod
        self._son_sirala_ters = ters

        def _safe(v, default):
            return v if v is not None else default

        anahtarlar = {
            'no':           lambda r: 0,
            'kaynak':       lambda r: r.get('Kaynak') or '',
            'rxid':         lambda r: _safe(r.get('RxId'), 0),
            'islem':        lambda r: _safe(r.get('IslemTarihi'), datetime.min),
            'kayit':        lambda r: _safe(r.get('KayitTarihi'), datetime.min),
            'kayit_fark':   lambda r: _safe(r.get('FarkSn'), 0),
            'kontrol':      lambda r: _safe(r.get('KontrolTarihi'), datetime.min),
            'kontrol_fark': lambda r: _safe(r.get('KontrolFarkSn'), 0),
            'donusum':      lambda r: 1 if r.get('DonusumMu') else 0,
            'antibiyotik':  lambda r: 1 if r.get('AntibiyotikMu') else 0,
            'isbankasi':    lambda r: 1 if r.get('IsBankasiMu') else 0,
            'bez':          lambda r: 1 if r.get('BezMidir') else 0,
            'cezaevi':      lambda r: 1 if r.get('CezaeviMidir') else 0,
        }
        key_fn = anahtarlar.get(sutun_kod, anahtarlar['islem'])
        try:
            self.sonuclar = sorted(self.sonuclar, key=key_fn, reverse=ters)
        except TypeError:
            self.sonuclar = sorted(
                self.sonuclar, key=lambda r: str(key_fn(r)), reverse=ters,
            )
        # Başlığa ↓/↑ ekle
        ok = ' ↓' if ters else ' ↑'
        for c, orj in self._anomali_basliklar_orj.items():
            yeni = orj + (ok if c == sutun_kod else '')
            try:
                self.tree.heading(c, text=yeni)
            except Exception:
                pass
        self._tabloyu_doldur_anomali()

    def _sonuclari_goster(self, rows: List[Dict], bas: datetime, bit: datetime,
                          esik_sn: int, sure: float):
        self.progress.stop()
        self.tara_btn.config(state='normal')
        self.sonuclar = rows or []

        _sn_okunabilir = _saniye_okunabilir

        # İstatistikler (sıradan bağımsız — tek sefer)
        from collections import Counter
        kaynak_say = Counter()
        yil_say = Counter()
        fark_kat = Counter()
        tipi_say = Counter()  # 'sn' / 'gun'
        negatif_say = 0
        max_fark = 0
        min_fark = 0
        toplam_fark = 0
        # Kontrol fark istatistikleri
        kontrol_anomali_say = 0
        kontrol_negatif_say = 0
        kontrol_max = 0
        kontrol_min = 0

        for k in self.sonuclar:
            fark = k.get('FarkSn')
            if fark is None:
                continue
            fark_tipi = k.get('FarkTipi') or 'sn'
            tipi_say[fark_tipi] += 1
            kaynak_say[k.get('Kaynak')] += 1
            islem = k.get('IslemTarihi')
            if hasattr(islem, 'year'):
                yil_say[islem.year] += 1
            toplam_fark += fark
            if fark > max_fark:
                max_fark = fark
            if fark < min_fark:
                min_fark = fark
            if fark < 0:
                negatif_say += 1
                kat = 'Negatif (kayıt işlemden önce)'
            elif fark_tipi == 'gun':
                gun = int(round(fark / 86400))
                if gun >= 30:
                    kat = '> 30 gün (saatsiz)'
                elif gun >= 7:
                    kat = '1 hafta – 30 gün (saatsiz)'
                elif gun >= 1:
                    kat = '1 – 7 gün (saatsiz)'
                else:
                    kat = 'Saatsiz, ≥1 gün'
            elif fark <= 3600:
                kat = '10 dk – 1 sa'
            elif fark <= 24 * 3600:
                kat = '1 sa – 24 sa'
            elif fark <= 7 * 24 * 3600:
                kat = '1 – 7 gün'
            elif fark <= 30 * 24 * 3600:
                kat = '1 hafta – 30 gün'
            else:
                kat = '> 30 gün'
            fark_kat[kat] += 1

            # Kontrol fark istatistiği
            kf = k.get('KontrolFarkSn')
            if kf is not None and (kf < 0 or kf > 600):
                kontrol_anomali_say += 1
                if kf < 0:
                    kontrol_negatif_say += 1
                if kf > kontrol_max:
                    kontrol_max = kf
                if kf < kontrol_min:
                    kontrol_min = kf

        # Tabloyu doldur (DB sıralı: |fark| azalan)
        # Sıralama durumunu sıfırla (yeni tarama)
        self._son_sirala_sutun = None
        self._son_sirala_ters = False
        for c, orj in self._anomali_basliklar_orj.items():
            try:
                self.tree.heading(c, text=orj)
            except Exception:
                pass
        self._tabloyu_doldur_anomali()

        n = len(self.sonuclar or [])
        # SQL hatası tarama bitince kontrol et — n>0 olsa bile uyar
        sql_hata = getattr(self.db, 'son_sorgu_hatasi', None)
        self.ozet_text.delete('1.0', tk.END)
        if n == 0:
            mesaj = (f"✓ Tarama tamam ({sure:.1f}sn). "
                     f"Aralık {bas.date()} → {bit.date()}, eşik {esik_sn}sn. "
                     f"\nEşik üstünde anomali bulunamadı.")
            if sql_hata:
                mesaj = (f"❌ SQL HATASI ({sure:.1f}sn): {sql_hata}\n\n"
                         f"Tarama gerçekleşemedi. Lütfen log'a bakın "
                         f"veya eşiği '0 sn' yapıp tekrar deneyin.")
                logger.error(f"zaman_anomali_tarama SQL hatası: {sql_hata}")
            else:
                mesaj += ("\n\nHiç fark olmadığından emin değilseniz: "
                          "Eşik dropdown'unu '0 sn — fark olan TÜM kayıtlar' "
                          "yapıp tekrar tarayın.")
            self.ozet_text.insert('1.0', mesaj)
            self.durum_lbl.config(
                text=f"✓ Tarama tamam: 0 anomali bulundu ({sure:.1f}sn).",
                fg='#B71C1C' if sql_hata else '#2E7D32',
            )
            return

        ort = toplam_fark / n if n > 0 else 0
        sn_say = tipi_say.get('sn', 0)
        gun_say = tipi_say.get('gun', 0)
        ozet = (
            f"✓ Tarama tamam ({sure:.1f}sn).  "
            f"Aralık: {bas.date()} → {bit.date()}  •  Eşik: {esik_sn}sn "
            f"({_sn_okunabilir(esik_sn)})\n"
            f"\n"
            f"📊 Toplam anomali: {n}  "
            f"(Reçete: {kaynak_say.get('RECETE', 0)}, "
            f"Elden: {kaynak_say.get('ELDEN', 0)})\n"
        )
        if gun_say:
            ozet += (f"   ⏱ Hassasiyet: {sn_say} saniye bazlı, "
                     f"{gun_say} gün bazlı (kayıt saatsiz).\n")
        ozet += (
            f"   📅 KAYIT farkı (RxKayitTarihi vs RxIslemTarihi):\n"
            f"      Min: {min_fark} sn ({_sn_okunabilir(min_fark)})    "
            f"Max: {max_fark} sn ({_sn_okunabilir(max_fark)})    "
            f"Ort: {ort:.0f} sn ({_sn_okunabilir(int(ort))})\n"
            f"      Negatif (kayıt < işlem): {negatif_say} kayıt\n"
        )
        if kontrol_anomali_say:
            ozet += (
                f"   🔒 KONTROL farkı (RxKontrolTarihi vs RxIslemTarihi, "
                f"sadece Reçete):\n"
                f"      Anomali: {kontrol_anomali_say} kayıt   "
                f"Negatif: {kontrol_negatif_say}\n"
                f"      Min: {kontrol_min} sn ({_sn_okunabilir(kontrol_min)})    "
                f"Max: {kontrol_max} sn ({_sn_okunabilir(kontrol_max)})\n"
            )

        # 5 bayrak özeti (filtre amaçlı işaretleme)
        n_don = sum(1 for k in self.sonuclar if k.get('DonusumMu'))
        n_anti = sum(1 for k in self.sonuclar if k.get('AntibiyotikMu'))
        n_bank = sum(1 for k in self.sonuclar if k.get('IsBankasiMu'))
        n_bez = sum(1 for k in self.sonuclar if k.get('BezMidir'))
        n_cez = sum(1 for k in self.sonuclar if k.get('CezaeviMidir'))
        n_aciklanmamis = sum(
            1 for k in self.sonuclar
            if not k.get('DonusumMu') and not k.get('AntibiyotikMu')
            and not k.get('IsBankasiMu') and not k.get('BezMidir')
            and not k.get('CezaeviMidir')
        )
        ozet += (
            f"\n"
            f"🏷 KATEGORİLER (gizleme filtreleri için):\n"
            f"   • 🔄 Perakende→Reçete dönüşümü : {n_don:>6} "
            f"({(n_don/n)*100:.1f}%)\n"
            f"   • 🦠 Antibiyotik içeren        : {n_anti:>6} "
            f"({(n_anti/n)*100:.1f}%)\n"
            f"   • 🏦 İş Bankası reçetesi       : {n_bank:>6} "
            f"({(n_bank/n)*100:.1f}%)\n"
            f"   • 🩹 Bez içeren                : {n_bez:>6} "
            f"({(n_bez/n)*100:.1f}%)\n"
            f"   • 🔒 Cezaevi reçetesi          : {n_cez:>6} "
            f"({(n_cez/n)*100:.1f}%)\n"
            f"   • 🚨 Açıklanmamış (hiçbiri)    : {n_aciklanmamis:>6} "
            f"({(n_aciklanmamis/n)*100:.1f}%)\n"
        )

        ozet += (
            f"\n"
            f"📈 Fark kategorisi:\n"
        )
        for kat in ['Negatif (kayıt işlemden önce)',
                    '10 dk – 1 sa', '1 sa – 24 sa',
                    '1 – 7 gün', '1 hafta – 30 gün', '> 30 gün',
                    '1 – 7 gün (saatsiz)', '1 hafta – 30 gün (saatsiz)',
                    '> 30 gün (saatsiz)']:
            sayi = fark_kat.get(kat, 0)
            if sayi:
                yuzde = (sayi / n) * 100
                ozet += f"   • {kat:<32} : {sayi:>6} ({yuzde:.1f}%)\n"
        if yil_say:
            ozet += "\n📅 Yıllara göre:\n"
            for yil in sorted(yil_say.keys()):
                ozet += f"   • {yil} : {yil_say[yil]} anomali\n"

        # Üst sınırı uyarısı
        if n >= 200000:
            ozet += ("\n⚠ Maksimum kayıt sınırı (200000) aşıldı — "
                     "gerçek sayı daha yüksek olabilir. Eşiği yükseltin veya "
                     "aralığı daraltın.")

        self.ozet_text.insert('1.0', ozet)
        self.durum_lbl.config(
            text=f"✓ Tarama tamam: {n} anomali ({sure:.1f}sn).",
            fg='#2E7D32',
        )

    def _excel_aktar(self):
        if not self.sonuclar:
            messagebox.showinfo("Boş", "Önce taramayı çalıştırın.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("openpyxl yok",
                "Excel export için openpyxl gerekli.")
            return
        from tkinter import filedialog as fd
        dosya = fd.asksaveasfilename(
            title="Anomali raporu kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"zaman_anomali_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not dosya:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Anomali"
        basliklar = ['No', 'Kaynak', 'RxId', 'İşlem Tarihi',
                     'Kayıt Tarihi', 'Kayıt Fark (gün)',
                     'Kontrol Tarihi', 'Kontrol Fark (sn)',
                     'Kontrol Fark (sa:dk:sn)',
                     'Dönüşüm', 'Antibiyotik', 'İş Bankası', 'Bez',
                     'Cezaevi']
        bf = Font(bold=True, color='FFFFFF')
        bg = PatternFill(start_color='5E35B1', end_color='5E35B1', fill_type='solid')
        for ci, b in enumerate(basliklar, 1):
            c = ws.cell(row=1, column=ci, value=b)
            c.font = bf
            c.fill = bg
            c.alignment = Alignment(horizontal='center')

        for i, k in enumerate(self.sonuclar, 1):
            islem = k.get('IslemTarihi')
            kayit = k.get('KayitTarihi')
            kontrol = k.get('KontrolTarihi')
            fark = k.get('FarkSn')
            fark_tipi = k.get('FarkTipi') or 'sn'
            kontrol_fark = k.get('KontrolFarkSn')

            ws.cell(row=i+1, column=1, value=i)
            ws.cell(row=i+1, column=2, value=k.get('Kaynak'))
            ws.cell(row=i+1, column=3, value=k.get('RxId'))
            ws.cell(row=i+1, column=4, value=islem)
            ws.cell(row=i+1, column=5, value=kayit)
            # Kayıt fark = gün
            if fark is not None and fark_tipi == 'gun':
                ws.cell(row=i+1, column=6, value=int(round(fark / 86400)))
            elif fark is not None:
                # saatli senaryoda saniye/86400 ile gün cinsinden de gösterilebilir
                ws.cell(row=i+1, column=6, value=round(fark / 86400, 2))
            ws.cell(row=i+1, column=7, value=kontrol)
            if kontrol_fark is not None:
                ws.cell(row=i+1, column=8, value=int(kontrol_fark))
                a = abs(int(kontrol_fark))
                sa = a // 3600
                dk = (a % 3600) // 60
                ksn = a % 60
                sgn = '-' if kontrol_fark < 0 else ''
                ws.cell(row=i+1, column=9,
                        value=f"{sgn}{sa:02d}:{dk:02d}:{ksn:02d}")
            ws.cell(row=i+1, column=10,
                    value='Evet' if k.get('DonusumMu') else 'Hayır')
            ws.cell(row=i+1, column=11,
                    value='Evet' if k.get('AntibiyotikMu') else 'Hayır')
            ws.cell(row=i+1, column=12,
                    value='Evet' if k.get('IsBankasiMu') else 'Hayır')
            ws.cell(row=i+1, column=13,
                    value='Evet' if k.get('BezMidir') else 'Hayır')
            ws.cell(row=i+1, column=14,
                    value='Evet' if k.get('CezaeviMidir') else 'Hayır')

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 18
        try:
            wb.save(dosya)
            messagebox.showinfo("Kaydedildi",
                f"{len(self.sonuclar)} anomali kaydı kaydedildi:\n{dosya}")
        except Exception as e:
            messagebox.showerror("Hata", str(e))


class LoggerForensikPopup:
    """Logger tablosu kullanarak adli zaman analizi:
    1. Manuel tarih değişikliği tespit (RxIslemTarihi ≠ Logger orijinal kayıt)
    2. LoggerMakina başına ortalama saat sapması (çoklu PC analizi)
    """

    def __init__(self, parent, db):
        self.parent = parent
        self.db = db
        self.son_manuel: List[Dict] = []
        self.son_makina: List[Dict] = []
        self._son_sirala_sutun: Optional[str] = None
        self._son_sirala_ters: bool = False

        self.top = tk.Toplevel(parent)
        self.top.title("🔬 Logger Forensik Analiz")
        self.top.geometry("1400x800")
        try:
            self.top.state('zoomed')
        except Exception:
            pass
        self.top.transient(parent)

        try:
            self._arayuz()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"LoggerForensikPopup HATA: {e}\n{tb}")
            print(f"[Logger Forensik] HATA: {e}\n{tb}")

    def _arayuz(self):
        # Header
        header = tk.Frame(self.top, bg='#00838F', height=46)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header,
                 text="🔬 Logger Forensik — Manuel Değişiklik + PC Sapma Analizi",
                 font=("Segoe UI", 11, "bold"), bg='#00838F', fg='white',
                 ).pack(side="left", padx=15, pady=12)
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self.top.destroy).pack(side="right", padx=10, pady=8)
        tk.Button(header, text="📥 Excel", bg='#2E7D32', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._excel_aktar).pack(side="right", padx=5, pady=8)
        tk.Button(header, text="🔎 Detay Gör", bg='#FF8F00', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._secili_detayi_ac).pack(side="right", padx=5, pady=8)

        # Açıklama
        aciklama = tk.Label(self.top,
            text="ℹ Logger tablosundaki orijinal kayıt zamanı vs ReceteAna/EldenAna "
                 "RxIslemTarihi karşılaştırması. Fark varsa: manuel tarih değişikliği "
                 "(insert-only log korunur, RxIslemTarihi sonradan elle değiştirildi). "
                 "Ayrıca LoggerMakina ile her PC'nin ortalama saat sapması analizi.",
            font=("Segoe UI", 9, "italic"), fg='#37474F', bg='#FFF8E1',
            anchor='w', justify='left', padx=15, pady=8, wraplength=1350)
        aciklama.pack(fill="x")

        # Parametre paneli
        filt = ttk.LabelFrame(self.top, text="Tarama Parametreleri", padding=6)
        filt.pack(fill="x", padx=5, pady=4)
        s1 = tk.Frame(filt)
        s1.pack(fill="x", pady=3)
        tk.Label(s1, text="Başlangıç:", font=("Segoe UI", 9, "bold")
                 ).pack(side="left", padx=4)
        self.bas_entry = ttk.Entry(s1, width=14)
        self.bas_entry.insert(0, "2017-05-23")
        self.bas_entry.pack(side="left", padx=4)
        tk.Label(s1, text="→  Bitiş:",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        self.bit_entry = ttk.Entry(s1, width=14)
        self.bit_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.bit_entry.pack(side="left", padx=4)
        tk.Label(s1, text="    Manuel değişiklik eşiği (sn):",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(15, 4))
        self.esik_entry = ttk.Entry(s1, width=8)
        self.esik_entry.insert(0, "5")
        self.esik_entry.pack(side="left", padx=4)

        for et, gun in [("Son 90 gün", 90), ("Son 1 yıl", 365),
                         ("Tümü (2017→)", -1)]:
            ttk.Button(s1, text=et, width=14,
                       command=lambda g=gun: self._tarih_preset(g)
                       ).pack(side="left", padx=2)

        self.tara_btn = ttk.Button(s1, text="▶ Taramayı Başlat", width=20,
                                    command=self._taramayi_baslat)
        self.tara_btn.pack(side="right", padx=8)

        # Satır 2: filtre checkboxları (taramadan sonra etkili)
        s2 = tk.Frame(filt)
        s2.pack(fill="x", pady=(2, 1))
        self.donusum_haric_var = tk.BooleanVar(value=False)
        self.antibiyotik_haric_var = tk.BooleanVar(value=False)
        self.is_bankasi_haric_var = tk.BooleanVar(value=False)
        self.bez_haric_var = tk.BooleanVar(value=False)
        self.cezaevi_haric_var = tk.BooleanVar(value=False)
        tk.Label(s2, text="Filtreler:",
                 font=("Segoe UI", 9, "bold"), fg='#37474F'
                 ).pack(side="left", padx=(4, 8))
        ttk.Checkbutton(
            s2,
            text="🔄 Perakende→Reçete dönüşümlerini gizle "
                 "(Logger Turu=2/AltTuru=3 izi var)",
            variable=self.donusum_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=8)
        ttk.Checkbutton(
            s2,
            text="🦠 Antibiyotik içerenleri gizle (ATC J01/J02/J04/J05)",
            variable=self.antibiyotik_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=8)
        ttk.Checkbutton(
            s2,
            text="🏦 İş Bankası reçetelerini gizle "
                 "(banka mensup sigortası özel akışı)",
            variable=self.is_bankasi_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=8)
        ttk.Checkbutton(
            s2,
            text="🩹 Bez içerenleri gizle "
                 "(hasta bezi / gazlı bez / mesane pedi vb.)",
            variable=self.bez_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=8)
        ttk.Checkbutton(
            s2,
            text="🔒 Cezaevi reçetelerini gizle "
                 "(KurumCezaeviId / KurumAdi CEZA/İNFAZ)",
            variable=self.cezaevi_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=8)

        # Durum
        durum = tk.Frame(self.top, bg='#F5F5F5', height=28)
        durum.pack(fill="x")
        durum.pack_propagate(False)
        self.durum_lbl = tk.Label(durum,
            text="Hazır. 'Taramayı Başlat'a basın.",
            bg='#F5F5F5', fg='#1565C0', font=("Segoe UI", 9, "italic"),
            anchor='w', padx=10)
        self.durum_lbl.pack(side="left", fill="x", expand=True)
        self.progress = ttk.Progressbar(durum, mode='indeterminate', length=200)
        self.progress.pack(side="right", padx=10, pady=3)

        # Üst panel: Manuel değişiklik tablosu
        ust = ttk.LabelFrame(self.top,
            text="📝 1. MUHTEMEL MANUEL TARİH DEĞİŞİKLİKLERİ "
                 "(RxIslemTarihi ≠ Logger orijinal kayıt)", padding=4)
        ust.pack(fill="both", expand=True, padx=5, pady=3)

        cols = ('no', 'kaynak', 'rxid', 'islem', 'logger', 'fark_sn',
                'fark', 'makina', 'personel',
                'donusum', 'antibiyotik', 'isbankasi', 'bez', 'cezaevi')
        bash = {'no': '#', 'kaynak': 'Kaynak', 'rxid': 'RxId',
                'islem': 'RxIslemTarihi (değişen)',
                'logger': 'Logger Orijinal',
                'fark_sn': 'Fark (sn)', 'fark': 'Fark (okunabilir)',
                'makina': 'Logger Makina',
                'personel': 'Personel',
                'donusum': '🔄 Dönüşüm',
                'antibiyotik': '🦠 Antibiyotik',
                'isbankasi': '🏦 İş Bankası',
                'bez': '🩹 Bez',
                'cezaevi': '🔒 Cezaevi'}
        widths = {'no': 50, 'kaynak': 70, 'rxid': 80,
                  'islem': 180, 'logger': 180,
                  'fark_sn': 130, 'fark': 130,
                  'makina': 150, 'personel': 80,
                  'donusum': 90, 'antibiyotik': 100, 'isbankasi': 100,
                  'bez': 80, 'cezaevi': 90}
        anchors = {'no': 'center', 'kaynak': 'center', 'rxid': 'center',
                   'personel': 'center',
                   'donusum': 'center', 'antibiyotik': 'center',
                   'isbankasi': 'center', 'bez': 'center',
                   'cezaevi': 'center'}
        self.manuel_tree = ttk.Treeview(ust, columns=cols, show='headings', height=12)
        self._manuel_basliklar_orj = dict(bash)
        for c in cols:
            self.manuel_tree.heading(c, text=bash[c],
                                       command=lambda col=c: self._sirala(col))
            self.manuel_tree.column(c, width=widths[c],
                                     anchor=anchors.get(c, 'w'))
        self.manuel_tree.pack(side="left", fill="both", expand=True)
        sb1 = ttk.Scrollbar(ust, orient="vertical",
                             command=self.manuel_tree.yview)
        sb1.pack(side="right", fill="y")
        self.manuel_tree.configure(yscrollcommand=sb1.set)
        self.manuel_tree.tag_configure('negatif', background='#FFCDD2')
        self.manuel_tree.tag_configure('buyuk', background='#FFAB91')
        self.manuel_tree.tag_configure('normal', background='white')
        self.manuel_tree.bind('<Double-1>', self._secili_detayi_ac)

        # Alt panel: PC (LoggerMakina) sapma özeti
        alt = ttk.LabelFrame(self.top,
            text="🖥 2. LOGGER MAKİNA (PC) BAŞINA ORTALAMA SAPMA",
            padding=4)
        alt.pack(fill="x", padx=5, pady=3)

        m_cols = ('makina', 'sayi', 'fark_ort', 'mutlak_ort',
                  'fark_min', 'fark_max', 'supheli')
        m_bash = {'makina': 'PC / Makina',
                  'sayi': 'İşlem Sayısı',
                  'fark_ort': 'Ort. Fark (sn)',
                  'mutlak_ort': '|Fark| Ortalama (sn)',
                  'fark_min': 'Min Fark',
                  'fark_max': 'Max Fark',
                  'supheli': '|Fark|>5sn Sayı'}
        m_widths = {'makina': 250, 'sayi': 100,
                    'fark_ort': 130, 'mutlak_ort': 150,
                    'fark_min': 100, 'fark_max': 100, 'supheli': 130}
        self.makina_tree = ttk.Treeview(alt, columns=m_cols,
                                          show='headings', height=6)
        for c in m_cols:
            self.makina_tree.heading(c, text=m_bash[c])
            self.makina_tree.column(c, width=m_widths[c],
                                      anchor=('w' if c == 'makina' else 'e'))
        self.makina_tree.pack(side="left", fill="both", expand=True)
        sb2 = ttk.Scrollbar(alt, orient="vertical",
                             command=self.makina_tree.yview)
        sb2.pack(side="right", fill="y")
        self.makina_tree.configure(yscrollcommand=sb2.set)
        self.makina_tree.tag_configure('supheli', background='#FFE0B2')

    def _tarih_preset(self, gun: int):
        bugun = datetime.now()
        if gun == -1:
            bas = datetime(2017, 5, 23)
        else:
            bas = bugun - timedelta(days=gun)
        self.bas_entry.delete(0, tk.END)
        self.bas_entry.insert(0, bas.strftime('%Y-%m-%d'))
        self.bit_entry.delete(0, tk.END)
        self.bit_entry.insert(0, bugun.strftime('%Y-%m-%d'))

    def _tarih_oku(self, widget) -> Optional[datetime]:
        metin = (widget.get() or '').strip()
        for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(metin, fmt)
            except ValueError:
                continue
        return None

    def _taramayi_baslat(self):
        bas = self._tarih_oku(self.bas_entry)
        bit = self._tarih_oku(self.bit_entry)
        if bas is None or bit is None:
            messagebox.showerror("Tarih hatası",
                "Başlangıç ve bitiş tarihini 'YYYY-MM-DD' formatında girin.")
            return
        if bas > bit:
            messagebox.showerror("Tarih hatası",
                "Başlangıç bitişten sonra olamaz.")
            return
        bit = bit.replace(hour=23, minute=59, second=59)
        try:
            esik = int((self.esik_entry.get() or '5').strip())
        except ValueError:
            esik = 5

        self.tara_btn.config(state='disabled')
        self.progress.start(10)
        self.durum_lbl.config(
            text=f"⏳ Taranıyor: {bas.date()} → {bit.date()}, eşik {esik}sn",
            fg='#C62828')
        for c in self.manuel_tree.get_children():
            self.manuel_tree.delete(c)
        for c in self.makina_tree.get_children():
            self.makina_tree.delete(c)

        import time as _time
        t0 = _time.time()

        def _calis():
            try:
                d = self.db.logger_manuel_degisiklik_tarama(
                    bas, bit, esik_sn=esik,
                )
                dt = _time.time() - t0
                self.top.after(0, lambda: self._sonuclari_goster(d, dt))
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Logger forensik hatası: {e}\n{tb}")
                self.top.after(0, lambda: self._hata(str(e)))

        threading.Thread(target=_calis, daemon=True).start()

    def _hata(self, mesaj: str):
        self.progress.stop()
        self.tara_btn.config(state='normal')
        self.durum_lbl.config(text=f"❌ Hata: {mesaj}", fg='#B71C1C')
        messagebox.showerror("Tarama hatası", mesaj)

    def _sonuclari_goster(self, d: Dict, sure: float):
        self.progress.stop()
        self.tara_btn.config(state='normal')

        self.son_manuel = d.get('manuel_kayitlar') or []
        self.son_makina = d.get('makina_ozeti') or []
        self._son_sure = sure

        # Filtre uygulayıp tabloyu doldur (özet etiketi de güncelleniyor)
        self._filtreleri_uygula()

        # Makina özeti (filtreden etkilenmez)
        for m in self.son_makina:
            mutlak = m.get('MutlakOrt', 0)
            tag = 'supheli' if mutlak > 60 else ''
            self.makina_tree.insert('', 'end', values=(
                m.get('LoggerMakina'),
                f"{m.get('IslemSayisi'):,}",
                f"{m.get('FarkOrtalama', 0):+.1f}",
                f"{mutlak:.1f}",
                f"{m.get('FarkMin', 0):+,d}",
                f"{m.get('FarkMax', 0):+,d}",
                f"{m.get('SuphliSayi', 0):,}",
            ), tags=(tag,))

    @staticmethod
    def _dt_fmt(v):
        if v is None:
            return ''
        if hasattr(v, 'strftime') and hasattr(v, 'hour'):
            return v.strftime('%Y-%m-%d %H:%M:%S')
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)

    def _tabloyu_doldur(self, kayitlar: List[Dict]):
        """Manuel değişiklik tablosunu verilen kayıt listesi ile doldurur.
        (Filtrelenmiş veya tam liste — sıralama uygulanmış olabilir.)
        """
        for c in self.manuel_tree.get_children():
            self.manuel_tree.delete(c)
        for i, k in enumerate(kayitlar, 1):
            fark = k.get('FarkSn')
            tag = 'normal'
            if fark is not None:
                if fark < 0:
                    tag = 'negatif'
                elif abs(fark) > 3600:
                    tag = 'buyuk'
            fark_sn_str = (f"{fark:+,d}" if fark is not None else '?')
            fark_str = _saniye_okunabilir(fark) if fark is not None else '?'
            don_str = '🔄 Evet' if k.get('DonusumMu') else ''
            anti_str = '🦠 Evet' if k.get('AntibiyotikMu') else ''
            bank_str = '🏦 Evet' if k.get('IsBankasiMu') else ''
            bez_str = '🩹 Evet' if k.get('BezMidir') else ''
            cez_str = '🔒 Evet' if k.get('CezaeviMidir') else ''
            self.manuel_tree.insert('', 'end', values=(
                i, k.get('Kaynak'), k.get('RxId'),
                self._dt_fmt(k.get('RxIslemTarihi')),
                self._dt_fmt(k.get('LoggerTarihi')),
                fark_sn_str, fark_str,
                k.get('LoggerMakina') or '(?)',
                k.get('LoggerPersonelId') or '',
                don_str, anti_str, bank_str, bez_str, cez_str,
            ), tags=(tag,))

    def _filtreleri_uygula(self):
        """Checkbox değişince çağrılır. son_manuel listesinden gizlenecekleri
        çıkarır, tabloyu yeniler ve özet durum etiketini günceller."""
        don_haric = bool(self.donusum_haric_var.get())
        anti_haric = bool(self.antibiyotik_haric_var.get())
        bank_haric = bool(self.is_bankasi_haric_var.get())
        bez_haric = bool(self.bez_haric_var.get())
        cez_haric = bool(self.cezaevi_haric_var.get())

        gosterilecek = [
            k for k in self.son_manuel
            if not (don_haric and k.get('DonusumMu'))
            and not (anti_haric and k.get('AntibiyotikMu'))
            and not (bank_haric and k.get('IsBankasiMu'))
            and not (bez_haric and k.get('BezMidir'))
            and not (cez_haric and k.get('CezaeviMidir'))
        ]
        # Mevcut sıralama varsa koru
        if self._son_sirala_sutun:
            self._uygula_sirala(gosterilecek,
                                self._son_sirala_sutun,
                                self._son_sirala_ters)
        self._tabloyu_doldur(gosterilecek)

        # Durum etiketi: özet bilgi
        n_top = len(self.son_manuel)
        n_don = sum(1 for k in self.son_manuel if k.get('DonusumMu'))
        n_anti = sum(1 for k in self.son_manuel if k.get('AntibiyotikMu'))
        n_bank = sum(1 for k in self.son_manuel if k.get('IsBankasiMu'))
        n_bez = sum(1 for k in self.son_manuel if k.get('BezMidir'))
        n_cez = sum(1 for k in self.son_manuel if k.get('CezaeviMidir'))
        n_aciklanmamis = sum(
            1 for k in self.son_manuel
            if not k.get('DonusumMu')
            and not k.get('AntibiyotikMu')
            and not k.get('IsBankasiMu')
            and not k.get('BezMidir')
            and not k.get('CezaeviMidir')
        )
        n_gor = len(gosterilecek)
        sure = getattr(self, '_son_sure', 0.0)
        ozet = (f"✓ ({sure:.1f}sn) Toplam {n_top} manuel "
                f"| 🔄 Dönüşüm: {n_don} | 🦠 Antibiyotik: {n_anti} "
                f"| 🏦 İş Bankası: {n_bank} | 🩹 Bez: {n_bez} "
                f"| 🔒 Cezaevi: {n_cez} "
                f"| 🚨 Açıklanmamış: {n_aciklanmamis}")
        if don_haric or anti_haric or bank_haric or bez_haric or cez_haric:
            ozet += f"  →  GÖSTERİLEN: {n_gor}"
        self.durum_lbl.config(text=ozet, fg='#2E7D32')

    def _uygula_sirala(self, kayitlar: List[Dict], sutun_kod: str, ters: bool):
        """In-place sıralama yapar. _sirala ve _filtreleri_uygula'dan çağrılır."""
        def _safe(v, default):
            return v if v is not None else default

        anahtarlar = {
            'no':         lambda r: 0,
            'kaynak':     lambda r: r.get('Kaynak') or '',
            'rxid':       lambda r: _safe(r.get('RxId'), 0),
            'islem':      lambda r: _safe(r.get('RxIslemTarihi'), datetime.min),
            'logger':     lambda r: _safe(r.get('LoggerTarihi'), datetime.min),
            'fark_sn':    lambda r: _safe(r.get('FarkSn'), 0),
            'fark':       lambda r: _safe(r.get('FarkSn'), 0),
            'makina':     lambda r: r.get('LoggerMakina') or '',
            'personel':   lambda r: _safe(r.get('LoggerPersonelId'), 0),
            'donusum':    lambda r: 1 if r.get('DonusumMu') else 0,
            'antibiyotik': lambda r: 1 if r.get('AntibiyotikMu') else 0,
            'isbankasi':  lambda r: 1 if r.get('IsBankasiMu') else 0,
            'bez':        lambda r: 1 if r.get('BezMidir') else 0,
            'cezaevi':    lambda r: 1 if r.get('CezaeviMidir') else 0,
        }
        key_fn = anahtarlar.get(sutun_kod, anahtarlar['fark_sn'])
        try:
            kayitlar.sort(key=key_fn, reverse=ters)
        except TypeError:
            kayitlar.sort(key=lambda r: str(key_fn(r)), reverse=ters)

    def _sirala(self, sutun_kod: str):
        if not self.son_manuel:
            return
        ters = (self._son_sirala_sutun == sutun_kod) and not self._son_sirala_ters
        self._son_sirala_sutun = sutun_kod
        self._son_sirala_ters = ters

        # Sıralamayı tam listeye uygula, sonra filtreyi tekrar çalıştır (tablo
        # yeniden dolar)
        self._uygula_sirala(self.son_manuel, sutun_kod, ters)
        self._filtreleri_uygula()

        # Başlık ok'unu güncelle
        ok = ' ↓' if ters else ' ↑'
        for c, orj in self._manuel_basliklar_orj.items():
            yeni = orj + (ok if c == sutun_kod else '')
            try:
                self.manuel_tree.heading(c, text=yeni)
            except Exception:
                pass

    def _secili_detayi_ac(self, _event=None):
        sec = self.manuel_tree.selection()
        if not sec:
            messagebox.showinfo("Seçim yok",
                "Önce manuel değişiklik tablosundan bir satır seçin.")
            return
        vals = self.manuel_tree.item(sec[0], 'values')
        if not vals or len(vals) < 3:
            return
        kaynak = (vals[1] or '').strip().upper()
        rx_id = (vals[2] or '').strip()
        if kaynak not in ('RECETE', 'ELDEN') or not rx_id:
            return
        SatisDetayPopup(parent=self.top, db=self.db,
                        kaynak=kaynak, rx_id=rx_id)

    def _excel_aktar(self):
        if not self.son_manuel and not self.son_makina:
            messagebox.showinfo("Boş", "Önce taramayı çalıştırın.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("openpyxl yok",
                "Excel export için openpyxl gerekli.")
            return
        from tkinter import filedialog as fd
        dosya = fd.asksaveasfilename(
            title="Logger Forensik raporu kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"logger_forensik_"
                        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not dosya:
            return

        wb = Workbook()
        # Sheet 1: Manuel değişiklikler
        ws1 = wb.active
        ws1.title = "Manuel Değişiklik"
        bf = Font(bold=True, color='FFFFFF')
        bg = PatternFill(start_color='00838F', end_color='00838F',
                          fill_type='solid')
        basliklar1 = ['No', 'Kaynak', 'RxId', 'RxIslemTarihi (değişen)',
                      'Logger Orijinal', 'Fark (sn)', 'Fark (sa:dk:sn)',
                      'Logger Makina', 'Personel',
                      'Dönüşüm (Perakende→Reçete)', 'Antibiyotik (ATC J)',
                      'İş Bankası (Kurum)', 'Bez (Hasta/Gazlı/Sargı)',
                      'Cezaevi (Kurum)']
        for ci, b in enumerate(basliklar1, 1):
            c = ws1.cell(row=1, column=ci, value=b)
            c.font = bf
            c.fill = bg
            c.alignment = Alignment(horizontal='center')
        for i, k in enumerate(self.son_manuel, 1):
            ws1.cell(row=i+1, column=1, value=i)
            ws1.cell(row=i+1, column=2, value=k.get('Kaynak'))
            ws1.cell(row=i+1, column=3, value=k.get('RxId'))
            ws1.cell(row=i+1, column=4, value=k.get('RxIslemTarihi'))
            ws1.cell(row=i+1, column=5, value=k.get('LoggerTarihi'))
            ws1.cell(row=i+1, column=6, value=k.get('FarkSn'))
            fark = k.get('FarkSn') or 0
            a = abs(int(fark))
            sa = a // 3600
            dk = (a % 3600) // 60
            ksn = a % 60
            sgn = '-' if fark < 0 else ''
            ws1.cell(row=i+1, column=7,
                     value=f"{sgn}{sa:02d}:{dk:02d}:{ksn:02d}")
            ws1.cell(row=i+1, column=8, value=k.get('LoggerMakina'))
            ws1.cell(row=i+1, column=9, value=k.get('LoggerPersonelId'))
            ws1.cell(row=i+1, column=10,
                     value='Evet' if k.get('DonusumMu') else 'Hayır')
            ws1.cell(row=i+1, column=11,
                     value='Evet' if k.get('AntibiyotikMu') else 'Hayır')
            ws1.cell(row=i+1, column=12,
                     value='Evet' if k.get('IsBankasiMu') else 'Hayır')
            ws1.cell(row=i+1, column=13,
                     value='Evet' if k.get('BezMidir') else 'Hayır')
            ws1.cell(row=i+1, column=14,
                     value='Evet' if k.get('CezaeviMidir') else 'Hayır')

        # Sheet 2: Makina özet
        ws2 = wb.create_sheet("PC Makina Özet")
        basliklar2 = ['Makina', 'İşlem Sayısı', 'Ort Fark (sn)',
                      '|Fark| Ort (sn)', 'Min Fark', 'Max Fark',
                      '|Fark|>5sn Sayı']
        for ci, b in enumerate(basliklar2, 1):
            c = ws2.cell(row=1, column=ci, value=b)
            c.font = bf
            c.fill = bg
            c.alignment = Alignment(horizontal='center')
        for i, m in enumerate(self.son_makina, 2):
            ws2.cell(row=i, column=1, value=m.get('LoggerMakina'))
            ws2.cell(row=i, column=2, value=m.get('IslemSayisi'))
            ws2.cell(row=i, column=3, value=round(m.get('FarkOrtalama', 0), 2))
            ws2.cell(row=i, column=4, value=round(m.get('MutlakOrt', 0), 2))
            ws2.cell(row=i, column=5, value=m.get('FarkMin'))
            ws2.cell(row=i, column=6, value=m.get('FarkMax'))
            ws2.cell(row=i, column=7, value=m.get('SuphliSayi'))

        try:
            wb.save(dosya)
            messagebox.showinfo("Kaydedildi", f"Rapor kaydedildi:\n{dosya}")
        except Exception as e:
            messagebox.showerror("Hata", str(e))


class BirlesikZamanAnaliziPopup:
    """Üç zaman analizini tek pencerede gösterir:
    - Kayıt Fark: RxIslemTarihi vs RxKayitTarihi (Zaman Tutarlılığı)
    - Logger Fark: RxIslemTarihi vs Logger.LoggerTarihi (manuel müdahale)
    - Kontrol Fark: RxKontrolTarihi vs RxIslemTarihi (sadece reçete)
    + Dönüşüm/Antibiyotik/İş Bankası filtreleri
    """

    def __init__(self, parent, db):
        self.parent = parent
        self.db = db
        self.son_kayitlar: List[Dict] = []
        self.son_makina: List[Dict] = []
        self._son_sirala_sutun: Optional[str] = None
        self._son_sirala_ters: bool = False
        self._son_sure: float = 0.0

        self.top = tk.Toplevel(parent)
        self.top.title("📊 Birleşik Zaman Analizi")
        self.top.geometry("1500x900")
        try:
            self.top.state('zoomed')
        except Exception:
            pass
        self.top.transient(parent)

        try:
            self._arayuz()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            logger.error(f"BirlesikZamanAnaliziPopup HATA: {e}\n{tb}")
            print(f"[Birleşik Zaman] HATA: {e}\n{tb}")

    def _arayuz(self):
        # Header
        header = tk.Frame(self.top, bg='#6A1B9A', height=46)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header,
                 text="📊 Birleşik Zaman Analizi  —  Kayıt + Logger + Kontrol",
                 font=("Segoe UI", 11, "bold"), bg='#6A1B9A', fg='white',
                 ).pack(side="left", padx=15, pady=12)
        tk.Button(header, text="✕ Kapat", bg='#C62828', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self.top.destroy).pack(side="right", padx=10, pady=8)
        tk.Button(header, text="📥 Excel", bg='#2E7D32', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._excel_aktar).pack(side="right", padx=5, pady=8)
        tk.Button(header, text="🔎 Detay Gör", bg='#FF8F00', fg='white', bd=0,
                  cursor='hand2', font=("Segoe UI", 9), padx=10,
                  command=self._secili_detayi_ac).pack(side="right", padx=5, pady=8)

        # Açıklama
        aciklama = tk.Label(self.top,
            text="ℹ Üç farklı zaman karşılaştırması tek pencerede:  "
                 "🔸 Kayıt Fark (RxIslemTarihi − RxKayitTarihi)  •  "
                 "🔸 Logger Fark (RxIslemTarihi − Logger.LoggerTarihi)  •  "
                 "🔸 Kontrol Fark (RxKontrolTarihi − RxIslemTarihi).  "
                 "En güçlü manuel müdahale kanıtı LOGGER farkı çünkü "
                 "Logger insert-only.",
            font=("Segoe UI", 9, "italic"), fg='#37474F', bg='#F3E5F5',
            anchor='w', justify='left', padx=15, pady=8, wraplength=1450)
        aciklama.pack(fill="x")

        # Parametre paneli
        filt = ttk.LabelFrame(self.top, text="Tarama Parametreleri", padding=6)
        filt.pack(fill="x", padx=5, pady=4)

        # Satır 1: Tarih
        s1 = tk.Frame(filt)
        s1.pack(fill="x", pady=3)
        tk.Label(s1, text="Başlangıç:", font=("Segoe UI", 9, "bold")
                 ).pack(side="left", padx=4)
        self.bas_entry = ttk.Entry(s1, width=14)
        self.bas_entry.insert(0, "2017-05-23")
        self.bas_entry.pack(side="left", padx=4)
        tk.Label(s1, text="→  Bitiş:",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=4)
        self.bit_entry = ttk.Entry(s1, width=14)
        self.bit_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.bit_entry.pack(side="left", padx=4)
        for et, gun in [("Son 90 gün", 90), ("Son 1 yıl", 365),
                         ("Tümü (2017→)", -1)]:
            ttk.Button(s1, text=et, width=14,
                       command=lambda g=gun: self._tarih_preset(g)
                       ).pack(side="left", padx=2)
        self.tara_btn = ttk.Button(s1, text="▶ Taramayı Başlat", width=20,
                                    command=self._taramayi_baslat)
        self.tara_btn.pack(side="right", padx=8)

        # Satır 2: Eşikler + mantık
        s2 = tk.Frame(filt)
        s2.pack(fill="x", pady=3)
        tk.Label(s2, text="🔸 Kayıt eşik (sn):",
                 font=("Segoe UI", 9, "bold"), fg='#6A1B9A'
                 ).pack(side="left", padx=(4, 4))
        self.kayit_esik_entry = ttk.Entry(s2, width=8)
        self.kayit_esik_entry.insert(0, "86400")  # 1 gün (RxKayitTarihi DATE)
        self.kayit_esik_entry.pack(side="left", padx=2)

        tk.Label(s2, text="🔸 Logger eşik (sn):",
                 font=("Segoe UI", 9, "bold"), fg='#00838F'
                 ).pack(side="left", padx=(15, 4))
        self.logger_esik_entry = ttk.Entry(s2, width=8)
        self.logger_esik_entry.insert(0, "60")
        self.logger_esik_entry.pack(side="left", padx=2)

        tk.Label(s2, text="🔸 Kontrol eşik (sn):",
                 font=("Segoe UI", 9, "bold"), fg='#1565C0'
                 ).pack(side="left", padx=(15, 4))
        self.kontrol_esik_entry = ttk.Entry(s2, width=8)
        self.kontrol_esik_entry.insert(0, "3600")  # 1 saat
        self.kontrol_esik_entry.pack(side="left", padx=2)

        tk.Label(s2, text="    Mantık:",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(15, 4))
        self.mantik_var = tk.StringVar(value="OR")
        ttk.Radiobutton(s2, text="En az birinde fark (OR)",
                        variable=self.mantik_var, value="OR"
                        ).pack(side="left", padx=4)
        ttk.Radiobutton(s2, text="Hepsinde fark (AND)",
                        variable=self.mantik_var, value="AND"
                        ).pack(side="left", padx=4)

        # Satır 3: filtre checkbox'ları
        s3 = tk.Frame(filt)
        s3.pack(fill="x", pady=(2, 1))
        self.donusum_haric_var = tk.BooleanVar(value=False)
        self.antibiyotik_haric_var = tk.BooleanVar(value=False)
        self.is_bankasi_haric_var = tk.BooleanVar(value=False)
        self.bez_haric_var = tk.BooleanVar(value=False)
        self.cezaevi_haric_var = tk.BooleanVar(value=False)
        tk.Label(s3, text="Filtreler:",
                 font=("Segoe UI", 9, "bold"), fg='#37474F'
                 ).pack(side="left", padx=(4, 8))
        ttk.Checkbutton(
            s3, text="🔄 Dönüşümleri gizle",
            variable=self.donusum_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s3, text="🦠 Antibiyotik gizle",
            variable=self.antibiyotik_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s3, text="🏦 İş Bankası gizle",
            variable=self.is_bankasi_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s3, text="🩹 Bez gizle",
            variable=self.bez_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)
        ttk.Checkbutton(
            s3, text="🔒 Cezaevi gizle",
            variable=self.cezaevi_haric_var,
            command=self._filtreleri_uygula,
        ).pack(side="left", padx=6)

        # Durum
        durum = tk.Frame(self.top, bg='#F5F5F5', height=28)
        durum.pack(fill="x")
        durum.pack_propagate(False)
        self.durum_lbl = tk.Label(durum,
            text="Hazır. 'Taramayı Başlat'a basın.",
            bg='#F5F5F5', fg='#1565C0', font=("Segoe UI", 9, "italic"),
            anchor='w', padx=10)
        self.durum_lbl.pack(side="left", fill="x", expand=True)
        self.progress = ttk.Progressbar(durum, mode='indeterminate', length=200)
        self.progress.pack(side="right", padx=10, pady=3)

        # Ana tablo
        ust = ttk.LabelFrame(self.top,
            text="📋 BİRLEŞİK KAYIT TABLOSU "
                 "(her satırda 3 fark görünür — eşik üstü vurgulanır)",
            padding=4)
        ust.pack(fill="both", expand=True, padx=5, pady=3)

        cols = ('no', 'kaynak', 'rxid',
                'islem', 'kayit', 'logger', 'kontrol',
                'kayit_fark', 'logger_fark', 'kontrol_fark',
                'makina', 'personel',
                'donusum', 'antibiyotik', 'isbankasi', 'bez', 'cezaevi')
        bash = {'no': '#', 'kaynak': 'Kaynak', 'rxid': 'RxId',
                'islem': 'RxIslemTarihi',
                'kayit': 'RxKayitTarihi',
                'logger': 'LoggerTarihi',
                'kontrol': 'RxKontrolTarihi',
                'kayit_fark': '🔸 Kayıt Fark',
                'logger_fark': '🔸 Logger Fark',
                'kontrol_fark': '🔸 Kontrol Fark',
                'makina': 'Makina', 'personel': 'Pers',
                'donusum': '🔄', 'antibiyotik': '🦠', 'isbankasi': '🏦',
                'bez': '🩹', 'cezaevi': '🔒'}
        widths = {'no': 45, 'kaynak': 65, 'rxid': 70,
                  'islem': 145, 'kayit': 100, 'logger': 145, 'kontrol': 145,
                  'kayit_fark': 110, 'logger_fark': 110, 'kontrol_fark': 110,
                  'makina': 110, 'personel': 50,
                  'donusum': 45, 'antibiyotik': 45, 'isbankasi': 45,
                  'bez': 45, 'cezaevi': 45}
        anchors = {'no': 'center', 'kaynak': 'center', 'rxid': 'center',
                   'personel': 'center',
                   'kayit_fark': 'e', 'logger_fark': 'e', 'kontrol_fark': 'e',
                   'donusum': 'center', 'antibiyotik': 'center',
                   'isbankasi': 'center', 'bez': 'center',
                   'cezaevi': 'center'}
        self.tree = ttk.Treeview(ust, columns=cols, show='headings', height=18)
        self._basliklar_orj = dict(bash)
        for c in cols:
            self.tree.heading(c, text=bash[c],
                               command=lambda col=c: self._sirala(col))
            self.tree.column(c, width=widths[c],
                              anchor=anchors.get(c, 'w'))
        self.tree.pack(side="left", fill="both", expand=True)
        sb1 = ttk.Scrollbar(ust, orient="vertical", command=self.tree.yview)
        sb1.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb1.set)
        self.tree.tag_configure('logger_buyuk', background='#FFCDD2')
        self.tree.tag_configure('kayit_buyuk', background='#FFE0B2')
        self.tree.tag_configure('hem_buyuk', background='#EF9A9A')
        self.tree.tag_configure('normal', background='white')
        self.tree.bind('<Double-1>', self._secili_detayi_ac)

        # Alt panel: PC sapma özeti (Logger'a göre)
        alt = ttk.LabelFrame(self.top,
            text="🖥 PC (LoggerMakina) BAŞINA LOGGER SAPMA ÖZETİ",
            padding=4)
        alt.pack(fill="x", padx=5, pady=3)
        m_cols = ('makina', 'sayi', 'fark_ort', 'mutlak_ort',
                  'fark_min', 'fark_max', 'supheli')
        m_bash = {'makina': 'PC / Makina',
                  'sayi': 'İşlem Sayısı',
                  'fark_ort': 'Ort. Fark (sn)',
                  'mutlak_ort': '|Fark| Ort. (sn)',
                  'fark_min': 'Min Fark', 'fark_max': 'Max Fark',
                  'supheli': '|Fark|>5sn'}
        m_widths = {'makina': 250, 'sayi': 100,
                    'fark_ort': 130, 'mutlak_ort': 150,
                    'fark_min': 100, 'fark_max': 100, 'supheli': 130}
        self.makina_tree = ttk.Treeview(alt, columns=m_cols,
                                          show='headings', height=5)
        for c in m_cols:
            self.makina_tree.heading(c, text=m_bash[c])
            self.makina_tree.column(c, width=m_widths[c],
                                      anchor=('w' if c == 'makina' else 'e'))
        self.makina_tree.pack(side="left", fill="both", expand=True)
        sb2 = ttk.Scrollbar(alt, orient="vertical",
                             command=self.makina_tree.yview)
        sb2.pack(side="right", fill="y")
        self.makina_tree.configure(yscrollcommand=sb2.set)
        self.makina_tree.tag_configure('supheli', background='#FFE0B2')

    def _tarih_preset(self, gun: int):
        bugun = datetime.now()
        if gun == -1:
            bas = datetime(2017, 5, 23)
        else:
            bas = bugun - timedelta(days=gun)
        self.bas_entry.delete(0, tk.END)
        self.bas_entry.insert(0, bas.strftime('%Y-%m-%d'))
        self.bit_entry.delete(0, tk.END)
        self.bit_entry.insert(0, bugun.strftime('%Y-%m-%d'))

    def _tarih_oku(self, widget) -> Optional[datetime]:
        metin = (widget.get() or '').strip()
        for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(metin, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _esik_oku(entry, default: int) -> int:
        try:
            return int((entry.get() or str(default)).strip())
        except ValueError:
            return default

    def _taramayi_baslat(self):
        bas = self._tarih_oku(self.bas_entry)
        bit = self._tarih_oku(self.bit_entry)
        if bas is None or bit is None:
            messagebox.showerror("Tarih hatası",
                "Tarihleri 'YYYY-MM-DD' formatında girin.")
            return
        if bas > bit:
            messagebox.showerror("Tarih hatası",
                "Başlangıç bitişten sonra olamaz.")
            return
        bit = bit.replace(hour=23, minute=59, second=59)
        k_esik = self._esik_oku(self.kayit_esik_entry, 86400)
        l_esik = self._esik_oku(self.logger_esik_entry, 60)
        c_esik = self._esik_oku(self.kontrol_esik_entry, 3600)
        en_az = (self.mantik_var.get() == 'OR')

        self.tara_btn.config(state='disabled')
        self.progress.start(10)
        op = 'OR' if en_az else 'AND'
        self.durum_lbl.config(
            text=f"⏳ Taranıyor: {bas.date()} → {bit.date()} | "
                 f"Kayıt {k_esik}sn {op} Logger {l_esik}sn {op} Kontrol {c_esik}sn",
            fg='#C62828')
        for c in self.tree.get_children():
            self.tree.delete(c)
        for c in self.makina_tree.get_children():
            self.makina_tree.delete(c)

        import time as _time
        t0 = _time.time()

        def _calis():
            try:
                d = self.db.birlesik_zaman_analizi(
                    bas, bit,
                    kayit_esik_sn=k_esik,
                    logger_esik_sn=l_esik,
                    kontrol_esik_sn=c_esik,
                    en_az_birinde=en_az,
                )
                dt = _time.time() - t0
                self.top.after(0, lambda: self._sonuclari_goster(d, dt))
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                logger.error(f"Birleşik zaman hatası: {e}\n{tb}")
                self.top.after(0, lambda: self._hata(str(e)))

        threading.Thread(target=_calis, daemon=True).start()

    def _hata(self, mesaj: str):
        self.progress.stop()
        self.tara_btn.config(state='normal')
        self.durum_lbl.config(text=f"❌ Hata: {mesaj}", fg='#B71C1C')
        messagebox.showerror("Tarama hatası", mesaj)

    def _sonuclari_goster(self, d: Dict, sure: float):
        self.progress.stop()
        self.tara_btn.config(state='normal')
        self.son_kayitlar = d.get('kayitlar') or []
        self.son_makina = d.get('makina_ozeti') or []
        self._son_sure = sure

        # PC sapma özeti
        for m in self.son_makina:
            mutlak = m.get('MutlakOrt', 0)
            tag = 'supheli' if mutlak > 60 else ''
            self.makina_tree.insert('', 'end', values=(
                m.get('LoggerMakina'),
                f"{m.get('IslemSayisi'):,}",
                f"{m.get('FarkOrtalama', 0):+.1f}",
                f"{mutlak:.1f}",
                f"{m.get('FarkMin', 0):+,d}",
                f"{m.get('FarkMax', 0):+,d}",
                f"{m.get('SuphliSayi', 0):,}",
            ), tags=(tag,))

        self._filtreleri_uygula()

    @staticmethod
    def _dt_fmt(v):
        if v is None:
            return ''
        if hasattr(v, 'strftime') and hasattr(v, 'hour'):
            return v.strftime('%Y-%m-%d %H:%M:%S')
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)

    @staticmethod
    def _fark_fmt(v):
        if v is None:
            return '—'
        return f"{v:+,d}"

    def _tabloyu_doldur(self, kayitlar: List[Dict]):
        for c in self.tree.get_children():
            self.tree.delete(c)
        k_esik = self._esik_oku(self.kayit_esik_entry, 86400)
        l_esik = self._esik_oku(self.logger_esik_entry, 60)
        for i, k in enumerate(kayitlar, 1):
            kf = k.get('KayitFark')
            lf = k.get('LoggerFark')
            cf = k.get('KontrolFark')
            kf_buyuk = (kf is not None and abs(kf) > k_esik)
            lf_buyuk = (lf is not None and abs(lf) > l_esik)
            if lf_buyuk and kf_buyuk:
                tag = 'hem_buyuk'
            elif lf_buyuk:
                tag = 'logger_buyuk'
            elif kf_buyuk:
                tag = 'kayit_buyuk'
            else:
                tag = 'normal'
            don_str = '🔄' if k.get('DonusumMu') else ''
            anti_str = '🦠' if k.get('AntibiyotikMu') else ''
            bank_str = '🏦' if k.get('IsBankasiMu') else ''
            bez_str = '🩹' if k.get('BezMidir') else ''
            cez_str = '🔒' if k.get('CezaeviMidir') else ''
            self.tree.insert('', 'end', values=(
                i, k.get('Kaynak'), k.get('RxId'),
                self._dt_fmt(k.get('RxIslemTarihi')),
                self._dt_fmt(k.get('RxKayitTarihi')),
                self._dt_fmt(k.get('LoggerTarihi')),
                self._dt_fmt(k.get('RxKontrolTarihi')),
                self._fark_fmt(kf),
                self._fark_fmt(lf),
                self._fark_fmt(cf),
                k.get('LoggerMakina') or '(?)',
                k.get('LoggerPersonelId') or '',
                don_str, anti_str, bank_str, bez_str, cez_str,
            ), tags=(tag,))

    def _filtreleri_uygula(self):
        don_haric = bool(self.donusum_haric_var.get())
        anti_haric = bool(self.antibiyotik_haric_var.get())
        bank_haric = bool(self.is_bankasi_haric_var.get())
        bez_haric = bool(self.bez_haric_var.get())
        cez_haric = bool(self.cezaevi_haric_var.get())
        gosterilecek = [
            k for k in self.son_kayitlar
            if not (don_haric and k.get('DonusumMu'))
            and not (anti_haric and k.get('AntibiyotikMu'))
            and not (bank_haric and k.get('IsBankasiMu'))
            and not (bez_haric and k.get('BezMidir'))
            and not (cez_haric and k.get('CezaeviMidir'))
        ]
        if self._son_sirala_sutun:
            self._uygula_sirala(gosterilecek,
                                self._son_sirala_sutun,
                                self._son_sirala_ters)
        self._tabloyu_doldur(gosterilecek)

        n_top = len(self.son_kayitlar)
        n_don = sum(1 for k in self.son_kayitlar if k.get('DonusumMu'))
        n_anti = sum(1 for k in self.son_kayitlar if k.get('AntibiyotikMu'))
        n_bank = sum(1 for k in self.son_kayitlar if k.get('IsBankasiMu'))
        n_bez = sum(1 for k in self.son_kayitlar if k.get('BezMidir'))
        n_cez = sum(1 for k in self.son_kayitlar if k.get('CezaeviMidir'))
        n_lf = sum(1 for k in self.son_kayitlar
                   if (k.get('LoggerFark') or 0) and abs(k.get('LoggerFark') or 0) > 60)
        n_aciklanmamis = sum(
            1 for k in self.son_kayitlar
            if not k.get('DonusumMu') and not k.get('AntibiyotikMu')
            and not k.get('IsBankasiMu') and not k.get('BezMidir')
            and not k.get('CezaeviMidir')
        )
        n_gor = len(gosterilecek)
        ozet = (f"✓ ({self._son_sure:.1f}sn) Toplam {n_top} kayıt "
                f"| 🔴 Logger>60sn: {n_lf} "
                f"| 🔄 Dönüşüm: {n_don} | 🦠 Antibiyotik: {n_anti} "
                f"| 🏦 İş Bankası: {n_bank} | 🩹 Bez: {n_bez} "
                f"| 🔒 Cezaevi: {n_cez} "
                f"| 🚨 Açıklanmamış: {n_aciklanmamis}")
        if don_haric or anti_haric or bank_haric or bez_haric or cez_haric:
            ozet += f"  →  GÖSTERİLEN: {n_gor}"
        self.durum_lbl.config(text=ozet, fg='#2E7D32')

    def _uygula_sirala(self, kayitlar: List[Dict], sutun_kod: str, ters: bool):
        def _safe(v, default):
            return v if v is not None else default
        anahtarlar = {
            'no':           lambda r: 0,
            'kaynak':       lambda r: r.get('Kaynak') or '',
            'rxid':         lambda r: _safe(r.get('RxId'), 0),
            'islem':        lambda r: _safe(r.get('RxIslemTarihi'), datetime.min),
            'kayit':        lambda r: _safe(r.get('RxKayitTarihi'), datetime.min),
            'logger':       lambda r: _safe(r.get('LoggerTarihi'), datetime.min),
            'kontrol':      lambda r: _safe(r.get('RxKontrolTarihi'), datetime.min),
            'kayit_fark':   lambda r: abs(r.get('KayitFark') or 0),
            'logger_fark':  lambda r: abs(r.get('LoggerFark') or 0),
            'kontrol_fark': lambda r: abs(r.get('KontrolFark') or 0),
            'makina':       lambda r: r.get('LoggerMakina') or '',
            'personel':     lambda r: _safe(r.get('LoggerPersonelId'), 0),
            'donusum':      lambda r: 1 if r.get('DonusumMu') else 0,
            'antibiyotik':  lambda r: 1 if r.get('AntibiyotikMu') else 0,
            'isbankasi':    lambda r: 1 if r.get('IsBankasiMu') else 0,
            'bez':          lambda r: 1 if r.get('BezMidir') else 0,
            'cezaevi':      lambda r: 1 if r.get('CezaeviMidir') else 0,
        }
        key_fn = anahtarlar.get(sutun_kod, anahtarlar['logger_fark'])
        try:
            kayitlar.sort(key=key_fn, reverse=ters)
        except TypeError:
            kayitlar.sort(key=lambda r: str(key_fn(r)), reverse=ters)

    def _sirala(self, sutun_kod: str):
        if not self.son_kayitlar:
            return
        ters = (self._son_sirala_sutun == sutun_kod) and not self._son_sirala_ters
        self._son_sirala_sutun = sutun_kod
        self._son_sirala_ters = ters
        self._uygula_sirala(self.son_kayitlar, sutun_kod, ters)
        self._filtreleri_uygula()
        ok = ' ↓' if ters else ' ↑'
        for c, orj in self._basliklar_orj.items():
            yeni = orj + (ok if c == sutun_kod else '')
            try:
                self.tree.heading(c, text=yeni)
            except Exception:
                pass

    def _secili_detayi_ac(self, _event=None):
        sec = self.tree.selection()
        if not sec:
            messagebox.showinfo("Seçim yok",
                "Önce tablodan bir satır seçin.")
            return
        vals = self.tree.item(sec[0], 'values')
        if not vals or len(vals) < 3:
            return
        kaynak = (vals[1] or '').strip().upper()
        rx_id = (vals[2] or '').strip()
        if kaynak not in ('RECETE', 'ELDEN') or not rx_id:
            return
        SatisDetayPopup(parent=self.top, db=self.db,
                        kaynak=kaynak, rx_id=rx_id)

    def _excel_aktar(self):
        if not self.son_kayitlar and not self.son_makina:
            messagebox.showinfo("Boş", "Önce taramayı çalıştırın.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("openpyxl yok",
                "Excel export için openpyxl gerekli.")
            return
        from tkinter import filedialog as fd
        dosya = fd.asksaveasfilename(
            title="Birleşik Zaman Analizi raporu kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"birlesik_zaman_"
                        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not dosya:
            return

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Birleşik Kayıtlar"
        bf = Font(bold=True, color='FFFFFF')
        bg = PatternFill(start_color='6A1B9A', end_color='6A1B9A',
                          fill_type='solid')
        basliklar = ['No', 'Kaynak', 'RxId',
                     'RxIslemTarihi', 'RxKayitTarihi',
                     'LoggerTarihi', 'RxKontrolTarihi',
                     'Kayıt Fark (sn)', 'Logger Fark (sn)',
                     'Kontrol Fark (sn)',
                     'Logger Makina', 'Personel',
                     'Dönüşüm', 'Antibiyotik', 'İş Bankası', 'Bez',
                     'Cezaevi']
        for ci, b in enumerate(basliklar, 1):
            c = ws1.cell(row=1, column=ci, value=b)
            c.font = bf
            c.fill = bg
            c.alignment = Alignment(horizontal='center')
        for i, k in enumerate(self.son_kayitlar, 1):
            ws1.cell(row=i+1, column=1, value=i)
            ws1.cell(row=i+1, column=2, value=k.get('Kaynak'))
            ws1.cell(row=i+1, column=3, value=k.get('RxId'))
            ws1.cell(row=i+1, column=4, value=k.get('RxIslemTarihi'))
            ws1.cell(row=i+1, column=5, value=k.get('RxKayitTarihi'))
            ws1.cell(row=i+1, column=6, value=k.get('LoggerTarihi'))
            ws1.cell(row=i+1, column=7, value=k.get('RxKontrolTarihi'))
            ws1.cell(row=i+1, column=8, value=k.get('KayitFark'))
            ws1.cell(row=i+1, column=9, value=k.get('LoggerFark'))
            ws1.cell(row=i+1, column=10, value=k.get('KontrolFark'))
            ws1.cell(row=i+1, column=11, value=k.get('LoggerMakina'))
            ws1.cell(row=i+1, column=12, value=k.get('LoggerPersonelId'))
            ws1.cell(row=i+1, column=13,
                     value='Evet' if k.get('DonusumMu') else 'Hayır')
            ws1.cell(row=i+1, column=14,
                     value='Evet' if k.get('AntibiyotikMu') else 'Hayır')
            ws1.cell(row=i+1, column=15,
                     value='Evet' if k.get('IsBankasiMu') else 'Hayır')
            ws1.cell(row=i+1, column=16,
                     value='Evet' if k.get('BezMidir') else 'Hayır')
            ws1.cell(row=i+1, column=17,
                     value='Evet' if k.get('CezaeviMidir') else 'Hayır')

        ws2 = wb.create_sheet("PC Makina Özet")
        m_bash = ['Makina', 'İşlem Sayısı', 'Ort Fark (sn)',
                  '|Fark| Ort (sn)', 'Min Fark', 'Max Fark', '|Fark|>5sn']
        for ci, b in enumerate(m_bash, 1):
            c = ws2.cell(row=1, column=ci, value=b)
            c.font = bf
            c.fill = bg
            c.alignment = Alignment(horizontal='center')
        for i, m in enumerate(self.son_makina, 2):
            ws2.cell(row=i, column=1, value=m.get('LoggerMakina'))
            ws2.cell(row=i, column=2, value=m.get('IslemSayisi'))
            ws2.cell(row=i, column=3, value=round(m.get('FarkOrtalama', 0), 2))
            ws2.cell(row=i, column=4, value=round(m.get('MutlakOrt', 0), 2))
            ws2.cell(row=i, column=5, value=m.get('FarkMin'))
            ws2.cell(row=i, column=6, value=m.get('FarkMax'))
            ws2.cell(row=i, column=7, value=m.get('SuphliSayi'))

        try:
            wb.save(dosya)
            messagebox.showinfo("Kaydedildi", f"Rapor kaydedildi:\n{dosya}")
        except Exception as e:
            messagebox.showerror("Hata", str(e))


def satis_raporlari_ac(parent=None, ana_menu_callback=None):
    """Modül launcher."""
    if parent is None:
        root = tk.Tk()
    else:
        root = tk.Toplevel(parent)
    SatisRaporlariGUI(root, ana_menu_callback=ana_menu_callback)
    return root


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
    root = tk.Tk()
    SatisRaporlariGUI(root)
    root.mainloop()
