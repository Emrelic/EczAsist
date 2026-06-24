# -*- coding: utf-8 -*-
"""Yabancı Uyruklu Hasta Tespit Raporu — Ek Raporlar modülü.

98/99 ile başlayan TC kimlik (yabancı uyruklu / YKN) hastaların reçetelerini
tarar; her birinin Botanik EOS kapsamını gösterir ve geçici koruma kapsamında
OLMAYANLARI ("Topkapı SGK kağıdı gerekli") işaretler.

İki sekme:
  • Detay (Dönem)  : seçilen tarih aralığındaki tek tek reçeteler.
  • Aylık Özet     : geçmiş aylar — her ay kaç yabancı reçete karşılanmış,
                     kaçı Topkapı kağıdı gerektiriyor / kaçı geçici koruma.

🚨 Botanik EOS'a yalnızca SELECT yapar (BotanikDB güvenlik filtresi üzerinden).

A/B/C grubu notu: A/B/C ayrımı Botanik EOS'ta reçete bazında saklanmaz (SGK
fatura kesimi sırasında belirlenir). Bu rapor güvenilir biçimde kapsam (geçici
koruma vs. yabancı SGK'lı) ayrımını verir; A/B/C tahmini içermez.
"""

import logging
import threading
import tkinter as tk
from datetime import date
from tkinter import messagebox, ttk

import yabanci_hasta_tespit as yht

logger = logging.getLogger(__name__)

_AYLAR = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz",
          "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]


def _ay_once(d: date, n: int) -> date:
    """d tarihinden n ay önceki ayın ilk günü."""
    ay = d.month - 1 - n
    yil = d.year + ay // 12
    return date(yil, ay % 12 + 1, 1)


class YabanciHastaRaporGUI:
    """Yabancı uyruklu hasta tespit raporu penceresi."""

    DETAY_SUTUNLAR = [
        ("tarih", "Tarih", 95),
        ("rec_no", "Reçete No", 130),
        ("hasta", "Hasta", 240),
        ("tc", "TC / YKN", 110),
        ("kapsam", "Kapsam (EOS)", 200),
        ("durum", "Durum", 360),
    ]
    AYLIK_SUTUNLAR = [
        ("donem", "Dönem", 160),
        ("toplam", "Toplam yabancı reçete", 180),
        ("topkapi", "🔴 Topkapı SGK kağıdı gerekli", 240),
        ("gk", "🟢 Geçici koruma", 160),
    ]

    def __init__(self, parent: tk.Toplevel):
        self.parent = parent
        self.parent.title("Yabancı Uyruklu Hasta Tespit Raporu")
        self.parent.geometry("1300x740")
        self.parent.configure(bg="#ECEFF1")

        self.db = None
        self.detay_veriler = []   # Detay sekmesi son sorgu
        self.aylik_veriler = []   # Aylık özet son sorgu (dict listesi)

        self._arayuz_olustur()
        self._baglanti_kur()

    # ── Arayüz ──────────────────────────────────────────────────────────
    def _arayuz_olustur(self):
        tk.Label(self.parent, text="🌍  Yabancı Uyruklu Hasta Tespit Raporu",
                 bg="#37474F", fg="white", font=("Segoe UI", 15, "bold"),
                 pady=12).pack(fill="x")

        self.nb = ttk.Notebook(self.parent)
        self.nb.pack(fill="both", expand=True, padx=8, pady=6)

        self._detay_sekmesi_olustur()
        self._aylik_sekmesi_olustur()

        self.status = tk.Label(self.parent, text="Hazır", bg="#CFD8DC",
                               fg="#263238", anchor="w", padx=8)
        self.status.pack(fill="x", side="bottom")

    def _detay_sekmesi_olustur(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="  Detay (Dönem)  ")

        flt = ttk.LabelFrame(tab, text="Dönem (reçete kayıt tarihi)", padding=8)
        flt.pack(fill="x", padx=6, pady=(6, 4))

        bugun = date.today()
        ttk.Label(flt, text="Başlangıç (YYYY-AA-GG):").grid(
            row=0, column=0, padx=4, sticky="w")
        self.e_bas = ttk.Entry(flt, width=14)
        self.e_bas.insert(0, bugun.replace(day=1).isoformat())
        self.e_bas.grid(row=0, column=1, padx=4)
        ttk.Label(flt, text="Bitiş (YYYY-AA-GG):").grid(
            row=0, column=2, padx=4, sticky="w")
        self.e_bit = ttk.Entry(flt, width=14)
        self.e_bit.insert(0, bugun.isoformat())
        self.e_bit.grid(row=0, column=3, padx=4)
        self.btn_sorgu = ttk.Button(flt, text="🔍 Sorgula",
                                    command=self._detay_sorgula)
        self.btn_sorgu.grid(row=0, column=4, padx=12)
        self.btn_excel = ttk.Button(flt, text="📊 Excel'e Aktar",
                                    command=self._detay_excele_aktar,
                                    state="disabled")
        self.btn_excel.grid(row=0, column=5, padx=4)

        tcer = ttk.Frame(tab)
        tcer.pack(fill="both", expand=True, padx=6, pady=4)
        cols = [c[0] for c in self.DETAY_SUTUNLAR]
        self.tree = ttk.Treeview(tcer, columns=cols, show="headings")
        for key, baslik, gen in self.DETAY_SUTUNLAR:
            self.tree.heading(key, text=baslik)
            self.tree.column(key, width=gen, anchor="w")
        self.tree.tag_configure("uyari", background="#FFCDD2")
        self.tree.tag_configure("gk", background="#C8E6C9")
        sb = ttk.Scrollbar(tcer, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.ozet = tk.Label(tab, text="", bg="#ECEFF1", fg="#263238",
                             font=("Segoe UI", 10), justify="left", anchor="w")
        self.ozet.pack(fill="x", padx=6, pady=(2, 4))

    def _aylik_sekmesi_olustur(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="  Aylık Özet (Geçmiş Aylar)  ")

        flt = ttk.LabelFrame(tab, text="Aralık (reçete kayıt tarihi)", padding=8)
        flt.pack(fill="x", padx=6, pady=(6, 4))
        bugun = date.today()
        ttk.Label(flt, text="Başlangıç (YYYY-AA-GG):").grid(
            row=0, column=0, padx=4, sticky="w")
        self.e_abas = ttk.Entry(flt, width=14)
        self.e_abas.insert(0, _ay_once(bugun, 11).isoformat())  # son 12 ay
        self.e_abas.grid(row=0, column=1, padx=4)
        ttk.Label(flt, text="Bitiş (YYYY-AA-GG):").grid(
            row=0, column=2, padx=4, sticky="w")
        self.e_abit = ttk.Entry(flt, width=14)
        self.e_abit.insert(0, bugun.isoformat())
        self.e_abit.grid(row=0, column=3, padx=4)
        self.btn_asorgu = ttk.Button(flt, text="🔍 Aylık Sorgula",
                                     command=self._aylik_sorgula)
        self.btn_asorgu.grid(row=0, column=4, padx=12)
        self.btn_aexcel = ttk.Button(flt, text="📊 Excel'e Aktar",
                                     command=self._aylik_excele_aktar,
                                     state="disabled")
        self.btn_aexcel.grid(row=0, column=5, padx=4)

        tcer = ttk.Frame(tab)
        tcer.pack(fill="both", expand=True, padx=6, pady=4)
        cols = [c[0] for c in self.AYLIK_SUTUNLAR]
        self.atree = ttk.Treeview(tcer, columns=cols, show="headings")
        for key, baslik, gen in self.AYLIK_SUTUNLAR:
            self.atree.heading(key, text=baslik)
            anchor = "w" if key == "donem" else "center"
            self.atree.column(key, width=gen, anchor=anchor)
        self.atree.tag_configure("varsa", background="#FFCDD2")  # Topkapı>0
        self.atree.tag_configure("toplam", background="#CFD8DC",
                                 font=("Segoe UI", 10, "bold"))
        sb = ttk.Scrollbar(tcer, orient="vertical", command=self.atree.yview)
        self.atree.configure(yscrollcommand=sb.set)
        self.atree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.aozet = tk.Label(
            tab,
            text=("Bu sekme geçmiş ayları gösterir: her ay karşılanan "
                  "(silinmemiş) yabancı uyruklu reçetelerin kaçı Topkapı SGK "
                  "kağıdı gerektiriyor. 🔴 satırlar = o ay Topkapı gerekli "
                  "reçete karşılanmış."),
            bg="#ECEFF1", fg="#263238", font=("Segoe UI", 9),
            justify="left", anchor="w", wraplength=1200)
        self.aozet.pack(fill="x", padx=6, pady=(2, 4))

    def _durum(self, msg: str):
        try:
            self.status.config(text=msg)
        except Exception:
            pass

    # ── Veritabanı ──────────────────────────────────────────────────────
    def _baglanti_kur(self):
        try:
            from botanik_db import BotanikDB
            self.db = BotanikDB()
            if self.db.baglan():
                self._durum("Botanik EOS'a bağlanıldı (salt-okuma).")
            else:
                self._durum("Bağlantı hatası!")
                self.db = None
        except Exception as e:
            logger.error("DB bağlantı hatası: %s", e)
            self._durum(f"Hata: {e}")
            self.db = None

    def _tarih_dogrula(self, bas, bit) -> bool:
        try:
            date.fromisoformat(bas)
            date.fromisoformat(bit)
            return True
        except ValueError:
            messagebox.showerror("Tarih", "Tarihler YYYY-AA-GG biçiminde olmalı.",
                                 parent=self.parent)
            return False

    # ── Detay sekmesi ───────────────────────────────────────────────────
    def _detay_sorgula(self):
        if not self.db:
            messagebox.showwarning("Bağlantı", "Botanik EOS bağlantısı yok.",
                                   parent=self.parent)
            return
        bas, bit = self.e_bas.get().strip(), self.e_bit.get().strip()
        if not self._tarih_dogrula(bas, bit):
            return
        self.btn_sorgu.config(state="disabled")
        self._durum("Sorgulanıyor…")
        threading.Thread(target=self._detay_calistir, args=(bas, bit),
                         daemon=True).start()

    def _detay_calistir(self, bas, bit):
        # ReceteAna = reçete bazlı. MusteriTCKN '98%'/'99%' = yabancı. SADECE SELECT.
        sql = """
            SELECT
                CAST(ra.RxKayitTarihi AS date) AS Tarih,
                ra.RxEReceteNo                 AS ReceteNo,
                m.MusteriTCKN                  AS TC,
                m.MusteriAdiSoyadi             AS Hasta,
                k.KapsamAdi                    AS Kapsam
            FROM ReceteAna ra
            LEFT JOIN Musteri m ON ra.RxMusteriId = m.MusteriId
            LEFT JOIN Kapsam  k ON m.MusteriKapsamId = k.KapsamId
            WHERE ra.RxSilme = 0
              AND (m.MusteriTCKN LIKE '98%' OR m.MusteriTCKN LIKE '99%')
              AND ra.RxKayitTarihi >= ?
              AND ra.RxKayitTarihi < DATEADD(day, 1, CAST(? AS date))
            ORDER BY ra.RxKayitTarihi DESC, m.MusteriAdiSoyadi
        """
        try:
            rows = self.db.sorgu_calistir(sql, (bas, bit))
        except Exception as e:
            logger.error("Detay sorgu hatası: %s", e)
            self.parent.after(0, lambda: self._detay_bitti(None, str(e)))
            return
        self.parent.after(0, lambda: self._detay_bitti(rows or [], None))

    def _detay_bitti(self, rows, hata):
        self.btn_sorgu.config(state="normal")
        if hata is not None:
            self._durum(f"Hata: {hata}")
            messagebox.showerror("Sorgu Hatası", hata, parent=self.parent)
            return
        self.detay_veriler = rows
        self.tree.delete(*self.tree.get_children())
        topkapi = gk = 0
        kapsam_set = {}
        for r in rows:
            tc = (r.get("TC") or "").strip()
            kapsam = (r.get("Kapsam") or "").strip()
            uyari = yht.topkapi_kagit_uyarisi_gerekir_mi(tc, kapsam)
            if uyari:
                topkapi += 1
                tag = "uyari"
            else:
                gk += 1
                tag = "gk"
            kapsam_set[kapsam or "(boş)"] = kapsam_set.get(kapsam or "(boş)", 0) + 1
            tar = r.get("Tarih")
            self.tree.insert("", "end", tags=(tag,), values=(
                tar.isoformat() if hasattr(tar, "isoformat") else (tar or ""),
                r.get("ReceteNo") or "", r.get("Hasta") or "", tc,
                kapsam or "(boş)", yht.durum_etiketi(tc, kapsam),
            ))
        kapsam_ozet = ", ".join(
            f"{ad} ({adet})" for ad, adet in sorted(
                kapsam_set.items(), key=lambda x: -x[1]))
        self.ozet.config(text=(
            f"Toplam yabancı uyruklu reçete: {len(rows)}    |    "
            f"🔴 Topkapı SGK kağıdı gerekli: {topkapi}    |    "
            f"🟢 Geçici koruma: {gk}\n"
            f"Kapsam dağılımı: {kapsam_ozet or '—'}\n"
            f"Not: A/B/C grubu Botanik EOS'ta reçete bazında saklanmaz "
            f"(SGK fatura kesiminde belirlenir); bu rapor kapsam ayrımını verir."
        ))
        self.btn_excel.config(state="normal" if rows else "disabled")
        self._durum(f"{len(rows)} reçete listelendi.")

    # ── Aylık özet sekmesi ──────────────────────────────────────────────
    def _aylik_sorgula(self):
        if not self.db:
            messagebox.showwarning("Bağlantı", "Botanik EOS bağlantısı yok.",
                                   parent=self.parent)
            return
        bas, bit = self.e_abas.get().strip(), self.e_abit.get().strip()
        if not self._tarih_dogrula(bas, bit):
            return
        self.btn_asorgu.config(state="disabled")
        self._durum("Aylık özet sorgulanıyor…")
        threading.Thread(target=self._aylik_calistir, args=(bas, bit),
                         daemon=True).start()

    def _aylik_calistir(self, bas, bit):
        # Yıl/ay + kapsam bazında say; Topkapı vs GK ayrımı kapsam adından
        # (Python tarafında, tüm satırlar zaten 98/99 = yabancı).
        sql = """
            SELECT YEAR(ra.RxKayitTarihi)  AS Yil,
                   MONTH(ra.RxKayitTarihi) AS Ay,
                   k.KapsamAdi             AS Kapsam,
                   COUNT(*)                AS Adet
            FROM ReceteAna ra
            LEFT JOIN Musteri m ON ra.RxMusteriId = m.MusteriId
            LEFT JOIN Kapsam  k ON m.MusteriKapsamId = k.KapsamId
            WHERE ra.RxSilme = 0
              AND (m.MusteriTCKN LIKE '98%' OR m.MusteriTCKN LIKE '99%')
              AND ra.RxKayitTarihi >= ?
              AND ra.RxKayitTarihi < DATEADD(day, 1, CAST(? AS date))
            GROUP BY YEAR(ra.RxKayitTarihi), MONTH(ra.RxKayitTarihi), k.KapsamAdi
            ORDER BY YEAR(ra.RxKayitTarihi) DESC, MONTH(ra.RxKayitTarihi) DESC
        """
        try:
            rows = self.db.sorgu_calistir(sql, (bas, bit))
        except Exception as e:
            logger.error("Aylık sorgu hatası: %s", e)
            self.parent.after(0, lambda: self._aylik_bitti(None, str(e)))
            return
        self.parent.after(0, lambda: self._aylik_bitti(rows or [], None))

    def _aylik_bitti(self, rows, hata):
        self.btn_asorgu.config(state="normal")
        if hata is not None:
            self._durum(f"Hata: {hata}")
            messagebox.showerror("Sorgu Hatası", hata, parent=self.parent)
            return
        # (yil, ay) -> {toplam, topkapi, gk}
        aylar = {}
        for r in rows:
            yil = int(r.get("Yil") or 0)
            ay = int(r.get("Ay") or 0)
            adet = int(r.get("Adet") or 0)
            kapsam = (r.get("Kapsam") or "").strip()
            d = aylar.setdefault((yil, ay), {"toplam": 0, "topkapi": 0, "gk": 0})
            d["toplam"] += adet
            if yht.kapsam_gecici_koruma_mi(kapsam):
                d["gk"] += adet
            else:
                d["topkapi"] += adet

        self.aylik_veriler = []
        self.atree.delete(*self.atree.get_children())
        t_top = t_kapi = t_gk = 0
        for (yil, ay) in sorted(aylar.keys(), reverse=True):
            d = aylar[(yil, ay)]
            donem = f"{yil} {_AYLAR[ay] if 1 <= ay <= 12 else ay}"
            tag = "varsa" if d["topkapi"] > 0 else ""
            self.atree.insert("", "end", tags=(tag,), values=(
                donem, d["toplam"], d["topkapi"], d["gk"]))
            self.aylik_veriler.append(
                {"donem": donem, **d})
            t_top += d["toplam"]
            t_kapi += d["topkapi"]
            t_gk += d["gk"]
        if aylar:
            self.atree.insert("", "end", tags=("toplam",), values=(
                "TOPLAM", t_top, t_kapi, t_gk))
        self.btn_aexcel.config(state="normal" if aylar else "disabled")
        self._durum(
            f"{len(aylar)} ay listelendi — toplam {t_top} yabancı reçete, "
            f"{t_kapi} Topkapı kağıdı gerekli.")

    # ── Excel ───────────────────────────────────────────────────────────
    def _excel_kaydet(self, basliklar, satirlar, renk_fn, dosya_adi):
        try:
            import os
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Excel", "openpyxl yüklü değil.",
                                 parent=self.parent)
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(basliklar)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="37474F")
            for deger_satiri, renk in satirlar:
                ws.append(deger_satiri)
                if renk:
                    fill = PatternFill("solid", fgColor=renk)
                    for c in ws[ws.max_row]:
                        c.fill = fill
            for i in range(1, len(basliklar) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 24
            klasor = os.path.join(os.path.expanduser("~"), "Desktop",
                                  "Reçete Kontrol")
            os.makedirs(klasor, exist_ok=True)
            yol = os.path.join(klasor, dosya_adi)
            wb.save(yol)
            messagebox.showinfo("Excel", f"Rapor kaydedildi:\n{yol}",
                                parent=self.parent)
            self._durum(f"Excel: {yol}")
        except Exception as e:
            logger.error("Excel aktarım hatası: %s", e)
            messagebox.showerror("Excel", f"Aktarılamadı:\n{e}",
                                 parent=self.parent)

    def _detay_excele_aktar(self):
        if not self.detay_veriler:
            return
        satirlar = []
        for r in self.detay_veriler:
            tc = (r.get("TC") or "").strip()
            kapsam = (r.get("Kapsam") or "").strip()
            uyari = yht.topkapi_kagit_uyarisi_gerekir_mi(tc, kapsam)
            tar = r.get("Tarih")
            satirlar.append(([
                tar.isoformat() if hasattr(tar, "isoformat") else (tar or ""),
                r.get("ReceteNo") or "", r.get("Hasta") or "", tc,
                kapsam or "(boş)", yht.durum_etiketi(tc, kapsam),
            ], "FFCDD2" if uyari else "C8E6C9"))
        self._excel_kaydet(
            ["Tarih", "Reçete No", "Hasta", "TC / YKN", "Kapsam (EOS)", "Durum"],
            satirlar,
            None,
            f"yabanci_uyruklu_detay_{self.e_bas.get().strip()}_"
            f"{self.e_bit.get().strip()}.xlsx")

    def _aylik_excele_aktar(self):
        if not self.aylik_veriler:
            return
        satirlar = [([d["donem"], d["toplam"], d["topkapi"], d["gk"]],
                     "FFCDD2" if d["topkapi"] > 0 else None)
                    for d in self.aylik_veriler]
        self._excel_kaydet(
            ["Dönem", "Toplam yabancı reçete",
             "Topkapı SGK kağıdı gerekli", "Geçici koruma"],
            satirlar, None,
            f"yabanci_uyruklu_aylik_{self.e_abas.get().strip()}_"
            f"{self.e_abit.get().strip()}.xlsx")
