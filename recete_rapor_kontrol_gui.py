# -*- coding: utf-8 -*-
"""
Reçete / Rapor Kontrol Modülü GUI

Ana menüden açılan, grup bazlı reçete kontrol ekranı.
Her grup için kaldığı yerden devam edebilir.
Medula bağlantısı kurarak reçeteleri kontrol eder.

MEDULA ELEMENT ID'LERİ (Öğrenildi 2026-03-24):
================================================
Giriş Penceresi (BotanikEOS 2.1.223.0 (T)):
  - Kullanıcı combo: auto_id="cmbKullanicilar"
  - Şifre: auto_id="txtSifre"
  - Giriş butonu: auto_id="btnGirisYap"

Sol Menü:
  - e-Reçete Sorgu: auto_id="form1:menuHtmlCommandExButton11"
  - Reçete Giriş: auto_id="form1:menuHtmlCommandExButton21"
  - Reçete Listesi: auto_id="form1:menuHtmlCommandExButton31"
  - Reçete Listesi (Günlük): auto_id="form1:menuHtmlCommandExButton41"
  - Reçete Sorgu: auto_id="form1:menuHtmlCommandExButton51"
  - İade Reçete: auto_id="form1:menuHtmlCommandExButton61"
  - İlaç Bilgisi: auto_id="form1:menuHtmlCommandExButton71"
  - Fatura Sonlandırma: auto_id="form1:menuHtmlCommandExButton121"

Reçete Listesi Sayfası:
  - Fatura Türü dropdown: auto_id="form1:menu1"
  - Dönem dropdown: auto_id="form1:menu2"
  - Sorgula butonu: auto_id="form1:buttonSonlandirilmamisReceteler"
  - Grup sekmeleri (DataItem, title ile erişilir):
      "A", "B", "C Sıralı", "C Kan", "GKKKOY", "Yurtdışı"
"""

import tkinter as tk
from tkinter import ttk, messagebox
import json
import logging
import os
import re
import subprocess
import time
import threading
from datetime import datetime

# pywinauto/UIA backend'in arka plan thread'lerinden çağrılması COM apartment
# initialize edilmediğinde access violation/0x80040155 veriyor. Her arka plan
# thread'inde bu helper kullan: with _com_thread(): ... şeklinde.
try:
    import pythoncom
    _PYTHONCOM_VAR = True
except ImportError:
    _PYTHONCOM_VAR = False

try:
    import win32process as _w32proc
    import win32api as _w32api
except ImportError:
    _w32proc = None
    _w32api = None


def _hwnd_process_adi(hwnd):
    """HWND → process exe adı (lowercase, .exe siz). Hata olursa boş string."""
    if _w32proc is None or _w32api is None:
        return ""
    try:
        _tid, pid = _w32proc.GetWindowThreadProcessId(int(hwnd))
        if not pid:
            return ""
        h = _w32api.OpenProcess(0x1000, False, pid)  # PROCESS_QUERY_LIMITED_INFORMATION
        try:
            exe = _w32proc.GetModuleFileNameEx(h, 0)
        finally:
            _w32api.CloseHandle(h)
        return os.path.splitext(os.path.basename(exe or ""))[0].lower()
    except Exception:
        return ""


class _com_thread:
    """COM apartment'ı thread için initialize/uninitialize eden context manager."""
    def __enter__(self):
        if _PYTHONCOM_VAR:
            try:
                pythoncom.CoInitialize()
            except Exception:
                pass
        return self
    def __exit__(self, *exc):
        if _PYTHONCOM_VAR:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return False

logger = logging.getLogger(__name__)

# Proje dizini
PROJE_DIZINI = os.path.dirname(os.path.abspath(__file__))
DURUM_DOSYASI = os.path.join(PROJE_DIZINI, "recete_kontrol_durumlari.json")
RENKLI_CACHE_DOSYASI = os.path.join(PROJE_DIZINI, "renkli_recete_cache.json")
MESAJ_YOKSAY_DOSYASI = os.path.join(PROJE_DIZINI, "mesaj_yoksay_listesi.json")

# Medula exe yolu
MEDULA_EXE = r"C:\BotanikEczane\BotanikMedula.exe"

# Medula giriş bilgileri
MEDULA_KULLANICI = "16-botan "
MEDULA_SIFRE = "152634"

# Grup tanımları - sıralı
# kod: bizim kodumuzu, medula_tab: Medula'daki sekme adı
# TÜMÜNÜ KONTROL ET sıralaması: C → A → B → GK → CK (Geçici Koruma kan üründen önce)
GRUP_TANIMLARI = [
    {"kod": "C", "ad": "C Grubu", "medula_tab": "C Sıralı", "renk": "#4CAF50", "hover": "#388E3C"},
    {"kod": "A", "ad": "A Grubu", "medula_tab": "A", "renk": "#2196F3", "hover": "#1976D2"},
    {"kod": "B", "ad": "B Grubu", "medula_tab": "B", "renk": "#FF9800", "hover": "#F57C00"},
    {"kod": "GK", "ad": "Geçici Koruma", "medula_tab": "GKKKOY", "renk": "#9C27B0", "hover": "#7B1FA2"},
    {"kod": "CK", "ad": "C Grubu Kan Ürünü", "medula_tab": "C Kan", "renk": "#F44336", "hover": "#D32F2F"},
]

# Medula AutomationID sabitleri
MEDULA_IDS = {
    # Giriş penceresi
    "giris_kullanici_combo": "cmbKullanicilar",
    "giris_sifre": "txtSifre",
    "giris_butonu": "btnGirisYap",
    # Sol menü
    "menu_recete_listesi": "form1:menuHtmlCommandExButton31",
    "menu_erecete_sorgu": "form1:menuHtmlCommandExButton11",
    "menu_recete_giris": "form1:menuHtmlCommandExButton21",
    "menu_recete_sorgu": "form1:menuHtmlCommandExButton51",
    "menu_ilac_bilgisi": "form1:menuHtmlCommandExButton71",
    # Reçete Listesi sayfası
    "fatura_turu_combo": "form1:menu1",
    "donem_combo": "form1:menu2",
    "sorgula_butonu": "form1:buttonSonlandirilmamisReceteler",
}


class MedulaBaglanti:
    """
    Medula (BotanikEOS) pencere bağlantısı ve kontrolü.

    Bağlantı akışı:
    1. MEDULA penceresi açık mı? → Direkt bağlan
    2. Giriş penceresi açık mı? → Kullanıcı seç, şifre gir, giriş butonuna birkaç kez bas
    3. Hiçbiri yok? → Exe'den başlat → Giriş yap
    4. Giriş başarısız? → taskkill ile kapat → Exe'den tekrar başlat → Giriş yap
    """

    MAX_GIRIS_DENEME = 4       # Giriş butonuna kaç kez basılacak
    GIRIS_BEKLEME = 2          # Her giriş denemesi arası bekleme (sn)
    EXE_ACILMA_BEKLEME = 5     # Exe başlatıldıktan sonra minimum bekleme (sn)
    KEEPALIVE_ARALIK = 30      # Oturum canlı tutma aralığı (saniye)

    def __init__(self, log_callback=None):
        self.main_window = None
        self.medula_hwnd = None
        self.bagli = False
        self.log = log_callback or (lambda msg, tag="info": logger.info(msg))
        self._keepalive_thread = None
        self._keepalive_aktif = False
        self._son_aktivite = time.time()
        # Bağlantı/giriş işlemlerini serileştirmek için lock
        # (otomatik_baglan + tumu_thread vb. aynı anda çalışırsa COM çakışır)
        self._baglan_lock = threading.Lock()

    # === PENCERE TARAMA ===

    def _pencereleri_tara(self):
        """
        Tüm pencereleri tara ve Medula ile ilgili olanları döndür.
        Returns:
            dict: {"medula": [(hwnd, title), ...], "giris": [(hwnd, title), ...]}

        Tespit önceliği: process adı BotanikMedula.exe → MEDULA başlık → 3+ kelime fallback.
        """
        import win32gui
        medula = []
        giris = []

        def callback(hwnd, _):
            if not win32gui.IsWindowVisible(hwnd):
                return
            t = win32gui.GetWindowText(hwnd)
            if not t:
                return
            proc = _hwnd_process_adi(hwnd)
            if proc == "botanikmedula":
                medula.append((hwnd, t))
                return
            if "MEDULA" in t and "(T)" in t:
                medula.append((hwnd, t))
                return
            if "BotanikEOS" in t and "(T)" in t:
                parcalar = t.replace("(T)", "").strip().split()
                if len(parcalar) >= 3:
                    medula.append((hwnd, t))

        win32gui.EnumWindows(callback, None)
        return {"medula": medula, "giris": giris}

    def medula_acik_mi(self):
        """Medula durumunu kontrol et. Eski arayüz uyumluluğu için."""
        try:
            sonuc = self._pencereleri_tara()
            if sonuc["medula"]:
                return "bagli", sonuc["medula"]
            elif sonuc["giris"]:
                return "giris_bekliyor", sonuc["giris"]
            else:
                return "kapali", []
        except Exception as e:
            logger.error(f"Medula kontrol hatası: {e}")
            return "hata", []

    # === BAĞLANTI ===

    # BotanikEOS toolbar buton ID'leri (WinForms - Medula web elementlerinden farklı!)
    EOS_GIRIS_BTN = "btnMedulayaGirisYap"
    EOS_SONRA_BTN = "btnSonraki"
    EOS_ONCE_BTN = "btnOnceki"

    def _eos_penceresi_bul(self):
        """Medula çalışan ana pencereyi bul.
        Öncelik:
          1) Process adı BotanikMedula olan görünür pencere (en güvenilir;
             oturum düşse, MEDULA prefix gitse bile process aynı kalır)
          2) Title'da "MEDULA" olan pencere (oturum aktifken)
          3) BotanikEOS ana penceresi (kullanıcı adı olan — eski heuristic)
        Küçük şifre penceresi ("BotanikEOS 2.1.223.0 (T)" — kullanıcı adı yok) ve
        görünmez (hayalet) pencereler atlanır.
        Returns: (desktop, window, hwnd) veya (None, None, None)
        """
        from pywinauto import Desktop
        desktop = Desktop(backend="uia")

        # Pencereleri tek seferde topla — birden fazla yerde lazım
        adaylar = []
        for w in desktop.windows():
            try:
                title = w.window_text() or ""
                if not title:
                    continue
                handle = w.handle
                visible = False
                try:
                    visible = desktop.window(handle=handle).is_visible()
                except Exception:
                    visible = False
                if not visible:
                    continue
                proc = _hwnd_process_adi(handle)
                adaylar.append((handle, title, proc))
            except Exception:
                continue

        # 1. Process adı BotanikMedula
        for handle, title, proc in adaylar:
            if proc == "botanikmedula":
                try:
                    return desktop, desktop.window(handle=handle), handle
                except Exception:
                    continue

        # 2. MEDULA başlığı
        for handle, title, proc in adaylar:
            if "MEDULA" in title and "(T)" in title:
                try:
                    return desktop, desktop.window(handle=handle), handle
                except Exception:
                    continue

        # 3. Eski heuristic (3+ kelime) — fallback
        for handle, title, proc in adaylar:
            if "BotanikEOS" not in title or "(T)" not in title:
                continue
            parcalar = title.replace("(T)", "").strip().split()
            if len(parcalar) >= 3:
                try:
                    return desktop, desktop.window(handle=handle), handle
                except Exception:
                    continue

        return None, None, None

    def _eos_buton_bas(self, auto_id):
        """BotanikEOS WinForms butonuna bas (descendants ile güvenilir arama)."""
        if not self.main_window:
            return False
        for elem in self.main_window.descendants():
            try:
                if elem.element_info.automation_id == auto_id:
                    elem.click_input()
                    return True
            except:
                pass
        return False

    def _oturum_canli_mi(self):
        """Medula oturumu aktif mi kontrol et (tıklama yapmadan).
        Web elementleri (menü, reçete sayfası) veya giriş butonu varsa pencere canlı.
        İkisi de yoksa embedded browser boş = oturum düşmüş.

        OPT/STABILITY: descendants() yerine hedefli child_window araması kullanılıyor.
        IE-embed pencerelerde tüm UIA ağacını yürümek E_INTERFACE_NOT_REGISTERED
        hatası tetikleyip pywinauto'yu çökertebiliyor (gözlenmiş crash).
        """
        if not self.main_window:
            return False
        aids = (
            "f:tbl1", "f:buttonSonraki", "f:buttonGeriDon",
            MEDULA_IDS["menu_recete_listesi"], "btnMedulayaGirisYap",
        )
        for aid in aids:
            try:
                cw = self.main_window.child_window(auto_id=aid)
                if cw.exists(timeout=0.2):
                    return True
            except Exception:
                continue
        return False

    def _oturumu_yenile(self):
        """Oturum düşmüşse giriş butonuna basarak yenile.
        Menü yoksa önce geri butonuna bas, sonra giriş."""
        self.log("Oturum düşmüş, yenileniyor...", "warning")

        # Menü görünmüyor olabilir — Geri Dön + Giriş dene
        if not self.oturum_aktif_mi():
            # Geri Dön butonu (web sayfası içi) — f:buttonGeriDon
            # STABILITY: descendants() yerine hedefli child_window araması
            for aid in ("f:buttonGeriDon", "form1:buttonGeriDon"):
                try:
                    cw = self.main_window.child_window(auto_id=aid)
                    if cw.exists(timeout=0.3):
                        cw.invoke()
                        self.log("Geri Dön basıldı", "info")
                        time.sleep(1)
                        break
                except Exception:
                    continue

        # Giriş butonuna 3 kez bas (1sn arayla)
        for i in range(3):
            self._eos_buton_bas(self.EOS_GIRIS_BTN)
            self.log(f"Giriş butonu basıldı ({i+1}/3)", "info")
            time.sleep(3)
            if self.oturum_aktif_mi():
                self.log("Oturum yenilendi!", "success")
                return True

        self.log("Oturum yenilenemedi", "error")
        return False

    def medula_baglan(self):
        """Tek noktadan Medula bağlantı akışı.

        Adımlar:
          1. Zaten bağlı + oturum aktif → True dön
          2. Medula penceresi var mı?
             a) VAR + oturum aktif → True dön
             b) VAR + oturum düşmüş → Giriş butonuna 3 kez bas (1 sn arayla)
                - Oturum kalktıysa → True dön
                - Hâlâ aktif değilse → taskkill + exe yeniden başlat + kullanıcı/şifre ile giriş
             c) YOK + sadece şifre penceresi açık → şifre ile giriş
             d) YOK + hiçbir pencere yok → exe başlat + kullanıcı/şifre ile giriş

        Lock: aynı anda iki thread çağırırsa ikincisi birincinin bitmesini bekler.

        Returns: True bağlanıldı / False bağlanılamadı
        """
        with self._baglan_lock:
            # 1) Zaten bağlı ve oturum sağlam mı?
            if self.bagli and self.main_window is not None:
                try:
                    if self._oturum_canli_mi():
                        return True
                except Exception:
                    pass

            try:
                from pywinauto import Desktop
                desktop = Desktop(backend="uia")

                # 2) Medula ana penceresi var mı?
                _desktop, win, hwnd = self._eos_penceresi_bul()

                if win:
                    self.main_window = win
                    self.medula_hwnd = hwnd
                    self.bagli = True

                    # 2a) Oturum aktif mi?
                    if self._oturum_canli_mi():
                        self.log("Medula penceresi bulundu, mevcut pencere kullanılıyor", "success")
                        return True

                    # 2b) Oturum düşmüş — Giriş butonuna 3 kez basarak yenilemeyi dene
                    self.log("Medula penceresi var ama oturum düşmüş, Giriş butonuna basılıyor...", "warning")
                    if self._oturumu_yenile():
                        return True

                    # 3 deneme yetmedi → taskkill + exe yeniden başlat + giriş
                    self.log("Giriş butonu 3 denemede başarılı olmadı, taskkill yapılıyor...", "warning")
                    self._medula_kapat()
                    time.sleep(3)
                    return self._exe_baslat_ve_giris_yap()

                # 2c) Sadece şifre penceresi açık mı?
                for w in desktop.windows():
                    try:
                        if w.element_info.automation_id == "SifreSorForm":
                            self.log("Şifre penceresi zaten açık, giriş yapılıyor...", "info")
                            return self._sifre_ile_giris_yap(w)
                    except:
                        pass

                # 2d) Hiçbir pencere yok → exe başlat + giriş
                self.log("Medula penceresi bulunamadı, açılıyor...", "info")
                return self._exe_baslat_ve_giris_yap()

            except Exception as e:
                self.log(f"Bağlantı hatası: {e}", "error")
                return False

    def medula_ac_ve_baglan(self):
        """Medula'yı aç ve bağlan — tek giriş noktası."""
        return self.medula_baglan()

    def medula_oturum_kurtarma(self):
        """Oturum düştüğünde çağrılır: taskkill + yeniden başlat + giriş.
        Tarama sırasında tıklama tepki vermezse bu fonksiyon kullanılır.
        """
        self.log("Oturum düşmüş — taskkill + yeniden başlatılıyor...", "warning")
        self._medula_kapat()
        time.sleep(3)
        return self._exe_baslat_ve_giris_yap()

    # === EXE BAŞLATMA + GİRİŞ ===

    def _exe_baslat_ve_giris_yap(self):
        """BotanikMedula.exe aç → şifre ekranı bekle → giriş yap → bağlan."""
        import subprocess as sp
        from pywinauto import Desktop

        if not os.path.exists(MEDULA_EXE):
            self.log(f"Exe bulunamadı: {MEDULA_EXE}", "error")
            return False

        # Önce eski process'leri temizle (çift açılma önleme)
        try:
            sp.run(["taskkill", "/F", "/IM", "BotanikMedula.exe"],
                   capture_output=True, text=True, timeout=5)
            time.sleep(2)
        except:
            pass

        # Exe başlat
        self.log("BotanikMedula.exe başlatılıyor...", "info")
        sp.Popen([MEDULA_EXE], cwd=os.path.dirname(MEDULA_EXE))

        # Şifre ekranı (SifreSorForm) bekle - max 20 sn
        giris_win = None
        for i in range(20):
            time.sleep(1)
            try:
                desktop = Desktop(backend="uia")
                for w in desktop.windows():
                    try:
                        if w.element_info.automation_id == "SifreSorForm":
                            giris_win = w
                            self.log(f"Şifre ekranı açıldı ({i+1} sn)", "info")
                            break
                    except:
                        pass
                if giris_win:
                    break
            except:
                pass

        if not giris_win:
            self.log("Şifre ekranı açılmadı!", "error")
            return False

        return self._sifre_ile_giris_yap(giris_win)

    def _sifre_ile_giris_yap(self, giris_win):
        """Açık şifre penceresinden giriş yap → ana pencere bekle → Medula oturum aç."""
        import pyautogui

        try:
            giris_win.set_focus()
            time.sleep(0.3)

            # Kullanıcı seç
            combo = None
            for elem in giris_win.descendants():
                try:
                    if elem.element_info.automation_id == "cmbKullanicilar":
                        combo = elem
                        break
                except:
                    pass
            if combo:
                combo.click_input()
                time.sleep(0.3)
                for item in combo.descendants(control_type="ListItem"):
                    try:
                        if "botan" in item.window_text().lower():
                            item.click_input()
                            self.log("Kullanıcı: botan", "info")
                            break
                    except:
                        pass
                time.sleep(0.2)

            # Şifre gir
            sifre = None
            for elem in giris_win.descendants():
                try:
                    if elem.element_info.automation_id == "txtSifre":
                        sifre = elem
                        break
                except:
                    pass
            if sifre:
                sifre.click_input()
                time.sleep(0.1)
                pyautogui.hotkey("ctrl", "a")
                time.sleep(0.05)
                sifre.type_keys(MEDULA_SIFRE, with_spaces=True)

            # Giriş butonu
            for elem in giris_win.descendants():
                try:
                    if elem.element_info.automation_id == "btnGirisYap":
                        elem.click_input()
                        self.log("Giriş butonuna basıldı", "info")
                        break
                except:
                    pass
        except Exception as e:
            self.log(f"Giriş hatası: {e}", "error")
            return False

        # BotanikEOS ana pencere açılmasını bekle ve bağlan - max 30 sn
        for i in range(30):
            time.sleep(1)
            desktop_new, win, hwnd = self._eos_penceresi_bul()
            if win:
                self.main_window = win
                self.medula_hwnd = hwnd
                self.bagli = True
                self.log(f"BotanikEOS açıldı ({i+1} sn)", "info")

                # Medula oturumunu başlat — btnMedulayaGirisYap'a bas
                # Bazen ilk giriş başarısız olur ("sisteme girilemedi" sayfası),
                # bu durumda tekrar basılması gerekir
                for deneme in range(3):
                    # Giriş butonuna bas
                    for elem in win.descendants():
                        try:
                            if elem.element_info.automation_id == "btnMedulayaGirisYap":
                                elem.click_input()
                                self.log(f"Medula portalı açılıyor... (deneme {deneme+1}/3)", "info")
                                break
                        except:
                            pass

                    # Oturum aktif olmasını bekle
                    for j in range(10):
                        time.sleep(1)
                        if self.oturum_aktif_mi():
                            self.log("Medula bağlantısı kuruldu!", "success")
                            return True

                self.log("Medula oturumu başlatılamadı", "error")
                return False

        self.log("BotanikEOS açılamadı!", "error")
        return False

    # === KAPATMA ===

    def _medula_kapat(self):
        """BotanikMedula.exe'yi taskkill ile kapat"""
        try:
            self.keepalive_durdur()

            result = subprocess.run(
                ["taskkill", "/F", "/IM", "BotanikMedula.exe"],
                capture_output=True, text=True, timeout=10
            )
            if result.returncode == 0:
                self.log("BotanikMedula.exe kapatıldı (taskkill)", "info")
            else:
                self.log("BotanikMedula.exe bulunamadı veya kapatılamadı", "warning")

            self.main_window = None
            self.medula_hwnd = None
            self.bagli = False
        except Exception as e:
            self.log(f"Taskkill hatası: {e}", "error")

    # === YARDIMCI ===

    def pencereyi_aktifle(self):
        """Medula penceresini öne getir"""
        try:
            import win32gui, win32con
            if self.medula_hwnd:
                win32gui.ShowWindow(self.medula_hwnd, win32con.SW_RESTORE)
                time.sleep(0.2)
                win32gui.SetForegroundWindow(self.medula_hwnd)
                time.sleep(0.3)
                return True
        except Exception as e:
            logger.error(f"Pencere aktifleştirme hatası: {e}")
        return False

    # === OTURUM CANLI TUTMA (KEEPALIVE) ===

    KEEPALIVE_ESIK = 90  # Saniye - bu süre boyunca Medula'ya tıklanmamışsa keepalive tetiklenir
    AKTIVITE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "medula_aktivite.tmp")

    def _son_aktivite_zamani(self):
        """Son Medula aktivite zamanını oku.
        Hem kendi _son_aktivite değişkenine hem de subprocess'in yazdığı dosyaya bakar.
        En yeni olanı döndürür."""
        en_yeni = self._son_aktivite

        # Subprocess'in yazdığı aktivite dosyasını oku
        try:
            if os.path.exists(self.AKTIVITE_FILE):
                with open(self.AKTIVITE_FILE, "r") as f:
                    dosya_zamani = float(f.read().strip())
                    if dosya_zamani > en_yeni:
                        en_yeni = dosya_zamani
        except:
            pass

        return en_yeni

    def aktivite_bildir(self):
        """Medula'ya tıklama yapıldığında çağrılır - keepalive zamanlayıcısını sıfırlar"""
        self._son_aktivite = time.time()
        # Dosyaya da yaz (subprocess de okuyabilsin)
        try:
            with open(self.AKTIVITE_FILE, "w") as f:
                f.write(str(self._son_aktivite))
        except:
            pass

    def keepalive_baslat(self):
        """
        Oturum canlı tutma - DEVRE DIŞI BIRAKILDI.
        Giriş butonuna otomatik basma, aktif tarama akışını bozuyordu.
        """
        logger.info("Keepalive devre dışı - otomatik giriş yapılmayacak")
        return

        if self._keepalive_aktif:
            return

        self._keepalive_aktif = True
        self._son_aktivite = time.time()

        def keepalive_loop():
            while self._keepalive_aktif and self.bagli:
                try:
                    time.sleep(10)  # Her 10 saniyede kontrol et
                    if not self._keepalive_aktif or not self.bagli:
                        break

                    if not self.medula_hwnd or not self.main_window:
                        continue

                    # Son aktiviteden bu yana geçen süre
                    son_zaman = self._son_aktivite_zamani()
                    gecen = time.time() - son_zaman
                    if gecen < self.KEEPALIVE_ESIK:
                        continue

                    # 90 saniye geçti, oturum kontrol et
                    logger.info(f"Keepalive: {gecen:.0f}sn aktivite yok, oturum kontrol ediliyor")

                    try:
                        if not self.oturum_aktif_mi():
                            logger.info("Keepalive: Oturum düşmüş, Giriş butonuna basılıyor")
                            for elem in self.main_window.descendants(control_type="Button"):
                                try:
                                    if elem.element_info.automation_id == "btnMedulayaGirisYap":
                                        elem.click_input()
                                        logger.info("Keepalive: Giriş butonu tıklandı")
                                        break
                                except:
                                    pass
                        else:
                            logger.debug("Keepalive: Oturum hala aktif")
                    except Exception:
                        pass

                    # Aktivite zamanını güncelle (tekrar 90sn beklesin)
                    self._son_aktivite = time.time()

                except Exception as e:
                    logger.debug(f"Keepalive hatası: {e}")

        self._keepalive_thread = threading.Thread(target=keepalive_loop, daemon=True)
        self._keepalive_thread.start()
        self.log("Oturum canlı tutma başlatıldı (90sn inaktivite eşiği)", "info")

    def keepalive_durdur(self):
        """Keepalive döngüsünü durdur"""
        self._keepalive_aktif = False
        self._keepalive_thread = None

    # === NAVİGASYON FONKSİYONLARI ===

    def _element_bul_ve_tikla(self, auto_id, aciklama=""):
        """AutomationID ile element bul ve tıkla (descendants ile güvenilir arama)"""
        try:
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == auto_id:
                        elem.click_input()
                        if aciklama:
                            self.log(aciklama, "success")
                        return True
                except:
                    pass
            self.log(f"Element bulunamadı: {auto_id}", "error")
            return False
        except Exception as e:
            self.log(f"Element tıklama hatası ({auto_id}): {e}", "error")
            return False

    def recete_listesine_git(self):
        """Sol menüdeki 'Reçete Listesi' butonuna tıkla"""
        if not self.bagli:
            self.log("Medula'ya bağlı değil!", "error")
            return False

        basarili = self._element_bul_ve_tikla(
            MEDULA_IDS["menu_recete_listesi"],
            "Reçete Listesi sayfasına gidildi"
        )
        if basarili:
            time.sleep(2)
        return basarili

    def grup_sekmesine_tikla(self, grup_kodu):
        """
        Reçete Listesi sayfasındaki grup sekmesine tıkla.
        grup_kodu: "C", "A", "B", "CK", "GK"

        Sekme tespiti: Tüm sekme adları ("A","B","C Sıralı","C Kan","GKKKOY","Yurtdışı")
        aynı y koordinatında DataItem olarak bulunur. Önce "C Sıralı" referans sekmesini
        bulup aynı y'deki hedef sekmeyi tıklar (tek harfli "A" gibi sekmelerde
        yanlış elemente tıklamayı önler).
        """
        if not self.bagli:
            return False

        grup_bilgi = next((g for g in GRUP_TANIMLARI if g["kod"] == grup_kodu), None)
        if not grup_bilgi:
            self.log(f"Bilinmeyen grup: {grup_kodu}", "error")
            return False

        medula_tab = grup_bilgi["medula_tab"]

        try:
            # 1. Referans sekmesini bul - GKKKOY kullan (ASCII, encoding sorunu yok)
            referans_y = None
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    if txt and txt.strip() == "GKKKOY":
                        referans_y = elem.rectangle().top
                        break
                except:
                    pass

            if referans_y is None:
                # Alternatif: "B" dene (tek karakter, encoding sorunu yok)
                for elem in self.main_window.descendants(control_type="DataItem"):
                    try:
                        txt = elem.window_text()
                        if txt and txt.strip() == "B":
                            r = elem.rectangle()
                            # B sekmesi çok belirgin olmayabilir - GKKKOY komşusu olmalı
                            referans_y = r.top
                            break
                    except:
                        pass

            if referans_y is None:
                self.log("Sekme satırı bulunamadı (GKKKOY/B referans yok)", "error")
                return False

            # 2. Hedef sekmeyi bul (aynı y satırında)
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    if not txt:
                        continue
                    stripped = txt.strip()

                    # Hedef sekme mi? (encoding-safe karşılaştırma)
                    eslesme = False
                    if medula_tab == stripped:
                        eslesme = True
                    elif medula_tab in ["C Sıralı", "C Sirali"] and "Sıralı" in stripped:
                        eslesme = True
                    elif medula_tab == "C Kan" and "Kan" in stripped and "C" in stripped:
                        eslesme = True
                    elif medula_tab in ["Yurtdışı", "Yurtdisi"] and ("Yurt" in stripped):
                        eslesme = True

                    if eslesme:
                        r = elem.rectangle()
                        if abs(r.top - referans_y) < 5:
                            elem.click_input()
                            self.log(f"'{medula_tab}' sekmesine tıklandı", "success")
                            # Sayfa yüklenmesi için bekle
                            time.sleep(3)
                            return True
                except:
                    pass

            self.log(f"'{medula_tab}' sekmesi bulunamadı (referans_y={referans_y})", "error")
            return False
        except Exception as e:
            self.log(f"Sekme tıklama hatası: {e}", "error")
            return False

    def sorgula(self):
        """Sorgula butonuna tıkla ve reçete listesinin yüklenmesini bekle"""
        if not self.bagli:
            return False

        basarili = self._element_bul_ve_tikla(
            MEDULA_IDS["sorgula_butonu"],
            "Sorgula butonuna tıklandı"
        )
        if basarili:
            # Reçete listesinin yüklenmesini bekle (3L ile başlayan DataItem görünene kadar)
            for bekle in range(10):
                time.sleep(1)
                for elem in self.main_window.descendants(control_type="DataItem"):
                    try:
                        txt = elem.window_text()
                        if txt and len(txt.strip()) == 7 and txt.strip()[0].isdigit() and txt.strip().isalnum():
                            self.log("Reçete listesi yüklendi", "info")
                            time.sleep(1)
                            return True
                    except:
                        pass
            # Timeout - reçete bulunamadı ama sayfa yüklendi
            time.sleep(2)
        return basarili

    # === OTURUM KONTROLÜ ===

    def oturum_aktif_mi(self):
        """Sol menü görünüyor mu? (oturum düşmüşse menü kaybolur)

        STABILITY: descendants() çağrısı IE-embed Medula penceresinde
        E_INTERFACE_NOT_REGISTERED COM hatası fırlatıp pywinauto'yu
        çökertebiliyor (access violation + stack overflow gözlemlendi).
        Hedefli child_window kullanarak tüm tree yürümeyi atlıyoruz.
        """
        if not self.bagli or not self.main_window:
            return False
        try:
            cw = self.main_window.child_window(auto_id=MEDULA_IDS["menu_recete_listesi"])
            return bool(cw.exists(timeout=0.3))
        except Exception:
            return False

    def oturumu_yenile(self):
        """
        Oturum düşmüşse BotanikEOS toolbar'daki Giriş butonuna
        birkaç kez basarak yenile.

        STABILITY: descendants() yerine hedefli child_window — IE-embed
        Medula penceresinde tree walk crash riski azaltıldı.
        """
        self.log("Oturum düşmüş, yenileniyor...", "warning")
        for deneme in range(3):
            try:
                cw = self.main_window.child_window(auto_id="btnMedulayaGirisYap")
                if cw.exists(timeout=0.3):
                    cw.click_input()
                    self.log(f"Giriş butonu tıklandı ({deneme+1}/3)", "info")
            except Exception:
                pass

            time.sleep(4)

            if self.oturum_aktif_mi():
                self.log("Oturum yenilendi!", "success")
                return True

        self.log("Oturum yenilenemedi", "error")
        return False

    def oturumu_kontrol_et_ve_yenile(self):
        """Oturum aktif mi kontrol et, değilse yenile. True dönerse kullanılabilir."""
        if self.oturum_aktif_mi():
            return True
        return self.oturumu_yenile()

    # === REÇETE LİSTESİ OKUMA ===

    def recete_listesi_oku(self):
        """
        Reçete Listesi sayfasındaki reçete numaralarını oku.
        Reçete no formatı: 3LE + 4 alfanümerik = 7 karakter (ör: 3LEO9QB)
        Returns: list[str] - reçete numaraları
        """
        if not self.bagli:
            return []

        receteler = []
        try:
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    if not txt:
                        continue
                    txt = txt.strip()
                    # Reçete numaraları: 3L ile başlar + 5 alfanümerik = 7 karakter
                    if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                        if txt not in receteler:  # Tekrar engelle
                            receteler.append(txt)
                except:
                    pass
        except Exception as e:
            self.log(f"Reçete listesi okuma hatası: {e}", "error")

        self.log(f"Reçete listesinde {len(receteler)} reçete bulundu", "info")
        return receteler

    def receteye_tikla(self, recete_no):
        """
        Reçete listesindeki bir reçete numarasına tıkla ve
        reçete sayfasının yüklenmesini bekle (f:buttonGeriDon görünene kadar).
        """
        if not self.bagli:
            return False

        try:
            tiklandi = False
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    if elem.window_text() == recete_no:
                        elem.click_input()
                        tiklandi = True
                        break
                except:
                    pass

            if not tiklandi:
                self.log(f"Reçete bulunamadı: {recete_no}", "error")
                return False

            # Reçete sayfasının ve ilaç tablosunun yüklenmesini bekle
            # Önce f:buttonIlacBilgiGorme, sonra f:tbl1:0:checkbox7 bekle
            sayfa_yuklendi = False
            for bekle in range(15):
                time.sleep(1)
                for elem in self.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == "f:buttonIlacBilgiGorme":
                            sayfa_yuklendi = True
                            break
                    except:
                        pass
                if sayfa_yuklendi:
                    break

            if not sayfa_yuklendi:
                self.log(f"Reçete sayfası yüklenemedi: {recete_no}", "error")
                return False

            # İlaç tablosunun yüklenmesini bekle (checkbox görünene kadar)
            for bekle in range(10):
                time.sleep(1)
                for elem in self.main_window.descendants(control_type="CheckBox"):
                    try:
                        if "tbl1" in str(elem.element_info.automation_id):
                            self.log(f"Reçete açıldı: {recete_no}", "info")
                            time.sleep(1)
                            return True
                    except:
                        pass

            # Checkbox bulunamadı ama sayfa yüklendi - devam et
            self.log(f"Reçete açıldı (tablo gecikmeli): {recete_no}", "warning")
            time.sleep(3)
            return True
        except Exception as e:
            self.log(f"Reçete tıklama hatası: {e}", "error")
            return False

    # === REÇETE DETAY OKUMA ===

    def recete_bilgileri_oku(self):
        """
        Açık reçetenin temel bilgilerini oku.
        DataItem'lardan reçete no, hasta adı, kapsam, fatura türü vb.

        Returns:
            dict: Reçete bilgileri
        """
        bilgiler = {
            "recete_no": "",
            "hasta_adi": "",
            "kapsam": "",
            "fatura_turu": "",
            "karekod_durumu": "",
            "e_recete_no": "",
        }

        try:
            items = self.main_window.descendants(control_type="DataItem")
            texts = []
            for item in items:
                try:
                    txt = item.window_text()
                    y = item.rectangle().top
                    if txt.strip() and y > 200 and y < 470:
                        texts.append((txt.strip(), y))
                except:
                    pass

            # Reçete No: 7 karakter alfanümerik (3M6WLIF, 3LExxxx vb.)
            for txt, y in texts:
                if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                    bilgiler["recete_no"] = txt
                elif "Sigortalı" in txt or txt.startswith("4"):
                    bilgiler["kapsam"] = txt
                elif "Grubu" in txt:
                    bilgiler["fatura_turu"] = txt
                elif "Sonlandırıldı" in txt:
                    bilgiler["karekod_durumu"] = txt

            # e-Reçete No (kısa format, reçete no'dan farklı)
            for txt, y in texts:
                if len(txt) <= 4 and txt.isalnum() and txt[0].isdigit():
                    bilgiler["e_recete_no"] = txt

        except Exception as e:
            self.log(f"Reçete bilgileri okuma hatası: {e}", "error")

        return bilgiler

    def ilac_tablosu_oku(self):
        """
        Açık reçetedeki ilaç tablosunu AutomationID bazlı oku.

        İlaç tablosu yapısı (keşfedildi 2026-03-28):
        AutomationID deseni: f:tbl1:{row}:xxx
        - f:tbl1:{row}:box32  → Satır container (Table)
        - f:tbl1:{row}:t2     → Adet (Edit)
        - f:tbl1:{row}:t5     → Periyot sayısı (Edit)
        - f:tbl1:{row}:m1     → Periyot birimi (ComboBox: 3=Günde,4=Haftada,5=Ayda)
        - f:tbl1:{row}:t3     → Çarpan (Edit)
        - f:tbl1:{row}:t4     → Doz (Edit, "1,0" formatı)
        - İlaç adı: DataItem (row container altında, uzun text)
        - Stk/Raf + Etkin madde: Doz satırı altında DataItem ("SGKFNL-SALBUTAMOL")
        - Tutar, Fark, Rapor, Verilebileceği, Msj: aynı row'da DataItem

        Returns:
            list[dict]: Her ilaç için bilgi sözlüğü
        """
        ilaclar = []

        try:
            # Kaç satır var? f:tbl1:{row}:box32 sayarak bul
            max_row = -1
            for elem in self.main_window.descendants():
                try:
                    aid = elem.element_info.automation_id
                    if aid and aid.startswith("f:tbl1:") and ":box32" in aid:
                        # f:tbl1:0:box32 → row=0
                        parts = aid.split(":")
                        row = int(parts[2])
                        if row > max_row:
                            max_row = row
                except:
                    pass

            if max_row < 0:
                return ilaclar

            self.log(f"İlaç tablosu: {max_row + 1} satır bulundu", "info")

            # Her element'in automation_id'sini ve text'ini topla
            elem_map = {}  # aid -> text
            for elem in self.main_window.descendants():
                try:
                    aid = elem.element_info.automation_id
                    if aid and "tbl1" in aid:
                        txt = elem.window_text()
                        if txt and txt.strip():
                            elem_map[aid] = txt.strip()
                except:
                    pass

            # Ayrıca DataItem'lardan ilaç adı, tutar, fark vb. oku
            # Her row için DataItem'ları topla
            all_data_items = []
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    aid = elem.element_info.automation_id or ""
                    r = elem.rectangle()
                    if txt and txt.strip():
                        all_data_items.append({
                            "txt": txt.strip(),
                            "aid": aid,
                            "y": r.top,
                            "x": r.left,
                        })
                except:
                    pass

            # Her satırı işle
            for row in range(max_row + 1):
                prefix = f"f:tbl1:{row}:"
                ilac = {
                    "satir": row,
                    "ilac_adi": "",
                    "etkin_madde": "",
                    "sgk_kodu": "",
                    "stk_raf": "",
                    "tutar": "",
                    "fark": "",
                    "rapor_kodu": "",
                    "verilebilecegi": "",
                    "msj": "",
                    "doz": "",
                    "adet": "",
                    "carpan": "",
                    "periyot": "",
                }

                # Edit alanlarını oku (doz bilgileri)
                ilac["adet"] = elem_map.get(f"{prefix}t2", "")
                ilac["periyot"] = elem_map.get(f"{prefix}t5", "")
                ilac["carpan"] = elem_map.get(f"{prefix}t3", "")
                doz_val = elem_map.get(f"{prefix}t4", "")
                if doz_val:
                    ilac["doz"] = doz_val

                # box32 konteynerinin y pozisyonunu bul (satır bazlı gruplama için)
                box_y = None
                for elem in self.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == f"{prefix}box32":
                            box_y = elem.rectangle().top
                            break
                    except:
                        pass

                if box_y is None:
                    continue

                # Bu satırdaki DataItem'ları bul (box_y ile ±40 piksel aralığında)
                row_items = [d for d in all_data_items
                             if abs(d["y"] - box_y) < 40]
                row_items.sort(key=lambda d: d["x"])

                for item in row_items:
                    txt = item["txt"]

                    # Hata/uyarı mesajlarını atla (ilaç adı değil)
                    if any(k in txt.upper() for k in
                           ["VERİLEMEZ", "KONTROL", "UZMAN", "UYUMLU ICD",
                            "DEĞER TEMİZLE", "İNCELEMEYE"]):
                        continue

                    # İlaç adı: uzun text, ilaç formatı veya en uzun text
                    ilac_aday = (
                        len(txt) > 10 and any(k in txt.upper() for k in
                            ["MG", "ML", "TABLET", "KAPSUL", "KAPSÜL", "DOZ",
                             "INH", "GARGARA", "ŞURUP", "DAMLA", "AMPUL",
                             "FLAKON", "KREM", "JEL", "POMAD", "ENJEKTÖR",
                             "ŞASE", "SURUP", "SPRAY", "FITIL", "SOLÜSYON",
                             "SÜSPANSIYON", "ŞIRINGAS", "KALEM", "PATCH",
                             "TOZU", "LOZANJ", "SACHET", "GRANÜL", "MERHEM"])
                    )
                    if ilac_aday:
                        ilac["ilac_adi"] = txt
                    # İlaç adı bulunamadıysa, satırdaki en uzun text'i kullan
                    elif not ilac["ilac_adi"] and len(txt) > 15 and not txt.replace(",", "").replace(".", "").replace(" ", "").isdigit():
                        ilac["_uzun_aday"] = txt
                    # Msj: var/yok
                    elif txt.lower() in ["var", "yok"]:
                        ilac["msj"] = txt.lower()
                    # Verilebileceği: tarih formatı
                    elif "/" in txt and len(txt) == 10 and txt[2] == "/":
                        ilac["verilebilecegi"] = txt
                    # Rapor kodu: XX.XX formatı
                    elif "." in txt and len(txt) <= 8 and txt[0].isdigit():
                        ilac["rapor_kodu"] = txt
                    # Tutar/Fark: sayısal değer (virgüllü)
                    elif txt.replace(",", "").replace(".", "").replace(" ", "").isdigit():
                        if not ilac["tutar"]:
                            ilac["tutar"] = txt
                        elif not ilac["fark"]:
                            ilac["fark"] = txt
                    # Stk/Raf: satır no + depo kodu
                    elif "\n" in txt or ("(" in txt and ")" in txt and len(txt) < 15):
                        ilac["stk_raf"] = txt.replace("\n", " ")

                # Doz satırından etkin madde oku (row+1 piksel altında)
                doz_items = [d for d in all_data_items
                             if d["y"] > box_y + 30 and d["y"] < box_y + 70]
                for item in doz_items:
                    txt = item["txt"]
                    # Doz string: "1 Günde 4 x 2,00 - Adet"
                    if "Günde" in txt or "Haftada" in txt or "Ayda" in txt:
                        ilac["doz"] = txt
                    # Etkin madde: "SGKFNL-SALBUTAMOL" veya "SGKFK3-PIOGLITAZON HCL"
                    elif txt.startswith("SGK") and "-" in txt:
                        parts = txt.split("-", 1)
                        ilac["sgk_kodu"] = parts[0].strip()
                        ilac["etkin_madde"] = parts[1].strip() if len(parts) > 1 else ""

                # Uyarı mesajları (daha altta)
                uyari_items = [d for d in all_data_items
                               if d["y"] > box_y + 60 and d["y"] < box_y + 100]
                for item in uyari_items:
                    txt = item["txt"]
                    if "VERİLEMEZ" in txt or "KONTROL" in txt or "UZMAN" in txt:
                        if "uyari" not in ilac:
                            ilac["uyari"] = []
                        ilac["uyari"].append(txt)

                # İlaç adı bulunamadıysa uzun aday'ı kullan
                if not ilac["ilac_adi"] and ilac.get("_uzun_aday"):
                    ilac["ilac_adi"] = ilac["_uzun_aday"]
                ilac.pop("_uzun_aday", None)

                if ilac["ilac_adi"]:
                    ilaclar.append(ilac)

        except Exception as e:
            self.log(f"İlaç tablosu okuma hatası: {e}", "error")

        return ilaclar

    # === REÇETE NAVİGASYON ===

    def sonraki_recete(self):
        """Sonraki Reçete butonuna tıkla"""
        try:
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "f:buttonSonraki":
                        elem.click_input()
                        time.sleep(3)
                        return True
                except:
                    pass
        except:
            pass
        return False

    def onceki_recete(self):
        """Önceki Reçete butonuna tıkla"""
        try:
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "f:buttonOncekiRecete":
                        elem.click_input()
                        time.sleep(3)
                        return True
                except:
                    pass
        except:
            pass
        return False

    def geri_don(self):
        """Geri Dön butonuna tıkla"""
        try:
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "f:buttonGeriDon":
                        elem.click_input()
                        time.sleep(2)
                        return True
                except:
                    pass
        except:
            pass
        return False

    # === İLAÇ BİLGİ OKUMA (f:buttonIlacBilgiGorme) ===

    def ilac_bilgi_oku(self, satir_index):
        """
        Belirli bir ilaç satırının İlaç Bilgi sayfasını aç ve oku.
        Etkin madde, SGK kodu, mesaj, maks doz bilgilerini döndürür.

        AutomationID'ler (keşfedildi 2026-03-07 / güncellendi 2026-03-28):
        - Checkbox: f:tbl1:{row}:checkbox7
        - İlaç Bilgi butonu: f:buttonIlacBilgiGorme
        - Etkin madde: form1:text35
        - SGK kodu: form1:text2
        - Mesaj başlığı: form1:tableExIlacMesajListesi:{idx}:text19
        - Mesaj metni: form1:textarea1
        - Kapat butonu: form1:buttonKapat

        Args:
            satir_index: İlaç tablosundaki satır indexi (0-4)

        Returns:
            dict: {etkin_madde, sgk_kodu, mesaj_basligi, mesaj_metni,
                   ayaktan_maks_doz, raporlu_maks_doz, rapor_turu, sut_bilgi}
        """
        bilgi = {
            "etkin_madde": "", "sgk_kodu": "", "mesaj_basligi": "",
            "mesaj_metni": "", "ayaktan_maks_doz": "", "raporlu_maks_doz": "",
            "rapor_turu": "", "sut_bilgi": "",
        }

        try:
            # 1. Checkbox seç (AutomationID bazlı)
            cb_id = f"f:tbl1:{satir_index}:checkbox7"
            self._element_bul_ve_tikla(cb_id)
            time.sleep(0.3)

            # 2. İlaç Bilgi butonuna tıkla
            self._element_bul_ve_tikla("f:buttonIlacBilgiGorme")
            time.sleep(3)

            # 3. Sayfanın yüklenmesini bekle (form1:buttonKapat görünene kadar)
            for bekle in range(10):
                for elem in self.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == "form1:buttonKapat":
                            break
                    except:
                        pass
                else:
                    time.sleep(1)
                    continue
                break

            # 4. AutomationID bazlı bilgi oku
            for elem in self.main_window.descendants():
                try:
                    aid = elem.element_info.automation_id
                    if not aid:
                        continue
                    txt = elem.window_text()
                    if not txt or not txt.strip():
                        continue
                    txt = txt.strip()

                    # Etkin madde
                    if aid == "form1:text35":
                        bilgi["etkin_madde"] = txt
                    # SGK kodu
                    elif aid == "form1:text2":
                        bilgi["sgk_kodu"] = txt
                    # Mesaj başlığı (ilk mesaj)
                    elif "tableExIlacMesajListesi" in aid and "text19" in aid:
                        if not bilgi["mesaj_basligi"]:
                            bilgi["mesaj_basligi"] = txt
                    # Mesaj metni
                    elif aid == "form1:textarea1":
                        bilgi["mesaj_metni"] = txt
                except:
                    pass

            # 5. DataItem'lardan ek bilgileri oku (doz, rapor türü, SUT)
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    if not txt or not txt.strip():
                        continue
                    txt = txt.strip()

                    if "Günde" in txt and "x" in txt:
                        if not bilgi["ayaktan_maks_doz"]:
                            bilgi["ayaktan_maks_doz"] = txt
                        elif not bilgi["raporlu_maks_doz"]:
                            bilgi["raporlu_maks_doz"] = txt
                    elif "Uzman Hekim Raporu" in txt:
                        bilgi["rapor_turu"] = txt
                    elif "Raporlu" in txt or "Raporsuz" in txt:
                        bilgi["sut_bilgi"] = txt
                except:
                    pass

            self.log(f"İlaç Bilgi okundu: {bilgi.get('etkin_madde', '?')}", "info")

            # 6. Kapat butonu ile geri dön (form1:buttonKapat)
            self._element_bul_ve_tikla("form1:buttonKapat")
            time.sleep(2)

        except Exception as e:
            self.log(f"İlaç Bilgi okuma hatası: {e}", "error")

        return bilgi

    # === RAPOR OKUMA (f:buttonRaporGoruntule) ===

    def rapor_bilgileri_oku(self, satir_index):
        """
        Belirli bir ilaç satırının Rapor sayfasını aç ve oku.

        AutomationID'ler:
        - Checkbox: f:tbl1:{row}:checkbox7
        - Rapor butonu: f:buttonRaporGoruntule
        - Rapor Listesi: f:buttonRaporListesi
        - Geri Dön: form1:buttonGeriDon

        Returns:
            dict: {rapor_no, rapor_tarihi, bitis_tarihi, tesis,
                   doktor, brans, tani_kodu, tani_adi,
                   etkin_maddeler: [{sgk_kodu, etkin_madde, doz}]}
        """
        rapor = {
            "rapor_no": "", "rapor_tarihi": "", "bitis_tarihi": "",
            "tesis": "", "doktor": "", "brans": "",
            "tani_kodu": "", "tani_adi": "",
            "etkin_maddeler": [],
        }

        try:
            # 1. Checkbox seç
            cb_id = f"f:tbl1:{satir_index}:checkbox7"
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == cb_id:
                        elem.click_input()
                        time.sleep(0.3)
                        break
                except:
                    pass

            # 2. Rapor butonuna tıkla
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "f:buttonRaporGoruntule":
                        elem.click_input()
                        time.sleep(4)
                        break
                except:
                    pass

            # 3. Rapor sayfası DataItem'larını oku
            items = []
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    r = elem.rectangle()
                    if txt.strip() and r.top > 100 and len(txt) < 200:
                        items.append((txt.strip(), r.top, r.left))
                except:
                    pass

            # Parse et
            for txt, y, x in items:
                # Rapor numarası
                if y > 290 and y < 320 and txt.isdigit() and len(txt) <= 6:
                    rapor["rapor_no"] = txt
                # Tarihler
                elif "/" in txt and len(txt) == 10 and txt[2] == "/":
                    if not rapor["rapor_tarihi"]:
                        rapor["rapor_tarihi"] = txt
                    elif not rapor["bitis_tarihi"]:
                        rapor["bitis_tarihi"] = txt
                # Tesis
                elif "HASTANE" in txt.upper() or "DEVLET" in txt.upper():
                    rapor["tesis"] = txt
                # Tanı kodu ve adı
                elif txt.startswith("J") or txt.startswith("E") or txt.startswith("N") or txt.startswith("I"):
                    if "." in txt and len(txt) <= 6:
                        rapor["tani_kodu"] = txt
                    elif len(txt) > 10:
                        rapor["tani_adi"] = txt
                # Rapor kodu + tanı açıklaması (ör: "05.02 - Kronik...")
                elif " - " in txt and txt[0].isdigit():
                    parts = txt.split(" - ", 1)
                    rapor["tani_kodu"] = rapor["tani_kodu"] or parts[0].strip()
                    rapor["tani_adi"] = rapor["tani_adi"] or (parts[1].strip() if len(parts) > 1 else "")
                # Branş
                elif "Hastalıkları" in txt or "Hekimi" in txt:
                    rapor["brans"] = txt
                # Etkin madde + doz (rapor etkin madde bilgileri bölümü)
                elif txt.startswith("SGK"):
                    rapor["etkin_maddeler"].append({
                        "sgk_kodu": txt,
                        "etkin_madde": "",
                        "doz": "",
                    })
                elif "Günde" in txt and "x" in txt and rapor["etkin_maddeler"]:
                    rapor["etkin_maddeler"][-1]["doz"] = txt

            # Etkin madde isimlerini eşleştir
            etkin_madde_texts = [txt for txt, y, x in items
                                if len(txt) > 10 and any(k in txt.upper() for k in
                                ["FORMOTEROL", "TIOTROPIUM", "SALBUTAMOL", "PIOGLITAZON",
                                 "METFORMIN", "INSULIN", "BEKLOMETAZON", "VILANTEROL",
                                 "FLUTIKAZON", "TEOFILIN", "UMEKLIDINYUM"])]
            for i, em in enumerate(rapor["etkin_maddeler"]):
                if i < len(etkin_madde_texts):
                    em["etkin_madde"] = etkin_madde_texts[i]

            self.log(f"Rapor okundu: {rapor['tani_kodu']} - {len(rapor['etkin_maddeler'])} etkin madde", "info")

            # 4. Geri dön
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "form1:buttonGeriDon":
                        elem.click_input()
                        time.sleep(2)
                        break
                except:
                    pass

        except Exception as e:
            self.log(f"Rapor okuma hatası: {e}", "error")

        return rapor

    # === İLAÇ GEÇMİŞİ OKUMA (f:buttonIlacListesi) ===

    def ilac_gecmisi_oku(self):
        """
        İlaç Geçmişi (Kullanılan İlaç Listesi) sayfasını aç ve oku.

        AutomationID'ler:
        - İlaç butonu: f:buttonIlacListesi
        - Geri Dön: form1:buttonGeriDon
        - Göz İlaçları: form1:buttonGozIlacListesi

        Returns:
            list[dict]: {recete_no, ilac_adi, barkod, recete_tarihi,
                        ilac_alim_tarihi, verilebilecegi, doz, rapor_kodu,
                        etkin_madde}
        """
        gecmis = []

        try:
            # 1. İlaç butonuna tıkla
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "f:buttonIlacListesi":
                        elem.click_input()
                        time.sleep(4)
                        break
                except:
                    pass

            # 2. DataItem'ları oku
            items = []
            for elem in self.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    r = elem.rectangle()
                    if txt.strip() and r.top > 450 and len(txt) < 200:
                        items.append((txt.strip(), r.top, r.left))
                except:
                    pass

            # Y koordinatına göre grupla
            items.sort(key=lambda x: (x[1], x[2]))

            # Her satır bir ilaç kaydı (reçete no ile başlar)
            current = None
            for txt, y, x in items:
                # Reçete no (7 karakter alfanümerik)
                if len(txt) == 7 and txt[0].isdigit() and txt.isalnum() and x < 700:
                    if current:
                        gecmis.append(current)
                    current = {
                        "recete_no": txt, "ilac_adi": "", "barkod": "",
                        "recete_tarihi": "", "ilac_alim_tarihi": "",
                        "verilebilecegi": "", "doz": "", "rapor_kodu": "",
                        "etkin_madde": "",
                    }
                elif current:
                    # İlaç adı (uzun, parantez içinde barkod)
                    if "(" in txt and ")" in txt and len(txt) > 20:
                        # "VENTOLIN INHALER 200 DOZ (SABA)(8699522521456)"
                        current["ilac_adi"] = txt
                    # Tarih
                    elif "/" in txt and len(txt) == 10 and txt[2] == "/":
                        if not current["recete_tarihi"]:
                            current["recete_tarihi"] = txt
                        elif not current["ilac_alim_tarihi"]:
                            current["ilac_alim_tarihi"] = txt
                        elif not current["verilebilecegi"]:
                            current["verilebilecegi"] = txt
                    # Doz
                    elif "Günde" in txt and "x" in txt:
                        current["doz"] = txt
                    # Rapor kodu
                    elif " - " in txt and txt[:5].replace(".", "").isdigit():
                        current["rapor_kodu"] = txt
                    # Etkin madde (SGK kodu veya uzun isim)
                    elif txt.startswith("SGK") or (len(txt) > 15 and any(k in txt.upper() for k in
                            ["FORMOTEROL", "TIOTROPIUM", "SALBUTAMOL", "PIOGLITAZON",
                             "METFORMIN", "BROMUR", "FUMARAT"])):
                        current["etkin_madde"] = txt

            if current:
                gecmis.append(current)

            self.log(f"İlaç geçmişi: {len(gecmis)} kayıt", "info")

            # 3. Geri dön
            for elem in self.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "form1:buttonGeriDon":
                        elem.click_input()
                        time.sleep(2)
                        break
                except:
                    pass

        except Exception as e:
            self.log(f"İlaç geçmişi okuma hatası: {e}", "error")

        return gecmis


class MultiSelectDonemDropdown(tk.Frame):
    """Checkbox'lı popup açan çoklu dönem seçim widget'ı.

    Bir butona benzer şekilde davranır; tıklayınca açılır panelde her dönemin
    yanında onay kutusu çıkar. Seçili dönem sayısına göre buton etiketi
    'YYYY Ay' (tek) ya da 'YYYY Ay  +N' (çoklu) formatında güncellenir.
    on_change callback'i seçim değiştikçe tetiklenir.
    """

    def __init__(self, parent, donemler, varsayilan=None, on_change=None, **btn_kwargs):
        try:
            pbg = parent.cget("bg")
        except Exception:
            pbg = None
        if pbg:
            super().__init__(parent, bg=pbg)
        else:
            super().__init__(parent)

        self.donemler = list(donemler)
        self.on_change = on_change

        if varsayilan is None:
            self.secili = [self.donemler[0]] if self.donemler else []
        elif isinstance(varsayilan, (list, tuple, set)):
            self.secili = [d for d in self.donemler if d in varsayilan]
            if not self.secili and self.donemler:
                self.secili = [self.donemler[0]]
        else:
            self.secili = ([varsayilan] if varsayilan in self.donemler
                           else ([self.donemler[0]] if self.donemler else []))

        defaults = dict(font=("Segoe UI", 9), relief="sunken",
                        bg="white", fg="black", bd=1, cursor="hand2")
        defaults.update(btn_kwargs)
        self._buton = tk.Button(self, text=self._etiket(),
                                 command=self._popup_toggle, **defaults)
        self._buton.pack(fill="both", expand=True)
        self._popup = None
        self._popup_varlari = {}
        self._global_click_funcid = None

    def _etiket(self):
        if not self.secili:
            return "Dönem seç... ▼"
        if len(self.secili) == 1:
            return f"{self.secili[0]}  ▼"
        return f"{self.secili[0]}  +{len(self.secili) - 1}  ▼"

    def _popup_toggle(self):
        if self._popup is not None and self._popup.winfo_exists():
            self._popup_kapat()
            return
        self._popup_ac()

    def _popup_ac(self):
        x = self._buton.winfo_rootx()
        y = self._buton.winfo_rooty() + self._buton.winfo_height()
        popup = tk.Toplevel(self)
        popup.wm_overrideredirect(True)
        popup.geometry(f"+{x}+{y}")
        popup.configure(bg="#37474F", bd=1, relief="solid",
                         highlightbackground="#90CAF9", highlightthickness=1)
        try:
            popup.transient(self.winfo_toplevel())
        except Exception:
            pass
        try:
            popup.attributes("-topmost", True)
        except Exception:
            pass

        tk.Label(popup, text="Dönem seçin (çoklu)",
                 font=("Segoe UI", 9, "bold"),
                 fg="#FFFFFF", bg="#263238", anchor="w",
                 padx=8, pady=3).pack(fill="x")

        self._popup_varlari = {}
        for d in self.donemler:
            var = tk.BooleanVar(value=(d in self.secili))
            self._popup_varlari[d] = var
            tk.Checkbutton(popup, text=d, variable=var, anchor="w",
                           bg="#37474F", fg="#FFFFFF",
                           selectcolor="#263238",
                           activebackground="#455A64",
                           activeforeground="#FFFFFF",
                           font=("Segoe UI", 9), padx=8,
                           command=lambda dd=d, vv=var: self._toggle(dd, vv)
                           ).pack(fill="x", padx=2, pady=1)

        alt = tk.Frame(popup, bg="#263238")
        alt.pack(fill="x")
        tk.Button(alt, text="Tümünü Seç", font=("Segoe UI", 8),
                  fg="white", bg="#1565C0", activebackground="#0D47A1",
                  bd=0, padx=8, pady=2, command=self._tumunu_sec
                  ).pack(side="left", padx=4, pady=3)
        tk.Button(alt, text="Temizle", font=("Segoe UI", 8),
                  fg="white", bg="#6A1B9A", activebackground="#4A148C",
                  bd=0, padx=8, pady=2, command=self._temizle
                  ).pack(side="left", padx=4, pady=3)
        tk.Button(alt, text="Kapat", font=("Segoe UI", 8),
                  fg="white", bg="#37474F", activebackground="#263238",
                  bd=0, padx=8, pady=2, command=self._popup_kapat
                  ).pack(side="right", padx=4, pady=3)

        self._popup = popup
        # Popup açık dururken: dışarıya tıklayınca kapat (after_idle ile bind
        # edildi ki popup'ı açan ilk tıklama tetiklemesin). Escape ile de
        # kapatılabilir.
        popup.bind("<Escape>", lambda e: self._popup_kapat())
        try:
            popup.focus_set()
        except Exception:
            pass
        self.after_idle(self._global_click_bind)

    def _global_click_bind(self):
        try:
            self._global_click_funcid = self.winfo_toplevel().bind(
                "<Button-1>", self._on_global_click, add="+"
            )
        except Exception:
            self._global_click_funcid = None

    def _on_global_click(self, event):
        """Popup açıkken: tıklama popup veya butonu içinde değilse kapat."""
        if self._popup is None:
            return
        try:
            x, y = event.x_root, event.y_root
            # Popup bbox
            px, py = self._popup.winfo_rootx(), self._popup.winfo_rooty()
            pw, ph = self._popup.winfo_width(), self._popup.winfo_height()
            if px <= x <= px + pw and py <= y <= py + ph:
                return  # popup içine tıklandı
            # Dropdown butonu bbox (butona tıklamak _popup_toggle ile zaten
            # işleniyor; biz burada kapatırsak çakışır)
            bx, by = self._buton.winfo_rootx(), self._buton.winfo_rooty()
            bw, bh = self._buton.winfo_width(), self._buton.winfo_height()
            if bx <= x <= bx + bw and by <= y <= by + bh:
                return
        except Exception:
            pass
        self._popup_kapat()

    def _popup_kapat(self, *args):
        if self._global_click_funcid:
            try:
                self.winfo_toplevel().unbind("<Button-1>", self._global_click_funcid)
            except Exception:
                pass
            self._global_click_funcid = None
        if self._popup is not None:
            try:
                if self._popup.winfo_exists():
                    self._popup.destroy()
            except Exception:
                pass
            self._popup = None
        self._popup_varlari = {}

    def _toggle(self, donem, var):
        if var.get():
            if donem not in self.secili:
                # Orijinal sırada tut
                self.secili = [d for d in self.donemler
                               if d in self.secili or d == donem]
        else:
            self.secili = [d for d in self.secili if d != donem]
        # En az bir dönem zorunlu
        if not self.secili and self.donemler:
            self.secili = [donem]
            var.set(True)
        self._buton.config(text=self._etiket())
        if self.on_change:
            try:
                self.on_change(list(self.secili))
            except Exception:
                pass

    def _tumunu_sec(self):
        self.secili = list(self.donemler)
        for var in self._popup_varlari.values():
            var.set(True)
        self._buton.config(text=self._etiket())
        if self.on_change:
            try:
                self.on_change(list(self.secili))
            except Exception:
                pass

    def _temizle(self):
        ilk = self.donemler[0] if self.donemler else None
        self.secili = [ilk] if ilk else []
        for d, var in self._popup_varlari.items():
            var.set(d == ilk)
        self._buton.config(text=self._etiket())
        if self.on_change:
            try:
                self.on_change(list(self.secili))
            except Exception:
                pass

    def get_secilenler(self):
        return list(self.secili)

    def set_secilenler(self, donemler):
        valid = [d for d in self.donemler if d in donemler]
        if not valid and self.donemler:
            valid = [self.donemler[0]]
        self.secili = valid
        self._buton.config(text=self._etiket())


class ReceteRaporKontrolGUI:
    """Reçete/Rapor Kontrol ana ekranı"""

    def __init__(self, root, ana_menu_callback=None):
        self.root = root
        self.ana_menu_callback = ana_menu_callback
        self.root.title("Reçete / Rapor Kontrol")
        ekran_w = root.winfo_screenwidth()
        pencere_w = 600
        pencere_h = 950
        pencere_x = ekran_w - pencere_w
        self.root.geometry(f"{pencere_w}x{pencere_h}+{pencere_x}+0")
        self.root.resizable(True, True)
        self.root.configure(bg="#1E3A5F")

        # Log alanı (önce oluştur ki MedulaBaglanti kullanabilsin)
        self.log_text = None

        # Medula bağlantısı
        self.medula = MedulaBaglanti(log_callback=self.log_yaz)

        # Kontrol durumu
        self.kontrol_aktif = False
        self.aktif_grup = None
        self._subprocess_proc = None
        self._coklu_donem_iptal = False

        # Escape kısayolu - taramayı durdur
        self.root.bind_all("<Escape>", lambda e: self._taramayi_durdur())

        # Grup durumlarını yükle
        self.grup_durumlari = self._durumlari_yukle()

        # GUI referansları
        self.grup_labels = {}
        self.grup_butonlari = {}
        self.durum_labels = {}
        self.recete_combolar = {}  # kod → Combobox widget
        self.recete_listeleri = {}  # kod → [(sira, recete_no), ...]
        self.medula_durum_label = None
        self.timing_labels = {}    # kod → Label (ortalama süre göstergesi)
        self._recete_sureleri = []  # [(sure_sn), ...] aktif tarama süreleri
        self._recete_baslangic = None  # son reçete başlangıç zamanı

        self._gui_olustur()

        # Combobox hafızasını yükle
        self._recete_combobox_hafiza_yukle()

        # Başlangıçta Medula durumunu kontrol et
        self.root.after(500, self._medula_durumu_kontrol)

    def _durumlari_yukle(self):
        """Grup durumlarını JSON'dan yükle"""
        varsayilan = {}
        for g in GRUP_TANIMLARI:
            varsayilan[g["kod"]] = {
                "son_recete": "",
                "toplam_kontrol": 0,
                "son_kontrol_tarihi": None,
            }

        if os.path.exists(DURUM_DOSYASI):
            try:
                with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                    kaydedilen = json.load(f)
                for kod in varsayilan:
                    if kod not in kaydedilen:
                        kaydedilen[kod] = varsayilan[kod]
                return kaydedilen
            except Exception as e:
                logger.error(f"Durum dosyası yükleme hatası: {e}")

        return varsayilan

    def _durumlari_kaydet(self):
        """Grup durumlarını JSON'a kaydet"""
        try:
            with open(DURUM_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(self.grup_durumlari, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"Durum kaydetme hatası: {e}")

    def _gui_olustur(self):
        """Ana GUI'yi oluştur"""
        # === BAŞLIK ===
        baslik_frame = tk.Frame(self.root, bg="#0D2137", height=50)
        baslik_frame.pack(fill="x")
        baslik_frame.pack_propagate(False)

        tk.Label(
            baslik_frame, text="Reçete / Rapor Kontrol",
            font=("Segoe UI", 16, "bold"), fg="white", bg="#0D2137"
        ).pack(side="left", padx=15, pady=10)

        geri_btn = tk.Button(
            baslik_frame, text="<- Ana Menü", font=("Segoe UI", 10),
            fg="white", bg="#455A64", activebackground="#37474F",
            bd=0, padx=10, pady=5, command=self._ana_menuye_don
        )
        geri_btn.pack(side="right", padx=15, pady=10)

        # Sayfa Yerleşimi butonları (sağ üst, Ana Menü solu)
        yer_uygula_btn = tk.Button(
            baslik_frame, text="📐 Yerleşimi Uygula",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#1565C0", activebackground="#0D47A1",
            bd=0, padx=10, pady=5, cursor="hand2",
            command=self._yerlesim_uygula_tikla
        )
        yer_uygula_btn.pack(side="right", padx=(0, 5), pady=10)

        yer_kaydet_btn = tk.Button(
            baslik_frame, text="💾 Yerleşimi Kaydet",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#2E7D32", activebackground="#1B5E20",
            bd=0, padx=10, pady=5, cursor="hand2",
            command=self._yerlesim_kaydet_tikla
        )
        yer_kaydet_btn.pack(side="right", padx=(0, 5), pady=10)

        # SUT Matrisi (DB tabanlı, ayrı pencere)
        sut_matrisi_btn = tk.Button(
            baslik_frame, text="🎯 SUT Matrisi",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#3F51B5", activebackground="#303F9F",
            bd=0, padx=10, pady=5, cursor="hand2",
            command=self._sut_matrisi_ac
        )
        sut_matrisi_btn.pack(side="right", padx=(0, 5), pady=10)

        # === MEDULA BAĞLANTI BÖLÜMÜ ===
        medula_frame = tk.Frame(self.root, bg="#263238")
        medula_frame.pack(fill="x", padx=15, pady=(10, 5))

        # Sol: MEDULA Bağlan butonu + üzerinde "Bağlantıyı Açık Tut" checkbox'ı
        medula_btn_frame = tk.Frame(medula_frame, bg="#263238")
        medula_btn_frame.pack(side="left", padx=5, pady=5)

        # Checkbox + geri sayım satırı
        cb_satir = tk.Frame(medula_btn_frame, bg="#263238")
        cb_satir.pack(anchor="w", pady=(0, 2))

        self.canli_tut_var = tk.BooleanVar(value=self._canli_tut_tercihi_yukle())
        self.canli_tut_cb = tk.Checkbutton(
            cb_satir, text="Bağlantıyı Açık Tut",
            variable=self.canli_tut_var,
            font=("Segoe UI", 8), fg="#90CAF9", bg="#263238",
            selectcolor="#0D2137", activebackground="#263238",
            activeforeground="#90CAF9",
            command=self._canli_tut_toggle,
        )
        self.canli_tut_cb.pack(side="left")

        self.lbl_canli_tut_sayac = tk.Label(
            cb_satir, text="", bg="#263238", fg="#80CBC4",
            font=("Segoe UI", 8),
        )
        self.lbl_canli_tut_sayac.pack(side="left", padx=(4, 0))

        # Eğer tercih açıksa sayacı başlat
        if self.canli_tut_var.get():
            self.root.after(1000, self._canli_tut_sayac_guncelle)

        medula_btn = tk.Button(
            medula_btn_frame, text="MEDULA Bağlan",
            font=("Segoe UI", 10, "bold"),
            fg="white", bg="#0277BD", activebackground="#01579B",
            bd=0, padx=12, pady=6, cursor="hand2",
            command=self._medula_baglan_tikla
        )
        medula_btn.pack(side="left")
        self.medula_btn = medula_btn

        self.medula_durum_label = tk.Label(
            medula_frame, text="Durum: Kontrol ediliyor...",
            font=("Segoe UI", 9), fg="#90A4AE", bg="#263238"
        )
        self.medula_durum_label.pack(side="left", padx=10)

        # === İÇERİK ===
        icerik_frame = tk.Frame(self.root, bg="#1E3A5F")
        icerik_frame.pack(fill="both", expand=True, padx=15, pady=5)

        # === RENKLI REÇETE + AY SEÇİMİ SATIRI ===
        ust_ayar_frame = tk.Frame(icerik_frame, bg="#1E3A5F")
        ust_ayar_frame.pack(fill="x", pady=(0, 5))

        # Ay seçimi (önce pack - sağda sabit kalması için)
        self.renkli_recete_dosya = None
        self.renkli_recete_listesi = []

        # İlaç mesajı yoksay listesi (substring patternleri)
        self.mesaj_yoksay_patternleri = []
        self._mesaj_yoksay_yukle()

        ay_frame = tk.Frame(ust_ayar_frame, bg="#37474F", bd=1, relief="groove")
        ay_frame.pack(side="right", padx=(5, 0))

        # Renkli Reçete Excel Yükle butonu + hatırla checkbox (kalan alan)
        renkli_frame = tk.Frame(ust_ayar_frame, bg="#37474F", bd=1, relief="groove")
        renkli_frame.pack(side="left", fill="x", expand=True, padx=(0, 5))

        self.renkli_btn = tk.Button(
            renkli_frame, text="📋 Renkli Reçete Excel Yükle",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#6A1B9A", activebackground="#4A148C",
            bd=0, padx=10, pady=6, cursor="hand2",
            command=self._renkli_recete_yukle
        )
        self.renkli_btn.pack(side="left", padx=3, pady=3)

        self.renkli_durum_label = tk.Label(
            renkli_frame, text="Yüklenmedi",
            font=("Segoe UI", 8), fg="#FF9800", bg="#37474F"
        )
        self.renkli_durum_label.pack(side="left", padx=3)

        # En son tarih label'ı (Excel'de ne tarihe kadarki reçeteler var)
        self.renkli_tarih_label = tk.Label(
            renkli_frame, text="", font=("Segoe UI", 8, "bold"),
            fg="#FFD54F", bg="#37474F"
        )
        self.renkli_tarih_label.pack(side="left", padx=8)

        # "Hatırla" checkbox
        self.renkli_hatirla_var = tk.BooleanVar(value=True)
        self.renkli_hatirla_cb = tk.Checkbutton(
            renkli_frame, text="Hatırla",
            variable=self.renkli_hatirla_var,
            font=("Segoe UI", 9), fg="#90CAF9", bg="#37474F",
            selectcolor="#263238", activebackground="#37474F",
            activeforeground="#90CAF9",
        )
        self.renkli_hatirla_cb.pack(side="left", padx=4)

        # Otomatik yükleme: kaydedilmiş excel dosyası varsa aç
        self._renkli_excel_hatirla()

        tk.Label(ay_frame, text="Dönem:", font=("Segoe UI", 9),
                 fg="#87CEEB", bg="#37474F").pack(side="left", padx=(5, 2), pady=3)

        from datetime import date
        bugun = date.today()
        aylar = []
        for i in range(6):
            ay = bugun.month - i
            yil = bugun.year
            if ay <= 0:
                ay += 12
                yil -= 1
            ay_adi = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                       "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"][ay]
            aylar.append(f"{yil} {ay_adi}")

        # Dönem seçimi — çoklu seçim destekli (checkbox'lı açılır liste).
        # self.donem_var, eski tek-dönem akışlarıyla geri uyum için tutulur ve
        # her zaman seçili dönemlerin ilkini gösterir. Çoklu dönem ihtiyacı
        # olan yerler self.donem_widget.get_secilenler() kullanır.
        self.donem_var = tk.StringVar(value=aylar[0])
        self.donem_widget = MultiSelectDonemDropdown(
            ay_frame, aylar, varsayilan=aylar[0],
            on_change=self._donemler_degisti, width=18,
        )
        self.donem_widget.pack(side="left", padx=3, pady=3)

        # Ayırıcı
        tk.Frame(icerik_frame, bg="#3D5A80", height=2).pack(fill="x", pady=3)

        # Tümünü Kontrol Et / Durdur toggle butonu
        tumunu_btn = tk.Button(
            icerik_frame,
            text="TÜMÜNÜ KONTROL ET",
            font=("Segoe UI", 14, "bold"),
            fg="white", bg="#00695C", activebackground="#004D40",
            bd=0, padx=20, pady=12, cursor="hand2",
            command=self._tumunu_toggle
        )
        tumunu_btn.pack(fill="x", pady=(0, 8))
        self.tumunu_btn = tumunu_btn

        # Duraklat / Devam Et toggle butonu
        self.duraklatildi = False
        self.duraklat_btn = tk.Button(
            icerik_frame,
            text="⏸ DURAKLAT",
            font=("Segoe UI", 11, "bold"),
            fg="white", bg="#F57C00", activebackground="#E65100",
            bd=0, padx=15, pady=8, cursor="hand2",
            state="disabled",
            command=self._duraklat_toggle
        )
        self.duraklat_btn.pack(fill="x", pady=(0, 8))

        # Ayırıcı
        tk.Frame(icerik_frame, bg="#3D5A80", height=2).pack(fill="x", pady=3)

        # Grup butonları
        for grup in GRUP_TANIMLARI:
            self._grup_satiri_olustur(icerik_frame, grup)

        # Ayırıcı
        tk.Frame(icerik_frame, bg="#3D5A80", height=2).pack(fill="x", pady=3)

        # Seçenekler satırı
        secenek_frame = tk.Frame(icerik_frame, bg="#1E3A5F")
        secenek_frame.pack(fill="x", pady=3)

        self.bastan_var = tk.BooleanVar(value=False)
        bastan_cb = tk.Checkbutton(
            secenek_frame, text="Baştan başla",
            variable=self.bastan_var,
            fg="#87CEEB", bg="#1E3A5F", selectcolor="#2C4A6E",
            activebackground="#1E3A5F", activeforeground="#87CEEB",
            font=("Segoe UI", 10)
        )
        bastan_cb.pack(side="left")

        self.kontrol_edilmisleri_atla_var = tk.BooleanVar(value=False)
        atla_cb = tk.Checkbutton(
            secenek_frame, text="Kontrol edilmişleri atla",
            variable=self.kontrol_edilmisleri_atla_var,
            fg="#87CEEB", bg="#1E3A5F", selectcolor="#2C4A6E",
            activebackground="#1E3A5F", activeforeground="#87CEEB",
            font=("Segoe UI", 10)
        )
        atla_cb.pack(side="left", padx=(10, 0))

        temizle_btn = tk.Button(
            secenek_frame, text="Hafızayı Temizle",
            font=("Segoe UI", 9), fg="white", bg="#B71C1C",
            activebackground="#7F0000", bd=0, padx=10, pady=3,
            command=self._hafizayi_temizle
        )
        temizle_btn.pack(side="right")

        rapor_btn = tk.Button(
            secenek_frame, text="Raporu Göster",
            font=("Segoe UI", 9, "bold"), fg="white", bg="#1565C0",
            activebackground="#0D47A1", bd=0, padx=10, pady=3,
            command=self._rapor_tablosu_goster
        )
        rapor_btn.pack(side="right", padx=(0, 5))

        ayar_btn = tk.Button(
            secenek_frame, text="Kontrol Ayarları",
            font=("Segoe UI", 9), fg="white", bg="#6A1B9A",
            activebackground="#4A148C", bd=0, padx=10, pady=3,
            command=self._kontrol_ayarlari_penceresi
        )
        ayar_btn.pack(side="right", padx=(0, 5))

        acik_btn = tk.Button(
            secenek_frame, text="Açık reçeteden başla",
            font=("Segoe UI", 9), fg="white", bg="#00695C",
            activebackground="#004D40", bd=0, padx=10, pady=3,
            command=self._acik_receteden_basla
        )
        acik_btn.pack(side="right", padx=(0, 5))

        # === LOG ALANI (Notebook: Sistem Logu + Kontrol Sonuçları) ===
        log_frame = tk.Frame(icerik_frame, bg="#0D2137", bd=1, relief="sunken")
        log_frame.pack(fill="both", expand=True, pady=(8, 0))

        notebook = ttk.Notebook(log_frame)
        notebook.pack(fill="both", expand=True, padx=2, pady=2)

        # --- Sekme 1: Kontrol Sonuçları (kullanıcının asıl baktığı) ---
        kontrol_tab = tk.Frame(notebook, bg="#0D2137")
        notebook.add(kontrol_tab, text="Kontrol Sonuçları")

        self.kontrol_text = tk.Text(
            kontrol_tab, font=("Consolas", 10), fg="#E0E0E0", bg="#0D2137",
            insertbackground="white", wrap="word", height=10,
            state="disabled", bd=0
        )
        kontrol_sb = tk.Scrollbar(kontrol_tab, command=self.kontrol_text.yview)
        self.kontrol_text.configure(yscrollcommand=kontrol_sb.set)
        kontrol_sb.pack(side="right", fill="y")
        self.kontrol_text.pack(fill="both", expand=True, padx=5, pady=2)
        self.kontrol_text.bind("<Control-c>", self._kontrol_kopyala)
        self.kontrol_text.bind("<Control-a>",
                                lambda e: (self.kontrol_text.tag_add("sel", "1.0", "end"), "break")[1])
        # Sağ-tık → Reçete/Rapor detay penceresi (log satırından recete_no parse edilir)
        self.kontrol_text.bind("<Button-3>", self._kontrol_text_sag_tik)

        # --- Sekme 2: Sistem Logu (akış / debug) ---
        log_tab = tk.Frame(notebook, bg="#0D2137")
        notebook.add(log_tab, text="Sistem Logu")

        self.log_text = tk.Text(
            log_tab, font=("Consolas", 9), fg="#E0E0E0", bg="#0D2137",
            insertbackground="white", wrap="word", height=10,
            state="disabled", bd=0
        )
        scrollbar = tk.Scrollbar(log_tab, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True, padx=5, pady=2)
        self.log_text.bind("<Control-c>", self._log_kopyala)
        self.log_text.bind("<Control-a>",
                            lambda e: (self.log_text.tag_add("sel", "1.0", "end"), "break")[1])
        self.log_text.bind("<Button-3>", self._kontrol_text_sag_tik)

        # Tag'ler - her iki widget'a aynı tag'leri ekle
        for w in (self.log_text, self.kontrol_text):
            w.tag_configure("info", foreground="#87CEEB")           # Nötr - açık mavi
            w.tag_configure("success", foreground="#0D2137", background="#A5D6A7")  # Uygun - yeşil
            w.tag_configure("warning", foreground="#0D2137", background="#FFE082")  # Sarı
            w.tag_configure("error", foreground="#FFFFFF", background="#E53935")    # Kırmızı
            w.tag_configure("sorun", foreground="#FFFFFF", background="#D84315")    # Koyu turuncu
            w.tag_configure("header", foreground="#FFFFFF", font=("Consolas", 10, "bold"))
            w.tag_configure("neutral", foreground="#B0BEC5")        # Gri
            w.tag_configure("ilac", foreground="#FFFFFF", font=("Consolas", 9, "bold"))

    def _grup_satiri_olustur(self, parent, grup):
        """Tek bir grup için buton + durum label satırı"""
        satir = tk.Frame(parent, bg="#1E3A5F")
        satir.pack(fill="x", pady=3)

        kod = grup["kod"]
        durum = self.grup_durumlari.get(kod, {})
        son_recete = durum.get("son_recete", "")
        toplam = durum.get("toplam_kontrol", 0)

        btn = tk.Button(
            satir,
            text=f"  {grup['ad']}",
            font=("Segoe UI", 11, "bold"),
            fg="white", bg=grup["renk"], activebackground=grup["hover"],
            bd=0, padx=15, pady=8, width=22, anchor="w", cursor="hand2",
            command=lambda k=kod: self._grup_toggle(k)
        )
        btn.pack(side="left")
        self.grup_butonlari[kod] = btn
        # Orijinal rengi sakla
        btn._renk_normal = grup["renk"]
        btn._renk_hover = grup["hover"]
        btn._ad = grup["ad"]

        # Reçete combobox (kontrol edilen reçetelerin listesi)
        from tkinter import ttk
        combo_var = tk.StringVar()
        combo = ttk.Combobox(satir, textvariable=combo_var, state="readonly",
                              font=("Consolas", 9), width=28)
        combo.pack(side="left", padx=(10, 5))
        combo.set("Henüz başlanmadı" if not son_recete else f"Son: {son_recete}")
        combo.bind("<<ComboboxSelected>>", lambda e, k=kod: self._recete_secildi(k))
        self.recete_combolar[kod] = combo
        self.recete_listeleri[kod] = []

        durum_text = f"{toplam} reçete" if toplam > 0 else ""
        durum_lbl = tk.Label(
            satir, text=durum_text,
            font=("Segoe UI", 9), fg="#87CEEB", bg="#1E3A5F"
        )
        durum_lbl.pack(side="left", padx=5)
        self.durum_labels[kod] = durum_lbl

        # Süre göstergesi label'ı
        timing_lbl = tk.Label(
            satir, text="",
            font=("Consolas", 9), fg="#FFB74D", bg="#1E3A5F"
        )
        timing_lbl.pack(side="left", padx=5)
        self.timing_labels[kod] = timing_lbl

    def _durum_label_guncelle(self, kod):
        """Bir grubun durum label'ını ve combobox'ını güncelle"""
        if kod not in self.durum_labels:
            return
        durum = self.grup_durumlari.get(kod, {})
        toplam = durum.get("toplam_kontrol", 0)
        durum_text = f"{toplam} reçete" if toplam > 0 else ""
        self.durum_labels[kod].config(text=durum_text)

    def _sure_guncelle(self, son_sure, ortalama, toplam_n):
        """Aktif grubun süre göstergesini güncelle"""
        kod = self.aktif_grup
        if not kod or kod not in self.timing_labels:
            return
        son_str = f"{son_sure:.1f}s"
        ort_str = f"{ortalama:.1f}s"
        self.timing_labels[kod].config(text=f"Son: {son_str} | Ort: {ort_str} ({toplam_n} rx)")

    COMBOBOX_HAFIZA = os.path.join(PROJE_DIZINI, "recete_combobox_hafiza.json")

    def _recete_combobox_hafiza_yukle(self):
        """Kaydedilmiş combobox listelerini dosyadan yükle"""
        try:
            if os.path.exists(self.COMBOBOX_HAFIZA):
                with open(self.COMBOBOX_HAFIZA, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for kod, liste in data.items():
                    if kod in self.recete_combolar:
                        self.recete_listeleri[kod] = [(s, r) for s, r in liste]
                        combo = self.recete_combolar[kod]
                        combo["values"] = [f"#{s} - {r}" for s, r in self.recete_listeleri[kod]]
                        if self.recete_listeleri[kod]:
                            combo.current(len(self.recete_listeleri[kod]) - 1)
        except Exception:
            pass

    def _recete_combobox_hafiza_kaydet(self):
        """Combobox listelerini dosyaya kaydet"""
        try:
            data = {}
            for kod, liste in self.recete_listeleri.items():
                if liste:
                    data[kod] = liste
            with open(self.COMBOBOX_HAFIZA, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
        except Exception:
            pass

    def _recete_combobox_ekle(self, kod, recete_no, sira):
        """Combobox'a yeni kontrol edilen reçeteyi ekle"""
        if kod not in self.recete_combolar:
            return
        combo = self.recete_combolar[kod]
        liste = self.recete_listeleri[kod]

        # Tekrar ekleme engelleme
        if any(r == recete_no for _, r in liste):
            return

        liste.append((sira, recete_no))
        combo["values"] = [f"#{s} - {r}" for s, r in liste]
        combo.current(len(liste) - 1)
        # Her ekleme sonrası hafızaya kaydet
        self._recete_combobox_hafiza_kaydet()

    def _recete_combolar_temizle(self, kod=None):
        """Combobox'ları temizle (yeni tarama başlangıcı)"""
        kodlar = [kod] if kod else list(self.recete_combolar.keys())
        for k in kodlar:
            if k in self.recete_combolar:
                self.recete_combolar[k].set("")
                self.recete_combolar[k]["values"] = []
                self.recete_listeleri[k] = []
        self._recete_combobox_hafiza_kaydet()

    def _recete_secildi(self, kod):
        """Combobox'tan bir reçete seçildiğinde - goto dosyası ile subprocess'e bildir"""
        if kod not in self.recete_combolar:
            return
        combo = self.recete_combolar[kod]
        secim = combo.get()
        if not secim or "#" not in secim:
            return
        # "#25 - 3LGXYZ" formatını parse et
        try:
            recete_no = secim.split(" - ", 1)[1].strip()
        except (IndexError, ValueError):
            return

        if not self.kontrol_aktif:
            self.log_yaz(f"Tarama aktif değil. Önce {kod} grubu taramasını başlatın.", "warning")
            return

        # Goto dosyası ile subprocess'e bildir
        goto_file = os.path.join(PROJE_DIZINI, "tarama_goto.flag")
        try:
            with open(goto_file, "w") as f:
                f.write(recete_no)
            self.log_yaz(f"Reçete {recete_no}'ya geri dönülüyor...", "info")
        except Exception as e:
            self.log_yaz(f"Goto dosyası yazılamadı: {e}", "error")

    # === MEDULA BAĞLANTI ===

    def _medula_durumu_kontrol(self):
        """
        Başlangıçta Medula durumunu kontrol et ve otomatik bağlan.
        Bağlandıktan sonra oturum aktif mi kontrol et.
        Düşmüşse taskkill + yeniden başlat.
        """
        self.log_yaz("Medula kontrol ediliyor...", "info")
        self.medula_durum_label.config(text="Durum: Kontrol ediliyor...", fg="#FF9800")
        self.root.update()

        def otomatik_baglan():
            with _com_thread():
                # CRASH KORUMA: medula_baglan COM tree walk yapan fonksiyonları
                # tetikleyebilir; IE-embed pencerede Windows fatal exception
                # (E_INTERFACE_NOT_REGISTERED) çökertme riski var. Tüm akışı
                # try/except ile sar — Python COMError'ı yakalar, fatal native
                # crash'leri yakalayamaz ama child_window'a geçtikten sonra
                # bu risk minimize.
                basarili = False
                try:
                    basarili = self.medula.medula_ac_ve_baglan()
                except Exception as e:
                    try:
                        self.root.after(0, lambda em=str(e): self.log_yaz(
                            f"Otomatik bağlanma hatası: {em[:120]}", "error"))
                    except Exception:
                        pass
                self.root.after(0, lambda b=basarili: self._medula_baglanti_sonuc(b))

        threading.Thread(target=otomatik_baglan, daemon=True).start()

    def _medula_baglan_tikla(self):
        """Medula bağlan butonuna tıklandı. Basit akış:
        1. medula_baglan() dene (BotanikEOS açıksa bağlan + oturum aç)
        2. Başarısızsa taskkill + exe'den yeniden başlat
        """
        self.log_yaz("Medula'ya bağlanılıyor...", "info")
        self.medula_durum_label.config(text="Durum: İşlem yapılıyor...", fg="#FF9800")
        self.root.update()

        def islem():
            import subprocess
            with _com_thread():
                # 1. Doğrudan bağlanmayı dene
                if self.medula.medula_baglan():
                    self.root.after(0, lambda: self._medula_baglanti_sonuc(True))
                    return

                # 2. Başarısız - taskkill + yeniden başlat
                self.root.after(0, lambda: self.log_yaz("Bağlantı başarısız, yeniden başlatılıyor...", "warning"))
                self.medula.keepalive_durdur()
                try:
                    subprocess.run(["taskkill", "/F", "/IM", "BotanikMedula.exe"],
                                   capture_output=True, text=True, timeout=10)
                except:
                    pass
                self.medula.bagli = False
                self.medula.main_window = None
                self.medula.medula_hwnd = None
                time.sleep(3)

                basarili = self.medula._exe_baslat_ve_giris_yap()
                self.root.after(0, lambda: self._medula_baglanti_sonuc(basarili))

        threading.Thread(target=islem, daemon=True).start()

    def _medula_baglanti_sonuc(self, basarili):
        """Medula bağlantı sonucunu GUI'ye yansıt + bağlanınca yerleşim uygula
        + 'Bağlantıyı Açık Tut' işaretliyse keepalive servisini başlat."""
        if basarili:
            self.medula_durum_label.config(text="Durum: Bağlı", fg="#4CAF50")
            self.medula_btn.config(text="MEDULA Bağlı", bg="#2E7D32")
            self.log_yaz("Medula'ya bağlantı kuruldu!", "success")
            # Bağlanır bağlanmaz yerleşim planını uygula (Medula sol, Hasta Takip sağ)
            try:
                self._pencereleri_konumlandir()
            except Exception as e:
                logger.debug(f"Bağlantı sonrası yerleşim hatası: {e}")
            # Bağlantıyı Açık Tut işaretliyse servisi başlat (kayıtlı tip ile)
            try:
                if getattr(self, "canli_tut_var", None) and self.canli_tut_var.get():
                    tip = self._canli_tut_tip_yukle()
                    self._canli_tut_uygula(True, tip)
            except Exception:
                pass
        else:
            self.medula_durum_label.config(text="Durum: Bağlantı başarısız", fg="#F44336")
            self.medula_btn.config(text="MEDULA Bağlan", bg="#D32F2F")
            self.log_yaz("Medula bağlantısı kurulamadı! Butona basarak tekrar deneyin.", "error")

    # === BAĞLANTIYI AÇIK TUT (keepalive servisi) ===

    def _canli_tut_tercihi_yukle(self):
        """DURUM_DOSYASI'ndan 'canli_tut' tercihini yükle (varsayılan False)."""
        try:
            if os.path.exists(DURUM_DOSYASI):
                with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                    return bool(json.load(f).get("_canli_tut", False))
        except Exception:
            pass
        return False

    def _canli_tut_tercihi_kaydet(self, deger: bool):
        """'canli_tut' tercihini DURUM_DOSYASI'na yaz."""
        try:
            durum = {}
            if os.path.exists(DURUM_DOSYASI):
                with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                    durum = json.load(f)
            durum["_canli_tut"] = bool(deger)
            with open(DURUM_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(durum, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def _canli_tut_tip_yukle(self) -> str:
        """DURUM_DOSYASI'ndan 'canli_tut_tip' tercihini yükle (varsayılan 'A')."""
        try:
            if os.path.exists(DURUM_DOSYASI):
                with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                    return json.load(f).get("_canli_tut_tip", "A")
        except Exception:
            pass
        return "A"

    def _canli_tut_tip_kaydet(self, tip: str):
        """'canli_tut_tip' tercihini DURUM_DOSYASI'na yaz."""
        try:
            durum = {}
            if os.path.exists(DURUM_DOSYASI):
                with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                    durum = json.load(f)
            durum["_canli_tut_tip"] = tip
            with open(DURUM_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(durum, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def _canli_tut_uygula(self, basla: bool, tip: str = "A"):
        """Keepalive servisini başlat veya durdur."""
        try:
            from medula_oturum_canli import (
                get_servis, CANLI_TUT_TIP_A, CANLI_TUT_TIP_B,
                CANLI_TUT_TIP_C, CANLI_TUT_TIP_D,
            )
            _TIP_MAP = {
                "A": CANLI_TUT_TIP_A,
                "B": CANLI_TUT_TIP_B,
                "C": CANLI_TUT_TIP_C,
                "D": CANLI_TUT_TIP_D,
            }
            _TIP_AD = {
                "A": "F5+Yenile",
                "B": "Sonraki Butonu",
                "C": "E-Reçete Sorgula",
                "D": "Giriş Butonu",
            }
            servis = get_servis()
            if basla:
                servis.tip = _TIP_MAP.get(tip, CANLI_TUT_TIP_A)
                if not servis.aktif_mi():
                    if servis.basla():
                        tip_ad = _TIP_AD.get(tip, tip)
                        self.log_yaz(f"✓ Bağlantıyı açık tutma başladı (119s, Tip {tip}: {tip_ad})", "success")
                        self.root.after(1000, self._canli_tut_sayac_guncelle)
                    else:
                        self.log_yaz("Bağlantıyı açık tutma başlatılamadı (pynput?)", "error")
                else:
                    servis.tip = _TIP_MAP.get(tip, CANLI_TUT_TIP_A)
            else:
                if servis.aktif_mi():
                    servis.dur()
                    self.log_yaz("Bağlantıyı açık tutma durduruldu", "info")
                if hasattr(self, "lbl_canli_tut_sayac"):
                    self.lbl_canli_tut_sayac.config(text="")
        except Exception as e:
            self.log_yaz(f"Bağlantıyı açık tutma hatası: {e}", "error")

    def _canli_tut_sayac_guncelle(self):
        """Her saniye geri sayım labelını güncelle."""
        try:
            from medula_oturum_canli import get_servis, IDLE_ESIK_SN
            servis = get_servis()
            if not servis.aktif_mi() or not self.canli_tut_var.get():
                if hasattr(self, "lbl_canli_tut_sayac"):
                    self.lbl_canli_tut_sayac.config(text="")
                return
            kalan = max(0, int(IDLE_ESIK_SN - servis.idle_saniye()))
            renk = "#EF9A9A" if kalan <= 15 else ("#FFCC80" if kalan <= 30 else "#80CBC4")
            tip = getattr(servis, "tip", "A")
            etiket = f"[{tip}] ⏱{kalan}s"
            if hasattr(self, "lbl_canli_tut_sayac"):
                self.lbl_canli_tut_sayac.config(text=etiket, fg=renk)
            self.root.after(1000, self._canli_tut_sayac_guncelle)
        except Exception:
            pass

    def _canli_tut_ayarlar_popup(self, onceki_tip: str) -> str:
        """A/B tip seçim popup'ı. Seçilen tipi döndürür (iptal edilirse onceki_tip)."""
        popup = tk.Toplevel(self.root)
        popup.title("Canlı Tut Ayarları")
        popup.resizable(False, False)
        popup.grab_set()
        popup.configure(bg="#263238")

        # Pencereyi ortala
        popup.update_idletasks()
        pw, ph = 380, 220
        rx = self.root.winfo_x() + (self.root.winfo_width() - pw) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - ph) // 2
        popup.geometry(f"{pw}x{ph}+{rx}+{ry}")

        tk.Label(
            popup, text="Canlı Tutma Yöntemi", bg="#263238", fg="#90CAF9",
            font=("Segoe UI", 10, "bold"),
        ).pack(pady=(14, 6))

        secim_var = tk.StringVar(value=onceki_tip)

        for tip, aciklama in [
            ("A", "A  —  F5 (Yenile) + Yeniden Dene kapat + Enter"),
            ("B", "B  —  \"Sonra >\" butonuna bas"),
            ("C", "C  —  E-Reçete Sorgula sayfasına tıkla"),
            ("D", "D  —  Giriş butonuna tıkla (oturum süresi dolunca)"),
        ]:
            tk.Radiobutton(
                popup, text=aciklama, variable=secim_var, value=tip,
                bg="#263238", fg="#CFD8DC", selectcolor="#0D2137",
                activebackground="#263238", activeforeground="#90CAF9",
                font=("Segoe UI", 9),
            ).pack(anchor="w", padx=24, pady=2)

        sonuc = {"tip": onceki_tip}

        def _tamam():
            sonuc["tip"] = secim_var.get()
            popup.destroy()

        def _iptal():
            popup.destroy()

        btn_fr = tk.Frame(popup, bg="#263238")
        btn_fr.pack(pady=(12, 0))
        tk.Button(
            btn_fr, text="Tamam", command=_tamam,
            bg="#0277BD", fg="white", font=("Segoe UI", 9, "bold"),
            relief="flat", padx=12,
        ).pack(side="left", padx=6)
        tk.Button(
            btn_fr, text="İptal", command=_iptal,
            bg="#455A64", fg="white", font=("Segoe UI", 9),
            relief="flat", padx=12,
        ).pack(side="left", padx=6)

        popup.wait_window()
        return sonuc["tip"]

    def _canli_tut_toggle(self):
        """Checkbox tıklandı: ayarlar popup göster, tercihi kaydet + servisi başlat/durdur."""
        deger = bool(self.canli_tut_var.get())
        # Her durumda A/B seçim popup'ı göster
        onceki_tip = self._canli_tut_tip_yukle()
        secilen_tip = self._canli_tut_ayarlar_popup(onceki_tip)
        self._canli_tut_tip_kaydet(secilen_tip)
        self._canli_tut_tercihi_kaydet(deger)
        self._canli_tut_uygula(deger, secilen_tip)

    # === RENKLİ REÇETE EXCEL YÜKLEME ===

    def _renkli_recete_yukle(self):
        """Renkli reçete Excel dosyası yükle (Gözat dialog)"""
        from tkinter import filedialog
        dosya = filedialog.askopenfilename(
            title="Renkli Reçete Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")],
            initialdir=os.path.expanduser("~/OneDrive/Desktop")
        )
        if dosya:
            self._renkli_excel_isle(dosya)

    def _renkli_excel_isle(self, dosya_yolu):
        """Excel dosyasını oku ve renkli reçete listesini oluştur.

        Beklenen format (esnek):
          - 1. sütun (A): reçete no(ları), "ABC/DEF" gibi /'li olabilir
          - 2. sütun (B): hasta (opsiyonel)
          - 3. sütun (C): hekim (opsiyonel)
          - 5. sütun (E): tarih (opsiyonel)
        İlk satır başlık olarak atlanır. Eksik sütunlar boş kabul edilir.
        """
        if not os.path.exists(dosya_yolu):
            self.log_yaz(f"Dosya bulunamadı: {dosya_yolu}", "error")
            self.renkli_durum_label.config(text="Dosya yok!", fg="#F44336")
            return

        try:
            import openpyxl
        except ImportError:
            self.log_yaz("openpyxl yüklü değil! 'pip install openpyxl' çalıştırın.", "error")
            self.renkli_durum_label.config(text="openpyxl yok!", fg="#F44336")
            return

        try:
            wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
        except Exception as e:
            import traceback
            self.log_yaz(f"Excel açılamadı ({type(e).__name__}): {e}", "error")
            logger.error(f"openpyxl load_workbook hatası:\n{traceback.format_exc()}")
            self.renkli_durum_label.config(text="Excel açılamadı!", fg="#F44336")
            return

        try:
            ws = wb.active
            if ws is None:
                ws = wb.worksheets[0] if wb.worksheets else None
            if ws is None:
                self.log_yaz("Excel'de aktif sayfa yok!", "error")
                self.renkli_durum_label.config(text="Sayfa yok!", fg="#F44336")
                wb.close()
                return

            self.renkli_recete_listesi = []
            atlanan = 0
            for satir_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
                # İlk satır başlık - atla
                if satir_no == 1:
                    continue
                # Boş satır
                if not row or row[0] in (None, ""):
                    continue
                try:
                    recete_no_raw = str(row[0]).strip()
                    if not recete_no_raw:
                        continue
                    # "ZTHPW2PU / 2MBOFVQ" formatı - her iki no'yu da kaydet
                    recete_nolar = [n.strip() for n in recete_no_raw.split("/") if n.strip()]
                    hasta = str(row[1] or "").strip() if len(row) > 1 else ""
                    hekim = str(row[2] or "").strip() if len(row) > 2 else ""
                    tarih_raw = row[4] if len(row) > 4 else ""
                    # tarih datetime objesi olabilir; isoformat'a normalize et
                    if hasattr(tarih_raw, "isoformat"):
                        tarih = tarih_raw.isoformat()
                    else:
                        tarih = str(tarih_raw or "").strip()

                    self.renkli_recete_listesi.append({
                        "recete_nolar": recete_nolar,
                        "recete_no_raw": recete_no_raw,
                        "hasta": hasta,
                        "hekim": hekim,
                        "tarih": tarih,
                    })
                except Exception as satir_e:
                    atlanan += 1
                    logger.debug(f"Satır {satir_no} atlandı: {satir_e}")
                    continue
            wb.close()

            self.renkli_recete_dosya = dosya_yolu
            sayi = len(self.renkli_recete_listesi)
            dosya_adi = os.path.basename(dosya_yolu)

            if sayi == 0:
                self.log_yaz(
                    f"Excel okundu ama hiç reçete bulunamadı! "
                    f"İlk sütunda reçete no var mı kontrol edin. ({dosya_adi})",
                    "warning")
                self.renkli_durum_label.config(text="Boş liste!", fg="#FF9800")
                return

            # En son tarihi hesapla ve label'a yaz
            max_tarih_str = self._renkli_max_tarih_hesapla()
            self.renkli_durum_label.config(
                text=f"✓ {dosya_adi} ({sayi} reçete)", fg="#4CAF50"
            )
            if max_tarih_str:
                self.renkli_tarih_label.config(text=f"📅 {max_tarih_str} tarihine kadar")
            else:
                self.renkli_tarih_label.config(text="")
            self.renkli_btn.config(bg="#4A148C")
            ek = f", {atlanan} satır atlandı" if atlanan else ""
            self.log_yaz(f"Renkli reçete listesi yüklendi: {sayi} reçete ({dosya_adi}){ek}", "success")

            # Reçete numaralarını JSON'a kaydet (subprocess okuyabilsin)
            self._renkli_json_guncelle()

            # Tüm veriyi cache JSON'a yaz (yeniden başlatınca Excel okumadan yüklenebilsin)
            self._renkli_cache_kaydet(dosya_yolu, max_tarih_str)

            # Dosya yolunu hatırla (checkbox işaretliyse JSON'a kaydet)
            if self.renkli_hatirla_var.get():
                self._renkli_excel_kaydet(dosya_yolu)

        except Exception as e:
            import traceback
            self.log_yaz(f"Renkli reçete işleme hatası ({type(e).__name__}): {e}", "error")
            logger.error(f"Excel parse hatası:\n{traceback.format_exc()}")
            self.renkli_durum_label.config(text="İşleme hatası!", fg="#F44336")

    def _renkli_json_guncelle(self):
        """Renkli reçete numaralarını renkli_recete_listesi.json'a kaydet.
        Subprocess (recete_tarama.py) bu dosyayı okur."""
        try:
            tum_nolar = set()
            for kayit in self.renkli_recete_listesi:
                for no in kayit.get("recete_nolar", []):
                    if no and len(no) >= 5:
                        tum_nolar.add(no)
            renkli_json = os.path.join(PROJE_DIZINI, "renkli_recete_listesi.json")
            data = {
                "receteler": list(tum_nolar),
                "yukleme_tarihi": datetime.now().isoformat(),
                "sayi": len(tum_nolar),
            }
            with open(renkli_json, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            self.log_yaz(f"Renkli reçete JSON güncellendi: {len(tum_nolar)} reçete no", "info")
        except Exception as e:
            self.log_yaz(f"Renkli JSON güncelleme hatası: {e}", "error")

    def _renkli_excel_kaydet(self, dosya_yolu):
        """Renkli reçete excel dosya yolunu JSON'a kaydet"""
        try:
            durum = {}
            if os.path.exists(DURUM_DOSYASI):
                with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                    durum = json.load(f)
            durum["_renkli_excel_dosya"] = dosya_yolu
            with open(DURUM_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(durum, f, indent=2, ensure_ascii=False)
        except:
            pass

    def _renkli_excel_hatirla(self):
        """Önce cache JSON'dan yükle. Cache yoksa veya güncel değilse Excel'i oku.

        Cache geçerlilik kuralı: cache'teki excel dosyası mevcut ve mtime aynı ise
        cache kullanılır (Excel hiç açılmaz). Aksi halde Excel yeniden okunur.
        """
        try:
            # 1) Cache'i dene
            if self._renkli_cache_yukle():
                return

            # 2) Cache yok / geçersiz - hatırlanan Excel yolunu dene
            if not os.path.exists(DURUM_DOSYASI):
                return
            with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                durum = json.load(f)
            dosya = durum.get("_renkli_excel_dosya", "")
            if not dosya:
                return
            if os.path.exists(dosya):
                self._renkli_excel_isle(dosya)
            else:
                # Kayıtlı yol artık geçersiz - temizle
                logger.info(f"Hatırlanan renkli excel yolu geçersiz, temizleniyor: {dosya}")
                durum.pop("_renkli_excel_dosya", None)
                try:
                    with open(DURUM_DOSYASI, "w", encoding="utf-8") as f:
                        json.dump(durum, f, indent=2, ensure_ascii=False)
                except Exception:
                    pass
        except Exception as e:
            logger.debug(f"Renkli excel hatırlama hatası: {e}")

    def _renkli_max_tarih_hesapla(self):
        """renkli_recete_listesi içindeki en geç tarihi 'dd.mm.yyyy' formatında döndür."""
        from datetime import datetime as _dt
        if not self.renkli_recete_listesi:
            return ""
        en_son = None
        formatlar = (
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d.%m.%Y",
            "%d-%m-%Y",
            "%m/%d/%Y",
        )
        for kayit in self.renkli_recete_listesi:
            t = (kayit.get("tarih") or "").strip()
            if not t:
                continue
            parsed = None
            for fmt in formatlar:
                try:
                    parsed = _dt.strptime(t[:len(fmt) + 4], fmt)
                    break
                except ValueError:
                    pass
            if parsed and (en_son is None or parsed > en_son):
                en_son = parsed
        if not en_son:
            return ""
        return en_son.strftime("%d.%m.%Y")

    def _renkli_cache_kaydet(self, dosya_yolu, max_tarih_str):
        """Tüm renkli reçete listesini cache JSON'a yaz.

        Bir sonraki açılışta Excel okunmadan bu dosyadan yüklenecek.
        """
        try:
            try:
                mtime = os.path.getmtime(dosya_yolu) if dosya_yolu and os.path.exists(dosya_yolu) else 0
            except Exception:
                mtime = 0
            cache_data = {
                "kaynak_dosya": dosya_yolu or "",
                "kaynak_mtime": mtime,
                "yukleme_tarihi": datetime.now().isoformat(),
                "max_tarih": max_tarih_str or "",
                "kayitlar": self.renkli_recete_listesi,
            }
            with open(RENKLI_CACHE_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(cache_data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.debug(f"Renkli cache kaydetme hatası: {e}")

    def _renkli_cache_yukle(self):
        """Cache JSON'dan renkli reçete listesini yükle. Başarılı ise True döner."""
        try:
            if not os.path.exists(RENKLI_CACHE_DOSYASI):
                return False
            with open(RENKLI_CACHE_DOSYASI, "r", encoding="utf-8") as f:
                cache = json.load(f)
            kayitlar = cache.get("kayitlar", [])
            if not kayitlar:
                return False

            # Cache'teki excel dosyası varsa ve mtime değişmişse cache geçersiz
            kaynak = cache.get("kaynak_dosya", "")
            kaynak_mtime = cache.get("kaynak_mtime", 0)
            if kaynak and os.path.exists(kaynak):
                try:
                    if abs(os.path.getmtime(kaynak) - kaynak_mtime) > 1:
                        # Excel değişmiş - cache kullanma, yeniden okut
                        return False
                except Exception:
                    pass

            self.renkli_recete_listesi = kayitlar
            self.renkli_recete_dosya = kaynak
            sayi = len(kayitlar)
            max_tarih_str = cache.get("max_tarih", "") or self._renkli_max_tarih_hesapla()
            dosya_adi = os.path.basename(kaynak) if kaynak else "cache"
            self.renkli_durum_label.config(
                text=f"✓ {dosya_adi} ({sayi} reçete)", fg="#4CAF50"
            )
            if max_tarih_str:
                self.renkli_tarih_label.config(text=f"📅 {max_tarih_str} tarihine kadar")
            else:
                self.renkli_tarih_label.config(text="")
            self.renkli_btn.config(bg="#4A148C")
            self.log_yaz(
                f"Renkli reçete listesi cache'ten yüklendi: {sayi} reçete (Excel okunmadı)",
                "info",
            )
            # JSON özet de güncel olsun
            self._renkli_json_guncelle()
            return True
        except Exception as e:
            logger.debug(f"Renkli cache yükleme hatası: {e}")
            return False

    def _mesaj_yoksay_yukle(self):
        """İlaç mesajı yoksay listesini JSON'dan yükle."""
        try:
            if not os.path.exists(MESAJ_YOKSAY_DOSYASI):
                self.mesaj_yoksay_patternleri = []
                return
            with open(MESAJ_YOKSAY_DOSYASI, "r", encoding="utf-8") as f:
                data = json.load(f)
            patternler = data.get("ilac_adi_patternleri", [])
            self.mesaj_yoksay_patternleri = [
                p.strip().upper() for p in patternler if p and p.strip()
            ]
        except Exception as e:
            logger.debug(f"Mesaj yoksay listesi yüklenemedi: {e}")
            self.mesaj_yoksay_patternleri = []

    def mesaj_yoksay_eslesti(self, ilac_adi):
        """İlaç adı yoksay listesindeki bir patternle eşleşiyor mu?"""
        if not ilac_adi or not self.mesaj_yoksay_patternleri:
            return False
        ilac_upper = ilac_adi.upper()
        return any(p in ilac_upper for p in self.mesaj_yoksay_patternleri)

    def renkli_recete_kontrol(self, recete_no):
        """
        Bir reçete no'sunun renkli reçete listesinde olup olmadığını kontrol et.
        Returns: (bulundu: bool, kayit: dict or None)
        """
        if not self.renkli_recete_listesi:
            return False, None

        for kayit in self.renkli_recete_listesi:
            for no in kayit["recete_nolar"]:
                if no == recete_no:
                    return True, kayit
        return False, None

    # === LOG ===

    def _log_kopyala(self, event=None):
        """Disabled text widget'tan seçili metni kopyala"""
        try:
            secili = self.log_text.get("sel.first", "sel.last")
            self.root.clipboard_clear()
            self.root.clipboard_append(secili)
        except tk.TclError:
            pass
        return "break"

    def log_yaz(self, mesaj, tag="info"):
        """Sistem Logu sekmesine mesaj yaz (akış / debug)"""
        if not self.log_text:
            return
        self.log_text.config(state="normal")
        zaman = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{zaman}] {mesaj}\n", tag)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _kontrol_kopyala(self, event=None):
        """Kontrol Sonuçları widget'ından seçili metni kopyala"""
        try:
            secili = self.kontrol_text.get("sel.first", "sel.last")
            self.root.clipboard_clear()
            self.root.clipboard_append(secili)
        except tk.TclError:
            pass
        return "break"

    def kontrol_yaz(self, mesaj, tag="info", zaman=False):
        """Kontrol Sonuçları sekmesine kullanıcıya yönelik satır yaz"""
        if not getattr(self, "kontrol_text", None):
            return
        self.kontrol_text.config(state="normal")
        if zaman:
            t = datetime.now().strftime("%H:%M:%S")
            self.kontrol_text.insert("end", f"[{t}] {mesaj}\n", tag)
        else:
            self.kontrol_text.insert("end", f"{mesaj}\n", tag)
        self.kontrol_text.see("end")
        self.kontrol_text.config(state="disabled")

    # === KONTROL FONKSİYONLARI ===

    def _medula_hazir_mi(self):
        """Medula bağlı mı kontrol et, değilse otomatik bağlan. True dönerse hazır."""
        if self.medula.bagli:
            return True

        self.log_yaz("Medula bağlı değil, otomatik bağlanılıyor...", "warning")
        self.medula_durum_label.config(text="Durum: Bağlanılıyor...", fg="#FF9800")
        self.root.update()

        basarili = self.medula.medula_ac_ve_baglan()
        self._medula_baglanti_sonuc(basarili)
        return basarili

    def _tumunu_kontrol_et(self):
        """Tüm grupları sırayla kontrol et: C -> A -> B -> GK -> CK.

        Her grup için tek tek grup butonunun yaptığı subprocess akışını çağırır
        (recete_tarama.py — uyarı kodu, SUT, teşhis kontrolleri dahil).
        Kritik hata olursa Medula otomatik kapatılıp yeniden açılır ve aynı
        gruba kaldığı yerden (--devam) tekrar başlatılır (max 3 deneme).
        """
        if self.kontrol_aktif:
            self.log_yaz("Zaten bir kontrol devam ediyor!", "warning")
            return

        self.log_yaz("=" * 40, "header")
        self.log_yaz("TÜMÜNÜ KONTROL ET başlatılıyor...", "header")
        self.log_yaz("Sıra: C -> A -> B -> Geçici Koruma -> C Kan Ürünü", "info")
        self.log_yaz("Kritik hata: Medula otomatik yeniden açılır, kaldığı yerden devam", "info")
        self.log_yaz("=" * 40, "header")

        # TÜMÜNÜ modu flag'ı: subprocess thread'leri buton resetlemesini atlasın
        self._tumu_aktif = True
        self._tumu_iptal = False
        self.aktif_grup = "TUMU"
        self.tumunu_btn.config(text="DURDUR (Esc)", bg="#B71C1C", activebackground="#7F0000")
        for btn in self.grup_butonlari.values():
            btn.config(state="disabled")

        # Otomatik kurtarma için maksimum deneme sayısı (her grup için ayrı)
        MAX_RECOVERY = 3

        def tumu_thread():
            # Recovery sırasında medula.medula_oturum_kurtarma() pywinauto/COM kullanır.
            # Worker thread'de COM bağlamı başlatılmalı.
            if _PYTHONCOM_VAR:
                try: pythoncom.CoInitialize()
                except Exception: pass
            try:
                # Çoklu dönem: dış döngü dönemlerin üzerinden geçer; her dönem
                # için donem_var güncellenir (subprocess donem_var.get()'ten
                # offset hesapladığı için bu yeterli). _grup_dongusu() bir
                # dönem içinde tüm grupları işler — hard-stop durumunda False
                # döner ve dış döngü kırılır.
                def _grup_dongusu():
                    for index, grup in enumerate(GRUP_TANIMLARI):
                        if self._tumu_iptal:
                            return False
                        # _tumu_grup_index: ilk grup (0) için --devam, sonrakiler için --bastan
                        self._tumu_grup_index = index
                        grup_kodu = grup["kod"]
                        grup_adi = grup.get("ad", grup_kodu)

                        # Bu grup için recovery döngüsü.
                        # Kritik hata gelirse: Medula taskkill + yeniden aç + login,
                        # ardından kaldığı reçeteden (--devam) tekrar başla.
                        self._tumu_recovery_devam = False  # ilk denemede normal akış
                        recovery_sayaci = 0
                        grup_basarili = False

                        while not self._tumu_iptal and recovery_sayaci <= MAX_RECOVERY:
                            # Grup başında kritik hata flag'ini ve reçete sayacını sıfırla.
                            # Subprocess çıktısı bu değerleri set eder; deneme sonrası kontrol ederiz.
                            self._tumu_son_grup_hata = None
                            self._tumu_son_grup_recete_sayisi = 0
                            self._tumu_son_grup_bos_normal = False

                            bitti_evt = threading.Event()
                            self.root.after(
                                0,
                                lambda gk=grup_kodu, evt=bitti_evt:
                                    self._grup_kontrol_baslat(gk, bittiginde=evt.set),
                            )

                            # Subprocess'in başlama + bitişini bekle.
                            # Her 0.5 sn'de bir iptal kontrolü.
                            while not bitti_evt.is_set() and not self._tumu_iptal:
                                time.sleep(0.5)

                            if self._tumu_iptal:
                                self.root.after(0, lambda ga=grup_adi: self.log_yaz(
                                    f"TÜMÜNÜ iptal edildi ({ga} sırasında)", "warning"))
                                return False

                            hata_mesaji = getattr(self, "_tumu_son_grup_hata", None)
                            recete_sayisi = getattr(self, "_tumu_son_grup_recete_sayisi", 0)
                            bos_normal = getattr(self, "_tumu_son_grup_bos_normal", False)

                            # 1) Kritik hata varsa: Medula recovery + aynı gruba --devam ile retry
                            if hata_mesaji:
                                recovery_sayaci += 1
                                if recovery_sayaci > MAX_RECOVERY:
                                    self.root.after(0, lambda h=hata_mesaji, ga=grup_adi: self.log_yaz(
                                        f"⚠ {ga}: {MAX_RECOVERY} kurtarma denemesi başarısız ({h}) — TÜMÜNÜ durduruldu",
                                        "error"))
                                    return False
                                self.root.after(0, lambda h=hata_mesaji, ga=grup_adi, n=recovery_sayaci:
                                    self.log_yaz(
                                        f"⚠ {ga} kritik hata ({h}) — Medula yeniden açılıyor "
                                        f"(deneme {n}/{MAX_RECOVERY})", "warning"))
                                # taskkill + exe yeniden başlat + login
                                try:
                                    kurtuldu = self.medula.medula_oturum_kurtarma()
                                except Exception as ex:
                                    self.root.after(0, lambda e=ex: self.log_yaz(
                                        f"Medula kurtarma istisna: {e}", "error"))
                                    kurtuldu = False
                                if not kurtuldu:
                                    self.root.after(0, lambda: self.log_yaz(
                                        "Medula kurtarma başarısız — bekleniyor 5 sn ve tekrar denenecek",
                                        "warning"))
                                    time.sleep(5)
                                # Recovery sonrası mutlaka --devam: kaldığı reçeteden başla
                                self._tumu_recovery_devam = True
                                continue  # while döngüsünde tekrar dene

                            # 2) Boş grup (Medula "Reçete kaydı bulunamadı") → sıradaki gruba geç
                            if bos_normal and recete_sayisi == 0:
                                self.root.after(0, lambda ga=grup_adi: self.log_yaz(
                                    f"  {ga} grubu boş ('Reçete kaydı bulunamadı.') — sonraki gruba geçiliyor",
                                    "info"))
                                grup_basarili = True
                                break

                            # 3) 0 reçete + kritik hata yok + boş normal değil:
                            #    Medula açıkta ama bir şey okunmamış. Recovery dene.
                            if recete_sayisi == 0:
                                recovery_sayaci += 1
                                if recovery_sayaci > MAX_RECOVERY:
                                    self.root.after(0, lambda ga=grup_adi: self.log_yaz(
                                        f"⚠ {ga}: {MAX_RECOVERY} denemede 0 reçete işlendi — TÜMÜNÜ durduruldu",
                                        "warning"))
                                    return False
                                self.root.after(0, lambda ga=grup_adi, n=recovery_sayaci: self.log_yaz(
                                    f"⚠ {ga}: 0 reçete işlendi — Medula yeniden açılıyor "
                                    f"(deneme {n}/{MAX_RECOVERY})", "warning"))
                                try:
                                    self.medula.medula_oturum_kurtarma()
                                except Exception:
                                    pass
                                self._tumu_recovery_devam = True
                                continue

                            # 4) Başarılı: bir sonraki gruba geç
                            grup_basarili = True
                            break

                        if not grup_basarili and not self._tumu_iptal:
                            # MAX_RECOVERY aşıldı — TÜMÜNÜ'yü durdur
                            return False
                    return True

                donemler = self._secilen_donemler() or [self.donem_var.get()]
                coklu_donem = len(donemler) > 1
                if coklu_donem:
                    self.root.after(0, lambda dl=donemler: self.log_yaz(
                        f"Çoklu dönem aktif: {', '.join(dl)}", "info"))
                for donem_idx, donem in enumerate(donemler, 1):
                    if self._tumu_iptal:
                        break
                    # Her dönem başında donem_var'ı güncelle (subprocess offset
                    # bunu okur). UI thread'inden set, kısa bir bekleme ile uygulanır.
                    self.root.after(0, lambda d=donem: self.donem_var.set(d))
                    time.sleep(0.15)
                    if coklu_donem:
                        self.root.after(
                            0,
                            lambda d=donem, i=donem_idx, n=len(donemler): self.log_yaz(
                                f"════ DÖNEM {i}/{n}: {d} ════", "header"))
                    if not _grup_dongusu():
                        break
            finally:
                self._tumu_aktif = False
                self._tumu_recovery_devam = False
                self.aktif_grup = None
                self.root.after(0, self._butonlari_normal_yap)
                self.root.after(0, lambda: self.log_yaz(
                    "TÜMÜNÜ KONTROL ET tamamlandı", "header"))
                if _PYTHONCOM_VAR:
                    try: pythoncom.CoUninitialize()
                    except Exception: pass

        threading.Thread(target=tumu_thread, daemon=True).start()

    def _medula_hazirla(self):
        """Medula bağlantısını kontrol et.
        Bağlıysa → doğrudan True dön (subprocess kendi oturum kontrolünü yapar)
        Bağlı değilse → medula_baglan() ile kur.
        Returns: True=hazır, False=bağlanılamadı
        """
        # Zaten bağlıysa, mevcut bağlantıyı kullan
        # Oturum durumu subprocess (recete_tarama.py) tarafında kontrol edilir
        if self.medula.bagli:
            self.root.after(0, lambda: self._medula_baglanti_sonuc(True))
            return True

        # Bağlı değilse → medula_baglan() ile kur
        basarili = self.medula.medula_baglan()
        self.root.after(0, lambda: self._medula_baglanti_sonuc(basarili))
        if not basarili:
            self.root.after(0, lambda: self.log_yaz("Medula bağlantısı kurulamadı!", "error"))
        return basarili

    def _rapor_tablosu_goster(self):
        """Tarama raporu JSON'ını oku ve yeni pencerede tablo olarak göster"""
        rapor_json = os.path.join(PROJE_DIZINI, "tarama_rapor.json")
        if not os.path.exists(rapor_json):
            self.log_yaz("Henüz rapor yok! Önce bir tarama yapın.", "warning")
            return

        try:
            with open(rapor_json, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            self.log_yaz(f"Rapor dosyası okunamadı: {e}", "error")
            return

        satirlar = data.get("satirlar", [])
        grup = data.get("grup", "?")
        tarih = data.get("tarih", "")[:16].replace("T", " ")

        if not satirlar:
            self.log_yaz("Rapor boş!", "warning")
            return

        # Yeni pencere
        rapor_win = tk.Toplevel(self.root)
        rapor_win.title(f"Kontrol Raporu - {grup} Grubu ({tarih})")
        rapor_win.geometry("1200x600")
        rapor_win.configure(bg="#1E3A5F")

        # Üst bilgi
        ust = tk.Frame(rapor_win, bg="#1E3A5F")
        ust.pack(fill="x", padx=5, pady=5)
        tk.Label(ust, text=f"Kontrol Raporu - {grup} Grubu  |  {tarih}  |  {len(satirlar)} satır",
                 font=("Segoe UI", 12, "bold"), fg="white", bg="#1E3A5F").pack(side="left")

        # Excel'e aktar butonu
        def excel_ac():
            masaustu = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
            import glob as g
            dosyalar = sorted(g.glob(os.path.join(masaustu, f"Kontrol_Raporu_{grup}_*.xlsx")))
            if dosyalar:
                os.startfile(dosyalar[-1])
            else:
                self.log_yaz("Excel dosyası bulunamadı!", "warning")

        tk.Button(ust, text="Excel Aç", font=("Segoe UI", 10, "bold"),
                  fg="white", bg="#1565C0", activebackground="#0D47A1",
                  bd=0, padx=10, pady=3, command=excel_ac).pack(side="right")

        # Treeview tablo
        from tkinter import ttk
        sutunlar = ("recete_no", "recete_turu", "ilac_adi", "etkin_madde",
                     "rapor_kodu", "msj", "renkli_kontrol", "rapor_kontrol",
                     "sut_kontrol", "sonuc", "aciklama")
        basliklar = ("Reçete No", "Tür", "İlaç Adı", "Etkin Madde",
                     "Rapor", "Msj", "Renkli", "Rapor K.", "SUT K.", "Sonuç", "Açıklama")
        genislikler = (85, 70, 280, 170, 60, 35, 90, 90, 80, 65, 300)

        style = ttk.Style()
        style.configure("Rapor.Treeview", font=("Segoe UI", 9), rowheight=24)
        style.configure("Rapor.Treeview.Heading", font=("Segoe UI", 9, "bold"))

        tablo_frame = tk.Frame(rapor_win, bg="#1E3A5F")
        tablo_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        tree = ttk.Treeview(tablo_frame, columns=sutunlar, show="headings",
                            style="Rapor.Treeview", selectmode="browse")

        for col, baslik, gen in zip(sutunlar, basliklar, genislikler):
            tree.heading(col, text=baslik)
            tree.column(col, width=gen, minwidth=30)

        # Scrollbar
        yscroll = ttk.Scrollbar(tablo_frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(tablo_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.pack(side="right", fill="y")
        xscroll.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        # Tag renkler
        tree.tag_configure("OK", background="#C8E6C9")
        tree.tag_configure("SORUN", background="#FFCDD2", foreground="#B71C1C")
        tree.tag_configure("UYARI", background="#FFF9C4")
        tree.tag_configure("YENİ", background="#B3E5FC")
        tree.tag_configure("recete", background="#E8EAF6", font=("Segoe UI", 9, "bold"))

        # Veri ekle (iid → satir map: sağ-tık menüsünde detay penceresi için)
        satir_map = {}
        son_recete_no = ""  # Boş "recete_no" olan satırlar bir önceki reçeteye ait
        onceki_recete = None
        for satir in satirlar:
            recete_no = satir.get("recete_no", "")
            yeni_recete = recete_no != onceki_recete
            onceki_recete = recete_no
            if recete_no:
                son_recete_no = recete_no

            degerler = (
                recete_no if yeni_recete else "",
                satir.get("recete_turu", "") if yeni_recete else "",
                satir.get("ilac_adi", ""),
                satir.get("etkin_madde", ""),
                satir.get("rapor_kodu", ""),
                satir.get("msj", ""),
                satir.get("renkli_kontrol", "-"),
                satir.get("rapor_kontrol", "-"),
                satir.get("sut_kontrol", "-"),
                satir.get("sonuc", ""),
                satir.get("aciklama", ""),
            )

            sonuc = satir.get("sonuc", "")
            tag = sonuc if sonuc in ("OK", "SORUN", "UYARI") else "YENİ" if sonuc == "YENİ" else ""
            iid = tree.insert("", "end", values=degerler, tags=(tag,))
            # Map'e tam satır + çözümlenmiş recete_no kaydet
            satir_map[iid] = dict(satir)
            if not satir_map[iid].get("recete_no"):
                satir_map[iid]["recete_no"] = son_recete_no

        # Sağ-tık menüsü: Reçeteyi Göster / Raporu Göster
        tree.bind(
            "<Button-3>",
            lambda e, t=tree, w=rapor_win, sm=satir_map: self._rapor_sag_tik_menu(e, t, w, sm),
        )

        # Alt bilgi
        sorun_sayisi = sum(1 for s in satirlar if s.get("sonuc") == "SORUN")
        uyari_sayisi = sum(1 for s in satirlar if s.get("sonuc") == "UYARI")
        yeni_sayisi = sum(1 for s in satirlar if s.get("sonuc") == "YENİ")
        ok_sayisi = sum(1 for s in satirlar if s.get("sonuc") == "OK")

        alt = tk.Frame(rapor_win, bg="#1E3A5F")
        alt.pack(fill="x", padx=5, pady=3)
        tk.Label(alt, text=f"OK: {ok_sayisi}", fg="#4CAF50", bg="#1E3A5F",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=5)
        tk.Label(alt, text=f"Sorun: {sorun_sayisi}", fg="#F44336", bg="#1E3A5F",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=5)
        tk.Label(alt, text=f"Uyarı: {uyari_sayisi}", fg="#FF9800", bg="#1E3A5F",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=5)
        tk.Label(alt, text=f"Yeni: {yeni_sayisi}", fg="#03A9F4", bg="#1E3A5F",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=5)

    # ───────────────────────── Sağ-tık menüsü + detay pencereleri ─────────────────────────
    def _kontrol_text_sag_tik(self, event):
        """Kontrol Sonuçları / Sistem Logu Text widget'ında sağ-tık → menü.

        Tıklanan satırdan başlayarak yukarı doğru yürüyüp şu desenleri arar:
        - "━━━ Reçete #N: XXXXXXX ━━━"  → recete_no
        - "╔══ İLAÇ #N: <ad> ══"       → ilac_adi (en yakın)
        - "║ Rapor kodu: <kod>"         → rapor_kodu (en yakın, yok/- ise yoksay)
        """
        widget = event.widget
        try:
            idx = widget.index(f"@{event.x},{event.y}")
            line_no = int(idx.split(".")[0])
        except Exception:
            return

        recete_no = ""
        ilac_adi = ""
        rapor_kodu = ""

        recete_re = re.compile(r"Reçete\s*#\d+\s*:\s*([A-Z0-9]{5,})", re.IGNORECASE)
        ilac_re = re.compile(r"İLAÇ\s*#\d+\s*:\s*(.+?)(?:\s*══|\s*$)", re.IGNORECASE)
        rapor_re = re.compile(r"Rapor\s+kodu\s*:\s*(\S+)", re.IGNORECASE)

        # Tıklanan satırdan max 200 satır yukarı yürü; ilk Reçete #'da dur
        for ln in range(line_no, max(0, line_no - 200), -1):
            try:
                line = widget.get(f"{ln}.0", f"{ln}.end")
            except Exception:
                continue
            if not line:
                continue
            if not ilac_adi:
                m = ilac_re.search(line)
                if m:
                    ilac_adi = m.group(1).strip()
            if not rapor_kodu:
                m = rapor_re.search(line)
                if m:
                    val = m.group(1).strip()
                    if val.lower() not in ("yok", "-", "none"):
                        rapor_kodu = val
            m = recete_re.search(line)
            if m:
                recete_no = m.group(1).strip()
                break

        # Tıklanan satırdan aşağı doğru da bak (reçete bulunmamış olabilir
        # — log akışında ilk satıra denk geldiyse bir sonraki reçete başlığı altta)
        if not recete_no:
            try:
                last_line = int(widget.index("end-1c").split(".")[0])
            except Exception:
                last_line = line_no + 50
            for ln in range(line_no + 1, min(last_line + 1, line_no + 100)):
                try:
                    line = widget.get(f"{ln}.0", f"{ln}.end")
                except Exception:
                    continue
                m = recete_re.search(line)
                if m:
                    recete_no = m.group(1).strip()
                    break

        if not recete_no and not ilac_adi:
            # Hiç bağlam çıkmadı — sessizce çık
            return

        satir = {"recete_no": recete_no, "rapor_kodu": rapor_kodu,
                 "ilac_adi": ilac_adi, "etkin_madde": ""}

        m = tk.Menu(self.root, tearoff=0)
        if recete_no:
            m.add_command(
                label=f"🔍 Reçete Detayını Göster ({recete_no})",
                command=lambda r=recete_no: self._recete_detay_penceresi(self.root, r),
            )
        else:
            m.add_command(label="🔍 Reçete Detayını Göster (reçete no bulunamadı)",
                          state="disabled")
        if rapor_kodu and recete_no:
            m.add_command(
                label=f"📄 Rapor Detayını Göster ({rapor_kodu})",
                command=lambda r=recete_no, s=dict(satir):
                    self._rapor_detay_penceresi(self.root, r, s),
            )
        else:
            m.add_command(label="📄 Rapor Detayını Göster (rapor yok)",
                          state="disabled")
        try:
            m.tk_popup(event.x_root, event.y_root)
        finally:
            m.grab_release()
        return "break"

    def _rapor_sag_tik_menu(self, event, tree, parent_win, satir_map):
        """Rapor tablosu satırına sağ-tık → Reçeteyi/Raporu Göster menüsü."""
        iid = tree.identify_row(event.y)
        if not iid:
            return
        tree.selection_set(iid)
        satir = satir_map.get(iid) or {}
        recete_no = (satir.get("recete_no") or "").strip()
        rapor_kodu = (satir.get("rapor_kodu") or "").strip()

        m = tk.Menu(parent_win, tearoff=0)
        if recete_no:
            m.add_command(
                label=f"🔍 Reçete Detayını Göster ({recete_no})",
                command=lambda r=recete_no, p=parent_win: self._recete_detay_penceresi(p, r),
            )
        else:
            m.add_command(label="🔍 Reçete Detayını Göster (reçete no yok)", state="disabled")
        if rapor_kodu and recete_no:
            m.add_command(
                label=f"📄 Rapor Detayını Göster ({rapor_kodu})",
                command=lambda r=recete_no, s=dict(satir), p=parent_win:
                    self._rapor_detay_penceresi(p, r, s),
            )
        else:
            m.add_command(label="📄 Rapor Detayını Göster (rapor yok)", state="disabled")
        try:
            m.tk_popup(event.x_root, event.y_root)
        finally:
            m.grab_release()

    def _detay_db_baglan(self):
        """Botanik EOS DB bağlantısı (salt-okunur). Hata durumunda mesaj göster, None dön."""
        try:
            from botanik_db import BotanikDB
        except Exception as e:
            messagebox.showerror("Hata", f"Botanik DB modülü yüklenemedi:\n{e}")
            return None
        try:
            db = BotanikDB(production=True)
            if not db.baglan():
                messagebox.showerror("Bağlantı Hatası",
                                     "Botanik EOS veritabanına bağlanılamadı.")
                return None
            return db
        except Exception as e:
            messagebox.showerror("Bağlantı Hatası",
                                 f"Botanik EOS bağlantısı kurulamadı:\n{e}")
            return None

    def _recete_detay_verileri_cek(self, db, recete_no):
        """Bir reçetenin tüm detayını DB'den çek. Returns: dict veya None."""
        # 1) Header + ilaçlar (tek sorgu, ReceteAna LEFT JOIN tüm referanslar)
        sql = """
            SELECT
                ra.RxId, ra.RxEReceteNo, ra.RxSgkIslemNo,
                ra.RxIslemTarihi, ra.RxKayitTarihi, ra.RxReceteTarihi,
                ra.RxBransId, ra.RxKurumId, ra.RxHastaneId,
                ra.RxMusteriId, ra.RxDoktorId,
                ra.RxReceteRenkId, ra.RxReceteAltTuruId, ra.RxProvizyonTipId,
                m.MusteriAdiSoyadi, m.MusteriTCKN, m.MusteriDogumTarihi,
                m.MusteriCinsiyet, m.MusteriKapsamId, m.MusteriEmeklilik,
                d.DoktorAdiSoyadi,
                h.HastaneAdi, h.HastaneKodu,
                k.KurumAdi,
                ri.RIId, ri.RIUrunId, ri.RIRaporKodId, ri.RIRaporNo,
                ri.RIAdet, ri.RIDoz, ri.RITekrar, ri.RIAralik, ri.RIPeriyotId,
                ri.RIToplam, ri.RIFiyatFarki,
                u.UrunAdi,
                (SELECT TOP 1 b.BarkodAdi FROM Barkod b
                  WHERE b.BarkodUrunId = u.UrunId
                  ORDER BY b.BarkodSilme ASC) AS UrunBarkodu,
                atc.ATCKodu, atc.ATCTurkce
            FROM ReceteAna ra
            LEFT JOIN Musteri m ON m.MusteriId = ra.RxMusteriId
            LEFT JOIN Doktor d ON d.DoktorId = ra.RxDoktorId
            LEFT JOIN Hastane h ON h.HastaneId = ra.RxHastaneId
            LEFT JOIN Kurum k ON k.KurumId = ra.RxKurumId
            INNER JOIN ReceteIlaclari ri ON ri.RIRxId = ra.RxId
                                          AND (ri.RISilme IS NULL OR ri.RISilme = 0)
            LEFT JOIN Urun u ON u.UrunId = ri.RIUrunId
            LEFT JOIN ATC atc ON atc.ATCId = u.UrunATCId
            WHERE ra.RxEReceteNo = ? AND (ra.RxSilme IS NULL OR ra.RxSilme = 0)
            ORDER BY ri.RIId
        """
        rows = db.sorgu_calistir(sql, (recete_no,))
        if not rows:
            return None
        rx_id = rows[0]["RxId"]
        musteri_id = rows[0].get("RxMusteriId")
        doktor_id = rows[0].get("RxDoktorId")

        # 2) Teşhisler — ReceteICD + ICD
        teshisler = []
        try:
            ricd = db.sorgu_calistir(
                """SELECT icd.ICDKodu, icd.ICDAciklamasi
                   FROM ReceteICD ricd
                   LEFT JOIN ICD icd ON icd.ICDId = ricd.ReceteICDICDId
                   WHERE ricd.ReceteICDRxId = ?
                     AND (ricd.ReceteICDSilme IS NULL OR ricd.ReceteICDSilme = 0)""",
                (rx_id,))
            for r in ricd:
                kod = (r.get("ICDKodu") or "").strip()
                ack = (r.get("ICDAciklamasi") or "").strip()
                if kod and ack:
                    teshisler.append(f"{kod} — {ack}")
                elif kod:
                    teshisler.append(kod)
        except Exception as e:
            logger.debug(f"ReceteICD okunamadı: {e}")

        # 3) Eski Teshis sistemi — ReceteTeshis + Teshis
        try:
            rt = db.sorgu_calistir(
                """SELECT t.TeshisAciklama
                   FROM ReceteTeshis rt
                   LEFT JOIN Teshis t ON t.TeshisId = rt.RTTeshisId
                   WHERE rt.RTRxId = ?""",
                (rx_id,))
            for r in rt:
                ack = (r.get("TeshisAciklama") or "").strip()
                if ack and "Seçiniz" not in ack and ack not in teshisler:
                    teshisler.append(ack)
        except Exception as e:
            logger.debug(f"ReceteTeshis okunamadı: {e}")

        # 4) E-Reçete açıklamaları
        aciklamalar = []
        try:
            ea = db.sorgu_calistir(
                """SELECT eat.EReceteAciklamaTuruAdi, ea.EReceteAciklamaAdi
                   FROM ERecete er
                   INNER JOIN EReceteAciklamalari era ON era.ERAEReceteId = er.EReceteId
                   LEFT JOIN EReceteAciklama ea
                       ON ea.EReceteAciklamaId = era.ERAEReceteAciklamaId
                   LEFT JOIN EReceteAciklamaTuru eat
                       ON eat.EReceteAciklamaTuruId = era.ERAEReceteAciklamaTuruId
                   WHERE er.EReceteNo = ? AND (er.EReceteSilme IS NULL OR er.EReceteSilme = 0)""",
                (recete_no,))
            for r in ea:
                tur = (r.get("EReceteAciklamaTuruAdi") or "").strip()
                ad = (r.get("EReceteAciklamaAdi") or "").strip()
                if not ad or ad in (".", ",", "-", "--"):
                    continue
                aciklamalar.append(f"[{tur}] {ad}" if tur and tur != "Seçiniz" else ad)
        except Exception as e:
            logger.debug(f"EReceteAciklamalari okunamadı: {e}")

        # 5) Medula yanıt mesajları (RxUyarilari)
        medula_yanitlari = []
        try:
            ru = db.sorgu_calistir(
                "SELECT RUAciklama FROM RxUyarilari WHERE RxId = ?", (rx_id,))
            for r in ru:
                txt = (r.get("RUAciklama") or "").strip()
                if txt:
                    medula_yanitlari.append(txt)
        except Exception as e:
            logger.debug(f"RxUyarilari okunamadı: {e}")

        # 6) Doktor branşı
        doktor_brans = ""
        if doktor_id:
            try:
                db_rows = db.sorgu_calistir(
                    """SELECT b.BransAdi FROM DoktorBrans db
                       INNER JOIN Brans b ON b.BransId = db.DoktorBransBransId
                       WHERE db.DoktorBransDoktorId = ? AND (b.BransSilme IS NULL OR b.BransSilme = 0)""",
                    (doktor_id,))
                doktor_brans = ", ".join(
                    (r.get("BransAdi") or "").strip() for r in db_rows
                    if (r.get("BransAdi") or "").strip()
                )
            except Exception as e:
                logger.debug(f"DoktorBrans okunamadı: {e}")

        return {
            "header": rows[0],
            "ilaclar": rows,
            "teshisler": teshisler,
            "aciklamalar": aciklamalar,
            "medula_yanitlari": medula_yanitlari,
            "doktor_brans": doktor_brans,
            "musteri_id": musteri_id,
        }

    def _recete_detay_penceresi(self, parent_win, recete_no):
        """Reçete detayını Botanik EOS'tan çek ve ayrı pencerede göster."""
        if not recete_no:
            return
        db = self._detay_db_baglan()
        if not db:
            return
        try:
            data = self._recete_detay_verileri_cek(db, recete_no)
        except Exception as e:
            logger.exception("Reçete detay sorgu hatası")
            messagebox.showerror("Sorgu Hatası", f"Reçete detayı çekilemedi:\n{e}")
            return
        finally:
            try:
                db.kapat()
            except Exception:
                pass

        if not data:
            messagebox.showinfo(
                "Bulunamadı",
                f"'{recete_no}' numaralı reçete Botanik EOS veritabanında bulunamadı."
            )
            return

        h = data["header"]
        win = tk.Toplevel(parent_win)
        win.title(f"Reçete Detayı — {recete_no}")
        win.geometry("900x650")
        win.configure(bg="#1E3A5F")
        win.transient(parent_win)

        # Ana scrollable text
        text_frame = tk.Frame(win, bg="#1E3A5F")
        text_frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(text_frame, wrap="word", bg="#FFFFFF", fg="#000000",
                       font=("Segoe UI", 10), padx=10, pady=8)
        sb = ttk.Scrollbar(text_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)

        # Stiller
        txt.tag_configure("h1", font=("Segoe UI", 13, "bold"), foreground="#1E3A5F",
                          spacing3=6)
        txt.tag_configure("h2", font=("Segoe UI", 11, "bold"), foreground="#0D47A1",
                          spacing1=8, spacing3=4)
        txt.tag_configure("k", font=("Segoe UI", 10, "bold"), foreground="#37474F")
        txt.tag_configure("warn", foreground="#C62828")
        txt.tag_configure("mono", font=("Consolas", 10))

        def yaz(s, tag=None):
            txt.insert("end", s, tag) if tag else txt.insert("end", s)

        def alan(etiket, deger):
            if deger is None or str(deger).strip() == "":
                return
            yaz(f"  {etiket}: ", "k"); yaz(f"{deger}\n")

        # Başlık
        yaz(f"Reçete Detayı — {recete_no}\n", "h1")

        # Reçete bilgileri
        yaz("Reçete Bilgileri\n", "h2")
        alan("Reçete No (e-Reçete)", h.get("RxEReceteNo"))
        alan("SGK İşlem No", h.get("RxSgkIslemNo"))
        alan("Reçete Tarihi", h.get("RxReceteTarihi"))
        alan("Kayıt Tarihi", h.get("RxKayitTarihi"))
        alan("İşlem Tarihi", h.get("RxIslemTarihi"))
        alan("RxId (sistem)", h.get("RxId"))

        # Hasta
        yaz("\nHasta\n", "h2")
        alan("Ad Soyad", h.get("MusteriAdiSoyadi"))
        alan("TCKN", h.get("MusteriTCKN"))
        alan("Doğum Tarihi", h.get("MusteriDogumTarihi"))
        alan("Cinsiyet", h.get("MusteriCinsiyet"))
        alan("Emeklilik", h.get("MusteriEmeklilik"))

        # Doktor
        yaz("\nDoktor\n", "h2")
        alan("Ad Soyad", h.get("DoktorAdiSoyadi"))
        alan("Branş", data.get("doktor_brans") or "")

        # Tesis / Kurum
        yaz("\nTesis / Kurum\n", "h2")
        alan("Hastane", h.get("HastaneAdi"))
        alan("Hastane Kodu", h.get("HastaneKodu"))
        alan("Kurum", h.get("KurumAdi"))

        # Teşhisler
        yaz("\nTeşhisler\n", "h2")
        if data["teshisler"]:
            for t in data["teshisler"]:
                yaz(f"  • {t}\n")
        else:
            yaz("  (Teşhis kaydı yok)\n")

        # Reçete açıklamaları
        yaz("\nReçete Açıklamaları\n", "h2")
        if data["aciklamalar"]:
            for a in data["aciklamalar"]:
                yaz(f"  • {a}\n")
        else:
            yaz("  (Açıklama yok)\n")

        # Medula yanıtları (varsa)
        if data["medula_yanitlari"]:
            yaz("\nMedula Yanıt Mesajları\n", "h2")
            for y in data["medula_yanitlari"]:
                yaz(f"  • {y}\n", "warn")

        # İlaçlar
        yaz("\nİlaçlar\n", "h2")
        for i, ilac in enumerate(data["ilaclar"], 1):
            yaz(f"\n  {i}. {ilac.get('UrunAdi') or '(ürün adı yok)'}\n", "k")
            atc_kodu = (ilac.get("ATCKodu") or "").strip()
            atc_ad = (ilac.get("ATCTurkce") or "").strip()
            if atc_kodu or atc_ad:
                yaz(f"     ATC: {atc_kodu} {atc_ad}\n")
            yaz(f"     Adet: {ilac.get('RIAdet') or '-'}  |  "
                f"Doz: {ilac.get('RIDoz') or '-'}  |  "
                f"Tekrar: {ilac.get('RITekrar') or '-'}  |  "
                f"Aralık: {ilac.get('RIAralik') or '-'}\n")
            rip_no = (str(ilac.get("RIRaporNo") or "")).strip()
            rip_kod_id = ilac.get("RIRaporKodId") or 0
            if rip_no or rip_kod_id:
                yaz(f"     Rapor: kod_id={rip_kod_id}, takip_no={rip_no or '-'}\n")
            barkod = (ilac.get("UrunBarkodu") or "").strip()
            if barkod:
                yaz(f"     Barkod: {barkod}\n")

        txt.configure(state="disabled")

        # Alt: Kapat butonu
        alt = tk.Frame(win, bg="#1E3A5F")
        alt.pack(fill="x", padx=8, pady=(0, 8))
        tk.Button(alt, text="Kapat", font=("Segoe UI", 10, "bold"),
                  fg="white", bg="#455A64", activebackground="#37474F",
                  bd=0, padx=14, pady=4,
                  command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda e: win.destroy())

    def _rapor_detay_verileri_cek(self, db, recete_no, ilac_adi="", etkin_madde="", rapor_kodu=""):
        """Belirli bir satırın rapor_kodu'na ait raporu hasta üzerinden bul ve detayını çek."""
        # Önce reçeteyi bulup hastayı al
        rx = db.sorgu_calistir(
            """SELECT TOP 1 ra.RxId, ra.RxMusteriId, ra.RxKayitTarihi
               FROM ReceteAna ra
               WHERE ra.RxEReceteNo = ? AND (ra.RxSilme IS NULL OR ra.RxSilme = 0)""",
            (recete_no,))
        if not rx:
            return None, "Reçete bulunamadı"
        rx_id = rx[0]["RxId"]
        musteri_id = rx[0].get("RxMusteriId")
        rx_kayit_tarihi = rx[0].get("RxKayitTarihi")
        if not musteri_id:
            return None, "Reçetenin hastası bulunamadı"

        # Rapor kodu üzerinden hastanın aktif raporlarından eşleşeni bul
        rapor_ana = None
        rapor_secim_kaynagi = ""

        # 1) Önce aynı reçetede (RxId) bu rapor_kodu'nu kullanan ilac satırının
        #    RIRaporKodId/RIRaporNo'sundan eşleşen RaporAna'yı bul (en isabetli)
        try:
            ri_rows = db.sorgu_calistir(
                """SELECT ri.RIRaporKodId, ri.RIRaporNo, u.UrunAdi
                   FROM ReceteIlaclari ri
                   LEFT JOIN Urun u ON u.UrunId = ri.RIUrunId
                   WHERE ri.RIRxId = ? AND (ri.RISilme IS NULL OR ri.RISilme = 0)""",
                (rx_id,))
            # ilac_adi/etkin_madde ile eşleştir
            il_norm = (ilac_adi or "").strip().upper()
            sec_kod_id = 0
            sec_rapor_no = ""
            for ri in ri_rows:
                ri_urun = (ri.get("UrunAdi") or "").strip().upper()
                if il_norm and ri_urun and (il_norm[:20] in ri_urun or ri_urun[:20] in il_norm):
                    sec_kod_id = ri.get("RIRaporKodId") or 0
                    sec_rapor_no = (ri.get("RIRaporNo") or "").strip()
                    break
            # Eşleşme bulunamadıysa, ilk raporlu ilacı dene
            if not sec_kod_id and not sec_rapor_no:
                for ri in ri_rows:
                    if (ri.get("RIRaporKodId") or 0) > 0 or (ri.get("RIRaporNo") or "").strip():
                        sec_kod_id = ri.get("RIRaporKodId") or 0
                        sec_rapor_no = (ri.get("RIRaporNo") or "").strip()
                        break

            # Önce takip_no eşleşmesi (en kesin)
            if sec_rapor_no:
                ra_rows = db.sorgu_calistir(
                    """SELECT TOP 1 ra.*, rt.RaporTuruAdi, h.HastaneAdi, h.HastaneKodu
                       FROM RaporAna ra
                       LEFT JOIN RaporTuru rt ON rt.RaporTuruId = ra.RaporAnaRaporTuruId
                       LEFT JOIN Hastane h ON h.HastaneId = ra.RaporAnaHastaneId
                       WHERE ra.RaporAnaMusteriId = ?
                         AND ra.RaporAnaRaporNo = ?
                         AND (ra.RaporAnaSilme IS NULL OR ra.RaporAnaSilme = 0)""",
                    (musteri_id, sec_rapor_no))
                if ra_rows:
                    rapor_ana = ra_rows[0]
                    rapor_secim_kaynagi = "RIRaporNo (takip no eşleşmesi)"

            # Sonra RaporKodId üzerinden
            if not rapor_ana and sec_kod_id:
                ra_rows = db.sorgu_calistir(
                    """SELECT TOP 1 ra.*, rt.RaporTuruAdi, h.HastaneAdi, h.HastaneKodu
                       FROM RaporAna ra
                       INNER JOIN RaporRaporKodlariICD rrki ON rrki.RRKIRaporAnaId = ra.RaporAnaId
                       LEFT JOIN RaporTuru rt ON rt.RaporTuruId = ra.RaporAnaRaporTuruId
                       LEFT JOIN Hastane h ON h.HastaneId = ra.RaporAnaHastaneId
                       WHERE ra.RaporAnaMusteriId = ?
                         AND rrki.RRKIRaporKodId = ?
                         AND (ra.RaporAnaSilme IS NULL OR ra.RaporAnaSilme = 0)
                         AND (rrki.RRKISilme IS NULL OR rrki.RRKISilme = 0)
                       ORDER BY ra.RaporAnaRaporTarihi DESC""",
                    (musteri_id, sec_kod_id))
                if ra_rows:
                    rapor_ana = ra_rows[0]
                    rapor_secim_kaynagi = "RIRaporKodId eşleşmesi"
        except Exception as e:
            logger.debug(f"ReceteIlaclari→RaporAna sorgusu fail: {e}")

        # 2) Hâlâ bulunamadıysa, ekrandaki rapor_kodu (örn. "10.04") metni üzerinden
        #    RaporKodlari.RaporKodu LIKE eşleşmesi yap
        if not rapor_ana and rapor_kodu:
            try:
                ra_rows = db.sorgu_calistir(
                    """SELECT TOP 1 ra.*, rt.RaporTuruAdi, h.HastaneAdi, h.HastaneKodu
                       FROM RaporAna ra
                       INNER JOIN RaporRaporKodlariICD rrki ON rrki.RRKIRaporAnaId = ra.RaporAnaId
                       INNER JOIN RaporKodlari rk ON rk.RaporKodId = rrki.RRKIRaporKodId
                       LEFT JOIN RaporTuru rt ON rt.RaporTuruId = ra.RaporAnaRaporTuruId
                       LEFT JOIN Hastane h ON h.HastaneId = ra.RaporAnaHastaneId
                       WHERE ra.RaporAnaMusteriId = ?
                         AND (rk.RaporKodu = ? OR rk.RaporKodu LIKE ?)
                         AND (ra.RaporAnaSilme IS NULL OR ra.RaporAnaSilme = 0)
                         AND (rrki.RRKISilme IS NULL OR rrki.RRKISilme = 0)
                         AND (rk.RaporKodSilme IS NULL OR rk.RaporKodSilme = 0)
                       ORDER BY ra.RaporAnaRaporTarihi DESC""",
                    (musteri_id, rapor_kodu, rapor_kodu + "%"))
                if ra_rows:
                    rapor_ana = ra_rows[0]
                    rapor_secim_kaynagi = f"RaporKodu='{rapor_kodu}' eşleşmesi"
            except Exception as e:
                logger.debug(f"RaporKodu eşleşme sorgusu fail: {e}")

        if not rapor_ana:
            return None, "Hastanın bu rapor koduna ait raporu bulunamadı"

        rapor_ana_id = rapor_ana["RaporAnaId"]

        # Hasta bilgisi
        hasta = db.sorgu_calistir(
            "SELECT MusteriAdiSoyadi, MusteriTCKN, MusteriDogumTarihi FROM Musteri WHERE MusteriId = ?",
            (musteri_id,))
        hasta_bilgi = hasta[0] if hasta else {}

        # ICD'ler
        icdler = []
        try:
            rows_icd = db.sorgu_calistir(
                """SELECT icd1.ICDKodu AS K1, icd1.ICDAciklamasi AS A1,
                          icd2.ICDKodu AS K2, icd2.ICDAciklamasi AS A2,
                          icd3.ICDKodu AS K3, icd3.ICDAciklamasi AS A3,
                          icd4.ICDKodu AS K4, icd4.ICDAciklamasi AS A4,
                          icd5.ICDKodu AS K5, icd5.ICDAciklamasi AS A5,
                          rk.RaporKodu, rk.RaporKodAciklama,
                          rrki.RRKIBaslamaTarihi, rrki.RRKIBitisTarihi
                   FROM RaporRaporKodlariICD rrki
                   LEFT JOIN ICD icd1 ON icd1.ICDId = rrki.RRKIICDId
                   LEFT JOIN ICD icd2 ON icd2.ICDId = rrki.RRKIICDId2
                   LEFT JOIN ICD icd3 ON icd3.ICDId = rrki.RRKIICDId3
                   LEFT JOIN ICD icd4 ON icd4.ICDId = rrki.RRKIICDId4
                   LEFT JOIN ICD icd5 ON icd5.ICDId = rrki.RRKIICDId5
                   LEFT JOIN RaporKodlari rk ON rk.RaporKodId = rrki.RRKIRaporKodId
                   WHERE rrki.RRKIRaporAnaId = ?
                     AND (rrki.RRKISilme IS NULL OR rrki.RRKISilme = 0)""",
                (rapor_ana_id,))
            for r in rows_icd:
                grup = []
                for n in (1, 2, 3, 4, 5):
                    kod = (r.get(f"K{n}") or "").strip()
                    ack = (r.get(f"A{n}") or "").strip()
                    if kod and ack:
                        grup.append(f"{kod} — {ack}")
                    elif kod:
                        grup.append(kod)
                icdler.append({
                    "rapor_kodu": (r.get("RaporKodu") or "").strip(),
                    "rapor_kodu_aciklama": (r.get("RaporKodAciklama") or "").strip(),
                    "icd_listesi": grup,
                    "baslama": r.get("RRKIBaslamaTarihi"),
                    "bitis": r.get("RRKIBitisTarihi"),
                })
        except Exception as e:
            logger.debug(f"RaporRaporKodlariICD okunamadı: {e}")

        # Ek bilgiler
        ek_bilgiler = []
        try:
            rows_eb = db.sorgu_calistir(
                """SELECT REBTuru, REBDeger, REBAciklama
                   FROM RaporEkBilgi WHERE REBRaporAnaId = ?""",
                (rapor_ana_id,))
            for r in rows_eb:
                parts = []
                if r.get("REBTuru"):
                    parts.append(str(r["REBTuru"]))
                if r.get("REBDeger"):
                    parts.append(str(r["REBDeger"]))
                if r.get("REBAciklama"):
                    parts.append(str(r["REBAciklama"]))
                if parts:
                    ek_bilgiler.append(": ".join(parts))
        except Exception as e:
            logger.debug(f"RaporEkBilgi okunamadı: {e}")

        # Etkin maddeler
        etkin_maddeler = []
        try:
            rows_em = db.sorgu_calistir(
                """SELECT em.EtkinMaddeAdi, em.EtkinMaddeSGKKodu,
                          re.EtkinMaddeDoz, re.EtkinMaddeAdetMiktar,
                          re.EtkinMaddeTekrar, re.EtkinMaddeAralik
                   FROM RaporEtkinMadde re
                   LEFT JOIN EtkinMadde em ON em.EtkinMaddeId = re.EtkinMaddeId
                   WHERE re.EtkinMaddeRaporAnaId = ?
                     AND (re.EtkinMaddeSilme IS NULL OR re.EtkinMaddeSilme = 0)""",
                (rapor_ana_id,))
            for r in rows_em:
                etkin_maddeler.append({
                    "ad": (r.get("EtkinMaddeAdi") or "").strip(),
                    "sgk": (r.get("EtkinMaddeSGKKodu") or "").strip(),
                    "doz": r.get("EtkinMaddeDoz"),
                    "adet": r.get("EtkinMaddeAdetMiktar"),
                    "tekrar": r.get("EtkinMaddeTekrar"),
                    "aralik": r.get("EtkinMaddeAralik"),
                })
        except Exception as e:
            logger.debug(f"RaporEtkinMadde okunamadı: {e}")

        return {
            "rapor": rapor_ana,
            "hasta": hasta_bilgi,
            "icdler": icdler,
            "ek_bilgiler": ek_bilgiler,
            "etkin_maddeler": etkin_maddeler,
            "secim_kaynagi": rapor_secim_kaynagi,
        }, None

    def _rapor_detay_penceresi(self, parent_win, recete_no, satir):
        """Satırın rapor_kodu'na ait raporu Botanik EOS'tan çek ve göster."""
        rapor_kodu = (satir.get("rapor_kodu") or "").strip()
        ilac_adi = (satir.get("ilac_adi") or "").strip()
        etkin = (satir.get("etkin_madde") or "").strip()

        db = self._detay_db_baglan()
        if not db:
            return
        try:
            data, err = self._rapor_detay_verileri_cek(
                db, recete_no, ilac_adi=ilac_adi, etkin_madde=etkin,
                rapor_kodu=rapor_kodu)
        except Exception as e:
            logger.exception("Rapor detay sorgu hatası")
            messagebox.showerror("Sorgu Hatası", f"Rapor detayı çekilemedi:\n{e}")
            return
        finally:
            try:
                db.kapat()
            except Exception:
                pass

        if not data:
            messagebox.showinfo("Bulunamadı", err or "Rapor bulunamadı.")
            return

        r = data["rapor"]
        h = data["hasta"]
        win = tk.Toplevel(parent_win)
        win.title(f"Rapor Detayı — {r.get('RaporAnaRaporNo') or '?'}")
        win.geometry("900x650")
        win.configure(bg="#1E3A5F")
        win.transient(parent_win)

        text_frame = tk.Frame(win, bg="#1E3A5F")
        text_frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(text_frame, wrap="word", bg="#FFFFFF", fg="#000000",
                       font=("Segoe UI", 10), padx=10, pady=8)
        sb = ttk.Scrollbar(text_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)

        txt.tag_configure("h1", font=("Segoe UI", 13, "bold"), foreground="#1E3A5F",
                          spacing3=6)
        txt.tag_configure("h2", font=("Segoe UI", 11, "bold"), foreground="#0D47A1",
                          spacing1=8, spacing3=4)
        txt.tag_configure("k", font=("Segoe UI", 10, "bold"), foreground="#37474F")
        txt.tag_configure("muted", foreground="#6B7280")

        def yaz(s, tag=None):
            txt.insert("end", s, tag) if tag else txt.insert("end", s)
        def alan(etiket, deger):
            if deger is None or str(deger).strip() == "":
                return
            yaz(f"  {etiket}: ", "k"); yaz(f"{deger}\n")

        yaz(f"Rapor Detayı — {r.get('RaporAnaRaporNo') or '?'}\n", "h1")
        yaz(f"  (Reçete: {recete_no} | Satır rapor kodu: {rapor_kodu or '-'})\n",
            "muted")
        if data.get("secim_kaynagi"):
            yaz(f"  (Eşleşme: {data['secim_kaynagi']})\n", "muted")

        # Rapor başlık
        yaz("\nRapor Bilgileri\n", "h2")
        alan("Rapor No", r.get("RaporAnaRaporNo"))
        alan("Takip No", r.get("RaporAnaRaporTakipNo"))
        alan("Rapor Tarihi", r.get("RaporAnaRaporTarihi"))
        alan("Tür", r.get("RaporTuruAdi"))
        alan("Açıklamalar", r.get("RaporAnaAciklamalar"))

        # Hasta
        yaz("\nHasta\n", "h2")
        alan("Ad Soyad", h.get("MusteriAdiSoyadi"))
        alan("TCKN", h.get("MusteriTCKN"))
        alan("Doğum Tarihi", h.get("MusteriDogumTarihi"))

        # Tesis
        yaz("\nTesis\n", "h2")
        alan("Hastane", r.get("HastaneAdi"))
        alan("Hastane Kodu", r.get("HastaneKodu"))

        # ICD listesi (rapor kodu ile gruplu)
        yaz("\nRapor Kodları + ICD Teşhisleri\n", "h2")
        if data["icdler"]:
            for grp in data["icdler"]:
                kod = grp.get("rapor_kodu") or "-"
                kod_ack = grp.get("rapor_kodu_aciklama") or ""
                bas = grp.get("baslama") or ""
                bit = grp.get("bitis") or ""
                yaz(f"  • {kod} {kod_ack}\n", "k")
                if bas or bit:
                    yaz(f"      ({bas} → {bit})\n", "muted")
                for ic in grp.get("icd_listesi", []):
                    yaz(f"      ICD: {ic}\n")
        else:
            yaz("  (Rapor kodu/ICD kaydı yok)\n")

        # Etkin maddeler
        yaz("\nRapor Etkin Maddeleri (tedavi)\n", "h2")
        if data["etkin_maddeler"]:
            for em in data["etkin_maddeler"]:
                ad = em.get("ad") or "-"
                sgk = em.get("sgk") or ""
                yaz(f"  • {ad}", "k")
                if sgk:
                    yaz(f"  [{sgk}]")
                doz_ozet = []
                if em.get("doz") is not None:
                    doz_ozet.append(f"Doz:{em['doz']}")
                if em.get("adet") is not None:
                    doz_ozet.append(f"Adet:{em['adet']}")
                if em.get("tekrar") is not None:
                    doz_ozet.append(f"Tekrar:{em['tekrar']}")
                if em.get("aralik") is not None:
                    doz_ozet.append(f"Aralık:{em['aralik']}")
                if doz_ozet:
                    yaz(f"  ({', '.join(doz_ozet)})")
                yaz("\n")
        else:
            yaz("  (Etkin madde tanımı yok)\n")

        # Ek bilgiler
        if data["ek_bilgiler"]:
            yaz("\nEk Bilgiler\n", "h2")
            for eb in data["ek_bilgiler"]:
                yaz(f"  • {eb}\n")

        txt.configure(state="disabled")

        alt = tk.Frame(win, bg="#1E3A5F")
        alt.pack(fill="x", padx=8, pady=(0, 8))
        tk.Button(alt, text="Kapat", font=("Segoe UI", 10, "bold"),
                  fg="white", bg="#455A64", activebackground="#37474F",
                  bd=0, padx=14, pady=4,
                  command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda e: win.destroy())

    def _pencereleri_konumlandir(self):
        """Tarama başladığında pencereleri konumlandır.

        Öncelik: pencere_yerlesim.json'da kayıtlı yerleşim varsa onu uygula.
        Yoksa fallback: Medula sol %60, Botanik Takip sağ %40."""
        # Önce kayıtlı yerleşimi dene
        try:
            import pencere_yerlesim as py_yerlesim
            kayitli = py_yerlesim.yerlesim_yukle()
            if kayitli.get("medula") or kayitli.get("hasta_takip"):
                if kayitli.get("medula"):
                    py_yerlesim.medulaya_uygula()
                if kayitli.get("hasta_takip"):
                    self.root.after(0, lambda: py_yerlesim.hasta_takibe_uygula(self.root))
                return
        except Exception as e:
            logger.debug(f"Kayıtlı yerleşim uygulanamadı, varsayılana dönülüyor: {e}")

        # Fallback: hard-coded yerleşim
        try:
            import win32gui, win32con

            # Ekran boyutu (taskbar hariç çalışma alanı)
            import ctypes
            user32 = ctypes.windll.user32

            # Çalışma alanı (taskbar hariç)
            from ctypes import wintypes
            rect = wintypes.RECT()
            user32.SystemParametersInfoW(0x0030, 0, ctypes.byref(rect), 0)  # SPI_GETWORKAREA
            ekran_w = rect.right - rect.left
            ekran_h = rect.bottom - rect.top

            # Medula: sol %60
            medula_w = int(ekran_w * 0.60)
            def enum_cb(hwnd, _):
                if win32gui.IsWindowVisible(hwnd):
                    title = win32gui.GetWindowText(hwnd)
                    if "MEDULA" in title:
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                        win32gui.MoveWindow(hwnd, 0, 0, medula_w, ekran_h, True)
            win32gui.EnumWindows(enum_cb, None)

            # Botanik Takip: sağ %40, yükseklik 100px kısa (görev çubuğu üstünde kalması için)
            bt_w = ekran_w - medula_w
            bt_x = medula_w
            bt_h = ekran_h - 50
            self.root.after(0, lambda: self.root.geometry(
                f"{bt_w}x{bt_h}+{bt_x}+0"))

        except Exception as e:
            logger.debug(f"Pencere konumlandırma hatası: {e}")

    def _sut_matrisi_ac(self):
        """🎯 SUT Matrisi penceresini aç (Botanik EOS DB tabanlı, ayrı Toplevel)."""
        try:
            from sut_matrisi_gui import SUTMatrisiGUI
            pencere = tk.Toplevel(self.root)
            SUTMatrisiGUI(pencere, kullanici_id=None)
        except Exception as e:
            logger.exception("SUT Matrisi açma hatası: %s", e)
            from tkinter import messagebox
            messagebox.showerror(
                "Hata",
                f"SUT Matrisi açılamadı:\n{e}",
            )

    def _yerlesim_kaydet_tikla(self):
        """'Sayfa Yerleşimini Kaydet' butonu — şu anki MEDULA + Hasta Takip
        konum/boyutunu pencere_yerlesim.json'a yazar."""
        try:
            import pencere_yerlesim as py_yerlesim
            from tkinter import messagebox
            ok = py_yerlesim.yerlesimi_kaydet_simdi(tk_root=self.root)
            if ok:
                data = py_yerlesim.yerlesim_yukle()
                msg_parts = []
                if "medula" in data:
                    m = data["medula"]
                    msg_parts.append(f"Medula: {m['x']},{m['y']}  {m['width']}x{m['height']}")
                if "hasta_takip" in data:
                    h = data["hasta_takip"]
                    msg_parts.append(f"Hasta Takip: {h['x']},{h['y']}  {h['width']}x{h['height']}")
                detay = "\n".join(msg_parts) if msg_parts else "(boş)"
                self.log_yaz(f"✓ Sayfa yerleşimi kaydedildi", "success")
                messagebox.showinfo("Yerleşim Kaydedildi",
                                    f"Mevcut pencere yerleşimi kaydedildi:\n\n{detay}")
            else:
                self.log_yaz("Yerleşim kaydedilemedi (MEDULA bulunamadı?)", "error")
                messagebox.showwarning("Kaydedilemedi",
                                       "Yerleşim kaydedilemedi.\nMEDULA penceresi açık mı?")
        except Exception as e:
            logger.error(f"Yerleşim kaydetme hatası: {e}")
            self.log_yaz(f"Yerleşim kaydetme hatası: {e}", "error")

    def _yerlesim_uygula_tikla(self):
        """'Sayfa Yerleşimini Uygula' butonu — kayıtlı yerleşimi MEDULA ve
        Hasta Takip pencerelerine uygular."""
        try:
            import pencere_yerlesim as py_yerlesim
            from tkinter import messagebox
            data = py_yerlesim.yerlesim_yukle()
            if not data:
                messagebox.showwarning("Yerleşim Yok",
                                       "Kayıtlı yerleşim bulunamadı.\n"
                                       "Önce 'Yerleşimi Kaydet' butonunu kullanın.")
                return
            uygulanan = []
            if data.get("medula"):
                if py_yerlesim.medulaya_uygula():
                    uygulanan.append("Medula")
            if data.get("hasta_takip"):
                if py_yerlesim.hasta_takibe_uygula(self.root):
                    uygulanan.append("Hasta Takip")
            if uygulanan:
                self.log_yaz(f"✓ Yerleşim uygulandı: {', '.join(uygulanan)}", "success")
            else:
                self.log_yaz("Yerleşim uygulanamadı (MEDULA açık mı?)", "warning")
                messagebox.showwarning("Uygulanamadı",
                                       "Yerleşim uygulanamadı.\nMEDULA penceresi açık mı?")
        except Exception as e:
            logger.error(f"Yerleşim uygulama hatası: {e}")
            self.log_yaz(f"Yerleşim uygulama hatası: {e}", "error")

    def _taramayi_durdur(self):
        """Taramayı durdur - STOP dosyası oluştur ve subprocess'i terminate et"""
        # TÜMÜNÜ modundaysa: iptal flag'ı set et (loop görsün)
        if getattr(self, "_tumu_aktif", False):
            self._tumu_iptal = True
        # Çoklu dönem akışındaysa onu da iptal et (dış dönem döngüsünden çıkar)
        self._coklu_donem_iptal = True

        if not self.kontrol_aktif:
            return

        self.log_yaz("TARAMA DURDURULUYOR...", "warning")

        # 1. STOP dosyası oluştur (subprocess kendi döngüsünde kontrol eder)
        try:
            stop_path = os.path.join(PROJE_DIZINI, "tarama_stop.flag")
            with open(stop_path, "w") as f:
                f.write("stop")
        except:
            pass

        # 2. Subprocess'i terminate et
        if self._subprocess_proc:
            try:
                self._subprocess_proc.terminate()
                self.log_yaz("Subprocess durduruldu", "warning")
            except:
                pass
            self._subprocess_proc = None

        self.kontrol_aktif = False
        self._butonlari_normal_yap()

    def _butonlari_durdur_yap(self, aktif_kod=None):
        """Aktif grubun butonunu DURDUR haline getir, diğerlerini devre dışı bırak"""
        for kod, btn in self.grup_butonlari.items():
            if kod == aktif_kod:
                btn.config(text="  DURDUR (Esc)", bg="#B71C1C", activebackground="#7F0000")
            else:
                btn.config(state="disabled")
        self.tumunu_btn.config(state="disabled")
        # Duraklat butonu aktif
        try:
            self.duraklat_btn.config(state="normal", text="⏸ DURAKLAT",
                                     bg="#F57C00", activebackground="#E65100")
            self.duraklatildi = False
        except Exception:
            pass

    def _butonlari_normal_yap(self):
        """Tüm butonları normal haline döndür"""
        for kod, btn in self.grup_butonlari.items():
            btn.config(text=f"  {btn._ad}", bg=btn._renk_normal,
                      activebackground=btn._renk_hover, state="normal")
        self.tumunu_btn.config(text="TÜMÜNÜ KONTROL ET", bg="#00695C",
                              activebackground="#004D40", state="normal")
        # Duraklat butonu pasif
        try:
            self.duraklat_btn.config(state="disabled", text="⏸ DURAKLAT",
                                     bg="#F57C00", activebackground="#E65100")
            self.duraklatildi = False
        except Exception:
            pass

    def _duraklat_toggle(self):
        """Subprocess'i duraklat/devam ettir (psutil.suspend/resume)"""
        if not self._subprocess_proc:
            self.log_yaz("Duraklatacak aktif tarama yok", "warning")
            return
        try:
            import psutil
            p = psutil.Process(self._subprocess_proc.pid)
            if not self.duraklatildi:
                # Alt süreçleri de duraklat
                for c in p.children(recursive=True):
                    try: c.suspend()
                    except: pass
                p.suspend()
                self.duraklatildi = True
                self.duraklat_btn.config(text="▶ DEVAM ET", bg="#388E3C",
                                         activebackground="#1B5E20")
                self.log_yaz("⏸ Tarama duraklatıldı", "warning")
            else:
                p.resume()
                for c in p.children(recursive=True):
                    try: c.resume()
                    except: pass
                self.duraklatildi = False
                self.duraklat_btn.config(text="⏸ DURAKLAT", bg="#F57C00",
                                         activebackground="#E65100")
                self.log_yaz("▶ Tarama devam ediyor", "info")
        except Exception as e:
            self.log_yaz(f"Duraklat/Devam hatası: {e}", "error")

    def _grup_toggle(self, grup_kodu):
        """Grup butonu toggle - basınca başlat, tekrar basınca durdur"""
        if self.kontrol_aktif and self.aktif_grup == grup_kodu:
            self._taramayi_durdur()
            self._coklu_donem_iptal = True
        elif not self.kontrol_aktif:
            donemler = self._secilen_donemler()
            if len(donemler) <= 1:
                # Tek dönem: mevcut akışı koru (donem_var doğru ayarlı)
                self._grup_kontrol_baslat(grup_kodu)
            else:
                self._grup_donemler_calistir(grup_kodu, donemler)

    def _grup_donemler_calistir(self, grup_kodu, donemler):
        """Tek grup için birden fazla dönemi sırayla işle.

        Her dönem öncesinde self.donem_var güncellenir; subprocess akışı
        donem_var.get() üzerinden offset hesapladığı için doğru dönem seçilir.
        Excel raporu her dönem için ayrı dosya olarak yazılır (mevcut
        Kontrol_Raporu_<grup>_<donem>.xlsx adlandırması otomatik çalışır).
        """
        self._coklu_donem_iptal = False
        grup_adi = next((g["ad"] for g in GRUP_TANIMLARI
                         if g["kod"] == grup_kodu), grup_kodu)

        def thread_fn():
            self.root.after(0, lambda: self.log_yaz(
                f"{'═' * 50}", "header"))
            self.root.after(0, lambda: self.log_yaz(
                f"{grup_adi}: {len(donemler)} dönem sırayla işlenecek "
                f"→ {', '.join(donemler)}", "header"))
            self.root.after(0, lambda: self.log_yaz(
                f"{'═' * 50}", "header"))

            for idx, donem in enumerate(donemler, 1):
                if self._coklu_donem_iptal:
                    self.root.after(0, lambda: self.log_yaz(
                        "Çoklu dönem kontrolü iptal edildi", "warning"))
                    break

                # donem_var'ı güncelle (UI thread'inde)
                self.root.after(0, lambda d=donem: self.donem_var.set(d))
                # Set'in işlenmesini bekle
                time.sleep(0.15)

                self.root.after(0, lambda d=donem, i=idx: self.log_yaz(
                    f"--- Dönem {i}/{len(donemler)}: {d} ---", "header"))

                evt = threading.Event()
                self.root.after(0, lambda gk=grup_kodu, e=evt:
                                self._grup_kontrol_baslat(gk, bittiginde=e.set))

                # Subprocess'in bitişini bekle (her 0.5 sn iptal kontrolü)
                while not evt.is_set() and not self._coklu_donem_iptal:
                    time.sleep(0.5)

            self.root.after(0, lambda: self.log_yaz(
                f"Çoklu dönem kontrolü tamamlandı ({len(donemler)} dönem)",
                "header"))

        threading.Thread(target=thread_fn, daemon=True).start()

    def _tumunu_toggle(self):
        """Tümünü Kontrol Et toggle"""
        if self.kontrol_aktif:
            self._taramayi_durdur()
        else:
            self._tumunu_kontrol_et()

    def _grup_kontrol_baslat(self, grup_kodu, acik_receteden=False, bittiginde=None):
        """Belirli bir grubun kontrolünü subprocess ile başlat.

        Args:
            grup_kodu: Grup kodu (C/A/B/CK/GK)
            acik_receteden: Açık reçeteden başla
            bittiginde: Subprocess bitince çağrılacak callback (TÜMÜNÜ akışı için)
        """
        if self.kontrol_aktif:
            self.log_yaz("Zaten bir kontrol devam ediyor!", "warning")
            return

        donem_offset = self._donem_offset_hesapla()
        secilen_donem = self.donem_var.get()
        grup_adi = next((g["ad"] for g in GRUP_TANIMLARI if g["kod"] == grup_kodu), grup_kodu)

        # Combobox'ta seçili reçete var mı? (geri dönme)
        # "Baştan başla" işaretliyse combobox seçimini yoksay
        # TÜMÜNÜ akışında ikinci+ gruplar (tumu_first değil) için combobox seçimi
        # ASLA kullanılmaz — önceki taramadan kalma yanlış kayıtlar (ör: C son
        # reçetesinin A/B/GK/CK comboboxlarına yanlışlıkla yazılması) yüzünden
        # her grupta o eski reçete aranıp bulunamayıp atlanıyor olabilir.
        tumu_modu_aktif = getattr(self, "_tumu_aktif", False)
        tumu_first_grup = (getattr(self, "_tumu_grup_index", 0) == 0)
        tumu_recovery_aktif = getattr(self, "_tumu_recovery_devam", False)
        # TÜMÜNÜ ikinci+ grup VE recovery değilse: combobox seçimini bypass et + temizle
        tumu_skip_combobox = (tumu_modu_aktif and not tumu_first_grup
                                and not tumu_recovery_aktif)

        secili_recete = None
        if (not acik_receteden and not self.bastan_var.get()
                and not tumu_skip_combobox
                and grup_kodu in self.recete_combolar):
            secim = self.recete_combolar[grup_kodu].get()
            if secim and "#" in secim:
                try:
                    secili_recete = secim.split(" - ", 1)[1].strip()
                except (IndexError, ValueError):
                    pass

        self.log_yaz(f"{'━' * 40}", "header")
        if acik_receteden:
            self.log_yaz(f"{grup_adi} - Açık reçeteden başlanıyor - {secilen_donem}", "header")
        elif tumu_skip_combobox:
            # TÜMÜNÜ ikinci+ grup: combobox eski kayıtları yoksayılıyor, baştan başla
            self.log_yaz(
                f"{grup_adi} - TÜMÜNÜ akışı: ilk reçeteden baştan başlatılıyor "
                f"(combobox yoksayıldı) - {secilen_donem}", "header")
            self._recete_combolar_temizle(grup_kodu)
        elif secili_recete:
            self.log_yaz(f"{grup_adi} - {secili_recete}'den devam - {secilen_donem}", "header")
        else:
            self.log_yaz(f"{grup_adi} kontrolü başlatılıyor - {secilen_donem}", "header")
            self._recete_combolar_temizle(grup_kodu)

        self.aktif_grup = grup_kodu
        # TÜMÜNÜ modunda bireysel buton kırmızılaştırmaz - tek tek tetiklenmiş gibi görünmesin
        if getattr(self, "_tumu_aktif", False):
            # TÜMÜNÜ butonunda aktif grubu göster
            try:
                self.tumunu_btn.config(text=f"DURDUR ({grup_adi})")
            except Exception:
                pass
        else:
            self._butonlari_durdur_yap(grup_kodu)

        def kontrol_thread():
            self.kontrol_aktif = True
            self._recete_sureleri = []
            self._recete_baslangic = None
            if _PYTHONCOM_VAR:
                try: pythoncom.CoInitialize()
                except Exception: pass
            try:
                # Önce Medula hazır mı kontrol et, değilse aç ve bağlan
                self.root.after(0, lambda: self.log_yaz("Medula bağlantısı kontrol ediliyor...", "info"))
                if not self._medula_hazirla():
                    self.root.after(0, lambda: self.log_yaz(
                        "Medula bağlantısı kurulamadı! Kontrol iptal edildi.", "error"))
                    return

                # Pencereleri konumlandır: Medula sol, Botanik Takip sağ
                self._pencereleri_konumlandir()

                # Medula hazır, subprocess başlat
                self.root.after(0, lambda: self.log_yaz(
                    "Subprocess ile recete_tarama.py çalıştırılıyor...", "info"))

                import subprocess as sp
                tarama_script = os.path.join(PROJE_DIZINI, "recete_tarama.py")
                cmd_args = ["python", "-u", tarama_script, grup_kodu, str(donem_offset)]
                # TÜMÜNÜ KONTROL ET: gruplar arası ekran durumu önceki gruba ait
                # olabileceğinden sonraki gruplar için --bastan ile zorla navigasyon.
                # İlk grup (index=0) için --devam: kullanıcı listeyi önceden hazırladıysa
                # tekrar yükleme yapmasın (regression önleme).
                tumu_modu = getattr(self, "_tumu_aktif", False)
                tumu_first = (getattr(self, "_tumu_grup_index", 0) == 0)
                # Recovery sonrası: kaldığı reçeteden devam etmesi için --devam zorla
                tumu_recovery = getattr(self, "_tumu_recovery_devam", False)
                if acik_receteden:
                    cmd_args.append("--acik")
                elif secili_recete:
                    cmd_args.extend(["--goto", secili_recete])
                elif self.bastan_var.get():
                    cmd_args.append("--bastan")
                elif tumu_modu and tumu_recovery:
                    # Medula yeniden açıldıktan sonra: kaldığı reçeteden başla
                    cmd_args.append("--devam")
                elif tumu_modu and not tumu_first:
                    cmd_args.append("--bastan")
                else:
                    cmd_args.append("--devam")
                if self.kontrol_edilmisleri_atla_var.get():
                    cmd_args.append("--atla")
                proc = sp.Popen(
                    cmd_args,
                    stdout=sp.PIPE, stderr=sp.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    cwd=PROJE_DIZINI,
                    bufsize=1  # Satır-bazlı tamponlama — her satır geldiği anda GUI'ye aksın
                )
                self._subprocess_proc = proc

                # TÜMÜNÜ sağlık tespiti için bilinen kritik hata pattern'ları.
                # Bu satırlardan biri çıkarsa subprocess fail oldu sayılır;
                # tumu_thread bu flag'i kontrol edip bir sonraki gruba geçmez.
                kritik_hata_patterns = (
                    ("Reçete Listesi menüsü bulunamadı", "Navigasyon: Reçete Listesi menüsü bulunamadı (Medula login/sayfa hazır değil)"),
                    ("MEDULA penceresi bulunamadı", "Medula penceresi bulunamadı"),
                    ("Medula penceresi bulunamadı", "Medula penceresi bulunamadı"),
                    ("Navigasyon başarısız", "Navigasyon başarısız oldu"),
                    ("Medula yeniden başlatılamadı", "Medula yeniden başlatılamadı"),
                )

                # stdout'u satır satır oku ve GUI log'a + terminale yaz
                # iter(readline, '') — default for-loop'un block-buffer davranışını önler
                for line in iter(proc.stdout.readline, ''):
                    line = line.rstrip()
                    if not line:
                        continue
                    # Terminale yaz
                    print(line, flush=True)
                    # Kritik hata tespiti — TÜMÜNÜ flag'i set
                    if getattr(self, "_tumu_aktif", False):
                        for patt, aciklama in kritik_hata_patterns:
                            if patt in line:
                                self._tumu_son_grup_hata = aciklama
                                break
                    # Boş grup veya tamamlanmış grup sinyali — "TÜMÜNÜ" modunda
                    # bu grup atlansın, durdurma yapılmasın (Reçete kaydı bulunamadı)
                    if "[GRUP_BOS]" in line or "[GRUP_TAMAMLANDI]" in line:
                        # 0 reçete olsa bile "kritik hata" sayma → tümünü devam etsin
                        self._tumu_son_grup_bos_normal = True
                        continue

                    # Reçete başlangıç zamanı (süre ölçümü)
                    if "[RECETE_BASLADI]" in line:
                        import time as _time
                        simdi = _time.time()
                        # TÜMÜNÜ: bu grup için reçete sayısı tut
                        if getattr(self, "_tumu_aktif", False):
                            self._tumu_son_grup_recete_sayisi = getattr(
                                self, "_tumu_son_grup_recete_sayisi", 0) + 1
                        if self._recete_baslangic is not None:
                            sure = simdi - self._recete_baslangic
                            self._recete_sureleri.append(sure)
                            ort = sum(self._recete_sureleri) / len(self._recete_sureleri)
                            n = len(self._recete_sureleri)
                            self.root.after(0, lambda s=sure, o=ort, nn=n: self._sure_guncelle(s, o, nn))
                        self._recete_baslangic = simdi
                        continue

                    # Son reçete bilgisini parse et (label güncelleme)
                    if "[SON_RECETE]" in line:
                        try:
                            parts = line.split("[SON_RECETE]")[1].strip().split(":")
                            if len(parts) >= 2:
                                grp = parts[0].strip()
                                rno = parts[1].strip()
                                sira = int(parts[2].strip()) if len(parts) >= 3 else 0
                                self.root.after(0, lambda g=grp, r=rno: self.grup_durumu_guncelle(g, r))
                                self.root.after(0, lambda g=grp, r=rno, s=sira: self._recete_combobox_ekle(g, r, s))
                        except:
                            pass
                        continue  # Bu satırı loga yazma

                    # Log seviyesini parse et - kontrol sonuçlarına göre renk
                    tag = "info"
                    if "[SORUN]" in line or "Doz aşımı" in line or "UYGUNSUZ" in line or "RAPORSUZ yazılmış" in line:
                        tag = "error"       # Kırmızı - uygunsuz
                    elif "[HATA!]" in line:
                        tag = "sorun"       # Koyu turuncu - sistem hatası
                    elif "[????]" in line or "KontrolEdilemedi" in line or "KONTROLEDİLEMEDİ" in line:
                        tag = "warning"     # Sarı - bilinemeyen
                    elif "[  OK  ]" in line or "[  OK ]" in line or "UYGUN" in line:
                        tag = "success"     # Yeşil - uygun
                    elif "[GEÇ ]" in line:
                        tag = "neutral"     # Gri - nötr (raporsuz mesajsız)
                    elif "[=====]" in line:
                        tag = "header"
                    # İlaç adı satırı: "[OKU ] İLAÇ ADI → ..." veya "[D+M ]" pattern
                    if "[OKU ]" in line and "→" in line and ("Raporlu" in line or "Raporsuz" in line or "Mesaj VAR" in line):
                        tag = "ilac"
                    elif "[GEÇ ]" in line and "→" in line:
                        tag = "ilac"
                    self.root.after(0, lambda l=line, t=tag: self.log_yaz(l, t))

                    # Kullanıcıya yönelik satırlar Kontrol Sonuçları sekmesine de yansısın
                    # (sistem akışı / debug satırları log'da kalır)
                    if tag in ("success", "error", "sorun", "warning", "ilac", "header"):
                        self.root.after(0, lambda l=line, t=tag: self.kontrol_yaz(l, t))
                    elif "Reçete:" in line or "Reçete #" in line or "Hasta:" in line:
                        self.root.after(0, lambda l=line: self.kontrol_yaz(l, "header"))

                proc.wait()
                # Son reçetenin süresini de hesapla
                import time as _time
                if self._recete_baslangic is not None:
                    sure = _time.time() - self._recete_baslangic
                    self._recete_sureleri.append(sure)
                    ort = sum(self._recete_sureleri) / len(self._recete_sureleri)
                    n = len(self._recete_sureleri)
                    self.root.after(0, lambda s=sure, o=ort, nn=n: self._sure_guncelle(s, o, nn))
                    self._recete_baslangic = None
                self.root.after(0, lambda: self.log_yaz(
                    f"Subprocess tamamlandı (exit: {proc.returncode})",
                    "success" if proc.returncode == 0 else "error"))
                # TÜMÜNÜ modunda: subprocess crash (returncode != 0) → kritik hata olarak işaretle
                # Kullanıcı durdurma ile subprocess.terminate() çağrılırsa _tumu_iptal True olur,
                # bu durumda recovery tetiklenmez (zaten manuel iptal).
                if (getattr(self, "_tumu_aktif", False)
                        and proc.returncode not in (0, None)
                        and not getattr(self, "_tumu_iptal", False)
                        and not self._tumu_son_grup_hata):
                    self._tumu_son_grup_hata = (
                        f"Subprocess çıkış kodu {proc.returncode} "
                        f"(beklenmeyen hata / pencere kapanması)"
                    )
            except Exception as e:
                self.root.after(0, lambda e=e: self.log_yaz(f"Subprocess hatası: {e}", "error"))
                # TÜMÜNÜ modunda: GUI tarafında istisna → kritik hata
                if getattr(self, "_tumu_aktif", False) and not self._tumu_son_grup_hata:
                    self._tumu_son_grup_hata = f"GUI subprocess istisna: {e}"
            finally:
                self.kontrol_aktif = False
                self._subprocess_proc = None
                # TÜMÜNÜ modunda: butonları normalleştirme (bir sonraki grup gelecek)
                if not getattr(self, "_tumu_aktif", False):
                    self.aktif_grup = None
                    self.root.after(0, self._butonlari_normal_yap)
                if _PYTHONCOM_VAR:
                    try: pythoncom.CoUninitialize()
                    except Exception: pass
                # Bittiğinde callback'i çağır (TÜMÜNÜ akışını uyandırmak için)
                if bittiginde is not None:
                    try:
                        bittiginde()
                    except Exception:
                        pass

        threading.Thread(target=kontrol_thread, daemon=True).start()

    def _donem_offset_hesapla(self):
        """GUI'deki dönem seçimine göre Medula dropdown offset hesapla"""
        from datetime import date
        secilen = self.donem_var.get()  # "2026 Mart" gibi
        bugun = date.today()
        ay_isimleri = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                       "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
        # Seçilen ayı parse et
        parts = secilen.split()
        if len(parts) == 2:
            secilen_yil = int(parts[0])
            secilen_ay = ay_isimleri.index(parts[1]) if parts[1] in ay_isimleri else bugun.month
        else:
            secilen_yil = bugun.year
            secilen_ay = bugun.month
        # Offset = bu ay - seçilen ay (ay farkı)
        offset = (bugun.year * 12 + bugun.month) - (secilen_yil * 12 + secilen_ay)
        return max(0, offset)

    def _donemler_degisti(self, secilenler):
        """Çoklu dönem widget'ı değiştiğinde — donem_var (compat shim)
        her zaman seçili dönemlerin ilkini gösterir."""
        if secilenler:
            self.donem_var.set(secilenler[0])

    def _secilen_donemler(self):
        """Aktif çoklu dönem seçimini döndür. Widget hazır değilse donem_var
        tek elemanlı listeye fallback eder."""
        w = getattr(self, "donem_widget", None)
        if w is not None:
            secilen = w.get_secilenler()
            if secilen:
                return secilen
        deger = self.donem_var.get()
        return [deger] if deger else []

    def _grup_kontrol_islemi(self, grup_kodu):
        """
        Bir grubun tam kontrol akışı (thread içinde çalışır).
        Yeni akış:
        1. Medula bağlan + oturum kontrol
        2. Reçete Listesi → Dönem seç → Sorgula → Grup sekmesi
        3. İlk reçeteye tıkla
        4. Her reçete için:
           a. Reçete türü kontrol (beyaz/kırmızı/yeşil/mor)
           b. Renkli ise → Excel listesinde var mı kontrol
           c. İlaç satırlarını oku (DataItem'lardan)
           d. Her ilaç için algoritmik SUT kontrol
           e. Rapor tablosuna yaz
           f. Sonraki Reçete (invoke ile)
        5. Reçete bulunamadı gelince dur → Excel rapor oluştur
        """
        import pyautogui
        pyautogui.FAILSAFE = False
        from recete_kontrol_motoru import ReceteKontrolMotoru, KontrolSonuc

        grup_adi = next((g["ad"] for g in GRUP_TANIMLARI if g["kod"] == grup_kodu), grup_kodu)
        donem_offset = self._donem_offset_hesapla()
        secilen_donem = self.donem_var.get()

        def log(msg, tag="info"):
            self.root.after(0, lambda m=msg, t=tag: self.log_yaz(m, t))

        def kontrol(msg, tag="info"):
            """Kontrol Sonuçları sekmesine kullanıcıya yönelik satır yaz"""
            self.root.after(0, lambda m=msg, t=tag: self.kontrol_yaz(m, t))

        log(f"{'━' * 40}", "header")
        log(f"{grup_adi} kontrolü başlatılıyor - {secilen_donem}", "header")
        log(f"{'━' * 40}", "header")
        kontrol(f"{'═' * 60}", "header")
        kontrol(f"{grup_adi} kontrolü - {secilen_donem}", "header")
        kontrol(f"{'═' * 60}", "header")

        # === 1. MEDULA BAĞLANTI ===
        # Bağlı değilse bağlan
        if not self.medula.bagli:
            log("Medula'ya bağlanılıyor...", "info")
            self.medula.medula_baglan()

        if not self.medula.bagli:
            log("Medula penceresi bulunamadı, açılıyor...", "warning")
            self.medula.medula_ac_ve_baglan()

        if not self.medula.bagli or not self.medula.main_window:
            log("Medula'ya bağlanılamadı!", "error")
            self.root.after(0, lambda: self._medula_baglanti_sonuc(False))
            return

        # Oturum kontrol - thread-safe try/except
        oturum_ok = False
        try:
            oturum_ok = self.medula.oturum_aktif_mi()
        except Exception:
            # Access violation veya benzeri hata - devam et, invoke denenecek
            oturum_ok = True  # Aktif kabul et, sorun olursa invoke hata verir

        if not oturum_ok:
            log("Oturum düşmüş, yenileniyor...", "warning")
            for _ in range(5):
                try:
                    for elem in self.medula.main_window.descendants():
                        try:
                            if elem.element_info.automation_id == "btnMedulayaGirisYap":
                                elem.click_input()
                                break
                        except:
                            pass
                except Exception:
                    pass
                time.sleep(4)
                try:
                    if self.medula.oturum_aktif_mi():
                        oturum_ok = True
                        break
                except Exception:
                    oturum_ok = True  # Hata durumunda devam et
                    break

        if not oturum_ok:
            log("Medula oturumu başlatılamadı! Önce MEDULA Bağlan butonunu kullanın.", "error")
            return

        self.root.after(0, lambda: self._medula_baglanti_sonuc(True))
        log("Medula bağlı ve oturum aktif", "success")

        # === 2. REÇETE LİSTESİ → DÖNEM → FATURA TÜRÜ → SORGULA ===
        # NOT: Web elementleri y=0,x=0 olabilir, bu yüzden:
        #   - Menü linkleri → invoke() ile
        #   - Combobox'lar → click_input() + keyboard ile
        #   - Butonlar → invoke() ile

        log("Reçete Listesi sayfasına gidiliyor...", "info")
        for elem in self.medula.main_window.descendants():
            try:
                if elem.element_info.automation_id == "form1:menuHtmlCommandExButton31":
                    elem.invoke()
                    log("Reçete Listesi menüsü açıldı", "success")
                    break
            except:
                pass

        # Sorgula butonunu bekle (sayfa yüklendi mi?) - 30 saniye timeout
        sorgula_var = False
        for bekle in range(30):
            time.sleep(1)
            try:
                for elem in self.medula.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == "form1:buttonSonlandirilmamisReceteler":
                            sorgula_var = True
                            break
                    except:
                        pass
            except Exception:
                pass  # access violation durumunda devam
            if sorgula_var:
                break

        if not sorgula_var:
            # 2. deneme - invoke tekrar
            log("Sayfa yüklenemedi, tekrar deneniyor...", "warning")
            for elem in self.medula.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "form1:menuHtmlCommandExButton31":
                        elem.invoke()
                        break
                except:
                    pass
            time.sleep(10)
            try:
                for elem in self.medula.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == "form1:buttonSonlandirilmamisReceteler":
                            sorgula_var = True
                            break
                    except:
                        pass
            except Exception:
                pass

        if not sorgula_var:
            log("Reçete Listesi sayfası yüklenemedi!", "error")
            return

        # ADIM 1: Dönem seçimi (form1:menu2)
        # click_input ile combobox'a odaklan, down ile eski aya git
        if donem_offset > 0:
            log(f"Dönem: {secilen_donem} seçiliyor...", "info")
            for elem in self.medula.main_window.descendants():
                try:
                    if elem.element_info.automation_id == "form1:menu2":
                        elem.click_input()
                        time.sleep(0.5)
                        for _ in range(donem_offset):
                            pyautogui.press("down")
                            time.sleep(0.15)
                        pyautogui.press("enter")
                        log(f"Dönem seçildi: {secilen_donem}", "success")
                        break
                except:
                    pass
            time.sleep(1)

        # ADIM 2: Fatura Türü combobox (form1:menu1)
        # click_input → 20x up (en başa) → N down → enter
        grup_combo_index = {
            "A": 1,    # A Grubu
            "B": 4,    # B Grubu
            "C": 7,    # C Grubu Sıralı Dağıtım
            "CK": 10,  # C Grubu Kan Ürünü
            "GK": 16,  # Yeşil Kart Normal
        }
        combo_idx = grup_combo_index.get(grup_kodu, 1)

        log(f"Fatura Türü: {grup_adi} seçiliyor...", "info")
        for elem in self.medula.main_window.descendants():
            try:
                if elem.element_info.automation_id == "form1:menu1":
                    elem.click_input()
                    time.sleep(0.5)
                    # En başa git (20x up yeterli)
                    for _ in range(20):
                        pyautogui.press("up")
                        time.sleep(0.03)
                    time.sleep(0.2)
                    # İstenen gruba git
                    for _ in range(combo_idx):
                        pyautogui.press("down")
                        time.sleep(0.1)
                    pyautogui.press("enter")
                    log(f"Fatura Türü seçildi: {grup_adi}", "success")
                    break
            except:
                pass
        time.sleep(1)

        # ADIM 3: Sorgula (invoke ile)
        log("Sorgula butonuna basılıyor...", "info")
        for elem in self.medula.main_window.descendants():
            try:
                if elem.element_info.automation_id == "form1:buttonSonlandirilmamisReceteler":
                    elem.invoke()
                    log("Sorgula tıklandı", "success")
                    break
            except:
                pass
        time.sleep(8)

        # === 3. İLK REÇETEYE TIKLA ===
        ilk_recete = None
        try:
            for elem in self.medula.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    if txt:
                        txt = txt.strip()
                        if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                            elem.click_input()
                            ilk_recete = txt
                            break
                except:
                    pass
        except Exception:
            pass

        if not ilk_recete:
            log(f"{grup_adi}: Reçete bulunamadı!", "warning")
            return

        log(f"İlk reçete açılıyor: {ilk_recete}", "info")

        # İlaç tablosu yüklenene kadar bekle (ilaç formu anahtar kelimesi görünene kadar)
        for bekle in range(15):
            time.sleep(1)
            try:
                for elem in self.medula.main_window.descendants(control_type="DataItem"):
                    try:
                        txt = elem.window_text()
                        if txt and any(k in txt.upper() for k in ["MG", "TABLET", "FTB", "KAPSUL", "FLAKON", "DOZ", "KREM"]):
                            break
                    except:
                        pass
                else:
                    continue
                break
            except Exception:
                pass
        time.sleep(2)

        # === 4. KONTROL MOTORU + RAPOR TABLOSU ===
        motor = ReceteKontrolMotoru(log_callback=lambda msg, tag="info": log(msg, tag))
        rapor_satirlari = []  # Excel rapor için

        sayac = 0
        onceki_ilaclar_str = None  # İlaç listesi değişimi ile tekrar tespiti
        tekrar_sayaci = 0

        while self.kontrol_aktif and sayac < 300:
            sayac += 1

            # Reçete no oku - koordinat kullanmadan, tüm 7 haneli alfanümerik text'lerden
            # İlk bulunan (genelde e-Reçete no veya takip no)
            recete_no = None
            try:
                for elem in self.medula.main_window.descendants(control_type="DataItem"):
                    try:
                        txt = elem.window_text()
                        if txt:
                            txt = txt.strip()
                            if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                                recete_no = txt
                                break
                    except:
                        pass
            except Exception:
                pass

            if not recete_no:
                # Sayfa henüz yüklenmemiş olabilir, biraz bekle
                time.sleep(3)
                try:
                    for elem in self.medula.main_window.descendants(control_type="DataItem"):
                        try:
                            txt = elem.window_text()
                            if txt:
                                txt = txt.strip()
                                if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                                    recete_no = txt
                                    break
                        except:
                            pass
                except Exception:
                    pass

            if not recete_no:
                log("Reçete no okunamadı - tarama tamamlandı", "info")
                break

            # Tekrar kontrolü - aynı ilaç listesi mi?
            if recete_no == onceki_ilaclar_str:
                tekrar_sayaci += 1
                if tekrar_sayaci >= 3:
                    log(f"Reçete değişmiyor ({recete_no}) - son sayfaya ulaşıldı", "info")
                    break
                # Tekrar dene
                for elem in self.medula.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == "f:buttonSonraki":
                            elem.invoke()
                            break
                    except:
                        pass
                time.sleep(5)
                continue
            else:
                tekrar_sayaci = 0
                onceki_ilaclar_str = recete_no

            log(f"", "info")
            log(f"{'━' * 35}", "header")
            log(f"Reçete #{sayac}: {recete_no}", "header")

            # === 4a. REÇETE TÜRÜ KONTROL ===
            recete_turu = "Normal"
            for elem in self.medula.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    r = elem.rectangle()
                    if txt and r.top > 310 and r.top < 340:
                        txt = txt.strip()
                        if txt in ["Normal", "Kırmızı", "Yeşil", "Turuncu", "Mor"]:
                            recete_turu = txt
                            break
                except:
                    pass
            log(f"  Reçete türü: {recete_turu}", "info")

            # === 4a-2. HASTA ADI OKU (best-effort, üst alandan) ===
            hasta_adi = ""
            HASTA_KARA_LISTE = (
                "müstehaklık", "mua.", "reç.kat", "il.kat", "toplam",
                "ödenecek", "sayfaya", "incelemeye", "botrastan",
                "kontrol", "uzman", "sorgu", "verilemez", "uyumlu",
                "öncelik", "raporsuz", "raporda", "girilebilcek",
                "endikasyon", "günde", "haftada", "ayda", "saatte",
                "katılım", "kapsam", "fatura", "tarih", "kalem",
                "sigortalı", "kurum", "doktor", "hekim", "branş",
                "tesis", "tahsil",
            )
            try:
                ad_adaylari = []
                for elem in self.medula.main_window.descendants(control_type="DataItem"):
                    try:
                        txt = elem.window_text()
                        r = elem.rectangle()
                        if not txt:
                            continue
                        txt = txt.strip()
                        if not txt or r.top >= 310:
                            continue
                        # Ad-soyad heuristiği:
                        if " " not in txt or len(txt) < 5 or len(txt) > 60:
                            continue
                        if any(ch.isdigit() for ch in txt):
                            continue
                        if any(ch in txt for ch in (":", "(", ")", "/", "=", "%", ",")):
                            continue
                        txt_lower = txt.lower()
                        if any(k in txt_lower for k in HASTA_KARA_LISTE):
                            continue
                        if txt in ["Normal", "Kırmızı", "Yeşil", "Turuncu", "Mor"]:
                            continue
                        # Adın çoğunlukla büyük harfli olduğunu varsay
                        harf_sayisi = sum(1 for c in txt if c.isalpha())
                        buyuk_sayi = sum(1 for c in txt if c.isalpha() and c.isupper())
                        if harf_sayisi == 0 or buyuk_sayi / harf_sayisi < 0.6:
                            continue
                        ad_adaylari.append((r.top, r.left, txt))
                    except:
                        pass
                if ad_adaylari:
                    ad_adaylari.sort(key=lambda a: (a[0], a[1]))
                    hasta_adi = ad_adaylari[0][2]
            except Exception:
                pass

            # === KONTROL SEKMESİ HEADER ===
            kontrol(f"", "info")
            kontrol(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", "header")
            _hasta_str = hasta_adi if hasta_adi else "—"
            kontrol(f"Reçete #{sayac}: {recete_no}  |  Hasta: {_hasta_str}  |  Türü: {recete_turu}", "header")
            kontrol(f"────────────────────────────────────────────────────────", "info")

            # === 4b. RENKLİ REÇETE KONTROL ===
            renkli_durum = ""
            if recete_turu in ["Kırmızı", "Yeşil", "Mor"]:
                if not self.renkli_recete_listesi:
                    renkli_durum = "UYARI: Renkli reçete listesi yüklenmemiş!"
                    log(f"  ⚠ {renkli_durum}", "error")
                    kontrol(f"  ⚠ Renkli reçete listesi: YÜKLENMEMİŞ", "error")
                else:
                    # e-Reçete no ile ara
                    bulundu, kayit = self.renkli_recete_kontrol(recete_no)
                    if not bulundu:
                        # Medula'daki e-Reçete no'yu da dene
                        erecete_no = None
                        for elem in self.medula.main_window.descendants(control_type="DataItem"):
                            try:
                                txt = elem.window_text()
                                r = elem.rectangle()
                                if txt and r.top > 290 and r.top < 310 and r.left > 870:
                                    txt = txt.strip()
                                    if len(txt) >= 5 and txt.isalnum():
                                        erecete_no = txt
                                        break
                            except:
                                pass
                        if erecete_no:
                            bulundu, kayit = self.renkli_recete_kontrol(erecete_no)

                    if bulundu:
                        renkli_durum = "Renkli reçete sistemine işlenmiş ✓"
                        log(f"  ✓ {renkli_durum}", "success")
                        kontrol(f"  ✓ Renkli reçete listesi: KAYITLI ({recete_turu})", "success")
                    else:
                        renkli_durum = "SORUN: Renkli reçete sistemine İŞLENMEMİŞ!"
                        log(f"  ✗ {renkli_durum}", "error")
                        kontrol(f"  ✗ Renkli reçete listesi: KAYITSIZ ({recete_turu}) — eksik!", "error")
            else:
                kontrol(f"  — Renkli reçete kontrolü gerekmiyor (Normal reçete)", "neutral")

            # === 4c. İLAÇ SATIRLARINI OKU (satır bazlı) ===
            ilaclar = []
            items = []
            for elem in self.medula.main_window.descendants(control_type="DataItem"):
                try:
                    txt = elem.window_text()
                    r = elem.rectangle()
                    if txt and txt.strip() and r.top > 450 and r.top < 920:
                        items.append((txt.strip()[:100], r.top, r.left))
                except:
                    pass
            items.sort(key=lambda x: (x[1], x[2]))

            ATLA = ["Maksimum", "Topl=", "Toplam Tutar", "Sayfaya",
                    "İncelemeye", "Botrastan", "Reçete Tutar", "YAZMIŞ",
                    "YAZMAMIŞ", "VERİLEMEZ", "KONTROL EDİNİZ", "UZMAN",
                    "Uyumlu ICD", "Raporsuz", "Raporda DOZ", "Öncelik",
                    "Girilebilcek", "endikasyon", "Maksimum Süre",
                    "Reç.Kat", "İl.Kat", "Mua.Kat", "Ödenecek",
                    "Günde", "Haftada", "Ayda", "Saatte", " x ",
                    "GETİRİN"]

            import re as _re
            TARIH_RE = _re.compile(r"^\d{2}/\d{2}/\d{4}$")
            FIYAT_RE = _re.compile(r"^\d+(?:[.,]\d+)?$")
            RAPOR_KODU_RE = _re.compile(r"^\d{1,2}\.\d{1,2}$")

            # Items'ları y'ye göre satırlara grupla (12px tolerans)
            satirlar = []
            mevcut_satir = []
            mevcut_y = None
            for txt, y, x in items:
                if mevcut_y is None or abs(y - mevcut_y) <= 12:
                    mevcut_satir.append((txt, y, x))
                    if mevcut_y is None:
                        mevcut_y = y
                else:
                    if mevcut_satir:
                        satirlar.append((mevcut_y, mevcut_satir))
                    mevcut_satir = [(txt, y, x)]
                    mevcut_y = y
            if mevcut_satir:
                satirlar.append((mevcut_y, mevcut_satir))

            # Her satırı analiz et
            for satir_y, satir in satirlar:
                texts = [t for t, _, _ in satir]
                has_var_yok = any(t.lower() in ("var", "yok") for t in texts)
                has_tarih = any(TARIH_RE.match(t) for t in texts)
                if not (has_var_yok and has_tarih):
                    continue
                # İlaç adı: en uzun, sayı/tarih/var-yok değil, ATLA listesinde yok
                ad_aday = ""
                for t in texts:
                    if (len(t) > 10
                            and not TARIH_RE.match(t)
                            and not FIYAT_RE.match(t)
                            and not RAPOR_KODU_RE.match(t)
                            and t.lower() not in ("var", "yok")
                            and not t.startswith("SGK")
                            and not any(a in t for a in ATLA)):
                        if len(t) > len(ad_aday):
                            ad_aday = t
                if not ad_aday:
                    continue

                rapor_kodu = ""
                msj = ""
                for t in texts:
                    if RAPOR_KODU_RE.match(t):
                        rapor_kodu = t
                    if t.lower() in ("var", "yok"):
                        msj = t.lower()

                ilaclar.append({
                    "ilac_adi": ad_aday,
                    "etkin_madde": "", "sgk_kodu": "",
                    "rapor_kodu": rapor_kodu, "msj": msj,
                    "_y": satir_y,
                })

            # Etkin madde / SGK kodu: ilaç satırının altındaki "SGKXXX-..." satırını eşle
            for txt, y, x in items:
                if not (txt.startswith("SGK") and "-" in txt):
                    continue
                # Bu y'den küçük en yakın ilaç satırını bul
                aday_ilac = None
                aday_fark = 99999
                for il in ilaclar:
                    iy = il.get("_y", 0)
                    fark = y - iy
                    if 0 < fark < 80 and fark < aday_fark:
                        aday_fark = fark
                        aday_ilac = il
                if aday_ilac is not None:
                    parts = txt.split("-", 1)
                    aday_ilac["sgk_kodu"] = parts[0].strip()
                    aday_ilac["etkin_madde"] = parts[1].strip() if len(parts) > 1 else ""

            # Geçici _y alanını temizle
            for il in ilaclar:
                il.pop("_y", None)

            if not ilaclar:
                log(f"  İlaç okunamadı, sonrakine geçiliyor", "warning")
                kontrol(f"  ⚠ İlaç tablosu okunamadı — manuel kontrol gerekli", "warning")
                rapor_satirlari.append({
                    "recete_no": recete_no, "recete_turu": recete_turu,
                    "renkli_durum": renkli_durum, "ilac_adi": "-",
                    "etkin_madde": "-", "rapor_kodu": "-",
                    "sonuc": "Bakılamadı", "aciklama": "İlaç okunamadı"
                })
            else:
                log(f"  {len(ilaclar)} ilaç bulundu:", "info")

                # === KONTROL ÇEKLİSTİ (Kontrol Sonuçları sekmesi) ===
                # Yoksay listesindeki ilaçların msj'sini etkisiz hale getir
                for _il in ilaclar:
                    if _il.get("msj") == "var" and self.mesaj_yoksay_eslesti(_il.get("ilac_adi", "")):
                        _il["msj_yoksayildi"] = True

                raporlu_ilaclar = [i for i in ilaclar if i.get("rapor_kodu")]
                mesajli_ilaclar = [
                    i for i in ilaclar
                    if i.get("msj") == "var" and not i.get("msj_yoksayildi")
                ]
                # Aynı ilaç hem raporlu hem mesajlı olabilir; tekrar saymayalım
                kontrolsuz_sayi = sum(
                    1 for i in ilaclar
                    if not i.get("rapor_kodu")
                    and (i.get("msj") != "var" or i.get("msj_yoksayildi"))
                )
                kontrol(f"  Toplam {len(ilaclar)} ilaç: "
                        f"{len(raporlu_ilaclar)} raporlu, "
                        f"{len(mesajli_ilaclar)} mesajlı, "
                        f"{kontrolsuz_sayi} kontrolsüz", "info")

                if not raporlu_ilaclar:
                    kontrol(f"  — Raporlu ilaç doz kontrolü gerekmiyor", "neutral")
                if not mesajli_ilaclar:
                    kontrol(f"  — İlaç mesajı kontrolü gerekmiyor", "neutral")

                # TODO: Uyarı kodu ve endikasyon dışı rapor tespitleri eklenecek
                kontrol(f"  — Uyarı kodu kontrolü: tespit henüz aktif değil", "neutral")
                kontrol(f"  — Endikasyon dışı rapor kontrolü: tespit henüz aktif değil", "neutral")

                # === 4d. HER İLAÇ İÇİN ALGORİTMİK KONTROL ===
                for ilac in ilaclar:
                    etkin = ilac.get("etkin_madde", "")
                    rapor_kodu = ilac.get("rapor_kodu", "")
                    msj = ilac.get("msj", "")

                    # Yoksay listesindeki ilaçlar için mesaj VAR olsa bile yok say
                    if ilac.get("msj_yoksayildi"):
                        log(f"    [YOKSAY] {ilac['ilac_adi'][:35]} → Mesaj yoksay listesinde", "info")
                        if not rapor_kodu:
                            continue
                        msj = ""  # raporlu kontrole devam ama mesajı görmezden gel

                    # Raporsuz + mesajsız → kontrole gerek yok, atla
                    if not rapor_kodu and msj != "var":
                        log(f"    [SKIP] {ilac['ilac_adi'][:35]} → Raporsuz, mesajsız", "info")
                        continue

                    # DB'de kural var mı?
                    kural = None
                    if etkin:
                        kural = motor.kural_bul(etkin)
                        if not kural and ilac.get("sgk_kodu"):
                            kural = motor.kural_bul_sgk_kodu(ilac["sgk_kodu"])

                    # Yeni kural öğren
                    if not kural and etkin:
                        raporlu = 1 if rapor_kodu else 0
                        tip = "rapor_kontrolu" if raporlu else "raporsuz_verilebilir"
                        motor.yeni_kural_ekle(
                            etkin_madde=etkin,
                            sgk_kodu=ilac.get("sgk_kodu", ""),
                            rapor_kodu=rapor_kodu,
                            rapor_gerekli=raporlu,
                            kontrol_tipi=tip,
                            aciklama=f"Otomatik: {ilac['ilac_adi']}"
                        )
                        log(f"    [YENİ] {ilac['ilac_adi'][:40]} → {etkin}", "warning")

                    # Kontrol sonucu belirle
                    sonuc = "Uygun"
                    aciklama_str = ""
                    kontrol_tag = "info"

                    if kural:
                        if kural["rapor_gerekli"] and not rapor_kodu:
                            sonuc = "Uygun Değil"
                            aciklama_str = f"RAPOR GEREKLİ! SUT {kural.get('sut_maddesi', '?')}"
                            kontrol_tag = "error"
                            log(f"    [SORUN] {ilac['ilac_adi'][:35]} → {aciklama_str}", "error")
                        elif kural["rapor_gerekli"] and rapor_kodu:
                            sonuc = "Uygun"
                            aciklama_str = f"Raporlu ({rapor_kodu})"
                            kontrol_tag = "success"
                            log(f"    [  OK ] {ilac['ilac_adi'][:35]} → {aciklama_str}", "info")
                        else:
                            sonuc = "Uygun"
                            aciklama_str = "Raporsuz verilebilir"
                            kontrol_tag = "success"
                            log(f"    [  OK ] {ilac['ilac_adi'][:35]} → {aciklama_str}", "info")
                    elif etkin:
                        sonuc = "Şüpheli"
                        aciklama_str = "Yeni öğrenildi, SUT kontrolü yapılmadı"
                        kontrol_tag = "warning"
                        log(f"    [  ?  ] {ilac['ilac_adi'][:35]} → {aciklama_str}", "warning")
                    else:
                        sonuc = "Bakılamadı"
                        aciklama_str = "Etkin madde okunamadı"
                        kontrol_tag = "warning"
                        log(f"    [ --- ] {ilac['ilac_adi'][:35]} → {aciklama_str}", "warning")

                    if msj == "var":
                        aciklama_str += " | Mesaj VAR"

                    # Kontrol Sonuçları sekmesine yaz
                    isaret = "✓" if sonuc == "Uygun" else ("✗" if sonuc == "Uygun Değil" else "?")
                    etiketler = []
                    if rapor_kodu:
                        etiketler.append(f"rapor: {rapor_kodu}")
                    if msj == "var":
                        etiketler.append("mesaj: VAR")
                    etiket_str = " | ".join(etiketler) if etiketler else ""
                    kontrol(
                        f"  {isaret} {ilac['ilac_adi'][:50]}"
                        + (f"  ({etiket_str})" if etiket_str else "")
                        + f"  → {aciklama_str}",
                        kontrol_tag,
                    )

                    # === 4e. RAPOR TABLOSUNA YAZ ===
                    rapor_satirlari.append({
                        "recete_no": recete_no,
                        "recete_turu": recete_turu,
                        "renkli_durum": renkli_durum,
                        "ilac_adi": ilac["ilac_adi"][:50],
                        "etkin_madde": etkin,
                        "sgk_kodu": ilac.get("sgk_kodu", ""),
                        "rapor_kodu": rapor_kodu,
                        "msj": msj,
                        "sonuc": sonuc,
                        "aciklama": aciklama_str,
                    })

            # Durumu kaydet
            self.root.after(0, lambda r=recete_no, g=grup_kodu: self.grup_durumu_guncelle(g, r))

            # === 4f. SONRAKİ REÇETEYE GEÇ (invoke ile) ===
            sonraki_ok = False
            try:
                for elem in self.medula.main_window.descendants():
                    try:
                        if elem.element_info.automation_id == "f:buttonSonraki":
                            elem.invoke()
                            sonraki_ok = True
                            break
                    except:
                        pass
            except Exception:
                pass
            if not sonraki_ok:
                log("Sonraki butonu bulunamadı - tarama bitti", "info")
                break

            # Sayfa yüklenene kadar bekle (ilaç formu kelimesi görünene kadar)
            for bekle in range(12):
                time.sleep(1)
                try:
                    for elem in self.medula.main_window.descendants(control_type="DataItem"):
                        try:
                            txt = elem.window_text()
                            if txt and any(k in txt.upper() for k in ["MG", "TABLET", "FTB", "KAPSUL", "KREM"]):
                                break
                        except:
                            pass
                    else:
                        continue
                    break
                except Exception:
                    pass
            time.sleep(1)

        motor.kapat()

        # === 5. EXCEL RAPOR OLUŞTUR ===
        log(f"", "header")
        log(f"{'━' * 40}", "header")
        log(f"TARAMA TAMAMLANDI - {sayac} reçete kontrol edildi", "header")

        if rapor_satirlari:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"{grup_adi} Kontrol"

                # Başlıklar
                basliklar = ["Reçete No", "Tür", "Renkli Reçete", "İlaç Adı",
                             "Etkin Madde", "SGK Kodu", "Rapor Kodu", "Msj",
                             "Sonuç", "Açıklama"]
                for col, b in enumerate(basliklar, 1):
                    c = ws.cell(row=1, column=col, value=b)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill(start_color="1E3A5F", fill_type="solid")

                # Veriler
                renk_map = {
                    "Uygun": "C8E6C9",        # Yeşil
                    "Uygun Değil": "FFCDD2",   # Kırmızı
                    "Şüpheli": "FFF9C4",       # Sarı
                    "Bakılamadı": "E0E0E0",    # Gri
                }
                for row_idx, satir in enumerate(rapor_satirlari, 2):
                    ws.cell(row=row_idx, column=1, value=satir["recete_no"])
                    ws.cell(row=row_idx, column=2, value=satir["recete_turu"])
                    ws.cell(row=row_idx, column=3, value=satir["renkli_durum"])
                    ws.cell(row=row_idx, column=4, value=satir["ilac_adi"])
                    ws.cell(row=row_idx, column=5, value=satir["etkin_madde"])
                    ws.cell(row=row_idx, column=6, value=satir.get("sgk_kodu", ""))
                    ws.cell(row=row_idx, column=7, value=satir["rapor_kodu"])
                    ws.cell(row=row_idx, column=8, value=satir.get("msj", ""))
                    sonuc_cell = ws.cell(row=row_idx, column=9, value=satir["sonuc"])
                    ws.cell(row=row_idx, column=10, value=satir["aciklama"])
                    # Renklendirme
                    renk = renk_map.get(satir["sonuc"], "FFFFFF")
                    sonuc_cell.fill = PatternFill(start_color=renk, fill_type="solid")

                # Sütun genişlikleri
                for col_letter, width in [("A", 12), ("B", 8), ("C", 30), ("D", 45),
                                          ("E", 25), ("F", 12), ("G", 10), ("H", 5),
                                          ("I", 12), ("J", 45)]:
                    ws.column_dimensions[col_letter].width = width

                rapor_dosya = os.path.join(PROJE_DIZINI,
                    f"Kontrol_Raporu_{grup_kodu}_{secilen_donem.replace(' ', '_')}.xlsx")
                wb.save(rapor_dosya)
                log(f"Rapor kaydedildi: {rapor_dosya}", "success")

                # İstatistik
                uygun = sum(1 for s in rapor_satirlari if s["sonuc"] == "Uygun")
                sorunlu = sum(1 for s in rapor_satirlari if s["sonuc"] == "Uygun Değil")
                supheli = sum(1 for s in rapor_satirlari if s["sonuc"] == "Şüpheli")
                log(f"  Toplam ilaç: {len(rapor_satirlari)}", "info")
                log(f"  ✓ Uygun: {uygun}", "success")
                if sorunlu > 0:
                    log(f"  ✗ Sorunlu: {sorunlu}", "error")
                if supheli > 0:
                    log(f"  ? Şüpheli: {supheli}", "warning")
            except Exception as e:
                log(f"Rapor oluşturma hatası: {e}", "error")

        log(f"{'━' * 40}", "header")

    def _kontrol_ayarlari_penceresi(self):
        """Kontrol bypass ayarları penceresi.
        İlaç/etkin madde/uyarı kodu bazında kontrol atla/yap ayarları."""
        from kontrol_kurallari import get_kontrol_ayarlari, HEDEF_TIPI_ILAC, \
            HEDEF_TIPI_ETKIN_MADDE, HEDEF_TIPI_ATC_GRUP, HEDEF_TIPI_UYARI_KODU, \
            KONTROL_TIPI_UYARI_KODU, KONTROL_TIPI_SUT, KONTROL_TIPI_ILAC_MESAJI, \
            KONTROL_TIPI_DOZ, KONTROL_TIPI_HEPSI

        ayarlar = get_kontrol_ayarlari()

        win = tk.Toplevel(self.root)
        win.title("Kontrol Ayarları (Bypass)")
        win.geometry("750x550")
        win.configure(bg="#1E3A5F")
        win.transient(self.root)

        # === ÜST: Yeni Ayar Ekleme ===
        ekle_frame = tk.LabelFrame(win, text="Yeni Ayar Ekle", font=("Segoe UI", 10, "bold"),
                                    fg="#87CEEB", bg="#1E3A5F", padx=10, pady=8)
        ekle_frame.pack(fill="x", padx=10, pady=(10, 5))

        # Satır 1: Hedef tipi + değer
        s1 = tk.Frame(ekle_frame, bg="#1E3A5F")
        s1.pack(fill="x", pady=2)

        tk.Label(s1, text="Hedef:", fg="white", bg="#1E3A5F", font=("Segoe UI", 9)).pack(side="left")
        hedef_tipi_var = tk.StringVar(value="ilac")
        hedef_combo = tk.OptionMenu(s1, hedef_tipi_var,
                                     "ilac", "etkin_madde", "uyari_kodu", "atc_grup")
        hedef_combo.config(font=("Segoe UI", 9), bg="#2C4A6E", fg="white", width=12)
        hedef_combo.pack(side="left", padx=5)

        tk.Label(s1, text="Değer:", fg="white", bg="#1E3A5F", font=("Segoe UI", 9)).pack(side="left")
        from tkinter import ttk as _ttk
        hedef_deger_var = tk.StringVar()
        hedef_deger_combo = _ttk.Combobox(s1, textvariable=hedef_deger_var,
                                           font=("Segoe UI", 9), width=35)
        hedef_deger_combo.pack(side="left", padx=5)

        # Öğrenilen ilaçlar DB'sinden değerleri yükle
        _ogrenilen_ilaclar = []
        _ogrenilen_etkin_maddeler = []
        try:
            from kontrol_kurallari import get_ogrenilen_ilaclar
            ogrenilen = get_ogrenilen_ilaclar()
            tum = ogrenilen.tum_ilaclar()
            _ogrenilen_ilaclar = sorted(set(i["ilac_adi"] for i in tum if i.get("ilac_adi")))
            _ogrenilen_etkin_maddeler = sorted(set(
                i["etkin_madde"] for i in tum if i.get("etkin_madde")))
        except Exception:
            pass

        # Öğrenilen uyarı kodlarını DB'den yükle
        _ogrenilen_uyari_kodlari = []
        try:
            import sqlite3 as _sql
            _db_path = os.path.join(PROJE_DIZINI, "kontrol_kurallari.db")
            _conn = _sql.connect(_db_path)
            _cur = _conn.cursor()
            _uyari_set = set()
            # ogrenilen_uyari_kodlari tablosundan (taramada tespit edilenler)
            try:
                _cur.execute("SELECT kod, aciklama FROM ogrenilen_uyari_kodlari ORDER BY gorulme_sayisi DESC")
                for row in _cur.fetchall():
                    if row[0]:
                        _uyari_set.add(f"{row[0]} - {row[1]}" if row[1] else row[0])
            except Exception:
                pass
            # Mevcut ayarlardaki uyarı kodlarını da ekle
            _cur.execute("SELECT DISTINCT hedef_deger FROM ilac_kontrol_ayarlari WHERE hedef_tipi='uyari_kodu'")
            for row in _cur.fetchall():
                if row[0]:
                    _uyari_set.add(row[0])
            _conn.close()
            _ogrenilen_uyari_kodlari = sorted(_uyari_set, key=lambda x: int(x) if x.isdigit() else 0)
        except Exception:
            pass

        def hedef_tipi_degisti(*_args):
            tip = hedef_tipi_var.get()
            if tip == "ilac":
                hedef_deger_combo["values"] = _ogrenilen_ilaclar
            elif tip == "etkin_madde":
                hedef_deger_combo["values"] = _ogrenilen_etkin_maddeler
            elif tip == "uyari_kodu":
                hedef_deger_combo["values"] = _ogrenilen_uyari_kodlari
            else:
                hedef_deger_combo["values"] = []
            hedef_deger_var.set("")
            # Uyarı kodu seçilince kontrol tipini otomatik uyari_kodu yap
            if tip == "uyari_kodu":
                kontrol_tipi_var.set("uyari_kodu")

        hedef_tipi_var.trace_add("write", hedef_tipi_degisti)
        hedef_tipi_degisti()  # İlk yükleme

        # Satır 2: Kontrol tipi + kaydet butonu
        s2 = tk.Frame(ekle_frame, bg="#1E3A5F")
        s2.pack(fill="x", pady=2)

        tk.Label(s2, text="Kontrol:", fg="white", bg="#1E3A5F", font=("Segoe UI", 9)).pack(side="left")
        kontrol_tipi_var = tk.StringVar(value="hepsi")
        kontrol_combo = tk.OptionMenu(s2, kontrol_tipi_var,
                                       "hepsi", "uyari_kodu", "sut", "ilac_mesaji", "doz")
        kontrol_combo.config(font=("Segoe UI", 9), bg="#2C4A6E", fg="white", width=12)
        kontrol_combo.pack(side="left", padx=5)

        tk.Label(s2, text="İşlem:", fg="white", bg="#1E3A5F", font=("Segoe UI", 9)).pack(side="left")
        islem_var = tk.StringVar(value="pasif")
        islem_combo = tk.OptionMenu(s2, islem_var, "pasif", "aktif")
        islem_combo.config(font=("Segoe UI", 9), bg="#2C4A6E", fg="white", width=8)
        islem_combo.pack(side="left", padx=5)

        # Açıklama
        aciklama_entry = tk.Entry(s2, font=("Segoe UI", 9), width=20, bg="#0D2137", fg="white",
                                   insertbackground="white")
        aciklama_entry.insert(0, "Açıklama (opsiyonel)")
        aciklama_entry.bind("<FocusIn>", lambda e: aciklama_entry.delete(0, "end")
                            if aciklama_entry.get() == "Açıklama (opsiyonel)" else None)
        aciklama_entry.pack(side="left", padx=5)

        def kaydet():
            h_tipi = hedef_tipi_var.get()
            h_deger = hedef_deger_var.get().strip()
            k_tipi = kontrol_tipi_var.get()
            aktif = islem_var.get() == "aktif"
            aciklama = aciklama_entry.get().strip()
            if aciklama == "Açıklama (opsiyonel)":
                aciklama = ""
            if not h_deger:
                messagebox.showwarning("Uyarı", "Hedef değer boş olamaz!", parent=win)
                return
            # Uyarı kodu "256 - Benign prostat" formatında → sadece kodu al
            if h_tipi == "uyari_kodu" and " - " in h_deger:
                h_deger = h_deger.split(" - ")[0].strip()
            ayarlar.ayar_kaydet(h_tipi, h_deger, k_tipi, aktif=aktif, aciklama=aciklama or None)
            hedef_deger_var.set("")
            tabloyu_yenile()
            durum = "PASİF (bypass)" if not aktif else "AKTİF"
            self.log_yaz(f"Ayar kaydedildi: {h_tipi}={h_deger} → {k_tipi} = {durum}", "success")

        tk.Button(s2, text="Kaydet", font=("Segoe UI", 9, "bold"), fg="white", bg="#2E7D32",
                  activebackground="#1B5E20", bd=0, padx=15, pady=3, command=kaydet).pack(side="right")

        # === TABLO: Mevcut Ayarlar ===
        tablo_frame = tk.LabelFrame(win, text="Mevcut Ayarlar", font=("Segoe UI", 10, "bold"),
                                     fg="#87CEEB", bg="#1E3A5F", padx=5, pady=5)
        tablo_frame.pack(fill="both", expand=True, padx=10, pady=5)

        from tkinter import ttk
        style = ttk.Style()
        style.configure("Ayar.Treeview", background="#0D2137", foreground="white",
                         fieldbackground="#0D2137", font=("Segoe UI", 9))
        style.configure("Ayar.Treeview.Heading", font=("Segoe UI", 9, "bold"),
                         background="#1E3A5F", foreground="white")

        sutunlar = ("hedef_tipi", "hedef_deger", "kontrol_tipi", "durum", "aciklama")
        tree = ttk.Treeview(tablo_frame, columns=sutunlar, show="headings",
                             style="Ayar.Treeview", height=15)
        tree.heading("hedef_tipi", text="Hedef Tipi")
        tree.heading("hedef_deger", text="Değer")
        tree.heading("kontrol_tipi", text="Kontrol")
        tree.heading("durum", text="Durum")
        tree.heading("aciklama", text="Açıklama")
        tree.column("hedef_tipi", width=100)
        tree.column("hedef_deger", width=200)
        tree.column("kontrol_tipi", width=100)
        tree.column("durum", width=80)
        tree.column("aciklama", width=200)

        scrollbar = ttk.Scrollbar(tablo_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def tabloyu_yenile():
            tree.delete(*tree.get_children())
            for a in ayarlar.tum_ayarlar():
                durum = "PASİF ✕" if not a["aktif"] else "AKTİF ✓"
                tree.insert("", "end", values=(
                    a["hedef_tipi"], a["hedef_deger"], a["kontrol_tipi"],
                    durum, a.get("aciklama", "") or ""
                ))

        # Alt butonlar
        alt_frame = tk.Frame(win, bg="#1E3A5F")
        alt_frame.pack(fill="x", padx=10, pady=(0, 10))

        def secili_sil():
            secim = tree.selection()
            if not secim:
                messagebox.showwarning("Uyarı", "Bir ayar seçin!", parent=win)
                return
            for item in secim:
                vals = tree.item(item, "values")
                ayarlar.ayar_sil(vals[0], vals[1], vals[2])
            tabloyu_yenile()
            self.log_yaz(f"{len(secim)} ayar silindi", "info")

        def secili_toggle():
            secim = tree.selection()
            if not secim:
                return
            for item in secim:
                vals = tree.item(item, "values")
                yeni_aktif = "PASİF" in vals[3]  # PASİF ise aktif yap, AKTİF ise pasif yap
                ayarlar.ayar_kaydet(vals[0], vals[1], vals[2], aktif=yeni_aktif)
            tabloyu_yenile()

        tk.Button(alt_frame, text="Seçili Sil", font=("Segoe UI", 9), fg="white", bg="#B71C1C",
                  activebackground="#7F0000", bd=0, padx=12, pady=3, command=secili_sil).pack(side="left")
        tk.Button(alt_frame, text="Aktif/Pasif Değiştir", font=("Segoe UI", 9), fg="white", bg="#E65100",
                  activebackground="#BF360C", bd=0, padx=12, pady=3, command=secili_toggle).pack(side="left", padx=5)

        # Bilgi etiketi
        tk.Label(alt_frame, text="PASİF = kontrol atlanır (bypass) | AKTİF = kontrol yapılır",
                 fg="#90CAF9", bg="#1E3A5F", font=("Segoe UI", 8)).pack(side="right")

        tabloyu_yenile()

    def _acik_receteden_basla(self):
        """Medula'da açık olan reçeteden kontrole başla. Hangi grup seçileceğini sor."""
        if self.kontrol_aktif:
            self.log_yaz("Zaten bir kontrol devam ediyor!", "warning")
            return

        # Grup seçim penceresi
        secim_win = tk.Toplevel(self.root)
        secim_win.title("Grup Seç")
        secim_win.geometry("250x200")
        secim_win.configure(bg="#1E3A5F")
        secim_win.transient(self.root)
        secim_win.grab_set()

        tk.Label(secim_win, text="Hangi grup için kontrol?",
                 font=("Segoe UI", 11, "bold"), fg="white", bg="#1E3A5F").pack(pady=10)

        for grup in GRUP_TANIMLARI:
            tk.Button(
                secim_win, text=grup["ad"],
                font=("Segoe UI", 10, "bold"), fg="white", bg=grup["renk"],
                activebackground=grup["hover"], bd=0, padx=15, pady=5, width=20,
                command=lambda k=grup["kod"]: (secim_win.destroy(),
                                                self._grup_kontrol_baslat(k, acik_receteden=True))
            ).pack(pady=2)

    def _hafizayi_temizle(self):
        """Tüm grup hafızalarını temizle"""
        cevap = messagebox.askyesno(
            "Hafızayı Temizle",
            "Tüm grup kontrol hafızaları temizlenecek.\nDevam etmek istiyor musunuz?"
        )
        if cevap:
            for kod in self.grup_durumlari:
                self.grup_durumlari[kod] = {
                    "son_recete": "",
                    "toplam_kontrol": 0,
                    "son_kontrol_tarihi": None,
                }
            self._durumlari_kaydet()
            self._recete_combolar_temizle()
            for kod in self.durum_labels:
                self._durum_label_guncelle(kod)
            self.log_yaz("Hafıza temizlendi (combobox'lar dahil)", "success")

    def _ana_menuye_don(self):
        """Ana menüye geri dön"""
        self.root.destroy()
        if self.ana_menu_callback:
            self.ana_menu_callback()

    def grup_durumu_guncelle(self, grup_kodu, son_recete, toplam_kontrol_arttir=1):
        """Bir grubun durumunu güncelle ve kaydet"""
        if grup_kodu not in self.grup_durumlari:
            self.grup_durumlari[grup_kodu] = {"son_recete": "", "toplam_kontrol": 0, "son_kontrol_tarihi": None}

        self.grup_durumlari[grup_kodu]["son_recete"] = son_recete
        self.grup_durumlari[grup_kodu]["toplam_kontrol"] += toplam_kontrol_arttir
        self.grup_durumlari[grup_kodu]["son_kontrol_tarihi"] = datetime.now().isoformat()
        self._durumlari_kaydet()
        self._durum_label_guncelle(grup_kodu)
