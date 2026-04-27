# -*- coding: utf-8 -*-
"""
Reçete Tarama Scripti - Subprocess olarak GUI'den çağrılır
Medula'daki reçeteleri tarar, ilaçları okur, SUT kurallarını kontrol eder.
Kullanım: python recete_tarama.py <grup> <donem_offset>
  grup: A, B, C, CK, GK
  donem_offset: 0=bu ay, 1=önceki ay, 2=iki ay önce...

Durdurma: tarama_stop.flag dosyası oluşturulursa tarama durur.
"""

import sys
import os
import time
import json
import sqlite3
import threading
import atexit
import pyautogui
from datetime import datetime
from pywinauto import Desktop

try:
    import win32process
    import win32api
    import win32con
except ImportError:
    win32process = None
    win32api = None
    win32con = None

sys.stdout.reconfigure(encoding='utf-8')
pyautogui.FAILSAFE = False

PROJE_DIZINI = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(PROJE_DIZINI, "kontrol_kurallari.db")
STOP_FILE = os.path.join(PROJE_DIZINI, "tarama_stop.flag")
GOTO_FILE = os.path.join(PROJE_DIZINI, "tarama_goto.flag")
AKTIVITE_FILE = os.path.join(PROJE_DIZINI, "medula_aktivite.tmp")
RENKLI_LISTE = os.path.join(PROJE_DIZINI, "renkli_recete_listesi.json")
LOG_DOSYA = os.path.join(PROJE_DIZINI, "tarama_log.txt")
_log_file = None


def aktivite_bildir():
    """Medula'ya her tıklamada/etkileşimde zaman damgası yaz.
    Keepalive bu dosyayı okuyarak son aktiviteyi takip eder."""
    try:
        with open(AKTIVITE_FILE, "w") as f:
            f.write(str(time.time()))
    except:
        pass


# ========== STOP KONTROL ==========
def durduruldu_mu():
    """Stop dosyası var mı kontrol et"""
    return os.path.exists(STOP_FILE)


def stop_temizle():
    """Eski stop ve goto dosyalarını sil"""
    try:
        if os.path.exists(STOP_FILE):
            os.remove(STOP_FILE)
    except:
        pass
    try:
        if os.path.exists(GOTO_FILE):
            os.remove(GOTO_FILE)
    except:
        pass


def goto_recete_kontrol():
    """GUI'den geri dönme isteği var mı kontrol et.
    Returns: reçete no (str) veya None"""
    try:
        if os.path.exists(GOTO_FILE):
            with open(GOTO_FILE, "r") as f:
                recete_no = f.read().strip()
            os.remove(GOTO_FILE)
            return recete_no if recete_no else None
    except:
        pass
    return None


# ========== LOG ==========
def log(msg, level="info"):
    global _log_file
    zaman = datetime.now().strftime("%H:%M:%S")
    prefix = {"info": "[BİLGİ]", "ok": "[  OK  ]", "warn": "[UYARI]",
              "error": "[HATA!]", "header": "[=====]",
              "sorun": "[SORUN]", "yeni": "[YENİ]"}.get(level, "[?]")
    satir = f"{zaman} {prefix} {msg}"
    print(satir, flush=True)
    try:
        if _log_file is None:
            _log_file = open(LOG_DOSYA, "w", encoding="utf-8")
        _log_file.write(satir + "\n")
        _log_file.flush()
    except:
        pass


# ========== MEDULA ==========
def _process_adi(handle):
    """Bir HWND'in process exe adını döndürür ('botanikmedula', 'botanikeczane', ...).
    pywin32 yoksa veya hata olursa boş string döner."""
    if win32process is None or win32api is None:
        return ""
    try:
        _tid, pid = win32process.GetWindowThreadProcessId(int(handle))
        if not pid:
            return ""
        # PROCESS_QUERY_LIMITED_INFORMATION = 0x1000 — düşük yetkiyle exe path okumaya yeter
        h = win32api.OpenProcess(0x1000, False, pid)
        try:
            exe_path = win32process.GetModuleFileNameEx(h, 0)
        finally:
            win32api.CloseHandle(h)
        return os.path.splitext(os.path.basename(exe_path or ""))[0].lower()
    except Exception:
        return ""


def medula_bul(diag=False):
    """BotanikEOS embedded Medula tarayıcı penceresini bul.

    Doğru hedef BotanikMedula.exe process'idir — BotanikEczane.exe ana uygulaması
    da "BotanikEOS ... (T)" başlığı taşır ama Medula otomasyonu için yanlış pencere.

    Eşleştirme önceliği:
      1. Process adı "BotanikMedula" olan görünür pencere (en güvenilir)
      2. Title'da "MEDULA" olan görünür pencere (oturum aktifken)
      3. Eski heuristic: BotanikEOS + (T) + 3+ kelime (kullanıcı adı içeren)

    diag=True olursa enumeration sırasında neyi atladığı loglanır.
    """
    desktop = Desktop(backend="uia")
    aday_pencereler = []
    try:
        for w in desktop.windows():
            try:
                title = (w.window_text() or "")
                if not title:
                    continue
                handle = w.handle
                visible = False
                try:
                    visible = desktop.window(handle=handle).is_visible()
                except Exception:
                    visible = False
                proc = _process_adi(handle)
                aday_pencereler.append((handle, title, visible, proc))
            except Exception:
                continue
    except Exception as e:
        if diag:
            log(f"  [TANI] desktop.windows() hatası: {e}", "warn")
        return None

    # 1. Process adıyla eşleşen — en güvenilir
    for handle, title, visible, proc in aday_pencereler:
        if proc == "botanikmedula" and visible:
            try:
                return desktop.window(handle=handle)
            except Exception:
                pass

    # 2. Title'da MEDULA varsa
    for handle, title, visible, proc in aday_pencereler:
        if "MEDULA" in title and visible:
            try:
                return desktop.window(handle=handle)
            except Exception:
                pass

    # 3. Kullanıcı adı içeren BotanikEOS penceresi (BotanikEczane bu yola düşer)
    for handle, title, visible, proc in aday_pencereler:
        if "BotanikEOS" in title and "(T)" in title and visible:
            parcalar = title.replace("(T)", "").strip().split()
            if len(parcalar) >= 3:
                try:
                    return desktop.window(handle=handle)
                except Exception:
                    pass

    if diag:
        log("  [TANI] medula_bul() eşleşme bulamadı. Görünen pencereler:", "warn")
        for handle, title, visible, proc in aday_pencereler:
            if not visible:
                continue
            t = title if "Botanik" in title or "MEDULA" in title else None
            if t or proc.startswith("botanik"):
                log(f"    - hwnd={handle} proc={proc or '?'} visible={visible} title={title!r}", "info")
    return None


def popup_kapat():
    """BotanikEOS popup'larını kontrol et ve Tamam/OK butonuyla kapat.
    Örn: 'Etken madde çakışması bulamadım...' penceresi.
    Sistem takılıyorsa çağrılır. Bulup kapattıysa True döner.

    GÜVENLİK:
    - BotanikMedula ana penceresi ASLA kapatılmaz (process adıyla filtre).
    - Sadece küçük (≤700×500 px) modal/dialog tipi pencereler hedeflenir;
      tam ekran tarayıcı içeriği popup sayılmaz.
    - "Uyarı" / "Dikkat" gibi kısa kelimeler artık tek başına eşleşme tetiklemez —
      gerçek popup'larda imzalar daha bağlamlı ifadelerle yer alır.
    - Kapatma butonu sadece ControlType "Button" ve caption tam eşleşme; pencerenin
      köşesindeki sistem "Kapat" butonu (genellikle isimli ama farklı parent'ta) skip.
    """
    # Bağlamlı popup imzaları — kısa kelimeler artık alone match etmez
    popup_imzalari = [
        "Etken madde çakışması",
        "HASTA İLAÇ BİLGİSİ",
        "Lütfen ilaç seçiniz",
        "ilaç seçiniz",
        "ilac seciniz",
        "Lütfen bekleyin",
    ]
    kapatma_captions = ("Tamam", "OK", "Evet")  # "Kapat"/"Close" çıkarıldı — pencere kromu butonu olabilir
    POPUP_MAX_W = 700
    POPUP_MAX_H = 500
    try:
        desktop = Desktop(backend="uia")
        for w in desktop.windows():
            try:
                handle = w.handle
                # 1) Ana BotanikMedula penceresine asla dokunma
                proc = _process_adi(handle)
                if proc == "botanikmedula":
                    continue
                if not w.is_visible():
                    continue
                title = (w.window_text() or "").strip()
                if len(title) > 80:
                    continue
                # 2) Pencere boyutu — tam ekran tarayıcı popup değildir
                try:
                    rect = w.rectangle()
                    genislik = rect.right - rect.left
                    yukseklik = rect.bottom - rect.top
                    if genislik > POPUP_MAX_W or yukseklik > POPUP_MAX_H:
                        continue
                    if genislik <= 0 or yukseklik <= 0:
                        continue
                except Exception:
                    continue
                # 3) Pencere içindeki text'leri topla
                text_parcalari = []
                tamam_btn = None
                try:
                    for d in w.descendants():
                        try:
                            ctrl = str(d.element_info.control_type)
                            cap = (d.window_text() or "").strip()
                            if not cap:
                                continue
                            if "Button" in ctrl and cap in kapatma_captions and tamam_btn is None:
                                tamam_btn = d
                            elif "Text" in ctrl or "Static" in ctrl or "Pane" in ctrl:
                                text_parcalari.append(cap)
                        except Exception:
                            continue
                except Exception:
                    continue
                if not tamam_btn:
                    continue
                # 4) Pencerenin içeriği bilinen imzalardan birini içeriyor mu?
                tum_metin = " ".join(text_parcalari) + " " + title
                if any(imza in tum_metin for imza in popup_imzalari):
                    try:
                        tamam_btn.invoke()
                        log(f"    [POPUP] Kapatıldı: '{title or tum_metin[:50]}' ({genislik}x{yukseklik})", "warn")
                        time.sleep(0.5)
                        return True
                    except Exception as e:
                        log(f"    [POPUP] Invoke hatası: {e} — Enter deneniyor", "warn")
                        pyautogui.press("enter")
                        time.sleep(0.5)
                        return True
            except Exception:
                continue
    except Exception as e:
        log(f"    [POPUP] Tarama hatası: {e}", "warn")
    return False


def sistem_dusmus_mu(medula):
    """Medula 'Sistem hatası' sayfasında mı kontrol et.
    Kırmızı hata sayfası görünüyorsa True döner."""
    try:
        for d in medula.descendants():
            try:
                txt = d.window_text()
                if txt and "Sistem hatası" in txt:
                    return True
            except:
                pass
    except:
        pass
    return False


MEDULA_EXE = r"C:\BotanikEczane\BotanikMedula.exe"
MEDULA_KULLANICI = "16-botan"
MEDULA_SIFRE = "152634"


def medula_yeniden_baslat():
    """Medula sistem hatası sonrası hızlı kurtarma."""
    import subprocess

    medula = medula_bul()

    # 1. Giriş butonu dene (en hızlı - 3sn)
    if medula:
        log("Sistem hatası - Giriş butonu deneniyor...", "error")
        element_tikla(medula, "btnMedulayaGirisYap", "click")
        time.sleep(3)
        medula = medula_bul()
        if medula and not sistem_dusmus_mu(medula):
            if element_bul(medula, "form1:menuHtmlCommandExButton31") or element_bul(medula, "f:tbl1"):
                log("Giriş butonu ile düzeldi", "ok")
                return medula

    # 2. Taskkill + exe + giriş (10-15sn)
    log("Taskkill + yeniden başlat...", "error")
    try:
        subprocess.run(["taskkill", "/F", "/IM", "BotanikMedula.exe"], capture_output=True, timeout=5)
    except:
        pass
    time.sleep(1)

    try:
        subprocess.Popen([MEDULA_EXE], cwd=os.path.dirname(MEDULA_EXE))
    except:
        log("BotanikMedula.exe başlatılamadı!", "error")
        return None

    # SifreSorForm bekle + hızlı giriş
    desktop = Desktop(backend="uia")
    for i in range(15):
        time.sleep(0.5)
        try:
            for w in desktop.windows():
                try:
                    if w.element_info.automation_id == "SifreSorForm":
                        w.set_focus()
                        time.sleep(0.2)
                        combo = w.child_window(auto_id="cmbKullanicilar")
                        combo.click_input()
                        time.sleep(0.2)
                        for item in combo.children():
                            try:
                                if "botan" in item.window_text().lower():
                                    item.click_input()
                                    break
                            except:
                                pass
                        time.sleep(0.2)
                        sifre = w.child_window(auto_id="txtSifre")
                        sifre.click_input()
                        time.sleep(0.1)
                        sifre.type_keys(MEDULA_SIFRE, with_spaces=True)
                        w.child_window(auto_id="btnGirisYap").click_input()
                        log("Giriş yapıldı", "info")
                        break
                except:
                    pass
            else:
                continue
            break
        except:
            pass

    # MEDULA penceresi bekle
    for i in range(15):
        time.sleep(1)
        medula = medula_bul()
        if medula and (element_bul(medula, "form1:menuHtmlCommandExButton31") or element_bul(medula, "f:tbl1")):
            log(f"Medula hazır ({i+1}s)", "ok")
            return medula

    log("Medula açılamadı!", "error")
    return None


_aid_cache_global = {}  # recete_tum_bilgi_topla'dan doldurulan cache

def element_bul(medula, auto_id):
    """Tek bir elementi automation_id ile bul (cache → child_window → descendants fallback)"""
    # 1. Global cache (en hızlı — 0ms)
    if auto_id in _aid_cache_global:
        return _aid_cache_global[auto_id]
    # 2. child_window (hızlı — WinForms elementleri için)
    try:
        cw = medula.child_window(auto_id=auto_id)
        if cw.exists(timeout=0.1):  # 0.3→0.1 (cache varsa buraya düşmemeli)
            found = cw.wrapper_object()
            _aid_cache_global[auto_id] = found
            return found
    except:
        pass
    # 3. Fallback: descendants tarama (yavaş — IE embedded browser elementleri için)
    try:
        for elem in medula.descendants():
            try:
                aid = elem.element_info.automation_id
                if aid:
                    _aid_cache_global[aid] = elem
                if aid == auto_id:
                    return elem
            except:
                pass
    except:
        pass
    return None


def element_tikla(medula, auto_id, yontem="invoke"):
    """Element bul + tıkla + aktivite bildir. Tüm Medula etkileşimleri bu fonksiyonla yapılmalı."""
    elem = element_bul(medula, auto_id)
    if not elem:
        return False
    try:
        if yontem == "invoke":
            elem.invoke()
        else:
            elem.click_input()
        aktivite_bildir()
        return True
    except:
        return False


def recete_no_oku(medula):
    """Açık reçetenin numarasını oku"""
    try:
        for elem in medula.descendants(control_type="DataItem"):
            try:
                txt = elem.window_text()
                if txt:
                    txt = txt.strip()
                    if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                        return txt
            except:
                pass
    except:
        pass
    return None


# ========== REÇETE TÜRÜ ==========
RECETE_TURLERI = {
    "1": "Normal", "2": "Kırmızı", "3": "Turuncu", "4": "Mor", "5": "Yeşil",
    "Normal": "Normal", "Kırmızı": "Kırmızı", "Turuncu": "Turuncu",
    "Mor": "Mor", "Yeşil": "Yeşil"
}

def recete_turu_oku(medula):
    """Reçete türünü oku - f:m4 SELECT combobox + DataItem fallback"""
    # Yöntem 1: f:m4 SELECT elementini oku
    try:
        elem = element_bul(medula, "f:m4")
        if elem:
            # selected_text() → "Normal", "Kırmızı", "Yeşil", "Mor", "Turuncu"
            try:
                selected = elem.selected_text()
                if selected and selected.strip() in RECETE_TURLERI:
                    return RECETE_TURLERI[selected.strip()]
            except:
                pass
            # Fallback: window_text
            val = elem.window_text().strip()
            if val in RECETE_TURLERI:
                return RECETE_TURLERI[val]
    except:
        pass

    # Yöntem 2: DataItem'lardan "Reçete Türü" yanındaki değeri oku
    try:
        for elem in medula.descendants(control_type="DataItem"):
            try:
                txt = elem.window_text().strip()
                r = elem.rectangle()
                # Reçete türü değeri genellikle y=362, x=865 civarında
                if txt in ["Normal", "Kırmızı", "Turuncu", "Mor", "Yeşil"]:
                    if r.top > 340 and r.top < 400:
                        return RECETE_TURLERI.get(txt, txt)
            except:
                pass
    except:
        pass

    # Yöntem 3: ComboBox'ların hepsini tara
    try:
        for elem in medula.descendants(control_type="ComboBox"):
            try:
                aid = elem.element_info.automation_id
                if aid == "f:m4":
                    # items() ile seçili değeri al
                    try:
                        for item in elem.children():
                            txt = item.window_text().strip()
                            if txt in RECETE_TURLERI:
                                return RECETE_TURLERI[txt]
                    except:
                        pass
            except:
                pass
    except:
        pass

    return "Normal"  # Varsayılan


# ========== RENKLİ REÇETE ==========
_renkli_liste_cache = None

_renkli_liste_tarih = None

def renkli_liste_yukle():
    """Renkli reçete listesini JSON'dan yükle (cache'le)"""
    global _renkli_liste_cache, _renkli_liste_tarih
    if _renkli_liste_cache is not None:
        return _renkli_liste_cache

    _renkli_liste_cache = set()
    try:
        if os.path.exists(RENKLI_LISTE):
            with open(RENKLI_LISTE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                _renkli_liste_cache = set(data.get('receteler', []))
                _renkli_liste_tarih = data.get('yukleme_tarihi', '')
                tarih_bilgi = ""
                if _renkli_liste_tarih:
                    tarih_bilgi = f" (yükleme: {_renkli_liste_tarih[:10]})"
                log(f"Renkli reçete listesi yüklendi: {len(_renkli_liste_cache)} reçete{tarih_bilgi}", "info")
    except Exception as e:
        log(f"Renkli liste yüklenemedi: {e}", "warn")

    return _renkli_liste_cache


def renkli_recete_kontrol(recete_no, recete_turu):
    """Renkli reçete sisteme işlenmiş mi kontrol et.
    Returns: (sorun_var: bool, mesaj: str)
    """
    if recete_turu == "Normal":
        return False, "Beyaz reçete"

    # Renkli reçete - listede var mı?
    liste = renkli_liste_yukle()
    if not liste:
        return False, f"{recete_turu} reçete (liste yüklü değil, kontrol edilemedi)"

    if recete_no in liste:
        return False, f"{recete_turu} reçete - sisteme işlenmiş"
    else:
        # Eski liste uyarısı: reçete no prefix'i listeden farklıysa
        liste_prefix = set()
        for r in list(liste)[:5]:
            if r:
                liste_prefix.add(r[0])
        recete_prefix = recete_no[0] if recete_no else ""
        if recete_prefix and liste_prefix and recete_prefix not in liste_prefix:
            return True, f"{recete_turu} reçete - SİSTEME İŞLENMEMİŞ! (⚠ Liste eski ay olabilir — liste: {liste_prefix}, reçete: {recete_prefix})"
        return True, f"{recete_turu} reçete - SİSTEME İŞLENMEMİŞ!"


# ========== İLAÇ OKUMA ==========
ILAC_ANAHTAR = [
    # === KONSANTRASYON BİRİMLERİ ===
    "MG", "ML", "MCG", "μG", "IU", "MIU", "KIU",
    "MG/", "ML/", "MG.", "ML.",
    "G.", "G/", "G(",          # "100 G(...)" gibi formatlar
    "GR", "GR.", "GR(",        # "5GRX20SASE" gibi
    "KCAL", "KAL", "KALORI",   # Beslenme ürünleri için kalori birimi

    # === KATI ORAL FORMLAR (tablet/kapsül) ===
    "TABLET", "TB", "TB.", "TBL", "TBL.",
    "FILM", "FTB", "FTB.", "FT.",
    "FILM TABLET", "ÇİĞNEME", "CIGNEME", "ÇÖZÜNÜR", "COZUNUR",
    "EFERVESAN", "EFFERVESAN", "EFF.", "EFF",
    "SUBLINGUAL", "SUBL", "SUBL.",
    "OROD", "OROD.",           # orodispersible
    "ENTERIK", "ENTERİK", "ENTERIC",
    "KAPSUL", "KAPSÜL", "KAP", "KAP.", "KP", "KP.",
    "MIKROPELLET", "MİKROPELLET", "PELLET",
    "PASTİL", "PASTIL", "LOZANJ", "LOZENGE",
    "GRANÜL", "GRANUL", "GRAN", "GRAN.",
    "TOZ", "TOZU", "POWDER",
    "ŞASE", "SASE", "SACHET", "SACH", "SACHE",  # SASE = ŞASE'nin ASCII'si

    # === SIVI ORAL FORMLAR ===
    "ŞURUP", "SURUP", "SİROP", "SIROP", "SIRUP", "SYRUP",
    "SÜSPANSİYON", "SÜSPANSIYON", "SUSPANSIYON", "SUSP", "SUSP.",
    "SOLÜSYON", "SOLUSYON", "SOLÜTİON", "SOLUTION", "SOL", "SOL.",
    "ÇÖZELTİ", "COZELTI", "ÇÖZELTI",
    "DAMLA", "DML", "DROPS",
    "ELİKSİR", "ELIKSIR", "ELIXIR",
    "EMÜLSİYON", "EMULSIYON", "EMÜLSİON", "EMULSION", "EMUL",
    "KONSANTRE", "CONCENTRATE", "KONS.",
    "AROMA", "AROMALI",        # bazı şuruplar için

    # === TOPİKAL FORMLAR ===
    "KREM", "KRM", "CREAM",
    "POMAD", "POM", "POM.",
    "MERHEM", "MRH",
    "JEL", "GEL",
    "LOSYON", "LOTION",
    "KÖPÜK", "KOPUK", "FOAM",
    "ŞAMPUAN", "SAMPUAN", "SHAMPOO",
    "SABUN", "SOAP",

    # === İNHALER / NEBULIZER ===
    "INH", "INH.", "INHALER",
    "NEBUL", "NEBÜL", "NEBÜLİZE", "NEB", "NEB.",
    "AEROSOL", "AEROZOL",
    "DİSKUS", "DISKUS",
    "TURBUHALER", "EASYHALER", "AEROLIZER", "SPINHALER",

    # === BURUN / GÖZ / KULAK ===
    "SPRAY", "SPREY",
    "NAZAL", "NASAL", "BURUN",
    "OFTALMİK", "OFTALMIK", "OPHTH", "GÖZ", "GOZ",
    "OTİK", "OTIK", "KULAK", "EAR",
    "GARGARA", "GARGLE",       # ağız çalkalama

    # === ENJEKTABL ===
    "AMPUL", "AMP", "AMP.", "AMPULE",
    "FLAKON", "FLK", "FLK.", "VIAL",
    "ENJEKT", "ENJ", "ENJ.", "INJ", "INJEKT",
    "ŞIRINGA", "SIRINGA", "SYRINGE", "ŞIRNG",
    "KALEM", "PEN",
    "İNFÜZYON", "INFUZYON", "INFUSION",
    "İNTRAVENÖZ", "INTRAVENOZ",
    "İNTRAMÜSKÜLER",
    "SUBKUTAN",

    # === REKTAL / VAJİNAL ===
    "FİTİL", "FITIL", "SUPP", "SUPP.", "SUPPOZITUVAR", "SUPPOZİTUVAR", "SUPPOSITORY",
    "OVÜL", "OVUL", "OVULE",
    "VAJİNAL", "VAJINAL", "VAGINAL",
    "REKTAL", "RECTAL",
    "LAVMAN", "ENEMA",

    # === TRANSDERMAL / YAMA ===
    "PATCH", "YAMA", "BANT", "BAND",
    "TRANSDERMAL", "DERMAL",
    "PLASTER",

    # === MODIFIYE SALIM / ETKİ MODÜLATÖRLERİ ===
    "FORTE", "FORT",
    "RAPID", "FAST", "QUICK",
    "RETARD", "DEPO", "DEPOT",
    "SR", "ER", "XR", "MR", "LA",       # slow/extended/modified/long-acting release
    "KONTROLLÜ", "KONTROLLU",
    "MODIFIYE", "MODİFİYE",
    "UZATILMIŞ", "UZATILMIS",
    "SUPRA",

    # === BESLENME ÜRÜNLERİ (mama, enteral) ===
    "ENTERAL", "PARENTERAL",
    "BESLENME", "MAMA", "FORMULA",
    "NUTRİSYON", "NUTRISYON", "NUTRITION",

    # === DİĞER ===
    "DOZ", "DOSE",
    "STİK", "STIK", "STICK",
    "KİT", "KIT", "SET",
    "ÜRÜN",
    "TIBBI",
]
ATLA = ["Maksimum", "Topl=", "KALEM", "Toplam Tutar", "Sayfaya",
        "İncelemeye", "Botrastan", "Reçete Tutar", "YAZMIŞ",
        "YAZMAMIŞ", "VERİLEMEZ", "KONTROL", "UZMAN",
        "Uyumlu ICD", "Raporsuz", "Raporda DOZ", "Öncelik",
        "Girilebilcek", "endikasyon",
        "Günde", "Haftada", "Ayda", "Saatte", " x "]


def _uyari_kodu_ekle(sonuc, txt, re):
    """Uyarı kodu pattern'ini kontrol et ve sonuc dict'ine ekle (tekrar engelleme dahil)."""
    match = re.match(r'(\d+)\s*-\s*(.+?)\s*=>\s*(.+)', txt)
    if match:
        anahtar = f"{match.group(1)}_{match.group(3).strip()}"
        mevcut = {f"{u['kod']}_{u['ilac_adi']}" for u in sonuc["uyari_kodlari"]}
        if anahtar not in mevcut:
            sonuc["uyari_kodlari"].append({
                "kod": match.group(1),
                "aciklama": match.group(2).strip(),
                "ilac_adi": match.group(3).strip(),
            })


def recete_tum_bilgi_topla(medula):
    """Reçete sayfasındaki TÜM bilgileri TEK BİR descendants() çağrısıyla topla.

    Mevcut akış 5-6 ayrı descendants()/element_bul() çağrısı yapıyordu:
      recete_no_oku, recete_turu_oku, ilaclari_oku, recete_teshisleri_oku,
      uyari_kodlari_oku → her biri ayrı UI ağacı taraması.

    Bu fonksiyon HEPSİNİ tek geçişte toplar.

    Returns: dict {
        "recete_no": str veya None,
        "recete_turu": str ("Normal", "Kırmızı", ...),
        "recete_alt_turu": str ("Ayaktan", "Acil", "Yatan", ...),
        "ilaclar": list[dict],      # ilaclari_oku() ile aynı format
        "teshisler": list[str],     # recete_teshisleri_oku() ile aynı format
        "uyari_kodlari": list[dict],# uyari_kodlari_oku() ile aynı format
        "doktor_uzmanligi": str,    # doktor branş/uzmanlık bilgisi
        "recete_aciklamalari": list[str], # Açıklama Listesi tablosu
        "_aid_cache": dict,         # automation_id → element cache (sonraki işlemler için)
    }
    """
    import re

    sonuc = {
        "recete_no": None,
        "recete_turu": "Normal",
        "recete_alt_turu": "Ayaktan",   # f:m8: Ayaktan/Yatan/Taburcu/Günübirlik/Acil/Yeşil Alan/Evde Bakım
        "ilaclar": [],
        "teshisler": [],
        "uyari_kodlari": [],
        "doktor_uzmanligi": "",
        "doktor_adi": "",
        "hasta_adi": "",
        "erecete_no": "",
        "fatura_turu": "",
        "recete_aciklamalari": [],
        "recete_teshisleri_input": [],
        "_aid_cache": {},
    }

    # ── 1. Tüm descendants'ları TEK SEFERDE topla ──
    t0 = time.time()
    tum_elementler = []
    try:
        tum_elementler = list(medula.descendants())
    except Exception as e:
        log(f"descendants() hatası: {e}", "error")
        return sonuc
    t_desc = time.time() - t0

    # ── 2. SINIFLANDIRMA: tek geçişte tip ayrımı + aid_cache + metadata cache ──
    # OPT: Her element için aid/ctrl_type/name TEK SEFERDE okunur, dict'te cache'lenir.
    # Sonraki phase'ler element_info'ya tekrar dokunmaz — yüzlerce COM çağrısı tasarrufu.
    # Ek olarak alakasız ctrl_type'lar (ScrollBar, Image, Separator, vb.) erkenden elenir.
    SKIP_CTRL = ("ScrollBar", "Image", "Separator", "ToolBar", "TitleBar",
                 "MenuBar", "StatusBar", "Slider", "Thumb", "ProgressBar")

    aid_cache = {}
    elem_meta = {}        # id(elem) → (aid, ctrl_type, name)
    data_item_refs = []   # [elem]
    text_custom_refs = [] # [elem]
    combo_items = []      # [elem]
    table_refs = []       # [elem]

    for elem in tum_elementler:
        try:
            info = elem.element_info
            ctrl_type = info.control_type or ""
            # ERKEN ELEME: ihtiyaç duymadığımız tipleri hızla at
            if any(s in ctrl_type for s in SKIP_CTRL):
                continue
            aid = info.automation_id or ""
            # name sadece Text/Custom için lazım (aciklama_keywords match için);
            # diğer tiplerde COM çağrısını yapmadan boş bırak
            name = ""
            if "Text" in ctrl_type or "Custom" in ctrl_type:
                name = (info.name or "").strip()

            elem_meta[id(elem)] = (aid, ctrl_type, name)

            if aid:
                aid_cache[aid] = elem

            if "DataItem" in ctrl_type:
                data_item_refs.append(elem)
            elif "Text" in ctrl_type or "Custom" in ctrl_type:
                text_custom_refs.append(elem)
                # Hızlı pattern tarama (rectangle/window_text çağırmadan)
                if name:
                    if "=>" in name:
                        _uyari_kodu_ekle(sonuc, name[:200], re)
                    if "Girilebilcek" in name or "Girilebilecek" in name:
                        sonuc["_uyari_alani_metni"] = name[:200]
            elif "ComboBox" in ctrl_type:
                combo_items.append(elem)
            elif "Table" in ctrl_type:
                table_refs.append(elem)
        except:
            pass

    sonuc["_aid_cache"] = aid_cache
    # Global cache'i güncelle (element_bul descendants fallback'ini engeller)
    global _aid_cache_global
    _aid_cache_global = aid_cache

    # f:tbl1 tablo pozisyonu (cache'den)
    tbl_top = 0
    tbl_bottom = 9999
    tbl_elem = aid_cache.get("f:tbl1")
    if tbl_elem:
        try:
            r = tbl_elem.rectangle()
            tbl_top = r.top - 30
            tbl_bottom = r.bottom + 50
        except:
            pass

    # ── 2b. DATAITEM İŞLEME: window_text + rectangle (ilaç tablosu + uyarı kodları için lazım) ──
    # OPT: aid Phase 2'de cache'lendi → elem_meta'dan alıyoruz, COM tekrar çağrılmıyor.
    data_items = []  # (txt_s, top, left, aid, elem)
    for elem in data_item_refs:
        try:
            txt = elem.window_text()
            if not txt or not txt.strip():
                continue
            r = elem.rectangle()
            txt_s = txt.strip()[:200]
            meta = elem_meta.get(id(elem))
            aid = meta[0] if meta else ""
            data_items.append((txt_s, r.top, r.left, aid, elem))
            if "=>" in txt_s:
                _uyari_kodu_ekle(sonuc, txt_s, re)
        except:
            pass

    # NOT: Text/Custom için name + "=>" / "Girilebilcek" tarama Phase 2'de yapıldı —
    # ek tek bir loop'a gerek yok. text_custom_refs deferred phase 2d için saklı.

    # ── 3. Reçete No ──
    for txt, y, x, aid, elem in data_items:
        if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
            sonuc["recete_no"] = txt
            break

    # ═══ OPT: FAST PATH — Reçete "atlanabilir" mi? ═══
    # Eğer hiçbir ilaçta rapor kodu yok, mesaj yok, uyarı kodu yok ise
    # teşhis/açıklama/doktor fallback'lerini ATLA (çoğu reçete için 10-30 sn tasarruf)
    _has_rapor = False
    _has_msj = False
    for _r in range(12):
        t9 = aid_cache.get(f"f:tbl1:{_r}:t9")
        if t9 is not None:
            try:
                _txt = (t9.window_text() or "").strip()
                if _txt and _txt[0].isdigit():
                    _has_rapor = True
                    break
            except Exception:
                pass
    if not _has_rapor:
        # İlaç mesajı bypass ayarlarını al (mesajı olan ilaç bypass listesindeyse mesaj yok say)
        try:
            from kontrol_kurallari import get_kontrol_ayarlari
            _ayarlar_fp = get_kontrol_ayarlari()
        except Exception:
            _ayarlar_fp = None
        for _r in range(12):
            t11 = aid_cache.get(f"f:tbl1:{_r}:t11")
            if t11 is None:
                continue
            try:
                _txt = (t11.window_text() or "").strip().lower()
                if _txt != "var":
                    continue
            except Exception:
                continue
            # Mesaj var — bu ilaç bypass mı?
            _bypass = False
            if _ayarlar_fp is not None:
                try:
                    t6 = aid_cache.get(f"f:tbl1:{_r}:t6")
                    _ilac_adi_row = ""
                    if t6 is not None:
                        try:
                            _ilac_adi_row = (t6.window_text() or "").strip()
                        except Exception:
                            pass
                    _bypass = not _ayarlar_fp.kontrol_aktif_mi(
                        "ilac_mesaji", ilac_adi=_ilac_adi_row, etkin_madde="",
                    )
                except Exception:
                    _bypass = False
            if not _bypass:
                _has_msj = True
                break
        if _has_msj is False:
            # Tüm "var" satırları bypass'ta — log bilgilendirme
            _bypass_msj_found = any(
                aid_cache.get(f"f:tbl1:{_r}:t11") is not None for _r in range(12)
            )
            if _bypass_msj_found:
                log(f"  [OPT] Mesajlı ilaçlar 'yoksay' listesinde — mesaj yok say", "info")
    _has_uyari = bool(sonuc.get("uyari_kodlari"))
    _fast_path = not (_has_rapor or _has_msj or _has_uyari)
    if _fast_path:
        log(f"  [OPT] Fast path: raporsuz/mesajsız/uyarısız reçete — detay toplamadan geçiliyor", "info")

    # ── 2d. DEFERRED: Text/Custom + Tables tam işleme (sadece fast_path DEĞİLSE) ──
    # Bu liste pozisyonel fallback'ler (tanı listesi, açıklama listesi, doktor branş) için.
    # Fast path reçetelerinde HİÇBİRİ kullanılmaz.
    # OPT: Phase 2'de cache'lenen elem_meta'yı kullanır — element_info'ya tekrar dokunmaz.
    text_items = []   # (txt_s, top, left, ctrl_type, elem)
    tables = []       # (name, top, bottom)
    if not _fast_path:
        for elem in text_custom_refs:
            meta = elem_meta.get(id(elem))
            if not meta:
                continue
            aid, ctrl_type, name_hint = meta
            # Hızlı filtre: hem name hem aid yoksa atla
            if not name_hint and not aid:
                continue
            txt_s = name_hint
            if not txt_s:
                try:
                    txt_s = (elem.window_text() or "").strip()
                except Exception:
                    continue
                if not txt_s:
                    continue
            # Toplama kriteri: anlamlı/kullanılacak metinler
            if len(txt_s) >= 3 or "=>" in txt_s:
                try:
                    r = elem.rectangle()
                    text_items.append((txt_s, r.top, r.left, ctrl_type, elem))
                except Exception:
                    pass
        # Table elementleri içinde Tanı Listesi / Açıklama Listesi
        # Table'ın name'i Phase 2'de henüz okunmadı — şimdi gerekirse oku
        for elem in table_refs:
            try:
                # Table için name Phase 2'de okunmamıştı (sadece Text/Custom için okuyoruz);
                # burada ihtiyaç anında oku
                tbl_name = (elem.element_info.name or "").strip()
                if tbl_name in ("Tanı Listesi", "Açıklama Listesi"):
                    r = elem.rectangle()
                    tables.append((tbl_name, r.top, r.bottom))
            except Exception:
                pass

    # ── 4. Reçete Türü (f:m4 cache'den) ──
    m4 = aid_cache.get("f:m4")
    if m4:
        try:
            selected = None
            try:
                selected = m4.selected_text()
            except:
                pass
            if selected and selected.strip() in RECETE_TURLERI:
                sonuc["recete_turu"] = RECETE_TURLERI[selected.strip()]
            else:
                val = (m4.window_text() or "").strip()
                if val in RECETE_TURLERI:
                    sonuc["recete_turu"] = RECETE_TURLERI[val]
        except:
            pass
    # Fallback: DataItem'lardan reçete türü
    if sonuc["recete_turu"] == "Normal" and not m4:
        for txt, y, x, aid, elem in data_items:
            if txt in ["Kırmızı", "Turuncu", "Mor", "Yeşil"]:
                try:
                    r = elem.rectangle()
                    if r.top > 340 and r.top < 400:
                        sonuc["recete_turu"] = RECETE_TURLERI.get(txt, txt)
                        break
                except:
                    pass

    # ── 4b. Reçete Alt Türü (f:m8 — Ayaktan/Acil/Yatan/Taburcu/Günübirlik) ──
    m8 = aid_cache.get("f:m8")
    if not m8:
        m8 = element_bul(medula, "f:m8")
    if m8:
        try:
            alt_tur = ""
            try:
                alt_tur = m8.selected_text() or ""
            except:
                pass
            if not alt_tur:
                alt_tur = (m8.window_text() or "").strip()
            if alt_tur:
                sonuc["recete_alt_turu"] = alt_tur.strip()
        except:
            pass

    # ── 5a. Reçete Teşhis INPUT'ları (f:tableEx1:{row}:text3) ──
    # Reçete sayfasındaki teşhis giriş alanları (INPUT/Edit tipi)
    for row in range(10):
        t3 = aid_cache.get(f"f:tableEx1:{row}:text3")
        if not t3:
            t3 = element_bul(medula, f"f:tableEx1:{row}:text3")
        if not t3:
            break
        try:
            txt = ""
            # INPUT elementinden değer oku
            try:
                txt = t3.get_value() or ""
            except:
                pass
            if not txt:
                try:
                    txt = t3.iface_value.CurrentValue or ""
                except:
                    pass
            if not txt:
                txt = (t3.window_text() or "")
            txt = txt.strip()
            if txt and len(txt) > 2:
                sonuc["recete_teshisleri_input"].append(txt)
        except:
            pass

    # ── 5b. Tanı Listesi (form1:tableEx3) ──
    # Ana reçete sayfasında E-Reçete tanı listesi:
    #   HTML: form1:tableEx3:{row}:text47 (ICD kodu), form1:tableEx3:{row}:text54 (tanı adı)
    # Önce cache'den dene (MSHTML bazen ID'yi gösterir)
    for row in range(20):
        icd_elem = aid_cache.get(f"form1:tableEx3:{row}:text47")
        tani_elem = aid_cache.get(f"form1:tableEx3:{row}:text54")
        if not icd_elem and not tani_elem:
            break
        tani_txt = ""
        icd_txt = ""
        if icd_elem:
            try:
                icd_txt = (icd_elem.window_text() or "").strip()
            except:
                pass
        if tani_elem:
            try:
                tani_txt = (tani_elem.window_text() or "").strip()
            except:
                pass
        if icd_txt or tani_txt:
            birlesik = f"{icd_txt} - {tani_txt}" if icd_txt and tani_txt else (tani_txt or icd_txt)
            sonuc["teshisler"].append(birlesik)

    # Fallback: Cache'de bulunamadıysa, "Tanı Listesi" tablosu altındaki
    # ICD + tanı çiftlerini pozisyon bazlı topla (OPT B: text_items + tables kullan)
    if not sonuc["teshisler"] and not _fast_path:
        _teshis_tablo_pozisyon = None
        _teshis_tablo_bitis = None
        icd_pattern = re.compile(r'^[A-Z]\d{1,2}(\.\d+)?\s*$')

        for name, top, bottom in tables:
            if name == "Tanı Listesi":
                _teshis_tablo_pozisyon = top
                _teshis_tablo_bitis = bottom
                break

        if _teshis_tablo_pozisyon is not None:
            tani_items = []
            for txt, top, left, ctrl, _e in text_items:
                if "Text" not in ctrl:
                    continue
                if len(txt) < 2:
                    continue
                if top < _teshis_tablo_pozisyon or top > _teshis_tablo_bitis:
                    continue
                if txt in ("ICD-10", "Tanı", "Tanı Listesi"):
                    continue
                tani_items.append((txt, top, left))

            # ICD + Tanı çiftlerini eşleştir (aynı y satırında, ICD solda, tanı sağda)
            tani_items.sort(key=lambda item: (item[1], item[2]))
            i = 0
            while i < len(tani_items):
                txt1, y1, x1 = tani_items[i]
                if icd_pattern.match(txt1) and i + 1 < len(tani_items):
                    txt2, y2, x2 = tani_items[i + 1]
                    if abs(y1 - y2) < 5:  # Aynı satır
                        sonuc["teshisler"].append(f"{txt1.strip()} - {txt2}")
                        i += 2
                        continue
                # Tek başına tanı veya ICD
                sonuc["teshisler"].append(txt1)
                i += 1

    # ── 6. Reçete Açıklamaları (Açıklama Listesi tablosu) ──
    # HTML: form1:tableEx1 (Açıklama Listesi)
    #   Açıklama türü: form1:tableEx1:{row}:text49 ("Teşhis/Tanı" vb.)
    #   Açıklama metni: form1:tableEx1:{row}:text50 (protokol bilgisi vb.)
    for row in range(20):
        aciklama_elem = aid_cache.get(f"form1:tableEx1:{row}:text50")
        if not aciklama_elem:
            break
        try:
            txt = (aciklama_elem.window_text() or "").strip()
            if txt:
                sonuc["recete_aciklamalari"].append(txt)
        except:
            pass

    # Fallback: "Açıklama Listesi" tablosundan pozisyon bazlı (OPT B)
    if not sonuc["recete_aciklamalari"] and not _fast_path:
        aciklama_tablo_top = None
        aciklama_tablo_bottom = None
        for name, top, bottom in tables:
            if name == "Açıklama Listesi":
                aciklama_tablo_top = top
                aciklama_tablo_bottom = bottom
                break

        if aciklama_tablo_top is not None:
            for txt, top, left, ctrl, _e in text_items:
                if len(txt) < 5:
                    continue
                if top < aciklama_tablo_top or top > aciklama_tablo_bottom:
                    continue
                if txt in ("Açıklama Listesi", "Açıklama Türü", "Reçete Açıklama"):
                    continue
                sonuc["recete_aciklamalari"].append(txt)

    # ── 7. Doktor Uzmanlığı ──
    # Önce HTML ID ile (en güvenilir): form1:text41
    doktor_elem = aid_cache.get("form1:text41")
    if doktor_elem:
        try:
            txt = (doktor_elem.window_text() or "").strip()
            if txt and len(txt) > 3:
                sonuc["doktor_uzmanligi"] = txt
        except:
            pass

    # Fallback 1: "Doktor Brans" label pozisyonundan (OPT B: text_items kullan)
    if not sonuc["doktor_uzmanligi"] and not _fast_path:
        doktor_brans_y = None
        for txt, top, left, ctrl, _e in text_items:
            if txt in ("Doktor Brans", "Doktor Branş"):
                doktor_brans_y = top
                break

        if doktor_brans_y is not None:
            for txt, top, left, ctrl, _e in text_items:
                if "Text" not in ctrl:
                    continue
                if txt in ("Doktor Brans", "Doktor Branş", ":"):
                    continue
                if abs(top - doktor_brans_y) < 5 and len(txt) > 3:
                    sonuc["doktor_uzmanligi"] = txt
                    break

    # Fallback 2: DataItem + Text'lerden anahtar kelime bazlı
    if not sonuc["doktor_uzmanligi"] and not _fast_path:
        uzmanlik_anahtar = ["Hastalıkları", "Kardiyoloji", "Nöroloji", "Psikiyatri",
                            "Göğüs", "Endokrin", "Ortopedi", "Üroloji", "Dermatoloji",
                            "Nefroloji", "Hematoloji", "Romatoloji", "Gastroenteroloji",
                            "Pratisyen", "Aile Hekimi", "Onkoloji", "Enfeksiyon",
                            "Fizik Tedavi", "Kulak Burun", "Göz", "Kadın Doğum",
                            "Çocuk", "Cerrahi", "Anestezi", "Radyoloji", "Hekim",
                            "Dahiliye", "Cildiye", "Nörolog", "Üroloj", "Jinekoloji"]
        for txt, y, x, aid, elem in data_items:
            if any(k in txt for k in uzmanlik_anahtar):
                if y > 200 and len(txt) < 80:
                    sonuc["doktor_uzmanligi"] = txt
                    break

    # ── 7. İlaçlar (tbl içindeki DataItem'lar + cache'den doğrulama) ──
    # Tablo alanındaki DataItem'ları sırala
    tablo_items = [(txt, y, x) for txt, y, x, aid, elem in data_items
                   if y >= tbl_top and y <= tbl_bottom]
    tablo_items.sort(key=lambda item: (item[1], item[2]))

    ilaclar = []
    # Y-yakınlık eşiği: aynı satırın metadata'sı (SGK, rapor kodu, msj, doz) ilaç adından
    # en fazla bu kadar piksel uzakta olabilir. Daha uzaktaysa başka bir satıra ait demektir.
    SATIR_Y_ESIGI = 60
    for txt, y, x in tablo_items:
        txt_upper = txt.upper()
        if (len(txt) > 12 and any(k in txt_upper for k in ILAC_ANAHTAR)
                and not txt.startswith("SGK") and not any(a in txt for a in ATLA)):
            ilaclar.append({"ilac_adi": txt, "etkin_madde": "", "sgk_kodu": "",
                           "rapor_kodu": "", "msj": "",
                           "eos_rapor_doz_metin": "", "_y": y})
            continue
        if not ilaclar:
            continue
        son = ilaclar[-1]
        # Son ilacın y'sinden çok uzaktaki metadata satırı başka bir kalemin verisi olabilir
        if abs(y - son.get("_y", y)) > SATIR_Y_ESIGI:
            continue
        if txt.startswith("SGK") and "-" in txt:
            parts = txt.split("-", 1)
            son["sgk_kodu"] = parts[0].strip()
            son["etkin_madde"] = parts[1].strip() if len(parts) > 1 else ""
        elif "." in txt and "/" not in txt and len(txt) <= 8 and txt[0].isdigit():
            son["rapor_kodu"] = txt
        elif txt.lower() in ["var", "yok"]:
            son["msj"] = txt.lower()
        elif _doz_satiri_mi(txt):
            if "eos_rapor_doz_listesi" not in son:
                son["eos_rapor_doz_listesi"] = []
            son["eos_rapor_doz_listesi"].append(txt)
            if not son["eos_rapor_doz_metin"]:
                son["eos_rapor_doz_metin"] = txt

    # ── 8. MEDULA TABLO SATIR İNDEKSİ EŞLEŞTİRME ──
    # tum_elementler'den f:tbl1:{row}:t6 automation_id'li elementleri bul
    # EK element_bul/descendants çağrısı YAPMADAN, zaten toplanan veriden çıkar
    t1_satir = time.time()
    medula_satir_haritasi = {}  # row → ilaç_adi_text

    # Yöntem 1: aid_cache'den (f:tbl1:{row}:t6)
    _bos_satir = 0
    for row in range(20):
        t6 = aid_cache.get(f"f:tbl1:{row}:t6")
        if not t6:
            _bos_satir += 1
            if _bos_satir >= 3:
                break  # 3 ardışık boş satırda dur (cache'de olmayan satırları atla)
            continue
        _bos_satir = 0
        try:
            t6_txt = (t6.window_text() or "").upper()
            if t6_txt:
                medula_satir_haritasi[row] = t6_txt
        except:
            pass

    # Yöntem 2: Cache'de yoksa tum_elementler'den ilaç adlarını pozisyon sırasıyla eşleştir
    # DataItem ilaç adları zaten sıralı toplandı — Medula tablo sırasıyla genellikle aynı
    # EK descendants() çağrısı YAPMIYORUZ
    if not medula_satir_haritasi:
        for idx_pos, ilac_item in enumerate(ilaclar):
            medula_satir_haritasi[idx_pos] = (ilac_item.get("ilac_adi", "") or "").upper().split()[0]

    log(f"  [SÜRE] Satır eşleştirme: {time.time()-t1_satir:.1f}s ({len(medula_satir_haritasi)} satır)", "info")

    # Eşleştirme: her algılanan ilaç için Medula tablosundaki gerçek satırı bul.
    # Çoklu kelime + etkin madde skorlu eşleşme (ticari ad Medula t6'da farklı yazılabilir,
    # örn: t6 "METFORMIN+LINAGLIPTIN" iken DataItem "LINATIN MET ...")
    kullanilan_satirlar = set()

    def _ilac_anahtar_kelimeler(ilac_item):
        """İlaç için t6'da aranacak anahtar kelime listesi (ticari ad + etkin madde)"""
        kelimeler = []
        ilac_adi = (ilac_item.get("ilac_adi", "") or "").upper()
        etkin = (ilac_item.get("etkin_madde", "") or "").upper()
        # Ticari ad kelimeleri (4+ harf, sayısal/dozaj olmayanlar)
        for k in ilac_adi.split():
            k2 = ''.join(c for c in k if c.isalpha())
            if len(k2) >= 4 and k2 not in ("TABLET", "TABLETS", "KAPSUL", "FILM", "KAPLI",
                                             "DRAJE", "FTB", "FORT", "RETARD"):
                kelimeler.append(k2)
        # Etkin madde kelimeleri (+ ile ayrılmış olabilir)
        for k in etkin.replace("+", " ").replace(",", " ").split():
            k2 = ''.join(c for c in k if c.isalpha())
            if len(k2) >= 5 and k2 not in kelimeler:
                kelimeler.append(k2)
        return kelimeler

    def _satir_skoru(t6_txt, anahtarlar):
        """t6 metninde anahtar kelimelerin kaç tanesi geçiyor"""
        if not t6_txt:
            return 0
        skor = 0
        for k in anahtarlar:
            if k in t6_txt:
                skor += 1
        return skor

    for i, ilac_item in enumerate(ilaclar):
        anahtarlar = _ilac_anahtar_kelimeler(ilac_item)
        ilac_item["medula_satir_idx"] = i  # Varsayılan: sıra numarası
        eslesti = False

        # Skorlu eşleşme: en yüksek skorlu satırı seç (≥1 anahtar eşleşmesi şart)
        if anahtarlar and medula_satir_haritasi:
            en_iyi_row = None
            en_iyi_skor = 0
            for row, t6_txt in medula_satir_haritasi.items():
                if row in kullanilan_satirlar:
                    continue
                skor = _satir_skoru(t6_txt, anahtarlar)
                if skor > en_iyi_skor:
                    en_iyi_skor = skor
                    en_iyi_row = row
            if en_iyi_row is not None and en_iyi_skor >= 1:
                ilac_item["medula_satir_idx"] = en_iyi_row
                kullanilan_satirlar.add(en_iyi_row)
                eslesti = True
                if en_iyi_row != i:
                    log(f"  [OKU ] Satır eşleşti: {anahtarlar[0]} → row {en_iyi_row} (skor {en_iyi_skor})", "info")

        # Haritada bulunamazsa: aid_cache'den tüm satırları tara
        if not eslesti and anahtarlar:
            for row in range(20):
                if row in kullanilan_satirlar or row in medula_satir_haritasi:
                    continue
                t6 = aid_cache.get(f"f:tbl1:{row}:t6")
                if not t6:
                    try:
                        t6 = element_bul(medula, f"f:tbl1:{row}:t6")
                    except:
                        pass
                if not t6:
                    continue
                try:
                    txt = (t6.window_text() or "").upper()
                    if txt and _satir_skoru(txt, anahtarlar) >= 1:
                        ilac_item["medula_satir_idx"] = row
                        kullanilan_satirlar.add(row)
                        medula_satir_haritasi[row] = txt
                        log(f"  [OKU ] Satır eşleştirme (genişletilmiş): {anahtarlar[0]} → row {row}", "info")
                        eslesti = True
                        break
                except:
                    pass

            if not eslesti:
                log(f"  [UYARI] {anahtarlar[0] if anahtarlar else '?'}: Medula satırı bulunamadı, varsayılan idx={i}", "warn")

    # Element ID ile doğrulama/tamamlama + REÇETE DOZU okuma
    # GERÇEK medula satır indeksini kullan (enumerate değil!)
    def _cache_input_oku(auto_id):
        """Cache'deki input elementinden değer oku (hızlı: sadece cache + window_text)"""
        el = aid_cache.get(auto_id)
        if not el:
            return 0.0  # Cache'te yoksa element_bul çağırma — çok yavaş
        val = ""
        # window_text en hızlı yöntem — önce dene
        try:
            val = el.window_text() or ""
        except:
            pass
        # Boşsa get_value dene (ama legacy_properties ÇAĞIRMA — çok yavaş)
        if not val:
            try:
                val = el.get_value() or ""
            except:
                pass
        if not val:
            try:
                val = el.iface_value.CurrentValue or ""
            except:
                pass
        val = val.strip().replace(",", ".")
        try:
            return float(val) if val else 0.0
        except:
            return 0.0

    for ilac_item in ilaclar:
        i = ilac_item.get("medula_satir_idx")
        if i is None:
            continue
        try:
            # Rapor kodu doğrulama
            rk_elem = aid_cache.get(f"f:tbl1:{i}:t9") or element_bul(medula, f"f:tbl1:{i}:t9")
            if rk_elem:
                rk_txt = (rk_elem.window_text() or "").strip()
                if rk_txt and rk_txt[0].isdigit():
                    ilac_item["rapor_kodu"] = rk_txt
            # Mesaj doğrulama
            msj_elem = aid_cache.get(f"f:tbl1:{i}:t11") or element_bul(medula, f"f:tbl1:{i}:t11")
            if msj_elem:
                msj_txt = (msj_elem.window_text() or "").strip().lower()
                if msj_txt in ["var", "yok"]:
                    ilac_item["msj"] = msj_txt

            # ── REÇETE DOZU okuma (cache'den — element_bul çağrısı YOK) ──
            adet = _cache_input_oku(f"f:tbl1:{i}:t2")
            carpan = _cache_input_oku(f"f:tbl1:{i}:t3")
            doz_val = _cache_input_oku(f"f:tbl1:{i}:t4")
            periyot_sayi = _cache_input_oku(f"f:tbl1:{i}:t5")

            periyot_birim = "Günde"
            m1_elem = aid_cache.get(f"f:tbl1:{i}:m1") or element_bul(medula, f"f:tbl1:{i}:m1")
            if m1_elem:
                try:
                    pv = ""
                    try:
                        pv = m1_elem.get_value() or ""
                    except:
                        pass
                    if not pv:
                        pv = m1_elem.window_text() or ""
                    if pv:
                        for k, v in [("3", "Günde"), ("4", "Haftada"), ("5", "Ayda"), ("6", "Yılda")]:
                            if pv.strip() == k or v.lower() in pv.lower():
                                periyot_birim = v
                                break
                except:
                    pass

            if doz_val > 0 or carpan > 0:
                c = carpan if carpan > 0 else 1.0
                ps = periyot_sayi if periyot_sayi > 0 else 1.0
                d = doz_val if doz_val > 0 else 1.0
                gunluk = _gunluk_doz_hesapla(d, c, ps, periyot_birim)
                ilac_item["recete_doz"] = {
                    "adet": adet, "carpan": c, "doz": d,
                    "periyot_sayi": ps, "periyot_birim": periyot_birim,
                    "gunluk_doz": round(gunluk, 2),
                    "metin": f"{int(ps)} {periyot_birim} {int(c)} x {d}"
                }
        except:
            pass

    # ── TOPLU DOZ KARŞILAŞTIRMA (cache'den okunan reçete dozları + EOS rapor dozları) ──
    for ilac in ilaclar:
        if not ilac.get("rapor_kodu"):
            ilac["doz_kontrol"] = "raporsuz"
            ilac["doz_aciklama"] = ""
            continue

        doz_listesi = ilac.get("eos_rapor_doz_listesi", [])
        if not doz_listesi:
            eos = ilac.get("eos_rapor_doz_metin", "")
            doz_listesi = [eos] if eos else []

        if not doz_listesi:
            ilac["doz_kontrol"] = "okunamadi"
            ilac["doz_aciklama"] = f"Raporlu ({ilac['rapor_kodu']}), doz şeridi yok"
            continue

        if len(doz_listesi) > 1:
            ilac["doz_kontrol"] = "uyari"
            ilac["doz_aciklama"] = f"Birden fazla rapor dozu ({len(doz_listesi)} şerit) — manuel kontrol"
            ilac["eos_rapor_doz_listesi"] = doz_listesi
            continue

        rapor_doz_metin = doz_listesi[0]
        rapor_doz_parsed = _doz_metin_parse(rapor_doz_metin)
        if not rapor_doz_parsed:
            ilac["doz_kontrol"] = "okunamadi"
            ilac["doz_aciklama"] = f"Rapor dozu parse edilemedi: {rapor_doz_metin}"
            continue

        ilac["rapor_doz_parsed"] = rapor_doz_parsed
        recete_doz = ilac.get("recete_doz")

        if not recete_doz:
            ilac["doz_kontrol"] = "uygun"
            ilac["doz_aciklama"] = f"Rapor doz: {rapor_doz_metin}, reçete doz okunamadı"
            continue

        rapor_gunluk = rapor_doz_parsed["gunluk_doz"]
        recete_gunluk = recete_doz["gunluk_doz"]

        # ═══ BİRİM DÖNÜŞÜM: Reçete ADET bazlı, Rapor MG bazlı olabilir ═══
        # İlaç adından tablet dozajını çıkar (ör: "JARDIANCE 10 MG" → 10.0)
        tablet_dozaj = _ilac_adından_dozaj_cikar(ilac.get("ilac_adi", ""))

        # Birim uyumsuzluğu tespiti:
        # Rapor doz değeri >> reçete doz değeri ise birimler farklı demektir
        # Örn: Rapor "1x40mg" (doz=40), Reçete "1x1 adet" (doz=1) → 40 vs 1
        # Veya: Rapor doz "Dozunda" ise mg bazlı, reçete "Adet" bazlı
        rapor_doz_val = rapor_doz_parsed["doz"]
        recete_doz_val = recete_doz["doz"]

        birim_farkli = False
        # Rapor birim bilgisini parse'dan al (Adet/Miligram/Mikrogram vb.)
        rapor_birim = rapor_doz_parsed.get("doz_birim", "Adet")
        # Reçete tarafı Medula girişinde her zaman Adet bazlı (t4 = doz alanı = adet)
        # Sadece rapor mg/mcg bazlı AND reçete adet bazlıysa çevir
        if tablet_dozaj and tablet_dozaj > 1 and rapor_birim in ("Miligram", "Mikrogram"):
            # Reçete adet → mg/mcg çevir
            recete_mg = recete_gunluk * tablet_dozaj
            birim_farkli = True
            log(f"    [DOZ ] Birim dönüşümü: reçete {recete_gunluk} adet × {tablet_dozaj}mg = {recete_mg}mg/gün | rapor {rapor_gunluk} {rapor_birim.lower()}/gün", "info")
            recete_gunluk = recete_mg
            recete_doz["gunluk_doz"] = round(recete_mg, 2)
            recete_doz["metin"] += f" (×{tablet_dozaj}mg={round(recete_mg, 2)}mg)"

        uygun, aciklama_doz = doz_karsilastir(recete_gunluk, rapor_gunluk, recete_doz, rapor_doz_parsed)
        if uygun:
            ilac["doz_kontrol"] = "uygun"
        else:
            ilac["doz_kontrol"] = "uygunsuz"
        ilac["doz_aciklama"] = aciklama_doz

    sonuc["ilaclar"] = ilaclar

    # ── 9. Ek bilgiler (DataItem'lardan pozisyon bazlı — ek maliyet yok) ──
    # Medula web elementleri automation_id taşımıyor, DataItem pozisyonundan okunur.
    # Label'ın sağındaki (aynı y, daha büyük x) değeri bulma yardımcısı
    def _label_yanindaki_deger(label_text, y_tolerans=10):
        """data_items'ten label'ın sağındaki ilk değer DataItem'ını bul"""
        label_y = None
        label_x = None
        for txt, y, x, aid, elem in data_items:
            if label_text in txt:
                label_y = y
                label_x = x
                break
        if label_y is None:
            return ""
        # Aynı y satırında, label'dan sağdaki ilk anlamlı değer
        adaylar = []
        for txt, y, x, aid, elem in data_items:
            if abs(y - label_y) <= y_tolerans and x > label_x and txt != ":" and txt != label_text:
                adaylar.append((x, txt))
        if adaylar:
            adaylar.sort(key=lambda t: t[0])
            # ":" olan ilk sonucu atla
            for _, txt in adaylar:
                if txt.strip() not in (":", ""):
                    return txt.strip()
        return ""

    # e-Reçete No (y=291 satırında "e-Reçete No" label'ından sonra)
    sonuc["erecete_no"] = _label_yanindaki_deger("e-Reçete No")
    # Kapsam / Fatura Türü (y=161 satırında "Kapsam:" label'ından sonra)
    kapsam = _label_yanindaki_deger("Kapsam")
    if kapsam:
        sonuc["fatura_turu"] = kapsam
    # Hasta adı (BotanikEOS'tan — cache'den)
    sonuc["hasta_adi"] = ""
    lbl_musteri = aid_cache.get("lblMusteriAdi")
    if lbl_musteri:
        try:
            sonuc["hasta_adi"] = (lbl_musteri.window_text() or "").strip()
        except:
            pass
    # Doktor adı (BotanikEOS label'ından)
    lbl_doktor = aid_cache.get("lblDoktorAdiSoyadi")
    if lbl_doktor:
        try:
            d_txt = (lbl_doktor.window_text() or "").strip()
            if d_txt:
                sonuc["doktor_adi"] = d_txt
        except:
            pass
    # Hasta TC — DataItem olarak genellikle görünmez (INPUT alanı)
    # BotanikEOS'un lblBilgi'sinden parse edilebilir ama TC yok
    # Tesis Kodu — aynı şekilde INPUT, DataItem'da label var ama değer yok
    # Bu alanlar şimdilik boş kalır — gerektiğinde element_bul ile okunabilir

    # Uyarı kodları log
    if sonuc["uyari_kodlari"]:
        log(f"  [OKU ] Uyarı kodları: {len(sonuc['uyari_kodlari'])} adet", "info")
        for uk in sonuc["uyari_kodlari"]:
            log(f"    [OKU ] Uyarı {uk['kod']}: {uk['aciklama']} => {uk['ilac_adi']}", "info")
    if sonuc.get("_uyari_alani_metni"):
        log(f"  [OKU ] Uyarı alanı: {sonuc['_uyari_alani_metni'][:100]}", "info")

    t_toplam = time.time() - t0
    t_isleme = t_toplam - t_desc
    _ilgili = len(elem_meta)  # SKIP_CTRL ile filtrelenmemiş kalan element sayısı
    log(f"  [SÜRE] Toplu: {t_toplam:.1f}s (descendants: {t_desc:.1f}s + işleme: {t_isleme:.1f}s) "
        f"| {_ilgili}/{len(tum_elementler)} ilgili element ({len(data_item_refs)} DataItem, "
        f"{len(text_custom_refs)} Text/Custom, {len(table_refs)} Table), {len(ilaclar)} ilaç", "info")
    return sonuc


def element_bul_cached(aid_cache, medula, auto_id):
    """Önce cache'e bak, yoksa element_bul() ile ara.
    recete_tum_bilgi_topla() sonrası kullanılır - descendants() çağrısı minimize edilir.
    """
    elem = aid_cache.get(auto_id)
    if elem:
        return elem
    return element_bul(medula, auto_id)


def ilaclari_oku(medula):
    """Reçete ilaç tablosundan ilaçları oku.
    Doz bilgisi de DataItem'lardan okunur ('1 Günde 4 x 2,00 - Adet' formatı).
    NOT: Mümkünse recete_tum_bilgi_topla() tercih edin - çok daha hızlıdır.
    """
    # f:tbl1 pozisyonunu bul - ilaç tablosunun başlangıcı
    tbl_top = 0
    tbl_bottom = 9999
    try:
        tbl = element_bul(medula, "f:tbl1")
        if tbl:
            r = tbl.rectangle()
            tbl_top = r.top - 30  # biraz üstten başla
            tbl_bottom = r.bottom + 50
    except:
        pass

    items = []
    try:
        for elem in medula.descendants(control_type="DataItem"):
            try:
                txt = elem.window_text()
                r = elem.rectangle()
                if txt and txt.strip() and r.top >= tbl_top and r.top <= tbl_bottom:
                    items.append((txt.strip()[:100], r.top, r.left))
            except:
                pass
    except:
        pass
    items.sort(key=lambda x: (x[1], x[2]))

    ilaclar = []
    for txt, y, x in items:
        txt_upper = txt.upper()
        if (len(txt) > 12 and any(k in txt_upper for k in ILAC_ANAHTAR)
                and not txt.startswith("SGK") and not any(a in txt for a in ATLA)):
            ilaclar.append({"ilac_adi": txt, "etkin_madde": "", "sgk_kodu": "",
                           "rapor_kodu": "", "msj": "",
                           "eos_rapor_doz_metin": ""})  # BotanikEOS yeşil şerit = RAPOR dozu
        elif txt.startswith("SGK") and "-" in txt and ilaclar:
            parts = txt.split("-", 1)
            ilaclar[-1]["sgk_kodu"] = parts[0].strip()
            ilaclar[-1]["etkin_madde"] = parts[1].strip() if len(parts) > 1 else ""
        elif "." in txt and "/" not in txt and len(txt) <= 8 and txt[0].isdigit() and ilaclar:
            ilaclar[-1]["rapor_kodu"] = txt
        elif txt.lower() in ["var", "yok"] and ilaclar:
            ilaclar[-1]["msj"] = txt.lower()
        elif _doz_satiri_mi(txt) and ilaclar:
            # Birden fazla doz şeridi olabilir (yeşil+kırmızı karışık)
            if "eos_rapor_doz_listesi" not in ilaclar[-1]:
                ilaclar[-1]["eos_rapor_doz_listesi"] = []
            ilaclar[-1]["eos_rapor_doz_listesi"].append(txt)
            if not ilaclar[-1]["eos_rapor_doz_metin"]:
                ilaclar[-1]["eos_rapor_doz_metin"] = txt  # İlk şerit (geriye uyumluluk)
    # === EK: Medula tablo ID'leri ile doğrulama/tamamlama ===
    # Rapor kodu ve mesaj bilgisini doğrudan element ID ile oku (DataItem sıralama hatasını önler)
    for i, ilac_item in enumerate(ilaclar):
        try:
            # Rapor kodu: f:tbl1:{i}:t9
            rk_elem = element_bul(medula, f"f:tbl1:{i}:t9")
            if rk_elem:
                rk_txt = (rk_elem.window_text() or "").strip()
                if rk_txt and rk_txt[0].isdigit():
                    ilac_item["rapor_kodu"] = rk_txt
            # Mesaj: f:tbl1:{i}:t11
            msj_elem = element_bul(medula, f"f:tbl1:{i}:t11")
            if msj_elem:
                msj_txt = (msj_elem.window_text() or "").strip().lower()
                if msj_txt in ["var", "yok"]:
                    ilac_item["msj"] = msj_txt
        except:
            pass

    # === TOPLU DOZ KARŞILAŞTIRMA ===
    # Tüm raporlu ilaçlar için reçete dozunu oku ve rapor dozu ile karşılaştır
    _toplu_doz_karsilastir(medula, ilaclar)

    return ilaclar


def _recete_dozu_element_oku(medula, satir_idx):
    """Reçete ilaç satırından doz bilgilerini element ID ile oku.
    recete_dozu_oku() ile aynı mantık ama sadece element ID kullanır.
    Returns: dict {adet, carpan, doz, periyot_sayi, periyot_birim, gunluk_doz, metin} veya None
    """
    def oku_input(auto_id):
        elem = element_bul(medula, auto_id)
        if not elem:
            return 0.0
        val = ""
        try:
            val = elem.get_value() or ""
        except:
            pass
        if not val:
            try:
                val = elem.iface_value.CurrentValue or ""
            except:
                pass
        if not val:
            try:
                lp = elem.legacy_properties()
                val = lp.get("Value", "") or ""
            except:
                pass
        if not val:
            val = elem.window_text() or ""
        val = val.strip().replace(",", ".")
        try:
            return float(val) if val else 0.0
        except:
            return 0.0

    adet = oku_input(f"f:tbl1:{satir_idx}:t2")
    carpan = oku_input(f"f:tbl1:{satir_idx}:t3")
    doz = oku_input(f"f:tbl1:{satir_idx}:t4")
    periyot_sayi = oku_input(f"f:tbl1:{satir_idx}:t5")

    periyot_birim = "Günde"
    elem = element_bul(medula, f"f:tbl1:{satir_idx}:m1")
    if elem:
        try:
            val = ""
            try:
                val = elem.get_value() or ""
            except:
                pass
            if not val:
                val = elem.window_text() or ""
            if val:
                for k, v in [("3", "Günde"), ("4", "Haftada"), ("5", "Ayda"), ("6", "Yılda")]:
                    if val.strip() == k or v.lower() in val.lower():
                        periyot_birim = v
                        break
        except:
            pass

    if doz == 0 and carpan == 0:
        return None

    carpan = carpan if carpan > 0 else 1.0
    periyot_sayi = periyot_sayi if periyot_sayi > 0 else 1.0
    doz = doz if doz > 0 else 1.0

    gunluk = _gunluk_doz_hesapla(doz, carpan, periyot_sayi, periyot_birim)

    return {
        "adet": adet, "carpan": carpan, "doz": doz,
        "periyot_sayi": periyot_sayi, "periyot_birim": periyot_birim,
        "gunluk_doz": round(gunluk, 2),
        "metin": f"{int(periyot_sayi)} {periyot_birim} {int(carpan)} x {doz}"
    }


def _ilac_adından_dozaj_cikar(ilac_adi):
    """İlaç adından tablet/kapsül dozajını çıkar.
    Örn: 'NEURONTIN 800 MG 50 KAPSUL' → 800.0
         'RISPERDAL 4 MG 20 TB' → 4.0
    NOT: 'DEPAKIN SURUP 150 ML' gibi şurup/jel şişelerinde ML = hacim,
    tablet dozaj değildir. Sadece MG/MCG/IU/ÜNİTE kabul edilir.
    Returns: float veya None
    """
    import re
    # Tablet/kapsül dozajı: sayı + MG/MCG/IU/Ünite (ML YOK — şişe hacmi olabilir)
    match = re.search(r'(\d+[,.]?\d*)\s*(MG|MCG|IU|ÜNITE|UNITE|MIKROGRAM)\b', ilac_adi.upper())
    if match:
        try:
            return float(match.group(1).replace(",", "."))
        except:
            pass
    return None


def _toplu_doz_karsilastir(medula, ilaclar):
    """Tüm ilaçların reçete ve rapor dozlarını tek seferde oku ve karşılaştır.
    Sonuçları ilac dict'lerine ekler:
      - ilac["doz_kontrol"]: "uygun" / "uygunsuz" / "uyari" / "raporsuz" / "okunamadi"
      - ilac["doz_aciklama"]: Detaylı açıklama
      - ilac["recete_doz"]: Parse edilmiş reçete dozu dict
      - ilac["rapor_doz_parsed"]: Parse edilmiş rapor dozu dict
    """
    for i, ilac in enumerate(ilaclar):
        # Raporsuz ilaçlar: doz kontrolü yok
        if not ilac.get("rapor_kodu"):
            ilac["doz_kontrol"] = "raporsuz"
            ilac["doz_aciklama"] = ""
            continue

        # Rapor dozu şeritleri
        doz_listesi = ilac.get("eos_rapor_doz_listesi", [])
        if not doz_listesi:
            eos = ilac.get("eos_rapor_doz_metin", "")
            doz_listesi = [eos] if eos else []

        if not doz_listesi:
            ilac["doz_kontrol"] = "okunamadi"
            ilac["doz_aciklama"] = f"Raporlu ({ilac['rapor_kodu']}), doz şeridi yok"
            continue

        # Birden fazla şerit varsa: manuel kontrol gerekir
        if len(doz_listesi) > 1:
            ilac["doz_kontrol"] = "uyari"
            ilac["doz_aciklama"] = f"Birden fazla rapor dozu ({len(doz_listesi)} şerit) — manuel kontrol"
            ilac["eos_rapor_doz_listesi"] = doz_listesi
            continue

        # Tek şerit — rapor dozunu parse et
        rapor_doz_metin = doz_listesi[0]
        rapor_doz_parsed = _doz_metin_parse(rapor_doz_metin)
        if not rapor_doz_parsed:
            ilac["doz_kontrol"] = "okunamadi"
            ilac["doz_aciklama"] = f"Rapor dozu parse edilemedi: {rapor_doz_metin}"
            continue

        ilac["rapor_doz_parsed"] = rapor_doz_parsed

        # Reçete dozunu oku (element ID ile)
        recete_doz = _recete_dozu_element_oku(medula, i)
        ilac["recete_doz"] = recete_doz

        if not recete_doz:
            ilac["doz_kontrol"] = "uygun"
            ilac["doz_aciklama"] = f"Rapor doz: {rapor_doz_metin}, reçete doz okunamadı"
            continue

        rapor_gunluk = rapor_doz_parsed["gunluk_doz"]
        recete_gunluk = recete_doz["gunluk_doz"]

        # ═══ BİRİM DÖNÜŞÜM: Reçete ADET, Rapor MG olabilir (veya tersi) ═══
        tablet_dozaj = _ilac_adından_dozaj_cikar(ilac.get("ilac_adi", ""))
        rapor_doz_val = rapor_doz_parsed["doz"]
        recete_doz_val = recete_doz["doz"]

        if tablet_dozaj and tablet_dozaj > 1:
            if rapor_doz_val >= tablet_dozaj * 0.5 and recete_doz_val <= 10:
                # Reçete adet bazlı, rapor mg bazlı → reçeteyi mg'ye çevir
                recete_mg = recete_gunluk * tablet_dozaj
                log(f"    [DOZ ] Birim dönüşümü: reçete {recete_gunluk} adet × {tablet_dozaj}mg = {recete_mg}mg/gün | rapor {rapor_gunluk}mg/gün", "info")
                recete_gunluk = recete_mg
                recete_doz["gunluk_doz"] = round(recete_mg, 2)
                recete_doz["metin"] += f" (×{tablet_dozaj}mg={round(recete_mg, 2)}mg)"
            elif recete_doz_val >= tablet_dozaj * 0.5 and rapor_doz_val <= 10:
                # Tam tersi: reçete mg bazlı, rapor adet bazlı
                rapor_mg = rapor_gunluk * tablet_dozaj
                log(f"    [DOZ ] Birim dönüşümü (ters): rapor {rapor_gunluk} adet × {tablet_dozaj}mg = {rapor_mg}mg/gün | reçete {recete_gunluk}mg/gün", "info")
                rapor_gunluk = rapor_mg

        # Periyot birim düzeltmesi (reçete birimi yanlış okunmuş olabilir)
        r_birim = recete_doz.get("periyot_birim", "Günde")
        rp_birim = rapor_doz_parsed.get("periyot_birim", "Günde")
        if r_birim == "Günde" and rp_birim in ("Ayda", "Haftada", "Yılda"):
            recete_gunluk = round(_gunluk_doz_hesapla(
                recete_doz["doz"], recete_doz["carpan"],
                recete_doz["periyot_sayi"], rp_birim), 2)

        # Karşılaştır
        uygun, doz_aciklama = doz_karsilastir(recete_gunluk, rapor_gunluk, recete_doz, rapor_doz_parsed)
        if not uygun:
            ilac["doz_kontrol"] = "uygunsuz"
        elif "⚠" in doz_aciklama or "manuel kontrol" in doz_aciklama:
            ilac["doz_kontrol"] = "uyari"  # rapor üstü ama 2× altında
        else:
            ilac["doz_kontrol"] = "uygun"
        ilac["doz_aciklama"] = doz_aciklama


def _doz_satiri_mi(txt):
    """DataItem metninin doz satırı olup olmadığını kontrol et.
    Örnekler: '1 Günde 4 x 2,00 - Adet', '( Günde 2 x 1,000 Dozunda)'"""
    import re
    return bool(re.search(r'(G.nde|Haftada|Ayda|Y.lda|Saatte)\s+\d+\s*x\s*[\d,\.]+', txt))


# ========== DB ==========
def db_baglanti():
    """DB bağlantısı aç, tablolar yoksa oluştur"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    # Tablo yoksa oluştur
    cur.execute('''CREATE TABLE IF NOT EXISTS etkin_madde_kurallari (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        etkin_madde TEXT UNIQUE,
        sgk_kodu TEXT DEFAULT '',
        sut_maddesi TEXT DEFAULT '',
        rapor_kodu TEXT DEFAULT '',
        rapor_gerekli INTEGER DEFAULT 0,
        raporlu_maks_doz TEXT DEFAULT '',
        kontrol_tipi TEXT DEFAULT 'bilinmiyor',
        birlikte_yasaklar TEXT DEFAULT '',
        aciklama TEXT DEFAULT '',
        olusturma_tarihi TEXT,
        aktif INTEGER DEFAULT 1
    )''')
    cur.execute('''CREATE TABLE IF NOT EXISTS ilac_mesaj_kurallari (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        etkin_madde TEXT DEFAULT '',
        ilac_adi_pattern TEXT DEFAULT '',
        mesaj_pattern TEXT DEFAULT '',
        sut_maddesi TEXT DEFAULT '',
        rapor_kodu TEXT DEFAULT '',
        aksiyon TEXT DEFAULT '',
        kosullar TEXT DEFAULT '',
        aciklama TEXT DEFAULT '',
        olusturma_tarihi TEXT,
        guncelleme_tarihi TEXT,
        aktif INTEGER DEFAULT 1
    )''')
    # Kontrol sonuçları tablosu (loglama)
    cur.execute('''CREATE TABLE IF NOT EXISTS kontrol_sonuclari (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT,
        grup TEXT,
        recete_no TEXT,
        recete_turu TEXT,
        ilac_adi TEXT,
        etkin_madde TEXT,
        sgk_kodu TEXT,
        rapor_kodu TEXT,
        msj TEXT,
        kontrol_tipi TEXT,
        renkli_kontrol TEXT,
        uyari_kodu_kontrol TEXT,
        doz_kontrol TEXT,
        sut_kontrol TEXT,
        genel_sonuc TEXT,
        aciklama TEXT,
        recete_dozu TEXT,
        rapor_dozu TEXT,
        mesaj_metni TEXT
    )''')
    # Öğrenilen uyarı kodları tablosu
    cur.execute('''CREATE TABLE IF NOT EXISTS ogrenilen_uyari_kodlari (
        kod TEXT PRIMARY KEY,
        aciklama TEXT DEFAULT '',
        ilk_gorulme TEXT,
        son_gorulme TEXT,
        gorulme_sayisi INTEGER DEFAULT 1
    )''')
    conn.commit()
    return conn, cur


def uyari_kodu_kaydet(cur, conn, kod, aciklama=""):
    """Tespit edilen uyarı kodunu DB'ye kaydet (combobox için)"""
    try:
        cur.execute("""INSERT INTO ogrenilen_uyari_kodlari (kod, aciklama, ilk_gorulme, son_gorulme)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(kod) DO UPDATE SET
                son_gorulme = excluded.son_gorulme,
                gorulme_sayisi = gorulme_sayisi + 1""",
            (str(kod), aciklama, datetime.now().isoformat(), datetime.now().isoformat()))
        conn.commit()
    except:
        pass


def db_kural_bul(cur, etkin_madde):
    """Etkin maddeye göre kural bul"""
    if not etkin_madde:
        return None
    cur.execute("SELECT * FROM etkin_madde_kurallari WHERE etkin_madde = ? AND aktif = 1",
                (etkin_madde.upper(),))
    row = cur.fetchone()
    if row:
        return dict(row)
    cur.execute("SELECT * FROM etkin_madde_kurallari WHERE ? LIKE '%' || etkin_madde || '%' AND aktif = 1",
                (etkin_madde.upper(),))
    row = cur.fetchone()
    return dict(row) if row else None


def db_kaydet(cur, conn, etkin, sgk, rapor_kodu="", rapor_gerekli=0, tip="bilinmiyor", aciklama=""):
    """Yeni etkin madde kuralı kaydet"""
    try:
        cur.execute('''INSERT OR IGNORE INTO etkin_madde_kurallari
            (etkin_madde, sgk_kodu, sut_maddesi, rapor_kodu, rapor_gerekli,
             raporlu_maks_doz, kontrol_tipi, birlikte_yasaklar, aciklama, olusturma_tarihi, aktif)
            VALUES (?, ?, '', ?, ?, '', ?, '', ?, ?, 1)''',
            (etkin.upper(), sgk, rapor_kodu, rapor_gerekli, tip, aciklama,
             datetime.now().isoformat()))
        conn.commit()
        return True
    except:
        return False


# ========== İLAÇ KONTROL ==========
def ilac_kontrol_et(cur, ilac):
    """Tek ilaç için SUT uygunluk kontrolü
    Returns: (durum, aciklama)
      durum: OK, UYARI, SORUN, YENİ
    """
    etkin = ilac.get("etkin_madde", "")
    rapor_kodu = ilac.get("rapor_kodu", "")
    msj = ilac.get("msj", "")

    kural = db_kural_bul(cur, etkin)
    if not kural:
        return "YENİ", "Veritabanında kural yok"

    sut = kural.get("sut_maddesi", "")
    sorunlar = []

    # 1. Rapor kontrolü
    if kural["rapor_gerekli"]:
        if not rapor_kodu:
            return "SORUN", f"RAPOR GEREKLİ ama yok! SUT {sut}"

    # 2. Rapor kodu uyumu
    beklenen_rapor = kural.get("rapor_kodu", "")
    if rapor_kodu and beklenen_rapor and rapor_kodu != beklenen_rapor:
        if rapor_kodu[:2] != beklenen_rapor[:2]:
            sorunlar.append(f"Rapor kodu uyumsuz: {rapor_kodu} (beklenen: {beklenen_rapor})")

    # 3. Mesaj kontrolü
    if msj == "var":
        sorunlar.append("Mesaj VAR - kontrol edilmeli")

    # Sonuç
    if sorunlar:
        detay = " | ".join(sorunlar)
        if any("RAPOR GEREKLİ" in s for s in sorunlar):
            return "SORUN", detay
        return "UYARI", f"Raporlu ({rapor_kodu}), SUT {sut} | {detay}"

    if kural["rapor_gerekli"]:
        return "OK", f"Raporlu ({rapor_kodu}), SUT {sut}"
    return "OK", f"Raporsuz verilebilir"


# ========== SONUÇ LOGLAMA (DB) ==========
def sonuc_logla(cur, conn, grup, recete_no, recete_turu, ilac, kontrol_tipi,
                renkli_k, uyari_k, doz_k, sut_k, genel, aciklama,
                recete_dozu="", rapor_dozu="", mesaj_metni=""):
    """Kontrol sonucunu veritabanına logla"""
    try:
        cur.execute('''INSERT INTO kontrol_sonuclari
            (tarih, grup, recete_no, recete_turu, ilac_adi, etkin_madde, sgk_kodu,
             rapor_kodu, msj, kontrol_tipi, renkli_kontrol, uyari_kodu_kontrol,
             doz_kontrol, sut_kontrol, genel_sonuc, aciklama, recete_dozu, rapor_dozu, mesaj_metni)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (datetime.now().isoformat(), grup, recete_no, recete_turu,
             ilac.get("ilac_adi", ""), ilac.get("etkin_madde", ""), ilac.get("sgk_kodu", ""),
             ilac.get("rapor_kodu", ""), ilac.get("msj", ""),
             kontrol_tipi, renkli_k, uyari_k, doz_k, sut_k, genel, aciklama,
             recete_dozu, rapor_dozu, mesaj_metni))
        conn.commit()
    except Exception as e:
        log(f"DB loglama hatası: {e}", "error")


# ========== CHECKBOX + İLAÇ BİLGİ PENCERESİ ==========
def checkbox_sec(medula, satir_idx):
    """İlaç satırının checkbox'ını seç (birden fazla yöntem dener).

    CLAUDE.md kuralı: Checkbox için click_input öncelikli — IE-embed Medula'da
    toggle/invoke sessiz başarısız olabilir (exception yok ama checkbox işaretlenmiyor),
    bu da "Lütfen ilaç seçiniz" popup'ına yol açar. click_input gerçek mouse click
    simüle eder, IE'nin DOM event handler'ları tetiklenir.
    """
    auto_id = f"f:tbl1:{satir_idx}:checkbox7"
    elem = element_bul(medula, auto_id)
    if not elem:
        return False
    # Yöntem 1 (öncelikli): click_input — IE-embed için en güvenilir
    try:
        elem.click_input()
        aktivite_bildir()
        time.sleep(0.3)
        return True
    except Exception:
        pass
    # Yöntem 2: toggle (UIA TogglePattern)
    try:
        elem.toggle()
        aktivite_bildir()
        time.sleep(0.3)
        return True
    except Exception:
        pass
    # Yöntem 3: invoke (last resort)
    try:
        elem.invoke()
        aktivite_bildir()
        time.sleep(0.3)
        return True
    except Exception:
        pass
    return False


def ilac_bilgi_ac(medula):
    """İlaç Bilgi butonuna tıkla ve pencere açılana kadar bekle"""
    if not element_tikla(medula, "f:buttonIlacBilgiGorme"):
        log("    İlaç Bilgi butonu bulunamadı", "warn")
        return False

    # Pencere açılana kadar bekle (form1:textarea1 veya form1:buttonKapat)
    for _ in range(10):
        time.sleep(1)
        if element_bul(medula, "form1:buttonKapat"):
            time.sleep(0.5)
            return True
    log("    İlaç Bilgi penceresi açılmadı", "warn")
    # Kurtarma: Escape tuşu ile olası popup'ı kapat
    pyautogui.press("escape")
    time.sleep(1)
    return False


def ilac_bilgi_mesaj_oku(medula):
    """İlaç Bilgi penceresinden mesaj başlığı ve metin oku.
    Returns: (baslik_listesi, mesaj_metni)
    """
    basliklar = []
    mesaj_metni = ""

    # Mesaj başlıkları (birden fazla olabilir)
    for idx in range(10):
        auto_id = f"form1:tableExIlacMesajListesi:{idx}:text19"
        elem = element_bul(medula, auto_id)
        if not elem:
            break
        try:
            txt = elem.window_text()
            if txt and txt.strip():
                basliklar.append(txt.strip())
        except:
            pass

    # İlk başlığa tıkla (mesaj metnini yüklemek için)
    if basliklar:
        element_tikla(medula, "form1:tableExIlacMesajListesi:0:text19", "click")
        time.sleep(1)

    # Mesaj metni (textarea)
    elem = element_bul(medula, "form1:textarea1")
    if elem:
        try:
            mesaj_metni = elem.window_text() or ""
            mesaj_metni = mesaj_metni.strip()
        except:
            pass

    return basliklar, mesaj_metni


def ilac_bilgi_kapat(medula):
    """İlaç Bilgi penceresini kapat"""
    element_tikla(medula, "form1:buttonKapat")
    time.sleep(1)


def ilac_bilgi_ac_oku_kapat(medula, satir_idx):
    """Tam akış: checkbox seç → İlaç Bilgi aç → mesaj oku → kapat
    Returns: (basliklar, mesaj_metni) veya ([], "")
    """
    if not checkbox_sec(medula, satir_idx):
        log(f"    Checkbox seçilemedi (satır {satir_idx})", "warn")
        return [], ""

    if not ilac_bilgi_ac(medula):
        return [], ""

    basliklar, mesaj_metni = ilac_bilgi_mesaj_oku(medula)
    ilac_bilgi_kapat(medula)
    return basliklar, mesaj_metni


# ========== REÇETE DOZU OKUMA ==========
def medula_satir_idx_bul(medula, ilac_adi):
    """Medula tablo satır indeksini ilaç adıyla eşleştir.
    f:tbl1:0:t6, f:tbl1:1:t6... satırlarını tarar.
    Returns: int (satır indeksi) veya None
    """
    ilac_kisa = (ilac_adi or "").upper().split()[0] if ilac_adi else ""
    if not ilac_kisa:
        return None
    for row in range(20):
        elem = element_bul(medula, f"f:tbl1:{row}:t6")
        if not elem:
            break
        try:
            txt = (elem.window_text() or "").upper()
            if ilac_kisa in txt:
                return row
        except:
            pass
    return None


def recete_dozu_oku(medula, satir_idx, doz_metin=""):
    """Reçete ilaç satırından doz bilgilerini oku.
    Önce DataItem doz_metin'den parse eder (güvenilir).
    Fallback olarak Edit elementlerini dener.
    Returns: dict {adet, carpan, doz, periyot_sayi, periyot_birim, gunluk_doz} veya None
    """
    import re

    # Yöntem 1: DataItem doz metni parse et ("1 Günde 4 x 2,00 - Adet")
    if doz_metin:
        parsed = _doz_metin_parse(doz_metin)
        if parsed:
            return parsed

    # Yöntem 2: Edit elementlerinden oku (get_value - IE embedded input'ları)
    def oku_input(auto_id):
        elem = element_bul(medula, auto_id)
        if not elem:
            return 0.0
        val = ""
        # get_value() IE embedded input'lardan değer okur
        try:
            val = elem.get_value() or ""
        except:
            pass
        if not val:
            try:
                val = elem.iface_value.CurrentValue or ""
            except:
                pass
        if not val:
            try:
                lp = elem.legacy_properties()
                val = lp.get("Value", "") or ""
            except:
                pass
        if not val:
            val = elem.window_text() or ""
        val = val.strip().replace(",", ".")
        try:
            return float(val) if val else 0.0
        except:
            return 0.0

    adet = oku_input(f"f:tbl1:{satir_idx}:t2")
    carpan = oku_input(f"f:tbl1:{satir_idx}:t3")
    doz = oku_input(f"f:tbl1:{satir_idx}:t4")
    periyot_sayi = oku_input(f"f:tbl1:{satir_idx}:t5")

    periyot_birim = None  # None = okunamadı (varsayılan Günde yerine)
    elem = element_bul(medula, f"f:tbl1:{satir_idx}:m1")
    if elem:
        try:
            val = ""
            try:
                val = elem.get_value() or ""
            except:
                pass
            if not val:
                val = elem.window_text() or ""
            val = val.strip()
            birim_map = {"3": "Günde", "4": "Haftada", "5": "Ayda", "6": "Yılda",
                         "Günde": "Günde", "Haftada": "Haftada", "Ayda": "Ayda", "Yılda": "Yılda"}
            periyot_birim = birim_map.get(val)
        except:
            pass
    if not periyot_birim:
        periyot_birim = "Günde"  # Varsayılan (rapor dozu varsa karşılaştırmada düzeltilecek)

    if adet == 0 and doz == 0 and carpan == 0:
        return None

    carpan = carpan if carpan > 0 else 1.0
    doz = doz if doz > 0 else 1.0
    periyot_sayi = periyot_sayi if periyot_sayi > 0 else 1.0
    gunluk_doz = _gunluk_doz_hesapla(doz, carpan, periyot_sayi, periyot_birim)

    return {
        "adet": adet, "carpan": carpan, "doz": doz,
        "periyot_sayi": periyot_sayi, "periyot_birim": periyot_birim,
        "gunluk_doz": round(gunluk_doz, 2),
        "metin": f"{periyot_birim} {carpan} x {doz}"
    }


def _doz_metin_parse(doz_metin):
    """DataItem doz metnini parse et.
    Formatlar:
      '1 Günde 4 x 2,00 - Adet'
      '( Günde 2 x  1,000 Dozunda)'
    Returns: dict veya None
    """
    import re
    # Periyot birimi bul
    birim_match = re.search(r'(G\S*nde|Haftada|Ayda|Y\S*lda|Saatte)', doz_metin)
    if not birim_match:
        return None

    periyot_birim_raw = birim_match.group(1)
    # Normalize
    if "nde" in periyot_birim_raw.lower():
        periyot_birim = "Günde"
    elif "Hafta" in periyot_birim_raw:
        periyot_birim = "Haftada"
    elif "Ay" in periyot_birim_raw:
        periyot_birim = "Ayda"
    elif "lda" in periyot_birim_raw.lower():
        periyot_birim = "Yılda"
    else:
        periyot_birim = "Saatte"

    # Birimden önceki sayı (periyot sayısı) ve sonraki "çarpan x doz"
    before = doz_metin[:birim_match.start()]
    after = doz_metin[birim_match.end():]

    # Periyot sayısı (birimden önceki son sayı)
    periyot_match = re.search(r'(\d+)\s*$', before.strip())
    periyot_sayi = float(periyot_match.group(1)) if periyot_match else 1.0

    # Çarpan x Doz (birimden sonraki kısım)
    xdoz_match = re.search(r'(\d+)\s*x\s*([\d,\.]+)', after)
    if not xdoz_match:
        return None

    carpan = float(xdoz_match.group(1))
    doz = float(xdoz_match.group(2).replace(",", "."))

    # Doz birimi (Adet / Miligram / Damla / Ünite / Mikrogram / ml)
    # Örn: "1 Günde 4 x 2,00 - Adet" | "Günde 2 x 1,000 Dozunda" | "Günde 1 x 60.0 Miligram"
    doz_birim = "Adet"  # Varsayılan
    after_xdoz = after[xdoz_match.end():].lower()
    if "miligram" in after_xdoz or "mg" in after_xdoz:
        doz_birim = "Miligram"
    elif "mikrogram" in after_xdoz or "mcg" in after_xdoz or "µg" in after_xdoz:
        doz_birim = "Mikrogram"
    elif "damla" in after_xdoz:
        doz_birim = "Damla"
    elif "ünite" in after_xdoz or "unite" in after_xdoz or "iu" in after_xdoz:
        doz_birim = "Ünite"
    elif "ml" in after_xdoz:
        doz_birim = "ml"
    elif "doz" in after_xdoz:  # "Dozunda" → birim belirsiz, adet gibi davran
        doz_birim = "Adet"

    gunluk_doz = _gunluk_doz_hesapla(doz, carpan, periyot_sayi, periyot_birim)

    return {
        "adet": 0, "carpan": carpan, "doz": doz,
        "periyot_sayi": periyot_sayi, "periyot_birim": periyot_birim,
        "doz_birim": doz_birim,
        "gunluk_doz": round(gunluk_doz, 2),
        "metin": f"{periyot_birim} {periyot_sayi} x {doz}"
    }


def _gunluk_doz_hesapla(doz, carpan, periyot_sayi, periyot_birim):
    """Günlük doz hesapla.
    Format: "{periyot_sayi} {periyot_birim} {carpan} x {doz}"
    Anlam: her `periyot_sayi × periyot_birim` sürede `carpan × doz` alınır.
    Örn: "90 Günde 1 x 100" = her 90 günde 1×100 birim → günlük = 100/90
    Örn: "1 Günde 3 x 1.0" = her 1 günde 3×1 birim → günlük = 3
    """
    carpan = carpan if carpan > 0 else 1.0
    periyot_sayi = periyot_sayi if periyot_sayi > 0 else 1.0
    toplam = doz * carpan
    # periyot_sayi gün cinsinden kaç günlük peryotta uygulandığını gösterir
    if "Hafta" in periyot_birim:
        periyot_gun = periyot_sayi * 7.0
    elif "Ay" in periyot_birim:
        periyot_gun = periyot_sayi * 30.0
    elif "Yıl" in periyot_birim or "Y\u0131l" in periyot_birim:
        periyot_gun = periyot_sayi * 365.0
    else:  # Günde
        periyot_gun = periyot_sayi
    return toplam / periyot_gun


# ========== RAPOR SAYFASI ==========
def rapor_ac(medula, satir_idx=None):
    """Rapor butonuna tıkla ve rapor sayfası açılana kadar bekle.
    2 deneme hakkı: ilk denemede Rapor butonuna tıkla → 4s bekle;
    açılmazsa popup kapat + checkbox tekrar seç + Rapor butonu + 6s bekle.

    satir_idx verilirse, ilk deneme başarısız olduğunda checkbox tekrar
    işaretlenir — IE-embed'de checkbox'ın sessiz başarısız olduğu durumlarda
    Medula 'Lütfen ilaç seçiniz' popup'ı gösterir, bu retry onu çözer.
    """
    global _aid_cache_global
    _aid_cache_global = {}  # Sayfa değişiyor, cache geçersiz

    for deneme in range(2):
        if not element_tikla(medula, "f:buttonRaporGoruntule"):
            if deneme == 0:
                # İlk denemede buton bulunamazsa Escape + tekrar dene
                pyautogui.press("escape")
                time.sleep(0.5)
                continue
            log("    Rapor butonu bulunamadı", "warn")
            return False

        # Bekleme: form1:buttonGeriDon veya tableEx1 görününce açılmıştır
        timeout = 4 if deneme == 0 else 6
        try:
            cw = medula.child_window(auto_id="form1:buttonGeriDon", found_index=0)
            if cw.exists(timeout=timeout):
                return True
        except Exception:
            pass
        if element_bul(medula, "form1:tableEx1"):
            return True

        # Açılmadı — popup ("Lütfen ilaç seçiniz") kapat + checkbox tekrar dene
        if deneme == 0:
            log("    Rapor sayfası açılmadı (1. deneme) — popup kontrolü + checkbox tekrar deneniyor", "warn")
            # 1. Popup'ı kapat (Lütfen ilaç seçiniz vb.)
            popup_kapatildi = popup_kapat()
            if not popup_kapatildi:
                pyautogui.press("escape")
                time.sleep(0.5)
            # 2. Checkbox tekrar işaretle (silent fail durumu için)
            if satir_idx is not None:
                if checkbox_sec(medula, satir_idx):
                    log(f"    Checkbox tekrar işaretlendi (satır {satir_idx})", "info")
                else:
                    log(f"    Checkbox tekrar işaretlenemedi (satır {satir_idx})", "warn")
            time.sleep(0.5)

    log("    Rapor sayfası açılmadı (2 deneme başarısız)", "warn")
    return False


def rapor_tedavi_semasi_doz_oku(medula, etkin_madde_aranan):
    """Rapor tedavi şemasından eşleşen ilaç satırının dozunu oku.
    Returns: (doz_metni, gunluk_doz) veya (None, None)
    """
    etkin_upper = (etkin_madde_aranan or "").upper()

    for row in range(20):
        # Etkin madde
        auto_id = f"form1:tableEx1:{row}:text63"
        elem = element_bul(medula, auto_id)
        if not elem:
            break
        try:
            etkin = (elem.window_text() or "").strip().upper()
        except:
            continue

        # Eşleşme kontrolü (tam veya kısmi)
        if not etkin:
            continue
        if etkin_upper not in etkin and etkin not in etkin_upper:
            # Kısmi kelime eşleşmesi dene
            kelimeler = etkin_upper.split()
            if not any(k in etkin for k in kelimeler if len(k) > 3):
                continue

        # Tedavi şeması dozunu oku (text76: "Günde 1 x 1.0 Adet")
        doz_elem = element_bul(medula, f"form1:tableEx1:{row}:text76")
        if doz_elem:
            try:
                doz_metni = (doz_elem.window_text() or "").strip()
                if doz_metni:
                    # "Günde 1 x 1.0 Adet" formatını parse et
                    gunluk_doz = _rapor_doz_parse(doz_metni)
                    return doz_metni, gunluk_doz
            except:
                pass

    return None, None


def _rapor_doz_parse(doz_metni):
    """Rapor doz metnini parse et: 'Günde 1 x 1.0 Adet' → günlük doz float
    Format: {Periyot} {sayı} x {doz} {birim}
    """
    import re
    match = re.search(r'(\w+)\s+([\d,\.]+)\s*x\s*([\d,\.]+)', doz_metni)
    if not match:
        return None

    periyot = match.group(1)
    sayi = float(match.group(2).replace(",", "."))
    doz = float(match.group(3).replace(",", "."))

    gunluk = sayi * doz
    if "Hafta" in periyot:
        gunluk /= 7.0
    elif "Ay" in periyot:
        gunluk /= 30.0
    elif "Yıl" in periyot:
        gunluk /= 365.0

    return round(gunluk, 2)


def botanik_eos_doz_penceresi_oku():
    """BotanikEOS'un rapor sayfasına girildiğinde açtığı küçük doz penceresi.
    WinForms tablosu: grdUrunEtkin... - Etkin, Reçetedeki ilaç, Doz, Rapor Kodu sütunları.
    Returns: list of dict [{etkin, ilac, doz, rapor_kodu}, ...]
    """
    satirlar = []
    try:
        desktop = Desktop(backend="uia")
        for w in desktop.windows():
            try:
                # WinForms penceresi - küçük popup
                for child in w.descendants(control_type="Table"):
                    try:
                        aid = child.element_info.automation_id or ""
                        if "grdUrunEtkin" in aid:
                            # Satırları oku
                            for item in child.descendants(control_type="ListItem"):
                                try:
                                    satir_data = {}
                                    for cell in item.children():
                                        try:
                                            name = cell.window_text() or ""
                                            cell_name = cell.element_info.name or ""
                                            if "Etkin" in cell_name and "satır" in cell_name:
                                                satir_data["etkin"] = name
                                            elif "ilaç" in cell_name and "satır" in cell_name:
                                                satir_data["ilac"] = name
                                            elif "Doz" in cell_name and "satır" in cell_name:
                                                satir_data["doz"] = name
                                            elif "Rapor Kodu" in cell_name:
                                                satir_data["rapor_kodu"] = name
                                        except:
                                            pass
                                    if satir_data:
                                        satirlar.append(satir_data)
                                except:
                                    pass
                            return satirlar
                    except:
                        pass
            except:
                pass
    except:
        pass
    return satirlar


def erecete_aciklama_oku(medula):
    """E-Reçete Görüntüle sayfasını aç, açıklama ve tanı metinlerini oku, geri dön.
    Returns: dict {aciklamalar, tanilar, tum_metin} veya None
    """
    import pyautogui
    pyautogui.FAILSAFE = False

    btn = element_bul(medula, "f:buttonEreceteGoruntule")
    if not btn:
        return None

    btn.invoke()
    time.sleep(1.5)

    # Optimize: TEK descendants() çağrısı + scroll sonrası sadece yeni elementleri oku
    sonuc = {"aciklamalar": [], "tanilar": [], "tum_metin": ""}
    parcalar_set = set()

    tani_keywords = ('diyabet', 'diabetes', 'mellitus', 'hipertansiyon',
                     'astim', 'koah', 'kronik', 'insülin',
                     'kby', 'böbrek', 'bobrek', 'hemodiyaliz', 'diyaliz',
                     'anemi', 'hemofili', 'romatoid', 'psoriazis')
    aciklama_keywords = ('metformin', 'sülfonil', 'sulfonil', 'glisemik',
                         'monoterapi', 'kontrol', 'tedavi edildi',
                         # Lab değerleri (ESA, IVIG, antiviral için kritik)
                         'hb', 'hemoglobin', 'ferritin', 'tsat', 'transferrin',
                         'hbv', 'hcv', 'cd4', 'alt', 'ast', 'igg', 'kreatinin',
                         'egfr', 'hba1c', 'ldl', 'trigliserid',
                         # ESA doz anahtarları
                         'iu/kg', 'mcg/kg', 'kuru kilo', 'idame', 'başlangıç',
                         'baslangic', 'uygulanacak doz', 'tedavi süresi',
                         # Nöropatik ağrı
                         'nöropati', 'noropati', 'fibromiyalji', 'phn')

    def _elementleri_tara():
        for d in medula.descendants():
            try:
                txt = d.window_text() or ''
                ts = txt.strip()
                if len(ts) < 5 or ts in parcalar_set:
                    continue
                r = d.rectangle()
                if r.top < 200:
                    continue

                parcalar_set.add(ts)

                tl = ts.lower()
                if any(k in tl for k in tani_keywords):
                    if ts not in sonuc["tanilar"]:
                        sonuc["tanilar"].append(ts)
                if any(k in tl for k in aciklama_keywords):
                    if ts not in sonuc["aciklamalar"]:
                        sonuc["aciklamalar"].append(ts)
            except:
                pass

    # İlk okuma
    _elementleri_tara()

    # Scroll ile ek içerik (sadece yeni element varsa descendants çağır)
    for _ in range(2):
        eski_boyut = len(parcalar_set)
        pyautogui.scroll(-5)
        time.sleep(0.3)
        _elementleri_tara()
        if len(parcalar_set) == eski_boyut:
            break  # Yeni element yok, scroll gereksiz

    sonuc["tum_metin"] = " ".join(parcalar_set)

    # Geri dön - E-Reçete sayfasında "Geri" caption'lı element
    # (buton, hyperlink, ya da başka ctrl type olabilir - hepsini tara)
    geri_tiklandi = False
    geri_caption_set = ("Geri", "Geri Dön", "Geri dön", "GERİ", "GERİ DÖN")
    try:
        for d in medula.descendants():
            try:
                cap = (d.window_text() or "").strip()
                if cap in geri_caption_set:
                    ctrl = str(d.element_info.control_type)
                    try:
                        d.invoke()
                        geri_tiklandi = True
                        log(f"    [OKU ] E-Reçete → Geri ({ctrl}) invoke edildi", "info")
                        time.sleep(0.8)
                        break
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception:
        pass

    if not geri_tiklandi:
        # Fallback: automation_id ile dene
        try:
            geri = element_bul(medula, "form1:buttonGeriDon")
            if geri:
                geri.invoke()
                time.sleep(0.5)
                geri_tiklandi = True
                log(f"    [OKU ] E-Reçete → form1:buttonGeriDon fallback", "info")
        except Exception:
            pass

    if not geri_tiklandi:
        # Tanı: ekrandaki butonları logla
        try:
            log(f"    [UYARI] E-Reçete'den çıkılamadı — ekrandaki elementler:", "warn")
            sayac_diag = 0
            for d in medula.descendants():
                try:
                    cap = (d.window_text() or "").strip()
                    if cap and len(cap) < 30:
                        ctrl = str(d.element_info.control_type).replace("ControlType.", "")
                        aid = d.element_info.automation_id or ""
                        log(f"      - [{ctrl}] '{cap}' id='{aid}'", "info")
                        sayac_diag += 1
                        if sayac_diag > 30:
                            break
                except Exception:
                    continue
        except Exception:
            pass
        # Son çare: Escape
        try:
            import pyautogui
            pyautogui.press("escape")
        except:
            pass

    return sonuc


def uyari_kodlari_oku(medula):
    """Reçete sayfasındaki uyarı kodlarını oku.
    Uyarı kodları reçete sayfasında ControlType.Text olarak görünüyor:
    '256 - Benign prostat hiperplazisi => XALFU XL 10 MG ...'
    Returns: list of dict [{kod, aciklama, ilac_adi}, ...]
    """
    import re
    kodlar = []
    bulunan = set()  # Tekrar engelleme
    try:
        # Uyarı kodu Text veya DataItem olabilir - ikisini de tara
        for d in medula.descendants():
            try:
                ctrl = d.element_info.control_type
                if ctrl not in ("ControlType.Text", "ControlType.DataItem", "ControlType.Custom"):
                    continue
                txt = d.window_text()
                if not txt or "=>" not in txt:
                    continue
                # Format: "256 - Açıklama => İLAÇ ADI"
                match = re.match(r'(\d+)\s*-\s*(.+?)\s*=>\s*(.+)', txt.strip())
                if match:
                    anahtar = f"{match.group(1)}_{match.group(3).strip()}"
                    if anahtar not in bulunan:
                        bulunan.add(anahtar)
                        kodlar.append({
                            "kod": match.group(1),
                            "aciklama": match.group(2).strip(),
                            "ilac_adi": match.group(3).strip(),
                        })
            except:
                pass
    except:
        pass
    if kodlar:
        log(f"  [OKU ] Uyarı kodları: {len(kodlar)} adet", "info")
        for uk in kodlar:
            log(f"    [OKU ] Uyarı {uk['kod']}: {uk['aciklama']} => {uk['ilac_adi']}", "info")
    return kodlar


def recete_teshisleri_oku(medula):
    """Reçete ana sayfasındaki teşhis/tanı satırlarını oku.
    HTML: form1:tableEx3 tablosu (Tanı Listesi)
      - ICD kodu: form1:tableEx3:{row}:text47
      - Tanı adı: form1:tableEx3:{row}:text54
    UI Automation'da automation_id genellikle boş gelir,
    bu yüzden "Tanı Listesi" Table'ının Name bazlı aranır.

    NOT: Mümkünse recete_tum_bilgi_topla() tercih edin — descendants() tekrar çağırmaz.
    Returns: list of str ("ICD - Tanı Adı" formatında)
    """
    import re
    teshisler = []

    # Yöntem 1: MSHTML automation_id ile (bazen çalışır)
    for row in range(20):
        icd_elem = element_bul(medula, f"form1:tableEx3:{row}:text47")
        tani_elem = element_bul(medula, f"form1:tableEx3:{row}:text54")
        if not icd_elem and not tani_elem:
            break
        icd_txt = ""
        tani_txt = ""
        if icd_elem:
            try:
                icd_txt = (icd_elem.window_text() or "").strip()
            except:
                pass
        if tani_elem:
            try:
                tani_txt = (tani_elem.window_text() or "").strip()
            except:
                pass
        if icd_txt or tani_txt:
            birlesik = f"{icd_txt} - {tani_txt}" if icd_txt and tani_txt else (tani_txt or icd_txt)
            teshisler.append(birlesik)

    if teshisler:
        return teshisler

    # Yöntem 2: "Tanı Listesi" tablosundan pozisyon bazlı okuma
    icd_pattern = re.compile(r'^[A-Z]\d{1,2}(\.\d+)?\s*$')
    tani_tablo_top = None
    tani_tablo_bottom = None

    try:
        for elem in medula.descendants(control_type="Table"):
            try:
                name = (elem.element_info.name or "").strip()
                if name == "Tanı Listesi":
                    r = elem.rectangle()
                    tani_tablo_top = r.top
                    tani_tablo_bottom = r.bottom
                    break
            except:
                pass
    except:
        pass

    if tani_tablo_top is None:
        return teshisler

    # Tablo alanındaki Text elementlerini topla
    tani_items = []
    try:
        for elem in medula.descendants(control_type="Text"):
            try:
                txt = (elem.window_text() or "").strip()
                if not txt or len(txt) < 2:
                    continue
                r = elem.rectangle()
                if r.top < tani_tablo_top or r.top > tani_tablo_bottom:
                    continue
                if txt in ("ICD-10", "Tanı", "Tanı Listesi"):
                    continue
                tani_items.append((txt, r.top, r.left))
            except:
                pass
    except:
        pass

    tani_items.sort(key=lambda item: (item[1], item[2]))
    i = 0
    while i < len(tani_items):
        txt1, y1, x1 = tani_items[i]
        if icd_pattern.match(txt1) and i + 1 < len(tani_items):
            txt2, y2, x2 = tani_items[i + 1]
            if abs(y1 - y2) < 5:
                teshisler.append(f"{txt1.strip()} - {txt2}")
                i += 2
                continue
        teshisler.append(txt1)
        i += 1

    return teshisler


def uyari_kodu_kontrol(uyari_kodlari, recete_teshisleri, rapor_aciklamalari, rapor_tanilari):
    """Uyarı kodlarının reçete/rapor teşhis ve açıklamalarıyla eşleşip eşleşmediğini kontrol et.

    Her uyarı kodu için sırayla şu kaynakları arar:
      1. Reçete teşhisleri (f:tableEx1 INPUT + form1:tableEx3 tanı listesi)
      2. Reçete açıklamaları (Açıklama Listesi)
      3. Rapor tanıları (rapor sayfasından okunan ICD + tanı)
      4. Rapor açıklamaları (rapor sayfasından okunan metinler)

    Özel kurallar:
      - Gabapentin/Pregabalin ilaçlarında "nöropatik ağrı" uyarı kodu →
        reçete/rapor açıklamalarında "nöropatik ağrı" ifadesi aranır

    Returns: list of dict [{kod, aciklama, ilac_adi, durum, eslesen_oran, eslesen_kaynak, eslesen_metin}, ...]
    """
    from recete_kontrol.sut_kontrolleri import _turkce_normalize

    # Gabapentin/Pregabalin ilaç isimleri (nöropatik ağrı özel kuralı)
    GABAPENTIN_PREGABALIN = [
        "GABAPENTIN", "PREGABALIN", "NERUDA", "LYRICA", "NEURONTIN",
        "GABATEVA", "GABANTIN", "GABAGAMMA", "GABABOZAN",
        "PREGABIN", "PREGENTA", "PREGOBIN", "PREGALIN",
    ]

    # Kaynakları ayrı ayrı hazırla
    kaynaklar = [
        ("Reçete teşhis", recete_teshisleri),
        ("Rapor açıklama", rapor_aciklamalari),
        ("Rapor tanı", rapor_tanilari),
    ]

    sonuclar = []
    for uk in uyari_kodlari:
        aciklama_norm = _turkce_normalize(uk["aciklama"])
        ilac_adi_upper = (uk.get("ilac_adi", "") or "").upper()

        # ── ÖZEL KURAL 1: Gabapentin/Pregabalin + "nöropatik ağrı" uyarı kodu ──
        noropatik_agri_kural = False
        if "noropatik" in aciklama_norm:
            if any(g in ilac_adi_upper for g in GABAPENTIN_PREGABALIN):
                noropatik_agri_kural = True

        # ── ÖZEL KURAL 2: Oksibutin intoleransı (Solifenacin/Fesoterodin) ──
        # Uyarı kodu 275 vb. "Oral Oksibutinine yanıt alınamayan ya da tolere edemeyen"
        # Bu bir hekim beyanıdır - Medula uyarı kodunu kabul ettiyse doktor onaylamış demektir
        OKSIBUTIN_ILACLAR = [
            "SOLIFENASIN", "FESOTERODIN", "KINZY", "VESICARE", "TOVIAZ",
            "SOLIDAR", "SOLIFENAX", "UROTOL",
        ]
        oksibutin_kural = False
        if "oksibutin" in aciklama_norm or "oksibutinin" in aciklama_norm:
            if any(g in ilac_adi_upper for g in OKSIBUTIN_ILACLAR):
                oksibutin_kural = True

        if oksibutin_kural:
            sonuclar.append({
                **uk, "durum": "UYGUN", "eslesen_oran": 1.0,
                "eslesen_kaynak": "Uyarı kodu (hekim beyanı)",
                "eslesen_metin": uk["aciklama"][:80],
                "_ozel_kural": "Oksibutin intoleransı: uyarı kodu girilmiş = hekim beyanı kabul edildi"
            })
            continue

        # ── ÖZEL KURAL 3: "Antidepresan tedavi" uyarı kodu ──
        # Antidepresan ilaçlarda uyarı kodu "antidepresan tedavi" ise
        # reçete teşhisi, rapor teşhisi veya açıklamalarında ilişkili kelimeler aranır
        antidepresan_kural = False
        if "antidepresan" in aciklama_norm:
            antidepresan_kural = True

        if antidepresan_kural:
            # Antidepresan tedavi kapsamındaki tanı/endikasyon kelimeleri
            antidepresan_ifadeler = [
                # Depresyon
                "depresyon", "depresif", "depresiv", "depression", "majordepresif",
                "major depresif", "major depresyon",
                # Anksiyete
                "anksiyete", "anksiete", "anxiety", "anksiyoz",
                "yaygin anksiyete", "sosyal anksiyete", "panik",
                # Duygudurum
                "duygudurum", "bipolar", "mani", "manik",
                # OKB
                "obsesif", "obsesif kompulsif", "okb", "ocd",
                # TSSB
                "travma sonrasi stres", "tssb", "ptsd",
                # Diğer psikiyatrik
                "kaygı", "kaygi", "korku", "fobi", "fobia", "agorafobi",
                "bulimia", "anoreksiya", "yeme bozuklugu",
                "somatoform", "somatizasyon",
                "distimi", "distimik", "siklotemik",
                "uyum bozuklugu", "uyku bozuklugu", "insomnia",
                # ICD kodları (F30-F49 arası psikiyatri)
                "f30", "f31", "f32", "f33", "f34", "f38", "f39",
                "f40", "f41", "f42", "f43", "f44", "f45", "f48",
                # Genel
                "psikiyatri", "ruhsal", "mental",
            ]

            en_iyi_kaynak = ""
            en_iyi_metin = ""
            bulundu = False

            for kaynak_adi, kaynak_metinler in kaynaklar:
                if not kaynak_metinler:
                    continue
                kaynak_birlesik = _turkce_normalize(" ".join(kaynak_metinler))
                for ifade in antidepresan_ifadeler:
                    if ifade in kaynak_birlesik:
                        en_iyi_kaynak = kaynak_adi
                        for km in kaynak_metinler:
                            if ifade in _turkce_normalize(km):
                                en_iyi_metin = km[:80]
                                break
                        bulundu = True
                        break
                if bulundu:
                    break

            if bulundu:
                sonuclar.append({
                    **uk, "durum": "UYGUN", "eslesen_oran": 1.0,
                    "eslesen_kaynak": en_iyi_kaynak, "eslesen_metin": en_iyi_metin
                })
            else:
                sonuclar.append({
                    **uk, "durum": "UYGUNSUZ", "eslesen_oran": 0,
                    "eslesen_kaynak": "", "eslesen_metin": "",
                    "_ozel_kural": "Antidepresan: ilişkili tanı/endikasyon ifadesi bulunamadı"
                })
            continue

        if noropatik_agri_kural:
            # Reçete/rapor AÇIKLAMALARINDA "nöropatik ağrı" ifadesini ara
            # NOT: Teşhiste değil, açıklamalarda aranır!
            aranan_alternatifler = ["noropatik agri", "neuropathic pain", "noropatik ag"]

            # Sadece açıklama kaynakları (teşhis hariç)
            aciklama_kaynaklari = [
                ("Reçete açıklama", rapor_aciklamalari),  # rapor_aciklamalari = reçete + rapor açıklamaları birleşik
                ("Reçete teşhis", recete_teshisleri),     # teşhiste de aranabilir ama son sırada
                ("Rapor tanı", rapor_tanilari),
            ]

            en_iyi_kaynak = ""
            en_iyi_metin = ""
            bulundu = False

            for kaynak_adi, kaynak_metinler in aciklama_kaynaklari:
                if not kaynak_metinler:
                    continue
                kaynak_birlesik = _turkce_normalize(" ".join(kaynak_metinler))
                for ifade in aranan_alternatifler:
                    if ifade in kaynak_birlesik:
                        en_iyi_kaynak = kaynak_adi
                        for km in kaynak_metinler:
                            if ifade in _turkce_normalize(km):
                                en_iyi_metin = km[:80]
                                break
                        bulundu = True
                        break
                if bulundu:
                    break

            if bulundu:
                sonuclar.append({
                    **uk, "durum": "UYGUN", "eslesen_oran": 1.0,
                    "eslesen_kaynak": en_iyi_kaynak, "eslesen_metin": en_iyi_metin
                })
            else:
                sonuclar.append({
                    **uk, "durum": "UYGUNSUZ", "eslesen_oran": 0,
                    "eslesen_kaynak": "", "eslesen_metin": "",
                    "_ozel_kural": "Gabapentin/Pregabalin: 'nöropatik ağrı' ifadesi reçete/rapor açıklamalarında bulunamadı"
                })
            continue

        # ── GENEL KURAL: Kelime bazlı eşleşme ──
        # Anahtar kelimeler: 3 harften uzun, anlamlı kelimeler
        kelimeler = [k for k in aciklama_norm.split() if len(k) > 3]

        en_iyi_oran = 0
        en_iyi_kaynak = ""
        en_iyi_metin = ""

        # Her kaynağı ayrı ayrı kontrol et
        for kaynak_adi, kaynak_metinler in kaynaklar:
            if not kaynak_metinler:
                continue
            kaynak_birlesik = _turkce_normalize(" ".join(kaynak_metinler))

            eslesme = 0
            eslesen_kelimeler = []
            for kelime in kelimeler:
                # Tam kelime eşleşmesi
                if kelime in kaynak_birlesik:
                    eslesme += 1
                    eslesen_kelimeler.append(kelime)
                # Kısmi kök eşleşmesi: kelimenin ilk 5+ harfi kaynakta geçiyor mu
                elif len(kelime) >= 5:
                    kok = kelime[:min(len(kelime)-2, 7)]  # İlk 5-7 harf (suffix'siz kök)
                    if len(kok) >= 4 and kok in kaynak_birlesik:
                        eslesme += 0.7  # Kısmi eşleşme katkısı
                        eslesen_kelimeler.append(f"{kelime}~{kok}")

            if kelimeler:
                oran = eslesme / len(kelimeler)
            else:
                oran = 0

            if oran > en_iyi_oran:
                en_iyi_oran = oran
                en_iyi_kaynak = kaynak_adi
                # Eşleşen metni bul (ilk eşleşen kelimeyi kaynak metninde ara)
                if eslesen_kelimeler:
                    for km in kaynak_metinler:
                        km_norm = _turkce_normalize(km)
                        aranan = eslesen_kelimeler[0].split("~")[0]  # Kök kısmını at
                        if aranan in km_norm or (len(aranan) >= 5 and aranan[:5] in km_norm):
                            en_iyi_metin = km[:80]
                            break

        if en_iyi_oran >= 0.3:  # %30 eşleşme yeterli (kısmi kök eşleşmeleri dahil)
            sonuclar.append({
                **uk, "durum": "UYGUN", "eslesen_oran": en_iyi_oran,
                "eslesen_kaynak": en_iyi_kaynak, "eslesen_metin": en_iyi_metin
            })
        else:
            sonuclar.append({
                **uk, "durum": "UYGUNSUZ", "eslesen_oran": en_iyi_oran,
                "eslesen_kaynak": en_iyi_kaynak, "eslesen_metin": en_iyi_metin
            })

    return sonuclar


def rapor_tum_metinleri_oku(medula):
    """Rapor sayfasındaki tüm açıklama, tanı, etkin madde, doz metinlerini topla.
    AutomationID yok - DataItem/Text/Custom elementlerinden pozisyon bazlı okur.
    Returns: dict {aciklamalar, tanilar, etkin_maddeler, dozlar, doktor_bransi, tum_metin}
    """
    import re
    sonuc = {
        "aciklamalar": [],
        "tanilar": [],
        "icd_kodlari": [],
        "etkin_maddeler": [],
        "dozlar": [],
        "doktor_bransi": "",
        "tum_metin": "",
    }

    tum_parcalar = []
    for d in medula.descendants():
        try:
            txt = d.window_text()
            if not txt or len(txt.strip()) < 3:
                continue
            r = d.rectangle()
            ctrl = d.element_info.control_type
            txt_s = txt.strip()

            # Sadece sayfa içeriği (y > 400, duyurular hariç)
            if r.top < 400:
                continue
            # Menü/duyuru filtreleme
            if txt_s.startswith("::") or txt_s.startswith("Tarih :") or "Konu :" in txt_s:
                continue

            tum_parcalar.append(txt_s)

            # Sınıflandır
            txt_upper = txt_s.upper()
            txt_lower = txt_s.lower()

            # Açıklama (uzun metin, SUT ibareleri)
            if len(txt_s) > 30 and any(k in txt_lower for k in
                    ["metformin", "sülfonil", "glisemik", "monoterapi", "kontrol",
                     "varfarin", "inr", "stent", "anjiografi", "ldl", "trigliserid",
                     "beta blok", "ejeksiyon", "dispne", "atak", "tedavi",
                     "nöropatik", "noropatik", "neuropat", "gabapentin", "pregabalin",
                     "epilepsi", "fibromiyalji", "ağrı", "agri"]):
                sonuc["aciklamalar"].append(txt_s)

            # Tanı
            if "DİYABETES" in txt_upper or "DIABETES" in txt_upper or "MELLİTÜS" in txt_upper:
                sonuc["tanilar"].append(txt_s)
            if "HİPERTANSİYON" in txt_upper or "HYPERTENS" in txt_upper:
                sonuc["tanilar"].append(txt_s)
            if "ASTIM" in txt_upper or "KOAH" in txt_upper or "COPD" in txt_upper:
                sonuc["tanilar"].append(txt_s)

            # ICD kodu (E11.9, I10, J44 vb.)
            if re.match(r'^[A-Z]\d{2}(\.\d+)?$', txt_s):
                sonuc["icd_kodlari"].append(txt_s)

            # Rapor kodu + tanı satırı
            if re.match(r'^\d{2}\.\d{2}', txt_s):
                sonuc["tanilar"].append(txt_s)

            # Etkin madde (SGK kodu ile birlikte veya tek)
            if txt_s.startswith("SGK"):
                sonuc["etkin_maddeler"].append(txt_s)
            if "HCL" in txt_upper or "SODYUM" in txt_upper or "KALSIYUM" in txt_upper:
                if len(txt_s) < 60 and not any(c.isdigit() for c in txt_s[:3]):
                    sonuc["etkin_maddeler"].append(txt_s)

            # Rapor dozu
            if re.search(r'G.nde|Haftada|Ayda', txt_s) and 'x' in txt_s:
                sonuc["dozlar"].append(txt_s)

            # Doktor branşı
            if "Hastalıkları" in txt_s or "Kardiyoloji" in txt_s or "Nöroloji" in txt_s \
               or "Psikiyatri" in txt_s or "Göğüs" in txt_s or "Endokrin" in txt_s:
                sonuc["doktor_bransi"] = txt_s

        except:
            pass

    sonuc["tum_metin"] = " ".join(tum_parcalar)
    return sonuc


def rapor_ac_oku_geri_don(medula, satir_idx):
    """Tam akış: checkbox seç → rapor aç → tüm metinleri oku + BotanikEOS doz penceresi → geri dön.
    Returns: dict (rapor_tum_metinleri_oku sonucu + eos_dozlar) veya None
    """
    if not checkbox_sec(medula, satir_idx):
        log(f"    Checkbox seçilemedi (satır {satir_idx})", "warn")
        return None

    if not rapor_ac(medula, satir_idx=satir_idx):
        return None

    # 1. Medula rapor sayfasından metinleri oku
    sonuc = rapor_tum_metinleri_oku(medula)

    # 2. BotanikEOS'un açtığı küçük doz penceresinden oku
    time.sleep(0.3)  # Pencere açılması için kısa bekleme
    eos_dozlar = botanik_eos_doz_penceresi_oku()
    if eos_dozlar:
        sonuc["eos_dozlar"] = eos_dozlar
        for ed in eos_dozlar:
            etkin_v = ed.get('etkin', '') or ''
            doz_v = ed.get('doz', '') or ''
            rk_v = ed.get('rapor_kodu', '') or ''
            if etkin_v.startswith('Etkin satır') and doz_v.startswith('Doz satır') and rk_v.startswith('Rapor Kodu satır'):
                continue
            log(f"    [OKU ] EOS Doz: {etkin_v[:20]} | Doz: {doz_v} | R:{rk_v}", "info")

    if sonuc["aciklamalar"]:
        log(f"    Rapor açıklama: {sonuc['aciklamalar'][0]}", "info")
    if sonuc["tanilar"]:
        log(f"    [OKU ] Rapor tanı: {sonuc['tanilar'][0]}", "info")
    if sonuc["dozlar"]:
        if len(sonuc["dozlar"]) == 1:
            log(f"    [OKU ] Rapor doz: {sonuc['dozlar'][0]}", "info")
        else:
            # Çoklu ilaçlı raporda [0] yanıltıcı — hepsini listele
            log(f"    [OKU ] Rapor dozları ({len(sonuc['dozlar'])}): {' | '.join(sonuc['dozlar'])}", "info")

    rapor_geri_don(medula)
    return sonuc


def rapor_geri_don(medula):
    """Rapor sayfasından reçete sayfasına geri dön"""
    global _aid_cache_global
    # Sadece rapor sayfası elementlerini temizle (form1: prefix), reçete elementlerini koru
    _aid_cache_global = {k: v for k, v in _aid_cache_global.items() if not k.startswith("form1:")}
    if element_bul(medula, "form1:buttonGeriDon"):
        element_tikla(medula, "form1:buttonGeriDon")
    elif element_bul(medula, "f:buttonGeriDon"):
        return True
    # Optimize: exists() ile aktif bekleme (eski: sleep(1) + 8×sleep(0.5) = 5sn)
    try:
        cw = medula.child_window(auto_id="f:tbl1", found_index=0)
        if cw.exists(timeout=3):
            return True
    except:
        pass
    # Fallback
    if element_bul(medula, "f:buttonSonraki"):
        return True
    return False


# ========== İLAÇ GEÇMİŞİ ==========
def ilac_gecmisi_oku(medula):
    """İlaç Geçmişi sayfasını aç, ilaç listesini oku, geri dön.
    Buton: f:buttonIlacListesi
    Tablo: form1:tableExKisiIlacList
    Geri: form1:buttonGeriDon (Geri Dön butonu)

    Returns: list of dict [{ilac_adi, recete_tar, adet, kullanim}, ...] veya []
    """
    # İlaç Geçmişi butonuna tıkla
    if not element_tikla(medula, "f:buttonIlacListesi"):
        log("    İlaç Geçmişi butonu bulunamadı", "warn")
        return []

    # Sayfa yüklenene kadar bekle
    for _ in range(8):
        time.sleep(0.5)
        if element_bul(medula, "form1:buttonGeriDon"):
            break
    time.sleep(0.5)

    # Tüm ilaç satırlarını oku (DataItem'lardan)
    ilaclar = []
    try:
        items = []
        for elem in medula.descendants(control_type="Text"):
            try:
                txt = (elem.window_text() or "").strip()
                r = elem.rectangle()
                if txt and len(txt) > 3 and r.top > 250:
                    items.append((txt, r.top, r.left))
            except:
                pass

        items.sort(key=lambda x: (x[1], x[2]))

        # Satırları grupla (aynı y pozisyonundakiler bir satır)
        current_row = []
        current_y = -1
        rows = []
        for txt, y, x in items:
            if current_y < 0 or abs(y - current_y) < 8:
                current_row.append(txt)
                current_y = y
            else:
                if current_row:
                    rows.append(current_row)
                current_row = [txt]
                current_y = y
        if current_row:
            rows.append(current_row)

        # İlaç adlarını çıkar (ilaç adı genellikle uzun ve MG/TABLET/KAPSUL içerir)
        for row_data in rows:
            for cell in row_data:
                cell_upper = cell.upper()
                if len(cell) > 15 and any(k in cell_upper for k in
                        ['MG', 'ML', 'TABLET', 'KAPSUL', 'KAPSÜL', 'TB', 'FTB',
                         'AMPUL', 'FLAKON', 'INH', 'ŞURUP', 'SURUP', 'DAMLA',
                         'KREM', 'JEL', 'KP.', 'NEBUL', 'SPRAY']):
                    ilaclar.append({"ilac_adi": cell})
                    break
    except Exception as e:
        log(f"    İlaç geçmişi okuma hatası: {e}", "error")

    # Geri dön
    rapor_geri_don(medula)

    if ilaclar:
        log(f"    [OKU ] İlaç geçmişi: {len(ilaclar)} ilaç bulundu", "info")
    return ilaclar


def ilac_gecmisinde_laba_ics_var_mi(medula):
    """İlaç geçmişinde LABA+ICS kombinasyonu kullanılmış mı kontrol et.
    Üçlü kombinasyon (LABA+ICS+LAMA) SUT koşulu: "en az 3 ay ICS+LABA ile tedavi edilmiş"

    Returns: (bool, str) — (bulundu_mu, açıklama)
    """
    gecmis = ilac_gecmisi_oku(medula)
    if not gecmis:
        return False, "İlaç geçmişi okunamadı"

    # LABA+ICS kombinasyon ilaçları
    laba_ics_ilaclar = ['SERETIDE', 'SYMBICORT', 'FOSTER', 'RELVAR', 'DUORESP',
                        'BUFOMIX', 'FOKUSAL', 'AIRFLUSAL', 'MIFLONIDE COMBI',
                        'SALMETEROL', 'FORMOTEROL', 'VILANTEROL']
    # ICS tek başına
    ics_ilaclar = ['PULMICORT', 'FLIXOTIDE', 'ALVESCO', 'MIFLONIDE', 'BUDESONID',
                   'FLUTIKAZON', 'BEKLOMETAZON', 'SIKLESONID']

    bulunan_laba_ics = []
    bulunan_ics = []

    for ilac in gecmis:
        adi = (ilac.get("ilac_adi") or "").upper()
        if any(k in adi for k in laba_ics_ilaclar):
            bulunan_laba_ics.append(adi[:40])
        elif any(k in adi for k in ics_ilaclar):
            bulunan_ics.append(adi[:40])

    if bulunan_laba_ics:
        return True, f"ICS+LABA geçmişi var: {', '.join(bulunan_laba_ics[:2])}"
    elif bulunan_ics:
        return True, f"ICS geçmişi var: {', '.join(bulunan_ics[:2])}"

    return False, "İlaç geçmişinde ICS+LABA kullanımı bulunamadı"


def rapor_ac_doz_oku_geri_don(medula, satir_idx, etkin_madde):
    """Tam akış: checkbox seç → rapor aç → doz oku → geri dön
    Returns: (rapor_doz_metni, rapor_gunluk_doz) veya (None, None)
    """
    if not checkbox_sec(medula, satir_idx):
        log(f"    Checkbox seçilemedi (satır {satir_idx})", "warn")
        return None, None

    if not rapor_ac(medula, satir_idx=satir_idx):
        return None, None

    doz_metni, gunluk_doz = rapor_tedavi_semasi_doz_oku(medula, etkin_madde)
    rapor_geri_don(medula)
    return doz_metni, gunluk_doz


# ========== DOZ KARŞILAŞTIRMA ==========
def doz_karsilastir(recete_gunluk, rapor_gunluk, recete_doz_dict=None, rapor_doz_dict=None):
    """Reçete dozu rapor dozunu geçiyor mu?
    recete_doz_dict/rapor_doz_dict: periyot_birim bilgisi için (opsiyonel)
    Returns: (uygun: bool, aciklama: str)
    """
    if recete_gunluk is None or rapor_gunluk is None:
        return True, "Doz bilgisi okunamadı, kontrol edilemedi"

    # Periyot birim uyumsuzluğu kontrolü:
    # Rapor "Ayda" reçete "Günde" ise büyük fark olur - muhtemelen SELECT okunamadı
    if recete_doz_dict and rapor_doz_dict:
        r_birim = recete_doz_dict.get("periyot_birim", "Günde")
        rp_birim = rapor_doz_dict.get("periyot_birim", "Günde")
        if r_birim == "Günde" and rp_birim in ("Ayda", "Haftada", "Yılda"):
            # Reçete birimi muhtemelen yanlış okundu (varsayılan Günde)
            # Rapor birimini kullanarak yeniden hesapla
            doz = recete_doz_dict.get("doz", 1.0)
            carpan = recete_doz_dict.get("carpan", 1.0)
            periyot_sayi = recete_doz_dict.get("periyot_sayi", 1.0)
            recete_gunluk = round(_gunluk_doz_hesapla(doz, carpan, periyot_sayi, rp_birim), 2)
            log(f"    [DOZ ] Periyot düzeltme: reçete 'Günde' → '{rp_birim}' (rapor referans)", "info")

    # Birim bilgisi (log için)
    r_metin = recete_doz_dict.get("metin", "") if recete_doz_dict else ""
    rp_metin = rapor_doz_dict.get("metin", "") if rapor_doz_dict else ""

    if recete_gunluk <= rapor_gunluk:
        return True, f"Doz uygun (reçete: {recete_gunluk} ≤ rapor: {rapor_gunluk}) [{r_metin} vs {rp_metin}]"
    # Reçete rapor dozundan fazla — fakat SUT'ta rapor dozu katı bir sınır değil;
    # hekim belirli aralıklarda arttırabilir. 2× eşiğine kadar UYARI, 2× üstü UYGUNSUZ.
    oran = recete_gunluk / rapor_gunluk if rapor_gunluk > 0 else 99
    if oran <= 2.0 + 1e-6:
        return True, f"⚠ Doz rapor üstü — manuel kontrol (reçete: {recete_gunluk} > rapor: {rapor_gunluk}, {oran:.2f}×) [{r_metin} vs {rp_metin}]"
    else:
        return False, f"Doz aşımı! (reçete: {recete_gunluk} > rapor: {rapor_gunluk}, {oran:.2f}×) [{r_metin} vs {rp_metin}]"


# ========== SUT MESAJ KONTROL (sut_kontrolleri entegrasyonu) ==========
def sut_mesaj_kontrol(ilac, mesaj_basliklar, mesaj_metni, recete_teshisleri=None,
                      doktor_uzmanligi="", recete_alt_turu="Ayaktan"):
    """Mesaj metni üzerinden SUT kontrolü yap.
    sut_kontrolleri modülünü kullanır.
    Returns: (sonuc: str, aciklama: str)  - "Uygun"/"UygunDegil"/"KontrolEdilemedi"
    """
    try:
        from recete_kontrol.sut_kontrolleri import sut_kontrol_yap

        ilac_sonuc = {
            "ilac_adi": ilac.get("ilac_adi", ""),
            "etkin_madde": ilac.get("etkin_madde", ""),
            "sgk_kodu": ilac.get("sgk_kodu", ""),
            "rapor_kodu": ilac.get("rapor_kodu", ""),
            "sut_maddesi": ilac.get("sut_maddesi", ""),
            "mesaj_metni": mesaj_metni,
            "mesaj_basliklar": mesaj_basliklar,
            "doktor_uzmanligi": doktor_uzmanligi,
            "recete_alt_turu": recete_alt_turu,
            "recete_teshisleri": recete_teshisleri or [],
            # Reçete açıklamaları — ESA Hb/Ferritin/TSAT vb. lab değerleri burada olabilir
            "recete_aciklamalari": ilac.get("_recete_aciklamalari", []) or [],
            # Rapor verisi — _esa_detayli_kontrol bu metinde de tarar
            "rapor_aciklamalari": ilac.get("_rapor_aciklamalari", []) or [],
            "rapor_tani_bilgileri": ilac.get("_rapor_tani_bilgileri", []) or [],
            # Reçete uyarı kodları (ESA için 217 kodu burada aranır)
            "_uyari_kodlari": ilac.get("_uyari_kodlari", []) or [],
            # Reçete dozu (enteral beslenme kalori/gün hesabı için gerekli)
            "recete_doz": ilac.get("recete_doz"),
            # E-Reçete Görüntüle ekranından okunan açıklama metni (217 kodu varsa
            # eagerly toplanır — TSAT/Ferritin/kuru kilo/doz bilgileri burada olur)
            "_erecete_aciklama_metni": ilac.get("_erecete_aciklama_metni", "") or "",
            "_erecete_aciklama_listesi": ilac.get("_erecete_aciklama_listesi", []) or [],
        }

        sonuc = sut_kontrol_yap(ilac_sonuc)
        if sonuc is None:
            return "KontrolEdilemedi", f"SUT kategorisi tespit edilemedi (mesaj: {mesaj_basliklar[:2]})"

        rapor = sonuc["kontrol_raporu"]
        kategori_adi = sonuc["kategori_adi"]

        # ═══ SUT KURAL EŞLEŞMESİ ÇERÇEVE LOG ═══
        ilac_adi_kisa = (ilac.get("ilac_adi", "") or "")[:40]
        sonuc_simge = "✓" if rapor.sonuc.value == "uygun" else ("✗" if rapor.sonuc.value == "uygun_degil" else "?")
        log(f"    ┌─── SUT Kontrol: {ilac_adi_kisa} ───", "info")
        if rapor.sut_kurali:
            log(f"    │ Kural : {rapor.sut_kurali}", "info")
        if rapor.aranan_ibare:
            log(f"    │ Aranan: {rapor.aranan_ibare}", "info")
        if rapor.bulunan_metin:
            log(f"    │ Buluna: {rapor.bulunan_metin[:80]}", "info")
        elif rapor.sonuc.value != "uygun":
            log(f"    │ Buluna: (metinde eşleşme bulunamadı)", "warn")
        log(f"    │ Sonuç : {sonuc_simge} {rapor.mesaj}", "ok" if rapor.sonuc.value == "uygun" else ("sorun" if rapor.sonuc.value == "uygun_degil" else "warn"))
        log(f"    └───────────────────────────────────", "info")

        if rapor.sonuc.value == "uygun":
            return "Uygun", f"SUT {kategori_adi}: {rapor.mesaj}"
        elif rapor.sonuc.value == "uygun_degil":
            return "UygunDegil", f"SUT {kategori_adi}: {rapor.mesaj}"
        else:
            return "KontrolEdilemedi", f"SUT {kategori_adi}: {rapor.mesaj}"

    except Exception as e:
        return "KontrolEdilemedi", f"SUT kontrol hatası: {e}"


# ========== İLAÇ KARAR AĞACI ==========
def ilac_detayli_kontrol(medula, cur, conn, grup, recete_no, recete_turu,
                          ilac, satir_idx, renkli_sonuc, recete_teshisleri=None,
                          doktor_uzmanligi="", recete_alt_turu="Ayaktan",
                          rapor_cache=None):
    """Tek ilaç için 4 durumlu karar ağacı (AA/BB/CC/DD).
    rapor_cache: aynı reçetedeki ilaçlar için rapor_kodu→rapor_verisi cache (OPT A).
    Returns: dict (rapor satırı)
    """
    if rapor_cache is None:
        rapor_cache = {}
    rapor_var = bool(ilac.get("rapor_kodu", ""))
    msj_var = (ilac.get("msj", "") == "var")
    etkin = ilac.get("etkin_madde", "")
    ilac_adi = ilac.get("ilac_adi", "")[:45]

    # Satır indeksi: recete_tum_bilgi_topla'da medula_satir_haritasi ile zaten eşleştirildi
    # medula_satir_idx kullan (ilac dict'e yazılmış), ek doğrulamaya gerek yok
    if ilac.get("medula_satir_idx") is not None and ilac["medula_satir_idx"] != satir_idx:
        log(f"    [OKU ] Satır düzeltme: idx {satir_idx} → {ilac['medula_satir_idx']} ({ilac_adi})", "info")
        satir_idx = ilac["medula_satir_idx"]

    # Mevcut DB kuralını da kontrol et
    kural = db_kural_bul(cur, etkin)

    # === BYPASS KONTROLÜ: İlaç kontrol ayarlarını sorgula ===
    try:
        from kontrol_kurallari import get_kontrol_ayarlari, get_ogrenilen_ilaclar
        ayarlar = get_kontrol_ayarlari()
        ogrenilen = get_ogrenilen_ilaclar()

        # İlacı öğrenilen veritabanına kaydet
        ogrenilen.ilac_kaydet(
            ilac.get("ilac_adi", ""), etkin_madde=etkin,
            sgk_kodu=ilac.get("sgk_kodu"), rapor_kodu=ilac.get("rapor_kodu")
        )

        # Öğrenilen ilaçtan ATC/farmakolojik grup bilgisi al
        ilac_kayit = ogrenilen.ilac_bul(ilac.get("ilac_adi", ""))
        atc_grup = ilac_kayit.get("atc_grup") if ilac_kayit else None
        farm_grup = ilac_kayit.get("farmakolojik_grup") if ilac_kayit else None

        bypass_sut = not ayarlar.kontrol_aktif_mi(
            "sut", ilac_adi=ilac.get("ilac_adi", ""), etkin_madde=etkin,
            atc_grup=atc_grup, farmakolojik_grup=farm_grup)
        bypass_doz = not ayarlar.kontrol_aktif_mi(
            "doz", ilac_adi=ilac.get("ilac_adi", ""), etkin_madde=etkin,
            atc_grup=atc_grup, farmakolojik_grup=farm_grup)
        bypass_mesaj = not ayarlar.kontrol_aktif_mi(
            "ilac_mesaji", ilac_adi=ilac.get("ilac_adi", ""), etkin_madde=etkin,
            atc_grup=atc_grup, farmakolojik_grup=farm_grup)

        # İlaç mesajı bypass ise msj_var'ı False yap (mesaj yokmuş gibi davran)
        if bypass_mesaj and msj_var:
            msj_var = False
            log(f"    [BYPASS] {ilac_adi} → İlaç mesajı kontrolü pasif (ayar)", "info")
    except Exception as e:
        bypass_sut = False
        bypass_doz = False
        bypass_mesaj = False
        atc_grup = None
        farm_grup = None
        logger.debug(f"Kontrol ayarları yüklenemedi: {e}")

    # Varsayılan sonuçlar
    kontrol_tipi = ""
    doz_k = "-"
    sut_k = "-"
    genel = ""
    aciklama = ""
    recete_dozu_str = ""
    rapor_dozu_str = ""
    mesaj_metni = ""
    aciklama_parcalari = []
    rapor_verisi = None
    erecete = None

    # ═══ DURUM AA: Raporsuz + Mesajsız → GEÇ ═══
    if not rapor_var and not msj_var:
        kontrol_tipi = "AA"
        genel = "GECİLDİ"
        aciklama = "Raporsuz, mesajsız - kontrol gerekmez"
        log(f"    [GEÇ ] {ilac_adi} → Raporsuz, mesajsız", "info")

    # ═══ DURUM BB: Raporsuz + Mesaj VAR → Reçete/E-Reçete açıklamaları oku + SUT kontrol ═══
    elif not rapor_var and msj_var:
        kontrol_tipi = "BB"
        log(f"    [OKU ] {ilac_adi} → Raporsuz ama mesaj VAR", "warn")

        # SUT bypass kontrolü
        if bypass_sut:
            sut_k = "Bypass"
            aciklama_parcalari.append("SUT kontrolü pasif (ayar)")
            genel = "GECİLDİ"
            log(f"    [BYPASS] {ilac_adi} → SUT kontrolü pasif (ayar)", "info")
        else:
            # Teşhis + açıklama bilgileri zaten recete_tum_bilgi_topla()'da toplandı
            # E-Reçete sayfasına gitmeye GEREK YOK → ~7-8 saniye tasarruf/ilaç
            recete_teshisleri = recete_teshisleri or []
            recete_metin = " ".join(recete_teshisleri)
            if recete_teshisleri:
                log(f"    [OKU ] Reçete teşhis: {recete_teshisleri[0]}", "info")

            # Reçete açıklamaları (Açıklama Listesi — ana sayfadan okundu)
            recete_aciklamalari = ilac.get("_recete_aciklamalari", [])
            if recete_aciklamalari:
                recete_metin += " " + " ".join(recete_aciklamalari)
                log(f"    [OKU ] Reçete açıklama: {recete_aciklamalari[0][:60]}", "info")

            sut_sonuc, sut_aciklama = sut_mesaj_kontrol(ilac, [], recete_metin,
                                                         recete_teshisleri=recete_teshisleri,
                                                         doktor_uzmanligi=doktor_uzmanligi,
                                                         recete_alt_turu=recete_alt_turu)
            sut_k = sut_sonuc
            aciklama_parcalari.append(sut_aciklama)

            if sut_sonuc == "Uygun":
                genel = "UYGUN"
                log(f"    [SUT ] {sut_aciklama}", "info")
                log(f"    [  OK ] SUT uygun", "ok")
            elif sut_sonuc == "UygunDegil":
                genel = "UYGUNSUZ"
                log(f"    [SUT ] {sut_aciklama}", "info")
                log(f"    [SORUN] SUT uygunsuz", "sorun")
            else:
                log(f"    [SUT ] {sut_aciklama}", "warn")
                if "kategorisi tespit edilemedi" in sut_aciklama:
                    # BB durumu: Raporsuz + mesaj var — senkron AI araştırması
                    log(f"    [AI] Bilinmeyen ilaç → senkron araştırma başlıyor...", "warn")
                    ai_sonuc = _ai_sut_implementasyon_iste(
                        ilac_adi, ilac.get("etkin_madde", ""),
                        rapor_kodu=ilac.get("rapor_kodu", ""),
                        senkron=True
                    )
                    if ai_sonuc and ai_sonuc.get("kategori"):
                        log(f"    [AI] Yeni kategori uygulanıyor: {ai_sonuc['kategori']} (kaynak: {ai_sonuc.get('kaynak')})", "ok")
                        ilac_yeni = dict(ilac)
                        if ai_sonuc.get("etkin_madde"):
                            ilac_yeni["etkin_madde"] = ai_sonuc["etkin_madde"]
                        sut_sonuc2, sut_aciklama2 = sut_mesaj_kontrol(
                            ilac_yeni, [], recete_metin,
                            recete_teshisleri=recete_teshisleri,
                            doktor_uzmanligi=doktor_uzmanligi,
                            recete_alt_turu=recete_alt_turu
                        )
                        if sut_sonuc2 != "KontrolEdilemedi":
                            sut_k = sut_sonuc2
                            aciklama_parcalari = [a for a in aciklama_parcalari
                                                   if "kategorisi tespit edilemedi" not in a]
                            aciklama_parcalari.append(f"[AI] {sut_aciklama2}")
                            if sut_sonuc2 == "Uygun":
                                genel = "UYGUN"
                            elif sut_sonuc2 == "UygunDegil":
                                genel = "UYGUNSUZ"
                            log(f"    [AI] Tekrar SUT kontrol: {sut_sonuc2} — {sut_aciklama2[:80]}", "ok" if sut_sonuc2 == "Uygun" else "warn")
                        else:
                            genel = "KONTROLEDİLEMEDİ"
                    else:
                        genel = "KONTROLEDİLEMEDİ"
                else:
                    genel = "KONTROLEDİLEMEDİ"

    # ═══ DURUM CC: Raporlu + Mesajsız → Doz karşılaştır (toplu sonucu kullan) ═══
    elif rapor_var and not msj_var:
        kontrol_tipi = "CC"

        if bypass_doz:
            doz_k = "Bypass"
            aciklama_parcalari.append("Doz kontrolü pasif (ayar)")
            genel = "GECİLDİ"
            log(f"    [BYPASS] {ilac_adi} → Doz kontrolü pasif (ayar)", "info")
        else:
            # Toplu doz karşılaştırma sonucunu kullan (ilaclari_oku'da hesaplandı)
            doz_sonuc = ilac.get("doz_kontrol", "okunamadi")
            doz_aciklama = ilac.get("doz_aciklama", "")
            eos_rapor_doz = ilac.get("eos_rapor_doz_metin", "")
            recete_doz = ilac.get("recete_doz")

            if recete_doz:
                recete_dozu_str = recete_doz.get("metin", "")
            if eos_rapor_doz:
                rapor_dozu_str = eos_rapor_doz

            if doz_sonuc == "uygun":
                doz_k = "Uygun"
                genel = "UYGUN"
                aciklama_parcalari.append(doz_aciklama)
                log(f"    [  OK ] {doz_aciklama}", "ok")
            elif doz_sonuc == "uygunsuz":
                doz_k = "UygunDegil"
                genel = "UYGUNSUZ"
                aciklama_parcalari.append(doz_aciklama)
                log(f"    [SORUN] {doz_aciklama}", "sorun")
            elif doz_sonuc == "uyari":
                doz_k = "Uyarı"
                genel = "UYARI"
                aciklama_parcalari.append(doz_aciklama)
                log(f"    [UYARI] {doz_aciklama}", "warn")
            else:
                doz_k = "Uygun"
                genel = "UYGUN"
                aciklama_parcalari.append(doz_aciklama or f"Raporlu ({ilac['rapor_kodu']}), doz okunamadı")
                log(f"    [OKU ] {doz_aciklama or 'Doz okunamadı'}", "info")

    # ═══ DURUM DD: Raporlu + Mesaj VAR → Doz karşılaştır (toplu) + SUT kontrol ═══
    elif rapor_var and msj_var:
        kontrol_tipi = "DD"
        log(f"    [OKU ] {ilac_adi} → Raporlu ({ilac['rapor_kodu']}) + Mesaj VAR", "warn")

        # 1. Doz karşılaştırma (toplu sonucu kullan)
        uygun_doz = True
        if bypass_doz:
            doz_k = "Bypass"
            aciklama_parcalari.append("Doz kontrolü pasif (ayar)")
            log(f"    [BYPASS] {ilac_adi} → Doz kontrolü pasif (ayar)", "info")
        else:
            doz_sonuc = ilac.get("doz_kontrol", "okunamadi")
            doz_aciklama = ilac.get("doz_aciklama", "")
            eos_rapor_doz = ilac.get("eos_rapor_doz_metin", "")
            recete_doz = ilac.get("recete_doz")

            if recete_doz:
                recete_dozu_str = recete_doz.get("metin", "")
            if eos_rapor_doz:
                rapor_dozu_str = eos_rapor_doz

            if doz_sonuc == "uygunsuz":
                uygun_doz = False
                doz_k = "UygunDegil"
                aciklama_parcalari.append(doz_aciklama)
                log(f"    [SORUN] Doz: {doz_aciklama}", "sorun")
            elif doz_sonuc == "uyari":
                doz_k = "Uyarı"
                aciklama_parcalari.append(doz_aciklama)
                log(f"    [UYARI] Doz: {doz_aciklama}", "warn")
            elif doz_sonuc == "uygun":
                doz_k = "Uygun"
                aciklama_parcalari.append(doz_aciklama)
                log(f"    [  OK ] Doz: {doz_aciklama}", "ok")
            else:
                doz_k = "Uygun"
                if eos_rapor_doz:
                    log(f"    [OKU ] Rapor doz: {eos_rapor_doz} | Reçete doz: okunamadı", "info")

        # 2. SUT kontrol
        if bypass_sut:
            sut_k = "Bypass"
            aciklama_parcalari.append("SUT kontrolü pasif (ayar)")
            log(f"    [BYPASS] {ilac_adi} → SUT kontrolü pasif (ayar)", "info")
        else:
            # OPT A: Rapor cache — aynı rapor_kodu için rapor 1 kez açılır
            rapor_kodu_cache = ilac.get("rapor_kodu", "")
            if rapor_kodu_cache and rapor_kodu_cache in rapor_cache:
                rapor_verisi = rapor_cache[rapor_kodu_cache]
                rapor_metin = rapor_verisi["tum_metin"] if rapor_verisi else ""
                log(f"    [CACHE] Rapor zaten okundu ({rapor_kodu_cache}) — tekrar açılmıyor", "info")
            else:
                rapor_verisi = rapor_ac_oku_geri_don(medula, satir_idx)
                rapor_metin = rapor_verisi["tum_metin"] if rapor_verisi else ""
                if rapor_verisi:
                    log(f"    [OKU ] Rapor açıklama okundu ({len(rapor_metin)} kr)", "info")
                if rapor_kodu_cache:
                    rapor_cache[rapor_kodu_cache] = rapor_verisi

            sut_sonuc, sut_aciklama = sut_mesaj_kontrol(ilac, [], rapor_metin,
                                                         recete_teshisleri=recete_teshisleri,
                                                         doktor_uzmanligi=doktor_uzmanligi,
                                                         recete_alt_turu=recete_alt_turu)

            # ESA / lab-tabanlı ilaçlar: Hb/Ferritin/TSAT rapor sayfasında değil,
            # e-reçete açıklamalarında olabilir. KONTROL_EDILEMEDI ise e-reçete aç ve tekrar dene.
            _esa_anahtar = ('ESA' in sut_aciklama or 'eritropoietin' in sut_aciklama.lower() or
                             'ferritin' in sut_aciklama.lower() or 'tsat' in sut_aciklama.lower() or
                             'hemoglobin' in sut_aciklama.lower() or 'hb/' in sut_aciklama.lower())
            if sut_sonuc == "KontrolEdilemedi" and _esa_anahtar:
                log(f"    [OKU ] Lab değerleri eksik — E-Reçete açıklamaları okunuyor...", "warn")
                try:
                    erecete_verisi = erecete_aciklama_oku(medula)
                    if erecete_verisi and erecete_verisi.get("tum_metin"):
                        ek_metin = erecete_verisi.get("tum_metin", "")
                        # Açıklamaları + tüm metni rapor metnine ekle ve tekrar kontrol
                        birlesik_metin = rapor_metin + " " + ek_metin
                        log(f"    [OKU ] E-Reçete metni eklendi ({len(ek_metin)} kr) — tekrar kontrol", "info")
                        sut_sonuc2, sut_aciklama2 = sut_mesaj_kontrol(
                            ilac, [], birlesik_metin,
                            recete_teshisleri=recete_teshisleri,
                            doktor_uzmanligi=doktor_uzmanligi,
                            recete_alt_turu=recete_alt_turu
                        )
                        if sut_sonuc2 != "KontrolEdilemedi":
                            sut_sonuc = sut_sonuc2
                            sut_aciklama = sut_aciklama2
                            log(f"    [OKU ] E-Reçete eklendikten sonra: {sut_sonuc2}", "ok" if sut_sonuc2 == "Uygun" else "warn")
                except Exception as _e:
                    log(f"    [UYARI] E-Reçete okuma hatası: {_e}", "warn")

            sut_k = sut_sonuc
            aciklama_parcalari.append(sut_aciklama)

            if sut_sonuc == "Uygun":
                log(f"    [SUT ] {sut_aciklama}", "info")
                log(f"    [  OK ] SUT uygun", "ok")
            elif sut_sonuc == "UygunDegil":
                log(f"    [SUT ] {sut_aciklama}", "info")
                log(f"    [SORUN] SUT uygunsuz", "sorun")
            else:
                log(f"    [SUT ] {sut_aciklama}", "warn")

                # ── Üçlü kombinasyon (LABA+ICS+LAMA) ise ilaç geçmişinden ICS+LABA kontrolü ──
                uclu_ticari = ['TRELEGY', 'TRIMBOW', 'ENERZAIR', 'BREQUAL', 'BREQAL']
                ilac_adi_upper = (ilac.get("ilac_adi") or "").upper()
                if any(t in ilac_adi_upper for t in uclu_ticari) and "ICS+LABA" in sut_aciklama:
                    log(f"    [OKU ] Üçlü kombinasyon — ilaç geçmişinden ICS+LABA kontrolü yapılıyor...", "info")
                    gecmis_var, gecmis_aciklama = ilac_gecmisinde_laba_ics_var_mi(medula)
                    if gecmis_var:
                        sut_k = "Uygun"
                        aciklama_parcalari.append(f"İlaç geçmişi: {gecmis_aciklama}")
                        log(f"    [  OK ] {gecmis_aciklama}", "ok")
                    else:
                        log(f"    [UYARI] {gecmis_aciklama}", "warn")

                if "kategorisi tespit edilemedi" in sut_aciklama or "Rapor/mesaj metni yok" in sut_aciklama:
                    # SENKRON AI: Bu reçete için AI araştırması yap + sonucu bu reçeteye uygula
                    log(f"    [AI] Bilinmeyen ilaç → senkron araştırma başlıyor...", "warn")
                    ai_sonuc = _ai_sut_implementasyon_iste(
                        ilac_adi, etkin, rapor_kodu=ilac.get("rapor_kodu", ""),
                        mesaj_metni=rapor_metin[:300] if rapor_metin else "",
                        senkron=True
                    )
                    if ai_sonuc and ai_sonuc.get("kategori"):
                        log(f"    [AI] Yeni kategori uygulanıyor: {ai_sonuc['kategori']} (kaynak: {ai_sonuc.get('kaynak')})", "ok")
                        # Yeni kategori ile SUT kontrolünü TEKRAR yap
                        ilac_yeni = dict(ilac)
                        if ai_sonuc.get("etkin_madde"):
                            ilac_yeni["etkin_madde"] = ai_sonuc["etkin_madde"]
                        sut_sonuc2, sut_aciklama2 = sut_mesaj_kontrol(
                            ilac_yeni, [], rapor_metin,
                            recete_teshisleri=recete_teshisleri,
                            doktor_uzmanligi=doktor_uzmanligi,
                            recete_alt_turu=recete_alt_turu
                        )
                        if sut_sonuc2 != "KontrolEdilemedi":
                            sut_k = sut_sonuc2
                            # Önceki "tespit edilemedi" mesajını değiştir
                            aciklama_parcalari = [a for a in aciklama_parcalari
                                                   if "kategorisi tespit edilemedi" not in a]
                            aciklama_parcalari.append(f"[AI] {sut_aciklama2}")
                            log(f"    [AI] Tekrar SUT kontrol: {sut_sonuc2} — {sut_aciklama2[:80]}", "ok" if sut_sonuc2 == "Uygun" else "warn")

        # Genel sonuç
        if not uygun_doz or sut_k == "UygunDegil":
            genel = "UYGUNSUZ"
        elif doz_k == "Uyarı":
            # Doz rapor üstü ama 2× altında — şüpheli
            genel = "UYARI"
        elif (sut_k in ("Uygun", "Bypass")) and uygun_doz:
            genel = "UYGUN" if sut_k == "Uygun" else "GECİLDİ"
        else:
            genel = "KONTROLEDİLEMEDİ"

    aciklama = " | ".join(aciklama_parcalari) if aciklama_parcalari else aciklama

    # DB'ye logla
    sonuc_logla(cur, conn, grup, recete_no, recete_turu, ilac, kontrol_tipi,
                renkli_sonuc, "-", doz_k, sut_k, genel, aciklama,
                recete_dozu_str, rapor_dozu_str, mesaj_metni)

    # Rapor/teşhis verisi (uyarı kodu kontrolünde kullanılacak)
    rapor_bilgi = {}
    if kontrol_tipi == "DD" and rapor_verisi:
        rapor_bilgi = rapor_verisi
    elif kontrol_tipi == "BB" and erecete:
        rapor_bilgi = erecete

    # Rapor satırı (JSON/Excel için)
    return {
        "recete_no": recete_no, "recete_turu": recete_turu,
        "ilac_adi": ilac["ilac_adi"], "etkin_madde": ilac.get("etkin_madde", ""),
        "rapor_kodu": ilac.get("rapor_kodu", ""), "msj": ilac.get("msj", ""),
        "renkli_kontrol": "-",
        "rapor_kontrol": doz_k if doz_k != "-" else "-",
        "sut_kontrol": sut_k if sut_k != "-" else "-",
        "sonuc": genel, "aciklama": aciklama,
        "_rapor_verisi": rapor_bilgi,  # iç kullanım: uyarı kodu kontrolü için
    }


# ========== MEDULA NAVİGASYON ==========
def medula_navigasyon(medula, grup, donem_offset):
    """Reçete Listesi → Dönem → Fatura Türü → Sorgula → İlk reçeteye tıkla"""

    # 0. webBrowser1'e focus ver (IE embedded browser)
    try:
        for c in medula.children():
            try:
                if c.element_info.automation_id == "webBrowser1":
                    c.set_focus()
                    time.sleep(0.3)
                    break
            except:
                pass
    except:
        pass

    # 1. Reçete Listesi menüsü (invoke)
    log("Reçete Listesi sayfasına gidiliyor...", "info")
    for deneme in range(3):
        if element_tikla(medula, "form1:menuHtmlCommandExButton31"):
            log("Reçete Listesi açıldı", "ok")
        else:
            # Menü görünmüyor — reçete sayfasında olabiliriz
            # BotanikEOS Giriş butonu ile ana sayfaya dön
            log("Menü görünmüyor, ana sayfaya dönülüyor...", "warn")
            element_tikla(medula, "btnMedulayaGirisYap", "click")
            time.sleep(5)
            # Tekrar dene
            if element_tikla(medula, "form1:menuHtmlCommandExButton31"):
                log("Reçete Listesi açıldı (giriş butonu sonrası)", "ok")
            else:
                # SAFETY NET 1: Menü bulunamadı ama Reçete Listesi sayfası zaten
                # ekrandaysa (Sorgula butonu var), navigasyonu başarılı say.
                # Kullanıcı listeyi manuel açtıysa veya bir önceki grup
                # transition'ı sonrası ekran hala liste sayfasındaysa olur.
                if element_bul(medula, "form1:buttonSonlandirilmamisReceteler"):
                    log("Menü tıklanamadı ama Reçete Listesi sayfası zaten açık — devam ediliyor", "warn")
                    break  # while döngüsünden çık, dönem/fatura/sorgula adımlarına geç
                # SAFETY NET 2: Reçete DETAY sayfasındaysak (f:tbl1 var + Sonraki
                # butonu var), kullanıcı bir reçeteyi açmış demektir; navigasyon
                # gereksiz — bunu acik_receteden moduna eşdeğer görüp başarılı say.
                if (element_bul(medula, "f:tbl1")
                        and element_bul(medula, "f:buttonSonraki")):
                    log("Menü tıklanamadı ama bir reçete detay sayfasında — bu reçeteden devam edilecek", "warn")
                    return True  # Direkt başarılı dön; tara() açık reçeteyi işleyecek

                # TANI: Hangi state'deyiz? Görünür buton ve linkleri logla ki
                # kullanıcı/geliştirici Medula'nın gerçekten hangi sayfada olduğunu görsün.
                log("Reçete Listesi menüsü bulunamadı! Ekrandaki state'i tanılıyorum...", "error")
                try:
                    sayac_diag = 0
                    for d in medula.descendants():
                        try:
                            ctrl = str(d.element_info.control_type or "")
                            if "Button" not in ctrl and "Hyperlink" not in ctrl:
                                continue
                            cap = (d.window_text() or "").strip()
                            aid = d.element_info.automation_id or ""
                            if cap or aid:
                                log(f"    [TANI] {ctrl.replace('ControlType.','')}: '{cap[:40]}' aid='{aid}'", "info")
                                sayac_diag += 1
                                if sayac_diag >= 25:
                                    log("    [TANI] (ilk 25 buton/link gösterildi)", "info")
                                    break
                        except Exception:
                            continue
                    if sayac_diag == 0:
                        log("    [TANI] Hiç buton/link yok — Medula muhtemelen oturum kapalı, login ekranında veya boş.", "error")
                        log("    [TANI] Çözüm: Medula tarayıcısını manuel olarak açın, kullanıcı girişi yapın, sonra TÜMÜNÜ KONTROL ET'e basın.", "info")
                except Exception as diag_err:
                    log(f"    [TANI] Element listeleme hatası: {diag_err}", "warn")
                return False

        # Sayfa yüklenene kadar bekle
        for bekle in range(15):
            time.sleep(1)
            if durduruldu_mu():
                log("DURDURULDU", "warn")
                return False
            if element_bul(medula, "form1:buttonSonlandirilmamisReceteler"):
                log("Sayfa yüklendi", "ok")
                break
        else:
            if deneme < 2:
                log(f"Sayfa yüklenemedi, tekrar deneniyor ({deneme+2}/3)...", "warn")
                time.sleep(2)
                continue
            log("Sayfa yüklenemedi!", "error")
            return False
        break  # Başarılı

    # 2. Dönem seçimi
    if donem_offset > 0:
        log(f"Dönem seçiliyor (offset: {donem_offset})...", "info")
        if element_tikla(medula, "form1:menu2", "click"):
            time.sleep(0.5)
            for _ in range(donem_offset):
                pyautogui.press("down")
                time.sleep(0.15)
            pyautogui.press("enter")
            log("Dönem seçildi", "ok")
        time.sleep(1)

    # 3. Fatura Türü
    grup_combo_index = {"A": 1, "B": 4, "C": 7, "CK": 10, "GK": 16}
    combo_idx = grup_combo_index.get(grup, 1)

    log(f"Fatura Türü: {grup} seçiliyor...", "info")
    if element_tikla(medula, "form1:menu1", "click"):
        time.sleep(0.5)
        for _ in range(20):
            pyautogui.press("up")
            time.sleep(0.03)
        time.sleep(0.2)
        for _ in range(combo_idx):
            pyautogui.press("down")
            time.sleep(0.1)
        pyautogui.press("enter")
        log(f"Fatura Türü seçildi: {grup}", "ok")
    time.sleep(1)

    # 4. Sorgula
    log("Sorgula butonuna basılıyor...", "info")
    if element_tikla(medula, "form1:buttonSonlandirilmamisReceteler"):
        log("Sorgula tıklandı", "ok")

    # Reçete listesi yüklenene kadar bekle (max 8 sn)
    for _ in range(16):
        time.sleep(0.5)
        try:
            for elem in medula.descendants(control_type="DataItem"):
                txt = (elem.window_text() or "").strip()
                if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                    break
            else:
                continue
            break
        except:
            pass

    # 5. İlk reçeteye tıkla
    ilk_recete_bulundu = False
    try:
        for elem in medula.descendants(control_type="DataItem"):
            try:
                txt = elem.window_text()
                if txt:
                    txt = txt.strip()
                    if len(txt) == 7 and txt[0].isdigit() and txt.isalnum():
                        elem.click_input()
                        aktivite_bildir()
                        log(f"İlk reçete: {txt}", "ok")
                        ilk_recete_bulundu = True
                        break
            except:
                pass
    except:
        pass
    if not ilk_recete_bulundu:
        log("Reçete bulunamadı!", "error")
        return False

    # İlaç tablosu yüklenene kadar bekle
    for bekle in range(6):
        time.sleep(0.5)
        if element_bul(medula, "f:tbl1:0:t6"):  # İlk ilaç adı hücresi
            break
    time.sleep(0.5)
    return True


# ========== MEDULA OTURUM ==========
def oturum_kontrol_ve_baglan(medula):
    """Oturum aktif mi kontrol et. Düşmüşse yeniden başlat.
    Kontrol sırası:
    1. Sistem hatası sayfası → yeniden başlat
    2. Web elementleri varsa (reçete sayfası veya menü) → oturum aktif
    3. Giriş butonu varsa → tıkla, bekle
    4. Hiçbiri yoksa (embedded browser boş) → taskkill + yeniden başlat
    """
    # Sistem hatası kontrolü
    if sistem_dusmus_mu(medula):
        log("Sistem hatası tespit edildi - yeniden başlatılıyor...", "error")
        return medula_yeniden_baslat()

    # Reçete sayfası elementleri varsa oturum aktiftir
    if element_bul(medula, "f:tbl1") or element_bul(medula, "f:buttonSonraki"):
        return medula

    # Menü görünüyor mu?
    if element_bul(medula, "form1:menuHtmlCommandExButton31"):
        return medula

    # Giriş butonu varsa tıkla
    if element_bul(medula, "btnMedulayaGirisYap"):
        log("Giriş butonu bulundu, tıklanıyor...", "warn")
        element_tikla(medula, "btnMedulayaGirisYap", "click")
        time.sleep(5)
        if element_bul(medula, "form1:menuHtmlCommandExButton31") or element_bul(medula, "f:tbl1"):
            log("Oturum yenilendi", "ok")
            return medula

    # Hiçbir web elementi yok — embedded browser yüklenmiyor olabilir
    # Kapatmadan bekle — sayfa yüklenebilir
    log("Web elementleri bulunamadı — sayfa yüklenmesi bekleniyor...", "warn")
    for bekle in range(10):
        time.sleep(2)
        if element_bul(medula, "form1:menuHtmlCommandExButton31") or element_bul(medula, "f:tbl1"):
            log(f"Sayfa yüklendi ({(bekle+1)*2}s)", "ok")
            return medula
        if element_bul(medula, "btnMedulayaGirisYap"):
            log("Giriş butonu göründü, tıklanıyor...", "warn")
            element_tikla(medula, "btnMedulayaGirisYap", "click")
            time.sleep(5)
            medula = medula_bul()
            if medula and (element_bul(medula, "form1:menuHtmlCommandExButton31") or element_bul(medula, "f:tbl1")):
                log("Oturum yenilendi", "ok")
                return medula

    # 20 saniye bekledik, hâlâ yüklenmediyse — sadece o zaman yeniden başlat
    log("20s beklendi, sayfa yüklenemedi — yeniden başlatılıyor...", "error")
    return medula_yeniden_baslat()


# ========== RAPOR ==========
DURUM_DOSYASI = os.path.join(PROJE_DIZINI, "recete_kontrol_durumlari.json")
AI_IMPL_DOSYA = os.path.join(PROJE_DIZINI, "ai_impl_kuyruk.json")  # AI implementasyon kuyruğu


_ai_istenen_ilaclar = set()  # Aynı ilaç için tekrar AI açma
_kuyruk_lock = threading.Lock()  # Kuyruk dosyası thread-safe erişim
_ai_threads = []  # Aktif AI thread'leri (atexit ile bekleme için)


def _ai_thread_bekle():
    """Process kapanırken aktif AI thread'lerini bekle (max 30sn)."""
    for t in _ai_threads:
        if t.is_alive():
            t.join(timeout=30)


atexit.register(_ai_thread_bekle)


def _kuyruk_guncelle(ilac_adi, etkin_madde, rapor_kodu, durum, **ekstra):
    """Kuyruk dosyasını thread-safe güncelle. Aynı ilaç varsa günceller, yoksa ekler."""
    with _kuyruk_lock:
        kuyruk = []
        if os.path.exists(AI_IMPL_DOSYA):
            try:
                with open(AI_IMPL_DOSYA, "r", encoding="utf-8") as f:
                    kuyruk = json.load(f)
            except:
                kuyruk = []

        # Aynı ilaç için mevcut kayıt varsa güncelle
        mevcut = None
        for k in kuyruk:
            if k.get("ilac_adi") == ilac_adi:
                mevcut = k
                break

        if mevcut:
            mevcut["durum"] = durum
            mevcut["guncelleme_tarihi"] = datetime.now().isoformat()
            mevcut.update(ekstra)
        else:
            kayit = {
                "ilac_adi": ilac_adi, "etkin_madde": etkin_madde,
                "rapor_kodu": rapor_kodu, "tarih": datetime.now().isoformat(),
                "durum": durum
            }
            kayit.update(ekstra)
            kuyruk.append(kayit)

        with open(AI_IMPL_DOSYA, "w", encoding="utf-8") as f:
            json.dump(kuyruk, f, indent=2, ensure_ascii=False)

_CLAUDE_CLI_YOK = False  # Bir kez bulunmazsa tekrar denememek için

def _ai_claude_calistir(prompt, timeout=180, dosya_yazabilir=False, web_arama=False):
    """Claude CLI çalıştır ve sonucu döndür. Bloklar.
    Claude CLI yoksa sessizce boş döner (her çağrıda tekrar hata basmaz)."""
    global _CLAUDE_CLI_YOK
    if _CLAUDE_CLI_YOK:
        return ""
    import subprocess as sp
    # Windows'ta claude.cmd veya claude.exe olabilir — shell=True gerekebilir
    try:
        cmd = ["claude", "-p", prompt, "--output-format", "text"]
        if dosya_yazabilir:
            cmd.insert(1, "--dangerously-skip-permissions")
        if web_arama:
            cmd.extend(["--allowedTools", "WebFetch", "WebSearch"])
        result = sp.run(
            cmd, capture_output=True, text=True, timeout=timeout,
            cwd=PROJE_DIZINI, encoding="utf-8", errors="replace",
            shell=True  # Windows'ta claude.cmd için
        )
        return result.stdout.strip() if result.stdout else ""
    except sp.TimeoutExpired:
        log(f"  [AI] Claude zaman aşımı ({timeout}sn)", "warn")
    except FileNotFoundError:
        _CLAUDE_CLI_YOK = True
        log(f"  [AI] Claude CLI bulunamadı — AI SUT analizi devre dışı (bu oturum için)", "warn")
    except Exception as e:
        log(f"  [AI] Claude hatası: {e}", "error")
    return ""


def _ai_sut_implementasyon_iste(ilac_adi, etkin_madde, rapor_kodu, mesaj_metni="",
                                  senkron=False):
    """SUT kuralı tespit edilemeyen ilaç için Claude AI ile araştırma.
    1. DB kontrol: son 6 ay içinde araştırıldıysa tekrar sorma (TTL)
    2. Kategori belirle (web aramalı) → DB'ye kaydet → runtime eşleştirme
    3. Mevcut kategoride kontrol fonksiyonu yoksa → sut_kontrolleri.py'ye kod yaz
    4. Yazılan kodu doğrula, hatalıysa geri al

    Args:
        senkron: True ise AI bitene kadar bekler (bu reçete için sonuç kullanılabilir).
                 False (varsayılan) arka plan thread — tarama bloklanmaz.
    Returns:
        senkron=True: dict {kategori, aciklama, etkin_madde, basarili} veya None
        senkron=False: None (her zaman)
    """
    # 1. DB'den TTL kontrolü (6 ay = 180 gün)
    try:
        from kontrol_kurallari import get_ogrenilen_ilaclar
        ogrenilen = get_ogrenilen_ilaclar()
        arastir, mevcut = ogrenilen.sut_arastirma_gerekli_mi(ilac_adi, ttl_gun=180)
        if not arastir and mevcut:
            log(f"  [AI] SUT kuralı DB'de mevcut (son 6 ay içinde araştırılmış): {ilac_adi} → {mevcut.get('sut_kategori','?')}", "info")
            if senkron:
                return {
                    "kategori": mevcut.get("sut_kategori"),
                    "aciklama": mevcut.get("sut_aciklama", ""),
                    "etkin_madde": mevcut.get("etkin_madde"),
                    "basarili": True,
                    "kaynak": "db_cache",
                }
            return None
    except Exception as e:
        log(f"  [AI] DB kontrol hatası: {e}", "warn")

    # Aynı ilaç için tekrar sorma (oturum içi)
    key = f"{ilac_adi}:{etkin_madde}"
    if key in _ai_istenen_ilaclar:
        if senkron:
            return None  # Zaten başka thread çalışıyor
        return
    _ai_istenen_ilaclar.add(key)

    log(f"  [AI] SUT analizi başlıyor: {ilac_adi} ({etkin_madde})", "warn")

    def ai_thread():
        try:
            # Kuyruğa başlangıç durumu yaz
            _kuyruk_guncelle(ilac_adi, etkin_madde, rapor_kodu, "ai_baslatildi")

            # ══════ AŞAMA 1: KATEGORİ BELİRLE (web aramalı) ══════
            kategoriler = (
                "KOMBINE_ANTIHIPERTANSIF, DIYABET_DPP4_SGLT2, KLOPIDOGREL, STATIN, FIBRAT, "
                "YOAK, IVABRADIN, EPLERENON, RANOLAZIN, PSIKIYATRI, SOLUNUM, ONKOLOJI, NOROLOJI, "
                "GOZ, ANTIVIRAL, GIS, RECETE_TURU_KIRMIZI, RECETE_TURU_YESIL, TRIMETAZIDIN, "
                "DMAH, KADIN_HORMON, ANTIBIYOTIK_FLOROKINOLON, RAPORSUZ_BILGILENDIRME, "
                "MONO_ANTIHIPERTANSIF, POTASYUM_SITRAT, GENEL_RAPORLU"
            )

            prompt1 = (
                f"Türkiye SGK SUT mevzuatına göre bu ilacın SUT kategorisini belirle.\n"
                f"İlaç: {ilac_adi}\n"
                f"Etkin Madde: {etkin_madde or 'BİLİNMİYOR'}\n"
                f"Rapor Kodu: {rapor_kodu or 'yok'}\n"
                f"Mesaj: {(mesaj_metni or 'yok')[:200]}\n\n"
                f"KRİTİK KURAL: Etkin madde BİLİNMİYOR ise KESİNLİKLE TAHMİN YAPMA!\n"
                f"İlaç adından etkin madde tahmin etme — yanlış sonuç verir.\n"
                f"Etkin madde bilinmiyorsa KATEGORI: GENEL_RAPORLU yaz ve ETKIN_MADDE: BILINMIYOR yaz.\n"
                f"Sadece etkin madde kesin olarak biliniyorsa kategori belirle.\n\n"
                f"Web araması yaparak ilacabak.com veya rxmedicinturkey.com adresinden "
                f"ilacın etkin maddesini doğrula.\n\n"
                f"Mevcut kategoriler:\n{kategoriler}\n\n"
                f"SADECE şu formatta cevap ver, başka bir şey yazma:\n"
                f"KATEGORI: <kategori_kodu>\n"
                f"ETKIN_MADDE: <etkin madde adı büyük harf>\n"
                f"ACIKLAMA: <kısa açıklama>\n\n"
                f"Eğer mevcut kategorilerden hiçbiri uymuyorsa KATEGORI: GENEL_RAPORLU yaz.\n"
                f"Eğer raporsuz verilebilen bir ilaç ise KATEGORI: RAPORSUZ_BILGILENDIRME yaz."
            )

            cikti = _ai_claude_calistir(prompt1, timeout=90, web_arama=True)
            if not cikti:
                log(f"  [AI] Aşama 1 başarısız ({ilac_adi})", "warn")
                _kuyruk_guncelle(ilac_adi, etkin_madde, rapor_kodu, "basarisiz",
                                 hata="Claude çıktı vermedi")
                return

            # Parse et
            kategori = None
            ai_etkin = None
            aciklama = ""
            for satir in cikti.split("\n"):
                satir = satir.strip()
                if satir.startswith("KATEGORI:"):
                    kategori = satir.split(":", 1)[1].strip()
                elif satir.startswith("ETKIN_MADDE:"):
                    ai_etkin = satir.split(":", 1)[1].strip()
                elif satir.startswith("ACIKLAMA:"):
                    aciklama = satir.split(":", 1)[1].strip()

            if not kategori:
                log(f"  [AI] Kategori parse edilemedi: {cikti[:100]}", "warn")
                _kuyruk_guncelle(ilac_adi, etkin_madde, rapor_kodu, "basarisiz",
                                 hata="Kategori parse edilemedi")
                return

            log(f"  [AI] Kategori: {ilac_adi} → {kategori} ({aciklama})", "ok")
            _kuyruk_guncelle(ilac_adi, ai_etkin or etkin_madde, rapor_kodu, "kategori_belirlendi",
                             kategori=kategori, aciklama=aciklama)

            # DB'ye kaydet — hem ilaç hem de SUT araştırma sonucu (6 ay TTL için)
            try:
                from kontrol_kurallari import get_ogrenilen_ilaclar
                ogrenilen = get_ogrenilen_ilaclar()
                ogrenilen.ilac_kaydet(
                    ilac_adi, etkin_madde=ai_etkin or etkin_madde,
                    farmakolojik_grup=kategori, rapor_kodu=rapor_kodu
                )
                # SUT araştırma damgası (6 ay TTL için)
                ogrenilen.sut_arastirma_kaydet(
                    ilac_adi, kategori=kategori, aciklama=aciklama,
                    etkin_madde=ai_etkin or etkin_madde
                )
            except Exception:
                pass

            # Runtime sözlüklere ekle
            try:
                from recete_kontrol import sut_kontrolleri
                ilac_kisa = ilac_adi.strip().upper().split()[0] if ilac_adi else ""
                if ilac_kisa and len(ilac_kisa) >= 3:
                    sut_kontrolleri.ILAC_ADI_KATEGORI[ilac_kisa] = kategori
                if ai_etkin:
                    sut_kontrolleri.ETKIN_MADDE_KATEGORI[ai_etkin.upper()] = kategori
            except Exception:
                pass

            # ══════ AŞAMA 2: KONTROL FONKSİYONU — YENİ YAZ VEYA MEVCUT GENİŞLET ══════
            try:
                from recete_kontrol import sut_kontrolleri
                fonk_var = kategori in sut_kontrolleri.KATEGORI_KONTROL_FONKSIYONU
            except Exception:
                fonk_var = False

            if fonk_var:
                log(f"  [AI] Kontrol fonksiyonu mevcut: {kategori} - genişletiliyor...", "info")
            else:
                log(f"  [AI] Kontrol fonksiyonu YOK: {kategori} - yeni yazılıyor...", "warn")

            kod_basarili = _ai_kod_yaz(kategori, ilac_adi, ai_etkin or etkin_madde,
                                        rapor_kodu, mesaj_metni, aciklama, guncelle=fonk_var)

            _kuyruk_guncelle(ilac_adi, ai_etkin or etkin_madde, rapor_kodu,
                             "tamamlandi" if kod_basarili else "kod_hatasi",
                             kategori=kategori, aciklama=aciklama,
                             kod_yazildi=kod_basarili)

        except Exception as e:
            log(f"  [AI] Thread beklenmeyen hata: {e}", "error")
            try:
                _kuyruk_guncelle(ilac_adi, etkin_madde, rapor_kodu, "hata",
                                 hata=str(e))
            except:
                pass

    if senkron:
        # Senkron mod: thread'i başlat VE bitene kadar bekle, sonucu DB'den al
        t = threading.Thread(target=ai_thread, daemon=False)
        t.start()
        t.join(timeout=300)  # Max 5 dk bekle
        # DB'den güncel araştırma sonucunu al
        try:
            from kontrol_kurallari import get_ogrenilen_ilaclar
            ogrenilen = get_ogrenilen_ilaclar()
            kayit = ogrenilen.ilac_bul(ilac_adi)
            if kayit and kayit.get("sut_kategori"):
                return {
                    "kategori": kayit["sut_kategori"],
                    "aciklama": kayit.get("sut_aciklama", ""),
                    "etkin_madde": kayit.get("etkin_madde"),
                    "basarili": True,
                    "kaynak": "ai_yeni",
                }
        except Exception:
            pass
        return None
    else:
        # Asenkron mod (varsayılan): arka planda çalışsın, tarama beklemesin
        t = threading.Thread(target=ai_thread, daemon=False)
        t.start()
        _ai_threads.append(t)
        return None


def _dosya_yedekle(dosya_yolu):
    """Dosyayı .bak.timestamp uzantısıyla yedekle. Son 3 yedeği tut."""
    import shutil
    import glob as _glob
    yedek = dosya_yolu + f".bak.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(dosya_yolu, yedek)
    log(f"  [AI] Yedek alındı: {os.path.basename(yedek)}", "info")
    # Eski yedekleri temizle (son 3 tut)
    yedekler = sorted(_glob.glob(dosya_yolu + ".bak.*"))
    for eski in yedekler[:-3]:
        try:
            os.remove(eski)
        except:
            pass
    return yedek


def _yedekten_geri_al(dosya_yolu):
    """En son yedekten geri yükle ve modülü yeniden yükle."""
    import shutil
    import glob as _glob
    yedekler = sorted(_glob.glob(dosya_yolu + ".bak.*"))
    if yedekler:
        shutil.copy2(yedekler[-1], dosya_yolu)
        log(f"  [AI] Yedekten geri yüklendi: {os.path.basename(yedekler[-1])}", "warn")
        try:
            import importlib
            from recete_kontrol import sut_kontrolleri
            importlib.reload(sut_kontrolleri)
        except:
            pass
        return True
    return False


def _ai_kod_dogrula(sut_dosya, kategori):
    """Yazılan kodun syntax, import ve çalışabilirlik kontrolü.
    Returns: (basarili: bool, hata_mesaji: str)
    """
    # 1. Syntax kontrolü
    import py_compile
    try:
        py_compile.compile(sut_dosya, doraise=True)
    except py_compile.PyCompileError as e:
        return False, f"Syntax hatası: {e}"

    # 2. Import kontrolü — modülü yeniden yükle
    try:
        import importlib
        from recete_kontrol import sut_kontrolleri
        importlib.reload(sut_kontrolleri)
    except Exception as e:
        return False, f"Import hatası: {e}"

    # 3. Fonksiyon varlığı kontrolü
    fonk_adi = f"kontrol_{kategori.lower()}"
    if not hasattr(sut_kontrolleri, fonk_adi):
        return False, f"Fonksiyon bulunamadı: {fonk_adi}"

    fonk = getattr(sut_kontrolleri, fonk_adi)
    if not callable(fonk):
        return False, f"{fonk_adi} çağrılabilir değil"

    # 4. Basit test — boş input ile çağır, exception fırlatmamalı
    try:
        test_ilac = {"ilac_adi": "TEST", "rapor_kodu": "", "mesajlar": [],
                     "rapor_aciklamalari": [], "rapor_tani_bilgileri": []}
        sonuc = fonk(test_ilac)
        if not isinstance(sonuc, sut_kontrolleri.KontrolRaporu):
            return False, f"Yanlış dönüş tipi: {type(sonuc)}"
    except Exception as e:
        return False, f"Çalışma hatası: {e}"

    # 5. KATEGORI_KONTROL_FONKSIYONU'na eklenmiş mi?
    if kategori not in sut_kontrolleri.KATEGORI_KONTROL_FONKSIYONU:
        return False, f"{kategori} KATEGORI_KONTROL_FONKSIYONU'na eklenmemiş"

    return True, "OK"


def _ai_kod_yaz(kategori, ilac_adi, etkin_madde, rapor_kodu, mesaj_metni, aciklama, guncelle=False):
    """Claude'a sut_kontrolleri.py'ye kontrol fonksiyonu yazdır veya mevcut olanı genişlet.
    Yazmadan önce yedekler, yazdıktan sonra doğrular, hatalıysa geri alır.
    Returns: True=başarılı, False=başarısız
    """
    sut_dosya = os.path.join(PROJE_DIZINI, "recete_kontrol", "sut_kontrolleri.py")

    # 1. YEDEKLE
    _dosya_yedekle(sut_dosya)

    # 2. Mevcut fonksiyon kodunu oku (güncelleme için)
    mevcut_fonk_kodu = ""
    if guncelle:
        try:
            with open(sut_dosya, "r", encoding="utf-8") as f:
                icerik = f.read()
            fonk_adi = f"kontrol_{kategori.lower()}"
            import re as _re
            match = _re.search(
                rf'(def {fonk_adi}\(.*?\n(?:(?:    .*|)\n)*)',
                icerik
            )
            if match:
                mevcut_fonk_kodu = match.group(1)
        except Exception:
            pass

    # 3. Prompt oluştur
    if guncelle and mevcut_fonk_kodu:
        gorev = (
            f"Bu dosyada MEVCUT kontrol fonksiyonunu GENİŞLET ve İYİLEŞTİR.\n"
            f"Mevcut fonksiyon basit bir keyword kontrolü yapıyor. SUT mevzuatının\n"
            f"TÜM kurallarını kapsayacak şekilde genişlet.\n\n"
            f"MEVCUT FONKSİYON:\n```python\n{mevcut_fonk_kodu}```\n\n"
            f"GENİŞLETME TALİMATLARI:\n"
            f"- Mevcut fonksiyonu SİL ve yerine detaylı versiyonunu yaz\n"
            f"- SUT mevzuatındaki TÜM kuralları kontrol et:\n"
            f"  * Hangi uzman yazabilir?\n"
            f"  * Rapor süresi ne kadar?\n"
            f"  * Doz kısıtlamaları var mı?\n"
            f"  * Kombinasyon kısıtlamaları var mı?\n"
            f"  * İlk reçete/devam reçete farkı var mı?\n"
            f"  * Raporsuz yazılabilir mi?\n"
            f"- Her kural için ayrı kontrol yap ve detaylı mesaj döndür\n"
        )
    else:
        gorev = (
            f"Bu dosyada YENİ bir SUT kontrol fonksiyonu implemente et.\n\n"
            f"YAPILACAKLAR:\n"
            f"1. İnternetten (ilacabak.com, SGK SUT) bu ilacın SUT kuralını araştır\n"
            f"2. recete_kontrol/sut_kontrolleri.py dosyasına:\n"
            f"   a) ETKIN_MADDE_KATEGORI sözlüğüne '{etkin_madde}': '{kategori}' ekle\n"
            f"   b) def kontrol_{kategori.lower()}(ilac_sonuc) fonksiyonu yaz\n"
            f"   c) KATEGORI_KONTROL_FONKSIYONU sözlüğüne '{kategori}': kontrol_{kategori.lower()} ekle\n"
            f"   d) KATEGORI_ISIMLERI sözlüğüne açıklayıcı isim ekle\n"
        )

    prompt = (
        f"Proje dizini: {PROJE_DIZINI}\n"
        f"Dosya: recete_kontrol/sut_kontrolleri.py\n\n"
        f"{gorev}\n"
        f"İlaç: {ilac_adi}\n"
        f"Etkin Madde: {etkin_madde}\n"
        f"Rapor Kodu: {rapor_kodu or 'yok'}\n"
        f"Mesaj örneği: {(mesaj_metni or 'yok')[:300]}\n"
        f"Kategori kodu: {kategori}\n"
        f"Açıklama: {aciklama}\n\n"
        f"Kontrol fonksiyonu şablonu:\n"
        f"```python\n"
        f"def kontrol_{kategori.lower()}(ilac_sonuc: Dict) -> KontrolRaporu:\n"
        f"    metin = _tum_metinleri_birlesir(ilac_sonuc)\n"
        f"    rapor_kodu = ilac_sonuc.get('rapor_kodu', '')\n"
        f"    if not metin and not rapor_kodu:\n"
        f"        return KontrolRaporu(KontrolSonucu.KONTROL_EDILEMEDI, 'Rapor/mesaj metni yok')\n"
        f"    # SUT kuralına göre DETAYLI kontrol...\n"
        f"    # _turkce_ara(metin, 'aranan ibare') kullan (Türkçe karakter güvenli)\n"
        f"    return KontrolRaporu(KontrolSonucu.UYGUN, 'Detaylı açıklama')\n"
        f"```\n\n"
        f"YARDIMCI FONKSİYONLAR:\n"
        f"- _turkce_ara(metin, aranan) → bool: İ/i/ı fark etmez\n"
        f"- _tum_metinleri_birlesir(ilac_sonuc) → str: Tüm metin kaynakları\n"
        f"- KontrolSonucu.UYGUN / UYGUN_DEGIL / KONTROL_EDILEMEDI\n"
        f"- KontrolRaporu(sonuc, mesaj, detaylar={{}}, uyari='')\n\n"
        f"KRİTİK KURALLAR:\n"
        f"- SADECE recete_kontrol/sut_kontrolleri.py dosyasını düzenle\n"
        f"- Mevcut DİĞER fonksiyonları BOZMA\n"
        f"- Medula verileri değiştirme YASAK\n"
        f"- DB'ye INSERT/UPDATE/DELETE YASAK\n"
    )

    # 4. Claude'a yazdır
    log(f"  [AI] Claude kod yazıyor: {kategori}...", "info")
    cikti = _ai_claude_calistir(prompt, timeout=900, dosya_yazabilir=True)

    if not cikti:
        log(f"  [AI] Kod yazma başarısız: {kategori}", "warn")
        _yedekten_geri_al(sut_dosya)
        return False

    # 5. DOĞRULA
    basarili, hata = _ai_kod_dogrula(sut_dosya, kategori)
    if not basarili:
        log(f"  [AI] Kod HATALI, geri alınıyor: {hata}", "error")
        _yedekten_geri_al(sut_dosya)
        return False

    log(f"  [AI] Kod başarıyla yazıldı ve doğrulandı: {kategori}", "ok")
    return True
RAPOR_JSON = os.path.join(PROJE_DIZINI, "tarama_rapor.json")


def _recete_sorgu_ile_git(medula, recete_no):
    """Reçete Sorgu sayfasından belirli bir reçeteye git.
    Önce ana sayfaya dön (menü görünür olsun) → Reçete Sorgu → No yaz → Sorgula
    Returns: True başarılı, False başarısız
    """
    log(f"  Reçete Sorgu ile {recete_no}'ya gidiliyor...", "info")

    # webBrowser1 focus
    try:
        for c in medula.children():
            try:
                if c.element_info.automation_id == "webBrowser1":
                    c.set_focus()
                    break
            except:
                pass
    except:
        pass
    time.sleep(0.3)

    # Önce menü görünür mü kontrol et
    menu_var = element_bul(medula, "form1:menuHtmlCommandExButton51")
    if not menu_var:
        # Menü yok - Giriş butonuyla ana sayfaya dön
        log("  Menü görünmüyor, Giriş butonuyla ana sayfaya dönülüyor...", "info")
        element_tikla(medula, "btnMedulayaGirisYap", "click")
        time.sleep(5)
        menu_var = element_bul(medula, "form1:menuHtmlCommandExButton51")

    if not menu_var:
        log("  Reçete Sorgu menüsü bulunamadı", "warn")
        return False

    # 1. Reçete Sorgu menüsüne tıkla
    if not element_tikla(medula, "form1:menuHtmlCommandExButton51"):
        log("  Reçete Sorgu tıklanamadı", "warn")
        return False

    # Sayfa yüklenene kadar bekle (text2 textbox görünene kadar)
    for i in range(10):
        time.sleep(1)
        if element_bul(medula, "form1:text2"):
            break
    else:
        log("  Reçete Sorgu sayfası yüklenemedi", "warn")
        return False

    # 2. Reçete numarasını yaz
    txt_elem = element_bul(medula, "form1:text2")
    if not txt_elem:
        log("  Reçete No textbox bulunamadı", "warn")
        return False

    try:
        txt_elem.click_input()
        time.sleep(0.2)
        import pyautogui
        pyautogui.hotkey("ctrl", "a")
        time.sleep(0.1)
        pyautogui.typewrite(recete_no, interval=0.03)
        log(f"  Reçete No yazıldı: {recete_no}", "info")
    except Exception as e:
        log(f"  Reçete No yazılamadı: {e}", "warn")
        return False

    time.sleep(0.5)

    # 3. Sorgula butonuna bas
    if not element_tikla(medula, "form1:buttonReceteNoSorgula"):
        log("  Sorgula butonu bulunamadı", "warn")
        return False

    log("  Sorgula tıklandı, reçete açılıyor...", "info")

    # 4. Reçete sayfasının yüklenmesini bekle
    for i in range(10):
        time.sleep(1)
        if element_bul(medula, "f:tbl1"):
            rno = recete_no_oku(medula)
            log(f"  Reçete açıldı: {rno}", "ok")
            return True

    log("  Reçete açılamadı", "warn")
    return False


def _son_recete_kaydet(grup, recete_no, sira=0):
    """Son kontrol edilen reçete no'yu JSON'a kaydet"""
    try:
        durum = {}
        if os.path.exists(DURUM_DOSYASI):
            with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                durum = json.load(f)
        if grup not in durum:
            durum[grup] = {"son_recete": "", "toplam_kontrol": 0, "son_kontrol_tarihi": None}
        durum[grup]["son_recete"] = recete_no
        durum[grup]["toplam_kontrol"] = durum[grup].get("toplam_kontrol", 0) + 1
        durum[grup]["son_kontrol_tarihi"] = datetime.now().isoformat()
        with open(DURUM_DOSYASI, "w", encoding="utf-8") as f:
            json.dump(durum, f, indent=2, ensure_ascii=False)
        # stdout'a da yaz (GUI combobox + durum güncellemesi için)
        print(f"[SON_RECETE] {grup}:{recete_no}:{sira}", flush=True)
    except:
        pass


def _son_recete_al(grup):
    """Kaldığı yerden devam için son reçete no'yu oku"""
    try:
        if os.path.exists(DURUM_DOSYASI):
            with open(DURUM_DOSYASI, "r", encoding="utf-8") as f:
                durum = json.load(f)
            return durum.get(grup, {}).get("son_recete", "")
    except:
        pass
    return ""

RAPOR_SUTUNLAR = [
    ("Reçete No", 12),
    ("Reçete Türü", 12),
    ("İlaç Adı", 40),
    ("Etkin Madde", 25),
    ("Rapor Kodu", 10),
    ("Msj", 5),
    ("Renkli Reçete", 14),
    ("Doz Kontrol", 14),        # Reçete dozu vs rapor dozu
    ("SUT Kontrol", 12),        # SUT kuralları kontrolü
    ("Sonuç", 14),
    ("Açıklama", 60),
]

def rapor_kaydet(satirlar, grup):
    """Raporu JSON + Excel olarak kaydet"""
    if not satirlar:
        log("Rapor verisi yok, kayıt atlanıyor", "warn")
        return

    # 1. JSON kaydet (GUI tablo için)
    try:
        with open(RAPOR_JSON, "w", encoding="utf-8") as f:
            json.dump({"grup": grup, "tarih": datetime.now().isoformat(),
                       "satirlar": satirlar}, f, ensure_ascii=False, indent=2)
        log(f"Rapor JSON kaydedildi: {RAPOR_JSON}", "ok")
    except Exception as e:
        log(f"JSON kayıt hatası: {e}", "error")

    # 2. Excel kaydet
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"{grup} Grubu Kontrol"

        # Stiller (renkler: yeşil=uygun, kırmızı=uygunsuz, turuncu=şüpheli, sarı=edilemedi, gri=geçildi)
        baslik_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        baslik_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        recete_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        uygun_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")     # yeşil
        sorun_fill = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")     # kırmızı
        uyari_fill = PatternFill(start_color="FFB74D", end_color="FFB74D", fill_type="solid")     # turuncu (şüpheli)
        edilemedi_fill = PatternFill(start_color="FFF176", end_color="FFF176", fill_type="solid")  # sarı
        gecildi_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")    # gri
        yeni_fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
        ince_border = Border(
            left=Side(style="thin", color="BDBDBD"),
            right=Side(style="thin", color="BDBDBD"),
            top=Side(style="thin", color="BDBDBD"),
            bottom=Side(style="thin", color="BDBDBD")
        )

        # Başlık satırı
        for col_idx, (baslik, genislik) in enumerate(RAPOR_SUTUNLAR, 1):
            cell = ws.cell(row=1, column=col_idx, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = ince_border
            ws.column_dimensions[cell.column_letter].width = genislik

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:K{len(satirlar) + 1}"

        # Veri satırları
        onceki_recete = None
        for row_idx, satir in enumerate(satirlar, 2):
            recete_no = satir.get("recete_no", "")
            yeni_recete = recete_no != onceki_recete
            onceki_recete = recete_no

            _sonuc_raw = satir.get("sonuc", "")
            _SONUC_DISPLAY = {
                "UYGUN": "✓ Uygun",
                "UYGUNSUZ": "✗ Uygunsuz",
                "UYARI": "⚠ Şüpheli",
                "KONTROLEDİLEMEDİ": "? Kontrol Edilemedi",
                "KONTROLEDILEMEDI": "? Kontrol Edilemedi",
                "GECİLDİ": "— Geçildi",
                "GECILDI": "— Geçildi",
                "YENİ": "+ Yeni",
            }
            _sonuc_display = _SONUC_DISPLAY.get(_sonuc_raw, _sonuc_raw)
            _DOZ_DISPLAY = {
                "Uygun": "✓ Uygun", "UygunDegil": "✗ Aşım", "Uyarı": "⚠ Şüpheli",
                "Bypass": "— Bypass", "-": "-",
            }
            _doz_raw = satir.get("rapor_kontrol", "-")
            _doz_display = _DOZ_DISPLAY.get(_doz_raw, _doz_raw)
            _SUT_DISPLAY = {
                "Uygun": "✓ Uygun", "UygunDegil": "✗ Uygunsuz", "KontrolEdilemedi": "? Edilemedi",
                "Bypass": "— Bypass", "-": "-",
            }
            _sut_raw = satir.get("sut_kontrol", "-")
            _sut_display = _SUT_DISPLAY.get(_sut_raw, _sut_raw)

            degerler = [
                recete_no if yeni_recete else "",
                satir.get("recete_turu", "") if yeni_recete else "",
                satir.get("ilac_adi", ""),
                satir.get("etkin_madde", ""),
                satir.get("rapor_kodu", ""),
                satir.get("msj", ""),
                satir.get("renkli_kontrol", "-"),
                _doz_display,
                _sut_display,
                _sonuc_display,
                satir.get("aciklama", ""),
            ]

            for col_idx, deger in enumerate(degerler, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=deger)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = ince_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                # Reçete satırı arka plan
                if yeni_recete and col_idx <= 2:
                    cell.fill = recete_fill
                    cell.font = Font(name="Segoe UI", size=10, bold=True)

            # Kontrol sütunlarını renklendir — raw değerlere göre
            sonuc = satir.get("sonuc", "")
            _raw_by_col = {7: satir.get("renkli_kontrol", "-"), 8: _doz_raw, 9: _sut_raw}
            for col in [7, 8, 9]:
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                val = _raw_by_col.get(col, "-")
                if val == "Uygun":
                    cell.fill = uygun_fill
                elif val in ("UygunDegil", "Uygun Değil"):
                    cell.fill = sorun_fill
                elif val == "Uyarı":
                    cell.fill = uyari_fill
                elif val == "KontrolEdilemedi":
                    cell.fill = edilemedi_fill
                elif val in ("Bypass", "raporsuz", "okunamadi", "-", ""):
                    cell.fill = gecildi_fill
                elif val == "Yeni":
                    cell.fill = yeni_fill

            # Sonuç sütunu (10) — ana verdict
            sonuc_cell = ws.cell(row=row_idx, column=10)
            sonuc_cell.font = Font(name="Segoe UI", size=10, bold=True)
            sonuc_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if sonuc == "UYGUN":
                sonuc_cell.fill = uygun_fill
                sonuc_cell.font = Font(name="Segoe UI", size=10, bold=True, color="1B5E20")
            elif sonuc == "UYGUNSUZ":
                sonuc_cell.fill = sorun_fill
                sonuc_cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            elif sonuc == "UYARI":
                sonuc_cell.fill = uyari_fill
                sonuc_cell.font = Font(name="Segoe UI", size=10, bold=True, color="E65100")
            elif sonuc in ("KONTROLEDİLEMEDİ", "KONTROLEDILEMEDI"):
                sonuc_cell.fill = edilemedi_fill
                sonuc_cell.font = Font(name="Segoe UI", size=10, bold=True, color="5D4037")
            elif sonuc in ("GECİLDİ", "GECILDI"):
                sonuc_cell.fill = gecildi_fill
                sonuc_cell.font = Font(name="Segoe UI", size=10, italic=True, color="616161")
            elif sonuc == "YENİ":
                sonuc_cell.fill = yeni_fill

        # Kaydet
        masaustu = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        tarih = datetime.now().strftime("%Y%m%d_%H%M")
        excel_dosya = os.path.join(masaustu, f"Kontrol_Raporu_{grup}_{tarih}.xlsx")
        wb.save(excel_dosya)
        log(f"Excel rapor kaydedildi: {excel_dosya}", "ok")

    except Exception as e:
        log(f"Excel rapor hatası: {e}", "error")


# ========== ANA TARAMA ==========
def tara(grup="A", donem_offset=0, bastan_basla=False, kontrol_edilmisleri_atla=False,
         acik_receteden=False, goto_recete_no=None):
    stop_temizle()

    log(f"{'='*50}", "header")
    log(f"REÇETE TARAMA - {grup} Grubu (offset: {donem_offset})", "header")
    log(f"Durdurma: Escape tuşu veya Durdur butonu", "info")
    log(f"{'='*50}", "header")

    # Medula bul
    medula = medula_bul()
    if not medula:
        log("MEDULA penceresi bulunamadı! Tanı bilgisi toplanıyor...", "error")
        medula_bul(diag=True)
        log("Çözüm: BotanikEOS açık mı? Medula sekmesi yüklü mü? Pencere minimize/gizli olabilir.", "warn")
        return

    # Oturum kontrol
    medula = oturum_kontrol_ve_baglan(medula)
    if not medula:
        return

    log("Medula bağlı ve oturum aktif", "ok")

    if durduruldu_mu():
        log("DURDURULDU", "warn")
        return

    # Renkli reçete listesini yükle
    renkli_liste_yukle()

    # === BAŞLANGIÇ NOKTASI BELİRLE ===
    son_kayitli = _son_recete_al(grup)

    # "Kontrol edilmişleri atla" seçeneği: DB'den kontrol edilmiş reçeteleri al
    _kontrol_edilmis = set()
    if kontrol_edilmisleri_atla:
        try:
            _tmp_conn, _tmp_cur = db_baglanti()
            _tmp_cur.execute(
                "SELECT DISTINCT recete_no FROM kontrol_sonuclari WHERE grup = ?", (grup,))
            _kontrol_edilmis = {row[0] for row in _tmp_cur.fetchall()}
            _tmp_conn.close()
            if _kontrol_edilmis:
                log(f"DB'de {len(_kontrol_edilmis)} kontrol edilmiş reçete var - bunlar atlanacak", "info")
        except Exception:
            pass

    # === BAŞLANGIÇ MODU ===
    if acik_receteden:
        # Ekranda açık olan reçeteden başla (navigasyon yapma)
        if element_bul(medula, "f:tbl1"):
            mevcut = recete_no_oku(medula)
            log(f"Açık reçeteden başlanıyor: {mevcut}", "ok")
        else:
            log("Ekranda açık reçete yok! Navigasyon yapılıyor...", "warn")
            if not medula_navigasyon(medula, grup, donem_offset):
                return
    elif goto_recete_no:
        # Belirli bir reçeteye git (combobox'tan seçildi)
        log(f"Reçete {goto_recete_no}'ya gidiliyor...", "info")
        if _recete_sorgu_ile_git(medula, goto_recete_no):
            log(f"Reçete {goto_recete_no}'dan devam ediliyor", "ok")
        else:
            log(f"Reçete {goto_recete_no}'ya gidilemedi, navigasyon yapılıyor", "warn")
            if not medula_navigasyon(medula, grup, donem_offset):
                return
    elif element_bul(medula, "f:tbl1") and not bastan_basla:
        log("Reçete sayfası zaten açık", "ok")
    else:
        if not medula_navigasyon(medula, grup, donem_offset):
            return

    conn, cur = db_baglanti()

    sayac = 0
    onceki_recete = None
    tekrar_sayaci = 0
    gorulen_receteler = set()  # Döngü tespiti için (A→B→A sorunu)
    toplam_ilac = 0
    toplam_sorun = 0
    toplam_renkli_sorun = 0
    yeni_kural = 0
    rapor_satirlari = []  # Excel/GUI raporu için veri toplama

    while sayac < 300:
        # === STOP KONTROL ===
        if durduruldu_mu():
            log("", "header")
            log("TARAMA DURDURULDU (kullanıcı tarafından)", "warn")
            break

        # === SİSTEM HATASI KONTROL ===
        if sistem_dusmus_mu(medula):
            log("SİSTEM HATASI TESPİT EDİLDİ - Medula yeniden başlatılıyor...", "error")
            medula = medula_yeniden_baslat()
            if not medula:
                log("Medula yeniden başlatılamadı - tarama durduruluyor", "error")
                break
            # Navigasyonu tekrar yap
            if not medula_navigasyon(medula, grup, donem_offset):
                log("Navigasyon başarısız - tarama durduruluyor", "error")
                break
            # Son kaldığımız reçeteye dönmeye çalış
            log(f"Sistem hatası sonrası yeniden başlatıldı, tarama devam ediyor", "ok")
            continue

        # === GOTO KONTROL (GUI'den reçete seçildi mi?) ===
        goto_recete = goto_recete_kontrol()
        if goto_recete:
            log(f"GUI'den geri dönme isteği: {goto_recete}", "info")
            if _recete_sorgu_ile_git(medula, goto_recete):
                log(f"Reçete {goto_recete}'ya geri dönüldü", "ok")
                onceki_recete = None
                tekrar_sayaci = 0
            else:
                log(f"Reçete {goto_recete}'ya gidilemedi, devam ediliyor", "warn")

        sayac += 1

        # ═══ TEK SEFERDE TÜM BİLGİLERİ TOPLA (descendants() 1 kez çağrılır) ═══
        toplu = recete_tum_bilgi_topla(medula)
        recete_no = toplu["recete_no"]

        # Tekrar kontrolü (art arda aynı reçete)
        if recete_no == onceki_recete:
            tekrar_sayaci += 1
            if tekrar_sayaci >= 3:
                log(f"Reçete değişmiyor ({recete_no}) - tarama bitti", "info")
                break
            element_tikla(medula, "f:buttonSonraki")
            time.sleep(2)
            continue
        else:
            tekrar_sayaci = 0
            onceki_recete = recete_no

        # Döngü tespiti (A→B→A→B gibi) — aynı reçete ikinci kez görüldüyse liste başa döndü
        if recete_no and recete_no in gorulen_receteler:
            log(f"Reçete {recete_no} 2. kez görüldü — liste başa döndü, tarama bitti", "info")
            break
        if recete_no:
            gorulen_receteler.add(recete_no)

        if not recete_no:
            # Popup engellemiş olabilir — kapat + tekrar dene
            popup_kapat()
            time.sleep(1)
            toplu = recete_tum_bilgi_topla(medula)
            recete_no = toplu["recete_no"]
            if not recete_no:
                log("Reçete no okunamadı - tarama bitti", "info")
                break

        # Daha önce kontrol edilmiş reçeteyi atla (kaldığı yerden devam)
        if recete_no in _kontrol_edilmis:
            log(f"  {recete_no} zaten kontrol edilmiş - atlanıyor", "info")
            element_tikla(medula, "f:buttonSonraki")
            time.sleep(0.5)
            continue

        recete_turu = toplu["recete_turu"]
        recete_alt_turu = toplu.get("recete_alt_turu", "Ayaktan")
        uyari_kodlari = toplu["uyari_kodlari"]
        recete_teshisleri = toplu["teshisler"]
        recete_aciklamalari = toplu["recete_aciklamalari"]
        doktor_uzmanligi = toplu["doktor_uzmanligi"]
        doktor_adi = toplu.get("doktor_adi", "")
        erecete_no = toplu.get("erecete_no", "")
        fatura_turu = toplu.get("fatura_turu", "")
        hasta_adi = toplu.get("hasta_adi", "")

        log(f"", "info")
        print(f"[RECETE_BASLADI] {sayac}", flush=True)
        log(f"━━━ Reçete #{sayac}: {recete_no} ━━━", "header")

        # OPT A: Rapor cache — her reçete başında temizle
        rapor_cache = {}

        # === ADIM 1: REÇETE TÜRÜ + ALT TÜR + DOKTOR ===
        tur_bilgi = f"[{recete_turu}]" if recete_turu != "Normal" else "[Beyaz]"
        if recete_alt_turu and recete_alt_turu != "Ayaktan":
            tur_bilgi += f" ({recete_alt_turu})"
        doktor_str = doktor_uzmanligi
        if doktor_adi and doktor_adi != doktor_uzmanligi:
            doktor_str = f"{doktor_uzmanligi} ({doktor_adi})" if doktor_uzmanligi else doktor_adi
        if doktor_str:
            log(f"  Tür: {tur_bilgi} | Doktor: {doktor_str}", "info")
        else:
            log(f"  Tür: {tur_bilgi}", "info")
        # Ek bilgiler (varsa)
        ek_bilgiler = []
        if hasta_adi:
            ek_bilgiler.append(f"Hasta: {hasta_adi}")
        if erecete_no:
            ek_bilgiler.append(f"eRx: {erecete_no}")
        if fatura_turu:
            ek_bilgiler.append(f"Kapsam: {fatura_turu}")
        if ek_bilgiler:
            log(f"  {' | '.join(ek_bilgiler)}", "info")

        # === ADIM 2: RENKLİ REÇETE KONTROLÜ ===
        renkli_sonuc = "-"
        renkli_aciklama = ""
        if recete_turu != "Normal":
            sorun, mesaj = renkli_recete_kontrol(recete_no, recete_turu)
            renkli_sonuc = "Uygun Değil" if sorun else "Uygun"
            renkli_aciklama = mesaj
            if sorun:
                toplam_renkli_sorun += 1
                log(f"  {mesaj}", "sorun")
            else:
                log(f"  {mesaj}", "ok")

        # === ADIM 2B: UYARI KODLARI TESPİT (kontrol en sonda yapılacak) ===
        recete_teshisleri_input = toplu.get("recete_teshisleri_input", [])
        if uyari_kodlari:
            # Bypass kontrolü
            try:
                from kontrol_kurallari import get_kontrol_ayarlari
                _uk_ayarlar = get_kontrol_ayarlari()
                filtreli_uk = []
                for uk in uyari_kodlari:
                    uk_kod = uk.get("kod", "") if isinstance(uk, dict) else str(uk)
                    uk_aciklama = uk.get("aciklama", "") if isinstance(uk, dict) else ""
                    if _uk_ayarlar.kontrol_aktif_mi("uyari_kodu", uyari_kodu=uk_kod):
                        filtreli_uk.append(uk)
                    else:
                        log(f"    [BYPASS] Uyarı kodu {uk_kod}: {uk_aciklama} → kontrol pasif (ayar)", "info")
                uyari_kodlari = filtreli_uk
            except Exception:
                pass
            if uyari_kodlari:
                log(f"  {len(uyari_kodlari)} uyarı kodu tespit edildi → en son kontrol edilecek", "info")

        # Teşhis ve açıklama bilgisi log
        if recete_teshisleri_input:
            log(f"  Reçete teşhis: {', '.join(recete_teshisleri_input[:3])}", "info")
        if recete_teshisleri:
            log(f"  Tanı listesi: {', '.join(recete_teshisleri[:3])}", "info")
        if recete_aciklamalari:
            log(f"  Açıklamalar: {', '.join(recete_aciklamalari[:2])}", "info")

        # === ADIM 2C: 217 KODU İÇİN E-REÇETE AÇIKLAMASINI PROAKTİF OKU ===
        # 217 hekimin TSAT/Ferritin değerlerini E-Reçete açıklamasına yazdığını
        # belirtir. Reçete sayfasındaki Açıklama Listesi (form1:tableEx1) bu
        # değerleri her zaman içermez; gerçek metin E-Reçete Görüntüle ekranındadır.
        # Eagerly çekiyoruz çünkü ESA SUT kontrolünden ÖNCE lazım. Sonuç cache'lenir,
        # uyari_kodu_kontrol fallback'i (line ~4994) tekrar çekmeyecek.
        erecete_metni = ""
        erecete_aciklama_listesi = []
        erecete_tani_listesi = []
        erecete_okundu = False
        has_217_in_uyari = any(
            str(uk.get("kod", "")).strip() == "217"
            for uk in uyari_kodlari if isinstance(uk, dict)
        )
        if has_217_in_uyari:
            log(f"  [217] Uyarı kodu 217 var — E-Reçete Görüntüle açılıyor (TSAT/Ferritin için)", "info")
            try:
                erecete = erecete_aciklama_oku(medula)
                if erecete:
                    erecete_metni = erecete.get("tum_metin", "") or ""
                    erecete_aciklama_listesi = erecete.get("aciklamalar", []) or []
                    erecete_tani_listesi = erecete.get("tanilar", []) or []
                    erecete_okundu = True
                    if erecete_metni:
                        # Lab değerleri (Hgb/Ferritin/TSAT) bu metinden okunacak — kullanıcının
                        # doktorun gerçekte ne yazdığını görmesi için snippet'i daha uzun ver.
                        ozet = erecete_metni[:300].replace("\n", " ")
                        log(f"  [217] E-Reçete açıklama okundu ({len(erecete_metni)} kr): {ozet}...", "info")
                        # Lab anahtar kelimeleri var mı? Hızlı sağlık tespiti
                        _ml = erecete_metni.lower()
                        _bulunan = [k for k in ('ferritin', 'tsat', 'transferrin', 'satürasyon',
                                                  'saturasyon', 'doygunluk', 'hgb', 'hemoglobin')
                                    if k in _ml]
                        if _bulunan:
                            log(f"  [217] Tespit edilen lab anahtarları: {', '.join(_bulunan)}", "info")
                        else:
                            log(f"  [217] DİKKAT: Metinde lab anahtar kelimesi bulunamadı (Ferritin/TSAT/Hb yok)", "warn")
                    else:
                        log(f"  [217] E-Reçete açıldı ama metin toplanamadı", "warn")
            except Exception as e:
                log(f"  [217] E-Reçete okuma hatası: {e}", "warn")

        # === ADIM 3: İLAÇ TABLOSU KONTROLÜ (zaten toplandı, doz karşılaştırma dahil) ===
        ilaclar = toplu["ilaclar"]
        # Reçete açıklamaları + cache'i + uyarı kodlarını + e-Reçete metnini her ilaca ekle
        _cache = toplu.get("_aid_cache", {})
        for _ilac in ilaclar:
            _ilac["_recete_aciklamalari"] = recete_aciklamalari
            _ilac["_aid_cache"] = _cache
            _ilac["_uyari_kodlari"] = uyari_kodlari
            _ilac["_erecete_aciklama_metni"] = erecete_metni
            _ilac["_erecete_aciklama_listesi"] = erecete_aciklama_listesi

        # İlaç listesi boşsa sayfa tam yüklenmemiş olabilir — 2 saniye bekleyip yeniden dene
        if not ilaclar:
            log("  İlaç okunamadı, 2 sn bekleyip tekrar deneniyor...", "warn")
            time.sleep(2)
            try:
                toplu_retry = recete_tum_bilgi_topla(medula)
                if toplu_retry.get("ilaclar"):
                    toplu = toplu_retry
                    ilaclar = toplu["ilaclar"]
                    _cache = toplu.get("_aid_cache", {})
                    for _ilac in ilaclar:
                        _ilac["_recete_aciklamalari"] = recete_aciklamalari
                        _ilac["_aid_cache"] = _cache
                        _ilac["_uyari_kodlari"] = uyari_kodlari
                        _ilac["_erecete_aciklama_metni"] = erecete_metni
                        _ilac["_erecete_aciklama_listesi"] = erecete_aciklama_listesi
                    log(f"  [OKU ] Retry başarılı: {len(ilaclar)} ilaç bulundu", "info")
            except Exception as retry_err:
                log(f"  [UYARI] Retry hatası: {retry_err}", "warn")

        if not ilaclar:
            log("  İlaç okunamadı (retry sonrası da boş), sonrakine geçiliyor", "warn")
            rapor_satirlari.append({
                "recete_no": recete_no, "recete_turu": recete_turu,
                "ilac_adi": "", "etkin_madde": "", "rapor_kodu": "", "msj": "",
                "renkli_kontrol": renkli_sonuc, "rapor_kontrol": "-",
                "sut_kontrol": "-", "sonuc": "UYARI", "aciklama": "İlaç okunamadı"
            })
        else:
            log(f"  {len(ilaclar)} ilaç bulundu - detaylı kontrol başlıyor:", "info")

            # AA ilaçları (raporsuz + mesajsız) say ama detaylı kontrole sokma
            aa_sayisi = sum(1 for il in ilaclar if not il.get("rapor_kodu") and il.get("msj", "") != "var")
            kontrol_gereken = len(ilaclar) - aa_sayisi
            if aa_sayisi > 0:
                log(f"  {aa_sayisi} raporsuz/mesajsız ilaç atlandı, {kontrol_gereken} ilaç kontrol ediliyor:", "info")

            for idx, ilac in enumerate(ilaclar):
                toplam_ilac += 1
                rapor_var = bool(ilac.get("rapor_kodu", ""))
                msj_var = (ilac.get("msj", "") == "var")

                # AA: Raporsuz + Mesajsız → doğrudan geç (detaylı kontrole girme)
                if not rapor_var and not msj_var:
                    rapor_satirlari.append({
                        "recete_no": recete_no, "recete_turu": recete_turu,
                        "ilac_adi": ilac["ilac_adi"], "etkin_madde": ilac.get("etkin_madde", ""),
                        "rapor_kodu": "", "msj": "",
                        "renkli_kontrol": renkli_sonuc if idx == 0 else "-",
                        "rapor_kontrol": "-", "sut_kontrol": "-",
                        "sonuc": "GECİLDİ", "aciklama": "Raporsuz, mesajsız"
                    })
                    continue

                # Yeni etkin madde ise DB'ye kaydet
                etkin = ilac.get("etkin_madde", "")
                kural = db_kural_bul(cur, etkin)
                if not kural and etkin:
                    rapor_kodu_db = ilac.get("rapor_kodu", "")
                    raporlu = 1 if rapor_kodu_db else 0
                    tip = "rapor_kontrolu" if raporlu else "raporsuz_verilebilir"
                    if db_kaydet(cur, conn, etkin, ilac.get("sgk_kodu", ""),
                                rapor_kodu_db, raporlu, tip,
                                f"Otomatik: {ilac['ilac_adi']}"):
                        yeni_kural += 1
                    log(f"    [YENİ] {ilac['ilac_adi'][:40]} -> DB'ye eklendi", "yeni")

                # BB/CC/DD karar ağacı
                gercek_idx = ilac.get("medula_satir_idx", idx)
                t_ilac = time.time()

                # ── İlaç başlık bloğu ──
                _ilac_adi_tam = ilac.get("ilac_adi", "")
                _rk = ilac.get("rapor_kodu", "") or "yok"
                _msj = "VAR" if ilac.get("msj", "") == "var" else "yok"
                log(f"    ╔══ İLAÇ #{idx+1}: {_ilac_adi_tam} ══", "header")
                log(f"    ║ Rapor kodu: {_rk} | Mesaj: {_msj}", "info")

                satir = ilac_detayli_kontrol(
                    medula, cur, conn, grup, recete_no, recete_turu,
                    ilac, gercek_idx, renkli_sonuc, recete_teshisleri=recete_teshisleri,
                    doktor_uzmanligi=doktor_uzmanligi,
                    recete_alt_turu=recete_alt_turu,
                    rapor_cache=rapor_cache
                )

                # ── Nihai görüş bloğu: önce gerekçe, sonra sonuç ──
                _sonuc = satir.get("sonuc", "")
                _aciklama_tam = satir.get("aciklama", "")
                if _aciklama_tam:
                    log(f"    ║ Gerekçe  : {_aciklama_tam}", "info")
                if _sonuc == "UYGUN":
                    log(f"    ║ NİHAİ    : ✓ UYGUN", "ok")
                elif _sonuc == "UYGUNSUZ":
                    log(f"    ║ NİHAİ    : ✗ UYGUNSUZ", "sorun")
                elif _sonuc == "GECİLDİ":
                    log(f"    ║ NİHAİ    : ⊘ GEÇİLDİ", "info")
                else:
                    log(f"    ║ NİHAİ    : ? {_sonuc}", "warn")
                log(f"    ╚══ İlaç kontrol: {time.time()-t_ilac:.1f}s ══", "info")

                # Renkli kontrol bilgisini ilk ilaca ekle
                if idx == 0:
                    satir["renkli_kontrol"] = renkli_sonuc

                rapor_satirlari.append(satir)

                # Sonuç sayacı
                if satir["sonuc"] == "UYGUNSUZ":
                    toplam_sorun += 1

        # ═══════════════════════════════════════════════════════════════
        # === ADIM 5: UYARI KODLARI KONTROLÜ ===
        # Sıralı arama:
        #   1. Reçete teşhisi (INPUT + tanı listesi)
        #   2. Rapor teşhisi + rapor açıklamaları (yoksa rapor sayfasını aç)
        #   3. E-Reçete sayfası açıklamaları (hala eşleşme yoksa)
        # ═══════════════════════════════════════════════════════════════
        if uyari_kodlari:
            # Tespit edilen uyarı kodlarını DB'ye kaydet (combobox için)
            for uk in uyari_kodlari:
                uk_kod = uk.get("kod", "") if isinstance(uk, dict) else str(uk)
                uk_acik = uk.get("aciklama", "") if isinstance(uk, dict) else ""
                if uk_kod:
                    uyari_kodu_kaydet(cur, conn, uk_kod, uk_acik)

            log(f"  ═══ UYARI KODU KONTROLÜ ({len(uyari_kodlari)} adet) ═══", "header")

            # ── KAYNAK 1: Reçete teşhisi ──
            kaynak_teshis = recete_teshisleri_input + recete_teshisleri
            if kaynak_teshis:
                log(f"    [1] Reçete teşhis: {', '.join(kaynak_teshis[:3])}", "info")

            # İlk kontrol: reçete teşhisi + reçete açıklamaları
            uk_sonuclar = uyari_kodu_kontrol(uyari_kodlari, kaynak_teshis, recete_aciklamalari, [])
            eslesmeyen = [uks for uks in uk_sonuclar if uks["durum"] == "UYGUNSUZ"]

            # ── KAYNAK 2: Rapor teşhisi + rapor açıklamaları ──
            rapor_aciklamalari_toplam = []
            rapor_tanilari_toplam = []
            if eslesmeyen:
                # Önce ilaç kontrollerinde okunmuş rapor verisi var mı?
                for s in rapor_satirlari:
                    rv = s.get("_rapor_verisi", {})
                    if rv:
                        rapor_aciklamalari_toplam.extend(rv.get("aciklamalar", []))
                        rapor_tanilari_toplam.extend(rv.get("tanilar", []))
                        rapor_tanilari_toplam.extend(rv.get("icd_kodlari", []))
                        # tum_metin: filtre uygulanmamış tam rapor metni (uyarı kodu eşleşmesi için)
                        tm = rv.get("tum_metin", "")
                        # Çöp metin kontrolü — rapor sayfası açılamadıysa reçete sayfası metni gelir
                        if tm and "Fatura Sonland" not in tm and "Reçete No" not in tm:
                            rapor_aciklamalari_toplam.append(tm)

                # Rapor verisi yoksa (hiçbir ilaç DD değilse) → rapor sayfasını aç
                if not rapor_tanilari_toplam and not rapor_aciklamalari_toplam:
                    # İlk raporlu ilacın checkbox'ını seç ve rapor aç
                    raporlu_idx = None
                    for il in ilaclar:
                        if il.get("rapor_kodu"):
                            raporlu_idx = il.get("medula_satir_idx", 0)
                            break
                    if raporlu_idx is not None:
                        log(f"    [2] Rapor verisi yok — rapor sayfasından okunuyor...", "info")
                        rapor_verisi = rapor_ac_oku_geri_don(medula, raporlu_idx)
                        if rapor_verisi:
                            rapor_tanilari_toplam.extend(rapor_verisi.get("tanilar", []))
                            rapor_tanilari_toplam.extend(rapor_verisi.get("icd_kodlari", []))
                            rapor_aciklamalari_toplam.extend(rapor_verisi.get("aciklamalar", []))
                            if rapor_verisi.get("tum_metin"):
                                rapor_aciklamalari_toplam.append(rapor_verisi["tum_metin"])

                if rapor_tanilari_toplam:
                    log(f"    [2] Rapor teşhis: {', '.join(rapor_tanilari_toplam[:2])}", "info")
                if rapor_aciklamalari_toplam:
                    log(f"    [2] Rapor açıklama: {rapor_aciklamalari_toplam[0][:60]}", "info")

                # Tekrar kontrol: reçete teşhis + rapor verileriyle
                tum_teshisler = kaynak_teshis + rapor_tanilari_toplam
                tum_aciklamalar = recete_aciklamalari + rapor_aciklamalari_toplam
                uk_sonuclar = uyari_kodu_kontrol(uyari_kodlari, tum_teshisler, tum_aciklamalar, [])
                eslesmeyen = [uks for uks in uk_sonuclar if uks["durum"] == "UYGUNSUZ"]

            # ── KAYNAK 3: E-Reçete sayfası açıklamaları (hala eşleşmeyen varsa) ──
            if eslesmeyen:
                # 217 için zaten eagerly okuduysak tekrar açma — yoksa şimdi aç
                if erecete_okundu:
                    log(f"    [3] {len(eslesmeyen)} uyarı kodu eşleşmedi — E-Reçete metni cache'den kullanılıyor", "info")
                    erecete = {
                        "tum_metin": erecete_metni,
                        "aciklamalar": erecete_aciklama_listesi,
                        "tanilar": erecete_tani_listesi,
                    }
                else:
                    log(f"    [3] {len(eslesmeyen)} uyarı kodu eşleşmedi — E-Reçete sayfasından okunuyor...", "warn")
                    erecete = erecete_aciklama_oku(medula)
                if erecete:
                    tum_teshisler_ek = kaynak_teshis + rapor_tanilari_toplam
                    tum_aciklamalar_ek = recete_aciklamalari + rapor_aciklamalari_toplam
                    if erecete.get("tanilar"):
                        tum_teshisler_ek.extend(erecete["tanilar"])
                        log(f"    [3] E-Reçete tanı: {erecete['tanilar'][0][:60]}", "info")
                    if erecete.get("aciklamalar"):
                        tum_aciklamalar_ek.extend(erecete["aciklamalar"])
                    if erecete.get("tum_metin"):
                        tum_aciklamalar_ek.append(erecete["tum_metin"])
                    # Son kontrol
                    uk_sonuclar = uyari_kodu_kontrol(uyari_kodlari, tum_teshisler_ek, tum_aciklamalar_ek, [])

            # Sonuçları logla (çerçeveli)
            for uks in uk_sonuclar:
                kod = uks["kod"]
                acik = uks["aciklama"]
                ilac = uks.get("ilac_adi", "")
                oran = uks.get("eslesen_oran", 0)
                kaynak = uks.get("eslesen_kaynak", "")
                eslesen = uks.get("eslesen_metin", "")
                if uks["durum"] == "UYGUN":
                    log(f"    ┌─── Uyarı Kodu {kod} ───", "info")
                    log(f"    │ Uyarı  : {acik} => {ilac}", "info")
                    log(f"    │ Kaynak : {kaynak}", "info")
                    if eslesen:
                        log(f"    │ Buluna : {eslesen[:70]}", "info")
                    log(f"    │ Sonuç  : ✓ UYGUN (%{int(oran*100)} eşleşme)", "ok")
                    log(f"    └───────────────────────────────────", "info")
                else:
                    log(f"    ┌─── Uyarı Kodu {kod} ───", "info")
                    log(f"    │ Uyarı  : {acik} => {ilac}", "info")
                    ozel = uks.get("_ozel_kural", "")
                    if ozel:
                        log(f"    │ Kural  : {ozel}", "info")
                    else:
                        log(f"    │ Aranan : {acik}", "info")
                    log(f"    │ Sonuç  : ✗ EŞLEŞME BULUNAMADI (%{int(oran*100)})", "sorun")
                    log(f"    └───────────────────────────────────", "info")
                    toplam_sorun += 1

        # _rapor_verisi iç kullanım alanını temizle (Excel'e yazılmasın)
        for s in rapor_satirlari:
            s.pop("_rapor_verisi", None)

        # Son kontrol edilen reçeteyi kaydet (kaldığı yerden devam için)
        if recete_no:
            _son_recete_kaydet(grup, recete_no, sayac)

        # Sonraki reçete
        t_nav = time.time()
        sonraki_ok = False
        # Önce Medula penceresini öne getir (BotanikEOS öne kaymış olabilir)
        try:
            medula.set_focus()
            time.sleep(0.3)
        except Exception:
            pass
        # 1. Direkt Sonraki
        if element_tikla(medula, "f:buttonSonraki"):
            sonraki_ok = True
        # 1b. Popup'ı kapat (etken madde çakışması vs.) + tekrar Sonraki
        if not sonraki_ok and popup_kapat():
            if element_tikla(medula, "f:buttonSonraki"):
                sonraki_ok = True
        # 2. Medula referansını yeniden bul + tekrar Sonraki
        if not sonraki_ok:
            yeni_medula = medula_bul()
            if yeni_medula:
                medula = yeni_medula
                try:
                    medula.set_focus()
                    time.sleep(0.3)
                except Exception:
                    pass
                log("    [OKU ] Medula penceresi yeniden bulundu", "info")
                if element_tikla(medula, "f:buttonSonraki"):
                    sonraki_ok = True
        # 3. Alt sayfadaysak Geri Dön + Sonraki
        if not sonraki_ok:
            element_tikla(medula, "form1:buttonGeriDon")
            time.sleep(0.8)
            if element_tikla(medula, "f:buttonSonraki"):
                sonraki_ok = True
        # 4. Son çare: Escape + Sonraki
        if not sonraki_ok:
            pyautogui.press("escape")
            time.sleep(0.5)
            if element_tikla(medula, "f:buttonSonraki"):
                sonraki_ok = True
        # 5. Hiçbiri işe yaramadı — tanı + dur
        if not sonraki_ok:
            try:
                log("    [TANI] Ekrandaki butonlar:", "warn")
                for d in medula.descendants(control_type="Button"):
                    try:
                        cap = (d.window_text() or "").strip()
                        aid = d.element_info.automation_id or ""
                        if cap or aid:
                            log(f"      - caption='{cap}' | auto_id='{aid}'", "info")
                    except Exception:
                        continue
            except Exception as diag_err:
                log(f"    [TANI] Buton listeleme hatası: {diag_err}", "warn")
            # Son çare: Reçete Listesi'ne dönüp navigasyonu yeniden çalıştır,
            # ardından son işlenen reçeteden sonrakine sorgu ile zıpla.
            log(f"  [KURTARMA] Sonraki bulunamadı — liste yeniden yükleniyor (son: {recete_no})", "warn")
            try:
                if medula_navigasyon(medula, grup, donem_offset):
                    if recete_no:
                        # Son işlenen reçeteden sonra başlayacak şekilde — şimdilik
                        # listeden devam ediyoruz; aynı reçete tekrar görülürse döngü
                        # tespiti (gorulen_receteler) zaten break atacak.
                        log(f"  [KURTARMA] Liste yeniden yüklendi, taramaya devam", "ok")
                        continue
            except Exception as recover_err:
                log(f"  [KURTARMA] Yeniden navigasyon hatası: {recover_err}", "warn")
            log(f"Sonraki butonu bulunamadı + kurtarma başarısız — taranan: {sayac} reçete (son: {recete_no})", "warn")
            break

        # Sayfa yüklenene kadar bekle (optimize: tek exists çağrısı)
        try:
            cw = medula.child_window(auto_id="f:tbl1", found_index=0)
            cw.exists(timeout=2)
        except:
            pass
        log(f"  [SÜRE] Navigasyon (Sonraki): {time.time()-t_nav:.1f}s", "info")

    # Özet
    try:
        cur.execute("SELECT COUNT(*) FROM etkin_madde_kurallari")
        toplam_db = cur.fetchone()[0]
    except:
        toplam_db = "?"

    # Kontrol sonuçları özeti (mevcut tarama — rapor_satirlari'ndan)
    toplam_uygun = 0
    toplam_uygunsuz = 0
    toplam_gecildi = 0
    for s in rapor_satirlari:
        sonuc_v = s.get("sonuc", "")
        if sonuc_v == "UYGUN":
            toplam_uygun += 1
        elif sonuc_v == "UYGUNSUZ":
            toplam_uygunsuz += 1
        elif sonuc_v == "GECİLDİ":
            toplam_gecildi += 1

    conn.close()

    log(f"", "header")
    log(f"{'='*50}", "header")
    log(f"TARAMA TAMAMLANDI", "header")
    log(f"  Taranan reçete    : {sayac}", "info")
    log(f"  Toplam ilaç       : {toplam_ilac}", "info")
    log(f"  Uygun             : {toplam_uygun}", "ok")
    log(f"  Uygunsuz          : {toplam_uygunsuz}", "error" if toplam_uygunsuz > 0 else "info")
    log(f"  Geçildi           : {toplam_gecildi}", "info")
    log(f"  Sorunlu ilaç      : {toplam_sorun}", "error" if toplam_sorun > 0 else "info")
    log(f"  Renkli reçete sor.: {toplam_renkli_sorun}", "error" if toplam_renkli_sorun > 0 else "info")
    log(f"  Yeni öğrenilen    : {yeni_kural}", "info")
    log(f"  DB toplam kural   : {toplam_db}", "info")
    log(f"{'='*50}", "header")

    # Rapor kaydet (JSON + Excel)
    rapor_kaydet(rapor_satirlari, grup)

    stop_temizle()


if __name__ == "__main__":
    grup = sys.argv[1] if len(sys.argv) > 1 else "A"
    donem_offset = int(sys.argv[2]) if len(sys.argv) > 2 else 0
    bastan = "--bastan" in sys.argv
    acik_receteden = "--acik" in sys.argv
    kontrol_edilmisleri_atla = "--atla" in sys.argv
    goto_recete_no = None
    if "--goto" in sys.argv:
        idx = sys.argv.index("--goto")
        if idx + 1 < len(sys.argv):
            goto_recete_no = sys.argv[idx + 1]
    tara(grup=grup, donem_offset=donem_offset, bastan_basla=bastan,
         kontrol_edilmisleri_atla=kontrol_edilmisleri_atla,
         acik_receteden=acik_receteden, goto_recete_no=goto_recete_no)
