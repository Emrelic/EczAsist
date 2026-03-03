"""
Prim Raporlama Modulu
Prim verilen urunlerin satis raporlamasi - personel bazli
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import calendar
from tkcalendar import DateEntry
import threading
from decimal import Decimal

from botanik_db import BotanikDB

logger = logging.getLogger(__name__)


class PrimRaporlamaGUI:
    """Prim Raporlama Modulu"""

    def __init__(self, parent, ana_menu_callback=None):
        self.parent = parent
        self.ana_menu_callback = ana_menu_callback

        self.parent.title("Prim Raporlama")
        self.parent.state('zoomed')
        self.parent.configure(bg='#ECEFF1')

        # Database
        self.db = None
        self._db_baglan()

        # Data
        self.rapor_verileri = []
        self.personel_listesi = []

        # UI
        self._arayuz_olustur()

        # Personel listesini yukle
        self._personel_yukle()

    def _db_baglan(self):
        """Veritabanina baglan"""
        try:
            self.db = BotanikDB()
            if not self.db.baglan():
                self.db = None
                logger.warning("Veritabanina baglanamadi")
        except Exception as e:
            logger.error(f"DB baglanti hatasi: {e}")
            self.db = None

    def _arayuz_olustur(self):
        """Ana arayuzu olustur"""
        # Header
        header = tk.Frame(self.parent, bg='#5D4037', height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        header_inner = tk.Frame(header, bg='#5D4037')
        header_inner.pack(fill=tk.X, padx=15)

        tk.Label(header_inner, text="Prim Raporlama",
                 bg='#5D4037', fg='white',
                 font=('Segoe UI', 14, 'bold')).pack(side=tk.LEFT, pady=12)

        if self.ana_menu_callback:
            tk.Button(header_inner, text="Ana Menu",
                      bg='#8D6E63', fg='white', font=('Segoe UI', 10),
                      relief=tk.FLAT, padx=15, pady=4,
                      command=self._ana_menuye_don).pack(side=tk.RIGHT, pady=10)

        # Main frame
        main_frame = tk.Frame(self.parent, bg='#ECEFF1', padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Filtre paneli
        self._filtre_panel_olustur(main_frame)

        # Tablo
        self._tablo_olustur(main_frame)

        # Ozet paneli
        self._ozet_panel_olustur(main_frame)

    def _filtre_panel_olustur(self, parent):
        """Filtre paneli"""
        filtre_frame = tk.LabelFrame(parent, text=" Filtreler ",
                                     font=('Segoe UI', 10, 'bold'),
                                     bg='#FFF3E0', padx=10, pady=8)
        filtre_frame.pack(fill=tk.X, pady=(0, 8))

        row = tk.Frame(filtre_frame, bg='#FFF3E0')
        row.pack(fill=tk.X)

        # Onceki Ay butonu
        tk.Button(row, text="<< Onceki Ay", font=('Segoe UI', 9),
                  bg='#8D6E63', fg='white', relief=tk.FLAT, padx=8, pady=3,
                  command=self._onceki_ay).pack(side=tk.LEFT, padx=(0, 5))

        # Baslangic tarihi
        tk.Label(row, text="Baslangic:", font=('Segoe UI', 10),
                 bg='#FFF3E0').pack(side=tk.LEFT, padx=(5, 3))

        bugun = date.today()
        ay_basi = bugun.replace(day=1)

        self.baslangic_entry = DateEntry(row, width=12, background='#5D4037',
                                         foreground='white', borderwidth=1,
                                         date_pattern='dd/MM/yyyy',
                                         font=('Segoe UI', 10))
        self.baslangic_entry.set_date(ay_basi)
        self.baslangic_entry.pack(side=tk.LEFT, padx=(0, 10))

        # Bitis tarihi
        tk.Label(row, text="Bitis:", font=('Segoe UI', 10),
                 bg='#FFF3E0').pack(side=tk.LEFT, padx=(5, 3))

        self.bitis_entry = DateEntry(row, width=12, background='#5D4037',
                                     foreground='white', borderwidth=1,
                                     date_pattern='dd/MM/yyyy',
                                     font=('Segoe UI', 10))
        self.bitis_entry.set_date(bugun)
        self.bitis_entry.pack(side=tk.LEFT, padx=(0, 10))

        # Sonraki Ay butonu
        tk.Button(row, text="Sonraki Ay >>", font=('Segoe UI', 9),
                  bg='#8D6E63', fg='white', relief=tk.FLAT, padx=8, pady=3,
                  command=self._sonraki_ay).pack(side=tk.LEFT, padx=(0, 15))

        # Personel secimi
        tk.Label(row, text="Personel:", font=('Segoe UI', 10),
                 bg='#FFF3E0').pack(side=tk.LEFT, padx=(10, 3))

        self.personel_var = tk.StringVar(value="Tumu")
        self.personel_combo = ttk.Combobox(row, textvariable=self.personel_var,
                                           width=20, state='readonly',
                                           font=('Segoe UI', 10))
        self.personel_combo['values'] = ["Tumu"]
        self.personel_combo.pack(side=tk.LEFT, padx=(0, 15))

        # Listele butonu
        self.listele_btn = tk.Button(row, text="Listele",
                                     font=('Segoe UI', 11, 'bold'),
                                     bg='#1565C0', fg='white',
                                     relief=tk.FLAT, padx=20, pady=5,
                                     command=self._listele)
        self.listele_btn.pack(side=tk.LEFT, padx=(10, 5))

        # Excel Aktar butonu
        tk.Button(row, text="Excel Aktar",
                  font=('Segoe UI', 10),
                  bg='#2E7D32', fg='white',
                  relief=tk.FLAT, padx=15, pady=5,
                  command=self._excel_aktar).pack(side=tk.LEFT, padx=(5, 0))

        # Prim Verilen Urun Listesi butonu
        tk.Button(row, text="Prim Verilen Urun Listesi",
                  font=('Segoe UI', 10),
                  bg='#7B1FA2', fg='white',
                  relief=tk.FLAT, padx=15, pady=5,
                  command=self._prim_urunleri_goster).pack(side=tk.LEFT, padx=(10, 0))

        # Ilac disi checkbox
        self.ilac_disi_var = tk.BooleanVar(value=False)
        self.ilac_disi_cb = tk.Checkbutton(row, text="Tum Ilac Disi Urunler",
                                            variable=self.ilac_disi_var,
                                            font=('Segoe UI', 9, 'bold'),
                                            bg='#FFF3E0', fg='#D84315',
                                            activebackground='#FFF3E0',
                                            selectcolor='#FFE0B2')
        self.ilac_disi_cb.pack(side=tk.RIGHT, padx=(10, 5))

        # Durum labeli
        self.durum_label = tk.Label(row, text="Hazir",
                                    font=('Segoe UI', 9), bg='#FFF3E0', fg='#666')
        self.durum_label.pack(side=tk.RIGHT, padx=10)

    def _tablo_olustur(self, parent):
        """Sonuc tablosu"""
        tablo_frame = tk.Frame(parent, bg='#ECEFF1')
        tablo_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        # Sutun tanimlari
        self.sutunlar = [
            ("UrunAdi", "Urun Adi", 240, tk.W),
            ("Barkod", "Barkod", 110, tk.W),
            ("Tarih", "Tarih", 85, tk.CENTER),
            ("Etiket", "Etiket", 75, tk.E),
            ("SatisFiy", "Satis Fiy", 75, tk.E),
            ("Adet", "Adet", 45, tk.CENTER),
            ("Indirimler", "Indirimler", 75, tk.E),
            ("Tutar", "Tutar", 85, tk.E),
            ("AlisFiy", "Alis Fiy", 75, tk.E),
            ("BrutKar", "Brut Kar", 85, tk.E),
            ("BrutKarYuzde", "Brut Kar %", 70, tk.E),
            ("Altidabirlik", "1/6 Prim", 80, tk.E),
            ("AltidabirlikOran", "1/6 Oran %", 75, tk.E),
            ("Personel", "Personel", 130, tk.W),
        ]

        self.tree = ttk.Treeview(tablo_frame,
                                  columns=[c[0] for c in self.sutunlar],
                                  show='headings', height=25)

        # Sutun ayarlari
        for col_id, baslik, width, anchor in self.sutunlar:
            self.tree.heading(col_id, text=baslik,
                              command=lambda c=col_id: self._sutun_sirala(c))
            self.tree.column(col_id, width=width, minwidth=40, anchor=anchor)

        # Renk tagleri - 1/6 prim brut kar % baremine gore 8 kademe
        self.tree.tag_configure('kar_15_ustu', background='#2E7D32', foreground='white')  # Koyu yesil >=15%
        self.tree.tag_configure('kar_12_15', background='#4CAF50', foreground='white')     # Orta yesil 12-15%
        self.tree.tag_configure('kar_10_12', background='#81C784')    # Acik yesil 10-12%
        self.tree.tag_configure('kar_7_10', background='#C5E1A5')     # Limoni sarimsi yesil 7-10%
        self.tree.tag_configure('kar_5_7', background='#FFF9C4')      # Sari 5-7%
        self.tree.tag_configure('kar_4_5', background='#FFE0B2')      # Acik turuncu 4-5%
        self.tree.tag_configure('kar_0_4', background='#FFB74D')      # Turuncu 0-4%
        self.tree.tag_configure('kar_negatif', background='#EF5350', foreground='white')   # Kirmizi <0%
        self.tree.tag_configure('alis_tahmini', foreground='#D32F2F')  # Kirmizi yazi - tahmini alis fiyati
        self.tree.tag_configure('toplam', background='#B0BEC5',
                                font=('Segoe UI', 10, 'bold'))

        # Scrollbar
        vsb = ttk.Scrollbar(tablo_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tablo_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        tablo_frame.grid_rowconfigure(0, weight=1)
        tablo_frame.grid_columnconfigure(0, weight=1)

        # Siralama durumu
        self._siralama_sutun = None
        self._siralama_ters = False

        # F1: Barkod kopyala, F2: Barkod yapistir
        self.tree.bind('<F1>', self._barkod_kopyala)
        self.parent.bind_all('<F2>', self._barkod_yapistir)

    def _ozet_panel_olustur(self, parent):
        """Alt ozet paneli"""
        ozet_frame = tk.LabelFrame(parent, text=" Ozet ",
                                   font=('Segoe UI', 10, 'bold'),
                                   bg='#E3F2FD', padx=10, pady=8)
        ozet_frame.pack(fill=tk.X, pady=(0, 0))

        # Ust satir - genel ozet
        row1 = tk.Frame(ozet_frame, bg='#E3F2FD')
        row1.pack(fill=tk.X, pady=(0, 5))

        self.ozet_labels = {}

        ozet_items = [
            ("satir", "Satir:"),
            ("adet", "Top. Adet:"),
            ("indirim", "Top. İndirim:"),
            ("tutar", "Net Tutar:"),
            ("brut_kar", "Top. Brut Kar:"),
            ("altidabirlik", "Top. 1/6 Prim:"),
        ]

        for key, baslik in ozet_items:
            tk.Label(row1, text=baslik, font=('Segoe UI', 10, 'bold'),
                     bg='#E3F2FD').pack(side=tk.LEFT, padx=(10, 2))
            lbl = tk.Label(row1, text="0", font=('Segoe UI', 10),
                           bg='#E3F2FD', fg='#1565C0')
            lbl.pack(side=tk.LEFT, padx=(0, 15))
            self.ozet_labels[key] = lbl

        # Alt satir - personel bazli ozet
        self.personel_ozet_frame = tk.Frame(ozet_frame, bg='#E3F2FD')
        self.personel_ozet_frame.pack(fill=tk.X, pady=(5, 0))

    # ==================== ISLEMLER ====================

    def _personel_yukle(self):
        """Personel listesini veritabanindan yukle"""
        if not self.db:
            return

        try:
            self.personel_listesi = self.db.personel_listesi_getir()
            degerler = ["Tumu"]
            for p in self.personel_listesi:
                degerler.append(p['PersonelAdiSoyadi'].strip())
            self.personel_combo['values'] = degerler
        except Exception as e:
            logger.error(f"Personel yukleme hatasi: {e}")

    def _onceki_ay(self):
        """Onceki aya git"""
        baslangic = self.baslangic_entry.get_date()
        yeni_baslangic = baslangic - relativedelta(months=1)
        yeni_baslangic = yeni_baslangic.replace(day=1)
        son_gun = calendar.monthrange(yeni_baslangic.year, yeni_baslangic.month)[1]
        yeni_bitis = yeni_baslangic.replace(day=son_gun)

        self.baslangic_entry.set_date(yeni_baslangic)
        self.bitis_entry.set_date(yeni_bitis)

    def _sonraki_ay(self):
        """Sonraki aya git"""
        baslangic = self.baslangic_entry.get_date()
        yeni_baslangic = baslangic + relativedelta(months=1)
        yeni_baslangic = yeni_baslangic.replace(day=1)
        son_gun = calendar.monthrange(yeni_baslangic.year, yeni_baslangic.month)[1]
        yeni_bitis = yeni_baslangic.replace(day=son_gun)

        self.baslangic_entry.set_date(yeni_baslangic)
        self.bitis_entry.set_date(yeni_bitis)

    def _listele(self):
        """Raporu listele"""
        if not self.db:
            messagebox.showerror("Hata", "Veritabani baglantisi yok!")
            return

        self.listele_btn.config(state=tk.DISABLED)
        self.durum_label.config(text="Veriler yukleniyor...")
        self.parent.update()

        threading.Thread(target=self._sorgu_thread, daemon=True).start()

    def _sorgu_thread(self):
        """Arka planda sorgu"""
        try:
            baslangic = self.baslangic_entry.get_date().strftime('%Y-%m-%d')
            bitis = self.bitis_entry.get_date().strftime('%Y-%m-%d')

            # Personel filtresi
            personel_id = None
            secilen = self.personel_var.get()
            if secilen != "Tumu":
                for p in self.personel_listesi:
                    if p['PersonelAdiSoyadi'].strip() == secilen:
                        personel_id = p['PersonelId']
                        break

            if self.ilac_disi_var.get():
                sonuclar = self.db.ilac_disi_raporu_getir(baslangic, bitis, personel_id)
            else:
                sonuclar = self.db.prim_raporu_getir(baslangic, bitis, personel_id)
            self.rapor_verileri = sonuclar

            self.parent.after(0, lambda: self._sonuclari_goster(sonuclar))

        except Exception as e:
            logger.error(f"Sorgu hatasi: {e}")
            self.parent.after(0, lambda: self._sorgu_hata(str(e)))

    def _sorgu_hata(self, hata_msg):
        """Sorgu hatasi"""
        self.listele_btn.config(state=tk.NORMAL)
        self.durum_label.config(text=f"Hata: {hata_msg[:60]}")
        messagebox.showerror("Sorgu Hatasi", f"Veri cekilemedi:\n{hata_msg}")

    def _sonuclari_goster(self, sonuclar):
        """Sonuclari tabloda goster"""
        # Tabloyu temizle
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Toplamlar
        toplam_adet = 0
        toplam_indirim = 0.0
        toplam_tutar = 0.0
        toplam_brut_kar = 0.0
        toplam_altidabirlik = 0.0
        personel_toplamlari = {}

        for s in sonuclar:
            etiket = float(s.get('Etiket') or 0)
            adet = int(s.get('Adet') or 0)
            indirimler = float(s.get('Indirimler') or 0)
            tutar = float(s.get('Tutar') or 0)
            alis_fiyati = float(s.get('AlisFiyati') or 0)
            alis_tahmini = int(s.get('AlisTahmini') or 0)
            personel = str(s.get('Personel') or 'ATANMAMIS')
            iade = int(s.get('Iade') or 0)

            # Tarih formatlama
            tarih_raw = s.get('Tarih')
            if tarih_raw:
                try:
                    if isinstance(tarih_raw, datetime):
                        tarih_str = tarih_raw.strftime('%d/%m/%Y')
                    else:
                        tarih_str = datetime.strptime(str(tarih_raw)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
                except (ValueError, TypeError):
                    tarih_str = str(tarih_raw)[:10]
            else:
                tarih_str = "-"

            # RIToplam = brut satis tutari (SatisFiy * Adet, indirim ONCESI)
            # RIIskonto = indirim tutari
            # Net tutar = RIToplam - RIIskonto (indirim sonrasi gercek gelir)
            net_tutar = tutar - indirimler

            # Satis fiyati = Indirim sonrasi birim fiyat
            satis_fiy = (net_tutar / adet) if adet != 0 else etiket

            # Brut kar hesaplama (net tutar uzerinden)
            brut_kar = net_tutar - (alis_fiyati * adet) if alis_fiyati > 0 else 0
            if iade:
                brut_kar_yuzde = (brut_kar / abs(net_tutar) * 100) if net_tutar != 0 else 0
                altidabirlik = brut_kar / 6.0
                altidabirlik_oran = (altidabirlik / abs(net_tutar) * 100) if net_tutar != 0 else 0
            else:
                brut_kar_yuzde = (brut_kar / net_tutar * 100) if net_tutar > 0 else 0
                # Altidabirlik prim = Brut Kar / 6
                altidabirlik = brut_kar / 6.0 if brut_kar > 0 else 0
                # 1/6 primin satis tutarina orani
                altidabirlik_oran = (altidabirlik / net_tutar * 100) if net_tutar > 0 and altidabirlik > 0 else 0

            # Renk tagi - 1/6 oran % baremine gore 8 kademe
            if altidabirlik_oran >= 15:
                tag = 'kar_15_ustu'
            elif altidabirlik_oran >= 12:
                tag = 'kar_12_15'
            elif altidabirlik_oran >= 10:
                tag = 'kar_10_12'
            elif altidabirlik_oran >= 7:
                tag = 'kar_7_10'
            elif altidabirlik_oran >= 5:
                tag = 'kar_5_7'
            elif altidabirlik_oran >= 4:
                tag = 'kar_4_5'
            elif altidabirlik_oran >= 0:
                tag = 'kar_0_4'
            else:
                tag = 'kar_negatif'

            barkod = str(s.get('Barkod') or '')

            # Tag listesi: kar rengi + tahmini alis fiyati ise kirmizi yazi
            tags = [tag]
            if alis_tahmini and not iade:
                tags.append('alis_tahmini')

            urun_adi = f"İADE: {s.get('UrunAdi', '')[:55]}" if iade else s.get('UrunAdi', '')[:60]

            self.tree.insert('', 'end', values=(
                urun_adi,
                barkod,
                tarih_str,
                f"{etiket:,.2f}",
                f"{satis_fiy:,.2f}",
                adet,
                f"{indirimler:,.2f}",
                f"{net_tutar:,.2f}",
                f"~{alis_fiyati:,.2f}" if alis_tahmini else (f"{alis_fiyati:,.2f}" if alis_fiyati > 0 else "-"),
                f"{brut_kar:,.2f}" if alis_fiyati > 0 else "-",
                f"%{brut_kar_yuzde:,.1f}" if alis_fiyati > 0 else "-",
                f"{altidabirlik:,.2f}" if altidabirlik != 0 else "-",
                f"%{altidabirlik_oran:,.1f}" if altidabirlik_oran != 0 else "-",
                personel,
            ), tags=tuple(tags))

            # Toplamlar
            toplam_adet += adet
            toplam_indirim += indirimler
            toplam_tutar += net_tutar
            toplam_brut_kar += brut_kar
            toplam_altidabirlik += altidabirlik

            # Personel bazli toplamlar
            if personel not in personel_toplamlari:
                personel_toplamlari[personel] = {
                    'adet': 0, 'tutar': 0.0, 'brut_kar': 0.0, 'altidabirlik': 0.0
                }
            personel_toplamlari[personel]['adet'] += adet
            personel_toplamlari[personel]['tutar'] += net_tutar
            personel_toplamlari[personel]['brut_kar'] += brut_kar
            personel_toplamlari[personel]['altidabirlik'] += altidabirlik

        # Ozet guncelle
        self.ozet_labels['satir'].config(text=str(len(sonuclar)))
        self.ozet_labels['adet'].config(text=str(toplam_adet))
        self.ozet_labels['indirim'].config(text=f"{toplam_indirim:,.2f} TL")
        self.ozet_labels['tutar'].config(text=f"{toplam_tutar:,.2f} TL")
        self.ozet_labels['brut_kar'].config(text=f"{toplam_brut_kar:,.2f} TL")
        self.ozet_labels['altidabirlik'].config(text=f"{toplam_altidabirlik:,.2f} TL")

        # Personel bazli ozet
        for widget in self.personel_ozet_frame.winfo_children():
            widget.destroy()

        if personel_toplamlari:
            tk.Label(self.personel_ozet_frame, text="Personel Bazli:",
                     font=('Segoe UI', 10, 'bold'), bg='#E3F2FD').pack(side=tk.LEFT, padx=(10, 10))

            for personel, toplamlar in sorted(personel_toplamlari.items()):
                metin = f"{personel}: {toplamlar['adet']} ad. | {toplamlar['tutar']:,.2f} TL | 1/6 Prim: {toplamlar['altidabirlik']:,.2f} TL"
                tk.Label(self.personel_ozet_frame, text=metin,
                         font=('Segoe UI', 9), bg='#E3F2FD', fg='#333',
                         relief=tk.GROOVE, padx=8, pady=2).pack(side=tk.LEFT, padx=(0, 8))

        # Durum guncelle
        baslangic = self.baslangic_entry.get_date().strftime('%d/%m/%Y')
        bitis = self.bitis_entry.get_date().strftime('%d/%m/%Y')
        self.durum_label.config(
            text=f"{len(sonuclar)} kayit | {baslangic} - {bitis}")
        self.listele_btn.config(state=tk.NORMAL)

    def _sutun_sirala(self, sutun_id):
        """Sutuna gore sirala"""
        if self._siralama_sutun == sutun_id:
            self._siralama_ters = not self._siralama_ters
        else:
            self._siralama_sutun = sutun_id
            self._siralama_ters = False

        items = [(self.tree.set(item, sutun_id), item)
                 for item in self.tree.get_children('')]

        # Sayisal siralama denemesi
        try:
            items.sort(key=lambda t: float(t[0].replace(',', '').replace('%', '').replace(' TL', '').replace('-', '0')),
                       reverse=self._siralama_ters)
        except (ValueError, TypeError):
            items.sort(key=lambda t: t[0], reverse=self._siralama_ters)

        for index, (val, item) in enumerate(items):
            self.tree.move(item, '', index)

    def _barkod_kopyala(self, event=None):
        """F1: Secili satirin barkodunu panoya kopyala"""
        secili = self.tree.selection()
        if not secili:
            self.durum_label.config(text="Barkod kopyalamak icin bir satir secin")
            return

        item = secili[0]
        barkod = self.tree.set(item, 'Barkod')
        urun_adi = self.tree.set(item, 'UrunAdi')[:25]

        if barkod:
            self.parent.clipboard_clear()
            self.parent.clipboard_append(barkod)
            self.durum_label.config(text=f"Barkod kopyalandi: {barkod} ({urun_adi})")
        else:
            self.durum_label.config(text="Bu urunun barkodu yok!")

    def _barkod_yapistir(self, event=None):
        """F2: Panodaki barkodu aktif metin alanina yapistir"""
        try:
            barkod = self.parent.clipboard_get()
        except Exception:
            self.durum_label.config(text="Panoda barkod yok!")
            return

        if not barkod:
            self.durum_label.config(text="Panoda barkod yok!")
            return

        focused = self.parent.focus_get()
        if focused and isinstance(focused, (tk.Entry, ttk.Entry)):
            focused.delete(0, tk.END)
            focused.insert(0, barkod)
            self.durum_label.config(text=f"Barkod yapistirildi: {barkod}")
        elif focused and isinstance(focused, tk.Text):
            focused.insert(tk.INSERT, barkod)
            self.durum_label.config(text=f"Barkod yapistirildi: {barkod}")
        else:
            self.durum_label.config(text=f"Barkod panoda: {barkod} (bir metin alanina tiklayip F2)")

    def _prim_urunleri_goster(self):
        """Prim verilen urunlerin listesini popup pencerede goster"""
        if not self.db:
            messagebox.showerror("Hata", "Veritabani baglantisi yok!")
            return

        try:
            urunler = self.db.prim_urunleri_listesi_getir()
        except Exception as e:
            messagebox.showerror("Hata", f"Veri cekilemedi:\n{e}")
            return

        if not urunler:
            messagebox.showinfo("Bilgi", "Prim verilen urun bulunamadi.")
            return

        # Popup pencere
        popup = tk.Toplevel(self.parent)
        popup.title(f"Prim Verilen Urun Listesi ({len(urunler)} urun)")
        popup.geometry("900x600")
        popup.configure(bg='#ECEFF1')
        popup.transient(self.parent)
        popup.grab_set()

        # Baslik
        tk.Label(popup, text=f"Prim Verilen Urun Listesi - {len(urunler)} urun",
                 font=('Segoe UI', 13, 'bold'), bg='#7B1FA2', fg='white',
                 pady=10).pack(fill=tk.X)

        # Tablo
        tablo_frame = tk.Frame(popup, bg='#ECEFF1')
        tablo_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        sutunlar = [
            ("UrunAdi", "Urun Adi", 350, tk.W),
            ("EtiketFiyat", "Etiket Fiyat", 100, tk.E),
            ("PrimYuzde", "Prim %", 70, tk.E),
            ("Stok", "Stok", 60, tk.CENTER),
            ("PrimTipAdi", "Prim Tipi", 100, tk.W),
        ]

        tree = ttk.Treeview(tablo_frame,
                             columns=[c[0] for c in sutunlar],
                             show='headings', height=22)

        for col_id, baslik, width, anchor in sutunlar:
            tree.heading(col_id, text=baslik)
            tree.column(col_id, width=width, minwidth=40, anchor=anchor)

        tree.tag_configure('tanimsiz', background='#FFF9C4')
        tree.tag_configure('aktif', background='#FFFFFF')

        for u in urunler:
            prim_yuzde = float(u.get('PrimYuzde') or 0)
            stok = int(u.get('Stok') or 0)
            etiket = float(u.get('EtiketFiyat') or 0)
            tag = 'aktif' if prim_yuzde > 0 else 'tanimsiz'

            tree.insert('', 'end', values=(
                u.get('UrunAdi', ''),
                f"{etiket:,.2f}" if etiket > 0 else "-",
                f"%{prim_yuzde:,.0f}" if prim_yuzde > 0 else "Tanimsiz",
                stok,
                u.get('PrimTipAdi', ''),
            ), tags=(tag,))

        vsb = ttk.Scrollbar(tablo_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Ozet
        tanimli = sum(1 for u in urunler if float(u.get('PrimYuzde') or 0) > 0)
        tanimsiz = len(urunler) - tanimli
        tk.Label(popup,
                 text=f"Toplam: {len(urunler)} urun | Prim Tanimli: {tanimli} | Tanimsiz: {tanimsiz}",
                 font=('Segoe UI', 10), bg='#ECEFF1', fg='#333',
                 pady=5).pack(fill=tk.X)

        # Kapat butonu
        tk.Button(popup, text="Kapat", font=('Segoe UI', 10),
                  bg='#5D4037', fg='white', relief=tk.FLAT, padx=20, pady=5,
                  command=popup.destroy).pack(pady=(0, 10))

    def _excel_aktar(self):
        """Excel'e aktar"""
        if not self.rapor_verileri:
            messagebox.showwarning("Uyari", "Once rapor listeleyin!")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            baslangic = self.baslangic_entry.get_date().strftime('%d%m%Y')
            bitis = self.bitis_entry.get_date().strftime('%d%m%Y')

            dosya = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel dosyasi", "*.xlsx")],
                initialfile=f"prim_raporu_{baslangic}_{bitis}.xlsx"
            )

            if not dosya:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Prim Raporu"

            # Stiller
            baslik_font = Font(bold=True, color="FFFFFF", size=11)
            baslik_fill = PatternFill(start_color="5D4037", end_color="5D4037", fill_type="solid")
            baslik_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # 8 kademe renk (altidabirlik_oran bazli)
            fill_15_ustu = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")   # >=15%
            fill_12_15 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")      # 12-15%
            fill_10_12 = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")      # 10-12%
            fill_7_10 = PatternFill(start_color="C5E1A5", end_color="C5E1A5", fill_type="solid")       # 7-10%
            fill_5_7 = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")        # 5-7%
            fill_4_5 = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")        # 4-5%
            fill_0_4 = PatternFill(start_color="FFB74D", end_color="FFB74D", fill_type="solid")        # 0-4%
            fill_negatif = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")    # <0%
            iade_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            toplam_fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
            toplam_font = Font(bold=True, size=11)
            sayi_align = Alignment(horizontal="right")

            # Baslik satiri (tarih araligi)
            baslangic_str = self.baslangic_entry.get_date().strftime('%d/%m/%Y')
            bitis_str = self.bitis_entry.get_date().strftime('%d/%m/%Y')
            ws.merge_cells('A1:N1')
            title_cell = ws.cell(row=1, column=1,
                                 value=f"{baslangic_str} - {bitis_str} Tarih Araligindaki Prim Raporu")
            title_cell.font = Font(bold=True, size=13)
            title_cell.alignment = Alignment(horizontal="center")

            # Sutun basliklari
            basliklar = ["Urun Adi", "Barkod", "Tarih", "Etiket", "Satis Fiy", "Adet", "Indirimler",
                         "Tutar", "Alis Fiy", "Brut Kar", "Brut Kar %",
                         "1/6 Prim", "1/6 Oran %", "Personel"]

            for col_idx, baslik in enumerate(basliklar, 1):
                cell = ws.cell(row=2, column=col_idx, value=baslik)
                cell.font = baslik_font
                cell.fill = baslik_fill
                cell.alignment = baslik_align
                cell.border = thin_border

            # Toplamlar
            toplam_adet = 0
            toplam_indirim_xl = 0.0
            toplam_tutar = 0.0
            toplam_brut_kar = 0.0
            toplam_altidabirlik_xl = 0.0

            # Veri satirlari
            tarih_align = Alignment(horizontal="center")

            for row_idx, s in enumerate(self.rapor_verileri, 3):
                etiket = float(s.get('Etiket') or 0)
                adet = int(s.get('Adet') or 0)
                indirimler = float(s.get('Indirimler') or 0)
                tutar = float(s.get('Tutar') or 0)
                alis_fiyati = float(s.get('AlisFiyati') or 0)
                personel = str(s.get('Personel') or 'ATANMAMIS')
                iade = int(s.get('Iade') or 0)

                # Tarih formatlama
                tarih_raw = s.get('Tarih')
                if tarih_raw:
                    try:
                        if isinstance(tarih_raw, datetime):
                            tarih_str = tarih_raw.strftime('%d/%m/%Y')
                        else:
                            tarih_str = datetime.strptime(str(tarih_raw)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
                    except (ValueError, TypeError):
                        tarih_str = str(tarih_raw)[:10]
                else:
                    tarih_str = "-"

                # RIToplam = brut satis tutari (indirim ONCESI)
                # RIIskonto = indirim tutari (DB'den)
                # Net tutar = RIToplam - RIIskonto
                net_tutar = tutar - indirimler
                satis_fiy = (net_tutar / adet) if adet != 0 else etiket
                brut_kar = net_tutar - (alis_fiyati * adet) if alis_fiyati > 0 else 0
                if iade:
                    brut_kar_yuzde = (brut_kar / abs(net_tutar) * 100) if net_tutar != 0 else 0
                    altidabirlik = brut_kar / 6.0
                    altidabirlik_oran = (altidabirlik / abs(net_tutar) * 100) if net_tutar != 0 else 0
                else:
                    brut_kar_yuzde = (brut_kar / net_tutar * 100) if net_tutar > 0 else 0
                    altidabirlik = brut_kar / 6.0 if brut_kar > 0 else 0
                    altidabirlik_oran = (altidabirlik / net_tutar * 100) if net_tutar > 0 and altidabirlik > 0 else 0

                barkod_xl = str(s.get('Barkod') or '')
                urun_adi_xl = f"İADE: {s.get('UrunAdi', '')}" if iade else s.get('UrunAdi', '')

                degerler = [
                    urun_adi_xl,
                    barkod_xl,
                    tarih_str,
                    round(etiket, 2),
                    round(satis_fiy, 2),
                    adet,
                    round(indirimler, 2),
                    round(net_tutar, 2),
                    round(alis_fiyati, 2) if alis_fiyati > 0 else None,
                    round(brut_kar, 2) if alis_fiyati > 0 else None,
                    round(brut_kar_yuzde, 1) if alis_fiyati > 0 else None,
                    round(altidabirlik, 2) if altidabirlik != 0 else None,
                    round(altidabirlik_oran, 1) if altidabirlik_oran != 0 else None,
                    personel,
                ]

                for col_idx, val in enumerate(degerler, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx == 3:  # Tarih
                        cell.alignment = tarih_align
                    elif col_idx >= 4 and col_idx <= 13:  # Sayisal sutunlar
                        cell.alignment = sayi_align
                        if isinstance(val, (int, float)) and val is not None:
                            cell.number_format = '#,##0.00' if col_idx != 6 else '#,##0'

                # Renk - 8 kademe altidabirlik_oran bazli
                if iade:
                    row_fill = iade_fill
                elif altidabirlik_oran >= 15:
                    row_fill = fill_15_ustu
                elif altidabirlik_oran >= 12:
                    row_fill = fill_12_15
                elif altidabirlik_oran >= 10:
                    row_fill = fill_10_12
                elif altidabirlik_oran >= 7:
                    row_fill = fill_7_10
                elif altidabirlik_oran >= 5:
                    row_fill = fill_5_7
                elif altidabirlik_oran >= 4:
                    row_fill = fill_4_5
                elif altidabirlik_oran >= 0:
                    row_fill = fill_0_4
                else:
                    row_fill = fill_negatif

                for col_idx in range(1, 15):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = row_fill
                    # Koyu arka plan icin beyaz yazi
                    if row_fill in (fill_15_ustu, fill_12_15, fill_negatif):
                        cell.font = Font(color="FFFFFF")

                toplam_adet += adet
                toplam_indirim_xl += indirimler
                toplam_tutar += net_tutar
                toplam_brut_kar += brut_kar
                toplam_altidabirlik_xl += altidabirlik

            # Toplam satiri
            toplam_row = len(self.rapor_verileri) + 3
            toplam_degerler = [
                "TOPLAM", None, None, None, None, toplam_adet, round(toplam_indirim_xl, 2),
                round(toplam_tutar, 2), None, round(toplam_brut_kar, 2), None,
                round(toplam_altidabirlik_xl, 2), None, None
            ]
            for col_idx, val in enumerate(toplam_degerler, 1):
                cell = ws.cell(row=toplam_row, column=col_idx, value=val)
                cell.font = toplam_font
                cell.fill = toplam_fill
                cell.border = thin_border
                if col_idx >= 2:
                    cell.alignment = sayi_align

            # Sutun genislikleri
            genislikler = [40, 15, 12, 12, 12, 8, 12, 14, 12, 14, 12, 12, 12, 18]
            for i, g in enumerate(genislikler, 1):
                ws.column_dimensions[get_column_letter(i)].width = g

            # Baslik satirini dondur
            ws.freeze_panes = 'A3'

            wb.save(dosya)
            messagebox.showinfo("Basarili", f"Excel dosyasi kaydedildi:\n{dosya}")

        except ImportError:
            messagebox.showerror("Hata",
                                 "openpyxl modulu gerekli!\npip install openpyxl")
        except Exception as e:
            logger.error(f"Excel aktarim hatasi: {e}")
            messagebox.showerror("Hata", f"Excel aktarim hatasi:\n{e}")

    def _ana_menuye_don(self):
        """Ana menuye don"""
        if self.db:
            try:
                self.db.kapat()
            except Exception:
                pass
        self.parent.destroy()
        if self.ana_menu_callback:
            self.ana_menu_callback()


def main():
    root = tk.Tk()
    app = PrimRaporlamaGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
