"""
Aylık İnceleme & Filtreleme Tablosu (Botanik EOS)

Kullanıcının manuel inceleme akışı için tek tablo + çoklu filtre + boyama.
Her satır = bir reçete-ilaç kombinasyonu.

GÜVENLİK: Yalnızca SELECT — BotanikDB.sorgu_calistir güvenlik filtresi kullanılır.
Hiçbir INSERT/UPDATE/DELETE yapılmaz.

Akış:
1. Yıl-Ay seç → Sorgula → tüm aylık reçete-ilaç satırları tabloya gelir (BEYAZ)
2. Üstteki sütun arama kutuları ve sağdaki hızlı filtre butonları ile filtre uygula
3. Filtreden geçen satırları "Filtreliyi Yeşile Boya" ile toplu boya
4. Renk filtre (Yeşil/Sarı/Kırmızı/Beyaz/Turuncu checkbox'ları) ile gösterimi daralt
5. Çember daraldıkça → kalan satırlar gerçek manuel kontrol gerektirenler
6. Excel export ile renkleri koruyarak çıktı al

State: aylik_inceleme_state.json — {dönem: {ri_id: renk}} mapping'i.
"""

import json
import logging
import os
import re
import threading
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, ttk
from typing import Dict, List

from botanik_db import BotanikDB
from recete_kontrol.sut_kontrolleri import _tr_lower

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
# RENK SİSTEMİ
# ═══════════════════════════════════════════════════════════════════════
RENK_BEYAZ = "beyaz"
RENK_YESIL = "yesil"
RENK_SARI = "sari"
RENK_TURUNCU = "turuncu"
RENK_KIRMIZI = "kirmizi"
RENK_GRI = "gri"          # Kontrol gerekmez — bu ilacı dışlama listesine ekler

RENK_BG = {
    RENK_BEYAZ:   "#FFFFFF",
    RENK_YESIL:   "#C8E6C9",
    RENK_SARI:    "#FFF9C4",
    RENK_TURUNCU: "#FFE0B2",
    RENK_KIRMIZI: "#FFCDD2",
    RENK_GRI:     "#E0E0E0",
}
RENK_FG = {
    RENK_BEYAZ:   "#212121",
    RENK_YESIL:   "#1B5E20",
    RENK_SARI:    "#827717",
    RENK_TURUNCU: "#E65100",
    RENK_KIRMIZI: "#B71C1C",
    RENK_GRI:     "#616161",
}
RENK_ETIKET = {
    RENK_BEYAZ:   "Beklemede",
    RENK_YESIL:   "Uygun (Elendi)",
    RENK_SARI:    "Manuel Kontrol",
    RENK_TURUNCU: "Şüpheli",
    RENK_KIRMIZI: "Uygunsuz",
    RENK_GRI:     "Kontrol Gerekmez (dışla)",
}


# ═══════════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════════
UYARI_KODU_REGEX = re.compile(r"\((\d{3,7})\)")


def uyari_kodlarini_ayikla(metin: str) -> list:
    if not metin:
        return []
    return list(dict.fromkeys(UYARI_KODU_REGEX.findall(metin)))


def _tarihi_parse(d):
    """DB'den gelen tarih (datetime/date/str) → date veya None."""
    if not d:
        return None
    if hasattr(d, "year"):
        return d
    # String — "YYYY-MM-DD" veya "YYYY-MM-DD HH:MM:SS"
    try:
        from datetime import datetime
        s = str(d)[:10]
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def tarih_format(d, fmt="%d.%m.%Y") -> str:
    """DB'den gelen tarihi (datetime/date/str) güvenli şekilde formatla."""
    if not d:
        return ""
    if hasattr(d, "strftime"):
        try:
            return d.strftime(fmt)
        except Exception:
            pass
    # String fallback
    s = str(d)
    if len(s) >= 10:
        # "YYYY-MM-DD" → "DD.MM.YYYY" varsayılan
        if fmt == "%d.%m.%Y":
            try:
                return f"{s[8:10]}.{s[5:7]}.{s[0:4]}"
            except Exception:
                return s[:10]
        elif fmt == "%Y-%m":
            return s[:7]
        return s[:10]
    return s


def yas_hesapla(dogum_tarihi) -> str:
    dt = _tarihi_parse(dogum_tarihi)
    if not dt:
        return ""
    try:
        bugun = date.today()
        yas = bugun.year - dt.year - (
            (bugun.month, bugun.day) < (dt.month, dt.day)
        )
        return str(yas)
    except Exception:
        return ""


def cinsiyet_etiket(c) -> str:
    if not c:
        return ""
    s = str(c).strip().upper()
    return {"E": "Erkek", "K": "Kadın", "M": "Erkek", "F": "Kadın"}.get(s, s)


_PERIYOT_AD = {
    3: "günde",
    4: "haftada",
    5: "ayda",
    6: "yılda",
    1: "saatte",
    2: "saatte",
}


def _sayi_kisa(d) -> str:
    """Decimal/float'ı kısa formatta göster — 1.00 → 1, 0.50 → 0.5"""
    if d is None or d == "":
        return ""
    try:
        f = float(d)
        if f == int(f):
            return str(int(f))
        return f"{f:g}"
    except Exception:
        return str(d)


def doz_metin(aralik, periyot_id, tekrar, doz) -> str:
    """Reçete dozunu '<aralık> <periyot> <tekrar> x <doz>' formatında göster.

    Örnek:
      aralik=1, periyot=3 (Günde), tekrar=2, doz=1   → '1 günde 2 x 1'
      aralik=7, periyot=3 (Günde), tekrar=1, doz=2   → '7 günde 1 x 2'
      aralik=1, periyot=5 (Ayda),  tekrar=1, doz=1   → '1 ayda 1 x 1'
    """
    aralik_s = _sayi_kisa(aralik) or "1"
    tekrar_s = _sayi_kisa(tekrar) or "1"
    doz_s = _sayi_kisa(doz) or "1"
    try:
        pid = int(periyot_id) if periyot_id is not None else 3
    except Exception:
        pid = 3
    periyot_ad = _PERIYOT_AD.get(pid, "günde")
    return f"{aralik_s} {periyot_ad} {tekrar_s} x {doz_s}"


# ═══════════════════════════════════════════════════════════════════════
# SUT KATEGORİLERİ
# ═══════════════════════════════════════════════════════════════════════
def sut_madde_tespit(ilac_adi: str, etkin_madde: str, rapor_kodu: str) -> str:
    """sut_kontrolleri'nden SUT maddesini tespit et (4.2.x formatında)."""
    try:
        from recete_kontrol.sut_kontrolleri import (
            sut_kategorisi_tespit_et, KATEGORI_ISIMLERI,
        )
        kategori = sut_kategorisi_tespit_et({
            "ilac_adi": ilac_adi or "",
            "etkin_madde": etkin_madde or "",
            "rapor_kodu": rapor_kodu or "",
        })
        if kategori:
            return KATEGORI_ISIMLERI.get(kategori, kategori)
    except Exception:
        pass
    return ""


# ═══════════════════════════════════════════════════════════════════════
# SÜTUN TANIMLARI (26 sütun)
# ═══════════════════════════════════════════════════════════════════════
SUTUNLAR = [
    # (kod, başlık, genişlik, tooltip)
    ("secim",      "✓",            32,  "Seçim — tıklayınca işaretle/kaldır"),
    ("grup",       "Grup",         45,  "A/B/C/GK/CK"),
    ("donem",      "Dönem",        70,  "Yıl-Ay"),
    ("rec_tar",    "Reç.Tarih",    80,  "Reçete tarihi"),
    ("rec_no",     "Reçete No",    80,  "E-reçete numarası"),
    ("rec_tip",    "Reç.Türü",     75,  "Reçete türü (Normal/Kırmızı/Yeşil/Mor)"),
    ("rec_alttur", "Reç.Alt Türü", 95,  "Reçete alt türü (Ayaktan/Yatan/A/B/C/GK/Kan)"),
    ("hasta",      "Hasta Adı",    140, "Hasta adı"),
    ("tc",         "TC No",        95,  "Hasta TC kimlik no"),
    ("yas",        "Yaş",          35,  "Hasta yaşı"),
    ("cins",       "Cin.",         38,  "Cinsiyet"),
    ("hasta_tip",  "Hasta Tipi",   90,  "Yeşil Kart/SSK/Emekli/Çalışan"),
    ("doktor",     "Doktor",       130, "Doktor adı"),
    ("brans",      "Branş",        110, "Doktor branşı"),
    ("ilac",       "İlaç",         180, "İlaç adı"),
    ("etkin",      "Etken Madde",  130, "Etken madde"),
    ("atc",        "ATC",          70,  "ATC kodu"),
    ("esdeger",    "Eşdeğer",      70,  "Eşdeğer grubu"),
    ("kutu",       "Kutu",         42,  "Kutu sayısı (kaç kutu verildiği)"),
    ("sut",        "SUT Maddesi",  130, "SUT 4.2.x kategorisi"),
    ("rap_kod",    "Rapor Kod",    65,  "Rapor kodu"),
    ("rap_tesh_tak", "Rap.Teşhis/Rap.Tak.No", 145,
     "Rapor teşhis kodu - Rapor takip numarası "
     "(ör: '04.05 - 512432126'). Tak.No önceliği RIRaporNo, "
     "boşsa eşleşen rapordan RaporAnaRaporNo."),
    ("rec_doz",    "Reçete Doz",   90,  "Reçete dozu"),
    ("rap_doz",    "Rapor Doz",    80,  "Rapor dozu"),
    ("msj",        "Msj",          35,  "İlaç mesaj durumu"),
    ("uyari",      "Uyarı Kod",    150, "Bu ilaç için reçeteye girilen uyarı kodları (ReceteTeshis)"),
    ("medula_msj", "Medula Msj",   220, "Medula provizyon yanıt metni (RxUyarilari)"),
    ("rec_tesh",   "Reç.Teşhis",   150, "Reçete teşhisleri"),
    ("rap_tesh",   "Rap.Teşhis",   150, "Rapor teşhisleri"),
    ("rec_ack",    "Reç.Açk",      180, "Reçete açıklamaları"),
    ("rap_ack",    "Rap.Açk",      180, "Rapor açıklamaları"),
    ("verdict_doz", "Doz Karş.",   100,
     "DOZ KARŞILAŞTIR butonu çalıştırıldığında doldurulur: "
     "UYGUN / UYGUN DEĞİL / ŞÜPHELİ / ATLANDI — Reçete günlük dozu rapor günlük dozunu geçiyor mu?"),
    ("verdict_uyari_kontrol", "Uyarı Kod Karş.", 130,
     "UYARI KODU KONTROL butonu doldurur: reçete uyarı kodlarının "
     "teşhis/açıklamayla eşleşmesi → UYGUN / UYGUN DEĞİL / ŞÜPHELİ"),
    ("verdict",    "SONUÇ",        130,
     "STATİN KONTROL butonu çalıştırıldığında doldurulur: "
     "UYGUN / UYGUN DEĞİL / ŞÜPHELİ / ATLANDI"),
    ("renkli_rr",  "Renkli RR",    85,
     "RENKLİ REÇETE butonu doldurur: Kırmızı/Yeşil/Mor reçetelerin "
     "renkli reçete sisteminde kaydı VAR / YOK / — (kapsam dışı)"),
    ("verdict_sart_raporu", "Şart Raporu", 460,
     "SUT kontrolü sonrası şart raporu: kullanılan SUT maddesi, ana mesaj, "
     "aranan/bulunan ibareler ve uyarılar. Excel kontrol raporundaki "
     "özetin tablo görünümü."),
]
SUTUN_KOD = [s[0] for s in SUTUNLAR]


# ═══════════════════════════════════════════════════════════════════════
# HIZLI FİLTRE TANIMLARI
# ═══════════════════════════════════════════════════════════════════════
# Her filtre: (etiket, fonksiyon(satir) -> bool, açıklama)
def _f_raporlu(s: dict) -> bool:
    return (s.get("msj") or "").strip().lower() == "var"


def _f_raporsuz(s: dict) -> bool:
    return (s.get("msj") or "").strip().lower() == "yok"


def _f_mesajsiz(s: dict) -> bool:
    """Raporsuz ilaç mesajı yoktur — Mesajsız raporsuz ile aynıdır."""
    return (s.get("msj") or "").strip().lower() == "yok"


def _f_doz_uygun(s: dict) -> bool:
    rd = s.get("rec_doz_sayi") or 0
    rpd = s.get("rap_doz_sayi") or 0
    if not rpd:
        return False
    return rd > 0 and rd <= rpd


def _f_klasik_oad(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    klasikler = ["METFORMIN", "GLIKLAZID", "GLIMEPIRID", "GLIBENKLAMID",
                 "GLIPIZID", "REPAGLINID", "NATEGLINID", "PIOGLITAZON",
                 "AKARBOZ"]
    ticari = ["GLIFOR", "GLUKOFEN", "DIAFORMIN", "DIAMICRON", "AMARYL",
              "GLIFIX", "GLUCOBAY", "GLIMAX", "ACTOS", "GLIBEDAL",
              "NOVONORM", "STARLIX"]
    return any(k in et for k in klasikler) or any(t in ad for t in ticari)


def _f_dpp4_sglt2_glp1(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    sinif = ["SITAGLIPTIN", "VILDAGLIPTIN", "SAKSAGLIPTIN", "LINAGLIPTIN",
             "ALOGLIPTIN", "EMPAGLIFLOZIN", "DAPAGLIFLOZIN", "KANAGLIFLOZIN",
             "ERTUGLIFLOZIN", "LIRAGLUTID", "SEMAGLUTID", "DULAGLUTID"]
    ticari = ["JANUVIA", "GALVUS", "ONGLYZA", "TRAJENTA", "NESINA",
              "JARDIANCE", "FORZIGA", "INVOKANA", "JANUMET", "GALVUSMET",
              "VICTOZA", "OZEMPIC", "TRULICITY"]
    return any(k in et for k in sinif) or any(t in ad for t in ticari)


def _f_klopidogrel(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("KLOPIDOGREL" in et or "PRASUGREL" in et or "TIKAGRELOR" in et
            or any(t in ad for t in ["PLAVIX", "PLANOR", "KARUM", "EFFIENT", "BRILINTA"]))


def _f_yoak(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("RIVAROKSABAN" in et or "APIKSABAN" in et or "EDOKSABAN" in et
            or "DABIGATRAN" in et
            or any(t in ad for t in ["XARELTO", "ELIQUIS", "LIXIANA", "PRADAXA"]))


def _f_statin(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("STATIN" in et or "ATORVASTATIN" in et or "ROSUVASTATIN" in et
            or "SIMVASTATIN" in et or "PRAVASTATIN" in et or "FLUVASTATIN" in et
            or any(t in ad for t in ["LIPITOR", "CRESTOR", "ATOR", "ULTROX",
                                       "ZOCOR", "LIPVAS", "ALVASTIN"]))


def _f_hepatit(s: dict) -> bool:
    """Hepatit B/C ilaçları (HIV ilaçları HARİÇ)."""
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    arama = ad + " " + et
    hbv_hcv_etken = ("ENTEKAVIR", "TENOFOVIR", "LAMIVUDIN", "TELBIVUDIN",
                     "ADEFOVIR", "SOFOSBUVIR", "LEDIPASVIR", "VELPATASVIR",
                     "VOXILAPREVIR", "GLEKAPREVIR", "GLECAPREVIR",
                     "PIBRENTASVIR", "OMBITASVIR", "PARITAPREVIR",
                     "DASABUVIR", "DAKLATASVIR", "ELBASVIR", "GRAZOPREVIR",
                     "PEGINTERFERON", "RIBAVIRIN")
    hbv_hcv_ticari = ("BARACLUDE", "VIREAD", "VEMLIDY", "ZEFFIX", "SEBIVO",
                      "HEPSERA", "SOVALDI", "HARVONI", "EPCLUSA", "VOSEVI",
                      "MAVYRET", "MAVIRET", "VIEKIRAX", "EXVIERA", "DAKLINZA",
                      "ZEPATIER", "PEGASYS", "PEGINTRON", "COPEGUS", "REBETOL")
    return (any(e in arama for e in hbv_hcv_etken)
            or any(t in ad for t in hbv_hcv_ticari))


def _f_noropatik(s: dict) -> bool:
    """Nöropatik ağrı uyarı kodu içeren reçete."""
    uy = (s.get("uyari") or "")
    return ("noropatik" in uy.lower() or "nöropatik" in uy.lower()
            or _f_gabapentinoid(s))


def _f_gabapentinoid(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("GABAPENTIN" in et or "PREGABALIN" in et
            or any(t in ad for t in ["NEURONTIN", "LYRICA", "NERUDA",
                                       "GABAGAMMA", "PREGALIN", "PREGABEX"]))


def _f_uriner_antibiyotik(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("FOSFOMISIN" in et or "NITROFURANTOIN" in et
            or any(t in ad for t in ["MONUROL", "UROMSIN"]))


def _f_immunsupresif(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("TAKROLIMUS" in et or "MIKOFENOLAT" in et or "SIROLIMUS" in et
            or "EVEROLIMUS" in et or "SIKLOSPORIN" in et or "AZATIYOPRIN" in et
            or any(t in ad for t in ["ADVAGRAF", "PROGRAF", "MYFORTIC",
                                       "CELLCEPT", "RAPAMUNE", "CERTICAN",
                                       "SANDIMMUN", "IMURAN"]))


def _f_ms_dmt(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return ("DIMETIL FUMARAT" in et or "INTERFERON BETA" in et or "FINGOLIMOD" in et
            or "GLATIRAMER" in et or "TERIFLUNOMID" in et or "OKRELIZUMAB" in et
            or any(t in ad for t in ["TENIPRA", "TECFIDERA", "AVONEX", "REBIF",
                                       "GILENYA", "AUBAGIO", "OCREVUS"]))


def _f_solunum(s: dict) -> bool:
    et = (s.get("etkin") or "").upper()
    ad = (s.get("ilac") or "").upper()
    return any(k in et for k in ["FORMOTEROL", "SALMETEROL", "VILANTEROL",
                                    "INDAKATEROL", "TIOTROPIUM", "GLIKOPIRONYUM",
                                    "BUDESONID", "BUDEZONID", "FLUTIKAZON",
                                    "MOMETAZON", "MONTELUKAST"]) or \
           any(t in ad for t in ["SERETIDE", "SYMBICORT", "FOSTER", "RELVAR",
                                   "TRELEGY", "TRIMBOW", "BREZTRI", "SPIRIVA",
                                   "ANORO", "ULTIBRO", "SINGULAIR"])


# Hızlı filtre listesi (sıralı buton düzeni)
HIZLI_FILTRELER = [
    ("● Raporlu ilaçlar",         _f_raporlu),
    ("○ Raporsuz ilaçlar",        _f_raporsuz),
    ("Mesajsız",                  _f_mesajsiz),
    ("Doz Uygun (≤Rapor)",        _f_doz_uygun),
    ("Klasik OAD",                _f_klasik_oad),
    ("DPP-4/SGLT-2/GLP-1",        _f_dpp4_sglt2_glp1),
    ("Klopidogrel/Antiplatelet",  _f_klopidogrel),
    ("YOAK",                      _f_yoak),
    ("Statin",                    _f_statin),
    ("Hepatit B/C",               _f_hepatit),
    ("Gabapentinoid/Nöropatik",   _f_noropatik),
    ("Üriner Antibiyotik",        _f_uriner_antibiyotik),
    ("İmmünsüpresif",             _f_immunsupresif),
    ("MS DMT",                    _f_ms_dmt),
    ("Solunum (LABA/ICS/LTRA)",   _f_solunum),
]


# ═══════════════════════════════════════════════════════════════════════
# TOOLTIP HELPER (1 sn gecikmeli hover ipucu)
# ═══════════════════════════════════════════════════════════════════════
class _Tooltip:
    """Widget üzerinde 1 saniye bekleyince çıkan açıklama balonu."""

    def __init__(self, widget, text: str, delay_ms: int = 1000):
        self.widget = widget
        self.text = text
        self.delay = delay_ms
        self._tip = None
        self._after_id = None
        widget.bind("<Enter>", self._enter, add="+")
        widget.bind("<Leave>", self._leave, add="+")
        widget.bind("<ButtonPress>", self._leave, add="+")

    def _enter(self, _event=None):
        self._cancel()
        self._after_id = self.widget.after(self.delay, self._show)

    def _leave(self, _event=None):
        self._cancel()
        if self._tip is not None:
            try:
                self._tip.destroy()
            except Exception:
                pass
            self._tip = None

    def _cancel(self):
        if self._after_id:
            try:
                self.widget.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

    def _show(self):
        if self._tip is not None:
            return
        try:
            x = self.widget.winfo_rootx() + 12
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        except Exception:
            return
        self._tip = tk.Toplevel(self.widget)
        self._tip.wm_overrideredirect(True)
        self._tip.wm_geometry(f"+{x}+{y}")
        try:
            self._tip.attributes("-topmost", True)
        except Exception:
            pass
        tk.Label(self._tip, text=self.text,
                  bg="#FFFDE7", fg="#37474F",
                  relief="solid", bd=1, padx=8, pady=4,
                  font=("Segoe UI", 9), wraplength=340, justify="left"
                  ).pack()


def _tt(widget, text: str):
    """Kısa kısayol — widget'a tooltip ekler ve widget'ı geri döndürür."""
    _Tooltip(widget, text)
    return widget


# ═══════════════════════════════════════════════════════════════════════
# ANA SINIF
# ═══════════════════════════════════════════════════════════════════════
class AylikReceteSorguGUI:
    """Aylık reçete-ilaç inceleme & filtreleme tablosu."""

    STATE_DOSYASI = "aylik_inceleme_state.json"
    SUTUN_AYAR_DOSYASI = "aylik_inceleme_sutun_ayarlari.json"
    SUTUN_SABLON_DOSYASI = "aylik_inceleme_sutun_sablonlari.json"

    def __init__(self, root: tk.Tk, ana_menu_callback=None):
        self.root = root
        self.ana_menu_callback = ana_menu_callback

        self.root.title("📋  AYLIK İNCELEME & FİLTRELEME  ·  BOTANİK EOS  ·  SALT OKUNUR")
        # Geçici varsayılan boyut (zoomed çalışmazsa fallback)
        self.root.geometry("1700x920")
        # Pencere açılır açılmaz tam ekrana (maximized) gelsin
        try:
            # Windows ve çoğu modern Tk için
            self.root.state("zoomed")
        except tk.TclError:
            try:
                # Linux varyantı
                self.root.attributes("-zoomed", True)
            except Exception:
                # Son fallback: ekran boyutuna oturt
                try:
                    sw = self.root.winfo_screenwidth()
                    sh = self.root.winfo_screenheight()
                    self.root.geometry(f"{sw}x{sh}+0+0")
                except Exception:
                    pass
        try:
            from eczasist_ikon import ikon_uygula
            ikon_uygula(self.root)
        except Exception:
            pass

        self.db: BotanikDB = None

        # Lookup cache
        self._lookup_kapsam = {}
        self._lookup_brans = {}
        self._lookup_etkin_madde = {}
        self._lookup_alttur = {}
        self._lookup_renk = {}
        self._lookup_provizyon = {}
        self._lookup_rapor_kodu = {}
        self._lookup_kurum = {}
        self._lookup_hastane_kodu = {}  # HastaneId → HastaneKodu (tesis kodu)

        # Veri
        self.tum_satirlar = []        # tüm sorgu sonucu (filtresiz)
        self.gosterilen_iids = set()  # şu anda görünen iid'ler
        self.satir_indeks = {}        # {iid: satir_dict}
        self.satir_renkleri = {}      # {iid: renk}
        self.secili_iidler = set()    # checkbox ile seçilmiş satırlar

        # Filtre durumu
        self.aktif_sutun_filtre = {}  # {sutun_kod: arama_metni}
        self.aktif_deger_filtre = {}  # {sutun_kod: set(secili_degerler)}  Excel-benzeri
        # Çoklu koşul: {sutun_kod: [{"baglac": None|"ve"|"veya",
        #                            "op": "icerir"|..., "deger": "..."}, ...]}
        # İlk koşulun baglac'ı None. VEYA grup ayırıcı, VE grup içi.
        # Eşleşme = HERHANGİ bir grubun TÜM koşulları sağlanır.
        # (Geriye dönük: eski (op, deger) tuple tek-koşul listeye normalize edilir.)
        self.aktif_metin_filtre = {}
        self.aktif_renk_filtre = {RENK_BEYAZ, RENK_YESIL, RENK_SARI,
                                    RENK_TURUNCU, RENK_KIRMIZI, RENK_GRI}
        # "Boş satırları gizle" — uyarı kodu / mesaj / rapor kodu olmayan
        # satırları liste dışı tutar (kontrol gerektirmeyen temiz satırlar).
        # Varsayılan AÇIK: kullanıcı kontrol gerektirebilecek satırlara
        # odaklansın diye temiz satırlar baştan elenir.
        self.gizle_bos_satirlar = tk.BooleanVar(value=True)

        # Detaylı filtre ayarları (⚙ butonundan açılan pencere ile yönetilir)
        try:
            import aylik_filtre_ayarlari as fa
            self._filtre_ayarlari_aktif = fa.ayarlari_yukle()
        except Exception:
            self._filtre_ayarlari_aktif = None

        # Aktif dönem (yıl-ay) — state için anahtar
        self.aktif_donem = ""

        # Sütun görünüm ayarı — varsayılan tüm sütunlar açık
        self.sutun_gosterim = {kod: True for kod in SUTUN_KOD}
        self._sutun_ayarlarini_yukle()

        # Sıralama durumu (Excel benzeri tıkla-sırala)
        self._siralama_kolonu = None
        self._siralama_yonu = None  # "asc" | "desc"

        # Arayüz
        self._arayuz_olustur()

        # Son aktif sütun şablonunu uygula (varsa) — tablo kuruldu, widths hazır
        try:
            self._aktif_sablonu_uygula()
        except Exception as e:
            logger.debug("Açılışta şablon uygulama: %s", e)

        # Async başlat
        self.root.after(100, self._db_baglan_async)

    # ------------------------------------------------------------------ DB
    def _db_baglan_async(self):
        threading.Thread(target=self._db_baglan_ve_lookup, daemon=True).start()

    def _db_baglan_ve_lookup(self):
        try:
            self.db = BotanikDB(production=True)
            if not self.db.baglan():
                raise RuntimeError("Botanik veritabanına bağlanılamadı")
            self.root.after(0, self._durum_yaz, "DB bağlandı (salt-okunur). Lookup yükleniyor…")
            self._lookup_yukle()
            self.root.after(0, self._lookup_bitti)
        except Exception as e:
            logger.error("DB bağlantı hatası: %s", e)
            self.root.after(0, self._durum_yaz, f"DB hatası: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Bağlantı Hatası",
                f"Botanik EOS veritabanına bağlanılamadı:\n{e}"))

    def _lookup_yukle(self):
        if not self.db:
            return
        try:
            for r in self.db.sorgu_calistir("SELECT KapsamId, KapsamAdi FROM Kapsam"):
                self._lookup_kapsam[r["KapsamId"]] = r["KapsamAdi"]
            for r in self.db.sorgu_calistir(
                "SELECT BransId, BransAdi FROM Brans WHERE BransSilme=0"):
                self._lookup_brans[r["BransId"]] = r["BransAdi"]
            for r in self.db.sorgu_calistir(
                "SELECT EtkinMaddeId, EtkinMaddeSGKKodu, EtkinMaddeAdi "
                "FROM EtkinMadde WHERE EtkinMaddeSilme=0"):
                self._lookup_etkin_madde[r["EtkinMaddeId"]] = (
                    r["EtkinMaddeSGKKodu"] or "", r["EtkinMaddeAdi"] or "")
            for r in self.db.sorgu_calistir(
                "SELECT ReceteAltTuruId, ReceteAltTuruAdi FROM ReceteAltTuru "
                "WHERE ReceteAltTuruSilme=0"):
                self._lookup_alttur[r["ReceteAltTuruId"]] = r["ReceteAltTuruAdi"]
            for r in self.db.sorgu_calistir(
                "SELECT ReceteRenkId, ReceteRenkAdi FROM ReceteRenk "
                "WHERE ReceteRenkSilme=0"):
                self._lookup_renk[r["ReceteRenkId"]] = r["ReceteRenkAdi"]
            # Kırmızı/Yeşil/Mor reçete tipi ID setleri (renkli reçete kontrolü
            # için her zaman gelmeli — boş satır filtresinden muaf)
            self._renkli_recete_idler = {
                rid: ad for rid, ad in self._lookup_renk.items()
                if ad and ad.strip().lower() in ("kırmızı", "kirmizi",
                                                    "yeşil", "yesil", "mor")
            }
            for r in self.db.sorgu_calistir(
                "SELECT ProvizyonTipId, ProvizyonTipAdi FROM ProvizyonTipi "
                "WHERE ProvizyonTipSilme=0"):
                self._lookup_provizyon[r["ProvizyonTipId"]] = r["ProvizyonTipAdi"]
            for r in self.db.sorgu_calistir(
                "SELECT RaporKodId, RaporKodu, RaporKodAciklama FROM RaporKodlari "
                "WHERE RaporKodSilme=0"):
                self._lookup_rapor_kodu[r["RaporKodId"]] = (
                    r["RaporKodu"] or "", r["RaporKodAciklama"] or "")
            for r in self.db.sorgu_calistir(
                "SELECT KurumId, KurumAdi FROM Kurum WHERE KurumSilme=0"):
                self._lookup_kurum[r["KurumId"]] = r["KurumAdi"]
            for r in self.db.sorgu_calistir(
                "SELECT HastaneId, HastaneKodu FROM Hastane "
                "WHERE HastaneSilme=0 AND HastaneKodu IS NOT NULL"):
                self._lookup_hastane_kodu[r["HastaneId"]] = (
                    r["HastaneKodu"] or "")
        except Exception as e:
            logger.exception("Lookup hatası: %s", e)

    def _lookup_bitti(self):
        self._durum_yaz(
            f"Hazır — {len(self._lookup_etkin_madde)} etkin madde, "
            f"{len(self._lookup_brans)} branş, "
            f"{len(self._lookup_rapor_kodu)} rapor kodu yüklendi"
        )
        self._ay_listesini_yukle()

    # --------------------------------------------------------------- LAYOUT
    def _arayuz_olustur(self):
        # ═══════════════════════════════════════════════════════════════════
        # 2 KATMAN — ÜST: Veri/Sayım/Araçlar | ALT: Renkli kontrol panelleri
        # ═══════════════════════════════════════════════════════════════════

        HEADER_BG = "#37474F"
        HEADER_FG = "#FFFFFF"
        HEADER_LBL_FG = "#B0BEC5"
        FONT_PRIMARY = ("Segoe UI", 11, "bold")
        FONT_GROUP = ("Segoe UI", 9, "bold")
        FONT_BUTON = ("Segoe UI", 9, "bold")
        FONT_LABEL_SM = ("Segoe UI", 8, "bold")

        renk_kisa = {
            RENK_YESIL:   "🟢", RENK_SARI:    "🟡",
            RENK_TURUNCU: "🟠", RENK_KIRMIZI: "🔴",
            RENK_BEYAZ:   "⚪", RENK_GRI:     "🩶",
        }
        renk_aciklama = {
            RENK_YESIL:   "YEŞİL — Uygun (elendi)",
            RENK_SARI:    "SARI — Manuel kontrol gerekli",
            RENK_TURUNCU: "TURUNCU — Şüpheli",
            RENK_KIRMIZI: "KIRMIZI — Uygunsuz",
            RENK_BEYAZ:   "BEYAZ — Renksiz / sıfırla",
            RENK_GRI:     "GRİ — Kontrol gerekmez "
                            "(ilaç dışlama listesine eklenir)",
        }
        renk_etiket = {
            RENK_BEYAZ:   "⚪ Beyaz",
            RENK_YESIL:   "🟢 Yeşil",
            RENK_SARI:    "🟡 Sarı",
            RENK_TURUNCU: "🟠 Turuncu",
            RENK_KIRMIZI: "🔴 Kırmızı",
            RENK_GRI:     "🩶 Gri",
        }

        # ════════════════════════════════════════════════════════════════
        # ÜST SATIR — DATA + STATUS + TOOLS (koyu zemin)
        # ════════════════════════════════════════════════════════════════
        row1 = tk.Frame(self.root, bg=HEADER_BG, height=44)
        row1.pack(fill="x")
        row1.pack_propagate(False)

        # ── SOL: DÖNEM grubu (Yıl + Ay + SORGULA) ──
        sol = tk.Frame(row1, bg=HEADER_BG)
        sol.pack(side="left", padx=10, pady=5)

        tk.Label(sol, text="📅 DÖNEM", bg=HEADER_BG, fg=HEADER_LBL_FG,
                 font=FONT_LABEL_SM).pack(side="left", padx=(0, 8))

        tk.Label(sol, text="Yıl", bg=HEADER_BG, fg=HEADER_FG,
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 3))
        self.var_yil = tk.StringVar()
        self.cb_yil = ttk.Combobox(sol, textvariable=self.var_yil,
                                    width=6, state="readonly")
        self.cb_yil.pack(side="left", padx=2)
        self.cb_yil.bind("<<ComboboxSelected>>", self._yil_degisti)
        _Tooltip(self.cb_yil,
                  "Yıl seçimi.\n'Tümü' = tüm yıllar (büyük veri seti)")

        tk.Label(sol, text="Ay", bg=HEADER_BG, fg=HEADER_FG,
                 font=("Segoe UI", 9)).pack(side="left", padx=(8, 3))
        self.var_ay = tk.StringVar()
        self.cb_ay = ttk.Combobox(sol, textvariable=self.var_ay,
                                   width=12, state="readonly")
        self.cb_ay.pack(side="left", padx=2)
        _Tooltip(self.cb_ay,
                  "Ay seçimi.\n'Tümü' = seçili yıldaki tüm aylar")

        btn = tk.Button(sol, text="🔍 SORGULA", bg="#1976D2", fg="white",
                         activebackground="#1565C0", bd=0,
                         command=self._receteleri_sorgula,
                         padx=14, pady=4, font=FONT_PRIMARY)
        btn.pack(side="left", padx=(10, 0))
        _Tooltip(btn,
                  "Seçili yıl/ay için Botanik EOS'tan reçete-ilaç satırlarını "
                  "yükle (sadece okuma).")

        # ── SAĞ: ARAÇLAR (Excel + Sütunlar) ──
        sag = tk.Frame(row1, bg=HEADER_BG)
        sag.pack(side="right", padx=10, pady=5)

        btn = tk.Button(sag, text="📊 Excel", bg="#43A047", fg="white",
                         activebackground="#388E3C", bd=0,
                         command=self._excel_export, padx=12, pady=3,
                         font=FONT_BUTON)
        btn.pack(side="right", padx=2)
        _Tooltip(btn, "Tabloyu (renk + sütunlarla) Excel dosyasına aktar")

        btn = tk.Button(sag, text="⚙ Sütunlar", bg="#546E7A", fg="white",
                         activebackground="#455A64", bd=0,
                         command=self._sutun_ayar_penceresi, padx=12, pady=3,
                         font=FONT_BUTON)
        btn.pack(side="right", padx=2)
        _Tooltip(btn, "Görünür sütunları ayarla (göster/gizle)")

        # ── ORTA: SAYIM (lbl_sayim) ──
        self.lbl_sayim = tk.Label(row1, text="", bg=HEADER_BG, fg=HEADER_FG,
                                    font=("Segoe UI", 10, "bold"))
        self.lbl_sayim.pack(side="left", padx=20, expand=True)
        _Tooltip(self.lbl_sayim,
                  "Toplam yüklenen / filtrelenmiş satır sayısı")

        # ════════════════════════════════════════════════════════════════
        # ALT SATIR(LAR) — KONTROL PANELLERİ + HÜCRE GÖSTERGESİ
        # Sol kolon (2 sıra):  Sıra A → Seçim + Boyama
        #                      Sıra B → Göster + Filtreleri Sıfırla
        # Sağ kolon (2 sıra yüksekliğinde): Hücre içeriği göstergesi
        # ════════════════════════════════════════════════════════════════
        ust_alan = tk.Frame(self.root, bg="#FAFAFA")
        ust_alan.pack(fill="x", padx=4, pady=(2, 2))

        sol_kol = tk.Frame(ust_alan, bg="#FAFAFA")
        sol_kol.pack(side="left", fill="y")

        row2 = tk.Frame(sol_kol, bg="#FAFAFA")
        row2.pack(fill="x", pady=(0, 2))

        row3 = tk.Frame(sol_kol, bg="#FAFAFA")
        row3.pack(fill="x", pady=(0, 0))

        sag_kol = tk.Frame(ust_alan, bg="#FAFAFA")
        sag_kol.pack(side="left", fill="both", expand=True, padx=(2, 0))

        def _panel_olustur(parent, baslik, baslik_fg, panel_bg):
            """Renkli arka planlı, başlık etiketli kompakt panel."""
            p = tk.Frame(parent, bg=panel_bg, bd=1, relief="solid")
            p.pack(side="left", padx=2, pady=1, fill="y")
            tk.Label(p, text=baslik, bg=panel_bg, fg=baslik_fg,
                     font=FONT_GROUP).pack(side="left", padx=(6, 6),
                                              pady=4)
            return p

        # ─── PANEL 1: SEÇİM (açık mavi) ───
        P_SECIM_BG = "#E3F2FD"
        p1 = _panel_olustur(row2, "🎯 Seçim", "#0D47A1", P_SECIM_BG)
        for txt, cmd, tip in [
            ("☑ Tümü", self._tumunu_sec,
             "Filtreden geçen tüm satırları SEÇ"),
            ("☐ Hiçbiri", self._hicbirini_secme,
             "Filtreden geçen satırların seçimini KALDIR"),
            ("↻ Tersine", self._secimi_tersine_cevir,
             "Seçimi tersine çevir (filtreli satırlarda)"),
            ("✓ Vurgulu", self._vurgulanmislari_isaretle,
             "Shift/Ctrl ile vurgulanmış satırların\n"
             "checkbox'larını işaretle"),
        ]:
            b = tk.Button(p1, text=txt, bg="white", bd=1,
                          command=cmd, padx=8, pady=3, font=FONT_BUTON)
            b.pack(side="left", padx=2, pady=4)
            _Tooltip(b, tip)
        tk.Frame(p1, bg=P_SECIM_BG, width=4).pack(side="left")

        # ─── PANEL 2: BOYAMA (açık sarı) — TEK PARÇA + radio mod selector ───
        P_BOYA_BG = "#FFF8E1"
        p2 = _panel_olustur(row2, "🎨 Boyama", "#E65100", P_BOYA_BG)

        # Mod seçici: Seçili / Filtreli (radiobutton — yer kazanır)
        self.boya_mod = tk.StringVar(value="secili")
        mod_frame = tk.Frame(p2, bg=P_BOYA_BG)
        mod_frame.pack(side="left", padx=(0, 6), pady=2)
        rb_secili = tk.Radiobutton(
            mod_frame, text="Seçili",
            variable=self.boya_mod, value="secili",
            bg=P_BOYA_BG, fg="#5D4037",
            selectcolor="#FFFFFF",
            font=("Segoe UI", 8, "bold"),
            bd=0, anchor="w")
        rb_secili.pack(side="top", anchor="w")
        rb_filtreli = tk.Radiobutton(
            mod_frame, text="Filtreli",
            variable=self.boya_mod, value="filtreli",
            bg=P_BOYA_BG, fg="#5D4037",
            selectcolor="#FFFFFF",
            font=("Segoe UI", 8, "bold"),
            bd=0, anchor="w")
        rb_filtreli.pack(side="top", anchor="w")
        _Tooltip(rb_secili, "Seçili: SADECE tabloda işaretli satırlar boyanır")
        _Tooltip(rb_filtreli,
                  "Filtreli: filtre sonucunda görünen TÜM satırlar boyanır\n"
                  "(scroll dışındakiler dahil)")

        # Tek satır 6 renk butonu — radio'ya göre çalışır
        btn_frame = tk.Frame(p2, bg=P_BOYA_BG)
        btn_frame.pack(side="left", padx=(0, 6), pady=4)
        for renk in [RENK_YESIL, RENK_SARI, RENK_TURUNCU,
                       RENK_KIRMIZI, RENK_BEYAZ, RENK_GRI]:
            b = tk.Button(btn_frame, text=renk_kisa[renk],
                          bg=RENK_BG[renk], fg=RENK_FG[renk], bd=1,
                          command=lambda r=renk: self._boya_dispatch(r),
                          padx=4, width=2, font=("Segoe UI", 11))
            b.pack(side="left", padx=1)
            _Tooltip(b,
                      f"Boya: {renk_aciklama[renk]}\n"
                      "Mod (sol): Seçili = işaretli satırlar / "
                      "Filtreli = görünen tümü")

        # ─── PANEL 3: FİLTRE (açık yeşil) — alt sıraya geçiyor ───
        P_FLT_BG = "#E8F5E9"
        p3 = _panel_olustur(row3, "👁 Göster", "#1B5E20", P_FLT_BG)

        # Label'sız renkli kareler (yer kazanmak için — UX karar 2026-05-08)
        # Renk kendisi bilgi taşır (Gestalt similarity). Tooltip ile ad açıklanır.
        self.var_renk = {}
        for renk in [RENK_BEYAZ, RENK_YESIL, RENK_SARI,
                       RENK_TURUNCU, RENK_KIRMIZI, RENK_GRI]:
            v = tk.BooleanVar(value=True)
            cb = tk.Checkbutton(p3, text="",  # label kaldırıldı (yer için)
                                  variable=v, bg=RENK_BG[renk],
                                  fg=RENK_FG[renk],
                                  selectcolor=RENK_BG[renk],
                                  font=("Segoe UI", 9, "bold"),
                                  padx=2, bd=1, relief="solid", width=2,
                                  command=self._renk_filtre_degisti)
            cb.pack(side="left", padx=1, pady=4)
            self.var_renk[renk] = v
            _Tooltip(cb,
                      f"{renk_etiket[renk]}\n"
                      f"{renk_aciklama[renk]}\n"
                      "renkli satırları tabloda GÖSTER / GİZLE")

        # 🧹 Filtreleri Sıfırla — Göster panelinin İÇİNE taşındı (kullanıcı isteği)
        btn_sifirla_p3 = tk.Button(
            p3, text="🧹 Sıfırla",
            bg="#FFE082", fg="#5D4037",
            activebackground="#FFD54F", bd=1,
            command=self._tum_filtreleri_temizle,
            font=FONT_BUTON, padx=8, pady=2)
        btn_sifirla_p3.pack(side="left", padx=(8, 6), pady=4)
        _Tooltip(btn_sifirla_p3,
                  "Tüm filtreleri SIFIRLA:\n"
                  "• Sütun arama kutuları\n"
                  "• Renk filtresi (6 renk açık)\n"
                  "• Excel-benzeri değer/metin filtreleri\n"
                  "• Sıralama")

        # 🚫 Boş Satırları Gizle + ⚙ Ayar butonu KALDIRILDI — yerine
        # ayrı bir "📁 Filtrelemeler" kompozit dikdörtgen butonu eklendi.

        # Reçete türü filtresi kaldırıldı; ilgili kod yolları boş dict ile çalışır.
        # Hücre içeriği göstergesi sağ kolonda (sag_kol) oluşturulur — aşağıda.
        self.var_recete_turu = {}

        # ─── PANEL: FİLTRELEMELER kompozit (Kullan + Filtrelemeler + Sıfırla)
        # Tek bir mavi dikdörtgen panel — soldan sağa 3 widget
        P_FLT_BG2 = "#E3F2FD"
        p_flt = tk.Frame(row3, bg=P_FLT_BG2, bd=1, relief="solid")
        p_flt.pack(side="left", padx=2, pady=1, fill="y")

        cb_flt = tk.Checkbutton(
            p_flt, text="Kullan",
            variable=self.gizle_bos_satirlar,
            bg=P_FLT_BG2, fg="#1565C0",
            selectcolor="#FFFFFF",
            font=FONT_GROUP, padx=4, bd=0,
            command=self._bos_satir_filtre_degisti)
        cb_flt.pack(side="left", padx=(6, 0), pady=4)
        _Tooltip(cb_flt,
                  "Filtre kurallarının AÇIK/KAPALI anahtarı.\n\n"
                  "✅ İŞARETLİ: tanımlanan kurallar SQL'e uygulanır\n"
                  "  (uyarı/mesaj/rapor + ilaç/etken/ATC/farma/tesis filtreleri)\n\n"
                  "☐ İŞARETSİZ: hiçbir filtre uygulanmaz")

        btn_flt_ayar = tk.Button(
            p_flt, text="📁 Filtrelemeler",
            bg="#1976D2", fg="white",
            activebackground="#1565C0",
            bd=0, padx=10, pady=4,
            font=FONT_BUTON, cursor="hand2",
            command=self._filtre_ayar_penceresi_ac)
        btn_flt_ayar.pack(side="left", padx=(2, 4), pady=4)
        _Tooltip(btn_flt_ayar,
                  "Tıklayınca filtre ayar penceresi açılır:\n"
                  "• Hangi içerikler gelsin (renkli reçete / mesaj / uyarı / rapor)\n"
                  "• İlaç / etken / ATC / farma / tesis / eşdeğer kuralları\n"
                  "• 🚫 Kurum (Dışlamalar) sekmesi\n\n"
                  "Üst checkbox: kuralları aktif/pasif yapar.")

        # 🧹 Sıfırla — Göster panelinin içine taşındı (yukarıda)

        # ─── PANEL: KONTROL BUTONLARI — Filtrelemeler'in YANINDA, sağ tarafta ───
        # İçinde: ▢ Sadece seçilenler checkbox + 🎯 Kontrol Butonları butonu
        P_KTR_BG = "#E8EAF6"
        p_kontrol = tk.Frame(row3, bg=P_KTR_BG, bd=1, relief="solid")
        p_kontrol.pack(side="left", padx=2, pady=1, fill="y")

        # Checkbox: "Sadece seçilenler" — 2 satır (yer kazanmak için)
        self.kontrol_sadece_secili = tk.BooleanVar(value=False)
        cb_secili = tk.Checkbutton(
            p_kontrol, text="Sadece\nseçilenler",
            variable=self.kontrol_sadece_secili,
            bg=P_KTR_BG, fg="#1A237E",
            selectcolor="#FFFFFF",
            font=("Segoe UI", 8, "bold"),
            padx=2, bd=0, justify="left", anchor="w")
        cb_secili.pack(side="left", padx=(6, 0), pady=2)
        _Tooltip(cb_secili,
                  "✅ İŞARETLİ: kontrol sadece TABLODA SEÇİLİ reçetelere\n"
                  "  uygulanır (tablo satırlarını ☑ ile işaretle ya da\n"
                  "  Ctrl+Click ile çoklu seç)\n\n"
                  "☐ İŞARETSİZ: kontrol tablodaki TÜM reçetelere uygulanır")

        btn_kontrol = tk.Button(p_kontrol, text="🎯 Kontrol Butonları",
                                 bg="#283593", fg="white",
                                 activebackground="#1A237E", bd=1,
                                 command=self._kontrol_butonlari_popup_ac,
                                 font=FONT_BUTON, padx=10, pady=4,
                                 cursor="hand2")
        btn_kontrol.pack(side="left", padx=(2, 6), pady=4)
        _Tooltip(btn_kontrol,
                 "15 SUT/uyarı kontrol butonunu içeren popup açılır.\n"
                 "Bir kontrol seç → otomatik çalışır.\n\n"
                 "▢ 'Sadece seçilenler' işaretliyse SEÇİLİ satırlara uygulanır.")

        # ─── HÜCRE İÇERİĞİ GÖSTERGESİ (sağ kolon, 2 sıra yüksekliğinde) ───
        # Tabloda bir hücreye tıklanınca içeriği burada gösterilir.
        # Yatayda Boyama / Filtreleri Sıfırla bittiği yerden sağ kenara kadar,
        # düşeyde row2 + row3 yüksekliği boyunca uzanır.
        self._hucre_frame = tk.Frame(sag_kol, bg="#FFF9C4",
                                       bd=1, relief="solid")
        self._hucre_frame.pack(fill="both", expand=True)

        ust_satir = tk.Frame(self._hucre_frame, bg="#FFF9C4")
        ust_satir.pack(fill="x", side="top")
        tk.Label(ust_satir, text="📋",
                 bg="#FFF9C4", fg="#5D4037",
                 font=("Segoe UI", 10, "bold"),
                 padx=6).pack(side="left")
        self.lbl_hucre_baslik = tk.Label(
            ust_satir, text="—",
            bg="#FFF9C4", fg="#1565C0",
            font=("Segoe UI", 9, "bold"), padx=4, anchor="w")
        self.lbl_hucre_baslik.pack(side="left", fill="x", expand=True)

        self.lbl_hucre_icerik = tk.Label(
            self._hucre_frame,
            text="(Hücreye tıkla → içeriği burada görüntülenir, "
                 "uzun metinler alt satırlara sarmalanır)",
            bg="#FFF9C4", fg="#37474F",
            font=("Segoe UI", 9), anchor="nw", padx=8, pady=4,
            justify="left", wraplength=600)
        self.lbl_hucre_icerik.pack(fill="both", expand=True,
                                     padx=2, pady=(0, 2))

        def _hucre_wrap_guncelle(_event=None):
            try:
                w = self._hucre_frame.winfo_width() - 30
                if w > 100:
                    self.lbl_hucre_icerik.config(wraplength=w)
            except Exception:
                pass
        self._hucre_frame.bind("<Configure>", _hucre_wrap_guncelle)

        # ───── ANA TABLO (artık tüm genişlikte) ─────
        tablo_frame = tk.Frame(self.root, bg="white")
        tablo_frame.pack(fill="both", expand=True, padx=6, pady=(2, 4))
        self._tablo_kur(tablo_frame)

        # ───── DURUM ÇUBUĞU ─────
        # NOT: 3 buton satırı (row_kontrol/sut/diger) ARTIK GİZLİ — popup'a alındı.
        # Sadece row_durum (durum mesajı + sayaç) ekranda görünür.
        # Butonlar widget olarak alive kalır (popup invoke için).
        self._durum_frame = tk.Frame(self.root, bg="#ECEFF1")
        self._durum_frame.pack(fill="x", side="bottom")

        # 3 buton satırı (gizli — popup'tan invoke edilir) + 1 durum mesajı
        self._row_kontrol_btn = tk.Frame(self._durum_frame, bg="#ECEFF1")
        # NOT: pack EDİLMEZ — gizli kalır, butonlar oluşturulur
        self._row_sut_btn = tk.Frame(self._durum_frame, bg="#ECEFF1")
        self._row_diger_btn = tk.Frame(self._durum_frame, bg="#ECEFF1")
        self._row_diger2_btn = tk.Frame(self._durum_frame, bg="#ECEFF1")
        # Backward-compat: eski kod row_kontrol vs adlarını kullanıyor
        row_kontrol = self._row_kontrol_btn
        row_sut = self._row_sut_btn
        row_diger = self._row_diger_btn
        row_diger2 = self._row_diger2_btn
        # Durum mesajı satırı (görünür kalır)
        row_durum = tk.Frame(self._durum_frame, bg="#ECEFF1")
        row_durum.pack(fill="x", side="top")

        # ── SATIR 1: UYARI KODU KONTROL butonu ──
        # Yüklenen tüm satırlardaki reçete uyarı kodlarını (256, 280 vb.)
        # alır; her kodun açıklamasındaki anahtar kelimeleri reçete teşhisi,
        # rapor teşhisi, reçete açıklaması ve rapor açıklaması içinde arar.
        # Sonuç "Uyarı Kod Karş." sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ
        # olarak yazılır; Excel raporu üretilir.
        self.btn_uyari_kod = tk.Button(
            row_kontrol, text="🔍 UYARI KODU KONTROL",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#F57F17", activebackground="#E65100",  # amber — uyarı
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._uyari_kod_kontrol_baslat
        )
        self.btn_uyari_kod.pack(side="left", padx=(4, 4), pady=2)

        # ── DOZ KARŞILAŞTIR butonu ──
        # Yüklenen tüm satırlar için reçete günlük dozu ile rapor günlük
        # dozunu (RaporEtkinMadde + ek bilgi parse) karşılaştırır.
        # Sonuç "Doz Karş." sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ olarak
        # yazılır; Excel raporu üretilir.
        self.btn_doz_kontrol = tk.Button(
            row_kontrol, text="📏 DOZ KARŞILAŞTIR",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#FF8F00", activebackground="#E65100",  # turuncu — doz
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._doz_kontrol_baslat
        )
        self.btn_doz_kontrol.pack(side="left", padx=(0, 4), pady=2)

        # ── RENKLİ REÇETE SİSTEMİ KONTROL butonu ──
        # Yüklenen tüm satırlardan Kırmızı/Yeşil/Mor reçete türündekileri
        # alır; ana ekranda yüklenmiş "Renkli Reçete Listesi" (PDF/Excel)
        # ile karşılaştırır. Sonuç "Renkli RR" sütununa VAR / YOK / —
        # olarak yazılır; YOK satırlar kırmızıya boyanır; Excel raporu üretilir.
        self.btn_renkli_rr = tk.Button(
            row_kontrol, text="🔴🟢🟣 RENKLİ REÇETE",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#7B1FA2", activebackground="#4A148C",  # mor — renkli
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._renkli_recete_kontrol_baslat
        )
        self.btn_renkli_rr.pack(side="left", padx=(0, 4), pady=2)

        # ── SATIR 2: STATİN / LİPİD KONTROL butonu ──
        # Buton tıklanınca yüklenen satırlardan ATC C10* (statin + non-statin
        # lipid) olanlar SUT 4.2.28.A/B kuralına göre denetlenir; sonuç en sağdaki
        # SONUÇ sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ olarak yazılır.
        self.btn_statin = tk.Button(
            row_sut, text="🩺 STATİN KONTROL",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#C62828", activebackground="#8E0000",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._statin_kontrol_baslat
        )
        self.btn_statin.pack(side="left", padx=(4, 4), pady=2)

        # ── DİYABET KONTROL butonu (SUT 4.2.38) ──
        # ATC A10* (insülin + oral antidiyabetikler) — DPP-4/SGLT-2/GLP-1 RA
        # + klasik OAD + insülin SUT 4.2.38'e göre denetlenir.
        self.btn_diyabet = tk.Button(
            row_sut, text="💉 DİYABET KONTROL",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#0D47A1", activebackground="#002171",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._diyabet_kontrol_baslat
        )
        self.btn_diyabet.pack(side="left", padx=(0, 4), pady=2)

        # ── KLOPİDOGREL KONTROL butonu (SUT 4.2.15) ──
        # ATC B01AC04/22/24 — klopidogrel/prasugrel/tikagrelor (P2Y12).
        # Endikasyon (stent/AKS/MI/KAH/inme/PAH) + ASA intoleransı + 12 ay
        # stent kuralı + kombinasyon yasağı (P2Y12+YOAK).
        self.btn_klopidogrel = tk.Button(
            row_sut, text="💊 KLOPİDOGREL",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#6A1B9A", activebackground="#4A148C",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._klopidogrel_kontrol_baslat
        )
        self.btn_klopidogrel.pack(side="left", padx=(0, 4), pady=2)

        # ── ARB KONTROL butonu (SUT EK-4/F Madde 51 / 1300/51) ──
        # ATC C09C* (mono ARB) / C09D* (ARB kombinasyonları) / C02AC*
        # (Rilmeniden/Moksonidin) — mono raporlu / kombi monoterapi ibaresi /
        # raporsuz aile hekimi + 1 kutu kuralları.
        self.btn_arb = tk.Button(
            row_sut, text="🩸 ARB (M.51)",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#00695C", activebackground="#004D40",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._arb_kontrol_baslat
        )
        self.btn_arb.pack(side="left", padx=(0, 4), pady=2)

        # ── KEMİK ERİMESİ / OSTEOPOROZ KONTROL butonu ──
        # ATC M05BA/BB/BX (bifosfonat + biyolojik) + H05AA (teriparatid) +
        # G03XC (raloksifen) — SUT 4.2.17 (T-skor/kırık eşiği) ve SUT 4.2.28.C
        # (biyolojik için uzman raporu) kurallarına göre denetlenir.
        self.btn_osteo = tk.Button(
            row_sut, text="🦴 KEMİK ERİMESİ",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#5D4037", activebackground="#3E2723",  # kahve — kemik
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._osteo_kontrol_baslat
        )
        self.btn_osteo.pack(side="left", padx=(0, 4), pady=2)

        # ── SATIR 3: YOAK / DOAK KONTROL butonu (SUT 4.2.15.D-1 / D-2) ──
        # ATC B01AE07 (dabigatran) / B01AF01 (rivaroksaban) / B01AF02 (apiksaban)
        # / B01AF03 (edoksaban). D-1: Non-valvüler AF + varfarin/INR şartı.
        # D-2: DVT/PE tedavisi/profilaksisi (varfarin şartı aranmaz). Kombine
        # YOAK+YOAK ve YOAK+P2Y12 yasak → otomatik UYGUN DEĞİL.
        self.btn_yoak = tk.Button(
            row_diger, text="🩸 YOAK (D-1/D-2)",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#283593", activebackground="#1A237E",  # indigo — koagülasyon
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._yoak_kontrol_baslat
        )
        self.btn_yoak.pack(side="left", padx=(4, 4), pady=2)

        # ── ÇEŞİTLİ İLAÇLAR butonu (SUT M.45 / M.2 / BPH α-bloker) ──
        # Tek buton, 3 alt grup dispatcher:
        #   1. Üriner inkontinans (M.45) — antimuskarinik / Mirabegron / Duloksetin
        #   2. Suni gözyaşı (M.2)        — keratitis sicca / kuru göz
        #   3. BPH α-bloker              — Alfuzosin/Tamsulosin/Terazosin/Doksazosin/Silodosin
        self.btn_cesitli = tk.Button(
            row_diger, text="🧪 ÇEŞİTLİ",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#6A1B9A", activebackground="#4A0072",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._cesitli_kontrol_baslat
        )
        self.btn_cesitli.pack(side="left", padx=(0, 4), pady=2)

        # ── PSİKİYATRİ / NÖROLOJİ KONTROL butonu (SUT 4.2.2 + 4.2.25) ──
        # ATC N05A* (antipsikotik), N06A* (antidepresan), N05B* (anksiyolitik),
        # N03A* (antiepileptik). Endikasyon-bazlı dispatch:
        #   - Antiepileptik kategorisi → kontrol_antiepileptik_4_2_25
        #   - Psikiyatri kategorisi   → kontrol_psikiyatri
        #   - Lamotrijin/Valproat bipolar → psikiyatri fallback
        self.btn_psikiyatri = tk.Button(
            row_diger, text="🧠 PSİKİYATRİ/NÖROLOJİ",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#7B1FA2", activebackground="#4A148C",  # mor
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._psikiyatri_kontrol_baslat
        )
        self.btn_psikiyatri.pack(side="left", padx=(0, 4), pady=2)

        # ── ENTERAL BESLENME KONTROL butonu ──
        # ATC V06D* + ürün ailesi (Resource/Nutridrink/Nutren/Fresubin/
        # Evolvia/Ensure/Peptamen/Nepro/Impact/Prosure/Modulen/Glucerna/
        # Diasip/Cubitan/Abound/Juven/Fortimel/Pediasure).
        # Endikasyon + uzman raporu + kalori planı kontrolü.
        self.btn_enteral = tk.Button(
            row_diger, text="🥛 ENTERAL BESLENME",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#5D4037", activebackground="#3E2723",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._enteral_kontrol_baslat
        )
        self.btn_enteral.pack(side="left", padx=(0, 4), pady=2)

        # ── HEPATİT B/C KONTROL butonu (SUT 4.2.13.3 / 4.2.13.4) ──
        # ATC J05AF (HBV nükleos(t)id) / J05AP (HCV DAA) / J05AB04 (Ribavirin)
        # / L03AB10/11 (Peginterferon alfa). HIV ilaçları KAPSAM DIŞI.
        # Endikasyon (HBV/HCV) + viral yük (HBV DNA / HCV RNA) + uzman branş
        # (Gastroenteroloji/Enfeksiyon/Hepatoloji) kontrolü.
        self.btn_hepatit = tk.Button(
            row_diger, text="🦠 HEPATİT B/C",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#1B5E20", activebackground="#0F3D14",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._hepatit_kontrol_baslat
        )
        self.btn_hepatit.pack(side="left", padx=(0, 4), pady=2)

        # ── ASTIM / KOAH KONTROL butonu (SUT 4.2.24 / 4.2.24.B) ──
        # ATC R03* (solunum sistemi obstrüktif hastalık ilaçları): SABA / SAMA
        # / ICS / LABA / LAMA / LABA+ICS / LABA+LAMA / üçlü (LABA+ICS+LAMA) /
        # LTRA (Montelukast/Zafirlukast) / Omalizumab (Anti-IgE) / Anti-IL5
        # (Mepo/Benral) / Anti-IL4 (Dupilumab) / Roflumilast / Teofilin /
        # Kromolin. Astım/KOAH tanısı + uzman branş + üçlü için 3 ay ICS+LABA
        # başarısızlığı + atak/mMRC/CAT şartları kontrol_solunum() ile denetlenir.
        self.btn_astim_koah = tk.Button(
            row_diger, text="🌬️ ASTIM/KOAH",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#00838F", activebackground="#005662",
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._astim_koah_kontrol_baslat
        )
        self.btn_astim_koah.pack(side="left", padx=(0, 4), pady=2)

        # ── NÖROPATİK / FİBROMİYALJİ KONTROL butonu (SUT 4.2.35.A + B) ──
        # Pregabalin / Gabapentin / Duloksetin / Alfa lipoik / Kapsaisin krem.
        # Endikasyon ayrımı (rap_ack + teshisler):
        #   - Fibromiyalji (M79.7)            → 4.2.35.B
        #   - Nöropatik / PHN / Diyabetik     → 4.2.35.A
        #   - Epilepsi (Pregab/Gabap)         → ATLANDI (4.2.25 antiepileptik)
        #   - Depresyon (Duloksetin)          → ATLANDI (4.2.2 psikiyatri)
        # Kombi yasağı: Pregabalin + Gabapentin AYNI REÇETE (cross-reçete YOK).
        # Branş tespit edilemezse "manuel doğrulanmalı" uyarısı.
        self.btn_noropatik_435 = tk.Button(
            row_diger, text="🧠 NÖROPATİK 4.2.35",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#AD1457", activebackground="#880E4F",  # pembe
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._noropatik_435_kontrol_baslat
        )
        self.btn_noropatik_435.pack(side="left", padx=(0, 8), pady=2)

        # ── SATIR 4: ANTİHT (mono+kombi) butonu (SUT 4.2.12.B) ──
        # ATC C09A* (ACE-i), C09C* (ARB), C03* (diüretik), C07* (BB), C08*
        # (KKB) — mono. C09B*/C09D* — kombi. Endikasyon-bazlı dispatch:
        #   - MONO_ANTIHIPERTANSIF   → kontrol_mono_antihipertansif
        #     (6 alt grup: ACE-i / ARB / KKB / BB / Diüretik / Alfa bloker;
        #      raporsuz uyumluluk + ACE+ARB kontrendikasyon + doz aşımı)
        #   - KOMBINE_ANTIHIPERTANSIF → kontrol_kombine_antihipertansif
        #     (raporsuz/04.05/monoterapi ibaresi 3-yollu kontrol)
        # Kategori tespiti: sut_kategorisi_tespit_et() (ATC + etken madde +
        # SUT_MADDESI_KATEGORI mapping) — kontrolü kategoriye göre dispatch eder.
        self.btn_antiht = tk.Button(
            row_diger2, text="🩸 ANTİHT (mono+kombi)",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#01579B", activebackground="#01428A",  # koyu mavi
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._antiht_kontrol_baslat
        )
        self.btn_antiht.pack(side="left", padx=(4, 4), pady=2)

        # ── SATIR 4: İSKEMİK KALP butonu (SUT 4.2.15.C / 4.2.15.F) ──
        # ATC C01EB17 (İvabradin) / C01EB18 (Ranolazin) / C03DA04 (Eplerenon).
        # Endikasyon-bazlı dispatch:
        #   - IVABRADIN/EPLERENON  → kontrol_ivabradin (NYHA + EF≤45% + BB
        #                            intoleransı + sinüs ritmi + angina/KY)
        #   - RANOLAZIN            → kontrol_ranolazin (kronik stabil angina
        #                            + BB/CCB intoleransı veya yetersiz yanıt)
        self.btn_iskemik_kalp = tk.Button(
            row_diger2, text="❤️ İSKEMİK KALP",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#B71C1C", activebackground="#7F0000",  # koyu kırmızı
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._iskemik_kalp_kontrol_baslat
        )
        self.btn_iskemik_kalp.pack(side="left", padx=(0, 4), pady=2)

        # ── K-SİTRAT butonu (Potasyum Sitrat — SUT EK-4/F) ──
        # ATC A12BA02. Üroloji/nefroloji uzmanı raporu zorunlu.
        # ICD: N20.0/N20.1 (böbrek taşı), N25.8/N25.9 (RTA — renal tübüler
        # asidoz). Raporlu + uzman branş + ICD eşleşmesi kontrolü.
        self.btn_potasyum_sitrat = tk.Button(
            row_diger2, text="🧂 K-SİTRAT",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#0277BD", activebackground="#01579B",  # mavi
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._potasyum_sitrat_kontrol_baslat
        )
        self.btn_potasyum_sitrat.pack(side="left", padx=(0, 4), pady=2)

        # ── RAPORSUZ BİLGİ butonu (Raporsuz verilebilen ilaçlar) ──
        # kontrol_raporsuz_bilgilendirme — 28 ilaç sınıfı dispatcher.
        # Filter: RAPORSUZ_BILGILENDIRME + BENZODIAZEPIN + GOZ_LUBRIKAN
        # kategorileri (hepsi aynı fn'ye dispatch edilir).
        # NOT: Bu buton sadece sut_kategorisi_tespit_et()'in mapping'inde
        # tanıdığı ilaçları kapsar — bazı raporsuz ilaçlar (Augmentin kombi,
        # PPI'lar vb.) None döndüğünde bu butona dahil olmaz.
        self.btn_raporsuz_bilgi = tk.Button(
            row_diger2, text="📋 RAPORSUZ BİLGİ",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#455A64", activebackground="#37474F",  # gri
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._raporsuz_bilgi_kontrol_baslat
        )
        self.btn_raporsuz_bilgi.pack(side="left", padx=(0, 4), pady=2)

        # ── BASIT SUT butonu (4 mikro kontrol — rapor var/yok binary) ──
        # 4 fonksiyon dispatch:
        #   - DMAH                   → kontrol_dmah (4.2.7)
        #   - KADIN_HORMON           → kontrol_kadin_hormon (4.2.29)
        #   - TRIMETAZIDIN           → kontrol_trimetazidin (4.2.14.C)
        #   - ANTIBIYOTIK_FLOROKINOLON → kontrol_antibiyotik_florokinolon (EK-4/E)
        # Hepsi <15 satırlık binary kontroller (rapor var/yok + uyarı kodu)
        # — tek tabloda gruplanır.
        self.btn_basit_sut = tk.Button(
            row_diger2, text="🔍 BASIT SUT",
            font=("Segoe UI", 9, "bold"),
            fg="white", bg="#558B2F", activebackground="#33691E",  # yeşil
            bd=0, padx=12, pady=3, cursor="hand2",
            command=self._basit_sut_kontrol_baslat
        )
        self.btn_basit_sut.pack(side="left", padx=(0, 8), pady=2)

        # ── SATIR 5: Durum mesajı + sağda renk dağılımı sayacı ──
        self.durum_bar = tk.Label(row_durum, text="Hazır", anchor="w",
                                    bg="#ECEFF1", fg="#37474F", padx=10)
        self.durum_bar.pack(fill="x", side="left", expand=True)
        self.lbl_durum_dagilim = tk.Label(row_durum, text="",
                                            bg="#ECEFF1", fg="#37474F", padx=10,
                                            font=("Segoe UI", 9))
        self.lbl_durum_dagilim.pack(side="right")

    def _tablo_kur(self, parent):
        """Tablo + sütun-hizalı filtre satırı (tek yatay scroll ile birlikte)."""
        cont = tk.Frame(parent, bg="white")
        cont.pack(fill="both", expand=True, padx=4, pady=4)

        # Style: kompakt satırlar
        style = ttk.Style()
        try:
            style.configure("Inceleme.Treeview", rowheight=22, font=("Segoe UI", 9))
            style.configure("Inceleme.Treeview.Heading",
                            font=("Segoe UI", 9, "bold"))
        except Exception:
            pass

        # Yatay ölçüler — filtre ve tablo görünür alanı eşitlemek için
        FILTRE_H = 28           # filtre satırı yüksekliği (px)
        SAG_PAD = 18            # ttk Scrollbar genişliği (yaklaşık ysb)

        # ─────────── Yatay scrollbar (en altta) ───────────
        # Önce alta sabitliyoruz ki filtre + tablo onun üstünde dizilsin.
        xsb = ttk.Scrollbar(cont, orient="horizontal")
        xsb.pack(fill="x", side="bottom", padx=(0, SAG_PAD))

        # ─────────── Filtre satırı (sütunların ÜSTÜNDE — Excel benzeri) ───────────
        # Tablonun ysb genişliği kadar sağdan boşluk bırakıyoruz ki filtre
        # alanı tablo görünür alanıyla pixel-pixel aynı olsun.
        filtre_satir = tk.Frame(cont, bg="#ECEFF1", height=FILTRE_H)
        filtre_satir.pack(fill="x", side="top", pady=(0, 2))
        filtre_satir.pack_propagate(False)

        tk.Frame(filtre_satir, width=SAG_PAD, bg="#ECEFF1"
                 ).pack(side="right", fill="y")

        self._filtre_canvas = tk.Canvas(filtre_satir, bg="#ECEFF1",
                                         highlightthickness=0,
                                         height=FILTRE_H)
        self._filtre_canvas.pack(side="left", fill="both", expand=True)

        filtre_inner = tk.Frame(self._filtre_canvas, bg="#ECEFF1")
        self._filtre_canvas.create_window((0, 0), window=filtre_inner,
                                            anchor="nw")

        self.arama_varlari = {}
        # Placeholder durum izleyici — True iken filtreye etki etmez
        self._placeholder_aktif = {}
        self._placeholder_metni = {}
        toplam_gen = 0
        PH_FG = "#9E9E9E"      # silik gri
        AKTIF_FG = "#000000"   # kullanıcı yazısı
        for kod, baslik, gen, _tip in SUTUNLAR:
            slot = tk.Frame(filtre_inner, bg="#ECEFF1",
                            width=gen, height=FILTRE_H)
            slot.pack(side="left")
            slot.pack_propagate(False)
            v = tk.StringVar()
            ent = tk.Entry(slot, textvariable=v,
                            font=("Segoe UI", 8, "italic"),
                            relief="solid", bd=1, fg=PH_FG)
            ent.place(x=1, y=4, width=gen - 2, height=20)

            # Placeholder mantığı — başlık silik gri/italik olarak içeride
            self._placeholder_metni[kod] = baslik

            def _ph_goster(k=kod, e=ent, vv=v, ph=baslik):
                if not vv.get():
                    self._placeholder_aktif[k] = True
                    vv.set(ph)
                    e.config(fg=PH_FG, font=("Segoe UI", 8, "italic"))

            def _ph_gizle(k=kod, e=ent, vv=v):
                if self._placeholder_aktif.get(k):
                    self._placeholder_aktif[k] = False
                    vv.set("")
                    e.config(fg=AKTIF_FG, font=("Segoe UI", 8))

            ent.bind("<FocusIn>", lambda _e, fn=_ph_gizle: fn())
            ent.bind("<FocusOut>", lambda _e, fn=_ph_goster: fn())

            v.trace_add("write",
                         lambda *a, k=kod: self._sutun_filtre_degisti(k))
            self.arama_varlari[kod] = v
            # Başlangıçta placeholder göster
            _ph_goster()
            toplam_gen += gen

        filtre_inner.update_idletasks()
        self._filtre_canvas.config(scrollregion=(0, 0, toplam_gen, FILTRE_H))

        # ─── Body: tablo (üst) + devre şeması paneli (alt) ───────────────
        body_frame = tk.Frame(cont, bg="white")
        body_frame.pack(fill="both", expand=True, side="top")

        # Devre şeması paneli ALTTA, yatay akış için tam genişlik + sabit yükseklik
        # Açılır/kapanır: tablonun 1/3'ü (380px) ↔ 1/2'si (genişletilmiş)
        self._sema_yukseklikler = {"normal": 380, "buyuk": 600, "kucuk": 50}
        self._sema_durum = "normal"
        self.sema_frame = tk.Frame(body_frame, bg="#FAFBFC",
                                    height=self._sema_yukseklikler["normal"],
                                    relief="solid", bd=2)
        self.sema_frame.pack(side="bottom", fill="x", pady=(4, 0))
        self.sema_frame.pack_propagate(False)
        try:
            self._sema_kur(self.sema_frame)
            import logging
            logging.getLogger(__name__).info(
                "SEMA: sema_frame oluştu, _sema_kur tamamlandı (h=%d)",
                self._sema_yukseklikler["normal"])
            # Geometry'i ölç (pack uygulandıktan sonra)
            def _olc():
                try:
                    bf = body_frame
                    sf = self.sema_frame
                    logging.getLogger(__name__).info(
                        "SEMA GEO: body=%dx%d (y=%d), sema=%dx%d (y=%d) "
                        "ismapped=%s",
                        bf.winfo_width(), bf.winfo_height(), bf.winfo_y(),
                        sf.winfo_width(), sf.winfo_height(), sf.winfo_y(),
                        sf.winfo_ismapped())
                except Exception as e2:
                    logging.getLogger(__name__).error("SEMA GEO HATA: %s", e2)
            self.root.after(800, _olc)
        except Exception as e:
            import logging, traceback
            logging.getLogger(__name__).error(
                "SEMA HATASI: _sema_kur fail: %s\n%s", e, traceback.format_exc())

        tree_frame = tk.Frame(body_frame, bg="white")
        tree_frame.pack(fill="both", expand=True, side="top")

        kolonlar = tuple(SUTUN_KOD)
        self.tv = ttk.Treeview(tree_frame, columns=kolonlar, show="headings",
                                selectmode="extended", style="Inceleme.Treeview")
        for kod, baslik, gen, _tip in SUTUNLAR:
            self.tv.heading(kod, text=baslik,
                            command=lambda k=kod: self._sutuna_gore_sirala(k))
            self.tv.column(kod, width=gen, minwidth=30, stretch=False)

        ysb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tv.yview)
        ysb.pack(side="right", fill="y")
        self.tv.pack(side="left", fill="both", expand=True)
        self.tv.configure(yscrollcommand=ysb.set)

        # Renk tag'leri
        for renk, bg in RENK_BG.items():
            fg = RENK_FG[renk]
            self.tv.tag_configure(renk, background=bg, foreground=fg)

        # Olay bind'leri
        self.tv.bind("<Button-3>", self._sag_tik_dispatch)
        self.tv.bind("<Double-1>", self._satir_detay)
        self.tv.bind("<Button-1>", self._sol_tik_dispatch, add="+")
        # Tek tıklamada şema panelini güncelle
        self.tv.bind("<<TreeviewSelect>>", self._sema_secim_guncelle, add="+")
        # Sütun genişlikleri değiştiğinde filtre slotlarını yeniden hizala.
        # Treeview'in column-resize özel event'i yok; mouse release sonrası
        # küçük bir gecikme ile güncel genişlikleri okuruz.
        def _siralama_sonrasi_hizala(_e=None):
            self.root.after(60, self._filtre_slotlarini_hizala)
        self.tv.bind("<ButtonRelease-1>", _siralama_sonrasi_hizala, add="+")
        self.tv.bind("<Configure>",
                       lambda _e: self.root.after(60,
                                                    self._filtre_slotlarini_hizala),
                       add="+")

        # ─────────── Yatay scroll senkronu (filtre + tablo birlikte) ───────────

        def _xview_birlikte(*args):
            self.tv.xview(*args)
            try:
                self._filtre_canvas.xview(*args)
            except Exception:
                pass
        xsb.configure(command=_xview_birlikte)

        def _on_tv_x_set(first, last):
            xsb.set(first, last)
            try:
                self._filtre_canvas.xview_moveto(float(first))
            except Exception:
                pass
        self.tv.configure(xscrollcommand=_on_tv_x_set)

        # Sütun yeniden boyutlandırma sonrası filtre slotlarını da güncellemek
        # için gecikmeli yenileme — başlık tıklamaları sonrası tetiklenir
        self._filtre_inner = filtre_inner
        self._filtre_slotlar = filtre_inner.winfo_children()
        self._filtre_h = FILTRE_H

    # ─────────────────────────────────────────────────────────────────
    # ATOMİK ŞART ŞEMASI PANELİ (sağ kolon, 1/3)
    # ─────────────────────────────────────────────────────────────────
    # Renkler: VAR yeşil, YOK kırmızı, KE sarı, NA gri
    _SEMA_DURUM_RENK = {
        "var": ("#1B5E20", "#E8F5E9", "✓"),
        "yok": ("#B71C1C", "#FFEBEE", "✗"),
        "kontrol_edilemedi": ("#E65100", "#FFF3E0", "?"),
        "na": ("#616161", "#ECEFF1", "—"),
    }
    _SEMA_SONUC_RENK = {
        "UYGUN": ("#1B5E20", "#C8E6C9"),
        "UYGUN DEĞİL": ("#B71C1C", "#FFCDD2"),
        "ŞÜPHELİ": ("#E65100", "#FFE0B2"),
        "ATLANDI": ("#616161", "#ECEFF1"),
    }

    # Devre şeması renk pallet (kullanıcı isteği — yeşil/sarı/kırmızı/koyu gri)
    # YEŞİL = eşleşti (veri var ve uygun)
    # SARI  = şüpheli (KE — manuel doğrulama)
    # KIRMIZI = uygunsuz (AND zorunlu şart karşılanmıyor)
    # KOYU GRİ = eşleşmedi (VEYA grupta alternatif var, bu yol seçilmedi)
    # AÇIK GRİ = bilgi (manuel — hesap dışı)
    _DEVRE_RENK = {
        "var":    ("#66BB6A", "#1B5E20"),  # YEŞİL — eşleşti
        "yok":    ("#EF5350", "#B71C1C"),  # KIRMIZI — uygunsuz
        "or_yok": ("#757575", "#424242"),  # KOYU GRİ — eşleşmedi (alt yol)
        "ke":     ("#FFEE58", "#F57F17"),  # SARI — şüpheli
        "bilgi":  ("#E0E0E0", "#9E9E9E"),  # AÇIK GRİ — bilgi
    }
    _DEVRE_SEMBOL = {"var": "✓", "yok": "✗", "or_yok": "—",
                     "ke": "?", "bilgi": "i"}
    _DEVRE_KABLO_RENK = "#90A4AE"          # akımsız kablo (gri)
    _DEVRE_KABLO_W = 2
    _DEVRE_AKIM_RENK = "#43A047"           # akımlı kablo (yeşil, kalın)
    _DEVRE_AKIM_W = 4

    def _sema_kur(self, parent: tk.Frame) -> None:
        """Atomik şart paneli — yatay devre şeması (alt panel, sol⊕→sağ⊖).

        Başlıkta 3 toggle butonu: ▼ küçült (sadece başlık 50px) /
                                     ▬ normal (380px, tablonun 1/3'ü) /
                                     ▲ büyüt (600px, tablonun 1/2'si)
        """
        # Başlık + toggle butonları
        baslik = tk.Frame(parent, bg="#37474F", height=30)
        baslik.pack(side="top", fill="x")
        baslik.pack_propagate(False)
        tk.Label(baslik, text="📋  Atomik SUT Devre Şeması",
                 bg="#37474F", fg="white",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=10)

        # Toggle butonları (sağda)
        for sembol, durum, ipucu in [("▼", "kucuk", "Küçült (sadece başlık)"),
                                       ("▬", "normal", "Normal (1/3)"),
                                       ("▲", "buyuk", "Büyüt (1/2)")]:
            btn = tk.Button(baslik, text=sembol,
                             bg="#455A64", fg="white",
                             font=("Segoe UI", 10, "bold"),
                             relief="flat", bd=0, padx=8, pady=2,
                             cursor="hand2",
                             command=lambda d=durum: self._sema_boyut_degistir(d))
            btn.pack(side="right", padx=2, pady=4)

        # Üst özet bandı — YATAY tek satır (yolak | sayım | sonuç)
        self._sema_ozet = tk.Frame(parent, bg="#FAFBFC", height=44)
        self._sema_ozet.pack(side="top", fill="x", padx=4, pady=(4, 0))
        self._sema_ozet.pack_propagate(False)

        # Canvas alanı — yatay scroll (devre uzun olabilir) + dikey scroll (bilgi listesi)
        body = tk.Frame(parent, bg="#FAFBFC")
        body.pack(side="top", fill="both", expand=True, padx=4, pady=4)

        self._sema_canvas = tk.Canvas(body, bg="white",
                                       highlightthickness=0)
        # Yatay scroll alt
        xsb = ttk.Scrollbar(body, orient="horizontal",
                             command=self._sema_canvas.xview)
        xsb.pack(side="bottom", fill="x")
        # Dikey scroll sağ
        ysb = ttk.Scrollbar(body, orient="vertical",
                             command=self._sema_canvas.yview)
        ysb.pack(side="right", fill="y")
        self._sema_canvas.pack(side="left", fill="both", expand=True)
        self._sema_canvas.configure(xscrollcommand=xsb.set,
                                     yscrollcommand=ysb.set)

        # _sema_inner kart-modu için yedek (Canvas dışında ayrı bir frame)
        # Devre modu canvas'ı doğrudan kullanır.
        self._sema_inner = tk.Frame(self._sema_canvas, bg="#FAFBFC")

        # Hover tooltip için bind
        self._sema_neden_map = {}
        self._sema_canvas.bind("<Motion>", self._sema_canvas_hover)
        self._sema_canvas.bind("<Leave>", lambda _e: self._sema_hide_tip())

        # Mouse wheel — yatay scroll için Shift+Wheel, dikey için Wheel
        def _on_wheel(event):
            self._sema_canvas.yview_scroll(int(-1 * (event.delta / 120)),
                                             "units")
        def _on_shift_wheel(event):
            self._sema_canvas.xview_scroll(int(-1 * (event.delta / 120)),
                                             "units")
        self._sema_canvas.bind("<MouseWheel>", _on_wheel)
        self._sema_canvas.bind("<Shift-MouseWheel>", _on_shift_wheel)

        self._sema_temizle("Bir reçete satırı seçin")

    # ─────────────────────────────────────────────────────────────────
    # KONTROL BUTONLARI POPUP — 15 buton tek bir popup grid'inde
    # ─────────────────────────────────────────────────────────────────
    def _kontrol_butonlari_topla(self) -> list:
        """Gizli buton frame'lerinden tüm Button widget'larını topla.

        Returns: List[(text, bg, button_widget)] — invoke için widget kendisi.
        """
        butonlar = []
        for frame_attr in ("_row_kontrol_btn", "_row_sut_btn",
                            "_row_diger_btn", "_row_diger2_btn"):
            f = getattr(self, frame_attr, None)
            if f is None:
                continue
            for ch in f.winfo_children():
                if isinstance(ch, tk.Button):
                    try:
                        butonlar.append((ch.cget("text"),
                                          ch.cget("bg"), ch))
                    except Exception:
                        pass
        return butonlar

    def _kontrol_butonlari_popup_ac(self) -> None:
        """Popup pencere açar — 15 SUT/uyarı butonunu 3×5 grid'de gösterir."""
        try:
            butonlar = self._kontrol_butonlari_topla()
        except Exception as e:
            butonlar = []
        if not butonlar:
            try:
                from tkinter import messagebox
                messagebox.showinfo("Kontrol Butonları",
                                     "Henüz buton yüklenmemiş.")
            except Exception:
                pass
            return

        pop = tk.Toplevel(self.root)
        pop.title("🎯 Kontrol Butonları")
        pop.configure(bg="#FAFBFC")
        pop.transient(self.root)
        # Pencere boyutu: butona göre dinamik (3x5 = ~860x320)
        n = len(butonlar)
        sutun = 5
        satir = (n + sutun - 1) // sutun
        pop.geometry(f"{160 * sutun + 40}x{60 * satir + 100}")

        # Üstte "TÜMÜ" butonu
        tk.Button(pop, text="🎯  TÜMÜNÜ KONTROL ET (15 buton sırayla)",
                   bg="#1A237E", fg="white",
                   font=("Segoe UI", 11, "bold"),
                   activebackground="#283593",
                   relief="flat", bd=0, padx=20, pady=10, cursor="hand2",
                   command=lambda: self._kontrol_tumunu_calistir(pop)
                   ).pack(side="top", fill="x", padx=8, pady=(8, 4))

        tk.Label(pop, text="—— veya tek bir kontrol seç ——",
                 bg="#FAFBFC", fg="#90A4AE",
                 font=("Segoe UI", 8, "italic")
                 ).pack(side="top", pady=(4, 6))

        # Grid alanı
        grid_frame = tk.Frame(pop, bg="#FAFBFC")
        grid_frame.pack(side="top", fill="both", expand=True, padx=8, pady=4)

        for i, (text, bg, widget) in enumerate(butonlar):
            r, col = divmod(i, sutun)
            try:
                fg = widget.cget("fg")
            except Exception:
                fg = "white"
            b = tk.Button(grid_frame, text=text,
                           bg=bg, fg=fg,
                           font=("Segoe UI", 9, "bold"),
                           activebackground=bg,
                           relief="flat", bd=0, padx=4, pady=10,
                           cursor="hand2", wraplength=140,
                           command=lambda w=widget, p=pop:
                               self._kontrol_buton_secildi(w, p))
            b.grid(row=r, column=col, padx=3, pady=3, sticky="nsew")

        for col in range(sutun):
            grid_frame.columnconfigure(col, weight=1, uniform="kontrol")
        for r in range(satir):
            grid_frame.rowconfigure(r, weight=1)

        # Klavye: Esc kapatır
        pop.bind("<Escape>", lambda _e: pop.destroy())
        pop.focus_set()

    def _kontrol_buton_secildi(self, widget: tk.Button,
                                 popup: tk.Toplevel) -> None:
        """Popup'tan bir buton seçildiğinde — popup'ı kapat + butonu invoke et.

        Çift tıklama koruması: aynı butonun art arda iki kez çağrılmasını
        engeller (yoksa 2 Excel raporu üretilir).
        """
        if getattr(self, "_kontrol_calistiriliyor", False):
            return
        self._kontrol_calistiriliyor = True
        try:
            popup.destroy()
        except Exception:
            pass

        def _calistir_ve_resetle():
            try:
                widget.invoke()
            finally:
                # Bir sonraki kontrol için flag'i sıfırla
                self._kontrol_calistiriliyor = False
        try:
            self.root.after(80, _calistir_ve_resetle)
        except Exception:
            self._kontrol_calistiriliyor = False

    def _kontrol_tumunu_calistir(self, popup: tk.Toplevel) -> None:
        """TÜMÜ butonu — tüm kontrolleri sırayla çalıştırır."""
        try:
            popup.destroy()
        except Exception:
            pass
        butonlar = self._kontrol_butonlari_topla()
        # Her butonu sırayla invoke et (yeterince gecikmeli — UI bloklamasın)
        gecikme = 100
        for i, (_text, _bg, widget) in enumerate(butonlar):
            try:
                self.root.after(gecikme + i * 80, widget.invoke)
            except Exception:
                pass

    def _sema_boyut_degistir(self, durum: str) -> None:
        """Şema panelinin yüksekliğini değiştir (kucuk/normal/buyuk).

        kucuk: sadece başlık (50px) — devre gizlenir
        normal: 380px (tablonun 1/3'ü)
        buyuk: 600px (tablonun 1/2'si)
        """
        if not hasattr(self, "_sema_yukseklikler"):
            return
        yeni_h = self._sema_yukseklikler.get(durum, 380)
        self._sema_durum = durum
        try:
            self.sema_frame.configure(height=yeni_h)
        except Exception:
            pass
        # Render'ı yeni yüksekliğe göre tekrar çiz
        self.root.after(120, self._sema_secim_guncelle)

    def _sema_canvas_hover(self, event) -> None:
        """Canvas üzerinde hover — ampule yakınsa neden tooltip aç."""
        try:
            c = self._sema_canvas
            cx = c.canvasx(event.x)
            cy = c.canvasy(event.y)
            items = c.find_overlapping(cx - 1, cy - 1, cx + 1, cy + 1)
            for it in items:
                if it in self._sema_neden_map:
                    if getattr(self, "_sema_tip_item", None) == it:
                        return
                    self._sema_hide_tip()
                    neden = self._sema_neden_map[it]
                    x = c.winfo_rootx() + event.x + 16
                    y = c.winfo_rooty() + event.y + 16
                    tip = tk.Toplevel(self.root)
                    tip.wm_overrideredirect(True)
                    tip.wm_geometry(f"+{x}+{y}")
                    tk.Label(tip, text=neden, bg="#FFFDE7", fg="#37474F",
                             font=("Segoe UI", 9), justify="left",
                             relief="solid", bd=1, padx=6, pady=3,
                             wraplength=420).pack()
                    self._sema_tip = tip
                    self._sema_tip_item = it
                    return
            self._sema_hide_tip()
        except Exception:
            pass

    def _sema_temizle(self, mesaj: str = "") -> None:
        """Şema panelini temizle, opsiyonel placeholder mesaj."""
        for w in self._sema_ozet.winfo_children():
            w.destroy()
        try:
            for w in self._sema_inner.winfo_children():
                w.destroy()
        except Exception:
            pass
        if mesaj:
            tk.Label(self._sema_ozet, text=mesaj, bg="#FAFBFC",
                     fg="#90A4AE",
                     font=("Segoe UI", 9, "italic")).pack(pady=20)

    def _sema_secim_guncelle(self, _event=None) -> None:
        """Treeview'de seçilen reçete için şema panelini günceller."""
        try:
            secimler = self.tv.selection()
            if not secimler:
                self._sema_temizle("Bir reçete satırı seçin")
                return
            iid = secimler[0]  # ilk seçili — iid = str(s["ri_id"])
            # tum_satirlar içinde ri_id eşleşmesi
            satir = None
            for s in (self.tum_satirlar or []):
                if str(s.get("ri_id")) == str(iid):
                    satir = s
                    break
            if satir is None:
                self._sema_temizle("Satır bulunamadı")
                return
            self._sema_render(satir)
        except Exception as e:
            self._sema_temizle(f"Hata: {e}")

    @staticmethod
    def _sema_grup_matematigi(sartlar: list) -> dict:
        """VEYA-aware şart matematiği — kaç gerekli, kaç sağlanıyor.

        Kurallar:
          • "(bilgi)" suffix grup → tamamen hesap dışı
          • veya_grubu=True olan grup (≥1 yeterli):
              gerekli = 1
              sağlanan = min(1, VAR sayısı)
          • Diğer (AND) gruplar:
              gerekli = grup uzunluğu
              sağlanan = VAR sayısı
          • Üst-VEYA çiftleri ("(a)" VEYA "(b)"; "Varfarin (D-2)" VEYA
            "İstisna grubu (D-2)"):
              en az birinin saglanan==gerekli olduğu yol seçilir;
              ikisi de tam değilse → daha tam olan seçilir; eşitse minimum
              gerekli olan yol baz alınır.

        Returns: {
            'gerekli_toplam': int,
            'saglanan_toplam': int,
            'gruplar': [{'grup', 'gerekli', 'saglanan', 'toplam',
                         'veya', 'durum_kismi'}],
        }
        """
        # Gruplara ayır (sıra korunur)
        gruplar_dict: dict = {}
        sira = []
        for s in sartlar:
            g = s.get("grup", "") or "(grupsuz)"
            if g not in gruplar_dict:
                gruplar_dict[g] = []
                sira.append(g)
            gruplar_dict[g].append(s)

        # İlk pass — her grubun gerekli/saglanan'ını hesapla
        ham_gruplar = {}
        for g in sira:
            if "(bilgi)" in g:
                continue
            gs = gruplar_dict[g]
            veya = any(s.get("veya_grubu") for s in gs)
            var_sayi = sum(1 for s in gs if s.get("durum") == "var")
            toplam_sayi = len(gs)
            if veya:
                gerekli, saglanan = 1, min(1, var_sayi)
            else:
                gerekli, saglanan = toplam_sayi, var_sayi
            ham_gruplar[g] = {
                "grup": g, "gerekli": gerekli, "saglanan": saglanan,
                "toplam": toplam_sayi, "veya": veya,
                "var_sayi": var_sayi,
            }

        # Üst-VEYA çiftleri — birleştir
        cift_prefix_listesi = [
            ("Varfarin yolu (a)", "Varfarin yolu (b)",
             "Varfarin yolu [(a) VEYA (b)]"),
            ("Varfarin yolu (D-2", "İstisna grubu (D-2",
             "Varfarin VEYA istisna [(1)(b) VEYA (2)]"),
        ]
        for pa, pb, birlesik in cift_prefix_listesi:
            ka = next((k for k in ham_gruplar if k.startswith(pa)), None)
            kb = next((k for k in ham_gruplar if k.startswith(pb)), None)
            if ka and kb:
                ga, gb = ham_gruplar[ka], ham_gruplar[kb]
                # En kısa yol = düşük gerekli; tamamlanmış olan tercih
                a_tam = ga["saglanan"] == ga["gerekli"]
                b_tam = gb["saglanan"] == gb["gerekli"]
                if a_tam and not b_tam:
                    secilen = ga
                elif b_tam and not a_tam:
                    secilen = gb
                elif a_tam and b_tam:
                    # ikisi de tam → daha kısa olanı seç
                    secilen = ga if ga["gerekli"] <= gb["gerekli"] else gb
                else:
                    # ikisi de eksik → daha çok ilerlemiş veya minimum gerekli
                    if ga["saglanan"] > gb["saglanan"]:
                        secilen = ga
                    elif gb["saglanan"] > ga["saglanan"]:
                        secilen = gb
                    else:
                        # eşit → minimum gerekli
                        secilen = ga if ga["gerekli"] <= gb["gerekli"] else gb
                ham_gruplar[birlesik] = {
                    "grup": birlesik,
                    "gerekli": secilen["gerekli"],
                    "saglanan": secilen["saglanan"],
                    "toplam": ga["toplam"] + gb["toplam"],
                    "veya": True,
                    "var_sayi": ga["var_sayi"] + gb["var_sayi"],
                    "secilen_yol": secilen["grup"],
                }
                ham_gruplar.pop(ka)
                ham_gruplar.pop(kb)

        sonuc_listesi = list(ham_gruplar.values())
        gerekli_t = sum(g["gerekli"] for g in sonuc_listesi)
        saglanan_t = sum(g["saglanan"] for g in sonuc_listesi)
        return {
            "gerekli_toplam": gerekli_t,
            "saglanan_toplam": saglanan_t,
            "gruplar": sonuc_listesi,
        }

    # Render modu: "devre" (Canvas devre şeması — pil/ampul/kablo)
    #              "kart"  (kart-bazlı 2-kolonlu liste — yedek)
    _SEMA_MOD_DEFAULT = "devre"

    def _sema_render(self, satir: dict) -> None:
        """Aktif moda göre kart-bazlı veya devre şeması render dispatcher."""
        mod = getattr(self, "_sema_mod", self._SEMA_MOD_DEFAULT)
        if mod == "kart":
            self._sema_render_kart(satir)
        else:
            self._sema_render_devre(satir)

    def _sema_render_kart(self, satir: dict) -> None:
        """[YEDEK] Kart-bazlı render — SUT lafzı sol/açık gri + durum sağ/renkli.

        Devre şeması yapılamazsa veya kullanıcı tercih ederse bu kullanılır.
        """
        self._sema_temizle()

        verdict = (satir.get("verdict") or "").strip()
        sartlar_json = satir.get("verdict_sartlar") or ""
        detaylar_json = satir.get("verdict_detaylar") or ""
        if not verdict and not sartlar_json:
            tk.Label(self._sema_ozet,
                     text="Bu reçete için kontrol yapılmamış",
                     bg="#FAFBFC", fg="#90A4AE",
                     font=("Segoe UI", 9, "italic")).pack(pady=20)
            return

        # Şart listesini parse et
        try:
            sartlar = json.loads(sartlar_json) if sartlar_json else []
        except Exception:
            sartlar = []
        try:
            detaylar = json.loads(detaylar_json) if detaylar_json else {}
        except Exception:
            detaylar = {}

        # Yolak: alt_dal var mı (örn. "4.2.15.D-1" / "4.2.15.D-2")?
        alt_dal = (detaylar.get("alt_dal") or "").strip()
        endikasyon = (detaylar.get("endikasyon") or "").strip()
        # SUT madde (verdict_sut) içinden de yolak bilgisi çıkarılır
        sut_kurali = (satir.get("verdict_sut") or "").strip()

        yolak_text = ""
        if alt_dal:
            if "D-1" in alt_dal:
                yolak_text = f"📍 SUT {alt_dal} — AF YOLAĞI"
            elif "D-2" in alt_dal:
                yolak_text = f"📍 SUT {alt_dal} — DVT/PE YOLAĞI"
            else:
                yolak_text = f"📍 SUT {alt_dal}"
        elif sut_kurali:
            yolak_text = f"📍 {sut_kurali[:50]}"

        # Hesaplama: VEYA-aware matematik
        # AND grup: gerekli=N, sağlanan=VAR sayısı
        # OR (≥1) grup: gerekli=1, sağlanan=min(1, VAR)
        # Üst-VEYA (örn. (a) VEYA (b)): en kısa yol seçilir, hangisi
        # tamamlanmışsa o saglanan sayılır; ikisi de değilse minimum yol gerekli
        sayilar = {"var": 0, "yok": 0, "kontrol_edilemedi": 0, "na": 0}
        for s in sartlar:
            grup = s.get("grup", "") or ""
            if "(bilgi)" in grup:
                continue
            d = s.get("durum", "")
            if d in sayilar:
                sayilar[d] += 1
        gmat = self._sema_grup_matematigi(sartlar)
        toplam_hesap = gmat["gerekli_toplam"]
        sayilan_var = gmat["saglanan_toplam"]

        # ── Üst özet bandı (3 katman) ─────────────────────────────────
        # 1. Yolak (büyük başlık)
        if yolak_text:
            tk.Label(self._sema_ozet, text=yolak_text,
                     bg="#FAFBFC", fg="#1565C0",
                     font=("Segoe UI", 9, "bold"),
                     anchor="w", justify="left",
                     wraplength=400).pack(side="top", fill="x", padx=6,
                                          pady=(2, 0))
        if endikasyon:
            tk.Label(self._sema_ozet, text=f"Endikasyon: {endikasyon}",
                     bg="#FAFBFC", fg="#455A64",
                     font=("Segoe UI", 8),
                     anchor="w").pack(side="top", fill="x", padx=8)

        # 2. Sayım — büyük, ana bilgi (VEYA-aware: gerekli vs sağlanan)
        sayim_buyuk = f"  {sayilan_var} / {toplam_hesap}  ŞART KARŞILANIYOR"
        # Renk: %≥80 yeşil, %50-80 turuncu, <50 kırmızı
        oran = (sayilan_var / toplam_hesap) if toplam_hesap else 0
        if oran >= 0.8:
            sayim_fg, sayim_bg = "#1B5E20", "#C8E6C9"
        elif oran >= 0.5:
            sayim_fg, sayim_bg = "#E65100", "#FFE0B2"
        else:
            sayim_fg, sayim_bg = "#B71C1C", "#FFCDD2"
        tk.Label(self._sema_ozet, text=sayim_buyuk,
                 bg=sayim_bg, fg=sayim_fg,
                 font=("Segoe UI", 10, "bold"),
                 anchor="w").pack(side="top", fill="x", padx=4, pady=(2, 0))

        # 3. Detay sayım + sonuç etiketi (küçük, ikincil)
        renk_fg, renk_bg = self._SEMA_SONUC_RENK.get(
            verdict.upper(), ("#37474F", "#ECEFF1"))
        alt_satir = tk.Frame(self._sema_ozet, bg="#FAFBFC")
        alt_satir.pack(side="top", fill="x", padx=4, pady=(2, 0))
        tk.Label(alt_satir,
                 text=f"✓{sayilar['var']}  ✗{sayilar['yok']}  "
                      f"?{sayilar['kontrol_edilemedi']}",
                 bg="#FAFBFC", fg="#455A64",
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=4)
        tk.Label(alt_satir, text=f" {verdict or '—'} ",
                 bg=renk_bg, fg=renk_fg,
                 font=("Segoe UI", 8, "bold")).pack(side="right", padx=4)

        ilac = (satir.get("ilac") or "")[:50]
        if ilac:
            tk.Label(self._sema_ozet, text=ilac, bg="#FAFBFC", fg="#90A4AE",
                     font=("Segoe UI", 7)).pack(anchor="w", padx=6)

        if not sartlar:
            tk.Label(self._sema_ozet,
                     text="(atomik şart listesi yok — eski format)",
                     bg="#FAFBFC", fg="#90A4AE",
                     font=("Segoe UI", 8, "italic")).pack(pady=10, padx=8)
            return

        # ── Gruplara ayır ─────────────────────────────────────────────
        gruplar: dict = {}
        sira = []
        for s in sartlar:
            g = s.get("grup", "") or "(grupsuz)"
            if g not in gruplar:
                gruplar[g] = []
                sira.append(g)
            gruplar[g].append(s)

        # Her grubun gerekli/sağlanan eşlemesi (gmat'tan)
        grup_mat = {gm["grup"]: gm for gm in gmat.get("gruplar", [])}

        # ── Kart-bazlı render: SUT lafzı (sol, açık gri) + durum (sağ, renkli)
        # Algoritmik akış: tepeden alta — grup başlıkları + indented şartlar
        for g in sira:
            gs = gruplar[g]
            bilgi = "(bilgi)" in g
            veya = any(s.get("veya_grubu") for s in gs)

            # Grup başlığı kartı
            if bilgi:
                grup_bg, grup_fg = "#ECEFF1", "#616161"
                sayim_text = "—"
                sayim_bg, sayim_fg = "#ECEFF1", "#9E9E9E"
                grup_text = g
            else:
                gm = grup_mat.get(g, {"saglanan": 0, "gerekli": len(gs)})
                gerekli = gm["gerekli"]
                saglanan = gm["saglanan"]
                ke_count = sum(1 for s in gs
                               if s.get("durum") == "kontrol_edilemedi")
                if saglanan >= gerekli:
                    grup_bg, grup_fg = "#1B5E20", "white"
                    sayim_bg, sayim_fg = "#A5D6A7", "#1B5E20"
                elif ke_count > 0 and (saglanan + ke_count) >= gerekli:
                    grup_bg, grup_fg = "#E65100", "white"
                    sayim_bg, sayim_fg = "#FFCC80", "#BF360C"
                else:
                    grup_bg, grup_fg = "#B71C1C", "white"
                    sayim_bg, sayim_fg = "#EF9A9A", "#B71C1C"
                sayim_text = f"{saglanan}/{gerekli}"
                grup_text = g + ("  [VEYA ≥1]" if veya else "")

            grup_kart = tk.Frame(self._sema_inner, bg=grup_bg)
            grup_kart.pack(side="top", fill="x", padx=4, pady=(8, 0))
            tk.Label(grup_kart, text=" " + grup_text, bg=grup_bg,
                     fg=grup_fg, font=("Segoe UI", 8, "bold"),
                     anchor="w", justify="left",
                     wraplength=300).pack(side="left", fill="x", expand=True,
                                           padx=2, pady=2)
            tk.Label(grup_kart, text=f" {sayim_text} ", bg=sayim_bg,
                     fg=sayim_fg, font=("Segoe UI", 9, "bold")
                     ).pack(side="right", padx=4, pady=2)

            # ── Atomik şart satırları (indented, 2-kolonlu kart) ───────
            for s in gs:
                ad = s.get("ad", "")
                d = s.get("durum", "na")
                neden = s.get("neden", "")
                veya_uye = bool(s.get("veya_grubu"))

                # Durum badge: SAĞ kolon (renkli) — Sol kolon AÇIK GRİ
                if bilgi:
                    bdg_bg, bdg_fg, bdg_text = "#ECEFF1", "#9E9E9E", "i  bilgi"
                elif veya_uye:
                    if d == "var":
                        bdg_bg, bdg_fg, bdg_text = "#A5D6A7", "#1B5E20", "✓ VAR"
                    elif d == "yok":
                        # VEYA grubunda YOK = SARI (kritik değil, alternatif)
                        bdg_bg, bdg_fg, bdg_text = "#FFF59D", "#F57F17", "○ yok"
                    elif d == "kontrol_edilemedi":
                        bdg_bg, bdg_fg, bdg_text = "#FFCC80", "#E65100", "? KE"
                    else:
                        bdg_bg, bdg_fg, bdg_text = "#ECEFF1", "#9E9E9E", "—"
                else:
                    if d == "var":
                        bdg_bg, bdg_fg, bdg_text = "#A5D6A7", "#1B5E20", "✓ VAR"
                    elif d == "yok":
                        bdg_bg, bdg_fg, bdg_text = "#EF9A9A", "#B71C1C", "✗ YOK"
                    elif d == "kontrol_edilemedi":
                        bdg_bg, bdg_fg, bdg_text = "#FFCC80", "#E65100", "? KE"
                    else:
                        bdg_bg, bdg_fg, bdg_text = "#ECEFF1", "#9E9E9E", "—"

                # Şart kartı — 20px sol indent (tree hissi)
                sart_kart = tk.Frame(self._sema_inner, bg="#FAFBFC")
                sart_kart.pack(side="top", fill="x", padx=(20, 4), pady=1)

                # SAĞ — durum badge (önce pack edilince sağda kalır)
                bdg = tk.Label(sart_kart, text=f" {bdg_text} ",
                                bg=bdg_bg, fg=bdg_fg,
                                font=("Segoe UI", 8, "bold"))
                bdg.pack(side="right", padx=2)

                # SOL — SUT lafzı, AÇIK GRİ italik (mevzuat referansı)
                lafz = tk.Label(sart_kart, text="• " + ad,
                                 bg="#FAFBFC", fg="#9E9E9E",
                                 font=("Segoe UI", 8, "italic"),
                                 anchor="w", justify="left",
                                 wraplength=260)
                lafz.pack(side="left", fill="x", expand=True, padx=2)

                # Hover tooltip — neden
                if neden:
                    def _bind_tip(widgets, text=neden, parent_w=sart_kart):
                        def _show(_e=None, t=text, w=parent_w):
                            try:
                                self._sema_show_tip(w, t)
                            except Exception:
                                pass
                        def _hide(_e=None):
                            self._sema_hide_tip()
                        for w in widgets:
                            w.bind("<Enter>", _show)
                            w.bind("<Leave>", _hide)
                    _bind_tip([sart_kart, lafz, bdg])

    # ─────────────────────────────────────────────────────────────────
    # DEVRE ŞEMASI (Canvas) — pil/ampul/kablo metaforu
    # ⊕ üstte, ⊖ altta. Seri (AND) → düz hat. Paralel (OR) → Y kavşak.
    # Ampul rengi: yeşil VAR / kırmızı YOK / sarı VEYA-YOK / turuncu KE / gri bilgi
    # ─────────────────────────────────────────────────────────────────
    def _sema_render_devre(self, satir: dict) -> None:
        """Devre şeması render — Canvas üzerinde pil/ampul/kablo çizer."""
        self._sema_temizle()

        verdict = (satir.get("verdict") or "").strip()
        sartlar_json = satir.get("verdict_sartlar") or ""
        detaylar_json = satir.get("verdict_detaylar") or ""

        # DEBUG — sema render giriş
        try:
            import logging
            logging.getLogger(__name__).info(
                "SEMA RENDER: ri_id=%s, verdict=%r, "
                "sartlar_json_len=%d, ilac=%r",
                satir.get("ri_id"), verdict, len(sartlar_json),
                (satir.get("ilac") or "")[:40])
            if sartlar_json:
                logging.getLogger(__name__).info(
                    "SEMA SARTLAR JSON: %s", sartlar_json[:400])
        except Exception:
            pass

        if not verdict and not sartlar_json:
            tk.Label(self._sema_ozet,
                     text="Bu reçete için kontrol yapılmamış",
                     bg="#FAFBFC", fg="#90A4AE",
                     font=("Segoe UI", 9, "italic")).pack(pady=20)
            return

        try:
            sartlar = json.loads(sartlar_json) if sartlar_json else []
        except Exception:
            sartlar = []
        try:
            detaylar = json.loads(detaylar_json) if detaylar_json else {}
        except Exception:
            detaylar = {}

        gmat = self._sema_grup_matematigi(sartlar)
        gerekli_t = gmat["gerekli_toplam"]
        saglanan_t = gmat["saglanan_toplam"]

        # Üst banda (devre + kart için ortak)
        self._sema_ust_banda_yaz(satir, sartlar, detaylar, verdict,
                                  gerekli_t, saglanan_t)

        if not sartlar:
            return

        # Canvas üzerinde devre şeması
        self.root.update_idletasks()  # canvas genişliğini al
        self._devre_ciz_canvas(sartlar, gmat)

    def _sema_ust_banda_yaz(self, satir: dict, sartlar: list, detaylar: dict,
                             verdict: str, gerekli_t: int, saglanan_t: int) -> None:
        """Üst özet bandını yatay tek satır olarak yazar (alt panel için)."""
        alt_dal = (detaylar.get("alt_dal") or "").strip()
        endikasyon = (detaylar.get("endikasyon") or "").strip()
        sut_kurali = (satir.get("verdict_sut") or "").strip()

        yolak_text = ""
        if alt_dal:
            if "D-1" in alt_dal:
                yolak_text = f"📍 {alt_dal} AF"
            elif "D-2" in alt_dal:
                yolak_text = f"📍 {alt_dal} DVT/PE"
            else:
                yolak_text = f"📍 {alt_dal}"
        elif sut_kurali:
            yolak_text = f"📍 {sut_kurali[:30]}"

        # Yatay tek satır: yolak | sayım (büyük) | sonuç badge | ilaç
        sayim_text = f"  {saglanan_t} / {gerekli_t}  ŞART KARŞILANIYOR  "
        oran = (saglanan_t / gerekli_t) if gerekli_t else 0
        if oran >= 0.99:
            sayim_fg, sayim_bg = "#1B5E20", "#C8E6C9"
        elif oran >= 0.5:
            sayim_fg, sayim_bg = "#E65100", "#FFE0B2"
        else:
            sayim_fg, sayim_bg = "#B71C1C", "#FFCDD2"
        renk_fg, renk_bg = self._SEMA_SONUC_RENK.get(
            verdict.upper(), ("#37474F", "#ECEFF1"))

        if yolak_text:
            tk.Label(self._sema_ozet, text=yolak_text,
                     bg="#FAFBFC", fg="#1565C0",
                     font=("Segoe UI", 10, "bold")
                     ).pack(side="left", padx=(8, 12), pady=8)
        tk.Label(self._sema_ozet, text=sayim_text,
                 bg=sayim_bg, fg=sayim_fg,
                 font=("Segoe UI", 11, "bold")
                 ).pack(side="left", padx=4, pady=8)
        tk.Label(self._sema_ozet,
                 text=f" ✓{sum(1 for s in sartlar if s.get('durum')=='var')}"
                      f"  ✗{sum(1 for s in sartlar if s.get('durum')=='yok')}"
                      f"  ?{sum(1 for s in sartlar if s.get('durum')=='kontrol_edilemedi')} ",
                 bg="#FAFBFC", fg="#455A64",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=8, pady=8)
        tk.Label(self._sema_ozet, text=f"  {verdict or '—'}  ",
                 bg=renk_bg, fg=renk_fg,
                 font=("Segoe UI", 10, "bold")).pack(side="right", padx=8, pady=8)
        if endikasyon:
            tk.Label(self._sema_ozet, text=f"({endikasyon})",
                     bg="#FAFBFC", fg="#455A64",
                     font=("Segoe UI", 9)).pack(side="left", padx=4, pady=8)
        ilac = (satir.get("ilac") or "")[:30]
        if ilac:
            tk.Label(self._sema_ozet, text=ilac,
                     bg="#FAFBFC", fg="#90A4AE",
                     font=("Segoe UI", 8)).pack(side="right", padx=8, pady=8)

    def _devre_kablo(self, c: tk.Canvas, x1: int, y1: int, x2: int, y2: int,
                      akim: bool = False) -> None:
        """Kablo çiz — akım varsa yeşil kalın, yoksa gri normal."""
        if akim:
            c.create_line(x1, y1, x2, y2,
                          fill=self._DEVRE_AKIM_RENK,
                          width=self._DEVRE_AKIM_W, tags="devre")
        else:
            c.create_line(x1, y1, x2, y2,
                          fill=self._DEVRE_KABLO_RENK,
                          width=self._DEVRE_KABLO_W, tags="devre")

    def _devre_ciz_canvas(self, sartlar: list, gmat: dict) -> None:
        """Yatay devre şeması: ⊕ sol → ⊖ sağ. AND yatay seri, OR dikey paralel.

        Layout:
          ⊕ ── ●AF ── ●Mitral ── ●Mekanik ──┬─●i─┐         ┬─ ... ─ ⊖
                                              ├─●T─┤         │
                                              ├─●75─┤        │
                                              └─●H──┘
        """
        c = self._sema_canvas
        c.delete("all")
        self._sema_neden_map = {}

        canvas_h = max(c.winfo_height(), 320)
        y_merkez = canvas_h // 2

        AMPUL_R = 14                           # küçültüldü 18→14 (yazı sığsın)
        AMPUL_X_GAP = 165                      # 130→165 (yazılar çakışmasın)
        PARALEL_Y_GAP = 75                     # 60→75 (üst+alt etiket için yer)
        KENAR = 36
        PIL_FONT = ("Segoe UI", 18, "bold")
        SEMBOL_FONT = ("Segoe UI", 11, "bold")
        ETIKET_FONT = ("Segoe UI", 8)          # 9→8 (sığsın)
        BASLIK_FONT = ("Segoe UI", 9, "bold")
        SAYIM_FONT = ("Segoe UI", 9, "bold")

        # Gruplara ayır
        gruplar_dict: dict = {}
        sira = []
        for s in sartlar:
            g = s.get("grup", "") or "(grupsuz)"
            if g not in gruplar_dict:
                gruplar_dict[g] = []
                sira.append(g)
            gruplar_dict[g].append(s)
        grup_mat = {gm["grup"]: gm for gm in gmat.get("gruplar", [])}

        # ── Üst-VEYA çiftleri — eşleştir (akım hesabı için) ───────────
        ust_veya_ciftleri = [
            ("Varfarin yolu (a)", "Varfarin yolu (b)",
             "Varfarin yolu [(a) VEYA (b)]"),
            ("Varfarin yolu (D-2", "İstisna grubu (D-2",
             "Varfarin VEYA istisna [(1)(b) VEYA (2)]"),
        ]
        ust_veya_eslesme = {}  # alt grup adı -> birleşik grup adı
        for prefix_a, prefix_b, birlesik in ust_veya_ciftleri:
            a_grup = next((g for g in gruplar_dict
                           if g.startswith(prefix_a)
                           and "(bilgi)" not in g), None)
            b_grup = next((g for g in gruplar_dict
                           if g.startswith(prefix_b)
                           and "(bilgi)" not in g), None)
            if a_grup and b_grup:
                ust_veya_eslesme[a_grup] = (birlesik, prefix_b)

        def grup_tam(g_ad):
            gm = grup_mat.get(g_ad, {})
            return (gm.get("saglanan", 0) >= gm.get("gerekli", 999) and
                    gm.get("gerekli", 0) > 0)

        def grup_birlesik_tam(birlesik_ad):
            gm = grup_mat.get(birlesik_ad, {})
            return (gm.get("saglanan", 0) >= gm.get("gerekli", 999) and
                    gm.get("gerekli", 0) > 0)

        # Akım track değişkenleri
        akim_aktif = True   # sol pilden akım gelir, kesilmediyse devam
        cift_baslangic_akim = None  # üst-VEYA çiftin başında prev_akim'i sakla

        # ⊕ Pil — SOL (akım kaynağı)
        x = KENAR
        c.create_text(x, y_merkez, text="⊕", fill="#1B5E20",
                      font=PIL_FONT, tags="devre")
        c.create_text(x, y_merkez + 28, text="GİRİŞ", fill="#37474F",
                      font=("Segoe UI", 8, "bold"), tags="devre")
        son_x = x + 14
        x += 35

        # Gruplar yatay olarak ardışık çizilir
        for g_ad in sira:
            if "(bilgi)" in g_ad:
                continue
            gs = gruplar_dict[g_ad]
            veya = any(s.get("veya_grubu") for s in gs)
            gm = grup_mat.get(g_ad, {})

            # Üst-VEYA çift kontrolü — bu grup çift parçası mı?
            cift_par = ust_veya_eslesme.get(g_ad)
            if cift_par:
                birlesik_ad, prefix_b = cift_par
                cift_son_grup = g_ad.startswith(prefix_b)
                if not cift_son_grup and cift_baslangic_akim is None:
                    cift_baslangic_akim = akim_aktif
                cift_tam = grup_birlesik_tam(birlesik_ad)

            kendi_tam = grup_tam(g_ad)
            # Bu grubun yolunda akım var mı?
            yol_akim = akim_aktif and kendi_tam

            if veya and len(gs) > 1:
                # ── DİKEY PARALEL (OR ≥1) ──
                n = len(gs)
                ampul_x = x + 60
                sol_kavsak_x = x + 5
                sag_kavsak_x = ampul_x + 60
                ampul_ys = [y_merkez + (i - (n - 1) / 2) * PARALEL_Y_GAP
                            for i in range(n)]

                # Ana hat → sol kavşak (akım giriş varsa yeşil)
                self._devre_kablo(c, son_x, y_merkez, sol_kavsak_x, y_merkez,
                                   akim=akim_aktif)
                # Sol dikey kavşak (var olan ampuller arası akım yayılır)
                self._devre_kablo(c, sol_kavsak_x, min(ampul_ys),
                                   sol_kavsak_x, max(ampul_ys),
                                   akim=akim_aktif)
                # Her ampul: sol kavşak → ampul → sağ kavşak
                # Sadece VAR olan ampul yolu yeşil
                for ay, s in zip(ampul_ys, gs):
                    bu_ampul_var = (s.get("durum") == "var")
                    bu_yol_akim = akim_aktif and bu_ampul_var
                    self._devre_kablo(c, sol_kavsak_x, ay,
                                       ampul_x - AMPUL_R, ay,
                                       akim=bu_yol_akim)
                    self._devre_ampul_at(c, ampul_x, ay, s, AMPUL_R,
                                          veya_grubu=True,
                                          font_sembol=SEMBOL_FONT,
                                          font_etiket=ETIKET_FONT)
                    self._devre_kablo(c, ampul_x + AMPUL_R, ay,
                                       sag_kavsak_x, ay,
                                       akim=bu_yol_akim)
                # Sağ dikey kavşak (akım çıkışı varsa yeşil)
                cikis_akim = akim_aktif and (kendi_tam)
                self._devre_kablo(c, sag_kavsak_x, min(ampul_ys),
                                   sag_kavsak_x, max(ampul_ys),
                                   akim=cikis_akim)
                son_x = sag_kavsak_x
                x = sag_kavsak_x + 25

                # Akım güncellemesi
                if cift_par:
                    if cift_son_grup:
                        # Çift sonu — birleşik tamlığa göre akım
                        akim_aktif = (cift_baslangic_akim if cift_baslangic_akim is not None else True) and cift_tam
                        cift_baslangic_akim = None
                else:
                    akim_aktif = akim_aktif and kendi_tam
            else:
                # ── YATAY SERİ (AND) ──
                # Ampuller yan yana, her ampul kendi durumuna göre akım keser
                grup_akim_baslangic = akim_aktif
                local_akim = akim_aktif
                for s in gs:
                    ampul_x = x + AMPUL_R + 5
                    self._devre_kablo(c, son_x, y_merkez,
                                       ampul_x - AMPUL_R, y_merkez,
                                       akim=local_akim)
                    self._devre_ampul_at(c, ampul_x, y_merkez, s, AMPUL_R,
                                          veya_grubu=False,
                                          font_sembol=SEMBOL_FONT,
                                          font_etiket=ETIKET_FONT)
                    # Bu ampul VAR değilse sonraki kablolarda akım kesilir
                    if s.get("durum") != "var":
                        local_akim = False
                    son_x = ampul_x + AMPUL_R
                    x = son_x + AMPUL_X_GAP - 2 * AMPUL_R - 5
                x += 5

                # Akım güncellemesi
                if cift_par:
                    if cift_son_grup:
                        akim_aktif = (cift_baslangic_akim if cift_baslangic_akim is not None else True) and cift_tam
                        cift_baslangic_akim = None
                    # else: çift başlangıcı, akım değişmez (son grupta belirlenir)
                else:
                    akim_aktif = local_akim

        # ⊖ Pil — SAĞ (akım çıkışı varsa yeşil kablo)
        x += 25
        self._devre_kablo(c, son_x, y_merkez, x - 14, y_merkez,
                           akim=akim_aktif)
        cikis_renk = "#1B5E20" if akim_aktif else "#37474F"
        c.create_text(x, y_merkez, text="⊖", fill=cikis_renk,
                      font=PIL_FONT, tags="devre")
        c.create_text(x, y_merkez + 28, text="ÇIKIŞ", fill="#37474F",
                      font=("Segoe UI", 8, "bold"), tags="devre")
        toplam_w = x + 30

        # Bilgi grupları — devre dışında alt sağ köşeye yatay liste
        bilgi_var = any("(bilgi)" in g for g in sira)
        if bilgi_var:
            bilgi_y = canvas_h - 60
            bilgi_x = KENAR
            c.create_text(bilgi_x, bilgi_y,
                          text="ⓘ Manuel doğrulama (devre dışı):",
                          anchor="w", fill="#9E9E9E",
                          font=("Segoe UI", 8, "bold"), tags="devre")
            bilgi_y += 16
            for g_ad in sira:
                if "(bilgi)" not in g_ad:
                    continue
                for s in gruplar_dict[g_ad]:
                    d = s.get("durum", "na")
                    sembol = "?" if d == "kontrol_edilemedi" else "i"
                    ad = s.get("ad", "")[:48]
                    c.create_text(bilgi_x + 10, bilgi_y,
                                  text=f"{sembol}  {ad}",
                                  anchor="w", fill="#BDBDBD",
                                  font=("Segoe UI", 8), tags="devre")
                    bilgi_y += 14

        # Scroll region — toplam genişlik baz al
        c.configure(scrollregion=(0, 0, max(toplam_w, c.winfo_width()),
                                    canvas_h))

    def _devre_ampul_at(self, canvas: tk.Canvas, x: int, y: int,
                         sart: dict, r: int, veya_grubu: bool = False,
                         font_sembol=("Segoe UI", 14, "bold"),
                         font_etiket=("Segoe UI", 9)) -> None:
        """Tek bir ampul + ÜST etiket (SUT lafzı, gri) + ALT etiket (durum, renkli).

        Üst: SUT kuralı lafzı, açık gri (sart.ad)
        Alt: hastanın reçete/rapor durumu, duruma göre yeşil/kırmızı/sarı (sart.neden)
        """
        d = sart.get("durum", "na")
        if d == "var":
            renk_key = "var"
        elif d == "yok":
            renk_key = "or_yok" if veya_grubu else "yok"
        elif d == "kontrol_edilemedi":
            renk_key = "ke"
        else:
            renk_key = "bilgi"

        fill, outline = self._DEVRE_RENK[renk_key]
        sembol = self._DEVRE_SEMBOL[renk_key]

        # Ampul halo (parlama efekti)
        canvas.create_oval(x - r - 3, y - r - 3, x + r + 3, y + r + 3,
                           fill="", outline=fill, width=1, tags="devre")
        item = canvas.create_oval(x - r, y - r, x + r, y + r,
                                    fill=fill, outline=outline, width=2,
                                    tags=("devre", "ampul"))
        # Sembol (büyük) ampul içinde
        canvas.create_text(x, y, text=sembol, fill="white",
                            font=font_sembol, tags="devre")

        # ── ÜST etiket: SUT lafzı (ne aranır) — açık gri ──────────────
        sut_lafzi = sart.get("ad", "")
        sut_kisa = sut_lafzi if len(sut_lafzi) <= 28 else sut_lafzi[:26] + ".."
        canvas.create_text(x, y - r - 14, text=sut_kisa,
                            fill="#90A4AE",  # açık gri (mevzuat referansı)
                            font=font_etiket, tags="devre",
                            width=145, anchor="s")

        # ── ALT etiket: hasta durumu (gerçek değer/sebep) — duruma göre renk ──
        durum_renk_map = {
            "var":    "#1B5E20",  # koyu yeşil
            "yok":    "#B71C1C",  # koyu kırmızı
            "or_yok": "#F57F17",  # koyu sarı/turuncu
            "ke":     "#E65100",  # turuncu
            "bilgi":  "#9E9E9E",  # gri
        }
        durum_fg = durum_renk_map.get(renk_key, "#37474F")
        # Alt etiket: neden alanı (rapor/teşhis'ten gelen gerçek bilgi)
        neden = sart.get("neden", "") or "—"
        # Sembol önek olarak ekle
        bdg_sembol = {"var": "✓", "yok": "✗", "or_yok": "○",
                       "ke": "?", "bilgi": "i"}.get(renk_key, "•")
        neden_kisa = neden if len(neden) <= 30 else neden[:28] + ".."
        canvas.create_text(x, y + r + 14,
                            text=f"{bdg_sembol} {neden_kisa}",
                            fill=durum_fg,
                            font=("Segoe UI", 8, "bold"),
                            tags="devre", width=145, anchor="n")

        if neden:
            self._sema_neden_map[item] = sart.get("neden", "")

    def _sema_show_tip(self, widget: tk.Widget, text: str) -> None:
        """Basit tooltip — sadece sebep metni."""
        self._sema_hide_tip()
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + widget.winfo_height() + 4
        tip = tk.Toplevel(self.root)
        tip.wm_overrideredirect(True)
        tip.wm_geometry(f"+{x}+{y}")
        tk.Label(tip, text=text, bg="#FFFDE7", fg="#37474F",
                 font=("Segoe UI", 8), justify="left",
                 relief="solid", bd=1, padx=6, pady=3,
                 wraplength=380).pack()
        self._sema_tip = tip

    def _sema_hide_tip(self) -> None:
        tip = getattr(self, "_sema_tip", None)
        if tip is not None:
            try:
                tip.destroy()
            except Exception:
                pass
            self._sema_tip = None
            self._sema_tip_row = None

    # ----------------------------------------------------------- AY/YIL
    def _ay_listesini_yukle(self):
        if not self.db:
            return
        try:
            rows = self.db.sorgu_calistir(
                "SELECT DISTINCT YEAR(RxKayitTarihi) AS Y FROM ReceteAna "
                "WHERE RxSilme=0 ORDER BY Y DESC")
            yillar = [str(r["Y"]) for r in rows if r["Y"]]
            self.cb_yil["values"] = ["Tümü"] + yillar
            if yillar:
                bugun = date.today()
                hedef = str(bugun.year)
                self.var_yil.set(hedef if hedef in yillar else yillar[0])
                self._yil_degisti()
        except Exception as e:
            logger.exception("Yıl listesi: %s", e)

    def _yil_degisti(self, *args):
        self.cb_ay["values"] = ["Tümü"] + [self._ay_etiketi(i)
                                              for i in range(1, 13)]
        bugun = date.today()
        self.var_ay.set(self._ay_etiketi(bugun.month))

    @staticmethod
    def _ay_etiketi(no: int) -> str:
        adlar = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                 "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
        return f"{no:02d}-{adlar[no - 1]}"

    @staticmethod
    def _ay_no_cek(et: str) -> int:
        try:
            return int(et.split("-")[0])
        except Exception:
            return 0

    # ----------------------------------------------------------- SORGU
    def _receteleri_sorgula(self):
        if not self.db:
            self._durum_yaz("DB bağlı değil")
            return
        yil_str = (self.var_yil.get() or "").strip()
        ay_str = (self.var_ay.get() or "").strip()

        # "Tümü" → None (filtre yok)
        if yil_str == "" or yil_str == "Tümü":
            yil = None
        else:
            try:
                yil = int(yil_str)
            except ValueError:
                messagebox.showwarning("Uyarı", "Geçersiz yıl")
                return
        if ay_str == "" or ay_str == "Tümü":
            ay = None
        else:
            ay = self._ay_no_cek(ay_str)
            if ay == 0:
                ay = None

        # Hem yıl hem ay 'Tümü' ise kullanıcıyı uyar (büyük veri seti)
        if yil is None and ay is None:
            if not messagebox.askyesno(
                "Tüm Reçeteler",
                "Yıl ve ay 'Tümü' — TÜM reçeteler yüklenecek.\n"
                "Bu uzun sürebilir ve çok bellek kullanabilir.\n\n"
                "Devam edilsin mi?"):
                return

        # Önce mevcut state kaydet
        self._state_kaydet()
        # Tabloyu temizle
        self.tum_satirlar = []
        self.satir_indeks = {}
        self.satir_renkleri = {}
        self.tv.delete(*self.tv.get_children())
        # Dönem etiketi (renk state'i bu anahtara kaydedilir)
        if yil is not None and ay is not None:
            self.aktif_donem = f"{yil}-{ay:02d}"
        elif yil is not None:
            self.aktif_donem = f"{yil}-TUM"
        elif ay is not None:
            self.aktif_donem = f"TUM-{ay:02d}"
        else:
            self.aktif_donem = "TUM"
        self.lbl_sayim.config(text="Sorgulanıyor…")
        self._durum_yaz(f"{self.aktif_donem} dönemi sorgulanıyor (ilaç bazlı)…")
        threading.Thread(target=self._sorgu_threadi, args=(yil, ay),
                          daemon=True).start()

    def _sorgu_threadi(self, yil, ay):
        """yil/ay None ise o filtre uygulanmaz (Tümü).
        gizle_bos_satirlar AÇIKsa: SQL düzeyinde uyarı kodu / mesaj / rapor
        kodu hiçbiri olmayan satırlar baştan elenir → daha az veri, daha
        hızlı yükleme."""
        try:
            # Ana SQL: ReceteAna × ReceteIlaclari × Musteri × Doktor × Urun × ATC × RaporAna
            # İlaç-rapor bağlantısı: RIRaporKodId + RRKIRaporKodId + Hasta + tarih aralığı
            # OUTER APPLY ile her ilaç için en uygun (en yeni) aktif raporu seç
            where_parts = ["ra.RxSilme = 0"]
            params = []
            if yil is not None:
                where_parts.append("YEAR(ra.RxKayitTarihi) = ?")
                params.append(int(yil))
            if ay is not None:
                where_parts.append("MONTH(ra.RxKayitTarihi) = ?")
                params.append(int(ay))

            # 🚫 Kurum dışlama (özel sigorta vb. — toggle'dan bağımsız her zaman)
            # Eski "Dışlamalar" butonu yerine: Filtre Ayarları > Kurum sekmesi.
            try:
                import aylik_filtre_ayarlari as fa
                disla_ay = self._filtre_ayarlari_aktif or fa.ayarlari_yukle()
                disla_kos = fa.sql_kurum_kosullari(disla_ay)
                if disla_kos:
                    where_parts.append(disla_kos)
            except Exception as e:
                logger.warning("Kurum dışlama SQL'e eklenemedi: %s", e)

            # 🚫 Boş Satırları Gizle — detaylı filtre ayarlarını uygula
            try:
                bos_gizle = self.gizle_bos_satirlar.get()
            except Exception:
                bos_gizle = False
            if bos_gizle:
                try:
                    import aylik_filtre_ayarlari as fa
                    ay = self._filtre_ayarlari_aktif or fa.ayarlari_yukle()
                    renkli_idler = sorted(
                        (self._renkli_recete_idler or {}).keys())
                    icerik_kos = fa.sql_icerik_kosullari(ay, renkli_idler)
                    if icerik_kos:
                        where_parts.append(icerik_kos)
                    liste_kos = fa.sql_liste_kosullari(ay)
                    if liste_kos:
                        where_parts.append(liste_kos)
                except Exception as e:
                    logger.warning("Filtre ayarları SQL'e dönüştürülemedi: %s", e)

            where_sql = " AND ".join(where_parts)

            sql = f"""
                SELECT
                    ra.RxId, ra.RxEReceteNo, ra.RxSgkIslemNo,
                    ra.RxIslemTarihi, ra.RxKayitTarihi,
                    ra.RxBransId, ra.RxKurumId, ra.RxHastaneId,
                    ra.RxMusteriId, ra.RxDoktorId,
                    ra.RxReceteRenkId, ra.RxReceteAltTuruId, ra.RxProvizyonTipId,
                    m.MusteriAdiSoyadi, m.MusteriTCKN, m.MusteriDogumTarihi,
                    m.MusteriCinsiyet, m.MusteriKapsamId, m.MusteriEmeklilik,
                    d.DoktorAdiSoyadi,
                    ri.RIId, ri.RIUrunId, ri.RIRaporKodId, ri.RIRaporNo,
                    ri.RIAdet, ri.RIDoz, ri.RITekrar, ri.RIAralik, ri.RIPeriyotId,
                    ri.RIToplam, ri.RIFiyatFarki,
                    u.UrunAdi, u.UrunATCId, u.UrunEsdegerId,
                    atc.ATCKodu, atc.ATCTurkce,
                    rapinfo.RaporAnaId, rapinfo.RaporAnaRaporNo,
                    rapinfo.RaporAnaRaporTarihi, rapinfo.RaporAnaAciklamalar,
                    rapinfo.rapor_secim_kaynagi
                FROM ReceteAna ra
                INNER JOIN ReceteIlaclari ri ON ri.RIRxId = ra.RxId
                                              AND ri.RISilme = 0
                LEFT JOIN Musteri m ON m.MusteriId = ra.RxMusteriId
                LEFT JOIN Doktor d ON d.DoktorId = ra.RxDoktorId
                LEFT JOIN Urun u ON u.UrunId = ri.RIUrunId
                LEFT JOIN ATC atc ON atc.ATCId = u.UrunATCId
                OUTER APPLY (
                    -- 2-KATMANLI RAPOR SEÇİMİ:
                    -- Tier 1: RIRaporKodId tam eşleşme (doktorun seçtiği rapor
                    --         kategorisi). Bunun içinde ATC'ye uygun ICD'si
                    --         olan raporu öne al.
                    -- Tier 2: Tier 1 boşsa → ATC'ye uygun ICD'si olan herhangi
                    --         bir aktif rapor (örn. RIRaporKodId=7 ile
                    --         eşleşen rapor yok ama hastanın E78 ICD'li
                    --         lipid raporu var → onu seç + uyarı bayrağı).
                    -- ATC mapping: Statin (C10) → E78, ARB/HT (C09/C02) → I10-I15,
                    --              Diyabet (A10) → E10-E14
                    SELECT TOP 1
                        rap.RaporAnaId, rap.RaporAnaRaporNo,
                        rap.RaporAnaRaporTarihi, rap.RaporAnaAciklamalar,
                        CASE WHEN rrki.RRKIRaporKodId = ri.RIRaporKodId
                             THEN 'tier1' ELSE 'tier2' END AS rapor_secim_kaynagi
                    FROM RaporRaporKodlariICD rrki
                    INNER JOIN RaporAna rap ON rap.RaporAnaId = rrki.RRKIRaporAnaId
                                            AND rap.RaporAnaMusteriId = ra.RxMusteriId
                                            AND (rap.RaporAnaSilme IS NULL
                                                 OR rap.RaporAnaSilme = 0)
                    LEFT JOIN ICD i_t ON i_t.ICDId = rrki.RRKIICDId
                    WHERE (rrki.RRKISilme IS NULL OR rrki.RRKISilme = 0)
                      AND (rrki.RRKIBaslamaTarihi IS NULL
                           OR rrki.RRKIBaslamaTarihi <= ra.RxKayitTarihi)
                      AND (rrki.RRKIBitisTarihi IS NULL
                           OR rrki.RRKIBitisTarihi >= ra.RxKayitTarihi)
                      AND (
                        rrki.RRKIRaporKodId = ri.RIRaporKodId
                        OR (
                          (atc.ATCKodu LIKE 'C10%' AND i_t.ICDKodu LIKE 'E78%')
                          OR (atc.ATCKodu LIKE 'C09%' AND
                              (i_t.ICDKodu LIKE 'I10%' OR i_t.ICDKodu LIKE 'I11%'
                            OR i_t.ICDKodu LIKE 'I12%' OR i_t.ICDKodu LIKE 'I13%'
                            OR i_t.ICDKodu LIKE 'I15%'))
                          OR (atc.ATCKodu LIKE 'C02%' AND
                              (i_t.ICDKodu LIKE 'I10%' OR i_t.ICDKodu LIKE 'I11%'
                            OR i_t.ICDKodu LIKE 'I12%' OR i_t.ICDKodu LIKE 'I13%'
                            OR i_t.ICDKodu LIKE 'I15%'))
                          OR (atc.ATCKodu LIKE 'A10%' AND
                              (i_t.ICDKodu LIKE 'E10%' OR i_t.ICDKodu LIKE 'E11%'
                            OR i_t.ICDKodu LIKE 'E12%' OR i_t.ICDKodu LIKE 'E13%'
                            OR i_t.ICDKodu LIKE 'E14%'))
                        )
                      )
                    ORDER BY
                        -- 1) Tier 1 (RIRaporKodId tam eşleşme) her zaman önce
                        CASE WHEN rrki.RRKIRaporKodId = ri.RIRaporKodId
                             THEN 0 ELSE 1 END,
                        -- 2) Tier 1 içinde ATC'ye uygun ICD'si olanı öne al
                        CASE WHEN EXISTS (
                            SELECT 1 FROM RaporRaporKodlariICD rrki2
                            INNER JOIN ICD i2 ON i2.ICDId = rrki2.RRKIICDId
                            WHERE rrki2.RRKIRaporAnaId = rap.RaporAnaId
                              AND (rrki2.RRKISilme IS NULL OR rrki2.RRKISilme = 0)
                              AND (
                                (atc.ATCKodu LIKE 'C10%' AND i2.ICDKodu LIKE 'E78%')
                                OR (atc.ATCKodu LIKE 'C09%' AND
                                    (i2.ICDKodu LIKE 'I10%' OR i2.ICDKodu LIKE 'I11%'
                                  OR i2.ICDKodu LIKE 'I12%' OR i2.ICDKodu LIKE 'I13%'
                                  OR i2.ICDKodu LIKE 'I15%'))
                                OR (atc.ATCKodu LIKE 'C02%' AND
                                    (i2.ICDKodu LIKE 'I10%' OR i2.ICDKodu LIKE 'I11%'
                                  OR i2.ICDKodu LIKE 'I12%' OR i2.ICDKodu LIKE 'I13%'
                                  OR i2.ICDKodu LIKE 'I15%'))
                                OR (atc.ATCKodu LIKE 'A10%' AND
                                    (i2.ICDKodu LIKE 'E10%' OR i2.ICDKodu LIKE 'E11%'
                                  OR i2.ICDKodu LIKE 'E12%' OR i2.ICDKodu LIKE 'E13%'
                                  OR i2.ICDKodu LIKE 'E14%'))
                              )
                        ) THEN 0 ELSE 1 END,
                        rap.RaporAnaRaporTarihi DESC
                ) rapinfo
                WHERE {where_sql}
                ORDER BY ra.RxKayitTarihi DESC, ri.RIId
            """
            try:
                rows = self.db.sorgu_calistir(sql, tuple(params))
                logger.info(f"Sorgu OK: {len(rows)} ilaç satırı")
            except Exception as e_sql:
                logger.error("Tam sorgu fail: %s", e_sql)
                # Hata kullanıcıya göster
                hata_msg = str(e_sql)
                self.root.after(0, lambda: messagebox.showerror(
                    "SQL Hatası", f"Sorgu fail:\n{hata_msg[:300]}"))
                self.root.after(0, lambda: self.lbl_sayim.config(text=f"SQL Hata"))
                self.root.after(0, self._durum_yaz, f"SQL Hata: {hata_msg[:120]}")
                return

            # Doktor branş bilgisi
            doktor_idleri = list({r["RxDoktorId"] for r in rows if r.get("RxDoktorId")})
            doktor_brans = self._doktor_branslarini_getir(doktor_idleri)

            # Reçete bazlı toplu sorgular
            rx_idler = list({r["RxId"] for r in rows})
            recete_teshis = self._toplu_recete_teshis_getir(rx_idler)
            recete_aciklama = self._toplu_recete_aciklama_getir(rx_idler)
            medula_yaniti = self._toplu_medula_yanit_getir(rx_idler)
            uyari_per_ilac, uyari_recete_geneli = (
                self._toplu_uyari_kodu_per_ilac_getir(rx_idler))

            # Rapor bazlı toplu sorgular (ana SQL'den gelen RaporAnaId'ler)
            rapor_ana_idler = list({r["RaporAnaId"] for r in rows
                                       if r.get("RaporAnaId")})
            rapor_detay = self._toplu_rapor_detay_getir(rapor_ana_idler)

            # Ürün bazlı UMTMesaj (her ilacın gerçek ilaç mesajı)
            urun_idler = list({r["RIUrunId"] for r in rows if r.get("RIUrunId")})
            urun_mesaj = self._toplu_urun_mesaj_getir(urun_idler)

            # Sonuçları yapıya çevir
            satirlar = []
            for r in rows:
                satir = self._satir_olustur(
                    r, doktor_brans, recete_teshis, recete_aciklama,
                    medula_yaniti, uyari_per_ilac, uyari_recete_geneli,
                    rapor_detay, urun_mesaj)
                satirlar.append(satir)

            self.root.after(0, self._sorgu_bitti, satirlar)
        except Exception as e:
            logger.exception("Sorgu hatası: %s", e)
            self.root.after(0, self._durum_yaz, f"Sorgu hatası: {e}")
            self.root.after(0, lambda: self.lbl_sayim.config(text="Hata"))

    def _doktor_branslarini_getir(self, doktor_idleri: list) -> dict:
        if not doktor_idleri:
            return {}
        try:
            ph = ",".join("?" * len(doktor_idleri))
            rows = self.db.sorgu_calistir(
                f"""SELECT DoktorBransDoktorId, DoktorBransBransId
                    FROM DoktorBrans WHERE DoktorBransDoktorId IN ({ph})""",
                tuple(doktor_idleri))
            result = {}
            for r in rows:
                did = r["DoktorBransDoktorId"]
                ad = self._lookup_brans.get(r["DoktorBransBransId"], "")
                if ad:
                    result.setdefault(did, []).append(ad)
            return {k: ", ".join(v) for k, v in result.items()}
        except Exception:
            return {}

    def _toplu_recete_teshis_getir(self, rx_idler: list) -> dict:
        """Reçete teşhislerini RxId bazında topla.
        İki kaynak:
        1) ReceteICD + ICD tablosu (ICD kodu + açıklama)
        2) ReceteTeshis + Teshis tablosu (eski Teshis sistemi)
        """
        if not rx_idler:
            return {}
        result = {}
        # 1) ReceteICD üzerinden (yeni sistem — ICD-10 kodu)
        try:
            for i in range(0, len(rx_idler), 1000):
                chunk = rx_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT ricd.ReceteICDRxId, icd.ICDKodu, icd.ICDAciklamasi
                        FROM ReceteICD ricd
                        LEFT JOIN ICD icd ON icd.ICDId = ricd.ReceteICDICDId
                        WHERE ricd.ReceteICDRxId IN ({ph})
                          AND ricd.ReceteICDSilme = 0""",
                    tuple(chunk))
                for r in rows:
                    rxid = r["ReceteICDRxId"]
                    kod = (r.get("ICDKodu") or "").strip()
                    aciklama = (r.get("ICDAciklamasi") or "").strip()
                    if kod and aciklama:
                        result.setdefault(rxid, []).append(f"{kod} {aciklama}")
                    elif kod:
                        result.setdefault(rxid, []).append(kod)
        except Exception as e:
            logger.warning("ReceteICD sorgu fail: %s", e)

        # 2) ReceteTeshis + Teshis (eski sistem — TeshisId üzerinden)
        try:
            for i in range(0, len(rx_idler), 1000):
                chunk = rx_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT rt.RTRxId, rt.RTTeshisKodu, t.TeshisAciklama
                        FROM ReceteTeshis rt
                        LEFT JOIN Teshis t ON t.TeshisId = rt.RTTeshisId
                        WHERE rt.RTRxId IN ({ph})""",
                    tuple(chunk))
                for r in rows:
                    rxid = r["RTRxId"]
                    aciklama = (r.get("TeshisAciklama") or "").strip()
                    kod = (r.get("RTTeshisKodu") or "").strip()
                    if aciklama:
                        # "001 - Teşhisi yok" formatında geliyor; "Seçiniz" gibi anlamsızları atla
                        if "Seçiniz" not in aciklama:
                            result.setdefault(rxid, []).append(aciklama)
        except Exception as e:
            logger.warning("ReceteTeshis sorgu fail: %s", e)
        return {k: " | ".join(dict.fromkeys(v)) for k, v in result.items()}

    def _toplu_recete_aciklama_getir(self, rx_idler: list) -> dict:
        """Reçete açıklamalarını RxId bazında topla.
        Bağlantı zinciri: ReceteAna.RxEReceteNo = ERecete.EReceteNo,
                            ERecete.EReceteId = EReceteAciklamalari.ERAEReceteId
        Yani RxId → EReceteId → açıklamalar (e-reçete ID'si üzerinden)
        """
        if not rx_idler:
            return {}
        result = {}
        try:
            for i in range(0, len(rx_idler), 1000):
                chunk = rx_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT ra.RxId,
                               eat.EReceteAciklamaTuruAdi,
                               ea.EReceteAciklamaAdi
                        FROM ReceteAna ra
                        INNER JOIN ERecete er ON er.EReceteNo = ra.RxEReceteNo
                                              AND er.EReceteSilme = 0
                        INNER JOIN EReceteAciklamalari era
                                 ON era.ERAEReceteId = er.EReceteId
                        LEFT JOIN EReceteAciklama ea
                                 ON ea.EReceteAciklamaId = era.ERAEReceteAciklamaId
                        LEFT JOIN EReceteAciklamaTuru eat
                                 ON eat.EReceteAciklamaTuruId = era.ERAEReceteAciklamaTuruId
                        WHERE ra.RxId IN ({ph})""",
                    tuple(chunk))
                for r in rows:
                    rxid = r["RxId"]
                    tur = r.get("EReceteAciklamaTuruAdi") or ""
                    ad = (r.get("EReceteAciklamaAdi") or "").strip()
                    # Boş veya anlamsız ("." gibi) açıklamaları atla
                    if not ad or ad in (".", ",", "-", "--"):
                        continue
                    if tur and tur != "Seçiniz":
                        txt = f"[{tur}] {ad}"
                    else:
                        txt = ad
                    result.setdefault(rxid, []).append(txt)
        except Exception as e:
            logger.warning("Reçete açıklama sorgu fail: %s", e)
        return {k: " | ".join(v) for k, v in result.items()}

    def _toplu_urun_mesaj_getir(self, urun_idler: list) -> dict:
        """UrunId → [mesaj1, mesaj2, ...] mapping (UMTUrunMesaj + UMTMesaj).
        Bu Botanik EOS'un ilaç mesaj sistemi — Medula'nın bir ürün için
        gösterdiği SUT mesajlarını (272, EK-4/E, 4.2.14.A vb.) içerir.
        """
        if not urun_idler:
            return {}
        result = {}
        try:
            for i in range(0, len(urun_idler), 1000):
                chunk = urun_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT umt.UMTUMUrunId, m.UMTMMesaj, m.UMTMSutKodu
                        FROM UMTUrunMesaj umt
                        LEFT JOIN UMTMesaj m ON m.UMTMId = umt.UMTUMUMTMesajId
                        WHERE umt.UMTUMUrunId IN ({ph})""",
                    tuple(chunk))
                for r in rows:
                    uid = r["UMTUMUrunId"]
                    mesaj = (r.get("UMTMMesaj") or "").strip()
                    sut = (r.get("UMTMSutKodu") or "").strip()
                    if mesaj:
                        result.setdefault(uid, []).append({
                            "mesaj": mesaj,
                            "sut": sut,
                        })
        except Exception as e:
            logger.warning("UMTUrunMesaj sorgu fail: %s", e)
        return result

    def _toplu_medula_yanit_getir(self, rx_idler: list) -> dict:
        """Medula provizyon yanıt mesajları (RxUyarilari.RUAciklama).
        Bunlar Medula'nın reçeteye verdiği yanıt mesajlarıdır
        (örn: "uyarı kodları uyumsuz (367)").
        Returns: {rxid: "mesaj1 | mesaj2"}
        """
        if not rx_idler:
            return {}
        result = {}
        try:
            for i in range(0, len(rx_idler), 1000):
                chunk = rx_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT RxId, RUAciklama FROM RxUyarilari
                        WHERE RxId IN ({ph})""",
                    tuple(chunk))
                for r in rows:
                    rxid = r["RxId"]
                    txt = r.get("RUAciklama") or ""
                    if txt:
                        result.setdefault(rxid, []).append(txt)
        except Exception as e:
            logger.warning("Medula yanıt sorgu fail: %s", e)
        return {k: " | ".join(v) for k, v in result.items()}

    def _toplu_uyari_kodu_per_ilac_getir(self, rx_idler: list):
        """Reçeteye GİRİLMİŞ uyarı kodları (ReceteTeshis tablosu).
        TAMPROST → 256 gibi: kullanıcının reçete kaydı sırasında
        ilaca atadığı uyarı kodları burada tutulur.
        Returns:
            per_ilac: {(rxid, urun_id): "256 - Benign prostat hiperplazisi"}
            recete_geneli: {rxid: "..."}  # RTUrunId boş olanlar (reçete genel)
        """
        per_ilac: dict = {}
        recete_geneli: dict = {}
        if not rx_idler:
            return per_ilac, recete_geneli
        try:
            for i in range(0, len(rx_idler), 1000):
                chunk = rx_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT rt.RTRxId, rt.RTUrunId, rt.RTTeshisKodu,
                               t.TeshisAciklama
                        FROM ReceteTeshis rt
                        LEFT JOIN Teshis t ON t.TeshisId = rt.RTTeshisId
                        WHERE rt.RTRxId IN ({ph})""",
                    tuple(chunk))
                for r in rows:
                    rxid = r["RTRxId"]
                    urun_id = r.get("RTUrunId")
                    kod = (r.get("RTTeshisKodu") or "").strip()
                    aciklama = (r.get("TeshisAciklama") or "").strip()
                    if not kod:
                        continue
                    # Aciklama zaten "256 - Benign prostat..." formatında
                    txt = aciklama if aciklama else kod
                    if "Seçiniz" in txt:
                        continue
                    if urun_id:
                        per_ilac.setdefault((rxid, urun_id), []).append(txt)
                    else:
                        recete_geneli.setdefault(rxid, []).append(txt)
        except Exception as e:
            logger.warning("ReceteTeshis uyarı kodu sorgu fail: %s", e)
        per_ilac = {k: " | ".join(dict.fromkeys(v)) for k, v in per_ilac.items()}
        recete_geneli = {k: " | ".join(dict.fromkeys(v))
                          for k, v in recete_geneli.items()}
        return per_ilac, recete_geneli

    def _toplu_rapor_detay_getir(self, rapor_ana_idler: list) -> dict:
        """RaporAnaId → {teshisler, etkin_madde_doz, ek_bilgiler}.
        Tek sorgu yerine 3 ayrı toplu sorgu (ek bilgi, etken madde, ICD).
        """
        if not rapor_ana_idler:
            return {}
        result = {rid: {"ek_bilgi": [], "etkin_doz": [], "icd": []}
                   for rid in rapor_ana_idler}

        # 0) Rapor ICD/teşhis (RaporRaporKodlariICD + ICD)
        try:
            for i in range(0, len(rapor_ana_idler), 1000):
                chunk = rapor_ana_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT rrki.RRKIRaporAnaId,
                               icd1.ICDKodu AS K1, icd1.ICDAciklamasi AS A1,
                               icd2.ICDKodu AS K2, icd2.ICDAciklamasi AS A2,
                               icd3.ICDKodu AS K3, icd3.ICDAciklamasi AS A3,
                               icd4.ICDKodu AS K4, icd4.ICDAciklamasi AS A4,
                               icd5.ICDKodu AS K5, icd5.ICDAciklamasi AS A5
                        FROM RaporRaporKodlariICD rrki
                        LEFT JOIN ICD icd1 ON icd1.ICDId = rrki.RRKIICDId
                        LEFT JOIN ICD icd2 ON icd2.ICDId = rrki.RRKIICDId2
                        LEFT JOIN ICD icd3 ON icd3.ICDId = rrki.RRKIICDId3
                        LEFT JOIN ICD icd4 ON icd4.ICDId = rrki.RRKIICDId4
                        LEFT JOIN ICD icd5 ON icd5.ICDId = rrki.RRKIICDId5
                        WHERE rrki.RRKIRaporAnaId IN ({ph})
                          AND rrki.RRKISilme = 0""",
                    tuple(chunk))
                for r in rows:
                    rid = r["RRKIRaporAnaId"]
                    if rid not in result:
                        result[rid] = {"ek_bilgi": [], "etkin_doz": [], "icd": []}
                    for n in (1, 2, 3, 4, 5):
                        kod = (r.get(f"K{n}") or "").strip()
                        ack = (r.get(f"A{n}") or "").strip()
                        if kod and ack:
                            result[rid]["icd"].append(f"{kod} {ack}")
                        elif kod:
                            result[rid]["icd"].append(kod)
        except Exception as e:
            logger.warning("RaporICD sorgu fail: %s", e)

        # 1) RaporEkBilgi — ek açıklamalar
        try:
            for i in range(0, len(rapor_ana_idler), 1000):
                chunk = rapor_ana_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT REBRaporAnaId, REBTuru, REBDeger, REBAciklama
                        FROM RaporEkBilgi WHERE REBRaporAnaId IN ({ph})""",
                    tuple(chunk))
                for r in rows:
                    rid = r["REBRaporAnaId"]
                    parts = []
                    if r.get("REBTuru"):
                        parts.append(str(r["REBTuru"]))
                    if r.get("REBDeger"):
                        parts.append(str(r["REBDeger"]))
                    if r.get("REBAciklama"):
                        parts.append(str(r["REBAciklama"]))
                    if parts:
                        result[rid]["ek_bilgi"].append(": ".join(parts))
        except Exception as e:
            logger.warning("RaporEkBilgi sorgu fail: %s", e)

        # 2) RaporEtkinMadde — rapor doz bilgisi
        try:
            for i in range(0, len(rapor_ana_idler), 1000):
                chunk = rapor_ana_idler[i:i + 1000]
                ph = ",".join("?" * len(chunk))
                rows = self.db.sorgu_calistir(
                    f"""SELECT EtkinMaddeRaporAnaId, EtkinMaddeId,
                               EtkinMaddeDoz, EtkinMaddeAdetMiktar,
                               EtkinMaddeTekrar, EtkinMaddeAralik,
                               EtkinMaddePeriyotId
                        FROM RaporEtkinMadde
                        WHERE EtkinMaddeRaporAnaId IN ({ph})
                          AND EtkinMaddeSilme = 0""",
                    tuple(chunk))
                for r in rows:
                    rid = r["EtkinMaddeRaporAnaId"]
                    em_id = r.get("EtkinMaddeId")
                    em_kod, em_ad = self._lookup_etkin_madde.get(em_id, ("", ""))
                    parts = []
                    if em_ad:
                        parts.append(em_ad)
                    doz = r.get("EtkinMaddeDoz")
                    adet = r.get("EtkinMaddeAdetMiktar")
                    tekrar = r.get("EtkinMaddeTekrar")
                    if doz:
                        parts.append(f"Doz:{doz}")
                    if adet:
                        parts.append(f"Adt:{adet}")
                    if tekrar:
                        parts.append(f"x{tekrar}")
                    if parts:
                        result[rid]["etkin_doz"].append({
                            "etkin_id": em_id,
                            "etkin_ad": em_ad,
                            "metin": " ".join(parts),
                            "doz": doz, "adet": adet, "tekrar": tekrar,
                            "aralik": r.get("EtkinMaddeAralik"),
                            "periyot_id": r.get("EtkinMaddePeriyotId"),
                        })
        except Exception as e:
            logger.warning("RaporEtkinMadde sorgu fail: %s", e)

        return result

    def _satir_olustur(self, r, doktor_brans, recete_teshis, recete_aciklama,
                        medula_yaniti, uyari_per_ilac, uyari_recete_geneli,
                        rapor_detay, urun_mesaj=None) -> dict:
        urun_mesaj = urun_mesaj or {}
        uyari_per_ilac = uyari_per_ilac or {}
        uyari_recete_geneli = uyari_recete_geneli or {}
        medula_yaniti = medula_yaniti or {}
        rxid = r.get("RxId")
        riid = r.get("RIId")

        # Rapor kodu (RIRaporKodId üzerinden lookup)
        rk_id = r.get("RIRaporKodId") or 0
        rap_kod_str, rap_kod_acik = self._lookup_rapor_kodu.get(rk_id, ("", ""))

        # Rapor takip numarası: RIRaporNo öncelikli (doktorun reçeteye yazdığı),
        # boşsa eşleşen rapordan (RaporAnaRaporNo) — RIRaporNo bug'ında düşmez.
        rap_tak_no = (str(r.get("RIRaporNo") or "").strip()
                      or str(r.get("RaporAnaRaporNo") or "").strip())
        # "Rap.Teşhis/Rap.Tak.No" birleşik gösterim: "04.05 - 512432126"
        if rap_kod_str and rap_tak_no:
            rap_tesh_tak = f"{rap_kod_str} - {rap_tak_no}"
        else:
            rap_tesh_tak = rap_kod_str or rap_tak_no

        # ATC (ana sorgudan, ATC tablosu join'lendi)
        atc_kodu = r.get("ATCKodu") or ""
        atc_ad = r.get("ATCTurkce") or ""

        # ───────────────────────────────────────────────────────────────
        # ETKEN MADDE — ATC.ATCTurkce her satırda dolu (raporlu/raporsuz fark etmez).
        # Raporlu ilaçlarda ek olarak RaporEtkinMadde'deki etken maddeler de gösterilir.
        # ───────────────────────────────────────────────────────────────
        atc_turkce = (r.get("ATCTurkce") or "").strip()
        rapor_ana_id = r.get("RaporAnaId")
        em_ad_rapor = ""
        rap_doz_metin = ""
        rap_doz_sayi = 0.0
        rapor_aciklamalari = []
        rap_tesh_listesi = []

        if rapor_ana_id and rapor_ana_id in rapor_detay:
            rd = rapor_detay[rapor_ana_id]
            # Rapor ICD/teşhisleri
            rap_tesh_listesi = rd.get("icd", [])
            # Etkin madde bilgisi rapor üzerinden
            em_listesi = rd.get("etkin_doz", [])
            if em_listesi:
                em_ad_rapor = " + ".join(e.get("etkin_ad", "") for e in em_listesi
                                           if e.get("etkin_ad"))
                doz_metinleri = [e.get("metin", "") for e in em_listesi
                                  if e.get("metin")]
                rap_doz_metin = " | ".join(doz_metinleri)
                ilk = em_listesi[0]
                try:
                    d = float(ilk.get("doz") or 0)
                    a_ = float(ilk.get("adet") or 0)
                    t_ = float(ilk.get("tekrar") or 1)
                    rap_doz_sayi = max(d, a_) * t_
                except Exception:
                    pass
            # Rapor açıklamaları: ana açıklama + ek bilgi listesi.
            # NOT: Botanik EOS'ta bir RaporAna.RaporAnaAciklamalar alanı
            # birden fazla açıklama satırı içerebilir (newline ile ayrılmış).
            # Örn: "Açıklama 1 (Ekleme=...)\nldl: 149 mg/dl (Ekleme=...)".
            # Her satırı ayrı bir açıklama olarak ele al ki kontrol modülleri
            # ikinci satırdaki LDL/risk bilgisini de yakalayabilsin.
            ana_ack = r.get("RaporAnaAciklamalar") or ""
            if ana_ack:
                for parca in ana_ack.replace("\r\n", "\n").split("\n"):
                    parca = parca.strip()
                    if parca:
                        rapor_aciklamalari.append(parca)
            rapor_aciklamalari.extend(rd.get("ek_bilgi", []))

        # Etken madde finali: ÖNCE bu ilacın kendi ATC'si (drug-specific).
        # Raporun etken maddeleri (em_ad_rapor) raporun TÜM ilaçları için
        # ortaktır — her bir ilaç satırına kopyalandığında karışıklık
        # yapar (ör: BENIPIN satırında ŞEKER ÇUBUĞU görünmesi). Bu yüzden
        # em_ad SADECE ATC.ATCTurkce'den alınır; ATC boşsa rapor etkenine
        # düşer.
        if atc_turkce:
            em_ad = atc_turkce
        elif em_ad_rapor:
            em_ad = em_ad_rapor
        else:
            em_ad = ""

        rap_ack = " | ".join(rapor_aciklamalari) if rapor_aciklamalari else ""
        rap_tesh = " | ".join(dict.fromkeys(rap_tesh_listesi)) if rap_tesh_listesi else ""

        # Tarih
        tar = r.get("RxIslemTarihi") or r.get("RxKayitTarihi")
        tar_str = tarih_format(tar, "%d.%m.%Y")

        # Hasta tipi
        kapsam_id = r.get("MusteriKapsamId")
        hasta_tip = self._lookup_kapsam.get(kapsam_id, "")

        # Reçete türü (Normal/Kırmızı/Yeşil/Mor) ve Alt türü (Ayaktan/Yatan/A/B/C/GK)
        renk_id = r.get("RxReceteRenkId")
        rec_tip = self._lookup_renk.get(renk_id, "")
        alttur_id = r.get("RxReceteAltTuruId")
        rec_alttur = self._lookup_alttur.get(alttur_id, "")

        # Grup belirleme
        grup = self._grup_belirle(r)

        # Reçete dozu — "<aralık> <periyot> <tekrar> x <doz>" formatında
        # ÖRNEKLER: "1 günde 2 x 1" / "7 günde 1 x 2" / "1 ayda 1 x 1"
        # NOT: RIAdet (kutu sayısı) BURAYA dahil edilmez — ayrı "kutu" sütununda gösterilir.
        rd_metin = doz_metin(
            r.get("RIAralik"),     # aralık (1, 7, vs.)
            r.get("RIPeriyotId"),  # periyot türü (3=günde, 4=haftada, 5=ayda)
            r.get("RITekrar"),     # tekrar (kez)
            r.get("RIDoz"),        # doz (her keferinde alınan miktar)
        )
        rd_sayi = float(r.get("RIDoz") or 0) * float(r.get("RITekrar") or 1)

        # SUT maddesi tespiti
        sut_madde = sut_madde_tespit(r.get("UrunAdi") or "", em_ad,
                                       rap_kod_str)

        # Dönem
        donem = (r.get("RxKayitTarihi") or r.get("RxIslemTarihi"))
        donem_str = tarih_format(donem, "%Y-%m")

        # Mesaj durumu: bu ürünün UMTUrunMesaj'da kayıtlı SUT mesajı var mı?
        # (Medula reçete kayıt sırasında gösterdiği "Msj: var/yok" bu tabloya geliyor)
        urun_id = r.get("RIUrunId")
        ilac_mesajlari = urun_mesaj.get(urun_id, []) if urun_id else []
        msj = "var" if ilac_mesajlari else "yok"

        # Uyarı kodu: bu reçetede bu ilaç için ReceteTeshis'e girilmiş kodlar.
        # Öncelik:
        #  1) Reçete-ilaç bazlı kod (RTUrunId = bu ilaç)
        #  2) Reçete geneline yazılmış kod (RTUrunId boş)
        ilac_kodu = uyari_per_ilac.get((rxid, urun_id), "") if urun_id else ""
        recete_kodu = uyari_recete_geneli.get(rxid, "")
        if ilac_kodu and recete_kodu:
            uyari_text = f"{ilac_kodu} | (Reç: {recete_kodu})"
        elif ilac_kodu:
            uyari_text = ilac_kodu
        elif recete_kodu:
            uyari_text = f"(Reç: {recete_kodu})"
        else:
            uyari_text = ""

        # Medula provizyon yanıt metni (RxUyarilari) — ayrı sütun
        medula_msj_text = medula_yaniti.get(rxid, "")

        return {
            "ri_id": riid, "rx_id": rxid,
            "rapor_ana_id": rapor_ana_id,
            # tier1 = doktorun seçtiği RIRaporKodId ile eşleşen rapor
            # tier2 = ATC bazlı ICD eşleşmesi (kod uyuşmadı, fallback)
            "rapor_secim_kaynagi": (r.get("rapor_secim_kaynagi") or ""),
            "musteri_id": r.get("RxMusteriId"),
            "secim": "☐",  # tablo render sırasında secili_iidler'e göre güncellenir
            "grup": grup,
            "donem": donem_str,
            "rec_tar": tar_str,
            "rec_no": r.get("RxEReceteNo") or "",
            # Medula "Reçete Sorgu" sayfasında sorgulanan sistem reçete no
            # (RxSgkIslemNo, ör: 3N0HR6D). E-Reçete'den ayrıdır.
            "sistem_recete_no": (r.get("RxSgkIslemNo") or "").strip(),
            "rec_tip": rec_tip,
            "rec_alttur": rec_alttur,
            "hasta": r.get("MusteriAdiSoyadi") or "",
            "tc": r.get("MusteriTCKN") or "",
            "yas": yas_hesapla(r.get("MusteriDogumTarihi")),
            "cins": cinsiyet_etiket(r.get("MusteriCinsiyet")),
            "hasta_tip": hasta_tip,
            "doktor": r.get("DoktorAdiSoyadi") or "",
            "brans": doktor_brans.get(r.get("RxDoktorId"), ""),
            "kurum_adi": self._lookup_kurum.get(r.get("RxKurumId"), ""),
            "tesis_kodu": self._lookup_hastane_kodu.get(
                r.get("RxHastaneId"), ""),
            "ilac": r.get("UrunAdi") or "",
            "etkin": em_ad,
            "atc": atc_kodu,
            "esdeger": str(r.get("UrunEsdegerId") or ""),
            "kutu": str(r.get("RIAdet") or ""),
            "sut": sut_madde,
            "rap_kod": rap_kod_str,
            "rap_tak_no": rap_tak_no,
            "rap_tesh_tak": rap_tesh_tak,
            "rec_doz": rd_metin,
            "rec_doz_sayi": rd_sayi,
            "rap_doz": rap_doz_metin,
            "rap_doz_sayi": rap_doz_sayi,
            "msj": msj,
            "uyari": uyari_text,
            "medula_msj": medula_msj_text,
            "rec_tesh": recete_teshis.get(rxid, ""),
            "rap_tesh": rap_tesh,
            "rec_ack": recete_aciklama.get(rxid, ""),
            "rap_ack": rap_ack,
            # Doz karşılaştırma için ham değerler (doz_kontrol modülü kullanır)
            "rec_doz_raw": {
                "doz": r.get("RIDoz"),
                "tekrar": r.get("RITekrar"),
                "aralik": r.get("RIAralik"),
                "periyot_id": r.get("RIPeriyotId"),
            },
            "rap_doz_listesi": (
                rapor_detay.get(rapor_ana_id, {}).get("etkin_doz", [])
                if rapor_ana_id else []
            ),
            "renkli_rr": "",
        }

    def _grup_belirle(self, r) -> str:
        """A/B/C/GK/CK tespiti (kural-bazlı tahmin).

        NOT: A/B/C/CK/GK grupları Botanik EOS DB'sinde reçete bazında
        doğrudan kayıtlı değil (SGK fatura kesimi sırasında belirleniyor).
        Bu fonksiyon kural-bazlı tahmin yapar:
          - Kan ürünü ATC kodları (J06, B05) → CK
          - Hasta kapsamı 60/c1 (yeşil kart benzeri) → GK
          - Reçete tipi 50 (Magistral) → "M"
          - Diğer → boş (eczacı manuel kategorize eder)
        """
        atc_kodu = (r.get("ATCKodu") or "").upper()
        kapsam_id = r.get("MusteriKapsamId")
        kapsam_ad = (self._lookup_kapsam.get(kapsam_id, "") or "").upper()
        recete_tipi = r.get("RxReceteTipi")

        # 1) CK — Kan ürünü tespiti (immünglobulin, kan/plazma türevleri)
        if atc_kodu:
            if (atc_kodu.startswith("J06A") or       # İmmünglobulin
                    atc_kodu.startswith("B05A") or   # Plazma türevleri (faktörler)
                    atc_kodu.startswith("B02") or    # Antihemorrhagic
                    atc_kodu.startswith("B03X")):    # ESA + diğer hematolojikler
                return "CK"

        # 2) GK — Geçici Koruma (genelde sığınmacı/yabancı kapsamı)
        # Botanik EOS'ta direkt "Geçici Koruma" kapsamı yok ama bazı
        # özel kodlar olabilir
        if "GEÇICI KORUMA" in kapsam_ad or "GEÇİCİ KORUMA" in kapsam_ad:
            return "GK"
        if "MÜLTECİ" in kapsam_ad or "SIĞINMACI" in kapsam_ad:
            return "GK"

        # 3) Magistral
        if recete_tipi == 50:
            return "M"

        # 4) A/B/C ayrımı için güvenilir alan yok — eczacı manuel kategorize eder
        return ""

    def _sorgu_bitti(self, satirlar: list):
        """SQL sonucu geldi, tabloyu doldur."""
        self.tum_satirlar = satirlar
        self.satir_indeks = {}
        # State'i yükle (önceki boyamalar)
        kayitli_renkler = self._state_yukle_donem(self.aktif_donem)

        for s in satirlar:
            riid = s["ri_id"]
            iid = str(riid)
            self.satir_indeks[iid] = s
            renk = kayitli_renkler.get(iid, RENK_BEYAZ)
            self.satir_renkleri[iid] = renk

        # Sütun görünüm ayarını uygula (kullanıcının tercihi varsa)
        self._sutun_gorunumunu_uygula()
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        self._durum_yaz(f"{len(satirlar)} satır geldi (ilaç bazlı)")

    # ----------------------------------------------------------- TABLO RENDER
    @staticmethod
    def _sart_raporu_metni(s: dict) -> str:
        """Satırın verdict_* alanlarından tabloya şart raporu üretir.

        Önce yapısal `verdict_sartlar` (JSON: [{ad, durum, neden, ...}, ...])
        varsa CLAUDE.md disiplinine uygun gruplandırma (✓/✗/?). Yoksa eski
        verdict_detay/uyari/sut/aranan/bulunan birleşimine düşer.
        """
        verdict = (s.get("verdict") or "").strip()
        if not verdict:
            return ""

        # Yapısal şart listesi varsa onu kullan
        sartlar_raw = s.get("verdict_sartlar")
        sartlar = None
        if sartlar_raw:
            if isinstance(sartlar_raw, str):
                try:
                    sartlar = json.loads(sartlar_raw)
                except Exception:
                    sartlar = None
            elif isinstance(sartlar_raw, list):
                sartlar = sartlar_raw

        if sartlar:
            saglanan = [p for p in sartlar if p.get("durum") == "var"]
            saglanmayan = [p for p in sartlar if p.get("durum") == "yok"]
            ke = [p for p in sartlar if p.get("durum") == "kontrol_edilemedi"]

            def _ozet(p):
                ad = (p.get("ad") or "").strip()
                neden = (p.get("neden") or "").strip()
                return f"{ad}: {neden}" if neden else ad

            parcalar = [f"📋 {verdict}"]
            sut_k = (s.get("verdict_sut") or "").strip()
            if sut_k:
                parcalar.append(sut_k)
            if saglanan:
                parcalar.append("✓ " + " | ".join(_ozet(p) for p in saglanan))
            if saglanmayan:
                parcalar.append("✗ " + " | ".join(_ozet(p) for p in saglanmayan))
            if ke:
                parcalar.append("? " + " | ".join(_ozet(p) for p in ke))
            uyari = (s.get("verdict_uyari") or "").strip()
            if uyari:
                parcalar.append(f"⚠ {uyari}")
            return " · ".join(parcalar)

        # Eski özet (yapısal şart listesi yoksa)
        parcalar = []
        sut = (s.get("verdict_sut") or "").strip()
        if sut:
            parcalar.append(sut)
        detay = (s.get("verdict_detay") or "").strip()
        if detay:
            parcalar.append(detay)
        aranan = (s.get("verdict_aranan") or "").strip()
        bulunan = (s.get("verdict_bulunan") or "").strip()
        if aranan and bulunan:
            parcalar.append(f"Aranan: {aranan} → Bulunan: {bulunan}")
        elif aranan:
            parcalar.append(f"Aranan: {aranan}")
        elif bulunan:
            parcalar.append(f"Bulunan: {bulunan}")
        uyari = (s.get("verdict_uyari") or "").strip()
        if uyari:
            parcalar.append(f"⚠ {uyari}")
        return " · ".join(parcalar)

    @staticmethod
    def _kontrol_raporunu_satira_yaz(s: dict, rapor, *, kategori: str = "",
                                      alt_sinif: str = "") -> None:
        """KontrolRaporu'u satırın verdict_* alanlarına yazar (yapısal sartlar dahil).

        Tüm SUT kontrol fonksiyonları (statin/diyabet/arb/klopidogrel/...)
        bu helper ile satıra yazsa kod tekrarı azalır.
        """
        from recete_kontrol.base_kontrol import KontrolSonucu
        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN: "UYGUN",
            KontrolSonucu.UYGUN_DEGIL: "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI: "ATLANDI",
        }
        s["verdict"] = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
        s["verdict_detay"] = rapor.mesaj or ""
        s["verdict_kategori"] = kategori
        s["verdict_alt_sinif"] = alt_sinif
        s["verdict_uyari"] = rapor.uyari or ""
        s["verdict_sut"] = rapor.sut_kurali or ""
        s["verdict_aranan"] = rapor.aranan_ibare or ""
        s["verdict_bulunan"] = rapor.bulunan_metin or ""
        try:
            s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                ensure_ascii=False)
        except Exception:
            s["verdict_detaylar"] = str(rapor.detaylar or {})
        # Yapısal şart listesi (CLAUDE.md disiplini)
        # grup + veya_grubu: atomik şart paneli için
        sartlar_obj = getattr(rapor, "sartlar", None) or []
        try:
            s["verdict_sartlar"] = json.dumps([
                {"ad": p.ad,
                 "durum": p.durum.value if hasattr(p.durum, "value") else str(p.durum),
                 "neden": p.neden,
                 "kaynak": getattr(p, "kaynak", ""),
                 "grup": getattr(p, "grup", ""),
                 "veya_grubu": bool(getattr(p, "veya_grubu", False))}
                for p in sartlar_obj
            ], ensure_ascii=False)
        except Exception:
            s["verdict_sartlar"] = ""

    def _tabloyu_yenile(self):
        """tum_satirlar + filtreler + renk filtresi → tv'ye doldur.
        Filtreyle gizlenen satırlar seçim setinden de düşer; bu sayede sonraki
        boyama işlemlerinde sadece ekrandaki seçimler işleme alınır."""
        self.tv.delete(*self.tv.get_children())
        self.gosterilen_iids = set()

        # Reçete türü filtresi — SADECE kapatılan tipler dışlanır.
        # Default (hepsi açık) ya da bilinmeyen tip → satır geçer.
        gizli_recete_turleri = {
            ad for ad, v in (self.var_recete_turu or {}).items()
            if not v.get()
        } if hasattr(self, "var_recete_turu") and self.var_recete_turu else set()

        for s in self.tum_satirlar:
            iid = str(s["ri_id"])
            if not self._satir_filtreden_geciyor_mu(s):
                continue
            renk = self.satir_renkleri.get(iid, RENK_BEYAZ)
            if renk not in self.aktif_renk_filtre:
                continue
            # Reçete türü filtresi: sadece kullanıcının kapattığı tipleri ele
            if gizli_recete_turleri:
                rec_tip = (s.get("rec_tip") or "").strip()
                if rec_tip in gizli_recete_turleri:
                    continue
            # Seçim göstergesi (☐/☑)
            s["secim"] = "☑" if iid in self.secili_iidler else "☐"
            # Şart raporu sütunu: verdict_* alanlarından kompakt özet
            s["verdict_sart_raporu"] = self._sart_raporu_metni(s)
            values = tuple(str(s.get(k, "")) for k in SUTUN_KOD)
            self.tv.insert("", "end", iid=iid, values=values, tags=(renk,))
            self.gosterilen_iids.add(iid)

        # Filtre ile gizlenen satırların seçimini de düşür — kullanıcının
        # "sadece görünenleri boyuyorum" beklentisini garanti eder.
        if self.secili_iidler:
            self.secili_iidler &= self.gosterilen_iids

    def _satir_filtreden_geciyor_mu(self, s: dict, haric_kod: str = None) -> bool:
        """Sütun arama kutuları + Excel-benzeri değer filtreleri ile filtre.

        haric_kod: Verildiğinde o sütunun kendi filtresi yoksayılır.
                   (Kaskad filtre popup'ı için — bir sütunun filtre popup'ı
                   açıldığında o sütunun filtresi diğer sütunlardaki seçenekleri
                   kısıtlamamalı, ama diğer sütunların filtreleri uygulanmalı.)
        """
        # 1) Alt sütun arama kutuları (içerir) — placeholder metni filtreye
        # etki etmesin diye atla.
        # Boşlukla ayrılmış kelimeler AND ile birleşir; aralarındaki diğer
        # karakterler/sözcükler önemsiz. Her kelime hücre içinde substring
        # olarak geçmeli. Örn: "metfor sülf maks" araması, hücredeki
        # "...metformin ve sülfonil ürelerin maksimum tolere..." metniyle
        # eşleşir.
        for kod, var in self.arama_varlari.items():
            if kod == haric_kod:
                continue
            if (getattr(self, "_placeholder_aktif", {})
                    .get(kod)):
                continue
            ara = _tr_lower((var.get() or "").strip())
            if not ara:
                continue
            deger = _tr_lower(str(s.get(kod, "")))
            parcalar = ara.split()
            if not all(p in deger for p in parcalar):
                return False
        # 2) Excel-benzeri değer filtreleri (set içinde mi?)
        for kod, secili in self.aktif_deger_filtre.items():
            if kod == haric_kod:
                continue
            if not secili:
                continue
            if str(s.get(kod, "")) not in secili:
                return False
        # 3) Excel-tarzı metin operatör filtreleri (başlar/biter/içerir...)
        #    Çoklu koşul: VEYA grup ayırıcı, VE grup içi. Eşleşme = bir grubun
        #    tüm koşulları sağlanır (gruplar OR ile birleşir).
        for kod, kosullar in self.aktif_metin_filtre.items():
            if kod == haric_kod:
                continue
            kosul_listesi = self._metin_filtre_normalize(kosullar)
            if not kosul_listesi:
                continue
            s_str = _tr_lower(str(s.get(kod, "") or ""))
            # VEYA'lara göre gruplara böl
            gruplar = []
            mevcut = []
            for i, k in enumerate(kosul_listesi):
                if i > 0 and k.get("baglac") == "veya":
                    if mevcut:
                        gruplar.append(mevcut)
                    mevcut = []
                mevcut.append(k)
            if mevcut:
                gruplar.append(mevcut)
            # Anlamlı koşulu olmayan grupları (ör. değer boş & op=içerir) at
            anlamli_gruplar = [
                g for g in gruplar
                if any(self._metin_kosulu_anlamli(k) for k in g)
            ]
            if not anlamli_gruplar:
                continue
            # En az bir grup tamamen eşleşmeli
            if not any(
                all(self._metin_kosulu_test(s_str, k) for k in g
                    if self._metin_kosulu_anlamli(k))
                for g in anlamli_gruplar
            ):
                return False
        return True

    @staticmethod
    def _metin_filtre_normalize(kosullar):
        """Geriye dönük uyum: eski (op, deger) tuple → tek-koşul liste.
        Yeni format zaten liste; aynen döner."""
        if isinstance(kosullar, tuple) and len(kosullar) == 2:
            op, deger = kosullar
            return [{"baglac": None, "op": op, "deger": deger}]
        if isinstance(kosullar, list):
            return kosullar
        return []

    @staticmethod
    def _metin_kosulu_anlamli(k):
        """Boş değer + boş-olmayan op = anlamsız (filtre kapalı sayılır)."""
        op = k.get("op", "icerir")
        if op in ("bos", "bos_degil"):
            return True
        return bool((k.get("deger") or "").strip())

    @staticmethod
    def _metin_kosulu_test(s_str, k):
        """Tek koşul s_str'e uyuyor mu? s_str ZATEN lowercase olmalı."""
        op = k.get("op", "icerir")
        d_str = _tr_lower(k.get("deger") or "")
        if op == "icerir":
            return d_str in s_str
        if op == "icermez":
            return d_str not in s_str
        if op == "baslar":
            return s_str.startswith(d_str)
        if op == "biter":
            return s_str.endswith(d_str)
        if op == "esit":
            return s_str == d_str
        if op == "esit_degil":
            return s_str != d_str
        if op == "bos":
            return not s_str
        if op == "bos_degil":
            return bool(s_str)
        if op == "regex":
            try:
                import re as _re
                return bool(_re.search(d_str, s_str, _re.IGNORECASE))
            except _re.error:
                return True  # geçersiz regex → koşulu yutkun
        return True

    def _sutun_filtre_degisti(self, kod: str):
        # Placeholder yazımları filtre tetiklemesin
        if (getattr(self, "_placeholder_aktif", {})
                .get(kod)):
            return
        # Kısa debounce ile çağırılabilir; şimdilik direkt
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    def _renk_filtre_degisti(self):
        self.aktif_renk_filtre = {r for r, v in self.var_renk.items()
                                    if v.get()}
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    def _bos_satir_filtre_degisti(self):
        """🚫 Boş Satırları Gizle değişti → SQL düzeyinde filtre değişti
        demektir, sorguyu yeniden çalıştır. Henüz veri çekilmediyse hiç
        bir şey yapma."""
        if not self.tum_satirlar:
            return
        self._durum_yaz("🚫 Boş satır filtresi değişti — sorgu yenileniyor…")
        self._receteleri_sorgula()

    def _filtre_ayar_penceresi_ac(self):
        """⚙ butonu — detaylı filtre ayar penceresini açar."""
        try:
            import aylik_filtre_ayarlari as fa
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Hata",
                                  f"Filtre ayar modülü yüklenemedi:\n{e}")
            return

        def _on_save(yeni_ayarlar):
            # Yeni ayarları memory'e al ve sorguyu yenile
            self._filtre_ayarlari_aktif = yeni_ayarlar
            if self.tum_satirlar:
                # Yalnızca bos_gizle açıkken yeniden sorgu mantıklı
                self._durum_yaz(
                    "⚙ Filtre ayarları güncellendi — sorgu yenileniyor…")
                self._receteleri_sorgula()
            else:
                self._durum_yaz("⚙ Filtre ayarları kaydedildi "
                                  "(sonraki Sorgula çalışmasında uygulanır)")

        fa.ayar_penceresini_ac(self.root, db=self.db, on_save=_on_save)

    def _recete_turu_filtre_degisti(self):
        """📋 Reçete türü checkbox'ı (Beyaz/Kırmızı/Yeşil/Mor) değişti.
        Yüklü veri arasında filtreleme — SQL'e gitmeden tablo yenilenir."""
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    def _tum_filtreleri_temizle(self):
        """Tüm aktif filtreleri sıfırla:
        - Sütun arama kutuları (alt satır)
        - Excel-benzeri değer filtreleri (sağ tık popup'tan)
        - Renk filtreleri (5 renk de açık)
        - Sıralama
        """
        # Sütun arama kutularını temizle (placeholder geri gelsin)
        for kod, var in self.arama_varlari.items():
            try:
                # Placeholder durumunu sıfırla — temiz başla
                if hasattr(self, "_placeholder_aktif"):
                    self._placeholder_aktif[kod] = False
                var.set("")
            except Exception:
                pass
        # Entry'lere placeholder geri yerleştir (focus out gibi davran)
        try:
            self._filtre_inner.focus_set()
        except Exception:
            pass
        # Excel-benzeri değer filtrelerini sıfırla
        self.aktif_deger_filtre.clear()
        # Metin filtrelerini sıfırla (başlar/biter/içerir...)
        self.aktif_metin_filtre.clear()
        # Renk filtrelerini hepsini aç
        for v in self.var_renk.values():
            try:
                v.set(True)
            except Exception:
                pass
        self.aktif_renk_filtre = {RENK_BEYAZ, RENK_YESIL, RENK_SARI,
                                    RENK_TURUNCU, RENK_KIRMIZI, RENK_GRI}
        # "Boş Satırları Gizle" — varsayılana (AÇIK) döndür
        try:
            self.gizle_bos_satirlar.set(True)
        except Exception:
            pass
        # Reçete türü filtresi → hepsi açık (varsayılan)
        try:
            for v in (self.var_recete_turu or {}).values():
                v.set(True)
        except Exception:
            pass
        # Sıralamayı sıfırla
        self._siralama_kolonu = None
        self._siralama_yonu = None
        # Tabloyu ve göstergeleri yenile
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        self._siralama_gostergesini_guncelle()
        self._durum_yaz("Tüm filtreler sıfırlandı — tüm satırlar görünür")

    def _sayaclari_guncelle(self):
        # Renk dağılımı (tek satırda kompakt)
        say = {r: 0 for r in RENK_BG.keys()}
        for r in self.satir_renkleri.values():
            say[r] = say.get(r, 0) + 1
        toplam = len(self.tum_satirlar)
        gor = len(self.gosterilen_iids)
        sec = len(self.secili_iidler)
        txt = (f"Toplam: {toplam}  |  Filtreli: {gor}  |  Seçili: {sec}  ||  "
               f"⚪ {say[RENK_BEYAZ]}  🟢 {say[RENK_YESIL]}  "
               f"🟡 {say[RENK_SARI]}  🟠 {say[RENK_TURUNCU]}  🔴 {say[RENK_KIRMIZI]}")
        try:
            self.lbl_durum_dagilim.config(text=txt)
        except Exception:
            pass
        try:
            if toplam:
                self.lbl_sayim.config(text=f"{toplam} satır  |  Filtrelenmiş: {gor}")
            else:
                self.lbl_sayim.config(text="")
        except Exception:
            pass

    # ----------------------------------------------------------- BOYAMA
    def _gri_iclac_disla(self, iidler):
        """RENK_GRI ile boyanan satırların ilaç adlarını filtre ayar
        dosyasındaki 'ilac' dışlama listesine ekler. Aynı SORGUDA
        bir daha gelmez (ana sorgu yenilenince devre dışı kalır)."""
        try:
            import aylik_filtre_ayarlari as fa
        except Exception as e:
            logger.warning("Filtre ayar modülü yok: %s", e)
            return 0
        try:
            ay = fa.ayarlari_yukle()
        except Exception as e:
            logger.warning("Filtre ayarları okunamadı: %s", e)
            return 0
        ilac_kurallari = ay.get("ilac") or []
        mevcut = {(k.get("deger") or "").strip().upper()
                   for k in ilac_kurallari
                   if k.get("mod") == "icermez"}
        eklendi = 0
        for iid in iidler:
            s = self.satir_indeks.get(iid)
            if not s:
                continue
            ilac_adi = (s.get("ilac") or "").strip()
            if not ilac_adi:
                continue
            if ilac_adi.upper() in mevcut:
                continue
            ilac_kurallari.append({"deger": ilac_adi, "mod": "icermez"})
            mevcut.add(ilac_adi.upper())
            eklendi += 1
        if eklendi:
            ay["ilac"] = ilac_kurallari
            try:
                fa.ayarlari_kaydet(ay)
                self._filtre_ayarlari_aktif = ay
            except Exception as e:
                logger.warning("Filtre ayarları yazılamadı: %s", e)
        return eklendi

    def _boya_dispatch(self, renk: str) -> None:
        """Tek-buton boyama: radio mod'a göre Seçili veya Filtreli boyar."""
        mod = (self.boya_mod.get() if hasattr(self, "boya_mod")
               else "secili")
        if mod == "filtreli":
            self._gorunenleri_boya(renk)
        else:
            self._secilenleri_renge_boya(renk)

    def _secilenleri_renge_boya(self, renk: str):
        """secili_iidler set'indeki satırları renge boya.
        İşlem sonrası seçim seti TEMİZLENİR. Renk RENK_GRI ise ilaç adı
        otomatik olarak dışlama listesine eklenir."""
        if not self.secili_iidler:
            messagebox.showinfo("Bilgi", "Önce satırları seç (☐ kutucukları)")
            return
        n = len(self.secili_iidler)
        # Gri ise dışlama listesine ekle (boyamadan önce — iidler üzerinden)
        eklendi = 0
        if renk == RENK_GRI:
            eklendi = self._gri_iclac_disla(list(self.secili_iidler))
        for iid in self.secili_iidler:
            self.satir_renkleri[iid] = renk
        self.secili_iidler.clear()
        self._state_kaydet()
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        ek = (f" — {eklendi} ilaç dışlama listesine eklendi"
              if renk == RENK_GRI and eklendi else "")
        self._durum_yaz(f"{n} satır → {RENK_ETIKET[renk]} "
                          f"(seçim temizlendi){ek}")

    def _hizli_filtre_menu(self, event=None):
        """Hızlı filtre seçim menüsü (dropdown)."""
        m = tk.Menu(self.root, tearoff=0)
        m.add_command(label="── Hızlı Filtreler ──", state="disabled")
        for etiket, fn in HIZLI_FILTRELER:
            m.add_command(label=etiket,
                            command=lambda f=fn: self._hizli_filtre_uygula(f))
        m.add_separator()
        m.add_command(label="(Filtreyi temizle)",
                        command=self._tabloyu_yenile)
        try:
            x = self.root.winfo_pointerx()
            y = self.root.winfo_pointery()
            m.tk_popup(x, y)
        finally:
            m.grab_release()

    def _gorunenleri_boya(self, renk: str):
        """Filtreden geçip listede kalan tüm satırları boyar
        (ekrana sığan değil — scroll dışındakiler de dahil).
        Renk RENK_GRI ise ilaçlar dışlama listesine eklenir."""
        if not self.gosterilen_iids:
            self._durum_yaz("Filtreli satır yok")
            return
        n = len(self.gosterilen_iids)
        ek_uyari = ""
        if renk == RENK_GRI:
            ek_uyari = ("\n\n⚠ Bu satırlardaki ilaçlar dışlama listesine "
                         "(ilaç → İçermez) eklenecektir.")
        if not messagebox.askyesno(
                "Filtreli Satırları Boya",
                f"Filtreden geçen {n} satır '{RENK_ETIKET[renk]}' rengine "
                f"boyanacak.\n(Listedeki tüm satırlar — scroll dışındakiler "
                f"dahil){ek_uyari}\n\nEmin misin?"):
            return
        eklendi = 0
        if renk == RENK_GRI:
            eklendi = self._gri_iclac_disla(list(self.gosterilen_iids))
        for iid in list(self.gosterilen_iids):
            self.satir_renkleri[iid] = renk
            self.secili_iidler.discard(iid)
        self._state_kaydet()
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        ek = (f" — {eklendi} ilaç dışlama listesine eklendi"
              if renk == RENK_GRI and eklendi else "")
        self._durum_yaz(f"{n} satır → {RENK_ETIKET[renk]}{ek}")

    def _secilenleri_boya(self):
        secimler = list(self.tv.selection())
        if not secimler:
            self._durum_yaz("Seçili satır yok")
            return
        # Renk seçim popup
        win = tk.Toplevel(self.root)
        win.title("Renk Seç")
        win.geometry("260x200")
        win.transient(self.root)
        tk.Label(win, text=f"{len(secimler)} satır için renk:",
                 font=("Segoe UI", 10, "bold")).pack(pady=8)
        for renk in [RENK_YESIL, RENK_SARI, RENK_TURUNCU, RENK_KIRMIZI, RENK_BEYAZ]:
            tk.Button(win, text=RENK_ETIKET[renk],
                      bg=RENK_BG[renk], fg=RENK_FG[renk],
                      command=lambda r=renk: (self._secilenleri_boya_uygula(secimler, r),
                                                win.destroy())
                      ).pack(fill="x", padx=20, pady=2)

    def _secilenleri_boya_uygula(self, iids, renk):
        eklendi = 0
        if renk == RENK_GRI:
            eklendi = self._gri_iclac_disla(iids)
        for iid in iids:
            self.satir_renkleri[iid] = renk
            self.secili_iidler.discard(iid)
        self._state_kaydet()
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        ek = (f" — {eklendi} ilaç dışlama listesine eklendi"
              if renk == RENK_GRI and eklendi else "")
        self._durum_yaz(f"{len(iids)} seçili satır → {RENK_ETIKET[renk]}{ek}")

    def _tum_renkleri_sifirla(self):
        if not messagebox.askyesno("Tüm Renkleri Sıfırla",
                                     "Tüm satırlar BEYAZ olacak. Emin misin?"):
            return
        for iid in self.satir_renkleri:
            self.satir_renkleri[iid] = RENK_BEYAZ
        self._state_kaydet()
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    # ----------------------------------------------------------- HIZLI FİLTRE
    def _hizli_filtre_uygula(self, fn):
        """Hızlı filtreyi uygula (sadece görünenleri filtrele)."""
        # Sütun filtrelerini koruyarak ek bir filtre uygulamak için:
        # Kullanıcı arama kutularını temizleyip sadece bu filtreyi uygulasın
        # Daha kullanıcı dostu: tablodaki iid'lerden filtrelenmiş kümeyi göster
        secilenler = set()
        for s in self.tum_satirlar:
            try:
                if fn(s):
                    secilenler.add(str(s["ri_id"]))
            except Exception:
                pass
        # Renk filtreden de geç
        gor = []
        self.tv.delete(*self.tv.get_children())
        self.gosterilen_iids = set()
        for iid in secilenler:
            s = self.satir_indeks.get(iid)
            if not s:
                continue
            renk = self.satir_renkleri.get(iid, RENK_BEYAZ)
            if renk not in self.aktif_renk_filtre:
                continue
            values = tuple(str(s.get(k, "")) for k in SUTUN_KOD)
            self.tv.insert("", "end", iid=iid, values=values, tags=(renk,))
            self.gosterilen_iids.add(iid)
        self._sayaclari_guncelle()
        self._durum_yaz(f"Hızlı filtre: {len(self.gosterilen_iids)} satır")

    def _hizli_filtre_uygula_ve_boya(self, fn):
        self._hizli_filtre_uygula(fn)
        if self.gosterilen_iids:
            n = len(self.gosterilen_iids)
            if messagebox.askyesno(
                    "Hızlı Filtre + Boya",
                    f"{n} satır filtreden geçti. Yeşile boyansın mı?"):
                self._gorunenleri_boya(RENK_YESIL)

    # ----------------------------------------------------------- SOL TIK (CHECKBOX TOGGLE)
    def _sol_tik_dispatch(self, event):
        """Sol tık — secim sütununda checkbox toggle, başka sütunlarda
        tıklanan hücrenin tam içeriğini alttaki gösterge satırında yansıt."""
        try:
            region = self.tv.identify_region(event.x, event.y)
        except Exception:
            return
        if region != "cell":
            return
        col_id = self.tv.identify_column(event.x)
        kod = self._col_id_to_kod(col_id)
        iid = self.tv.identify_row(event.y)
        if not iid:
            return

        # Secim sütunu → checkbox toggle (eski davranış)
        if kod == "secim":
            if iid in self.secili_iidler:
                self.secili_iidler.discard(iid)
                self.tv.set(iid, "secim", "☐")
            else:
                self.secili_iidler.add(iid)
                self.tv.set(iid, "secim", "☑")
            s = self.satir_indeks.get(iid)
            if s:
                s["secim"] = "☑" if iid in self.secili_iidler else "☐"
            self._sayaclari_guncelle()
            return

        # Diğer sütunlar → hücre içeriğini alt label'a bas
        try:
            deger = ""
            if kod:
                deger = self.tv.set(iid, kod)
            # Sütun başlığı
            baslik = ""
            for k, b, _g, _t in SUTUNLAR:
                if k == kod:
                    baslik = b
                    break
            self.lbl_hucre_baslik.config(text=f"[{baslik}]")
            self.lbl_hucre_icerik.config(
                text=deger if deger else "(boş)")
        except Exception as e:
            logger.debug("hücre içeriği gösterimi: %s", e)

    def _tumunu_sec(self):
        """Görünen tüm satırları işaretle."""
        for iid in self.gosterilen_iids:
            self.secili_iidler.add(iid)
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    def _hicbirini_secme(self):
        """Görünen satırların seçimini kaldır (gizli satırların seçimi korunur).
        Bu Excel benzeri tutarlı davranış için: 'Tümü Seç' görünenleri seçer,
        'Hiçbiri' de görünenlerin işaretini kaldırır.
        """
        for iid in self.gosterilen_iids:
            self.secili_iidler.discard(iid)
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    def _secimi_tersine_cevir(self):
        """Görünen satırlardan seçili olanları kaldır, olmayanları işaretle."""
        for iid in self.gosterilen_iids:
            if iid in self.secili_iidler:
                self.secili_iidler.discard(iid)
            else:
                self.secili_iidler.add(iid)
        self._tabloyu_yenile()
        self._sayaclari_guncelle()

    def _vurgulanmislari_isaretle(self):
        """Treeview'da Shift/Ctrl ile vurgulanmış satırların checkbox'ını işaretle."""
        vurgulu = list(self.tv.selection())
        if not vurgulu:
            self._durum_yaz("Vurgulanmış satır yok "
                            "(Shift/Ctrl ile satır seçin)")
            return
        eklendi = 0
        for iid in vurgulu:
            if iid not in self.secili_iidler:
                self.secili_iidler.add(iid)
                eklendi += 1
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        self._durum_yaz(f"{len(vurgulu)} vurgulu satır işaretlendi "
                        f"({eklendi} yeni)")

    # ----------------------------------------------------------- SAĞ TIK & DETAY
    def _sag_tik_dispatch(self, event):
        """Başlığa mı satıra mı tıklandı? Doğru menüyü aç."""
        try:
            region = self.tv.identify_region(event.x, event.y)
        except Exception:
            region = ""
        if region == "heading" or region == "separator":
            col_id = self.tv.identify_column(event.x)
            kod = self._col_id_to_kod(col_id)
            if kod:
                self._sutun_basligi_menu(event, kod)
                return
        # Satır
        self._sag_tik_menu(event)

    def _col_id_to_kod(self, col_id: str):
        """Treeview '#1', '#2' formatından sütun kodunu çıkart."""
        if not col_id or not col_id.startswith("#"):
            return None
        try:
            idx = int(col_id[1:]) - 1
            # Görünen sütunlar (displaycolumns) varsa ona göre
            disp = self.tv["displaycolumns"]
            if disp and disp != "#all":
                if isinstance(disp, str):
                    return None
                if 0 <= idx < len(disp):
                    return disp[idx]
            else:
                if 0 <= idx < len(SUTUN_KOD):
                    return SUTUN_KOD[idx]
        except Exception:
            pass
        return None

    def _sutun_basligi_menu(self, event, kod: str):
        """Excel-benzeri sütun başlığı sağ tık menüsü."""
        baslik = next((b for k, b, _g, _t in SUTUNLAR if k == kod), kod)
        m = tk.Menu(self.root, tearoff=0)
        # Sıralama
        m.add_command(label="▲ Artan sırala (A→Z)",
                        command=lambda: self._kolon_sirala_belirli(kod, "asc"))
        m.add_command(label="▼ Azalan sırala (Z→A)",
                        command=lambda: self._kolon_sirala_belirli(kod, "desc"))
        m.add_command(label="↕ Sıralamayı temizle",
                        command=lambda: self._kolon_sirala_belirli(kod, None))
        m.add_separator()
        # Excel-benzeri değer filtresi
        m.add_command(label=f"🔍 '{baslik}' sütununda filtrele (değer listesi)…",
                        command=lambda: self._sutun_deger_filtre_popup(kod))
        # Metin filtresi (başlar/biter/içerir/regex...)
        m.add_command(label=f"🔎 '{baslik}' metin filtresi (başlar/biter/içerir...)…",
                        command=lambda: self._metin_filtre_popup(kod))
        if kod in self.aktif_deger_filtre or kod in self.aktif_metin_filtre:
            m.add_command(label=f"⊘ '{baslik}' filtresini temizle",
                            command=lambda: self._sutun_filtresini_temizle(kod))
        m.add_separator()
        # Hızlı bu değere göre filtrele
        m.add_command(label=f"⚙ '{baslik}' sütununu gizle",
                        command=lambda: self._sutunu_hizli_gizle(kod))
        m.add_command(label="⚙ Sütun ayarları...",
                        command=self._sutun_ayar_penceresi)
        try:
            m.tk_popup(event.x_root, event.y_root)
        finally:
            m.grab_release()

    def _kolon_sirala_belirli(self, kod, yon):
        """Belirli yönde sırala."""
        if yon is None:
            self._siralama_kolonu = None
            self._siralama_yonu = None
        else:
            self._siralama_kolonu = kod
            self._siralama_yonu = yon
        self._siralama_gostergesini_guncelle()
        if self._siralama_kolonu:
            ornekler = [s.get(kod, "") for s in self.tum_satirlar[:50] if s.get(kod)]
            sayisal = bool(ornekler) and all(
                self._sayiya_cevrilebilir_mi(o) for o in ornekler[:10])
            rev = (yon == "desc")
            if sayisal:
                self.tum_satirlar.sort(
                    key=lambda s: self._sayiya_cevir(s.get(kod, "")), reverse=rev)
            else:
                self.tum_satirlar.sort(
                    key=lambda s: str(s.get(kod, "") or "").lower(), reverse=rev)
        self._tabloyu_yenile()

    def _sutun_deger_filtre_popup(self, kod: str):
        """Excel-benzeri sütun değer filtre popup'ı.
        Tüm benzersiz değerler checkbox listesi olarak gösterilir.
        """
        baslik = next((b for k, b, _g, _t in SUTUNLAR if k == kod), kod)
        win = tk.Toplevel(self.root)
        win.title(f"Filtrele — {baslik}")
        win.geometry("440x560")
        win.transient(self.root)

        # KASKAD FİLTRE — sadece DİĞER filtreler uygulandıktan sonra GÖRÜNECEK
        # satırların değerleri listelensin. Bu sütunun kendi filtresi yoksayılır
        # (yoksa popup tek seçili değerle gelir).
        from collections import Counter
        sayim = Counter()
        for s in self.tum_satirlar:
            # Diğer sütun filtrelerini uygula (bu sütun hariç)
            if not self._satir_filtreden_geciyor_mu(s, haric_kod=kod):
                continue
            # Renk filtresini de uygula
            iid = str(s.get("ri_id"))
            renk = self.satir_renkleri.get(iid, RENK_BEYAZ)
            if renk not in self.aktif_renk_filtre:
                continue
            sayim[str(s.get(kod, "") or "")] += 1
        # Sıralı liste — boşları sona, geri kalanı alfabetik
        degerler = sorted(sayim.keys(),
                            key=lambda d: (d == "", d.lower()))

        tk.Label(win, text=f"📋 {baslik}",
                 font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=8, pady=(8, 2))
        tk.Label(win,
                 text=f"{len(degerler)} farklı değer | Diğer filtrelerden geçen: "
                       f"{sum(sayim.values())} satır",
                 fg="#546E7A", font=("Segoe UI", 8)).pack(anchor="w", padx=8)

        # Arama kutusu — değerleri filtrele
        ara_frame = tk.Frame(win)
        ara_frame.pack(fill="x", padx=8, pady=4)
        tk.Label(ara_frame, text="🔍").pack(side="left")
        var_ara = tk.StringVar()
        ent = tk.Entry(ara_frame, textvariable=var_ara)
        ent.pack(side="left", fill="x", expand=True, padx=4)
        ent.focus_set()

        # Tümü Seç / Tümü Kaldır
        ust = tk.Frame(win)
        ust.pack(fill="x", padx=8)
        tk.Button(ust, text="Tümünü Seç", bg="#C8E6C9",
                  command=lambda: self._popup_tumu(var_dict, True)
                  ).pack(side="left", padx=2)
        tk.Button(ust, text="Tümünü Kaldır", bg="#FFCDD2",
                  command=lambda: self._popup_tumu(var_dict, False)
                  ).pack(side="left", padx=2)
        tk.Button(ust, text="Aramaya Uyanları Seç", bg="#BBDEFB",
                  command=lambda: self._popup_aranana_uygula(var_dict, var_ara, True)
                  ).pack(side="left", padx=2)

        # Liste (scrollable, checkbox'lı)
        list_frame = tk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=8, pady=4)
        canvas = tk.Canvas(list_frame, highlightthickness=0, bg="white")
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        inner = tk.Frame(canvas, bg="white")
        canvas_win = canvas.create_window((0, 0), window=inner, anchor="nw")

        # Mevcut filtre durumu
        secili = self.aktif_deger_filtre.get(kod, set(degerler))
        var_dict = {}
        cb_dict = {}

        def _populate(filtre_metni=""):
            # Mevcut widget'ları temizle
            for w in inner.winfo_children():
                w.destroy()
            cb_dict.clear()
            f = _tr_lower(filtre_metni)
            for d in degerler:
                if f and f not in _tr_lower(str(d)):
                    continue
                v = var_dict.get(d)
                if v is None:
                    v = tk.BooleanVar(value=(d in secili))
                    var_dict[d] = v
                etiket = d if d else "(boş)"
                cnt = sayim[d]
                cb = tk.Checkbutton(
                    inner, text=f"{etiket}   ({cnt})",
                    variable=v, anchor="w", bg="white",
                    font=("Segoe UI", 9), padx=4)
                cb.pack(fill="x", anchor="w")
                cb_dict[d] = cb
            inner.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))

        _populate()

        # Arama değişikliğinde liste güncelle
        var_ara.trace_add("write", lambda *a: _populate(var_ara.get()))

        # Uygula / İptal
        alt = tk.Frame(win)
        alt.pack(fill="x", padx=8, pady=8)

        def _uygula():
            secimler = {d for d, v in var_dict.items() if v.get()}
            # Tümü seçiliyse filtre yok
            if len(secimler) == len(degerler):
                self.aktif_deger_filtre.pop(kod, None)
            elif not secimler:
                # Hiçbiri seçili değil — kullanıcı tüm satırları gizliyor
                self.aktif_deger_filtre[kod] = set()
            else:
                self.aktif_deger_filtre[kod] = secimler
            self._tabloyu_yenile()
            self._sayaclari_guncelle()
            self._siralama_gostergesini_guncelle()
            win.destroy()

        tk.Button(alt, text="Uygula", bg="#1976D2", fg="white",
                  command=_uygula, padx=14, font=("Segoe UI", 9, "bold")
                  ).pack(side="right", padx=4)
        tk.Button(alt, text="İptal", command=win.destroy, padx=12
                  ).pack(side="right", padx=4)

    def _popup_tumu(self, var_dict, deger):
        for v in var_dict.values():
            v.set(deger)

    def _popup_aranana_uygula(self, var_dict, var_ara, deger):
        f = _tr_lower(var_ara.get() or "")
        for d, v in var_dict.items():
            if not f or f in _tr_lower(str(d)):
                v.set(deger)

    def _metin_filtre_popup(self, kod: str):
        """Çoklu koşul + ve/veya bağlaçlı Excel-tarzı metin filtre popup'ı.

        Her satır: [bağlaç (ilk hariç)] [operatör] [değer] [×]
        VEYA satırı yeni grup başlatır, VE satırı grubu sürdürür.
        Eşleşme = bir grubun tüm koşulları sağlanır (gruplar OR'lanır).
        """
        baslik = next((b for k, b, _g, _t in SUTUNLAR if k == kod), kod)
        win = tk.Toplevel(self.root)
        win.title(f"Metin Filtresi — {baslik}")
        win.geometry("720x560")
        win.minsize(560, 420)
        win.transient(self.root)

        # Başlık (üstte sabit)
        tk.Label(win, text=f"📋 {baslik}",
                 font=("Segoe UI", 11, "bold")
                 ).pack(side="top", anchor="w", padx=10, pady=(10, 4))

        op_listesi = [
            ("İçerir",       "icerir"),
            ("İçermez",      "icermez"),
            ("Başlar",       "baslar"),
            ("Biter",        "biter"),
            ("Eşittir",      "esit"),
            ("Eşit değil",   "esit_degil"),
            ("Boş",          "bos"),
            ("Boş değil",    "bos_degil"),
            ("Regex (gelişmiş)", "regex"),
        ]
        op_disp_listesi = [d for d, _k in op_listesi]
        kod_to_disp = {k: d for d, k in op_listesi}
        disp_to_kod = {d: k for d, k in op_listesi}

        baglac_disp_to_kod = {"VE": "ve", "VEYA": "veya"}
        baglac_kod_to_disp = {"ve": "VE", "veya": "VEYA"}

        # Mevcut koşulları yükle (eski tuple format → tek koşullu listeye çevir)
        mevcut = self.aktif_metin_filtre.get(kod)
        baslangic = self._metin_filtre_normalize(mevcut)
        if not baslangic:
            baslangic = [{"baglac": None, "op": "icerir", "deger": ""}]

        # Liste çerçevesini oluşturuyoruz ama PACK etmeyeceğiz —
        # alt panelden (butonlar/bilgi/ekle) sonra en sona pack edilecek ki
        # buttonlar her zaman pencerenin altında, görünür kalsın.
        liste_cerceve = tk.Frame(win, bd=1, relief="solid")

        canvas = tk.Canvas(liste_cerceve, highlightthickness=0)
        scroll = ttk.Scrollbar(liste_cerceve, orient="vertical",
                                command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        ic = tk.Frame(canvas)
        ic_id = canvas.create_window((0, 0), window=ic, anchor="nw")

        def _ic_resize(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(ic_id, width=canvas.winfo_width())
        ic.bind("<Configure>", _ic_resize)
        canvas.bind("<Configure>", _ic_resize)

        # Tek satırın widget'larını tutar: list of dicts
        satirlar = []

        def _satir_ciz():
            # Önce tüm satır widget'larını temizle
            for w in ic.winfo_children():
                w.destroy()
            for i, s in enumerate(satirlar):
                fr = tk.Frame(ic)
                fr.pack(fill="x", padx=4, pady=2)
                # Bağlaç sütunu (ilk satırda boş yer tutucu)
                if i == 0:
                    tk.Label(fr, text="", width=6).pack(side="left")
                else:
                    bg_cb = ttk.Combobox(fr, values=["VE", "VEYA"],
                                          width=5, state="readonly")
                    bg_cb.set(baglac_kod_to_disp.get(s["baglac_var"].get(),
                                                      "VE"))
                    bg_cb.pack(side="left", padx=(0, 4))

                    def _on_baglac(_e=None, idx=i, cb=bg_cb):
                        satirlar[idx]["baglac_var"].set(
                            baglac_disp_to_kod.get(cb.get(), "ve"))
                    bg_cb.bind("<<ComboboxSelected>>", _on_baglac)
                # Operatör
                op_cb = ttk.Combobox(fr, values=op_disp_listesi,
                                      width=18, state="readonly")
                op_cb.set(kod_to_disp.get(s["op_var"].get(), "İçerir"))
                op_cb.pack(side="left", padx=2)

                def _on_op(_e=None, idx=i, cb=op_cb):
                    satirlar[idx]["op_var"].set(
                        disp_to_kod.get(cb.get(), "icerir"))
                op_cb.bind("<<ComboboxSelected>>", _on_op)
                # Değer
                ent = tk.Entry(fr, textvariable=s["deger_var"],
                                font=("Segoe UI", 10))
                ent.pack(side="left", padx=2, fill="x", expand=True)
                # Sil
                if len(satirlar) > 1:
                    tk.Button(fr, text="×", width=2,
                               command=lambda idx=i: _satir_sil(idx)
                               ).pack(side="left", padx=(2, 0))
                else:
                    tk.Label(fr, text="", width=3).pack(side="left")
            ic.update_idletasks()
            _ic_resize()

        def _satir_ekle(baglac="ve", op="icerir", deger=""):
            satirlar.append({
                "baglac_var": tk.StringVar(value=baglac or "ve"),
                "op_var": tk.StringVar(value=op or "icerir"),
                "deger_var": tk.StringVar(value=deger or ""),
            })
            _satir_ciz()

        def _satir_sil(idx):
            if 0 <= idx < len(satirlar) and len(satirlar) > 1:
                del satirlar[idx]
                # İlk satırın bağlacı her zaman None sayılır; tek tutmak için
                # kalan satırlar arasında bir sorun yok (görsel olarak ilk
                # satırda bağlaç gösterilmiyor zaten).
                _satir_ciz()

        # Mevcut koşulları yükle
        for i, k in enumerate(baslangic):
            satirlar.append({
                "baglac_var": tk.StringVar(value=k.get("baglac") or "ve"),
                "op_var": tk.StringVar(value=k.get("op", "icerir")),
                "deger_var": tk.StringVar(value=k.get("deger", "") or ""),
            })
        _satir_ciz()

        # === ALT PANEL — side="bottom" ile sabit, listeden ÖNCE pack ===
        # Pack sırası: alt → bilgi → ekle_fr → (en son) liste_cerceve.
        # side="bottom" ilk packlenen en alta yapışır; sonrakiler üstüne.

        def _uygula():
            yeni = []
            for i, s in enumerate(satirlar):
                op = s["op_var"].get() or "icerir"
                deger = s["deger_var"].get() or ""
                # Boş-değer + değer-gerektiren op → bu koşulu at
                if not deger.strip() and op not in ("bos", "bos_degil"):
                    continue
                baglac = None if i == 0 else (s["baglac_var"].get() or "ve")
                yeni.append({"baglac": baglac, "op": op, "deger": deger})
            # İlk anlamlı koşulun bağlacını None'a sabitle
            if yeni:
                yeni[0]["baglac"] = None
            if not yeni:
                self.aktif_metin_filtre.pop(kod, None)
            else:
                self.aktif_metin_filtre[kod] = yeni
            self._tabloyu_yenile()
            self._sayaclari_guncelle()
            self._siralama_gostergesini_guncelle()
            win.destroy()

        def _temizle():
            self.aktif_metin_filtre.pop(kod, None)
            self._tabloyu_yenile()
            self._sayaclari_guncelle()
            self._siralama_gostergesini_guncelle()
            win.destroy()

        # 1) En alta sabitlen → butonlar
        alt = tk.Frame(win)
        alt.pack(side="bottom", fill="x", padx=10, pady=(6, 10))
        tk.Button(alt, text="Uygula", bg="#1976D2", fg="white",
                  command=_uygula, padx=14, font=("Segoe UI", 9, "bold")
                  ).pack(side="right", padx=4)
        tk.Button(alt, text="Filtreyi Kaldır",
                  command=_temizle, padx=10
                  ).pack(side="right", padx=4)
        tk.Button(alt, text="İptal",
                  command=win.destroy, padx=10
                  ).pack(side="right", padx=4)

        # 2) Butonların üstüne → bilgi
        tk.Label(win,
                  text="• VE = grup içi (hepsi sağlanmalı), "
                       "VEYA = yeni grup başlatır.\n"
                       "• Eşleşme: en az bir grubun TÜM koşulları sağlanır.\n"
                       "• Boş/Boş değil için değer alanı yok sayılır. "
                       "Regex: ^A.*B$",
                  fg="#546E7A", font=("Segoe UI", 8),
                  justify="left", anchor="w"
                  ).pack(side="bottom", fill="x", padx=10, pady=(2, 4))

        # 3) Bilginin üstüne → + Koşul ekle
        ekle_fr = tk.Frame(win)
        ekle_fr.pack(side="bottom", fill="x", padx=10, pady=(2, 4))
        tk.Button(ekle_fr, text="+ Koşul ekle (VE)",
                   command=lambda: _satir_ekle("ve")
                   ).pack(side="left", padx=(0, 4))
        tk.Button(ekle_fr, text="+ Koşul ekle (VEYA)",
                   command=lambda: _satir_ekle("veya")
                   ).pack(side="left")

        # 4) En son: liste çerçevesi → ortayı doldurur (expand)
        liste_cerceve.pack(fill="both", expand=True, padx=10, pady=4)

        win.bind("<Return>", lambda e: _uygula())
        win.bind("<Escape>", lambda e: win.destroy())

    def _sutun_filtresini_temizle(self, kod):
        self.aktif_deger_filtre.pop(kod, None)
        self.aktif_metin_filtre.pop(kod, None)
        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        self._siralama_gostergesini_guncelle()

    def _sutunu_hizli_gizle(self, kod):
        self.sutun_gosterim[kod] = False
        self._sutun_ayarlarini_kaydet()
        self._sutun_gorunumunu_uygula()

    def _sag_tik_menu(self, event):
        iid = self.tv.identify_row(event.y)
        if iid and iid not in self.tv.selection():
            self.tv.selection_set(iid)
        if not self.tv.selection():
            return
        m = tk.Menu(self.root, tearoff=0)
        # En üst: Reçete / Rapor detay pencereleri (Botanik EOS'tan canlı çek)
        s_aktif = self.satir_indeks.get(self.tv.selection()[0]) or {}
        sistem_no_aktif = (s_aktif.get("sistem_recete_no") or "").strip()
        rec_no_aktif = (s_aktif.get("rec_no") or "").strip()
        rapor_ana_id_aktif = s_aktif.get("rapor_ana_id") or 0
        # Etiketde önce sistem reçete no (Medula'da kullanılan), yoksa e-reçete no
        if sistem_no_aktif:
            m.add_command(
                label=f"🔍 Reçete Detayını Göster ({sistem_no_aktif})",
                command=self._recete_detay_goster)
        elif rec_no_aktif:
            m.add_command(
                label=f"🔍 Reçete Detayını Göster ({rec_no_aktif})",
                command=self._recete_detay_goster)
        else:
            m.add_command(label="🔍 Reçete Detayını Göster (reçete no yok)",
                          state="disabled")
        if rapor_ana_id_aktif:
            m.add_command(label="📄 Rapor Detayını Göster",
                          command=self._rapor_detay_goster)
        else:
            m.add_command(label="📄 Rapor Detayını Göster (rapor yok)",
                          state="disabled")
        m.add_separator()
        # Üst grup: kopyala / Medula'da aç (tek satır işlemleri)
        m.add_command(label="📋 Hasta TC'sini Kopyala",
                      command=self._kopyala_hasta_tc)
        m.add_command(label="📋 E-Reçete No'yu Kopyala",
                      command=self._kopyala_recete_no)
        m.add_command(label="🔢 Sistem Reçete No'yu Kopyala",
                      command=self._kopyala_sistem_recete_no)
        m.add_command(label="📄 Rapor Açıklamasını Kopyala",
                      command=self._kopyala_rapor_aciklamasi)
        m.add_command(label="🪪 Tüm Künyeyi Kopyala",
                      command=self._kopyala_tum_kunye)
        m.add_command(label="🔓 Reçeteyi Medula'da Aç",
                      command=self._medulada_ac)
        m.add_separator()
        for renk in [RENK_YESIL, RENK_SARI, RENK_TURUNCU, RENK_KIRMIZI, RENK_BEYAZ]:
            m.add_command(label=f"→ {RENK_ETIKET[renk]}",
                            command=lambda r=renk: self._secilenleri_boya_uygula(
                                list(self.tv.selection()), r))
        m.add_separator()
        m.add_command(label="Detay Göster", command=self._satir_detay)
        try:
            m.tk_popup(event.x_root, event.y_root)
        finally:
            m.grab_release()

    # ----------------------------------------------------- Sağ tık: kopyala / Medula
    def _aktif_satir(self):
        """Sağ tık menüsünden çağrılan komutlar için aktif (ilk seçili) satır."""
        sec = self.tv.selection()
        if not sec:
            return None
        return self.satir_indeks.get(sec[0])

    def _panoya_yaz(self, metin: str):
        """Verilen metni panoya yaz (Tk clipboard)."""
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(metin)
            # Pano içeriğinin uygulama kapansa bile kalması için update
            self.root.update()
        except Exception as e:
            logger.warning(f"Panoya yazma hatası: {e}")

    def _kopyala_hasta_tc(self):
        s = self._aktif_satir()
        if not s:
            return
        tc = (s.get("tc") or "").strip()
        if not tc:
            messagebox.showwarning("Uyarı", "Bu satırda hasta TC bilgisi yok.")
            return
        self._panoya_yaz(tc)
        logger.info(f"Hasta TC panoya kopyalandı: {tc}")

    def _kopyala_recete_no(self):
        s = self._aktif_satir()
        if not s:
            return
        rec_no = (s.get("rec_no") or "").strip()
        if not rec_no:
            messagebox.showwarning("Uyarı", "Bu satırda E-reçete numarası yok.")
            return
        self._panoya_yaz(rec_no)
        logger.info(f"E-Reçete no panoya kopyalandı: {rec_no}")

    def _kopyala_sistem_recete_no(self):
        s = self._aktif_satir()
        if not s:
            return
        sistem_no = (s.get("sistem_recete_no") or "").strip()
        if not sistem_no:
            messagebox.showwarning(
                "Uyarı",
                "Bu satırda sistem reçete no (SGK İşlem No) yok.\n"
                "Eski/farklı tipte bir reçete olabilir.")
            return
        self._panoya_yaz(sistem_no)
        logger.info(f"Sistem reçete no panoya kopyalandı: {sistem_no}")

    def _kopyala_rapor_aciklamasi(self):
        """Sağ tık → seçili satırın ilintilendirilen rapor açıklamasını
        panoya kopyala. Rapor açıklaması birden fazla satır içeriyorsa
        her birini yeni satıra ayırıp birleştirir."""
        s = self._aktif_satir()
        if not s:
            return
        rap_ack = (s.get("rap_ack") or "").strip()
        if not rap_ack:
            messagebox.showwarning(
                "Uyarı",
                "Bu satırda rapor açıklaması yok.\n"
                "(Reçete raporsuz olabilir veya rapor eşleşmesi başarısız.)",
                parent=self.root)
            return
        # rap_ack zaten " | " ile ayrılmış birden fazla parça olabilir →
        # panoya yazarken \n ile ayır (kullanıcı için daha okunaklı).
        kopyalanan = rap_ack.replace(" | ", "\n")
        self._panoya_yaz(kopyalanan)
        rapor_no = (s.get("rap_no") or "").strip()
        ek = f" (Rapor No: {rapor_no})" if rapor_no else ""
        logger.info(f"Rapor açıklaması panoya kopyalandı{ek}")
        self._durum_yaz(
            f"📄 Rapor açıklaması panoya kopyalandı{ek} "
            f"({len(kopyalanan)} karakter)")

    def _kopyala_tum_kunye(self):
        """Sağ tık → seçili satırın tüm künyesini panoya kopyala.
        TC, hasta adı, reçete tarihi, sistem reçete no, rapor açıklaması,
        reçete açıklaması — her biri ayrı satırda."""
        s = self._aktif_satir()
        if not s:
            return
        tc        = (s.get("tc")               or "").strip()
        hasta     = (s.get("hasta")            or "").strip()
        rec_tar   = (s.get("rec_tar")          or "").strip()
        sistem_no = (s.get("sistem_recete_no") or "").strip()
        rap_ack   = (s.get("rap_ack")          or "").strip().replace(" | ", " / ")
        rec_ack   = (s.get("rec_ack")          or "").strip().replace(" | ", " / ")
        satirlar = [
            f"TC: {tc}",
            f"Hasta: {hasta}",
            f"Reçete Tarihi: {rec_tar}",
            f"Sistem Reçete No: {sistem_no}",
            f"Rapor Açıklaması: {rap_ack}",
            f"Reçete Açıklaması: {rec_ack}",
        ]
        kunye = "\n".join(satirlar)
        self._panoya_yaz(kunye)
        logger.info(f"Tüm künye panoya kopyalandı (TC: {tc})")
        self._durum_yaz(
            f"🪪 Künye panoya kopyalandı ({len(kunye)} karakter, 6 satır)")

    def _medulada_ac(self):
        """Seçili reçeteyi Medula'da aç:
        Medula'yı öne getir → Reçete Sorgu menüsüne tıkla → SGK İşlem No
        (sistem reçete no) yaz → Sorgula. Sistem no yoksa fallback E-reçete no.
        Ağır iş thread'de çalışır; GUI'yi bloklamaz.
        """
        s = self._aktif_satir()
        if not s:
            return
        # Medula "Reçete Sorgu" sayfası SGK İşlem No (sistem reçete no) ister
        sistem_no = (s.get("sistem_recete_no") or "").strip()
        rec_no = sistem_no or (s.get("rec_no") or "").strip()
        if not rec_no:
            messagebox.showwarning("Uyarı", "Bu satırda reçete numarası yok.")
            return
        threading.Thread(
            target=self._medulada_ac_worker,
            args=(rec_no,),
            daemon=True,
        ).start()

    def _medulada_ac_worker(self, recete_no: str):
        """Thread: Medula'ya bağlan, reçete sorguyu aç ve sorgula."""
        try:
            from botanik_bot import BotanikBot
        except Exception as e:
            logger.error(f"BotanikBot import hatası: {e}")
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Hata", f"Medula otomasyon modülü yüklenemedi:\n{e}"),
            )
            return

        try:
            bot = BotanikBot()
            if not bot.baglanti_kur("MEDULA", ilk_baglanti=False):
                self.root.after(
                    0,
                    lambda: messagebox.showwarning(
                        "Medula bulunamadı",
                        "Medula penceresi açık değil veya bulunamadı.\n"
                        "Önce Medula'yı açıp giriş yapın."),
                )
                return

            # Pencereyi öne getir
            try:
                import ctypes
                import win32con
                import win32gui
                hwnd = bot.medula_hwnd or getattr(bot.main_window, "handle", None)
                if hwnd:
                    if win32gui.IsIconic(hwnd):
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                    user32 = ctypes.windll.user32
                    # ALT trick (foreground lock için)
                    user32.keybd_event(0x12, 0, 0, 0)
                    user32.keybd_event(0x12, 0, 0x0002, 0)
                    user32.SetForegroundWindow(hwnd)
            except Exception as e:
                logger.debug(f"Medula öne getirme hatası: {e}")

            import time as _t
            _t.sleep(0.25)

            if not bot.recete_sorgu_ac():
                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "Hata",
                        "Reçete Sorgu menüsü açılamadı. Medula ana sayfada mı?"),
                )
                return

            _t.sleep(0.4)
            # Pencereyi yenile (sayfa değişti)
            bot.baglanti_kur("MEDULA", ilk_baglanti=False)

            if not bot.recete_no_yaz(recete_no):
                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "Hata", f"Reçete numarası yazılamadı: {recete_no}"),
                )
                return

            _t.sleep(0.2)
            if not bot.sorgula_butonuna_tikla():
                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "Hata", "Sorgula butonu tıklanamadı."),
                )
                return

            logger.info(f"✓ Medula'da reçete sorgulandı: {recete_no}")
        except Exception as e:
            logger.error(f"Medula'da aç hatası: {e}", exc_info=True)
            self.root.after(
                0,
                lambda err=e: messagebox.showerror(
                    "Hata", f"Medula'da açma sırasında hata:\n{err}"),
            )

    def _satir_detay(self, event=None):
        sec = self.tv.selection()
        if not sec:
            return
        iid = sec[0]
        s = self.satir_indeks.get(iid)
        if not s:
            return
        win = tk.Toplevel(self.root)
        win.title(f"Detay — {s.get('rec_no')} / {s.get('ilac')}")
        win.geometry("780x600")
        txt = tk.Text(win, wrap="word", font=("Segoe UI", 9))
        txt.pack(fill="both", expand=True)
        for kod, baslik, _gen, _tip in SUTUNLAR:
            txt.insert("end", f"{baslik}: ", "lbl")
            txt.insert("end", f"{s.get(kod, '')}\n\n")
        txt.tag_configure("lbl", font=("Segoe UI", 9, "bold"), foreground="#1976D2")
        txt.config(state="disabled")

    # ───────────────────────── Reçete / Rapor detay pencereleri (Botanik EOS canlı) ─────────────────────────
    def _recete_detay_goster(self):
        """Aktif satırın reçetesinin detayını Botanik EOS'tan çekip yeni pencerede göster.

        Arama önceliği:
        1. sistem_recete_no (RxSgkIslemNo) — Medula'da reçete açarken kullanılan no
        2. rx_id (PK fallback)
        3. rec_no (RxEReceteNo fallback)
        """
        s = self._aktif_satir()
        if not s:
            return
        sistem_recete_no = (s.get("sistem_recete_no") or "").strip()
        rx_id = s.get("rx_id")
        rec_no = (s.get("rec_no") or "").strip()
        if not sistem_recete_no and not rx_id and not rec_no:
            messagebox.showwarning("Uyarı", "Bu satırda reçete kimliği yok.", parent=self.root)
            return
        if not self.db:
            messagebox.showerror("Hata", "Botanik EOS bağlantısı yok.", parent=self.root)
            return
        try:
            data = self._recete_detay_verileri_cek(
                sistem_recete_no=sistem_recete_no, rx_id=rx_id, rec_no=rec_no)
        except Exception as e:
            logger.exception("Reçete detay sorgu hatası")
            messagebox.showerror("Sorgu Hatası", f"Reçete detayı çekilemedi:\n{e}",
                                 parent=self.root)
            return
        if not data:
            diag = getattr(self, "_son_recete_diag", "") or "(diagnostic yok)"
            messagebox.showinfo(
                "Bulunamadı",
                f"Reçete Botanik EOS'ta bulunamadı.\n\n"
                f"Aktif satır:\n"
                f"  sistem_recete_no = {sistem_recete_no!r}\n"
                f"  rec_no = {rec_no!r}\n"
                f"  rx_id = {rx_id!r}\n\n"
                f"Denemeler:\n{diag}",
                parent=self.root)
            return
        self._recete_detay_pencere_olustur(sistem_recete_no or rec_no, data)

    def _rapor_detay_goster(self):
        """Aktif satırın rapor_ana_id'sine ait raporu Botanik EOS'tan çekip göster."""
        s = self._aktif_satir()
        if not s:
            return
        rapor_ana_id = s.get("rapor_ana_id") or 0
        if not rapor_ana_id:
            messagebox.showwarning(
                "Uyarı",
                "Bu satıra bağlı rapor yok (rapor_ana_id=0).\n"
                "Reçete raporsuz olabilir veya rapor eşleşmesi yapılamamış.",
                parent=self.root)
            return
        if not self.db:
            messagebox.showerror("Hata", "Botanik EOS bağlantısı yok.", parent=self.root)
            return
        try:
            data = self._rapor_detay_verileri_cek(rapor_ana_id, s.get("musteri_id"))
        except Exception as e:
            logger.exception("Rapor detay sorgu hatası")
            messagebox.showerror("Sorgu Hatası", f"Rapor detayı çekilemedi:\n{e}",
                                 parent=self.root)
            return
        if not data:
            messagebox.showinfo("Bulunamadı",
                                "Rapor Botanik EOS'ta bulunamadı.", parent=self.root)
            return
        self._rapor_detay_pencere_olustur(s, data)

    def _recete_detay_verileri_cek(self, sistem_recete_no="", rx_id=None, rec_no=""):
        """Tek reçete + ilaçları + teşhisleri + açıklamaları + medula yanıtlarını çek.

        Sıralı fallback: sistem_recete_no (RxSgkIslemNo) →
        rx_id (RxId, PK) → rec_no (RxEReceteNo).
        Birinde bulunmazsa sonrakine geçer.

        Geri dönüş: bulunduysa dict; bulunamadıysa
        ('NOT_FOUND', diagnostic_str) tuple'ı (caller diagnostic'i kullanıcıya göstersin).
        """
        # 1) Header — ReceteAna + LEFT JOIN'ler (INNER JOIN YOK, böylece ilaç
        #    yoksa veya silinmiş bile olsa header çekilebilir)
        header_kalibi = """
            SELECT
                ra.RxId, ra.RxEReceteNo, ra.RxSgkIslemNo,
                ra.RxIslemTarihi, ra.RxKayitTarihi, ra.RxReceteTarihi,
                ra.RxBransId, ra.RxKurumId, ra.RxHastaneId,
                ra.RxMusteriId, ra.RxDoktorId, ra.RxSilme,
                ra.RxReceteRenkId, ra.RxReceteAltTuruId, ra.RxProvizyonTipId,
                m.MusteriAdiSoyadi, m.MusteriTCKN, m.MusteriDogumTarihi,
                m.MusteriCinsiyet, m.MusteriKapsamId, m.MusteriEmeklilik,
                d.DoktorAdiSoyadi,
                h.HastaneAdi, h.HastaneKodu,
                k.KurumAdi
            FROM ReceteAna ra
            LEFT JOIN Musteri m ON m.MusteriId = ra.RxMusteriId
            LEFT JOIN Doktor d ON d.DoktorId = ra.RxDoktorId
            LEFT JOIN Hastane h ON h.HastaneId = ra.RxHastaneId
            LEFT JOIN Kurum k ON k.KurumId = ra.RxKurumId
            WHERE {where_klozu}
        """
        denemeler = []
        if sistem_recete_no:
            denemeler.append(("RxSgkIslemNo", "ra.RxSgkIslemNo = ?", sistem_recete_no))
        if rx_id:
            denemeler.append(("RxId", "ra.RxId = ?", rx_id))
        if rec_no:
            denemeler.append(("RxEReceteNo", "ra.RxEReceteNo = ?", rec_no))
        header_rows = []
        diag_satirlari = []
        for alan_adi, where_klozu, param in denemeler:
            sql = header_kalibi.format(where_klozu=where_klozu)
            try:
                header_rows = self.db.sorgu_calistir(sql, (param,))
                diag_satirlari.append(f"  {alan_adi}={param!r} → {len(header_rows)} satır")
            except Exception as e:
                diag_satirlari.append(f"  {alan_adi}={param!r} → HATA: {e}")
                logger.warning(f"Reçete detay sorgu hatası ({alan_adi}={param}): {e}")
                header_rows = []
            if header_rows:
                logger.info(f"Reçete detay {alan_adi}={param} ile bulundu")
                break
        if not header_rows:
            self._son_recete_diag = "\n".join(diag_satirlari)
            return None
        self._son_recete_diag = ""
        header = header_rows[0]
        rx_id_resolved = header["RxId"]
        doktor_id = header.get("RxDoktorId")

        # 2) İlaçlar — ayrı sorgu (silinmiş olanları da göster, render'da işaretle)
        ilaclar = []
        try:
            ilaclar = self.db.sorgu_calistir(
                """SELECT
                       ri.RIId, ri.RIUrunId, ri.RIRaporKodId, ri.RIRaporNo,
                       ri.RIAdet, ri.RIDoz, ri.RITekrar, ri.RIAralik, ri.RIPeriyotId,
                       ri.RIToplam, ri.RIFiyatFarki, ri.RISilme,
                       u.UrunAdi, u.UrunBarkodu,
                       atc.ATCKodu, atc.ATCTurkce
                   FROM ReceteIlaclari ri
                   LEFT JOIN Urun u ON u.UrunId = ri.RIUrunId
                   LEFT JOIN ATC atc ON atc.ATCId = u.UrunATCId
                   WHERE ri.RIRxId = ?
                   ORDER BY ri.RIId""",
                (rx_id_resolved,))
        except Exception as e:
            logger.warning(f"ReceteIlaclari sorgu hatası: {e}")
            ilaclar = []

        # 2) Teşhisler (ReceteICD + ICD)
        teshisler = []
        try:
            ricd = self.db.sorgu_calistir(
                """SELECT icd.ICDKodu, icd.ICDAciklamasi
                   FROM ReceteICD ricd
                   LEFT JOIN ICD icd ON icd.ICDId = ricd.ReceteICDICDId
                   WHERE ricd.ReceteICDRxId = ?
                     AND (ricd.ReceteICDSilme IS NULL OR ricd.ReceteICDSilme = 0)""",
                (rx_id_resolved,))
            for r in ricd:
                kod = (r.get("ICDKodu") or "").strip()
                ack = (r.get("ICDAciklamasi") or "").strip()
                if kod and ack:
                    teshisler.append(f"{kod} — {ack}")
                elif kod:
                    teshisler.append(kod)
        except Exception as e:
            logger.debug(f"ReceteICD okunamadı: {e}")

        # 3) ReceteTeshis (eski sistem)
        try:
            rt = self.db.sorgu_calistir(
                """SELECT t.TeshisAciklama
                   FROM ReceteTeshis rt
                   LEFT JOIN Teshis t ON t.TeshisId = rt.RTTeshisId
                   WHERE rt.RTRxId = ?""",
                (rx_id_resolved,))
            for r in rt:
                ack = (r.get("TeshisAciklama") or "").strip()
                if ack and "Seçiniz" not in ack and ack not in teshisler:
                    teshisler.append(ack)
        except Exception as e:
            logger.debug(f"ReceteTeshis okunamadı: {e}")

        # 4) E-Reçete açıklamaları
        aciklamalar = []
        try:
            ea = self.db.sorgu_calistir(
                """SELECT eat.EReceteAciklamaTuruAdi, ea.EReceteAciklamaAdi
                   FROM ERecete er
                   INNER JOIN EReceteAciklamalari era ON era.ERAEReceteId = er.EReceteId
                   LEFT JOIN EReceteAciklama ea
                       ON ea.EReceteAciklamaId = era.ERAEReceteAciklamaId
                   LEFT JOIN EReceteAciklamaTuru eat
                       ON eat.EReceteAciklamaTuruId = era.ERAEReceteAciklamaTuruId
                   WHERE er.EReceteNo = ? AND (er.EReceteSilme IS NULL OR er.EReceteSilme = 0)""",
                (header.get("RxEReceteNo") or "",))
            for r in ea:
                tur = (r.get("EReceteAciklamaTuruAdi") or "").strip()
                ad = (r.get("EReceteAciklamaAdi") or "").strip()
                if not ad or ad in (".", ",", "-", "--"):
                    continue
                aciklamalar.append(f"[{tur}] {ad}" if tur and tur != "Seçiniz" else ad)
        except Exception as e:
            logger.debug(f"EReceteAciklamalari okunamadı: {e}")

        # 5) Medula yanıt mesajları (RxUyarilari)
        medula_yanitlari = []
        try:
            ru = self.db.sorgu_calistir(
                "SELECT RUAciklama FROM RxUyarilari WHERE RxId = ?", (rx_id_resolved,))
            for r in ru:
                txt = (r.get("RUAciklama") or "").strip()
                if txt:
                    medula_yanitlari.append(txt)
        except Exception as e:
            logger.debug(f"RxUyarilari okunamadı: {e}")

        # 6) Doktor branşı
        doktor_brans = ""
        if doktor_id:
            try:
                db_rows = self.db.sorgu_calistir(
                    """SELECT b.BransAdi FROM DoktorBrans db
                       INNER JOIN Brans b ON b.BransId = db.DoktorBransBransId
                       WHERE db.DoktorBransDoktorId = ?
                         AND (b.BransSilme IS NULL OR b.BransSilme = 0)""",
                    (doktor_id,))
                doktor_brans = ", ".join(
                    (r.get("BransAdi") or "").strip() for r in db_rows
                    if (r.get("BransAdi") or "").strip())
            except Exception as e:
                logger.debug(f"DoktorBrans okunamadı: {e}")

        return {
            "header": header,
            "ilaclar": ilaclar,
            "teshisler": teshisler,
            "aciklamalar": aciklamalar,
            "medula_yanitlari": medula_yanitlari,
            "doktor_brans": doktor_brans,
        }

    def _rapor_detay_verileri_cek(self, rapor_ana_id, musteri_id=None):
        """rapor_ana_id'ye ait raporu + ICD'leri + ek bilgileri + etkin maddeleri çek."""
        ra_rows = self.db.sorgu_calistir(
            """SELECT TOP 1 ra.*, rt.RaporTuruAdi, h.HastaneAdi, h.HastaneKodu
               FROM RaporAna ra
               LEFT JOIN RaporTuru rt ON rt.RaporTuruId = ra.RaporAnaRaporTuruId
               LEFT JOIN Hastane h ON h.HastaneId = ra.RaporAnaHastaneId
               WHERE ra.RaporAnaId = ?
                 AND (ra.RaporAnaSilme IS NULL OR ra.RaporAnaSilme = 0)""",
            (rapor_ana_id,))
        if not ra_rows:
            return None
        rapor = ra_rows[0]
        if not musteri_id:
            musteri_id = rapor.get("RaporAnaMusteriId")

        # Hasta bilgisi
        hasta_bilgi = {}
        if musteri_id:
            hb = self.db.sorgu_calistir(
                "SELECT MusteriAdiSoyadi, MusteriTCKN, MusteriDogumTarihi FROM Musteri WHERE MusteriId = ?",
                (musteri_id,))
            if hb:
                hasta_bilgi = hb[0]

        # ICD listesi (rapor kodu × ICD1..5)
        icdler = []
        try:
            rows_icd = self.db.sorgu_calistir(
                """SELECT icd1.ICDKodu AS K1, icd1.ICDAciklamasi AS A1,
                          icd2.ICDKodu AS K2, icd2.ICDAciklamasi AS A2,
                          icd3.ICDKodu AS K3, icd3.ICDAciklamasi AS A3,
                          icd4.ICDKodu AS K4, icd4.ICDAciklamasi AS A4,
                          icd5.ICDKodu AS K5, icd5.ICDAciklamasi AS A5,
                          rk.RaporKodu, rk.RaporKodAciklama,
                          rrki.RRKIBaslamaTarihi, rrki.RRKIBitisTarihi
                   FROM RaporRaporKodlariICD rrki
                   LEFT JOIN ICD icd1 ON icd1.ICDId = rrki.RRKIICDId
                   LEFT JOIN ICD icd2 ON icd2.ICDId = rrki.RRKIICDId2
                   LEFT JOIN ICD icd3 ON icd3.ICDId = rrki.RRKIICDId3
                   LEFT JOIN ICD icd4 ON icd4.ICDId = rrki.RRKIICDId4
                   LEFT JOIN ICD icd5 ON icd5.ICDId = rrki.RRKIICDId5
                   LEFT JOIN RaporKodlari rk ON rk.RaporKodId = rrki.RRKIRaporKodId
                   WHERE rrki.RRKIRaporAnaId = ?
                     AND (rrki.RRKISilme IS NULL OR rrki.RRKISilme = 0)""",
                (rapor_ana_id,))
            for r in rows_icd:
                grup = []
                for n in (1, 2, 3, 4, 5):
                    kod = (r.get(f"K{n}") or "").strip()
                    ack = (r.get(f"A{n}") or "").strip()
                    if kod and ack:
                        grup.append(f"{kod} — {ack}")
                    elif kod:
                        grup.append(kod)
                icdler.append({
                    "rapor_kodu": (r.get("RaporKodu") or "").strip(),
                    "rapor_kodu_aciklama": (r.get("RaporKodAciklama") or "").strip(),
                    "icd_listesi": grup,
                    "baslama": r.get("RRKIBaslamaTarihi"),
                    "bitis": r.get("RRKIBitisTarihi"),
                })
        except Exception as e:
            logger.debug(f"RaporRaporKodlariICD okunamadı: {e}")

        # Ek bilgiler
        ek_bilgiler = []
        try:
            rows_eb = self.db.sorgu_calistir(
                """SELECT REBTuru, REBDeger, REBAciklama
                   FROM RaporEkBilgi WHERE REBRaporAnaId = ?""",
                (rapor_ana_id,))
            for r in rows_eb:
                parts = []
                if r.get("REBTuru"):
                    parts.append(str(r["REBTuru"]))
                if r.get("REBDeger"):
                    parts.append(str(r["REBDeger"]))
                if r.get("REBAciklama"):
                    parts.append(str(r["REBAciklama"]))
                if parts:
                    ek_bilgiler.append(": ".join(parts))
        except Exception as e:
            logger.debug(f"RaporEkBilgi okunamadı: {e}")

        # Etkin maddeler (rapor tedavisi)
        etkin_maddeler = []
        try:
            rows_em = self.db.sorgu_calistir(
                """SELECT em.EtkinMaddeAdi, em.EtkinMaddeSGKKodu,
                          re.EtkinMaddeDoz, re.EtkinMaddeAdetMiktar,
                          re.EtkinMaddeTekrar, re.EtkinMaddeAralik
                   FROM RaporEtkinMadde re
                   LEFT JOIN EtkinMadde em ON em.EtkinMaddeId = re.EtkinMaddeId
                   WHERE re.EtkinMaddeRaporAnaId = ?
                     AND (re.EtkinMaddeSilme IS NULL OR re.EtkinMaddeSilme = 0)""",
                (rapor_ana_id,))
            for r in rows_em:
                etkin_maddeler.append({
                    "ad": (r.get("EtkinMaddeAdi") or "").strip(),
                    "sgk": (r.get("EtkinMaddeSGKKodu") or "").strip(),
                    "doz": r.get("EtkinMaddeDoz"),
                    "adet": r.get("EtkinMaddeAdetMiktar"),
                    "tekrar": r.get("EtkinMaddeTekrar"),
                    "aralik": r.get("EtkinMaddeAralik"),
                })
        except Exception as e:
            logger.debug(f"RaporEtkinMadde okunamadı: {e}")

        return {
            "rapor": rapor,
            "hasta": hasta_bilgi,
            "icdler": icdler,
            "ek_bilgiler": ek_bilgiler,
            "etkin_maddeler": etkin_maddeler,
        }

    def _recete_detay_pencere_olustur(self, rec_no, data):
        h = data["header"]
        win = tk.Toplevel(self.root)
        win.title(f"Reçete Detayı — {rec_no or h.get('RxEReceteNo') or h.get('RxId')}")
        win.geometry("900x650")
        win.configure(bg="#1E3A5F")
        win.transient(self.root)

        text_frame = tk.Frame(win, bg="#1E3A5F")
        text_frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(text_frame, wrap="word", bg="#FFFFFF", fg="#000000",
                       font=("Segoe UI", 10), padx=10, pady=8)
        sb = ttk.Scrollbar(text_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)

        txt.tag_configure("h1", font=("Segoe UI", 13, "bold"), foreground="#1E3A5F",
                          spacing3=6)
        txt.tag_configure("h2", font=("Segoe UI", 11, "bold"), foreground="#0D47A1",
                          spacing1=8, spacing3=4)
        txt.tag_configure("k", font=("Segoe UI", 10, "bold"), foreground="#37474F")
        txt.tag_configure("warn", foreground="#C62828")

        def yaz(s, tag=None):
            txt.insert("end", s, tag) if tag else txt.insert("end", s)
        def alan(et, dg):
            if dg is None or str(dg).strip() == "":
                return
            yaz(f"  {et}: ", "k"); yaz(f"{dg}\n")

        yaz(f"Reçete Detayı — {rec_no or h.get('RxEReceteNo') or '?'}\n", "h1")

        yaz("Reçete Bilgileri\n", "h2")
        alan("Reçete No (e-Reçete)", h.get("RxEReceteNo"))
        alan("SGK İşlem No", h.get("RxSgkIslemNo"))
        alan("Reçete Tarihi", h.get("RxReceteTarihi"))
        alan("Kayıt Tarihi", h.get("RxKayitTarihi"))
        alan("İşlem Tarihi", h.get("RxIslemTarihi"))
        alan("RxId (sistem)", h.get("RxId"))

        yaz("\nHasta\n", "h2")
        alan("Ad Soyad", h.get("MusteriAdiSoyadi"))
        alan("TCKN", h.get("MusteriTCKN"))
        alan("Doğum Tarihi", h.get("MusteriDogumTarihi"))
        alan("Cinsiyet", h.get("MusteriCinsiyet"))
        alan("Emeklilik", h.get("MusteriEmeklilik"))

        yaz("\nDoktor\n", "h2")
        alan("Ad Soyad", h.get("DoktorAdiSoyadi"))
        alan("Branş", data.get("doktor_brans") or "")

        yaz("\nTesis / Kurum\n", "h2")
        alan("Hastane", h.get("HastaneAdi"))
        alan("Hastane Kodu", h.get("HastaneKodu"))
        alan("Kurum", h.get("KurumAdi"))

        yaz("\nTeşhisler\n", "h2")
        if data["teshisler"]:
            for t in data["teshisler"]:
                yaz(f"  • {t}\n")
        else:
            yaz("  (Teşhis kaydı yok)\n")

        yaz("\nReçete Açıklamaları\n", "h2")
        if data["aciklamalar"]:
            for a in data["aciklamalar"]:
                yaz(f"  • {a}\n")
        else:
            yaz("  (Açıklama yok)\n")

        if data["medula_yanitlari"]:
            yaz("\nMedula Yanıt Mesajları\n", "h2")
            for y in data["medula_yanitlari"]:
                yaz(f"  • {y}\n", "warn")

        yaz("\nİlaçlar\n", "h2")
        if not data["ilaclar"]:
            yaz("  (ReceteIlaclari kaydı yok)\n")
        for i, ilac in enumerate(data["ilaclar"], 1):
            silinmis = bool(ilac.get("RISilme"))
            silme_etiketi = "  [SİLİNMİŞ]" if silinmis else ""
            yaz(f"\n  {i}. {ilac.get('UrunAdi') or '(ürün adı yok)'}{silme_etiketi}\n",
                "warn" if silinmis else "k")
            atc_kodu = (ilac.get("ATCKodu") or "").strip()
            atc_ad = (ilac.get("ATCTurkce") or "").strip()
            if atc_kodu or atc_ad:
                yaz(f"     ATC: {atc_kodu} {atc_ad}\n")
            yaz(f"     Adet: {ilac.get('RIAdet') or '-'}  |  "
                f"Doz: {ilac.get('RIDoz') or '-'}  |  "
                f"Tekrar: {ilac.get('RITekrar') or '-'}  |  "
                f"Aralık: {ilac.get('RIAralik') or '-'}\n")
            rip_no = (str(ilac.get("RIRaporNo") or "")).strip()
            rip_kod_id = ilac.get("RIRaporKodId") or 0
            if rip_no or rip_kod_id:
                yaz(f"     Rapor: kod_id={rip_kod_id}, takip_no={rip_no or '-'}\n")
            barkod = (ilac.get("UrunBarkodu") or "").strip()
            if barkod:
                yaz(f"     Barkod: {barkod}\n")

        txt.configure(state="disabled")

        alt = tk.Frame(win, bg="#1E3A5F")
        alt.pack(fill="x", padx=8, pady=(0, 8))
        tk.Button(alt, text="Kapat", font=("Segoe UI", 10, "bold"),
                  fg="white", bg="#455A64", activebackground="#37474F",
                  bd=0, padx=14, pady=4, command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda e: win.destroy())

    def _rapor_detay_pencere_olustur(self, satir, data):
        r = data["rapor"]
        h = data["hasta"]
        win = tk.Toplevel(self.root)
        win.title(f"Rapor Detayı — {r.get('RaporAnaRaporNo') or '?'}")
        win.geometry("900x650")
        win.configure(bg="#1E3A5F")
        win.transient(self.root)

        text_frame = tk.Frame(win, bg="#1E3A5F")
        text_frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(text_frame, wrap="word", bg="#FFFFFF", fg="#000000",
                       font=("Segoe UI", 10), padx=10, pady=8)
        sb = ttk.Scrollbar(text_frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)

        txt.tag_configure("h1", font=("Segoe UI", 13, "bold"), foreground="#1E3A5F",
                          spacing3=6)
        txt.tag_configure("h2", font=("Segoe UI", 11, "bold"), foreground="#0D47A1",
                          spacing1=8, spacing3=4)
        txt.tag_configure("k", font=("Segoe UI", 10, "bold"), foreground="#37474F")
        txt.tag_configure("muted", foreground="#6B7280")

        def yaz(s, tag=None):
            txt.insert("end", s, tag) if tag else txt.insert("end", s)
        def alan(et, dg):
            if dg is None or str(dg).strip() == "":
                return
            yaz(f"  {et}: ", "k"); yaz(f"{dg}\n")

        yaz(f"Rapor Detayı — {r.get('RaporAnaRaporNo') or '?'}\n", "h1")
        yaz(f"  (Reçete: {satir.get('rec_no') or '-'} | "
            f"Satır rapor kodu: {satir.get('rap_kod') or '-'} | "
            f"Eşleşme kaynağı: {satir.get('rapor_secim_kaynagi') or '-'})\n", "muted")

        yaz("\nRapor Bilgileri\n", "h2")
        alan("Rapor No", r.get("RaporAnaRaporNo"))
        alan("Takip No", r.get("RaporAnaRaporTakipNo"))
        alan("Rapor Tarihi", r.get("RaporAnaRaporTarihi"))
        alan("Tür", r.get("RaporTuruAdi"))
        alan("Açıklamalar", r.get("RaporAnaAciklamalar"))

        yaz("\nHasta\n", "h2")
        alan("Ad Soyad", h.get("MusteriAdiSoyadi"))
        alan("TCKN", h.get("MusteriTCKN"))
        alan("Doğum Tarihi", h.get("MusteriDogumTarihi"))

        yaz("\nTesis\n", "h2")
        alan("Hastane", r.get("HastaneAdi"))
        alan("Hastane Kodu", r.get("HastaneKodu"))

        yaz("\nRapor Kodları + ICD Teşhisleri\n", "h2")
        if data["icdler"]:
            for grp in data["icdler"]:
                kod = grp.get("rapor_kodu") or "-"
                kod_ack = grp.get("rapor_kodu_aciklama") or ""
                bas = grp.get("baslama") or ""
                bit = grp.get("bitis") or ""
                yaz(f"  • {kod} {kod_ack}\n", "k")
                if bas or bit:
                    yaz(f"      ({bas} → {bit})\n", "muted")
                for ic in grp.get("icd_listesi", []):
                    yaz(f"      ICD: {ic}\n")
        else:
            yaz("  (Rapor kodu/ICD kaydı yok)\n")

        yaz("\nRapor Etkin Maddeleri (tedavi)\n", "h2")
        if data["etkin_maddeler"]:
            for em in data["etkin_maddeler"]:
                ad = em.get("ad") or "-"
                sgk = em.get("sgk") or ""
                yaz(f"  • {ad}", "k")
                if sgk:
                    yaz(f"  [{sgk}]")
                doz_ozet = []
                if em.get("doz") is not None:
                    doz_ozet.append(f"Doz:{em['doz']}")
                if em.get("adet") is not None:
                    doz_ozet.append(f"Adet:{em['adet']}")
                if em.get("tekrar") is not None:
                    doz_ozet.append(f"Tekrar:{em['tekrar']}")
                if em.get("aralik") is not None:
                    doz_ozet.append(f"Aralık:{em['aralik']}")
                if doz_ozet:
                    yaz(f"  ({', '.join(doz_ozet)})")
                yaz("\n")
        else:
            yaz("  (Etkin madde tanımı yok)\n")

        if data["ek_bilgiler"]:
            yaz("\nEk Bilgiler\n", "h2")
            for eb in data["ek_bilgiler"]:
                yaz(f"  • {eb}\n")

        txt.configure(state="disabled")

        alt = tk.Frame(win, bg="#1E3A5F")
        alt.pack(fill="x", padx=8, pady=(0, 8))
        tk.Button(alt, text="Kapat", font=("Segoe UI", 10, "bold"),
                  fg="white", bg="#455A64", activebackground="#37474F",
                  bd=0, padx=14, pady=4, command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda e: win.destroy())

    def _sutuna_gore_sirala(self, kod: str):
        """Excel benzeri sütun başlığına tıklayarak sırala (asc/desc/reset toggle)."""
        # 3-aşamalı toggle: önce asc, sonra desc, sonra orijinal
        if self._siralama_kolonu == kod:
            if self._siralama_yonu == "asc":
                self._siralama_yonu = "desc"
            elif self._siralama_yonu == "desc":
                # Reset
                self._siralama_kolonu = None
                self._siralama_yonu = None
            else:
                self._siralama_yonu = "asc"
        else:
            self._siralama_kolonu = kod
            self._siralama_yonu = "asc"

        # Başlıklarda göstergeyi güncelle (▲ ▼ ↕)
        self._siralama_gostergesini_guncelle()

        # Sıralama uygula
        if self._siralama_kolonu is None:
            # Orijinal sıraya dön — DB'den yüklenen sıra zaten korundu mu?
            # Güvenlik: tum_satirlar yeni sıraya göre değişmiş olabilir.
            # Sadece tabloyu yenile (filtreleri uygulayarak)
            self._tabloyu_yenile()
            return

        # Sayısal vs string kararı (ilk dolu değere bak)
        ornekler = [s.get(kod, "") for s in self.tum_satirlar[:50] if s.get(kod)]
        sayisal_mi = bool(ornekler) and all(
            self._sayiya_cevrilebilir_mi(o) for o in ornekler[:10])

        rev = (self._siralama_yonu == "desc")
        if sayisal_mi:
            self.tum_satirlar.sort(
                key=lambda s: self._sayiya_cevir(s.get(kod, "")), reverse=rev)
        else:
            self.tum_satirlar.sort(
                key=lambda s: str(s.get(kod, "") or "").lower(), reverse=rev)
        self._tabloyu_yenile()

    @staticmethod
    def _sayiya_cevrilebilir_mi(v) -> bool:
        try:
            float(str(v).replace(",", "."))
            return True
        except Exception:
            return False

    @staticmethod
    def _sayiya_cevir(v) -> float:
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return -float("inf")

    def _siralama_gostergesini_guncelle(self):
        """Sütun başlıklarına ▲/▼ (sıralama) ve 🔽 (filtre) göstergeleri ekle."""
        try:
            for kod, baslik, _gen, _tip in SUTUNLAR:
                etiket = baslik
                # Sıralama göstergesi
                if kod == self._siralama_kolonu:
                    etiket += " ▲" if self._siralama_yonu == "asc" else " ▼"
                # Filtre göstergesi (Excel'deki gibi) — değer veya metin filtresi
                if kod in self.aktif_deger_filtre or kod in self.aktif_metin_filtre:
                    etiket += " 🔽"
                self.tv.heading(kod, text=etiket)
        except Exception:
            pass

    # ----------------------------------------------------------- SÜTUN AYAR
    def _sutun_ayarlarini_yukle(self):
        """Hangi sütunların görünür olduğunu JSON'dan yükle."""
        try:
            if os.path.exists(self.SUTUN_AYAR_DOSYASI):
                with open(self.SUTUN_AYAR_DOSYASI, "r", encoding="utf-8") as f:
                    kayit = json.load(f)
                for kod in SUTUN_KOD:
                    if kod in kayit:
                        self.sutun_gosterim[kod] = bool(kayit[kod])
        except Exception:
            pass

    def _sutun_ayarlarini_kaydet(self):
        try:
            with open(self.SUTUN_AYAR_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(self.sutun_gosterim, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # ------------------------------------------------------- SÜTUN ŞABLONLARI
    def _sablonlari_yukle(self) -> dict:
        """Şablon dosyasını oku. Format:
        {"sablonlar": {ad: {"gosterim": {kod: bool}, "genislik": {kod: int}}},
         "son_aktif": ad|null}
        """
        try:
            if os.path.exists(self.SUTUN_SABLON_DOSYASI):
                with open(self.SUTUN_SABLON_DOSYASI, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
                if "sablonlar" not in data:
                    data["sablonlar"] = {}
                return data
        except Exception as e:
            logger.warning("Şablon dosyası okunamadı: %s", e)
        return {"sablonlar": {}, "son_aktif": None}

    def _sablonlari_kaydet_dosyaya(self, data: dict) -> bool:
        try:
            with open(self.SUTUN_SABLON_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            logger.error("Şablon kaydedilemedi: %s", e)
            return False

    def _sablon_olarak_kaydet(self, ad: str, gecici_gosterim: dict = None) -> bool:
        """Anki sütun gösterim + tablo genişliklerini 'ad' adıyla kaydet.
        gecici_gosterim verilirse popup'taki henüz uygulanmamış checkbox
        durumu kullanılır; yoksa self.sutun_gosterim okunur."""
        ad = (ad or "").strip()
        if not ad:
            return False
        gosterim = {}
        if gecici_gosterim is not None:
            for kod in SUTUN_KOD:
                v = gecici_gosterim.get(kod)
                gosterim[kod] = bool(v.get()) if hasattr(v, "get") else bool(v)
        else:
            for kod in SUTUN_KOD:
                gosterim[kod] = bool(self.sutun_gosterim.get(kod, True))

        genislik = {}
        try:
            for kod in SUTUN_KOD:
                try:
                    genislik[kod] = int(self.tv.column(kod, "width"))
                except Exception:
                    pass
        except Exception:
            pass

        data = self._sablonlari_yukle()
        data.setdefault("sablonlar", {})[ad] = {
            "gosterim": gosterim,
            "genislik": genislik,
        }
        data["son_aktif"] = ad
        return self._sablonlari_kaydet_dosyaya(data)

    def _sablon_uygula(self, ad: str) -> bool:
        """Verilen şablonu sütun görünümüne ve tablo genişliklerine uygula."""
        ad = (ad or "").strip()
        if not ad:
            return False
        data = self._sablonlari_yukle()
        sablon = data.get("sablonlar", {}).get(ad)
        if not sablon:
            return False

        gosterim = sablon.get("gosterim") or {}
        for kod in SUTUN_KOD:
            if kod in gosterim:
                self.sutun_gosterim[kod] = bool(gosterim[kod])
        self._sutun_ayarlarini_kaydet()

        genislik = sablon.get("genislik") or {}
        try:
            for kod, w in genislik.items():
                if kod in SUTUN_KOD:
                    try:
                        self.tv.column(kod, width=int(w))
                    except Exception:
                        pass
        except Exception:
            pass

        try:
            self._sutun_gorunumunu_uygula()
        except Exception:
            pass
        try:
            self._filtre_slotlarini_hizala()
        except Exception:
            pass

        data["son_aktif"] = ad
        self._sablonlari_kaydet_dosyaya(data)
        return True

    def _sablon_sil(self, ad: str) -> bool:
        ad = (ad or "").strip()
        if not ad:
            return False
        data = self._sablonlari_yukle()
        if ad in data.get("sablonlar", {}):
            del data["sablonlar"][ad]
            if data.get("son_aktif") == ad:
                data["son_aktif"] = None
            return self._sablonlari_kaydet_dosyaya(data)
        return False

    def _aktif_sablonu_uygula(self):
        """Açılışta son aktif şablonu (varsa) uygula. Tablo zaten kurulmuş
        olmalı — _arayuz_olustur sonrası çağrılır."""
        try:
            data = self._sablonlari_yukle()
            ad = data.get("son_aktif")
            if ad and ad in data.get("sablonlar", {}):
                self._sablon_uygula(ad)
        except Exception as e:
            logger.debug("Aktif şablon yükleme: %s", e)

    def _sutun_ayar_penceresi(self):
        """Sütun göster/gizle popup penceresi."""
        win = tk.Toplevel(self.root)
        win.title("Sütun Ayarları")
        win.geometry("460x780")
        win.transient(self.root)

        # ─────────── ŞABLONLAR BÖLÜMÜ (üstte) ───────────
        sablon_frame = tk.LabelFrame(
            win, text=" 💾 Şablonlar ",
            font=("Segoe UI", 9, "bold"),
            fg="#0D47A1", bd=1, relief="groove"
        )
        sablon_frame.pack(fill="x", padx=8, pady=(8, 4))

        tk.Label(
            sablon_frame,
            text="Mevcut sutun gösterimi + tablo genişliklerini ad ile kaydet, "
                 "sonra istediğin zaman tek tıkla uygula.",
            font=("Segoe UI", 8), fg="#455A64", wraplength=420, justify="left"
        ).pack(fill="x", padx=6, pady=(4, 2))

        sablon_secim_frame = tk.Frame(sablon_frame)
        sablon_secim_frame.pack(fill="x", padx=6, pady=(2, 2))

        tk.Label(sablon_secim_frame, text="Şablon:",
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))

        sablon_data = self._sablonlari_yukle()
        sablon_adlari = sorted(sablon_data.get("sablonlar", {}).keys())
        son_aktif = sablon_data.get("son_aktif") or ""

        sablon_var = tk.StringVar(value=son_aktif if son_aktif in sablon_adlari else "")
        sablon_cb = ttk.Combobox(
            sablon_secim_frame, textvariable=sablon_var,
            values=sablon_adlari, state="readonly", width=28
        )
        sablon_cb.pack(side="left", fill="x", expand=True, padx=2)

        def _sablonlari_yenile():
            data = self._sablonlari_yukle()
            adlar = sorted(data.get("sablonlar", {}).keys())
            sablon_cb["values"] = adlar
            if sablon_var.get() not in adlar:
                sablon_var.set(data.get("son_aktif") or "")

        sablon_btn_frame = tk.Frame(sablon_frame)
        sablon_btn_frame.pack(fill="x", padx=6, pady=(2, 6))

        def _sablon_kaydet_tikla():
            from tkinter import simpledialog, messagebox
            mevcut = sablon_var.get().strip()
            ad = simpledialog.askstring(
                "Şablon Kaydet",
                "Şablon adı:\n(Aynı adla mevcutsa üzerine yazılır)",
                initialvalue=mevcut, parent=win
            )
            if not ad:
                return
            ad = ad.strip()
            if not ad:
                return
            mevcut_adlar = self._sablonlari_yukle().get("sablonlar", {})
            if ad in mevcut_adlar:
                if not messagebox.askyesno(
                    "Üzerine Yaz",
                    f"'{ad}' adlı şablon zaten var.\nÜzerine yazılsın mı?",
                    parent=win):
                    return
            ok = self._sablon_olarak_kaydet(ad, gecici_gosterim=gecici)
            if ok:
                _sablonlari_yenile()
                sablon_var.set(ad)
                messagebox.showinfo(
                    "Şablon Kaydedildi",
                    f"'{ad}' şablonu kaydedildi.\n"
                    f"({sum(1 for v in gecici.values() if v.get())} sütun görünür, "
                    f"genişlikler tablodan alındı.)",
                    parent=win)
            else:
                messagebox.showerror("Hata", "Şablon kaydedilemedi.", parent=win)

        def _sablon_uygula_tikla():
            from tkinter import messagebox
            ad = sablon_var.get().strip()
            if not ad:
                messagebox.showwarning("Şablon Seçilmedi",
                                        "Önce listeden bir şablon seç.",
                                        parent=win)
                return
            ok = self._sablon_uygula(ad)
            if ok:
                # Popup'taki checkbox'ları da senkronla
                for kod in SUTUN_KOD:
                    if kod in gecici:
                        gecici[kod].set(self.sutun_gosterim.get(kod, True))
                messagebox.showinfo("Şablon Uygulandı",
                                     f"'{ad}' şablonu uygulandı.",
                                     parent=win)
            else:
                messagebox.showerror("Hata",
                                      f"'{ad}' şablonu uygulanamadı.",
                                      parent=win)

        def _sablon_sil_tikla():
            from tkinter import messagebox
            ad = sablon_var.get().strip()
            if not ad:
                return
            if not messagebox.askyesno(
                "Şablonu Sil",
                f"'{ad}' şablonunu silmek istediğine emin misin?",
                parent=win):
                return
            if self._sablon_sil(ad):
                _sablonlari_yenile()
                sablon_var.set("")

        tk.Button(sablon_btn_frame, text="💾 Şablon Olarak Kaydet",
                  bg="#2E7D32", fg="white", activebackground="#1B5E20",
                  bd=0, padx=8, pady=3, font=("Segoe UI", 9, "bold"),
                  command=_sablon_kaydet_tikla
                  ).pack(side="left", padx=2)
        tk.Button(sablon_btn_frame, text="✓ Şablonu Uygula",
                  bg="#1565C0", fg="white", activebackground="#0D47A1",
                  bd=0, padx=8, pady=3, font=("Segoe UI", 9, "bold"),
                  command=_sablon_uygula_tikla
                  ).pack(side="left", padx=2)
        tk.Button(sablon_btn_frame, text="🗑 Sil",
                  bg="#C62828", fg="white", activebackground="#B71C1C",
                  bd=0, padx=8, pady=3, font=("Segoe UI", 9, "bold"),
                  command=_sablon_sil_tikla
                  ).pack(side="left", padx=2)

        # ─────────── SÜTUN GÖSTER/GİZLE BÖLÜMÜ ───────────
        tk.Label(win, text="Görünmesini istediğin sütunları işaretle:",
                 font=("Segoe UI", 10, "bold")).pack(pady=(6, 4), padx=8, anchor="w")

        # Scroll
        cont = tk.Frame(win)
        cont.pack(fill="both", expand=True, padx=8)
        canvas = tk.Canvas(cont, highlightthickness=0)
        sb = ttk.Scrollbar(cont, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        inner = tk.Frame(canvas)
        canvas.create_window((0, 0), window=inner, anchor="nw")

        gecici = {kod: tk.BooleanVar(value=self.sutun_gosterim.get(kod, True))
                   for kod in SUTUN_KOD}
        for kod, baslik, _gen, tip in SUTUNLAR:
            row = tk.Frame(inner)
            row.pack(fill="x", padx=6, pady=1)
            tk.Checkbutton(row, text=f"{baslik}", variable=gecici[kod],
                            anchor="w", font=("Segoe UI", 9)
                            ).pack(side="left", padx=4)
            tk.Label(row, text=f"({tip[:50]})", fg="#9E9E9E",
                     font=("Segoe UI", 8)).pack(side="left", padx=4)
        inner.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        # Hızlı işlemler
        ust = tk.Frame(win)
        ust.pack(fill="x", padx=8, pady=4)
        tk.Button(ust, text="Tümü Göster", bg="#C8E6C9",
                  command=lambda: [v.set(True) for v in gecici.values()]
                  ).pack(side="left", padx=2)
        tk.Button(ust, text="Tümü Gizle", bg="#FFCDD2",
                  command=lambda: [v.set(False) for v in gecici.values()]
                  ).pack(side="left", padx=2)
        tk.Button(ust, text="Varsayılan",
                  command=lambda: [v.set(True) for v in gecici.values()]
                  ).pack(side="left", padx=2)

        # Kaydet/iptal
        alt = tk.Frame(win)
        alt.pack(fill="x", padx=8, pady=8)

        def _uygula():
            for kod, var in gecici.items():
                self.sutun_gosterim[kod] = var.get()
            self._sutun_ayarlarini_kaydet()
            self._sutun_gorunumunu_uygula()
            win.destroy()

        tk.Button(alt, text="Uygula", bg="#1976D2", fg="white",
                  command=_uygula, padx=14
                  ).pack(side="right", padx=4)
        tk.Button(alt, text="İptal", command=win.destroy
                  ).pack(side="right", padx=4)

    def _sutun_gorunumunu_uygula(self):
        """sutun_gosterim'e göre Treeview'da kolon görünürlüğünü ve
        filtre satırındaki slot'ları senkron ayarla."""
        try:
            gorunenler = [k for k in SUTUN_KOD if self.sutun_gosterim.get(k, True)]
            self.tv["displaycolumns"] = gorunenler
        except Exception:
            pass
        # Filtre satırı slot'larını gizle/göster + güncel sütun genişliklerine
        # göre boyutlandır (sütun yeniden boyutlandırma sonrası hizalı kalsın)
        try:
            slotlar = getattr(self, "_filtre_slotlar", []) or []
            toplam_gen = 0
            for (kod, _baslik, gen, _tip), slot in zip(SUTUNLAR, slotlar):
                if self.sutun_gosterim.get(kod, True):
                    # Tablodaki anlık genişliği oku
                    try:
                        w = int(self.tv.column(kod, "width"))
                    except Exception:
                        w = gen
                    slot.config(width=w)
                    # İçerideki Entry'yi de güncelle
                    for child in slot.winfo_children():
                        if isinstance(child, tk.Entry):
                            child.place_configure(width=max(10, w - 2))
                    slot.pack(side="left")
                    toplam_gen += w
                else:
                    slot.pack_forget()
            if hasattr(self, "_filtre_canvas"):
                self._filtre_canvas.config(
                    scrollregion=(0, 0, toplam_gen,
                                   getattr(self, "_filtre_h", 28)))
        except Exception as e:
            logger.debug("filtre slot görünürlüğü: %s", e)

    def _filtre_slotlarini_hizala(self):
        """Tablonun anlık sütun genişliklerine göre filtre slotlarını
        yeniden boyutlandır. Sütun manuel resize edildiğinde çağrılır."""
        try:
            slotlar = getattr(self, "_filtre_slotlar", []) or []
            if not slotlar:
                return
            toplam = 0
            for (kod, _baslik, gen, _tip), slot in zip(SUTUNLAR, slotlar):
                if not self.sutun_gosterim.get(kod, True):
                    continue
                try:
                    w = int(self.tv.column(kod, "width"))
                except Exception:
                    w = gen
                slot.config(width=w)
                for child in slot.winfo_children():
                    if isinstance(child, tk.Entry):
                        child.place_configure(width=max(10, w - 2))
                toplam += w
            if hasattr(self, "_filtre_canvas"):
                self._filtre_canvas.config(
                    scrollregion=(0, 0, toplam,
                                   getattr(self, "_filtre_h", 28)))
        except Exception as e:
            logger.debug("filtre slot hizalama: %s", e)

    # ----------------------------------------------------------- STATE
    def _state_kaydet(self):
        """Tüm renk durumlarını dönem bazında JSON'a yaz."""
        if not self.aktif_donem:
            return
        try:
            tum = {}
            if os.path.exists(self.STATE_DOSYASI):
                with open(self.STATE_DOSYASI, "r", encoding="utf-8") as f:
                    tum = json.load(f)
            # Sadece beyaz olmayanları kaydet (yer tasarrufu)
            donem_state = {iid: r for iid, r in self.satir_renkleri.items()
                            if r != RENK_BEYAZ}
            if donem_state:
                tum[self.aktif_donem] = donem_state
            elif self.aktif_donem in tum:
                del tum[self.aktif_donem]
            with open(self.STATE_DOSYASI, "w", encoding="utf-8") as f:
                json.dump(tum, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning("State kaydet: %s", e)

    def _state_yukle_donem(self, donem: str) -> dict:
        try:
            if not os.path.exists(self.STATE_DOSYASI):
                return {}
            with open(self.STATE_DOSYASI, "r", encoding="utf-8") as f:
                tum = json.load(f)
            return tum.get(donem, {}) or {}
        except Exception:
            return {}

    # ----------------------------------------------------------- EXCEL EXPORT
    def _excel_export(self):
        if not self.tum_satirlar:
            messagebox.showwarning("Uyarı", "Aktarılacak veri yok")
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            messagebox.showerror("Hata", "openpyxl yüklü değil")
            return

        donem = self.aktif_donem or "tum"
        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.expanduser("~")
        path = filedialog.asksaveasfilename(
            initialdir=masa,
            initialfile=f"Aylik_Inceleme_{donem}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = donem

        # Başlıklar
        for c, (kod, baslik, _gen, _tip) in enumerate(SUTUNLAR, 1):
            cell = ws.cell(row=1, column=c, value=baslik)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="263238")
            cell.alignment = Alignment(horizontal="center")

        # Renk fill mapping
        excel_fill = {
            RENK_BEYAZ:   None,
            RENK_YESIL:   PatternFill("solid", fgColor="C8E6C9"),
            RENK_SARI:    PatternFill("solid", fgColor="FFF9C4"),
            RENK_TURUNCU: PatternFill("solid", fgColor="FFE0B2"),
            RENK_KIRMIZI: PatternFill("solid", fgColor="FFCDD2"),
        }

        for ri, s in enumerate(self.tum_satirlar, 2):
            iid = str(s["ri_id"])
            renk = self.satir_renkleri.get(iid, RENK_BEYAZ)
            fill = excel_fill.get(renk)
            for ci, kod in enumerate(SUTUN_KOD, 1):
                cell = ws.cell(row=ri, column=ci, value=str(s.get(kod, "")))
                if fill:
                    cell.fill = fill

        # Sütun genişlikleri
        for ci, (_kod, _baslik, gen, _tip) in enumerate(SUTUNLAR, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(8, gen / 7)

        try:
            wb.save(path)
            self._durum_yaz(f"Excel kaydedildi: {path}")
            if messagebox.askyesno("Tamamlandı",
                                     f"Excel kaydedildi:\n{path}\n\nAçılsın mı?"):
                os.startfile(path)
        except Exception as e:
            messagebox.showerror("Hata", f"Excel kaydedilemedi: {e}")

    # ───────────────────────────────────────────────────────────────────
    # STATİN / LİPİD SUT KONTROLÜ
    # ───────────────────────────────────────────────────────────────────
    # Lipid sınıfı tespiti için ATC C10* öncelikli; ATC dolu olmayan eski
    # kayıtlar için ticari isim fallback'leri.
    _LIPID_AD_FALLBACK = (
        "STATIN", "ATORVAS", "ROSUVAS", "SIMVAS", "PRAVAS", "FLUVAS",
        "PITAVAS", "EZETIMIB", "EZETROL", "INEGY",
        "LIPITOR", "CRESTOR", "ROZACT", "ROSUVA", "ATOR", "LIPVAS",
        "ULTROX", "ZOCOR", "ALVASTIN", "KOLESTER", "PRAVATOR",
        "FENOFIB", "GEMFIB", "BEZAFIB", "SIPROFIB",
        "LIPANTHYL", "LIPANTIL", "TRALIP", "LIPOFEN", "LOPID",
        "OMACOR", "REPATHA", "PRALUENT", "EVOLOKUMAB", "ALIROKUMAB",
    )

    @staticmethod
    def _lipid_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın STATIN / FIBRAT / DIGER_LIPID / NONE olarak sınıflandırması.

        ATC önceliklidir (WHO ATC C10):
          - C10AA / C10BA / C10BX → STATIN (statin veya statin-kombin)
          - C10AB                 → FIBRAT
          - C10AC / C10AD / C10AX → DIGER_LIPID (ezetimib tek başına dahil)
        ATC boşsa ticari isim/etken ipuçları kullanılır.
        """
        a = (atc or "").upper().strip()
        if a.startswith("C10AA") or a.startswith("C10BA") or a.startswith("C10BX"):
            return "STATIN"
        if a.startswith("C10AB"):
            return "FIBRAT"
        if a.startswith("C10AC") or a.startswith("C10AD") or a.startswith("C10AX"):
            return "DIGER_LIPID"
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        statin_ipuc = ("STATIN", "ATORVAS", "ROSUVAS", "SIMVAS", "PRAVAS",
                       "FLUVAS", "PITAVAS", "LIPITOR", "CRESTOR", "ROZACT",
                       "EZETIMIB", "EZETROL", "INEGY")
        fibrat_ipuc = ("FENOFIB", "GEMFIB", "BEZAFIB", "SIPROFIB",
                       "LIPANTHYL", "LIPANTIL", "TRALIP", "LIPOFEN", "LOPID")
        if any(s in ad for s in statin_ipuc) or any(s in et for s in statin_ipuc):
            return "STATIN"
        if any(s in ad for s in fibrat_ipuc) or any(s in et for s in fibrat_ipuc):
            return "FIBRAT"
        return "NONE"

    @staticmethod
    def _ilac_sonuc_olustur(s: dict) -> dict:
        """Satır dict'inden kontrol_statin/kontrol_fibrat'ın beklediği
        ilac_sonuc dict'ini üret."""
        def _bol(metin):
            if not metin:
                return []
            # Bu modülün satır birleştiricisi " | " kullanıyor.
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        # Rapor teşhisleri de rapor metnine dahil — LDL/risk faktörü
        # ICD'leri burada olabilir (RaporRaporKodlariICD)
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
        }

    # ───────────────────────────────────────────────────────────────────
    # DİYABET (SUT 4.2.38) KATEGORİ TESPİTİ + ilac_sonuc üreticisi
    # ───────────────────────────────────────────────────────────────────
    @staticmethod
    def _diyabet_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın diyabet ilacı olup olmadığını ATC A10* önceliği ve
        ad/etken fallback'i ile sınıflandırır.

        Dönüş: "DIYABET" / "NONE"
        """
        a = (atc or "").upper().strip()
        if a.startswith("A10"):
            return "DIYABET"
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et
        diyabet_etken = (
            "METFORMIN", "GLIKLAZID", "GLICLAZID", "GLIMEPIRID",
            "GLIBENKLAMID", "GLIPIZID", "REPAGLINID", "NATEGLINID",
            "PIOGLITAZON", "AKARBOZ", "ACARBOSE",
            "SITAGLIPTIN", "VILDAGLIPTIN", "SAKSAGLIPTIN", "LINAGLIPTIN",
            "ALOGLIPTIN",
            "EMPAGLIFLOZIN", "DAPAGLIFLOZIN", "KANAGLIFLOZIN",
            "CANAGLIFLOZIN", "ERTUGLIFLOZIN",
            "LIRAGLUTID", "SEMAGLUTID", "DULAGLUTID", "EKSENATID",
            "EXENATID", "LIKSISENATID", "LIXISENATID",
            "INSULIN", "INSÜLIN", "GLARGIN", "DETEMIR", "DEGLUDEC",
            "ASPART", "LISPRO", "GLULIZIN",
        )
        diyabet_ticari = (
            "GLIFOR", "GLUKOFEN", "DIAFORMIN", "GLUCOPHAGE", "MATOFIN",
            "DIAMICRON", "BETANORM", "DIAMERID", "AMARYL", "GLIMAX",
            "MERIDIA", "GLIBEDAL", "DAONIL", "GLUKOMID",
            "MINIDIAB", "GLUCOTROL", "NOVONORM", "STARLIX",
            "ACTOS", "GLUSTIN", "PIONORM", "GLUCOBAY",
            "JANUVIA", "GALVUS", "ONGLYZA", "TRAJENTA", "NESINA",
            "JANUMET", "GALVUSMET", "KOMBOGLYZE", "JENTADUETO", "VIPDOMET",
            "JARDIANCE", "FORZIGA", "FORXIGA", "INVOKANA", "STEGLATRO",
            "SYNJARDY", "XIGDUO", "VOKANAMET", "SEGLUROMET",
            "GLYXAMBI", "QTERN", "STEGLUJAN",
            "VICTOZA", "OZEMPIC", "RYBELSUS", "TRULICITY", "BYETTA",
            "BYDUREON", "SAXENDA", "LYXUMIA", "WEGOVY",
            "LANTUS", "TOUJEO", "TRESIBA", "LEVEMIR", "BASAGLAR",
            "ABASAGLAR", "NOVORAPID", "HUMALOG", "APIDRA", "ACTRAPID",
            "HUMULIN", "NOVOMIX", "RYZODEG", "XULTOPHY", "SOLIQUA",
        )
        if any(k in arama for k in diyabet_etken):
            return "DIYABET"
        if any(k in arama for k in diyabet_ticari):
            return "DIYABET"
        return "NONE"

    @staticmethod
    def _diyabet_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """Diyabet ilacı için alt sınıf (rapor tablosu kategori sütunu).

        ATC A10* alt-kodlarına göre:
          A10A*  → INSULIN
          A10BA  → BIGUANID (metformin)
          A10BB  → SULFONILURE
          A10BD  → KOMBI_OAD (sabit kombi: DPP4+met, SGLT2+met, vb.)
          A10BG  → TZD (pioglitazon)
          A10BH  → DPP4
          A10BJ  → GLP1
          A10BK  → SGLT2
          A10BX  → DIGER (glinid, alfa-glukozidaz vb.)
        ATC boşsa ad/etken bazlı fallback uygulanır.
        """
        a = (atc or "").upper().strip()
        if a.startswith("A10A"):
            return "INSULIN"
        if a.startswith("A10BA"): return "BIGUANID"
        if a.startswith("A10BB"): return "SULFONILURE"
        if a.startswith("A10BD"): return "KOMBI_OAD"
        if a.startswith("A10BG"): return "TZD"
        if a.startswith("A10BH"): return "DPP4"
        if a.startswith("A10BJ"): return "GLP1"
        if a.startswith("A10BK"): return "SGLT2"
        if a.startswith("A10BX"): return "DIGER"
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et
        if any(k in arama for k in ("DAPAGLIFLOZIN", "EMPAGLIFLOZIN",
                                     "KANAGLIFLOZIN", "CANAGLIFLOZIN",
                                     "ERTUGLIFLOZIN", "JARDIANCE", "FORZIGA",
                                     "FORXIGA", "INVOKANA", "STEGLATRO")):
            # SGLT-2 sabit kombiler
            if any(k in arama for k in ("SYNJARDY", "XIGDUO", "VOKANAMET",
                                          "SEGLUROMET", "GLYXAMBI", "QTERN",
                                          "STEGLUJAN")):
                return "KOMBI_OAD"
            return "SGLT2"
        if any(k in arama for k in ("SITAGLIPTIN", "VILDAGLIPTIN",
                                     "SAKSAGLIPTIN", "LINAGLIPTIN", "ALOGLIPTIN",
                                     "JANUVIA", "GALVUS", "ONGLYZA", "TRAJENTA",
                                     "NESINA")):
            if any(k in arama for k in ("JANUMET", "GALVUSMET", "KOMBOGLYZE",
                                          "JENTADUETO", "VIPDOMET")):
                return "KOMBI_OAD"
            return "DPP4"
        if any(k in arama for k in ("LIRAGLUTID", "SEMAGLUTID", "DULAGLUTID",
                                     "EKSENATID", "EXENATID", "LIKSISENATID",
                                     "LIXISENATID", "VICTOZA", "OZEMPIC",
                                     "RYBELSUS", "TRULICITY", "BYETTA",
                                     "BYDUREON", "SAXENDA", "WEGOVY", "LYXUMIA")):
            return "GLP1"
        if any(k in arama for k in ("METFORMIN", "GLIFOR", "GLUKOFEN",
                                     "DIAFORMIN", "GLUCOPHAGE", "MATOFIN")):
            return "BIGUANID"
        if any(k in arama for k in ("GLIKLAZID", "GLICLAZID", "GLIMEPIRID",
                                     "GLIBENKLAMID", "GLIPIZID", "DIAMICRON",
                                     "AMARYL", "GLIMAX", "DAONIL", "GLUKOMID",
                                     "MINIDIAB", "GLUCOTROL", "BETANORM",
                                     "DIAMERID", "GLIBEDAL")):
            return "SULFONILURE"
        if any(k in arama for k in ("REPAGLINID", "NATEGLINID", "NOVONORM",
                                     "STARLIX")):
            return "GLINID"
        if any(k in arama for k in ("PIOGLITAZON", "ACTOS", "GLUSTIN", "PIONORM")):
            return "TZD"
        if any(k in arama for k in ("AKARBOZ", "ACARBOSE", "GLUCOBAY")):
            return "AKARBOZ"
        if any(k in arama for k in ("INSULIN", "INSÜLIN", "GLARGIN", "DETEMIR",
                                     "DEGLUDEC", "ASPART", "LISPRO", "GLULIZIN",
                                     "LANTUS", "TOUJEO", "TRESIBA", "LEVEMIR",
                                     "BASAGLAR", "ABASAGLAR", "NOVORAPID",
                                     "HUMALOG", "APIDRA", "ACTRAPID", "HUMULIN",
                                     "NOVOMIX", "RYZODEG", "XULTOPHY", "SOLIQUA")):
            return "INSULIN"
        return "DIGER"

    @staticmethod
    def _ilac_sonuc_olustur_diyabet(s: dict, diger_ilac_adlari: list) -> dict:
        """Satır dict'inden kontrol_diyabet_dpp4_sglt2'nin beklediği
        ilac_sonuc dict'ini üret.

        diger_ilac_adlari: aynı reçetedeki DİĞER ilaçların adları
                            (kombinasyon yasağı kontrolü — DPP-4 + GLP-1 vb.)
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
            "hasta_yasi": s.get("yas") or "",
            "recete_dozu": s.get("rec_doz") or "",
            "recete_ilaclari": [{"ad": x} for x in (diger_ilac_adlari or []) if x],
        }

    # ───────────────────────────────────────────────────────────────────
    # ARB / SUT EK-4/F MADDE 51 — KATEGORİ + ilac_sonuc üreticisi
    # ───────────────────────────────────────────────────────────────────
    _ARB_MONO_ETKEN = (
        "IRBESARTAN", "İRBESARTAN",
        "KANDESARTAN", "CANDESARTAN",
        "LOSARTAN",
        "TELMISARTAN",
        "VALSARTAN",
        "OLMESARTAN",
        "EPROSARTAN",
        "RILMENIDEN", "RILMENIDIN", "RİLMENİDİN",
        "MOKSONIDIN", "MOXONIDIN", "MOKSONİDİN",
    )
    _ARB_TICARI_FALLBACK = (
        # Valsartan
        "DIOVAN", "VALEXA", "VALSACOR", "TARK", "VALSARTIL",
        # Telmisartan
        "MICARDIS", "PRITOR", "TELVAS", "TEVELOX", "TWENTACOR",
        # Olmesartan
        "OLMETEC", "BENICAR", "VOTUM",
        # Kandesartan
        "ATACAND", "BLOPRESS", "RATACAND",
        # Irbesartan
        "APROVEL", "KARVEA", "IRBES", "IRDA",
        # Losartan
        "COZAAR", "EKSEFOR", "LARITON", "LOSAR",
        # Eprosartan
        "TEVETEN",
        # Rilmeniden
        "HYPERIUM", "TENAXUM",
        # Moksonidin
        "CYNT", "PHYSIOTENS",
        # Kombinasyonlar (ARB + diüretik / + CCB / + 3'lü)
        "EXFORGE", "SEVIKAR", "TWYNSTA", "MICARDISPLUS", "MICARDIS PLUS",
        "CO-DIOVAN", "CODIOVAN", "CO-APROVEL", "COAPROVEL",
        "CO-OLMETEC", "HYZAAR", "KARVEZIDE", "FORZATEN",
        "CO-IRDA", "TRIVERAM", "ATACAND PLUS",
    )
    _ARB_KOMBI_TICARI = (
        "EXFORGE", "SEVIKAR", "TWYNSTA", "MICARDISPLUS", "MICARDIS PLUS",
        "CO-DIOVAN", "CODIOVAN", "CO-APROVEL", "COAPROVEL",
        "CO-OLMETEC", "HYZAAR", "KARVEZIDE", "FORZATEN",
        "CO-IRDA", "TRIVERAM", "ATACAND PLUS",
    )

    @staticmethod
    def _arb_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın ARB veya ARB-kombinasyonu olup olmadığını sınıflandırır
        (SUT EK-4/F Madde 51 kapsamı).

        ATC önceliği:
          - C09C   → ARB tek başına (mono)
          - C09DA  → ARB + tiazid (HCT) — SGK 17.10.2016 istisnası kapsamı
          - C09DB  → ARB + CCB
          - C09DX  → ARB + diğer / 3'lü (genelde CCB+HCT)
          - C09D*  → diğer ARB kombinasyonları
          - C02AC  → Santral etkili (Rilmeniden/Moksonidin)

        ATC yoksa: etken madde + ticari isim fallback.

        Dönüş: "ARB_MONO" / "ARB_KOMBI_HCT" / "ARB_KOMBI" / "NONE"
        """
        a = (atc or "").upper().strip()
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et

        # HCT içeriği tespiti (etken madde + ilaç adı)
        hct_var = (
            'HIDROKLOROTIAZID' in arama
            or 'HİDROKLOROTİAZİD' in arama
            or 'HIDROKLORTIAZID' in arama
            or 'HYDROCHLOROTHIAZID' in arama
            or 'HCTZ' in arama
            or ' HCT' in (' ' + arama + ' ')
            or '/HCT' in arama
        )
        # CCB içeriği — HCT ile birlikte ise 3'lü demektir, istisna geçersiz
        ccb_var = any(k in arama for k in (
            'AMLODIPIN', 'AMLODIPINE', 'LERKANIDIPIN', 'LERKANIDIPINE',
            'FELODIPIN', 'FELODIPINE', 'NIFEDIPIN', 'NIFEDIPINE',
            'NITRENDIPIN', 'BARNIDIPIN', 'NIKARDIPIN', 'ISRADIPIN',
        ))

        # ATC bazlı sınıflandırma
        if a.startswith("C09DA"):
            # ARB + tiazid (sadece HCT) — SGK 17.10.2016 istisnası
            return "ARB_KOMBI_HCT"
        if a.startswith("C09DB"):
            # ARB + CCB — istisna yok
            return "ARB_KOMBI"
        if a.startswith("C09DX"):
            # 3'lü / diğer kombi — HCT olsa bile CCB içeriyorsa istisna yok
            if hct_var and not ccb_var:
                return "ARB_KOMBI_HCT"
            return "ARB_KOMBI"
        if a.startswith("C09CA"):
            return "ARB_MONO"
        if a.startswith("C09D"):  # Diğer C09D alt grupları
            if hct_var and not ccb_var:
                return "ARB_KOMBI_HCT"
            return "ARB_KOMBI"
        if a.startswith("C09C"):
            return "ARB_MONO"
        if a.startswith("C02AC"):
            # Santral etkili antihipertansifler (Rilmeniden/Moksonidin) —
            # SUT'ta mono kabul ediliyor, kombi formu nadirdir.
            return "ARB_MONO"

        # ATC yok — etken madde + ticari isim fallback
        arb_var = (
            any(e in et for e in AylikReceteSorguGUI._ARB_MONO_ETKEN)
            or any(e in ad for e in AylikReceteSorguGUI._ARB_MONO_ETKEN)
            or any(t in ad for t in AylikReceteSorguGUI._ARB_TICARI_FALLBACK)
        )
        if not arb_var:
            return "NONE"

        # Kombi tespiti: etkin madde içinde "/" varsa veya ticari isim kombi
        is_kombi = ('/' in et) or any(
            t in ad for t in AylikReceteSorguGUI._ARB_KOMBI_TICARI)
        if not is_kombi:
            return "ARB_MONO"
        # HCT-only kombi mi? (CCB içermiyor → istisna)
        if hct_var and not ccb_var:
            return "ARB_KOMBI_HCT"
        return "ARB_KOMBI"

    @staticmethod
    def _ilac_sonuc_olustur_arb(s: dict) -> dict:
        """Satır dict'inden kontrol_arb_ek4f_m51'in beklediği ilac_sonuc
        dict'ini üret.

        Statin'den farklı olarak ek alanlar:
          - doktor_uzmanligi → branş listesi (sertifika dahil — "Aile Hek." aranır)
          - kurum_adi        → ASM tespiti (raporsuz reçete kurum bazlı yetki)
          - tesis_kodu       → Hastane.HastaneKodu (Medula tesis kodu — bilinen
                                aile hekimliği kodları AILE_HEKIMLIGI_TESIS_KODLARI
                                listesinden eşleşirse aile hekimi yetkisi)
          - kutu_sayisi      → ayda 1 kutu sınırı kontrolü
          - etkin_madde      → mono/kombi ayrımı
          - atc_kodu         → fallback için
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
        }

    # ───────────────────────────────────────────────────────────────────
    # ÇEŞİTLİ İLAÇLAR (SUT M.45 üriner / M.2 gözyaşı / BPH α-bloker)
    # KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    @staticmethod
    def _cesitli_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın ÇEŞİTLİ kapsamına girip girmediğini ve hangi alt gruba ait
        olduğunu sınıflandırır. Tüm liste eşlemesi sut_kontrolleri.py içinde
        (DRY) — buradan delege edilir.

        Dönüş: 'URINER' / 'GOZYASI' / 'BPH' / 'NONE'
        """
        try:
            from recete_kontrol.sut_kontrolleri import _cesitli_alt_grup_tespit
        except Exception:
            return "NONE"
        return _cesitli_alt_grup_tespit(ilac_adi or "", etkin or "", atc or "")

    @staticmethod
    def _ilac_sonuc_olustur_cesitli(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden kontrol_cesitli'nin beklediği ilac_sonuc dict'ini
        üret.

        ARB'den farklı olarak ek alanlar:
          - cinsiyet         → Duloksetin (kadın + stres SUI kontrolü)
          - diger_etken_maddeler / diger_ilac_adlari
                              → Antimuskarinik / α-bloker kombinasyon yasağı
          - mesaj_metni      → Medula uyarı mesajı (tum_metin'e dahil)
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        diger_etken = []
        diger_ilac = []
        if diger_satirlar:
            for d in diger_satirlar:
                if d is s:
                    continue
                et = (d.get("etkin") or "").strip()
                ad = (d.get("ilac") or "").strip()
                if et:
                    diger_etken.append(et)
                if ad:
                    diger_ilac.append(ad)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "cinsiyet": s.get("cins") or "",
            "diger_etken_maddeler": diger_etken,
            "diger_ilac_adlari": diger_ilac,
        }

    # ───────────────────────────────────────────────────────────────────
    # ANTİHT (SUT 4.2.12.B Mono + Kombi Antihipertansif) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    # ARB etken madde markerleri (ANTİHT'ten dışlanır — ARB butonu var)
    _ARB_ETKEN_MARKERLAR = ("ARTAN", "RILMENIDEN", "MOKSONIDIN")
    # Kombi etken madde göstergeleri ("RAMIPRIL/HCT" gibi)
    _ANTIHT_KOMBI_MARKERLAR = (
        "PRIL", "DIPIN", "OLOL", "TIAZID",
        "FUROSEMID", "SPIRONOLAKTON", "INDAPAMID",
        "HIDROKLOROTIAZID",
    )

    @staticmethod
    def _antiht_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın ANTİHT (mono/kombi antihipertansif) kapsamına girip
        girmediğini sınıflandırır.

        Önce ATC bazlı net karar; ardından etken madde fallback
        (sut_kategorisi_tespit_et kombi etken maddeyi MONO sayma hatasını
        önlemek için).

        ARB (C09C*/C09D*) ve ARB etken maddeleri (sartan/rilmeniden/
        moksonidin) → 'NONE' döner çünkü ARB için ayrı buton var
        (kontrol_arb_ek4f_m51, 1300/51 istisnası ile).

        Dönüş: 'MONO' / 'KOMBI' / 'NONE'
        """
        a = (atc or "").upper().strip()
        et = (etkin or "").upper()
        cls = AylikReceteSorguGUI

        # 1) ARB'yi tamamen dışla (kendi butonu var)
        if a.startswith("C09C") or a.startswith("C09D"):
            return "NONE"
        if any(m in et for m in cls._ARB_ETKEN_MARKERLAR):
            return "NONE"

        # 2) ATC bazlı net karar
        if a.startswith("C09B"):  # ACE-i + diüretik/KKB kombi
            return "KOMBI"
        if a.startswith("C09A"):  # ACE-i mono
            return "MONO"
        if a.startswith("C03") or a.startswith("C07") \
                or a.startswith("C08") or a.startswith("C02AC"):
            return "MONO"  # Diüretik / BB / KKB / santral

        # 3) Kombi etken madde göstergesi ("RAMIPRIL/HCT" gibi)
        if "/" in et or ("+" in et and len(et) > 3):
            import re as _re
            sub_etkenler = [s.strip() for s in _re.split(r"[/+]", et)
                            if s.strip()]
            if len(sub_etkenler) >= 2 and any(
                    any(m in p for m in cls._ANTIHT_KOMBI_MARKERLAR)
                    for p in sub_etkenler):
                return "KOMBI"

        # 4) Son çare: sut_kategorisi_tespit_et fallback
        try:
            from recete_kontrol.sut_kontrolleri import sut_kategorisi_tespit_et
        except Exception:
            return "NONE"
        kategori = sut_kategorisi_tespit_et({
            "ilac_adi": ilac_adi or "",
            "etkin_madde": etkin or "",
            "atc_kodu": atc or "",
        })
        if kategori == "MONO_ANTIHIPERTANSIF":
            return "MONO"
        if kategori == "KOMBINE_ANTIHIPERTANSIF":
            return "KOMBI"
        return "NONE"

    @staticmethod
    def _ilac_sonuc_olustur_antiht(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden ANTİHT kontrol fn'inin (mono/kombi) beklediği
        ilac_sonuc dict'ini üret. Aynı reçetedeki diğer satırlar
        diger_etken_maddeler/diger_ilac_adlari olarak iletilir
        (ACE+ARB kombi yasağı için)."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        diger_etken = []
        diger_ilac = []
        if diger_satirlar:
            for d in diger_satirlar:
                if d is s:
                    continue
                et = (d.get("etkin") or "").strip()
                ad = (d.get("ilac") or "").strip()
                if et:
                    diger_etken.append(et)
                if ad:
                    diger_ilac.append(ad)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "cinsiyet": s.get("cins") or "",
            "yas": s.get("yas") or "",
            "diger_etken_maddeler": diger_etken,
            "diger_ilac_adlari": diger_ilac,
        }

    # ───────────────────────────────────────────────────────────────────
    # İSKEMİK KALP (SUT 4.2.15.C İvabradin + 4.2.15.F Ranolazin) KATEGORİ
    # ───────────────────────────────────────────────────────────────────
    @staticmethod
    def _iskemik_kalp_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın İSKEMİK KALP (ivabradin/ranolazin/eplerenon) kapsamına
        girip girmediğini sınıflandırır.

        ATC öncelikli (C01EB17 / C01EB18 / C03DA04); ATC yoksa etken madde
        fallback. Her üçü de tek etken maddeli ilaçlar — kombi parsing
        gerekmez.

        Dönüş: 'IVABRADIN' / 'RANOLAZIN' / 'EPLERENON' / 'NONE'
        """
        a = (atc or "").upper().strip()
        et = (etkin or "").upper()

        if a.startswith("C01EB17") or "IVABRADIN" in et:
            return "IVABRADIN"
        if a.startswith("C01EB18") or "RANOLAZIN" in et:
            return "RANOLAZIN"
        if a.startswith("C03DA04") or "EPLERENON" in et:
            return "EPLERENON"

        # Fallback — sut_kategorisi_tespit_et
        try:
            from recete_kontrol.sut_kontrolleri import sut_kategorisi_tespit_et
        except Exception:
            return "NONE"
        kategori = sut_kategorisi_tespit_et({
            "ilac_adi": ilac_adi or "", "etkin_madde": etkin or "",
            "atc_kodu": atc or "",
        })
        if kategori in ("IVABRADIN", "RANOLAZIN", "EPLERENON"):
            return kategori
        return "NONE"

    @staticmethod
    def _ilac_sonuc_olustur_iskemik_kalp(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden ivabradin/ranolazin/eplerenon kontrol fn'inin
        beklediği ilac_sonuc dict'ini üret.

        Aynı reçetenin diğer satırları diger_etken_maddeler/
        diger_ilac_adlari olarak iletilir (BB intoleransı kontrolünde
        beraber yazılan BB tespiti için)."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        diger_etken = []
        diger_ilac = []
        if diger_satirlar:
            for d in diger_satirlar:
                if d is s:
                    continue
                et = (d.get("etkin") or "").strip()
                ad = (d.get("ilac") or "").strip()
                if et:
                    diger_etken.append(et)
                if ad:
                    diger_ilac.append(ad)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "cinsiyet": s.get("cins") or "",
            "yas": s.get("yas") or "",
            "diger_etken_maddeler": diger_etken,
            "diger_ilac_adlari": diger_ilac,
        }

    # ───────────────────────────────────────────────────────────────────
    # K-SİTRAT (Potasyum Sitrat — SUT EK-4/F) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    @staticmethod
    def _potasyum_sitrat_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Potasyum sitrat kapsam kontrolü.

        ATC A12BA02 veya etken madde 'POTASYUM SITRAT' içeriyorsa POTASYUM,
        aksi halde NONE.

        Dönüş: 'POTASYUM' / 'NONE'
        """
        a = (atc or "").upper().strip()
        et = (etkin or "").upper()
        if a.startswith("A12BA02") or "POTASYUM SITRAT" in et \
                or "POTASYUM SİTRAT" in et:
            return "POTASYUM"

        try:
            from recete_kontrol.sut_kontrolleri import sut_kategorisi_tespit_et
        except Exception:
            return "NONE"
        kategori = sut_kategorisi_tespit_et({
            "ilac_adi": ilac_adi or "", "etkin_madde": etkin or "",
            "atc_kodu": atc or "",
        })
        return "POTASYUM" if kategori == "POTASYUM_SITRAT" else "NONE"

    @staticmethod
    def _ilac_sonuc_olustur_potasyum_sitrat(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden kontrol_potasyum_sitrat'ın beklediği ilac_sonuc
        dict'ini üret."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "cinsiyet": s.get("cins") or "",
            "yas": s.get("yas") or "",
        }

    # ───────────────────────────────────────────────────────────────────
    # RAPORSUZ BİLGİ (Raporsuz verilebilen ilaçlar) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    # 3 kategori birleşir → kontrol_raporsuz_bilgilendirme dispatcher'ı
    # 28 ilaç sınıfı kontrolü yapar (NSAİD, parasetamol, makrolid, penisilin,
    # antihistaminik, dekonjestan, MAO-B, PPI, H2RA, antidiyareik, laksatif,
    # antispazmodik, kas gevşetici, topikal NSAİD, vitamin, vb.)
    _RAPORSUZ_BILGI_KATEGORILER = frozenset({
        "RAPORSUZ_BILGILENDIRME", "BENZODIAZEPIN", "GOZ_LUBRIKAN",
    })

    @staticmethod
    def _raporsuz_bilgi_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın raporsuz bilgilendirme kapsamına girip girmediğini
        sınıflandırır.

        sut_kategorisi_tespit_et() ile RAPORSUZ_BILGILENDIRME / BENZODIAZEPIN
        / GOZ_LUBRIKAN'dan birine düşerse 'RAPORSUZ' döner. Aksi halde NONE.

        Dönüş: 'RAPORSUZ' / 'NONE'
        """
        try:
            from recete_kontrol.sut_kontrolleri import sut_kategorisi_tespit_et
        except Exception:
            return "NONE"
        kategori = sut_kategorisi_tespit_et({
            "ilac_adi": ilac_adi or "", "etkin_madde": etkin or "",
            "atc_kodu": atc or "",
        })
        cls = AylikReceteSorguGUI
        if kategori in cls._RAPORSUZ_BILGI_KATEGORILER:
            return "RAPORSUZ"
        return "NONE"

    @staticmethod
    def _ilac_sonuc_olustur_raporsuz_bilgi(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden kontrol_raporsuz_bilgilendirme'nin beklediği
        ilac_sonuc dict'ini üret.

        Aynı reçetenin diğer satırları diger_etken_maddeler/diger_ilac_adlari
        olarak iletilir (gizli parasetamol kombinasyonu kontrolü için)."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        diger_etken = []
        diger_ilac = []
        if diger_satirlar:
            for d in diger_satirlar:
                if d is s:
                    continue
                et = (d.get("etkin") or "").strip()
                ad = (d.get("ilac") or "").strip()
                if et:
                    diger_etken.append(et)
                if ad:
                    diger_ilac.append(ad)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "cinsiyet": s.get("cins") or "",
            "yas": s.get("yas") or "",
            "diger_etken_maddeler": diger_etken,
            "diger_ilac_adlari": diger_ilac,
        }

    # ───────────────────────────────────────────────────────────────────
    # BASIT SUT (DMAH+Kadın+Trimetazidin+Kinolon mikro kontroller) KATEGORİ
    # ───────────────────────────────────────────────────────────────────
    # 4 farklı SUT kategorisi → 4 farklı kontrol fonksiyonu dispatch
    _BASIT_SUT_KATEGORILER = frozenset({
        "DMAH", "KADIN_HORMON", "TRIMETAZIDIN", "ANTIBIYOTIK_FLOROKINOLON",
    })

    @staticmethod
    def _basit_sut_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın BASIT SUT (DMAH/Kadın hormon/Trimetazidin/Florokinolon)
        kapsamına girip girmediğini sınıflandırır.

        Dönüş: 'DMAH' / 'KADIN_HORMON' / 'TRIMETAZIDIN' /
               'ANTIBIYOTIK_FLOROKINOLON' / 'NONE'
        """
        a = (atc or "").upper().strip()
        et = (etkin or "").upper()

        # ATC öncelikli net kararlar
        if a.startswith("B01AB"):  # DMAH ATC ailesi
            return "DMAH"
        if a.startswith("C01EB15") or "TRIMETAZIDIN" in et:
            return "TRIMETAZIDIN"
        if a.startswith("J01MA"):  # Florokinolon ailesi
            return "ANTIBIYOTIK_FLOROKINOLON"

        # Kadın hormon ATC ailesi (estradiol/estriol/progesteron + kombiler)
        if a.startswith("G03C") or a.startswith("G03D") \
                or a.startswith("G03F"):
            return "KADIN_HORMON"

        # DMAH etken madde fallback (mapping eksikse)
        if any(em in et for em in (
                "ENOKSAPARIN", "TINZAPARIN", "DALTEPARIN",
                "NADROPARIN", "BEMIPARIN", "PARNAPARIN")):
            return "DMAH"

        # Florokinolon etken madde fallback
        if any(em in et for em in (
                "SIPROFLOKSASIN", "CIPROFLOKSASIN",
                "LEVOFLOKSASIN", "MOKSIFLOKSASIN",
                "NORFLOKSASIN", "OFLOKSASIN", "GEMIFLOKSASIN")):
            return "ANTIBIYOTIK_FLOROKINOLON"

        # sut_kategorisi_tespit_et fallback (mapping varsa)
        try:
            from recete_kontrol.sut_kontrolleri import sut_kategorisi_tespit_et
        except Exception:
            return "NONE"
        kategori = sut_kategorisi_tespit_et({
            "ilac_adi": ilac_adi or "", "etkin_madde": etkin or "",
            "atc_kodu": atc or "",
        })
        cls = AylikReceteSorguGUI
        if kategori in cls._BASIT_SUT_KATEGORILER:
            return kategori
        return "NONE"

    @staticmethod
    def _ilac_sonuc_olustur_basit_sut(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden BASIT SUT kontrol fn'lerinin beklediği
        ilac_sonuc dict'ini üret."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "cinsiyet": s.get("cins") or "",
            "yas": s.get("yas") or "",
        }

    # ───────────────────────────────────────────────────────────────────
    # PSİKİYATRİ + ANTİEPİLEPTİK (SUT 4.2.2 + 4.2.25) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    # Antidepresanlar — ATC N06A*
    _PSI_ANTIDEPRESAN_ETKEN = (
        # SSRI
        "ESSITALOPRAM", "ESCITALOPRAM", "SERTRALIN", "FLUOKSETIN", "FLUOXETIN",
        "PAROKSETIN", "PAROXETIN", "FLUVOKSAMIN", "SITALOPRAM", "CITALOPRAM",
        # SNRI / NaSSA / SARI
        "DULOKSETIN", "VENLAFAKSIN", "VENLAFAXIN", "DESVENLAFAKSIN",
        "MILNASIPRAN", "MILNACIPRAN",
        "MIRTAZAPIN", "MIRTAZAPINE",
        "TRAZODON",
        # Trisiklik / Tetrasiklik
        "AMITRIPTILIN", "IMIPRAMIN", "KLOMIPRAMIN", "CLOMIPRAMIN",
        "NORTRIPTILIN", "DOKSEPIN", "DOXEPIN", "MAPROTILIN", "OPIPRAMOL",
        # Atipik / RIMA
        "BUPROPION", "BUPROPIYON", "VORTIOKSETIN", "VORTIOXETIN",
        "AGOMELATINE", "AGOMELATIN", "AGOMELATAIN",
        "TIANEPTIN", "MOKLOBEMID", "MOCLOBEMID",
    )
    _PSI_ANTIDEPRESAN_TICARI = (
        "CIPRALEX", "CIPRAM", "CITOLES", "ESCITOL", "SECITA", "SELECTRA",
        "LUSTRAL", "MISOL", "ZOLOFT",
        "FULSAC", "PROZAC",
        "SEROXAT", "PAXIL",
        "FAVERIN", "FEVARIN",
        "DUXET", "CYMBALTA",
        "EFEXOR", "EFFEXOR",
        "REMERON", "MIRTARON", "REDEPRES",
        "DESYREL", "TRITTICO",
        "LAROXYL", "TRYPTIZOL",
        "ANAFRANIL",
        "AURORIX",
        "WELLBUTRIN", "WELBOX",
        "BRINTELLIX",
        "VALDOXAN", "THYMANAX",
    )
    # Antipsikotikler — ATC N05A*
    _PSI_ANTIPSIKOTIK_ETKEN = (
        # Atipik
        "KLOZAPIN", "CLOZAPIN", "OLANZAPIN", "RISPERIDON", "AMISÜLPRID",
        "AMISULPIRID", "AMISULPRID", "KETIAPIN", "QUETIAPIN", "ZIPRASIDON",
        "ARIPIPRAZOL", "ZOTEPIN", "SERTINDOL", "BREKSPIPRAZOL",
        "BREXPIPRAZOL", "LURASIDON", "KARIPRAZIN", "CARIPRAZIN", "PALIPERIDON",
        # Tipik
        "HALOPERIDOL", "KLORPROMAZIN", "CHLORPROMAZIN", "LEVOMEPROMAZIN",
        "SÜLPIRID", "SULPIRID", "ZUKLOPENTIKSOL", "ZUCLOPENTIXOL",
        "FLUFENAZIN", "FLUPHENAZIN", "TRIFLUOPERAZIN", "PIMOZID",
    )
    _PSI_ANTIPSIKOTIK_TICARI = (
        "LEPONEX", "CLOZARIL",
        "ZYPREXA", "OLAXINN", "OLEANZ", "ZALASTA",
        "RISPERDAL", "RISPOLEPT",
        "SOLIAN",
        "SEROQUEL", "KETILEPT", "CEDRINA", "QUETIA",
        "ZELDOX", "GEODON",
        "ABILIFY", "ABIZOL",
        "ZOLEPTIL",
        "SERDOLECT",
        "REXULTI",
        "LATUDA",
        "REAGILA", "VRAYLAR",
        "INVEGA", "TREVICTA",
        "NORODOL", "HALDOL",
        "LARGACTIL",
        "NOZINAN",
        "DOGMATIL",
        "CLOPIXOL", "CISORDINOL",
        "MODITEN", "PROLIXIN",
        "STELAZINE",
        "ORAP",
    )
    # Mood stabilizatörler (psikiyatri akışı)
    _PSI_MOOD_STAB_ETKEN = (
        "LITYUM", "LITHIUM",
        # Lamotrijin / Valproat — ENDIKASYON BAZLI dispatch (bipolar→psi, epi→aep)
    )
    _PSI_MOOD_STAB_TICARI = ("LITHURIL",)

    # Benzodiazepinler — yeşil reçete kapsamı
    _PSI_BENZO_ETKEN = (
        "DIAZEPAM", "ALPRAZOLAM", "LORAZEPAM", "BROMAZEPAM",
        "KLORDIAZEPOKSIT", "CHLORDIAZEPOXIT", "KLORAZEPAT",
        "MIDAZOLAM", "KLONAZEPAM", "CLONAZEPAM", "ESTAZOLAM", "OKSAZEPAM",
    )
    _PSI_BENZO_TICARI = (
        "DIAZEM", "VALIUM", "NERVIUM",
        "XANAX", "ALPRAX",
        "ATIVAN",
        "BROXAN", "LEXOTAN",
        "RIVOTRIL",
        "DORMICUM",
    )
    # Antiepileptikler — ATC N03A*
    _AEP_YENI_NESIL_ETKEN = (
        "LAMOTRIJIN", "LAMOTRIGIN", "TOPIRAMAT", "VIGABATRIN",
        "LEVETIRASETAM", "LEVATIRASETAM", "LEVETIRACETAM",
    )
    _AEP_YENI_NESIL_TICARI = (
        "LAMICTAL", "LAMOTRIX",
        "TOPAMAX", "TOPAMAC",
        "SABRIL",
        "KEPPRA", "LEVEBON", "EPITERRA", "TIRATAM",
    )
    _AEP_ZONISAMIT_ETKEN = ("ZONISAMID", "ZONISAMIT")
    _AEP_ZONISAMIT_TICARI = ("ZONEGRAN",)
    _AEP_PREGABALIN_ETKEN = ("PREGABALIN",)
    _AEP_PREGABALIN_TICARI = ("LYRICA", "PREGABA", "ALYSE")
    _AEP_LAKOZAMID_ETKEN = ("LAKOSAMID", "LAKOZAMID", "LACOSAMID")
    _AEP_LAKOZAMID_TICARI = ("VIMPAT",)
    _AEP_GABAPENTIN_ETKEN = ("GABAPENTIN",)
    _AEP_GABAPENTIN_TICARI = ("NEURONTIN", "GABANTIN", "GABALEPT")
    _AEP_VALPROAT_ETKEN = ("VALPROAT", "VALPROIK")
    _AEP_VALPROAT_TICARI = ("DEPAKIN", "CONVULEX")

    @staticmethod
    def _psikiyatri_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın SUT 4.2.2 (Psikiyatri) veya 4.2.25 (Antiepileptik) kapsamına
        girip girmediğini sınıflandırır.

        ATC önceliği:
          - N06A*  → Antidepresan (4.2.2)
          - N05A*  → Antipsikotik (4.2.2)
          - N05B*  → Anksiyolitik (4.2.2 — benzodiazepin yeşil reçete)
          - N03A*  → Antiepileptik (4.2.25)
          - N05AN  → Lityum (mood stabilizatör — 4.2.2)
        ATC yoksa: etken madde + ticari isim fallback.

        Dönüş: "PSIKIYATRI" / "ANTIEPILEPTIK" / "NONE"
        """
        a = (atc or "").upper().strip()
        if a.startswith("N03A"):
            return "ANTIEPILEPTIK"
        if a.startswith("N06A") or a.startswith("N05A") or a.startswith("N05B"):
            return "PSIKIYATRI"

        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        cls = AylikReceteSorguGUI

        # Antiepileptik tespiti
        aep_etken_listeleri = (
            cls._AEP_YENI_NESIL_ETKEN, cls._AEP_ZONISAMIT_ETKEN,
            cls._AEP_PREGABALIN_ETKEN, cls._AEP_LAKOZAMID_ETKEN,
            cls._AEP_GABAPENTIN_ETKEN, cls._AEP_VALPROAT_ETKEN,
        )
        aep_ticari_listeleri = (
            cls._AEP_YENI_NESIL_TICARI, cls._AEP_ZONISAMIT_TICARI,
            cls._AEP_PREGABALIN_TICARI, cls._AEP_LAKOZAMID_TICARI,
            cls._AEP_GABAPENTIN_TICARI, cls._AEP_VALPROAT_TICARI,
        )
        for lst in aep_etken_listeleri:
            if any(e in et for e in lst):
                return "ANTIEPILEPTIK"
        for lst in aep_ticari_listeleri:
            if any(t in ad for t in lst):
                return "ANTIEPILEPTIK"

        # Psikiyatri tespiti
        psi_etken_listeleri = (
            cls._PSI_ANTIDEPRESAN_ETKEN, cls._PSI_ANTIPSIKOTIK_ETKEN,
            cls._PSI_MOOD_STAB_ETKEN, cls._PSI_BENZO_ETKEN,
        )
        psi_ticari_listeleri = (
            cls._PSI_ANTIDEPRESAN_TICARI, cls._PSI_ANTIPSIKOTIK_TICARI,
            cls._PSI_MOOD_STAB_TICARI, cls._PSI_BENZO_TICARI,
        )
        for lst in psi_etken_listeleri:
            if any(e in et for e in lst):
                return "PSIKIYATRI"
        for lst in psi_ticari_listeleri:
            if any(t in ad for t in lst):
                return "PSIKIYATRI"

        return "NONE"

    @staticmethod
    def _psikiyatri_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """Daha ayrıntılı alt sınıflandırma — Excel raporu ve dispatch için.

        Dönüşler:
          PSIKIYATRI grubu:
            ANTIDEPRESAN_SSRI / ANTIDEPRESAN_SNRI / ANTIDEPRESAN_TRISIKLIK
            ANTIDEPRESAN_ATIPIK / ANTIDEPRESAN_BVA (Bupropion/Vortioksetin/Agomelatin)
            ANTIPSIKOTIK_TIPIK / ANTIPSIKOTIK_ATIPIK
            BENZODIAZEPIN / MOOD_STAB_LITYUM
          ANTIEPILEPTIK grubu:
            AEP_YENI_NESIL / AEP_ZONISAMIT / AEP_PREGABALIN
            AEP_LAKOZAMID / AEP_GABAPENTIN / AEP_VALPROAT
          DIGER / NONE
        """
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        cls = AylikReceteSorguGUI

        # Antiepileptik alt sınıf
        if any(e in et for e in cls._AEP_PREGABALIN_ETKEN) or \
           any(t in ad for t in cls._AEP_PREGABALIN_TICARI):
            return "AEP_PREGABALIN"
        if any(e in et for e in cls._AEP_GABAPENTIN_ETKEN) or \
           any(t in ad for t in cls._AEP_GABAPENTIN_TICARI):
            return "AEP_GABAPENTIN"
        if any(e in et for e in cls._AEP_LAKOZAMID_ETKEN) or \
           any(t in ad for t in cls._AEP_LAKOZAMID_TICARI):
            return "AEP_LAKOZAMID"
        if any(e in et for e in cls._AEP_ZONISAMIT_ETKEN) or \
           any(t in ad for t in cls._AEP_ZONISAMIT_TICARI):
            return "AEP_ZONISAMIT"
        if any(e in et for e in cls._AEP_VALPROAT_ETKEN) or \
           any(t in ad for t in cls._AEP_VALPROAT_TICARI):
            return "AEP_VALPROAT"
        if any(e in et for e in cls._AEP_YENI_NESIL_ETKEN) or \
           any(t in ad for t in cls._AEP_YENI_NESIL_TICARI):
            return "AEP_YENI_NESIL"

        # Antipsikotik alt sınıf
        tipik_etken = ("HALOPERIDOL", "KLORPROMAZIN", "LEVOMEPROMAZIN",
                       "SÜLPIRID", "SULPIRID", "ZUKLOPENTIKSOL",
                       "FLUFENAZIN", "TRIFLUOPERAZIN", "PIMOZID")
        tipik_ticari = ("NORODOL", "HALDOL", "LARGACTIL", "NOZINAN",
                        "DOGMATIL", "CLOPIXOL", "MODITEN", "STELAZINE", "ORAP")
        if any(e in et for e in tipik_etken) or any(t in ad for t in tipik_ticari):
            return "ANTIPSIKOTIK_TIPIK"
        if any(e in et for e in cls._PSI_ANTIPSIKOTIK_ETKEN) or \
           any(t in ad for t in cls._PSI_ANTIPSIKOTIK_TICARI):
            return "ANTIPSIKOTIK_ATIPIK"

        # Antidepresan alt sınıf
        ssri = ("ESSITALOPRAM", "ESCITALOPRAM", "SERTRALIN", "FLUOKSETIN",
                "FLUOXETIN", "PAROKSETIN", "PAROXETIN", "FLUVOKSAMIN",
                "SITALOPRAM", "CITALOPRAM")
        snri = ("DULOKSETIN", "VENLAFAKSIN", "VENLAFAXIN", "DESVENLAFAKSIN",
                "MILNASIPRAN", "MILNACIPRAN", "MIRTAZAPIN")
        trisik = ("AMITRIPTILIN", "IMIPRAMIN", "KLOMIPRAMIN", "CLOMIPRAMIN",
                  "NORTRIPTILIN", "DOKSEPIN", "DOXEPIN", "MAPROTILIN", "OPIPRAMOL")
        bva = ("BUPROPION", "BUPROPIYON", "VORTIOKSETIN", "VORTIOXETIN",
               "AGOMELATINE", "AGOMELATIN", "AGOMELATAIN")
        if any(e in et for e in ssri):
            return "ANTIDEPRESAN_SSRI"
        if any(e in et for e in snri):
            return "ANTIDEPRESAN_SNRI"
        if any(e in et for e in trisik):
            return "ANTIDEPRESAN_TRISIKLIK"
        if any(e in et for e in bva):
            return "ANTIDEPRESAN_BVA"
        if any(e in et for e in cls._PSI_ANTIDEPRESAN_ETKEN) or \
           any(t in ad for t in cls._PSI_ANTIDEPRESAN_TICARI):
            return "ANTIDEPRESAN_ATIPIK"

        # Benzodiazepin
        if any(e in et for e in cls._PSI_BENZO_ETKEN) or \
           any(t in ad for t in cls._PSI_BENZO_TICARI):
            return "BENZODIAZEPIN"

        # Mood stabilizatör (Lityum)
        if any(e in et for e in cls._PSI_MOOD_STAB_ETKEN) or \
           any(t in ad for t in cls._PSI_MOOD_STAB_TICARI):
            return "MOOD_STAB_LITYUM"

        return "DIGER"

    @staticmethod
    def _ilac_sonuc_olustur_psikiyatri(s: dict, diger_satirlar: list = None) -> dict:
        """Satır dict'inden kontrol_psikiyatri / kontrol_antiepileptik_4_2_25
        fonksiyonlarının beklediği ilac_sonuc dict'ini üret.

        Ek alanlar (ARB'den farklı):
          - yas             → Lakozamid 16 yaş kontrolü
          - recete_turu     → Benzodiazepin yeşil reçete kontrolü
          - recete_alt_turu → Acil servis kontrolü (antipsikotik fıkra 5)
          - diger_etken_maddeler / diger_ilac_adlari → kombinasyon yasağı
            (Pregabalin + Gabapentin)
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        diger_etken = []
        diger_ilac = []
        if diger_satirlar:
            for d in diger_satirlar:
                if d is s:
                    continue
                et = (d.get("etkin") or "").strip()
                ad = (d.get("ilac") or "").strip()
                if et:
                    diger_etken.append(et)
                if ad:
                    diger_ilac.append(ad)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "yas": s.get("yas") or "",
            "recete_turu": s.get("rec_turu") or "Normal",
            "recete_alt_turu": s.get("rec_alt_turu") or "",
            "diger_etken_maddeler": diger_etken,
            "diger_ilac_adlari": diger_ilac,
        }

    # ───────────────────────────────────────────────────────────────────
    # NÖROPATİK / FİBROMİYALJİ İLAÇLARI (SUT 4.2.35.A + B)
    # ───────────────────────────────────────────────────────────────────
    # Plan v2'ye göre 5 alt grup:
    #   PREGABALIN, GABAPENTIN, DULOKSETIN, ALFA_LIPOIK, KAPSAISIN
    _NOR435_PREGABALIN_ETKEN = ("PREGABALIN",)
    _NOR435_PREGABALIN_TICARI = (
        "LYRICA", "GABRICA", "PREGALIN", "PREGABEX", "PREJUNTIN", "PREGABA",
    )
    _NOR435_GABAPENTIN_ETKEN = ("GABAPENTIN",)
    _NOR435_GABAPENTIN_TICARI = (
        "NEURONTIN", "NERUDA", "GABATEVA", "GABALEPT", "GABANTIN", "GABAGAMMA",
    )
    _NOR435_DULOKSETIN_ETKEN = ("DULOKSETIN", "DULOXETINE")
    _NOR435_DULOKSETIN_TICARI = ("CYMBALTA", "DUXET", "DULOXIN", "DULOX")
    _NOR435_ALFA_LIPOIK_ETKEN = (
        "TIOKTIK", "TİOKTİK", "ALFA LIPOIK", "ALFA-LIPOIK",
        "ALPHA LIPOIC", "ALFA LİPOİK",
    )
    _NOR435_ALFA_LIPOIK_TICARI = (
        "THIOCTACID", "NOREXIA", "LIPOIK", "TIOXIDAL",
    )
    _NOR435_KAPSAISIN_ETKEN = ("KAPSAISIN", "KAPSAİSİN", "CAPSAICIN")
    _NOR435_KAPSAISIN_TICARI = ("CAPSIN", "ZOSTRIX")

    @staticmethod
    def _noropatik_435_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın SUT 4.2.35 (nöropatik/fibromiyalji) kapsamına girip
        girmediğini sınıflandırır.

        ATC önceliği yok — bu kategoride etken madde + ticari isim ile
        tespit yapılır (Duloksetin N06AX21 olduğu için psikiyatri ATC'siyle
        çakışır; ayrımı endikasyon yapar).

        Dönüş: "NOROPATIK_435" / "NONE"
        """
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        cls = AylikReceteSorguGUI

        gruplar = (
            (cls._NOR435_PREGABALIN_ETKEN, cls._NOR435_PREGABALIN_TICARI),
            (cls._NOR435_GABAPENTIN_ETKEN, cls._NOR435_GABAPENTIN_TICARI),
            (cls._NOR435_DULOKSETIN_ETKEN, cls._NOR435_DULOKSETIN_TICARI),
            (cls._NOR435_ALFA_LIPOIK_ETKEN, cls._NOR435_ALFA_LIPOIK_TICARI),
            (cls._NOR435_KAPSAISIN_ETKEN, cls._NOR435_KAPSAISIN_TICARI),
        )
        for etk_lst, tic_lst in gruplar:
            if any(e in et for e in etk_lst) or any(t in ad for t in tic_lst):
                return "NOROPATIK_435"
        return "NONE"

    @staticmethod
    def _noropatik_435_alt_grup(ilac_adi: str, etkin: str) -> str:
        """Alt grup tespiti — Excel raporu ve dispatch için.

        Dönüşler:
          PREGABALIN / GABAPENTIN / DULOKSETIN / ALFA_LIPOIK / KAPSAISIN / DIGER
        """
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        cls = AylikReceteSorguGUI

        if any(e in et for e in cls._NOR435_PREGABALIN_ETKEN) or \
           any(t in ad for t in cls._NOR435_PREGABALIN_TICARI):
            return "PREGABALIN"
        if any(e in et for e in cls._NOR435_GABAPENTIN_ETKEN) or \
           any(t in ad for t in cls._NOR435_GABAPENTIN_TICARI):
            return "GABAPENTIN"
        if any(e in et for e in cls._NOR435_DULOKSETIN_ETKEN) or \
           any(t in ad for t in cls._NOR435_DULOKSETIN_TICARI):
            return "DULOKSETIN"
        if any(e in et for e in cls._NOR435_ALFA_LIPOIK_ETKEN) or \
           any(t in ad for t in cls._NOR435_ALFA_LIPOIK_TICARI):
            return "ALFA_LIPOIK"
        if any(e in et for e in cls._NOR435_KAPSAISIN_ETKEN) or \
           any(t in ad for t in cls._NOR435_KAPSAISIN_TICARI):
            return "KAPSAISIN"
        return "DIGER"

    @staticmethod
    def _ilac_sonuc_olustur_noropatik_435(s: dict,
                                            diger_satirlar: list = None) -> dict:
        """Satır dict'inden kontrol_noropatik_4_2_35'in beklediği ilac_sonuc
        dict'ini üret. Pregabalin+Gabapentin kombi yasağı için aynı reçetenin
        diğer satırlarındaki etken/ilaç adları diger_etken_maddeler /
        diger_ilac_adlari olarak iletilir.
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        diger_etken = []
        diger_ilac = []
        if diger_satirlar:
            for d in diger_satirlar:
                if d is s:
                    continue
                et = (d.get("etkin") or "").strip()
                ad = (d.get("ilac") or "").strip()
                if et:
                    diger_etken.append(et)
                if ad:
                    diger_ilac.append(ad)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": s.get("medula_msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "kurum_adi": s.get("kurum_adi") or "",
            "tesis_kodu": s.get("tesis_kodu") or "",
            "kutu_sayisi": s.get("kutu") or "",
            "yas": s.get("yas") or "",
            "diger_etken_maddeler": diger_etken,
            "diger_ilac_adlari": diger_ilac,
        }

    # ───────────────────────────────────────────────────────────────────
    # HEPATİT B/C İLAÇLARI (SUT 4.2.13.3 / 4.2.13.4) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    # HBV nükleos(t)id analogları (ATC J05AF) — kronik HBV tedavisi
    _HEPATIT_HBV_ETKEN = (
        "ENTEKAVIR", "TENOFOVIR", "TENOFOVIR DISOPROKSIL",
        "TENOFOVIR DISOPROKSIL FUMARAT", "TENOFOVIR ALAFENAMID",
        "LAMIVUDIN", "TELBIVUDIN", "ADEFOVIR", "ADEFOVIR DIPIVOKSIL",
    )
    _HEPATIT_HBV_TICARI = (
        "BARACLUDE", "ENTAVIR", "ENTECAVIR",
        "VIREAD", "TENOF", "VEMLIDY",
        "ZEFFIX", "EPIVIR HBV", "SEBIVO", "TYZEKA",
        "HEPSERA",
    )

    # HCV DAA (Direct-Acting Antivirals, ATC J05AP)
    _HEPATIT_HCV_ETKEN = (
        "SOFOSBUVIR", "LEDIPASVIR", "VELPATASVIR", "VOXILAPREVIR",
        "GLEKAPREVIR", "GLECAPREVIR", "PIBRENTASVIR",
        "OMBITASVIR", "PARITAPREVIR", "DASABUVIR",
        "DAKLATASVIR", "ELBASVIR", "GRAZOPREVIR",
    )
    _HEPATIT_HCV_TICARI = (
        "SOVALDI", "HARVONI", "EPCLUSA", "VOSEVI",
        "MAVYRET", "MAVIRET", "VIEKIRAX", "EXVIERA",
        "DAKLINZA", "ZEPATIER",
    )

    # Klasik tedavi (peginterferon + ribavirin) — HBV/HCV ortak
    _HEPATIT_KLASIK_ETKEN = (
        "PEGINTERFERON", "PEGINTERFERON ALFA",
        "PEGINTERFERON ALFA-2A", "PEGINTERFERON ALFA-2B",
        "INTERFERON ALFA", "RIBAVIRIN",
    )
    _HEPATIT_KLASIK_TICARI = (
        "PEGASYS", "PEGINTRON", "COPEGUS", "REBETOL", "VIRAZOLE",
    )

    @staticmethod
    def _hepatit_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın hepatit B/C ilacı olup olmadığını ATC J05AF/J05AP
        önceliği ve etken/ticari isim fallback'i ile sınıflandırır.

        ATC önceliği:
          - J05AF  → HBV nükleos(t)id analogları
          - J05AP  → HCV DAA (Direct-Acting Antivirals)
          - J05AB04 → Ribavirin (klasik HCV tedavisi yardımcısı)
          - L03AB10/11 → Peginterferon alfa-2a/2b

        HIV ilaçları (J05AR/J05AE/J05AG/J05AJ/J05AX) bu butonun
        KAPSAMI DIŞINDA — onlar ANTIVIRAL kategorisinde kalır.

        Dönüş: "HBV" / "HCV" / "KLASIK" / "NONE"
        """
        a = (atc or "").upper().strip()
        if a.startswith("J05AF"):
            return "HBV"
        if a.startswith("J05AP"):
            return "HCV"
        if a.startswith("J05AB04"):
            return "KLASIK"  # Ribavirin
        if a.startswith("L03AB10") or a.startswith("L03AB11"):
            return "KLASIK"  # Peginterferon alfa

        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et

        # HCV ilk önce (DAA → kombinasyon ad/etken örtüşebilir)
        if any(e in arama for e in AylikReceteSorguGUI._HEPATIT_HCV_ETKEN) or \
                any(t in ad for t in AylikReceteSorguGUI._HEPATIT_HCV_TICARI):
            return "HCV"
        if any(e in arama for e in AylikReceteSorguGUI._HEPATIT_HBV_ETKEN) or \
                any(t in ad for t in AylikReceteSorguGUI._HEPATIT_HBV_TICARI):
            return "HBV"
        if any(e in arama for e in AylikReceteSorguGUI._HEPATIT_KLASIK_ETKEN) or \
                any(t in ad for t in AylikReceteSorguGUI._HEPATIT_KLASIK_TICARI):
            return "KLASIK"
        return "NONE"

    @staticmethod
    def _hepatit_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """Hepatit alt-sınıfı (rapor için ek detay):
          HBV_NUKLEOSID  → entekavir, tenofovir, lamivudin, telbivudin, adefovir
          HCV_DAA        → DAA tek tek veya kombinasyon
          KLASIK_PEGIFN  → peginterferon (HBV veya HCV)
          KLASIK_RIBA    → ribavirin (HCV kombi)
          DIGER          → kategoride ama alt-sınıf belirsiz
        """
        kategori = AylikReceteSorguGUI._hepatit_kategori(ilac_adi, etkin, atc)
        if kategori == "HBV":
            return "HBV_NUKLEOSID"
        if kategori == "HCV":
            return "HCV_DAA"
        if kategori == "KLASIK":
            ad_et = (ilac_adi or "").upper() + " " + (etkin or "").upper()
            if "RIBAVIRIN" in ad_et or "COPEGUS" in ad_et or "REBETOL" in ad_et:
                return "KLASIK_RIBA"
            return "KLASIK_PEGIFN"
        return "DIGER"

    @staticmethod
    def _ilac_sonuc_olustur_hepatit(s: dict) -> dict:
        """Satır dict'inden kontrol_hepatit'in beklediği ilac_sonuc dict'ini
        üret. Mesaj/rapor metnine ek olarak doktor branşı da gönderilir
        (uzman branş kontrolü için)."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        # Branş bilgisini de rapor metnine kat — uzman ibaresi için
        brans = s.get("brans")
        if brans:
            rapor_aciklamalari.append(str(brans).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
        }

    # ───────────────────────────────────────────────────────────────────
    # KLOPİDOGREL / PRASUGREL / TIKAGRELOR (SUT 4.2.15) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    _KLOP_AD_FALLBACK = (
        "KLOPIDOGREL", "CLOPIDOGREL", "PRASUGREL", "TIKAGRELOR", "TICAGRELOR",
        "PLAVIX", "PLANOR", "KARUM", "AYRINEX", "KLOPIRA", "OPIROL",
        "EFFIENT", "EFIENT", "PRASIBLOCK",
        "BRILINTA", "BRILIQUE",
    )

    @staticmethod
    def _klopidogrel_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın klopidogrel/prasugrel/tikagrelor (P2Y12 inhibitörü) olup
        olmadığını ATC B01AC04/22/24 önceliği ve ad/etken fallback ile
        sınıflandırır.

        Dönüş: "KLOPIDOGREL" / "NONE"
        """
        a = (atc or "").upper().strip()
        # B01AC04 = klopidogrel, B01AC22 = prasugrel, B01AC24 = tikagrelor
        if a.startswith("B01AC04") or a.startswith("B01AC22") or a.startswith("B01AC24"):
            return "KLOPIDOGREL"
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et
        if any(k in arama for k in AylikReceteSorguGUI._KLOP_AD_FALLBACK):
            return "KLOPIDOGREL"
        return "NONE"

    @staticmethod
    def _klopidogrel_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """P2Y12 alt-sınıfı (rapor için): KLOPIDOGREL / PRASUGREL / TIKAGRELOR."""
        a = (atc or "").upper().strip()
        if a.startswith("B01AC04"):
            return "KLOPIDOGREL"
        if a.startswith("B01AC22"):
            return "PRASUGREL"
        if a.startswith("B01AC24"):
            return "TIKAGRELOR"
        ad_et = (ilac_adi or "").upper() + " " + (etkin or "").upper()
        if "PRASUGREL" in ad_et or "EFFIENT" in ad_et or "EFIENT" in ad_et \
                or "PRASIBLOCK" in ad_et:
            return "PRASUGREL"
        if "TIKAGRELOR" in ad_et or "TICAGRELOR" in ad_et \
                or "BRILINTA" in ad_et or "BRILIQUE" in ad_et:
            return "TIKAGRELOR"
        return "KLOPIDOGREL"

    @staticmethod
    def _ilac_sonuc_olustur_klopidogrel(s: dict, diger_ilac_adlari: list) -> dict:
        """Satır dict'inden kontrol_klopidogrel'in beklediği ilac_sonuc dict'ini
        üret. diger_ilac_adlari aynı reçetedeki diğer ilaçların adları
        (kombinasyon yasağı tespiti için: klopidogrel+prasugrel+tikagrelor+YOAK
        birlikte karşılanmaz)."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
            "hasta_yasi": s.get("yas") or "",
            "recete_dozu": s.get("rec_doz") or "",
            "recete_ilaclari": [{"ad": x} for x in (diger_ilac_adlari or []) if x],
        }

    # ───────────────────────────────────────────────────────────────────
    # YOAK / DOAK (SUT 4.2.15.D-1) KATEGORİ + ilac_sonuc üreticisi
    # Dabigatran / Rivaroksaban / Apiksaban / Edoksaban
    # ───────────────────────────────────────────────────────────────────
    _YOAK_AD_FALLBACK = (
        "DABIGATRAN", "DABIGATRAN ETEKSILAT",
        "RIVAROKSABAN", "RIVAROXABAN",
        "APIKSABAN", "APIXABAN",
        "EDOKSABAN", "EDOXABAN", "EDOKSABAN TOSILAT",
        "PRADAXA", "XARELTO", "ELIQUIS", "LIXIANA",
    )

    @staticmethod
    def _yoak_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın YOAK (Yeni Oral Antikoagülan) olup olmadığını ATC B01AE/B01AF
        önceliği ve ad/etken fallback ile sınıflandırır.

        ATC kodları:
          B01AE07 = dabigatran
          B01AF01 = rivaroksaban
          B01AF02 = apiksaban
          B01AF03 = edoksaban

        Dönüş: "YOAK" / "NONE"
        """
        a = (atc or "").upper().strip()
        if (a.startswith("B01AE07") or a.startswith("B01AF01")
                or a.startswith("B01AF02") or a.startswith("B01AF03")):
            return "YOAK"
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et
        if any(k in arama for k in AylikReceteSorguGUI._YOAK_AD_FALLBACK):
            return "YOAK"
        return "NONE"

    @staticmethod
    def _yoak_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """YOAK alt-etken: DABIGATRAN / RIVAROKSABAN / APIKSABAN / EDOKSABAN."""
        a = (atc or "").upper().strip()
        if a.startswith("B01AE07"): return "DABIGATRAN"
        if a.startswith("B01AF01"): return "RIVAROKSABAN"
        if a.startswith("B01AF02"): return "APIKSABAN"
        if a.startswith("B01AF03"): return "EDOKSABAN"
        ad_et = (ilac_adi or "").upper() + " " + (etkin or "").upper()
        if "DABIGATRAN" in ad_et or "PRADAXA" in ad_et:
            return "DABIGATRAN"
        if "RIVAROKSABAN" in ad_et or "RIVAROXABAN" in ad_et or "XARELTO" in ad_et:
            return "RIVAROKSABAN"
        if "APIKSABAN" in ad_et or "APIXABAN" in ad_et or "ELIQUIS" in ad_et:
            return "APIKSABAN"
        if "EDOKSABAN" in ad_et or "EDOXABAN" in ad_et or "LIXIANA" in ad_et:
            return "EDOKSABAN"
        return "YOAK"

    @staticmethod
    def _ilac_sonuc_olustur_yoak(s: dict, diger_ilac_adlari: list) -> dict:
        """Satır dict'inden kontrol_yoak'ın beklediği ilac_sonuc dict'ini üret.
        diger_ilac_adlari aynı reçetedeki diğer ilaçların adları
        (kombinasyon yasağı tespiti için: YOAK+YOAK ya da YOAK+P2Y12 birlikte
        karşılanmaz)."""
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
            "hasta_yasi": s.get("yas") or "",
            "recete_dozu": s.get("rec_doz") or "",
            "recete_ilaclari": [{"ad": x} for x in (diger_ilac_adlari or []) if x],
        }

    # ───────────────────────────────────────────────────────────────────
    # OSTEOPOROZ / KEMİK ERİMESİ (SUT 4.2.17 + 4.2.28.C) KATEGORİ + ilac_sonuc
    # ───────────────────────────────────────────────────────────────────
    # Bifosfonat (oral/IV) + biyolojik (denosumab/teriparatid/romosozumab/
    # stronsiyum) + SERM (raloksifen) — hepsi tek butonda denetlenir.
    _OSTEO_BIFOSFONAT_ETKEN = (
        "ALENDRONAT", "ALENDRONIK", "ALENDRONATE",
        "RISEDRONAT", "RISEDRONIK", "RISEDRONATE",
        "IBANDRONAT", "İBANDRONAT", "IBANDRONIK", "IBANDRONATE",
        "ZOLEDRONIK ASIT", "ZOLEDRONAT", "ZOLEDRONATE", "ZOLEDRONIC",
        "PAMIDRONAT", "PAMIDRONATE",
        # Bifosfonat + D vit kombileri (Fosavance vb.)
        "KOLEKALSIFEROL",
    )
    _OSTEO_BIFOSFONAT_TICARI = (
        "FOSAMAX", "FOSAVANCE", "FOSALAN", "ALENDRO",
        "ACTONEL", "OPTINATE", "OSTEONAT", "OSTEOFOS", "RISTABEN",
        "BONVIVA", "BONDRONAT", "IBADRON", "IBANDRO",
        "ZOMETA", "ACLASTA", "AKLASTA", "ZOLENAT", "ZOLEDRON",
        "AREDIA",
    )
    _OSTEO_BIYOLOJIK_ETKEN = (
        "DENOSUMAB",
        "TERIPARATID", "TERIPARATIDE",
        "ROMOSOZUMAB",
        "STRONSIYUM RANELAT", "STRONTIUM RANELATE",
    )
    _OSTEO_BIYOLOJIK_TICARI = (
        "PROLIA", "XGEVA",                     # denosumab
        "FORTEO", "FORSTEO", "MOVYMIA", "TERROSA",  # teriparatid
        "EVENITY",                             # romosozumab
        "OSSEOR", "PROTELOS",                  # stronsiyum ranelat
    )
    _OSTEO_SERM_ETKEN = (
        "RALOKSIFEN", "RALOXIFENE", "BAZEDOKSIFEN", "BAZEDOXIFENE",
    )
    _OSTEO_SERM_TICARI = (
        "EVISTA", "OPTRUMA", "CONBRIZA", "DUAVIVE",
    )

    # ───────────────────────────────────────────────────────────────────
    # ENTERAL BESLENME SOLÜSYONLARI — KATEGORİ + ad/etken fallback
    # ───────────────────────────────────────────────────────────────────
    # SUT: Enteral beslenme — endikasyon (malnütrisyon/disfaji/kanser/
    # kistik fibroz/IBD/inek sütü alerjisi/nörolojik/demans) + uzman raporu
    # (İç Hast/GE/Onkoloji/Geriatri/Pediatri/Nöroloji) + kalori planı.
    # Tespit altyapısı: recete_kontrol/sut_kontrolleri.py içindeki
    # TICARI_AD_KATEGORI sözlüğüyle aynı ürün ailelerini kullanır.
    _ENTERAL_ETKEN = (
        "KAZEIN", "MALTODEKSTRIN", "PEPTON",
        "SOYA PROTEIN IZOLATI", "AMINO ASIT KARISIMI",
        "BESIN AMACLI ENTERAL", "ENTERAL NUTRISYON",
    )
    _ENTERAL_TICARI = (
        # Resource ailesi
        "RESOURCE",
        # Nutridrink ailesi
        "NUTRIDRINK",
        # Nutren ailesi
        "NUTREN",
        # Fresubin ailesi
        "FRESUBIN",
        # Evolvia ailesi
        "EVOLVIA",
        # Ensure ailesi
        "ENSURE",
        # Peptamen ailesi
        "PEPTAMEN",
        # Nepro ailesi (renal)
        "NEPRO",
        # Impact ailesi
        "IMPACT ORAL", "IMPACT ENTERAL",
        # Tek ürünler
        "PROSURE", "MODULEN IBD", "GLUCERNA", "DIASIP", "CUBITAN",
        "ABOUND", "JUVEN",
        # Diğer yaygın enteral solüsyonlar
        "FORTIMEL", "NUTRISON", "PEDIASURE", "INFATRINI",
        "NOVASOURCE", "ISOSOURCE", "FRESUBIN INTENSIVE",
    )

    @staticmethod
    def _enteral_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın enteral beslenme solüsyonu olup olmadığını sınıflandırır.

        ATC önceliği:
          - V06D* (besin maddeleri — enteral nütrisyon)
          - V06DD (amino asit kombinasyonları)
          - V06DC (karbonhidrat içerenler)
          - B05BA (parenteral nütrisyon — bu butonun kapsamı dışında ama
                  enteral=True olarak işaretlenip raporda not düşülebilir.
                  Şu an dahil edilmiyor; B05BA parenteral olduğu için
                  enteral kontrol fonksiyonuyla eşleşmez)

        ATC yoksa: ticari isim/etken fallback.

        Dönüş: "ENTERAL" / "NONE"
        """
        a = (atc or "").upper().strip()
        if a.startswith("V06D"):
            return "ENTERAL"

        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et

        # Ticari isim — ürün ailesi adı reçete adında geçiyorsa ENTERAL
        if any(t in ad for t in AylikReceteSorguGUI._ENTERAL_TICARI):
            return "ENTERAL"
        # Etken madde fallback (kazein/maltodekstrin/pepton vb.)
        if any(k in arama for k in AylikReceteSorguGUI._ENTERAL_ETKEN):
            return "ENTERAL"
        return "NONE"

    @staticmethod
    def _enteral_alt_sinif(ilac_adi: str) -> str:
        """Enteral solüsyon alt-sınıfı (raporlama için).

        STANDART / DIYABETIK / RENAL / PEDIATRIK / IMMUNONUTRIYON / DIGER
        """
        ad = (ilac_adi or "").upper()
        if any(k in ad for k in ("DIABET", "DIASIP", "GLUCERNA",
                                   "FRESUBIN DB")):
            return "DIYABETIK"
        if any(k in ad for k in ("NEPRO", "RENAL")):
            return "RENAL"
        if any(k in ad for k in ("JUNIOR", "PEDIASURE", "INFATRINI",
                                   "MODULEN IBD")):
            return "PEDIATRIK"
        if any(k in ad for k in ("IMPACT", "PROSURE", "CUBITAN", "ABOUND",
                                   "JUVEN")):
            return "IMMUNONUTRIYON"
        return "STANDART"

    @staticmethod
    def _ilac_sonuc_olustur_enteral(s: dict) -> dict:
        """Satır dict'inden kontrol_enteral_beslenme'nin beklediği ilac_sonuc
        dict'ini üret.

        Kalori hesabı için recete_doz['gunluk_doz'] doldurulur (rec_doz_sayi'den).
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        # Günlük doz — _enteral_beslenme_detayli_kontrol kalori hesabı için
        # recete_doz dict'i içinde 'gunluk_doz' bekliyor.
        rec_doz_sayi = s.get("rec_doz_sayi")
        try:
            gunluk_doz = float(rec_doz_sayi) if rec_doz_sayi else None
        except (TypeError, ValueError):
            gunluk_doz = None

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
            "hasta_yasi": s.get("yas") or "",
            "recete_doz": {"gunluk_doz": gunluk_doz} if gunluk_doz else {},
        }

    # ───────────────────────────────────────────────────────────────────
    # ASTIM / KOAH — SOLUNUM SİSTEMİ İLAÇLARI (SUT 4.2.24 / 4.2.24.B)
    # ───────────────────────────────────────────────────────────────────
    # ATC R03* (solunum sistemi obstrüktif hastalık ilaçları) önceliği,
    # ardından etken madde + ticari isim fallback.
    # Alt sınıflar: SABA / SAMA / ICS / LABA / LAMA / LABA+ICS / LABA+LAMA
    # / UCLU (LABA+ICS+LAMA) / LTRA / OMALIZUMAB / ANTI_IL / ROFLUMILAST /
    # TEOFILIN / KROMOLIN / DIGER
    _SOLUNUM_ETKEN = (
        # SABA (kısa etkili β2 agonist)
        "SALBUTAMOL", "TERBUTALIN", "FENOTEROL",
        # SAMA (kısa etkili antikolinerjik)
        "IPRATROPIUM", "IPRATROPY",
        # LABA (uzun etkili β2 agonist)
        "FORMOTEROL", "SALMETEROL", "VILANTEROL",
        "INDAKATEROL", "INDACATEROL", "OLODATEROL",
        # LAMA (uzun etkili antikolinerjik)
        "TIOTROPIUM", "GLIKOPIRONYUM", "GLYCOPYRRONIUM",
        "GLICOPIRONYUM", "UMEKLIDINYUM", "UMECLIDINIUM",
        "AKLIDINYUM", "ACLIDINIUM", "REVEFENACIN",
        # ICS (inhale kortikosteroid)
        "BUDESONID", "BUDEZONID", "FLUTIKAZON", "FLUTICASONE",
        "BEKLOMETAZON", "BECLOMETHASONE", "SIKLESONID",
        "CICLESONIDE", "MOMETAZON", "MOMETASONE",
        # LTRA
        "MONTELUKAST", "ZAFIRLUKAST",
        # Anti-IgE / Anti-IL5 / IL4
        "OMALIZUMAB", "MEPOLIZUMAB", "BENRALIZUMAB", "DUPILUMAB",
        "RESLIZUMAB",
        # PDE4
        "ROFLUMILAST",
        # Metilksantin
        "TEOFILIN", "THEOPHYLLINE", "AMINOFILIN", "AMINOPHYLLINE",
        # Mast hücre stabilizatörleri
        "KROMOGLISIK", "CROMOGLICATE", "CROMOLYN", "NEDOKROMIL",
        "NEDOCROMIL",
    )
    _SOLUNUM_TICARI = (
        # SABA
        "VENTOLIN", "BUVENTOL", "BRICANYL", "AIROMIR", "ASMOL",
        # SAMA
        "ATROVENT",
        # SABA+SAMA kombi
        "COMBIVENT", "IPRAMOL", "DUOVENT", "BERODUAL",
        # ICS tek
        "PULMICORT", "FLIXOTIDE", "BECLOFORTE", "ALVESCO", "MIFLONIDE",
        "CORTAIR", "BUDICORT", "NEUMOCORT", "BUDECORT", "BECLATE",
        "CLENIL", "ASMANEX", "BECLOJET",
        # LABA tek
        "FORADIL", "OXIS", "SEREVENT", "ONBREZ", "STRIVERDI",
        # LABA+ICS
        "SERETIDE", "SYMBICORT", "FOSTER", "RELVAR", "DUORESP",
        "AIRFLUSAL", "BUFOMIX", "FOKUSAL", "VANNAIR", "WIXELA",
        "AIRDUO", "BREO", "FLUTIFORM", "BREQUAL", "BREQAL",
        "INUVAIR", "BUFEX", "BUFOM", "AIRPLUS",
        # LAMA tek
        "SPIRIVA", "INCRUSE", "SEEBRI", "BRETARIS", "EKLIRA", "TUDORZA",
        # LABA+LAMA
        "ANORO", "ULTIBRO", "SPIOLTO", "DUAKLIR", "BEVESPI", "STIOLTO",
        # Üçlü (LABA+ICS+LAMA)
        "TRELEGY", "TRIMBOW", "ENERZAIR", "BREZTRI", "TRIXEO",
        "AIRSUPRA", "TRIMUS", "TRELE",
        # LTRA
        "SINGULAIR", "ONCEAIR", "LUKASM", "NOTTA", "DESMONT",
        "AIRLUKAST", "ACCOLATE", "LEVMONT", "LEVOKAST", "MONKAST",
        "MONTELAIR", "MONLAS", "NOLEMON", "MUSTAIR",
        # Biyolojikler
        "XOLAIR", "NUCALA", "FASENRA", "DUPIXENT", "CINQAIR",
        # PDE4
        "DAXAS", "DALIRESP",
        # Teofilin
        "TEOBID", "TALOTREN", "AMINOCARDOL",
        # Kromolin
        "INTAL", "TILADE",
    )
    # Üçlü inhaler (tek başına LABA+ICS+LAMA içerenler)
    _SOLUNUM_UCLU_TICARI = (
        "TRELEGY", "TRIMBOW", "ENERZAIR", "BREZTRI", "TRIXEO",
        "AIRSUPRA", "TRIMUS",
    )
    # LABA+ICS ikili kombiler
    _SOLUNUM_LABA_ICS_TICARI = (
        "SERETIDE", "SYMBICORT", "FOSTER", "RELVAR", "DUORESP",
        "AIRFLUSAL", "BUFOMIX", "FOKUSAL", "VANNAIR", "WIXELA",
        "AIRDUO", "BREO", "FLUTIFORM", "BREQUAL", "BREQAL",
        "INUVAIR", "BUFEX", "BUFOM", "AIRPLUS",
    )
    # LABA+LAMA ikili kombiler
    _SOLUNUM_LABA_LAMA_TICARI = (
        "ANORO", "ULTIBRO", "SPIOLTO", "DUAKLIR", "BEVESPI", "STIOLTO",
    )
    _SOLUNUM_ICS_ETKEN = (
        "BUDESONID", "BUDEZONID", "FLUTIKAZON", "FLUTICASONE",
        "BEKLOMETAZON", "BECLOMETHASONE", "SIKLESONID", "CICLESONIDE",
        "MOMETAZON", "MOMETASONE",
    )
    _SOLUNUM_LABA_ETKEN = (
        "FORMOTEROL", "SALMETEROL", "VILANTEROL", "INDAKATEROL",
        "INDACATEROL", "OLODATEROL",
    )
    _SOLUNUM_LAMA_ETKEN = (
        "TIOTROPIUM", "GLIKOPIRONYUM", "GLYCOPYRRONIUM", "GLICOPIRONYUM",
        "UMEKLIDINYUM", "UMECLIDINIUM", "AKLIDINYUM", "ACLIDINIUM",
        "REVEFENACIN",
    )

    @staticmethod
    def _solunum_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın astım/KOAH grubuna giren bir solunum sistemi ilacı olup
        olmadığını sınıflandırır.

        ATC önceliği (WHO ATC R03 — solunum sistemi obstrüktif hastalıkları):
          - R03A* → β2 agonist (SABA / LABA / kombiler)
          - R03B* → diğer inhalerler (SAMA / LAMA / ICS / kromolin)
          - R03C* → sistemik β2 agonist (oral/parenteral)
          - R03D* → sistemik diğer (montelukast, omalizumab, roflumilast,
                    teofilin, mepolizumab, benralizumab, dupilumab dahil)

        ATC yoksa: etken madde + ticari isim fallback.

        Dönüş: "SOLUNUM" / "NONE"
        """
        a = (atc or "").upper().strip()
        if a.startswith("R03"):
            return "SOLUNUM"

        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et

        if any(t in ad for t in AylikReceteSorguGUI._SOLUNUM_TICARI):
            return "SOLUNUM"
        if any(k in arama for k in AylikReceteSorguGUI._SOLUNUM_ETKEN):
            return "SOLUNUM"
        return "NONE"

    @staticmethod
    def _solunum_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """Solunum ilacı alt-sınıfı (raporlama için).

        SABA / SAMA / SABA+SAMA / ICS / LABA / LAMA / LABA+ICS /
        LABA+LAMA / UCLU / LTRA / OMALIZUMAB / ANTI_IL / ROFLUMILAST /
        TEOFILIN / KROMOLIN / DIGER
        """
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et

        # ── Biyolojikler / sistemik özel ilaçlar ──
        if "OMALIZUMAB" in arama or "XOLAIR" in arama:
            return "OMALIZUMAB"
        if any(k in arama for k in ("MEPOLIZUMAB", "NUCALA",
                                     "BENRALIZUMAB", "FASENRA",
                                     "DUPILUMAB", "DUPIXENT",
                                     "RESLIZUMAB", "CINQAIR")):
            return "ANTI_IL"
        if "ROFLUMILAST" in arama or "DAXAS" in arama or "DALIRESP" in arama:
            return "ROFLUMILAST"
        if any(k in arama for k in ("TEOFILIN", "THEOPHYLLINE",
                                     "AMINOFILIN", "AMINOPHYLLINE",
                                     "TEOBID", "TALOTREN", "AMINOCARDOL")):
            return "TEOFILIN"
        if any(k in arama for k in ("KROMOGLISIK", "CROMOGLICATE",
                                     "CROMOLYN", "NEDOKROMIL", "NEDOCROMIL",
                                     "INTAL", "TILADE")):
            return "KROMOLIN"
        if "MONTELUKAST" in arama or "ZAFIRLUKAST" in arama or \
                any(t in ad for t in ("SINGULAIR", "ONCEAIR", "LUKASM",
                                       "NOTTA", "DESMONT", "AIRLUKAST",
                                       "ACCOLATE", "LEVMONT", "LEVOKAST",
                                       "MONKAST", "MONTELAIR", "MONLAS",
                                       "NOLEMON", "MUSTAIR")):
            return "LTRA"

        # ── İnhaler kombi tespiti ──
        # Üçlü inhaler tek ürün
        if any(t in ad for t in AylikReceteSorguGUI._SOLUNUM_UCLU_TICARI):
            return "UCLU"
        # LABA+ICS kombi
        if any(t in ad for t in AylikReceteSorguGUI._SOLUNUM_LABA_ICS_TICARI):
            return "LABA+ICS"
        # LABA+LAMA kombi
        if any(t in ad for t in AylikReceteSorguGUI._SOLUNUM_LABA_LAMA_TICARI):
            return "LABA+LAMA"

        # ── Etken madde bazlı (tek bileşen) ──
        has_ics = any(m in et for m in AylikReceteSorguGUI._SOLUNUM_ICS_ETKEN)
        has_laba = any(m in et for m in AylikReceteSorguGUI._SOLUNUM_LABA_ETKEN)
        has_lama = any(m in et for m in AylikReceteSorguGUI._SOLUNUM_LAMA_ETKEN)

        # Etken madde 3'lü içeriyorsa (nadir — kombi etken yazılışı)
        if has_laba and has_ics and has_lama:
            return "UCLU"
        if has_laba and has_ics:
            return "LABA+ICS"
        if has_laba and has_lama:
            return "LABA+LAMA"
        if has_lama:
            return "LAMA"
        if has_laba:
            return "LABA"
        if has_ics:
            return "ICS"

        # ── SABA / SAMA / SABA+SAMA ──
        saba_etken = any(m in et for m in ("SALBUTAMOL", "TERBUTALIN",
                                             "FENOTEROL"))
        sama_etken = "IPRATROPIUM" in et or "IPRATROPY" in et
        saba_ticari = any(t in ad for t in ("VENTOLIN", "BUVENTOL",
                                              "BRICANYL", "AIROMIR", "ASMOL"))
        sama_ticari = "ATROVENT" in ad
        saba_sama_kombi = any(t in ad for t in ("COMBIVENT", "IPRAMOL",
                                                  "DUOVENT", "BERODUAL")) or \
                          (saba_etken and sama_etken)
        if saba_sama_kombi:
            return "SABA+SAMA"
        if saba_etken or saba_ticari:
            return "SABA"
        if sama_etken or sama_ticari:
            return "SAMA"

        return "DIGER"

    @staticmethod
    def _ilac_sonuc_olustur_solunum(s: dict, diger_ilac_adlari: list) -> dict:
        """Satır dict'inden kontrol_solunum'un beklediği ilac_sonuc dict'ini
        üret.

        diger_ilac_adlari aynı reçetedeki diğer ilaçların adları —
        kontrol_solunum üçlü kullanım tespitinde (LABA+ICS+LAMA aynı
        reçetede 3 ayrı ilaç) bu listeyi okuyor.
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "msj_durumu": s.get("msj") or "",
            "doktor_uzmanligi": s.get("brans") or "",
            "hasta_yasi": s.get("yas") or "",
            "recete_dozu": s.get("rec_doz") or "",
            "recete_ilaclari": [{"ad": x} for x in (diger_ilac_adlari or []) if x],
        }

    @staticmethod
    def _osteo_kategori(ilac_adi: str, etkin: str, atc: str) -> str:
        """Satırın osteoporoz ilacı olup olmadığını sınıflandırır.

        ATC önceliği (WHO ATC):
          - M05BA / M05BB → BIFOSFONAT (oral/IV bifosfonat + D vit kombi)
          - M05BX         → BIYOLOJIK  (denosumab, romosozumab, stronsiyum)
          - H05AA02       → BIYOLOJIK  (teriparatid — paratiroid hormonu)
          - G03XC01/02    → SERM       (raloksifen, bazedoksifen)

        ATC yoksa: etken madde + ticari isim fallback.

        Dönüş: "BIFOSFONAT" / "BIYOLOJIK" / "SERM" / "NONE"
        """
        a = (atc or "").upper().strip()
        if a.startswith("M05BA") or a.startswith("M05BB"):
            return "BIFOSFONAT"
        if a.startswith("M05BX"):
            return "BIYOLOJIK"
        if a.startswith("H05AA"):
            # H05AA02 = teriparatid; H05AA01 = paratiroid hormonu
            return "BIYOLOJIK"
        if a.startswith("G03XC"):
            return "SERM"

        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et

        # Etken madde / ticari isim fallback
        if any(k in arama for k in AylikReceteSorguGUI._OSTEO_BIYOLOJIK_ETKEN) or \
                any(k in ad for k in AylikReceteSorguGUI._OSTEO_BIYOLOJIK_TICARI):
            return "BIYOLOJIK"
        if any(k in arama for k in AylikReceteSorguGUI._OSTEO_SERM_ETKEN) or \
                any(k in ad for k in AylikReceteSorguGUI._OSTEO_SERM_TICARI):
            return "SERM"
        # Bifosfonat fallback (KOLEKALSIFEROL tek başına D vitamini = osteoporoz
        # değil; sadece alendronat ile birlikte ise dahil edilmeli — ama burada
        # ATC ile zaten yakalandığı için fallback'te KOLEKALSIFEROL'ü dahil
        # ETMİYORUZ ki D vit ilaçları yanlışlıkla bifosfonat sayılmasın)
        bifos_etken_safe = tuple(
            x for x in AylikReceteSorguGUI._OSTEO_BIFOSFONAT_ETKEN
            if x != "KOLEKALSIFEROL"
        )
        if any(k in arama for k in bifos_etken_safe) or \
                any(k in ad for k in AylikReceteSorguGUI._OSTEO_BIFOSFONAT_TICARI):
            return "BIFOSFONAT"
        return "NONE"

    @staticmethod
    def _osteo_alt_sinif(ilac_adi: str, etkin: str, atc: str) -> str:
        """Osteoporoz alt-sınıfı (raporlama için).

        BIFOSFONAT_ORAL / BIFOSFONAT_IV / DENOSUMAB / TERIPARATID /
        ROMOSOZUMAB / STRONSIYUM / RALOKSIFEN / DIGER
        """
        a = (atc or "").upper().strip()
        ad = (ilac_adi or "").upper()
        et = (etkin or "").upper()
        arama = ad + " " + et
        if "DENOSUMAB" in arama or "PROLIA" in arama or "XGEVA" in arama:
            return "DENOSUMAB"
        if any(k in arama for k in ("TERIPARATID", "TERIPARATIDE",
                                     "FORTEO", "FORSTEO", "MOVYMIA")):
            return "TERIPARATID"
        if "ROMOSOZUMAB" in arama or "EVENITY" in arama:
            return "ROMOSOZUMAB"
        if "STRONSIYUM" in arama or "STRONTIUM" in arama or \
                "OSSEOR" in arama or "PROTELOS" in arama:
            return "STRONSIYUM"
        if "RALOKSIFEN" in arama or "RALOXIFENE" in arama or \
                "EVISTA" in arama or "OPTRUMA" in arama:
            return "RALOKSIFEN"
        # Bifosfonat oral vs IV
        # IV: zoledronik (Zometa/Aclasta), pamidronat (Aredia), ibandronat
        # IV form (Bondronat) — oral/iv ayrımı için ad/dozaj ipucu kullan
        if any(k in arama for k in ("ZOLEDRONIK", "ZOLEDRONAT", "ZOLEDRONIC",
                                     "ZOMETA", "ACLASTA", "AKLASTA",
                                     "PAMIDRONAT", "AREDIA")):
            return "BIFOSFONAT_IV"
        if "BONDRONAT" in arama:
            return "BIFOSFONAT_IV"
        if any(k in arama for k in ("ALENDRONAT", "RISEDRONAT", "IBANDRONAT",
                                     "İBANDRONAT", "FOSAMAX", "FOSAVANCE",
                                     "FOSALAN", "ACTONEL", "OPTINATE",
                                     "OSTEONAT", "BONVIVA")):
            return "BIFOSFONAT_ORAL"
        return "DIGER"

    @staticmethod
    def _ilac_sonuc_olustur_osteo(s: dict) -> dict:
        """Satır dict'inden kontrol_bifosfonat / kontrol_osteoporoz_biyolojik
        fonksiyonlarının beklediği ilac_sonuc dict'ini üret.

        Statin'den farklı olarak ek alanlar:
          - etkin_madde / atc_kodu  → biyolojik alt grup tespiti
          - doktor_uzmanligi        → uzman branş kontrolü (Endokrin/Romatoloji/FTR)
          - hasta_yasi              → 65 yaş eşik kontrolü (T-skor kuralı)
        """
        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        rapor_aciklamalari = []
        rap_ack = s.get("rap_ack")
        if rap_ack:
            rapor_aciklamalari.append(str(rap_ack).strip())
        for t in _bol(s.get("rap_tesh")):
            rapor_aciklamalari.append(t)

        return {
            "ilac_adi": s.get("ilac") or "",
            "etkin_madde": s.get("etkin") or "",
            "atc_kodu": s.get("atc") or "",
            "rapor_kodu": (s.get("rap_kod") or "").strip(),
            "rapor_kodu_aciklama": "",
            "recete_teshisleri": _bol(s.get("rec_tesh")),
            "rapor_aciklamalari": rapor_aciklamalari,
            "recete_aciklamalari": _bol(s.get("rec_ack")),
            "mesaj_metni": "",
            "doktor_uzmanligi": s.get("brans") or "",
            "hasta_yasi": s.get("yas") or "",
        }

    def _hasta_tum_icd_kodlarini_topla(self, musteri_idler: List[int],
                                         kontrol_tarihi=None) -> Dict[int, List[str]]:
        """Verilen hastaların TÜM aktif raporlarındaki ICD kodlarını + rapor
        kodlarını + ICD açıklamalarını toplu olarak çek.

        Returns: {musteri_id: ['E11.9 Tip 2 DM', 'I25 Koroner', ...]}

        Bu sayede statin denetimi sırasında, ilgili hastanın bir BAŞKA
        raporunda DM/KAH varsa o tanı statin satırına da yansıtılır → risk
        faktörü algoritmik olarak doğru değerlendirilir.

        kontrol_tarihi: rapor süresinin aktif sayılacağı referans tarih
        (None ise SQL Server GETDATE() kullanılır). Süresi dolmuş ama
        silinmemiş "hayalet" raporlar bu filtreyle elenir — aksi halde
        eski I50 (KY) / N18 (KBH) / E11 (DM) ICD'leri yıllar sonra hâlâ
        aktifmiş gibi sayılıyordu (HAMDULLAH AKSU 3JEQJ4C, ZELİHA
        ÇELİKTENYILDIZ 3JBO64A — 2026-05-07).
        """
        if not musteri_idler or not self.db:
            return {}
        result: Dict[int, List[str]] = {}
        try:
            for i in range(0, len(musteri_idler), 500):
                chunk = [m for m in musteri_idler[i:i + 500] if m]
                if not chunk:
                    continue
                ph = ",".join("?" * len(chunk))
                # Tarih filtresi: kontrol_tarihi verilmişse ?, yoksa GETDATE().
                if kontrol_tarihi is not None:
                    tarih_filtre = ("AND (rrki.RRKIBaslamaTarihi IS NULL "
                                    "     OR rrki.RRKIBaslamaTarihi <= ?) "
                                    "AND (rrki.RRKIBitisTarihi IS NULL "
                                    "     OR rrki.RRKIBitisTarihi >= ?) ")
                    params = tuple(chunk) + (kontrol_tarihi, kontrol_tarihi)
                else:
                    tarih_filtre = ("AND (rrki.RRKIBaslamaTarihi IS NULL "
                                    "     OR rrki.RRKIBaslamaTarihi <= GETDATE()) "
                                    "AND (rrki.RRKIBitisTarihi IS NULL "
                                    "     OR rrki.RRKIBitisTarihi >= GETDATE()) ")
                    params = tuple(chunk)
                rows = self.db.sorgu_calistir(
                    f"""SELECT
                            ra.RaporAnaMusteriId AS musteri_id,
                            i.ICDKodu             AS icd_kodu,
                            i.ICDAciklamasi       AS icd_aciklamasi,
                            rk.RaporKodu          AS rapor_kodu,
                            rk.RaporKodAciklama   AS rapor_kod_aciklama
                        FROM RaporRaporKodlariICD rrki
                        INNER JOIN RaporAna ra
                                ON ra.RaporAnaId = rrki.RRKIRaporAnaId
                        LEFT JOIN ICD i
                                ON i.ICDId = rrki.RRKIICDId
                        LEFT JOIN RaporKodlari rk
                                ON rk.RaporKodId = rrki.RRKIRaporKodId
                        WHERE ra.RaporAnaMusteriId IN ({ph})
                          AND (rrki.RRKISilme IS NULL OR rrki.RRKISilme = 0)
                          AND (ra.RaporAnaSilme IS NULL OR ra.RaporAnaSilme = 0)
                          {tarih_filtre}""",
                    params)
                for r in rows:
                    mid = r.get("musteri_id")
                    if not mid:
                        continue
                    parcalar = []
                    icd_k = (r.get("icd_kodu") or "").strip()
                    icd_a = (r.get("icd_aciklamasi") or "").strip()
                    if icd_k and icd_a:
                        parcalar.append(f"{icd_k} {icd_a}")
                    elif icd_k:
                        parcalar.append(icd_k)
                    rk_k = (r.get("rapor_kodu") or "").strip()
                    rk_a = (r.get("rapor_kod_aciklama") or "").strip()
                    if rk_k and rk_a:
                        parcalar.append(f"[Rap {rk_k}] {rk_a}")
                    for p in parcalar:
                        if p:
                            result.setdefault(mid, []).append(p)
        except Exception as e:
            logger.warning("Hasta ICD toplu sorgu fail: %s", e)
        # Tekrarsız + temiz
        return {k: list(dict.fromkeys(v)) for k, v in result.items()}

    def _arb_kontrol_baslat(self):
        """ARB KONTROL butonu — yüklenen satırlardan ARB veya ARB-kombinasyonu
        (ATC C09C* / C09D* / C02AC*) olanları SUT EK-4/F Madde 51 (1300/51)
        kuralına göre denetler ve sonucu en sağdaki SONUÇ sütununa yazar.

        Kapsam: İrbesartan, Kandesartan, Losartan, Telmisartan, Valsartan,
                Olmesartan, Eprosartan, Rilmeniden, Moksonidin + bunların
                diğer antihipertansiflerle kombinasyonları.

        Kurallar:
          - Mono ARB raporlu: doz/plan/süre belirtme zorunluluğu yok → UYGUN
          - Kombi ARB raporlu: "monoterapi yetersizliği" ibaresi şart
          - Raporsuz: aile hekimi + ayda 1 kutu sınırı
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "ARB Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import kontrol_arb_ek4f_m51
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_arb_disi": 0}
        kategori_sayac = {"ARB_MONO": 0, "ARB_KOMBI": 0, "ARB_KOMBI_HCT": 0}
        denetlenen_satirlar = []

        # Önceki çalıştırmadan kalan ARB verdict'lerini temizle
        for s in self.tum_satirlar:
            kategori = self._arb_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") in ("ARB_MONO", "ARB_KOMBI", "ARB_KOMBI_HCT"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_arb_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1
            ilac_sonuc = self._ilac_sonuc_olustur_arb(s)
            try:
                rapor = kontrol_arb_ek4f_m51(ilac_sonuc)
            except Exception as e:
                logger.warning("ARB kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"ARB SUT (EK-4/F M.51) kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"[Mono {kategori_sayac['ARB_MONO']} / "
            f"Kombi {kategori_sayac['ARB_KOMBI']} / "
            f"HCT-istisna {kategori_sayac['ARB_KOMBI_HCT']}]  "
            f"(ARB dışı {sayac['_arb_disi']} satır boş bırakıldı)"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam_arb = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                       + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam_arb == 0:
            messagebox.showinfo(
                "ARB Kontrol",
                "Bu dönemde ARB grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"ARB SUT EK-4/F M.51 kontrolü tamamlandı.\n\n"
            f"Toplam ARB satırı       : {toplam_arb}\n"
            f"  ✓ UYGUN              : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL        : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ            : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI            : {sayac['ATLANDI']}\n\n"
            f"Kategori dağılımı:\n"
            f"  Mono ARB             : {kategori_sayac['ARB_MONO']}\n"
            f"  Kombi (CCB/ACE/3'lü) : {kategori_sayac['ARB_KOMBI']}\n"
            f"  Kombi HCT (istisna)  : {kategori_sayac['ARB_KOMBI_HCT']}\n"
            f"      (SGK 17.10.2016 — diüretik kombileri kapsam dışı)\n\n"
            f"ARB dışı (atlanan)      : {sayac['_arb_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._arb_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("ARB rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── ÇEŞİTLİ İLAÇLAR KONTROL (SUT M.45 / M.2 / BPH α-bloker) ──────
    def _cesitli_kontrol_baslat(self):
        """ÇEŞİTLİ KONTROL butonu — yüklenen satırlardan üriner inkontinans
        (M.45) / suni gözyaşı (M.2) / BPH α-bloker kapsamına girenleri tek
        dispatcher (kontrol_cesitli) üzerinden denetler ve sonucu en sağdaki
        SONUÇ sütununa yazar.

        ARB kalıbına benzer ama 3 alt grup dispatcher altında.
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Çeşitli Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import kontrol_cesitli
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"URINER": 0, "GOZYASI": 0, "BPH": 0}
        denetlenen_satirlar = []

        # Aynı reçetenin diğer satırlarını grupla (kombi yasağı için)
        rec_grup = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no") or ""
            rec_grup.setdefault(rno, []).append(s)

        # Önceki çalıştırmadan kalan ÇEŞİTLİ verdict'lerini temizle
        for s in self.tum_satirlar:
            kategori = self._cesitli_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") in ("URINER", "GOZYASI", "BPH"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1

            rno = s.get("rec_no") or ""
            ayni_recete = rec_grup.get(rno, [])
            ilac_sonuc = self._ilac_sonuc_olustur_cesitli(s, ayni_recete)

            try:
                rapor = kontrol_cesitli(ilac_sonuc)
            except Exception as e:
                logger.warning("Çeşitli kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"ÇEŞİTLİ SUT (M.45/M.2/BPH) kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "Çeşitli Kontrol",
                "Bu dönemde Çeşitli (M.45/M.2/BPH) grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"ÇEŞİTLİ SUT (M.45/M.2/BPH) kontrolü tamamlandı.\n\n"
            f"Toplam ÇEŞİTLİ satır : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"Alt grup dağılımı:\n"
            f"  • Üriner (M.45)  : {kategori_sayac.get('URINER', 0)}\n"
            f"  • Gözyaşı (M.2)  : {kategori_sayac.get('GOZYASI', 0)}\n"
            f"  • BPH α-bloker   : {kategori_sayac.get('BPH', 0)}\n\n"
            f"Kapsam dışı (atlanan): {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._cesitli_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Çeşitli rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── ANTİHT KONTROL (SUT 4.2.12.B mono + kombi antihipertansif) ──
    def _antiht_kontrol_baslat(self):
        """ANTİHT KONTROL butonu — yüklenen satırlardan mono/kombi
        antihipertansif kapsamına girenleri sut_kategorisi_tespit_et()
        üzerinden sınıflar ve uygun kontrol fn'ine dispatch eder.

        Dispatch:
          - MONO  → kontrol_mono_antihipertansif
                    (ACE-i / ARB / KKB / BB / Diüretik / Alfa bloker;
                     raporsuz uyumluluk + ACE+ARB kontrendikasyon)
          - KOMBI → kontrol_kombine_antihipertansif
                    (raporsuz / 04.05 / monoterapi ibaresi 3-yollu)
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "ANTİHT Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_mono_antihipertansif,
                kontrol_kombine_antihipertansif,
            )
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"MONO": 0, "KOMBI": 0}
        denetlenen_satirlar = []

        # Aynı reçetenin diğer satırlarını grupla (ACE+ARB kombi yasağı için)
        rec_grup = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no") or ""
            rec_grup.setdefault(rno, []).append(s)

        # Önceki çalıştırmadan kalan ANTİHT verdict'lerini temizle
        for s in self.tum_satirlar:
            kategori = self._antiht_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") in ("MONO", "KOMBI"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1

            rno = s.get("rec_no") or ""
            ayni_recete = rec_grup.get(rno, [])
            ilac_sonuc = self._ilac_sonuc_olustur_antiht(s, ayni_recete)

            try:
                if kategori == "MONO":
                    rapor = kontrol_mono_antihipertansif(ilac_sonuc)
                else:  # KOMBI
                    rapor = kontrol_kombine_antihipertansif(ilac_sonuc)
            except Exception as e:
                logger.warning("ANTİHT kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"ANTİHT (4.2.12.B) kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "ANTİHT Kontrol",
                "Bu dönemde mono/kombi antihipertansif grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"ANTİHT (4.2.12.B mono+kombi) kontrolü tamamlandı.\n\n"
            f"Toplam ANTİHT satır : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"Tip dağılımı:\n"
            f"  • Mono  : {kategori_sayac.get('MONO', 0)}\n"
            f"  • Kombi : {kategori_sayac.get('KOMBI', 0)}\n\n"
            f"Kapsam dışı (atlanan): {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._antiht_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("ANTİHT rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── İSKEMİK KALP KONTROL (SUT 4.2.15.C İvabradin + 4.2.15.F Ranolazin) ──
    def _iskemik_kalp_kontrol_baslat(self):
        """İSKEMİK KALP KONTROL butonu — yüklenen satırlardan ivabradin/
        ranolazin/eplerenon kapsamına girenleri kategoriye göre dispatch eder.

        Dispatch:
          - IVABRADIN / EPLERENON → kontrol_ivabradin (NYHA + EF≤45% + BB
                                    intoleransı + sinüs ritmi)
          - RANOLAZIN             → kontrol_ranolazin (kronik stabil angina
                                    + BB/CCB intoleransı / yetersiz yanıt)
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "İSKEMİK KALP Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_ivabradin, kontrol_ranolazin,
            )
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"IVABRADIN": 0, "RANOLAZIN": 0, "EPLERENON": 0}
        denetlenen_satirlar = []

        rec_grup = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no") or ""
            rec_grup.setdefault(rno, []).append(s)

        for s in self.tum_satirlar:
            kategori = self._iskemik_kalp_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") in (
                        "IVABRADIN", "RANOLAZIN", "EPLERENON"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1

            rno = s.get("rec_no") or ""
            ayni_recete = rec_grup.get(rno, [])
            ilac_sonuc = self._ilac_sonuc_olustur_iskemik_kalp(s, ayni_recete)

            try:
                if kategori == "RANOLAZIN":
                    rapor = kontrol_ranolazin(ilac_sonuc)
                else:  # IVABRADIN veya EPLERENON
                    rapor = kontrol_ivabradin(ilac_sonuc)
            except Exception as e:
                logger.warning("İSKEMİK KALP kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"İSKEMİK KALP (4.2.15.C/F) kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "İSKEMİK KALP Kontrol",
                "Bu dönemde ivabradin/ranolazin/eplerenon grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"İSKEMİK KALP (4.2.15.C/F) kontrolü tamamlandı.\n\n"
            f"Toplam İSKEMİK KALP satır : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"İlaç dağılımı:\n"
            f"  • İvabradin  : {kategori_sayac.get('IVABRADIN', 0)}\n"
            f"  • Ranolazin  : {kategori_sayac.get('RANOLAZIN', 0)}\n"
            f"  • Eplerenon  : {kategori_sayac.get('EPLERENON', 0)}\n\n"
            f"Kapsam dışı (atlanan): {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._iskemik_kalp_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("İSKEMİK KALP rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── K-SİTRAT KONTROL (Potasyum Sitrat — SUT EK-4/F) ─────────────
    def _potasyum_sitrat_kontrol_baslat(self):
        """K-SİTRAT KONTROL butonu — yüklenen satırlardan potasyum sitrat
        kapsamına girenleri kontrol_potasyum_sitrat'a yollar.

        Üroloji/nefroloji uzmanı raporu zorunlu, ICD eşleşmesi (N20.0/N20.1
        böbrek taşı, N25.8/N25.9 RTA) kontrolü yapılır."""
        if not self.tum_satirlar:
            messagebox.showinfo(
                "K-SİTRAT Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import kontrol_potasyum_sitrat
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"POTASYUM": 0}
        denetlenen_satirlar = []

        for s in self.tum_satirlar:
            kategori = self._potasyum_sitrat_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") == "POTASYUM":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1

            ilac_sonuc = self._ilac_sonuc_olustur_potasyum_sitrat(s)
            try:
                rapor = kontrol_potasyum_sitrat(ilac_sonuc)
            except Exception as e:
                logger.warning("K-SİTRAT kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"K-SİTRAT (EK-4/F) kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "K-SİTRAT Kontrol",
                "Bu dönemde potasyum sitrat reçetesi bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"K-SİTRAT (EK-4/F) kontrolü tamamlandı.\n\n"
            f"Toplam K-SİTRAT satır : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"Kapsam dışı (atlanan): {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._potasyum_sitrat_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("K-SİTRAT rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── RAPORSUZ BİLGİ KONTROL (Raporsuz verilebilen ilaçlar) ────────
    def _raporsuz_bilgi_kontrol_baslat(self):
        """RAPORSUZ BİLGİ KONTROL butonu — raporsuz verilebilen ilaçları
        kontrol_raporsuz_bilgilendirme'ye yollar (28 ilaç sınıfı dispatcher).

        Kapsam: RAPORSUZ_BILGILENDIRME / BENZODIAZEPIN / GOZ_LUBRIKAN
        kategorilerinin tümü."""
        if not self.tum_satirlar:
            messagebox.showinfo(
                "RAPORSUZ BİLGİ Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import kontrol_raporsuz_bilgilendirme
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"RAPORSUZ": 0}
        denetlenen_satirlar = []

        rec_grup = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no") or ""
            rec_grup.setdefault(rno, []).append(s)

        for s in self.tum_satirlar:
            kategori = self._raporsuz_bilgi_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") == "RAPORSUZ":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1

            rno = s.get("rec_no") or ""
            ayni_recete = rec_grup.get(rno, [])
            ilac_sonuc = self._ilac_sonuc_olustur_raporsuz_bilgi(s, ayni_recete)

            try:
                rapor = kontrol_raporsuz_bilgilendirme(ilac_sonuc)
            except Exception as e:
                logger.warning("RAPORSUZ BİLGİ kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"RAPORSUZ BİLGİ kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "RAPORSUZ BİLGİ Kontrol",
                "Bu dönemde raporsuz bilgilendirme kapsamına giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"RAPORSUZ BİLGİ kontrolü tamamlandı.\n\n"
            f"Toplam RAPORSUZ satır : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"Kapsam dışı (atlanan): {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._raporsuz_bilgi_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("RAPORSUZ BİLGİ rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── BASIT SUT KONTROL (DMAH+Kadın+Trimetazidin+Kinolon) ──────────
    def _basit_sut_kontrol_baslat(self):
        """BASIT SUT KONTROL butonu — 4 mikro fonksiyonu kategoriye göre
        dispatch eder.

        Dispatch:
          - DMAH                     → kontrol_dmah
          - KADIN_HORMON             → kontrol_kadin_hormon
          - TRIMETAZIDIN             → kontrol_trimetazidin
          - ANTIBIYOTIK_FLOROKINOLON → kontrol_antibiyotik_florokinolon
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "BASIT SUT Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_dmah, kontrol_kadin_hormon, kontrol_trimetazidin,
                kontrol_antibiyotik_florokinolon,
            )
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        DISPATCH = {
            "DMAH":                     kontrol_dmah,
            "KADIN_HORMON":             kontrol_kadin_hormon,
            "TRIMETAZIDIN":             kontrol_trimetazidin,
            "ANTIBIYOTIK_FLOROKINOLON": kontrol_antibiyotik_florokinolon,
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {k: 0 for k in DISPATCH}
        denetlenen_satirlar = []

        for s in self.tum_satirlar:
            kategori = self._basit_sut_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") in DISPATCH:
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1

            ilac_sonuc = self._ilac_sonuc_olustur_basit_sut(s)
            kontrol_fn = DISPATCH[kategori]

            try:
                rapor = kontrol_fn(ilac_sonuc)
            except Exception as e:
                logger.warning("BASIT SUT kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"BASIT SUT kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "BASIT SUT Kontrol",
                "Bu dönemde DMAH/Kadın hormon/Trimetazidin/Florokinolon kapsamına giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"BASIT SUT kontrolü tamamlandı.\n\n"
            f"Toplam BASIT SUT satır : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"Kategori dağılımı:\n"
            f"  • DMAH             : {kategori_sayac.get('DMAH', 0)}\n"
            f"  • Kadın hormon     : {kategori_sayac.get('KADIN_HORMON', 0)}\n"
            f"  • Trimetazidin     : {kategori_sayac.get('TRIMETAZIDIN', 0)}\n"
            f"  • Florokinolon     : {kategori_sayac.get('ANTIBIYOTIK_FLOROKINOLON', 0)}\n\n"
            f"Kapsam dışı (atlanan): {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._basit_sut_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("BASIT SUT rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── PSİKİYATRİ / NÖROLOJİ KONTROL (SUT 4.2.2 + 4.2.25) ───────────
    def _psikiyatri_kontrol_baslat(self):
        """PSİKİYATRİ/NÖROLOJİ KONTROL butonu — SUT 4.2.2 (Antipsikotik/
        Antidepresan/Mood Stab/Benzodiazepin) + SUT 4.2.25 (Antiepileptik)
        kapsamındaki satırları denetler.

        Dispatch:
          - Kategori "ANTIEPILEPTIK"  → kontrol_antiepileptik_4_2_25
          - Kategori "PSIKIYATRI"     → kontrol_psikiyatri
          - Lamotrijin/Valproat bipolar → AEP fn ATLANDI döner, fallback PSI fn
          - Aynı reçetenin diğer satırları diger_etken_maddeler/diger_ilac_adlari
            olarak iletilir (Pregabalin+Gabapentin kombi yasağı için)
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Psikiyatri/Nöroloji Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_psikiyatri, kontrol_antiepileptik_4_2_25,
            )
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"PSIKIYATRI": 0, "ANTIEPILEPTIK": 0}
        alt_sayac = {}
        denetlenen_satirlar = []

        # Aynı reçeteye ait satırları gruplama (kombinasyon yasağı için)
        # Reçete numarasına göre dict
        rec_grup = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no") or ""
            rec_grup.setdefault(rno, []).append(s)

        # Önceki çalıştırmadan kalan PSI/AEP verdict'lerini temizle
        for s in self.tum_satirlar:
            kategori = self._psikiyatri_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") in ("PSIKIYATRI", "ANTIEPILEPTIK"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue

            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1
            alt_sinif = self._psikiyatri_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            alt_sayac[alt_sinif] = alt_sayac.get(alt_sinif, 0) + 1

            rno = s.get("rec_no") or ""
            ayni_recete = rec_grup.get(rno, [])
            ilac_sonuc = self._ilac_sonuc_olustur_psikiyatri(s, ayni_recete)

            try:
                if kategori == "ANTIEPILEPTIK":
                    rapor = kontrol_antiepileptik_4_2_25(ilac_sonuc)
                    # Lamotrijin/Valproat bipolar → AEP ATLANDI döner, PSI'ye fallback
                    if rapor.sonuc == KontrolSonucu.ATLANDI:
                        rapor = kontrol_psikiyatri(ilac_sonuc)
                else:
                    rapor = kontrol_psikiyatri(ilac_sonuc)
            except Exception as e:
                logger.warning("Psikiyatri/AEP kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"Psikiyatri/Nöroloji SUT (4.2.2 + 4.2.25) kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']} satır boş bırakıldı)"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "Psikiyatri/Nöroloji Kontrol",
                "Bu dönemde psikiyatri/nöroloji (4.2.2 / 4.2.25) "
                "grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Psikiyatri/Nöroloji SUT (4.2.2 + 4.2.25) kontrolü tamamlandı.\n\n"
            f"Toplam denetlenen satır : {toplam}\n"
            f"  ✓ UYGUN              : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL        : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ            : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI            : {sayac['ATLANDI']}\n"
            f"  Psikiyatri (4.2.2)   : {kategori_sayac.get('PSIKIYATRI', 0)}\n"
            f"  Antiepileptik (4.2.25): {kategori_sayac.get('ANTIEPILEPTIK', 0)}\n"
            f"Kapsam dışı (atlanan)  : {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._psikiyatri_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                alt_sayac=alt_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Psikiyatri/Nöroloji rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── HEPATİT B/C KONTROL (SUT 4.2.13.3 / 4.2.13.4) ────────────────
    def _hepatit_kontrol_baslat(self):
        """HEPATİT B/C KONTROL butonu — yüklenen satırlardan hepatit B
        (J05AF) / HCV DAA (J05AP) / klasik (peginterferon + ribavirin)
        ilaçlarını SUT 4.2.13.3 ve 4.2.13.4 kuralına göre denetler ve
        sonucu en sağdaki SONUÇ sütununa yazar.

        Kapsam:
          • HBV: Entekavir, Tenofovir TDF/TAF, Lamivudin, Telbivudin, Adefovir
          • HCV DAA: Sofosbuvir, Ledipasvir, Velpatasvir, Voxilaprevir,
                     Glecaprevir, Pibrentasvir, Ombitasvir, Paritaprevir,
                     Dasabuvir, Daklatasvir, Elbasvir, Grazoprevir
          • Klasik: Peginterferon alfa, Ribavirin

        HIV ilaçları (Dolutegravir/Abacavir/Efavirenz vb.) bu butonun
        KAPSAMI DIŞINDA — onlar ANTIVIRAL kategorisinde kalır.

        Kurallar:
          • Uzman branş: Gastroenteroloji / Enfeksiyon Hastalıkları /
                         Hepatoloji (HBV için İç Hastalıkları da kabul)
          • HBV: HBsAg pozitif + HBV DNA + ALT yüksek
          • HCV: Anti-HCV pozitif + HCV RNA + genotip + fibrozis (METAVIR)
          • Rapor kodu pragmatik fallback (06.01 / 14.* / B*) → Medula
            şart kontrolünü yapmış kabul edilir
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Hepatit B/C Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_hepatit, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        # Hastaların TÜM aktif raporlarındaki ICD kodlarını topla
        # (B16/B17/B18 hepatit ICD'leri ayrı raporda yazılı olabilir)
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"Hepatit B/C kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_hepatit_disi": 0}
        kategori_sayac = {"HBV": 0, "HCV": 0, "KLASIK": 0}
        denetlenen_satirlar = []

        # Önceki çalıştırmadan kalan kendi verdict'lerimizi temizle
        for s in self.tum_satirlar:
            kategori = self._hepatit_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin (HEPATIT) önceki artığını sil —
                # statin/diyabet/klopidogrel/arb/osteo verdict'lerine dokunma.
                if s.get("verdict_kategori") in ("HBV", "HCV", "KLASIK"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_alt_sinif"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_hepatit_disi"] += 1
                continue

            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1
            alt_sinif = self._hepatit_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            ilac_sonuc = self._ilac_sonuc_olustur_hepatit(s)

            # Hastanın diğer raporlarındaki ICD'leri ekle
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                rapor = kontrol_hepatit(ilac_sonuc)
            except Exception as e:
                logger.warning("Hepatit kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(rapor, ek_icd, ['DM', 'KY'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = kategori
            s["verdict_alt_sinif"] = alt_sinif
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"Hepatit B/C SUT kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(hepatit dışı {sayac['_hepatit_disi']} satır boş bırakıldı)"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam_hep = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                      + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam_hep == 0:
            messagebox.showinfo(
                "Hepatit B/C Kontrol",
                "Bu dönemde hepatit B/C grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Hepatit B/C SUT kontrolü tamamlandı.\n\n"
            f"Toplam hepatit satırı: {toplam_hep}\n"
            f"  HBV               : {kategori_sayac['HBV']}\n"
            f"  HCV               : {kategori_sayac['HCV']}\n"
            f"  Klasik (PegIFN/RBV): {kategori_sayac['KLASIK']}\n\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Hepatit dışı (atlanan): {sayac['_hepatit_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._hepatit_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Hepatit rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── ASTIM / KOAH KONTROLÜ (verdict sütunu) ───────────────────────
    def _astim_koah_kontrol_baslat(self):
        """ASTIM/KOAH KONTROL butonu — yüklenen satırlardan ATC R03* (solunum
        sistemi obstrüktif hastalık ilaçları) olanları SUT 4.2.24 / 4.2.24.B
        kuralına göre denetler ve sonucu en sağdaki SONUÇ sütununa yazar.

        ÖNEMLİ:
          - Hastanın bu reçetede sadece solunum ilacının kendi raporu değil,
            TÜM aktif raporlarındaki ICD kodları da tanı tespiti için kullanılır
            (astım/KOAH ayrı bir raporda olabilir).
          - Aynı reçetedeki diğer ilaçların adları kontrol_solunum'a iletilir;
            LABA+ICS+LAMA üçlü kullanımı (3 ayrı ilaç olarak) bu sayede tespit
            edilir → SUT 4.2.24.B özel hükmü uygulanır.
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Astım/KOAH Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_solunum, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        # Hastaların TÜM raporlarındaki ICD kodlarını topla (cross-rapor tanı)
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"Astım/KOAH kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        # Reçete bazında diğer ilaçların adlarını topla — kontrol_solunum
        # üçlü kullanım (LABA+ICS+LAMA) tespiti için bunları okuyor.
        rec_no_ilaclar: dict = {}
        for s in self.tum_satirlar:
            rno = (s.get("rec_no") or "").strip()
            if not rno:
                continue
            ad = s.get("ilac") or ""
            if ad:
                rec_no_ilaclar.setdefault(rno, []).append(ad)

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_solunum_disi": 0}
        # Alt-sınıf bazlı sayım (rapor için)
        ALT_SINIFLAR = ("SABA", "SAMA", "SABA+SAMA", "ICS", "LABA", "LAMA",
                        "LABA+ICS", "LABA+LAMA", "UCLU", "LTRA", "OMALIZUMAB",
                        "ANTI_IL", "ROFLUMILAST", "TEOFILIN", "KROMOLIN",
                        "DIGER")
        kategori_sayac = {k: 0 for k in ALT_SINIFLAR}
        denetlenen_satirlar = []

        # Önceki çalıştırmadan kalan kendi verdict'lerimizi temizle
        for s in self.tum_satirlar:
            kategori = self._solunum_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin (SOLUNUM) önceki artığını sil —
                # statin/diyabet/klopidogrel/arb/osteo/hepatit verdict'lerine
                # dokunma.
                if s.get("verdict_kategori") == "SOLUNUM":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_alt_sinif"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_solunum_disi"] += 1
                continue

            alt_sinif = self._solunum_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            kategori_sayac[alt_sinif] = kategori_sayac.get(alt_sinif, 0) + 1

            # Aynı reçetedeki diğer ilaçların adları (kendi adı hariç)
            rno = (s.get("rec_no") or "").strip()
            kendi_ad = (s.get("ilac") or "").upper()
            diger = []
            for ad in rec_no_ilaclar.get(rno, []):
                if ad and ad.upper() != kendi_ad:
                    diger.append(ad)

            ilac_sonuc = self._ilac_sonuc_olustur_solunum(s, diger)

            # Hastanın diğer raporlarındaki ICD'leri ekle
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                rapor = kontrol_solunum(ilac_sonuc)
            except Exception as e:
                logger.warning("Solunum kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "SOLUNUM"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(rapor, ek_icd, ['KOAH', 'ASTIM'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = "SOLUNUM"
            s["verdict_alt_sinif"] = alt_sinif
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"Astım/KOAH SUT kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(solunum dışı {sayac['_solunum_disi']} satır boş bırakıldı)"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam_sol = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                      + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam_sol == 0:
            messagebox.showinfo(
                "Astım/KOAH Kontrol",
                "Bu dönemde astım/KOAH (solunum) grubuna giren reçete "
                "bulunamadı.",
                parent=self.root)
            return

        # En sık görülen 3 alt sınıfı özette gösterelim
        en_sik = sorted(
            ((k, v) for k, v in kategori_sayac.items() if v > 0),
            key=lambda x: x[1], reverse=True)[:5]
        en_sik_str = "\n".join(f"  {k:<14}: {v}" for k, v in en_sik) \
                       or "  (yok)"

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Astım/KOAH SUT kontrolü tamamlandı.\n\n"
            f"Toplam solunum satırı: {toplam_sol}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Solunum dışı (atlanan): {sayac['_solunum_disi']}\n\n"
            f"En sık alt sınıflar:\n{en_sik_str}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._solunum_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Astım/KOAH rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── DOZ KARŞILAŞTIRMA (verdict_doz sütunu) ────────────────────────
    def _doz_kontrol_baslat(self):
        """DOZ KARŞILAŞTIR butonu — yüklenen tüm satırlar için reçete
        günlük dozu ile rapor günlük dozunu (RaporEtkinMadde + ek bilgi
        parse) karşılaştırır. Sonuç 'Doz Karş.' sütununa yazılır,
        ardından ayrıntılı Excel raporu üretilir.
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Doz Karşılaştır",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.doz_kontrol import kontrol_doz
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"Doz kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol.doz_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0}
        denetlenen = []

        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        for s in self.tum_satirlar:
            # Önceki çalıştırmadan kalan doz verdict'ini temizle
            for k in ("verdict_doz", "verdict_doz_detay",
                      "verdict_doz_aranan", "verdict_doz_bulunan",
                      "verdict_doz_kaynak", "verdict_doz_rec_g",
                      "verdict_doz_rap_g"):
                s[k] = ""

            ilac_sonuc = {
                "ilac_adi": s.get("ilac") or "",
                "etkin_madde": s.get("etkin") or "",
                "atc": s.get("atc") or "",
                "rec_doz_raw": s.get("rec_doz_raw") or {},
                "rap_doz_listesi": s.get("rap_doz_listesi") or [],
                "rapor_aciklamalari": _bol(s.get("rap_ack")),
                "rapor_kodu": (s.get("rap_kod") or "").strip(),
                # Raporsuz ilaçlar doz kontrolü kapsamı dışında — rapor
                # yoksa rapor dozu da yok, karşılaştırma anlamsız.
                "raporlu": bool(s.get("rapor_ana_id")),
            }
            try:
                rapor = kontrol_doz(ilac_sonuc)
            except Exception as e:
                logger.warning("Doz kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict_doz"] = "ŞÜPHELİ"
                s["verdict_doz_detay"] = f"Hata: {e}"
                sayac["ŞÜPHELİ"] += 1
                denetlenen.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict_doz"] = etiket
            s["verdict_doz_detay"] = rapor.mesaj or ""
            s["verdict_doz_aranan"] = rapor.aranan_ibare or ""
            s["verdict_doz_bulunan"] = rapor.bulunan_metin or ""
            d = rapor.detaylar or {}
            s["verdict_doz_kaynak"] = d.get("kaynak", "")
            s["verdict_doz_rec_g"] = (
                f"{d.get('rec_gunluk'):g}"
                if d.get("rec_gunluk") is not None else "")
            s["verdict_doz_rap_g"] = (
                f"{d.get('rap_gunluk'):g}"
                if d.get("rap_gunluk") is not None else "")
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen.append(s)

        self._tabloyu_yenile()
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        self._durum_yaz(
            f"Doz karşılaştırma: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(toplam {toplam})"
        )

        if toplam == 0:
            messagebox.showinfo(
                "Doz Karşılaştır",
                "Bu dönemde değerlendirilebilecek reçete satırı bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Doz karşılaştırma tamamlandı.\n\n"
            f"Toplam satır       : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._doz_rapor_excel_olustur(
                sayac=sayac, denetlenen_satirlar=denetlenen)
        except Exception as e:
            logger.exception("Doz kontrol rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    def _doz_rapor_excel_olustur(self, *, sayac: dict,
                                   denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne doz karşılaştırma Excel
        raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, dağılım, kaynak istatistiği
          - Doz Detay : Her satır + reçete/rapor günlük doz + kaynak +
                        aranan/bulunan + sonuç (renkli)
          - Şüpheli & Eşleşmeyen : Karar verilemeyen satırların özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Doz_Karsilastirma_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam = sum(sayac.values())

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"
        ws1.cell(row=1, column=1,
                 value="DOZ KARŞILAŞTIRMA RAPORU "
                       "(reçete günlük dozu ≤ rapor günlük dozu)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="E65100")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        # Kaynak istatistiği
        kaynak_sayac = {"RaporEtkinMadde": 0, "RaporEkBilgi (parse)": 0,
                        "yok": 0}
        for s in denetlenen_satirlar:
            k = s.get("verdict_doz_kaynak") or "yok"
            kaynak_sayac[k] = kaynak_sayac.get(k, 0) + 1

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Değerlendirilen Satır", str(toplam)),
            ("", ""),
            ("Kontrol Yöntemi",
             "Reçete (RIDoz/RITekrar/RIAralik/RIPeriyot) → günlük doz; "
             "Rapor: RaporEtkinMadde → günlük doz; "
             "yapısal yoksa RaporEkBilgi metninden parse."),
            ("Karşılaştırma",
             "Reçete günlük ≤ Rapor günlük × 1.001 → UYGUN; aşılırsa UYGUN DEĞİL."),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="E65100")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac.get(etiket, 0)
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="RAPOR DOZU KAYNAĞI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Kaynak", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kaynak_aciklama = {
            "RaporEtkinMadde":      "Yapısal alan (doz/adet/tekrar/aralık/periyot)",
            "RaporEkBilgi (parse)": "Açıklama metninden NxN parse edildi",
            "yok":                  "Rapor dozu çıkarılamadı (raporsuz / eşleşme yok)",
        }
        for i, k in enumerate(["RaporEtkinMadde", "RaporEkBilgi (parse)",
                                "yok"], start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kaynak_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kaynak_aciklama[k])

        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• Günlük doz = (doz × tekrar) / (aralık × periyot_gün). "
            "Periyot: 3=günde, 4=haftada(×7), 5=ayda(×30), 6=yılda(×365).",
            "• UYGUN = reçete günlük ≤ rapor günlük (≤ %0.1 tolerans).",
            "• UYGUN DEĞİL = reçete günlük > rapor günlük → manuel inceleme önerilir.",
            "• ŞÜPHELİ = rapor dozu yapısal alanda yok ve açıklama metninde "
            "ilaç/etken bağlamlı NxN ifade bulunamadı.",
            "• ATLANDI = raporsuz ilaç (msj=yok ve hiç rapor verisi yok).",
            "• Reçete birimi (Adet) ile rapor birimi farklıysa (mg/ml) v1'de "
            "ŞÜPHELİ — mg↔tablet dönüşümü v2 işi.",
            "• Çoklu etken raporlarda eşleşme: ATC > etkin_id > etkin_ad. "
            "Tek etkenli raporda otomatik eşleşir.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: DOZ DETAY ──────────
        ws2 = wb.create_sheet("Doz Detay")
        kolonlar = [
            ("rec_tar",            "Reç.Tarih",       12),
            ("rec_no",             "Reçete No",       18),
            ("hasta",              "Hasta",           24),
            ("tc",                 "TC",              13),
            ("ilac",               "İlaç",            28),
            ("etkin",              "Etken Madde",     22),
            ("atc",                "ATC",             10),
            ("rap_kod",            "Rapor Kod",       11),
            ("rec_doz",            "Reçete Doz (metin)", 18),
            ("verdict_doz_rec_g",  "Reç. Günlük",     12),
            ("rap_doz",            "Rapor Doz (metin)",  20),
            ("verdict_doz_rap_g",  "Rapor Günlük",    12),
            ("verdict_doz_kaynak", "Kaynak",          22),
            ("msj",                "Msj",             7),
            ("verdict_doz_aranan", "Aranan İbare",    24),
            ("verdict_doz_bulunan","Bulunan Metin",   24),
            ("rap_ack",            "Rap.Açıklama",    32),
            ("verdict_doz_detay",  "Açıklama",        50),
            ("verdict_doz",        "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict_doz") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center",
                                             vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: ŞÜPHELİ ──────────
        ws3 = wb.create_sheet("Şüpheli & Eşleşmeyen")
        ws3.cell(row=1, column=1,
                 value=("Aşağıdaki satırlarda doz karşılaştırması "
                        "yapılamadı (rapor dozu yapısal alanda yok ve "
                        "açıklamada ilaç/etken bağlamlı NxN bulunamadı).")
                 ).font = Font(italic=True, color="546E7A")
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        atl_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("rec_doz", "Reçete Doz", 18),
            ("rap_doz", "Rapor Doz",  18),
            ("verdict_doz_detay", "Sebep", 50),
        ]
        for c, (_k, b, _g) in enumerate(atl_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in denetlenen_satirlar:
            if (s.get("verdict_doz") or "") != "ŞÜPHELİ":
                continue
            for ci, (kod, _b, _g) in enumerate(atl_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atl_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── RENKLİ REÇETE SİSTEMİ KONTROL (renkli_rr sütunu) ──────────────
    def _renkli_recete_kontrol_baslat(self):
        """RENKLİ REÇETE butonu — yüklenen tüm satırlardan Kırmızı/Yeşil/Mor
        reçete türündekileri alır; ana ekranda yüklenmiş 'Renkli Reçete
        Listesi' (PDF/Excel/Manuel) ile karşılaştırır.

        Sonuç 'Renkli RR' sütununa:
          • VAR — reçete numarası listede bulundu
          • YOK — kapsamda ama listede bulunamadı (sisteme işlenmemiş)
          • —   — kapsam dışı (Beyaz veya bilinmeyen tür)

        YOK satırlar kırmızıya boyanır; ardından Excel raporu üretilir.
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Renkli Reçete Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return

        try:
            from recete_kontrol import get_renkli_recete_kontrol
        except Exception as e:
            self._durum_yaz(f"Renkli reçete modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol.renkli_recete modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        rk = get_renkli_recete_kontrol()
        if not rk.pdf_yuklu or not rk.pdf_receteler:
            messagebox.showwarning(
                "Renkli Reçete Listesi Yok",
                "Renkli reçete listesi yüklü değil.\n\n"
                "Ana ekrandaki '🔴🟢 Renkli Reçete Yükle' butonu ile "
                "PDF/Excel listesini önce yükleyin.",
                parent=self.root)
            return

        # Önceki çalıştırmadan kalan değerleri temizle
        for s in self.tum_satirlar:
            s["renkli_rr"] = ""

        # Reçete bazında dedup — aynı reçetenin birden çok ilaç satırı için
        # tek kontrol; sonuç tüm satırlara yazılır.
        KAPSAM = {"Kırmızı", "Yeşil", "Mor"}
        rec_no_kapsam: dict = {}        # {rec_no: rec_tip}
        rec_no_satirlar: dict = {}      # {rec_no: [satir, ...]}
        kapsam_disi = 0
        for s in self.tum_satirlar:
            rec_no = (s.get("rec_no") or "").strip()
            rec_tip = (s.get("rec_tip") or "").strip()
            if rec_tip in KAPSAM and rec_no:
                rec_no_kapsam[rec_no] = rec_tip
                rec_no_satirlar.setdefault(rec_no, []).append(s)
            else:
                s["renkli_rr"] = "—"
                kapsam_disi += 1

        # Kontrol — recete_pdf_de_var_mi() ile (kontrol_et bypass edilir,
        # zira o fonksiyon Mor reçeteyi kapsam dışı bırakıyor).
        sayac = {"VAR": 0, "YOK": 0}
        renk_sayac = {"Kırmızı": {"VAR": 0, "YOK": 0},
                      "Yeşil":   {"VAR": 0, "YOK": 0},
                      "Mor":     {"VAR": 0, "YOK": 0}}
        yok_listesi: list = []   # [(rec_no, rec_tip, [satirlar])]
        for rec_no, rec_tip in rec_no_kapsam.items():
            var_mi = rk.recete_pdf_de_var_mi(rec_no)
            etiket = "VAR" if var_mi else "YOK"
            sayac[etiket] += 1
            renk_sayac[rec_tip][etiket] += 1
            for s in rec_no_satirlar[rec_no]:
                s["renkli_rr"] = etiket
            if not var_mi:
                yok_listesi.append((rec_no, rec_tip, rec_no_satirlar[rec_no]))
                # Satırları kırmızıya boya — sadece beyaz/sarı olanlar
                # (kullanıcının diğer kontrollerden boyamış olduklarına dokunma)
                for s in rec_no_satirlar[rec_no]:
                    iid = str(s["ri_id"])
                    mevcut = self.satir_renkleri.get(iid, RENK_BEYAZ)
                    if mevcut in (RENK_BEYAZ, RENK_SARI):
                        self.satir_renkleri[iid] = RENK_KIRMIZI

        self._tabloyu_yenile()
        self._sayaclari_guncelle()
        toplam = sayac["VAR"] + sayac["YOK"]
        self._durum_yaz(
            f"Renkli reçete kontrolü: "
            f"✓ VAR {sayac['VAR']}  "
            f"✗ YOK {sayac['YOK']}  "
            f"(toplam {toplam} renkli reçete; "
            f"kapsam dışı {kapsam_disi} satır)"
        )

        if toplam == 0:
            messagebox.showinfo(
                "Renkli Reçete Kontrol",
                "Bu dönemde Kırmızı/Yeşil/Mor reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Renkli reçete kontrolü tamamlandı.\n\n"
            f"Toplam renkli reçete : {toplam}\n"
            f"  ✓ Sistemde VAR    : {sayac['VAR']}\n"
            f"  ✗ Sistemde YOK    : {sayac['YOK']}\n\n"
            f"Renge göre:\n"
            f"  🔴 Kırmızı: VAR {renk_sayac['Kırmızı']['VAR']}  "
            f"YOK {renk_sayac['Kırmızı']['YOK']}\n"
            f"  🟢 Yeşil  : VAR {renk_sayac['Yeşil']['VAR']}  "
            f"YOK {renk_sayac['Yeşil']['YOK']}\n"
            f"  🟣 Mor    : VAR {renk_sayac['Mor']['VAR']}  "
            f"YOK {renk_sayac['Mor']['YOK']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._renkli_rr_rapor_excel_olustur(
                sayac=sayac,
                renk_sayac=renk_sayac,
                kapsam_disi=kapsam_disi,
                yok_listesi=yok_listesi,
                rec_no_kapsam=rec_no_kapsam,
                rec_no_satirlar=rec_no_satirlar,
                liste_kaynagi=rk.pdf_dosya_adi or "(manuel)",
                liste_sayisi=len(rk.pdf_receteler),
            )
        except Exception as e:
            logger.exception("Renkli reçete rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    def _renkli_rr_rapor_excel_olustur(self, *, sayac: dict,
                                          renk_sayac: dict,
                                          kapsam_disi: int,
                                          yok_listesi: list,
                                          rec_no_kapsam: dict,
                                          rec_no_satirlar: dict,
                                          liste_kaynagi: str,
                                          liste_sayisi: int) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne renkli reçete denetim raporu.

        3 sayfa:
          - Özet         : Toplam, renge göre dağılım, liste kaynağı
          - YOK Listesi  : Renkli reçete sistemine işlenmemiş reçeteler
          - Tüm Renkli   : Kapsamdaki tüm reçete-ilaç satırları + VAR/YOK
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Renkli_Recete_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        VAR_RENK = "C8E6C9"
        YOK_RENK = "FFCDD2"
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")

        toplam = sayac["VAR"] + sayac["YOK"]

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"
        ws1.cell(row=1, column=1,
                 value="RENKLİ REÇETE SİSTEMİ KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="4A148C")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Renkli Reçete (Kırmızı/Yeşil/Mor)", str(toplam)),
            ("  ✓ Sistemde VAR", str(sayac["VAR"])),
            ("  ✗ Sistemde YOK", str(sayac["YOK"])),
            ("Kapsam Dışı (Beyaz vb.) Satır", str(kapsam_disi)),
            ("", ""),
            ("Liste Kaynağı (yüklenmiş)", liste_kaynagi),
            ("Listedeki Reçete Sayısı", str(liste_sayisi)),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="4A148C")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="GENEL DAĞILIM")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, (etiket, renk) in enumerate(
                [("VAR", VAR_RENK), ("YOK", YOK_RENK)], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=renk)
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Renge göre dağılım
        bas2 = bas + 5
        ws1.cell(row=bas2, column=1, value="RENGE GÖRE DAĞILIM")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Reçete Türü", "VAR", "YOK", "Toplam"],
                                  start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        renk_zemin = {"Kırmızı": "FFCDD2", "Yeşil": "C8E6C9",
                      "Mor": "E1BEE7"}
        for i, tip in enumerate(["Kırmızı", "Yeşil", "Mor"], start=bas2 + 1):
            v = renk_sayac[tip]["VAR"]
            y = renk_sayac[tip]["YOK"]
            t = v + y
            c1 = ws1.cell(row=i, column=1, value=tip)
            c2 = ws1.cell(row=i, column=2, value=v)
            c3 = ws1.cell(row=i, column=3, value=y)
            c4 = ws1.cell(row=i, column=4, value=t)
            c1.fill = PatternFill("solid", fgColor=renk_zemin[tip])
            c1.font = Font(bold=True)
            for c in (c2, c3, c4):
                c.alignment = Alignment(horizontal="center")

        # Notlar
        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• Karşılaştırma reçete numarası bazlıdır (rec_no).",
            "• 'YOK' = reçete yüklenen renkli reçete listesinde bulunamadı "
            "→ renkli reçete sistemine işlenmemiş olabilir; manuel kontrol önerilir.",
            "• 'VAR' = reçete listede bulundu (sisteme işlenmiş).",
            "• '—' = Beyaz veya bilinmeyen tür (kapsam dışı).",
            "• Turuncu reçete bu sürüm kapsamında değildir.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 18, 14], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: YOK LİSTESİ ──────────
        ws2 = wb.create_sheet("YOK Listesi")
        ws2.cell(row=1, column=1,
                 value="Renkli reçete sistemine İŞLENMEMİŞ reçeteler "
                       "(reçete numarası listede bulunamadı):").font = (
            Font(italic=True, bold=True, color="C62828"))
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        yok_kolonlar = [
            ("rec_tar",  "Reç.Tarih",   12),
            ("rec_no",   "Reçete No",   18),
            ("rec_tip",  "Türü",        10),
            ("hasta",    "Hasta",       24),
            ("tc",       "TC",          13),
            ("doktor",   "Doktor",      22),
            ("ilac",     "İlaç(lar)",   60),
        ]
        for c, (_k, b, _g) in enumerate(yok_kolonlar, 1):
            cell = ws2.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        renk_zemin_satir = {"Kırmızı": "FFEBEE", "Yeşil": "E8F5E9",
                              "Mor": "F3E5F5"}
        for rec_no, rec_tip, satirlar in yok_listesi:
            ilk = satirlar[0]
            ilac_birlesik = " | ".join(
                dict.fromkeys((s.get("ilac") or "") for s in satirlar
                              if s.get("ilac")))
            satir_zemin = PatternFill(
                "solid", fgColor=renk_zemin_satir.get(rec_tip, "FFFFFF"))
            for ci, (kod, _b, _g) in enumerate(yok_kolonlar, 1):
                if kod == "ilac":
                    deger = ilac_birlesik
                else:
                    deger = ilk.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = satir_zemin
            ri += 1
        for ci, (_k, _b, gen) in enumerate(yok_kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.freeze_panes = "A4"
        if ri > 4:
            ws2.auto_filter.ref = f"A3:{get_column_letter(len(yok_kolonlar))}{ri-1}"

        # ────────── SAYFA 3: TÜM RENKLİ DETAY ──────────
        ws3 = wb.create_sheet("Tüm Renkli")
        kolonlar = [
            ("rec_tar",   "Reç.Tarih",   12),
            ("rec_no",    "Reçete No",   18),
            ("rec_tip",   "Türü",        10),
            ("hasta",     "Hasta",       24),
            ("tc",        "TC",          13),
            ("doktor",    "Doktor",      22),
            ("brans",     "Branş",       18),
            ("ilac",      "İlaç",        28),
            ("etkin",     "Etken Madde", 22),
            ("atc",       "ATC",         10),
            ("rap_kod",   "Rapor Kod",   11),
            ("rec_doz",   "Reçete Doz",  14),
            ("kutu",      "Kutu",        6),
            ("renkli_rr", "Renkli RR",   12),
        ]
        for c, (_k, b, _g) in enumerate(kolonlar, 1):
            cell = ws3.cell(row=1, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws3.row_dimensions[1].height = 28
        ws3.freeze_panes = "A2"

        ri = 2
        son_col = len(kolonlar)
        for rec_no in rec_no_kapsam.keys():
            for s in rec_no_satirlar[rec_no]:
                for ci, (kod, _b, _g) in enumerate(kolonlar, 1):
                    cell = ws3.cell(row=ri, column=ci,
                                    value=str(s.get(kod, "")))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                etiket = s.get("renkli_rr") or ""
                vcell = ws3.cell(row=ri, column=son_col)
                if etiket == "VAR":
                    vcell.fill = PatternFill("solid", fgColor=VAR_RENK)
                elif etiket == "YOK":
                    vcell.fill = PatternFill("solid", fgColor=YOK_RENK)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")
                ri += 1

        for ci, (_k, _b, gen) in enumerate(kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        if ri > 2:
            ws3.auto_filter.ref = ws3.dimensions

        wb.save(path)
        return path

    # ── UYARI KODU KONTROL (verdict_uyari_kontrol sütunu) ─────────────
    def _uyari_kod_kontrol_baslat(self):
        """UYARI KODU KONTROL butonu — yüklenen tüm satırlardaki reçete
        uyarı kodlarını (256, 280 vb.) alır; her kodun açıklamasındaki
        anahtar kelimeleri reçete teşhisi, rapor teşhisi, reçete
        açıklaması ve rapor açıklaması içinde arar. Sonuç 'Uyarı Kod
        Karş.' sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ olarak yazılır.

        Karar mantığı (bir satırda birden çok kod olabilir):
          - Hepsi UYGUN              → UYGUN
          - En az 1 UYGUN DEĞİL      → UYGUN DEĞİL
          - Karışık (UYGUN + ŞÜPHELİ veya hepsi ŞÜPHELİ) → ŞÜPHELİ

        Tek kod için:
          - durum=UYGUN              → UYGUN
          - durum=UYGUNSUZ + oran=0  → UYGUN DEĞİL (hiç eşleşme yok)
          - durum=UYGUNSUZ + oran>0  → ŞÜPHELİ (kısmi eşleşme, eşik altı)
          - durum=EDILEMEDI (ör 272) → ŞÜPHELİ
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Uyarı Kodu Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_tarama import uyari_kodu_kontrol
        except Exception as e:
            self._durum_yaz(f"Uyarı kodu kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_tarama.uyari_kodu_kontrol yüklenemedi:\n{e}",
                parent=self.root)
            return

        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "_kodsuz": 0}
        denetlenen = []  # rapor için (uyarı kodu olanlar)

        def _bol(metin):
            if not metin:
                return []
            return [p.strip() for p in str(metin).split(" | ") if p.strip()]

        def _uyari_parse(uyari_text: str):
            """'256 - Benign prostat | 280 - Hipertansiyon | (Reç: 367 - X)'
            → [{'kod': '256', 'aciklama': 'Benign prostat'},
               {'kod': '280', 'aciklama': 'Hipertansiyon'},
               {'kod': '367', 'aciklama': 'X'}]
            """
            if not uyari_text:
                return []
            sonuc = []
            for parca in str(uyari_text).split(" | "):
                p = parca.strip()
                if not p:
                    continue
                # "(Reç: ...)" sarmasını kaldır
                if p.startswith("(Reç:"):
                    p = p[len("(Reç:"):].strip()
                    if p.endswith(")"):
                        p = p[:-1].strip()
                # "256 - Açıklama" formatı
                if " - " in p:
                    kod, _, aciklama = p.partition(" - ")
                    kod = kod.strip()
                    aciklama = aciklama.strip()
                else:
                    # Sadece kod (sayı) veya sadece açıklama
                    kod = p if p.isdigit() else ""
                    aciklama = "" if p.isdigit() else p
                if kod or aciklama:
                    sonuc.append({"kod": kod, "aciklama": aciklama})
            return sonuc

        def _karar_birlestir(durumlar: list) -> str:
            """Bir satırdaki birden çok uyarı kodunun kararlarını birleştir."""
            if not durumlar:
                return ""
            if "UYGUN DEĞİL" in durumlar:
                return "UYGUN DEĞİL"
            if all(d == "UYGUN" for d in durumlar):
                return "UYGUN"
            return "ŞÜPHELİ"

        for s in self.tum_satirlar:
            # Önceki çalıştırmadan kalan alanları temizle
            for k in ("verdict_uyari_kontrol", "verdict_uk_detay",
                      "verdict_uk_kodlar", "verdict_uk_kaynaklar",
                      "verdict_uk_eslesenler", "verdict_uk_oran",
                      "verdict_uk_ozel"):
                s[k] = ""

            uyari_text = s.get("uyari") or ""
            uyari_listesi = _uyari_parse(uyari_text)
            if not uyari_listesi:
                sayac["_kodsuz"] += 1
                continue

            # recete_tarama.uyari_kodu_kontrol formatına çevir
            uyari_kodlari_param = []
            for uk in uyari_listesi:
                uyari_kodlari_param.append({
                    "kod": uk["kod"],
                    "aciklama": uk["aciklama"],
                    "ilac_adi": s.get("ilac") or "",
                    "etkin_madde": s.get("etkin") or "",
                })

            recete_teshisleri = _bol(s.get("rec_tesh"))
            rapor_tanilari = _bol(s.get("rap_tesh"))
            # Reçete açıklaması + Rapor açıklaması birleşik
            aciklamalar = _bol(s.get("rec_ack")) + _bol(s.get("rap_ack"))
            doktor_brans = s.get("brans") or ""

            try:
                sonuclar = uyari_kodu_kontrol(
                    uyari_kodlari=uyari_kodlari_param,
                    recete_teshisleri=recete_teshisleri,
                    rapor_aciklamalari=aciklamalar,
                    rapor_tanilari=rapor_tanilari,
                    doktor_uzmanligi=doktor_brans,
                )
            except Exception as e:
                logger.warning("Uyarı kodu kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict_uyari_kontrol"] = "ŞÜPHELİ"
                s["verdict_uk_detay"] = f"Hata: {e}"
                sayac["ŞÜPHELİ"] += 1
                denetlenen.append(s)
                continue

            # Her kod için satır-bazlı karar
            kod_durumlar = []
            kod_detaylar = []
            kodlar_str = []
            kaynaklar_str = []
            eslesenler_str = []
            oranlar_str = []
            ozel_str = []
            for r in sonuclar:
                kod = r.get("kod", "")
                aciklama = r.get("aciklama", "")
                durum = r.get("durum", "")
                oran = float(r.get("eslesen_oran") or 0)
                kaynak = r.get("eslesen_kaynak", "") or ""
                eslesen_metin = r.get("eslesen_metin", "") or ""
                ozel = r.get("_ozel_kural", "") or ""

                if durum == "UYGUN":
                    kd = "UYGUN"
                elif durum == "EDILEMEDI":
                    kd = "ŞÜPHELİ"
                elif durum == "UYGUNSUZ":
                    kd = "UYGUN DEĞİL" if oran <= 0 else "ŞÜPHELİ"
                else:
                    kd = "ŞÜPHELİ"
                kod_durumlar.append(kd)
                kod_detaylar.append(
                    f"[{kod}] {aciklama[:50]} → {kd} "
                    f"(oran %{oran*100:.0f}, kaynak: {kaynak or '-'})"
                )
                kodlar_str.append(kod or "?")
                kaynaklar_str.append(kaynak or "-")
                eslesenler_str.append(eslesen_metin or "-")
                oranlar_str.append(f"%{oran*100:.0f}")
                if ozel:
                    ozel_str.append(f"[{kod}] {ozel}")

            karar = _karar_birlestir(kod_durumlar)
            s["verdict_uyari_kontrol"] = karar
            s["verdict_uk_detay"] = " ; ".join(kod_detaylar)
            s["verdict_uk_kodlar"] = ", ".join(kodlar_str)
            s["verdict_uk_kaynaklar"] = " | ".join(kaynaklar_str)
            s["verdict_uk_eslesenler"] = " | ".join(eslesenler_str)
            s["verdict_uk_oran"] = " | ".join(oranlar_str)
            s["verdict_uk_ozel"] = " ; ".join(ozel_str)
            sayac[karar] = sayac.get(karar, 0) + 1
            denetlenen.append(s)

        self._tabloyu_yenile()
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"])
        self._durum_yaz(
            f"Uyarı kodu kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"(uyarı kodsuz {sayac['_kodsuz']} satır boş bırakıldı)"
        )

        if toplam == 0:
            messagebox.showinfo(
                "Uyarı Kodu Kontrol",
                "Bu dönemde uyarı kodu girilmiş reçete satırı bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Uyarı kodu kontrolü tamamlandı.\n\n"
            f"Toplam denetlenen   : {toplam}\n"
            f"  ✓ UYGUN           : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL     : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ         : {sayac['ŞÜPHELİ']}\n"
            f"Uyarı kodsuz (atlanan): {sayac['_kodsuz']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._uyari_kod_rapor_excel_olustur(
                sayac=sayac, denetlenen_satirlar=denetlenen)
        except Exception as e:
            logger.exception("Uyarı kodu rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    def _uyari_kod_rapor_excel_olustur(self, *, sayac: dict,
                                          denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne uyarı kodu kontrol Excel
        raporu yazar.

        3 sayfa:
          - Özet : Toplam, sonuç dağılımı, en sık görülen kodlar, notlar
          - Detaylı Kontrol : Her satır + uyarı kodları + aranan ibareler +
                              eşleşen kaynak/metin + sonuç (renkli)
          - Şüpheli & Eşleşmeyen : Manuel inceleme önerilen satırlar
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Uyari_Kodu_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"])

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"
        ws1.cell(row=1, column=1,
                 value="UYARI KODU KONTROL RAPORU "
                       "(reçete uyarı kodu ↔ teşhis/açıklama eşleşmesi)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="E65100")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        # En sık görülen kodlar
        kod_sayac: dict = {}
        for s in denetlenen_satirlar:
            for k in (s.get("verdict_uk_kodlar") or "").split(","):
                k = k.strip()
                if k:
                    kod_sayac[k] = kod_sayac.get(k, 0) + 1
        en_sik_kodlar = sorted(kod_sayac.items(),
                                key=lambda x: x[1], reverse=True)[:10]

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Uyarı Kodu Olan Satır", str(toplam)),
            ("Uyarı Kodu Olmayan (Atlanan)", str(sayac["_kodsuz"])),
            ("", ""),
            ("Kontrol Yöntemi",
             "recete_tarama.uyari_kodu_kontrol — kodun açıklamasındaki "
             "anahtar kelimeler reçete teşhisi / rapor teşhisi / reçete "
             "açıklaması / rapor açıklaması içinde aranır."),
            ("Eşleşme Eşiği",
             "Tam kelime eşleşmesi 1.0, kök eşleşmesi 0.7 puan; "
             "ortalama oran ≥ 0.30 → UYGUN, "
             "0 < oran < 0.30 → ŞÜPHELİ, oran = 0 → UYGUN DEĞİL."),
            ("Özel Kurallar",
             "Kod 272 (TİTCK EK-4/A endikasyon dışı), "
             "Gabapentin/Pregabalin 'nöropatik ağrı' kuralı, "
             "Solifenasin/Fesoterodin oksibutin intoleransı, "
             "Antidepresan F30-F49 ICD eşleşmesi."),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="E65100")
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # SONUÇ DAĞILIMI
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL", "ŞÜPHELİ"],
                                     start=bas + 1):
            adet = sayac.get(etiket, 0)
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # EN SIK GÖRÜLEN KODLAR
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="EN SIK GÖRÜLEN UYARI KODLARI (TOP 10)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Kod", "Görülme Sayısı", "", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, (kod, adet) in enumerate(en_sik_kodlar, start=bas2 + 1):
            ws1.cell(row=i, column=1, value=kod).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=adet
                     ).alignment = Alignment(horizontal="center")

        # NOTLAR
        bas3 = bas2 + len(en_sik_kodlar) + 3
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = Uyarı kodu açıklamasındaki anahtar kelimeler "
            "reçete/rapor teşhisi veya açıklamasında bulundu.",
            "• UYGUN DEĞİL = Hiçbir kaynakta eşleşme yok — manuel inceleme önerilir.",
            "• ŞÜPHELİ = Kısmi eşleşme var ama eşik altı; veya özel kural "
            "(272 EK-4/A, antidepresan vb.) ile karar verilemedi.",
            "• 'Detaylı Kontrol' sayfasında her satırın hangi kodu "
            "kontrol ettiği, neyi aradığı, hangi kaynakta ne bulduğu yazılır.",
            "• Bir satırda birden çok uyarı kodu varsa: en az 1 UYGUN DEĞİL "
            "→ UYGUN DEĞİL; hepsi UYGUN → UYGUN; karışık → ŞÜPHELİ.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            c = ws1.cell(row=i, column=1, value=n)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: DETAYLI KONTROL ──────────
        ws2 = wb.create_sheet("Detaylı Kontrol")
        kolonlar = [
            ("rec_tar",                 "Reç.Tarih",       12),
            ("rec_no",                  "Reçete No",       18),
            ("hasta",                   "Hasta",           24),
            ("tc",                      "TC",              13),
            ("doktor",                  "Doktor",          22),
            ("brans",                   "Branş",           18),
            ("ilac",                    "İlaç",            28),
            ("etkin",                   "Etken Madde",     22),
            ("atc",                     "ATC",             10),
            ("rap_kod",                 "Rapor Kod",       11),
            ("uyari",                   "Uyarı Kodları (girilmiş)", 30),
            ("verdict_uk_kodlar",       "Kod Listesi",     16),
            ("rec_tesh",                "Reçete Teşhis",   28),
            ("rap_tesh",                "Rapor Teşhis",    28),
            ("rec_ack",                 "Reçete Açıklama", 28),
            ("rap_ack",                 "Rapor Açıklama",  28),
            ("verdict_uk_kaynaklar",    "Eşleşen Kaynak",  20),
            ("verdict_uk_eslesenler",   "Bulunan Metin",   28),
            ("verdict_uk_oran",         "Eşleşme Oranı",   13),
            ("verdict_uk_ozel",         "Özel Kural",      28),
            ("verdict_uk_detay",        "Karar Detayı",    50),
            ("verdict_uyari_kontrol",   "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict_uyari_kontrol") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center",
                                             vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: ŞÜPHELİ & UYGUN DEĞİL ──────────
        ws3 = wb.create_sheet("Şüpheli & Uygun Değil")
        ws3.cell(row=1, column=1,
                 value=("Aşağıdaki satırlarda manuel inceleme önerilir: "
                        "uyarı kodunun açıklamasındaki anahtar kelimeler "
                        "teşhis/açıklamada bulunamadı veya kısmi eşleşme "
                        "eşik altında kaldı.")
                 ).font = Font(italic=True, color="546E7A")
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        atl_kolonlar = [
            ("rec_tar",               "Reç.Tarih", 12),
            ("rec_no",                "Reçete No", 18),
            ("hasta",                 "Hasta",     24),
            ("ilac",                  "İlaç",      28),
            ("verdict_uk_kodlar",     "Kodlar",    14),
            ("uyari",                 "Açıklamaları", 30),
            ("verdict_uk_kaynaklar",  "Kaynak",    18),
            ("verdict_uk_detay",      "Karar Detayı", 50),
            ("verdict_uyari_kontrol", "SONUÇ",     14),
        ]
        for c, (_k, b, _g) in enumerate(atl_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in denetlenen_satirlar:
            v = s.get("verdict_uyari_kontrol") or ""
            if v not in ("ŞÜPHELİ", "UYGUN DEĞİL"):
                continue
            for ci, (kod, _b, _g) in enumerate(atl_kolonlar, 1):
                cell = ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            renk = VERDICT_RENK.get(v)
            if renk:
                vc = ws3.cell(row=ri, column=len(atl_kolonlar))
                vc.fill = PatternFill("solid", fgColor=renk)
                vc.font = Font(bold=True)
                vc.alignment = Alignment(horizontal="center")
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atl_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── NÖROPATİK / FİBROMİYALJİ KONTROLÜ (SUT 4.2.35.A + B) ─────────
    def _noropatik_435_kontrol_baslat(self):
        """NÖROPATİK 4.2.35 KONTROL butonu — yüklenen satırlardan
        Pregabalin / Gabapentin / Duloksetin / Alfa lipoik / Kapsaisin
        krem'i SUT 4.2.35.A (nöropatik ağrı) ve 4.2.35.B (fibromiyalji)
        kuralına göre denetler.

        Endikasyon-bazlı dispatch (kontrol_noropatik_4_2_35 içinde):
          - Pregab/Gabap + epilepsi   → ATLANDI (4.2.25 antiepileptik butonu)
          - Duloksetin + depresyon    → ATLANDI (4.2.2 psikiyatri butonu)
          - Pregab + Gabap aynı reçete → UYGUN_DEGIL (kombi yasağı)
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Nöropatik 4.2.35 Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_noropatik_4_2_35, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_kapsam_disi": 0}
        kategori_sayac = {"PREGABALIN": 0, "GABAPENTIN": 0, "DULOKSETIN": 0,
                          "ALFA_LIPOIK": 0, "KAPSAISIN": 0}
        denetlenen_satirlar = []

        # Aynı reçeteye ait satırları gruplama (Pregab+Gabap kombi yasağı için)
        rec_grup = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no") or ""
            rec_grup.setdefault(rno, []).append(s)

        for s in self.tum_satirlar:
            kategori = self._noropatik_435_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin önceki artığını sil
                if s.get("verdict_kategori") == "NOROPATIK_435":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_kapsam_disi"] += 1
                continue

            alt_grup = self._noropatik_435_alt_grup(
                s.get("ilac"), s.get("etkin"))
            kategori_sayac[alt_grup] = kategori_sayac.get(alt_grup, 0) + 1

            rno = s.get("rec_no") or ""
            ayni_recete = rec_grup.get(rno, [])
            ilac_sonuc = self._ilac_sonuc_olustur_noropatik_435(s, ayni_recete)

            try:
                rapor = kontrol_noropatik_4_2_35(ilac_sonuc)
            except Exception as e:
                logger.warning("Nöropatik 4.2.35 kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "NOROPATIK_435"
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(
                rapor, ek_icd, ['DM', 'INME'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = "NOROPATIK_435"
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        self._tabloyu_yenile()
        self._durum_yaz(
            f"Nöropatik SUT 4.2.35 kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_kapsam_disi']})"
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "Nöropatik 4.2.35 Kontrol",
                "Bu dönemde SUT 4.2.35 (nöropatik/fibromiyalji) "
                "grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Nöropatik SUT 4.2.35 kontrolü tamamlandı.\n\n"
            f"Toplam denetlenen satır : {toplam}\n"
            f"  ✓ UYGUN              : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL        : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ            : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI            : {sayac['ATLANDI']}\n\n"
            f"Alt grup dağılımı:\n"
            f"  • Pregabalin    : {kategori_sayac.get('PREGABALIN', 0)}\n"
            f"  • Gabapentin    : {kategori_sayac.get('GABAPENTIN', 0)}\n"
            f"  • Duloksetin    : {kategori_sayac.get('DULOKSETIN', 0)}\n"
            f"  • Alfa lipoik   : {kategori_sayac.get('ALFA_LIPOIK', 0)}\n"
            f"  • Kapsaisin     : {kategori_sayac.get('KAPSAISIN', 0)}\n\n"
            f"Kapsam dışı (atlanan) : {sayac['_kapsam_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._noropatik_435_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Nöropatik 4.2.35 rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    def _statin_kontrol_baslat(self):
        """STATİN KONTROL butonu — yüklenen satırlardan statin/lipid (ATC C10*)
        olanları SUT 4.2.28.A/B kuralına göre denetler ve sonucu en sağdaki
        SONUÇ sütununa yazar.

        ÖNEMLİ: Hasta'nın bu reçetede sadece statinin kendi raporu değil, TÜM
        aktif raporlarındaki ICD kodları da risk faktörü tespiti için
        kullanılır (DM/KAH/inme ayrı bir raporda olabilir)."""
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Statin Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_statin, kontrol_fibrat,
                _diger_rapor_notunu_uyariya_ekle,
            )
            from recete_kontrol.base_kontrol import KontrolSonucu
            from recete_kontrol.eski_rapor_kontrol import (
                eski_rapor_statin_kontrol_calistir,
            )
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        # Hastaların TÜM raporlarındaki ICD kodlarını topla (cross-rapor risk)
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"Statin kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        # Eski rapor (istinaden) sayaçları
        eski_rapor_istatistik = {
            "kontrol_edilen": 0,    # rap_ack'te referans bulunan satır sayısı
            "bulunan": 0,           # eski rapor EOS'ta bulundu
            "bulunamadi": 0,        # bulunamadı (Medula kontrolü gerekir)
            "override_uygun": 0,
            "override_uygun_degil": 0,
            "override_supheli": 0,
        }

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_lipid_disi": 0}
        # Rapor için kategoriye göre ayrım da tutulur
        kategori_sayac = {"STATIN": 0, "FIBRAT": 0, "DIGER_LIPID": 0}
        denetlenen_satirlar = []  # rapor için (lipid olanlar)

        # Önceki çalıştırmadan kalan verdict'leri temizle (lipid olmayanlar
        # her zaman boş kalsın, lipid olanlar yeniden hesaplansın)
        for s in self.tum_satirlar:
            kategori = self._lipid_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin (lipid) önceki artığını sil —
                # diyabet vb. başka kontrollerin verdict'lerine dokunma.
                if s.get("verdict_kategori") in ("STATIN", "FIBRAT", "DIGER_LIPID"):
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_lipid_disi"] += 1
                continue
            kategori_sayac[kategori] = kategori_sayac.get(kategori, 0) + 1
            ilac_sonuc = self._ilac_sonuc_olustur(s)
            # Hastanın diğer aktif raporlarındaki ICD/rapor kodlarını AYRI tut.
            # Ana karar reçeteye ilintili rapora göre verilir; diğer
            # raporlardaki bilgi sadece UYGUN_DEĞİL/ŞÜPHELİ sonuçlarda eczacı
            # için bilgi notu olarak rapora eklenir (kullanıcı kuralı 2026-05-07).
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)
            try:
                if kategori == "FIBRAT":
                    rapor = kontrol_fibrat(ilac_sonuc)
                else:
                    # STATIN ve DIGER_LIPID için aynı kural (LDL/risk mantığı)
                    rapor = kontrol_statin(ilac_sonuc)
            except Exception as e:
                logger.warning("Statin kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = kategori
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue
            _diger_rapor_notunu_uyariya_ekle(
                rapor, ek_icd,
                ['LIPID', 'DM', 'KAH', 'INME', 'PAH', 'KY', 'KBH'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            detay = rapor.mesaj or ""
            uyari = rapor.uyari or ""

            # ── TIER 2 UYARISI: rapor kodu eşleşmedi, ATC fallback'i kullanıldı
            if s.get("rapor_secim_kaynagi") == "tier2":
                tier2_uyari = (
                    "Reçetedeki rapor kodu hastanın aktif raporlarıyla "
                    "doğrudan eşleşmedi — ATC'ye uygun ICD'li rapor "
                    "kullanıldı (manuel kontrol önerilir)"
                )
                uyari = (uyari + " | " if uyari else "") + tier2_uyari

            # ── ESKİ RAPOR (İSTİNADEN) KONTROLÜ — sadece STATIN/DIGER_LIPID ──
            # FIBRAT bu kapsam dışında (kullanıcı: SADECE statin için).
            #
            # KOŞULLAR (kullanıcı kuralı):
            #   1) Şimdiki rapor zaten UYGUN ise → eski rapora HİÇ bakma
            #      (LDL ve risk faktörleri yeterli, idame iddiası anlamsız).
            #   2) Şimdiki rapor UYGUN DEĞİL veya ŞÜPHELİ ise → rapor
            #      açıklamasında "idame / devam raporu / istinaden / tarihli
            #      rapor / eski rapor" gibi anahtar kelime varsa eski rapora
            #      bakılır. Anahtar kelime yoksa pattern fonksiyonu zaten
            #      boş döner.
            simdiki_uygun = (rapor.sonuc == KontrolSonucu.UYGUN)
            if (kategori in ("STATIN", "DIGER_LIPID")
                    and not simdiki_uygun):
                rap_ack = s.get("rap_ack") or ""
                if rap_ack and self.db:
                    try:
                        eski_sonuc = eski_rapor_statin_kontrol_calistir(
                            db=self.db,
                            musteri_id=s.get("musteri_id") or 0,
                            rap_ack=rap_ack,
                            yeni_ilac_adi=s.get("ilac") or "",
                            yeni_rapor_ana_id=s.get("rapor_ana_id"),
                            yeni_rapor_tarihi=None,  # rap_tarih satırda yok
                            tolerans_gun=3,
                        )
                    except Exception as e:
                        logger.warning(
                            "Eski rapor kontrol fail (rx %s): %s",
                            s.get("rec_no"), e)
                        eski_sonuc = None

                    if eski_sonuc and eski_sonuc.get("referans_var"):
                        eski_rapor_istatistik["kontrol_edilen"] += 1
                        fd = eski_sonuc["final_durum"]
                        eski_mesaj = eski_sonuc.get("mesaj") or ""

                        if fd == "bulunamadi":
                            eski_rapor_istatistik["bulunamadi"] += 1
                            etiket = "ŞÜPHELİ"
                        elif fd == "uygun":
                            eski_rapor_istatistik["bulunan"] += 1
                            eski_rapor_istatistik["override_uygun"] += 1
                            etiket = "UYGUN"
                        elif fd == "uygun_degil":
                            eski_rapor_istatistik["bulunan"] += 1
                            eski_rapor_istatistik["override_uygun_degil"] += 1
                            etiket = "UYGUN DEĞİL"
                        elif fd == "supheli":
                            eski_rapor_istatistik["bulunan"] += 1
                            eski_rapor_istatistik["override_supheli"] += 1
                            etiket = "ŞÜPHELİ"

                        # Açıklamayı yapılandır: önce şimdiki, sonra eski rapor
                        # (eski rapor referansı tespit edildiğinde
                        # ŞİMDİKİ/ESKİ etiketleriyle ayır — kullanıcı her iki
                        # raporun durumunu da net görsün)
                        parcalar = []
                        if detay:
                            parcalar.append(f"[ŞİMDİKİ RAPOR] {detay}")
                        if eski_mesaj:
                            parcalar.append(eski_mesaj)
                        detay = " || ".join(parcalar)

            s["verdict"] = etiket
            s["verdict_detay"] = detay
            s["verdict_kategori"] = kategori
            s["verdict_uyari"] = uyari
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            # detaylar dict — string'e çevir (Excel'e güvenle yazılsın)
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        # Tabloyu yenile (mevcut filtre/renk durumuna saygı duyarak)
        self._tabloyu_yenile()
        eski_ek = ""
        if eski_rapor_istatistik["kontrol_edilen"] > 0:
            eski_ek = (
                f" | Eski rapor: {eski_rapor_istatistik['kontrol_edilen']} "
                f"satırda referans bulundu "
                f"(✓{eski_rapor_istatistik['bulunan']} bulundu, "
                f"✗{eski_rapor_istatistik['bulunamadi']} bulunamadı)"
            )
        self._durum_yaz(
            f"Statin/Lipid SUT kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(lipid dışı {sayac['_lipid_disi']} satır boş bırakıldı)"
            + eski_ek
        )

        # ── KONTROL RAPORU ÜRET ──
        toplam_lipid = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                        + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam_lipid == 0:
            messagebox.showinfo(
                "Statin Kontrol",
                "Bu dönemde statin/lipid grubuna giren reçete bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Statin/Lipid SUT kontrolü tamamlandı.\n\n"
            f"Toplam lipid satırı : {toplam_lipid}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Lipid dışı (atlanan): {sayac['_lipid_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._statin_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Statin rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        # Aç
        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── ARB KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────────────────
    def _arb_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                   denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne ARB SUT denetim raporu yazar.

        3 sayfa:
          - Özet         : Toplam sayım, mono/kombi dağılımı, çalışma zamanı
          - ARB Reçeteleri : Denetlenen satır + SUT detayları + verdict
          - ARB Dışı     : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"ARB_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam_arb = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                       + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="ARB SUT EK-4/F MADDE 51 (1300/51) KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="0D47A1")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("ARB Olarak Tespit Edilen", str(toplam_arb)),
            ("ARB Dışı (Atlanan)", str(sayac["_arb_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı",
             "SUT EK-4/F Madde 51 (1300/51) — ARB ve kombinasyonları"),
            ("Filtreleme",
             "ATC C09C* (mono ARB) + C09D* (ARB kombi) + C02AC* (santral) "
             "+ etken/ticari isim fallback"),
            ("Kapsam",
             "İrbesartan, Kandesartan, Losartan, Telmisartan, Valsartan, "
             "Olmesartan, Eprosartan, Rilmeniden, Moksonidin + kombiler"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="0D47A1")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_arb * 100) if toplam_arb else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Mono / Kombi dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ARB TİPİ DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Tip", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "ARB_MONO": "ATC C09CA / C02AC — tek başına ARB veya santral antihipertansif "
                        "(rapor varsa doz/plan/süre yazma zorunluluğu yok)",
            "ARB_KOMBI": "ATC C09D* — ARB + diüretik / + CCB / + 3'lü "
                         "(raporda monoterapi yetersizliği ibaresi gerekli)",
        }
        for i, k in enumerate(["ARB_MONO", "ARB_KOMBI"], start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + 4
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• Mono ARB raporlu → UYGUN (doz/plan/süre yazma zorunluluğu yok)",
            "• Kombi ARB raporlu → 'monoterapi yetersizliği' ibaresi raporda olmalı",
            "• Raporsuz ARB → AYDA EN FAZLA 1 KUTU + AİLE HEKİMLERİNCE yazılabilir",
            "• Raporsuz + uzman hekim ya da >1 kutu → UYGUN DEĞİL",
            "• ŞÜPHELİ = monoterapi ibaresi parse edilemedi veya branş bilinmiyor — manuel kontrol",
            "• ARB dışı = ATC C09C/C09D/C02AC dışında olduğu için bu butonun kapsamına girmiyor",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: ARB REÇETELERİ ──────────
        ws2 = wb.create_sheet("ARB Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           20),
            ("kurum_adi",        "Kurum",           28),
            ("tesis_kodu",       "Tesis Kodu",      11),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_kategori", "Tip",             10),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar",    38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: ARB DIŞI ──────────
        ws3 = wb.create_sheet("ARB Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC C09C/C09D/C02AC dışında olduğu "
                       "için ARB butonu KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._arb_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── ANTİHT KONTROL RAPORU EXCEL ÜRETİCİ ──────────────────────────
    def _antiht_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                      denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne ANTİHT SUT denetim raporu yazar.

        3 sayfa:
          - Özet           : Toplam sayım, mono/kombi dağılımı, çalışma zamanı
          - ANTİHT Reçeteleri : Denetlenen satır + SUT detayları + verdict
          - ANTİHT Dışı   : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"ANTIHT_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="01579B")  # koyu mavi
        toplam_antiht = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                          + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="ANTİHT SUT 4.2.12.B (Mono + Kombi Antihipertansif) KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="01579B")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("ANTİHT Olarak Tespit Edilen", str(toplam_antiht)),
            ("ANTİHT Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı",
             "SUT 4.2.12.B — Mono ve kombine antihipertansif kullanım"),
            ("Filtreleme",
             "ATC C09A* (ACE-i) / C09C* (ARB) / C03* (Diüretik) / "
             "C07* (BB) / C08* (KKB) — mono. C09B*/C09D* — kombi. "
             "Etken madde + SUT_MADDESI_KATEGORI mapping fallback."),
            ("Kapsam (Mono)",
             "ACE-i (ramipril/perindopril/lisinopril vb.), ARB (telmisartan/"
             "valsartan/losartan vb.), KKB (amlodipin/lerkanidipin vb.), "
             "BB (bisoprolol/metoprolol vb.), Diüretik, Alfa bloker"),
            ("Kapsam (Kombi)",
             "ACE-i+Diüretik / ARB+Diüretik / ACE-i+KKB / ARB+KKB / "
             "üçlü kombiler (C09B*/C09D*)"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="01579B")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_antiht * 100) if toplam_antiht else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Mono / Kombi dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ANTİHT TİPİ DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Tip", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "MONO":  "Tek başına antihipertansif (ACE-i / ARB / KKB / BB / "
                     "Diüretik / Alfa bloker) — raporsuz uyumluluk + ACE+ARB "
                     "kontrendikasyon + doz aşımı kontrolü",
            "KOMBI": "Sabit kombine antihipertansif (ATC C09B*/C09D*) — "
                     "raporsuz / rapor 04.05 / monoterapi yetersizliği "
                     "ibaresi 3-yollu kontrol",
        }
        for i, k in enumerate(["MONO", "KOMBI"], start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + 4
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• Mono raporsuz → genelde UYGUN (alt grup-bazlı uyarılar var: "
            "ACE-i/ARB doz aşımı, ACE+ARB kontrendikasyon)",
            "• Mono raporlu (rapor kodu 04.05) → UYGUN — antihipertansiyon raporu",
            "• Kombi raporsuz → UYGUN (kombine etken madde monoterapi "
            "yetersizliğinin örtük kanıtı sayılır)",
            "• Kombi raporlu (04.05) → UYGUN — antihipertansiyon raporu mevcut",
            "• Kombi raporlu (başka kod) → 'monoterapi' / 'tek ilaç yetersiz' "
            "ibaresi raporda olmalı; yoksa ŞÜPHELİ",
            "• ŞÜPHELİ = monoterapi ibaresi parse edilemedi veya rapor kodu "
            "antihipertansiyon değil — manuel kontrol",
            "• Kapsam dışı = ATC C09/C03/C07/C08 dışında olduğu için bu butonun "
            "kapsamına girmiyor",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: ANTİHT REÇETELERİ ──────────
        ws2 = wb.create_sheet("ANTİHT Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           20),
            ("kurum_adi",        "Kurum",           28),
            ("tesis_kodu",       "Tesis Kodu",      11),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_kategori", "Tip",             10),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar",    38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: ANTİHT DIŞI ──────────
        ws3 = wb.create_sheet("ANTİHT Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar mono/kombi antihipertansif "
                       "kapsamında olmadığı için ANTİHT butonu KAPSAMI "
                       "DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._antiht_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── İSKEMİK KALP KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────
    def _iskemik_kalp_rapor_excel_olustur(
            self, *, sayac: dict, kategori_sayac: dict,
            denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne İSKEMİK KALP SUT denetim
        raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, ivabradin/ranolazin/eplerenon dağılımı
          - İSKEMİK KALP Reçeteleri : Denetlenen satır + SUT detay + verdict
          - İSKEMİK KALP Dışı : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"ISKEMIK_KALP_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="B71C1C")  # koyu kırmızı
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="İSKEMİK KALP SUT 4.2.15.C/F (İvabradin + Ranolazin + Eplerenon) KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="B71C1C")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("İSKEMİK KALP Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı",
             "SUT 4.2.15.C (İvabradin/Eplerenon) + SUT 4.2.15.F (Ranolazin)"),
            ("Filtreleme",
             "ATC C01EB17 (İvabradin) / C01EB18 (Ranolazin) / "
             "C03DA04 (Eplerenon) — etken madde fallback."),
            ("Kapsam (İvabradin)",
             "NYHA II-IV + EF≤45% + sinüs ritmi + BB intoleransı veya "
             "yetersiz yanıt + uzman branş raporu"),
            ("Kapsam (Ranolazin)",
             "Kronik stabil angina + BB veya CCB intoleransı/yetersiz yanıt"),
            ("Kapsam (Eplerenon)",
             "İvabradin'le benzer koşullar (KY + EF≤45%) — kontrol_ivabradin "
             "üzerinden delege"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="B71C1C")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="İLAÇ DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["İlaç", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "IVABRADIN": "ATC C01EB17 — NYHA II-IV + EF≤45% + sinüs ritmi + "
                         "BB intoleransı/yetersiz yanıt (kontrol_ivabradin)",
            "RANOLAZIN": "ATC C01EB18 — kronik stabil angina + BB/CCB "
                         "intoleransı veya yetersiz yanıt (kontrol_ranolazin)",
            "EPLERENON": "ATC C03DA04 — KY + EF≤45% (kontrol_ivabradin'e delege "
                         "edilir, koşullar benzer)",
        }
        for i, k in enumerate(["IVABRADIN", "RANOLAZIN", "EPLERENON"],
                              start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• İvabradin → kalp yetmezliği (NYHA II-IV) + EF≤45% + sinüs "
            "ritmi + BB intoleransı veya yetersiz yanıt (KH ≥75 atım/dk)",
            "• Ranolazin → kronik stabil angina + BB/CCB intoleransı veya "
            "yetersiz semptom kontrolü",
            "• Eplerenon → İvabradin koşullarına benzer (KY + EF≤45%); "
            "kontrol_ivabradin'e delege edilir",
            "• Tüm üçü için raporlu olmak ZORUNLU; raporsuz → UYGUN DEĞİL",
            "• ŞÜPHELİ = NYHA/EF/BB-intolerans ibaresi parse edilemedi — manuel kontrol",
            "• Kapsam dışı = ATC C01EB17/18 + C03DA04 dışında olduğu için bu "
            "butonun kapsamına girmiyor",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2 ──────────
        ws2 = wb.create_sheet("İSKEMİK KALP Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           20),
            ("kurum_adi",        "Kurum",           28),
            ("tesis_kodu",       "Tesis Kodu",      11),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_kategori", "Tip",             10),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar",    38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3 ──────────
        ws3 = wb.create_sheet("İSKEMİK KALP Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar İvabradin/Ranolazin/Eplerenon "
                       "kapsamında olmadığı için İSKEMİK KALP butonu "
                       "KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._iskemik_kalp_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── K-SİTRAT KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────────
    def _potasyum_sitrat_rapor_excel_olustur(
            self, *, sayac: dict, kategori_sayac: dict,
            denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne K-SİTRAT (Potasyum sitrat —
        SUT EK-4/F) denetim raporu yazar. 3 sayfa: Özet / Reçeteler /
        K-SİTRAT Dışı."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"K_SITRAT_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="0277BD")
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="K-SİTRAT (Potasyum Sitrat) SUT EK-4/F KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="0277BD")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("K-SİTRAT Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı", "SUT EK-4/F — Potasyum Sitrat"),
            ("Filtreleme",
             "ATC A12BA02 + etken madde 'POTASYUM SITRAT'"),
            ("Kapsam",
             "Üroloji veya Nefroloji uzmanı raporu zorunlu. ICD: N20.0/N20.1 "
             "(böbrek taşı), N25.8/N25.9 (renal tübüler asidoz). "
             "Aile hekimi yazımı UYGUN DEĞİL."),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="0277BD")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        bas3 = bas + 6
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• Üroloji veya Nefroloji uzmanı raporu ZORUNLU",
            "• Aile hekimi / pratisyen yazımı → UYGUN DEĞİL",
            "• ICD: N20.0/N20.1 (böbrek taşı) veya N25.8/N25.9 (RTA)",
            "• Raporsuz → ŞÜPHELİ (rapor metni okunamadığı için manuel)",
            "• Kapsam dışı = ATC A12BA02 / Potasyum Sitrat değil",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2 ──────────
        ws2 = wb.create_sheet("K-SİTRAT Reçeteleri")
        kolonlar = [
            ("rec_tar", "Reç.Tarih", 12), ("rec_no", "Reçete No", 18),
            ("hasta", "Hasta", 24), ("tc", "TC", 13),
            ("yas", "Yaş", 6), ("cins", "Cin.", 6),
            ("doktor", "Doktor", 22), ("brans", "Branş", 20),
            ("kurum_adi", "Kurum", 28), ("tesis_kodu", "Tesis Kodu", 11),
            ("ilac", "İlaç", 28), ("etkin", "Etken Madde", 22),
            ("atc", "ATC", 10), ("rap_kod", "Rapor Kod", 11),
            ("rec_doz", "Reçete Doz", 14), ("kutu", "Kutu", 6),
            ("medula_msj", "Medula Msj", 30),
            ("rec_tesh", "Reçete Teşhis", 30), ("rap_tesh", "Rapor Teşhis", 30),
            ("rec_ack", "Reçete Açıklama", 30), ("rap_ack", "Rapor Açıklama", 30),
            ("verdict_sut", "Uygulanan SUT", 30),
            ("verdict_aranan", "Aranan İbare", 28),
            ("verdict_bulunan", "Bulunan Metin", 28),
            ("verdict_detaylar", "SUT Detaylar", 38),
            ("verdict_uyari", "Uyarı", 34),
            ("verdict_detay", "Açıklama", 50),
            ("verdict", "SONUÇ", 14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3 ──────────
        ws3 = wb.create_sheet("K-SİTRAT Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar potasyum sitrat kapsamında "
                       "olmadığı için bu butonun KAPSAMI DIŞINDA "
                       "bırakıldı.").font = Font(italic=True, color="546E7A")
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12), ("rec_no", "Reçete No", 18),
            ("hasta", "Hasta", 24), ("ilac", "İlaç", 30),
            ("etkin", "Etken", 22), ("atc", "ATC", 10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._potasyum_sitrat_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── RAPORSUZ BİLGİ KONTROL RAPORU EXCEL ÜRETİCİ ──────────────────
    def _raporsuz_bilgi_rapor_excel_olustur(
            self, *, sayac: dict, kategori_sayac: dict,
            denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne RAPORSUZ BİLGİ denetim
        raporu yazar. 3 sayfa: Özet / Reçeteler / RAPORSUZ Dışı."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"RAPORSUZ_BILGI_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="455A64")  # gri
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="RAPORSUZ BİLGİ (Raporsuz verilebilen ilaçlar) KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="455A64")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("RAPORSUZ Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan Kontrol", "kontrol_raporsuz_bilgilendirme — 28 alt sınıf dispatcher"),
            ("Filtreleme",
             "RAPORSUZ_BILGILENDIRME / BENZODIAZEPIN / GOZ_LUBRIKAN "
             "kategorileri (sut_kategorisi_tespit_et mapping'inde tanınan)"),
            ("Kapsam",
             "NSAİD, parasetamol, makrolid, penisilin, sefalosporin, "
             "antihistaminik, dekonjestan, MAO-B, PPI, H2RA, antidiyareik, "
             "laksatif, antispazmodik, kas gevşetici, topikal NSAİD, "
             "topikal kortikosteroid, oftalmik/otik/nazal, mukolitik, "
             "antitüsif, vitamin, oral antifungal, antiemetik, vb."),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="455A64")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        bas3 = bas + 6
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• Bu butonun amacı RAPORSUZ verilebilen ilaçları toplu görmek + "
            "doz/etkileşim/yan etki uyarılarını listelemek (UYGUN olsalar bile)",
            "• kontrol_raporsuz_bilgilendirme her ilaç için detaylı uyarı "
            "döndürür — verdict_uyari sütununa bakın",
            "• Bazı raporsuz ilaçlar (Augmentin gibi kombi antibiyotikler, "
            "PPI'lar) sut_kategorisi_tespit_et mapping'inde tanınmadığı için "
            "bu butona dahil olmayabilir — kapsam sınırlı",
            "• ŞÜPHELİ = doz/etkileşim parse edilemedi — manuel kontrol",
            "• Kapsam dışı = RAPORSUZ_BILGILENDIRME / BENZODIAZEPIN / "
            "GOZ_LUBRIKAN kategorilerinden hiçbirine düşmedi",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2 ──────────
        ws2 = wb.create_sheet("RAPORSUZ Reçeteleri")
        kolonlar = [
            ("rec_tar", "Reç.Tarih", 12), ("rec_no", "Reçete No", 18),
            ("hasta", "Hasta", 24), ("tc", "TC", 13),
            ("yas", "Yaş", 6), ("cins", "Cin.", 6),
            ("doktor", "Doktor", 22), ("brans", "Branş", 20),
            ("kurum_adi", "Kurum", 28), ("ilac", "İlaç", 28),
            ("etkin", "Etken Madde", 22), ("atc", "ATC", 10),
            ("kutu", "Kutu", 6),
            ("medula_msj", "Medula Msj", 30),
            ("rec_tesh", "Reçete Teşhis", 30),
            ("rec_ack", "Reçete Açıklama", 30),
            ("verdict_sut", "Uygulanan Kural", 30),
            ("verdict_uyari", "Uyarı / Etkileşim / Doz", 60),
            ("verdict_detay", "Açıklama", 50),
            ("verdict", "SONUÇ", 14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3 ──────────
        ws3 = wb.create_sheet("RAPORSUZ Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar raporsuz bilgilendirme kapsamında "
                       "olmadığı (veya mapping'de tanınmadığı) için bu "
                       "butonun KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12), ("rec_no", "Reçete No", 18),
            ("hasta", "Hasta", 24), ("ilac", "İlaç", 30),
            ("etkin", "Etken", 22), ("atc", "ATC", 10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._raporsuz_bilgi_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── BASIT SUT KONTROL RAPORU EXCEL ÜRETİCİ ───────────────────────
    def _basit_sut_rapor_excel_olustur(
            self, *, sayac: dict, kategori_sayac: dict,
            denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne BASIT SUT (DMAH+Kadın+
        Trimetazidin+Kinolon) denetim raporu yazar. 3 sayfa: Özet /
        Reçeteler / BASIT SUT Dışı."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"BASIT_SUT_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="558B2F")  # yeşil
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="BASIT SUT (DMAH + Kadın Hormon + Trimetazidin + Florokinolon) KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="558B2F")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("BASIT SUT Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "DMAH (4.2.7) + Kadın Hormon (4.2.29) + Trimetazidin "
             "(4.2.14.C) + Florokinolon (EK-4/E)"),
            ("Filtreleme",
             "ATC: B01AB* (DMAH) / G03C/D/F* (Kadın hormon) / "
             "C01EB15 (Trimetazidin) / J01MA* (Florokinolon)"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="558B2F")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="KATEGORİ DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Kategori", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "DMAH":                     "DMAH/Enoksaparin (SUT 4.2.7) — rapor + uyarı kodu (280/281) kontrolü",
            "KADIN_HORMON":             "Kadın cinsiyet hormonları (SUT 4.2.29) — rapor var/yok + süre uyarısı",
            "TRIMETAZIDIN":             "Trimetazidin (SUT 4.2.14.C) — rapor zorunlu",
            "ANTIBIYOTIK_FLOROKINOLON": "Florokinolon (EK-4/E) — raporsuz verilebilir + kültür antibiyogram önerisi",
        }
        for i, k in enumerate(["DMAH", "KADIN_HORMON",
                                "TRIMETAZIDIN", "ANTIBIYOTIK_FLOROKINOLON"],
                              start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        bas3 = bas2 + 6
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• 4 farklı SUT kategorisi tek butonda — hepsi <15 satırlık binary kontroller",
            "• DMAH → rapor + uyarı kodu (280/281) zorunlu",
            "• Kadın hormon → rapor zorunlu, uzun süreli kullanım uyarısı",
            "• Trimetazidin → rapor zorunlu",
            "• Florokinolon → raporsuz verilebilir (ayaktan tedavi), kültür antibiyogram önerisi",
            "• ŞÜPHELİ = rapor metni okunamadı — manuel kontrol",
            "• Kapsam dışı = bu 4 kategoriden hiçbirine düşmedi",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2 ──────────
        ws2 = wb.create_sheet("BASIT SUT Reçeteleri")
        kolonlar = [
            ("rec_tar", "Reç.Tarih", 12), ("rec_no", "Reçete No", 18),
            ("hasta", "Hasta", 24), ("tc", "TC", 13),
            ("yas", "Yaş", 6), ("cins", "Cin.", 6),
            ("doktor", "Doktor", 22), ("brans", "Branş", 20),
            ("kurum_adi", "Kurum", 28), ("ilac", "İlaç", 28),
            ("etkin", "Etken Madde", 22), ("atc", "ATC", 10),
            ("verdict_kategori", "Kategori", 18),
            ("rap_kod", "Rapor Kod", 11),
            ("kutu", "Kutu", 6), ("uyari", "Uyarı Kod", 18),
            ("medula_msj", "Medula Msj", 30),
            ("rec_tesh", "Reçete Teşhis", 30),
            ("verdict_sut", "Uygulanan SUT", 30),
            ("verdict_uyari", "Uyarı", 34),
            ("verdict_detay", "Açıklama", 50),
            ("verdict", "SONUÇ", 14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3 ──────────
        ws3 = wb.create_sheet("BASIT SUT Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar DMAH/Kadın hormon/Trimetazidin/"
                       "Florokinolon kapsamında olmadığı için bu butonun "
                       "KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12), ("rec_no", "Reçete No", 18),
            ("hasta", "Hasta", 24), ("ilac", "İlaç", 30),
            ("etkin", "Etken", 22), ("atc", "ATC", 10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._basit_sut_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── ÇEŞİTLİ KONTROL RAPORU EXCEL ÜRETİCİ ─────────────────────────
    def _cesitli_rapor_excel_olustur(self, *, sayac: dict,
                                       kategori_sayac: dict,
                                       denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne ÇEŞİTLİ SUT denetim raporu yazar.

        3 sayfa:
          - Özet                : Toplam, alt grup dağılımı, sonuç dağılımı
          - Çeşitli Reçeteler   : Denetlenen satır + verdict + detay
          - Kapsam Dışı         : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Cesitli_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                   + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="ÇEŞİTLİ İLAÇLAR SUT KONTROL RAPORU "
                       "(M.45 üriner / M.2 gözyaşı / BPH α-bloker)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="4A148C")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Çeşitli Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "M.45 (üriner inkontinans) / M.2 (suni gözyaşı) / BPH α-bloker"),
            ("Filtreleme",
             "ATC G04BD* (üriner) + G04CA* (BPH) + S01XA/S01KA (gözyaşı) "
             "+ N06AX21 (Duloksetin) + etken/ticari isim fallback"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="4A148C")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt grup dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ALT GRUP DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Grup", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "URINER":  "M.45 — Antimuskarinik (Solifenasin/Tolterodin/...) + "
                       "Mirabegron + Duloksetin (kadın+SUI)",
            "GOZYASI": "M.2 — Suni gözyaşı / kuru göz / keratitis sicca "
                       "(Hyaluronat/Hipromelloz/Karmelloz vb.)",
            "BPH":     "BPH α-bloker — Alfuzosin/Tamsulosin/Terazosin/"
                       "Doksazosin/Silodosin",
        }
        for i, k in enumerate(["URINER", "GOZYASI", "BPH"], start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "── ÜRİNER (M.45) ──",
            "• Antimuskarinik+Mirabegron raporlu + Nöroloji/FTR/Üroloji/Pediatri/"
            "Geriatri/Kadın-Doğum uzm → UYGUN",
            "• Antimuskarinik kombi (Mirabegron HARİÇ) → UYGUN DEĞİL",
            "• Duloksetin: yalnız erişkin kadın + stres SUI/mikst SUI + raporlu",
            "  (yanlış endikasyon/cinsiyet → ŞÜPHELİ — Medula şart kontrolü)",
            "── GÖZYAŞI (M.2) ──",
            "• Göz Hastalıkları uzm raporsuz/raporlu → UYGUN",
            "• 1 yıl göz uzm raporu + diğer uzm → UYGUN (doz max 7 damla/gün — Medula)",
            "• Aile hekimi raporsuz + ayda ≤1 kutu → UYGUN; >1 kutu → UYGUN DEĞİL",
            "── BPH α-BLOKER ──",
            "• Üroloji uzm hekim → UYGUN",
            "• Üroloji 1 yıl raporu + tüm hekimler → UYGUN",
            "• Kombi yasağı: α-blokerler birlikte ödenmez (HT eşlikse istisna serbest)",
            "── GENEL ──",
            "• ŞÜPHELİ = endikasyon/branş tespit edilemedi — manuel kontrol gerekli",
            "• ATLANDI = ÇEŞİTLİ kapsamında olmayan ilaç (kategori boş bırakıldı)",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: ÇEŞİTLİ REÇETELERİ ──────────
        ws2 = wb.create_sheet("Çeşitli Reçeteler")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           20),
            ("kurum_adi",        "Kurum",           28),
            ("tesis_kodu",       "Tesis Kodu",      11),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_kategori", "Alt Grup",        12),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar",    38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ÇEŞİTLİ (M.45/M.2/BPH) kapsamında "
                       "olmadığı için bu butonun KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._cesitli_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── PSİKİYATRİ/NÖROLOJİ KONTROL RAPORU EXCEL ÜRETİCİ ─────────────
    def _psikiyatri_rapor_excel_olustur(self, *, sayac: dict,
                                          kategori_sayac: dict,
                                          alt_sayac: dict,
                                          denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne SUT 4.2.2 + 4.2.25 denetim
        raporu yazar.

        3 sayfa:
          - Özet                       : Toplam, kategori/alt-grup dağılımı
          - Psikiyatri/Nöroloji Reçeteleri : Denetlenen satır + verdict
          - Kapsam Dışı                : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Psikiyatri_Noroloji_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="4A148C")  # mor
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="PSİKİYATRİ / NÖROLOJİ KONTROL RAPORU "
                       "(SUT 4.2.2 + 4.2.25)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="4A148C")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Denetlenen Satır", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.2 (Antipsikotik/Antidepresan/Mood Stab/Benzodiazepin) + "
             "SUT 4.2.25 (Antiepileptik)"),
            ("Filtreleme",
             "ATC N06A* (antidepresan) + N05A* (antipsikotik) + "
             "N05B* (anksiyolitik) + N03A* (antiepileptik) + "
             "etken/ticari isim fallback"),
            ("Endikasyon Bazlı Dispatch",
             "Lamotrijin/Valproat: bipolar → 4.2.2 psikiyatri; "
             "epilepsi → 4.2.25 antiepileptik"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="4A148C")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Kategori dağılımı (PSIKIYATRI / ANTIEPILEPTIK)
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="KATEGORİ DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Kategori", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "PSIKIYATRI":   "ATC N05A/N06A/N05B — Antipsikotik, antidepresan, "
                            "anksiyolitik, mood stab (SUT 4.2.2)",
            "ANTIEPILEPTIK": "ATC N03A — Yeni nesil/Pregabalin/Lakozamid/"
                             "Gabapentin/Zonisamit (SUT 4.2.25)",
        }
        for i, k in enumerate(["PSIKIYATRI", "ANTIEPILEPTIK"],
                              start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Alt grup dağılımı
        bas3 = bas2 + 4
        ws1.cell(row=bas3, column=1, value="ALT GRUP DAĞILIMI")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        bas3 += 1
        for col, hd in enumerate(["Alt Grup", "Adet", "", ""], start=1):
            c = ws1.cell(row=bas3, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        # Sıralı alt grup listesi
        alt_sira = [
            "ANTIDEPRESAN_SSRI", "ANTIDEPRESAN_SNRI", "ANTIDEPRESAN_TRISIKLIK",
            "ANTIDEPRESAN_ATIPIK", "ANTIDEPRESAN_BVA",
            "ANTIPSIKOTIK_TIPIK", "ANTIPSIKOTIK_ATIPIK",
            "BENZODIAZEPIN", "MOOD_STAB_LITYUM",
            "AEP_YENI_NESIL", "AEP_ZONISAMIT", "AEP_PREGABALIN",
            "AEP_LAKOZAMID", "AEP_GABAPENTIN", "AEP_VALPROAT",
            "DIGER",
        ]
        ri = bas3 + 1
        for k in alt_sira:
            adet = alt_sayac.get(k, 0)
            if adet == 0:
                continue
            ws1.cell(row=ri, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=ri, column=2, value=adet
                     ).alignment = Alignment(horizontal="center")
            ri += 1

        # Notlar
        bas4 = ri + 2
        ws1.cell(row=bas4, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas4, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas4, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas4, start_column=1,
                        end_row=bas4, end_column=4)
        notlar = [
            "• SUT 4.2.2(1): Trisiklik/SSRI tüm hekimlerce — 1 yıl sonra psikiyatri raporu",
            "• SUT 4.2.2(1): SNRI/SSRE/RIMA/NASSA — psik/nöroloji/geriatri uzmanı/raporu",
            "• SUT 4.2.2(1): Bupropion/Vortioksetin/Agomelatin — sadece major depresif",
            "• SUT 4.2.2(2): Atipik antipsikotikler — psikiyatri uzmanı/raporu (klozapin maks 1 ay)",
            "• SUT 4.2.2(4): Tipik antipsikotikler — rapor kısıtlamasına tabi DEĞİL",
            "• SUT 4.2.2(5): Acil servis — atipik parenteral (depot hariç) tek doz tüm hekim",
            "• SUT 4.2.2(6): Demans → atipik antipsikotik psik/nöro/geriatri uzmanı/raporu",
            "• SUT 4.2.2(7): Valproat bipolar — psik/nöroloji uzmanı/raporu",
            "• SUT 4.2.2(8): Lamotrijin bipolar — psikiyatri uzmanı/raporu",
            "• SUT 4.2.25(1): Yeni nesil AEP (Lamotrijin/Topiramat/Vigabatrin/Levetirasetam) — nöro/beyin cer.",
            "• SUT 4.2.25(2): Zonisamit — nöroloji uzmanı + 1 YIL süreli rapor",
            "• SUT 4.2.25(3): Pregabalin — 2./3. basamak SHS + en az 1 nöroloji + 1 yıl SK raporu",
            "•                YAB endikasyonunda ÖDENMEZ; Gabapentin ile kombi YASAK",
            "• SUT 4.2.25(4): Lakozamid — 16+ yaş, 2 AEP 6 ay yetersiz, parsiyel epilepsi, nöro raporu",
            "• SUT 4.2.25(5): Gabapentin — 2./3. basamak SHS + en az 1 nöroloji + 1 yıl SK raporu",
            "•                Pregabalin ile kombi YASAK",
            "• ATLANDI = Lamotrijin/Valproat bipolar — psikiyatri akışında işlendi (çift sayım yok)",
            "• ŞÜPHELİ = SK raporu / branş / ek şartlar parse edilemedi — manuel kontrol",
            "• Kapsam dışı = ATC N03A/N05A/N05B/N06A dışında — bu butonun kapsamına girmiyor",
        ]
        for i, n in enumerate(notlar, start=bas4 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: PSIKIYATRI/NOROLOJI REÇETELERİ ──────────
        ws2 = wb.create_sheet("Psikiyatri Nöroloji Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           20),
            ("kurum_adi",        "Kurum",           28),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_kategori", "Kategori",        14),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("rec_turu",         "Reç.Türü",        10),
            ("rec_alt_turu",     "Reç.Alt",         10),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar",    38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri2, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri2, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri2, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC N03A/N05A/N05B/N06A dışında "
                       "olduğu için psikiyatri/nöroloji butonu KAPSAMI DIŞINDA "
                       "bırakıldı.").font = (Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri3 = 4
        for s in self.tum_satirlar:
            kategori = self._psikiyatri_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri3, column=ci, value=str(s.get(kod, "")))
            ri3 += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── NÖROPATİK 4.2.35 KONTROL RAPORU EXCEL ÜRETİCİ ────────────────
    def _noropatik_435_rapor_excel_olustur(self, *, sayac: dict,
                                             kategori_sayac: dict,
                                             denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne SUT 4.2.35 (nöropatik/
        fibromiyalji) denetim raporu yazar.

        3 sayfa:
          - Özet                : Toplam, alt grup dağılımı, SUT madde
          - Nöropatik Reçeteleri: Denetlenen satır + verdict
          - Kapsam Dışı         : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Noropatik_4_2_35_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="880E4F")  # koyu pembe
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="NÖROPATİK / FİBROMİYALJİ KONTROL RAPORU "
                       "(SUT 4.2.35.A + B)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="880E4F")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Denetlenen Satır", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_kapsam_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.35.A (Nöropatik ağrı) + 4.2.35.B (Fibromiyalji)"),
            ("Filtreleme",
             "Etken/ticari isim: Pregabalin / Gabapentin / Duloksetin / "
             "Alfa lipoik (Tioktik) / Kapsaisin krem"),
            ("Endikasyon Bazlı Dispatch",
             "Pregab/Gabap + epilepsi → 4.2.25 antiepileptik (ATLANDI); "
             "Duloksetin + depresyon → 4.2.2 psikiyatri (ATLANDI)"),
            ("Kombi Yasağı",
             "Pregabalin + Gabapentin AYNI REÇETE — UYGUN DEĞİL "
             "(cross-reçete tarama YOK)"),
            ("Rapor Süresi",
             "Fibromiyalji ≤ 6 ay; nöropatik ağrı ≤ 12 ay (regex parse)"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="880E4F")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt grup dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ALT GRUP DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Grup", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        alt_aciklama = {
            "PREGABALIN":  "LYRICA, GABRICA, PREGALIN, PREGABEX — SK raporu + "
                            "nöroloji uzmanı (4.2.35 kapsamında nöropatik/fibromiyalji)",
            "GABAPENTIN":  "NEURONTIN, NERUDA, GABATEVA, GABALEPT — SK raporu + "
                            "nöroloji uzmanı (Pregab ile kombi YASAK)",
            "DULOKSETIN":  "CYMBALTA, DUXET, DULOXIN, DULOX — nöroloji/algoloji/"
                            "FTR/psikiyatri/endokrin/dahiliye",
            "ALFA_LIPOIK": "Tioktik / α-lipoik asit (THIOCTACID, NOREXIA) — "
                            "endokrin/nöroloji/dahiliye/FTR (sadece nöropatik)",
            "KAPSAISIN":   "Krem (CAPSIN, ZOSTRIX) — nöroloji/algoloji/"
                            "dermatoloji/FTR (sadece nöropatik)",
        }
        ri = bas2 + 1
        for k in ("PREGABALIN", "GABAPENTIN", "DULOKSETIN",
                  "ALFA_LIPOIK", "KAPSAISIN"):
            adet = kategori_sayac.get(k, 0)
            ws1.cell(row=ri, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=ri, column=2, value=adet
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=ri, column=3, value=alt_aciklama[k])
            ri += 1

        # Notlar
        bas4 = ri + 2
        ws1.cell(row=bas4, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas4, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas4, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas4, start_column=1,
                        end_row=bas4, end_column=4)
        notlar = [
            "• SUT 4.2.35.A: Nöropatik ağrı (PHN, diyabetik nöropati, "
            "nöropatik ağrı, spinal kord yaralanması, kanser ağrısı)",
            "• SUT 4.2.35.B: Fibromiyalji (M79.7) — rapor süresi en fazla 6 ay",
            "• Pregabalin/Gabapentin: 2./3. basamak SHS + en az 1 nöroloji "
            "uzmanı dahil sağlık kurulu raporu",
            "• Pregabalin + Gabapentin AYNI REÇETEDE birlikte yazılamaz",
            "• Duloksetin (4.2.35.A): nöroloji/algoloji/FTR/psikiyatri/"
            "endokrin/dahiliye uzman raporu",
            "• Duloksetin (4.2.35.B): romatoloji/FTR/nöroloji/algoloji/"
            "psikiyatri uzman raporu",
            "• Alfa lipoik / Kapsaisin krem: SADECE nöropatik ağrı "
            "endikasyonunda ödenir (fibromiyalji için ödenmez)",
            "• ATLANDI = Pregab/Gabap + epilepsi → 4.2.25 antiepileptik butonu;"
            " Duloksetin + depresyon → 4.2.2 psikiyatri butonu",
            "• ŞÜPHELİ = Branş / SK raporu / endikasyon parse edilemedi — "
            "manuel kontrol gerekli",
            "• Kapsam dışı = SUT 4.2.35 kapsamına giren ilaç değil",
        ]
        for i, n in enumerate(notlar, start=bas4 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: NÖROPATİK REÇETELER ──────────
        ws2 = wb.create_sheet("Nöropatik 4.2.35 Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           20),
            ("kurum_adi",        "Kurum",           28),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar",    38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri2, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri2, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri2, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar SUT 4.2.35 (nöropatik/fibromiyalji) "
                       "kapsamı dışında olduğu için atlandı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri3 = 4
        for s in self.tum_satirlar:
            kategori = self._noropatik_435_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri3, column=ci, value=str(s.get(kod, "")))
            ri3 += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── HEPATİT B/C KONTROL RAPORU EXCEL ÜRETİCİ ─────────────────────
    def _hepatit_rapor_excel_olustur(self, *, sayac: dict,
                                       kategori_sayac: dict,
                                       denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne hepatit B/C SUT denetim
        raporu yazar.

        3 sayfa:
          - Özet         : Toplam sayım, HBV/HCV/Klasik dağılımı, SUT madde
          - Hepatit Reçeteleri : Denetlenen satır + SUT detayları + verdict
          - Hepatit Dışı : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Hepatit_BC_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam_hep = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                      + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="HEPATİT B/C SUT KONTROL RAPORU "
                       "(SUT 4.2.13.3 / 4.2.13.4)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="1B5E20")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Hepatit Olarak Tespit Edilen", str(toplam_hep)),
            ("Hepatit Dışı (Atlanan)", str(sayac["_hepatit_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.13.3 (Kronik Hepatit B) · "
             "SUT 4.2.13.4 (Kronik Hepatit C - DAA)"),
            ("Filtreleme",
             "ATC J05AF (HBV) + J05AP (HCV DAA) + J05AB04 (Ribavirin) "
             "+ L03AB10/11 (Peginterferon) + etken/ticari isim fallback"),
            ("Kapsam — HBV",
             "Entekavir, Tenofovir TDF/TAF, Lamivudin, Telbivudin, Adefovir"),
            ("Kapsam — HCV",
             "Sofosbuvir, Ledipasvir, Velpatasvir, Voxilaprevir, "
             "Glecaprevir, Pibrentasvir, Ombitasvir, Paritaprevir, "
             "Dasabuvir, Daklatasvir, Elbasvir, Grazoprevir"),
            ("Kapsam — Klasik",
             "Peginterferon alfa-2a/2b (Pegasys, PegIntron), Ribavirin"),
            ("KAPSAM DIŞI",
             "HIV ilaçları (Dolutegravir/Abacavir/Efavirenz vb.) — "
             "ANTIVIRAL kategorisinde kalır, ayrı kontrol"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="1B5E20")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı tablosu
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_hep * 100) if toplam_hep else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Kategori dağılımı (HBV / HCV / KLASIK)
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="HEPATİT TİPİ DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Tip", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "HBV":    "ATC J05AF — Kronik Hepatit B nükleos(t)id analogları "
                      "(SUT 4.2.13.3)",
            "HCV":    "ATC J05AP — HCV DAA (Direct-Acting Antivirals) "
                      "(SUT 4.2.13.4)",
            "KLASIK": "Peginterferon alfa + Ribavirin — klasik HBV/HCV "
                      "tedavisi (artık nadir)",
        }
        for i, k in enumerate(["HBV", "HCV", "KLASIK"], start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• HBV (4.2.13.3): HBsAg pozitif + HBV DNA "
            "(HBeAg+ ≥20.000 IU/mL, HBeAg- ≥2.000 IU/mL) + ALT yüksekliği "
            "veya karaciğer biyopsisi.",
            "• HCV (4.2.13.4): Anti-HCV pozitif + HCV RNA pozitif + genotip "
            "+ fibrozis evresi (METAVIR F0-F4). Tedavi süresi 8/12/16/24 hafta.",
            "• Uzman branş: Gastroenteroloji / Enfeksiyon Hastalıkları / "
            "Hepatoloji (HBV için İç Hastalıkları da kabul).",
            "• Rapor kodu pragmatik fallback (06.01 / 14.* / B*) → Medula "
            "endikasyon/lab şart kontrolünü yapmış kabul edilir.",
            "• ŞÜPHELİ = endikasyon var ama lab değeri/uzman branş ibaresi "
            "metinden parse edilemedi — manuel kontrol önerilir.",
            "• UYGUN DEĞİL = raporsuz veya zorunlu alan tamamen eksik.",
            "• Hepatit dışı = HBV/HCV/Klasik kapsamına girmeyen satırlar "
            "(HIV ilaçları dahil) — bu butonun KAPSAMI DIŞINDA.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: HEPATİT REÇETELERİ ──────────
        ws2 = wb.create_sheet("Hepatit Reçeteleri")
        kolonlar = [
            ("rec_tar",           "Reç.Tarih",       12),
            ("rec_no",            "Reçete No",       18),
            ("hasta",             "Hasta",           24),
            ("tc",                "TC",              13),
            ("yas",               "Yaş",             6),
            ("cins",              "Cin.",            6),
            ("doktor",            "Doktor",          22),
            ("brans",             "Branş",           20),
            ("kurum_adi",         "Kurum",           28),
            ("ilac",              "İlaç",            28),
            ("etkin",             "Etken Madde",     22),
            ("atc",               "ATC",             10),
            ("verdict_kategori",  "Tip",             8),
            ("verdict_alt_sinif", "Alt Sınıf",       16),
            ("rap_kod",           "Rapor Kod",       11),
            ("rec_doz",           "Reçete Doz",      14),
            ("rap_doz",           "Rapor Doz",       14),
            ("kutu",              "Kutu",            6),
            ("msj",               "Msj",             7),
            ("uyari",             "Uyarı Kod",       18),
            ("medula_msj",        "Medula Msj",      30),
            ("rec_tesh",          "Reçete Teşhis",   30),
            ("rap_tesh",          "Rapor Teşhis",    30),
            ("rec_ack",           "Reçete Açıklama", 30),
            ("rap_ack",           "Rapor Açıklama",  30),
            ("verdict_sut",       "Uygulanan SUT",   30),
            ("verdict_aranan",    "Aranan İbare",    32),
            ("verdict_bulunan",   "Bulunan Metin",   32),
            ("verdict_detaylar",  "SUT Detaylar",    38),
            ("verdict_uyari",     "Uyarı",           34),
            ("verdict_detay",     "Açıklama",        50),
            ("verdict",           "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center",
                                             vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: HEPATİT DIŞI ──────────
        ws3 = wb.create_sheet("Hepatit Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar HBV/HCV/Klasik kapsamına girmediği "
                       "için bu butonun KAPSAMI DIŞINDA bırakıldı "
                       "(HIV ilaçları dahil).").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._hepatit_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── ASTIM/KOAH KONTROL RAPORU EXCEL ÜRETİCİ ──────────────────────
    def _solunum_rapor_excel_olustur(self, *, sayac: dict,
                                       kategori_sayac: dict,
                                       denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne astım/KOAH SUT denetim
        raporu yazar.

        3 sayfa:
          - Özet            : Toplam sayım, alt-sınıf dağılımı, SUT 4.2.24
                              kuralı özeti, çalışma zamanı
          - Solunum Reçeteleri : Denetlenen satır + alt sınıf + SUT detayları
                                  + verdict (renkli)
          - Solunum Dışı    : Atlanan satırların kısa özeti
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Astim_KOAH_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam_sol = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                      + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="ASTIM / KOAH SUT KONTROL RAPORU "
                       "(SUT 4.2.24 / 4.2.24.B)")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="00838F")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Solunum Olarak Tespit Edilen", str(toplam_sol)),
            ("Solunum Dışı (Atlanan)", str(sayac["_solunum_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.24 (Solunum Sistemi İlaçları — Astım/KOAH) · "
             "SUT 4.2.24.B (Üçlü Kombinasyon — LABA+ICS+LAMA)"),
            ("Filtreleme",
             "ATC R03* (solunum sistemi obstrüktif hastalık ilaçları) "
             "+ etken/ticari isim fallback"),
            ("Kapsam — Kısa etkili",
             "SABA (Salbutamol/Terbutalin/Fenoterol — Ventolin, Bricanyl), "
             "SAMA (İpratropium — Atrovent), "
             "SABA+SAMA (Combivent, İpramol, Berodual)"),
            ("Kapsam — ICS/LABA/LAMA",
             "ICS tek (Pulmicort, Flixotide, Asmanex, Cortair neb), "
             "LABA tek (Foradil, Oxis, Onbrez), "
             "LABA+ICS (Seretide, Symbicort, Foster, Relvar, Inuvair, Airplus), "
             "LAMA (Spiriva, Incruse, Seebri, Eklira), "
             "LABA+LAMA (Anoro, Ultibro, Spiolto, Duaklir)"),
            ("Kapsam — Üçlü",
             "LABA+ICS+LAMA (Trelegy, Trimbow, Breztri, Trixeo, Enerzair) — "
             "SUT 4.2.24.B özel hüküm: 3 ay ICS+LABA başarısızlığı + ≥2 atak/yıl "
             "+ mMRC≥2 veya CAT≥10"),
            ("Kapsam — LTRA",
             "Montelukast/Zafirlukast (Singulair, Onceair, Notta, Levokast, "
             "Monkast) — astım veya alerjik rinit"),
            ("Kapsam — Biyolojik / sistemik",
             "Omalizumab/Anti-IgE (Xolair), Anti-IL5 (Nucala/Mepo, Fasenra/Benral), "
             "Anti-IL4/IL13 (Dupixent/Dupilumab), Roflumilast (Daxas), "
             "Teofilin/Aminofilin, Kromoglisik asit/Nedokromil"),
            ("Üçlü Tespiti",
             "Aynı reçetede LABA+ICS+LAMA üç ayrı ilaç olarak yazılmışsa da "
             "üçlü kullanım kabul edilir → SUT 4.2.24.B kuralı uygulanır"),
            ("Tanı Kaynağı",
             "Reçete teşhisi + rapor teşhisi + hastanın diğer aktif "
             "raporlarındaki ICD kodları (J45/J46 astım, J43/J44 KOAH, "
             "J30 alerjik rinit, J47 bronşiektazi)"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="00838F")
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_sol * 100) if toplam_sol else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt-sınıf dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ALT SINIF DAĞILIMI")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Sınıf", "Adet", "Açıklama", ""],
                                  start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        alt_aciklama = {
            "SABA":       "Kısa etkili β2 agonist — raporsuz (Salbutamol/Terbutalin)",
            "SAMA":       "Kısa etkili antikolinerjik — raporsuz (İpratropium)",
            "SABA+SAMA":  "Combivent / Berodual — raporsuz",
            "ICS":        "İnhale kortikosteroid tek — astım/KOAH + uzman raporu, 1 yıl",
            "LABA":       "Uzun etkili β2 agonist tek — astım/KOAH + uzman, 1 yıl "
                          "(astımda LABA tek ICS'siz önerilmez)",
            "LAMA":       "Uzun etkili antikolinerjik — YALNIZCA KOAH + göğüs hast., 1 yıl",
            "LABA+ICS":   "Seretide/Symbicort/Foster — astım/KOAH + uzman raporu, 1 yıl",
            "LABA+LAMA":  "Anoro/Ultibro/Spiolto — KOAH + göğüs hast., 1 yıl",
            "UCLU":       "LABA+ICS+LAMA üçlü (Trelegy/Trimbow) — KOAH + 3 ay "
                          "ICS+LABA yetersiz + ≥2 atak/yıl + mMRC≥2 veya CAT≥10, "
                          "göğüs hast., 1 yıl (SUT 4.2.24.B)",
            "LTRA":       "Montelukast/Zafirlukast — astım veya alerjik rinit + "
                          "iç hast./çocuk/göğüs/alerji, 1 yıl",
            "OMALIZUMAB": "Xolair (Anti-IgE) — ağır alerjik astım + IgE 30-1500 + "
                          "yüksek doz ICS+LABA başarısızlığı + ≥6 yaş + sağlık kurulu",
            "ANTI_IL":    "Nucala/Fasenra/Dupixent — ağır eozinofilik astım + "
                          "eozinofil eşik + sağlık kurulu (göğüs+alerji)",
            "ROFLUMILAST": "Daxas (PDE4 inh.) — KOAH + FEV1≤%50 + ≥2 atak/yıl + "
                           "kronik bronşit, göğüs hast., 6 ay rapor",
            "TEOFILIN":   "Teofilin/Aminofilin — raporsuz",
            "KROMOLIN":   "Kromoglisik asit/Nedokromil — raporsuz",
            "DIGER":      "Alt sınıf tespit edilemedi — manuel kontrol",
        }
        # Sadece adedi >0 olan alt sınıfları göster, en yüksekten düşüğe sırala
        sirali_alt = sorted(
            ((k, v) for k, v in kategori_sayac.items() if v > 0),
            key=lambda x: x[1], reverse=True)
        for i, (k, v) in enumerate(sirali_alt, start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=v
                     ).alignment = Alignment(horizontal="center")
            c3 = ws1.cell(row=i, column=3,
                          value=alt_aciklama.get(k, ""))
            c3.alignment = Alignment(wrap_text=True, vertical="top")

        # Notlar
        bas3 = bas2 + max(len(sirali_alt), 1) + 3
        ws1.cell(row=bas3, column=1, value="NOTLAR / SUT KURALI ÖZETİ")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• SABA/SAMA/Teofilin/Kromolin → raporsuz, tüm hekimler yazabilir.",
            "• ICS/LABA/LABA+ICS/LTRA → astım veya KOAH tanısı + uzman raporu "
            "(göğüs/alerji/iç hast./çocuk), 1 yıl.",
            "• LAMA tek ve LABA+LAMA ikili → SADECE KOAH endikasyonunda SGK "
            "kapsamındadır (astımda kabul EDİLMEZ).",
            "• LABA+ICS+LAMA üçlü (tek inhaler veya 3 ayrı ilaç): SUT 4.2.24.B "
            "→ 3 ay ICS+LABA yetersiz + ≥2 atak/yıl + mMRC≥2 veya CAT≥10, "
            "göğüs hast., 1 yıl.",
            "• Omalizumab → ağır alerjik astım + IgE 30-1500 IU/mL + yüksek doz "
            "ICS+LABA başarısızlığı + ≥6 yaş + sağlık kurulu (göğüs+alerji).",
            "• Anti-IL5 (Mepo eozinofil ≥150, Benral ≥300) / Anti-IL4/IL13 "
            "(Dupilumab) → ağır eozinofilik astım + sağlık kurulu.",
            "• Roflumilast (Daxas) → KOAH + FEV1≤%50 + ≥2 atak/yıl + kronik "
            "bronşit, göğüs hast., 6 AY rapor (diğer solunumdan farklı).",
            "• ŞÜPHELİ = tanı/uzman/eşik metinden parse edilemedi → manuel "
            "kontrol önerilir.",
            "• Solunum dışı = ATC R03* dışında olduğu için bu butonun "
            "KAPSAMI DIŞINDA bırakıldı.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            c = ws1.cell(row=i, column=1, value=n)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 70, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: SOLUNUM REÇETELERİ ──────────
        ws2 = wb.create_sheet("Solunum Reçeteleri")
        kolonlar = [
            ("rec_tar",           "Reç.Tarih",       12),
            ("rec_no",            "Reçete No",       18),
            ("hasta",             "Hasta",           24),
            ("tc",                "TC",              13),
            ("yas",               "Yaş",             6),
            ("cins",              "Cin.",            6),
            ("doktor",            "Doktor",          22),
            ("brans",             "Branş",           20),
            ("kurum_adi",         "Kurum",           28),
            ("ilac",              "İlaç",            28),
            ("etkin",             "Etken Madde",     22),
            ("atc",               "ATC",             10),
            ("verdict_alt_sinif", "Alt Sınıf",       14),
            ("rap_kod",           "Rapor Kod",       11),
            ("rec_doz",           "Reçete Doz",      14),
            ("rap_doz",           "Rapor Doz",       14),
            ("kutu",              "Kutu",            6),
            ("msj",               "Msj",             7),
            ("uyari",             "Uyarı Kod",       18),
            ("medula_msj",        "Medula Msj",      30),
            ("rec_tesh",          "Reçete Teşhis",   30),
            ("rap_tesh",          "Rapor Teşhis",    30),
            ("rec_ack",           "Reçete Açıklama", 30),
            ("rap_ack",           "Rapor Açıklama",  30),
            ("verdict_sut",       "Uygulanan SUT",   34),
            ("verdict_aranan",    "Aranan İbare",    32),
            ("verdict_bulunan",   "Bulunan Metin",   32),
            ("verdict_detaylar",  "SUT Detaylar",    44),
            ("verdict_uyari",     "Uyarı",           34),
            ("verdict_detay",     "Açıklama",        50),
            ("verdict",           "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center",
                                             vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: SOLUNUM DIŞI ──────────
        ws3 = wb.create_sheet("Solunum Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC R03* (solunum sistemi obstrüktif "
                       "hastalık ilaçları) kapsamı dışında olduğu için bu "
                       "butonun KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._solunum_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ── KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────────────────────
    def _statin_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                      denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne kapsamlı Excel raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, kategori dağılımı, çalışma zamanı, dönem
          - Lipid Reçeteleri : Denetlenen her satır + SUT detayları + verdict
          - Lipid Dışı : Atlanan satırların kısa özeti (sadece sayım/listeleme)

        Returns: oluşturulan dosyanın tam yolu.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        # Hedef klasör: ~/Desktop/Reçete Kontrol  (OneDrive\Desktop varsa o)
        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Statin_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",  # yeşil
            "UYGUN DEĞİL": "FFCDD2",  # kırmızı
            "ŞÜPHELİ":     "FFE0B2",  # turuncu
            "ATLANDI":     "ECEFF1",  # gri
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="263238")
        toplam_lipid = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                        + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1, value="STATİN / LİPİD SUT KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="0D47A1")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Lipid Olarak Tespit Edilen", str(toplam_lipid)),
            ("Lipid Dışı (Atlanan)", str(sayac["_lipid_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.28.A (Statin) · SUT 4.2.28.B (Fibrat)"),
            ("Filtreleme",
             "ATC C10* (lipid modifying agents) + ticari isim fallback"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="0D47A1")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı tablosu
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)

        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_lipid * 100) if toplam_lipid else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Kategori dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="KATEGORİ DAĞILIMI (lipid satırları)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Kategori", "Adet", "Açıklama", ""],
                                  start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "STATIN":       "ATC C10AA/C10BA/C10BX — atorvastatin, rosuvastatin, simvastatin, pravastatin, fluvastatin, pitavastatin + statin kombinleri",
            "FIBRAT":       "ATC C10AB — fenofibrat (Lipanthyl/Lipantil/Tralip/Lipofen), gemfibrozil, bezafibrat",
            "DIGER_LIPID":  "ATC C10AC/AD/AX — ezetimib tek başına, kolestiramin, omega-3, PCSK9 (Repatha/Praluent)",
        }
        for i, k in enumerate(["STATIN", "FIBRAT", "DIGER_LIPID"],
                               start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = SUT kuralı algoritmik olarak doğrulandı (LDL eşik + risk faktörü vb.)",
            "• UYGUN DEĞİL = Algoritmik kural net şekilde ihlal — manuel inceleme önerilir",
            "• ŞÜPHELİ = Karar verilemedi (eksik LDL/TG değeri, eksik teşhis vb.) — manuel kontrol",
            "• ATLANDI = SUT denetimi gerekmeyen satır",
            "• Lipid dışı = ATC C10* dışında olduğu için bu butonun kapsamına girmiyor",
            "• 'Lipid Reçeteleri' sayfasında her satırın hangi metni okuduğu, "
            "neyi aradığı ve ne bulduğu yazılır.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        # Sütun genişlikleri (özet)
        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: LİPİD REÇETELERİ (denetlenen) ──────────
        ws2 = wb.create_sheet("Lipid Reçeteleri")
        # Sütunlar: tabloya yansıyanların hepsi + SUT denetim detayları
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           18),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_kategori", "Kategori",        12),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar (LDL/risk/...)", 38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        # Veri satırları
        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            # SONUÇ sütununu (en sondaki) renge boya
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        # Sütun genişlikleri
        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        # Otomatik filtre
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: LİPİD DIŞI (atlanmış) ──────────
        ws3 = wb.create_sheet("Lipid Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC C10* (lipid) sınıfına girmediği "
                       "için statin/lipid butonu KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        # Atlananları topla
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._lipid_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ───────────────────────────────────────────────────────────────────
    # DİYABET (SUT 4.2.38) SUT KONTROLÜ — DPP-4 / SGLT-2 / GLP-1 + klasik OAD + insülin
    # ───────────────────────────────────────────────────────────────────
    def _diyabet_kontrol_baslat(self):
        """DİYABET KONTROL butonu — yüklenen satırlardan diyabet ilaçları
        (ATC A10*) SUT 4.2.38'e göre denetlenir; sonuç en sağdaki SONUÇ
        sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ olarak yazılır.

        Kapsanan ilaç sınıfları:
          • Biguanid (metformin), Sülfonilüre, Glinid, Akarboz, TZD
          • DPP-4 inhibitörleri (sitagliptin, vildagliptin, ...)
          • SGLT-2 inhibitörleri (empa/dapa/kana/ertugliflozin)
          • GLP-1 RA (liraglutid, semaglutid, dulaglutid, ...)
          • İnsülinler (kısa/orta/uzun etkili, kombi)
          • Sabit kombiler (DPP4+met, SGLT2+met, SGLT2+DPP4)

        SUT 4.2.38(7): GLP-1 RA → BMI ≥ 30 + HbA1c ≥ %7 + metformin maks doz
        SUT 4.2.38(8): DPP-4 + GLP-1 RA birlikte ödenmez (kombinasyon yasağı)
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Diyabet Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_diyabet_dpp4_sglt2, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_diyabet_disi": 0}
        # Alt sınıf bazlı sayım (raporlama için)
        kategori_sayac = {"INSULIN": 0, "BIGUANID": 0, "SULFONILURE": 0,
                          "GLINID": 0, "TZD": 0, "AKARBOZ": 0,
                          "DPP4": 0, "SGLT2": 0, "GLP1": 0,
                          "KOMBI_OAD": 0, "DIGER": 0}
        denetlenen_satirlar = []

        # ── Reçete bazlı ilaç gruplaması (kombinasyon yasağı kontrolü için) ──
        # Aynı rec_no'ya sahip diğer ilaçların adları → DPP-4 + GLP-1 yasağı vb.
        recete_ilac_grup: Dict[str, List[str]] = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no")
            if rno:
                recete_ilac_grup.setdefault(str(rno), []).append(
                    s.get("ilac") or "")

        # ── Hastanın diğer raporlarındaki ICD/rapor kodlarını çek ──
        # (KY/KBH varsa SGLT-2 kuralı buna göre değişir)
        musteri_idler = []
        for s in self.tum_satirlar:
            mid = s.get("musteri_id")
            if mid and mid not in musteri_idler:
                musteri_idler.append(mid)
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        for s in self.tum_satirlar:
            kategori = self._diyabet_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin (DIYABET) önceki artığını sil —
                # statin/lipid verdict'lerine dokunma.
                if s.get("verdict_kategori") == "DIYABET":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_diyabet_disi"] += 1
                continue

            alt_sinif = self._diyabet_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            kategori_sayac[alt_sinif] = kategori_sayac.get(alt_sinif, 0) + 1

            # Aynı reçetedeki diğer ilaçların adları (kendisi hariç)
            rno = str(s.get("rec_no") or "")
            kendi_ad = (s.get("ilac") or "").upper()
            diger_adlar = [x for x in recete_ilac_grup.get(rno, [])
                            if x and x.upper() != kendi_ad]

            ilac_sonuc = self._ilac_sonuc_olustur_diyabet(s, diger_adlar)

            # Hastanın diğer aktif raporlarındaki ICD'leri AYRI tut. Ana sonucu
            # reçeteye ilintili rapor belirler; diğer raporlardaki bilgi sadece
            # uyarı/açıklama olarak rapora not edilir (kullanıcı kuralı 2026-05-07).
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                rapor = kontrol_diyabet_dpp4_sglt2(ilac_sonuc)
            except Exception as e:
                logger.warning("Diyabet kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "DIYABET"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(rapor, ek_icd, ['DM', 'KY', 'KBH'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            self._kontrol_raporunu_satira_yaz(s, rapor,
                                                kategori="DIYABET",
                                                alt_sinif=alt_sinif)
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        # Tabloyu yenile (mevcut filtre/renk durumuna saygı duyarak)
        self._tabloyu_yenile()
        self._durum_yaz(
            f"Diyabet SUT 4.2.38 kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(diyabet dışı {sayac['_diyabet_disi']} satır boş bırakıldı)"
        )

        toplam_diyabet = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                          + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam_diyabet == 0:
            messagebox.showinfo(
                "Diyabet Kontrol",
                "Bu dönemde diyabet ilacı (ATC A10*) bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Diyabet SUT 4.2.38 kontrolü tamamlandı.\n\n"
            f"Toplam diyabet satırı : {toplam_diyabet}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Diyabet dışı (atlanan): {sayac['_diyabet_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._diyabet_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Diyabet rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── DİYABET KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────────────
    def _diyabet_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                       denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne kapsamlı Excel raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, alt sınıf dağılımı, çalışma zamanı, dönem
          - Diyabet Reçeteleri : Denetlenen her satır + SUT detayları + verdict
          - Diyabet Dışı : Atlanan satırların kısa özeti

        Returns: oluşturulan dosyanın tam yolu.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Diyabet_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="0D47A1")
        toplam_diyabet = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                          + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1, value="DİYABET SUT 4.2.38 KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="0D47A1")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Diyabet Olarak Tespit Edilen", str(toplam_diyabet)),
            ("Diyabet Dışı (Atlanan)", str(sayac["_diyabet_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı",
             "SUT 4.2.38 — Diyabet (DPP-4 / SGLT-2 / GLP-1 RA + klasik OAD + insülin)"),
            ("Kapsam",
             "ATC A10* (insülin + oral antidiyabetikler) + ad/etken fallback"),
            ("Aranan İbare (DPP-4/SGLT-2/GLP-1)",
             "Metformin ve sülfonilürelerin maksimum tolere edilebilir "
             "dozlarında yeterli glisemik kontrol sağlanamamıştır"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="0D47A1")
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1, end_row=bas, end_column=4)
        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_diyabet * 100) if toplam_diyabet else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt sınıf dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ALT SINIF DAĞILIMI (diyabet ilaçları)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1, end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Sınıf", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        sinif_aciklama = {
            "INSULIN":     "ATC A10A — bazal/hızlı/orta/uzun etkili insülinler + kombi",
            "BIGUANID":    "ATC A10BA — METFORMIN (Glifor, Glukofen, Diaformin)",
            "SULFONILURE": "ATC A10BB — gliklazid, glimepirid, glibenklamid, glipizid",
            "GLINID":      "ATC A10BX — repaglinid (Novonorm), nateglinid (Starlix)",
            "TZD":         "ATC A10BG — pioglitazon (Actos)",
            "AKARBOZ":     "ATC A10BF — alfa-glukozidaz inhibitörü (Glucobay)",
            "DPP4":        "ATC A10BH — sitagliptin/Januvia, vildagliptin/Galvus, linagliptin/Trajenta",
            "SGLT2":       "ATC A10BK — empagliflozin/Jardiance, dapagliflozin/Forziga, kanagliflozin/Invokana",
            "GLP1":        "ATC A10BJ — liraglutid/Victoza, semaglutid/Ozempic, dulaglutid/Trulicity",
            "KOMBI_OAD":   "ATC A10BD — sabit kombiler (Janumet, Galvusmet, Synjardy, Xigduo, Glyxambi)",
            "DIGER":       "Yukarıdaki sınıflara girmeyen diyabet ilaçları",
        }
        sinif_sira = ["INSULIN", "BIGUANID", "SULFONILURE", "GLINID", "TZD",
                      "AKARBOZ", "DPP4", "SGLT2", "GLP1", "KOMBI_OAD", "DIGER"]
        for i, k in enumerate(sinif_sira, start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=sinif_aciklama[k])

        # Notlar
        bas3 = bas2 + len(sinif_sira) + 3
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1, end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = SUT 4.2.38 algoritmik olarak doğrulandı (rapor + glisemik şart + uzman/branş + lab)",
            "• UYGUN DEĞİL = Algoritmik kural net ihlal — manuel inceleme önerilir",
            "• ŞÜPHELİ = Karar verilemedi (eksik HbA1c/BMI/eGFR, eksik teşhis, vb.) — manuel kontrol",
            "• ATLANDI = SUT denetimi gerekmeyen satır (klasik OAD + diyabet tanısı + raporsuz)",
            "• Diyabet dışı = ATC A10* dışında olduğu için bu butonun kapsamına girmiyor",
            "• DPP-4/SGLT-2/GLP-1 için aranan ibare: 'metformin ve sülfonilürelerin maksimum "
            "tolere edilebilir dozlarında yeterli glisemik kontrol sağlanamamıştır'",
            "• GLP-1 RA için ek şart: BMI ≥ 30 + HbA1c ≥ %7 (rapor metninden taranır)",
            "• DPP-4 + GLP-1 RA aynı reçetede ÖDENMEZ (SUT 4.2.38(8) — kombinasyon yasağı)",
            "• 'Diyabet Reçeteleri' sayfasında her satırın hangi metni okuduğu, "
            "neyi aradığı ve ne bulduğu yazılır.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: DİYABET REÇETELERİ (denetlenen) ──────────
        ws2 = wb.create_sheet("Diyabet Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           18),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_alt_sinif", "Alt Sınıf",      12),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar (HbA1c/BMI/eGFR/...)", 38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: DİYABET DIŞI (atlanmış) ──────────
        ws3 = wb.create_sheet("Diyabet Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC A10* (diyabet) sınıfına girmediği "
                       "için diyabet butonu KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._diyabet_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ───────────────────────────────────────────────────────────────────
    # KLOPİDOGREL / PRASUGREL / TIKAGRELOR (SUT 4.2.15) SUT KONTROLÜ
    # ───────────────────────────────────────────────────────────────────
    def _klopidogrel_kontrol_baslat(self):
        """KLOPİDOGREL KONTROL butonu — yüklenen satırlardan klopidogrel/prasugrel/
        tikagrelor (ATC B01AC04/22/24) olanları SUT 4.2.15'e göre denetler ve sonucu
        en sağdaki SONUÇ sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ olarak yazar.

        Endikasyonlar:
          A) Koroner stent (max 12 ay)
          B) AKS (STEMI/NSTEMI/unstabil angina), MI
          C) Anjiografik koroner arter hastalığı (ASA intoleransı şart)
          D) Tıkayıcı periferik arter hastalığı
          E) İskemik inme
          F) Kalp kapak biyoprotezi

        Kombinasyon yasağı: klopidogrel+prasugrel+tikagrelor+YOAK (rivaroksaban/
        apiksaban/edoksaban/dabigatran) birlikte KARŞILANMAZ.

        ÖNEMLİ: Hastanın TÜM aktif raporlarındaki ICD kodları da risk faktörü/
        endikasyon tespiti için kullanılır (KAH/inme/PAH ayrı raporda olabilir).
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Klopidogrel Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_klopidogrel, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu, KontrolRaporu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_klop_disi": 0}
        # Alt sınıf dağılımı (raporlama için)
        kategori_sayac = {"KLOPIDOGREL": 0, "PRASUGREL": 0, "TIKAGRELOR": 0}
        denetlenen_satirlar = []

        # ── Reçete bazlı ilaç gruplaması (kombinasyon yasağı kontrolü için) ──
        # Aynı rec_no'ya sahip diğer ilaçların adları → P2Y12 + YOAK / çift P2Y12
        recete_ilac_grup: Dict[str, List[str]] = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no")
            if rno:
                recete_ilac_grup.setdefault(str(rno), []).append(
                    s.get("ilac") or "")

        # YOAK ad ipuçları (kombinasyon yasağı için)
        YOAK_IPUC = ("RIVAROKSABAN", "APIKSABAN", "EDOKSABAN", "DABIGATRAN",
                     "XARELTO", "ELIQUIS", "LIXIANA", "PRADAXA")
        # Çift P2Y12 yasağı (klop+prasugrel+tikagrelor)
        P2Y12_IPUC = ("KLOPIDOGREL", "CLOPIDOGREL", "PRASUGREL", "TIKAGRELOR",
                       "TICAGRELOR", "PLAVIX", "EFFIENT", "EFIENT", "BRILINTA",
                       "BRILIQUE", "PLANOR", "KARUM", "AYRINEX", "KLOPIRA",
                       "OPIROL", "PRASIBLOCK")

        # ── Hastanın diğer raporlarındaki ICD/rapor kodlarını çek ──
        # (Endikasyon ICD'si ayrı bir raporda olabilir → cross-rapor tespit)
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"Klopidogrel kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        # Önceki çalıştırmadan kalan kendi verdict'lerimizi temizle
        for s in self.tum_satirlar:
            kategori = self._klopidogrel_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                if s.get("verdict_kategori") == "KLOPIDOGREL":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_klop_disi"] += 1
                continue

            alt_sinif = self._klopidogrel_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            kategori_sayac[alt_sinif] = kategori_sayac.get(alt_sinif, 0) + 1

            # Aynı reçetedeki diğer ilaçlar (kendisi hariç)
            rno = str(s.get("rec_no") or "")
            kendi_ad = (s.get("ilac") or "").upper()
            diger_adlar = [x for x in recete_ilac_grup.get(rno, [])
                            if x and x.upper() != kendi_ad]

            # ── KOMBİNASYON YASAĞI ÖN-KONTROLÜ ──
            # SUT 4.2.15: P2Y12 + YOAK veya çift P2Y12 → karşılanmaz
            diger_upper = " ".join(diger_adlar).upper()
            yoak_var = any(y in diger_upper for y in YOAK_IPUC)
            cift_p2y12 = any(p in diger_upper for p in P2Y12_IPUC)

            if yoak_var or cift_p2y12:
                ihlal_aciklama = []
                if yoak_var:
                    ihlal_aciklama.append("YOAK (oral antikoagülan) ile birlikte")
                if cift_p2y12:
                    ihlal_aciklama.append("başka bir P2Y12 inhibitörü ile birlikte")
                s["verdict"] = "UYGUN DEĞİL"
                s["verdict_detay"] = (
                    f"Kombinasyon yasağı: {', '.join(ihlal_aciklama)} reçete "
                    f"edilmiş. SUT 4.2.15: P2Y12+YOAK veya çift P2Y12 "
                    f"karşılanmaz."
                )
                s["verdict_kategori"] = "KLOPIDOGREL"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = "Kombinasyon yasağı ihlali"
                s["verdict_sut"] = "SUT 4.2.15 — Kombinasyon yasağı"
                s["verdict_aranan"] = "P2Y12 + YOAK veya çift P2Y12"
                s["verdict_bulunan"] = " | ".join(diger_adlar)[:200]
                s["verdict_detaylar"] = json.dumps(
                    {"diger_ilaclar": diger_adlar,
                     "yoak_var": yoak_var,
                     "cift_p2y12": cift_p2y12},
                    ensure_ascii=False)
                sayac["UYGUN DEĞİL"] += 1
                denetlenen_satirlar.append(s)
                continue

            ilac_sonuc = self._ilac_sonuc_olustur_klopidogrel(s, diger_adlar)

            # Hastanın diğer raporlarındaki ICD'leri ekle (KAH/inme/PAH cross-rapor)
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                rapor = kontrol_klopidogrel(ilac_sonuc)
            except Exception as e:
                logger.warning("Klopidogrel kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "KLOPIDOGREL"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(
                rapor, ek_icd, ['KAH', 'INME', 'PAH'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = "KLOPIDOGREL"
            s["verdict_alt_sinif"] = alt_sinif
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        # Tabloyu yenile
        self._tabloyu_yenile()
        self._durum_yaz(
            f"Klopidogrel SUT 4.2.15 kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(klopidogrel dışı {sayac['_klop_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "Klopidogrel Kontrol",
                "Bu dönemde klopidogrel/prasugrel/tikagrelor (ATC B01AC04/22/24) "
                "bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Klopidogrel SUT 4.2.15 kontrolü tamamlandı.\n\n"
            f"Toplam P2Y12 satırı  : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Klopidogrel dışı (atlanan): {sayac['_klop_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._klopidogrel_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Klopidogrel rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── KLOPİDOGREL KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────────
    def _klopidogrel_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                            denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne kapsamlı Excel raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, alt sınıf dağılımı, çalışma zamanı, dönem
          - Klopidogrel Reçeteleri : Denetlenen her satır + SUT detayları + verdict
          - Kapsam Dışı : ATC B01AC04/22/24 dışında kalanların kısa özeti

        Returns: oluşturulan dosyanın tam yolu.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Klopidogrel_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",  # yeşil
            "UYGUN DEĞİL": "FFCDD2",  # kırmızı
            "ŞÜPHELİ":     "FFE0B2",  # turuncu
            "ATLANDI":     "ECEFF1",  # gri
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="4A148C")  # mor (klop teması)
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1, value="KLOPİDOGREL / P2Y12 SUT KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="4A148C")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("P2Y12 Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_klop_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı", "SUT 4.2.15 — Klopidogrel/Prasugrel/Tikagrelor"),
            ("Filtreleme",
             "ATC B01AC04 (klopidogrel) / B01AC22 (prasugrel) / B01AC24 (tikagrelor) "
             "+ ticari isim fallback"),
            ("Kombinasyon Yasağı",
             "P2Y12 + YOAK (rivaroksaban/apiksaban/edoksaban/dabigatran) veya çift "
             "P2Y12 birlikte karşılanmaz → otomatik UYGUN DEĞİL"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="4A148C")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)

        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt sınıf dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ALT SINIF DAĞILIMI (P2Y12)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Sınıf", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        alt_aciklama = {
            "KLOPIDOGREL": "ATC B01AC04 — Plavix, Planor, Karum, Ayrinex, Klopira, Opirol",
            "PRASUGREL":   "ATC B01AC22 — Effient/Efient, Prasiblock",
            "TIKAGRELOR":  "ATC B01AC24 — Brilinta, Brilique",
        }
        for i, k in enumerate(["KLOPIDOGREL", "PRASUGREL", "TIKAGRELOR"],
                               start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=alt_aciklama[k])

        # Notlar
        bas3 = bas2 + 5
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = SUT 4.2.15 endikasyonu raporda/teşhiste algoritmik olarak doğrulandı",
            "• UYGUN DEĞİL = Kombinasyon yasağı (P2Y12+YOAK / çift P2Y12) veya kural ihlali",
            "• ŞÜPHELİ = Endikasyon ibaresi bulunamadı veya ASA intoleransı eksik — manuel kontrol",
            "• ATLANDI = SUT denetimi gerekmeyen satır",
            "• Kapsam dışı = ATC B01AC04/22/24 dışında — bu butonun kapsamına girmiyor",
            "• Cross-rapor ICD: Hastanın diğer aktif raporlarındaki tanılar da risk/endikasyon için kullanılır",
            "• 'Klopidogrel Reçeteleri' sayfasında her satırın aranan ibare ve bulunan metni yazılır",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: KLOPİDOGREL REÇETELERİ ──────────
        ws2 = wb.create_sheet("Klopidogrel Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           18),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_alt_sinif", "Alt Sınıf",      13),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    32),
            ("verdict_bulunan",  "Bulunan Metin",   32),
            ("verdict_detaylar", "Detaylar (endikasyon/ASA/...)", 38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC B01AC04/22/24 (P2Y12) sınıfına "
                       "girmediği için klopidogrel butonu KAPSAMI DIŞINDA "
                       "bırakıldı.").font = Font(italic=True, color="546E7A")
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._klopidogrel_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ───────────────────────────────────────────────────────────────────
    # YOAK / DOAK (SUT 4.2.15.D-1 / D-2) SUT KONTROLÜ
    # Dabigatran / Rivaroksaban / Apiksaban / Edoksaban
    # ───────────────────────────────────────────────────────────────────
    def _yoak_kontrol_baslat(self):
        """YOAK KONTROL butonu — yüklenen satırlardan dabigatran/rivaroksaban/
        apiksaban/edoksaban (ATC B01AE07/B01AF01/B01AF02/B01AF03) olanları
        SUT 4.2.15.D-1 (AF) ve D-2 (DVT/PE) kurallarına göre denetler ve sonucu
        en sağdaki SONUÇ sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ olarak yazar.

        Endikasyonlar:
          D-1 (Non-valvüler AF):
            • İnme/GİA öyküsü, ≥75 yaş, KY (NYHA ≥II), DM veya HT'den ≥1
            • En az 2 ay varfarin + son 5 INR'nin ≥3'ünde 2-3 aralığı sağlanamamış
              (veya varfarin altında SVO geçirilmiş → doğrudan YOAK)
          D-2 (DVT/PE — yetişkin):
            • DVT veya PE tedavisi / tekrarlayan VTE/PE önlenmesi
            • Yine 2 ay varfarin + INR şartı (idiopatik tekrarlayan PE,
              homozigot trombofili, aktif kanser+VTE, immobil hasta → varfarin
              şartı aranmaz)

        Kombinasyon yasakları (SUT 4.2.15.D-1 (4)):
          • YOAK + YOAK birlikte → KARŞILANMAZ
          • YOAK + Klopidogrel/Prasugrel/Tikagrelor → SUT 4.2.15: KARŞILANMAZ

        ÖNEMLİ: Hastanın TÜM aktif raporlarındaki ICD kodları da endikasyon
        tespiti için kullanılır (AF/DVT/PE/kanser ayrı bir raporda olabilir).
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "YOAK Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_yoak, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_yoak_disi": 0}
        kategori_sayac = {"DABIGATRAN": 0, "RIVAROKSABAN": 0,
                          "APIKSABAN": 0, "EDOKSABAN": 0}
        denetlenen_satirlar = []

        # ── Reçete bazlı ilaç gruplaması (kombinasyon yasağı kontrolü) ──
        recete_ilac_grup: Dict[str, List[str]] = {}
        for s in self.tum_satirlar:
            rno = s.get("rec_no")
            if rno:
                recete_ilac_grup.setdefault(str(rno), []).append(
                    s.get("ilac") or "")

        # YOAK ad ipuçları (kombine YOAK+YOAK tespiti için)
        YOAK_IPUC = ("DABIGATRAN", "RIVAROKSABAN", "RIVAROXABAN", "APIKSABAN",
                     "APIXABAN", "EDOKSABAN", "EDOXABAN",
                     "PRADAXA", "XARELTO", "ELIQUIS", "LIXIANA")
        # P2Y12 ad ipuçları (YOAK + P2Y12 yasağı için)
        P2Y12_IPUC = ("KLOPIDOGREL", "CLOPIDOGREL", "PRASUGREL", "TIKAGRELOR",
                      "TICAGRELOR", "PLAVIX", "EFFIENT", "EFIENT", "BRILINTA",
                      "BRILIQUE", "PLANOR", "KARUM", "AYRINEX", "KLOPIRA",
                      "OPIROL", "PRASIBLOCK")

        # ── Hastanın diğer raporlarındaki ICD kodlarını çek ──
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"YOAK kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        for s in self.tum_satirlar:
            kategori = self._yoak_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin önceki artığını sil — başka
                # kontrollerin verdict'lerine dokunma.
                if s.get("verdict_kategori") == "YOAK":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_alt_sinif"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_yoak_disi"] += 1
                continue

            alt_sinif = self._yoak_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            kategori_sayac[alt_sinif] = kategori_sayac.get(alt_sinif, 0) + 1

            # Aynı reçetedeki diğer ilaçlar (kendisi hariç)
            rno = str(s.get("rec_no") or "")
            kendi_ad = (s.get("ilac") or "").upper()
            diger_adlar = [x for x in recete_ilac_grup.get(rno, [])
                           if x and x.upper() != kendi_ad]
            diger_upper = " ".join(diger_adlar).upper()

            # ── KOMBİNASYON YASAĞI ÖN-KONTROLÜ ──
            # SUT 4.2.15.D-1 (4): YOAK+YOAK karşılanmaz
            # SUT 4.2.15: YOAK+P2Y12 karşılanmaz
            yoak_var = any(y in diger_upper for y in YOAK_IPUC)
            p2y12_var = any(p in diger_upper for p in P2Y12_IPUC)

            if yoak_var or p2y12_var:
                ihlal_aciklama = []
                if yoak_var:
                    ihlal_aciklama.append("başka bir YOAK ile birlikte")
                if p2y12_var:
                    ihlal_aciklama.append("P2Y12 inhibitörü (klop/prasug/tikagrelor) ile birlikte")
                s["verdict"] = "UYGUN DEĞİL"
                s["verdict_detay"] = (
                    f"Kombinasyon yasağı: {', '.join(ihlal_aciklama)} reçete "
                    f"edilmiş. SUT 4.2.15.D-1 (4): YOAK+YOAK; SUT 4.2.15: "
                    f"YOAK+P2Y12 karşılanmaz."
                )
                s["verdict_kategori"] = "YOAK"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = "Kombinasyon yasağı ihlali"
                s["verdict_sut"] = "SUT 4.2.15.D-1 (4) — Kombinasyon yasağı"
                s["verdict_aranan"] = "YOAK+YOAK veya YOAK+P2Y12"
                s["verdict_bulunan"] = " | ".join(diger_adlar)[:200]
                s["verdict_detaylar"] = json.dumps(
                    {"diger_ilaclar": diger_adlar,
                     "yoak_var": yoak_var,
                     "p2y12_var": p2y12_var},
                    ensure_ascii=False)
                sayac["UYGUN DEĞİL"] += 1
                denetlenen_satirlar.append(s)
                continue

            ilac_sonuc = self._ilac_sonuc_olustur_yoak(s, diger_adlar)

            # Cross-rapor ICD ekle (AF/DVT/PE/kanser ayrı raporda olabilir)
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                rapor = kontrol_yoak(ilac_sonuc)
            except Exception as e:
                logger.warning("YOAK kontrol hata (rx %s): %s",
                               s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "YOAK"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(
                rapor, ek_icd, ['KAH', 'INME', 'PAH'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = "YOAK"
            s["verdict_alt_sinif"] = alt_sinif
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                   ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            # Atomik şart listesi — atomik şema paneli için (BUG FIX 2026-05-08)
            sartlar_obj = getattr(rapor, "sartlar", None) or []
            try:
                s["verdict_sartlar"] = json.dumps([
                    {"ad": p.ad,
                     "durum": (p.durum.value if hasattr(p.durum, "value")
                                else str(p.durum)),
                     "neden": p.neden,
                     "kaynak": getattr(p, "kaynak", ""),
                     "grup": getattr(p, "grup", ""),
                     "veya_grubu": bool(getattr(p, "veya_grubu", False))}
                    for p in sartlar_obj
                ], ensure_ascii=False)
            except Exception:
                s["verdict_sartlar"] = ""
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        # Tabloyu yenile
        self._tabloyu_yenile()
        self._durum_yaz(
            f"YOAK SUT 4.2.15.D kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(YOAK dışı {sayac['_yoak_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "YOAK Kontrol",
                "Bu dönemde dabigatran/rivaroksaban/apiksaban/edoksaban "
                "(ATC B01AE07/B01AF01/B01AF02/B01AF03) bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"YOAK SUT 4.2.15.D kontrolü tamamlandı.\n\n"
            f"Toplam YOAK satırı   : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"YOAK dışı (atlanan)  : {sayac['_yoak_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._yoak_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("YOAK rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── YOAK KONTROL RAPORU EXCEL ÜRETİCİ ───────────────────────────────
    def _yoak_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                    denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne kapsamlı YOAK Excel raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, alt etken dağılımı, çalışma zamanı, dönem
          - YOAK Reçeteleri : Denetlenen her satır + SUT detayları + verdict
          - Kapsam Dışı : ATC B01AE07/B01AF01/02/03 dışında kalanların özeti

        Returns: oluşturulan dosyanın tam yolu.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"YOAK_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",  # yeşil
            "UYGUN DEĞİL": "FFCDD2",  # kırmızı
            "ŞÜPHELİ":     "FFE0B2",  # turuncu
            "ATLANDI":     "ECEFF1",  # gri
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="1A237E")  # indigo (YOAK teması)
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1, value="YOAK / DOAK SUT 4.2.15.D KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="1A237E")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("YOAK Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_yoak_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.15.D-1 (Non-valvüler AF) · SUT 4.2.15.D-2 (DVT/PE)"),
            ("Filtreleme",
             "ATC B01AE07 (dabigatran) / B01AF01 (rivaroksaban) / "
             "B01AF02 (apiksaban) / B01AF03 (edoksaban) + ticari isim fallback "
             "(Pradaxa/Xarelto/Eliquis/Lixiana)"),
            ("Kombinasyon Yasağı",
             "YOAK+YOAK (SUT 4.2.15.D-1 fıkra 4) ve YOAK+P2Y12 (klop/prasug/"
             "tikagrelor) birlikte karşılanmaz → otomatik UYGUN DEĞİL"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="1A237E")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)

        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                    "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt etken dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1, value="ALT ETKEN DAĞILIMI (YOAK)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Etken", "Adet", "Açıklama", ""], start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        alt_aciklama = {
            "DABIGATRAN":   "ATC B01AE07 — Pradaxa (direk trombin inhibitörü)",
            "RIVAROKSABAN": "ATC B01AF01 — Xarelto (Faktör Xa inhibitörü)",
            "APIKSABAN":    "ATC B01AF02 — Eliquis (Faktör Xa inhibitörü)",
            "EDOKSABAN":    "ATC B01AF03 — Lixiana (Faktör Xa inhibitörü)",
        }
        for i, k in enumerate(["DABIGATRAN", "RIVAROKSABAN",
                               "APIKSABAN", "EDOKSABAN"],
                              start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=alt_aciklama[k])

        # Notlar
        bas3 = bas2 + 6
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = SUT 4.2.15.D endikasyonu (AF / DVT / PE / kanser+VTE / ortopedik) algoritmik olarak doğrulandı",
            "• UYGUN DEĞİL = Kombinasyon yasağı (YOAK+YOAK / YOAK+P2Y12) veya kural ihlali",
            "• ŞÜPHELİ = Endikasyon ibaresi/ICD bulunamadı veya AF'de varfarin/INR bilgisi eksik — manuel kontrol",
            "• ATLANDI = SUT denetimi gerekmeyen satır",
            "• Kapsam dışı = ATC B01AE07/B01AF01/02/03 dışında — bu butonun kapsamına girmiyor",
            "• AF'de SUT şartı: en az 2 ay varfarin + son 5 INR'nin ≥3'ünde 2-3 sağlanamaması "
            "(veya varfarin altında SVO → doğrudan YOAK)",
            "• DVT/PE'de varfarin şartı YOK (D-2 fıkra 1.a/2)",
            "• Cross-rapor ICD: Hastanın diğer aktif raporlarındaki tanılar da endikasyon için kullanılır",
            "• 'YOAK Reçeteleri' sayfasında her satırın aranan ibare ve bulunan metni yazılır",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: YOAK REÇETELERİ ──────────
        ws2 = wb.create_sheet("YOAK Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           18),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_alt_sinif", "Alt Etken",      14),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    32),
            ("verdict_bulunan",  "Bulunan Metin",   32),
            ("verdict_detaylar", "Detaylar (endikasyon/INR/...)", 38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen

        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar ATC B01AE07/B01AF01/02/03 (YOAK) "
                       "sınıfına girmediği için YOAK butonu KAPSAMI DIŞINDA "
                       "bırakıldı.").font = Font(italic=True, color="546E7A")
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._yoak_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ───────────────────────────────────────────────────────────────────
    # OSTEOPOROZ / KEMİK ERİMESİ (SUT 4.2.17 + 4.2.28.C) SUT KONTROLÜ
    # ───────────────────────────────────────────────────────────────────
    def _osteo_kontrol_baslat(self):
        """KEMİK ERİMESİ KONTROL butonu — yüklenen satırlardan osteoporoz
        ilaçları (bifosfonat / biyolojik / SERM) SUT 4.2.17 ve SUT 4.2.28.C'ye
        göre denetlenir; sonuç en sağdaki SONUÇ sütununa UYGUN / UYGUN DEĞİL /
        ŞÜPHELİ olarak yazılır.

        Kapsanan ilaç sınıfları:
          • Oral bifosfonatlar (alendronat, risedronat, ibandronat oral)
          • IV bifosfonatlar (zoledronik asit, pamidronat, ibandronat IV)
          • Biyolojik osteoporoz (denosumab Prolia, teriparatid, romosozumab)
          • Stronsiyum ranelat (Osseor/Protelos — kullanımı kısıtlı)
          • SERM (raloksifen — Evista)
          • Onkolojik denosumab (Xgeva — ayrı SUT 12.7.X kuralı)

        Kurallar:
          • SUT 4.2.17: T-skoru eşikleri (kalça kırığı→KMY gerekmez,
            kırık+T≤-1, kortikosteroid+T≤-1, ≥65 yaş T≤-2.5, <65 yaş T≤-3)
          • SUT 4.2.28.C: Biyolojikler için Endokrin/Romatoloji/FTR uzman
            raporu zorunlu
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Kemik Erimesi Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_bifosfonat, kontrol_osteoporoz_biyolojik,
            )
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        # Hastaların TÜM aktif raporlarındaki ICD kodlarını topla
        # (Sekonder osteoporoz / kortikosteroid kullanımı / kırık öyküsü
        # ayrı raporda yazılı olabilir)
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"Kemik erimesi kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_osteo_disi": 0}
        kategori_sayac = {
            "BIFOSFONAT_ORAL": 0, "BIFOSFONAT_IV": 0,
            "DENOSUMAB": 0, "TERIPARATID": 0, "ROMOSOZUMAB": 0,
            "STRONSIYUM": 0, "RALOKSIFEN": 0, "DIGER": 0,
        }
        denetlenen_satirlar = []

        # Önceki çalıştırmadan kalan kendi verdict'lerimizi temizle
        for s in self.tum_satirlar:
            kategori = self._osteo_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin (OSTEOPOROZ) önceki artığını sil —
                # statin/diyabet/klopidogrel/arb verdict'lerine dokunma.
                if s.get("verdict_kategori") == "OSTEOPOROZ":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_osteo_disi"] += 1
                continue

            alt_sinif = self._osteo_alt_sinif(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            kategori_sayac[alt_sinif] = kategori_sayac.get(alt_sinif, 0) + 1

            ilac_sonuc = self._ilac_sonuc_olustur_osteo(s)

            # Hastanın diğer raporlarındaki ICD'leri ekle
            # (M80/M81/M82 osteoporoz; S22/S32/S72 kırık ICD'leri ayrı
            # raporda olabilir)
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                if kategori == "BIYOLOJIK":
                    rapor = kontrol_osteoporoz_biyolojik(ilac_sonuc)
                else:
                    # BIFOSFONAT ve SERM için ortak T-skoru/kırık algoritması
                    rapor = kontrol_bifosfonat(ilac_sonuc)
            except Exception as e:
                logger.warning("Kemik erimesi kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "OSTEOPOROZ"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = "OSTEOPOROZ"
            s["verdict_alt_sinif"] = alt_sinif
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        # Tabloyu yenile
        self._tabloyu_yenile()
        self._durum_yaz(
            f"Kemik erimesi SUT 4.2.17 + 4.2.28.C kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(kapsam dışı {sayac['_osteo_disi']} satır boş bırakıldı)"
        )

        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam == 0:
            messagebox.showinfo(
                "Kemik Erimesi Kontrol",
                "Bu dönemde osteoporoz/kemik erimesi ilacı (bifosfonat/biyolojik/SERM) "
                "bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Kemik erimesi SUT kontrolü tamamlandı.\n\n"
            f"Toplam osteoporoz satırı : {toplam}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Kapsam dışı (atlanan): {sayac['_osteo_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._osteo_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Kemik erimesi rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── KEMİK ERİMESİ KONTROL RAPORU EXCEL ÜRETİCİ ──────────────────────
    def _osteo_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                     denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne kapsamlı Excel raporu yazar.

        3 sayfa:
          - Özet : Toplam sayım, alt sınıf dağılımı, çalışma zamanı, dönem
          - Osteoporoz Reçeteleri : Denetlenen her satır + SUT detayları + verdict
          - Kapsam Dışı : Bifosfonat/biyolojik/SERM dışında kalanların özeti

        Returns: oluşturulan dosyanın tam yolu.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Kemik_Erimesi_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",  # yeşil
            "UYGUN DEĞİL": "FFCDD2",  # kırmızı
            "ŞÜPHELİ":     "FFE0B2",  # turuncu
            "ATLANDI":     "ECEFF1",  # gri
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="6A1B9A")  # mor — buton rengiyle uyumlu
        toplam = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                  + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1,
                 value="KEMİK ERİMESİ / OSTEOPOROZ SUT KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="6A1B9A")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Osteoporoz Olarak Tespit Edilen", str(toplam)),
            ("Kapsam Dışı (Atlanan)", str(sayac["_osteo_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralları",
             "SUT 4.2.17 (Bifosfonat) · SUT 4.2.28.C (Osteoporoz biyolojik)"),
            ("Filtreleme",
             "ATC M05BA/BB/BX + H05AA + G03XC + ticari isim fallback"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="6A1B9A")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı tablosu
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)

        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam * 100) if toplam else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt sınıf dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1,
                 value="ALT SINIF DAĞILIMI (osteoporoz satırları)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Sınıf", "Adet", "Açıklama", ""],
                                  start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "BIFOSFONAT_ORAL": "Alendronat, Risedronat, İbandronat oral (Fosamax/Fosavance/Actonel/Bonviva)",
            "BIFOSFONAT_IV":   "Zoledronik asit (Zometa/Aclasta), Pamidronat (Aredia), İbandronat IV (Bondronat)",
            "DENOSUMAB":       "Prolia (60 mg s.c. 6 ayda bir) — osteoporoz; Xgeva (120 mg) — onkoloji",
            "TERIPARATID":     "Forteo / Forsteo / Movymia — paratiroid hormonu (max 24 ay)",
            "ROMOSOZUMAB":     "Evenity — sklerostin inhibitörü (12 ay; KV olay öyküsü kontrendike)",
            "STRONSIYUM":      "Osseor / Protelos — KV risk nedeniyle kullanımı kısıtlı",
            "RALOKSIFEN":      "Evista / Optruma — SERM (postmenopozal osteoporoz)",
            "DIGER":           "Sınıflandırılamayan",
        }
        sira = ["BIFOSFONAT_ORAL", "BIFOSFONAT_IV", "DENOSUMAB",
                "TERIPARATID", "ROMOSOZUMAB", "STRONSIYUM",
                "RALOKSIFEN", "DIGER"]
        for i, k in enumerate(sira, start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + len(sira) + 2
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = SUT kuralı algoritmik olarak doğrulandı (T-skor + kırık + kortikosteroid + uzman raporu)",
            "• UYGUN DEĞİL = Algoritmik kural net şekilde ihlal — manuel inceleme önerilir",
            "  (Örn: T > -1; biyolojik ilaç raporsuz; XGEVA onkoloji raporsuz)",
            "• ŞÜPHELİ = Karar verilemedi (eksik T-skoru sayısal değer, eksik kırık/kortikosteroid bilgisi)",
            "• ATLANDI = SUT denetimi gerekmeyen satır",
            "• Kapsam dışı = ATC M05*/H05AA/G03XC dışında olduğu için bu butonun kapsamına girmiyor",
            "• 'Osteoporoz Reçeteleri' sayfasında her satırın hangi metni okuduğu, "
            "neyi aradığı ve ne bulduğu yazılır.",
            "• T-skoru eşikleri (SUT 4.2.17): Kalça kırığı→KMY gerekmez | "
            "T≤-1 (kırık veya kortikosteroid) | T≤-2.5 (≥65 yaş) | T≤-3 (<65 yaş)",
            "• Biyolojik (SUT 4.2.28.C): Endokrinoloji/Romatoloji/FTR uzman raporu zorunlu, "
            "bifosfonata intolerans/yetersiz yanıt belirtilmeli",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        # Sütun genişlikleri
        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: OSTEOPOROZ REÇETELERİ ──────────
        ws2 = wb.create_sheet("Osteoporoz Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",       12),
            ("rec_no",           "Reçete No",       18),
            ("hasta",            "Hasta",           24),
            ("tc",               "TC",              13),
            ("yas",              "Yaş",             6),
            ("cins",             "Cin.",            6),
            ("doktor",           "Doktor",          22),
            ("brans",            "Branş",           18),
            ("ilac",             "İlaç",            28),
            ("etkin",            "Etken Madde",     22),
            ("atc",              "ATC",             10),
            ("verdict_alt_sinif","Alt Sınıf",       18),
            ("rap_kod",          "Rapor Kod",       11),
            ("rec_doz",          "Reçete Doz",      14),
            ("rap_doz",          "Rapor Doz",       14),
            ("kutu",             "Kutu",            6),
            ("msj",              "Msj",             7),
            ("uyari",            "Uyarı Kod",       18),
            ("medula_msj",       "Medula Msj",      30),
            ("rec_tesh",         "Reçete Teşhis",   30),
            ("rap_tesh",         "Rapor Teşhis",    30),
            ("rec_ack",          "Reçete Açıklama", 30),
            ("rap_ack",          "Rapor Açıklama",  30),
            ("verdict_sut",      "Uygulanan SUT",   30),
            ("verdict_aranan",   "Aranan İbare",    28),
            ("verdict_bulunan",  "Bulunan Metin",   28),
            ("verdict_detaylar", "SUT Detaylar (T-skor/kırık/...)", 38),
            ("verdict_uyari",    "Uyarı",           34),
            ("verdict_detay",    "Açıklama",        50),
            ("verdict",          "SONUÇ",           14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI (atlanmış) ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar bifosfonat/biyolojik/SERM sınıfına "
                       "girmediği için kemik erimesi butonu KAPSAMI DIŞINDA "
                       "bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._osteo_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ───────────────────────────────────────────────────────────────────
    # ENTERAL BESLENME SUT KONTROLÜ
    # ───────────────────────────────────────────────────────────────────
    def _enteral_kontrol_baslat(self):
        """ENTERAL BESLENME KONTROL butonu — yüklenen satırlardan enteral
        beslenme solüsyonlarını (V06D* + ad fallback) SUT'a göre denetler;
        sonuç en sağdaki SONUÇ sütununa UYGUN / UYGUN DEĞİL / ŞÜPHELİ
        olarak yazılır.

        SUT kuralı (özet):
          • Endikasyon: Malnütrisyon / Disfaji / Kanser / Kistik fibroz /
            IBD-Crohn / İnek sütü alerjisi / Nörolojik / Demans
          • Uzman branş: İç Hast / Dahiliye / GE / Onkoloji / Geriatri /
            Pediatri / Nöroloji
          • Kalori planı: kcal/birim × günlük doz; hasta kilosu varsa
            kcal/kg/gün (hedef 25-35)
          • Rapor zorunlu — raporsuz UYGUN_DEĞİL
          • Rapor kodu 15.* veya 02.* başlıyorsa Medula örtük endikasyon

        Hastanın TÜM aktif raporlarındaki ICD'leri de teşhislere ekler
        (kanser/Crohn ayrı raporda olabilir).
        """
        if not self.tum_satirlar:
            messagebox.showinfo(
                "Enteral Beslenme Kontrol",
                "Önce DÖNEM seçip 🔍 SORGULA ile reçeteleri yükleyin.",
                parent=self.root)
            return
        try:
            from recete_kontrol.sut_kontrolleri import (
                kontrol_enteral_beslenme, _diger_rapor_notunu_uyariya_ekle)
            from recete_kontrol.base_kontrol import KontrolSonucu
        except Exception as e:
            self._durum_yaz(f"SUT kontrol modülü yüklenemedi: {e}")
            messagebox.showerror(
                "Modül Hatası",
                f"recete_kontrol modülü yüklenemedi:\n{e}",
                parent=self.root)
            return

        VERDICT_ETIKET = {
            KontrolSonucu.UYGUN:             "UYGUN",
            KontrolSonucu.UYGUN_DEGIL:       "UYGUN DEĞİL",
            KontrolSonucu.KONTROL_EDILEMEDI: "ŞÜPHELİ",
            KontrolSonucu.ATLANDI:           "ATLANDI",
        }
        sayac = {"UYGUN": 0, "UYGUN DEĞİL": 0,
                 "ŞÜPHELİ": 0, "ATLANDI": 0, "_enteral_disi": 0}
        kategori_sayac = {"STANDART": 0, "DIYABETIK": 0, "RENAL": 0,
                          "PEDIATRIK": 0, "IMMUNONUTRIYON": 0, "DIGER": 0}
        denetlenen_satirlar = []

        # Hastaların TÜM aktif raporlarındaki ICD'leri topla — kanser/Crohn
        # endikasyonu farklı raporda yazıyor olabilir.
        musteri_idler = list({
            s.get("musteri_id") for s in self.tum_satirlar
            if s.get("musteri_id")
        })
        self._durum_yaz(
            f"Enteral beslenme kontrol — {len(musteri_idler)} hastanın "
            "diğer raporları taranıyor…")
        self.root.update_idletasks()
        hasta_tum_icd = self._hasta_tum_icd_kodlarini_topla(musteri_idler)

        for s in self.tum_satirlar:
            kategori = self._enteral_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori == "NONE":
                # Sadece kendi kategorimizin (ENTERAL) önceki artığını sil —
                # statin/diyabet/klopidogrel/osteo verdict'lerine dokunma.
                if s.get("verdict_kategori") == "ENTERAL":
                    s["verdict"] = ""
                    s["verdict_detay"] = ""
                    s["verdict_kategori"] = ""
                    s["verdict_alt_sinif"] = ""
                    s["verdict_uyari"] = ""
                    s["verdict_sut"] = ""
                    s["verdict_aranan"] = ""
                    s["verdict_bulunan"] = ""
                    s["verdict_detaylar"] = ""
                sayac["_enteral_disi"] += 1
                continue

            alt_sinif = self._enteral_alt_sinif(s.get("ilac"))
            kategori_sayac[alt_sinif] = kategori_sayac.get(alt_sinif, 0) + 1

            ilac_sonuc = self._ilac_sonuc_olustur_enteral(s)

            # Hastanın diğer raporlarındaki ICD'leri ekle (kanser/Crohn vb.
            # endikasyon başka bir raporda yazıyor olabilir)
            mid = s.get("musteri_id")
            ek_icd = hasta_tum_icd.get(mid, []) if mid else []
            ilac_sonuc["diger_raporlar_icd"] = list(ek_icd)

            try:
                rapor = kontrol_enteral_beslenme(ilac_sonuc)
            except Exception as e:
                logger.warning("Enteral kontrol hata (rx %s): %s",
                                s.get("rec_no"), e)
                s["verdict"] = "ŞÜPHELİ"
                s["verdict_detay"] = f"Hata: {e}"
                s["verdict_kategori"] = "ENTERAL"
                s["verdict_alt_sinif"] = alt_sinif
                s["verdict_uyari"] = ""
                s["verdict_sut"] = ""
                s["verdict_aranan"] = ""
                s["verdict_bulunan"] = ""
                s["verdict_detaylar"] = ""
                sayac["ŞÜPHELİ"] += 1
                denetlenen_satirlar.append(s)
                continue

            _diger_rapor_notunu_uyariya_ekle(rapor, ek_icd, ['DM'])
            etiket = VERDICT_ETIKET.get(rapor.sonuc, "ŞÜPHELİ")
            s["verdict"] = etiket
            s["verdict_detay"] = rapor.mesaj or ""
            s["verdict_kategori"] = "ENTERAL"
            s["verdict_alt_sinif"] = alt_sinif
            s["verdict_uyari"] = rapor.uyari or ""
            s["verdict_sut"] = rapor.sut_kurali or ""
            s["verdict_aranan"] = rapor.aranan_ibare or ""
            s["verdict_bulunan"] = rapor.bulunan_metin or ""
            try:
                s["verdict_detaylar"] = json.dumps(rapor.detaylar or {},
                                                    ensure_ascii=False)
            except Exception:
                s["verdict_detaylar"] = str(rapor.detaylar or {})
            sayac[etiket] = sayac.get(etiket, 0) + 1
            denetlenen_satirlar.append(s)

        # ═══════════════════════════════════════════════════════════════
        # FAZ 2 — HASTA-BAZLI ÇOKLU MAMA TOPLAM KALORİ KONTROLÜ
        # SUT 4.2.8.A: Aynı hastanın aynı reçetede birden fazla enteral
        # solüsyonu varsa toplam günlük kcal raporda yazan ihtiyacı veya
        # 1200 kcal tavanını (yetişkin malnütrisyon) aşamaz.
        # ═══════════════════════════════════════════════════════════════
        gruplar: Dict[tuple, List[dict]] = {}
        for s in denetlenen_satirlar:
            mid = s.get("musteri_id")
            rno = s.get("rec_no")
            if mid is None or not rno:
                continue
            gruplar.setdefault((mid, str(rno)), []).append(s)

        coklu_mama_ihlal_grup = 0
        for (mid, rno), grup in gruplar.items():
            if len(grup) < 2:
                continue  # Tek satır — çoklu mama yok

            # Grup içindeki gunluk_kcal'leri topla (verdict_detaylar JSON'undan)
            toplam_kcal = 0.0
            kcal_olculen = 0
            rapor_ihtiyac = None
            yas_grubu = None
            ozel_durum_var = False
            for sat in grup:
                try:
                    det = json.loads(sat.get("verdict_detaylar") or "{}")
                except (ValueError, TypeError):
                    det = {}
                gk = det.get("gunluk_kcal")
                if gk:
                    try:
                        toplam_kcal += float(gk)
                        kcal_olculen += 1
                    except (TypeError, ValueError):
                        pass
                if rapor_ihtiyac is None:
                    rapor_ihtiyac = det.get("rapor_gunluk_kcal_ihtiyaci")
                if yas_grubu is None:
                    yas_grubu = det.get("yas_grubu")
                # Yetişkin özel durumlarda 1200 tavanı uygulanmaz
                if det.get("yetiskin_ozel_durumlar"):
                    ozel_durum_var = True

            if kcal_olculen < 2:
                continue  # Toplam kıyaslama yapılamaz, en az 2 satırda kcal lazım

            # Tavanlar
            tavanlar = []
            if rapor_ihtiyac is not None:
                tavanlar.append(("raporda yazan günlük ihtiyaç",
                                  float(rapor_ihtiyac)))
            if (yas_grubu in ("YETISKIN_70_ALT", "YETISKIN_70_USTU")
                    and not ozel_durum_var):
                tavanlar.append(("yetişkin malnütrisyon tavanı 1200 kcal",
                                  1200.0))

            if not tavanlar:
                continue

            for tavan_adi, tavan_deg in tavanlar:
                if toplam_kcal > tavan_deg:
                    # Tüm grup satırlarını UYGUN_DEĞİL'e çevir
                    coklu_mama_ihlal_grup += 1
                    for sat in grup:
                        eski_verdict = sat.get("verdict") or ""
                        eski_detay = sat.get("verdict_detay") or ""
                        ek_mesaj = (
                            f"ÇOKLU MAMA İHLALİ — Bu reçetedeki {len(grup)} "
                            f"enteral satırın toplam günlük kalorisi "
                            f"{toplam_kcal:.0f} kcal, {tavan_adi} ({tavan_deg:.0f} kcal) "
                            f"aşıyor"
                        )
                        sat["verdict"] = "UYGUN DEĞİL"
                        sat["verdict_detay"] = (
                            f"{ek_mesaj} | (tek-satır kararı: {eski_verdict}) "
                            f"{eski_detay}"
                        )
                        sat["verdict_uyari"] = (
                            "SUT 4.2.8.A: Aynı reçetedeki tüm enteral mamaların "
                            "toplam günlük kalorisi tavanı geçemez"
                        )
                        # Eski verdict UYGUN ise sayaçtan düş
                        if eski_verdict == "UYGUN":
                            sayac["UYGUN"] = max(0, sayac["UYGUN"] - 1)
                            sayac["UYGUN DEĞİL"] += 1
                        elif eski_verdict == "ŞÜPHELİ":
                            sayac["ŞÜPHELİ"] = max(0, sayac["ŞÜPHELİ"] - 1)
                            sayac["UYGUN DEĞİL"] += 1
                        # ATLANDI ve UYGUN DEĞİL kalanlar zaten doğru
                    break  # Bir tavan aşıldı, diğerini kontrol etmeye gerek yok

        # Tabloyu yenile (mevcut filtre/renk durumuna saygı duyarak)
        self._tabloyu_yenile()
        self._durum_yaz(
            f"Enteral beslenme SUT kontrolü: "
            f"✓ UYGUN {sayac['UYGUN']}  "
            f"✗ UYGUN DEĞİL {sayac['UYGUN DEĞİL']}  "
            f"? ŞÜPHELİ {sayac['ŞÜPHELİ']}  "
            f"− ATLANDI {sayac['ATLANDI']}  "
            f"(çoklu mama ihlal: {coklu_mama_ihlal_grup} grup, "
            f"enteral dışı {sayac['_enteral_disi']})"
        )

        toplam_enteral = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                          + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])
        if toplam_enteral == 0:
            messagebox.showinfo(
                "Enteral Beslenme Kontrol",
                "Bu dönemde enteral beslenme solüsyonu (V06D* / ürün ailesi) "
                "bulunamadı.",
                parent=self.root)
            return

        cevap = messagebox.askyesno(
            "Kontrol Raporu",
            f"Enteral beslenme SUT kontrolü tamamlandı.\n\n"
            f"Toplam enteral satırı : {toplam_enteral}\n"
            f"  ✓ UYGUN          : {sayac['UYGUN']}\n"
            f"  ✗ UYGUN DEĞİL    : {sayac['UYGUN DEĞİL']}\n"
            f"  ? ŞÜPHELİ        : {sayac['ŞÜPHELİ']}\n"
            f"  − ATLANDI        : {sayac['ATLANDI']}\n"
            f"Enteral dışı (atlanan): {sayac['_enteral_disi']}\n\n"
            f"Kontrol raporu Excel olarak masaüstündeki "
            f"'Reçete Kontrol' klasörüne kaydedilecek.\n\n"
            f"Rapor oluşturulup açılsın mı?",
            parent=self.root)
        if not cevap:
            return

        try:
            rapor_yolu = self._enteral_rapor_excel_olustur(
                sayac=sayac,
                kategori_sayac=kategori_sayac,
                denetlenen_satirlar=denetlenen_satirlar,
            )
        except Exception as e:
            logger.exception("Enteral rapor üretim hatası")
            messagebox.showerror(
                "Rapor Hatası",
                f"Kontrol raporu oluşturulamadı:\n{e}",
                parent=self.root)
            return

        try:
            os.startfile(rapor_yolu)
            self._durum_yaz(f"Kontrol raporu: {rapor_yolu}")
        except Exception as e:
            messagebox.showinfo(
                "Rapor Kaydedildi",
                f"Rapor kaydedildi ama otomatik açılamadı:\n{rapor_yolu}\n\n{e}",
                parent=self.root)

    # ── ENTERAL KONTROL RAPORU EXCEL ÜRETİCİ ────────────────────────────
    def _enteral_rapor_excel_olustur(self, *, sayac: dict, kategori_sayac: dict,
                                       denetlenen_satirlar: list) -> str:
        """Masaüstü/Reçete Kontrol/ klasörüne 3 sayfalı Excel raporu yazar.

          - Özet : Toplam sayım, alt sınıf dağılımı, çalışma zamanı, dönem
          - Enteral Reçeteleri : Denetlenen her satır + kalori/endikasyon/
                                  uzman/yol bilgileri + verdict
          - Kapsam Dışı (Atlanan) : Enteral olmayan satırların kısa özeti

        Returns: oluşturulan dosyanın tam yolu.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError("openpyxl yüklü değil") from e

        from datetime import datetime as _dt

        masa = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
        if not os.path.exists(masa):
            masa = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(masa):
                masa = os.path.expanduser("~")
        klasor = os.path.join(masa, "Reçete Kontrol")
        os.makedirs(klasor, exist_ok=True)

        donem = self.aktif_donem or "tum"
        zaman = _dt.now().strftime("%Y%m%d_%H%M%S")
        dosya_adi = f"Enteral_Beslenme_Kontrol_{donem}_{zaman}.xlsx"
        path = os.path.join(klasor, dosya_adi)

        wb = openpyxl.Workbook()

        # ────────── SAYFA 1: ÖZET ──────────
        ws1 = wb.active
        ws1.title = "Özet"

        VERDICT_RENK = {
            "UYGUN":       "C8E6C9",
            "UYGUN DEĞİL": "FFCDD2",
            "ŞÜPHELİ":     "FFE0B2",
            "ATLANDI":     "ECEFF1",
        }
        baslik_font = Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = PatternFill("solid", fgColor="00695C")  # teal
        toplam_enteral = (sayac["UYGUN"] + sayac["UYGUN DEĞİL"]
                          + sayac["ŞÜPHELİ"] + sayac["ATLANDI"])

        ws1.cell(row=1, column=1, value="ENTERAL BESLENME SUT KONTROL RAPORU")
        ws1.cell(row=1, column=1).font = Font(bold=True, size=14, color="00695C")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        bilgi_satirlari = [
            ("Dönem (Yıl-Ay)", donem),
            ("Rapor Üretim Tarihi", _dt.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Toplam Yüklenen Satır", str(len(self.tum_satirlar))),
            ("Enteral Olarak Tespit Edilen", str(toplam_enteral)),
            ("Enteral Dışı (Atlanan)", str(sayac["_enteral_disi"])),
            ("", ""),
            ("Uygulanan SUT Kuralı",
             "Enteral beslenme — endikasyon + uzman raporu + kalori planı"),
            ("Kapsam",
             "ATC V06D* + ürün ailesi (Resource/Nutridrink/Nutren/Fresubin/"
             "Evolvia/Ensure/Peptamen/Nepro/Impact/Prosure/Modulen/Glucerna/"
             "Diasip/Cubitan/Abound/Juven/Fortimel/Pediasure)"),
            ("Endikasyonlar",
             "Malnütrisyon / Disfaji / Kanser / Kistik fibroz / IBD-Crohn / "
             "İnek sütü alerjisi / Nörolojik / Demans"),
            ("Uzman Branşlar",
             "İç Hastalıkları / Dahiliye / Gastroenteroloji / Onkoloji / "
             "Geriatri / Pediatri / Nöroloji"),
            ("Kalori Hedefi",
             "25-35 kcal/kg/gün (hasta kilosu metinde varsa hesaplanır)"),
            ("Veri Kaynağı",
             "Botanik EOS DB (SADECE SELECT) — hiçbir veri değiştirilmedi"),
        ]
        for i, (etk, deg) in enumerate(bilgi_satirlari, start=3):
            c1 = ws1.cell(row=i, column=1, value=etk)
            c2 = ws1.cell(row=i, column=2, value=deg)
            if etk:
                c1.font = Font(bold=True, color="37474F")
            c2.font = Font(color="00695C")
            ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # Sonuç dağılımı
        bas = len(bilgi_satirlari) + 4
        ws1.cell(row=bas, column=1, value="SONUÇ DAĞILIMI")
        ws1.cell(row=bas, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas, start_column=1,
                        end_row=bas, end_column=4)

        bas += 1
        for col, hd in enumerate(["Sonuç", "Adet", "Yüzde", ""], start=1):
            c = ws1.cell(row=bas, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        for i, etiket in enumerate(["UYGUN", "UYGUN DEĞİL",
                                     "ŞÜPHELİ", "ATLANDI"], start=bas + 1):
            adet = sayac[etiket]
            yuzde = (adet / toplam_enteral * 100) if toplam_enteral else 0
            c1 = ws1.cell(row=i, column=1, value=etiket)
            c2 = ws1.cell(row=i, column=2, value=adet)
            c3 = ws1.cell(row=i, column=3, value=f"%{yuzde:.1f}")
            fill = PatternFill("solid", fgColor=VERDICT_RENK[etiket])
            for c in (c1, c2, c3):
                c.fill = fill
                c.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")

        # Alt sınıf dağılımı
        bas2 = bas + 6
        ws1.cell(row=bas2, column=1,
                  value="ALT SINIF DAĞILIMI (enteral satırları)")
        ws1.cell(row=bas2, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas2, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas2, start_column=1,
                        end_row=bas2, end_column=4)
        bas2 += 1
        for col, hd in enumerate(["Alt Sınıf", "Adet", "Açıklama", ""],
                                  start=1):
            c = ws1.cell(row=bas2, column=col, value=hd)
            c.font = baslik_font
            c.fill = baslik_fill
            c.alignment = Alignment(horizontal="center")
        kat_aciklama = {
            "STANDART":       "Genel enteral solüsyon (Resource, Nutridrink, Nutren, Fresubin, Evolvia, Ensure...)",
            "DIYABETIK":      "Diyabet hastasına uyarlanmış (Diasip, Glucerna, Fresubin DB)",
            "RENAL":          "Böbrek hastasına uyarlanmış (Nepro HP / LP)",
            "PEDIATRIK":      "Çocuk hastalar (Junior varyantları, Pediasure, Infatrini, Modulen IBD)",
            "IMMUNONUTRIYON": "Bağışıklık destekleyici (Impact, Prosure, Cubitan, Abound, Juven)",
            "DIGER":          "Yukarıdaki sınıflara girmeyen enteral ürün",
        }
        for i, k in enumerate(["STANDART", "DIYABETIK", "RENAL",
                                "PEDIATRIK", "IMMUNONUTRIYON", "DIGER"],
                               start=bas2 + 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=kategori_sayac.get(k, 0)
                     ).alignment = Alignment(horizontal="center")
            ws1.cell(row=i, column=3, value=kat_aciklama[k])

        # Notlar
        bas3 = bas2 + 8
        ws1.cell(row=bas3, column=1, value="NOTLAR")
        ws1.cell(row=bas3, column=1).font = Font(bold=True, color="FFFFFF")
        ws1.cell(row=bas3, column=1).fill = baslik_fill
        ws1.merge_cells(start_row=bas3, start_column=1,
                        end_row=bas3, end_column=4)
        notlar = [
            "• UYGUN = Endikasyon + uzman branş + (varsa) rapor kodu örtük endikasyon doğrulandı",
            "• UYGUN DEĞİL = Rapor yok veya endikasyon raporda bulunamadı — manuel inceleme",
            "• ŞÜPHELİ = Hesaplama hatası veya eksik veri — manuel kontrol önerilir",
            "• ATLANDI = SUT denetimi gerekmeyen satır",
            "• Enteral dışı = ATC V06D* / ürün ailesi dışı — bu butonun kapsamına girmiyor",
            "• Kalori hesabı: ilaç adından kcal/birim parse, reçete günlük dozu ile çarpım. "
            "Hasta kilosu reçete/rapor metninde varsa kcal/kg/gün hesaplanır (hedef 25-35).",
            "• Rapor kodu 15.* veya 02.* başlıyorsa Medula örtük endikasyon olarak değerlendirilir.",
        ]
        for i, n in enumerate(notlar, start=bas3 + 1):
            ws1.cell(row=i, column=1, value=n)
            ws1.merge_cells(start_row=i, start_column=1,
                            end_row=i, end_column=8)

        # Sütun genişlikleri (özet)
        for col, w in enumerate([34, 22, 60, 12], start=1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ────────── SAYFA 2: ENTERAL REÇETELERİ ──────────
        ws2 = wb.create_sheet("Enteral Reçeteleri")
        kolonlar = [
            ("rec_tar",          "Reç.Tarih",         12),
            ("rec_no",           "Reçete No",         18),
            ("hasta",            "Hasta",             24),
            ("tc",               "TC",                13),
            ("yas",              "Yaş",               6),
            ("cins",             "Cin.",              6),
            ("doktor",           "Doktor",            22),
            ("brans",            "Branş",             18),
            ("ilac",             "İlaç",              30),
            ("etkin",            "Etken Madde",       22),
            ("atc",              "ATC",               10),
            ("verdict_alt_sinif", "Alt Sınıf",         16),
            ("rap_kod",          "Rapor Kod",         11),
            ("rec_doz",          "Reçete Doz",        14),
            ("rap_doz",          "Rapor Doz",         14),
            ("kutu",             "Kutu",              6),
            ("msj",              "Msj",               7),
            ("uyari",            "Uyarı Kod",         18),
            ("medula_msj",       "Medula Msj",        30),
            ("rec_tesh",         "Reçete Teşhis",     30),
            ("rap_tesh",         "Rapor Teşhis",      30),
            ("rec_ack",          "Reçete Açıklama",   30),
            ("rap_ack",          "Rapor Açıklama",    30),
            ("verdict_sut",      "Uygulanan SUT",     30),
            ("verdict_aranan",   "Aranan İbare",      28),
            ("verdict_bulunan",  "Bulunan Metin",     28),
            ("verdict_detaylar", "Detaylar (kcal/endikasyon/uzman)", 42),
            ("verdict_uyari",    "Uyarı",             34),
            ("verdict_detay",    "Açıklama",          50),
            ("verdict",          "SONUÇ",             14),
        ]
        for c, (_kod, baslik, _g) in enumerate(kolonlar, 1):
            cell = ws2.cell(row=1, column=c, value=baslik)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for ri, s in enumerate(denetlenen_satirlar, start=2):
            for ci, (kod, _baslik, _g) in enumerate(kolonlar, start=1):
                deger = s.get(kod, "")
                cell = ws2.cell(row=ri, column=ci, value=str(deger))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            verdict = s.get("verdict") or ""
            renk = VERDICT_RENK.get(verdict)
            if renk:
                son_col = len(kolonlar)
                vcell = ws2.cell(row=ri, column=son_col)
                vcell.fill = PatternFill("solid", fgColor=renk)
                vcell.font = Font(bold=True)
                vcell.alignment = Alignment(horizontal="center", vertical="center")

        for ci, (_kod, _baslik, gen) in enumerate(kolonlar, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = gen
        ws2.auto_filter.ref = ws2.dimensions

        # ────────── SAYFA 3: KAPSAM DIŞI (atlanmış) ──────────
        ws3 = wb.create_sheet("Kapsam Dışı (Atlanan)")
        ws3.cell(row=1, column=1,
                 value="Aşağıdaki ilaçlar enteral beslenme solüsyonu "
                       "(V06D* / ürün ailesi) sınıfına girmediği için "
                       "enteral butonu KAPSAMI DIŞINDA bırakıldı.").font = (
            Font(italic=True, color="546E7A"))
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        atlanan_kolonlar = [
            ("rec_tar", "Reç.Tarih", 12),
            ("rec_no",  "Reçete No", 18),
            ("hasta",   "Hasta",     24),
            ("ilac",    "İlaç",      30),
            ("etkin",   "Etken",     22),
            ("atc",     "ATC",       10),
        ]
        for c, (_k, b, _g) in enumerate(atlanan_kolonlar, 1):
            cell = ws3.cell(row=3, column=c, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = Alignment(horizontal="center")
        ri = 4
        for s in self.tum_satirlar:
            kategori = self._enteral_kategori(
                s.get("ilac"), s.get("etkin"), s.get("atc"))
            if kategori != "NONE":
                continue
            for ci, (kod, _b, _g) in enumerate(atlanan_kolonlar, 1):
                ws3.cell(row=ri, column=ci, value=str(s.get(kod, "")))
            ri += 1
        for ci, (_k, _b, gen) in enumerate(atlanan_kolonlar, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = gen
        ws3.freeze_panes = "A4"

        wb.save(path)
        return path

    # ----------------------------------------------------------- DURUM
    def _durum_yaz(self, msg: str):
        try:
            self.durum_bar.config(text=msg)
        except Exception:
            pass

    def _kapat(self):
        try:
            self._state_kaydet()
        except Exception:
            pass
        try:
            if self.db:
                self.db.kapat()
        except Exception:
            pass
        if self.ana_menu_callback:
            try:
                self.root.destroy()
                self.ana_menu_callback()
                return
            except Exception:
                pass
        self.root.destroy()


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    root = tk.Tk()
    AylikReceteSorguGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
